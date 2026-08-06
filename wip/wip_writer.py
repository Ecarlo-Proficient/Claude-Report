#!/usr/bin/env python3
"""
wip_writer.py — the shared WIP report ENGINE.

Everything that turns CpRow objects into a formatted, edit-tracked, audited
'Test-*' tab lives here: the CpRow data model, the column layout (COLS), all
formatting, write_test_cp, the change audit, the owner-edit baselines, and the
QC check.

WHY THIS FILE EXISTS (2026-08-04): this engine used to live inside
cp_wip_reader.py, so every division tool that wanted the same formatting did
`import cp_wip_reader as CP` — a tool importing a tool, which the repo rules
forbid (shared/ is the only importable common code; tools never import tools).
That buried MFD and RP logic in a file named "cp" and let the layout drift.
The engine now has its own honest home; the readers import it, not each other:
  cp_wip_reader.py   — CP folder scan / draws / takeoffs
  rp_wip_reader.py   — the owner's RP WIP file
  master_wip_test.py — MFD off 'WIP Master' + orchestrates all three
This file is pure output machinery — it never reads a division's source.
"""
from __future__ import annotations

import datetime as dt
import logging
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name.*")
warnings.filterwarnings("ignore", message=".*LibreSSL.*")

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths
from shared import qbo_api

from wip_excel_guard import (
    ALLOWED_WRITE_SHEETS,
    WipWriteDenied,
    assert_write_allowed,
    open_wip_workbook_for_write,
)

WIP_EXCEL_PATH = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx",
)
TEST_TAB = "Test - CP"              # default write target (allow-listed)
MASTER_TITLE_SHEET = "WIP Master"   # read-only: the formatting reference sheet

# Realm (company id) captured by a reader's enrich_with_qbo so this writer can
# build QBO deep links on the number cells. Set via `wip_writer.QBO_REALM = …`.
QBO_REALM = ""

log = logging.getLogger("wip_writer")

class _Term:
    """Small ANSI helper — degrades gracefully when stdout isn't a TTY."""
    ENABLED = sys.stdout.isatty()

    RESET = "\033[0m"     if ENABLED else ""
    BOLD  = "\033[1m"     if ENABLED else ""
    DIM   = "\033[2m"     if ENABLED else ""
    CYAN  = "\033[36m"    if ENABLED else ""
    GREEN = "\033[32m"    if ENABLED else ""
    AMBER = "\033[33m"    if ENABLED else ""
    RED   = "\033[31m"    if ENABLED else ""
    GRAY  = "\033[90m"    if ENABLED else ""

    @staticmethod
    def color(code: str, text: str) -> str:
        return f"{code}{text}{_Term.RESET}"


_ANSI_RE = re.compile(r"\033\[[0-9;]*m")

def _vlen(s: str) -> int:
    """Visible length (strips ANSI escape codes so padding math works)."""
    return len(_ANSI_RE.sub("", s))

def _rpad(s: str, width: int) -> str:
    """Right-pad to `width` visible chars (for left-aligned columns)."""
    return s + " " * max(0, width - _vlen(s))

def _lpad(s: str, width: int) -> str:
    """Left-pad to `width` visible chars (for right-aligned columns)."""
    return " " * max(0, width - _vlen(s)) + s


def _section(title: str) -> None:
    """Print a section header."""
    print()
    print(_Term.color(_Term.BOLD + _Term.CYAN, f"▸ {title}"))
    print(_Term.color(_Term.DIM, "─" * (len(title) + 2)))


def _kv(label: str, value, indent: int = 2) -> None:
    """Print a key/value pair aligned."""
    print(f"{' ' * indent}{label:<18} {value}")


def _fmt_money(v: Optional[float]) -> str:
    """Money as a plain (uncolored) string. Coloring happens at pad time."""
    if v is None:
        return "—"
    return f"${v:,.2f}"


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "—"
    return f"{v*100:.1f}%"


def _dim_if_dash(s: str) -> str:
    """Dim the em-dash placeholder so real numbers pop; leave money plain."""
    return _Term.color(_Term.DIM, s) if s == "—" else s


# ─────────────────────── config / paths ────────────────────────────


@dataclass
class CpRow:
    """One row per CP project, ready to write to Test - CP."""
    project_num: str
    project_name: str                # from folder name after " - "
    is_completed: bool               # from Completed Projects/ subfolder
    base_contract: Optional[float]   # from Commercial Proposal Grand Total (audit)
    co_revenue: Optional[float]      # approved COs from the draw (G702 Line 2);
                                     # None until Draw #1 (no draw ⇒ no COs yet)
    base_etc: Optional[float]        # from Bid!AP1961 (audit — pre-CO)
    billed_to_date: Optional[float]  # QBO P&L income = GROSS billed (incl retainage)
    costs_to_date: Optional[float]   # from QBO P&L COGS + Expenses
    retainage_held: Optional[float] = None  # gross billed − net collectible (retainage receivable)
    status_flags: List[str] = field(default_factory=list)  # TRUE flags only: the script
                                          # hit something it could not confirm as fact and a
                                          # human must fix it (unreadable takeoff, QBO failure,
                                          # ambiguous proposal). NOT business observations —
                                          # the WIP report shows over-budget / CO $ itself.
    notes: List[str] = field(default_factory=list)  # informational; the script IS certain
                                          # (e.g. 'Draw #6…', 'No draw yet') — never a flag.
    takeoff_path: Optional[Path] = None   # audit trail: first included takeoff (hyperlink anchor)
    included_takeoffs: List[Path] = field(default_factory=list)  # takeoff(s) summed into this row
    folder_path: Optional[Path] = None   # explicit folder for the project-name link
                                          # (RP sets this; CP falls back to takeoff_path.parent)
    draw_num: Optional[int] = None       # latest draw # that sourced contract/billed/retainage
                                          # (None = pre-Draw#1 → takeoff proposal + QBO instead)
    draw_path: Optional[Path] = None     # the latest draw's G702/G703 workbook (audit trail)
    qbo_customer_id: Optional[str] = None  # QBO customer id → deep links on Billed/Costs cells
    needs_review: bool = False           # number doesn't look right / flagged → red font in Excel
    client: Optional[str] = None         # builder/client display name (RP tab)
    home_type: Optional[str] = None      # 'Tract' / 'Custom' (RP tab)
    why_link: Optional[str] = None       # path to the justification workbook (temp WHY column)
    why_fragment: Optional[str] = None   # "#'SHEET'!A<row>" — jump straight to this line's row
    src_link: Optional[str] = None       # source workbook the numbers came from (PROJECT # cell link)
    src_fragment: Optional[str] = None   # "#'SHEET'!C<row>" — exact source row (the user 2026-07-15)
    section: Optional[str] = None        # master-sheet grouping (SECTION column)
    co_cost_override: Optional[float] = None  # owner typed a CO cost onto the WIP

    @property
    def contract_price(self) -> Optional[float]:
        """Total Contract Price = base contract + summed COs. Returns None when
        the base is undetermined (e.g. Missing Grand Total / contract not yet
        decided) — a CO-only total would be misleading, so the cell stays blank
        and the flag explains why."""
        if self.base_contract is None:
            return None
        return self.base_contract + (self.co_revenue or 0.0)

    @property
    def co_cost_estimate(self) -> Optional[float]:
        """CO Cost is None until estimators add a real cost cell to the CO
        template. No proxy — the user's rule (2026-07-01): "no false numbers,
        don't populate if there is no source. If no CO costs, don't put
        it, just flag it."

        Property name kept as `co_cost_estimate` for column-mapping
        stability; semantics are now "CO cost sourced from the draw/template"
        and the value is None until such sourcing exists (blocked on a
        template change that adds a CO cost line).

        EXCEPTION (2026-08-03): if the owner types a CO cost straight onto the
        WIP, that override is carried in `co_cost_override` and wins — it is a
        real number from the person who owns the estimate, and until the CO
        template carries a cost line it is the only source there is."""
        return self.co_cost_override

    @property
    def etc(self) -> Optional[float]:
        """Revised ETC.
        - No COs → equals Base ETC (revised == base is truthful here).
        - COs present but no CO Cost data → defaults to Base ETC and the
          row is flagged provisional (the user 2026-07-02). This WIP is a
          MONITORING view, fully script-driven from takeoffs — no manual
          entry. Defaulting to Original ETC keeps the row alive and
          surfaces the over-budget signal (Costs > Original ETC) instead
          of going dark. The bounded error (excludes the CO's cost) is
          disclosed by the `% based on Original ETC — excludes CO cost`
          flag. We still never FABRICATE a CO cost — the base ETC is a
          real number, just scoped pre-CO. See [[project-cp-wip-takeoff-extraction]]."""
        if self.base_etc is None:
            return None
        if not self.co_revenue:
            return self.base_etc                # no COs → Revised = Base
        if self.co_cost_estimate is None:
            return self.base_etc                # CO Rev but no CO Cost → provisional (Original ETC)
        return self.base_etc + self.co_cost_estimate

    @property
    def has_missing_co_cost(self) -> bool:
        """True when the project has CO Revenue but no CO Cost data — a
        signal that Revised ETC (and the % / Earned / Over-Under derived
        from it) is PROVISIONAL: computed off Original ETC, which excludes
        the CO's cost. Drives the provisional flag."""
        return bool(self.co_revenue) and self.co_cost_estimate is None

    @property
    def status(self) -> str:
        if not self.status_flags:
            return "OK"
        return "; ".join(self.status_flags)

    @property
    def notes_text(self) -> str:
        """Informational notes joined for the NOTES column (blank when none)."""
        return "; ".join(self.notes)


# ─────────────────────── takeoff parsing ───────────────────────────


MASTER_FONT_NAME = "Tahoma"
MASTER_FONT_SIZE = 8
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")        # light gray header
HDR_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                bold=True, color="000000")
DATA_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE)
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")       # flag cell
FLAG_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                 italic=True, color="7F6000")
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")      # light yellow — SOURCED inputs
LINK_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                 color="0563C1", underline="single")     # Excel link
_SIDE = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
# Title block, copied cell-for-cell from the real 'WIP Master' sheet.
TITLE_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE, bold=True)
_MEDIUM = Side(style="medium", color="000000")
# Opens a column group: medium rule on the left, thin elsewhere.
_GROUP_BORDER = Border(left=_MEDIUM, right=_SIDE, top=_SIDE, bottom=_SIDE)
CURRENCY_FMT = '"$"#,##0_);[Red]\\("$"#,##0\\)'          # verbatim from 'WIP Master'
PCT_FMT = "0.00%"                                        # verbatim from 'WIP Master'

# Which fields render as currency vs percent (drives number_format in the write
# loop). Everything money-valued is currency; the two ratios are percent.
_MONEY_FIELDS = frozenset({
    "base_contract", "co_revenue", "contract_price", "base_etc",
    "co_cost_estimate", "etc", "original_profit", "costs_to_date",
    "cost_to_complete", "_earned_revenue", "profit_earned", "future_profit",
    "billed_to_date", "retainage_held", "left_to_bill", "overbillings",
    "underbillings", "job_borrow",
})
_PCT_FIELDS = frozenset({"_pct_complete", "gross_profit_pct"})

# SOURCED cells (raw numbers straight from the takeoff or QBO) get a yellow
# fill so they're visually distinct from the white CALCULATED cells (Excel
# formulas). The user 2026-07-02: "yellow = metrics that have sources, calculations
# leave white." Identifier/metadata columns (project #, name, status, flags,
# last synced) stay white — they're not metrics.
_SOURCE_FIELDS = frozenset({
    "base_contract",     # ORIGINAL CONTRACT
    "contract_price",    # TOTAL CONTRACT PRICE (Original + COs)
    "base_etc",          # ORIGINAL ESTIMATED COST
    "co_cost_estimate",  # CO COSTS — yellow + EMPTY = an input nobody filled
    "etc",               # ESTIMATED TOTAL COSTS (Revised ETC)
    "billed_to_date",    # QBO (gross income)
    "costs_to_date",     # QBO
    # trailing cross-check columns (also sourced → yellow)
    "co_revenue",        # APPROVED COs
    "retainage_held",    # QBO (gross − net collectible)
})


def _sheet_fragment(sheet_name: str, cell_ref: str = "A1") -> str:
    """Build the '#SheetName!CellRef' fragment appended to a file:// URI so
    Excel opens the workbook AND jumps to the specified cell.

    Sheet names with spaces or special chars need single-quote wrapping in
    the URL fragment (Excel convention). Cell ref defaults to A1."""
    return f"#'{sheet_name}'!{cell_ref}"


def _apply_hyperlink(cell, target: Optional[Path], fragment: str = "") -> None:
    """Attach a file:// hyperlink to a cell and apply link styling.

    `target` may be a Path to a file OR a directory. If None, this is a
    no-op so callers can safely pass None when the source is missing.
    `fragment` is optional (e.g. \"#'Bid'!AP1961\" to jump to a cell).

    **Never attaches a hyperlink to a cell whose value is None.** openpyxl's
    behavior in that case is to display the hyperlink URL as the cell text,
    which produces confusing "why is there a URL in this blank cell?"
    output. If the cell is empty, there's nothing to click anyway.

    The sheet/cell jump goes in the hyperlink's `location` ATTRIBUTE, never
    appended to the URI (2026-07-21): a fragment like #'Small Jobs'!C7 puts
    raw spaces into the .rels target URI — invalid XML that makes Excel
    demand a repair on open."""
    from openpyxl.worksheet.hyperlink import Hyperlink
    if target is None:
        return
    if cell.value is None or cell.value == "":
        return
    try:
        uri = target.as_uri()
    except (ValueError, OSError):
        # Path can't be converted to URI (unusual on macOS/Linux) — skip
        # rather than crash the whole run.
        return
    loc = fragment.lstrip("#") if fragment else None
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=uri,
                               location=loc or None)
    cell.font = LINK_FONT

# Full standard construction-WIP column set. Formulas verified 2026-07-02
# against Marcum LLP / Construction Executive, Wouch Maloney CPAs, CFMA/
# AICPA-referenced guidance. Billed to Date is GROSS (incl retainage); a
# separate RETAINAGE HELD column keeps the cash-collection timeline visible.
# Column order mirrors the team's WIP schedule — reads left→right as a story:
# the deal (contract, cost) → progress (billed, spent) → projection (cost to
# complete, profit) → recognition (% , earned) → billing position (over/under,
# left to bill) → profitability (GP%, future profit, job borrow). The four
# yellow inputs (contract, cost, billed, costs) + retainage lead; everything
# after is derived. Draw #/no-draw context lives in the NOTES column; FLAGS is
# reserved for genuine script must-fix issues (kept clean per the user 2026-07-02).
COLS = [
    # ── Identifiers ──
    ("PROJECT #",                                       12, "project_num"),
    ("PROJECT NAME",                                    30, "project_name"),
    # PROJECT FOLDER / DATA SOURCE link columns REMOVED (the user 2026-07-29:
    # file hyperlinks weigh the workbook down — the only links anywhere are
    # the QBO deep links on Billed/Costs).
    ("STATUS",                                          10, "_active_status"),
    # Columns are GROUPED left→right the way a WIP schedule reads (the user
    # 2026-08-03): CONTRACT → BUDGET → COSTS → PROFIT → BILLING → REMAINING →
    # ANALYSIS. `_COL_GROUPS` below maps each group to its first field so the
    # writer can draw a vertical rule at every boundary and print a column
    # guide at the bottom. Header NAMES are the 'WIP Master' vocabulary and
    # must not be renamed — five other tools read this tab by header name.
    # A change order moves BOTH sides — the contract AND the budget. The two
    # trios below mirror each other on purpose (researched 2026-08-03; Dean
    # Dorton lists "adding the change order to the contract amount but not
    # adjusting the project's cost budget" as a top WIP mistake). CO COSTS has
    # no source yet, so it sits EMPTY — a blank yellow input cell says
    # "nobody has costed this CO" more plainly than any flag would.
    # NOTE: 'TOTAL CONTRACT PRICE' IS the revised contract and 'ESTIMATED
    # TOTAL COSTS' IS the revised estimated cost. Those two names are read by
    # five other tools and must not be renamed.
    # ── CONTRACT ──
    ("ORIGINAL CONTRACT",                               15, "base_contract"),      # A
    ("APPROVED COs",                                    14, "co_revenue"),         # B
    ("TOTAL CONTRACT PRICE",                            16, "contract_price"),     # C = A+B
    # ── BUDGET ──
    ("ORIGINAL ESTIMATED COST",                         16, "base_etc"),           # D
    ("CO COSTS",                                        13, "co_cost_estimate"),   # E
    ("ESTIMATED TOTAL COSTS",                           16, "etc"),                # F = D+E
    # ── PROFIT ──
    ("ORIGINAL PROFIT",                                 14, "original_profit"),    # G = C-F
    ("GROSS PROFIT %",                                  12, "gross_profit_pct"),   # H = G/C
    # ── COSTS ──
    ("COSTS TO DATE",                                   15, "costs_to_date"),      # I
    ("COST TO COMPLETE",                                14, "cost_to_complete"),   # J = F-I
    ("PERCENT COMPLETE",                                11, "_pct_complete"),      # K = I/F
    # ── EARNED ──
    ("REVENUES EARNED TO DATE",                         16, "_earned_revenue"),    # L = C*K
    ("PROFIT EARNED TO DATE",                           15, "profit_earned"),      # M = G*K
    # ── BILLING ──
    ("BILLED TO DATE",                                  15, "billed_to_date"),     # N (gross, incl retainage)
    ("OVERBILLINGS",                                    14, "overbillings"),       # O = MAX(N-L,0)
    ("UNDERBILLINGS",                                   14, "underbillings"),      # P = MAX(L-N,0)
    ("RETAINAGE HELD",                                  14, "retainage_held"),     # Q
    ("LEFT TO BILL",                                    14, "left_to_bill"),       # R = C-N
    # ── ANALYSIS ──
    ("FUTURE PROFIT TO EARN",                           15, "future_profit"),      # S = G-M
    ("PURE JOB BORROW",                                 14, "job_borrow"),         # T
    ("LAST SYNCED",                                     18, "_last_synced"),
    # ONE commentary column (the user 2026-07-31: "stick to one" — NOTES and
    # FLAGS were two columns saying overlapping things). Informational notes
    # (Draw #, no-draw, the owner's ACTION text) and genuine script must-fix
    # flags now share this cell; the cell turns yellow/italic when it carries
    # a flag.
    ("NOTES",                                           60, "_notes_all"),
]

# Column groups → (group label, first field of the group). Drives the vertical
# rule between groups and the COLUMN GUIDE block at the bottom of the sheet.
_COL_GROUPS = [
    ("CONTRACT", "base_contract"),
    ("BUDGET",   "base_etc"),
    ("PROFIT",   "original_profit"),
    ("COSTS",    "costs_to_date"),
    ("EARNED",   "_earned_revenue"),
    # LEFT TO BILL sits INSIDE the billing box (the user 2026-08-03) — it is a
    # billing position, not a separate section.
    ("BILLING",  "billed_to_date"),
    ("ANALYSIS", "future_profit"),
]

# Column guide printed at the bottom: field → (letter, what it does). Mirrors
# the reference WIP the user supplied — every column says how it is derived so
# anyone reading the report can follow the maths (the user 2026-08-03).
_COL_LETTERS = {
    "base_contract": ("A", "original contract, before change orders"),
    "co_revenue": ("B", "approved change orders"),
    "contract_price": ("C", "C = A + B   the REVISED contract"),
    "base_etc": ("D", "original estimated cost — the bid budget"),
    "co_cost_estimate": ("E", "cost of the approved change orders "
                              "(blank = the CO has not been costed)"),
    "etc": ("F", "F = D + E   the REVISED estimated cost"),
    "original_profit": ("G", "G = C − F"),
    "gross_profit_pct": ("H", "H = G ÷ C"),
    "costs_to_date": ("I", "job costs booked in QBO"),
    "cost_to_complete": ("J", "J = F − I"),
    "_pct_complete": ("K", "K = I ÷ F"),
    "_earned_revenue": ("L", "L = C × K"),
    "profit_earned": ("M", "M = G × K"),
    "billed_to_date": ("N", "gross billed, retainage included"),
    "overbillings": ("O", "O = MAX(N − L, 0)  billed ahead of earned"),
    "underbillings": ("P", "P = MAX(L − N, 0)  earned ahead of billed"),
    "retainage_held": ("Q", "gross billed − net collectible"),
    "left_to_bill": ("R", "R = C − N"),
    "future_profit": ("S", "S = G − M"),
    "job_borrow": ("T", "T = MAX(J − R, 0)  cash this job borrows"),
}


# Fields written as Excel FORMULAS (not static values) so any input change
# auto-recalculates and the formula is visible in the formula bar for audit.
# Mirrors the project-pnl workbook's formula-driven design.
FORMULA_FIELDS: frozenset = frozenset({
    # Total Contract Price (contract_price) + Estimated Total Costs (etc) are
    # now written as VALUES (the yellow inputs, = Original+COs / Revised ETC via
    # the CpRow properties); the derived columns below reference those cells.
    "original_profit",   # ORIGINAL PROFIT   = Total Contract − Estimated Costs
    "gross_profit_pct",  # GROSS PROFIT %    = Original Profit / Revised Contract
    "cost_to_complete",  # COST TO COMPLETE  = Revised ETC − Costs to Date
    "_pct_complete",     # % COMPLETE        = Costs / Revised ETC
    "_earned_revenue",   # REVENUES EARNED   = Revised Contract × (Costs / Revised ETC)
    "profit_earned",     # PROFIT EARNED     = Earned Revenue − Costs to Date
    "future_profit",     # FUTURE PROFIT     = Original Profit − Profit Earned
    "left_to_bill",      # LEFT TO BILL      = Revised Contract − Billed
    "overbillings",      # OVERBILLINGS      = MAX(Billed − Earned, 0)
    "underbillings",     # UNDERBILLINGS     = MAX(Earned − Billed, 0)
    "job_borrow",        # PURE JOB BORROW   = MAX(Cost to Complete − Left to Bill, 0)
})

# The two ROLL-UPS. They are normally written as VALUES because four tools read
# this workbook with data_only=True, and an openpyxl-written formula carries no
# cached value — it would read back as None and silently zero them (exactly how
# MFD's budget went blank, 2026-08-03).
#
# On the tabs the owner EDITS they are written as live formulas instead (the
# user 2026-08-04: "etc total are not formulas so if i edit the og etc and the
# revised co etc it doesn't add up"), so typing a CO cost updates the revised
# total on the spot. Test-Master keeps values: it is locked, nobody types into
# it, and it is the tab every other tool reads.
_LIVE_ROLLUP_FIELDS = frozenset({"contract_price", "etc"})


def _build_formula(field_name: str, row_num: int,
                   col_letter_by_field: Dict[str, str]) -> str:
    """Build the Excel formula string for a derived-field cell in row `row_num`.

    Every reference is resolved by FIELD NAME → column letter (dynamic), so
    the formulas survive any column reordering. Blank inputs propagate to a
    blank output (never a false zero or #DIV/0), matching the Python
    None-cascades."""
    def ref(f: str) -> str:
        return col_letter_by_field[f] + str(row_num)

    F  = ref("contract_price")     # Total Contract Price (value)
    I  = ref("etc")                # Estimated Total Costs (value)
    J  = ref("billed_to_date")     # Billed (gross)
    K  = ref("costs_to_date")      # Costs to Date
    OP = ref("original_profit")    # Original Profit
    CTC = ref("cost_to_complete")  # Cost to Complete
    ER = ref("_earned_revenue")    # Revenues Earned to Date
    PE = ref("profit_earned")      # Profit Earned to Date
    LTB = ref("left_to_bill")      # Left to Bill

    if field_name == "contract_price":       # C = A + B (blank when both are)
        A, B = ref("base_contract"), ref("co_revenue")
        return f'=IF(AND({A}="",{B}=""),"",{A}+{B})'
    if field_name == "etc":                  # F = D + E (blank when both are)
        D, E = ref("base_etc"), ref("co_cost_estimate")
        return f'=IF(AND({D}="",{E}=""),"",{D}+{E})'
    if field_name == "original_profit":
        return f'=IF(OR({F}="",{I}=""),"",{F}-{I})'
    if field_name == "gross_profit_pct":
        return f'=IF(OR({OP}="",{F}="",{F}=0),"",{OP}/{F})'
    if field_name == "cost_to_complete":
        return f'=IF(OR({I}="",{K}=""),"",{I}-{K})'
    if field_name == "_pct_complete":
        return f'=IF(OR({K}="",{I}="",{I}=0),"",{K}/{I})'
    if field_name == "_earned_revenue":
        return f'=IF(OR({F}="",{K}="",{I}="",{I}=0),"",{F}*{K}/{I})'
    if field_name == "profit_earned":
        return f'=IF(OR({ER}="",{K}=""),"",{ER}-{K})'
    if field_name == "future_profit":
        return f'=IF(OR({OP}="",{PE}=""),"",{OP}-{PE})'
    if field_name == "left_to_bill":
        return f'=IF(OR({F}="",{J}=""),"",{F}-{J})'
    if field_name == "overbillings":
        # BIE — billings in excess of earned revenue (liability). 0 if underbilled.
        return f'=IF(OR({J}="",{ER}=""),"",MAX({J}-{ER},0))'
    if field_name == "underbillings":
        # CIE — earned revenue in excess of billings (asset). 0 if overbilled.
        return f'=IF(OR({ER}="",{J}=""),"",MAX({ER}-{J},0))'
    if field_name == "job_borrow":
        # Pure Job Borrow — remaining costs in excess of remaining billing
        # capacity (Wouch Maloney CPAs): costs-to-complete beyond what's left
        # to bill = cash this job must borrow from others. 0 if none.
        return f'=IF(OR({CTC}="",{LTB}=""),"",MAX({CTC}-{LTB},0))'
    raise ValueError(f"No formula defined for field {field_name!r}")


def _row_display_value(row: CpRow, field_name: str, sync_ts: str):
    """Return the value to write into the cell for a NON-formula `field_name`
    (formula fields are written via _build_formula, never here)."""
    if field_name == "_active_status":
        return "Closed" if row.is_completed else "Active"
    if field_name == "why_link":
        return "why ⇗" if row.why_link else None
    if field_name == "_folder_link":
        return "folder ⇗" if (row.folder_path or row.takeoff_path) else None
    if field_name == "_source_link":
        return "source ⇗" if row.src_link else None
    if field_name == "_last_synced":
        return sync_ts
    if field_name == "_notes_all":
        # ONE commentary cell: the owner's ACTION text, the script's
        # informational notes, then any genuine must-fix flag.
        #
        # De-duplication is per SEGMENT, not per string (2026-08-03): the
        # owner's ACTION text arrives as one string that already contains
        # ' · ' separators, so whole-string matching let the same sentence
        # come back twice. Split everything to segments first, compare on a
        # normalised key, and keep the FIRST wording seen — the message is
        # never reworded, only the redundant 'note:' label is dropped (the
        # column is already called NOTES).
        parts, seen = [], set()
        for chunk in ([getattr(row, "action_note", None)] + list(row.notes)
                      + list(row.status_flags)):
            for seg in str(chunk or "").split(" · "):
                seg = seg.strip()
                for lead in ("RED:", "note:", "NOTE:"):
                    if seg.startswith(lead):
                        seg = seg[len(lead):].strip()
                if _MUTED_NOTE_RE.search(seg):
                    continue                  # known/accepted — not a finding
                key = re.sub(r"\s+", " ", seg).lower()
                if seg and key not in seen:
                    seen.add(key)
                    parts.append(seg)
        return " · ".join(parts) or None
    if field_name == "status":
        # RED numbers must explain themselves (the user 2026-07-16): a red
        # row with no script flag showed "OK" — surface the classify reason
        # here instead (tabs without a NOTES column had nowhere else).
        if row.status_flags:
            return "; ".join(row.status_flags)
        if row.needs_review:
            red = next((n[len("RED: "):] for n in row.notes
                        if n.startswith("RED: ")), None)
            return red or (row.notes[-1] if row.notes else "needs review")
        return "OK"
    return getattr(row, field_name, None)


# ───────────────── change audit (the user 2026-07-31) ─────────────
# Every sync must state, per division, what moved: jobs added, jobs removed,
# original contract/ETC changes, revised contract/ETC changes. The owner
# audits this on every run — it is never optional and never summarised away.
AUDIT_XLSX = Path.home() / "Downloads" / "WIP Changes.xlsx"

# Note segments the SCRIPT writes. Anything else in a NOTES cell was typed by
# a human and is carried forward across the full-replace (the user
# 2026-07-31: "be sure to preserve any notes").
# Notes for KNOWN, ACCEPTED conditions — suppressed rather than shown, because
# they describe how the takeoffs are built, not something anyone will act on.
# "proposal quotes PIERS but no PR cost": pier costs sit inside the Piers
# takeoff sheet's overall costs and were never broken out per code (the user
# 2026-08-03: "i don't need to see that").
_MUTED_NOTE_RE = re.compile(r"quotes\s+PIERS\s+but\s+no\s+PR\s+cost", re.I)

_SCRIPT_NOTE_RE = re.compile(
    r"^(Draw #|No draw yet|QBO |No QBO project|Duplicate line in the RP file|"
    r"No budget \(ETC\)|Contract/ETC from |Data integrity:|On the .* schedule|"
    r"% based on Original ETC|RED: )", re.IGNORECASE)


# ─────────── owner edits: baseline + auto-colour (2026-08-03) ──────────
# The owner edits the WIP directly and wants three things: the cell to colour
# ITSELF the moment he changes it, his value to survive the next sync, and his
# notes/comments never to be lost.
#
# Mechanism: every editable input is mirrored into a HIDDEN column holding
# exactly what the script last wrote (the "baseline"). Excel conditional
# formatting compares the live cell to its baseline — differ ⇒ red. That runs
# inside Excel with no macro, so it fires the instant he types. On the next
# sync the same comparison tells the script which cells he overrode, and those
# values are carried forward instead of being overwritten.
#
# Overrides are limited to the SOURCED inputs. Everything else on the sheet is
# an Excel formula off these, so an override flows through the whole row and
# the arithmetic stays consistent.
_OVERRIDE_FIELDS = (
    "base_contract", "co_revenue", "base_etc", "co_cost_estimate",
    "billed_to_date", "costs_to_date", "retainage_held",
)
_BASE_PREFIX = "«base» "          # hidden baseline column header prefix
EDIT_FILL = PatternFill("solid", fgColor="FFC7CE")   # Excel's standard "bad"
EDIT_FONT = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                 color="FF0000", bold=True)          # red = the owner changed it


def _apply_edit_formatting(ws, cols_, hdr_row: int, first_row: int,
                           last_row: int, src_by_row: Dict[int, dict]) -> int:
    """Mirror each editable input into a hidden baseline column and add the
    conditional-format rule that reddens any cell differing from it.

    The baseline holds what the DATA SOURCES say — never the owner's override.
    If it held the override, the cell would match its baseline on the next run,
    the red would clear and the script would quietly put the source value back.
    Keeping the source in the baseline means an override stays marked and stays
    honoured until the source itself catches up with it.

    Returns the first column index used by the baseline block."""
    from openpyxl.formatting.rule import FormulaRule
    if last_row < first_row:
        return len(cols_) + 1
    base_col = len(cols_) + 2                   # one spacer column
    idx = {f: i + 1 for i, (_l, _w, f) in enumerate(cols_)}
    n = 0
    for field in _OVERRIDE_FIELDS + ("_notes_all",):
        c = idx.get(field)
        if not c:
            continue
        bcol = base_col + n
        n += 1
        bL, cL = get_column_letter(bcol), get_column_letter(c)
        ws.cell(hdr_row, bcol, f"{_BASE_PREFIX}{field}").font = DATA_FONT
        for r in range(first_row, last_row + 1):
            ws.cell(r, bcol, (src_by_row.get(r) or {}).get(field))
        ws.column_dimensions[bL].hidden = True
        if field == "_notes_all":
            continue                    # notes are preserved, not reddened
        ws.conditional_formatting.add(
            f"{cL}{first_row}:{cL}{last_row}",
            FormulaRule(formula=[f"{cL}{first_row}<>{bL}{first_row}"],
                        fill=EDIT_FILL, font=EDIT_FONT, stopIfTrue=False))
    return base_col


def read_owner_edits(ws, hdr_row: Optional[int],
                     cols_: Optional[List] = None) -> Dict[str, dict]:
    """Cells the owner changed since the last sync: {project #: {field: value}}.
    A cell differing from its hidden baseline IS an owner edit — that is the
    same test Excel used to colour it red, so what the script preserves and
    what he sees marked are always the same set."""
    if not hdr_row:
        return {}
    idx = {ws.cell(hdr_row, c).value: c
           for c in range(1, (ws.max_column or 0) + 1)}
    pcol = idx.get("PROJECT #")
    if not pcol:
        return {}
    label_of = {f: l for l, _w, f in COLS}
    pairs = []
    for f in _OVERRIDE_FIELDS:
        bcol = idx.get(f"{_BASE_PREFIX}{f}")
        ccol = idx.get(label_of.get(f))
        if bcol and ccol:
            pairs.append((f, bcol, ccol))
    # TRUST CHECK (2026-08-03): only believe the baselines if the tab was last
    # written by THIS layout. Several tools write these tabs, and after a
    # foreign write the numbers sitting under our headers came from a different
    # pipeline — every difference would then look like an owner edit and get
    # locked in as one. Seen for real: a run of rp_wip_reader left 'Test - RP'
    # in its own layout and the next sync reported 28 phantom edits.
    if cols_ is not None:
        want = {f for _l, _w, f in cols_ if f in _OVERRIDE_FIELDS}
        if want - {f for f, _b, _c in pairs}:
            print(_Term.color(_Term.AMBER,
                  "  ⚠ this tab was last written by a DIFFERENT layout — its "
                  "baselines aren't ours, so edit detection is skipped for "
                  "this run (your next edits will track normally)"))
            return {}
    if not pairs:
        return {}
    out = {}
    for r in range(hdr_row + 1, (ws.max_row or 0) + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip() == "TOTALS":
            break
        pnum = ws.cell(r, pcol).value
        if not pnum:
            continue
        for field, bcol, ccol in pairs:
            cur, base = ws.cell(r, ccol).value, ws.cell(r, bcol).value
            if isinstance(cur, str) and cur.startswith("="):
                continue                       # a formula, not a typed value
            cn, bn = _num(cur), _num(base)
            if cn is None and bn is None:
                continue
            if cn is None or bn is None or abs(cn - bn) >= 0.01:
                out.setdefault(str(pnum).strip().upper(), {})[field] = cn
    return out


def _division(pnum: str) -> str:
    """MFD / CP / RP from the project number (RP7234-FTW → RP)."""
    p = (pnum or "").strip().upper()
    for d in ("MFD", "CP", "RP"):
        if p.startswith(d):
            return d
    return "OTHER"


def _num(v):
    """Cell → float, or None (formulas/blanks/text read back as None)."""
    return float(v) if isinstance(v, (int, float)) else None


def _snapshot_tab(ws, hdr_row: Optional[int]) -> Dict[str, dict]:
    """The tab as it stands BEFORE the rewrite: per project #, the numbers the
    audit compares against plus its NOTES text. Read straight off the sheet so
    the baseline is what the owner last saw, not what any script remembers."""
    if not hdr_row:
        return {}
    idx = {ws.cell(hdr_row, c).value: c
           for c in range(1, (ws.max_column or 0) + 1)}
    pcol = idx.get("PROJECT #")
    if not pcol:
        return {}
    prior = {}
    for r in range(hdr_row + 1, (ws.max_row or 0) + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip() == "TOTALS":
            break                       # summary block — not data
        pnum = ws.cell(r, pcol).value
        if not pnum or not str(pnum).strip():
            continue
        def val(label):
            c = idx.get(label)
            return _num(ws.cell(r, c).value) if c else None
        cos = val("APPROVED COs") or 0.0
        rev_k = val("TOTAL CONTRACT PRICE")
        etc = val("ESTIMATED TOTAL COSTS")
        ncol = idx.get("NOTES")
        prior[str(pnum).strip().upper()] = {
            "name": ws.cell(r, idx["PROJECT NAME"]).value if idx.get("PROJECT NAME") else None,
            "rev_contract": rev_k,
            "orig_contract": (rev_k - cos) if rev_k is not None else None,
            "rev_etc": etc,
            "orig_etc": etc,      # identical until CO costs have a source
            "notes": (str(ws.cell(r, ncol).value).strip()
                      if ncol and ws.cell(r, ncol).value else ""),
        }
    return prior


def _changed(old, new) -> bool:
    if old is None and new is None:
        return False
    if old is None or new is None:
        return True
    return abs(float(old) - float(new)) >= 0.01


def audit_changes(prior: Dict[str, dict], rows: List["CpRow"]) -> Dict[str, dict]:
    """Diff the incoming rows against the tab's previous state, split by
    division. Returns {division: {added/removed/orig/revised}}."""
    out = {}
    def bucket(d):
        return out.setdefault(d, {"added": [], "removed": [],
                                  "orig": [], "revised": []})
    seen = set()
    for row in rows:
        p = row.project_num.strip().upper()
        seen.add(p)
        b = bucket(_division(p))
        was = prior.get(p)
        if was is None:
            b["added"].append((p, row.project_name, row.contract_price, row.etc))
            continue
        for key, label, old, new in (
                ("orig", "ORIGINAL CONTRACT", was["orig_contract"], row.base_contract),
                ("orig", "ORIGINAL ETC",      was["orig_etc"],      row.base_etc),
                ("revised", "REVISED CONTRACT", was["rev_contract"], row.contract_price),
                ("revised", "REVISED ETC",      was["rev_etc"],      row.etc)):
            if _changed(old, new):
                b[key].append((p, row.project_name, label, old, new))
    for p, was in prior.items():
        if p not in seen:
            bucket(_division(p))["removed"].append(
                (p, was.get("name"), was.get("rev_contract"), was.get("rev_etc")))
    return out


def _fmt(v) -> str:
    return "—" if v is None else f"${v:,.0f}"


def print_audit(audit: Dict[str, dict], tab_name: str) -> None:
    """The always-on change report. Loud, per division, never summarised away."""
    print()
    print(_Term.color(_Term.BOLD, f"  WIP CHANGE AUDIT — {tab_name}"))
    print("  " + "=" * 66)
    if not audit:
        print("  (no previous data on this tab — this run sets the baseline)")
        return
    total = 0
    for div in ("MFD", "CP", "RP", "OTHER"):
        b = audit.get(div)
        if not b or not any(b.values()):
            continue
        n = sum(len(v) for v in b.values())
        total += n
        print(_Term.color(_Term.BOLD + _Term.CYAN, f"\n  ── {div} ──"))
        for job, name, k, e in b["added"]:
            print(f"    + NEW      {job:<14} {str(name or '')[:30]:<30} "
                  f"contract {_fmt(k)} · ETC {_fmt(e)}")
        for job, name, k, e in b["removed"]:
            print(f"    − REMOVED  {job:<14} {str(name or '')[:30]:<30} "
                  f"was contract {_fmt(k)} · ETC {_fmt(e)}")
        for tag, key in (("ORIGINAL", "orig"), ("REVISED", "revised")):
            for job, name, label, old, new in b[key]:
                delta = (new or 0) - (old or 0)
                print(f"    ~ {label:<17} {job:<14} {_fmt(old)} → {_fmt(new)} "
                      f"({'+' if delta >= 0 else '−'}{_fmt(abs(delta))[1:]})")
        if not any(b.values()):
            print("    no changes")
    if total == 0:
        print("  No changes in any division since the last sync.")
    print()


def write_audit_xlsx(audit: Dict[str, dict], tab_name: str,
                     out_path: Path = AUDIT_XLSX) -> Optional[Path]:
    """One flat, plain sheet of every change so the owner can audit later.
    Overwrites the SAME file each run (never a v2). Skipped if it's open."""
    from openpyxl import Workbook
    if out_path.with_name("~$" + out_path.name).exists():
        print(_Term.color(_Term.AMBER,
              f"  ⚠ {out_path.name} is open in Excel — change file not written"))
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "CHANGES"
    ws.cell(1, 2, f"WIP CHANGES - {tab_name}").font = TITLE_FONT
    ws.cell(2, 2, f"REPORT DATE: {dt.date.today():%b %d, %Y}".upper()).font = TITLE_FONT
    hdr = ["DIVISION", "CHANGE", "JOB #", "JOB NAME", "FIELD", "OLD", "NEW", "CHANGE $"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(4, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = (10, 12, 16, 34, 20, 16, 16, 16)
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 5
    for div in ("MFD", "CP", "RP", "OTHER"):
        b = audit.get(div)
        if not b:
            continue
        for job, name, k, e in b["added"]:
            for c, v in enumerate([div, "NEW JOB", job, name, "CONTRACT",
                                   None, k, k], start=1):
                ws.cell(r, c, v)
            r += 1
            for c, v in enumerate([div, "NEW JOB", job, name, "ETC",
                                   None, e, e], start=1):
                ws.cell(r, c, v)
            r += 1
        for job, name, k, e in b["removed"]:
            for c, v in enumerate([div, "REMOVED", job, name, "CONTRACT / ETC", k, None, None], start=1):
                ws.cell(r, c, v)
            r += 1
        for key, tag in (("orig", "ORIGINAL"), ("revised", "REVISED")):
            for job, name, label, old, new in b[key]:
                delta = (new or 0) - (old or 0)
                for c, v in enumerate([div, tag + " CHANGE", job, name, label,
                                       old, new, delta], start=1):
                    ws.cell(r, c, v)
                r += 1
    for rr in range(5, r):
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(rr, c)
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
            if c >= 6:
                cell.number_format = CURRENCY_FMT
    if r == 5:
        ws.cell(5, 1, "No changes since the last sync.").font = DATA_FONT
    ws.freeze_panes = "A5"
    wb.save(out_path)
    return out_path


def _master_title_prefix(wb) -> str:
    """The company/report prefix from the real 'WIP Master' title cell (B1),
    e.g. 'PROFICIENT CONCRETE, LLC' out of '<company> - WIP REPORT'. Read at
    RUNTIME so the test tabs inherit the owner's own wording (and so no
    company name is hard-coded in this repo). '' when unreadable."""
    try:
        v = wb[MASTER_TITLE_SHEET].cell(row=1, column=2).value
    except (KeyError, AttributeError):
        return ""
    return str(v).split(" - ")[0].strip() if v else ""


def _find_header_row(ws) -> Optional[int]:
    """Locate the table header row (the one holding 'PROJECT #') in the first
    few rows — row 1 on plain tabs, below the title banner (and any legend
    block) on branded ones."""
    for r in range(1, min(ws.max_row or 0, 15) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            if ws.cell(r, c).value == "PROJECT #":
                return r
    return None


def _write_summary(ws, cols_, col_letter_by_field, data_start: int,
                   last_row: int, start_row: int) -> None:
    """WIP-master-style TOTALS row + FUTURE WIP CASH FLOW block (the user
    2026-07-16), written BELOW the table/appendix.

    TOTALS uses SUBTOTAL(109, …) over the table's data rows — it counts only
    VISIBLE rows, so filtering the table (e.g. hiding FTW BACKLOG or Closed)
    re-totals live. The cash-flow block derives everything from the TOTALS
    cells:  rev left = contract − earned · GP left = rev left − CTC ·
    net under/(over) = under − over · cash flow = GP left + net under."""
    L = col_letter_by_field.get
    tot = start_row

    # ── TOTALS row ──
    _sum_fields = [f for f in (
        "contract_price", "etc", "billed_to_date", "costs_to_date",
        "cost_to_complete", "original_profit", "_earned_revenue",
        "profit_earned", "overbillings", "underbillings", "left_to_bill",
        "future_profit", "job_borrow", "co_revenue", "retainage_held")
        if L(f)]
    for c in range(1, len(cols_) + 1):
        cell = ws.cell(row=tot, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = CELL_BORDER
    ws.cell(row=tot, column=1, value="TOTALS")
    for f in _sum_fields:
        cell = ws.cell(row=tot, column=column_index_from_string(L(f)))
        cell.value = f"=SUBTOTAL(109,{L(f)}{data_start}:{L(f)}{last_row})"
        cell.number_format = CURRENCY_FMT
    if L("_pct_complete") and L("costs_to_date") and L("etc"):
        c = ws.cell(row=tot, column=column_index_from_string(L("_pct_complete")))
        c.value = (f'=IF({L("etc")}{tot}=0,"",'
                   f'{L("costs_to_date")}{tot}/{L("etc")}{tot})')
        c.number_format = PCT_FMT
    if L("gross_profit_pct") and L("original_profit") and L("contract_price"):
        c = ws.cell(row=tot, column=column_index_from_string(L("gross_profit_pct")))
        c.value = (f'=IF({L("contract_price")}{tot}=0,"",'
                   f'{L("original_profit")}{tot}/{L("contract_price")}{tot})')
        c.number_format = PCT_FMT

    # ── FUTURE WIP CASH FLOW block ──
    # Label and amount sit SIDE BY SIDE in a bordered box (the user
    # 2026-08-03: "why is numbers far from this? where are the dividers and
    # lines?"). Label spans cols A:B, the amount C:D, so nothing is stranded
    # halfway across the sheet.
    _bold = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE, bold=True)
    vcol, vL = 3, get_column_letter(3)
    r0 = tot + 2
    lines = [
        ("FUTURE WIP CASH FLOW", None, True),
        ("TOTAL CONTRACT PRICE", f"={L('contract_price')}{tot}", False),
        ("REVENUE EARNED TO DATE", f"={L('_earned_revenue')}{tot}", False),
        ("REVENUE LEFT TO EARN", f"={vL}{r0 + 1}-{vL}{r0 + 2}", False),
        ("COST TO COMPLETE", f"={L('cost_to_complete')}{tot}", False),
        ("G.P. LEFT TO EARN", f"={vL}{r0 + 3}-{vL}{r0 + 4}", False),
        ("+/- NET UNDER/(OVERBILLINGS)",
         f"={L('underbillings')}{tot}-{L('overbillings')}{tot}", False),
        ("FUTURE WIP CASH FLOW", f"={vL}{r0 + 5}+{vL}{r0 + 6}", True),
    ]
    for k, (label, formula, bold) in enumerate(lines):
        r = r0 + k
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _bold if bold else DATA_FONT
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=vcol)
        if formula:
            vc.value = formula
            vc.number_format = CURRENCY_FMT
            vc.font = _bold if bold else DATA_FONT
            vc.alignment = Alignment(horizontal="right", vertical="center")
        if label == "FUTURE WIP CASH FLOW" and formula is None:
            for c in range(1, 5):                     # block title bar
                ws.cell(row=r, column=c).fill = HDR_FILL
        for c in range(1, 5):                         # box every cell
            ws.cell(row=r, column=c).border = CELL_BORDER
    # G.P. LEFT TO EARN margin % (of revenue left to earn), labelled so the
    # stray percentage isn't a mystery number floating beside the block.
    pr = r0 + 5
    pc = ws.cell(row=pr, column=5,
                 value=f'=IF({vL}{r0 + 3}=0,"",{vL}{r0 + 5}/{vL}{r0 + 3})')
    pc.number_format = PCT_FMT
    pc.font = DATA_FONT
    pc.alignment = Alignment(horizontal="right")
    pc.border = CELL_BORDER
    lbl = ws.cell(row=pr, column=6, value="of revenue left to earn")
    lbl.font = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE, italic=True)
    return r0 + len(lines)          # first free row under the block


def _write_bottom_notes(ws, cols_, start_row: int,
                        legend: Optional[List]) -> None:
    """The reference blocks that live UNDER the report (the user 2026-08-03:
    "put the rp legend in the bottom of the sheet"):

      COLUMN GUIDE — every money column, its letter and how it is derived, in
      the same CONTRACT/BUDGET/COSTS/… groups as the columns themselves, so a
      reader can follow the maths without asking.
      LEGEND — the caller's own rows (RP categories, colour meanings).

    Plain single-font cells throughout — inline rich text corrupts the file."""
    _bold = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE, bold=True)
    fields = [f for _l, _w, f in cols_]
    group_of = {}
    for gname, first in _COL_GROUPS:
        if first in fields:
            group_of[first] = gname
    r = start_row + 1
    ws.cell(row=r, column=1, value="COLUMN GUIDE").font = _bold
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = HDR_FILL
        ws.cell(row=r, column=c).border = CELL_BORDER
    r += 1
    for label, _w, field in cols_:
        if field not in _COL_LETTERS:
            continue
        if field in group_of:                       # group heading
            gc = ws.cell(row=r, column=1, value=f"── {group_of[field]} ──")
            gc.font = _bold
            r += 1
        letter, how = _COL_LETTERS[field]
        ws.cell(row=r, column=1, value=letter).font = _bold
        ws.cell(row=r, column=2, value=label).font = DATA_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(row=r, column=3, value=how).font = DATA_FONT
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = CELL_BORDER
        r += 1
    if legend:
        r += 1
        for txt, rgb, bold in legend:
            lc = ws.cell(row=r, column=1, value=txt)
            lc.font = Font(name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                           color=(rgb or "000000"), bold=bold)
            r += 1


def write_test_cp(rows: List[CpRow], wip_path: Path, dry_run: bool = False,
                  tab_name: str = TEST_TAB,
                  appendix: Optional[Tuple[str, List[CpRow]]] = None,
                  cols: Optional[List] = None,
                  default_filter_active: bool = False,
                  title: Optional[str] = None,
                  summary: bool = False,
                  qbo_links_only: bool = True,
                  legend: Optional[List] = None,
                  audit: bool = False,
                  audit_xlsx: Optional[Path] = None,
                  protect: bool = False,
                  live_formulas: bool = False) -> bool:
    """Write rows to the given WIP tab (default 'Test - CP'; RP passes
    'Test - RP'). Same structure/formatting for every division. Guarded by
    wip_excel_guard. Returns True if written, False if skipped (dry-run, or the
    file is open in Excel).

    `appendix` = (section title, rows): written BELOW the main table under a
    gray band — RP uses it for the FTW backlog (bid with the slab, not poured
    yet; the user 2026-07-14: separated at the bottom, they read as expected
    wins rather than in-progress jobs).

    `title` (the user 2026-07-16): report banner across the top ("WIP REPORT
    as of …") with the two banner rows reserved as logo space — embedded
    images (the logo) survive every sync (openpyxl round-trips them; the
    rewrite only touches cells). `summary` adds the WIP-master-style TOTALS
    row under the table (live SUBTOTALs — they follow the table filter) plus
    the FUTURE WIP CASH FLOW block derived from it.

    `qbo_links_only` (the user 2026-07-29, now the DEFAULT — file hyperlinks
    weigh the workbook down): suppress every file:// hyperlink (Synology
    folders, takeoffs, draws, source workbooks, WHY) — the only links on the
    report are the QBO deep links on the Billed/Costs cells. Pass False to
    restore the full click-to-verify link set.

    `legend` (the user 2026-07-31): rows of (text, font_rgb_or_None, bold)
    rendered under the banner, one plain single-font cell per row (rich text
    is banned — it corrupts the workbook). Rows carrying `cell_marks`
    ({field: rgb}) get the owner's colour re-applied to those cells — his
    verified/changed/verify-me marks survive the sync."""
    assert_write_allowed(tab_name)  # tripwire before we even open the workbook
    cols_ = cols or COLS

    if dry_run:
        _print_rows_table(rows, wip_path, tab_name)
        return False

    if not wip_path.exists():
        raise FileNotFoundError(
            f"WIP Excel not found at {wip_path}. "
            f"Check WIP_EXCEL_PATH env or verify OneDrive is synced."
        )

    # Never clobber a workbook that's open in Excel. Excel drops a
    # ~$<name> owner-lock file next to an open workbook (project-pnl
    # safe_save pattern, the user 2026-06-23). If present, skip with a clear
    # message rather than writing underneath the open file — a
    # last-writer-wins save would silently lose the sync or the user's edits.
    lock = wip_path.with_name("~$" + wip_path.name)
    if lock.exists():
        print(_Term.color(_Term.AMBER,
              f"  ⚠ {wip_path.name} looks OPEN in Excel — skipped the write to "
              f"avoid overwriting it. Close the file and re-run."))
        return False

    wb = open_wip_workbook_for_write(wip_path)
    try:
        if tab_name not in wb.sheetnames:
            wb.create_sheet(tab_name)
        ws = wb[tab_name]
        assert_write_allowed(ws.title)  # belt + suspenders

        # A SHEET-level AutoFilter cannot coexist with the Table's own filter:
        # Excel calls the workbook damaged ("We found a problem with some
        # content…") and repairs it on open. A previous tool's filter survives
        # the cell wipe — 'Test - RP' still carried A2:L69 from the old
        # 12-column layout (2026-07-31). Clearing the ref also drops the
        # hidden _xlnm._FilterDatabase defined name openpyxl derives from it.
        ws.auto_filter.ref = None
        # Conditional formatting is re-added below; without this reset the
        # rules would pile up one duplicate set per sync.
        from openpyxl.formatting.formatting import ConditionalFormattingList
        ws.conditional_formatting = ConditionalFormattingList()

        # PRESERVE USER CELL COMMENTS across the full-replace (the user
        # 2026-07-16: review notes typed on cells — e.g. "add the missing
        # $5k" on a contract price — must survive every sync). Harvest them
        # keyed by (PROJECT #, header label) before the wipe; re-attach to
        # the same project/column after the rewrite. A comment whose project
        # left the tab is PRINTED, never silently dropped.
        saved_comments = {}
        prior_hdr = _find_header_row(ws)
        # Baseline for the change audit + the owner's typed NOTES, both read
        # off the sheet BEFORE it is wiped (the user 2026-07-31).
        prior_state = _snapshot_tab(ws, prior_hdr)
        # Owner edits: any input differing from its hidden baseline. Same test
        # Excel used to redden the cell, so "what he sees marked" and "what the
        # script keeps" can never drift apart.
        owner_edits = read_owner_edits(ws, prior_hdr, cols_)
        # NOTES he typed = whatever the prior cell holds that the script did
        # NOT write last time (exact provenance from the baseline column, so a
        # note deleted at its source stays deleted instead of resurrecting).
        _bidx = ({ws.cell(prior_hdr, c).value: c
                  for c in range(1, (ws.max_column or 0) + 1)}
                 if prior_hdr else {})
        _nb = _bidx.get(f"{_BASE_PREFIX}_notes_all")
        _pcol = _bidx.get("PROJECT #")
        _script_notes = {}
        if _nb and _pcol:
            for r in range(prior_hdr + 1, (ws.max_row or 0) + 1):
                v = ws.cell(r, _pcol).value
                if v:
                    _script_notes[str(v).strip().upper()] = {
                        s.strip() for s in
                        str(ws.cell(r, _nb).value or "").split(" · ") if s.strip()}
        for _p, _s in prior_state.items():
            known = _script_notes.get(_p)
            # Exact provenance: we know precisely what the script wrote last
            # time, so anything else in the cell was typed by a human.
            _s["notes_exact"] = known is not None
            _s["manual_notes"] = [
                seg.strip() for seg in _s["notes"].split(" · ")
                if seg.strip()
                and (seg.strip() not in known if known is not None
                     else not _SCRIPT_NOTE_RE.match(seg.strip()))]
        if owner_edits:
            print(_Term.color(_Term.AMBER,
                  f"  ✎ {len(owner_edits)} job(s) carry cell edits you made on "
                  f"'{tab_name}' — keeping your values:"))
            for _p, _f in sorted(owner_edits.items()):
                print(f"      {_p}: " + ", ".join(
                    f"{k}={'blank' if v is None else format(v, ',.0f')}"
                    for k, v in _f.items()))
        _sect_rows = [r for _t, ap in ([appendix] if isinstance(appendix, tuple)
                                       else (appendix or []))
                      for r in (ap or [])]
        for row in list(rows) + _sect_rows:
            # Snapshot what the SOURCES say BEFORE any override is applied —
            # this is what goes in the baseline column.
            row.src_vals = {f: getattr(row, f, None) for f in _OVERRIDE_FIELDS}
            for _f, _v in (owner_edits.get(row.project_num.strip().upper())
                           or {}).items():
                setattr(row, "co_cost_override" if _f == "co_cost_estimate"
                        else _f, _v)
        hdr_labels = ({c: ws.cell(prior_hdr, c).value
                       for c in range(1, (ws.max_column or 0) + 1)}
                      if prior_hdr else {})
        pnum_col = next((c for c, v in hdr_labels.items() if v == "PROJECT #"), None)
        if pnum_col:
            for r in range(prior_hdr + 1, (ws.max_row or 0) + 1):
                pnum = ws.cell(r, pnum_col).value
                if not pnum:
                    continue
                for c, label in hdr_labels.items():
                    cm = ws.cell(r, c).comment
                    if cm and label:
                        saved_comments[(str(pnum).strip(), label)] = (cm.text, cm.author)

        # Full replace: clear existing rows AND explicitly wipe every cell
        # attribute (value, hyperlink, number_format, fill, font) up to the
        # prior extent so leftovers can't leak into new rows.
        # openpyxl's delete_rows sometimes leaves formatting or hyperlinks
        # behind on cells that were populated before — belt-and-suspenders
        # clear here to guarantee a clean slate.
        # Merged cells (a prior run's title banner) must be unmerged before
        # the clear — writing to a MergedCell raises in openpyxl.
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))
        prior_max_row = ws.max_row or 0
        prior_max_col = ws.max_column or 0
        off = 2 if title else 0      # banner rows above the header; the legend
                                     # lives at the BOTTOM (the user 2026-08-03)
        _sects = ([appendix] if isinstance(appendix, tuple) else (appendix or []))
        n_total = (off + len(rows) + sum(len(a[1]) + 3 for a in _sects if a[1])
                   + (14 if summary else 0)
                   + len(_COL_LETTERS) + len(_COL_GROUPS) + 4
                   + (len(legend) + 2 if legend else 0))
        for r in range(1, max(prior_max_row, n_total + 1) + 1):
            for c in range(1, max(prior_max_col, len(cols_)) + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.hyperlink = None
                cell.number_format = "General"
                cell.comment = None   # harvested above; stale ones must not linger
        if prior_max_row > 0:
            ws.delete_rows(1, prior_max_row)
        # Reset stale hidden flags — rows shift between runs, and a hidden
        # flag left on the wrong row would silently hide live data.
        # Row HEIGHTS reset too (2026-08-03): the header/banner moves between
        # layouts, and a leftover 30pt height on what is now a blank or legend
        # row shows up as a random tall gap.
        for _rd in ws.row_dimensions.values():
            _rd.hidden = False
            _rd.height = None

        # Title banner (the user 2026-07-16): "WIP REPORT as of …" across the
        # table width, two tall rows that double as the logo's parking space
        # (the logo image floats over the cells — the rewrite never touches
        # it, so it stays put run after run).
        hdr_row = 1 + off
        if title:
            # TITLE BLOCK — copied from the real 'WIP Master' sheet and
            # BINDING (the user 2026-07-31: "keep the same format as the
            # original WIP Master sheet"). B1 = company + report name,
            # B2 = REPORT DATE, both Tahoma 8 bold and LEFT-aligned, a medium
            # rule above row 1 and below row 2. No merge-and-center, no
            # oversized font, no custom row heights. Do not redesign this.
            prefix = _master_title_prefix(wb)
            ws.cell(row=1, column=2,
                    value=f"{prefix} - {title}" if prefix else title
                    ).font = TITLE_FONT
            ws.cell(row=2, column=2,
                    value=f"REPORT DATE: {dt.date.today():%b %d, %Y}".upper()
                    ).font = TITLE_FONT
            for c in range(1, len(cols_) + 1):
                ws.cell(row=1, column=c).border = Border(top=_MEDIUM)
                ws.cell(row=2, column=c).border = Border(bottom=_MEDIUM)

        # Header row — gray, bold, centered + wrapped, bordered. A MEDIUM
        # vertical rule opens each column group (CONTRACT / BUDGET / COSTS /
        # PROFIT / BILLING / REMAINING / ANALYSIS) so the groups read as boxes
        # the way the owner's reference WIP does (the user 2026-08-03).
        _group_starts = {f for _g, f in _COL_GROUPS}
        for c, (label, width, _key) in enumerate(cols_, start=1):
            cell = ws.cell(row=hdr_row, column=c, value=label)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = (_GROUP_BORDER if _key in _group_starts
                           else CELL_BORDER)
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.row_dimensions[hdr_row].height = 30
        ws.freeze_panes = f"A{hdr_row + 1}"

        sync_ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

        # Column indices for hyperlink attachment (1-based like openpyxl)
        # + column LETTERS for building cross-cell formulas.
        col_idx = {field: i + 1 for i, (_, _, field) in enumerate(cols_)}
        col_letter_by_field = {
            field: get_column_letter(i + 1)
            for i, (_, _, field) in enumerate(cols_)
        }

        src_by_row: Dict[int, dict] = {}

        def _emit(i: int, row: CpRow) -> None:
            # Baseline for NOTES = what the SCRIPT generates, before the
            # owner's own lines are folded back in. Otherwise his note would
            # look script-written next run and get dropped.
            src_by_row[i] = dict(getattr(row, "src_vals", None) or {})
            src_by_row[i]["_notes_all"] = _row_display_value(
                row, "_notes_all", sync_ts)
            # Re-attach any NOTES text a human typed on this job's row last
            # time (script segments regenerate themselves; only human text is
            # carried) — the user 2026-07-31: "be sure to preserve any notes".
            # With a baseline column we know EXACTLY what the script wrote last
            # time, so a note deleted at its source can't resurrect (it was in
            # the baseline) and a note he typed here is always kept. Only when
            # that provenance is missing do we fall back to skipping
            # file-sourced rows.
            _st = prior_state.get(row.project_num.strip().upper(), {})
            if _st.get("notes_exact") or not getattr(row, "notes_from_source",
                                                     False):
                for _seg in _st.get("manual_notes") or []:
                    if _seg not in row.notes:
                        row.notes.append(_seg)
            # Invariant guard: if CO Cost is populated but CO Revenue is
            # not, refuse to write the CO Cost. That's either a bug or a
            # corrupted state, and quietly writing a made-up cost would
            # be worse than surfacing the issue.
            if row.co_revenue is None and row.co_cost_estimate is not None:
                log.warning(
                    "%s: CO Cost populated ($%s) with no CO Revenue — "
                    "refusing to write CO Cost. Investigate parse logic.",
                    row.project_num, row.co_cost_estimate
                )
                row.status_flags.append("Data integrity: CO Cost without CO Rev — dropped")

            _formula_now = (FORMULA_FIELDS | _LIVE_ROLLUP_FIELDS
                            if live_formulas else FORMULA_FIELDS)
            for c, (_label, _width, field_name) in enumerate(cols_, start=1):
                if field_name in _formula_now:
                    # Derived cell — write an Excel formula referencing
                    # the input cells in the same row. Excel evaluates on
                    # open/edit so any input change auto-recalculates.
                    val = _build_formula(field_name, i, col_letter_by_field)
                else:
                    val = _row_display_value(row, field_name, sync_ts)
                cell = ws.cell(row=i, column=c, value=val)
                cell.border = (_GROUP_BORDER if field_name in _group_starts
                               else CELL_BORDER)
                cell.font = DATA_FONT          # master-sheet Tahoma 8 baseline
                if field_name in _MONEY_FIELDS:
                    cell.number_format = CURRENCY_FMT
                elif field_name in _PCT_FIELDS:
                    cell.number_format = PCT_FMT
                # Yellow = sourced input (raw from takeoff / QBO); white = calc.
                # A roll-up written as a live formula IS a calc — leave it white
                # so the yellow still means "a number came from a source".
                if field_name in _SOURCE_FIELDS and field_name not in _formula_now:
                    cell.fill = INPUT_FILL
                elif field_name in ("status", "_notes_all") and (
                        row.status_flags or row.needs_review):
                    cell.fill = FLAG_FILL
                    cell.font = FLAG_FONT

            # PROJECT FOLDER → project (Awarded Project) folder; DATA SOURCE →
            # where the numbers came from (GL row / master tab / takeoff). The
            # user 2026-07-16: links moved OFF the # and name cells into their
            # own columns so the identifiers can be selected/copied without
            # Excel navigating away. (file:// links still get rewritten by
            # Windows/OneDrive Excel — kept for the Mac, the trace point.)
            # Folder target: explicit folder_path (RP) else takeoff's parent (CP).
            if not qbo_links_only:
                link_folder = (row.folder_path if row.folder_path is not None
                               else (row.takeoff_path.parent if row.takeoff_path else None))
                if "_folder_link" in col_idx:
                    _apply_hyperlink(ws.cell(row=i, column=col_idx["_folder_link"]),
                                     link_folder)
                if "_source_link" in col_idx and row.src_link:
                    _apply_hyperlink(ws.cell(row=i, column=col_idx["_source_link"]),
                                     Path(row.src_link), row.src_fragment or "")

                # Number-cell links (the user 2026-07-13: every number click-to-verify):
                # Contract/COs → the draw workbook (takeoff pre-draw); ETC → takeoff;
                # Billed/Retainage → QBO customer page (all invoices on one screen);
                # Costs → QBO project-filtered P&L report.
                for _f, _tgt in (("contract_price", row.draw_path or row.takeoff_path),
                                 ("co_revenue", row.draw_path),
                                 ("etc", row.takeoff_path)):
                    if _f in col_idx:
                        _apply_hyperlink(ws.cell(row=i, column=col_idx[_f]), _tgt)
                if "why_link" in col_idx and row.why_link:
                    _apply_hyperlink(ws.cell(row=i, column=col_idx["why_link"]),
                                     Path(row.why_link),
                                     row.why_fragment or "")
            if row.qbo_customer_id and QBO_REALM:
                _cu = qbo_api.customer_url(row.qbo_customer_id, QBO_REALM)
                _pu = qbo_api.project_pl_url(row.qbo_customer_id, QBO_REALM)
                for _f, _u in (("billed_to_date", _cu), ("retainage_held", _cu),
                               ("costs_to_date", _pu)):
                    if _f not in col_idx:
                        continue
                    _c = ws.cell(row=i, column=col_idx[_f])
                    if _u and _c.value not in (None, ""):
                        _c.hyperlink = _u
                        _c.font = LINK_FONT

            # Review pass LAST so it wins over link styling: numbers that don't
            # look right (row.needs_review) render RED (the user 2026-07-13);
            # underline is kept where the cell carries a link.
            if row.needs_review:
                for _f in (_MONEY_FIELDS | _PCT_FIELDS):
                    if _f in col_idx:
                        _c = ws.cell(row=i, column=col_idx[_f])
                        _c.font = Font(
                            name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                            color="C00000",
                            underline=("single" if _c.hyperlink else None))

            # The owner's colour marks are applied LAST — they outrank link
            # and review styling (the user 2026-07-31: "keep all the notes
            # and colors since they mean something").
            for _f, _rgb in (getattr(row, "cell_marks", None) or {}).items():
                if _f in col_idx:
                    _c = ws.cell(row=i, column=col_idx[_f])
                    _c.font = Font(
                        name=MASTER_FONT_NAME, size=MASTER_FONT_SIZE,
                        color=_rgb, bold=True,
                        underline=("single" if _c.hyperlink else None))

        data_start = hdr_row + 1
        written_rows = {}                   # PROJECT # → sheet row (for comments)
        for i, row in enumerate(rows, start=data_start):
            _emit(i, row)
            written_rows[row.project_num] = i

        # Wrap the range in an Excel Table — gives filter/sort dropdowns and a
        # structured, clean look. Explicit gray header + borders above override
        # the table style, so it stays clean (no row stripes).
        last_row = len(rows) + hdr_row
        last_col = get_column_letter(len(cols_))
        for tname in list(ws.tables):
            del ws.tables[tname]          # drop any prior run's table first
        tbl_name = re.sub(r"[^A-Za-z0-9]", "", tab_name) or "WIP"  # unique per tab
        table = Table(displayName=tbl_name,
                      ref=f"A{hdr_row}:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        if default_filter_active and "_active_status" in col_idx:
            # Default view = Active only (the user 2026-07-14): apply the
            # STATUS filter on the table AND hide the Closed rows — Excel
            # shows them again with one filter click.
            from openpyxl.worksheet.filters import (AutoFilter, FilterColumn,
                                                     Filters)
            fc = FilterColumn(colId=col_idx["_active_status"] - 1,
                              filters=Filters(filter=["Active"]))
            table.autoFilter = AutoFilter(ref=f"A{hdr_row}:{last_col}{last_row}",
                                          filterColumn=[fc])
            for i, row in enumerate(rows, start=data_start):
                if row.is_completed:
                    ws.row_dimensions[i].hidden = True
        ws.add_table(table)

        # Appendix section BELOW the table (outside it, so its rows don't
        # pollute the table's filters): gray band title, then the same row
        # rendering as the main block.
        sections = ([appendix] if isinstance(appendix, tuple) else (appendix or []))
        next_row = last_row + 2             # one blank spacer row under the table
        for sect_title, ap_rows in sections:
            if not ap_rows:
                continue
            band = next_row
            for c in range(1, len(cols_) + 1):
                bc = ws.cell(row=band, column=c)
                bc.fill = HDR_FILL
                bc.border = CELL_BORDER
            t = ws.cell(row=band, column=1, value=sect_title)
            t.font = HDR_FONT
            ws.row_dimensions[band].height = 22
            for k, row in enumerate(ap_rows, start=band + 1):
                _emit(k, row)
                written_rows[row.project_num] = k
            next_row = band + len(ap_rows) + 2

        # Baseline mirror + the auto-colour rule (must run BEFORE the summary,
        # so it sees only real data rows).
        _apply_edit_formatting(ws, cols_, hdr_row, data_start, last_row,
                               src_by_row)

        if summary:
            next_row = _write_summary(ws, cols_, col_letter_by_field,
                                      data_start, last_row, next_row)
        _write_bottom_notes(ws, cols_, next_row, legend)

        # LOCK the finished report (the user 2026-08-03: Test-Master "should be
        # locked by default because it's a read only" roll-up of the CP/RP tabs
        # and the WIP Master MFD section). No password — Review ▸ Unprotect
        # Sheet is one click — so this stops accidental typing, it doesn't take
        # the sheet away. Filtering, sorting and selecting stay enabled.
        if protect:
            ws.protection.sheet = True         # no password: one-click unprotect
            ws.protection.autoFilter = False   # False = still allowed
            ws.protection.sort = False
            ws.protection.selectLockedCells = False
            ws.protection.selectUnlockedCells = False
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False

        # Re-attach the harvested user comments to their project/column.
        col_by_label = {label: c for c, (label, _w, _f) in enumerate(cols_, start=1)}
        for (pnum, label), (text, author) in saved_comments.items():
            r, c = written_rows.get(pnum), col_by_label.get(label)
            if r and c:
                ws.cell(row=r, column=c).comment = Comment(text, author or "")
            else:
                print(_Term.color(_Term.AMBER,
                      f"  ⚠ comment on {pnum} / {label} has no row this run "
                      f"(line left the tab) — text was: {text!r}"))

        # Atomic write — save to a temp file then os.replace() so a crash
        # or interruption can't leave a half-written WIP (safe_save pattern).
        tmp = wip_path.with_name(wip_path.name + ".tmp")
        wb.save(str(tmp))
        try:
            os.replace(str(tmp), str(wip_path))
        except OSError as e:
            # Keep the good data in the .tmp rather than losing it.
            log.error("Could not replace %s (%s); new data saved to %s",
                      wip_path.name, e, tmp.name)
            raise
        log.info("Wrote %d rows to %s in %s", len(rows), tab_name, wip_path)
        if audit:
            _sections = ([appendix] if isinstance(appendix, tuple)
                         else (appendix or []))
            _all = list(rows) + [r for _t, ap in _sections for r in (ap or [])]
            _audit = audit_changes(prior_state, _all)
            print_audit(_audit, tab_name)
            if audit_xlsx is not None:
                _p = write_audit_xlsx(_audit, tab_name, audit_xlsx)
                if _p:
                    print(_Term.color(_Term.DIM, f"  change file → {_p}"))
    finally:
        wb.close()
    _qc_check(wip_path, tab_name, expected_rows=len(rows) + sum(
        len(a[1]) for a in ([appendix] if isinstance(appendix, tuple)
                            else (appendix or [])) if a[1]),
              active_only=default_filter_active)
    return True


def _qc_check(wip_path: Path, tab_name: str, expected_rows: int,
              active_only: bool) -> None:
    """Visual QC after EVERY write (the user 2026-07-15): re-open the saved
    file and verify what the reader believes matches what Excel will show.
    Never raises — prints ✓/⚠ lines so a bad write is loud, not silent."""
    try:
        wb = load_workbook(wip_path)
        ws = wb[tab_name]
        hdr = _find_header_row(ws) or 1     # header sits below any title banner
        hix = {ws.cell(hdr, c).value: c for c in range(1, ws.max_column + 1)}
        pcol = hix.get("PROJECT #")
        scol = hix.get("STATUS")
        lcols = [c for lbl, c in hix.items()
                 if lbl in ("PROJECT FOLDER", "DATA SOURCE")]
        n = vis_closed = links = 0
        last_data = hdr
        for r in range(hdr + 1, ws.max_row + 1):
            # The summary block (TOTALS + FUTURE WIP CASH FLOW) is written
            # below the data in column 1 — on a tab whose first column IS
            # 'PROJECT #' its labels would otherwise be miscounted as data
            # rows. Everything from TOTALS down is the summary; stop there.
            first = ws.cell(r, 1).value
            if isinstance(first, str) and first.strip() == "TOTALS":
                break
            j = ws.cell(r, pcol).value if pcol else None
            if not j:
                continue
            n += 1
            last_data = r
            if any(ws.cell(r, c).hyperlink for c in lcols):
                links += 1
            if (scol and ws.cell(r, scol).value == "Closed"
                    and not ws.row_dimensions[r].hidden):
                vis_closed += 1
        tbl_ok = True
        if ws.tables:
            ref = list(ws.tables.values())[0].ref            # e.g. A1:U130
            tbl_end = int(re.findall(r"(\d+)$", ref)[0])
            tbl_ok = tbl_end >= last_data
        # Excel "repair on open" tripwire (2026-07-31): a sheet-level
        # AutoFilter alongside the Table's own filter makes Excel declare the
        # workbook damaged. Loud here beats a repair dialog on the user's desk.
        sheet_filter = ws.auto_filter.ref if ws.tables else None
        wb.close()
        probs = []
        if sheet_filter:
            probs.append(f"sheet AutoFilter {sheet_filter} coexists with the "
                         f"table filter — Excel will demand a repair")
        if n != expected_rows:
            probs.append(f"rows {n} ≠ expected {expected_rows}")
        if active_only and vis_closed:
            probs.append(f"{vis_closed} Closed row(s) VISIBLE in an active view")
        if not tbl_ok:
            probs.append("table does not span all data rows")
        if lcols and links == 0 and n:
            # Only meaningful when the layout HAS link columns — the master
            # tab dropped them (qbo_links_only, the user 2026-07-29).
            probs.append("no source links found")
        if probs:
            print(_Term.color(_Term.AMBER, "  ⚠ QC: " + " · ".join(probs)))
        else:
            print(_Term.color(_Term.GREEN,
                  f"  ✓ QC: {n} rows · closed hidden · table spans all · links ok"))
    except Exception as e:                                    # QC must never kill a run
        print(_Term.color(_Term.AMBER, f"  ⚠ QC check failed to run: {e}"))


# ─────────────────────── pretty run report ────────────────────────
def _shorten(path: Path, max_len: int = 70) -> str:
    """Shorten long paths by inserting an ellipsis in the middle."""
    s = str(path)
    if len(s) <= max_len:
        return s
    head = s[: max_len // 2 - 2]
    tail = s[-(max_len // 2 - 2):]
    return f"{head}…{tail}"


def _wip_metrics(r: CpRow) -> Dict[str, Optional[float]]:
    """Compute the derived WIP numbers in Python for TERMINAL DISPLAY only —
    mirrors the Excel formulas exactly (same blank-propagation), so the
    dry-run preview matches what the workbook will show after recalc."""
    F = r.contract_price          # Revised Contract
    I = r.etc                     # Revised ETC
    K = r.costs_to_date
    J = r.billed_to_date          # Billed (gross)

    def sub(a, b):
        return (a - b) if (a is not None and b is not None) else None

    orig_profit = sub(F, I)
    gp_pct = (orig_profit / F) if (orig_profit is not None and F not in (None, 0)) else None
    ctc = sub(I, K)
    pct = (K / I) if (K is not None and I not in (None, 0)) else None
    earned = (F * K / I) if (F is not None and K is not None and I not in (None, 0)) else None
    profit_earned = sub(earned, K)
    future_profit = sub(orig_profit, profit_earned)
    left_to_bill = sub(F, J)
    overbill = max(J - earned, 0.0) if (J is not None and earned is not None) else None
    underbill = max(earned - J, 0.0) if (earned is not None and J is not None) else None
    job_borrow = max(ctc - left_to_bill, 0.0) if (ctc is not None and left_to_bill is not None) else None
    return {
        "orig_profit": orig_profit, "gp_pct": gp_pct, "ctc": ctc, "pct": pct,
        "earned": earned, "profit_earned": profit_earned,
        "future_profit": future_profit, "left_to_bill": left_to_bill,
        "overbill": overbill, "underbill": underbill, "job_borrow": job_borrow,
    }


def _print_rows_table(rows: List[CpRow], wip_path: Path, tab_name: str = TEST_TAB) -> None:
    """Dry-run preview. Vertical per-project detail for small runs (the
    single-project verify case), compact table for bulk runs. Values match
    the Excel formulas via _wip_metrics."""
    _section(f"DRY RUN — would write {len(rows)} row(s) to {tab_name!r}")
    print(f"  target: {_Term.color(_Term.DIM, _shorten(wip_path))}")

    clean = sum(1 for r in rows if not r.status_flags)
    flagged = len(rows) - clean

    if len(rows) <= 3:
        # Full vertical WIP card per project — every column, easy to verify.
        for r in rows:
            m = _wip_metrics(r)
            status = "Closed" if r.is_completed else "Active"
            print()
            print(_Term.color(_Term.BOLD + _Term.CYAN,
                  f"  ▌ {r.project_num}  {r.project_name}") +
                  _Term.color(_Term.DIM, f"   [{status}]"))
            rows_out = [
                ("Original Contract",   _fmt_money(r.base_contract)),
                ("Approved COs",        _fmt_money(r.co_revenue)),
                ("Revised Contract",    _fmt_money(r.contract_price)),
                ("Original ETC",        _fmt_money(r.base_etc)),
                ("CO Cost",             _fmt_money(r.co_cost_estimate)),
                ("Revised ETC",         _fmt_money(r.etc)),
                ("Original Profit",     _fmt_money(m["orig_profit"])),
                ("Gross Profit %",      _fmt_pct(m["gp_pct"])),
                ("Costs to Date",       _fmt_money(r.costs_to_date)),
                ("Cost to Complete",    _fmt_money(m["ctc"])),
                ("% Complete",          _fmt_pct(m["pct"])),
                ("Revenues Earned",     _fmt_money(m["earned"])),
                ("Profit Earned",       _fmt_money(m["profit_earned"])),
                ("Future Profit",       _fmt_money(m["future_profit"])),
                ("Billed (gross)",      _fmt_money(r.billed_to_date)),
                ("Retainage Held",      _fmt_money(r.retainage_held)),
                ("Left to Bill",        _fmt_money(m["left_to_bill"])),
                ("Overbillings",        _fmt_money(m["overbill"])),
                ("Underbillings",       _fmt_money(m["underbill"])),
                ("Pure Job Borrow",     _fmt_money(m["job_borrow"])),
            ]
            for label, val in rows_out:
                print(f"    {label:<20} {_lpad(_dim_if_dash(val), 16)}")
            if r.notes:
                print(_Term.color(_Term.DIM, "    · " + "; ".join(r.notes)))
            if r.status_flags:
                print(_Term.color(_Term.AMBER, "    ⚑ " + "; ".join(r.status_flags)))
            else:
                print(_Term.color(_Term.GREEN, "    ✓ no flags"))
    else:
        # Compact table — the headline columns for scanning many jobs.
        # NOTES (informational: Draw #, no-draw) and FLAGS (true script must-fix
        # issues) are kept in separate trailing columns.
        W_P, W_N, W_M, W_PC, W_NOTE = 8, 24, 13, 6, 44
        SEP = _Term.color(_Term.DIM, " │ ")
        hdr = [
            _rpad("PROJECT", W_P), _rpad("NAME", W_N),
            _lpad("CONTRACT", W_M), _lpad("CO $", W_M), _lpad("ETC", W_M),
            _lpad("COSTS", W_M), _lpad("%", W_PC), _lpad("BILLED", W_M),
            _lpad("RETAIN", W_M), _lpad("OVER", W_M), _lpad("UNDER", W_M),
            _lpad("BORROW", W_M), _rpad("NOTES", W_NOTE), "FLAGS",
        ]
        print()
        print(_Term.color(_Term.BOLD, "  " + SEP.join(hdr)))
        print(_Term.color(_Term.DIM, "  " + "─" * 196))
        for r in rows:
            m = _wip_metrics(r)
            flag = (_Term.color(_Term.AMBER, "⚑ " + "; ".join(r.status_flags))
                    if r.status_flags else _Term.color(_Term.GREEN, "✓"))
            note_txt = "; ".join(r.notes)
            note = (_Term.color(_Term.DIM, _rpad(note_txt[:W_NOTE], W_NOTE))
                    if note_txt else _rpad("", W_NOTE))
            name = _rpad(r.project_name[:W_N], W_N)
            if r.is_completed:
                name = _Term.color(_Term.GREEN, name)
            cells = [
                _rpad(r.project_num, W_P), name,
                _lpad(_dim_if_dash(_fmt_money(r.contract_price)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.co_revenue)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.etc)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.costs_to_date)), W_M),
                _lpad(_dim_if_dash(_fmt_pct(m["pct"])), W_PC),
                _lpad(_dim_if_dash(_fmt_money(r.billed_to_date)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.retainage_held)), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["overbill"])), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["underbill"])), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["job_borrow"])), W_M),
                note, flag,
            ]
            print("  " + SEP.join(cells))

    print()
    print(f"  {_Term.color(_Term.BOLD, 'Summary')}:  "
          f"{len(rows)} total  ·  "
          f"{_Term.color(_Term.GREEN, str(clean) + ' clean')}  ·  "
          f"{_Term.color(_Term.AMBER, str(flagged) + ' flagged')}")

    # Audit trail — file paths the reader touched. Shown when small enough
    # to be useful (≤ 5 projects), so single-project runs always get it.
    if len(rows) <= 5:
        print()
        print(_Term.color(_Term.DIM, "  Audit trail:"))
        for r in rows:
            if r.draw_path:
                print(_Term.color(_Term.DIM,
                      f"    {r.project_num}  draw #{r.draw_num}:    {r.draw_path}"))
            if len(r.included_takeoffs) > 1:
                print(_Term.color(_Term.DIM, f"    {r.project_num}  takeoffs (summed, {len(r.included_takeoffs)}):"))
                for tk in r.included_takeoffs:
                    print(_Term.color(_Term.DIM, f"    {r.project_num}    · {tk.name}"))
            elif r.takeoff_path:
                print(_Term.color(_Term.DIM, f"    {r.project_num}  takeoff:      {r.takeoff_path}"))


# ─────────────────────── orchestration ─────────────────────────────
