#!/usr/bin/env python3
"""
wip_qc.py - the WIP report verifier. Runs OUR rules against any bank-format WIP
report and returns pass / fail, so "verified" means the same thing every time.

WHY THIS EXISTS (the owner, 2026-08-26: "we need the wip system to have our own
rules so that i can trust it will catch all")
The checks below were learned the hard way rebuilding the 12-31-25 and 3-31-26
reports, and they lived as PROSE in a memory file. Prose has to be remembered,
gets applied to one file at a time, leaves no record of what was actually run,
and gates nothing. Three defects reached the bank because of that:

  * the 8-7-26 report was never re-checked after it was sent - the working scope
    was "the file I am editing", never "every report the bank holds";
  * eight December rows carried an ETC set mechanically to cost x 1.025, which
    check ETC_FROM_COST exists to catch - it was found once and not closed out;
  * nothing stopped a send, so "verified" meant whatever that session got to.

Every rule here is executable, has an ID, and reports the rows it fired on.

    python3 wip/wip_qc.py <report.xlsx> [<report2.xlsx> ...]
    python3 wip/wip_qc.py --json <report.xlsx>          machine-readable
    python3 wip/wip_qc.py --strict <report.xlsx>        WARN also fails

Give it more than one file and it cross-checks them by report date: costs and
billing may never go backwards for the same job, contracts may not move without
a reason, margins may not swing. Exit code is non-zero when anything FAILs, so
it can gate a delivery.

SCOPE, STATED HONESTLY: this proves internal consistency, arithmetic, formula
correctness and cross-report coherence. It does NOT re-derive costs or billings
from QBO - a number can be internally perfect and still wrong at source. Passing
this is necessary, not sufficient.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from shared import xlsx_verify

# ── OUR RULES. Company conventions, not generic accounting. Tune here. ──────
MAX_PLAUSIBLE_MARGIN = 0.30    # concrete does not do 30%+; above this is a data
                               # problem, not good news. Flag as hard as a loss.
MIN_PLAUSIBLE_MARGIN = 0.0     # below zero is a loss job - real, but must be known
ALLOW_OVER_100_PCT = True      # the overrun convention: keep the estimate and let
                               # % complete run past 100. Reported, never a FAIL.
ETC_RATIO_TOLERANCE = 0.0005   # how close ETC/cost must be to a constant before
                               # it reads as mechanically derived
ETC_RATIO_MIN_ROWS = 3         # that many rows sharing one ratio = a pattern
PCT_REPEAT_MIN_ROWS = 4        # identical % complete on this many rows = ETC set
                               # from cost (the 97.56% fingerprint)
ZERO_BILLED_COST_FLOOR = 1000  # costs above this with no billing is worth reading
GP_SWING_POINTS = 5.0          # margin move between reports needing a reason
MATERIAL_COST_DROP = 0.50      # a cost regression under this is a rounding blip

# The 19-column bank layout. Header text is the contract with five other tools;
# order and spelling both matter.
STD_HEADERS = [
    "TYPE", "PROJECT #", "PROJECT NAME", "BONDED", "TOTAL CONTRACT PRICE",
    "ESTIMATED TOTAL COSTS", "ORIGINAL PROFIT", "GROSS PROFIT %",
    "COSTS TO DATE", "COST TO COMPLETE", "PERCENT COMPLETE",
    "REVENUES EARNED TO DATE", "PROFIT EARNED TO DATE", "BILLED TO DATE",
    "OVERBILLINGS", "UNDERBILLINGS", "LEFT TO BILL", "FUTURE PROFIT TO EARN",
    "PURE JOB BORROW",
]
C_TYPE, C_PROJ, C_NAME, C_BOND, C_CONTRACT, C_ETC = 1, 2, 3, 4, 5, 6
C_COSTS, C_BILLED = 9, 14

# Derived columns must be these formulas, exactly. Anything else is a hand edit.
STD_FORMULAS = {
    7:  '=IF(OR(E{r}="",F{r}=""),"",E{r}-F{r})',
    8:  '=IF(OR(G{r}="",E{r}="",E{r}=0),"",G{r}/E{r})',
    10: '=IF(OR(F{r}="",I{r}=""),"",F{r}-I{r})',
    11: '=IF(OR(I{r}="",F{r}="",F{r}=0),"",I{r}/F{r})',
    12: '=IF(OR(E{r}="",I{r}="",F{r}="",F{r}=0),"",E{r}*I{r}/F{r})',
    13: '=IF(OR(L{r}="",I{r}=""),"",L{r}-I{r})',
    15: '=IF(OR(N{r}="",L{r}=""),"",MAX(N{r}-L{r},0))',
    16: '=IF(OR(L{r}="",N{r}=""),"",MAX(L{r}-N{r},0))',
    17: '=IF(OR(E{r}="",N{r}=""),"",E{r}-N{r})',
    18: '=IF(OR(G{r}="",M{r}=""),"",G{r}-M{r})',
    19: '=IF(OR(J{r}="",Q{r}=""),"",MAX(J{r}-Q{r},0))',
}

HDR_ROW, FIRST_DATA_ROW = 3, 4


class Finding:
    __slots__ = ("check", "severity", "message", "rows", "accepted")

    def __init__(self, check: str, severity: str, message: str, rows=None):
        self.check, self.severity, self.message = check, severity, message
        self.rows = rows or []
        self.accepted: List[str] = []

    def as_dict(self):
        return {"check": self.check, "severity": self.severity,
                "message": self.message, "rows": self.rows,
                "accepted": self.accepted}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


_DATE_RE = re.compile(r"(\d{1,2})[-.](\d{1,2})[-.](\d{2,4})")


def report_date(path: Path) -> Optional[date]:
    """Report date from the filename (12-31-25, 3-31-26, 8.7.26). Used only to
    order files for the cross-report checks - never for arithmetic."""
    m = _DATE_RE.search(path.stem)
    if not m:
        return None
    mo, dy, yr = (int(x) for x in m.groups())
    if yr < 100:
        yr += 2000
    try:
        return date(yr, mo, dy)
    except ValueError:
        return None


class Report:
    """One WIP report, loaded once, both formula and value views."""

    def __init__(self, path: Path):
        self.path = Path(path)
        self.name = self.path.stem
        self.date = report_date(self.path)
        self.wbf = load_workbook(str(path))
        self.wbv = load_workbook(str(path), data_only=True)
        self.ws = self.wbf.worksheets[0]
        self.wsv = self.wbv.worksheets[0]
        self.last_row = self._last_data_row()
        self.rows = self._read_rows()

    def _last_data_row(self) -> int:
        last = HDR_ROW
        for r in range(FIRST_DATA_ROW, self.ws.max_row + 1):
            v = self.ws.cell(r, C_PROJ).value
            s = str(v or "").strip()
            if s and not s.upper().startswith("TOTAL"):
                last = r
        return last

    def _read_rows(self) -> List[dict]:
        out = []
        for r in range(FIRST_DATA_ROW, self.last_row + 1):
            proj = str(self.ws.cell(r, C_PROJ).value or "").strip()
            if not proj:
                continue
            out.append({
                "row": r, "proj": proj.upper(),
                "name": str(self.ws.cell(r, C_NAME).value or ""),
                "contract": _num(self.wsv.cell(r, C_CONTRACT).value),
                "etc": _num(self.wsv.cell(r, C_ETC).value),
                "costs": _num(self.wsv.cell(r, C_COSTS).value),
                "billed": _num(self.wsv.cell(r, C_BILLED).value),
            })
        return out

    def by_proj(self) -> Dict[str, dict]:
        return {r["proj"]: r for r in self.rows}

    def close(self):
        self.wbf.close()
        self.wbv.close()


# ─────────────────────── structural checks ───────────────────────

def check_file_integrity(rep: Report) -> List[Finding]:
    issues = xlsx_verify.verify_xlsx(rep.path)
    if issues:
        return [Finding("FILE_CORRUPT", "FAIL",
                        "Excel will show a repair prompt: " + "; ".join(issues))]
    return []


def check_headers(rep: Report) -> List[Finding]:
    """The 19-column layout is a contract with five other tools. A renamed or
    reordered header silently breaks every one of them."""
    out = []
    for i, want in enumerate(STD_HEADERS, start=1):
        got = str(rep.ws.cell(HDR_ROW, i).value or "").strip()
        if got.upper() != want.upper():
            out.append(Finding("HEADER_DRIFT", "FAIL",
                               f"column {i} should be {want!r}, found {got!r}"))
    return out


def check_derived_formulas(rep: Report) -> List[Finding]:
    """Every derived cell must be the standard formula for ITS OWN row. Catches
    hand-typed values pasted over a formula, and the classic delete_rows damage
    where formulas keep pointing at the row they came from."""
    hand, wrong = [], []
    for r in range(FIRST_DATA_ROW, rep.last_row + 1):
        if not str(rep.ws.cell(r, C_PROJ).value or "").strip():
            continue
        for col, tmpl in STD_FORMULAS.items():
            got = rep.ws.cell(r, col).value
            if got is None:
                continue
            if not (isinstance(got, str) and got.startswith("=")):
                hand.append(f"{rep.ws.cell(r, col).coordinate}={got!r}")
            elif got.replace(" ", "") != tmpl.format(r=r).replace(" ", ""):
                wrong.append(f"{rep.ws.cell(r, col).coordinate}={got}")
    out = []
    if hand:
        out.append(Finding("FORMULA_OVERWRITTEN", "FAIL",
                           f"{len(hand)} derived cell(s) hold a typed value, not a "
                           f"formula - the number will not move when inputs change",
                           hand[:12]))
    if wrong:
        out.append(Finding("FORMULA_WRONG_ROW", "FAIL",
                           f"{len(wrong)} derived formula(s) do not match the standard "
                           f"for their own row (row-shift damage?)", wrong[:12]))
    return out


def check_table_and_totals(rep: Report) -> List[Finding]:
    out = []
    for name in rep.ws.tables:
        t = rep.ws.tables[name]
        ref = t.ref if hasattr(t, "ref") else t
        _, r1, _, r2 = range_boundaries(str(ref))
        if r2 < rep.last_row:
            out.append(Finding("TABLE_TOO_SHORT", "FAIL",
                               f"table {name!r} ends at row {r2} but data runs to "
                               f"{rep.last_row} - {rep.last_row - r2} row(s) sit outside "
                               f"the table and are excluded from every total"))
        elif r2 > rep.last_row:
            out.append(Finding("TABLE_TOO_LONG", "WARN",
                               f"table {name!r} extends {r2 - rep.last_row} row(s) past "
                               f"the last job"))

    tot_row = None
    for r in range(rep.last_row + 1, min(rep.ws.max_row, rep.last_row + 6) + 1):
        vals = [rep.ws.cell(r, c).value for c in range(1, 20)]
        if any(isinstance(v, str) and ("TOTAL" in v.upper() or v.startswith("=SUBTOTAL"))
               for v in vals if v):
            tot_row = r
            break
    if tot_row is None:
        return out + [Finding("NO_TOTALS_ROW", "FAIL", "no totals row found below the data")]

    for col, label in ((C_CONTRACT, "contract"), (C_ETC, "ETC"),
                       (C_COSTS, "costs"), (C_BILLED, "billed")):
        f = rep.ws.cell(tot_row, col).value
        if not (isinstance(f, str) and f.startswith("=")):
            out.append(Finding("TOTAL_HARDCODED", "FAIL",
                               f"{label} total is a typed value, not a formula"))
            continue
        m = re.search(r"(\d+):[A-Z]+(\d+)", f)
        if m and (int(m.group(1)) != FIRST_DATA_ROW or int(m.group(2)) != rep.last_row):
            out.append(Finding("TOTAL_RANGE_WRONG", "FAIL",
                               f"{label} total sums rows {m.group(1)}..{m.group(2)} but the "
                               f"data is rows {FIRST_DATA_ROW}..{rep.last_row}"))
        cached = _num(rep.wsv.cell(tot_row, col).value)
        rowsum = sum(x for x in (r[label if label != "ETC" else "etc"]
                                 for r in rep.rows) if x)
        if cached is not None and abs(cached - rowsum) > 1:
            out.append(Finding("TOTAL_MISMATCH", "FAIL",
                               f"{label} total shows {cached:,.0f} but the rows sum to "
                               f"{rowsum:,.0f}"))
    rep.totals_row = tot_row
    return out


def check_presentation(rep: Report) -> List[Finding]:
    """Presentation defects reach the bank looking like carelessness. Every one
    of these shipped at least once (2026-08-25)."""
    out = []
    tot = getattr(rep, "totals_row", rep.last_row)
    pa = rep.ws.print_area
    if not pa:
        out.append(Finding("NO_PRINT_AREA", "WARN",
                           "no print area set - printing or PDF'ing gives an "
                           "undefined range"))
    else:
        m = re.search(r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", str(pa))
        if m and int(m.group(4)) < tot:
            out.append(Finding("PRINT_AREA_SHORT", "FAIL",
                               f"print area stops at row {m.group(4)} but the totals "
                               f"row is {tot} - the total will not print"))
    if not rep.ws.freeze_panes:
        out.append(Finding("NO_FREEZE_PANES", "WARN",
                           "header row is not frozen - it scrolls away on a long report"))
    return out


# ─────────────────────── per-row data checks ───────────────────────

def check_rows(rep: Report) -> List[Finding]:
    out = []
    missing_c, missing_e, negative = [], [], []
    over_100, over_contract, loss, rich, zero_billed = [], [], [], [], []

    for r in rep.rows:
        tag = f"{r['proj']} (row {r['row']})"
        c, e, co, b = r["contract"], r["etc"], r["costs"], r["billed"]
        if not c:
            missing_c.append(tag)
        if not e:
            missing_e.append(tag)
        for label, v in (("contract", c), ("ETC", e), ("costs", co), ("billed", b)):
            if v is not None and v < 0:
                negative.append(f"{tag} {label} {v:,.0f}")
        if c and e:
            gp = (c - e) / c
            if gp < MIN_PLAUSIBLE_MARGIN:
                loss.append(f"{tag} {gp:.1%}")
            elif gp > MAX_PLAUSIBLE_MARGIN:
                rich.append(f"{tag} {gp:.1%}")
        if e and co and co / e > 1.0:
            over_100.append(f"{tag} {co / e:.1%}")
        if c and co and co > c:
            over_contract.append(f"{tag} costs {co:,.0f} > contract {c:,.0f}")
        if (co or 0) > ZERO_BILLED_COST_FLOOR and not b:
            zero_billed.append(f"{tag} costs {co:,.0f}")

    if missing_c:
        out.append(Finding("CONTRACT_MISSING", "FAIL",
                           f"{len(missing_c)} job(s) have no contract - every derived "
                           f"column on those rows is blank or wrong", missing_c[:12]))
    if missing_e:
        out.append(Finding("ETC_MISSING", "FAIL",
                           f"{len(missing_e)} job(s) have no ETC - percent complete and "
                           f"earned revenue cannot compute", missing_e[:12]))
    if negative:
        out.append(Finding("NEGATIVE_VALUE", "FAIL",
                           f"{len(negative)} negative input(s)", negative[:12]))
    if rich:
        out.append(Finding("MARGIN_IMPLAUSIBLE", "FAIL",
                           f"{len(rich)} job(s) above {MAX_PLAUSIBLE_MARGIN:.0%} margin - "
                           f"concrete does not do that, so it is a data problem "
                           f"(usually a flatwork-only ETC), not good news", rich[:12]))
    if loss:
        out.append(Finding("MARGIN_NEGATIVE", "WARN",
                           f"{len(loss)} loss job(s) - real, but each needs to be known",
                           loss[:12]))
    if over_100:
        sev = "INFO" if ALLOW_OVER_100_PCT else "WARN"
        out.append(Finding("OVER_100_PCT", sev,
                           f"{len(over_100)} job(s) past 100% complete - expected under the "
                           f"overrun convention (keep the estimate, let % run over)",
                           over_100[:12]))
    if over_contract:
        out.append(Finding("COSTS_OVER_CONTRACT", "WARN",
                           f"{len(over_contract)} job(s) whose costs exceed the CONTRACT - "
                           f"that usually means a partial contract on the report, a "
                           f"different diagnosis from an overrun", over_contract[:12]))
    if zero_billed:
        out.append(Finding("ZERO_BILLED", "WARN",
                           f"{len(zero_billed)} job(s) with real costs and no billing - "
                           f"almost always parent-coded billing, not collections",
                           zero_billed[:12]))
    return out


def check_mechanical_etc(rep: Report) -> List[Finding]:
    """ETC set from cost instead of from a bid. Two fingerprints: a constant
    ETC/cost ratio shared by several rows, and repeated identical % complete.
    This is the cardinal-rule violation - once ETC comes from cost, the margin
    column stops showing the bid and starts echoing the outcome."""
    out = []
    ratios: Dict[float, List[str]] = {}
    pcts: Dict[float, List[str]] = {}
    for r in rep.rows:
        e, co = r["etc"], r["costs"]
        if not e or not co:
            continue
        key = round(e / co, 4)
        ratios.setdefault(key, []).append(f"{r['proj']} (row {r['row']})")
        pk = round(co / e, 4)
        pcts.setdefault(pk, []).append(f"{r['proj']} (row {r['row']})")

    for ratio, jobs in sorted(ratios.items(), key=lambda kv: -len(kv[1])):
        if len(jobs) >= ETC_RATIO_MIN_ROWS and abs(ratio - 1.0) > ETC_RATIO_TOLERANCE:
            out.append(Finding("ETC_FROM_COST", "FAIL",
                               f"{len(jobs)} job(s) share an ETC of exactly cost x "
                               f"{ratio:.4f} - the ETC was set from the cost, not bid. "
                               f"Their margins are manufactured", jobs[:12]))
    for pct, jobs in sorted(pcts.items(), key=lambda kv: -len(kv[1])):
        if len(jobs) >= PCT_REPEAT_MIN_ROWS and 0 < pct < 1:
            out.append(Finding("PCT_REPEATED", "WARN",
                               f"{len(jobs)} job(s) at exactly {pct:.2%} complete - the "
                               f"fingerprint of a mechanically set ETC", jobs[:12]))
    return out


# ─────────────────────── cross-report checks ───────────────────────

def check_across(earlier: Report, later: Report) -> List[Finding]:
    """The story has to hold across dates. A job's costs and billing can only go
    up; a drop means the two reports were built by different methods, not that
    something shrank. This is the check that catches an old report still sitting
    with the bank after a rebuild corrected the others."""
    out = []
    A, B = earlier.by_proj(), later.by_proj()
    both = sorted(set(A) & set(B))
    if not both:
        return [Finding("NO_OVERLAP", "INFO",
                        f"{earlier.name} and {later.name} share no jobs - nothing to compare")]

    cost_back, bill_back, contract_moved, gp_swing = [], [], [], []
    for p in both:
        a, b = A[p], B[p]
        if (a["costs"] or 0) - (b["costs"] or 0) > MATERIAL_COST_DROP:
            cost_back.append(f"{p} {a['costs']:,.0f} -> {b['costs']:,.0f}")
        if (a["billed"] or 0) - (b["billed"] or 0) > MATERIAL_COST_DROP:
            bill_back.append(f"{p} {a['billed']:,.0f} -> {b['billed']:,.0f}")
        if a["contract"] and b["contract"] and abs(a["contract"] - b["contract"]) > 1:
            contract_moved.append(f"{p} {a['contract']:,.0f} -> {b['contract']:,.0f}")
        if a["contract"] and a["etc"] and b["contract"] and b["etc"]:
            g1 = (a["contract"] - a["etc"]) / a["contract"]
            g2 = (b["contract"] - b["etc"]) / b["contract"]
            if abs(g2 - g1) * 100 >= GP_SWING_POINTS:
                gp_swing.append(f"{p} {g1:.1%} -> {g2:.1%}")

    if cost_back:
        out.append(Finding("COSTS_WENT_BACKWARDS", "FAIL",
                           f"{len(cost_back)} job(s) show LOWER costs on the later report - "
                           f"costs cannot shrink, so the two reports were built by "
                           f"different methods", cost_back[:12]))
    if bill_back:
        out.append(Finding("BILLING_WENT_BACKWARDS", "FAIL",
                           f"{len(bill_back)} job(s) show LOWER billing on the later report",
                           bill_back[:12]))
    if contract_moved:
        out.append(Finding("CONTRACT_MOVED", "WARN",
                           f"{len(contract_moved)} contract(s) changed between reports - "
                           f"each needs a stated reason (a CO, a scope cut)",
                           contract_moved[:12]))
    if gp_swing:
        out.append(Finding("MARGIN_SWING", "WARN",
                           f"{len(gp_swing)} job(s) whose margin moved {GP_SWING_POINTS}+ "
                           f"points between reports - each needs a reason", gp_swing[:12]))
    return out


def check_staleness(reports: List[Report]) -> List[Finding]:
    """The trap that let three defects reach the bank: the newest report by DATE
    was the oldest by BUILD, so it never received the corrections the others got.
    Whenever a report is dated later but built earlier, say so loudly."""
    out = []
    dated = [r for r in reports if r.date]
    if len(dated) < 2:
        return out
    for r in dated:
        r._built = r.path.stat().st_mtime
    newest_date = max(dated, key=lambda r: r.date)
    newest_build = max(dated, key=lambda r: r._built)
    if newest_date is not newest_build:
        out.append(Finding("STALE_LATEST", "FAIL",
                           f"{newest_date.name!r} is the most RECENT report by date but was "
                           f"BUILT before {newest_build.name!r}. The report the bank treats "
                           f"as current never received the later corrections."))
    return out


# ─────────────────────── run + report ───────────────────────

SINGLE = [check_file_integrity, check_headers, check_derived_formulas,
          check_table_and_totals, check_presentation, check_rows,
          check_mechanical_etc]

_ICON = {"FAIL": "FAIL", "WARN": "WARN", "INFO": "info", "ACCEPTED": " ok "}
_ORDER = {"FAIL": 0, "WARN": 1, "INFO": 2, "ACCEPTED": 3}

# The sign-off file lives beside the reports, NOT in this repo - it carries job
# numbers and reasons, which is business content.
DEFAULT_SIGNOFF = "wip-qc-signoff.json"


def load_signoff(path: Optional[Path]) -> dict:
    """{check_id: {"reason": ..., "by": ..., "on": ...}} - findings a human has
    already reviewed and accepted.

    WHY THIS MATTERS AS MUCH AS THE CHECKS: several findings are permanently
    correct and permanently accepted - a documented scope cut will trip
    CONTRACT_MOVED on every run forever. Without a way to retire them, the
    warning list only grows, everyone learns to skim it, and a real finding
    hides in the noise. A finding is then either OPEN or has a written reason
    with a name against it. Nothing sits in between."""
    if not path or not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except (ValueError, OSError) as e:
        print(f"  (sign-off file unreadable: {e})", file=sys.stderr)
        return {}


def apply_signoff(findings: List["Finding"], signoff: dict) -> None:
    """Downgrade any finding whose id carries a written reason. Matches on
    check + the specific row, so accepting one job never silences the check."""
    for f in findings:
        keep = []
        for row in f.rows:
            job = row.split(" ")[0]
            rec = signoff.get(f"{f.check}:{job}")
            if rec:
                f.accepted.append(f"{job} - {rec.get('reason', 'no reason given')} "
                                  f"[{rec.get('by', '?')} {rec.get('on', '?')}]")
            else:
                keep.append(row)
        f.rows = keep
        whole = signoff.get(f.check)
        if whole and not f.rows:
            f.accepted.append(f"whole check - {whole.get('reason', '')}")
        if f.accepted and not f.rows:
            f.severity = "ACCEPTED"
        elif f.accepted:
            f.message += f"  ({len(f.accepted)} already signed off)"


def run(paths: List[Path], strict: bool = False, signoff: dict = None):
    signoff = signoff or {}
    reports, results = [], []
    for p in paths:
        rep = Report(p)
        findings = []
        for fn in SINGLE:
            findings.extend(fn(rep))
        apply_signoff(findings, signoff)
        results.append((rep, findings))
        reports.append(rep)

    cross: List[Finding] = []
    dated = sorted([r for r in reports if r.date], key=lambda r: r.date)
    for a, b in zip(dated, dated[1:]):
        for f in check_across(a, b):
            f.message = f"{a.name} -> {b.name}: {f.message}"
            cross.append(f)
    cross.extend(check_staleness(reports))
    apply_signoff(cross, signoff)
    return reports, results, cross


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("files", nargs="+", type=Path)
    ap.add_argument("--json", action="store_true", help="machine-readable output")
    ap.add_argument("--strict", action="store_true", help="WARN also fails the run")
    ap.add_argument("--signoff", type=Path, default=None,
                    help=f"JSON of reviewed-and-accepted findings (default: "
                         f"{DEFAULT_SIGNOFF} beside the first report)")
    args = ap.parse_args()

    missing = [p for p in args.files if not p.exists()]
    if missing:
        print("not found: " + ", ".join(str(p) for p in missing))
        return 1

    sp = args.signoff or (args.files[0].parent / DEFAULT_SIGNOFF)
    signoff = load_signoff(sp)
    if signoff:
        # stderr, so --json output stays parseable
        print(f"  sign-off file: {sp}  ({len(signoff)} accepted finding(s))",
              file=sys.stderr if args.json else sys.stdout)
    reports, results, cross = run(args.files, args.strict, signoff)

    if args.json:
        print(json.dumps({
            "reports": [{"file": str(r.path), "name": r.name,
                         "date": r.date.isoformat() if r.date else None,
                         "jobs": len(r.rows),
                         "findings": [f.as_dict() for f in fs]}
                        for r, fs in results],
            "cross_report": [f.as_dict() for f in cross],
        }, indent=2))
    else:
        for rep, findings in results:
            print(f"\n{'=' * 74}\n{rep.name}   {len(rep.rows)} jobs"
                  f"{'   report date ' + rep.date.isoformat() if rep.date else ''}")
            print("=" * 74)
            if not findings:
                print("  all single-report checks passed")
            for f in sorted(findings, key=lambda x: _ORDER[x.severity]):
                print(f"  [{_ICON[f.severity]}] {f.check}: {f.message}")
                for row in f.rows:
                    print(f"          {row}")
                for row in f.accepted:
                    print(f"          signed off: {row}")
        if cross:
            print(f"\n{'=' * 74}\nCROSS-REPORT\n{'=' * 74}")
            for f in sorted(cross, key=lambda x: _ORDER[x.severity]):
                print(f"  [{_ICON[f.severity]}] {f.check}: {f.message}")
                for row in f.rows:
                    print(f"          {row}")
                for row in f.accepted:
                    print(f"          signed off: {row}")

    every = [f for _, fs in results for f in fs] + cross
    fails = [f for f in every if f.severity == "FAIL"]
    warns = [f for f in every if f.severity == "WARN"]
    accepted = [f for f in every if f.severity == "ACCEPTED"]
    out = sys.stderr if args.json else sys.stdout
    print(f"\n{'-' * 74}", file=out)
    print(f"  {len(fails)} FAIL · {len(warns)} WARN · "
          f"{len([f for f in every if f.severity == 'INFO'])} info · "
          f"{len(accepted)} accepted", file=out)
    print("  Scope: internal consistency, arithmetic, formulas, cross-report coherence.", file=out)
    print("  This does NOT re-derive costs or billing from QBO. Passing is necessary,", file=out)
    print("  not sufficient - a number can be internally perfect and wrong at source.", file=out)
    if fails:
        print("  RESULT: FAIL - do not send until these are fixed or signed off.", file=out)
    elif warns and args.strict:
        print("  RESULT: FAIL (strict) - warnings must be signed off.", file=out)
    elif warns:
        print("  RESULT: pass with warnings - each needs a stated reason.", file=out)
    else:
        print("  RESULT: pass.", file=out)
    for rep in reports:
        rep.close()
    return 1 if fails or (args.strict and warns) else 0


if __name__ == "__main__":
    raise SystemExit(main())
