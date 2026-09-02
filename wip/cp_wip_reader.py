#!/usr/bin/env python3
"""
cp_wip_reader.py — the CP (Commercial) WIP READER.

Scans the CP project folders on Synology, reads each job's latest G702 draw
(or, pre-Draw-#1, the signed proposal PDF then the takeoff), enriches with QBO
billed/costs, and writes the 'Test - CP' tab.

This file is ONLY the CP reader. The shared report engine — CpRow,
write_test_cp, formatting, the change audit, edit-tracking, QC — lives in
wip_writer.py; every division tool imports it from there (never from here).
"""
from __future__ import annotations

import argparse
import datetime as dt
import logging
import os
import re
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name.*")
warnings.filterwarnings("ignore", message=".*LibreSSL.*")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths
from shared import proposals  # signed bid-proposal PDF → contract price
from shared import qbo_api

# The shared report engine. CP is a reader; it feeds these, never redefines them.
import wip_writer as W
import wip_review_common as WR   # shared WIP-review diff/merge (ledger accept/merge flow)
# CpRow (the shared data model) and the terminal-output helpers are used
# pervasively here; the WRITER itself (write_test_cp / WipWriteDenied) is NOT
# re-exported — call it as W.write_test_cp so no other tool can ever reach the
# writer through cp_wip_reader again (that shortcut is what tangled these files).
from wip_writer import CpRow, WIP_EXCEL_PATH, TEST_TAB, _Term, _section, _kv, _shorten

CP_ACTIVE_DIR = Path(os.getenv(
    "CP_ACTIVE_DIR",
    "/Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects",
))
CP_COMPLETED_DIR = Path(os.getenv(
    "CP_COMPLETED_DIR",
    str(CP_ACTIVE_DIR / "Completed Projects"),
))

# Cell anchors on the takeoff (locked)
PROPOSAL_SHEET = "Commercial Proposal"
GRAND_TOTAL_LABEL = "GRAND TOTAL"
BID_SHEET = "Bid"
ETC_CELL = "AP1961"

from shared.draws import (                                    # noqa: E402
    G702_SHEET,
    coerce_float as _coerce_float,
    find_latest_draw,
    read_draw_g702,
)

# Project # from folder name — e.g. "CP672 - FIRESTONE RED OAK" → "CP672"
_CP_FOLDER_RE = re.compile(r"^(CP\d{3,4})\b", re.IGNORECASE)

log = logging.getLogger("cp_wip_reader")

def _find_label_cells(ws, label: str) -> List[tuple]:
    """Return ALL (row, col) matches for the given label (case-insensitive,
    trailing colon/whitespace stripped).

    Returns a list because CO templates have TWO 'TOTAL:' cells: a header
    label at the top of the description table AND the summary total at
    the bottom. Caller must try each until one yields a numeric value to
    the right (see _read_number_to_right)."""
    label_norm = label.strip().upper().rstrip(":").strip()
    results: List[tuple] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip().upper().rstrip(":").strip()
            if s == label_norm:
                results.append((cell.row, cell.column))
    return results


def _resolve_sheet_name(wb, target: str) -> Optional[str]:
    """Case-insensitive, whitespace-tolerant sheet name lookup. Returns the
    ACTUAL sheet name as it appears in the workbook, or None if not found.
    Handles: 'BID' vs 'Bid' vs 'bid ' vs '  Commercial Proposal '."""
    target_norm = target.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target_norm:
            return name
    return None


def _read_number_to_right(ws_data, ws_formula, row: int, start_col: int,
                          max_scan: int = 12) -> Optional[float]:
    """Scan rightward from `start_col` looking for the first numeric value.

    Skips empty cells (None), cells that are part of a merged region but
    aren't the top-left anchor, and cells that hold just a currency
    symbol like '$'. This handles two real-world CP takeoff patterns:

      1. TOTAL: label lives in a merged cell — the value sits past the
         merge's right edge, not one column right of the label anchor.
      2. Currency symbol is a separate cell from the number:
             [ ...merged label... ] [ $ ] [ 3,840.00 ]

    Returns the first cell that resolves to a number via
    `_read_number_smart` (which itself handles formula fallback). None
    if nothing found within `max_scan` columns."""
    for offset in range(1, max_scan + 1):
        col = start_col + offset
        try:
            ref = f"{get_column_letter(col)}{row}"
        except ValueError:
            return None
        # Try the smart reader first — handles cached values + formula refs.
        v = _read_number_smart(ws_data, ws_formula, ref)
        if v is not None:
            return v
        # If not numeric, check if the raw string is a currency marker; if
        # so, keep scanning. Any other non-empty non-numeric value → treat
        # as a stop signal (we've passed the value region).
        raw = ws_data[ref].value
        if raw is None or raw == "":
            continue
        s = str(raw).strip()
        if s in ("$", "USD", "€", "£", "-", "—"):
            continue
        # Non-numeric text (like a next-row label) — bail.
        return None
    return None


# ─────────────────────── formula-safe cell reader ────────────────
# openpyxl(data_only=True) returns cached formula results. If the takeoff was
# saved without recalc (common when files are edited via Google Sheets export,
# LibreOffice, or a Python tool), the cache is empty and formula cells come
# back as None. In that case we open a SECOND view of the workbook (data_only
# =False) to see the formula string, then evaluate it dynamically against the
# cached view — so we NEVER hardcode the specific cell-relationship (the
# estimators can change AP1961's formula without breaking us).
_CELL_REF_RE = re.compile(r"^[A-Z]+\d+$")
_TOKEN_RE    = re.compile(r"[A-Z]+\d+|[+\-]")


def _read_number_smart(ws_data, ws_formula, cell_ref: str, depth: int = 0) -> Optional[float]:
    """Read a numeric value from `cell_ref`. Order of attempts:
       1. Cached value from data_only workbook.
       2. If the cell holds a formula, evaluate simple +/- expressions of
          cell references (with recursive fallback for each ref).

    Recursion depth is bounded to prevent runaway on circular refs."""
    if depth > 4:
        return None
    try:
        cached = ws_data[cell_ref].value
    except (KeyError, ValueError):
        return None
    if cached is not None:
        coerced = _coerce_float(cached)
        if coerced is not None:
            return coerced

    # Cached is missing / non-numeric — try the formula view.
    try:
        raw = ws_formula[cell_ref].value
    except (KeyError, ValueError):
        return None
    if not isinstance(raw, str) or not raw.strip().startswith("="):
        return None

    expr = raw.strip().lstrip("=").replace(" ", "").upper()
    # Support only simple +/- chains of cell refs (e.g. "AM1948+AN1948+AO1948").
    # Anything more complex (SUM, ranges, other sheets) falls back to None +
    # the caller will flag the row for manual review.
    if not re.fullmatch(r"[A-Z]+\d+([+\-][A-Z]+\d+)*", expr):
        return None
    tokens = _TOKEN_RE.findall(expr)
    if not tokens:
        return None

    total = 0.0
    sign = 1.0
    expect_ref = True
    for tok in tokens:
        if tok in ("+", "-"):
            if expect_ref:
                return None  # malformed
            sign = 1.0 if tok == "+" else -1.0
            expect_ref = True
        else:
            if not expect_ref or not _CELL_REF_RE.match(tok):
                return None
            v = _read_number_smart(ws_data, ws_formula, tok, depth + 1)
            if v is None:
                return None
            total += sign * v
            expect_ref = False
            sign = 1.0
    return total


_WIP_TAG_RE = re.compile(r"\bwip\b", re.IGNORECASE)
# Auxiliary xlsx that live alongside takeoffs in a project folder and are NOT
# takeoffs (used only as a fallback when no file is named 'takeoff').
_AUX_XLSX_RE = re.compile(r"cost\s*code|explanation", re.IGNORECASE)


# Takeoff flags that only concern the CONTRACT figure — dropped when the
# proposal PDF supplied it, so a pre-draw job doesn't carry a complaint about a
# source we no longer needed.
_CONTRACT_FLAG_RE = re.compile(
    r"Grand/Sub Total|Contract Price|proposal (tab|sheet)", re.I)


def _find_contract_total(ws_data, ws_formula):
    """Read the contract total off a proposal sheet. Templates are inconsistent
    (the user 2026-07-02) — the overall total is labeled 'GRAND TOTAL', 'SUB TOTAL',
    or just 'TOTAL'. Match is EXACT (normalized, colon-stripped) so 'TOTAL SQFT',
    'TOTAL YARDS', etc. are NOT caught — only cells that are exactly one of the
    three dollar-total labels. Collect every value from all three and take the
    **largest** = the overall contract total (an overall is ≥ any section
    subtotal). Returns (value_or_None, label_used_or_None)."""
    best = None
    best_label = None
    for label in ("GRAND TOTAL", "SUB TOTAL", "TOTAL"):
        for (r, c) in _find_label_cells(ws_data, label):
            v = _read_number_to_right(ws_data, ws_formula, r, c)
            if v is not None and (best is None or v > best):
                best = v
                best_label = label
    return best, best_label


def _select_proposal_sheet(wb):
    """Pick which proposal tab to read the contract GRAND TOTAL from, when a
    takeoff has multiple proposal sheets (the user 2026-07-02):
      - a tab with 'final' in its name → the final proposal, use it;
      - else if exactly ONE proposal tab → use it;
      - else (multiple proposals, none marked FINAL) → don't guess, flag it.
    'Proposal' tabs = any sheet whose name contains 'proposal' (Commercial /
    Alternative / '11.14.25 Proposal' …). Non-proposal tabs (Bid, JMP
    Subcontract, Change Order#N, Cost Codes) are ignored here.

    RESIDENTIAL PROPOSAL TABS ARE NEVER READ HERE (the user 2026-08-04:
    "residential proposal is only used for RP, if CP use the commercial 100%").
    The takeoff template ships with both tabs, so a CP job would otherwise look
    like two competing proposals and get skipped — which is exactly why CP910
    came through with a blank contract.

    Returns (sheet_name_or_None, flag_or_None)."""
    prop_tabs = [s for s in wb.sheetnames if "proposal" in s.lower()]
    if not prop_tabs:
        return None, "Missing Proposal Sheet"
    cp_tabs = [s for s in prop_tabs if "residential" not in s.lower()]
    if not cp_tabs:
        return None, ("Only a RESIDENTIAL proposal tab in a CP takeoff — "
                      "commercial pricing belongs on the Commercial tab")
    prop_tabs = cp_tabs
    finals = [s for s in prop_tabs if "final" in s.lower()]
    if len(finals) == 1:
        return finals[0], None
    if len(finals) > 1:
        return finals[0], f"Multiple FINAL proposals — used '{finals[0]}'"
    if len(prop_tabs) == 1:
        return prop_tabs[0], None
    commercial = [s for s in prop_tabs if "commercial" in s.lower()]
    if len(commercial) == 1:
        return commercial[0], None          # CP ⇒ the Commercial tab, always
    return None, (f"Multiple proposals ({len(prop_tabs)}), none marked FINAL — "
                  f"mark the final one: {', '.join(prop_tabs[:4])}"
                  f"{'...' if len(prop_tabs) > 4 else ''}")


def _parse_one_takeoff(tk: Path):
    """Read Contract Price (final proposal Grand Total) + ETC (Bid!AP1961)
    from ONE takeoff file. Change Orders are NOT read here — approved COs only
    ever come from a draw, and a no-draw project (the only caller of this path)
    has no COs yet. Returns (contract, etc, flags). Never raises."""
    flags: List[str] = []
    # Two views: cached values (data_only) fast path + formulas (fallback for
    # cells saved without recalc). Random-access needed → not read_only.
    try:
        wb_data    = load_workbook(tk, data_only=True)
        wb_formula = load_workbook(tk, data_only=False)
    except Exception as e:
        return None, None, [f"Takeoff Read Failed: {type(e).__name__}"]

    contract = etc = None
    try:
        prop_sheet, prop_flag = _select_proposal_sheet(wb_data)
        if prop_flag:
            flags.append(prop_flag)
        if prop_sheet is not None:
            cp, label_used = _find_contract_total(wb_data[prop_sheet],
                                                  wb_formula[prop_sheet])
            if cp is None:
                has_label = (_find_label_cells(wb_data[prop_sheet], "GRAND TOTAL")
                             or _find_label_cells(wb_data[prop_sheet], "SUB TOTAL"))
                flags.append("Bad Contract Price" if has_label
                             else "Missing Grand/Sub Total")
            else:
                contract = cp

        bid_sheet = _resolve_sheet_name(wb_data, BID_SHEET)
        if not bid_sheet:
            flags.append("Missing Bid Sheet")
        else:
            e = _read_number_smart(wb_data[bid_sheet], wb_formula[bid_sheet], ETC_CELL)
            if e is None:
                flags.append(f"Bad ETC ({ETC_CELL})")
            else:
                etc = e

    finally:
        wb_data.close()
        wb_formula.close()
    return contract, etc, flags


def _select_takeoffs(folder: Path):
    """Identify which takeoff file(s) to read in a project folder. Shared by
    parse_takeoff (full contract+ETC read) and parse_takeoff_etc (ETC-only,
    used when a draw supplies the contract). Rules (the user 2026-07-02):
      - Only files with 'takeoff' in the name are takeoffs; auxiliary xlsx
        (Cost Codes, Explanation OH, …) are ignored. If none is named
        'takeoff', fall back to the single non-auxiliary xlsx.
      - ONE takeoff → use it; MULTIPLE → only the 'WIP'-tagged one(s), summed.
      - Multiple, none tagged WIP → don't guess.
    Returns (included_list, flag_or_None)."""
    xlsx_files = sorted([p for p in folder.iterdir()
                         if p.suffix.lower() in (".xlsx", ".xlsm")
                         and not p.name.startswith("~$")])
    if len(xlsx_files) == 0:
        return [], "No Takeoff"
    takeoffs = [p for p in xlsx_files if "takeoff" in p.name.lower()]
    if not takeoffs:
        non_aux = [p for p in xlsx_files if not _AUX_XLSX_RE.search(p.name)]
        if len(non_aux) == 1:
            takeoffs = non_aux
        else:
            return [], (f"No takeoff file identified ({len(xlsx_files)} xlsx, none "
                        f"named 'takeoff') — rename the takeoff to include 'takeoff'")
    if len(takeoffs) == 1:
        return takeoffs, None
    included = [p for p in takeoffs if _WIP_TAG_RE.search(p.name)]
    if not included:
        return [], (f"Multiple takeoffs ({len(takeoffs)}) — none tagged 'WIP'; "
                    f"estimator must tag the one(s) to include")
    return included, None


def _takeoff_label(included) -> str:
    """'takeoff · <file>' or 'takeoffs summed · a + b' - the review's source chip."""
    names = [Path(p).name for p in included]
    if len(names) == 1:
        return f"takeoff · {names[0]}"
    return f"takeoffs summed · {' + '.join(names)}"


def parse_takeoff(folder: Path, row: CpRow) -> None:
    """Extract Contract Price + ETC into `row` from the takeoff. Used when the
    project has NO draw yet (pre-Draw#1) — contract comes from the proposal
    Grand/Sub Total, ETC from Bid!AP1961. No COs are read: a project that
    hasn't started billing has no approved change orders yet (COs come from a
    draw). Appends to row.status_flags on any failure; never raises."""
    included, flag = _select_takeoffs(folder)
    if flag:
        row.status_flags.append(flag)
    if not included:
        return

    row.included_takeoffs = included
    row.takeoff_path = included[0]                  # hyperlink anchor

    contract_total = 0.0
    etc_total = 0.0
    got_contract = False
    got_etc = False
    multi = len(included) > 1
    for tk in included:
        c, e, fflags = _parse_one_takeoff(tk)
        prefix = f"{tk.name}: " if multi else ""
        for f in fflags:
            row.status_flags.append(prefix + f)
        if c is not None:
            contract_total += c
            got_contract = True
        if e is not None:
            etc_total += e
            got_etc = True

    takeoff_contract = contract_total if got_contract else None
    row.base_etc = etc_total if got_etc else None
    if got_etc:
        WR.set_source(row, "orig_etc", _takeoff_label(included), included[0])

    # CONTRACT SOURCE ORDER for a pre-draw job (the user 2026-08-04):
    # the signed proposal PDF FIRST, the takeoff only as a fallback. The PDF is
    # the document the customer agreed to; the takeoff is an internal file whose
    # template ships several proposal tabs, so it is often ambiguous — CP910 read
    # as a blank contract for exactly that reason while its PDF said $105,815.
    pdf_amt, pdf_note = proposals.contract_from_folder(folder)
    if pdf_amt is not None:
        row.base_contract = pdf_amt
        pdf_path, _ = proposals.find_proposal_pdf(folder)   # the review names the file
        WR.set_source(row, "orig_contract",
                      f"proposal PDF · {pdf_path.name}" if pdf_path else "proposal PDF",
                      pdf_path)
        # Cross-verify the two independent sources (the user 2026-08-04:
        # "verify with pdf or vice versa"). Agreement is worth saying out loud;
        # disagreement is a must-fix, because one of the two is stale.
        if takeoff_contract is None:
            row.notes.append(f"Contract from the {pdf_note}")
        elif abs(takeoff_contract - pdf_amt) < 1:
            row.notes.append(f"Contract from the {pdf_note} — "
                             f"matches the Commercial Proposal tab ✓")
        else:
            row.status_flags.append(
                f"Proposal PDF ${pdf_amt:,.0f} vs Commercial Proposal tab "
                f"${takeoff_contract:,.0f} — they disagree, verify")
            row.notes.append(f"Contract from the {pdf_note}")
        # The takeoff's contract complaints are moot once the PDF answered.
        row.status_flags = [f for f in row.status_flags
                            if not _CONTRACT_FLAG_RE.search(f)]
    else:
        row.base_contract = takeoff_contract
        if takeoff_contract is not None:
            row.notes.append(f"Contract from the takeoff — {pdf_note}")
            WR.set_source(row, "orig_contract", _takeoff_label(included), included[0])

    if multi:
        row.notes.append(
            f"WIP takeoffs summed ({len(included)}): "
            f"{', '.join(p.name for p in included)}")


# ─────────────────────── ETC-only takeoff read (draw path) ────────
def parse_takeoff_etc(folder: Path, row: CpRow) -> None:
    """Read ONLY the ETC (Bid!AP1961) from the project's takeoff. Used when a
    DRAW supplies contract/CO/billed/retainage but the cost estimate still
    lives in the takeoff (the user 2026-07-09: "ETC — still keep the takeoff
    costs"). The proposal/contract/CO parsing is skipped, so a draw-backed row
    isn't cluttered with contract-side takeoff flags. Never raises."""
    included, flag = _select_takeoffs(folder)
    if not included:
        # Missing takeoff only matters for ETC now; keep the message specific.
        row.status_flags.append(f"ETC: {flag}" if flag else "ETC: no takeoff")
        return
    if row.takeoff_path is None:
        row.takeoff_path = included[0]
    row.included_takeoffs = included

    etc_total = 0.0
    got_etc = False
    multi = len(included) > 1
    for tk in included:
        try:
            wb_data = load_workbook(tk, data_only=True)
            wb_formula = load_workbook(tk, data_only=False)
        except Exception as e:
            row.status_flags.append(f"ETC read failed ({tk.name}): {type(e).__name__}")
            continue
        try:
            bid_sheet = _resolve_sheet_name(wb_data, BID_SHEET)
            prefix = f"{tk.name}: " if multi else ""
            if not bid_sheet:
                row.status_flags.append(prefix + "Missing Bid Sheet (ETC)")
                continue
            e = _read_number_smart(wb_data[bid_sheet], wb_formula[bid_sheet], ETC_CELL)
            if e is None:
                row.status_flags.append(prefix + f"Bad ETC ({ETC_CELL})")
            else:
                etc_total += e
                got_etc = True
        finally:
            wb_data.close()
            wb_formula.close()
    if got_etc:
        row.base_etc = etc_total
        WR.set_source(row, "orig_etc", _takeoff_label(included), included[0])


# ─────────────────────── draw (G702/G703) read ────────────────────
def parse_draw(project_folder: Path, row: CpRow) -> bool:
    """If the project has a draw, read the LATEST one's G702 into `row`
    (contract, CO, billed, retainage) and return True. ETC still comes from the
    takeoff and costs from QBO — those are handled by the caller. Returns False
    when there's no draw yet (caller falls back to the takeoff proposal)."""
    found = find_latest_draw(project_folder)
    if not found:
        return False
    draw_num, draw_file = found
    row.draw_num = draw_num
    row.draw_path = draw_file

    data, flags = read_draw_g702(draw_file)
    for f in flags:
        row.status_flags.append(f"Draw #{draw_num}: {f}")
    if data is None or data["contract_to_date"] is None or data["billed"] is None:
        # A draw exists but is unreadable — do NOT fall through to takeoff/QBO
        # for billing (that would silently mix sources). Flag for triage.
        row.status_flags.append(
            f"Draw #{draw_num} unreadable — contract/billed left blank for review")
        return True

    row.base_contract = data["orig_contract"]
    row.co_revenue = data["net_co"]
    row.billed_to_date = data["billed"]
    row.retainage_held = data["retainage"]
    # The review names the document every one of these four came from - a CP
    # value that goes DOWN is usually a different draw file being read (a
    # revised copy, a deeper folder), and the owner must see which.
    draw_src = f"Draw #{draw_num} G702 · {draw_file.name}"
    for key in ("orig_contract", "approved_cos", "billed", "retainage"):
        WR.set_source(row, key, draw_src, draw_file)
    row.notes.append(
        f"Draw #{draw_num}: billed ${data['billed']:,.0f} (gross), "
        f"retainage ${(data['retainage'] or 0):,.0f}, "
        f"contract ${(data['contract_to_date'] or 0):,.0f}")
    return True


# ─────────────────────── folder scan ───────────────────────────────
def _project_num_from_folder(folder: Path) -> Optional[str]:
    m = _CP_FOLDER_RE.match(folder.name)
    return m.group(1).upper() if m else None


def _project_name_from_folder(folder: Path) -> str:
    """'CP672 - FIRESTONE RED OAK' → 'FIRESTONE RED OAK'"""
    if " - " in folder.name:
        return folder.name.split(" - ", 1)[1].strip()
    return folder.name


def scan_cp_folders(root: Path, is_completed: bool) -> List[CpRow]:
    """Iterate CP#### folders under root. Returns one CpRow per project.
    Skips non-CP folders silently (folder naming discipline is on the
    estimators). Handles Change Orders/ sub-folder = flag-and-skip."""
    rows: List[CpRow] = []
    if not root.exists():
        log.warning("CP root does not exist: %s (Synology unmounted?)", root)
        return rows

    for entry in sorted(root.iterdir()):
        if not entry.is_dir():
            continue
        if entry.name == "Completed Projects":  # handled in a separate pass
            continue

        proj_num = _project_num_from_folder(entry)
        if not proj_num:
            continue  # not a CP folder — skip silently

        row = CpRow(
            project_num=proj_num,
            project_name=_project_name_from_folder(entry),
            is_completed=is_completed,
            base_contract=None,
            co_revenue=None,
            base_etc=None,
            billed_to_date=None,
            costs_to_date=None,
        )

        # Draw-first (the user 2026-07-09): if the project has a draw (AIA G702/G703
        # payment application), the LATEST draw IS the WIP update — it supplies
        # Contract Price, Approved COs, Billed-to-Date (gross), and Retainage
        # Held. ETC still comes from the takeoff; Costs from QBO. Only before
        # Draw #1 lands do we fall back to the takeoff proposal for contract/CO
        # and QBO for billed/retainage.
        if parse_draw(entry, row):
            row.folder_path = entry                 # project-name link target
            parse_takeoff_etc(entry, row)           # ETC (Bid!AP1961) only
        else:
            # No draw yet — takeoff proposal drives contract; QBO drives
            # billed/retainage (in enrich_with_qbo). No COs: a project that
            # hasn't started its first draw has no approved change orders yet,
            # so the Change Orders/ sub-folder is intentionally not read.
            row.notes.append("No draw yet — contract from takeoff proposal")
            parse_takeoff(entry, row)

        rows.append(row)

    return rows


# ─────────────────────── QBO join ──────────────────────────────────
# Retainage detection — case-insensitive match on "Retainage Not Billed"
# anywhere in the invoice's PrivateNote (memo field). Pattern harvested
# from legacy wip/wip_sync.py; see [[reference-wip-qbo-aggregation-patterns]].
_RETAINAGE_NOT_BILLED_RE = re.compile(r"retainage\s+not\s+billed", re.IGNORECASE)


def _linked_txn_types(inv: dict) -> List[str]:
    """Normalized TxnType strings of every transaction linked to `inv`
    (payments, journal entries, credits) from QBO's LinkedTxn array.
    Spaces stripped + lowercased so 'JournalEntry' / 'Journal Entry' both
    normalize to 'journalentry'."""
    out = []
    for lt in (inv.get("LinkedTxn") or []):
        t = str(lt.get("TxnType", "")).replace(" ", "").lower()
        if t:
            out.append(t)
    return out


def _cleared_by_journal_entry(inv: dict) -> bool:
    """True if a Journal Entry is linked to this invoice — the reclass that
    moves the balance into the Retainage Receivable account (i.e. it was NOT
    settled by a real customer cash payment)."""
    return "journalentry" in _linked_txn_types(inv)


def _is_retainage_receivable_invoice(inv: dict) -> bool:
    """True if this is a standalone retainage-receivable invoice → its amount
    is retainage HELD (not collectible cash). It is excluded from Billed's
    net-collectible component and surfaces in the RETAINAGE HELD column.

    Keyed on the 'Retainage Not Billed' memo — the user's explicit tag for
    retainage moved to (or awaiting) the Retainage Receivable account.

    CORRECTED 2026-07-02 (live run): an earlier version ALSO required a
    Journal Entry on the invoice's LinkedTxn. That left retainage held
    understated by an order of magnitude on CP672 because the JE
    that reclasses AR → Retainage Receivable is applied to the invoice via a
    PAYMENT (the "Payment on 12/31/25" that carries the JE credit) — so the
    JE never appears directly in Invoice.LinkedTxn and the check missed it.
    The memo is the reliable signal and correctly handles BOTH methods:
      * standalone 'Retainage Not Billed' invoice → matched here → held.
      * 2026 draw-with-retainage-LINE → memo is 'Draw #N' (not matched); its
        withheld portion is captured via gross income − net TotalAmt.
    (`_cleared_by_journal_entry` is retained for audit logging only.)"""
    memo = inv.get("PrivateNote") or ""
    return bool(_RETAINAGE_NOT_BILLED_RE.search(memo))


def _fetch_billed_ex_retainage(pnl, access: str, company_id: str,
                               customer_id: str) -> float:
    """Sum TotalAmt of all invoices for a customer, EXCLUDING standalone
    retainage-receivable invoices (memo 'Retainage Not Billed'). This is the
    NET-COLLECTIBLE billed; the caller derives Retainage Held = gross − net.
    Returns 0.0 if the query fails or there are no invoices.

    Every retainage-memo invoice is logged with its LinkedTxn types + whether
    a JE is directly attached, so the retainage reclass path stays auditable."""
    try:
        invoices = pnl.fetch_customer_invoices(access, company_id, customer_id)
    except Exception as e:
        log.warning("Invoice fetch failed for customer %s: %s — falling back to 0",
                    customer_id, e)
        return 0.0

    total = 0.0
    skipped = 0
    for inv in invoices:
        if _is_retainage_receivable_invoice(inv):
            skipped += 1
            doc = inv.get("DocNumber", "?")
            amt = inv.get("TotalAmt", "?")
            types = _linked_txn_types(inv) or ["<none>"]
            je = "direct" if _cleared_by_journal_entry(inv) else "via-payment/none"
            log.info("  EXCL retainage-receivable inv #%s ($%s) — memo-tagged "
                     "(LinkedTxn=%s, JE=%s) → counts as Retainage Held",
                     doc, amt, ",".join(types), je)
            continue
        total += float(inv.get("TotalAmt", 0) or 0)

    if skipped:
        log.info("Customer %s: %d retainage-receivable invoice(s) → Retainage Held",
                 customer_id, skipped)
    return total


def enrich_with_qbo(rows: List[CpRow]) -> None:
    """Fetch QBO Billed/Costs per project and populate rows in-place.
    All-time window (start_date = 2019-01-01, end_date = today) — CP is
    slow-turn commercial work, worth the extra API cost. Never raises.
    Side effect: records the realm + per-row customer id so the Excel writer
    can attach QBO deep links to the Billed/Costs cells."""
    pnl = qbo_api
    try:
        access, company_id = pnl.load_credentials()
    except SystemExit:
        log.error("QBO auth failed — leaving Billed/Costs blank on all rows")
        for r in rows:
            r.status_flags.append("QBO Auth Failed")
        return
    except Exception as e:
        log.error("QBO auth error: %s", e)
        for r in rows:
            r.status_flags.append(f"QBO Auth Error: {type(e).__name__}")
        return

    try:
        proj_map = pnl.build_project_customer_map(access, company_id)
    except Exception as e:
        log.error("QBO customer map fetch failed: %s", e)
        for r in rows:
            r.status_flags.append(f"QBO Customer Map Failed: {type(e).__name__}")
        return

    W.QBO_REALM = company_id
    start_date = "2019-01-01"
    end_date = dt.date.today().isoformat()

    for row in rows:
        cust = proj_map.get(row.project_num)
        if not cust:
            row.status_flags.append("QBO Not Found")
            continue
        row.qbo_customer_id = cust["id"]

        try:
            report_data = pnl.fetch_project_pl(
                access, company_id, cust["id"], start_date, end_date
            )
            totals = pnl.extract_pl_totals(report_data)
            # Billed to Date + Retainage Held come from the DRAW when the
            # project has one (the user 2026-07-09: the latest G702 is the billing
            # source of record). Only PRE-Draw#1 projects fall back to QBO for
            # billing. This also saves the per-customer invoice fetch on
            # draw-backed jobs.
            if row.draw_num is None:
                # QBO fallback (no draw yet). Billed = GROSS, incl retainage
                # (Marcum/CFMA basis, verified 2026-07-02): QBO P&L Total Income
                # is gross billed. Net-collectible excludes 'Retainage Not
                # Billed' memo invoices; Retainage Held = gross − net.
                gross_billed = float(totals.get("income", 0.0) or 0.0)
                net_collectible = _fetch_billed_ex_retainage(
                    pnl, access, company_id, cust["id"]
                )
                row.billed_to_date = gross_billed
                row.retainage_held = max(gross_billed - net_collectible, 0.0)
                WR.set_source(row, "billed", "QuickBooks · invoices, gross incl. retainage")
                WR.set_source(row, "retainage", "QuickBooks · gross billed minus net collectible")
                log.info("  %s billed(gross)=%.2f net-collectible=%.2f retainage-held=%.2f (QBO)",
                         row.project_num, gross_billed, net_collectible, row.retainage_held)
            else:
                log.info("  %s billed/retainage from Draw #%s — QBO billing skipped",
                         row.project_num, row.draw_num)
            # Costs = COGS + Expenses. QBO Projects UI sums both; per GAAP
            # job costing, direct project spending is a project cost
            # regardless of which account category it lands in. Coding
            # errors in QBO (bills posted to Expenses instead of COGS)
            # should NOT hide costs from the WIP — the WIP should surface
            # true total spend so the coding error becomes visible.
            row.costs_to_date = (
                (totals.get("cogs", 0.0) or 0.0)
                + (totals.get("expenses", 0.0) or 0.0)
            )
            WR.set_source(row, "costs", "QuickBooks · project P&L, COGS + expenses")
            # Over-budget is NOT flagged: it's a business observation the WIP
            # report surfaces itself (Costs > ETC, and the uncapped % column).
            # Flags are reserved for things the script could not confirm as
            # fact and a human must fix — not report readings. % stays UNCAPPED
            # so the overage remains visible in the numbers.
        except Exception as e:
            row.status_flags.append(f"QBO P&L Failed: {type(e).__name__}")


# ─────────────────────── Excel write (Test - CP only) ──────────────
# Clean layout matching the team's WIP: light-gray bold header, thin grid
# borders on every cell, light-yellow ONLY on sourced inputs, white calcs,
# no green rows (the user 2026-07-02 "make it clean and easy to read").
# Font + number formats MATCH the real 'WIP Master' sheet (the user
# 2026-07-29: same font, size, formatting as the master): Tahoma 8


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse takeoffs + fetch QBO but don't write to Excel.")
    ap.add_argument("--project",
                    help="Filter to one project # (e.g. CP672). Case-insensitive.")
    ap.add_argument("--no-qbo", action="store_true",
                    help="Skip QBO join (fast local test of takeoff parsing).")
    ap.add_argument("--emit-review", metavar="JSON",
                    help="Ledger WIP Review: compute as usual, then write a "
                         "before/after diff of 'Test - CP' to this JSON and STOP "
                         "(no tab write).")
    ap.add_argument("--apply-review", metavar="JSON",
                    help="Ledger WIP Review: compute as usual, revert every "
                         "DISAPPROVED field in this decisions JSON to the current "
                         "tab value, then write 'Test - CP' normally.")
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="Show debug logs (timestamps + module names).")
    args = ap.parse_args()

    # Interactive mode uses pretty terminal output (no log prefixes).
    # --verbose reverts to the standard timestamped log format for debugging.
    # Route logging to stdout (not stderr) so pretty print order is preserved
    # when stdout/stderr are merged.
    if args.verbose:
        logging.basicConfig(
            level=logging.DEBUG,
            format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
            stream=sys.stdout,
        )
    else:
        logging.basicConfig(
            level=logging.WARNING,
            format="  %(levelname)s: %(message)s",
            stream=sys.stdout,
        )

    # ── Header ──
    print()
    print(_Term.color(_Term.BOLD + _Term.CYAN, "  CP WIP Reader"))
    print(_Term.color(_Term.DIM, "  " + dt.datetime.now().strftime("%Y-%m-%d %H:%M")))

    _section("Configuration")
    _kv("Active dir",    _shorten(CP_ACTIVE_DIR))
    _kv("Completed dir", _shorten(CP_COMPLETED_DIR))
    _kv("WIP target",    _shorten(WIP_EXCEL_PATH))
    import wip_excel_guard
    _kv("Write target",  f"{TEST_TAB!r} tab (guard allow-list: "
                         f"{', '.join(sorted(wip_excel_guard.ALLOWED_WRITE_SHEETS))})")

    # ── Scan ──
    _section("Scanning CP folders on Synology")
    active_rows = scan_cp_folders(CP_ACTIVE_DIR, is_completed=False)
    completed_rows = scan_cp_folders(CP_COMPLETED_DIR, is_completed=True)
    rows = active_rows + completed_rows
    _kv("Active",    f"{len(active_rows)} folder(s)")
    _kv("Completed", f"{len(completed_rows)} folder(s)")
    _kv("Total",     f"{len(rows)} folder(s) scanned")

    if args.project:
        pf = args.project.upper()
        rows = [r for r in rows if r.project_num == pf]
        _kv(f"--project {pf}", f"{len(rows)} row(s) after filter")

    if not rows:
        print()
        print(_Term.color(_Term.AMBER, "  No CP projects to process — exiting"))
        return 0

    # ── QBO enrichment ──
    if not args.no_qbo:
        _section("Enriching with QBO Billed/Costs")
        print(f"  {_Term.color(_Term.DIM, 'Fetching...')}")
        t0 = dt.datetime.now()
        enrich_with_qbo(rows)
        elapsed = (dt.datetime.now() - t0).total_seconds()
        print(_Term.color(_Term.GREEN, f"  ✓ {len(rows)} project(s) enriched") +
              _Term.color(_Term.DIM, f" ({elapsed:.1f}s)"))
    else:
        _section("Skipping QBO join (--no-qbo)")

    # Flagged rows = the script couldn't verify a number → red in Excel
    # (the user 2026-07-13: bad-looking numbers render red).
    for row in rows:
        row.needs_review = bool(row.status_flags)

    # ── Ledger WIP Review: emit the diff and STOP, or apply decisions then write ──
    if args.emit_review:
        prior = WR.snapshot_tab(WIP_EXCEL_PATH, TEST_TAB, "working")
        recs = WR.diff_rows(rows, prior, division="Commercial",
                            tab_name=TEST_TAB, tab_kind="working")
        WR.write_review_json(args.emit_review, "Commercial", TEST_TAB, recs)
        print(f"  ✓ WIP review emitted → {args.emit_review} ({WR.summarize(recs)})")
        return 0
    if args.apply_review:
        # The SAME carry rule the review applied: a PM field with no source this
        # run keeps the tab value, so an approved sync can never blank it.
        prior = WR.snapshot_tab(WIP_EXCEL_PATH, TEST_TAB, "working")
        rows = WR.apply_decisions(rows, WR.load_decisions(args.apply_review),
                                  prior=prior, tab_kind="working")

    # ── Write / dry-run report ──
    try:
        # Active-only default view (the user 2026-07-31: "don't want to see
        # Closed on default open") — same as the master tab.
        wrote = W.write_test_cp(rows, WIP_EXCEL_PATH, dry_run=args.dry_run,
                              default_filter_active=True,
                              title="CP WIP REPORT", summary=True,
                              # working tab → live roll-up formulas
                              live_formulas=True)
    except W.WipWriteDenied as e:
        print(_Term.color(_Term.RED, f"  ✗ Guard blocked write: {e}"))
        return 2
    except FileNotFoundError as e:
        print(_Term.color(_Term.RED, f"  ✗ {e}"))
        return 3

    if not args.dry_run and wrote:
        _section("Done")
        print(_Term.color(_Term.GREEN, f"  ✓ Wrote {len(rows)} row(s) to {TEST_TAB!r}"))
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())

