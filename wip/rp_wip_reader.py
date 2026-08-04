#!/usr/bin/env python3
"""
rp_wip_reader.py — RP (Residential) WIP reader — v2 (the user 2026-07-13).

MODEL (locked 2026-07-10, built 2026-07-13):
  • SOURCE = the General List workbook (READ-ONLY, never written): sheets
    'General list - Alpha order' + 'Small Jobs' (identical layout — Small Jobs
    holds the jobs kept out of Alpha to avoid clutter). Prices are entered
    there by hand: AI=SLAB BID $, AJ=SLAB COST $, AK=FLATWORK BID $,
    AL=FLATWORK COST $. Completion % in Z, job # in C.
  • The WIP AUTO-SPLITS each RP job into TWO lines: RP#### (slab — contract=AI,
    ETC=AJ) and RP####-FTW (flatwork — contract=AK, ETC=AL). Flatwork is
    pre-bid and assumed to follow the slab; RP####-FTW is its own standalone
    QBO project. CP#### jobs in the list are STANDALONE (never -FTW): one
    line, contract=AI+AK, ETC=AJ+AL, billed under the plain CP#.
  • POUR FLATWORK (col AF) = "OTHER" ⇒ another contractor won the flatwork
    (the user 2026-07-16): that scope NEVER enters the WIP — no -FTW line
    (and no flat $ in a CP standalone sum), even when prices were entered.
  • QBO enrich per line (read-only GET): Billed = P&L income of the line's
    own customer; Costs = COGS + expenses. A job is DONE when 100% complete
    AND billed = contract (the RP manager's rule — billing is the truth).
  • needs_review → RED numbers in Excel (the user 2026-07-13): billed over
    contract · 100% but not fully billed (punch work) · billed with no
    contract in the sheet · fully billed but <100% on the list · contract
    with no QBO project.
  • Writes ONLY the 'Test - RP' tab of the WIP master (wip_excel_guard).
    Links: project name → Residential folder; Contract/ETC → the General
    List; Billed → QBO customer page; Costs → QBO project P&L report.

Usage:
  python3 rp_wip_reader.py --dry-run               # preview, no write
  python3 rp_wip_reader.py --project RP7538        # one job (both lines)
  python3 rp_wip_reader.py --no-qbo --dry-run      # fast local test
  python3 rp_wip_reader.py                         # live run → Test - RP
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import load_workbook

# Reuse the CP writer so RP gets the exact same WIP structure/formatting/links,
# just written to the 'Test - RP' tab. Intra-folder import (allowed).
import cp_wip_reader as CP
from shared import paths, qbo_api

def _rp_cols():
    """RP column set — the SAME standard every other Test tab uses.

    It must stay aligned (2026-08-03): this tool and
    `master_wip_test --rp-from-file` both write 'Test - RP', so a layout that
    drifts here silently regresses the tab the moment anyone runs this script.
    That is exactly what happened — a run of this reader replaced the
    master-aligned tab with a header-on-row-1 layout that had no APPROVED COs
    column at all.

    APPROVED COs is kept (a change order is half the contract story), and the
    old WHY (TEMP) column is gone — it was a file:// link into a Downloads
    workbook, which the QBO-links-only rule no longer allows.
    """
    drop = {"retainage_held"}
    cols = []
    for label, width, field in CP.COLS:
        if field in drop:
            continue
        cols.append((label, width, field))
        if field == "project_name":
            cols.append(("TYPE", 9, "home_type"))
            cols.append(("BUILDER", 18, "client"))
    return cols

ALPHA_PATH = Path(os.getenv(
    "RP_ALPHA_PATH",
    "/Volumes/Common/OPERATIONS/GENERAL LIST/LISTA GENERAL AÑO 2026.xlsx",
))
ALPHA_SHEET = "General list - Alpha order"
SMALL_SHEET = "Small Jobs"
RP_ROOT = Path(os.getenv(
    "RP_ROOT",
    "/Volumes/Common/CURRENT PROJECTS/Residential",
))

# General List column map (1-based), header row 4, data from row 6.
COL_JOB, COL_HOUSE, COL_STREET, COL_CITY, COL_COMPLETION = 3, 4, 5, 6, 26
COL_SLAB_BID, COL_SLAB_COST, COL_FLAT_BID, COL_FLAT_COST = 35, 36, 37, 38
# AF = POUR FLATWORK: a date (we poured it), blank (not yet), or "OTHER" —
# someone ELSE won the flatwork (the user 2026-07-16) → that scope is NOT
# ours: no -FTW line, even when flatwork prices were entered on the row.
COL_POUR_FLAT = 32

# RP#### (with optional suffix like -FTW already in the list) or CP#### —
# CP jobs live here because the RP team runs them (standalone, never split).
_JOB_RE = re.compile(r"^(RP\d{4}|CP\d{3,4})\b", re.IGNORECASE)
_RP_RE = re.compile(r"RP\d{4}(?:-[A-Za-z]{2,6})?(?!\d)", re.IGNORECASE)

# Home type (the user 2026-07-14): tract = production builders (repeat volume
# work) — everyone else is custom. Edit these sets as builders come and go.
TRACT_CLIENTS = {"GRAND HOMES", "WILLIAM RYAN HOMES",
                 "DALLAS AREA HABITAT FOR HUMANITY",
                 # Camden — named by the team as a tract builder in the
                 # 2026-07-22 review (STATUS.md, "Team fixes round 1").
                 "CAMDEN HOMES"}
TRACT_CODES = {"GRAND", "WRYAN", "DAHH"}

# RP tab column layout (the user 2026-07-14): TYPE + CLIENT after the name;
# no Approved COs / Retainage (those are draw-model columns — CP only);
# temp WHY column links each row to the run's justification JSON dump.
RP_COLS = None   # built after CP import below

# Fully billed tolerance — within half a percent of contract counts as billed
# out (retainage/rounding noise).
_FULL_TOL = 0.005
_VARIANCE_CLOSE = 1000.0   # 100% + billed within $1K of contract → Closed w/ variance
                           # (the user 2026-07-14 materiality rule; covers WRH's 1.5% fee)


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").upper().strip())


def _money(v):
    try:
        f = float(v)
        return f if f else None
    except (TypeError, ValueError):
        return None




# ══════════════════════════════════════════════════════════════════════
# THE OWNER'S RP WIP FILE — the binding RP source of truth
# ══════════════════════════════════════════════════════════════════════
# The estimators' verified snapshot, not the General List. Lives HERE (not in
# master_wip_test) so `rp_wip_reader.py` run on its own and the unified writer
# produce the SAME 'Test - RP' tab from the SAME code — a copy in each was how
# the tab regressed on 2026-08-04.
RP_WIP_FILE = paths.get_path(
    "RP_WIP_FILE", paths.onedrive_base() / "RP WIP TO FIX_Final.xlsx")

# The owner's fixed RP WIP ('RP WIP' sheet) — band rows partition the lines.
# Band substring → (SECTION on the master, TYPE on Test - RP). None = excluded:
# CP jobs belong to the CP folder scan, never the RP section (the user
# 2026-07-29). TYPE wording is the user's (2026-07-31): the top section is
# just GOOD; the other three bands are their own types.
_RP_FILE_BANDS = (
    ("CP JOBS",                  None),
    ("DROPPED OFF THE SCHEDULE", ("RP — DROPPED, UNBILLED", "DROPPED OFF SCHEDULE")),
    ("FTW WITH COSTS",           ("FTW — OFF-SCHEDULE (COSTS)", "FTW WITH COSTS")),
    ("FTW BACKLOG",              ("FTW BACKLOG", "FTW BACKLOG")),
)
_RP_FILE_COL = {"job": 1, "addr": 2, "builder": 3, "contract": 4, "etc": 5,
                "billed": 6, "costs": 7, "sched": 9, "action": 11, "co": 12}

# The owner's colour contract (rp_wip_update.py, the user 2026-07-30) — his
# font-colour marks are authoritative and must SURVIVE into the test tabs
# (the user 2026-07-31: "keep all the notes and colors since they mean
# something"). theme 9 = the workbook's orange → written as literal orange.
_OWNER_RGB = {"00B050": "00B050",     # green — the owner verified this number
              "FF0000": "FF0000"}     # red   — the owner changed this number
_OWNER_ORANGE = "ED7D31"              # theme 9 — ops manager must verify


# Divisions are SPELLED OUT on the report (the user 2026-08-03) — the codes
# are internal shorthand, not something a reader of the WIP should decode.
SECTION_LABEL = {
    "MFD": "Multi-Family",
    "CP": "Commercial",
    "RP SLAB": "Residential — Slab",
    "FTW — ACTIVE": "Residential — Flatwork",
    "FTW — OFF-SCHEDULE (COSTS)": "Residential — Flatwork (off-schedule)",
    "RP — DROPPED, UNBILLED": "Residential — Dropped, Unbilled",
    "FTW BACKLOG": "Residential — Flatwork Backlog",
}


def _rp_category(row) -> str:
    """The row's TYPE for a line from the owner's top section, decided AFTER
    the QBO pass (the user 2026-07-31: "rp7234-ftw is not good … there are no
    costs" — the top band alone doesn't make a job good).

    GOOD means work is actually underway: QBO shows costs or billing.
    Anything with no money moved yet is NOT STARTED when it's still on the
    schedule, and — for flatwork — plain backlog when it isn't (the locked
    FTW backlog rule: no activity AND not scheduled)."""
    if row.billed_to_date or row.costs_to_date:
        return "GOOD"
    if getattr(row, "on_schedule", False):
        return "NOT STARTED"
    return "FTW BACKLOG" if row.project_num.endswith("-FTW") else "NOT STARTED"


def _owner_mark(cell):
    """The owner's colour on this cell ('00B050' / 'FF0000' / orange), or None."""
    col = cell.font.color if cell.font else None
    if col is None:
        return None
    rgb = getattr(col, "rgb", None)
    if isinstance(rgb, str) and rgb[-6:].upper() in _OWNER_RGB:
        return _OWNER_RGB[rgb[-6:].upper()]
    th = getattr(col, "theme", None)
    return _OWNER_ORANGE if th == 9 else None


def read_rp_from_file(xlsx_path: Path):
    """RP rows from the owner's verified RP WIP workbook (the user 2026-07-29:
    'for RP, use this excel'). His contract/ETC/CO values are taken as-is —
    the file IS the RP source of record. Billed/Costs are pre-seeded from the
    file, then refreshed from QBO (QBO wins; the file value only survives a
    failed lookup). CP-numbered lines are EXCLUDED (they belong to the CP
    folder scan). Sections come from the file's band rows; lines above the
    first band split RP SLAB vs FTW — ACTIVE by the -FTW suffix."""
    import re as _re
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["RP WIP"]
    C = _RP_FILE_COL
    rows, section, skipped_cp = [], "", []
    seen = {}                                          # project # → first row
    for r in range(3, ws.max_row + 1):
        raw = ws.cell(r, C["job"]).value
        if raw is None or not str(raw).strip():
            continue
        job = str(raw).strip().upper()
        if not _re.match(r"^(RP|CP)\d", job):          # band row → new section
            band = str(raw)
            for key, sect in _RP_FILE_BANDS:
                if key in band:
                    section = sect
                    break
            continue
        if job.startswith("CP"):
            skipped_cp.append(job)
            continue
        if section is None:                            # inside an excluded band
            skipped_cp.append(job)
            continue
        contract = _money(ws.cell(r, C["contract"]).value)
        etc = _money(ws.cell(r, C["etc"]).value)
        if job in seen:
            # The file lists the line twice (e.g. under two warning bands).
            # One WIP line per job — keep the FIRST (fuller) copy; if the
            # duplicate carries DIFFERENT numbers, flag it red so a human
            # settles which contract/ETC is real.
            first = seen[job]
            if (contract, etc) != (first.base_contract, first.base_etc):
                first.status_flags.append(
                    f"Duplicate line in the RP file with different numbers "
                    f"(row {r}: contract {contract}, ETC {etc}) — verify")
            print(f"    ⚠ {job}: duplicate line in the RP file (row {r}) — "
                  f"kept the first copy"
                  + ("" if (contract, etc) == (first.base_contract, first.base_etc)
                     else " (NUMBERS DIFFER — flagged)"))
            continue
        row = CP.CpRow(
            job, str(ws.cell(r, C["addr"]).value or job).strip(), False,
            contract,
            _money(ws.cell(r, C["co"]).value),
            etc,
            _money(ws.cell(r, C["billed"]).value),
            _money(ws.cell(r, C["costs"]).value))
        row.client = str(ws.cell(r, C["builder"]).value or "").strip() or None
        # TYPE = Tract / Custom (the user 2026-07-31 — it must not disappear
        # from the RP view). Same rule as the GL pipeline: production builders
        # (by name OR by the General-Lista code) are Tract, everyone else is
        # Custom.
        _b = _norm(row.client)
        row.home_type = ("Tract" if (_b in TRACT_CLIENTS
                                     or _b in TRACT_CODES) else "Custom")
        row.on_schedule = str(ws.cell(r, C["sched"]).value or "").strip() == "✓"
        if section:
            row.section, row.rp_type = section
        else:
            row.section = ("FTW — ACTIVE" if job.endswith("-FTW") else "RP SLAB")
            row.rp_type = None          # decided after QBO — see _rp_category()
        # The owner's notes + colour marks travel with the row (the user
        # 2026-07-31). A red/green/orange mark on Billed/Costs also means
        # HIS number stands — the QBO refresh must not overwrite it.
        row.action_note = str(ws.cell(r, C["action"]).value or "").strip() or None
        row.notes_from_source = True     # his file is authoritative for NOTES
        row.cell_marks, row.qbo_protect = {}, {}
        for fld, ccol in (("contract_price", "contract"), ("etc", "etc"),
                          ("billed_to_date", "billed"), ("costs_to_date", "costs"),
                          ("action_note", "action")):
            mark = _owner_mark(ws.cell(r, C[ccol]))
            if mark:
                row.cell_marks[fld] = mark
                if fld in ("billed_to_date", "costs_to_date"):
                    row.qbo_protect[fld] = getattr(row, fld)
        seen[job] = row
        rows.append(row)
    wb.close()
    if skipped_cp:
        print(f"  RP file: excluded {len(skipped_cp)} CP/band line(s): "
              f"{', '.join(skipped_cp)}")
    return rows



def classify_from_file(rows):
    """Post-QBO pass over rows from the owner's file → rows in report order.

    (1) An owner-marked Billed/Costs value stands — undo the QBO overwrite.
    (2) LOCKED backlog rule: a backlog line with ANY costs/billing is not
        backlog.  (3) The top section's CATEGORY is decided from the DATA, not
        from the band it sat in.  (4) Backlog lines never render red — having
        no QBO project is their normal state, not a must-fix.
    """
    order = ["RP SLAB", "FTW — ACTIVE", "FTW — OFF-SCHEDULE (COSTS)",
             "RP — DROPPED, UNBILLED", "FTW BACKLOG"]
    for row in rows:
        for fld, val in (getattr(row, "qbo_protect", None) or {}).items():
            if getattr(row, fld) != val:
                print(f"    ⚠ {row.project_num}: {fld} kept at the owner's "
                      f"marked value (QBO refresh discarded)")
                setattr(row, fld, val)
        if (row.section == "FTW BACKLOG"
                and (row.billed_to_date or row.costs_to_date)):
            row.section = "FTW — OFF-SCHEDULE (COSTS)"
            row.rp_type = "FTW WITH COSTS"
            print(f"    ⚠ {row.project_num}: QBO shows activity — "
                  f"reclassed out of FTW BACKLOG")
        if row.rp_type is None:                 # owner's top section
            row.rp_type = _rp_category(row)
            if row.rp_type != "GOOD":
                print(f"    • {row.project_num}: no costs/billing → "
                      f"{row.rp_type} (was in the top section)")
        if row.base_etc is None:
            row.status_flags.append("No budget (ETC) on the RP file")
        row.needs_review = (bool(row.status_flags)
                            and row.section != "FTW BACKLOG")
    return [r for s in order for r in rows if r.section == s]


def rp_tab_cols():
    """'Test - RP' columns — the master standard plus the RP-only extras.

    CATEGORY describes the row; TYPE stays Tract/Custom exactly as on the
    master layout. This is a WORKING tab, so unlike Test-Master it keeps the
    commentary and the sync stamp (money_bleeds reads both from here)."""
    drop = {"retainage_held", "notes_text", "_last_synced", "_notes_all"}
    cols = [("CATEGORY", 20, "rp_type")]
    for label, width, field in CP.COLS:
        if field in drop:
            continue
        cols.append((label, width, field))
        if field == "project_name":
            cols.append(("TYPE", 9, "home_type"))
            cols.append(("BUILDER", 24, "client"))
    return cols + [("LAST SYNCED", 18, "_last_synced"),
                   ("NOTES", 60, "_notes_all")]


RP_TAB_LEGEND = [
    ("LEGEND — CATEGORY (what each row's TYPE means):", None, True),
    ("GOOD — work is underway: QBO shows costs and/or billing", None, False),
    ("NOT STARTED — on the schedule / priced, but no costs and no billing yet", None, False),
    ("FTW WITH COSTS — flatwork started off-schedule (has costs) — belongs on the schedule", None, False),
    ("DROPPED OFF SCHEDULE — left the schedule with money still on the table", None, False),
    ("FTW BACKLOG — flatwork priced with the slab, no activity and not scheduled (expected wins)", None, False),
    ("COLORS:  GREEN = the owner verified this number", "00B050", True),
    ("RED = the owner changed this number", "FF0000", True),
    ("ORANGE = ops manager must verify", "ED7D31", True),
]


def write_rp_tab(rows, dry_run: bool = False) -> bool:
    """Write 'Test - RP'. The ONLY writer of that tab — both this script and
    master_wip_test come through here, so the two can't drift apart."""
    type_order = ["GOOD", "NOT STARTED", "FTW WITH COSTS",
                  "DROPPED OFF SCHEDULE", "FTW BACKLOG"]
    view = [r for t in type_order for r in rows
            if getattr(r, "rp_type", None) == t]
    wrote = CP.write_test_cp(
        view, CP.WIP_EXCEL_PATH, dry_run=dry_run, tab_name="Test - RP",
        cols=rp_tab_cols(), default_filter_active=True,
        title="RP WIP REPORT", summary=True, qbo_links_only=True,
        legend=RP_TAB_LEGEND,
        # A working tab: the roll-ups are live formulas so an edited
        # ORIGINAL COST / CO COST adds up on the spot.
        live_formulas=True)
    if not dry_run and wrote:
        print(f"  ✓ Wrote {len(view)} RP line(s) to 'Test - RP'")
    return wrote


# ─────────────────── General List: jobs + prices ───────────────────
def read_general_list(path: Path):
    """Read BOTH source sheets (Alpha + Small Jobs) READ-ONLY.
    Returns [{job, source, completion, house, street, city, slab_bid,
    slab_cost, flat_bid, flat_cost}] for every RP/CP job with any price
    entered OR completion < 100%. Alpha wins on duplicates."""
    wb = load_workbook(path, data_only=True)
    out, seen = [], set()
    missing_sheets = []
    for sheet in (ALPHA_SHEET, SMALL_SHEET):
        if sheet not in wb.sheetnames:
            missing_sheets.append(sheet)
            continue
        ws = wb[sheet]
        hdr = _norm(ws.cell(4, COL_SLAB_BID).value)
        if "SLAB BID" not in hdr:
            print(f"  ⚠ {sheet!r}: no SLAB BID header at col AI — "
                  f"price columns missing, sheet skipped")
            continue
        for r in range(6, ws.max_row + 1):
            job = ws.cell(r, COL_JOB).value
            if not job:
                continue
            m = _JOB_RE.match(str(job).strip())
            if not m:
                continue
            job = m.group(1).upper()
            if job in seen:
                continue
            comp = ws.cell(r, COL_COMPLETION).value
            pour_flat = ws.cell(r, COL_POUR_FLAT).value
            rec = {
                "job": job, "source": sheet, "gl_row": r,
                "flat_other": "OTHER" in str(pour_flat or "").upper(),
                "completion": comp if isinstance(comp, (int, float)) else None,
                "builder": ws.cell(r, 2).value,
                "house": ws.cell(r, COL_HOUSE).value,
                "street": ws.cell(r, COL_STREET).value,
                "city": ws.cell(r, COL_CITY).value,
                "slab_bid": _money(ws.cell(r, COL_SLAB_BID).value),
                "slab_cost": _money(ws.cell(r, COL_SLAB_COST).value),
                "flat_bid": _money(ws.cell(r, COL_FLAT_BID).value),
                "flat_cost": _money(ws.cell(r, COL_FLAT_COST).value),
            }
            has_price = any(rec[k] for k in
                            ("slab_bid", "slab_cost", "flat_bid", "flat_cost"))
            # PRICED ONLY (the user 2026-07-14): a job with no price for any
            # scope was SKIPPED on purpose ("unless we put the contract price
            # there just to be done") — it does not enter the WIP. Kills the
            # bloat of empty slab lines from Small Jobs / legacy Alpha rows.
            if has_price:
                seen.add(job)
                out.append(rec)
    wb.close()
    return out, missing_sheets


# ─────────────────── Today's schedule: flatwork crews ──────────────
SCHEDULE_DIR = Path(os.getenv(
    "RP_SCHEDULE_DIR", "/Volumes/Common/OPERATIONS/SCHEDULE"))
_SCHED_FILE_RE = re.compile(r"Schedule (\d{1,2})-(\d{1,2})-(\d{2})\.xlsx$",
                            re.IGNORECASE)


def read_schedule_flatwork(sched_dir: Path):
    """Latest daily schedule → the set of normalized ADDRESSES whose crew
    description mentions flatwork (the user 2026-07-14: anything on the
    schedule under Flatwork is won and being worked — never backlog).
    Returns (set_of_norm_addresses, schedule_label) — empty set if the
    schedule can't be found/read (backlog rule then falls back to $-only)."""
    best = None
    try:
        for year_dir in sched_dir.iterdir():
            if not (year_dir.is_dir() and year_dir.name.isdigit()):
                continue
            for month_dir in year_dir.iterdir():
                if not month_dir.is_dir():
                    continue
                for f in month_dir.iterdir():
                    m = _SCHED_FILE_RE.search(f.name)
                    if m:
                        mo, dy, yy = (int(g) for g in m.groups())
                        key = (2000 + yy, mo, dy)
                        if best is None or key > best[0]:
                            best = (key, f)
    except OSError:
        return set(), None
    if best is None:
        return set(), None
    _key, path = best
    label = f"{_key[1]}-{_key[2]}-{_key[0] % 100:02d}"
    addrs = set()
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet = next((s for s in wb.sheetnames
                      if s.strip().lower() == "daily schedule"), wb.sheetnames[0])
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=1, max_row=200):
            vals = [str(getattr(c, "value", "") or "") for c in row[:8]]
            desc = " ".join(vals).upper()
            if "FLATWORK" in desc or re.search(r"\bFTW\b", desc):
                # address is the first street-looking cell in the row
                for v in vals:
                    vn = _norm(v)
                    if re.match(r"^\d+[A-Z0-9 ,.'-]+$", vn) and len(vn) > 8:
                        addrs.add(vn)
                        break
        wb.close()
    except Exception:
        return set(), None
    return addrs, label


# ─────────────────── Residential: folder lookup ────────────────────
def _list_dir(path: Path):
    """One directory listing via scandir → (subdirs, filenames). Errors → empty."""
    subdirs, files = [], []
    try:
        with os.scandir(path) as it:
            for e in it:
                try:
                    if e.is_dir():
                        subdirs.append(Path(e.path))
                    elif e.is_file():
                        files.append(e.name)
                except OSError:
                    pass
    except OSError:
        pass
    return subdirs, files


def index_residential(root: Path, workers: int = 24):
    """RP#→folder index from filenames at client (depth 1) + address (depth 2)
    levels, listed in parallel (Synology round-trips overlap)."""
    root = Path(root)
    rp_to_folders = defaultdict(set)
    addr_folders = []

    clients, _root_files = _list_dir(root)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        client_listings = list(ex.map(_list_dir, clients))

    addr_dirs = []
    for client, (subs, files) in zip(clients, client_listings):
        for fn in files:
            for m in _RP_RE.findall(fn):
                rp_to_folders[m.upper()].add(client)
        addr_dirs.extend(subs)

    with ThreadPoolExecutor(max_workers=workers) as ex:
        addr_listings = list(ex.map(_list_dir, addr_dirs))

    for addr, (_subs, files) in zip(addr_dirs, addr_listings):
        addr_folders.append((_norm(addr.name), addr))
        for fn in files:
            for m in _RP_RE.findall(fn):
                rp_to_folders[m.upper()].add(addr)
    return rp_to_folders, addr_folders


def match_by_address(rec, addr_folders):
    """Fallback: match house+street to an address folder name."""
    house = _norm(rec["house"])
    street_words = [w for w in _norm(rec["street"]).split()
                    if w not in ("ROAD", "STREET", "ST", "DRIVE", "DR", "LANE",
                                 "LN", "AVENUE", "AVE", "COURT", "CT", "TRAIL",
                                 "CIRCLE", "WAY", "BLVD", "N", "S", "E", "W")]
    for norm_name, folder in addr_folders:
        if house and house in norm_name and any(w in norm_name for w in street_words):
            return folder
    return None


# ─────────────────── line building + classification ────────────────
def _classify(row: CP.CpRow, completion) -> None:
    """The RP done-rule (billing is the truth): appends notes/flags and sets
    needs_review (→ RED numbers in Excel) in-place."""
    K = row.base_contract
    B = row.billed_to_date
    c100 = isinstance(completion, (int, float)) and completion >= 0.999
    if completion is not None:
        row.notes.append(f"{completion * 100:.0f}% complete (list)")

    if K and B is None and row.qbo_customer_id is None and row.status_flags:
        # QBO enrich already flagged (no project / fetch failed) — red it.
        row.needs_review = True
        return
    if not K:
        # No contract on this line — ANY QBO activity is a red flag
        # (the user 2026-07-14, RP5542: slab line had no bid yet QBO carried
        # $9.6K of costs — either the bid was never entered or the costs are
        # coded to the wrong project/scope).
        activity = []
        if B:
            activity.append(f"billed ${B:,.0f}")
        if row.costs_to_date:
            activity.append(f"costs ${row.costs_to_date:,.0f}")
        if activity:
            row.status_flags.append(
                "No contract in the list but QBO has "
                + " · ".join(activity) + " — enter the bid or fix the coding")
            row.needs_review = True
        return
    if B is None:
        return                      # --no-qbo run: contract-only view
    if B > K * (1 + _FULL_TOL):
        row.notes.append(f"Billed OVER contract by ${B - K:,.0f}")
        row.needs_review = True
    elif B >= K * (1 - _FULL_TOL):
        if c100:
            # The done-rule (RP manager): 100% complete AND fully billed →
            # STATUS reads 'Closed' (the user 2026-07-14: STATUS was hardcoded
            # Active for every line; it must derive from this rule).
            row.notes.append("CLOSED — fully billed + 100%")
            row.is_completed = True
        else:
            row.notes.append("Fully billed but list <100% — update completion")
            row.needs_review = True
    elif B == 0:
        if c100:
            row.notes.append("100% but $0 billed — punch/flatwork backlog")
            row.needs_review = True
        else:
            row.notes.append("Not billed yet")
    else:
        if c100:
            gap = K - B
            if gap <= _VARIANCE_CLOSE:
                # Materiality rule (the user 2026-07-14): 100% + billed within
                # $1K of contract = close with a documented small variance
                # (builder fee like WRH's 1.5%, or an approved write-down) —
                # amber note, NOT red. Over the threshold: chase the billing.
                row.notes.append(
                    f"CLOSED — small variance ${gap:,.0f} (fee/write-down)")
                row.is_completed = True
            else:
                row.notes.append(f"100% but only ${B:,.0f} billed — bill the rest")
                row.needs_review = True
        else:
            row.notes.append("Partially billed — treat as draw")


def build_lines(records, rp_to_folders, addr_folders):
    """General List records → CpRow lines. RP splits slab/-FTW; CP stays one
    standalone line (never -FTW — the user 2026-07-13)."""
    rows = []
    # Learn builder-code → full client name from every record whose folder
    # DID resolve (the user 2026-07-14: 'MARVE' leaking through = a job with
    # no folder match fell back to the GL code; siblings know the real name).
    code_names = {}
    for rec in records:
        folders = sorted(rp_to_folders.get(rec["job"], ()),
                         key=lambda f: (f.parent.name, f.name))
        folder = folders[0] if folders else match_by_address(rec, addr_folders)
        if folder is not None and rec.get("builder"):
            cname = (folder.name if _norm(folder.parent.name) == "RESIDENTIAL"
                     else folder.parent.name)
            code_names.setdefault(_norm(rec["builder"]), cname)
    for rec in sorted(records, key=lambda x: x["job"]):
        name = _norm(f"{rec['house'] or ''} {rec['street'] or ''}") or rec["job"]
        folders = sorted(rp_to_folders.get(rec["job"], ()),
                         key=lambda f: (f.parent.name, f.name))
        folder = folders[0] if folders else match_by_address(rec, addr_folders)

        def _mk(line_num, contract, etc):
            row = CP.CpRow(line_num, name, False, contract, None, etc, None, None)
            row.folder_path = folder
            row.takeoff_path = ALPHA_PATH      # Contract/ETC cells link → the List
            # PROJECT # → the exact General List row it came from (the user
            # 2026-07-15: click the number, land where the data lives).
            row.src_link = str(ALPHA_PATH)
            row.src_fragment = CP._sheet_fragment(rec["source"], f"C{rec['gl_row']}")
            if rec["source"] == SMALL_SHEET:
                row.notes.append("Small Jobs list")
            # CLIENT = the Residential client folder (full name) when found,
            # else the GL builder code. TYPE = Tract for production builders.
            if folder is not None:
                cname = (folder.name if _norm(folder.parent.name) == "RESIDENTIAL"
                         else folder.parent.name)
            else:
                code = _norm(rec.get("builder"))
                cname = code_names.get(code) or (str(rec.get("builder") or "") or None)
            row.client = cname
            code = _norm(rec.get("builder"))
            row.home_type = ("Tract" if (_norm(cname) in TRACT_CLIENTS
                                         or code in TRACT_CODES) else "Custom")
            return row

        # POUR FLATWORK = OTHER (list col AF): the flatwork went to another
        # contractor (the user 2026-07-16) — the flatwork scope never enters
        # the WIP, priced or not. The slab line (if priced) is still ours.
        flat_other = rec.get("flat_other")

        if rec["job"].startswith("CP"):
            # CP standalone: whole contract on one line, bills under CP#.
            contract = ((rec["slab_bid"] or 0)
                        + (0 if flat_other else (rec["flat_bid"] or 0))) or None
            etc = ((rec["slab_cost"] or 0)
                   + (0 if flat_other else (rec["flat_cost"] or 0))) or None
            if contract is None and etc is None:
                continue            # only scope was flatwork and OTHER won it
            row = _mk(rec["job"], contract, etc)
            row.notes.append("CP standalone (never -FTW)")
            if flat_other:
                row.notes.append("Flatwork = OTHER on the list — excluded")
            rows.append((row, rec["completion"], rec))
            continue

        # A line exists ONLY for a scope that has pricing (the user
        # 2026-07-14): Small Jobs entries are usually flatwork-only — the
        # clerk confirmed what is real by pricing it. No slab price → no
        # slab line (e.g. RP5542: take only the -FTW line).
        if rec["slab_bid"] or rec["slab_cost"]:
            row = _mk(rec["job"], rec["slab_bid"], rec["slab_cost"])
            if flat_other:
                row.notes.append("Flatwork = OTHER on the list — no -FTW line")
            rows.append((row, rec["completion"], rec))
        if (rec["flat_bid"] or rec["flat_cost"]) and not flat_other:
            rows.append((_mk(rec["job"] + "-FTW", rec["flat_bid"],
                             rec["flat_cost"]), rec["completion"], rec))
    return rows


def enrich_with_qbo(pairs) -> None:
    """Billed (P&L income) + Costs (COGS+expenses) per LINE customer —
    read-only GETs. Sets CP.QBO_REALM so the writer builds QBO deep links."""
    try:
        access, company_id = qbo_api.load_credentials()
    except Exception as e:
        for row, *_ in pairs:
            row.status_flags.append(f"QBO Auth Failed: {type(e).__name__}")
        return
    try:
        proj_map = qbo_api.build_project_customer_map(access, company_id)
    except Exception as e:
        for row, *_ in pairs:
            row.status_flags.append(f"QBO Customer Map Failed: {type(e).__name__}")
        return
    CP.QBO_REALM = company_id
    start, end = "2019-01-01", dt.date.today().isoformat()
    for n, (row, *_rest) in enumerate(pairs, 1):
        cust = proj_map.get(row.project_num)
        if not cust:
            if row.base_contract:
                row.status_flags.append("No QBO project")
            continue
        row.qbo_customer_id = cust["id"]
        try:
            totals = qbo_api.extract_pl_totals(
                qbo_api.fetch_project_pl(access, company_id, cust["id"], start, end))
            row.billed_to_date = float(totals.get("income", 0.0) or 0.0)
            row.costs_to_date = ((totals.get("cogs", 0.0) or 0.0)
                                 + (totals.get("expenses", 0.0) or 0.0))
        except Exception as e:
            row.status_flags.append(f"QBO P&L Failed: {type(e).__name__}")
        if n % 20 == 0:
            print(f"    …{n}/{len(pairs)} lines enriched")


# ─────────────────────────── main ──────────────────────────────────
# ─────────────────── justification workbook ────────────────────────
def write_justification(pairs, backlog, out_path: Path) -> None:
    """One row per WIP line, ops-manager-readable: GL source row, completion,
    contract, QBO billed/costs, status and the sentence that defends it
    (the user 2026-07-14: every job must be justifiable one-by-one)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    GL_FILL = PatternFill("solid", fgColor="D9D9D9")
    GOOD = Font(color="006100", bold=True)
    BAD = Font(color="9C0006", bold=True)
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CUR = '"$"#,##0.00_);[Red]("$"#,##0.00)'

    def sentence(row, comp, rec, in_backlog):
        K, b = row.base_contract, row.billed_to_date
        cs = f"{comp * 100:.0f}%" if isinstance(comp, (int, float)) else "no %"
        sheet = "Alpha" if "Alpha" in rec["source"] else "Small Jobs"
        src = f"General List '{sheet}' row {rec['gl_row']}"
        if in_backlog:
            return (f"{src}: flatwork bid ${K:,.0f} entered with the slab — "
                    f"$0 billed AND $0 costs under -FTW, not on the schedule "
                    f"→ true backlog (expected win; bills when poured).")
        if row.is_completed:
            gap = (K - b) if (K and b is not None) else None
            if gap is not None and abs(gap) < K * _FULL_TOL:
                return (f"{src}: {cs} complete; contract ${K:,.2f}; QBO billed "
                        f"${b:,.2f} — EQUAL → done-rule met (100% + fully "
                        f"billed) → CLOSED.")
            return (f"{src}: {cs} complete; contract ${K:,.2f}; QBO billed "
                    f"${b:,.2f} — ${gap:,.0f} under (≤$1K materiality) → "
                    f"CLOSED with small variance (builder fee / write-down).")
        why = "; ".join(row.status_flags or row.notes or ["in progress"])
        base = f"{src}: {cs} complete"
        if K:
            base += f"; contract ${K:,.0f}"
        if b is not None:
            base += f"; billed ${b:,.0f}"
        if row.costs_to_date:
            base += f"; QBO costs ${row.costs_to_date:,.0f} (working)"
        return f"{base} → {why}"

    # Never clobber the workbook while it is open in Excel.
    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        print(f"  ⚠ {out_path.name} open in Excel — justification not refreshed")
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "JUSTIFICATION"
    ws["A1"] = ("RP WIP — LINE-BY-LINE JUSTIFICATION (regenerated every run). "
                "Source: General List (Alpha + Small Jobs) → QBO per line → "
                "Test - RP. Priced scopes only.")
    ws["A1"].font = Font(bold=True)
    ws.append([])
    HDR = ["SECTION", "LINE", "TYPE", "CLIENT", "GL SHEET", "GL ROW",
           "% COMPLETE", "CONTRACT $", "QBO BILLED $", "QBO COSTS $",
           "STATUS", "RED?", "JUSTIFICATION"]
    ws.append(HDR)
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(3, c)
        cell.font = Font(bold=True)
        cell.fill = GL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    backlog_set = {id(r) for r in backlog}
    ordered = sorted(pairs, key=lambda t: (id(t[0]) in backlog_set,
                                           not t[0].is_completed,
                                           t[0].project_num))
    line_rows = {}
    for row, comp, rec in ordered:
        in_bk = id(row) in backlog_set
        line_rows[row.project_num] = ws.max_row + 1
        ws.append([("FTW BACKLOG" if in_bk else "MAIN"), row.project_num,
                   row.home_type, row.client,
                   ("Alpha" if "Alpha" in rec["source"] else "Small Jobs"),
                   rec["gl_row"],
                   (comp if isinstance(comp, (int, float)) else None),
                   row.base_contract, row.billed_to_date, row.costs_to_date,
                   ("Closed" if row.is_completed else "Active"),
                   ("RED" if row.needs_review else ""),
                   sentence(row, comp, rec, in_bk)])
        r = ws.max_row
        for cc in range(1, len(HDR) + 1):
            ws.cell(r, cc).border = BORDER
            ws.cell(r, cc).alignment = Alignment(vertical="top",
                                                 wrap_text=(cc == len(HDR)))
        ws.cell(r, 7).number_format = "0%"
        for cc in (8, 9, 10):
            ws.cell(r, cc).number_format = CUR
        if row.is_completed:
            ws.cell(r, 11).font = GOOD
        if row.needs_review:
            ws.cell(r, 12).font = BAD
    for col, w in zip("ABCDEFGHIJKLM",
                      (13, 12, 8, 20, 10, 7, 11, 13, 13, 13, 9, 6, 92)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(out_path)
    print(f"  ✓ Justification → {out_path}")
    return line_rows



def main() -> int:
    ap = argparse.ArgumentParser(
        description="RP WIP — General List (Alpha + Small Jobs) → slab/-FTW "
                    "lines → QBO billed/costs → 'Test - RP' tab.")
    ap.add_argument("--alpha", help="override General List path")
    ap.add_argument("--root", help="override Residential root")
    ap.add_argument("--project", help="filter to one job # (both lines)")
    ap.add_argument("--no-qbo", action="store_true", help="skip QBO join")
    ap.add_argument("--dry-run", action="store_true", help="preview, no write")
    ap.add_argument("--general-list", action="store_true",
                    help="LEGACY: build RP from the General List instead of "
                         "the owner's RP WIP file. The file is the source of "
                         "truth — only use this to inspect the old pipeline.")
    ap.add_argument("--file", help="override the RP WIP file path")
    args = ap.parse_args()

    # DEFAULT: the owner's verified RP WIP file (the binding RP source of
    # truth). Running this script daily must produce exactly what
    # master_wip_test produces, so it shares that code rather than copying it.
    if not args.general_list:
        rp_file = Path(args.file) if args.file else RP_WIP_FILE
        print()
        print("  RP WIP Reader — the owner's RP WIP file → 'Test - RP'")
        print(f"  file:  {rp_file}")
        print("  " + "─" * 74)
        if not rp_file.exists():
            print(f"  ✗ RP WIP file not found: {rp_file}")
            print("    (pass --file <xlsx>, or --general-list for the legacy "
                  "pipeline)")
            return 1
        rows = read_rp_from_file(rp_file)
        print(f"  RP lines: {len(rows)}")
        if args.project:
            pf = args.project.upper()
            rows = [r for r in rows if r.project_num.upper().startswith(pf)]
            print(f"  --project {pf}: {len(rows)} line(s)")
        if not args.no_qbo:
            print("  Enriching with QBO Billed/Costs …")
            enrich_with_qbo([(r,) for r in rows])
        rows = classify_from_file(rows)
        try:
            write_rp_tab(rows, dry_run=args.dry_run)
        except CP.WipWriteDenied as e:
            print(f"  ✗ Guard blocked write: {e}")
            return 2
        return 0
    alpha = Path(args.alpha) if args.alpha else ALPHA_PATH
    root = Path(args.root) if args.root else RP_ROOT

    print()
    print("  RP WIP Reader — General List → slab/-FTW lines → Test - RP")
    print(f"  list:  {alpha}")
    print(f"  root:  {root}")
    print("  " + "─" * 74)

    if not alpha.exists():
        print(f"  ✗ General List not found: {alpha}  (Synology mounted?)")
        return 1
    records, missing_sheets = read_general_list(alpha)
    for s in missing_sheets:
        print(f"  ⚠ sheet {s!r} not in the workbook — skipped")
    print(f"  Jobs in scope (priced or active): {len(records)}")

    rp_to_folders, addr_folders = ({}, [])
    if root.exists():
        rp_to_folders, addr_folders = index_residential(root)
    else:
        print(f"  ⚠ Residential root not found: {root} — no folder links")

    pairs = build_lines(records, rp_to_folders, addr_folders)
    n_other = sum(1 for rec in records
                  if rec.get("flat_other") and (rec["flat_bid"] or rec["flat_cost"]))
    if n_other:
        print(f"  POUR FLATWORK = OTHER: {n_other} priced flatwork scope(s) "
              f"excluded (won by another contractor)")
    if args.project:
        pf = args.project.upper()
        pairs = [t for t in pairs
                 if t[0].project_num in (pf, f"{pf}-FTW")]
    print(f"  Lines (slab + -FTW + CP standalone): {len(pairs)}")
    if not pairs:
        print("  No RP lines to process — exiting")
        return 0

    if not args.no_qbo:
        print("  Enriching with QBO Billed/Costs …")
        enrich_with_qbo(pairs)
    for row, comp, _rec in pairs:
        _classify(row, comp)

    # FTW backlog → its own section at the BOTTOM of the tab (the user
    # 2026-07-14): flatwork bid together with the slab but not poured yet —
    # effectively unwon-but-expected work (~95% follows the slab). Not an
    # error state, so no red; invoiced under RP#-FTW when the pour lands.
    sched_addrs, sched_label = read_schedule_flatwork(SCHEDULE_DIR)
    if sched_label:
        print(f"  Schedule check: {sched_label} — "
              f"{len(sched_addrs)} flatwork address(es)")
    main_rows, backlog = [], []
    for row, comp, rec in pairs:
        if not row.project_num.endswith("-FTW"):
            main_rows.append(row)
            continue
        addr_n = _norm(f"{rec['house'] or ''} {rec['street'] or ''}")
        on_sched = bool(addr_n) and any(
            addr_n == a or a.startswith(addr_n) or addr_n.startswith(a)
            for a in sched_addrs)
        has_activity = bool(row.billed_to_date) or bool(row.costs_to_date)
        # WON + WORKING when there is ANY QBO activity (billed OR costs —
        # the user 2026-07-14: "how can this be a backlog if there are
        # costs? we won it!") or the job is on today's flatwork schedule.
        if has_activity or on_sched:
            if on_sched:
                row.notes.append(
                    f"On the {sched_label} schedule — flatwork crew")
            main_rows.append(row)
        else:
            row.needs_review = False   # expected pre-pour state, not an error
            backlog.append(row)
    rows = main_rows + backlog              # for counts/logs
    n_red = sum(1 for r in rows if r.needs_review)
    print(f"  Review (red): {n_red} line(s) · FTW backlog: {len(backlog)}")

    # Justification workbook (the user 2026-07-15: no JSON, human-readable,
    # in Downloads) — one row per line: where it was found in the General
    # List, completion, contract, QBO billed/costs, and the rule behind the
    # status. The WHY (TEMP) column links here. Overwritten every run.
    justify_path = Path(os.getenv(
        "RP_JUSTIFY_XLSX",
        str(Path.home() / "Downloads" / "RP WIP - Justification.xlsx")))
    just_rows = write_justification(pairs, backlog, justify_path) or {}
    for row, _comp, _rec in pairs:
        row.why_link = str(justify_path)
        jr = just_rows.get(row.project_num)
        # Jump straight to THIS line's row in the workbook (the user
        # 2026-07-15), not just the file.
        row.why_fragment = CP._sheet_fragment("JUSTIFICATION", f"A{jr}") if jr else None
        # Red must explain itself in NOTES (the user 2026-07-14) — flags carry
        # the must-fix reason; mirror it so the notes column reads standalone.
        if row.needs_review:
            reason = "; ".join(row.status_flags) or (row.notes[-1] if row.notes else "")
            if reason and f"RED: {reason}" not in row.notes:
                row.notes.append(f"RED: {reason}")

    try:
        wrote = CP.write_test_cp(
            main_rows, CP.WIP_EXCEL_PATH,
            dry_run=args.dry_run, tab_name="Test - RP",
            appendix=("FTW BACKLOG — flatwork bid with the slab, NOT poured "
                      "yet (expected wins; invoice under RP#-FTW when poured)",
                      backlog),
            cols=_rp_cols(), default_filter_active=True)
    except CP.WipWriteDenied as e:
        print(f"  ✗ Guard blocked write: {e}")
        return 2
    except FileNotFoundError as e:
        print(f"  ✗ {e}")
        return 3
    if not args.dry_run and wrote:
        print(f"  ✓ Wrote {len(rows)} line(s) to 'Test - RP'")
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
