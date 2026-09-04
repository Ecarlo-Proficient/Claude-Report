"""WIP pre-write AUDIT — the inspect-before-you-write report (the user
2026-08-07: "how can I verify what you're picking up before you update the WIP,
and I need an audit/log of the add/remove-jobs logic and the non-QBO values").

READ-ONLY. Given the rows the pipeline is ABOUT to write, the prior tab's job
set, and the full classified set, it writes one plain workbook — one row per
job — showing, for the NON-QBO parts you can't take on faith:

  · Δ vs the current report: ADDED / REMOVED / SAME, with the REASON
  · CONTRACT + the exact source cell it came from
  · ETC + its source (your RP-file cell, a takeoff file+cell, or blank)
  · QBO billed/costs alongside (the part you already trust) for context

It never touches the WIP workbook. The caller runs it under `--audit` and skips
the write, so you review this first and only then approve the real run.
"""
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
_thin = Side(style="thin", color="000000")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CUR = '"$"#,##0_);[Red]("$"#,##0)'
FONT = "Tahoma"

COLS = [
    ("PROJECT #", 14), ("NAME", 30), ("SECTION", 26),
    ("Δ VS CURRENT", 13), ("REASON", 40),
    ("CONTRACT", 14), ("CONTRACT SOURCE", 34),
    ("ETC", 14), ("ETC SOURCE", 40),
    ("QBO BILLED", 14), ("QBO COSTS", 14), ("FLAGS / NOTES", 44),
]


def _div(pn: str) -> str:
    pn = pn.upper()
    if pn.startswith("MFD"):
        return "Multi-Family"
    if pn.startswith("CP"):
        return "Commercial"
    return "Residential"


def _contract_src(row) -> str:
    s = getattr(row, "audit_contract_src", None)
    if s:
        return s
    d = _div(row.project_num)
    return ("CP folder draw (G702)" if d == "Commercial"
            else "'WIP Master' tab" if d == "Multi-Family" else "—")


def _etc_src(row) -> str:
    s = getattr(row, "audit_etc_src", None)
    if s:
        return s
    d = _div(row.project_num)
    return ("CP folder draw / proposal" if d == "Commercial"
            else "'WIP Master' tab" if d == "Multi-Family" else "—")


def build_rows(new_rows: List, prior: Dict[str, dict],
               all_classified: List, excluded_jobs) -> List[dict]:
    """Assemble audit rows: every job being written (SAME/ADDED) plus every job
    that was on the report last run but isn't now (REMOVED), each with a
    reason. `all_classified` (the full set before the bank cut) lets a REMOVED
    job say whether it was excluded by a rule vs. left the source entirely."""
    excluded_jobs = {j.upper() for j in (excluded_jobs or ())}
    classified_by = {r.project_num.upper(): r for r in (all_classified or [])}
    out, seen = [], set()
    for row in new_rows:
        pn = row.project_num.upper()
        seen.add(pn)
        was = prior.get(pn)
        status = "SAME" if was else "ADDED"
        reason = "" if was else f"new — {getattr(row, 'audit_origin', 'in source')}"
        out.append(dict(
            pn=row.project_num, name=row.project_name or "",
            section=getattr(row, "section", "") or "",
            status=status, reason=reason,
            contract=getattr(row, "contract_price", None),
            contract_src=_contract_src(row),
            etc=getattr(row, "etc", None), etc_src=_etc_src(row),
            billed=getattr(row, "billed_to_date", None),
            costs=getattr(row, "costs_to_date", None),
            flags="; ".join((getattr(row, "status_flags", None) or [])
                            + (getattr(row, "notes", None) or [])),
        ))
    for pn, was in prior.items():
        if pn in seen:
            continue
        if pn in excluded_jobs:
            reason = "REMOVED by rule — on the bank-exclude list"
        elif pn in classified_by:
            reason = (f"off the bank report — section "
                      f"'{classified_by[pn].section}' (kept on the working tab)")
        else:
            reason = "left the source — no longer in the RP file / folders"
        out.append(dict(
            pn=pn, name=was.get("name") or "", section="", status="REMOVED",
            reason=reason, contract=was.get("rev_contract"), contract_src="(prior value)",
            etc=was.get("rev_etc"), etc_src="(prior value)", billed=None, costs=None,
            flags=""))
    order = {"ADDED": 0, "REMOVED": 1, "SAME": 2}
    out.sort(key=lambda d: (order[d["status"]], _div(d["pn"]), d["pn"]))
    return out


def write_audit(new_rows: List, prior: Dict[str, dict], all_classified: List,
                excluded_jobs, out_path: Path, tab_name: str = "Test-Master") -> Path:
    """Write the audit workbook and return its path. Read-only on every source."""
    rows = build_rows(new_rows, prior, all_classified, excluded_jobs)
    wb = Workbook()
    ws = wb.active
    ws.title = "WIP Audit"
    ws.cell(1, 1, f"WIP PRE-WRITE AUDIT — {tab_name}").font = Font(name=FONT, size=10, bold=True)
    ws.cell(2, 1, f"{sum(r['status']=='ADDED' for r in rows)} added · "
                  f"{sum(r['status']=='REMOVED' for r in rows)} removed · "
                  f"{sum(r['status']=='SAME' for r in rows)} unchanged   "
                  f"(QBO billed/costs shown for context only)").font = Font(name=FONT, size=8)
    hdr = 4
    for c, (label, w) in enumerate(COLS, 1):
        cell = ws.cell(hdr, c, label)
        cell.fill = HDR_FILL
        cell.font = Font(name=FONT, size=8, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = w
    money = {"CONTRACT", "ETC", "QBO BILLED", "QBO COSTS"}
    keys = ["pn", "name", "section", "status", "reason", "contract",
            "contract_src", "etc", "etc_src", "billed", "costs", "flags"]
    for i, r in enumerate(rows, hdr + 1):
        for c, (key, (label, _w)) in enumerate(zip(keys, COLS), 1):
            cell = ws.cell(i, c, r[key])
            cell.font = Font(name=FONT, size=8, bold=(r["status"] != "SAME" and c <= 5))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(label in ("REASON", "ETC SOURCE",
                                                            "CONTRACT SOURCE", "FLAGS / NOTES")))
            if label in money:
                cell.number_format = CUR
    last = hdr + len(rows)
    if rows:
        tbl = Table(displayName="WIPAudit", ref=f"A{hdr}:L{last}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.tables.add(tbl)
    ws.freeze_panes = f"A{hdr+1}"
    out_path = Path(out_path)
    wb.save(out_path)
    return out_path
