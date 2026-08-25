"""
wip_excel_guard.py — hard-baked safety rail for writes to WIP - MASTER.xlsx.

The user's rule (2026-06-25): the script may only write to the "Test" sheet of
the team's live WIP Excel file. Every other sheet (WIP Master, WIP - CP,
WIP - MFD) is read-only at the code level until the user explicitly graduates
the script after verifying Test-sheet output is correct.

How to use:
    from wip_excel_guard import assert_write_allowed, open_wip_workbook_for_write

    wb = open_wip_workbook_for_write(path)       # raises if file missing
    ws = wb["Test"]                              # OK — Test is allow-listed
    assert_write_allowed(ws.title)               # belt-and-suspenders
    ws.cell(1, 1, value="hello")                 # safe

    ws2 = wb["WIP - CP"]
    assert_write_allowed(ws2.title)              # RAISES — not allowed

Reading any sheet is fine; this module only guards WRITES.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


# Hard-coded allow-list. Do NOT add sheets here without explicit user direction.
# To graduate the script to write live tabs, the user must say so and we add
# names here intentionally — never via config flag.
#
# 2026-06-30: expanded from {"Test"} to per-division test tabs. The user chose one
# tab per division so each division's WIP can be sanity-checked independently
# before any of them graduate to the live tabs (WIP - CP / WIP - MFD / WIP Master).
# 2026-08-25: 'WIP - MFD' GRADUATED to the live allow-list on the owner's explicit
# instruction, after 'Test - MFD' was diffed against it attribute by attribute
# (values, formulas, styles, merges, dimensions, comments, conditional formatting,
# validation, hyperlinks, images, filters AND sheet chrome) and confirmed a faithful
# superset. The two tabs were then merged into this one and 'Test - MFD' deleted;
# that name was dropped from this list on 2026-08-25 once nothing wrote to it.
# (mfd_wip_test still DELETES a stray 'Test - MFD' if a workbook carries one, but
# a sheet delete does not go through assert_write_allowed, so it needs no entry.)
# mfd_wip_test.py writes columns N..T of this sheet ONLY and never touches B..M.
# This is the documented graduation path, not a config flag.
ALLOWED_WRITE_SHEETS: frozenset[str] = frozenset({
    "Test",
    "Test-Master",         # legacy single-division sandbox — kept so existing tests pass
    "Test - CP",    # Commercial test tab
    "Test - RP",    # Residential test tab
    "WIP - MFD",    # LIVE MFD division tab — graduated 2026-08-25, N..T only
})


class WipWriteDenied(RuntimeError):
    """Raised when code attempts to write to a non-allow-listed WIP sheet."""


def assert_write_allowed(sheet_name: str) -> None:
    """Raise if writing to this sheet is not allowed.
    Call this immediately before any cell write or table modification.
    """
    if sheet_name not in ALLOWED_WRITE_SHEETS:
        raise WipWriteDenied(
            f"Refusing to write to sheet {sheet_name!r}. "
            f"The WIP Excel script may only write to: {sorted(ALLOWED_WRITE_SHEETS)}. "
            f"This rule is baked in at code level — change it only by editing "
            f"ALLOWED_WRITE_SHEETS in wip_excel_guard.py after the user's explicit OK."
        )


def open_wip_workbook_for_write(path: Path | str) -> Workbook:
    """Open the live WIP Excel for writing. Reading any sheet is fine; only
    writes to non-allow-listed sheets are blocked by assert_write_allowed.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"WIP Excel not found at {p}")
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"Expected .xlsx / .xlsm at {p}, got {p.suffix}. "
            f"openpyxl cannot write .xlsb. Save the workbook as .xlsx in Excel."
        )
    return load_workbook(p)


def list_writable_sheets() -> Iterable[str]:
    """Return the allow-listed sheet names (mostly for tests / logging)."""
    return sorted(ALLOWED_WRITE_SHEETS)
