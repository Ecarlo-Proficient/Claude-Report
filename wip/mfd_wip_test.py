#!/usr/bin/env python3
"""
mfd_wip_test.py - the MFD division WIP tab: an entry block MFD fills in, a QBO
block the script owns, and the metrics that read off both. Writes 'WIP - MFD'
in 'WIP - MASTER new.xlsx'.

COLUMN ORDER LIVES IN mfd_wip_cols.py, NOT HERE. Never write a column letter or
index in this file - ask that module. Reordering the sheet is editing its list.

HISTORY - WHY THIS WRITES A LIVE TAB
Built 2026-08-25 on a staging copy ('Test - MFD') because 'WIP - MFD' was
read-only at code level. The two were then diffed attribute by attribute and the
copy proved a faithful superset, so on the owner's instruction the tabs were
MERGED: 'WIP - MFD' was graduated in wip_excel_guard.ALLOWED_WRITE_SHEETS and the
staging tab deleted.

THE CONTRACT THAT MAKES WRITING A LIVE TAB SAFE
MFD owns the value of every 'carry' and 'input' column in the spec. This script
carries those values through a reorder and NEVER invents or overwrites one. It
owns the 'qbo' and 'calc' columns. That split is enforced in one place -
mfd_wip_cols.OWNED_BY_MFD - so it cannot drift.

LAYOUT (the owner, 2026-08-25): everything MFD types sits in ONE run so entry is a
single left-to-right pass with no calculated cell interrupting it; the QBO block
follows under its sync stamp; the metrics that drive decisions are furthest right.

FORMATTING NOTE (read before "fixing" this file)
Repo CLAUDE.md rail 5a freezes the GENERATED Test tabs to the 'WIP Master'
Tahoma-8 look. This tab is not one of them: it is MFD's own hand-kept sheet in its
original Calibri. Do not restyle it to Tahoma 8.

RETAINAGE - WHERE THE NUMBER COMES FROM (probed 2026-08-25, do not re-derive)
QBO tracks retainage properly: the invoice item '99 - Retainage' posts to a real
Other Current Asset account, 'Retainage Receivable'. A negative retainage line on
a draw DEBITS that account; billing the retainage later CREDITS it back out. So
the per-job balance of that account IS "what QBO has", pulled from the
GeneralLedger report filtered to that account.

Two traps, both handled below:
  * The GL report's account filter is 'account' (SINGULAR). Passing 'accounts' is
    silently IGNORED and returns the whole 66k-row general ledger, truncated.
  * Do NOT derive retainage from invoice lines the way cp_wip_reader does. That
    heuristic is built for CP and is wrong on MFD - on the largest job it missed
    by more than twice the retainage actually at stake - because retainage that
    has since been BILLED still sits in the invoice history.

The retainage column is EXPECTED to disagree with MFD's own Total Retainage, and
that disagreement is the point of it.

ONE JOB NUMBER CAN BE SEVERAL CONTRACTS
'WIP - MFD' carries THREE contract rows for job 192 (Hudsonwood 009, Offsite 010,
base 008) but QBO has ONE project MFD192, whose cost lines carry no contract
marker. Costs cannot be split, so QBO figures ANCHOR on the largest-contract row
of each job group and sibling rows get a muted 'see MFD192'. SUM() ignores text,
so the totals row counts each job exactly once. Every metric that needs COSTS is
therefore anchor-only, and blanks itself on the siblings rather than lying.

AUDIT TRAIL (the owner, 2026-08-25: "everything must be logged so that when
someone asks who did this we can trace back")
Every change to an MFD-owned value is appended to an immutable JSONL log in
~/Library/Logs/Proficient/mfd-wip/ with the old value, the new value, the run
timestamp and the workbook's last-modified time. HONEST LIMIT: Excel does not
record which PERSON typed a cell, so no script can read that off the file. The
log answers what changed, when, and on what basis; pair the workbook mtime with
SharePoint version history to put a name to it.

USAGE
    python3 wip/mfd_wip_test.py              refresh QBO + metrics
    python3 wip/mfd_wip_test.py --dry-run    show what would change
    python3 wip/mfd_wip_test.py --no-qbo     layout only, no QBO pull
    python3 wip/mfd_wip_test.py --history    print the change log and exit
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import mfd_wip_cols as C
from shared import paths, xlsx_verify
from wip_excel_guard import assert_write_allowed, open_wip_workbook_for_write

WIP_EXCEL_PATH = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx",
)

TARGET_TAB = "WIP - MFD"
RETIRED_TAB = "Test - MFD"       # the staging copy this replaced; deleted on sight

HDR_ROW = 6
FIRST_DATA_ROW = 7
BANNER_ROW = 5                   # the sync stamp merges across the QBO block

LOG_DIR = Path.home() / "Library/Logs/Proficient/mfd-wip"
CHANGE_LOG = LOG_DIR / "value-changes.jsonl"
SNAPSHOT = LOG_DIR / "last-snapshot.json"

# ── styles: lifted verbatim from the tab so it keeps looking like itself ──
FONT_NAME = "Calibri"
CURRENCY = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
PCT = "0.00%"
DATEFMT = "mm-dd-yy"

_THIN = Side(style="thin", color="000000")
_MED = Side(style="medium", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

INPUT_FILL = PatternFill("solid", fgColor="FFF2F2F2")
INPUT_FONT = Font(name=FONT_NAME, size=12, bold=True, color="FFFA7D00")
CALC_FONT = Font(name=FONT_NAME, size=11)
CARRY_FONT = Font(name=FONT_NAME, size=11)
HDR_FONT = Font(name=FONT_NAME, size=11, bold=True)
HDR_INPUT_FILL = PatternFill("solid", fgColor="FFF2F2F2")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

QBO_HDR_FILL = PatternFill("solid", fgColor="FF2CA01C")
QBO_HDR_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")
QBO_CELL_FILL = PatternFill("solid", fgColor="FFEBF4E8")
QBO_CELL_FONT = Font(name=FONT_NAME, size=11, color="FF375623")
BANNER_FILL = PatternFill("solid", fgColor="FFD9EAD3")
BANNER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FF1F4E20")

MUTED_FILL = PatternFill("solid", fgColor="FFF1EFE8")
MUTED_FONT = Font(name=FONT_NAME, size=10, italic=True, color="FF7F7F7F")

TOTAL_FONT = Font(name=FONT_NAME, size=11, bold=True)

QBO_NOTE = ("From QuickBooks Online (project P&L, all time).\n"
            "The sync overwrites this cell - do not type here.")

RETAINAGE_NOTE = (
    "From QuickBooks Online: the balance of the 'Retainage Receivable' account "
    "for this job.\nThe sync overwrites this cell - do not type here.\n\n"
    "This is EXPECTED to differ from the Total Retainage column. QBO stops "
    "counting retainage once it has been billed to the GC; this report keeps "
    "carrying it. A gap means retainage was invoiced and has not been taken off "
    "the WIP, or the reverse.")

REV_ETC_NOTE = ("Starts equal to ETC. Type the new budget over it when an "
                "approved change order moves the cost.")

EARNED_NOTE = (
    "Cost-to-cost earned revenue: revised contract x (QBO costs / revised ETC).\n"
    "This is the CPA and bank method, and the only basis on which BILLED AHEAD / "
    "BILLED BEHIND mean anything.\nBlank on a row with no QBO costs of its own.")

OVER_NOTE = ("Billed MORE than the work earned. You are holding the GC's money - "
             "good for cash, but it is borrowed against work still to do.")
UNDER_NOTE = ("Billed LESS than the work earned. You are financing this job out "
              "of pocket - revenue earned that has not been invoiced.")

_JOB_RE = re.compile(r"^\s*(\d{2,4})\b")

# Old header text -> spec key, so a reorder reads the tab by NAME and never by
# position. Anything not listed is a calculated column and gets rebuilt.
_HEADER_ALIASES = {
    "PROJECT": "project",
    "MOBE DATE": "mobe",
    "COMPLETION DATE SOG/PAVING": "completion",
    "CUSTOMER": "customer",
    "CONTRACT": "contract",
    "CHANGE ORDERS": "co",
    "ETC": "etc",
    "REVISED ETC": "rev_etc",
    "COMPLETED TO DATE": "completed",
    "EARNED LESS RET.": "earned_less",
    "EARNED LESS RET": "earned_less",
    "TOTAL RETAINAGE": "retainage",
    "COSTS TO DATE": "qbo_costs",
    "BILLED TO DATE": "qbo_billed",
    "RETAINAGE (QBO)": "qbo_retain",
}


def _norm(text) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip().upper()


def job_of(label) -> Optional[str]:
    """'192 - JPI MAYHILL - 008' -> 'MFD192'. None when the row has no job #."""
    m = _JOB_RE.match(str(label or ""))
    return f"MFD{m.group(1)}" if m else None


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ─────────────────────────── sheet geometry ───────────────────────────

def header_map(ws) -> Dict[str, int]:
    """{spec key -> column index} as the sheet CURRENTLY stands, matched by
    header TEXT. This is what makes a reorder safe: the tab is read by name and
    written by position, so the two never have to agree beforehand."""
    found: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _HEADER_ALIASES.get(_norm(ws.cell(row=HDR_ROW, column=col).value))
        if key and key not in found:
            found[key] = col
    return found


def data_rows(ws, project_col: int) -> List[int]:
    out = []
    r = FIRST_DATA_ROW
    while r <= ws.max_row:
        label = ws.cell(row=r, column=project_col).value
        if not str(label or "").strip():
            break
        if job_of(label):
            out.append(r)
        r += 1
    return out


def totals_rows(ws):
    """(label_row, sum_row) of the TOTALS block, or (None, None)."""
    for r in range(FIRST_DATA_ROW, min(ws.max_row, 60) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            if _norm(ws.cell(row=r, column=c).value) == "TOTALS:":
                return r, r + 1
    return None, None


def group_rows(ws, rows: List[int], project_col: int) -> Dict[str, List[int]]:
    groups: Dict[str, List[int]] = {}
    for r in rows:
        job = job_of(ws.cell(row=r, column=project_col).value)
        if job:
            groups.setdefault(job, []).append(r)
    return groups


def anchor_row(ws, rows: List[int], cols: Dict[str, int]) -> int:
    """The row a job's QBO figures land on: the biggest revised contract.
    For MFD192 that is the base 008 row, which is what the owner asked for."""
    def size(r):
        return (_num(ws.cell(row=r, column=cols["contract"]).value)
                + _num(ws.cell(row=r, column=cols["co"]).value))
    return max(rows, key=size)


# ─────────────────────────── read what MFD owns ───────────────────────────

def read_owned(ws, rows: List[int], cols: Dict[str, int]) -> Dict[int, Dict[str, object]]:
    """Every MFD-owned value on the sheet as it stands, keyed by row then spec
    key. Read BEFORE the rewrite; written back into the new positions after.
    Calculated and QBO columns are deliberately not read - they are rebuilt."""
    out: Dict[int, Dict[str, object]] = {}
    for r in rows:
        vals: Dict[str, object] = {}
        for key in C.OWNED_BY_MFD:
            col = cols.get(key)
            if not col:
                continue
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.startswith("="):
                continue                      # a formula is not a typed value
            if v is not None:
                vals[key] = v
        out[r] = vals
    return out


# ─────────────────────────── audit trail ───────────────────────────

def _snapshot_key(vals: Dict[str, object]) -> str:
    return str(vals.get("project", "")).strip()


def audit(owned: Dict[int, Dict[str, object]], stamp: str, dry: bool) -> List[dict]:
    """Diff MFD-owned values against the previous run and append every change to
    an immutable JSONL log. Answers "what changed, when, on what basis" - see the
    honest limit on WHO in the module docstring."""
    first_run = not SNAPSHOT.exists()
    prior = {}
    if not first_run:
        try:
            prior = json.loads(SNAPSHOT.read_text())
        except (ValueError, OSError):
            prior = {}

    current = {_snapshot_key(v): {k: _jsonable(x) for k, x in v.items()}
               for v in owned.values() if _snapshot_key(v)}

    try:
        mtime = dt.datetime.fromtimestamp(
            WIP_EXCEL_PATH.stat().st_mtime).strftime("%m/%d/%Y %-I:%M %p")
    except OSError:
        mtime = None

    # With no prior snapshot there is nothing to diff against - every value would
    # read as "new", which is noise, not history. Record one baseline entry so the
    # log says where the trail starts instead of inventing 50 changes.
    if first_run:
        if not dry:
            LOG_DIR.mkdir(parents=True, exist_ok=True)
            with CHANGE_LOG.open("a", encoding="utf-8") as fh:
                fh.write(json.dumps({
                    "at": stamp, "workbook_modified": mtime, "row": None,
                    "column": None, "key": None, "old": None,
                    "new": f"{len(current)} job(s)", "kind": "baseline",
                }) + "\n")
            SNAPSHOT.write_text(json.dumps(current, indent=1, sort_keys=True))
        return []

    changes: List[dict] = []
    for job, vals in current.items():
        was = prior.get(job, {})
        for key, new in vals.items():
            old = was.get(key)
            if old == new:
                continue
            changes.append({
                "at": stamp, "workbook_modified": mtime, "row": job,
                "column": C.header(key), "key": key,
                "old": old, "new": new,
                "kind": "added" if job not in prior else "changed",
            })
        for key in set(was) - set(vals):
            changes.append({
                "at": stamp, "workbook_modified": mtime, "row": job,
                "column": C.header(key), "key": key,
                "old": was[key], "new": None, "kind": "cleared",
            })
    for job in set(prior) - set(current):
        changes.append({"at": stamp, "workbook_modified": mtime, "row": job,
                        "column": None, "key": None, "old": None, "new": None,
                        "kind": "row removed"})

    if changes and not dry:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        with CHANGE_LOG.open("a", encoding="utf-8") as fh:
            for rec in changes:
                fh.write(json.dumps(rec) + "\n")
    if not dry:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        SNAPSHOT.write_text(json.dumps(current, indent=1, sort_keys=True))
    return changes


def _jsonable(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        return round(v, 2)
    return v


def print_history(limit: int = 40) -> int:
    if not CHANGE_LOG.exists():
        print(f"\n  no change log yet at {CHANGE_LOG}\n")
        return 0
    recs = [json.loads(l) for l in CHANGE_LOG.read_text().splitlines() if l.strip()]
    print(f"\n  {len(recs)} logged change(s) - {CHANGE_LOG}")
    print("  Excel does not record WHO typed a cell; pair 'workbook modified' "
          "with SharePoint version history for a name.\n")
    for rec in recs[-limit:]:
        old, new = rec.get("old"), rec.get("new")
        fmt = lambda v: ("-" if v is None else
                         f"{v:,.2f}" if isinstance(v, (int, float)) else str(v))
        print(f"  {rec['at']:>22}  {str(rec['row'])[:26]:26} "
              f"{str(rec.get('column'))[:22]:22} {fmt(old):>16} -> {fmt(new):>16}"
              f"   (workbook saved {rec.get('workbook_modified')})")
    print()
    return 0


# ─────────────────────────── write the layout ───────────────────────────

def _style(cell, font, fill=None, fmt=None, align=None, border=True):
    cell.font = font
    if fill is not None:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = BORDER


def _fmt_for(key: str) -> Optional[str]:
    if key in C.PCT_KEYS:
        return PCT
    if key in C.DATE_KEYS:
        return DATEFMT
    if key in C.TEXT_KEYS:
        return "General"
    return CURRENCY


def clear_block(ws, rows: List[int], last_row: int) -> None:
    """Wipe the whole report block so a reorder cannot leave a stale column
    behind. Rows 1-4 (the title block) are never touched; row 5 is cleared only
    across the data columns so the green 'WIP REPORT' label at B5 survives."""
    wide = max(ws.max_column, C.last_index())
    # Unmerge EVERYTHING first. A merged cell's value is read-only, so any merge
    # left standing makes the clear below raise; write_layout re-creates the two
    # merges the sheet actually needs (the group banners and the TOTALS: label).
    for merged in [str(m) for m in ws.merged_cells.ranges]:
        ws.unmerge_cells(merged)
    for r in range(BANNER_ROW, last_row + 1):
        for c in range(C.FIRST_COL, wide + 1):
            if r == BANNER_ROW and c == C.FIRST_COL:
                continue                          # B5 = the sheet's own title
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name=FONT_NAME, size=11)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"
            cell.comment = None


def write_layout(ws, rows: List[int], owned: Dict[int, Dict[str, object]],
                 label_row: int, sum_row: int) -> None:
    """Headers, group banners, MFD's carried values, and every formula - all
    positioned from the spec. The ONLY place the sheet's shape is decided."""
    for key in C.KEYS:
        ws.column_dimensions[C.letter(key)].width = C.BY_KEY[key][1]

    # group banners: the QBO block carries the sync stamp, the others a label
    for label, first_key in C.group_starts().items():
        keys = C.group_keys(label)
        span = f"{C.letter(keys[0])}{BANNER_ROW}:{C.letter(keys[-1])}{BANNER_ROW}"
        ws.merge_cells(span)
        cell = ws.cell(row=BANNER_ROW, column=C.index(keys[0]))
        is_qbo = C.kind(first_key) == "qbo"
        cell.value = label
        _style(cell, BANNER_FONT if is_qbo else Font(name=FONT_NAME, size=10, bold=True,
                                                     color="FF7F7F7F"),
               BANNER_FILL if is_qbo else PatternFill("solid", fgColor="FFF7F7F7"),
               align=Alignment(horizontal="center", vertical="center"))

    for key in C.KEYS:
        cell = ws.cell(row=HDR_ROW, column=C.index(key), value=C.header(key))
        k = C.kind(key)
        _style(cell,
               QBO_HDR_FONT if k == "qbo" else HDR_FONT,
               QBO_HDR_FILL if k == "qbo" else (HDR_INPUT_FILL if k == "input" else None),
               align=HDR_ALIGN)

    notes = {"rev_etc": REV_ETC_NOTE, "earned_rev": EARNED_NOTE,
             "over": OVER_NOTE, "under": UNDER_NOTE}

    for r in rows:
        vals = owned.get(r, {})
        for key in C.KEYS:
            cell = ws.cell(row=r, column=C.index(key))
            k = C.kind(key)
            if k in ("carry", "input"):
                cell.value = vals.get(key)
                if key == "rev_etc" and cell.value in (None, ""):
                    cell.value = f"={C.letter('etc')}{r}"
                _style(cell,
                       INPUT_FONT if k == "input" else CARRY_FONT,
                       INPUT_FILL if k == "input" else None,
                       _fmt_for(key),
                       Alignment(horizontal="center", vertical="center")
                       if key in C.DATE_KEYS else None)
            elif k == "calc":
                cell.value = C.formula(key, r)
                _style(cell, CALC_FONT, None, _fmt_for(key),
                       Alignment(horizontal="center", vertical="center")
                       if key in C.PCT_KEYS else None)
            else:                                     # qbo - value comes later
                _style(cell, QBO_CELL_FONT, QBO_CELL_FILL, CURRENCY)
            if key in notes:
                cell.comment = Comment(notes[key], "WIP")

    _write_totals(ws, rows, label_row, sum_row)
    _write_production_balance(ws, sum_row)
    widen_print_area(ws, sum_row)


def _write_totals(ws, rows: List[int], label_row: int, sum_row: int) -> None:
    first, last = rows[0], rows[-1]
    tot_cell = ws.cell(row=label_row, column=C.index("customer"), value="TOTALS:")
    _style(tot_cell, HDR_FONT, align=Alignment(horizontal="center", vertical="center"))
    ws.merge_cells(f"{C.letter('customer')}{label_row}:{C.letter('customer')}{sum_row}")

    for key in C.KEYS:
        if C.kind(key) == "carry":
            continue
        _style(ws.cell(row=label_row, column=C.index(key), value=C.header(key)),
               HDR_FONT, align=HDR_ALIGN)
        f = C.total_formula(key, sum_row, first, last)
        if not f:
            continue
        _style(ws.cell(row=sum_row, column=C.index(key), value=f),
               TOTAL_FONT, None, _fmt_for(key),
               Alignment(horizontal="center", vertical="center"))


def _write_production_balance(ws, sum_row: int) -> None:
    """The sheet's own closing figure: balance to finish less retainage held.
    It used to sit in a hardcoded cell; positioned from the spec now so a
    reorder carries it along instead of stranding it under a new column."""
    row = sum_row + 2
    lab = ws.cell(row=row, column=C.index("ctc"),
                  value="PRODUCTION BALANCE LESS RETAINAGE")
    lab.font = Font(name=FONT_NAME, size=11, bold=True, italic=True)
    lab.alignment = Alignment(horizontal="right", vertical="center")
    val = ws.cell(row=row + 1, column=C.index("ctc"),
                  value=f"={C.letter('balance')}{sum_row}-{C.letter('retainage')}{sum_row}")
    val.font = Font(name=FONT_NAME, size=11, bold=True)
    val.number_format = CURRENCY
    val.alignment = Alignment(horizontal="right", vertical="center")
    val.border = Border(top=_THIN, bottom=_MED)


def widen_print_area(ws, sum_row: int) -> None:
    """Cover every column the report owns. The tab's original area stopped at
    column L, predating both Total Retainage and everything added here."""
    ws.print_area = f"$B$2:${get_column_letter(C.last_index())}${sum_row}"


# ─────────────────────────── QBO ───────────────────────────

def _retainage_account_id(access: str, company_id: str, qbo_api) -> Optional[str]:
    """The 'Retainage Receivable' account. Matched on AccountSubType first
    (QBO has a dedicated 'Retainage' subtype) and only then on the name, so a
    rename in the chart of accounts does not silently zero the column."""
    try:
        accounts = qbo_api.query_all(access, company_id, "Account")
    except Exception:
        return None
    for a in accounts:
        if (a.get("AccountSubType") or "").lower() == "retainage":
            return a["Id"]
    for a in accounts:
        if (a.get("Name") or "").strip().lower() == "retainage receivable":
            return a["Id"]
    return None


def fetch_retainage(access: str, company_id: str, qbo_api) -> Dict[str, float]:
    """{job -> retainage still held} = the per-project balance of the
    'Retainage Receivable' account, walked out of the GeneralLedger report.

    The filter key is 'account', SINGULAR. 'accounts' is accepted and then
    silently ignored, which hands back the entire general ledger truncated to
    its first few accounts - looking like a clean empty result. Never raises."""
    acct = _retainage_account_id(access, company_id, qbo_api)
    if not acct:
        print("  no 'Retainage Receivable' account in QBO - retainage skipped")
        return {}
    try:
        rep = qbo_api.report(access, company_id, "GeneralLedger", {
            "start_date": "2019-01-01",
            "end_date": dt.date.today().isoformat(),
            "accounting_method": "Accrual",
            "account": acct,
            "columns": "tx_date,txn_type,doc_num,name,subt_nat_amount,rbal_nat_amount",
        })
    except Exception as e:
        print(f"  retainage GL failed ({type(e).__name__}) - retainage skipped")
        return {}

    lines: List[list] = []

    def walk(node):
        if node.get("Rows"):
            for child in node["Rows"].get("Row", []):
                walk(child)
        elif node.get("ColData"):
            lines.append([c.get("value") for c in node["ColData"]])

    for node in (rep.get("Rows") or {}).get("Row", []):
        walk(node)

    out: Dict[str, float] = {}
    for cols in lines:
        cols = (cols + [""] * 6)[:6]
        job = qbo_api.extract_proj(cols[3])
        if not job:
            continue
        try:
            out[job] = out.get(job, 0.0) + float(cols[4])
        except (TypeError, ValueError):
            continue
    return out


def fetch_qbo(jobs: List[str]) -> Dict[str, dict]:
    """{job -> {'costs': float, 'billed': float, 'retainage': float}}.
    Costs = COGS + Expenses (same basis as cp_wip_reader); billed = Total
    Income, i.e. GROSS billed with retainage included; retainage = the
    Retainage Receivable balance. Never raises."""
    from shared import qbo_api

    try:
        access, company_id = qbo_api.load_credentials()
    except Exception as e:
        print(f"  QBO auth failed ({type(e).__name__}) - leaving QBO columns as they are")
        return {}
    try:
        proj_map = qbo_api.build_project_customer_map(access, company_id)
    except Exception as e:
        print(f"  QBO customer map failed ({type(e).__name__})")
        return {}

    start, end = "2019-01-01", dt.date.today().isoformat()
    retainage = fetch_retainage(access, company_id, qbo_api)
    out: Dict[str, dict] = {}
    for job in jobs:
        cust = proj_map.get(job)
        if not cust:
            print(f"  {job}: not found in QBO - skipped")
            continue
        try:
            totals = qbo_api.extract_pl_totals(
                qbo_api.fetch_project_pl(access, company_id, cust["id"], start, end))
        except Exception as e:
            print(f"  {job}: P&L failed ({type(e).__name__}) - skipped")
            continue
        out[job] = {
            "costs": _num(totals.get("cogs")) + _num(totals.get("expenses")),
            "billed": _num(totals.get("income")),
            "retainage": retainage.get(job),
        }
        ret = out[job]["retainage"]
        ret_txt = f"{ret:>13,.2f}" if ret is not None else "          n/a"
        print(f"  {job}: costs {out[job]['costs']:>14,.2f}   "
              f"billed {out[job]['billed']:>14,.2f}   retainage {ret_txt}")
    return out


def _retainage_variance(ws, job_rows: List[int], qbo_amount: float) -> str:
    """The one line of the comment that does the comparing: QBO's balance
    against the sum of this job's own 'Total Retainage' cells (col M). Summed
    across the group because a job like MFD192 spreads its retainage over three
    contract rows while QBO holds one balance."""
    theirs = sum(_num(ws.cell(row=r, column=COL_TAB_RETAINAGE).value)
                 for r in job_rows)
    gap = qbo_amount - theirs
    if abs(gap) < 0.01:
        return f"This report says {theirs:,.2f} - they agree."
    return (f"This report says {theirs:,.2f}.\n"
            f"Difference: {gap:+,.2f} (QBO minus this report).")


def write_qbo(ws, rows: List[int], cols: Dict[str, int], data: Dict[str, dict],
              stamp: str) -> int:
    """Write each job's QBO figures onto its anchor row and stamp the banner.
    Touches ONLY the spec's qbo columns."""
    groups = group_rows(ws, rows, C.index("project"))
    written = 0
    for job, job_rows in groups.items():
        anchor = anchor_row(ws, job_rows, {k: C.index(k) for k in C.KEYS})
        for r in job_rows:
            if r == anchor:
                continue
            for key in C.QBO_KEYS:
                cell = ws.cell(row=r, column=C.index(key))
                cell.value = f"see {job}"
                cell.comment = None
                _style(cell, MUTED_FONT, MUTED_FILL, "General",
                       Alignment(horizontal="center", vertical="center"))
        vals = data.get(job)
        if not vals:
            continue
        for key, src in (("qbo_costs", "costs"), ("qbo_billed", "billed"),
                         ("qbo_retain", "retainage")):
            cell = ws.cell(row=anchor, column=C.index(key))
            amount = vals.get(src)
            _style(cell, QBO_CELL_FONT, QBO_CELL_FILL, CURRENCY)
            cell.value = None if amount is None else round(amount, 2)
            note = QBO_NOTE
            if key == "qbo_retain":
                note = RETAINAGE_NOTE
                if amount is not None:
                    note += "\n\n" + _retainage_variance(ws, job_rows, amount)
            cell.comment = Comment(note, "WIP")
        written += 1

    banner_keys = C.group_keys("FROM QBO")
    banner = ws.cell(row=BANNER_ROW, column=C.index(banner_keys[0]))
    banner.value = f"QBO - LAST SYNC {stamp}"
    _style(banner, BANNER_FONT, BANNER_FILL,
           align=Alignment(horizontal="center", vertical="center"))
    return written


def _retainage_variance(ws, job_rows: List[int], qbo_amount: float) -> str:
    """QBO's balance against the sum of this job's own Total Retainage cells.
    Summed across the group because MFD192 spreads retainage over three contract
    rows while QBO holds one balance."""
    theirs = sum(_num(ws.cell(row=r, column=C.index("retainage")).value)
                 for r in job_rows)
    gap = qbo_amount - theirs
    if abs(gap) < 0.01:
        return f"This report says {theirs:,.2f} - they agree."
    return (f"This report says {theirs:,.2f}.\n"
            f"Difference: {gap:+,.2f} (QBO minus this report).")


# ─────────────────────────── main ───────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--no-qbo", action="store_true", help="Skip the QBO pull.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Report what would change; write nothing.")
    ap.add_argument("--history", action="store_true",
                    help="Print the logged value changes and exit.")
    args = ap.parse_args()

    if args.history:
        return print_history()

    print(f"\n  MFD WIP -> '{TARGET_TAB}'")
    print(f"  workbook: {WIP_EXCEL_PATH}")
    if not WIP_EXCEL_PATH.exists():
        print("  workbook not found (OneDrive synced?)")
        return 1

    wb = open_wip_workbook_for_write(WIP_EXCEL_PATH)
    if TARGET_TAB not in wb.sheetnames:
        print(f"  tab {TARGET_TAB!r} missing")
        return 1
    ws = wb[TARGET_TAB]
    assert_write_allowed(ws.title)

    if RETIRED_TAB in wb.sheetnames and not args.dry_run:
        del wb[RETIRED_TAB]
        print(f"  removed the retired '{RETIRED_TAB}' tab")

    # Read the sheet as it stands, BY HEADER NAME - this is what lets the column
    # order change without the old and new layouts having to agree.
    cols = header_map(ws)
    missing = [k for k in C.OWNED_BY_MFD if k not in cols]
    if "project" not in cols:
        print("  cannot find the PROJECT column - header row changed?")
        return 1
    if missing:
        print(f"  note: no existing column for {', '.join(missing)} - will start blank")

    rows = data_rows(ws, cols["project"])
    if not rows:
        print("  no contract rows found")
        return 1
    owned = read_owned(ws, rows, cols)
    label_row, sum_row = totals_rows(ws)
    if not label_row:
        label_row, sum_row = rows[-1] + 2, rows[-1] + 3

    groups = group_rows(ws, rows, cols["project"])
    print(f"  {len(rows)} contract row(s), {len(groups)} QBO job(s)")
    for job, job_rows in sorted(groups.items()):
        anchor = anchor_row(ws, job_rows, cols)
        label = str(ws.cell(row=anchor, column=cols["project"]).value or "").strip()
        extra = (f"  (+{len(job_rows) - 1} sibling row(s) -> 'see {job}')"
                 if len(job_rows) > 1 else "")
        print(f"    {job}: anchor row {anchor} - {label}{extra}")

    order = [C.letter(k) for k in C.KEYS]
    print(f"  layout: {len(C.COLS)} columns {order[0]}..{order[-1]}  "
          f"| MFD enters {C.letter(C.INPUT_KEYS[0])}..{C.letter(C.INPUT_KEYS[-1])}"
          f" | QBO {C.letter(C.QBO_KEYS[0])}..{C.letter(C.QBO_KEYS[-1])}"
          f" | metrics {C.letter(C.CALC_KEYS[0])}..{C.letter(C.CALC_KEYS[-1])}")

    stamp = dt.datetime.now().strftime("%m/%d/%Y %-I:%M %p")
    changes = audit(owned, stamp, args.dry_run)
    if changes:
        print(f"  {len(changes)} value change(s) since the last run "
              f"-> {CHANGE_LOG.name}")
        for rec in changes[:8]:
            print(f"    {rec['row'][:28]:28} {str(rec['column'])[:20]:20} "
                  f"{rec['old']} -> {rec['new']}")
    else:
        print("  no MFD-entered value changed since the last run")

    data = {} if args.no_qbo else fetch_qbo(sorted(groups))

    if args.dry_run:
        print("  dry run - nothing written\n")
        return 0

    clear_block(ws, rows, sum_row + 3)
    write_layout(ws, rows, owned, label_row, sum_row)
    n = write_qbo(ws, rows, cols, data, stamp)
    print(f"  QBO figures written to {n} anchor row(s); banner stamped {stamp}")

    wb.save(WIP_EXCEL_PATH)
    xlsx_verify.assert_clean(WIP_EXCEL_PATH)
    print(f"  saved and verified clean: {WIP_EXCEL_PATH.name}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
