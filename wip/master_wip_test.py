#!/usr/bin/env python3
"""
master_wip_test.py — the UNIFIED WIP test sheet (the user 2026-07-15).

One run, one 'Test' tab in the WIP master (guard-allowed), all divisions
stacked in the user's order:
  main table:  MFD (from the 'WIP Master' tab: contract E / ETC F) then CP
               (live folder scan, draws model — ACTIVE projects only)
  sections:    RP SLABS — CUSTOM · RP SLABS — TRACT · FTW — ACTIVE ·
               FTW BACKLOG   (all from the General List via rp_wip_reader)
Billed + Costs always fresh from QBO per line. Read-only everywhere except
the 'Test' tab write.

Usage:  python3 master_wip_test.py [--dry-run]
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

import cp_wip_reader as CP        # CP READER: folder scan / draws
import wip_writer as W            # the shared report ENGINE
import rp_wip_reader as RP

MASTER_SHEET = "WIP Master"
# 'WIP Master' tab layout: header row 3, data row 4+.
MCOL_PROJ, MCOL_NAME, MCOL_CONTRACT, MCOL_ETC = 1, 2, 5, 6


_MFD_ETC_FORMULA = re.compile(r"^=\(?\s*E(\d+)\s*/\s*([0-9.]+)\s*\)?$", re.I)


def read_mfd_from_master(wip_path: Path):
    """MFD#### rows from the 'WIP Master' tab — the master sheet IS the MFD
    source for active jobs and pricing (the user 2026-07-15). Billed/Costs
    come from QBO, not from the tab.

    The ETC cell is a FORMULA on that sheet (`=(E4/1.17)` — contract ÷ markup),
    and openpyxl drops every cached formula result workbook-wide each time it
    saves. So after our first write the cached value is gone and a plain
    data_only read returns None — which silently blanked MFD's entire budget
    (caught by the change audit, 2026-08-03). Read the cached value, and when
    it is missing evaluate the sheet's own divisor formula instead."""
    wb = load_workbook(wip_path, data_only=True, read_only=True)
    fwb = load_workbook(wip_path, data_only=False, read_only=True)
    ws, fws = wb[MASTER_SHEET], fwb[MASTER_SHEET]
    rows = []
    for r in range(4, ws.max_row + 1):
        proj = ws.cell(r, MCOL_PROJ).value
        if not proj or not str(proj).strip().upper().startswith("MFD"):
            continue
        proj = str(proj).strip().upper()
        contract = RP._money(ws.cell(r, MCOL_CONTRACT).value)
        etc = RP._money(ws.cell(r, MCOL_ETC).value)
        if etc is None and contract:
            m = _MFD_ETC_FORMULA.match(
                str(fws.cell(r, MCOL_ETC).value or "").replace(" ", ""))
            if m and float(m.group(2)):
                etc = contract / float(m.group(2))
                print(f"    · {proj}: ETC recomputed from "
                      f"'{MASTER_SHEET}'!F{r} ({fws.cell(r, MCOL_ETC).value}) "
                      f"— Excel's cached value was stripped")
        row = W.CpRow(proj, str(ws.cell(r, MCOL_NAME).value or proj), False,
                       contract, None, etc, None, None)
        row.client = "Multi Family"
        row.takeoff_path = wip_path          # contract/ETC link → the master tab
        row.src_link = str(wip_path)
        row.src_fragment = W._sheet_fragment(MASTER_SHEET, f"A{r}")
        row.notes.append(f"Contract/ETC from '{MASTER_SHEET}' row {r}")
        rows.append(row)
    wb.close()
    fwb.close()
    return rows


def master_cols():
    """Column set for the bank-facing Test-Master report.

    Test-Master is the FINISHED WIP the owner sends to banks (the user
    2026-08-06), so it is the LEAN view:
    - The first column is the division, labelled **TYPE** (the user: "why not
      call it Section Type") — Multi-Family / Commercial / Residential — Slab /
      Residential — Flatwork.
    - The old Tract/Custom TYPE column is DROPPED (mostly empty, internal).
    - The change-order breakout is DROPPED — one **TOTAL CONTRACT PRICE** and
      one **ESTIMATED TOTAL COSTS**, not the ORIGINAL/CO/REVISED trios (the
      user: "just make it on Contract Price and One ETC").
    - No RETAINAGE / NOTES / LAST SYNCED (those live on the working tabs, which
      money_bleeds reads).
    project-pnl reads this tab for Contract/ETC/STATUS — those names are kept.
    """
    drop = {"retainage_held", "notes_text", "_last_synced", "_notes_all",
            "home_type",              # Tract/Custom — dropped for the bank report
            "base_contract", "co_revenue",       # collapse to TOTAL CONTRACT PRICE
            "base_etc", "co_cost_estimate"}      # collapse to ESTIMATED TOTAL COSTS
    cols = [("TYPE", 32, "section")]   # division + home type (Residential —
    #                                    Custom — Slab); wide enough for the
    #                                    longest folded label (the user 2026-08-07)
    for label, width, field in W.COLS:
        if field in drop:
            continue
        if field == "_active_status":
            # The bank report carries only active WIP (closed jobs are excluded
            # outright), so the STATUS column is repurposed as BONDED — "N" on
            # every job (none are bonded) — the user 2026-08-06.
            cols.append(("BONDED", 10, "_bonded"))
            continue
        cols.append((label, width, field))
    return cols


# Sections that DON'T belong on the bank-facing Test-Master (the user
# 2026-08-06): only clean active WIP goes to banks — no off-schedule, no
# backlog, no dropped/unbilled. These are the raw section keys (pre-label).
_BANK_EXCLUDE = {"FTW — OFF-SCHEDULE (COSTS)", "FTW BACKLOG",
                 "RP — DROPPED, UNBILLED"}

# Specific jobs kept OFF the bank report — billed out / done (the user
# 2026-08-06: "drop rp6901, already billed out"; 2026-08-07: also RP6586 and
# CP585). They still appear on the working 'Test - RP' / 'Test - CP' tabs.
_BANK_EXCLUDE_JOBS = {"RP6901", "RP6586", "CP585"}


def existing_project_nums(wip_path: Path, tab_name: str = "Test-Master"):
    """PROJECT #s already on the Test-Master tab (read-only). Used to lock
    the RP refresh to jobs that are ALREADY in the report (the user
    2026-07-22: 'RP just the projects already in the WIP report … don't
    add')."""
    from openpyxl import load_workbook
    if not wip_path.exists():
        return set()
    wb = load_workbook(wip_path, data_only=True, read_only=True)
    if tab_name not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[tab_name]
    hdr = next((r for r in range(1, 8)
                for c in range(1, (ws.max_column or 0) + 1)
                if ws.cell(r, c).value == "PROJECT #"), None)
    nums = set()
    if hdr:
        pcol = next(c for c in range(1, ws.max_column + 1)
                    if ws.cell(hdr, c).value == "PROJECT #")
        for r in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(r, pcol).value
            if v and str(v).strip().upper() != "PROJECT #":
                nums.add(str(v).strip().upper())
    wb.close()
    return nums


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--rp-existing-only", action="store_true",
                    help="RP: refresh only the project #s already in "
                         "Test-Master — never add new RP jobs (the user "
                         "2026-07-22)")
    ap.add_argument("--rp-from-file", metavar="XLSX",
                    help="RP: take the RP lines from the owner's verified "
                         "'RP WIP' workbook instead of the General List "
                         "pipeline (the user 2026-07-29); CP lines in it are "
                         "excluded, billed/costs still refresh from QBO")
    args = ap.parse_args()

    print("\n  UNIFIED WIP — MFD → CP → RP → 'Test' tab")
    print("  " + "─" * 60)

    # ── MFD (master sheet) ──
    mfd_rows = read_mfd_from_master(W.WIP_EXCEL_PATH)
    print(f"  MFD from '{MASTER_SHEET}': {len(mfd_rows)} job(s)")

    # ── CP (live folder scan, ACTIVE only — the unified sheet is a WIP) ──
    if not CP.CP_ACTIVE_DIR.exists():
        # Tripwire (2026-07-29): an unmounted Synology once produced a WIP
        # with ZERO CP lines — a gutted report is worse than no report.
        print(f"  ✗ CP root unreachable: {CP.CP_ACTIVE_DIR}\n"
              f"    (Synology unmounted?) Refusing to write a WIP report "
              f"without its CP section — reconnect the volume and re-run.")
        return 4
    cp_rows = CP.scan_cp_folders(CP.CP_ACTIVE_DIR, is_completed=False)
    print(f"  CP active folders: {len(cp_rows)} project(s)")

    # ── RP: the owner's verified workbook (--rp-from-file) OR the
    #        General List pipeline (legacy) ──
    if args.rp_from_file:
        rp_file = Path(args.rp_from_file).expanduser()
        if not rp_file.exists():
            print(f"  ✗ RP file not found: {rp_file}")
            return 3
        rp_rows = RP.read_rp_from_file(rp_file)
        pairs = [(row,) for row in rp_rows]     # RP.enrich_with_qbo unpacks row-first
        print(f"  RP lines from {rp_file.name!r}: {len(rp_rows)}")
    else:
        records, _missing = RP.read_general_list(RP.ALPHA_PATH)
        rp_to_folders, addr_folders = ({}, [])
        if RP.RP_ROOT.exists():
            rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)
        pairs = RP.build_lines(records, rp_to_folders, addr_folders)
        if args.rp_existing_only:
            keep = existing_project_nums(W.WIP_EXCEL_PATH)
            before = len(pairs)
            pairs = [t for t in pairs if t[0].project_num in keep]
            print(f"  RP locked to existing Test-Master lines: {len(pairs)} kept "
                  f"of {before} (no new RP jobs added)")
        else:
            print(f"  RP lines from the General List: {len(pairs)}")

    # ── ONE QBO pass across everything ──
    print("  Enriching with QBO Billed/Costs (all divisions) …")
    CP.enrich_with_qbo(mfd_rows + cp_rows)
    RP.enrich_with_qbo(pairs)
    for row in mfd_rows + cp_rows:
        row.needs_review = bool(row.status_flags)
    for row in mfd_rows:
        row.section = "MFD"
    for row in cp_rows:
        row.section = "CP"

    if args.rp_from_file:
        # Sections came from the file's bands; the post-QBO pass lives in
        # rp_wip_reader so this script and a standalone RP run classify
        # identically.
        rp_sorted = RP.classify_from_file([t[0] for t in pairs])
        order = ["RP SLAB", "FTW — ACTIVE", "FTW — OFF-SCHEDULE (COSTS)",
                 "RP — DROPPED, UNBILLED", "FTW BACKLOG"]
        all_rows = mfd_rows + cp_rows + rp_sorted
        counts = " · ".join(
            f"{s} {sum(1 for r in rp_sorted if r.section == s)}"
            for s in order)
        print(f"  Sections: MFD {len(mfd_rows)} · CP {len(cp_rows)} · {counts}")
    else:
        for row, comp, _rec in pairs:
            RP._classify(row, comp)

        # ── RP partition: custom/tract slabs, FTW active, FTW backlog ──
        sched_addrs, sched_label = RP.read_schedule_flatwork(RP.SCHEDULE_DIR)
        slabs_custom, slabs_tract, ftw_active, ftw_backlog = [], [], [], []
        for row, comp, rec in pairs:
            if row.project_num.endswith("-FTW"):
                addr_n = RP._norm(f"{rec['house'] or ''} {rec['street'] or ''}")
                on_sched = bool(addr_n) and any(
                    addr_n == a or a.startswith(addr_n) or addr_n.startswith(a)
                    for a in sched_addrs)
                if row.billed_to_date or row.costs_to_date or on_sched:
                    if on_sched:
                        row.notes.append(f"On the {sched_label} schedule — flatwork crew")
                    ftw_active.append(row)
                else:
                    row.needs_review = False
                    ftw_backlog.append(row)
            elif row.home_type == "Tract":
                slabs_tract.append(row)
            elif row.project_num.startswith("CP"):
                cp_rows.append(row)          # CP-standalone GL jobs (e.g. CP865)
                                             # belong with CP (the user 2026-07-15)
            else:
                slabs_custom.append(row)

        # RP justification + WHY row-jump links (same workbook as the RP tab uses)
        justify_path = Path.home() / "Downloads" / "RP WIP - Justification.xlsx"
        just_rows = RP.write_justification(pairs, ftw_backlog, justify_path) or {}
        for row, _comp, _rec in pairs:
            row.why_link = str(justify_path)
            jr = just_rows.get(row.project_num)
            row.why_fragment = W._sheet_fragment("JUSTIFICATION", f"A{jr}") if jr else None
            if row.needs_review:
                reason = "; ".join(row.status_flags) or (row.notes[-1] if row.notes else "")
                if reason and f"RED: {reason}" not in row.notes:
                    row.notes.append(f"RED: {reason}")

        print(f"  Sections: MFD {len(mfd_rows)} · CP {len(cp_rows)} · "
              f"RP custom {len(slabs_custom)} · tract {len(slabs_tract)} · "
              f"FTW {len(ftw_active)} · backlog {len(ftw_backlog)}")

        for rows_, sect in ((cp_rows, "CP"),      # re-stamp: partition may have
                                                  # appended CP-standalone GL rows
                            (slabs_custom, "RP SLAB — CUSTOM"),
                            (slabs_tract, "RP SLAB — TRACT"),
                            (ftw_active, "FTW — ACTIVE"),
                            (ftw_backlog, "FTW BACKLOG")):
            for row in rows_:
                row.section = sect
        all_rows = (mfd_rows + cp_rows + slabs_custom + slabs_tract
                    + ftw_active + ftw_backlog)
    # SECTION shows the division spelled out, not the internal code.
    # Bank-facing Test-Master shows only clean ACTIVE WIP — no off-schedule,
    # backlog, or dropped/unbilled sections, no CLOSED jobs, and no jobs the
    # owner has called done (RP6901) — the user 2026-08-06. Those rows stay on
    # the working 'Test - RP' tab (rp_sorted), just not on the bank report.
    bank_rows = [r for r in all_rows
                 if r.section not in _BANK_EXCLUDE
                 and not r.is_completed
                 and r.project_num.upper() not in _BANK_EXCLUDE_JOBS]
    for row in bank_rows:                         # division → spelled-out label
        label = RP.SECTION_LABEL.get(row.section, row.section)
        # Fold the home type into the residential TYPE, before the scope word
        # (the user 2026-08-07: "add custom/tract before slab to indicate what
        # type of home it is") → "Residential — Custom — Slab".
        ht = getattr(row, "home_type", None)
        if ht in ("Tract", "Custom") and label.startswith("Residential — "):
            label = label.replace("Residential — ", f"Residential — {ht} — ", 1)
        row.section = label

    import datetime as dt
    try:
        wrote = W.write_test_cp(
            bank_rows, W.WIP_EXCEL_PATH,
            dry_run=args.dry_run, tab_name="Test-Master",
            # No STATUS/active filter — closed jobs are excluded outright, so
            # every row on the bank report is active by construction.
            cols=master_cols(), default_filter_active=False,
            title="WIP REPORT",
            summary=True,
            qbo_links_only=True,
            # The change audit runs HERE — Test-Master carries all three
            # divisions, so one report covers MFD + CP + RP (the user
            # 2026-07-31: "I need to always audit these things").
            audit=True, audit_xlsx=W.AUDIT_XLSX,
            # Read-only roll-up — locked so it can't be typed into by accident.
            protect=True,
            # The finished bank report: clean, no colour/links/edit-tracking
            # (the user 2026-08-06).
            plain_report=True)
    except W.WipWriteDenied as e:
        print(f"  ✗ Guard blocked write: {e}")
        return 2
    except FileNotFoundError as e:
        print(f"  ✗ {e}")
        return 3
    total = len(bank_rows)
    if not args.dry_run and wrote:
        print(f"  ✓ Wrote {total} line(s) to 'Test-Master'")
        _drop_stale_test_tab()

    # 'Test - RP' is written by rp_wip_reader.write_rp_tab — the single
    # writer of that tab, so a standalone RP run and this unified run can
    # never produce different layouts (they did, on 2026-08-04).
    if args.rp_from_file:
        try:
            RP.write_rp_tab(rp_sorted, dry_run=args.dry_run)
        except W.WipWriteDenied as e:
            print(f"  ✗ Guard blocked the 'Test - RP' write: {e}")
            return 2
    return 0


def _drop_stale_test_tab() -> None:
    """Remove the superseded 'Test' tab (renamed to 'Test-Master' — the user
    2026-07-15). Only ever deletes 'Test'; guard-checked; live tabs untouched."""
    from openpyxl import load_workbook as _lw
    W.assert_write_allowed("Test")
    wb = _lw(W.WIP_EXCEL_PATH)
    if "Test" in wb.sheetnames:
        del wb["Test"]
        wb.save(W.WIP_EXCEL_PATH)
        print("  ✓ removed superseded 'Test' tab")
    wb.close()


if __name__ == "__main__":
    sys.exit(main())
