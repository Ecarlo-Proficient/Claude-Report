"""
wip_review_common.py - the shared diff/merge core for the ledger's WIP Review.

THE FLOW (the user 2026-08-25: "make the wip update give me the report to
accept/merge ... i can accept costs/billed to date and have pm answer on the
rest ... easy to see the before and afters"):

  1. EMIT  - each WIP tool (cp / rp / master) computes its rows exactly as it
             would before a real write, then instead of writing calls emit_review():
             snapshot the current tab, diff it against the fresh rows FIELD BY
             FIELD, and dump a review JSON. No tab is touched.
  2. REVIEW - the ledger reads those JSONs and shows every change as WAS -> NOW,
             split into the two blocks below. The owner approves/disapproves each.
  3. APPLY  - each tool recomputes, apply_decisions() reverts every DISAPPROVED
             field back to the current tab value (approved fields keep the fresh
             value), and the tool writes through its normal guarded writer.

THE TWO BLOCKS (this is the whole point of the split):
  * QBO   - costs / billed / retainage. Facts from QuickBooks; the owner ACCEPTS.
  * PM    - original contract / approved COs / original ETC / CO costs. Judgment
            the PM answers on.

THE DIRECTION RULES (2026-09-01 - "the WIP Review must stop proposing contract /
CO / ETC values that go backwards with no reason on screen"):
  * CARRY     - a PM field the reader could not source this run (None) while the
                tab holds a value KEEPS the tab value. carry_pm_fields() puts it
                back on the row itself, so the review, apply_decisions() and the
                guarded writers all read one number: a PM field never shows
                "cleared", and a later approved sync cannot blank it either.
  * REVERSED  - a PM field going DOWN (now < was) is flagged `reversed` and the
                job's status becomes REVERSED (sorted between CHANGED and ADDED).
                The UI leaves it unchecked - even under "Approve all shown" - and
                names the document the new value came from.
  * DECREASED - a QBO field going down stays an accepted fact (checked) but is
                marked `decreased` with a note: a bill was voided, deleted or
                re-coded, and the owner should see that.
  * SOURCE    - every field cell carries `source` (a few words naming the document:
                "Draw #4 G702 · <file>", "takeoff · <file>", "proposal PDF · <file>",
                "RP WIP file · 'RP WIP' row N", "QuickBooks · project P&L",
                "carried from tab") and `source_path`. Readers stamp them with
                set_source() the moment they read a document; row_source() falls
                back to the row's audit attributes; a PM change with no source
                says so in its `note`.

This module is the ONE place the field set, the tab-header lookup, the diff, the
carry and the revert live, so the CP / RP / Master tabs can never diff or merge
differently. It is pure (no QBO, no Excel writes of its own beyond reading a
workbook snapshot), so it is safe to import from any wip tool.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── the reviewable fields ────────────────────────────────────────────────────
# key         : stable id used in the JSON and the decisions file
# attr        : the CpRow attribute the value is read from / written back to
# block       : "qbo" (owner accepts) or "pm" (PM answers)
# working_hdr : column header on the working tabs (Test - CP / Test - RP)
# master_hdr  : column header on Test-Master (the lean bank view); None = absent
# For MFD, co_revenue is always None, so base_contract == TOTAL CONTRACT PRICE and
# base_etc == ESTIMATED TOTAL COSTS - which is why those two map cleanly to the
# master headers while the CO breakout columns do not exist there.
FIELDS = [
    dict(key="costs",         attr="costs_to_date",    block="qbo", label="Costs to date",
         working_hdr="COSTS TO DATE",         master_hdr="COSTS TO DATE"),
    dict(key="billed",        attr="billed_to_date",   block="qbo", label="Billed to date",
         working_hdr="BILLED TO DATE",        master_hdr="BILLED TO DATE"),
    dict(key="retainage",     attr="retainage_held",   block="qbo", label="Retainage held",
         working_hdr="RETAINAGE HELD",        master_hdr=None),
    dict(key="orig_contract", attr="base_contract",    block="pm",  label="Original contract",
         working_hdr="ORIGINAL CONTRACT",     master_hdr="TOTAL CONTRACT PRICE"),
    dict(key="approved_cos",  attr="co_revenue",       block="pm",  label="Approved COs",
         working_hdr="APPROVED COs",          master_hdr=None),
    dict(key="orig_etc",      attr="base_etc",         block="pm",  label="Original ETC",
         working_hdr="ORIGINAL ESTIMATED COST", master_hdr="ESTIMATED TOTAL COSTS"),
    dict(key="co_costs",      attr="co_cost_override", block="pm",  label="CO costs",
         working_hdr="CO COSTS",              master_hdr=None),
]
FIELD_BY_KEY = {f["key"]: f for f in FIELDS}
KEY_BY_ATTR = {f["attr"]: f["key"] for f in FIELDS}
PM_KEYS = tuple(f["key"] for f in FIELDS if f["block"] == "pm")
QBO_KEYS = tuple(f["key"] for f in FIELDS if f["block"] == "qbo")
_EPS = 0.01                                   # money equal within a cent

# Job statuses in display order. REVERSED sits between CHANGED and ADDED.
STATUS_ORDER = {"CHANGED": 0, "REVERSED": 1, "ADDED": 2, "REMOVED": 3, "SAME": 4}
CHANGED_STATUSES = ("CHANGED", "REVERSED", "ADDED", "REMOVED")

# The words the UI prints. One copy, here.
CARRIED_SOURCE = "carried from tab"
CARRY_NOTE = "source not found - kept the tab value"
NO_SOURCE_NOTE = "no document named for this value"
NO_QBO_NOTE = ("no QBO value this run - kept unchecked, approve only if the job "
               "truly has none")
DECREASE_NOTE = {
    "costs":     "decreased - a bill was voided, deleted or re-coded?",
    "billed":    "decreased - an invoice was voided, deleted or credited?",
    "retainage": "decreased - retainage billed out, released or re-coded?",
}


def _num(v) -> Optional[float]:
    """Any cell/attr -> float, or None. Tolerates '$1,234', '(123)', '', formulas."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _changed(old, new) -> bool:
    a, b = _num(old), _num(new)
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    return abs(a - b) >= _EPS


def _fields_for(tab_kind: str) -> List[dict]:
    """The review fields that exist as columns on this kind of tab."""
    hdr = "master_hdr" if tab_kind == "master" else "working_hdr"
    return [f for f in FIELDS if f[hdr]]


def _key(key_or_attr: str) -> str:
    """Accept either the review key ('orig_etc') or the CpRow attr ('base_etc')."""
    if key_or_attr in FIELD_BY_KEY:
        return key_or_attr
    if key_or_attr in KEY_BY_ATTR:
        return KEY_BY_ATTR[key_or_attr]
    raise KeyError(f"not a review field: {key_or_attr}")


# ── read the current tab (the "before") ──────────────────────────────────────
def snapshot_tab(wip_path: Path, tab_name: str, tab_kind: str) -> Dict[str, dict]:
    """{PROJECT# -> {field_key: value, '_name':, '_notes':}} straight off the tab
    as it stands now. tab_kind is 'working' or 'master' (picks the header set).
    Missing file/tab -> {} so a first-ever run reads as all-ADDED."""
    from openpyxl import load_workbook
    wip_path = Path(wip_path)
    if not wip_path.exists():
        return {}
    wb = load_workbook(wip_path, data_only=True, read_only=True)
    if tab_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[tab_name]
    hdr = next((r for r in range(1, 16)
                for c in range(1, (ws.max_column or 0) + 1)
                if ws.cell(r, c).value == "PROJECT #"), None)
    if not hdr:
        wb.close()
        return {}
    idx = {ws.cell(hdr, c).value: c for c in range(1, (ws.max_column or 0) + 1)}
    hdr_key = "master_hdr" if tab_kind == "master" else "working_hdr"
    pcol, ncol = idx.get("PROJECT #"), idx.get("PROJECT NAME")
    out: Dict[str, dict] = {}
    for r in range(hdr + 1, (ws.max_row or 0) + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip() in ("TOTALS", "TOTAL"):
            break                                  # summary block, not data
        pn = ws.cell(r, pcol).value if pcol else None
        if not pn or not str(pn).strip():
            continue
        rec = {"_name": (ws.cell(r, ncol).value if ncol else None)}
        for f in FIELDS:
            col = idx.get(f[hdr_key]) if f[hdr_key] else None
            rec[f["key"]] = _num(ws.cell(r, col).value) if col else None
        out[str(pn).strip().upper()] = rec
    wb.close()
    return out


def row_value(row, key: str):
    """The fresh value a CpRow carries for a review field (the "after")."""
    return _num(getattr(row, FIELD_BY_KEY[key]["attr"], None))


# ── where a value came from ──────────────────────────────────────────────────
def set_source(row, key: str, label: str, path=None) -> None:
    """Readers call this the moment they read a review field. `key` is the
    review key or the CpRow attr; `label` names the document in a few words;
    `path` is the file it came from (optional). Kept on the row as
    `review_sources` - CpRow is a plain dataclass and ad-hoc attributes are how
    the readers already carry audit provenance (audit_etc_src, cell_marks)."""
    srcs = getattr(row, "review_sources", None)
    if srcs is None:
        srcs = {}
        setattr(row, "review_sources", srcs)
    srcs[_key(key)] = (label, str(path) if path else None)


def row_source(row, key: str) -> Tuple[Optional[str], Optional[str]]:
    """(source, source_path) for a review field: the source the reader stamped,
    else what the row's audit attributes imply, else (None, None) - and the
    diff then says so on the cell instead of guessing."""
    srcs = getattr(row, "review_sources", None) or {}
    if key in srcs:
        return srcs[key]
    draw_num = getattr(row, "draw_num", None)
    draw_path = getattr(row, "draw_path", None)
    takeoff = getattr(row, "takeoff_path", None)
    if key in ("orig_contract", "approved_cos"):
        if draw_path:
            return (f"Draw #{draw_num} G702 · {Path(draw_path).name}", str(draw_path))
        s = getattr(row, "audit_contract_src", None)
        if s:
            return (s, getattr(row, "src_link", None))
        if key == "orig_contract" and takeoff:
            return (Path(takeoff).name, str(takeoff))
    elif key == "orig_etc":
        s = getattr(row, "audit_etc_src", None)
        if s:
            return (s, None)
        if takeoff:
            return (Path(takeoff).name, str(takeoff))
    return (None, None)


# ── THE CARRY RULE ───────────────────────────────────────────────────────────
def carry_pm_fields(rows: List, prior: Dict[str, dict], *,
                    tab_kind: str = "working") -> Dict[tuple, float]:
    """For every PM field on this tab where the fresh row has NO value (no
    document found, or an unreadable / ambiguous one) while the tab holds one,
    put the tab value back on the row and stamp its source "carried from tab".
    Returns {(PROJECT#, key): carried value}.

    Applied to the row OBJECT on purpose: the review (diff_rows), the merge
    (apply_decisions) and the guarded writers all read the same attribute, so
    the carried number is what gets written - never a blank. Runs in both
    --emit-review and --apply-review, against the tab as it stands at that
    moment. Idempotent.

    On the lean Test-Master the PM columns are the folded TOTAL CONTRACT PRICE /
    ESTIMATED TOTAL COSTS; the carry lands them in base_contract / base_etc,
    which is exact whenever the row carries no CO breakout - always for MFD,
    and for CP/RP rows at apply time (the CO cost override is harvested by the
    writer after this step)."""
    carried: Dict[tuple, float] = {}
    pm_here = [f for f in _fields_for(tab_kind) if f["block"] == "pm"]
    for row in rows:
        pn = row.project_num.strip().upper()
        was = prior.get(pn)
        if not was:
            continue
        for f in pm_here:
            old = _num(was.get(f["key"]))
            if old is None or row_value(row, f["key"]) is not None:
                continue
            setattr(row, f["attr"], old)
            set_source(row, f["key"], CARRIED_SOURCE)
            carried[(pn, f["key"])] = old
    return carried


# ── build the diff (the review payload for one tab) ──────────────────────────
def _cell(f: dict, was, now, **kw) -> dict:
    """One field cell. Every cell has every key so the JSON shape never varies."""
    d = dict(key=f["key"], label=f["label"], block=f["block"], was=was, now=now,
             changed=False, reversed=False, decreased=False, carried=False,
             source=None, source_path=None, note=None)
    d.update(kw)
    return d


def diff_rows(new_rows: List, prior: Dict[str, dict], *, division: str,
              tab_name: str, tab_kind: str) -> List[dict]:
    """One record per job: status CHANGED / REVERSED / ADDED / REMOVED / SAME
    and, for each field that exists on this tab, a cell:
      key · label · block · was · now · changed · reversed · decreased · carried
      · source · source_path · note
    QBO block first, PM block next. The carry rule runs FIRST and mutates the
    rows (see carry_pm_fields), so a carried PM field reads was == now,
    changed False, with its note - never "cleared"."""
    fields_here = _fields_for(tab_kind)
    carried = carry_pm_fields(new_rows, prior, tab_kind=tab_kind)
    records, seen = [], set()
    for row in new_rows:
        pn = row.project_num.strip().upper()
        seen.add(pn)
        was = prior.get(pn)
        cells, any_change, any_rev = [], False, False
        for f in fields_here:
            key = f["key"]
            now = row_value(row, key)
            old = was.get(key) if was else None
            src, src_path = row_source(row, key)
            cell = _cell(f, old, now, source=src, source_path=src_path)
            if (pn, key) in carried:
                cell["carried"] = True                # changed stays False
                cell["note"] = CARRY_NOTE
            elif not was:
                cell["changed"] = now is not None
            else:
                ch = _changed(old, now)
                cell["changed"] = ch
                if ch and old is not None and now is not None and now < old:
                    if f["block"] == "pm":
                        cell["reversed"] = True
                    else:
                        cell["decreased"] = True
                        cell["note"] = DECREASE_NOTE.get(key, "decreased")
                elif ch and now is None and f["block"] == "qbo":
                    cell["note"] = NO_QBO_NOTE
            if cell["changed"] and f["block"] == "pm" and not src:
                cell["note"] = NO_SOURCE_NOTE
            any_change = any_change or cell["changed"]
            any_rev = any_rev or cell["reversed"]
            cells.append(cell)
        status = ("ADDED" if not was else "REVERSED" if any_rev
                  else "CHANGED" if any_change else "SAME")
        records.append(dict(
            project_num=row.project_num, name=(row.project_name or ""),
            division=division, tab=tab_name, tab_kind=tab_kind,
            section=(getattr(row, "section", None) or getattr(row, "rp_type", None) or ""),
            status=status, fields=cells,
            flags="; ".join((getattr(row, "status_flags", None) or [])
                            + (getattr(row, "notes", None) or [])),
        ))
    for pn, was in prior.items():                  # on the tab last run, gone now
        if pn in seen:
            continue
        cells = [_cell(f, was.get(f["key"]), None,
                       changed=(was.get(f["key"]) is not None))
                 for f in fields_here]
        records.append(dict(
            project_num=pn, name=(was.get("_name") or ""), division=division,
            tab=tab_name, tab_kind=tab_kind, section="", status="REMOVED",
            fields=cells, flags=""))
    records.sort(key=lambda d: (STATUS_ORDER[d["status"]], d["project_num"]))
    return records


def summarize(records: List[dict]) -> str:
    """One line for the terminal: '34 changed of 51 · 2 reversed · 4 carried'."""
    changed = sum(r["status"] in CHANGED_STATUSES for r in records)
    reversed_ = sum(r["status"] == "REVERSED" for r in records)
    carried = sum(c["carried"] for r in records for c in r["fields"])
    return f"{changed} changed of {len(records)} · {reversed_} reversed · {carried} carried"


# ── apply the owner's decisions before a real write ──────────────────────────
def apply_decisions(rows: List, decisions: dict, prior: Optional[Dict[str, dict]] = None,
                    *, tab_kind: str = "working") -> List:
    """Return the row list to actually write, honouring the owner's marks.

    decisions = {
      "fields": {"<PN>": {"<field_key>": {"approved": bool, "revert": <num|null>}}},
      "drop_added": ["<PN>", ...],          # ADDED jobs the owner rejected
    }

    When `prior` (the tab snapshot) is given, THE CARRY RULE runs first: a PM
    field the reader could not source this run keeps the tab value, exactly as
    the review showed it. Without it a later approved sync could blank a value
    the review never proposed changing.

    A DISAPPROVED field is reverted to `revert` - the exact "was" value the owner
    saw in the review - so every tab reverts the SAME source number and can't
    drift (the lean Master tab folds Contract+COs into one column, so re-reading
    it here would be wrong; the carried value avoids that). An APPROVED or unmarked
    field keeps its freshly computed value. An ADDED job in drop_added is dropped.
    (Reverting the source fields - base_contract, co_revenue - also fixes the
    Master tab's derived TOTAL CONTRACT PRICE / ESTIMATED TOTAL COSTS for free.)"""
    if prior:
        carried = carry_pm_fields(rows, prior, tab_kind=tab_kind)
        if carried:
            print(f"  · carried {len(carried)} PM value(s) from the tab (no source "
                  f"this run): " + ", ".join(f"{pn} {key}" for pn, key in sorted(carried)))
    field_dec = (decisions or {}).get("fields", {}) or {}
    drop = {p.upper() for p in ((decisions or {}).get("drop_added", []) or [])}
    kept = []
    for row in rows:
        pn = row.project_num.strip().upper()
        if pn in drop:                              # rejected brand-new job
            continue
        for key, d in field_dec.get(pn, {}).items():
            if key not in FIELD_BY_KEY:
                continue
            approved = d.get("approved", True) if isinstance(d, dict) else bool(d)
            if approved:
                continue                            # approved -> keep fresh value
            revert = d.get("revert") if isinstance(d, dict) else None
            setattr(row, FIELD_BY_KEY[key]["attr"], revert)   # disapproved -> what they saw
        kept.append(row)
    return kept


# ── JSON on disk (what the tools write / the dashboard reads) ────────────────
def write_review_json(out_path: Path, division: str, tab_name: str,
                      records: List[dict]) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"division": division, "tab": tab_name, "count": len(records),
               "changed": sum(r["status"] in CHANGED_STATUSES for r in records),
               "reversed": sum(r["status"] == "REVERSED" for r in records),
               "carried": sum(c["carried"] for r in records for c in r["fields"]),
               "records": records}
    out_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return out_path


def load_decisions(path: Path) -> dict:
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
