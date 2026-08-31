#!/usr/bin/env python3
"""
project_pnl_export.py — Generate per-project P&L workbooks from QBO.

FINAL LAYOUT SPEC (the user, 2026-06-05): TWO sheets, PLAIN formatting — white
sheet, black text, bold + indent + borders only. No fills, no hidden rows.

  Sheet "P&L"     LEFT (cols A–B): QBO's Project P&L line items verbatim,
                  then Overhead Allocation, TRUE NET PROFIT / %, Billing &
                  Retainage. RIGHT (cols D–J): ACCUMULATING COSTS — NEXT DRAW
                  (sorted by Cost Code Sheet order), then the DRAW COVERAGE
                  table beneath it.
  Sheet "Labor" / "Concrete"
                  BUDGET vs ACTUAL BY DRAW for the two trades the PM and ops
                  manager track by scrutiny (the user 2026-07-29) — rows are
                  the takeoff's cost codes, columns are the draw windows, and
                  each code expands (outline ±) to the bills behind it. Sales
                  tax is pulled out of the comparison and summed at the bottom;
                  Concrete adds yards and $/yd vs the takeoff's implied rate.
  Sheet "Draws"   DETAIL ONLY (cols A–E): COSTS OUTSIDE DRAW WINDOWS first
                  (collapsed, sum on the header, click + to expand), then ONE
                  BLOCK PER DRAW newest-first (invoices, then costs grouped
                  cost code / account → sub-account → vendor → transactions,
                  subtotal ON each group row WITH a tracing line under it).
                  Untagged invoices block at the end.

Draw-period tagging comes from invoice PrivateNote in the format
"(Period: MM/DD/YY - MM/DD/YY)". Bills are NEVER period-tagged — costs are
bucketed into draw windows by TxnDate.

Account names that are cost codes (e.g. CS1, SL6, PV51) are broken out to
"CODE - Job Type Cost Name" with ONLY the code bold (rich text); see
_cost_code_label() / _cost_code_value().

USAGE
    python3 project_pnl_export.py MFD177 MFD325 CP672 CP745
    python3 project_pnl_export.py active cp          # every Active CP project
    python3 project_pnl_export.py active rp mfd      # Active RP + MFD batch
    python3 project_pnl_export.py --out "/path/to/folder" MFD177
    python3 project_pnl_export.py --dry-run MFD177

OPTIONS
    --out            Output folder. Defaults to:
                     ~/Library/CloudStorage/OneDrive-ProficientConcrete,LLC/
                       Automations-/PROJECT P&Ls
    --start-date     P&L start date (default: 2020-01-01)
    --end-date       P&L end date (default: today)
    --wip-master     Path to WIP master .xlsx for Contract/ETC/STATUS lookup
                     (default: the SharePoint 'WIP - MASTER new.xlsx' via
                     WIP_EXCEL_PATH / ACB_ONEDRIVE_BASE; reads the Test-Master tab)
    --dry-run        Print what would be written, don't save files.

DEPENDENCIES
    pip3 install --break-system-packages openpyxl requests xlrd
    (xlrd reads the legacy .xls G702 pay applications; without it the contract
     price falls back to the WIP master and the run says so.)
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
import time
import warnings

# RP takeoffs carry INDIRECT() print areas openpyxl can't keep — harmless
# (read-only budget pull), silence the noise (same as the WIP readers).
warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name.*")
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    # NOTE: no rich-text imports on purpose — multi-run inline strings are what
    # make Mac Excel offer to "repair" the file (shared/xlsx_verify.py). Style
    # the cell, never runs inside it.
except ImportError:
    print("✗  pip3 install --break-system-packages openpyxl requests")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import pnl_paths
from shared.draws import read_pay_app, learn_period_shape, infer_period_tag
from shared.qbo_api import (
    API_BASE, MINOR_VERSION, PROJ_RE,
    load_credentials, _api_get, query_all, report,
    extract_proj, build_project_customer_map,
    fetch_project_pl, _walk_pl_rows, extract_pl_totals,
    fetch_customer_invoices,
)
from shared.cost_lines import line_category, combine_bill_lines, CATEGORY_ORDER
# cost_leaf moved to shared/ (2026-08-08) — the ledger's load_costs.py needs the
# SAME resolver, so it can never drift from this tool's cost buckets.
from shared.qbo_costs import cost_leaf
from shared.job_lines import JobMatcher, discover_job_classes
from shared.xlsx_verify import assert_clean

# ── terminal output (styled like sync-ar / sync-ap; the user 2026-06-26) ──────────
# Colors auto-disable when piped/redirected or NO_COLOR is set.
_TTY = sys.stdout.isatty() and os.environ.get("NO_COLOR") is None


def _c(code: str) -> str:
    return code if _TTY else ""


_RESET, _BOLD, _DIM = _c("\033[0m"), _c("\033[1m"), _c("\033[90m")
_RED, _GREEN, _YEL = _c("\033[31m"), _c("\033[32m"), _c("\033[33m")
_CYAN, _BLUE, _BRCYAN = _c("\033[36m"), _c("\033[94m"), _c("\033[96m")
_CHECK = f"{_GREEN}✓{_RESET}"
_W = 54                                   # banner inner width


def ui_banner(title: str) -> None:
    inner = f" {title} "
    print(f"\n{_BOLD}{_BLUE}╭─{inner}{'─' * max(2, _W - len(inner))}╮{_RESET}")


def ui_close() -> None:
    print(f"{_BOLD}{_BLUE}╰{'─' * (_W + 1)}╯{_RESET}")


def ui_cfg(label: str, value) -> None:
    print(f"  {_DIM}{label:<11}{_RESET} {value}")


def ui_step(label: str, detail: str = "", done: bool = True) -> None:
    mark = _CHECK if done else f"{_CYAN}·{_RESET}"
    det = f"  {_BOLD}{_BRCYAN}{detail}{_RESET}" if detail else ""
    print(f"  {label}{det}  {mark}".rstrip())


def ui_proj(proj: str, name: str) -> None:
    print(f"\n{_BOLD}{_CYAN}▌ {proj}{_RESET}  {_DIM}{name}{_RESET}")


def ui_event(text: str, icon: str = "·", color: str = "") -> None:
    print(f"     {color or _DIM}{icon} {text}{_RESET}")


def ui_warn(text: str) -> None:
    print(f"     {_YEL}⚠ {text}{_RESET}")


def ui_done(text: str) -> None:
    print(f"     {_GREEN}✓ {text}{_RESET}")


def ui_fail(text: str) -> None:
    print(f"  {_RED}✗ {text}{_RESET}")


DEFAULT_OUT = paths.get_path(
    "ACB_PNL_OUT_DIR",
    paths.onedrive_base() / "Automations-/PROJECT P&Ls",
)
# The WIP master is the SharePoint workbook the wip/ readers maintain (Test-Master
# tab = the unified MFD+CP+RP table). Same key + default as cp_wip_reader.py, so
# one machine.env override moves both tools (the user 2026-07-16 — auto-pull
# Contract/ETC instead of leaving the yellow cells blank).
DEFAULT_WIP_MASTER = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report" / "WIP - MASTER new.xlsx",
)
# Per-project home folder (the user 2026-06-25): each P&L run creates <out>/<PROJ>/ and
# <out>/<PROJ>/rd-reports/ (skipped if they exist). The workbook lands in the
# project folder; the PM drops draw-cost reports in rd-reports/, and the P&L auto
# cross-checks each against QBO for the missed/underbilled total per draw.
DRAW_REPORTS_SUBDIR = "rd-reports"

# CP-only (the user 2026-07-02): Commercial P&Ls drop into the awarded-project folder on
# the Common drive — <CP project folder>/Profit and Loss/ — mirroring where the WIP
# report lives. If the drive isn't mounted (or no matching folder is found), fall
# back to the OneDrive PROJECT P&Ls tree. MFD/RP are unaffected.
CP_AWARDED_BASE = Path(
    "/Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects"
)
CP_PNL_SUBDIR = "Profit and Loss"


def _find_awarded_cp_folder(base: Path, proj: str) -> Optional[Path]:
    """The awarded-project folder for a CP job, matched by project #, inspired by
    how the WIP report locates a job folder. Prefers a folder whose name contains
    the full project # (e.g. 'CP672'); falls back to one containing the bare number
    ('672') on a digit boundary. Returns None if the base isn't reachable or nothing
    matches (caller then falls back to OneDrive)."""
    try:
        if not base.is_dir():
            return None
    except OSError:
        return None
    pu = proj.upper()
    num = re.sub(r"\D", "", proj)          # 'CP672' -> '672'
    numbered = None
    for child in sorted(base.iterdir()):
        if not child.is_dir():
            continue
        compact = re.sub(r"[\s\-_]+", "", child.name.upper())
        if pu in compact:                  # strongest: name carries 'CP672'
            return child
        if num and numbered is None and re.search(rf"(?<!\d){num}(?!\d)", child.name):
            numbered = child               # weaker: bare number match, keep first
    return numbered


def _resolve_project_out_dir(proj: str, out_dir: Path) -> Tuple[Path, Optional[str]]:
    """Where a project's workbook folder should live. CP → Common-drive awarded
    folder's 'Profit and Loss' subfolder; everything else → <out_dir>/<proj>.
    Returns (folder, note) where note explains any CP fallback (for the UI)."""
    if not proj.upper().startswith("CP"):
        # If this job has already been FILED under an archive subfolder
        # ("completed mfd project p&l"), regenerate it THERE — otherwise a
        # re-run silently creates a second copy at the top level and the two
        # drift apart (the user 2026-08-27).
        for _arch in pnl_paths._archive_dirs():
            if (_arch / proj).is_dir():
                return _arch / proj, f"filed under {_arch.name}"
        return out_dir / proj, None
    if not CP_AWARDED_BASE.exists():
        return out_dir / proj, "Common drive not mounted → OneDrive"
    folder = _find_awarded_cp_folder(CP_AWARDED_BASE, proj)
    if folder is None:
        return out_dir / proj, f"no awarded folder for {proj} → OneDrive"
    return folder / CP_PNL_SUBDIR, None


# ─────────── regexes ───────────
DRAW_PERIOD_RE = re.compile(
    r"\(\s*Period\s*:\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*\)",
    re.IGNORECASE,
)
# Draw NUMBER from the invoice memo: "Draw 11" / "Draw #11". 1–3 digits only,
# and NOT followed by another digit, so a year ("April Draw 2026") is ignored.
DRAW_NUM_RE = re.compile(r"Draw\s*#?\s*(\d{1,3})(?!\d)", re.IGNORECASE)


def extract_draw_number(memo: str) -> Optional[int]:
    """Read the draw # from an invoice memo (the user 2026-06-09: use the memo's
    number, never auto-assign). Returns None if no 'Draw N' is present."""
    if not memo:
        return None
    m = DRAW_NUM_RE.search(memo)
    return int(m.group(1)) if m else None
# Retainage detection: any of retainage/retainaged/retention (substring).
RETAINAGE_RE = re.compile(r"retainag|retention", re.IGNORECASE)
# The ONLY discriminator (the user 2026-06-19): a retainage invoice is NOT BILLED iff
# "not billed" appears ANYWHERE in its memo or any line description; otherwise
# it's a real (billed) retainage draw — "City" or anything else is irrelevant.
NOT_BILLED_RE = re.compile(r"not\s*billed", re.IGNORECASE)
RETAINAGE_NOT_BILLED_RE = re.compile(r"retainag.*not\s*billed", re.IGNORECASE)  # legacy (unused)


def _is_retainage_not_billed(inv: dict) -> bool:
    """True if a raw QBO invoice is a 'retainage NOT billed' doc — retainage moved
    to Retainage Receivable by JE, NOT earned revenue (the user 2026-07-02). Signal:
    both 'retainag/retention' AND 'not billed' appear in the memo or any line
    text. Used to keep these out of billed income on EVERY template (the Draw
    template routes them via __retainage; the RP template guards with this)."""
    parts = [inv.get("PrivateNote", "") or ""]
    for ln in inv.get("Line") or []:
        parts.append(ln.get("Description") or "")
        parts.append(((ln.get("SalesItemLineDetail") or {})
                      .get("ItemRef", {}) or {}).get("name") or "")
    text = " ".join(parts)
    return bool(RETAINAGE_RE.search(text) and NOT_BILLED_RE.search(text))
# RP (residential): a bill dated AFTER the invoice whose line says "wreck"
# (wrecking/demo labor) belongs to that job; a late non-wreck bill is suspect.
WRECK_RE = re.compile(r"wreck", re.IGNORECASE)
# SUB bills lag — a sub bill dated after the invoice is normal billing delay, so
# it's INCLUDED, not flagged (the user 2026-06-19). Subs are tagged "sub" in the memo.
SUB_RE = re.compile(r"\bsub\b|subcontract|sub\s*service", re.IGNORECASE)
# A work PERIOD anywhere in the memo ("Period 3/13/2026 - 3/19/2026"), colon
# optional. If the period STARTS on/before the invoice, the work predates the
# invoice → the late bill belongs to the job.
PERIOD_ANYWHERE_RE = re.compile(
    r"Period\s*:?\s*(\d{1,2}/\d{1,2}/\d{1,6})\s*[-–]\s*\d{1,2}/\d{1,2}/\d{1,6}",
    re.IGNORECASE)
# NOTE: sub bills ("sub" in PrivateNote) are INCLUDED in all cost figures.
# The tag exists for the separate Sub LOC tracker, not for P&L exclusion.

# ─────────── cost-memo display cleaning (the user 2026-07-16) ───────────
# Cost memos carry project identification the report doesn't need — the
# project # ("CP672"), the address ("E OVILLA RD"), and the GC/project name
# ("MADEWELL COMPANIES", "FIRESTONE RED OAK"). Show only the actual item.
# Segments are split on "-"; a segment is DROPPED when it is a project-#
# token, looks like an address (street-suffix word), is company-ish
# (COMPANIES/LLC/…), or shares a significant word with the customer/project
# name. Sub bills: the "Sub Service: Period" label goes, the DATES stay.
_PROJ_TOKEN_RE = re.compile(r"\b(?:MFD|CP|RP)\s?\d+(?:-FTW)?\b[.,]?", re.IGNORECASE)
_SUBSVC_LABEL_RE = re.compile(r"\bsub\s*service\s*:?\s*(?:period)?\s*:?\s*",
                              re.IGNORECASE)
_STREETISH_RE = re.compile(
    r"\b(?:RD|ROAD|ST|STREET|BLVD|AVE|AVENUE|LN|LANE|DR|DRIVE|HWY|PKWY|CT|WAY)\b\.?",
    re.IGNORECASE)
_COMPANYISH_RE = re.compile(
    r"\b(?:COMPANIES|COMPANY|CONSTRUCTION|BUILDERS|LLC|INC|CORP)\b\.?",
    re.IGNORECASE)
_TEXT_SEG_SPLIT_RE = re.compile(r"\s+[-–—]\s+")


def _project_name_words(cust_name: str) -> frozenset:
    """Significant words (≥4 letters) of 'Parent:Proj# Project Name' — used to
    recognize project/GC-name segments in memos regardless of spelling drift."""
    return frozenset(w for w in re.findall(r"[A-Za-z]{4,}", (cust_name or "").upper()))


def _clean_cost_text(text: str, known_words: frozenset = frozenset()) -> str:
    """Strip project identification from a cost memo/description, keep the item."""
    t = _SUBSVC_LABEL_RE.sub("", text or "")
    t = _PROJ_TOKEN_RE.sub("", t)
    kept = []
    for seg in _TEXT_SEG_SPLIT_RE.split(t):
        s = seg.strip(" ,-–—")
        if not s:
            continue
        seg_words = {w for w in re.findall(r"[A-Za-z]{4,}", s.upper())}
        if (_STREETISH_RE.search(s) or _COMPANYISH_RE.search(s)
                or (known_words and seg_words & known_words)):
            continue
        kept.append(s)
    return " - ".join(kept)

# ─────────── styling ───────────
# Sizes bumped 2026-05-28 — the user: "make a bit bigger overall, it shows a
# little too small when first opened". Base font 12 with proportional bumps.
BASE_SIZE = 12
HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=BASE_SIZE + 1)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHDR_FONT = Font(bold=True, color="1F3A5F", size=BASE_SIZE)
TOTAL_FILL = PatternFill("solid", fgColor="E7E6E6")
TOTAL_FONT = Font(bold=True, size=BASE_SIZE)
MANUAL_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
CURR_FMT = '#,##0.00;[Red]-#,##0.00'
# Accounting dollars for the Labor/Concrete sheets (the user 2026-08-04:
# "make all $ a number $$$ format") — $ pinned left, zeros as "-", red parens.
ACC_FMT = '_("$"* #,##0.00_);[Red]_("$"* (#,##0.00)_);_("$"* "-"_);_(@_)'
PCT_FMT = '0.00%'

# Control chars that are illegal in XML 1.0 / .xlsx. Real QBO strings
# (vendor names, memos, descriptions) occasionally carry these from imports
# or copy-paste. Older openpyxl writes them silently → Excel "found a problem
# with content, recover" on open. We strip them at the source so the output
# is valid regardless of the openpyxl version on the running machine.
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _xml_clean(s):
    """Strip XML-illegal control chars from a string (passes non-str through)."""
    if not isinstance(s, str):
        return s
    return _ILLEGAL_XML_RE.sub("", s)


# ─────────── cost-code → name mapping (the user 2026-06-09) ───────────
# Proficient switched from category-named accounts to cost codes partway
# through. A code = optional 2-letter JOB-TYPE prefix + a cost-code number.
# We display the COST NAME (the account it lands in) so old category names
# and new cost codes read the same and merge together.
_COST_CODE_NAMES = {
    "1": "Concrete",
    "2": "Rebar & Reinforcement",
    "3": "Formwork & Lumber",
    "4": "Aggregates",
    "5": "Equipment & Rentals",
    "51": "Pump",
    "52": "Saw Cutting",
    "6": "Labor",
    "7": "Specialty/Misc.",
    "8": "Fuel",
    "9": "Supplies",
}
# Job-type prefixes: SL PV FW PR WL CS MS. Number: 51/52 before single digits.
_COST_CODE_RE = re.compile(
    r"^(?:SL|PV|FW|PR|WL|CS|MS)?(51|52|[1-9])$", re.IGNORECASE
)


def _is_cost_code(name) -> bool:
    """True if the string is a cost code (e.g. 'CS1', 'SL6', 'PV51', '9'),
    optionally wrapped in 'Total '/'Total for '."""
    if not isinstance(name, str):
        return False
    s = name.strip()
    if _COST_CODE_RE.match(s):
        return True
    for pfx in ("Total for ", "Total "):
        if s.startswith(pfx) and _COST_CODE_RE.match(s[len(pfx):].strip()):
            return True
    return False


# Job-type prefix → full name
_JOB_TYPE_NAMES = {
    "SL": "Slab", "PV": "Paving", "FW": "Flatwork for Residential",
    "PR": "Piers", "WL": "Walls", "CS": "Commercial Sidewalks",
    "MS": "Miscellaneous",
}
# Capture prefix and number separately for the full break-out.
_COST_CODE_SPLIT_RE = re.compile(
    r"^(SL|PV|FW|PR|WL|CS|MS)?(51|52|[1-9])$", re.IGNORECASE
)
# Display order from the Cost Code Sheet: job-type prefix order, then
# cost-code number order (1,2,3,4,5,51,52,6,7,8,9).
_JOB_PREFIX_ORDER = {p: i for i, p in enumerate(
    ["SL", "PV", "FW", "PR", "WL", "CS", "MS"])}
_COST_NUM_ORDER = {n: i for i, n in enumerate(
    ["1", "2", "3", "4", "5", "51", "52", "6", "7", "8", "9"])}


# Job-type band colors (the user 2026-07-17) — one soft tint per prefix so the
# Budget vs Actual cost-code rows group visually by trade (Slab, Paving, …).
_JOB_BAND_HEX = {
    "SL": "DDEBF7",   # Slab — blue
    "PV": "E2EFDA",   # Paving — green
    "FW": "FCE4D6",   # Flatwork (Residential) — peach
    "PR": "E4DFEC",   # Piers — lavender
    "WL": "D6EAE6",   # Walls — teal
    "CS": "FFF6D5",   # Commercial Sidewalks — yellow
    "MS": "F2DEDE",   # Miscellaneous — rose
}
_JOB_BAND_DEFAULT = "ECECEC"   # non-prefixed / unclassified codes — gray


def _cost_band_fill(code) -> PatternFill:
    """Soft job-type band fill for a cost-code row (Budget vs Actual)."""
    pfx, _num = _split_code(code)
    return PatternFill("solid",
                       fgColor=_JOB_BAND_HEX.get((pfx or "").upper(),
                                                 _JOB_BAND_DEFAULT))


def _cost_code_sort_key(name):
    """Sort key matching the Cost Code Sheet: prefix order, then cost number.
    Cost codes sort first (grouped by job type); non-codes go last,
    alphabetically."""
    if isinstance(name, str):
        m = _COST_CODE_SPLIT_RE.match(name.strip())
        if m:
            pfx, num = m.group(1), m.group(2)
            return (0,
                    _JOB_PREFIX_ORDER.get((pfx or "").upper(), 99),
                    _COST_NUM_ORDER.get(num, 99))
    return (1, 99, 99, str(name).lower())


def _cost_code_label(name):
    """Break a cost code out into its full meaning, KEEPING the code:
    'CS1' -> 'CS1 - Commercial Sidewalks Concrete'
    'SL6' -> 'SL6 - Slab Labor'
    'PV51' -> 'PV51 - Paving Pump'
    '9'   -> '9 - Supplies'   (no job-type prefix)
    Non-codes pass through unchanged."""
    if not isinstance(name, str):
        return name
    s = name.strip()

    def expand(code: str) -> Optional[str]:
        m = _COST_CODE_SPLIT_RE.match(code)
        if not m:
            return None
        pfx, num = m.group(1), m.group(2)
        cost = _COST_CODE_NAMES.get(num, "")
        if pfx:
            job = _JOB_TYPE_NAMES.get(pfx.upper(), "")
            return f"{code} - {job} {cost}".strip()
        return f"{code} - {cost}".strip()

    full = expand(s)
    if full:
        return full
    for wrap in ("Total for ", "Total "):
        if s.startswith(wrap):
            inner = expand(s[len(wrap):].strip())
            if inner:
                return wrap + inner
    return name


def _cost_code_value(raw, indent: int = 0, size: Optional[int] = None,
                     color: str = "000000"):
    """Build the cell value for an account label: the indented
    'SL1 - Slab Concrete' string.

    PLAIN STRING, never rich text. This used to return a CellRichText with the
    code token bold and the description regular (the user 2026-06-09, "just the
    cost code bolded"), but multi-run inline strings are the exact thing
    `shared/xlsx_verify` refuses — they are what makes Mac Excel throw "we found
    a problem with some content" and offer to repair the file. A P&L whose
    accumulating-costs block happened to contain a cost code shipped a workbook
    at risk of that prompt (found on MFD172, 2026-08-24; the header of
    shared/xlsx_verify.py is the standing list). The `size`/`color` arguments
    are kept so every call site still works — style the CELL, not runs inside
    it."""
    pad = "    " * indent
    return pad + _xml_clean(_cost_code_label(raw))


# cost_leaf now lives in shared/qbo_costs.py (imported above) — the single
# resolver shared with the ledger. Do not re-add a local copy here.


def _split_code(name):
    """Return (PREFIX_or_None, NUM_or_None) for a cost code, else (None, None)."""
    if isinstance(name, str):
        m = _COST_CODE_SPLIT_RE.match(name.strip())
        if m:
            return (m.group(1) or "").upper() or None, m.group(2)
    return None, None


def _cost_name_only(code):
    """'SL2' -> 'SL2 - Rebar & Reinforcement' (drops the redundant job-type
    word; used UNDER a job-type group header that already names the job)."""
    pfx, num = _split_code(code)
    if num is None:
        return code
    return f"{code} - {_COST_CODE_NAMES.get(num, '')}".rstrip(" -")


def _cost_name_value(code, indent: int = 0, size: Optional[int] = None,
                     color: str = "000000"):
    """'SL2 - Rebar & Reinforcement' (no job-type word), indented. PLAIN
    STRING — see `_cost_code_value` for why rich text is banned here."""
    pad = "    " * indent
    return pad + _xml_clean(_cost_name_only(code))


# Cost CATEGORY = the account a cost lands in, merging cost codes by meaning
# (the user 2026-06-09): SL1/PV1/CS1… → "Concrete"; a real account name keeps
# itself. Used by the Draws sheet so there's ONE Concrete, ONE Labor, etc.
# Order = biggest money movers first (the user 2026-07-15): Concrete, then
# Labor, then the material/other codes in their original order.
_COST_NAME_ORDER = {name: i for i, name in enumerate(
    ["Concrete", "Labor"]
    + [n for n in _COST_CODE_NAMES.values() if n not in ("Concrete", "Labor")])}


def _cost_category(leaf):
    pfx, num = _split_code(leaf)
    if num is not None:
        return _COST_CODE_NAMES.get(num, leaf)
    return leaf


def _cost_category_sort_key(name):
    return (_COST_NAME_ORDER.get(name, 99), str(name).lower())


def _safe_cell(v):
    """Sanitize any value bound for a worksheet cell: strip illegal XML chars
    from strings, and convert non-finite floats (inf/nan) to 0.0 — both cause
    Excel's 'found a problem with content' recovery dialog otherwise."""
    if isinstance(v, str):
        return _ILLEGAL_XML_RE.sub("", v)
    if isinstance(v, float) and v != v:  # NaN
        return 0.0
    if isinstance(v, float) and v in (float("inf"), float("-inf")):
        return 0.0
    return v


def _write_cell(ws, row: int, column: int, value):
    """Create a cell with a sanitized value. Forces text data_type for any
    string so a value starting with '=' is NOT interpreted as a formula —
    openpyxl auto-treats leading-'=' strings as formulas, which Excel then
    strips ('Removed Records: Formula') and flags the workbook as corrupt."""
    v = _safe_cell(value)
    cell = ws.cell(row=row, column=column, value=v)
    if isinstance(v, str) and cell.data_type == "f":
        cell.data_type = "s"
    return cell


# ─── QBO auth + API helpers moved to shared/qbo_api.py (2026-07-13) ───
# load_credentials/_api_get/query_all/report/extract_proj/
# build_project_customer_map/fetch_project_pl/_walk_pl_rows/
# extract_pl_totals/fetch_customer_invoices are imported at the top.


# ────────── invoice / bill pulls ──────────

# ── legacy-job attribution (the user 2026-08-24) ──────────────────────
# Jobs that predate consistent project coding carry only PART of their cost on
# the project customer; the rest is named in the line description or the bill
# memo. `--legacy` installs a JobMatcher here for ONE job and every cost-line
# test in this file goes through `_line_belongs`. It is scoped by customer id,
# so any other project in the same run falls straight back to the strict test —
# behaviour for modern jobs is unchanged. See shared/job_lines.py.
_LEGACY_MATCH: Optional[JobMatcher] = None


def _set_legacy_matcher(proj: str, customer_id: str, legacy: bool,
                        aliases: Optional[List[str]] = None,
                        job_class: str = "", text_rules: bool = True,
                        class_ids=()) -> None:
    """Install (or clear) the legacy matcher for ONE project. Called per
    project so a batch can never leak one job's aliases into the next."""
    global _LEGACY_MATCH
    _LEGACY_MATCH = (JobMatcher(customer_id, proj, aliases or (), legacy=True,
                                class_prefix=job_class, text_rules=text_rules,
                                class_ids=class_ids)
                     if legacy else None)


def _line_belongs(det: dict, ln: dict, txn: dict, customer_id: str) -> bool:
    if _LEGACY_MATCH is not None and _LEGACY_MATCH.customer_id == customer_id:
        return _LEGACY_MATCH(det, ln, txn)
    return (det.get("CustomerRef") or {}).get("value") == customer_id


def fetch_customer_bills_and_purchases(
    access: str, company_id: str, customer_id: str,
    start_date: str, end_date: str,
) -> Tuple[List[dict], List[dict]]:
    """
    Fetch bills + purchases that touch this customer within the date range.

    Approach:
      - Filter QBO query by TxnDate range (start_date to end_date) — restricts
        payload to project window so the API doesn't time out on huge result
        sets. Then filter in Python to lines referencing this customer.

    Returns (bills, purchases).
    """
    def has_customer_line(txn: dict) -> bool:
        for ln in txn.get("Line") or []:
            det = (
                ln.get("AccountBasedExpenseLineDetail")
                or ln.get("ItemBasedExpenseLineDetail")
                or {}
            )
            if det and _line_belongs(det, ln, txn, customer_id):
                return True
        return False

    where = f"TxnDate >= '{start_date}' AND TxnDate <= '{end_date}'"
    bills = [b for b in query_all(access, company_id, "Bill", where=where) if has_customer_line(b)]
    purchases = [p for p in query_all(access, company_id, "Purchase", where=where) if has_customer_line(p)]
    return bills, purchases


# Pre-mobilization buffer on the Bill/Purchase pull. 18 months: long enough to
# catch real early cost, short enough to keep the payload sane. It was 180 days
# and silently CLIPPED real cost — MFD281's first bill is 2023-12-08, eleven
# months before its first invoice, so $237.50 of pre-mobilization rental fell
# outside and the workbook disagreed with QBO (caught by
# one-offs/pnl_line_level_audit.py, 2026-08-27).
_BILL_LOOKBACK_DAYS = 548


def _txn_touches_job(txn: dict, customer_id: str) -> bool:
    """Does any expense line of this txn belong to the job? (window check only)"""
    for ln in txn.get("Line") or []:
        det = (ln.get("AccountBasedExpenseLineDetail")
               or ln.get("ItemBasedExpenseLineDetail"))
        if det and _line_belongs(det, ln, txn, customer_id):
            return True
    return False


def _synth_pl_totals(bills: List[dict], purchases: List[dict],
                     income_groups: Dict[str, dict], customer_id: str,
                     acct_type: Dict[str, str], item_account: Dict[str, str],
                     pl_end: str) -> Dict[str, float]:
    """Stand in for QBO's project P&L report on a legacy job, built from the
    attributed lines + the job's invoices so every figure ties to the same
    source as the rest of the workbook. COGS vs Expense split by account type,
    exactly as the Transactions sheet does."""
    cogs = exp = 0.0
    for txn in list(bills) + list(purchases):
        if (txn.get("TxnDate") or "") > pl_end:
            continue
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            aid = ((det.get("AccountRef") or {}).get("value")
                   or item_account.get((det.get("ItemRef") or {}).get("value")))
            if (acct_type.get(aid) or "").lower().startswith("cost of goods"):
                cogs += amt
            else:
                exp += amt
    income = sum(g.get("net_billed", 0.0) for k, g in income_groups.items()
                 if k != "__retainage")
    gp = income - cogs
    return {"income": round(income, 2), "cogs": round(cogs, 2),
            "gross_profit": round(gp, 2), "expenses": round(exp, 2),
            "net_ordinary_income": round(gp - exp, 2),
            "net_income": round(gp - exp, 2)}


AMBER = "BF8F00"           # partially paid — encodes state, not decoration


def _pay_state(balance, total=None):
    """(label, colour) for a payment state. PARTIAL carries the OPEN amount,
    because 'UNPAID' on an invoice with a few hundred dollars left of a
    six-figure total reads as a
    collection problem when it is a rounding tail (the user 2026-08-27)."""
    try:
        b = float(balance or 0)
    except (TypeError, ValueError):
        return None, None
    if b <= 0.005:
        return "PAID", "008000"
    try:
        t = float(total) if total is not None else None
    except (TypeError, ValueError):
        t = None
    if t is not None and b < t - 0.005:
        return f"PARTIAL — {b:,.2f} open", AMBER
    return "UNPAID", "C00000"


def _proj_line_total(txn: dict, customer_id: str) -> float:
    """Sum of a transaction's line amounts that reference this project."""
    tot = 0.0
    for ln in txn.get("Line") or []:
        det = (ln.get("AccountBasedExpenseLineDetail")
               or ln.get("ItemBasedExpenseLineDetail") or {})
        if det and _line_belongs(det, ln, txn, customer_id):
            tot += float(ln.get("Amount", 0) or 0)
    return tot


def fetch_bill_payments(access: str, company_id: str,
                        start_date: str, end_date: str) -> List[dict]:
    """BillPayment txns in the window = ACTUAL AP cash-out (the user 2026-06-26 — bill
    DATE is cost accrual; the PAYMENT date is when cash left). Each links to the
    Bill(s) it paid. Fetched company-wide then attributed to project bills."""
    where = f"TxnDate >= '{start_date}' AND TxnDate <= '{end_date}'"
    return query_all(access, company_id, "BillPayment", where=where)


def fetch_customer_payments(access: str, company_id: str,
                            start_date: str, end_date: str) -> List[dict]:
    """Payment txns in the window = ACTUAL AR cash-in (when the GC's money landed,
    not the invoice date). Each links to the Invoice(s) it paid."""
    where = f"TxnDate >= '{start_date}' AND TxnDate <= '{end_date}'"
    return query_all(access, company_id, "Payment", where=where)


def build_cashflow_events(bills: list, invoices: list, bill_payments: list,
                          customer_payments: list, customer_id: str) -> list:
    """Merge AP cash-out (BillPayment→Bill) and AR cash-in (Payment→Invoice) into
    one chronological event list with a running balance. AP is scaled to the
    project's SHARE of a multi-project bill. Returns events sorted by date (cash
    OUT before IN on the same day), each with a cumulative 'running' position.
    The minimum running value = the peak cash requirement ('how far in the hole')."""
    bill_by_id = {}
    for b in bills:
        bid = b.get("Id")
        if not bid:
            continue
        total = sum(float(l.get("Amount", 0) or 0) for l in b.get("Line") or [])
        proj = _proj_line_total(b, customer_id)
        bill_by_id[bid] = {
            "ratio": (proj / total) if total else 1.0,
            "doc": _xml_clean(b.get("DocNumber", "") or bid),
            "vendor": _xml_clean(((b.get("VendorRef") or {}).get("name") or "")),
        }
    inv_by_id = {inv.get("Id"): _xml_clean(inv.get("DocNumber", "") or inv.get("Id", ""))
                 for inv in invoices if inv.get("Id")}

    events = []
    for bp in bill_payments or []:
        date = bp.get("TxnDate", "")
        vend = _xml_clean(((bp.get("VendorRef") or {}).get("name") or ""))
        for ln in bp.get("Line") or []:
            for lt in ln.get("LinkedTxn") or []:
                if lt.get("TxnType") == "Bill" and lt.get("TxnId") in bill_by_id:
                    info = bill_by_id[lt["TxnId"]]
                    amt = float(ln.get("Amount", 0) or 0) * info["ratio"]
                    if abs(amt) < 0.005:
                        continue
                    events.append({"date": date, "kind": "out",
                                   "party": vend or info["vendor"], "ref": info["doc"],
                                   "doc_id": lt["TxnId"], "page": "bill",
                                   "amount": round(amt, 2)})
    for pm in customer_payments or []:
        date = pm.get("TxnDate", "")
        cust = _xml_clean(((pm.get("CustomerRef") or {}).get("name") or ""))
        for ln in pm.get("Line") or []:
            for lt in ln.get("LinkedTxn") or []:
                if lt.get("TxnType") == "Invoice" and lt.get("TxnId") in inv_by_id:
                    amt = float(ln.get("Amount", 0) or 0)
                    if abs(amt) < 0.005:
                        continue
                    events.append({"date": date, "kind": "in", "party": cust,
                                   "ref": inv_by_id[lt["TxnId"]], "doc_id": lt["TxnId"],
                                   "page": "invoice", "amount": round(amt, 2)})

    # Within a day, apply RECEIPTS before PAYMENTS (the user 2026-06-26: "peak isn't
    # peak if money comes in the same day"). Cash is fungible within a day, so the
    # deepest point can't be below the day's CLOSING balance — ordering in-before-
    # out makes min(running) equal the day-close trough, not a phantom intra-day dip.
    events.sort(key=lambda e: ((_parse_date(e["date"]) or dt.date.min),
                               0 if e["kind"] == "in" else 1))
    run = 0.0
    for e in events:
        run += (e["amount"] if e["kind"] == "in" else -e["amount"])
        e["running"] = round(run, 2)
    return events


def fetch_customer_purchase_orders(
    access: str, company_id: str, customer_id: str,
    start_date: str, end_date: str,
) -> List[dict]:
    """PurchaseOrders that touch this project (line CustomerRef). Same
    company-wide-then-filter approach as bills — but POs are far fewer."""
    def has_customer_line(txn: dict) -> bool:
        for ln in txn.get("Line") or []:
            det = (ln.get("ItemBasedExpenseLineDetail")
                   or ln.get("AccountBasedExpenseLineDetail") or {})
            if (det.get("CustomerRef") or {}).get("value") == customer_id:
                return True
        return False

    where = f"TxnDate >= '{start_date}' AND TxnDate <= '{end_date}'"
    return [p for p in query_all(access, company_id, "PurchaseOrder", where=where)
            if has_customer_line(p)]


def match_pos_to_bills(pos: List[dict], bills: List[dict],
                       customer_id: str) -> Tuple[list, list]:
    """Split POs into (unused, used). A PO is USED if any Bill links to it
    (Bill.LinkedTxn TxnType='PurchaseOrder') or its POStatus is Closed.
    Returns lists of dicts: {po_num, po_date, vendor, po_amt, bills:[{ref,date,amt}], status}.
    """
    # po_id -> [bill, ...] via the bills' LinkedTxn back-references
    po_bills: Dict[str, list] = {}
    for b in bills:
        for lt in b.get("LinkedTxn") or []:
            if lt.get("TxnType") == "PurchaseOrder" and lt.get("TxnId"):
                po_bills.setdefault(lt["TxnId"], []).append(b)

    unused, used = [], []
    for po in pos:
        pid = po.get("Id")
        matched = po_bills.get(pid, [])
        rec = {
            "id": pid,
            "po_num": _xml_clean(po.get("DocNumber", "") or pid or ""),
            "po_date": po.get("TxnDate", ""),
            "vendor": _xml_clean(((po.get("VendorRef") or {}).get("name") or "")),
            "po_amt": _proj_line_total(po, customer_id),
            "status": po.get("POStatus", ""),
            "bills": [{
                "ref": _xml_clean(b.get("DocNumber", "") or b.get("Id", "")),
                "date": b.get("TxnDate", ""),
                "amt": _proj_line_total(b, customer_id),
                "txn_id": b.get("Id", ""), "tx_type": "Bill",
            } for b in matched],
        }
        is_used = bool(matched) or (po.get("POStatus", "").lower() == "closed")
        (used if is_used else unused).append(rec)
    return unused, used


# ────────────────────────── draw period grouping ──────────────────────────

def parse_draw_period(text: str) -> Optional[Tuple[dt.date, dt.date]]:
    if not text:
        return None
    m = DRAW_PERIOD_RE.search(text)
    if not m:
        return None
    s = _parse_date(m.group(1))
    e = _parse_date(m.group(2))
    if s and e and s <= e:
        return (s, e)
    return None


# Lenient: catches a Period tag even when a date is mistyped (e.g. a 5-digit
# year "5/16/20206"). Years 1–6 digits, optional space after "Period:".
LENIENT_PERIOD_RE = re.compile(
    r"Period\s*:\s*(\d{1,2}/\d{1,2}/\d{1,6})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{1,6})",
    re.IGNORECASE,
)


def _suggest_date_fix(token: str, sibling_year: Optional[int]) -> Optional[dt.date]:
    """Best guess for a mistyped date token. Tries: the sibling date's year
    (period spans ~1 month, so same year), then trimming the year to its
    first/last 4 digits, then 2000+yy."""
    parts = token.split("/")
    if len(parts) != 3:
        return None
    mo, da, yr = parts
    cands = []
    if sibling_year:
        cands.append(str(sibling_year))
    if len(yr) >= 4:
        cands += [yr[:4], yr[-4:]]
    if len(yr) == 2:
        cands.append("20" + yr)
    for cy in cands:
        d = _parse_date(f"{mo}/{da}/{cy}")
        if d:
            return d
    return None


def _resolve_one_date(ref: str, which: str, bad: str,
                      suggestion: Optional[dt.date], memo: str,
                      interactive: bool) -> Optional[dt.date]:
    """Resolve a single mistyped period date. Prompts the user in the terminal
    when interactive; otherwise warns and skips (won't silently guess)."""
    if not interactive:
        print(f"      ⚠ Invoice {ref}: period {which} '{bad}' is not a valid "
              f"date — invoice SKIPPED (run in a terminal to correct it).")
        return None
    print(f"\n  ⚠ Invoice {ref}: the period {which} date '{bad}' is not a "
          f"valid date.")
    print(f"     Full memo: {memo}")
    sug = suggestion.strftime("%m/%d/%Y") if suggestion else None
    while True:
        if sug:
            ans = input(f"     Did you mean {sug}?  [Enter = yes / type the "
                        f"correct MM/DD/YYYY / s = skip]: ").strip()
        else:
            ans = input(f"     Type the correct {which} date MM/DD/YYYY "
                        f"(or s = skip): ").strip()
        if ans == "" and sug:
            return suggestion
        if ans.lower() in ("s", "skip"):
            return None
        d = _parse_date(ans)
        if d:
            return d
        print("     ✗ Not a valid date — use MM/DD/YYYY.")


def resolve_draw_period(memo: str, ref: str,
                        interactive: bool) -> Optional[Tuple[dt.date, dt.date]]:
    """Strict parse first; if a Period tag is present but a date is mistyped,
    ask the user to confirm/correct it (the user 2026-06-09) instead of silently
    dropping the invoice to 'untagged'."""
    p = parse_draw_period(memo)
    if p:
        return p
    m = LENIENT_PERIOD_RE.search(memo or "")
    if not m:
        return None  # genuinely no Period tag → untagged
    raw_s, raw_e = m.group(1), m.group(2)
    ds, de = _parse_date(raw_s), _parse_date(raw_e)
    if ds is None:
        ds = _resolve_one_date(ref, "START", raw_s,
                               _suggest_date_fix(raw_s, de.year if de else None),
                               memo, interactive)
    if de is None:
        de = _resolve_one_date(ref, "END", raw_e,
                               _suggest_date_fix(raw_e, ds.year if ds else None),
                               memo, interactive)
    if ds and de and ds <= de:
        if (raw_s and not _parse_date(raw_s)) or (raw_e and not _parse_date(raw_e)):
            print(f"     ✓ Using period {ds.strftime('%m/%d/%Y')} – "
                  f"{de.strftime('%m/%d/%Y')} for invoice {ref}.")
        return (ds, de)
    return None


def _parse_date(s: str) -> Optional[dt.date]:
    if not s:
        return None
    s = str(s).strip()
    # QBO returns dates as ISO 8601 (YYYY-MM-DD). Try that first since most
    # transaction dates come from QBO. Then the common US human formats.
    for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def draw_label(period: Tuple[dt.date, dt.date]) -> str:
    s, e = period
    return f"{s.strftime('%m/%d/%y')}–{e.strftime('%m/%d/%y')}"


def group_invoices_by_draw(invoices: List[dict],
                           interactive: bool = False) -> Dict[str, dict]:
    """
    Returns { draw_label: {period, gross_income, retainage_held, net_billed,
                            invoice_count, invoices: [{doc_num, date, memo, amount}, ...]} }
    Skips retainage-only invoices (pre-2026 method). For 2026+ method
    where retainage is a negative line in the draw invoice, TotalAmt
    is already net. When `interactive`, a mistyped Period date prompts the
    user to correct it instead of silently dropping the invoice.
    """
    groups: Dict[str, dict] = {}
    untagged = {"label": "Untagged (no Period in PrivateNote)",
                "gross_income": 0.0, "retainage_held": 0.0,
                "retainage_billed": 0.0,
                "net_billed": 0.0, "invoice_count": 0, "invoices": []}
    # Standalone "retainage not billed" invoices — kept out of draw income but
    # collected for the Retainage breakdown at the bottom of the Draws sheet
    # (the user 2026-06-09: don't drop them).
    retainage_block = {"invoices": [], "total": 0.0}
    # Standalone BILLED retainage invoices (a real retainage invoice with no
    # Period — "City Retainage" etc.) — these ARE billed; show them as their own
    # retainage section, NOT lumped into "untagged" (the user 2026-06-19).
    ret_billed_block = {"label": "Retainage billed",
                        "gross_income": 0.0, "retainage_held": 0.0,
                        "retainage_billed": 0.0,
                        "net_billed": 0.0, "invoice_count": 0, "invoices": []}

    # Draw periods normally come from the invoice memo's "(Period: …)" tag. If a
    # project carries NONE, derive one draw per CALENDAR MONTH (1st→last day) from
    # the invoice DATE and bucket by date — same template, periods from dates
    # instead of memos (the user 2026-06-23).
    def _month_period(d: dt.date) -> Tuple[dt.date, dt.date]:
        first = d.replace(day=1)
        nxt = (first.replace(year=first.year + 1, month=1) if first.month == 12
               else first.replace(month=first.month + 1))
        return (first, nxt - dt.timedelta(days=1))

    # VOIDED invoices are QBO noise, not draws (the user 2026-08-05, MFD192's
    # two voided $0 retainage invoices cluttering the untagged block): QBO
    # zeroes a voided invoice and prefixes its memo "Voided - ". Dropped
    # everywhere — a voided invoice never belongs on a P&L.
    invoices = [inv for inv in invoices
                if not (abs(float(inv.get("TotalAmt", 0) or 0)) < 0.005
                        and str(inv.get("PrivateNote") or "")
                        .strip().lower().startswith("voided"))]

    any_period = any(
        resolve_draw_period(inv.get("PrivateNote", "") or "",
                            inv.get("DocNumber", "") or "", False)
        for inv in invoices)
    if not any_period:
        print("      no Period tags found — grouping draws by invoice MONTH "
              "(1st→last day) from the invoice date")

    for inv in invoices:
        pn = inv.get("PrivateNote", "") or ""
        ref = inv.get("DocNumber", "") or inv.get("Id", "")
        total = float(inv.get("TotalAmt", 0) or 0)
        # ── walk lines: split WORK vs WITHHELD retainage vs BILLED retainage.
        #   NEGATIVE retainage line = withheld (held back); POSITIVE = billed.
        gross = 0.0
        retainage = 0.0          # withheld (held back), positive
        retainage_billed = 0.0   # billed retainage, positive
        line_items = []          # for the not-billed display breakdown
        text_parts = [pn]
        for ln in inv.get("Line") or []:
            if ln.get("DetailType") in ("SubTotalLineDetail",):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            desc = ln.get("Description") or ""
            item_name = ((ln.get("SalesItemLineDetail") or {})
                         .get("ItemRef", {}).get("name") or "")
            text_parts.append(desc); text_parts.append(item_name)
            is_ret = bool(RETAINAGE_RE.search(desc) or RETAINAGE_RE.search(item_name))
            if is_ret and amt < 0:
                retainage += -amt
            elif is_ret and amt > 0:
                retainage_billed += amt
            else:
                gross += amt
            if amt:
                line_items.append({"desc": _xml_clean(desc or item_name or "Retainage"),
                                   "amt": amt})
        full_text = " ".join(text_parts)

        # ── STANDALONE RETAINAGE INVOICE = no real work (purely retainage).
        #   The ONLY thing that decides billed vs not-billed is whether
        #   "not billed" appears ANYWHERE in the memo/descriptions (the user
        #   2026-06-19) — "City" or any other text is irrelevant.
        is_ret_inv = (abs(gross) <= 0.005 and total != 0
                      and (retainage > 0.005 or retainage_billed > 0.005
                           or RETAINAGE_RE.search(full_text)))
        if is_ret_inv:
            if NOT_BILLED_RE.search(full_text):       # NOT BILLED → excluded
                retainage_block["invoices"].append({
                    "doc_num": _xml_clean(ref), "date": inv.get("TxnDate", ""),
                    "memo": _xml_clean(pn), "amount": total,
                    "lines": line_items or [{"desc": "Retainage", "amt": total}]})
                retainage_block["total"] += total
            else:                                     # real BILLED retainage draw
                ret_billed_block["retainage_billed"] += (retainage_billed or total)
                ret_billed_block["net_billed"] += total
                ret_billed_block["invoice_count"] += 1
                ret_billed_block["invoices"].append({
                    "doc_num": _xml_clean(ref), "date": inv.get("TxnDate", ""),
                    "memo": _xml_clean(pn), "amount": total,
                    "gross": 0.0, "retainage": 0.0,
                    "retainage_billed": retainage_billed or total})
            continue

        period = resolve_draw_period(pn, ref, interactive)
        if period is None and not any_period:
            d = _parse_date(inv.get("TxnDate", ""))
            if d:
                period = _month_period(d)   # 1st→last of the invoice's month
        inv_entry = {
            "doc_num": _xml_clean(inv.get("DocNumber", "") or inv.get("Id", "")),
            "id": inv.get("Id", ""),          # for the QBO invoice deep link
            "balance": float(inv.get("Balance", 0) or 0),   # 0 = PAID
            "date": inv.get("TxnDate", ""),
            "memo": _xml_clean(pn),
            "amount": total,
            "gross": gross,
            "retainage": retainage,
            "retainage_billed": retainage_billed,
        }

        if period is None:
            untagged["gross_income"] += gross
            untagged["retainage_held"] += retainage
            untagged["retainage_billed"] += retainage_billed
            untagged["net_billed"] += total
            untagged["invoice_count"] += 1
            untagged["invoices"].append(inv_entry)
            continue

        lbl = draw_label(period)
        if lbl not in groups:
            groups[lbl] = {
                "period": period,
                "gross_income": 0.0,
                "retainage_held": 0.0,
                "retainage_billed": 0.0,
                "net_billed": 0.0,
                "invoice_count": 0,
                "invoices": [],
            }
        groups[lbl]["gross_income"] += gross
        groups[lbl]["retainage_held"] += retainage
        groups[lbl]["retainage_billed"] += retainage_billed
        groups[lbl]["net_billed"] += total
        groups[lbl]["invoice_count"] += 1
        groups[lbl]["invoices"].append(inv_entry)

    # Sort groups by period start date
    sorted_groups = dict(sorted(groups.items(), key=lambda kv: kv[1]["period"][0]))
    if ret_billed_block["invoice_count"] > 0:
        sorted_groups["__retainage_billed"] = ret_billed_block
    if untagged["invoice_count"] > 0:
        sorted_groups["__untagged"] = untagged
    if retainage_block["invoices"]:
        sorted_groups["__retainage"] = retainage_block
    return sorted_groups


# ────────────────────────── retainage from Balance Sheet ──────────────────────────

def fetch_retainage_held(
    access: str, company_id: str, customer_id: str,
    end_date: str,
    accounts: Optional[List[dict]] = None,
) -> float:
    """
    Find Retainage Receivable balance for this customer as of end_date.

    Approach:
      1. Look up the actual 'Retainage Receivable' Account by name.
      2. Pull BalanceSheet filtered by customer=customer_id and look for
         the matching account name in the rows (preferring the data row
         over any 'Total X' or 'Less: X' row).
      3. Return abs() — Retainage Receivable is an asset; sign convention
         varies between QBO report and underlying GL.

    Per [[project_retainage_je_history]]: BalanceSheet is correct source
    of truth post-2026 cleanup JE.
    """
    # Step 1: find the Retainage Receivable account by name
    if accounts is not None:
        accts = accounts
    else:
        try:
            accts = query_all(access, company_id, "Account")
        except RuntimeError:
            accts = []
    retainage_acct_name = None
    for a in accts:
        name = a.get("Name", "") or ""
        if re.search(r"retainage\s*receivable", name, re.IGNORECASE):
            retainage_acct_name = name
            break
    if not retainage_acct_name:
        for a in accts:
            if re.search(r"retainage", a.get("Name", "") or "", re.IGNORECASE):
                retainage_acct_name = a["Name"]
                break

    # Step 2: customer-filtered Balance Sheet
    try:
        data = report(access, company_id, "BalanceSheet", params={
            "end_date": end_date,
            "accounting_method": "Accrual",
            "customer": customer_id,
        })
    except RuntimeError:
        return 0.0

    rows = _walk_pl_rows(data)

    # First pass: exact account name match on a Data row (not a Total or Less)
    if retainage_acct_name:
        for label, amt, _depth, kind in rows:
            if amt is None or kind != "data":
                continue
            if label.strip().lower() == retainage_acct_name.strip().lower():
                return abs(amt)

    # Second pass: any data row whose label contains 'retainage'
    for label, amt, _depth, kind in rows:
        if amt is None or kind != "data":
            continue
        if re.search(r"retainage", label, re.IGNORECASE) and not label.lower().startswith("less"):
            return abs(amt)

    # Third pass: total/section row containing 'retainage' (excluding 'less')
    for label, amt, _depth, _kind in rows:
        if amt is None:
            continue
        if re.search(r"retainage", label, re.IGNORECASE) and not label.lower().startswith("less"):
            return abs(amt)

    return 0.0


# ────────────────────────── WIP master lookup ──────────────────────────

def load_wip_master(path: Path) -> Dict[str, dict]:
    """Return { project#: {revised_contract, revised_etc, status, ...} } from the
    WIP master workbook.

    Two schemas are understood (the user 2026-07-16 — auto-pull Contract/ETC):
      - NEW: the readers' 'Test-Master' tab (unified MFD+CP+RP) — headers
        'TOTAL CONTRACT PRICE' / 'ESTIMATED TOTAL COSTS' / 'STATUS'. When this
        tab exists it is the ONLY one read (the workbook's live tabs have their
        own layouts and must not shadow it).
      - LEGACY: any tab whose name contains 'wip' with the original
        'Original/Revised Contract Price' / 'Original/Revised ETC' headers.
    Both feed the same keys the P&L templates already consume via
    _num("contract_saved","revised_contract","original_contract","contract")."""
    out: Dict[str, dict] = {}
    if not path.exists():
        return out
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:
        return out
    # Prefer the unified Test-Master tab; else fall back to legacy 'wip' tabs.
    sheet_names = (["Test-Master"] if "Test-Master" in wb.sheetnames
                   else [s for s in wb.sheetnames if "wip" in s.lower()])
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # Find header row (look for "Project #")
        header_row = None
        col_idx_map: Dict[str, int] = {}
        for r_idx in range(1, min(10, ws.max_row + 1)):
            row_vals = [
                str(ws.cell(row=r_idx, column=c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)
            ]
            if "project #" in row_vals:
                header_row = r_idx
                for c, v in enumerate(row_vals, start=1):
                    col_idx_map[v] = c
                break
        if header_row is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            proj_cell = ws.cell(row=r, column=col_idx_map["project #"]).value
            if not proj_cell:
                continue
            proj = str(proj_cell).strip().upper()
            # Skip repeated header rows / band rows (appendix blocks re-emit
            # the header; a project # always contains a digit).
            if proj == "PROJECT #" or not any(ch.isdigit() for ch in proj):
                continue

            def get(*col_names):
                for col_name in col_names:
                    idx = col_idx_map.get(col_name.lower())
                    if idx is not None:
                        v = ws.cell(row=r, column=idx).value
                        if v not in (None, ""):
                            return v
                return None

            out[proj] = {
                "description": get("Project Description", "PROJECT NAME"),
                "customer_gc": get("Customer / GC"),
                "original_contract": get("Original Contract Price"),
                "change_orders": get("Change Orders", "APPROVED COs"),
                # Test-Master's TOTAL CONTRACT PRICE = base + approved COs — that
                # IS the revised contract, so it feeds the same precedence slot.
                "revised_contract": get("Revised Contract Price",
                                        "TOTAL CONTRACT PRICE"),
                "original_etc": get("Original ETC"),
                "revised_etc": get("Revised ETC", "ESTIMATED TOTAL COSTS"),
                "super": get("Super"),
                # 'Closed' drives the WIP close-out (% forced to 100%).
                "status": get("STATUS"),
            }
    # 2026-08-07: Test-Master was restructured (bonding-style: TYPE/BONDED/
    # PROFIT, NO STATUS column) outside the repo's readers. The per-division
    # 'Test - CP' / 'Test - RP' tabs still carry STATUS, so when Test-Master
    # yields none, overlay it from there — otherwise `active cp|rp` finds
    # nothing and Closed handling goes dark.
    if out and not any(v.get("status") for v in out.values()):
        for div_tab in ("Test - CP", "Test - RP"):
            if div_tab not in wb.sheetnames:
                continue
            ws = wb[div_tab]
            hdr_row = proj_idx = status_idx = None
            for r in range(1, 8):
                vals = {str(ws.cell(row=r, column=c).value or "").strip().upper(): c
                        for c in range(1, 12)}
                if "PROJECT #" in vals and "STATUS" in vals:
                    hdr_row, proj_idx, status_idx = r, vals["PROJECT #"], vals["STATUS"]
                    break
            if hdr_row is None:
                continue
            for r in range(hdr_row + 1, ws.max_row + 1):
                proj = str(ws.cell(row=r, column=proj_idx).value or "").strip().upper()
                st = ws.cell(row=r, column=status_idx).value
                if proj and any(ch.isdigit() for ch in proj) and st:
                    if proj in out:
                        out[proj]["status"] = st
                    else:                     # project only on the division tab
                        out[proj] = {"description": None, "status": st}
    return out


_MARK_SUFFIX_RE = re.compile(r"\s*\(\d+ files\)\s*$")


def read_back_ledger_marks(path: Path,
                           sheets: Tuple[str, ...] = ("Labor", "Concrete")
                           ) -> Dict[str, Dict[tuple, str]]:
    """Manual row FILLS from the prior workbook's Labor/Concrete ledgers, so
    they survive a re-sync (the user 2026-07-31: a GREEN row = the PM
    confirmed that bill — crucial once this runs on routine). Same contract as
    read_back_inputs: the file is regenerated, the human's input is not.

    Bill rows are written with NO fill, so any solid fill on a row with a
    DATE is the estimator's mark. Keyed by (bill #, date, vendor, amount) —
    the color itself is preserved verbatim, so other color conventions keep
    working too. Returns {sheet: {key: argb}}."""
    out: Dict[str, Dict[tuple, str]] = {}
    if not path.exists():
        return out
    try:
        wb = openpyxl.load_workbook(str(path))
    except Exception:
        return out
    for nm in sheets:
        if nm not in wb.sheetnames:
            continue
        ws = wb[nm]
        # Column positions come from the LEDGER HEADER ROW, not fixed indices
        # (2026-08-10: the ↗ column shifted everything right by one; header-
        # driven reading handles both the old and new layouts).
        cols = {}
        for row in ws.iter_rows(min_col=1, max_col=12):
            hdrs = {str(c.value or "").strip().upper(): c.column for c in row}
            if "QBO #" in hdrs and "DATE" in hdrs and "AMOUNT" in hdrs:
                cols = {"qbo": hdrs["QBO #"], "date": hdrs["DATE"],
                        "vend": hdrs.get("VENDOR"), "amt": hdrs["AMOUNT"]}
                break
        if not cols:
            continue
        marks: Dict[tuple, str] = {}
        for row in ws.iter_rows(min_col=1, max_col=max(cols.values())):
            a = row[cols["qbo"] - 1]
            b = row[cols["date"] - 1]
            c = row[cols["vend"] - 1]
            f = row[cols["amt"] - 1]
            # Only LEDGER BILL rows: the DATE cell must be an actual date.
            # Scoreboard code rows carry band fills by design — without this
            # check they'd read as phantom marks.
            if not (a.value and re.match(r"\d{4}-\d{2}-\d{2}$",
                                         str(b.value or "").strip())):
                continue
            argb = None
            for cell in row:
                fl = cell.fill
                if (fl is not None and fl.patternType == "solid"
                        and fl.fgColor is not None
                        and isinstance(fl.fgColor.rgb, str)):
                    rgb = fl.fgColor.rgb.upper()
                    if rgb not in ("00000000", "FFFFFFFF"):
                        argb = rgb
                        break
            if not argb:
                continue
            try:
                amt = round(float(f.value), 2)
            except (TypeError, ValueError):
                continue
            key = (_MARK_SUFFIX_RE.sub("", str(a.value)).strip(),
                   str(b.value).strip(), str(c.value or "").strip(), amt)
            marks[key] = argb
        if marks:
            out[nm] = marks
    wb.close()
    return out


def read_back_inputs(path: Path, sheet: str = "P&L") -> Dict[str, float]:
    """Read the user-typed yellow input cells (Contract Price / ETC) from a
    PREVIOUSLY generated P&L so a re-sync preserves them (the user 2026-06-23).
    Locates the cells by their label in column A; returns {} if not found."""
    out: Dict[str, float] = {}
    if not path.exists():
        return out
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception:
        return out
    if sheet not in wb.sheetnames:
        return out
    ws = wb[sheet]
    # New labels (2026-07-16 WIP v2) + legacy ones so old sheets still carry
    # their typed values forward. CO cost is a manual input until the CO
    # template grows a cost line.
    wanted = {"original contract price": "contract",
              "contract price": "contract",                    # legacy label
              "original etc (estimated total cost)": "etc",
              "estimated total cost (etc)": "etc",             # legacy label
              "co costs (estimated)": "co_cost"}
    for r in range(1, min(ws.max_row, 80) + 1):
        lbl = str(ws.cell(row=r, column=1).value or "").strip().lower()
        key = wanted.get(lbl)
        if not key:
            continue
        v = ws.cell(row=r, column=2).value
        try:
            if isinstance(v, (int, float)):
                out[key] = float(v)
            elif isinstance(v, str) and v.strip() and not v.startswith("="):
                out[key] = float(v.replace(",", "").replace("$", "").strip())
        except (TypeError, ValueError):
            pass
    return out


def safe_save(wb: Workbook, out_path: Path) -> Optional[Path]:
    """Write atomically and NEVER clobber a workbook that's open in Excel.
    Excel drops a `~$<name>` owner-lock file next to an open workbook; if that
    exists we skip and warn. Otherwise write to a temp file and os.replace() it
    in (atomic) so a crash can't leave a half-written file (the user 2026-06-23)."""
    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        print(f"    ⚠ {out_path.name} looks OPEN in Excel — skipped to avoid "
              f"overwriting it. Close it and re-run.")
        return None
    tmp = out_path.with_name(out_path.name + ".tmp")
    wb.save(str(tmp))
    # Rule 5b: never hand over a workbook that hasn't passed the corruption
    # check as its LAST step. Run it on the TEMP file — a file that would make
    # Excel offer to "repair" it must never reach the real path (the user
    # 2026-08-17, "same errors still producing"; MFD172 tripped it 2026-08-24).
    try:
        assert_clean(tmp)
    except Exception as e:
        print(f"    ✗ {out_path.name} failed the xlsx corruption check — "
              f"NOT written:\n      {e}")
        try:
            tmp.unlink()
        except OSError:
            pass
        return None
    try:
        os.replace(str(tmp), str(out_path))
    except OSError as e:
        print(f"    ⚠ could not replace {out_path.name} ({e}); saved to "
              f"{tmp.name} instead.")
        return tmp
    return out_path


# ────────────────────────── account parent rollup ──────────────────────────

def build_account_parent_map(accounts: List[dict]) -> Dict[str, str]:
    """
    Map account Id → TOP-LEVEL ancestor account name — the level-2 account
    directly under Cost of Goods Sold on the P&L (e.g. a bill line coded to
    'Job Materials:Concrete' rolls up to 'Job Materials'). Top-level
    accounts map to themselves. Used by the Draw Audit grouping.
    """
    by_id = {a.get("Id"): a for a in accounts if a.get("Id")}
    out: Dict[str, str] = {}
    for a in accounts:
        aid = a.get("Id")
        if not aid:
            continue
        node = a
        for _ in range(10):  # parent-chain loop guard
            pid = (node.get("ParentRef") or {}).get("value")
            if not pid or pid not in by_id:
                break
            node = by_id[pid]
        out[aid] = node.get("Name") or a.get("Name") or "(Other)"
    return out


def build_account_type_map(accounts: List[dict]) -> Dict[str, str]:
    """Map account Id → AccountType ('Cost of Goods Sold', 'Expense', …) so a
    bill line can be split COGS vs operating Expense (the user 2026-06-22)."""
    return {a.get("Id"): (a.get("AccountType") or "")
            for a in accounts if a.get("Id")}


def _order_sheets(wb: Workbook, order: List[str]) -> None:
    """Reorder worksheets to match `order`; unlisted sheets keep their order
    at the end (the user 2026-06-22 sheet ordering)."""
    idx = {name: i for i, name in enumerate(order)}
    wb._sheets.sort(key=lambda s: idx.get(s.title, len(order)))


def gather_transactions(
    income_groups: Dict[str, dict], bills: List[dict], purchases: List[dict],
    customer_id: str, parent_map: Dict[str, str],
    account_names: Optional[Dict[str, str]] = None,
    acct_type: Optional[Dict[str, str]] = None,
    item_account: Optional[Dict[str, str]] = None,
) -> dict:
    """
    Every project transaction, for the Transactions sheet (the user 2026-06-22 — show
    the PM where each P&L number comes from). Income split into BILLED INCOME /
    RETAINAGE WITHHELD / RETAINAGE BILLED per invoice; bills split COGS vs
    EXPENSE by account type and grouped by vendor (newest→old).
    """
    account_names = account_names or {}
    acct_type = acct_type or {}
    item_account = item_account or {}

    income = []  # {doc, date, memo, billed, withheld, billed_ret}
    for k, grp in income_groups.items():
        if k == "__retainage":
            # "retainage not billed" = accumulated retainage moved to Retainage
            # Receivable by journal entry. It is NOT billed income (the user 2026-07-02:
            # "it clearly says retainage not billed … paid by JE to move it to
            # retainage receivable"), so it stays OUT of the billed / withheld /
            # billed_ret sums entirely. It IS retainage owed, so the amount is
            # carried separately (not_billed_ret) — shown in its own block on the
            # Transactions sheet and rolled into Total Retainage on the P&L.
            for inv in grp.get("invoices", []):
                amt = float(inv.get("amount", 0) or 0)
                income.append({"doc": inv.get("doc_num", ""), "id": inv.get("id", ""),
                               "date": inv.get("date", ""),
                               "memo": inv.get("memo", ""), "billed": 0.0,
                               "withheld": 0.0, "billed_ret": 0.0,
                               "not_billed_ret": amt})
            continue
        for inv in grp.get("invoices", []):
            income.append({
                "doc": inv.get("doc_num", ""), "id": inv.get("id", ""),
                "date": inv.get("date", ""),
                "memo": inv.get("memo", ""),
                "balance": float(inv.get("balance", 0) or 0),
                "billed": float(inv.get("gross", 0) or 0),
                "withheld": float(inv.get("retainage", 0) or 0),
                "billed_ret": float(inv.get("retainage_billed", 0) or 0)})
    income.sort(key=lambda x: _parse_date(x.get("date", "")) or dt.date.min, reverse=True)

    cogs: Dict[str, list] = {}
    exp: Dict[str, list] = {}
    cogs_accounts: Dict[str, float] = {}   # account name -> total (for P&L)
    exp_accounts: Dict[str, float] = {}

    def take(txn, tx_type, vendor_field):
        # ONE ROW PER LINE ITEM (the user 2026-07-02): every cost line is shown even when
        # multiple lines share the same bill ref# — costs must be seen, never summed
        # away. Only this project's lines are kept (OTHER-project lines on the same
        # bill are excluded). COGS vs Expense is decided by the line's ACCOUNT TYPE —
        # for item-based lines (no AccountRef) we resolve the item's expense account
        # first, so operating expenses no longer fall into COGS. Account TOTALS (for
        # the P&L per-account rollup) still aggregate by account name.
        vendor = _xml_clean(((txn.get(vendor_field) or {}).get("name") or "(no vendor)").strip())
        ref = _xml_clean(str(txn.get("DocNumber") or ""))
        txn_id = txn.get("Id", "")
        date = txn.get("TxnDate", "")
        memo = _xml_clean((txn.get("PrivateNote") or "").strip())
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if abs(amt) < 0.005:
                continue
            aid = (det.get("AccountRef") or {}).get("value")
            if not aid:   # item-based line → resolve item's expense account
                aid = item_account.get((det.get("ItemRef") or {}).get("value"))
            name = _xml_clean(account_names.get(aid) or "(unclassified)")
            atype = (acct_type.get(aid) or "").lower()
            if ("goods sold" in atype) or ("cogs" in atype):
                is_cogs = True
            elif "expense" in atype:
                is_cogs = False
            else:
                is_cogs = True   # unknown → treat as job cost (COGS)
            rec = {"ref": ref, "txn_id": txn_id, "tx_type": tx_type,
                   "date": date, "desc": _xml_clean((ln.get("Description") or memo or "").strip()),
                   "memo": memo,   # bill PrivateNote — its own column on the sheet
                   "account": name, "amount": amt}
            if is_cogs:
                cogs.setdefault(vendor, []).append(rec)
                cogs_accounts[name] = cogs_accounts.get(name, 0.0) + amt
            else:
                exp.setdefault(vendor, []).append(rec)
                exp_accounts[name] = exp_accounts.get(name, 0.0) + amt

    for b in bills:
        take(b, "Bill", "VendorRef")
    for p in purchases:
        take(p, "Expense", "EntityRef")

    def vsort(d):  # vendor groups; lines newest→old, same-bill lines kept together
        out = {}
        for v in sorted(d, key=lambda v: -sum(r["amount"] for r in d[v])):
            out[v] = sorted(d[v], key=lambda r: (_parse_date(r.get("date", ""))
                            or dt.date.min, str(r.get("ref", ""))), reverse=True)
        return out

    return {
        "income": income, "cogs": vsort(cogs), "exp": vsort(exp),
        "cogs_accounts": dict(sorted(cogs_accounts.items(), key=lambda x: -x[1])),
        "exp_accounts": dict(sorted(exp_accounts.items(), key=lambda x: -x[1])),
        "tot": {
            "billed": sum(i["billed"] for i in income),
            "withheld": sum(i["withheld"] for i in income),
            "billed_ret": sum(i["billed_ret"] for i in income),
            "not_billed_ret": sum(i.get("not_billed_ret", 0.0) for i in income),
            "cogs": sum(r["amount"] for v in cogs.values() for r in v),
            "exp": sum(r["amount"] for v in exp.values() for r in v),
        },
    }


# ────────────────────────── cost bucketing by draw window ──────────────────────────

def bucket_costs_by_draw_window(
    bills: List[dict],
    purchases: List[dict],
    customer_id: str,
    draw_periods: List[Tuple[str, dt.date, dt.date]],
    parent_map: Dict[str, str],
    account_names: Optional[Dict[str, str]] = None,
) -> Dict[str, dict]:
    """
    Bucket bill/purchase lines into draw windows by TxnDate (bills are never
    period-tagged — invoices only). Sub bills INCLUDED ("sub" tag = Sub LOC
    tracker, not P&L exclusion).

    Returns, per draw label:
      {"total": float,
       "groups": { parent_account: {
           "total": float,
           "subs": { leaf_account: {
               "total": float,
               "vendors": { vendor: {"total": float, "txns": [...] } } } } } } }

    parent → leaf (sub-account) → vendor → transactions, per the user 2026-06-05:
    "I need the SUB account listed... and group by vendor."
    Lines outside every window land under the special key "__outside".
    """
    account_names = account_names or {}
    out: Dict[str, dict] = {
        lbl: {"total": 0.0, "groups": {}} for lbl, _, _ in draw_periods
    }
    outside: dict = {"total": 0.0, "groups": {}}
    # Anchor at the FIRST period-tagged draw (the user 2026-07-16): history
    # dated before it is the pre-period-tagging era ("us not doing the new
    # process in the past") — it must NOT pour into the accumulating/next-draw
    # view and cloud the P&L. Disregarded here means "not in the draw views";
    # the P&L ② totals and the Transactions sheet still carry every dollar.
    anchor = min((s for _, s, _ in draw_periods), default=None)
    disregarded: dict = {"total": 0.0, "count": 0,
                         "anchor": anchor.isoformat() if anchor else None}

    def bucket_for(txn_date: Optional[dt.date]) -> Optional[dict]:
        if txn_date:
            for lbl, s, e in draw_periods:
                if s <= txn_date <= e:
                    return out[lbl]
            if anchor and txn_date < anchor:
                return None                    # pre-period history — disregard
        return outside

    def assign(txn: dict, tx_type: str, vendor_field: str) -> None:
        target = bucket_for(_parse_date(txn.get("TxnDate", "")))
        if target is None:                     # pre-anchor history — count, skip
            for ln in txn.get("Line") or []:
                det = (ln.get("AccountBasedExpenseLineDetail")
                       or ln.get("ItemBasedExpenseLineDetail") or {})
                if det and _line_belongs(det, ln, txn, customer_id):
                    disregarded["total"] += float(ln.get("Amount", 0) or 0)
                    disregarded["count"] += 1
            return
        vendor = _xml_clean(((txn.get(vendor_field) or {}).get("name") or "(no vendor)").strip())
        doc_num = _xml_clean(str(txn.get("DocNumber") or txn.get("Id") or ""))
        memo = _xml_clean((txn.get("PrivateNote") or "").strip())
        for ln in txn.get("Line") or []:
            det = (
                ln.get("AccountBasedExpenseLineDetail")
                or ln.get("ItemBasedExpenseLineDetail")
                or {}
            )
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if amt == 0:
                continue
            aref = det.get("AccountRef") or {}
            aid = aref.get("value")
            parent = _xml_clean(
                parent_map.get(aid)
                or aref.get("name")
                or (det.get("ItemRef") or {}).get("name")
                or "(Other)"
            )
            # Leaf = the cost code (item name) or account — via the ONE shared
            # resolver so this ties to costs_by_code / Budget vs Actual. Cost
            # codes (CS1, SL6, ...) kept VERBATIM; bolded at render (the user
            # 2026-06-09).
            leaf = cost_leaf(det, account_names, fallback=parent)
            pg = target["groups"].setdefault(parent, {"total": 0.0, "subs": {}})
            lg = pg["subs"].setdefault(leaf, {"total": 0.0, "vendors": {}})
            vg = lg["vendors"].setdefault(vendor, {"total": 0.0, "txns": []})
            vg["txns"].append({
                "doc_num": doc_num,
                "txn_id": txn.get("Id", ""),     # for the QBO bill/expense link
                "date": txn.get("TxnDate", ""),
                "tx_type": tx_type,
                "vendor": vendor,
                "desc": _xml_clean((ln.get("Description") or memo or "").strip()),
                "amount": amt,
            })
            vg["total"] += amt
            lg["total"] += amt
            pg["total"] += amt
            target["total"] += amt

    for b in bills:
        assign(b, "Bill", "VendorRef")
    for p in purchases:
        assign(p, "Expense", "EntityRef")

    if outside["total"] or outside["groups"]:
        out["__outside"] = outside
    if disregarded["count"]:
        out["__disregarded"] = disregarded
    return out


def gather_rp_costs(
    bills: List[dict],
    purchases: List[dict],
    customer_id: str,
    invoice_date: Optional[dt.date],
    parent_map: Dict[str, str],
    account_names: Optional[Dict[str, str]] = None,
    account_fqn: Optional[Dict[str, str]] = None,
    item_account: Optional[Dict[str, str]] = None,
) -> Tuple[dict, float, List[dict], List[dict]]:
    """
    RESIDENTIAL (RP) cost gathering — no draw windows (the user 2026-06-09).
    Every project bill/purchase line is either a JOB COST or goes to PENDING
    REVIEW:
      • line dated ON/BEFORE the invoice → job cost
      • line dated AFTER the invoice and the description says "wreck"
        (wrecking/demo labor) → job cost, flagged as wreck
      • line dated AFTER the invoice, NOT wreck → PENDING (likely an error)
      • zero/negative amount → PENDING
      • no account / no cost code → PENDING
    Duplicate bill ref #s (same ref on >1 transaction) are surfaced as an
    informational FLAGS list (still counted in costs — verify).

    Returns (job_groups, job_total, pending, dup_flags). job_groups uses the
    same parent→sub→vendor→txns shape as bucket_costs_by_draw_window so the
    existing nested writer can render it.
    """
    account_names = account_names or {}
    job = {"total": 0.0, "groups": {}}
    pending: List[dict] = []
    ref_ids: Dict[str, set] = {}          # ref# -> set of distinct txn Ids
    ref_rows: Dict[str, dict] = {}        # txn Id -> summary row (for dup list)

    def handle(txn: dict, tx_type: str, vendor_field: str) -> None:
        vendor = _xml_clean(((txn.get(vendor_field) or {}).get("name")
                             or "(no vendor)").strip())
        doc_num = _xml_clean(str(txn.get("DocNumber") or txn.get("Id") or ""))
        txn_id = str(txn.get("Id") or doc_num)
        memo = _xml_clean((txn.get("PrivateNote") or "").strip())
        tdate = _parse_date(txn.get("TxnDate", ""))
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if (det.get("CustomerRef") or {}).get("value") != customer_id:
                continue
            amt = float(ln.get("Amount", 0) or 0)
            desc = _xml_clean((ln.get("Description") or memo or "").strip())
            aref = det.get("AccountRef") or {}
            item_id = (det.get("ItemRef") or {}).get("value")
            has_item = bool((det.get("ItemRef") or {}).get("name") or item_id)
            # Resolve item-based lines to their EXPENSE ACCOUNT, then read the
            # ACCOUNT name (not the cost-code item) — the user 2026-06-26. parent = the
            # immediate parent account, leaf = the sub-account, both from the FQN.
            aid = aref.get("value") or (item_account or {}).get(item_id)
            fqn = _xml_clean((account_fqn or {}).get(aid) or account_names.get(aid)
                             or aref.get("name")
                             or (det.get("ItemRef") or {}).get("name") or "(Other)")
            segs = [s.strip() for s in fqn.split(":") if s.strip()]
            leaf = segs[-1] if segs else fqn
            parent = (segs[-2] if len(segs) >= 2
                      else (_xml_clean(parent_map.get(aid) or "") or leaf))
            rec = {"doc_num": doc_num, "date": txn.get("TxnDate", ""),
                   "tx_type": tx_type, "vendor": vendor, "desc": desc,
                   "amount": amt, "txn_id": txn_id}
            is_wreck = bool(invoice_date and tdate and tdate > invoice_date
                            and WRECK_RE.search(desc))
            # A late bill still BELONGS to the job when (a) it's a sub bill (subs
            # bill with a delay) or (b) its memo work PERIOD starts on/before the
            # invoice (work predates the invoice) — the user 2026-06-19.
            text = f"{memo} {desc}"
            is_sub = bool(SUB_RE.search(text))
            pm = PERIOD_ANYWHERE_RE.search(text)
            period_before = False
            if pm:
                pstart = _parse_date(pm.group(1))
                period_before = bool(pstart and invoice_date and pstart <= invoice_date)
            late = bool(invoice_date and tdate and tdate > invoice_date)
            # ── PENDING REVIEW filters (excluded from job cost) ──
            reason = None
            if amt <= 0:
                reason = "zero or negative amount"
            elif not aid and not has_item:
                reason = "uncategorized — no account / cost code"
            elif late and not is_wreck and not is_sub and not period_before:
                reason = "billed after the invoice, not wreck labor"
            if reason:
                pending.append({**rec, "reason": reason})
                continue
            # ── job cost ──
            rec["wreck"] = is_wreck
            pg = job["groups"].setdefault(parent, {"total": 0.0, "subs": {}})
            lg = pg["subs"].setdefault(leaf, {"total": 0.0, "vendors": {}})
            vg = lg["vendors"].setdefault(vendor, {"total": 0.0, "txns": []})
            vg["txns"].append(rec)
            vg["total"] += amt
            lg["total"] += amt
            pg["total"] += amt
            job["total"] += amt
            # track for duplicate-ref detection
            if doc_num:
                ref_ids.setdefault(doc_num, set()).add(txn_id)
                row = ref_rows.setdefault(txn_id, {
                    "doc_num": doc_num, "date": txn.get("TxnDate", ""),
                    "vendor": vendor, "amount": 0.0,
                    "txn_id": txn_id, "tx_type": tx_type, "desc": desc})
                row["amount"] += amt

    for b in bills:
        handle(b, "Bill", "VendorRef")
    for p in purchases:
        handle(p, "Expense", "EntityRef")

    dup_flags: List[dict] = []
    for ref, ids in ref_ids.items():
        if len(ids) > 1:
            for tid in ids:
                if tid in ref_rows:
                    dup_flags.append({**ref_rows[tid],
                                      "reason": f"ref # {ref} on {len(ids)} bills"})
    # return the GROUPS dict (parent→sub→vendor→txns), not the wrapper
    return job["groups"], job["total"], pending, dup_flags


def inject_pending_pos(
    job_groups: dict, pos: List[dict], unused_ids: set, customer_id: str,
    invoice_date: Optional[dt.date], parent_map: Dict[str, str],
    account_names: Optional[Dict[str, str]] = None,
    account_fqn: Optional[Dict[str, str]] = None,
    item_account: Optional[Dict[str, str]] = None,
) -> Tuple[float, int]:
    """
    An OPEN (unused) PO dated ON/BEFORE the invoice = a committed cost we're
    still WAITING for the bill on (the user 2026-06-19). Add those PO lines into the
    job-cost groups as `po_pending` transactions (rendered YELLOW), so the cost
    shows in the P&L even though the final bill hasn't arrived. Returns
    (pending_total, count). Mutates job_groups in place.
    """
    account_names = account_names or {}
    total, count = 0.0, 0
    for po in pos:
        if po.get("Id") not in unused_ids:
            continue
        pdate = _parse_date(po.get("TxnDate", ""))
        if invoice_date and pdate and pdate > invoice_date:
            continue                      # open PO after invoice = future buffer
        vendor = _xml_clean(((po.get("VendorRef") or {}).get("name") or "(no vendor)").strip())
        po_num = _xml_clean(po.get("DocNumber", "") or po.get("Id", ""))
        for ln in po.get("Line") or []:
            det = (ln.get("ItemBasedExpenseLineDetail")
                   or ln.get("AccountBasedExpenseLineDetail") or {})
            if (det.get("CustomerRef") or {}).get("value") != customer_id:
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if amt == 0:
                continue
            aref = det.get("AccountRef") or {}
            item_id = (det.get("ItemRef") or {}).get("value")
            aid = aref.get("value") or (item_account or {}).get(item_id)
            fqn = _xml_clean((account_fqn or {}).get(aid) or account_names.get(aid)
                             or aref.get("name")
                             or (det.get("ItemRef") or {}).get("name") or "(Other)")
            segs = [s.strip() for s in fqn.split(":") if s.strip()]
            leaf = segs[-1] if segs else fqn
            parent = (segs[-2] if len(segs) >= 2
                      else (_xml_clean(parent_map.get(aid) or "") or leaf))
            pg = job_groups.setdefault(parent, {"total": 0.0, "subs": {}})
            lg = pg["subs"].setdefault(leaf, {"total": 0.0, "vendors": {}})
            vg = lg["vendors"].setdefault(vendor, {"total": 0.0, "txns": []})
            vg["txns"].append({
                "doc_num": f"PO #{po_num}", "date": po.get("TxnDate", ""),
                "tx_type": "PO", "vendor": vendor,
                "desc": (_xml_clean((ln.get("Description") or "").strip())
                         or "bill pending"),
                "amount": amt, "po_pending": True})
            vg["total"] += amt; lg["total"] += amt; pg["total"] += amt
            total += amt; count += 1
    return total, count


# ────────────────────────── Excel builders ──────────────────────────

def _project_division(proj: str) -> str:
    if proj.startswith("MFD"):
        return "Multifamily (MFD)"
    if proj.startswith("CP"):
        return "Commercial (CP)"
    if proj.startswith("RP"):
        return "Residential (RP)"
    return "Unknown"


def _expected_class(proj: str) -> str:
    """The QBO Class a project's cost lines SHOULD carry (spelled out — the
    class-reconciliation check; see the QBO gotcha in CLAUDE.md)."""
    p = (proj or "").upper()
    if p.startswith("MFD"):
        return "Multi Family"
    if p.startswith("CP"):
        return "Commercial"
    if p.startswith("RP"):
        return "Residential"
    return ""


def _normalize_class(s: str) -> str:
    """Collapse a QBO class name to its top segment, lower/space-normalized,
    so 'Commercial:Sitework ' == 'commercial' (the user 2026-07-17)."""
    return re.sub(r"\s+", " ", (s or "").split(":")[0].strip().lower())


def _class_ok(actual: str, expected: str) -> bool:
    """True when the line's class matches the project's expected division class.
    A blank class is NOT ok — it's flagged so it gets a real class."""
    return bool(actual) and _normalize_class(actual) == _normalize_class(expected)


def _write_meta_block(ws, proj: str, cust_info: dict, wip_info: dict,
                      as_of: str, start_row: int = 1,
                      start_date: Optional[str] = None,
                      end_date: Optional[str] = None,
                      note: Optional[str] = None,
                      compact: bool = False) -> int:
    """
    Compact 2-line project header (title + one combined subtitle). Returns the
    first content row (start_row + 2 — no blank gap; the user 2026-06-09).
    """
    division = _project_division(proj)
    # Prefer the fully-qualified 'Customer:Project' name (it already contains
    # the project #), so we don't print the project number twice. Fall back
    # to GC name + project if needed.
    customer = _xml_clean(
        cust_info.get("fully_qualified_name", "")
        or (wip_info.get("customer_gc") if wip_info else None)
        or ""
    )
    if proj.upper() not in customer.upper():
        customer = f"{customer} : {proj}".strip(" :") if customer else proj

    title = f"PROJECT P&L — {customer}"
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=16, color="000000")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

    if compact:
        # ONE header row (the user 2026-08-31): the division line is noise on a
        # draw sheet and it costs a row of frozen height. "Generated" goes in
        # the cell straight after the merge — a merged cell can only carry one
        # format, so the grey italic needs its own cell.
        gen = ws.cell(row=start_row, column=9, value=f"Generated {as_of}")
        gen.font = Font(italic=True, size=BASE_SIZE, color="595959")
        gen.alignment = Alignment(vertical="center")
        return start_row + 1

    # Everything in ONE subtitle line (division · window · generated · note)
    sub_parts = [division]
    if start_date and end_date:
        sub_parts.append(f"Window {start_date} → {end_date}")
    sub_parts.append(f"Generated {as_of}")
    if note:
        sub_parts.append(note)
    sub_cell = ws.cell(row=start_row + 1, column=1, value="   ·   ".join(sub_parts))
    sub_cell.font = Font(italic=True, size=BASE_SIZE, color="595959")
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=8)

    return start_row + 2


# ── Sheet builders — PLAIN formatting (the user 2026-06-05: white sheet, black ──
# ── text; bold + indent + borders; color sparingly — navy headers, red ⚠) ──

TOP_BORDER = Border(top=Side(style="thin", color="000000"))
DOUBLE_TOP = Border(top=Side(style="double", color="000000"))
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="000000"))
_THICK = Side(style="thick", color="000000")
_MED = Side(style="medium", color="1F3A5F")
_HAIR = Side(style="hair", color="BFBFBF")
_THINB = Side(style="thin", color="808080")
NAVY = "1F3A5F"
RED = "C00000"
GREEN = "008000"
LINK = "0563C1"
# Light fill for the top number row of a group, so it pops (the user 2026-06-09).
ACCENT_FILL = PatternFill("solid", fgColor="DDEBF7")
# Section bands for P&L grouping (the user 2026-06-09, modeled on his example):
INCOME_FILL = PatternFill("solid", fgColor="C6E0B4")   # green  — income
COGS_FILL = PatternFill("solid", fgColor="FCE4D6")     # tan    — COGS
GP_FILL = PatternFill("solid", fgColor="FFF2CC")       # yellow — gross profit
SECT_FILL = PatternFill("solid", fgColor="E7E6E6")     # gray   — other sections


def _cov(net: float, costs: float, overhead_pct: float):
    """Return coverage / profit metrics for a draw.
      gross_profit   = Billed − Costs
      cost_pct       = Billed ÷ Costs              (Coverage %)
      net_profit     = Billed − Costs − overhead   (overhead = oh% × Billed)
      oh_pct         = Billed ÷ break-even         (Net Coverage %); break-even
                       = Costs ÷ (1 − oh%); net_profit ≥ 0 ⟺ oh_pct ≥ 100%.
    """
    needed = costs / (1 - overhead_pct / 100.0) if (costs and overhead_pct < 100) else costs
    return {
        "cost_dollar": net - costs,                       # Gross Profit
        "cost_pct": (net / costs) if costs else None,     # Coverage %
        "net_profit": net * (1 - overhead_pct / 100.0) - costs,  # Net Profit
        "oh_dollar": net - needed,
        "oh_pct": (net / needed) if needed else None,     # Net Coverage %
        "needed": needed,
    }


def build_sheet_transactions(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    tx: dict, as_of: str, realm: str = "", paid_map: Optional[dict] = None,
) -> Dict[str, str]:
    """
    TRANSACTIONS sheet — every invoice + bill behind the P&L (the user 2026-06-22:
    show the PM where each number comes from, no plain numbers). INCOME split
    into Billed income / Retainage withheld / Retainage billed; BILLS split COGS
    vs Expense, grouped by vendor newest→old. Returns sheet-qualified cell refs
    for the subtotals so the P&L can SUM-link to them.
    """
    SZ = BASE_SIZE - 1
    ws = wb.create_sheet("Transactions")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_properties.outlinePr.summaryBelow = False  # +/- sits on vendor row
    for col, w in (("A", 18), ("B", 15), ("C", 44), ("D", 18),
                   ("E", 18), ("F", 18), ("G", 18), ("H", 11)):
        ws.column_dimensions[col].width = w
    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    leg = ws.cell(row=r, column=1, value=(
        "Every invoice and bill behind the P&L. The P&L's Income, Retainage and "
        "COGS are SUM links to the totals on this sheet — click a P&L number to "
        "trace it here."))
    leg.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    def cell(rr, c, v, *, bold=False, fmt=None, color="000000", fill=None,
             border=None, size=SZ):
        if isinstance(v, str) and v.startswith("="):
            cc = ws.cell(row=rr, column=c, value=v)          # real formula
        elif isinstance(v, str):
            cc = _write_cell(ws, rr, c, v)                   # sanitized text
        else:
            cc = ws.cell(row=rr, column=c, value=v)
        cc.font = Font(bold=bold, size=size, color=color)
        if fmt:
            cc.number_format = fmt
        if fill is not None:
            cc.fill = fill
        if border is not None:
            cc.border = border
        return cc

    def ddate(rr, c, ds):
        dv = _parse_date(ds)
        x = ws.cell(row=rr, column=c, value=dv or ds)
        if dv:
            x.number_format = "mm/dd/yyyy"
        x.font = Font(size=SZ)
        x.alignment = Alignment(horizontal="left")  # tie date to its header

    def idcell(rr, c, s, *, indent=False):
        # numeric refs as real numbers → no "number stored as text" triangle;
        # left-aligned (with Excel indent) so they read like an ID, not a value.
        s = str(s or "")
        if s.isdigit() and len(s) <= 15:
            x = ws.cell(row=rr, column=c, value=int(s))
            x.number_format = "0"
        else:
            x = _write_cell(ws, rr, c, s)
        x.font = Font(size=SZ)
        x.alignment = Alignment(horizontal="left", indent=(1 if indent else 0))
        return x

    refs: Dict[str, str] = {}

    # ── INCOME ──
    # Gross → Retainage withheld → NET → Retainage billed (the user 2026-07-02): billed
    # income is GROSS; the PM needs the NET total of each invoice too. Net = gross −
    # retainage withheld + retainage billed (= the invoice TotalAmt). Invoices with
    # no retainage line simply show 0 in the retainage columns.
    # Purposeful color (the user 2026-07-15): navy section headers with white
    # text replace the pastel banners; categories get subtle tints below.
    SECTION_HDR = PatternFill("solid", fgColor=NAVY)
    CAT_FILLS = {"Concrete": PatternFill("solid", fgColor="E8EEF7"),   # light blue
                 "Labor": PatternFill("solid", fgColor="F7EFE1"),      # light tan
                 "Materials": PatternFill("solid", fgColor="F2F2F2")}  # light gray
    CAT_LABELS = {
        "Concrete": "CONCRETE",
        "Labor": "LABOR   (every line — subs paid weekly, never combined)",
        "Materials": "MATERIALS   (rebar · lumber · aggregates · pump · equipment)",
    }

    cell(r, 1, "INCOME  (invoices)", bold=True, size=BASE_SIZE, color="FFFFFF",
         fill=SECTION_HDR)
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = SECTION_HDR
    r += 1
    for c, h in ((1, "Inv #"), (2, "Date"), (3, "Memo"),
                 (4, "Gross income"), (5, "Retainage withheld"),
                 (6, "Net"), (7, "Retainage billed"), (8, "Paid?")):
        hc = cell(r, c, h, bold=True, color=NAVY); hc.border = BOTTOM_BORDER
    r += 1
    istart = r
    # retainage-not-billed rows are shown in their own block below the income
    # total (they are NOT billed income), so keep them out of the summed rows.
    nb_rows = [inv for inv in tx["income"] if inv.get("not_billed_ret")]
    for inv in tx["income"]:
        if inv.get("not_billed_ret"):
            continue
        ic = idcell(r, 1, inv.get("doc", ""))
        url = _qbo_txn_url("invoice", inv.get("id", ""), realm)
        if url:
            ic.hyperlink = url
            ic.font = Font(size=SZ, color=LINK, underline="single")
        ddate(r, 2, inv.get("date", ""))
        cell(r, 3, inv.get("memo", ""))
        cell(r, 4, float(inv.get("billed", 0) or 0), fmt=CURR_FMT)
        cell(r, 5, float(inv.get("withheld", 0) or 0), fmt=CURR_FMT, color="C0504D")
        cell(r, 6, f"=D{r}-E{r}+G{r}", bold=True, fmt=CURR_FMT)     # NET = TotalAmt
        cell(r, 7, float(inv.get("billed_ret", 0) or 0), fmt=CURR_FMT, color=GREEN)
        # AR payment state (the user 2026-08-05): Balance 0 = collected;
        # part-paid shows what is still open (the user 2026-08-27).
        _lbl, _col = _pay_state(inv.get("balance"),
                                float(inv.get("billed", 0) or 0)
                                + float(inv.get("billed_ret", 0) or 0)
                                - float(inv.get("withheld", 0) or 0))
        if _lbl:
            cell(r, 8, _lbl, bold=True, color=_col)
        r += 1
    cell(r, 1, "TOTAL INCOME", bold=True, border=TOP_BORDER)
    for c in (4, 5, 6, 7):
        t = cell(r, c, f"=SUM({get_column_letter(c)}{istart}:{get_column_letter(c)}{r-1})"
                 if r > istart else 0, bold=True, fmt=CURR_FMT, border=TOP_BORDER)
    refs["billed"] = f"Transactions!D{r}"
    refs["withheld"] = f"Transactions!E{r}"
    refs["net"] = f"Transactions!F{r}"
    refs["billed_ret"] = f"Transactions!G{r}"
    r += 2

    # ── RETAINAGE MOVED TO RECEIVABLE (not billed) ──
    # Shown for transparency but excluded from income — it's a JE move to
    # Retainage Receivable, not a draw (the user 2026-07-02). Rolled into Total
    # Retainage on the P&L via refs["not_billed_ret"].
    if nb_rows:
        cell(r, 1, "RETAINAGE MOVED TO RECEIVABLE  (not billed — excluded from income)",
             bold=True, size=SZ, color="FFFFFF", fill=SECTION_HDR)
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = SECTION_HDR
        r += 1
        nb_start = r
        for inv in nb_rows:
            ic = idcell(r, 1, inv.get("doc", ""))
            url = _qbo_txn_url("invoice", inv.get("id", ""), realm)
            if url:
                ic.hyperlink = url
                ic.font = Font(size=SZ, color=LINK, underline="single")
            ddate(r, 2, inv.get("date", ""))
            cell(r, 3, inv.get("memo", ""))
            cell(r, 5, float(inv.get("not_billed_ret", 0) or 0), fmt=CURR_FMT, color="C0504D")
            r += 1
        cell(r, 1, "Total retainage receivable (not billed)", bold=True, border=TOP_BORDER)
        cell(r, 5, f"=SUM(E{nb_start}:E{r-1})" if r > nb_start else 0,
             bold=True, fmt=CURR_FMT, border=TOP_BORDER, color="C0504D")
        refs["not_billed_ret"] = f"Transactions!E{r}"
        r += 2

    _known_words = _project_name_words(cust_info.get("name", ""))

    def _memo_text(ln):
        """Col C = bill memo (PrivateNote) first, line description appended when
        it adds info (the user 2026-07-15 — 'make sure to include the memo').
        Both are CLEANED of project #/address/GC-name noise (the user 2026-07-16
        — 'only what the actual item is'); sub bills keep their period dates."""
        memo = _clean_cost_text((ln.get("memo") or "").strip(), _known_words)
        desc = _clean_cost_text((ln.get("desc") or "").strip(), _known_words)
        if memo and desc and desc.lower() not in memo.lower():
            return f"{memo} — {desc}"
        return memo or desc

    def _vendor_lines(vendor, lines):
        """Vendor row (bold, SUM in E) + its line rows (collapsible). The
        vendor row carries NO account in col D, so the P&L's per-account
        SUMIF only ever matches the line rows. Returns the vendor row #."""
        nonlocal r
        vrow = r
        cell(r, 1, f"{vendor}  ({len(lines)})", bold=True)
        r += 1
        lstart = r
        for ln in lines:
            rc = idcell(r, 1, ln.get("ref", ""), indent=True)
            url = _qbo_txn_url(ln.get("tx_type", ""), ln.get("txn_id", ""), realm)
            if url:
                rc.hyperlink = url
                rc.font = Font(size=SZ, color=LINK, underline="single")
            ddate(r, 2, ln.get("date", ""))
            cell(r, 3, _memo_text(ln))
            cell(r, 4, ln.get("account", ""))
            cell(r, 5, float(ln.get("amount", 0) or 0), fmt=CURR_FMT)
            if paid_map is not None:
                _pd = paid_map.get(ln.get("txn_id"))
                if _pd is not None:
                    _lbl, _col = _pay_state(_pd[0], _pd[1])
                    if _lbl:
                        cell(r, 6, _lbl, bold=True, color=_col)
            ws.row_dimensions[r].outline_level = 1   # collapsible, open default
            r += 1
        vt = ws.cell(row=vrow, column=5,
                     value=f"=SUM(E{lstart}:E{r-1})" if r > lstart else 0)
        vt.number_format = CURR_FMT
        vt.font = Font(bold=True, size=SZ)
        return vrow

    def bills_block(title, groups, categories=False):
        """Render a bills section.

        categories=True (COGS): CONCRETE → LABOR → MATERIALS bands (biggest
        money movers first, the user 2026-07-15), vendors by size inside each,
        non-labor lines combined per (bill × account), labor never combined.
        TOTAL sums the category-subtotal cells; subtotals sum vendor rows;
        vendor rows sum line rows — each dollar counted exactly once.

        categories=False (EXPENSES): vendor grouping as before, same columns.
        """
        nonlocal r
        cell(r, 1, title, bold=True, size=BASE_SIZE, color="FFFFFF",
             fill=SECTION_HDR)
        for c in range(2, 7):                       # bills use cols A–F
            ws.cell(row=r, column=c).fill = SECTION_HDR
        r += 1
        for c, h in ((1, "Ref #"), (2, "Date"), (3, "Memo"),
                     (4, "Account"), (5, "Amount"), (6, "Paid?")):
            hc = cell(r, c, h, bold=True, color=NAVY); hc.border = BOTTOM_BORDER
        r += 1
        anchor_rows = []    # rows whose E cells the TOTAL row sums
        if not groups:
            cell(r, 1, "(none)", color="808080"); r += 1
        elif categories:
            # combine first (non-labor, per bill × account), then bucket by category
            by_cat: Dict[str, Dict[str, list]] = {}
            for vendor, lines in groups.items():
                for ln in combine_bill_lines(lines):
                    by_cat.setdefault(line_category(ln.get("account")), {}) \
                          .setdefault(vendor, []).append(ln)
            for cat in sorted(by_cat, key=lambda c_: CATEGORY_ORDER.get(c_, 99)):
                cfill = CAT_FILLS.get(cat)
                cell(r, 1, CAT_LABELS.get(cat, cat.upper()), bold=True,
                     color=NAVY, fill=cfill)
                for c in range(2, 6):
                    ws.cell(row=r, column=c).fill = cfill
                r += 1
                vmap = by_cat[cat]
                vend_rows = [
                    _vendor_lines(v, vmap[v])
                    for v in sorted(vmap, key=lambda v: -sum(
                        float(x.get("amount", 0) or 0) for x in vmap[v]))
                ]
                cell(r, 1, f"Subtotal — {cat}", bold=True, border=TOP_BORDER)
                cell(r, 5, "=" + "+".join(f"E{v}" for v in vend_rows),
                     bold=True, fmt=CURR_FMT, border=TOP_BORDER)
                anchor_rows.append(r)
                r += 1
        else:
            for vendor, lines in groups.items():
                anchor_rows.append(_vendor_lines(vendor, combine_bill_lines(lines)))
        cell(r, 1, "TOTAL " + title.split("—")[0].strip(), bold=True, border=TOP_BORDER)
        f = ("=" + "+".join(f"E{v}" for v in anchor_rows)) if anchor_rows else "=0"
        cell(r, 5, f, bold=True, fmt=CURR_FMT, border=TOP_BORDER)
        tot = r
        r += 2
        return tot

    # Simple full-width line separating INCOME (above) from COSTS (below) so the
    # wider income columns don't bleed into the narrower bill rows (the user 2026-07-02).
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=NAVY))
    r += 2
    cogs_tot = bills_block("COGS — JOB COSTS  (concrete → labor → materials)",
                           tx["cogs"], categories=True)
    refs["cogs"] = f"Transactions!E{cogs_tot}"
    exp_tot = bills_block("EXPENSES — bills (non-COGS)", tx["exp"])
    refs["exp"] = f"Transactions!E{exp_tot}"
    # for the P&L's per-account SUMIF mirror (account col D, amount col E)
    refs["acct_col"] = "Transactions!$D:$D"
    refs["amt_col"] = "Transactions!$E:$E"
    refs["cogs_accts"] = list(tx.get("cogs_accounts", {}).keys())
    refs["exp_accts"] = list(tx.get("exp_accounts", {}).keys())

    _setup_print(ws, 6)
    return refs


def build_sheet_by_account(wb: Workbook, proj: str, cust_info: dict,
                           wip_info: dict, tx: dict, as_of: str,
                           realm: str = "", paid_map: Optional[dict] = None
                           ) -> Dict[str, int]:
    """The SECOND view of the same money: ACCOUNT → VENDOR → the lines.

    Transactions groups by vendor, which answers "what did we buy from X".
    This answers "what is IN this account", which is the question you cannot
    otherwise ask — e.g. `*Job Material` carrying 3.1M on MFD133 (the user
    2026-08-27: "I need to audit this to make sure nothing is in there that
    shouldn't be"). Same records, same totals, pivoted.

    Returns {account name: row} so the P&L can link each of its account lines
    straight to the detail that adds up to it.
    """
    ws = wb.create_sheet("By Account")
    ws.sheet_view.showGridLines = False
    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    sub = ws.cell(row=r, column=1, value=(
        "COST BY ACCOUNT — every account, the vendors inside it, and every "
        "line. Same lines as Transactions, grouped the other way."))
    sub.font = Font(bold=True, size=BASE_SIZE, color="1F3A5F")
    r += 2

    SZ = BASE_SIZE
    anchors: Dict[str, int] = {}

    def cell(rr, cc, v, *, bold=False, fmt=None, color="000000",
             fill=None, indent=0, size=None):
        c = _write_cell(ws, rr, cc, v)
        c.font = Font(bold=bold, size=size or SZ, color=color)
        if fmt:
            c.number_format = fmt
        if fill is not None:
            c.fill = fill
        if indent:
            c.alignment = Alignment(indent=indent)
        return c

    for title, key, accts_key, fill in (
            ("COST OF GOODS SOLD", "cogs", "cogs_accounts", PatternFill("solid", fgColor="1F3A5F")),
            ("OPERATING EXPENSES (non-COGS)", "exp", "exp_accounts", PatternFill("solid", fgColor="7F6000"))):
        recs = [rec for v in (tx.get(key) or {}).values() for rec in v]
        if not recs:
            continue
        # account -> vendor -> lines. Vendor is recoverable because the source
        # dict is keyed by it.
        by_acct: Dict[str, Dict[str, list]] = {}
        for vendor, rows_ in (tx.get(key) or {}).items():
            for rec in rows_:
                by_acct.setdefault(rec["account"], {}).setdefault(vendor, []).append(rec)

        cell(r, 1, title, bold=True, color="FFFFFF", fill=fill)
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = fill
        cell(r, 5, sum(rec["amount"] for rec in recs), bold=True, fmt=CURR_FMT,
             color="FFFFFF", fill=fill)
        r += 1
        for c, h in ((1, "Account / Vendor / Line"), (2, "Date"), (3, "Doc #"),
                     (4, "Description"), (5, "Amount"), (6, "Paid?")):
            hc = cell(r, c, h, bold=True, color="1F3A5F")
            hc.border = BOTTOM_BORDER
        r += 1

        for acct in sorted(by_acct, key=lambda a: -sum(
                rec["amount"] for v in by_acct[a].values() for rec in v)):
            vendors = by_acct[acct]
            atot = sum(rec["amount"] for v in vendors.values() for rec in v)
            anchors[acct] = r
            cell(r, 1, acct, bold=True, color="1F3A5F")
            cell(r, 5, atot, bold=True, fmt=CURR_FMT, color="1F3A5F")
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="DDEBF7")
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1
            for vendor in sorted(vendors, key=lambda v: -sum(x["amount"] for x in vendors[v])):
                lines = vendors[vendor]
                cell(r, 1, vendor or "(no vendor)", bold=True, indent=1)
                cell(r, 5, sum(x["amount"] for x in lines), bold=True, fmt=CURR_FMT)
                ws.row_dimensions[r].outline_level = 1
                r += 1
                for rec in lines:
                    idc = cell(r, 3, rec.get("ref") or rec.get("txn_id") or "")
                    u = _qbo_txn_url(rec.get("tx_type", "Bill"), rec.get("txn_id", ""), realm)
                    if u:
                        idc.hyperlink = u
                        idc.font = Font(size=SZ, color="0563C1", underline="single")
                    _d = _parse_date(rec.get("date", ""))
                    dc = _write_cell(ws, r, 2, _d or rec.get("date", ""))
                    dc.number_format = "mm/dd/yyyy"
                    dc.font = Font(size=SZ)
                    cell(r, 4, rec.get("desc", ""), indent=2)
                    cell(r, 5, rec["amount"], fmt=CURR_FMT)
                    if paid_map is not None:
                        _pd = paid_map.get(rec.get("txn_id"))
                        if _pd is not None:
                            _lbl, _col = _pay_state(_pd[0], _pd[1])
                            if _lbl:
                                cell(r, 6, _lbl, bold=True, color=_col)
                    ws.row_dimensions[r].outline_level = 2
                    r += 1
        r += 1

    for col, w in zip("ABCDEF", (58, 13, 16, 62, 18, 24)):
        ws.column_dimensions[col].width = w
    ws.sheet_properties.outlinePr.summaryBelow = False
    _setup_print(ws, 6)
    return anchors


def _find_amount_match(target: float, candidates: list, tol: float = 0.01) -> list:
    """Transactions whose amount ≈ |target| (the user 2026-06-26 — find what a recon gap
    equals, fast). De-duped, largest first."""
    target = round(abs(float(target)), 2)
    if target < 0.01 or not candidates:
        return []
    seen, out = set(), []
    for c in candidates:
        if abs(round(abs(float(c.get("amount", 0) or 0)), 2) - target) <= tol:
            key = (c.get("ref"), round(abs(float(c.get("amount", 0) or 0)), 2),
                   c.get("txn_id"))
            if key in seen:
                continue
            seen.add(key)
            out.append(c)
    return sorted(out, key=lambda c: -abs(float(c.get("amount", 0) or 0)))


def build_sheet_reconciliations(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    qbo_income: float, qbo_cogs: float, qbo_exp: float,
    tx_refs: Dict[str, str], as_of: str, has_retainage: bool = True,
    orphans: Optional[list] = None, reports_relpath: str = "rd-reports",
    mismatches: Optional[list] = None, candidates: Optional[list] = None,
    tx_totals: Optional[dict] = None, realm: str = "",
) -> None:
    """
    RECONCILIATIONS sheet (the user 2026-06-22) — the checks-and-balances. The P&L is
    built from the Transactions sheet; this sheet proves it ties to QuickBooks.
    COGS and Expenses are summed bill-by-bill, so any gap is a real
    coding/missing-bill error. Income gaps are usually retainage timing.
    """
    SZ = BASE_SIZE
    ws = wb.create_sheet("Reconciliations")
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 38), ("B", 22), ("C", 22), ("D", 16), ("E", 18), ("F", 30)):
        ws.column_dimensions[col].width = w
    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    leg = ws.cell(row=r, column=1, value=(
        "The P&L is built from the Transactions sheet. This sheet checks those "
        "totals against QuickBooks. COGS / Expenses are summed bill-by-bill — any "
        "difference is a real coding or missing-bill error to chase."))
    leg.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    status_row = r           # reserve the verdict line; fill after the table
    r += 2

    # header
    for c, h in ((1, "Line"), (2, "QuickBooks"), (3, "P&L (Transactions)"),
                 (4, "Difference"), (5, "Status")):
        hc = _write_cell(ws, r, c, h)
        hc.font = Font(bold=True, size=SZ, color=NAVY)
        hc.border = BOTTOM_BORDER
    r += 1

    def recon(label, qbo_val, tx_formula):
        nonlocal r
        _write_cell(ws, r, 1, label).font = Font(size=SZ)
        b = ws.cell(row=r, column=2, value=float(qbo_val or 0))
        b.number_format = CURR_FMT; b.font = Font(size=SZ)
        c = ws.cell(row=r, column=3, value=tx_formula)
        c.number_format = CURR_FMT; c.font = Font(size=SZ, bold=True)
        d = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        d.number_format = CURR_FMT; d.font = Font(size=SZ)
        e = ws.cell(row=r, column=5,
                    value=f'=IF(ABS(D{r})<1,"✓ ties","⚠ off by "&TEXT(D{r},"#,##0.00"))')
        e.font = Font(size=SZ, bold=True, color="C0504D")
        rr = r
        r += 1
        return rr

    # Income ties to QBO on the RETAINAGE-INCLUSIVE figure (= P&L ② "Income
    # (incl. retainage)" = Btot): QBO income picks up the "retainage not billed"
    # JE invoice, so the Transactions side must include it too or it falsely
    # flags off by the not-billed retainage (the user 2026-07-17).
    _tx_income = f"{tx_refs['billed']}+{tx_refs['billed_ret']}"
    if tx_refs.get("not_billed_ret"):
        _tx_income += f"+{tx_refs['not_billed_ret']}"
    inc_row = recon("Income (incl. retainage)", qbo_income, f"={_tx_income}")
    if has_retainage:
        note = ws.cell(row=r, column=1, value=(
            "    note: matches P&L ② Income (incl. retainage); a gap here means a "
            "billed/retainage invoice is miscoded or missing — see Transactions"))
        note.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
        r += 1
    cogs_row = recon("Cost of Goods Sold", qbo_cogs, f"={tx_refs['cogs']}")
    exp_row = recon("Operating Expenses (non-COGS)", qbo_exp, f"={tx_refs['exp']}")
    r += 1
    foot = ws.cell(row=r, column=1, value=(
        "If COGS or Expenses show ⚠, a bill is miscoded, missing, or double-"
        "entered — open the Transactions sheet and compare to QuickBooks."))
    foot.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")

    # ── verdict line (driven by the deterministic COGS + Expenses checks) ──
    sc = _write_cell(ws, status_row, 1, "RECONCILIATION STATUS")
    sc.font = Font(bold=True, size=BASE_SIZE - 1, color=NAVY)
    v = ws.cell(row=status_row, column=2,
                value=(f'=IF(AND(ABS(D{cogs_row})<1,ABS(D{exp_row})<1),'
                       f'"✓ RECONCILED — bills tie to QuickBooks",'
                       f'"⚠ REVIEW NEEDED — see flagged rows below")'))
    v.font = Font(bold=True, size=BASE_SIZE + 2, color="375623")
    v.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=status_row, start_column=2, end_row=status_row, end_column=5)
    ws.row_dimensions[status_row].height = 24

    # ── DIFFERENCE FINDER — for each gap, hunt a transaction equaling it (the user
    #    2026-06-26: "I don't know why it's off by that amount — find it fast"). ──
    if tx_totals:
        tx_income = ((tx_totals.get("billed", 0.0) or 0.0)
                     + (tx_totals.get("billed_ret", 0.0) or 0.0)
                     + (tx_totals.get("not_billed_ret", 0.0) or 0.0))
        gaps = [("Income (incl. retainage)", round(tx_income - (qbo_income or 0.0), 2)),
                ("Cost of Goods Sold", round((tx_totals.get("cogs", 0.0) or 0.0) - (qbo_cogs or 0.0), 2)),
                ("Operating Expenses", round((tx_totals.get("exp", 0.0) or 0.0) - (qbo_exp or 0.0), 2))]
        gaps = [(lbl, d) for lbl, d in gaps if abs(d) >= 1.0]
        if gaps:
            r += 2
            hc = _write_cell(ws, r, 1, "DIFFERENCE FINDER — transactions that equal each gap")
            hc.font = Font(bold=True, size=BASE_SIZE + 1, color="BF8F00")
            for c in range(2, 6):
                ws.cell(row=r, column=c).border = BOTTOM_BORDER
            hc.border = BOTTOM_BORDER
            r += 1
            for lbl, d in gaps:
                g = _write_cell(ws, r, 1, f"{lbl} — off by {d:,.2f}")
                g.font = Font(bold=True, size=SZ, color=RED)
                r += 1
                matches = _find_amount_match(d, candidates or [])
                if matches:
                    note = _write_cell(ws, r, 1, "    a single transaction equals this gap:")
                    note.font = Font(italic=True, size=SZ, color="595959"); r += 1
                    for m in matches[:10]:
                        rc = _write_cell(ws, r, 2, str(m.get("ref", "")) or "(no #)")
                        url = _qbo_txn_url(m.get("tx_type", "Bill"), m.get("txn_id", ""), realm)
                        if url:
                            rc.hyperlink = url
                            rc.font = Font(size=SZ, color=LINK, underline="single")
                        else:
                            rc.font = Font(size=SZ)
                        _write_cell(ws, r, 3, m.get("party", "")).font = Font(size=SZ)
                        ac = ws.cell(row=r, column=4, value=round(abs(float(m["amount"])), 2))
                        ac.number_format = CURR_FMT; ac.font = Font(size=SZ, bold=True)
                        _write_cell(ws, r, 5, ("invoice" if m.get("tx_type") == "invoice"
                                               else "bill")).font = Font(size=SZ, color="595959")
                        r += 1
                else:
                    note = _write_cell(ws, r, 1,
                        "    no single transaction equals this gap — likely a "
                        "combination of bills, a partial payment, or retainage timing.")
                    note.font = Font(italic=True, size=SZ, color="595959"); r += 1

    # ── AMOUNT MISMATCHES — bill # IS in QBO, amount differs (the user 2026-06-26) ──
    if mismatches:
        r += 2
        hc = _write_cell(ws, r, 1, "AMOUNT MISMATCHES — bill # in QBO, different $ (could be partial, tax, combined)")
        hc.font = Font(bold=True, size=BASE_SIZE + 1, color="BF8F00")
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = BOTTOM_BORDER
        hc.border = BOTTOM_BORDER
        r += 1
        for c, h in ((1, "Bill #"), (2, "Vendor"), (3, "Report $"),
                     (4, "QBO $"), (5, "Difference"), (6, "Source report")):
            x = _write_cell(ws, r, c, h)
            x.font = Font(bold=True, size=SZ, color=NAVY); x.border = BOTTOM_BORDER
        r += 1
        for o in sorted(mismatches, key=lambda x: -abs(float(x.get("diff", 0) or 0))):
            bc = _write_cell(ws, r, 1, str(o["num"]) or "(no #)")
            url = _qbo_txn_url(o.get("qbo_tx_type", "Bill"), o.get("qbo_txn_id", ""), realm)
            if url:
                bc.hyperlink = url
                bc.font = Font(size=SZ, color=LINK, underline="single")
            else:
                bc.font = Font(size=SZ)
            _write_cell(ws, r, 2, o.get("vendor", "")).font = Font(size=SZ)
            ra = ws.cell(row=r, column=3, value=float(o["amount"] or 0))
            ra.number_format = CURR_FMT; ra.font = Font(size=SZ)
            qa = ws.cell(row=r, column=4, value=float(o.get("qbo_amount", 0) or 0))
            qa.number_format = CURR_FMT; qa.font = Font(size=SZ)
            dv = ws.cell(row=r, column=5, value=float(o.get("diff", 0) or 0))
            dv.number_format = CURR_FMT; dv.font = Font(size=SZ, bold=True, color=RED)
            sc2 = _write_cell(ws, r, 6, o["source"])
            sc2.hyperlink = f"{reports_relpath}/{o['source']}"
            sc2.font = Font(size=SZ, color=LINK, underline="single")
            r += 1

    # ── ORPHAN PM-REPORT LINES — final catcher (the user 2026-06-26) ──
    # Costs a PM listed on a draw report that match NO QBO bill anywhere (typo,
    # wrong amount, or a cost never entered in QBO). Each links to its source file.
    if orphans:
        r += 2
        hc = _write_cell(ws, r, 1, "ORPHAN REPORT LINES — on a PM report, not in QuickBooks")
        hc.font = Font(bold=True, size=BASE_SIZE + 1, color=RED)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BOTTOM_BORDER
        hc.border = BOTTOM_BORDER
        r += 1
        for c, h in ((1, "Bill # / Vendor"), (2, "Date"), (3, "Description"),
                     (4, "Amount"), (5, "Source report")):
            x = _write_cell(ws, r, c, h)
            x.font = Font(bold=True, size=SZ, color=NAVY); x.border = BOTTOM_BORDER
        r += 1
        obyv = {}                                         # group by vendor (the user 2026-06-26)
        for o in orphans:
            obyv.setdefault(o.get("vendor") or "(no vendor)", []).append(o)
        for vend in sorted(obyv, key=lambda v: -sum(float(o["amount"] or 0) for o in obyv[v])):
            vit = obyv[vend]
            vh = _write_cell(ws, r, 1, f"{vend}  ({len(vit)})")
            vh.font = Font(bold=True, size=SZ, color=RED)
            vt = ws.cell(row=r, column=4, value=round(sum(float(o["amount"] or 0) for o in vit), 2))
            vt.number_format = CURR_FMT; vt.font = Font(bold=True, size=SZ, color=RED)
            r += 1
            for o in sorted(vit, key=lambda x: -float(x["amount"] or 0)):
                # bill # links to the source report (no QBO txn for an orphan)
                bc = _write_cell(ws, r, 1, "    " + (str(o["num"]) or "(no #)"))
                bc.hyperlink = f"{reports_relpath}/{o['source']}"
                bc.font = Font(size=SZ, color=LINK, underline="single")
                dv = _parse_date(o["date"]) if isinstance(o["date"], str) else o["date"]
                dc = ws.cell(row=r, column=2, value=dv or o["date"])
                if isinstance(dv, dt.date):
                    dc.number_format = "mm/dd/yyyy"
                dc.font = Font(size=SZ)
                _write_cell(ws, r, 3, o.get("desc", "")).font = Font(size=SZ)
                ac = ws.cell(row=r, column=4, value=float(o["amount"] or 0))
                ac.number_format = CURR_FMT; ac.font = Font(size=SZ, color=RED)
                sc2 = ws.cell(row=r, column=5, value=o["source"])
                sc2.hyperlink = f"{reports_relpath}/{o['source']}"
                sc2.font = Font(size=SZ, color=LINK, underline="single")
                r += 1
        tc = ws.cell(row=r, column=1, value="Total orphan lines")
        tc.font = Font(bold=True, size=SZ)
        tt = ws.cell(row=r, column=4,
                     value=round(sum(float(o["amount"] or 0) for o in orphans), 2))
        tt.number_format = CURR_FMT; tt.font = Font(bold=True, size=SZ, color=RED)
        for cc in range(1, 6):
            ws.cell(row=r, column=cc).border = TOP_BORDER

    _setup_print(ws, 6)


def build_sheet_pl(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    pl_data: dict, totals: Dict[str, float],
    net_billed: float, retainage_held: float,
    as_of: str, overhead_pct: float = 10.0,
    pl_cutoff: Optional[str] = None,
    accum: Optional[dict] = None,
    draw_rows: Optional[List[Tuple[str, str, float, float, float, float]]] = None,
    draw_anchors: Optional[Dict[str, int]] = None,
    retainage_nb: float = 0.0,
    retainage_billed_total: float = 0.0,
    show_retainage_block: bool = True,
    tx_refs: Optional[Dict[str, str]] = None,
    alt_overhead_pct: Optional[float] = None,
    underbill_total: float = 0.0,
    underbill_count: int = 0,
    income_rows: Optional[List[dict]] = None,
    simple: bool = False,
    acct_anchors: Optional[Dict[str, int]] = None,
    realm: str = "",
) -> None:
    """
    P&L sheet — every DERIVED figure is a live Excel FORMULA referencing its
    source cells (the user 2026-06-09), so the math is traceable: section subtotals
    are SUM() of their lines; Gross Profit = Income − COGS; NOI, Overhead,
    True Net Profit, the coverage columns, accumulating totals, Draw needed and
    Labor-paid are all formulas. Only the raw QBO line amounts are constants.
    """
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22

    oh = overhead_pct / 100.0          # e.g. 0.11
    one_minus_oh = 1 - oh              # e.g. 0.89

    meta_note = (f"P&L through {pl_cutoff} (last draw period end)"
                 if pl_cutoff else None)
    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of, note=meta_note)
    # Header shortcut (the user 2026-08-06): one "Open Project in QBO" link in the
    # free top-right header cell → the project HOME page (customerdetail), distinct
    # from the per-figure Billed/Costs links below. Stored hyperlink (never
    # =HYPERLINK(), which fails in Mac Excel).
    _home_url = _qbo_customer_url(cust_info.get("id", ""), realm)
    if _home_url:
        _hc = ws.cell(row=2, column=9, value="Open Project in QBO  ↗")
        _hc.hyperlink = _home_url
        _hc.font = Font(size=BASE_SIZE, color="0563C1", underline="single")

    def row(label: str, amt=None, *, formula=None, bold=False, indent=0,
            border=None, fmt=CURR_FMT, color="000000", fill=None,
            size: Optional[int] = None) -> int:
        """Write a label (col A) + value (col B). value is a number (amt) or a
        live formula (formula='=...'). Returns the row written (for refs)."""
        nonlocal r
        sz = size or BASE_SIZE
        used = r
        c = _write_cell(ws, r, 1, "    " * indent + label)
        c.font = Font(bold=bold, size=sz, color=color)
        a = None
        if formula is not None:
            a = ws.cell(row=r, column=2, value=formula)   # real Excel formula
            a.number_format = fmt
            a.font = Font(bold=bold, size=sz, color=color)
        elif amt is not None:
            a = _write_cell(ws, r, 2, amt)
            a.number_format = fmt
            a.font = Font(bold=bold, size=sz, color=color)
        if border is not None:
            c.border = border
            if a is not None:
                a.border = border
        if fill is not None:
            c.fill = fill
            (a or ws.cell(row=r, column=2)).fill = fill
        r += 1
        return used

    # Income EXCLUDES retainage not billed (the user 2026-06-09): QBO's income
    # picks up standalone "retainage not billed" invoices, but that revenue
    # isn't truly billed/earned for this view — strip it so it cascades out
    # of Gross Profit / NOI / TRUE NET PROFIT.
    income = totals.get("income", 0.0) - retainage_nb
    hero_fill = PatternFill("solid", fgColor="1F3A5F")

    # ════════════════════════════════════════════════════════════════════
    #  STORY LAYOUT (the user 2026-07-16): ① WIP (beginning) → ② P&L TOTALS
    #  incl. retainage (the true totals) → ③ Snapshot REALIZED net-billed
    #  (MFD vs Company) → ④ Billing & Retainage. Draw coverage table right.
    # ════════════════════════════════════════════════════════════════════
    ws.column_dimensions["C"].width = 3   # thin gap before the coverage table
    grp_fill = PatternFill("solid", fgColor="DDEBF7")
    YEL = PatternFill("solid", fgColor="FFE699")
    _THK = Side(style="medium", color=NAVY)
    # MFD gets the dual overhead view (MFD % on costs + Company % on revenue).
    # Non-MFD (CP) gets ONLY the company overhead — MFD is a different player and
    # stays out of CP entirely (the user 2026-07-02).
    show_mfd = alt_overhead_pct is not None
    _alt = alt_overhead_pct if alt_overhead_pct is not None else overhead_pct
    _aoh = _alt / 100.0

    # QBO deep links (the user 2026-07-13): Billed totals → the customer page
    # (all invoices on one screen); Costs totals → the project-filtered P&L
    # report. Attached to the value cells, underlined so they read clickable.
    _cust_url = _qbo_customer_url(cust_info.get("id", ""), realm)
    _costs_url = _qbo_project_pl_url(cust_info.get("id", ""), realm)

    def _qbo_link(rr: Optional[int], url: Optional[str]) -> None:
        if not rr or not url:
            return
        c = ws.cell(row=rr, column=2)
        c.hyperlink = url
        f = c.font
        c.font = Font(bold=f.bold, size=f.size or BASE_SIZE,
                      color=f.color, underline="single")

    def box(top_row, bot_row, c0=1, c1=2):
        for rr in range(top_row, bot_row + 1):
            for cc in range(c0, c1 + 1):
                cur = ws.cell(row=rr, column=cc).border
                ws.cell(row=rr, column=cc).border = Border(
                    left=_THK if cc == c0 else cur.left,
                    right=_THK if cc == c1 else cur.right,
                    top=_THK if rr == top_row else cur.top,
                    bottom=_THK if rr == bot_row else cur.bottom)

    def sect_title(txt):
        nonlocal r
        c = _write_cell(ws, r, 1, txt)
        c.font = Font(bold=True, size=BASE_SIZE + 1, color="FFFFFF")
        c.fill = hero_fill
        ws.cell(row=r, column=2).fill = hero_fill
        used = r
        r += 1
        return used

    def acct_lines(header, names, total_label, fill):
        """Header bar CARRIES the total (the user 2026-08-27) — the section
        reads total-first and the accounts detail it underneath, instead of
        making you scroll to a total row at the bottom. `total_label` is kept
        in the signature for callers but is no longer written as its own row."""
        nonlocal r
        hdr_row = row(header, None, bold=True, color=NAVY, fill=fill)
        rows_ = []
        for nm in names:
            esc = str(nm).replace('"', '""')
            _rr = row(nm, formula=f'=SUMIF({tx_refs["acct_col"]},"{esc}",{tx_refs["amt_col"]})', indent=1)
            rows_.append(_rr)
            # click the account name → the lines that add up to it
            _a = (acct_anchors or {}).get(nm)
            if _a:
                _c = ws.cell(row=_rr, column=1)
                _c.hyperlink = f"#'By Account'!A{_a}"
                _c.font = Font(size=BASE_SIZE, color="0563C1", underline="single")
        f = ("=" + "+".join(f"B{x}" for x in rows_)) if rows_ else "=0"
        tc = ws.cell(row=hdr_row, column=2, value=f)
        tc.number_format = CURR_FMT
        tc.font = Font(bold=True, size=BASE_SIZE, color=NAVY)
        tc.fill = fill
        return hdr_row

    def snapshot(title, basis_label, inc_expr, gp_expr, costs, opex):
        nonlocal r
        # SINGLE COLUMN (the user 2026-06-22): shared facts once, then MFD block, then
        # COMPANY block STACKED below it — no separate column C (kills the empty
        # space). Only Overhead / Net Profit / Profit % differ between the two.
        t0 = sect_title(title)
        _write_cell(ws, r, 1, basis_label).font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
        r += 1

        def line(label, f, pct=False, hero=False, fill=None):
            nonlocal r
            lc = _write_cell(ws, r, 1, label)
            b = ws.cell(row=r, column=2, value=f)
            b.number_format = PCT_FMT if pct else CURR_FMT
            if hero:
                lc.font = Font(bold=True, size=BASE_SIZE, color="FFFFFF"); lc.fill = hero_fill
                b.font = Font(bold=True, size=BASE_SIZE, color="C6EFCE"); b.fill = hero_fill
            else:
                lc.font = Font(bold=True, size=BASE_SIZE)
                b.font = Font(bold=True, size=BASE_SIZE)
                if fill is not None:
                    lc.fill = fill; b.fill = fill
            b.alignment = Alignment(horizontal="right")
            r += 1

        def subhdr(txt):
            nonlocal r
            hc = _write_cell(ws, r, 1, txt)
            hc.font = Font(bold=True, size=BASE_SIZE, color=NAVY)
            ws.cell(row=r, column=1).fill = grp_fill
            ws.cell(row=r, column=2).fill = grp_fill
            r += 1

        line("Income", f"={inc_expr}")
        line("Costs (COGS)", f"={costs}")
        line("Gross Profit", f"={gp_expr}")
        line("Markup on costs %", f'=IF({costs}=0,"",({gp_expr})/{costs})', pct=True)
        co_np = f"({gp_expr})-{opex}-{oh}*({inc_expr})"
        # MFD-only: the MFD (% on costs) view sits above the company view.
        if show_mfd:
            mfd_np = f"({gp_expr})-{_aoh}*{costs}"
            subhdr(f"MFD — {_alt:.0f}% on costs")
            line(f"less: Overhead ({_alt:.0f}% on costs)", f"=-{_aoh}*{costs}")
            line("NET PROFIT", f"={mfd_np}", hero=True)
            line("Profit %", f'=IF({costs}=0,"",({mfd_np})/{costs})', pct=True, hero=True)
            subhdr(f"COMPANY — {overhead_pct:.0f}% on revenue")
        # % spelled out on the row itself (the user 2026-07-16: "so we can see
        # what % we're working with")
        line(f"less: Overhead ({overhead_pct:.0f}% of revenue + direct opex)",
             f"=-{opex}-{oh}*({inc_expr})")
        line("NET PROFIT", f"={co_np}", hero=True)
        line("Profit %", f'=IF(({inc_expr})=0,"",({co_np})/({inc_expr}))', pct=True, hero=True)
        box(t0, r - 1)
        r += 1

    if not tx_refs:
        income_row = row("Income", income, bold=True, fill=INCOME_FILL)
        _qbo_link(income_row, _cust_url)
        wip_contract_cell = None
    else:
        Binc = f"{tx_refs['billed']}-{tx_refs['withheld']}+{tx_refs['billed_ret']}"
        Wcell = tx_refs['withheld']
        Bgross = f"{tx_refs['billed']}+{tx_refs['billed_ret']}"
        # retainage moved to receivable by JE — not billed income, but retainage
        # OWED; folds into the retainage snapshots/total only (the user 2026-07-02).
        NBcell = tx_refs.get("not_billed_ret")
        # WIP "Billed to Date" wants GROSS contract billing incl. retainage held:
        # billed + billed_ret + not_billed_ret. On invoice-line-retainage jobs NBcell
        # is empty and Bgross already = gross; on CP "moved to receivable by JE" jobs
        # this adds the held retainage back in (the user 2026-07-15).
        Btot = f"{Bgross}+{NBcell}" if NBcell else Bgross

        # ── ① WIP / PROJECTION (rebuilt 2026-07-16, mock approved by the user):
        #    the BID story first (Original → Change Orders → Revised), the
        #    projection with its profit/overhead split, then the QBO ACTUALS
        #    band. Yellow inputs persist via read-back; auto-fill comes from
        #    the WIP master (typed value still wins on re-sync).
        wtop = sect_title("① WIP / PROJECTION   (yellow = your input)")
        def _num(*keys):
            for k in keys:
                v = wip_info.get(k)
                try:
                    if v not in (None, "") and float(v):
                        return float(v)
                except (TypeError, ValueError):
                    pass
            return None
        # ORIGINAL numbers are the inputs (before change orders). When the
        # master only carries the revised total, original = revised − COs.
        _cos = _num("change_orders")
        _rev_ctr = _num("revised_contract", "contract")
        # what the WIP master WOULD supply (before any typed override), so we
        # can flag a hand-entered value that no longer matches (the user 2026-07-16)
        _wip_ctr = _num("original_contract")
        if _wip_ctr is None and _rev_ctr is not None:
            _wip_ctr = round(_rev_ctr - (_cos or 0.0), 2)
        _wip_etc = _num("original_etc", "revised_etc")
        _typed_ctr = _num("contract_saved")
        _typed_etc = _num("etc_saved")
        _ctr = _typed_ctr if _typed_ctr is not None else _wip_ctr
        _etc = _typed_etc if _typed_etc is not None else _wip_etc
        _co_cost = _num("co_cost_saved")

        def _mismatch(typed, wip):
            return (typed is not None and wip is not None
                    and abs(typed - wip) > 1.0)

        _g702_src = wip_info.get("contract_g702")
        k_row = row(f"Original Contract Price  ({_g702_src})" if _g702_src
                    else "Original Contract Price", _ctr, bold=True)
        if _g702_src:
            _ovr = wip_info.get("contract_g702_typed")
            if _ovr is not None:
                row(f"⚑ a typed ${_ovr:,.0f} was overridden by the G702", None,
                    indent=1, size=BASE_SIZE - 2, color="9C5700")
        elif _mismatch(_typed_ctr, _wip_ctr):
            row(f"⚑ typed — differs from WIP master (${_wip_ctr:,.0f})", None,
                indent=1, size=BASE_SIZE - 2, color="9C5700")
        etc_in = row("Original ETC (Estimated Total Cost)", _etc, bold=True)
        if _mismatch(_typed_etc, _wip_etc):
            row(f"⚑ typed — differs from WIP master (${_wip_etc:,.0f})", None,
                indent=1, size=BASE_SIZE - 2, color="9C5700")
        co_row = row("Change Orders (approved, per G702)" if _g702_src
                     else "Change Orders (approved, from draw)",
                     _cos, color=GREEN)
        coc_row = row("CO Costs (estimated)", _co_cost,
                      color=("000000" if _co_cost else "C0504D"))
        # yellow inputs: the originals + CO cost (no QBO source yet — pending
        # the CO-template cost line; type it here and it survives re-syncs)
        # Yellow = YOU typed it. A G702-sourced contract is not an input, so it
        # stays white (the user 2026-07-29).
        for rr in ((etc_in, coc_row) if _g702_src else (k_row, etc_in, coc_row)):
            cc = ws.cell(row=rr, column=2)
            cc.fill = YEL
            cc.border = Border(left=_HAIR, right=_HAIR, top=_HAIR, bottom=_HAIR)
        rev_ctr_row = row("Revised Contract Price",
                          formula=f"=B{k_row}+B{co_row}", bold=True,
                          border=TOP_BORDER)
        rev_etc_row = row("Revised ETC", formula=f"=B{etc_in}+B{coc_row}",
                          bold=True)
        # everything downstream measures against the REVISED numbers
        c_ref = f"$B${rev_ctr_row}"
        e_ref = f"$B${rev_etc_row}"
        # 'Closed' in the WIP master (Test-Master STATUS) forces the close-out
        # view (the user 2026-07-16): % Complete = 100%, Cost to Complete = 0,
        # Earned = full contract. Closing a job in the WIP master closes it
        # here — no manual ETC massaging.
        wip_closed = str(wip_info.get("status") or "").strip().lower() in (
            "closed", "complete", "completed", "done")

        # ── projection + the profit/overhead split (the user 2026-07-16) ──
        pp_row = row("Projected Profit at Completion", formula=f"={c_ref}-{e_ref}",
                     bold=True, border=TOP_BORDER)
        row("Projected Margin %", formula=f'=IF({c_ref}=0,"",B{pp_row}/{c_ref})',
            fmt=PCT_FMT, bold=True, color=NAVY)
        poh_row = row(f"less: Overhead ({overhead_pct:.0f}% of revenue)",
                      formula=f"=-{oh}*{c_ref}", indent=1, color="595959")
        pnp_row = row("Projected NET Profit", formula=f"=B{pp_row}+B{poh_row}",
                      indent=1, bold=True, color=GREEN)
        row("Projected Net Margin %",
            formula=f'=IF({c_ref}=0,"",B{pnp_row}/{c_ref})',
            fmt=PCT_FMT, indent=1, bold=True, color=GREEN)

        # ── ACTUALS — QBO to date: billed first, then costs, then overhead
        #    sitting on top of the REAL profit ($ and %) — the user 2026-07-16 ──
        row("ACTUALS — QBO to date", None, bold=True, color="FFFFFF",
            fill=hero_fill)
        bd_row = row("Billed to Date", formula=f"={Btot}", bold=True)
        _qbo_link(bd_row, _cust_url)
        ctd_row = row("Costs to Date", None)
        gpa_row = row("Gross Profit (to date)",
                      formula=f"=B{bd_row}-B{ctd_row}",
                      bold=True, border=TOP_BORDER)
        aoh_row = row(f"less: Overhead ({overhead_pct:.0f}% of billed)",
                      formula=f"=-{oh}*B{bd_row}", indent=1, color="595959")
        rnp_row = row("REAL Net Profit (to date)",
                      formula=f"=B{gpa_row}+B{aoh_row}",
                      bold=True, border=TOP_BORDER)
        row("REAL Net Profit %",
            formula=f'=IF(B{bd_row}=0,"",B{rnp_row}/B{bd_row})',
            fmt=PCT_FMT, bold=True)
        # Two progress metrics (the user 2026-07-16): cost-based drives Earned
        # Revenue and reads >100% when costs blow past ETC (red = over budget,
        # never capped — the overage IS the signal); % Billed is billing
        # progress against the contract (retainage-inclusive Btot).
        pc_row = row("% Complete (cost ÷ ETC)", None, fmt=PCT_FMT, bold=True, color=NAVY)
        row("% Billed (billed ÷ contract)",
            formula=f'=IF({c_ref}=0,"",({Btot})/{c_ref})',
            fmt=PCT_FMT, bold=True, color=NAVY)
        earn_row = row("Earned Revenue (contract × %)", None)
        row("Over / (Under) Billing",
            formula=f'=IF({e_ref}=0,"",({Btot})-B{earn_row})', color="C0504D")
        ctc_row = row("Cost to Complete (remaining)", None)
        if wip_closed:
            row("closed per WIP master — % complete forced to 100%", None,
                size=BASE_SIZE - 2, color="595959")
        box(wtop, r - 1)
        # over-budget flag: cost-based % complete turns red past 100%
        ws.conditional_formatting.add(
            f"B{pc_row}",
            CellIsRule(operator="greaterThan", formula=["1"],
                       font=Font(bold=True, color="C00000")))
        wip_contract_cell = c_ref
        r += 1

        # ── ② PROFIT & LOSS TOTALS — TRUE totals, retainage included (the
        #    user 2026-07-16): income = gross work billed + retainage billed
        #    back + retainage receivable (JE) = the same Btot as ① Billed to
        #    Date. The realized/net view lives in ③.
        ftop = sect_title("② PROFIT & LOSS TOTALS")
        income_row = row("Income (incl. retainage)", formula=f"={Btot}",
                         bold=True, size=BASE_SIZE + 1, color="375623", fill=INCOME_FILL)
        _qbo_link(income_row, _cust_url)
        # Every invoice behind that total, newest first (the user 2026-08-27:
        # "show me the invoice #, memo then the amount under it so we see the
        # p&l totals"). Total on the bar, detail underneath — same shape as
        # COGS. Amounts are each invoice's CONTRIBUTION to the bar: gross work
        # billed + retainage billed back, or the retainage moved by JE.
        for _inv in sorted(income_rows or [],
                           key=lambda i: str(i.get("date") or ""), reverse=True):
            _amt = (float(_inv.get("billed", 0) or 0)
                    + float(_inv.get("billed_ret", 0) or 0)
                    + float(_inv.get("not_billed_ret", 0) or 0))
            if abs(_amt) < 0.005:
                continue
            _doc = str(_inv.get("doc") or _inv.get("id") or "")
            # One clean line: memos carry an embedded newline + the Period tag,
            # and repeat the project name on every invoice. Drop all three —
            # the tag is the draw's identity, already the row above it.
            _raw = DRAW_PERIOD_RE.sub("", str(_inv.get("memo") or ""))
            _raw = re.sub(r"\s+", " ", _raw).strip(" -–·")
            _memo = _clean_cost_text(_raw, _project_name_words(cust_info.get("name", "")))
            _lbl = f"#{_doc} — {_memo}" if _memo else f"#{_doc}"
            _ir = row(_lbl, _amt, indent=1, size=BASE_SIZE - 1, color="375623")
            _iu = _qbo_txn_url("invoice", _inv.get("id", ""), realm)
            if _iu:
                _c = ws.cell(row=_ir, column=1)
                _c.hyperlink = _iu
                _c.font = Font(size=BASE_SIZE - 1, color=LINK, underline="single")
        cogs_row = acct_lines("Cost of Goods Sold", tx_refs.get("cogs_accts") or [],
                              "Total Cost of Goods Sold", COGS_FILL)
        _qbo_link(cogs_row, _costs_url)
        gp_row = row("Gross Profit", formula=f"=B{income_row}-B{cogs_row}",
                     bold=True, border=TOP_BORDER, fill=GP_FILL)
        row("Gross Profit %", formula=f'=IF(B{income_row}=0,"",B{gp_row}/B{income_row})',
            fmt=PCT_FMT, bold=True, fill=GP_FILL)
        exp_names = tx_refs.get("exp_accts") or []
        if exp_names:
            exp_row = acct_lines("Operating Expenses (non-COGS)", exp_names,
                                 "Total Operating Expenses", SECT_FILL)
        else:
            exp_row = row("Operating Expenses (non-COGS)", formula=f"={tx_refs['exp']}",
                          bold=True, fill=SECT_FILL)
        noi_row = row("Net Operating Income", formula=f"=B{gp_row}-B{exp_row}",
                      bold=True, border=TOP_BORDER, fill=SECT_FILL)
        box(ftop, r - 1)
        r += 1

        # fill WIP cells now that COGS is known. A 'Closed' WIP-master status
        # overrides the cost-based math: the job is done, so % = 100%, Earned =
        # full contract, nothing left to spend (the user 2026-07-16).
        if wip_closed:
            _wip_fills = (
                (ctd_row, f"=B{cogs_row}", CURR_FMT, False),
                (pc_row, 1.0, PCT_FMT, True),
                (earn_row, f"={c_ref}", CURR_FMT, False),
                (ctc_row, 0.0, CURR_FMT, False))
        else:
            _wip_fills = (
                (ctd_row, f"=B{cogs_row}", CURR_FMT, False),
                (pc_row, f'=IF({e_ref}=0,"",B{cogs_row}/{e_ref})', PCT_FMT, True),
                (earn_row, f'=IF({e_ref}=0,"",{c_ref}*B{cogs_row}/{e_ref})', CURR_FMT, False),
                (ctc_row, f"={e_ref}-B{cogs_row}", CURR_FMT, False))
        for rr, frm, fmt, isb in _wip_fills:
            cc = ws.cell(row=rr, column=2, value=frm)
            cc.number_format = fmt
            cc.font = Font(bold=isb, size=BASE_SIZE, color=NAVY if isb else "000000")
        _qbo_link(ctd_row, _costs_url)

        costs = f"B{cogs_row}"
        opex = f"B{exp_row}"

        # ── ③ SNAPSHOT — REALIZED (net billed) — the cash-billed view, less
        #    retainage (the user 2026-07-16: ② is the true totals WITH
        #    retainage; ③ answers "what did we bill in cash"). The old
        #    with-retainage snapshot is gone — ② already answers it.
        _snap3 = ("③ SNAPSHOT — MFD vs COMPANY (realized, net billed)" if show_mfd
                  else "③ SNAPSHOT — REALIZED (net billed, less retainage)")
        snapshot(_snap3, "Realized — billed in cash, less retainage",
                 f"{Binc}", f"({Binc})-{costs}", costs, opex)

        # ── ④ BILLING & RETAINAGE (to date) ──
        btop = sect_title("④ BILLING & RETAINAGE (to date)")
        gb_row = row("Billed to Date (gross, incl. retainage)", formula=f"={Btot}",
                     indent=1)
        _qbo_link(gb_row, _cust_url)
        # ONE receivable row (the user 2026-07-16: withheld-on-draws and
        # moved-to-receivable-by-JE are the same money owed to us — two booking
        # styles, one bucket).
        _rec = f"{Wcell}+{NBcell}" if NBcell else Wcell
        rec_row = row("less: Retainage receivable (withheld + not billed)",
                      formula=f"=-({_rec})", indent=1, color="C0504D")
        row("Net Billed (to AR)", formula=f"=B{gb_row}+B{rec_row}", indent=1,
            bold=True, border=TOP_BORDER)
        row("Retainage billed (returned by GC)", formula=f"={tx_refs['billed_ret']}",
            indent=1, color=GREEN)
        row("Total retainage (all-time)",
            formula=f"=({_rec})+{tx_refs['billed_ret']}", indent=1,
            bold=True, border=TOP_BORDER, color=NAVY)
        if underbill_total > 0:
            ur = row(f"⚠ Underbilling risk — bills missed by PM "
                     f"({underbill_count} bill{'s' if underbill_count != 1 else ''})",
                     underbill_total, indent=1, bold=True, border=TOP_BORDER,
                     color=RED)
            n = row("see Draws → PM Draw Report Cross-Check", indent=2)
            ws.cell(row=n, column=1).font = Font(italic=True, size=BASE_SIZE - 2,
                                                 color="595959")
        box(btop, r - 1)
        r += 1

    # ── RIGHT SIDE: DRAW COVERAGE table (D–M) + ACCUMULATING COSTS (O–P) ──
    #   D Draw | E Period | F Draw Total | G Retained | H Net Billed |
    #   I Costs | J Gross Profit | K Coverage % | L Net Profit | M Net Cov %
    #   Draw Total (F) + Retained (G) are COLLAPSED by default (the user 2026-06-09)
    #   — click the [+] above to expand. Vertical rule on the right of Net
    #   Billed (H) separates the BILLED side from the COSTS side.
    for col, w in (("D", 22), ("E", 19), ("F", 14), ("G", 13), ("H", 15),
                   ("I", 12), ("J", 15), ("K", 13), ("L", 15), ("M", 13)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["N"].width = 11 if wip_contract_cell else 2  # % Compl / spacer
    ws.column_dimensions["O"].width = 44
    ws.column_dimensions["P"].width = 18
    # collapse the two retainage-detail columns (Draw Total, Retained) by
    # default — outline level 1 + hidden on BOTH so they form one collapsed
    # group; the [+] control sits to the right (over Net Billed). group()
    # alone only tags the first column, so set each explicitly.
    for _cl in ("F", "G"):
        ws.column_dimensions[_cl].outline_level = 1
        ws.column_dimensions[_cl].hidden = True
    ws.column_dimensions["H"].collapsed = True
    ws.sheet_properties.outlinePr.summaryRight = True

    COST_TXT = "C55A11"   # Costs column text — distinct color for visual flair
    _vrule = Side(style="thin", color="808080")

    def _clr(pct):
        return "000000" if pct is None else (GREEN if pct >= 1.0 else RED)

    # DRAW COVERAGE and ACCUMULATING COSTS both answer "what do we bill
    # next", which is settled on a finished job — `simple` drops both (the
    # user 2026-08-27: "these are completed projects, a more streamlined what
    # the project performed is all we need").
    if draw_rows and not simple:
        rc = 3
        cov_top = rc
        t = ws.cell(row=rc, column=4, value="DRAW COVERAGE")
        t.font = Font(bold=True, size=BASE_SIZE + 1, color=NAVY)
        # ── two-tier header (the user 2026-06-19): GROSS over (Gross Profit,
        #    Coverage %); AFTER OVERHEAD over (Net Profit, Net Cov %). The %
        #    lives IN the header (the user 2026-07-16), and MFD jobs use the
        #    MFD 9%-on-costs model here instead of %-of-revenue.
        _ao_hdr = (f"AFTER OVERHEAD — MFD {_alt:.0f}% on costs" if show_mfd
                   else f"AFTER OVERHEAD — {overhead_pct:.0f}% of revenue")
        for c0, c1, txt in ((10, 11, "GROSS"), (12, 13, _ao_hdr)):
            gb = ws.cell(row=rc, column=c0, value=txt)
            gb.font = Font(bold=True, size=BASE_SIZE - 1, color=NAVY)
            gb.alignment = Alignment(horizontal="center")
            for cc in range(c0, c1 + 1):
                ws.cell(row=rc, column=cc).fill = ACCENT_FILL
                ws.cell(row=rc, column=cc).border = BOTTOM_BORDER
            ws.merge_cells(start_row=rc, start_column=c0, end_row=rc, end_column=c1)
        rc += 1
        # cols: 4 Draw 5 Period | 6 Draw Total 7 Retained | 8 Net Billed |
        #       9 Costs 10 Gross Profit 11 Coverage % 12 Net Profit 13 Net Cov %
        _heads = ["Draw", "Period", "Draw Total", "Retained", "Net Billed",
                  "Costs", "Gross Profit", "Coverage %", "Net Profit", "Net Cov %"]
        if wip_contract_cell:
            _heads.append("% Compl")   # cumulative billed ÷ contract (the user 2026-06-22)
        for ci, h in enumerate(_heads, start=4):
            hc = _write_cell(ws, rc, ci, h)
            hc.font = SUBHDR_FONT
            hc.border = BOTTOM_BORDER
            hc.alignment = Alignment(horizontal="center", wrap_text=True)
        rc += 1
        first_draw_row = rc
        for name, lbl, net, costs, held, billed in draw_rows:
            m = _cov(net, costs, overhead_pct)
            pc, po = m["cost_pct"], m["oh_pct"]
            c = _write_cell(ws, rc, 4, name)
            if draw_anchors and name in draw_anchors:
                # draw_anchors now maps draw name → its own SHEET (the user 2026-06-26,
                # one draw = one sheet); the P&L coverage table is the index.
                c.hyperlink = f"#'{draw_anchors[name]}'!A1"
                c.font = Font(size=BASE_SIZE - 1, color=LINK, underline="single")
            else:
                c.font = Font(size=BASE_SIZE - 1)
            _write_cell(ws, rc, 5, lbl).alignment = Alignment(horizontal="center")
            ws.cell(row=rc, column=5).font = Font(size=BASE_SIZE - 1)
            # Draw Total (6) = Net Billed + retainage HELD (gross of withholding).
            # Retained (7) shows retainage on the draw — HELD (black) or, when the
            # GC PAID retainage back, BILLED (green). NEVER red (the user 2026-06-09).
            retained = held + billed
            ret_clr = GREEN if billed > 0.005 else "000000"
            dt_ = _write_cell(ws, rc, 6, net + held); dt_.number_format = CURR_FMT
            dt_.font = Font(size=BASE_SIZE - 1)
            rt = _write_cell(ws, rc, 7, retained); rt.number_format = CURR_FMT
            rt.font = Font(size=BASE_SIZE - 1, color=ret_clr)
            nb = _write_cell(ws, rc, 8, net); nb.number_format = CURR_FMT
            nb.font = Font(size=BASE_SIZE - 1)
            ct = _write_cell(ws, rc, 9, costs); ct.number_format = CURR_FMT
            ct.font = Font(size=BASE_SIZE - 1, color=COST_TXT)
            # formulas: Gross Profit, Coverage %, Net Profit, Net Coverage %
            fmls = [
                (10, f"=H{rc}-I{rc}", CURR_FMT, _clr(pc)),
                (11, f'=IF(I{rc}=0,"",H{rc}/I{rc})', '0.0%', _clr(pc)),
                # MFD nets overhead on COSTS (9%); company nets on REVENUE.
                (12, (f"=H{rc}-I{rc}*{1 + _aoh}" if show_mfd
                      else f"=H{rc}*{one_minus_oh}-I{rc}"), CURR_FMT, _clr(po)),
                (13, (f'=IF(I{rc}=0,"",H{rc}/(I{rc}*{1 + _aoh}))' if show_mfd
                      else f'=IF(I{rc}=0,"",H{rc}/(I{rc}/{one_minus_oh}))'),
                 '0.0%', _clr(po)),
            ]
            for col, f, nf, clr in fmls:
                cell = ws.cell(row=rc, column=col, value=f)
                cell.number_format = nf
                cell.font = Font(size=BASE_SIZE - 1, color=clr)
                if col in (11, 13):
                    cell.alignment = Alignment(horizontal="center")
            if wip_contract_cell:   # cumulative billed ÷ contract through this draw
                pcc = ws.cell(row=rc, column=14,
                              value=f'=IF({wip_contract_cell}=0,"",'
                                    f'SUM(H{first_draw_row}:H{rc})/{wip_contract_cell})')
                pcc.number_format = "0.0%"
                pcc.font = Font(size=BASE_SIZE - 1)
                pcc.alignment = Alignment(horizontal="center")
            rc += 1
        last_draw_row = rc - 1
        # TOTAL row — SUM the source columns, recompute derived from totals
        tc = _write_cell(ws, rc, 4, "TOTAL"); tc.font = Font(bold=True, size=BASE_SIZE - 1)
        tc.border = TOP_BORDER
        for col, f, nf, clr in (
                (6, f"=SUM(F{first_draw_row}:F{last_draw_row})", CURR_FMT, "000000"),
                (7, f"=SUM(G{first_draw_row}:G{last_draw_row})", CURR_FMT, "000000"),
                (8, f"=SUM(H{first_draw_row}:H{last_draw_row})", CURR_FMT, "000000"),
                (9, f"=SUM(I{first_draw_row}:I{last_draw_row})", CURR_FMT, COST_TXT),
                (10, f"=H{rc}-I{rc}", CURR_FMT, "000000"),
                (11, f'=IF(I{rc}=0,"",H{rc}/I{rc})', '0.0%', "000000"),
                (12, (f"=H{rc}-I{rc}*{1 + _aoh}" if show_mfd
                      else f"=H{rc}*{one_minus_oh}-I{rc}"), CURR_FMT, "000000"),
                (13, (f'=IF(I{rc}=0,"",H{rc}/(I{rc}*{1 + _aoh}))' if show_mfd
                      else f'=IF(I{rc}=0,"",H{rc}/(I{rc}/{one_minus_oh}))'),
                 '0.0%', "000000")):
            cell = ws.cell(row=rc, column=col, value=f)
            cell.number_format = nf
            cell.font = Font(bold=True, size=BASE_SIZE - 1, color=clr)
            cell.border = TOP_BORDER
            if col in (11, 13):
                cell.alignment = Alignment(horizontal="center")
        if wip_contract_cell:   # overall % complete = total billed ÷ contract
            tpc = ws.cell(row=rc, column=14,
                          value=f'=IF({wip_contract_cell}=0,"",H{rc}/{wip_contract_cell})')
            tpc.number_format = "0.0%"
            tpc.font = Font(bold=True, size=BASE_SIZE - 1)
            tpc.border = TOP_BORDER
            tpc.alignment = Alignment(horizontal="center")

        # ── vertical rules: after Costs (col 9) and between Coverage % and
        #    Net Profit (after col 11), full table height (the user 2026-06-19) ──
        for gr in range(cov_top, rc + 1):
            for col in (9, 11):
                cur = ws.cell(row=gr, column=col).border
                ws.cell(row=gr, column=col).border = Border(
                    left=cur.left, right=_vrule, top=cur.top, bottom=cur.bottom)

    if accum and not simple:
        ra = 3

        def side(label, amt=None, *, formula=None, bold=False, border=None,
                 fmt=CURR_FMT, color="000000", italic=False, size=None,
                 fill=None, indent=0) -> int:
            nonlocal ra
            used = ra
            val = ("    " * indent + label) if (isinstance(label, str) and indent) else label
            c = _write_cell(ws, ra, 15, val)
            c.font = Font(bold=bold, italic=italic, size=size or BASE_SIZE, color=color)
            a = None
            if formula is not None:
                a = ws.cell(row=ra, column=16, value=formula)
                a.number_format = fmt
                a.font = Font(bold=bold, size=size or BASE_SIZE, color=color)
            elif amt is not None:
                a = _write_cell(ws, ra, 16, amt)
                a.number_format = fmt
                a.font = Font(bold=bold, size=size or BASE_SIZE, color=color)
            if border is not None:
                c.border = border
                if a is not None:
                    a.border = border
            if fill is not None:
                c.fill = fill
                (a or ws.cell(row=ra, column=16)).fill = fill
            ra += 1
            return used

        h = ws.cell(row=ra, column=15,
                    value="ACCUMULATING COSTS — NEXT DRAW   ➜ see detail")
        out_sheet = (draw_anchors or {}).get("__outside")
        if out_sheet:
            h.hyperlink = f"#'{out_sheet}'!A1"
            h.font = Font(bold=True, size=BASE_SIZE, color=LINK, underline="single")
        else:
            h.font = Font(bold=True, size=BASE_SIZE, color=NAVY)
        h.border = BOTTOM_BORDER
        ws.cell(row=ra, column=16).border = BOTTOM_BORDER
        ra += 1
        side(f"bills + purchases outside any draw window (through "
             f"{accum['through']}) — same as Draws ➜", italic=True,
             color="595959", size=BASE_SIZE - 2)

        by_pfx: Dict[str, dict] = {}
        for code, tot in accum["groups"].items():
            pfx, _num = _split_code(code)
            key = pfx or "—"
            d = by_pfx.setdefault(key, {"total": 0.0, "codes": []})
            d["total"] += tot
            d["codes"].append(code)
        pfx_total_rows = []
        labor_code_rows = []
        for pfx in sorted(by_pfx, key=lambda k: _JOB_PREFIX_ORDER.get(k, 99)):
            d = by_pfx[pfx]
            jobname = (_JOB_TYPE_NAMES.get(pfx, "Other")
                       if pfx != "—" else "No Job Type")
            hdr_row = side(jobname, None, bold=True, fill=ACCENT_FILL)
            code_rows = []
            for code in sorted(d["codes"], key=_cost_code_sort_key):
                amt = accum["groups"][code]
                lbl_val = (_cost_name_value(code, indent=1, size=BASE_SIZE)
                           if _is_cost_code(code) else None)
                if lbl_val is not None:
                    c = ws.cell(row=ra, column=15, value=lbl_val)
                    c.font = Font(size=BASE_SIZE)
                    a = _write_cell(ws, ra, 16, amt); a.number_format = CURR_FMT
                    a.font = Font(size=BASE_SIZE)
                    cur = ra
                    ra += 1
                else:
                    cur = side(code, amt, indent=1)
                code_rows.append(cur)
                if _cost_category(code) == "Labor":
                    labor_code_rows.append(cur)
            # job-prefix total = SUM of its code cells
            f = ("=" + "+".join(f"P{cr}" for cr in code_rows)) if code_rows else "=0"
            ws.cell(row=hdr_row, column=16, value=f).number_format = CURR_FMT
            ws.cell(row=hdr_row, column=16).font = Font(bold=True, size=BASE_SIZE)
            ws.cell(row=hdr_row, column=16).fill = ACCENT_FILL
            pfx_total_rows.append(hdr_row)
        total_f = ("=" + "+".join(f"P{pr}" for pr in pfx_total_rows)) if pfx_total_rows else "=0"
        tot_row = side("Total accumulating costs", formula=total_f, bold=True,
                       border=TOP_BORDER)
        if labor_code_rows:
            lf = "=" + "+".join(f"P{lr}" for lr in labor_code_rows)
            side("  ↳ of which Labor already PAID out of pocket", formula=lf,
                 bold=True, color="C0504D")
            side("(subs/crews are paid weekly — this cash is already out the "
                 "door, awaiting the next draw)", italic=True, color="595959",
                 size=BASE_SIZE - 2)
        side("Draw needed (costs + overhead)",
             formula=(f"=P{tot_row}*{round(1 + _aoh, 4)}" if show_mfd
                      else f"=P{tot_row}/{one_minus_oh}"),
             bold=True, color=NAVY)
        side((f"(= costs × {round(1 + _aoh, 4)} — MFD {_alt:.0f}% on costs; "
              f"break-even only, no profit margin)") if show_mfd else
             f"(= costs ÷ {one_minus_oh}; break-even only, no profit margin)",
             italic=True, color="595959", size=BASE_SIZE - 2)


def _draw_flat_bills(draw_cost: dict) -> list:
    """Flatten one draw's bucketed costs (parent→leaf→vendor→txns) into a plain
    list of bills for the per-draw 'TRANSACTIONS' table."""
    rows = []
    for pg in (draw_cost.get("groups") or {}).values():
        for leaf, lg in pg.get("subs", {}).items():
            cat = _cost_category(leaf)
            for vend, vg in lg.get("vendors", {}).items():
                for t in vg.get("txns", []):
                    rows.append({"num": str(t.get("doc_num", "")), "date": t.get("date", ""),
                                 "vendor": vend, "cat": cat, "desc": t.get("desc", ""),
                                 "amount": float(t.get("amount", 0) or 0),
                                 "tx_type": t.get("tx_type", ""), "txn_id": t.get("txn_id", "")})
    return rows


def build_sheet_one_draw(wb, sheet_name, proj, cust_info, wip_info, name, lbl,
                         net, costs, held, billed, invoices, draw_cost,
                         matched_report, report_index, qbo_loc, period,
                         as_of, overhead_pct=10.0, realm="", alt_overhead_pct=None,
                         reports_relpath="rd-reports", paid_map=None):
    """ONE SHEET PER DRAW = a TWO-PERSPECTIVE reconciliation (the user 2026-06-26):
    the PM's version of the draw (their report, their costs, the profit they thought
    they had) SIDE BY SIDE with QBO truth (the director-revised period download).
    Same billed invoice on both sides, so the profit gap is purely the cost/period
    difference. Below: the bill-level reconciliation — QBO-only (in QBO this draw,
    not on the PM report) and PM-only (on the report, not in QBO this draw), each
    cross-referenced to where the bill actually landed. Returns
    (bottom_row, missed_total, missed_count) where missed = QBO bills on NO report
    anywhere (true underbilling)."""
    from collections import Counter
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    _known_words = _project_name_words(cust_info.get("name", ""))
    # C carries bill descriptions (was 22 and clipped most of them); E is the
    # short "where / status" note, so the width moves from E to C.
    for col, w in (("A", 26), ("B", 20), ("C", 34), ("D", 16), ("E", 30),
                   ("F", 11)):
        ws.column_dimensions[col].width = w

    def _font(bold=False, color="000000", size=None, under=False):
        return Font(bold=bold, size=size or BASE_SIZE, color=color,
                    underline=("single" if under else None))

    def wc(row, col, val, *, bold=False, color="000000", fmt=None, fill=None,
           indent=0, link=None, wrap=False, size=None):
        v = ("    " * indent + val) if isinstance(val, str) else val
        c = ws.cell(row=row, column=col, value=v)
        c.font = _font(bold, LINK if link else color, size, bool(link))
        if fmt:
            c.number_format = fmt
        if fill is not None:
            c.fill = fill
        if link:
            c.hyperlink = link
        if wrap:
            c.alignment = Alignment(wrap_text=True, vertical="top")
        return c

    def wdate(row, col, ds):
        d = _parse_date(ds) if isinstance(ds, str) else (ds if isinstance(ds, dt.date) else None)
        c = ws.cell(row=row, column=col, value=d or ds)
        if d:
            c.number_format = "mm/dd/yyyy"
        c.font = _font()

    def band(row, c0, c1, label, fill=HDR_FILL):
        wc(row, c0, label, bold=True, color="FFFFFF", fill=fill, size=BASE_SIZE)
        for c in range(c0, c1 + 1):
            ws.cell(row=row, column=c).fill = fill

    def keyb(x):
        return (str(x.get("num", "")).strip(), round(float(x.get("amount", 0) or 0), 2))

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of, compact=True)
    # PAID / UNPAID leads the title (the user 2026-08-05): the draw is PAID
    # when every invoice in it has a zero open balance in QBO.
    _inv_paid = bool(invoices) and all(
        float(i.get("balance", 0) or 0) <= 0.005 for i in invoices)
    wc(r, 1, f"{'PAID' if _inv_paid else 'UNPAID'} {name}  —  {lbl}",
       bold=True, color=(GREEN if _inv_paid else RED), size=BASE_SIZE + 3)
    r += 2

    # ── data ──
    pm_name, pm_rep = (matched_report if matched_report else (None, None))
    pm_lines = [l for l in (pm_rep["lines"] if pm_rep else []) if not l.get("prior")]
    pm_total = round(sum(l["amount"] for l in pm_lines), 2)
    qbo_bills = _draw_flat_bills(draw_cost)
    qbo_total = round(sum(b["amount"] for b in qbo_bills), 2)
    rev = net                                            # billed invoice (both sides)

    pm_ct = Counter(keyb(l) for l in pm_lines)
    qbo_ct = Counter(keyb(b) for b in qbo_bills)
    matched, qbo_only, used = [], [], Counter()
    for b in qbo_bills:
        k = keyb(b)
        if used[k] < min(pm_ct[k], qbo_ct[k]):
            matched.append(b); used[k] += 1
        else:
            qbo_only.append(b)
    pm_only, pused = [], Counter()
    for l in pm_lines:
        k = keyb(l)
        if pused[k] < min(pm_ct[k], qbo_ct[k]):
            pused[k] += 1
        else:
            pm_only.append(l)
    # The PM draw-report cross-check is the MFD workflow. CP (and any job with no PM
    # reports) has nothing to cross-check, so there's no "missed by PM" (the user 2026-07-02).
    has_pm = bool(report_index)
    # true underbilling = QBO-only bills on NO report anywhere (only when PM reports exist)
    missed = [b for b in qbo_only if not report_index.get(keyb(b))] if has_pm else []
    missed_total = round(sum(b["amount"] for b in missed), 2)

    # ── SUMMARY ── a HORIZONTAL headline strip (the user 2026-07-29): income →
    # retainage → net draw → costs → profit → overhead → REAL net, read left to
    # right, big, with the profit cells colored by sign. MFD (which has a PM
    # report to argue with) gets a second strip so both perspectives stay.
    # Each KPI is a MERGED PAIR of columns, so the strip carries its own width
    # and never dictates the bills table below it (the user 2026-08-31 — it was
    # forcing the table's Paid?/status columns to 18). Nine tiles over columns
    # A..R, contiguous, no gap between GROSS PROFIT and GROSS MARGIN %.
    KPI_COLS = [1, 3, 5, 7, 9, 11, 13, 15, 17]
    KPI_SPAN = 2
    _oh_label = (f"OVERHEAD\n{alt_overhead_pct:.0f}% of costs"
                 if alt_overhead_pct is not None
                 else f"OVERHEAD\n{overhead_pct:.0f}% of income")

    def kpi_strip(title, costs_val, n_bills, periodtxt):
        """One perspective as a label row + a value row. Returns the next row."""
        nonlocal r
        gp = round(rev - costs_val, 2)
        oh = round(((alt_overhead_pct / 100.0) * costs_val)
                   if alt_overhead_pct is not None
                   else (overhead_pct / 100.0) * rev, 2)
        npf = round(gp - oh, 2)
        KPI_FMT = '"$"#,##0;[Red]-"$"#,##0'
        cells = [("INCOME\nbilled this draw", rev, KPI_FMT, False),
                 ("RETAINAGE\nheld back", -abs(held), KPI_FMT, False),
                 ("NET DRAW\ncash to collect", round(rev - abs(held), 2), KPI_FMT, False),
                 ("COSTS\n" + f"{n_bills} bills", costs_val, KPI_FMT, False),
                 ("GROSS PROFIT\nincome − costs", gp, KPI_FMT, True),
                 ("GROSS MARGIN %", (gp / rev if rev else 0), "0.0%", True),
                 (_oh_label, -oh, KPI_FMT, False),
                 ("REAL NET PROFIT", npf, KPI_FMT, True),
                 ("REAL NET %", (npf / rev if rev else 0), "0.0%", True)]
        band(r, 1, KPI_COLS[-1] + KPI_SPAN - 1, f"{title}   ·   {periodtxt}")
        r += 1
        for col, (label, _v, _f, _s) in zip(KPI_COLS, cells):
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + KPI_SPAN - 1)
            c = ws.cell(row=r, column=col, value=label)
            c.font = Font(bold=True, size=BASE_SIZE, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="44546A")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            for _cc in range(col, col + KPI_SPAN):
                ws.cell(row=r, column=_cc).border = THIN_BORDER
                ws.cell(row=r, column=_cc).fill = PatternFill("solid", fgColor="44546A")
        ws.row_dimensions[r].height = 32
        r += 1
        # The derived figures are FORMULAS so the derivation is visible —
        # gross profit is income − costs, NOT net draw − costs (the user
        # 2026-08-31: "i want to see if you are getting the gross from total
        # income or net draw"). Column letters follow KPI_COLS.
        _L = [get_column_letter(c) for c in KPI_COLS]
        _INC, _RET, _NET, _CST, _GP, _GM, _OH, _NP, _NM = _L
        _formula = {
            _NET: f"={_INC}{r}+{_RET}{r}",
            _GP:  f"={_INC}{r}-{_CST}{r}",
            _GM:  f'=IF({_INC}{r}=0,"",{_GP}{r}/{_INC}{r})',
            _NP:  f"={_GP}{r}+{_OH}{r}",
            _NM:  f'=IF({_INC}{r}=0,"",{_NP}{r}/{_INC}{r})',
        }
        signed = []
        for col, (_l, value, fmt, sign) in zip(KPI_COLS, cells):
            _lt = get_column_letter(col)
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + KPI_SPAN - 1)
            c = ws.cell(row=r, column=col, value=_formula.get(_lt, value))
            c.number_format = fmt
            c.font = Font(bold=True, size=BASE_SIZE + 4)
            c.alignment = Alignment(horizontal="center", vertical="center")
            for _cc in range(col, col + KPI_SPAN):
                ws.cell(row=r, column=_cc).border = THIN_BORDER
            if sign:
                signed.append(get_column_letter(col) + str(r))
        ws.row_dimensions[r].height = 26
        if signed:
            ref = " ".join(signed)
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                 fill_type="solid"),
                font=Font(color="9C0006", bold=True, size=BASE_SIZE + 4)))
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                 fill_type="solid"),
                font=Font(color="006100", bold=True, size=BASE_SIZE + 4)))
        r += 1

    qbo_periodtxt = (f"{period[0]:%m/%d/%y}–{period[1]:%m/%d/%y}" if period else lbl)
    kpi_strip("DRAW SUMMARY  (QBO)", qbo_total, len(qbo_bills), qbo_periodtxt)
    if has_pm:
        r += 1
        if pm_rep:
            pm_periodtxt = (f"{pm_rep['period'][0]:%m/%d/%y}–{pm_rep['period'][1]:%m/%d/%y}"
                            if pm_rep.get("period") else "— no period —")
            kpi_strip("SAME DRAW, PM REPORT", pm_total, len(pm_lines), pm_periodtxt)
            if pm_name:
                wc(r, 1, "PM source:", color="595959")
                wc(r, 2, pm_name, color=LINK, link=f"{reports_relpath}/{pm_name}")
                r += 1
        else:
            wc(r, 1, "— no PM report matched this draw —", color="595959")
            r += 1
    # Only the strip stays pinned — everything below scrolls under it.
    ws.freeze_panes = ws.cell(row=r, column=1)
    r += 2

    # ── BILL-LEVEL RECONCILIATION ──
    def detail(title, items, color, kind):
        """kind: 'qbo' rows are QBO bills; 'pm' rows are report lines. GROUPED BY
        VENDOR (the user 2026-06-26 — every transaction listing groups by vendor); every
        bill links (QBO deep-link for QBO rows, the source PM report for PM rows)."""
        nonlocal r
        # Column A is an empty gutter so the vendor name is not jammed against
        # the sheet edge, and so the outline +/- controls have somewhere to sit
        # (the user 2026-08-31). Everything below shifts one column right.
        tot = round(sum(i["amount"] for i in items), 2)
        band(r, 2, 7, f"{title}", fill=(WARN_FILL if color == RED else SUBHDR_FILL))
        ws.cell(row=r, column=2).font = _font(bold=True, color=(RED if color == RED else NAVY))
        ws.cell(row=r, column=5).value = tot
        ws.cell(row=r, column=5).number_format = CURR_FMT
        ws.cell(row=r, column=5).font = _font(bold=True, color=(RED if color == RED else NAVY))
        r += 1
        for c, h in ((2, "Vendor / Bill #"), (3, "Date"), (4, "Description"),
                     (5, "Amount"), (6, "Where / status"), (7, "Paid?")):
            wc(r, c, h, bold=True, color=NAVY).border = BOTTOM_BORDER
        r += 1
        byv = {}
        for i in items:
            byv.setdefault(i.get("vendor") or "(no vendor)", []).append(i)
        for vend in sorted(byv, key=lambda v: -sum(i["amount"] for i in byv[v])):
            vit = byv[vend]
            wc(r, 2, f"{vend}  ({len(vit)})", bold=True, color=color)
            wc(r, 5, round(sum(i["amount"] for i in vit), 2), fmt=CURR_FMT,
               bold=True, color=color)
            r += 1
            for i in sorted(vit, key=lambda x: -x["amount"]):
                if kind == "plain":                      # no PM reports (CP): neutral
                    blink = _qbo_txn_url(i.get("tx_type", ""), i.get("txn_id", ""), realm)
                    note, ncol, nlink = ("in this draw", "595959", None)
                elif kind == "qbo":
                    blink = _qbo_txn_url(i.get("tx_type", ""), i.get("txn_id", ""), realm)
                    other = sorted(report_index.get(keyb(i), set())
                                   - ({pm_name} if pm_name else set()))
                    if other:
                        note, ncol, nlink = (f"on {other[0]}", "BF8F00",
                                             f"{reports_relpath}/{other[0]}")
                    elif not report_index.get(keyb(i)):
                        note, ncol, nlink = ("⚠ on NO report — underbilled", RED, None)
                    else:
                        note, ncol, nlink = ("on this report", "595959", None)
                else:                                    # pm-only report line
                    blink = f"{reports_relpath}/{pm_name}" if pm_name else None
                    loc = qbo_loc.get(keyb(i))
                    if loc and loc != name:
                        note, ncol, nlink = (f"→ in QBO {loc}", "375623", None)
                    elif loc == name:
                        note, ncol, nlink = ("in this draw (count differs)", "595959", None)
                    else:
                        note, ncol, nlink = ("⚠ not in QBO (orphan → Reconciliations)",
                                             RED, None)
                wc(r, 2, str(i["num"]) or "(no #)", indent=1, link=blink)
                wdate(r, 3, i.get("date", ""))
                wc(r, 4, _clean_cost_text(i.get("desc", ""), _known_words))
                wc(r, 5, i["amount"], fmt=CURR_FMT, color=color)
                wc(r, 6, note, color=ncol, link=nlink)
                # AP payment state (the user 2026-08-05); PM report lines have
                # no QBO bill to check.
                if kind != "pm" and paid_map is not None:
                    _pd = paid_map.get(i.get("txn_id"))
                    if _pd is not None:
                        # Use the colour _pay_state returns. `_pd` is a
                        # (balance, total) TUPLE — always truthy — so the old
                        # `GREEN if _pd else RED` painted UNPAID green.
                        _lbl, _col = _pay_state(_pd[0], _pd[1])
                        wc(r, 7, _lbl or "", bold=True, color=_col or RED)
                # COLLAPSED BY DEFAULT — the sheet opens on vendor totals, the
                # way the Project Ledger does; click + to open one vendor
                # (the user 2026-08-31).
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = True
                r += 1
        r += 1

    if has_pm:
        if matched:
            detail(f"MATCHED — on PM report AND in QBO this draw  ({len(matched)})",
                   matched, GREEN, "qbo")
        if qbo_only:
            detail(f"QBO ONLY — in QBO this draw, not on the PM report  ({len(qbo_only)})",
                   qbo_only, RED, "qbo")
        if pm_only:
            detail(f"PM ONLY — on the PM report, not in QBO this draw  ({len(pm_only)})",
                   pm_only, "BF8F00", "pm")
    elif qbo_bills:                                   # CP: just the draw's bills, by vendor
        detail(f"BILLS THIS DRAW — grouped by vendor  ({len(qbo_bills)})",
               qbo_bills, NAVY, "plain")

    # A = gutter for the outline +/-; the table is CONTIGUOUS B..G with no
    # reserved spill columns — those were a 64-character void on every row
    # (the user 2026-08-31: "this is what i mean by extra space"). H..R exist
    # only to give the merged KPI pairs above their width.
    for _c, _w in zip("ABCDEFGHIJKLMNOPQR",
                      (3, 30, 12, 32, 17, 20, 14,
                       15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15)):
        ws.column_dimensions[_c].width = _w
    ws.sheet_properties.outlinePr.summaryBelow = False
    _setup_print(ws, 7)
    return r, missed_total, len(missed)


def build_sheet_next_draw_retainage(wb, proj, cust_info, wip_info, income_groups,
                                    draw_costs, as_of, realm=""):
    """Costs OUTSIDE every draw window (accumulating toward the next draw) +
    untagged invoices. Retainage blocks were REMOVED (the user 2026-07-16 —
    that story already lives on the Transactions sheet)."""
    outside = draw_costs.get("__outside")
    untag = income_groups.get("__untagged")
    has = ((outside and (outside.get("total") or outside.get("groups")))
           or (untag and untag.get("invoices")))
    if not has:
        return None

    ws = wb.create_sheet("Next Draw")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    for col, w in (("A", 16), ("B", 12), ("C", 24), ("D", 16), ("E", 40), ("F", 14)):
        ws.column_dimensions[col].width = w

    def wc(row, col, val, *, bold=False, color="000000", fmt=None, fill=None,
           indent=0, link=None, wrap=False):
        v = ("    " * indent + val) if isinstance(val, str) else val
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(bold=bold, size=BASE_SIZE - 1, color=LINK if link else color,
                      underline=("single" if link else None))
        if fmt:
            c.number_format = fmt
        if fill is not None:
            c.fill = fill
        if link:
            c.hyperlink = link
        if wrap:
            c.alignment = Alignment(wrap_text=True, vertical="top")
        return c

    def wdate(row, col, ds):
        d = _parse_date(ds) if isinstance(ds, str) else (ds if isinstance(ds, dt.date) else None)
        c = ws.cell(row=row, column=col, value=d or ds)
        if d:
            c.number_format = "mm/dd/yyyy"
        c.font = Font(size=BASE_SIZE - 1)

    def band(row, label, color="FFFFFF", fill=HDR_FILL):
        wc(row, 1, label, bold=True, color=color, fill=fill)
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = fill

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    _known_words = _project_name_words(cust_info.get("name", ""))

    disregarded = draw_costs.get("__disregarded")
    if outside and (outside.get("total") or outside.get("groups")):
        band(r, "NEXT DRAW — costs since the last draw (accumulating)"); r += 1
        if disregarded:
            wc(r, 1, f"history before {disregarded.get('anchor') or 'the first tagged period'} "
                     f"disregarded here ({disregarded['count']} lines, "
                     f"${disregarded['total']:,.0f}) — pre-period process; "
                     f"still in P&L totals / Transactions", color="595959")
            r += 1
        for i, h in enumerate(["Bill # / Vendor", "Date", "Cost", "Description", "", "Amount"], 1):
            hc = wc(r, i, h, bold=True, color=NAVY); hc.border = BOTTOM_BORDER
        r += 1
        byv = {}                                          # group by vendor (the user 2026-06-26)
        for b in _draw_flat_bills(outside):
            byv.setdefault(b["vendor"] or "(no vendor)", []).append(b)
        for vend in sorted(byv, key=lambda v: -sum(b["amount"] for b in byv[v])):
            vit = byv[vend]
            wc(r, 1, f"{vend}  ({len(vit)})", bold=True, color="C55A11")
            wc(r, 6, round(sum(b["amount"] for b in vit), 2), fmt=CURR_FMT,
               bold=True, color="C55A11")
            r += 1
            for b in sorted(vit, key=lambda x: (_parse_date(x["date"]) if isinstance(x["date"], str)
                                                else x["date"]) or dt.date.min, reverse=True):
                wc(r, 1, b["num"] or "(no #)", indent=1,
                   link=_qbo_txn_url(b["tx_type"], b["txn_id"], realm))
                wdate(r, 2, b["date"]); wc(r, 3, b["cat"])
                wc(r, 4, _clean_cost_text(b["desc"], _known_words))
                wc(r, 6, b["amount"], fmt=CURR_FMT)
                r += 1
        wc(r, 1, "Total accumulating", bold=True)
        wc(r, 6, outside.get("total", 0.0), fmt=CURR_FMT, bold=True)
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = TOP_BORDER
        r += 2

    def invoice_section(grp, title, color):
        nonlocal r
        band(r, title, color="FFFFFF", fill=HDR_FILL); r += 1
        for inv in grp["invoices"]:
            for i, ln in enumerate(inv.get("lines") or [{"desc": inv.get("memo", ""),
                                                          "amt": inv.get("amount", 0.0)}]):
                wc(r, 1, f"Inv #{inv.get('doc_num', '')}" if i == 0 else "",
                   indent=1, color=color)
                if i == 0:
                    wdate(r, 2, inv.get("date", ""))
                wc(r, 3, ln.get("desc", ""), wrap=True)
                wc(r, 6, float(ln.get("amt", 0) or 0), fmt=CURR_FMT, color=color)
                r += 1
        wc(r, 1, f"Total — {title.split('—')[0].strip().title()}", bold=True, color=color)
        wc(r, 6, grp.get("net_billed", grp.get("total", 0.0)), fmt=CURR_FMT,
           bold=True, color=color)
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = TOP_BORDER
        r += 2

    if untag and untag.get("invoices"):
        # Pre-anchor untagged invoices are the pre-period-tagging era — roll
        # them into ONE line instead of itemizing (the user 2026-07-16: they
        # cloud the P&L; the Transactions sheet still lists each one).
        _anch = _parse_date((disregarded or {}).get("anchor") or "")
        if _anch:
            cur = [i for i in untag["invoices"]
                   if (_parse_date(i.get("date", "")) or dt.date.max) >= _anch]
            old = [i for i in untag["invoices"]
                   if (_parse_date(i.get("date", "")) or dt.date.max) < _anch]
        else:
            cur, old = untag["invoices"], []
        if cur:
            cur_grp = dict(untag)
            cur_grp["invoices"] = cur
            cur_grp["net_billed"] = sum(float(i.get("amount", 0) or 0) for i in cur)
            invoice_section(cur_grp, "UNTAGGED INVOICES (no Period in PrivateNote)", NAVY)
        if old:
            band(r, "PRE-PERIOD HISTORY (before the first tagged draw)"); r += 1
            wc(r, 1, f"{len(old)} invoice(s) before {(disregarded or {}).get('anchor')} "
                     f"— pre-period process, see Transactions", color="595959")
            wc(r, 6, round(sum(float(i.get('amount', 0) or 0) for i in old), 2),
               fmt=CURR_FMT, color="595959")
            r += 2

    _setup_print(ws, 6)
    return ws


# ────────────────────────── Budget vs Actual (the user 2026-07-16) ──────────────────────────

# CP takeoff discovery mirrors wip/cp_wip_reader.py's locked convention (tools
# never import tools, so the ~20 lines live here too): project folder under
# the Awarded-Projects root (active, then Completed Projects/), takeoff = the
# 'takeoff'-named xlsx in the folder ROOT; several takeoffs → the 'WIP'-tagged
# ones, summed. Budget = the 'Cost Code(s)' sheet: col A = code, col C = $.
_CP_ACTIVE_DIR = Path(os.getenv(
    "CP_ACTIVE_DIR",
    "/Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects"))
_AUX_XLSX_RE = re.compile(r"cost\s*code|explanation", re.IGNORECASE)
_WIP_TAG_RE = re.compile(r"\bwip\b", re.IGNORECASE)


def _find_cp_folder(proj: str) -> Optional[Path]:
    for root in (_CP_ACTIVE_DIR, _CP_ACTIVE_DIR / "Completed Projects"):
        try:
            for d in sorted(root.iterdir()):
                if d.is_dir() and d.name.upper().startswith(proj.upper()):
                    return d
        except OSError:
            continue
    return None


def _find_cp_takeoffs(proj: str) -> List[Path]:
    folder = _find_cp_folder(proj)
    if not folder:
        return []
    xl = [f for f in folder.iterdir()
          if f.suffix.lower() in (".xlsx", ".xlsm")
          and not f.name.startswith("~$")]
    takeoffs = [f for f in xl if "takeoff" in f.name.lower()]
    if not takeoffs:
        cands = [f for f in xl if not _AUX_XLSX_RE.search(f.name)]
        return cands if len(cands) == 1 else []
    if len(takeoffs) == 1:
        return takeoffs
    wip_tagged = [f for f in takeoffs if _WIP_TAG_RE.search(f.name)]
    return wip_tagged  # several with none tagged → don't guess


def _read_cost_code_sheet(path: Path, sheet_hint: str) -> Dict[str, float]:
    """Cost-code budget rows from ONE workbook sheet whose name contains
    `sheet_hint`: col A = cost code, col C = cached budget $. Stops at the
    first non-code row (total rows etc.); skips zero and uncached rows."""
    out: Dict[str, float] = {}
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return out
    # Several sheets can carry the hint ('Cost Codes V2' + the old descriptive
    # 'Cost Codes' in one workbook — CP831). V2 first, then first sheet that
    # actually yields codes wins; a descriptive sheet yields nothing and falls
    # through (the user 2026-07-31).
    matches = [nm for nm in wb.sheetnames if sheet_hint in nm.strip().lower()]
    matches.sort(key=lambda nm: (0 if "v2" in nm.lower() else 1))
    for target in matches:
        ws = wb[target]
        for row_vals in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            code = str(row_vals[0] or "").strip().upper()
            if not code or not _is_cost_code(code):
                break                    # data ends (blank / total row)
            v = row_vals[2]
            try:
                amt = float(v)
            except (TypeError, ValueError):
                continue                 # stale formula cache → skip, no guess
            if abs(amt) > 0.005:
                out[code] = out.get(code, 0.0) + round(amt, 2)
        if out:
            break
    wb.close()
    return out


def load_cp_budget(proj: str) -> Tuple[Dict[str, float], str]:
    """CP budget by cost code. Two homes (the user 2026-07-31):
      1. the takeoff's 'Cost Code(s)' sheet (col A = code, col C = $), else
      2. a 'Cost Codes*.xlsx' in the project folder ROOT — the newer takeoffs
         keep the coded budget there on a 'Cost Codes V2' sheet, and the
         takeoff's own 'Cost Codes' sheet is descriptive (no codes).
    Returns ({code: $}, source-note); empty dict when neither exists."""
    budget: Dict[str, float] = {}
    names = []
    for tk in _find_cp_takeoffs(proj):
        part = _read_cost_code_sheet(tk, "cost code")
        if part:
            names.append(tk.name)
            for k, v in part.items():
                budget[k] = budget.get(k, 0.0) + v
    if not budget:
        folder = _find_cp_folder(proj)
        if folder is not None:
            try:
                cc_books = sorted(
                    f for f in folder.iterdir()
                    if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm")
                    and not f.name.startswith("~$")
                    and re.match(r"cost\s*codes?", f.name, re.IGNORECASE))
            except OSError:
                cc_books = []
            for f in cc_books:
                part = _read_cost_code_sheet(f, "cost code")
                if part:
                    names.append(f.name)
                    for k, v in part.items():
                        budget[k] = budget.get(k, 0.0) + v
                    break                # one budget book per job — don't sum copies
    return budget, " + ".join(names)


# ───────── CP concrete yards + G702 pay application (the user 2026-07-29) ─────────

def _read_yards_sheet(path: Path) -> Tuple[float, Dict[str, float]]:
    """The takeoff's 'CONCRETE YARDS' sheet → (total_yards, {deliverable: yards}).
    Col A = deliverable, col C = yards, a 'TOTALS' row closes the block.

    The takeoff does NOT split yards by cost code — the deliverable rows roll
    into the SL/PV/CS/MS families in ways only the estimator knows (CP745 moves
    a ~$1.7k concrete line from the foundation group into paving). So the TOTAL
    is the only figure a $/yd comparison may lean on; per-family yards are
    deliberately NOT inferred."""
    total, rows = 0.0, {}
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return 0.0, {}
    target = next((n for n in wb.sheetnames
                   if "concrete yards" in n.strip().lower()), None)
    if not target:
        wb.close()
        return 0.0, {}
    ws = wb[target]
    for vals in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        label = str(vals[0] or "").strip()
        if not label:
            continue
        if label.upper().startswith("TOTAL"):
            try:
                total = round(float(vals[2]), 2)
            except (TypeError, ValueError):
                pass
            break
        try:
            y = float(vals[2])
        except (TypeError, ValueError):
            continue
        if abs(y) > 0.005:
            rows[label] = round(y, 2)
    wb.close()
    return (total or round(sum(rows.values()), 2)), rows


def load_cp_concrete_yards(proj: str) -> Tuple[float, Dict[str, float], str]:
    """Budget concrete YARDS from the takeoff → (total, {deliverable: yards}, source)."""
    for tk in _find_cp_takeoffs(proj):
        total, rows = _read_yards_sheet(tk)
        if total:
            return total, rows, tk.name
    return 0.0, {}, ""


# The G702 is what the GC certifies, so CONTRACT PRICE and APPROVED CHANGE
# ORDERS come from the signed pay application — NOT from the draw invoices
# (the user 2026-07-29). Discovery + parsing live in shared/draws.py (repo
# rule: tools never import tools), which also owns the modern 'G702'-sheet
# readers used by the WIP reader and health-dashboard.


def load_g702(proj: str) -> Dict[str, object]:
    """CONTRACT PRICE + APPROVED COs from this CP project's latest signed pay
    application. {} when there's no awarded folder or no pay app; {"error": …}
    when one exists but can't be read."""
    folder = _find_awarded_cp_folder(CP_AWARDED_BASE, proj)
    if folder is None:
        return {}
    return read_pay_app(folder)


# RP takeoffs live under client/address folders; the budget is the takeoff's
# 'Cost Gral' sheet (the last visible tab; older takeoff vintages don't have
# it — then there's no budget to show). Codes col A (trailing spaces!), qty
# col C (0/1), amount col D. FW codes belong to the -FTW project; SL/PR (and
# the rest) to the slab project — mirrors the WIP master's RP#/RP#-FTW split.
_RP_ROOT = Path(os.getenv("RP_ROOT", "/Volumes/Common/CURRENT PROJECTS/Residential"))
_RP_SKIP_RE = re.compile(r"flatwork|invoice|estimate|measure", re.IGNORECASE)


_RP_FILE_RE = re.compile(r"^(RP\d{4})_", re.IGNORECASE)
_rp_index_cache: Optional[Dict[str, List[Path]]] = None


def _rp_takeoff_index() -> Dict[str, List[Path]]:
    """ONE walk of the Residential tree (client/ + client/address/) → RP# →
    candidate takeoff files. Parallel scandir (the NAS is slow serially —
    same trick as rp_wip_reader) and cached for the process, so an
    `active rp` batch scans once, not 74 times."""
    global _rp_index_cache
    if _rp_index_cache is not None:
        return _rp_index_cache
    from concurrent.futures import ThreadPoolExecutor
    index: Dict[str, List[Path]] = {}

    def _scan(folder: Path):
        files: List[Path] = []
        subdirs: List[Path] = []
        try:
            with os.scandir(folder) as it:
                for e in it:
                    if e.is_dir(follow_symlinks=False):
                        subdirs.append(Path(e.path))
                    elif e.is_file(follow_symlinks=False):
                        files.append(Path(e.path))
        except OSError:
            pass
        return files, subdirs

    try:
        clients = [d for d in _RP_ROOT.iterdir() if d.is_dir()]
    except OSError:
        clients = []
    all_files: List[Path] = []
    with ThreadPoolExecutor(max_workers=24) as ex:
        level1 = list(ex.map(_scan, clients))
        addr_dirs = [d for _, subs in level1 for d in subs]
        all_files.extend(f for fs, _ in level1 for f in fs)
        for fs, _ in ex.map(_scan, addr_dirs):
            all_files.extend(fs)
    for f in all_files:
        if (f.suffix.lower() in (".xlsm", ".xlsx")
                and not f.name.startswith("~$")
                and not _RP_SKIP_RE.search(f.name)):
            mm = _RP_FILE_RE.match(f.name)
            if mm:
                index.setdefault(mm.group(1).upper(), []).append(f)
    _rp_index_cache = index
    return index


def _find_rp_takeoff(proj: str) -> Optional[Path]:
    base = proj.upper().replace("-FTW", "")
    cands = _rp_takeoff_index().get(base) or []
    if not cands:
        return None
    # the current takeoff is the "UPDATED" one; else the plainest name
    return sorted(cands, key=lambda f: (0 if "UPDATED" in f.name.upper() else 1,
                                        len(f.name)))[0]


def load_rp_budget(proj: str) -> Tuple[Dict[str, float], str]:
    """RP budget by cost code from the takeoff's 'Cost Gral' sheet, filtered
    to the project's side of the RP#/RP#-FTW split."""
    tk = _find_rp_takeoff(proj)
    if not tk:
        return {}, ""
    out: Dict[str, float] = {}
    try:
        wb = openpyxl.load_workbook(str(tk), read_only=True, data_only=True)
    except Exception:
        return {}, ""
    target = next((nm for nm in wb.sheetnames
                   if "cost" in nm.lower() and "gral" in nm.lower()), None)
    if not target:
        wb.close()
        return {}, ""
    is_ftw = proj.upper().endswith("-FTW")
    for row_vals in wb[target].iter_rows(min_row=2, max_col=4, values_only=True):
        code = str(row_vals[0] or "").strip().upper()
        if not code or not _is_cost_code(code):
            continue                      # section totals / headers interleave
        if code.startswith("FW") != is_ftw:
            continue                      # FW ↔ -FTW project; SL/PR ↔ slab
        try:
            qty = float(row_vals[2]) if row_vals[2] is not None else 1.0
            amt = float(row_vals[3])
        except (TypeError, ValueError):
            continue                      # '#N/A' cache etc. — skip, no guess
        eff = round(amt * (qty if qty in (0.0, 1.0) else 1.0), 2)
        if abs(eff) > 0.005:
            out[code] = out.get(code, 0.0) + eff
    wb.close()
    return out, tk.name

def costs_by_code(bills: List[dict], purchases: List[dict], customer_id: str,
                  parent_map: Dict[str, str],
                  account_names: Optional[Dict[str, str]] = None) -> Dict[str, dict]:
    """ALL of this project's cost lines grouped by their cost-code leaf via the
    shared `cost_leaf()` resolver — IDENTICAL keys to the accumulating-costs
    block, so item-based lines land on their cost code (SL1, PV6…) and join the
    takeoff budget. Non-code accounts keep their name → 'not budgeted' rows.

    Returns { code: {"total": float, "txns": [ {ref, vendor, date, amount,
    txn_id, tx_type, memo, desc} ]} } so Budget vs Actual can list each
    transaction under its code with a QBO deep-link (the user 2026-07-17)."""
    account_names = account_names or {}
    out: Dict[str, dict] = {}
    sources = ([(b, "Bill", "VendorRef") for b in bills]
               + [(p, "Expense", "EntityRef") for p in purchases])
    for txn, tx_type, vfield in sources:
        vendor = _xml_clean(((txn.get(vfield) or {}).get("name") or "(no vendor)").strip())
        ref = _xml_clean(str(txn.get("DocNumber") or txn.get("Id") or ""))
        date = txn.get("TxnDate", "")
        txn_id = txn.get("Id", "")
        memo = _xml_clean((txn.get("PrivateNote") or "").strip())
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if abs(amt) < 0.005:
                continue
            leaf = cost_leaf(det, account_names)
            # Class (Residential / Commercial / Multi Family) for the class-
            # reconciliation check — line ClassRef first, then the txn's.
            cls = ((det.get("ClassRef") or ln.get("ClassRef")
                    or txn.get("ClassRef") or {}).get("name") or "")
            g = out.setdefault(leaf, {"total": 0.0, "txns": []})
            g["total"] += amt
            g["txns"].append({
                "ref": ref, "vendor": vendor, "date": date, "amount": amt,
                "txn_id": txn_id, "tx_type": tx_type, "memo": memo,
                "class": _xml_clean(cls),
                "desc": _xml_clean((ln.get("Description") or "").strip())})
    for g in out.values():
        g["total"] = round(g["total"], 2)
    return out


# Ready-mix and rebar vendors bill sales tax on its OWN line ("TAXES"), so it
# has to come out of budget-vs-actual and be shown as its own sum — otherwise
# it silently inflates the actual against a takeoff budget that is pre-tax
# (the user 2026-07-29).
_TAX_LINE_RE = re.compile(r"\btax(es)?\b", re.IGNORECASE)
# The fuel surcharge SHOULD be its own line the way tax is. Today AP folds it
# into the per-yard rate, so this matches nothing on most jobs — the column is
# there so the moment the clerks code it correctly it lands where it belongs
# (the user 2026-07-29).
# Martin Marietta bills the surcharge as "SERVICE CHARGE" (the user
# 2026-08-01) — same animal, same column, lumped with fuel.
_FUEL_LINE_RE = re.compile(r"fuel|surcharge|energy\s*fee|service\s*charge",
                           re.IGNORECASE)
_BEFORE_KEY, _AFTER_KEY = "__before", "__after"


def code_costs_by_draw(
    bills: List[dict], purchases: List[dict], customer_id: str,
    draw_periods: List[Tuple[str, dt.date, dt.date]],
    account_names: Optional[Dict[str, str]] = None,
) -> Dict[str, dict]:
    """Every project cost LINE keyed by COST CODE **and** draw window, keeping
    each line's qty/rate so the Concrete sheet can do yards and $/yd, and
    tagging sales-tax and fuel-surcharge lines (the user 2026-07-29).

    Returns { code: {"draws":…, "tax":…, "fuel":…, "lines":[…]} } where
    key is a draw label, "__before" (dated before the first draw window — the
    pre-tagging era) or "__after" (past the last window — accumulating toward
    the next draw). Non-tax and tax dollars are kept apart so every column of
    the Labor/Concrete grid ties to the sum of its lines."""
    account_names = account_names or {}
    anchor = min((s for _, s, _ in draw_periods), default=None)

    def window_for(d: Optional[dt.date]) -> str:
        if not d:
            return _AFTER_KEY
        for lbl, s, e in draw_periods:
            if s <= d <= e:
                return lbl
        if anchor and d < anchor:
            return _BEFORE_KEY
        return _AFTER_KEY

    out: Dict[str, dict] = {}
    sources = ([(b, "Bill", "VendorRef") for b in bills]
               + [(p, "Expense", "EntityRef") for p in purchases])
    for txn, tx_type, vfield in sources:
        vendor = _xml_clean(((txn.get(vfield) or {}).get("name") or "(no vendor)").strip())
        date = txn.get("TxnDate", "")
        key = window_for(_parse_date(date))
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if abs(amt) < 0.005:
                continue
            desc = _xml_clean((ln.get("Description") or "").strip())
            is_fuel = bool(_FUEL_LINE_RE.search(desc))
            is_tax = not is_fuel and bool(_TAX_LINE_RE.search(desc))
            leaf = cost_leaf(det, account_names)
            g = out.setdefault(leaf, {"draws": {}, "tax": {}, "fuel": {},
                                      "lines": []})
            bucket = g["fuel"] if is_fuel else (g["tax"] if is_tax else g["draws"])
            bucket[key] = round(bucket.get(key, 0.0) + amt, 2)

            def _f(v):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return None

            g["lines"].append({
                "date": date, "vendor": vendor, "desc": desc, "draw": key,
                "amount": round(amt, 2), "is_tax": is_tax, "is_fuel": is_fuel,
                "qty": _f(det.get("Qty")), "rate": _f(det.get("UnitPrice")),
                "doc": _xml_clean(str(txn.get("DocNumber") or "")),
                "txn_id": txn.get("Id", ""), "tx_type": tx_type,
            })
    return out


# Which cost-code NUMBER each focus sheet tracks (SL6/PV6/CS6/MS6 = labor,
# SL1/PV1/CS1/MS1 = concrete). The PM and ops manager track these two by
# scrutiny — materials are bought as packages (the user 2026-07-29).
_FOCUS_NUM = {"Labor": "6", "Concrete": "1"}
# The P&L COGS line each trade's grand total must equal — the scoreboard says
# so explicitly, which is what makes the roll-up auditable (the user 2026-07-29
# "where is that adding up to?").
_FOCUS_TIE = {"Labor": "Subcontractors Expense: Labor",
              "Concrete": "Job Materials: Concrete"}
# A concrete line only counts toward YARDS when it was actually bought by the
# yard. Lump vendor bills (qty 1 × a ~$8.9k lump) would wreck the $/yd math, so
# they're carried in their own bucket and named on the sheet.
_YARD_DESC_RE = re.compile(r"\byd|yard", re.IGNORECASE)


def _is_yard_line(ln: dict) -> bool:
    q = ln.get("qty")
    if ln.get("is_tax") or not q or q <= 1:
        return False
    return bool(_YARD_DESC_RE.search(ln.get("desc") or "")) or q >= 5


def _merge_bill_lines(lines: List[dict]) -> List[dict]:
    """Fold a bill's SALES TAX and FUEL SURCHARGE lines onto the purchase line
    they belong to (the user 2026-07-29 — 'I want that amount on the same line
    item it came from'). The vendors bill them as separate QBO lines on the
    SAME bill, so the bill number is the join.

    A bill with one purchase line absorbs its tax/fuel. A bill with several
    can't be attributed, so its purchase lines stay clean and the tax/fuel
    rides one extra row — nothing is silently dropped or double-counted."""
    by_txn: Dict[str, List[dict]] = {}
    for ln in lines:
        by_txn.setdefault(ln["txn_id"], []).append(ln)
    out: List[dict] = []
    for group in by_txn.values():
        main = [l for l in group if not l["is_tax"] and not l["is_fuel"]]
        tax = round(sum(l["amount"] for l in group if l["is_tax"]), 2)
        fuel = round(sum(l["amount"] for l in group if l["is_fuel"]), 2)
        if len(main) == 1:
            m = dict(main[0]); m["tax"], m["fuel"] = tax, fuel
            out.append(m)
            continue
        for m in main:
            mm = dict(m); mm["tax"], mm["fuel"] = 0.0, 0.0
            out.append(mm)
        if tax or fuel:
            o = dict(group[0])
            o.update({"amount": 0.0, "qty": None, "rate": None,
                      "tax": tax, "fuel": fuel,
                      "desc": "(tax / surcharge on this bill)"})
            out.append(o)
    return sorted(out, key=lambda x: (x["date"], str(x.get("doc") or "")))


def _autofit(ws, first_col: int, last_col: int, header_rows: List[int],
             include: List[Tuple[int, int]],
             skip_cols: Tuple[int, ...] = (), min_w: float = 9.0,
             max_w: float = 46.0) -> None:
    """What Excel's double-click autofit would do, computed at write time
    (the user 2026-07-31 — the file must ARRIVE fitted; openpyxl has no
    renderer, so widths are estimated from the longest display line).

    Only rows inside the `include` (start, end) ranges are measured — the
    long single-cell note lines (fuel warning, tie-out sentence, lump notes)
    deliberately SPILL across columns and would otherwise blow their column
    to max width. Width unit ≈ one '0' at Calibri 11; body font is 12, so
    chars scale by 12/11 plus padding. Wrapped headers measure per-LINE and
    the header row height is set from the deepest line count. Formula cells
    display numbers, not formula text — estimated at 13 chars."""
    SCALE, PAD = 12.0 / 11.0, 2.0

    def _in(r):
        return any(a <= r <= b for a, b in include)

    deepest = {r: 1 for r in header_rows}
    for col in range(first_col, last_col + 1):
        if col in skip_cols:
            continue
        longest = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            cell = row[0]
            v = cell.value
            if v is None or not _in(cell.row):
                continue
            if isinstance(v, str):
                if v.startswith("="):
                    longest = max(longest, 13)
                    continue
                lines = v.split("\n")
                longest = max(longest, max(len(l) for l in lines))
                if cell.row in deepest and len(lines) > 1:
                    deepest[cell.row] = max(deepest[cell.row], len(lines))
            elif isinstance(v, (int, float)):
                longest = max(longest, len(f"{v:,.2f}"))
        if longest:
            ws.column_dimensions[get_column_letter(col)].width = round(
                min(max_w, max(min_w, longest * SCALE + PAD)), 1)
    for r, n in deepest.items():
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0,
                                          n * 19 + 5)


def build_sheet_labor_concrete(
    wb, kind: str, proj: str, cust_info: dict, wip_info: dict,
    budget: Dict[str, float], code_costs: Dict[str, dict],
    draw_cols: List[Tuple[str, str]], as_of: str,
    yards: Tuple[float, Dict[str, float], str] = (0.0, {}, ""),
    budget_source: str = "", realm: str = "",
    att_links: Optional[Dict[str, dict]] = None,
    marks: Optional[Dict[tuple, str]] = None,
) -> Optional[str]:
    """LABOR / CONCRETE — TWO blocks at different altitudes (the user
    2026-07-29: 'the metrics are a top-level data point'; one grid trying to be
    both is what produced the empty cells and hidden rows).

    SCOREBOARD (top, frozen): one row per cost code — budget, actual, balance,
    then the total per draw. Every cell populated. Concrete adds the sales-tax
    roll-up and the yards/$-per-yd strip, and the grand total names the P&L
    line it ties to.

    LEDGER (below, fully expanded): every bill — QBO # first (the user's
    identifier), then date, vendor, description, qty, rate, amount, its own
    tax (folded onto the bill row by bill #), and WHICH draw it landed in as a
    label. A bill belongs to exactly one draw, so a label column beats a draw
    matrix that guarantees blank cells on every row. Tax/fuel columns appear
    only on a trade that has such lines (labor subs bill neither).
    Returns the sheet name, or None when the trade has no budget and no cost."""
    num = _FOCUS_NUM[kind]
    codes = sorted(
        {c for c in budget if _split_code(c)[1] == num}
        | {c for c in code_costs if _split_code(c)[1] == num},
        key=_cost_code_sort_key)
    if not codes:
        return None

    def _grp(c):
        return code_costs.get(c) or {"draws": {}, "tax": {}, "fuel": {}, "lines": []}

    col_keys: List[Tuple[str, str]] = []
    if any(_grp(c)["draws"].get(_BEFORE_KEY) for c in codes):
        col_keys.append((_BEFORE_KEY, "Before draw 1"))
    col_keys += list(draw_cols)
    if any(_grp(c)["draws"].get(_AFTER_KEY) for c in codes):
        col_keys.append((_AFTER_KEY, "Outside draw\nwindows"))
    draw_name = {k: h.split("\n")[0] for k, h in col_keys}
    draw_name[_BEFORE_KEY] = "Before draw 1"
    draw_name[_AFTER_KEY] = "Outside windows"

    has_tax = any(_grp(c)["tax"] for c in codes)
    has_fuel = any(_grp(c)["fuel"] for c in codes)
    tie_line = _FOCUS_TIE.get(kind, "")

    ws = wb.create_sheet(kind)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    SZ = BASE_SIZE                                    # flat 12 — no 11pt anywhere
    ROW_H = 17

    # Shared column widths serve both blocks: A item/QBO#, B budget/date,
    # C actual/vendor, D balance/description, E bal%/qty, F../draws + rate,
    # amount, tax, draw-label.
    n_draws = len(col_keys)
    sb_tax = 7 + n_draws                     # scoreboard TAX col (after draws)
    sb_fuel = sb_tax + (1 if has_tax else 0)
    sb_incl = sb_fuel + (1 if has_fuel else 0)
    # Col 1 is the ledger's ↗ lane and must stay GENUINELY NARROW (the user
    # 2026-08-10 — a wide shared column with floating arrows is slop); the
    # scoreboard starts at col 2.
    widths = {1: 4.5, 2: 34, 3: 16, 4: 20, 5: 16, 6: 12}
    for i in range(n_draws):
        widths[7 + i] = 18                   # fits "01/26/26–02/25/26" unclipped
    if has_tax:
        widths[sb_tax] = 14
    if has_fuel:
        widths[sb_fuel] = 14
    if has_tax or has_fuel:
        widths[sb_incl] = 17
    # ledger fixed columns (overlap the same sheet columns); DESCRIPTION is
    # LAST so it can spill right over empty cells instead of owning a width.
    # Col 1 is the QBO-page arrow (the user 2026-08-10: the bill # opens the
    # SCAN, so the direct QBO link needed its own little column up front).
    L_ARROW = 1
    L_QBO, L_DATE, L_VEND, L_QTY, L_RATE, L_AMT = 2, 3, 4, 5, 6, 7
    L_TAX = 8 if has_tax else 0
    L_FUEL = (9 if has_tax else 8) if has_fuel else 0
    L_DRAW = max(L_AMT, L_TAX, L_FUEL) + 1
    L_DESC = L_DRAW + 1
    for col, w in ((L_QTY, 12), (L_RATE, 13), (L_AMT, 15), (L_DRAW, 16),
                   (L_DESC, 24)):
        widths[col] = max(widths.get(col, 0), w)
    if L_TAX:
        widths[L_TAX] = max(widths.get(L_TAX, 0), 13)
    if L_FUEL:
        widths[L_FUEL] = max(widths.get(L_FUEL, 0), 13)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    sub = ws.cell(row=r, column=2, value=(
        f"{kind.upper()} — SCOREBOARD (budget vs actual by cost code and draw), "
        f"then the LEDGER: every bill, fully expanded."))
    sub.font = Font(bold=True, size=SZ, color="1F3A5F")
    r += 1
    if budget_source:
        src = ws.cell(row=r, column=2, value=f"Budget source: {budget_source}")
        src.font = Font(italic=True, size=SZ, color="595959")
        r += 1
    r += 1

    # ───────────────────────── SCOREBOARD ─────────────────────────
    F_DRAW = HDR_FILL
    F_OUT = PatternFill("solid", fgColor="7F7F7F")
    F_TAXF = PatternFill("solid", fgColor="7F6000")
    hdr = r
    heads = ([("ITEM (cost code)", 2, F_DRAW), ("BUDGET", 3, F_DRAW),
              ("ACTUAL" + (" (ex-tax)" if has_tax else ""), 4, F_DRAW),
              ("BALANCE $", 5, F_DRAW), ("BALANCE %", 6, F_DRAW)]
             + [(h, 7 + i, F_OUT if k in (_BEFORE_KEY, _AFTER_KEY) else F_DRAW)
                for i, (k, h) in enumerate(col_keys)])
    if has_tax:
        heads.append(("SALES TAX", sb_tax, F_TAXF))
    if has_fuel:
        heads.append(("FUEL / SVC CHARGE", sb_fuel, F_TAXF))
    if has_tax or has_fuel:
        heads.append(("ACTUAL INCL.", sb_incl, F_TAXF))
    for text, col, fill in heads:
        c = ws.cell(row=hdr, column=col, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=SZ)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[hdr].height = 34
    r += 1

    first_code = r
    for code in codes:
        g = _grp(code)
        band = _cost_band_fill(code)
        ws.cell(row=r, column=2, value=_cost_code_label(code)
                ).font = Font(bold=True, size=SZ)
        bcell = ws.cell(row=r, column=3, value=round(budget.get(code, 0.0), 2))
        bcell.number_format = ACC_FMT
        bcell.font = Font(size=SZ)
        for i, (key, _h) in enumerate(col_keys):
            c = ws.cell(row=r, column=7 + i,
                        value=round(g["draws"].get(key, 0.0), 2))
            c.number_format = ACC_FMT
            c.font = Font(size=SZ)
        ac = ws.cell(row=r, column=4, value="=" + "+".join(
            f"{get_column_letter(7 + i)}{r}" for i in range(n_draws)))
        ac.number_format = ACC_FMT
        ac.font = Font(bold=True, size=SZ)
        blc = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        blc.number_format = ACC_FMT
        blc.font = Font(bold=True, size=SZ)
        pc = ws.cell(row=r, column=6, value=f'=IF(C{r}=0,"",E{r}/C{r})')
        pc.number_format = PCT_FMT
        pc.font = Font(bold=True, size=SZ)
        if has_tax:
            tcell = ws.cell(row=r, column=sb_tax,
                            value=round(sum(g["tax"].values()), 2))
            tcell.number_format = ACC_FMT
            tcell.font = Font(size=SZ)
        if has_fuel:
            fcell = ws.cell(row=r, column=sb_fuel,
                            value=round(sum(g["fuel"].values()), 2))
            fcell.number_format = ACC_FMT
            fcell.font = Font(size=SZ)
        if has_tax or has_fuel:
            parts = [f"D{r}"] + ([f"{get_column_letter(sb_tax)}{r}"] if has_tax else []) \
                    + ([f"{get_column_letter(sb_fuel)}{r}"] if has_fuel else [])
            ic = ws.cell(row=r, column=sb_incl, value="=" + "+".join(parts))
            ic.number_format = ACC_FMT
            ic.font = Font(size=SZ)
        last_sb = sb_incl if (has_tax or has_fuel) else 6 + n_draws
        for col in range(2, last_sb + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = band
            cell.border = THIN_BORDER
            if cell.font is None or cell.font.size is None:
                cell.font = Font(size=SZ)
        ws.row_dimensions[r].height = ROW_H
        r += 1
    last_code = r - 1

    tot = r
    ws.cell(row=tot, column=2, value=f"TOTAL {kind.upper()}").font = TOTAL_FONT
    last_sb = sb_incl if (has_tax or has_fuel) else 6 + n_draws
    for col in range(2, last_sb + 1):
        cell = ws.cell(row=tot, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
    for col in ([3, 4] + [7 + i for i in range(n_draws)]
                + ([sb_tax] if has_tax else [])
                + ([sb_fuel] if has_fuel else [])
                + ([sb_incl] if (has_tax or has_fuel) else [])):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{first_code}:{L}{last_code})")
        c.number_format = ACC_FMT
    ws.cell(row=tot, column=5, value=f"=C{tot}-D{tot}").number_format = ACC_FMT
    ws.cell(row=tot, column=6,
            value=f'=IF(C{tot}=0,"",E{tot}/C{tot})').number_format = PCT_FMT
    ws.row_dimensions[tot].height = ROW_H
    r += 1

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for col in (5, 6):
        L = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{L}{first_code}:{L}{tot}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=red, font=Font(color="9C0006", bold=True)))


    known = _project_name_words(cust_info.get("name", ""))
    if kind == "Concrete":
        # Top-right beside the title, the user's own arrangement (2026-08-04)
        # — the strip stopped being a band between scoreboard and ledger.
        _yards_strip(ws, 1, 10, codes, code_costs, budget, yards, SZ)

    r += 2

    # ───────────────────────── LEDGER ─────────────────────────
    n_marks_kept = [0]
    r_ledger_hdr = r + 1                       # the QBO#/DATE/… header row
    lh = ws.cell(row=r, column=2, value=(
        "LEDGER — every bill; the DRAW column says which draw window the bill "
        "date fell in. QBO # opens the uploaded bill file (attachments/ beside "
        "this workbook); '(N files)' opens that bill's own scan folder. "
        "No attachment → the QBO bill page. MARK A ROW GREEN when the PM "
        "confirms the bill — the mark survives every re-sync."))
    lh.font = Font(bold=True, size=SZ, color="1F3A5F")
    r += 1
    led_heads = [("↗", L_ARROW), ("QBO #", L_QBO), ("DATE", L_DATE),
                 ("VENDOR", L_VEND),
                 ("QTY", L_QTY), ("RATE", L_RATE), ("AMOUNT", L_AMT)]
    if L_TAX:
        led_heads.append(("SALES TAX", L_TAX))
    if L_FUEL:
        led_heads.append(("FUEL/SVC", L_FUEL))
    led_heads += [("DRAW", L_DRAW), ("DESCRIPTION", L_DESC)]
    for text, col in led_heads:
        c = ws.cell(row=r, column=col, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=SZ)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal=("left" if col == L_DESC else "center"),
                                vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[r].height = ROW_H + 3
    r += 1

    for code in codes:
        g = _grp(code)
        merged = _merge_bill_lines(g["lines"])
        if not merged:
            continue
        band = _cost_band_fill(code)
        bc = ws.cell(row=r, column=2, value=_cost_code_label(code))
        bc.font = Font(bold=True, size=SZ)
        sub_amt = round(sum(l["amount"] for l in merged), 2)
        sc = ws.cell(row=r, column=L_AMT, value=sub_amt)
        sc.number_format = ACC_FMT
        sc.font = Font(bold=True, size=SZ)
        if L_TAX:
            stx = ws.cell(row=r, column=L_TAX,
                          value=round(sum(l.get("tax", 0) for l in merged), 2))
            stx.number_format = ACC_FMT
            stx.font = Font(bold=True, size=SZ)
        if L_FUEL:
            sfl = ws.cell(row=r, column=L_FUEL,
                          value=round(sum(l.get("fuel", 0) for l in merged), 2))
            sfl.number_format = ACC_FMT
            sfl.font = Font(bold=True, size=SZ)
        for col in range(1, L_DESC + 1):
            ws.cell(row=r, column=col).fill = band
            ws.cell(row=r, column=col).border = THIN_BORDER
        ws.row_dimensions[r].height = ROW_H
        r += 1
        for ln in merged:
            # ↗ = the QBO bill page, always (the user 2026-08-10).
            _qurl = _qbo_txn_url(ln["tx_type"], ln["txn_id"], realm)
            arr = ws.cell(row=r, column=L_ARROW, value="↗")
            arr.alignment = Alignment(horizontal="center")
            if _qurl:
                arr.hyperlink = _qurl
                arr.font = Font(color="0563C1", underline="single", size=SZ)
            else:
                arr.font = Font(size=SZ, color="BFBFBF")
            # The uploaded bill file when QBO has one (the user 2026-07-31 —
            # "just want the straight attachment"); the QBO bill page
            # otherwise. Several scans → the attachments folder, count shown.
            # STORED RELATIVE link, deliberately (tested on the Mac
            # 2026-07-31): =HYPERLINK() formulas hard-fail in Mac Excel's
            # sandbox ("Cannot open the specified file", every URL form), a
            # HyperlinkBase property breaks resolution too, while a stored
            # relative target opens the file (one-time Grant File Access per
            # file) AND survives a Mac Excel save unrewritten — verified by
            # saving and re-reading the sheet rels. Windows resolves the same
            # relative target against the share path it opened from.
            _att = (att_links or {}).get(ln["txn_id"])
            url = (_att["link"] if _att
                   else _qbo_txn_url(ln["tx_type"], ln["txn_id"], realm))
            _lbl = str(ln["doc"] or ln["txn_id"] or "(no #)")
            if _att and _att["n"] > 1:
                _lbl += f"  ({_att['n']} files)"
            idc = ws.cell(row=r, column=L_QBO, value=_lbl)
            if url:
                idc.hyperlink = url
                idc.font = Font(color="0563C1", underline="single", size=SZ)
            else:
                idc.font = Font(size=SZ)
            ws.cell(row=r, column=L_DATE, value=ln["date"]).font = Font(size=SZ)
            ws.cell(row=r, column=L_VEND, value=ln["vendor"]).font = Font(size=SZ)
            dc = ws.cell(row=r, column=L_DESC,
                         value=_clean_cost_text(ln["desc"], known) or ln["desc"])
            dc.font = Font(size=SZ)                # last column — spills right
            if ln["qty"]:
                ws.cell(row=r, column=L_QTY,
                        value=ln["qty"]).number_format = "#,##0.00"
            if ln["rate"]:
                ws.cell(row=r, column=L_RATE,
                        value=ln["rate"]).number_format = ACC_FMT
            amt_c = ws.cell(row=r, column=L_AMT, value=ln["amount"])
            amt_c.number_format = ACC_FMT
            if L_TAX and ln.get("tax"):
                ws.cell(row=r, column=L_TAX,
                        value=ln["tax"]).number_format = ACC_FMT
            if L_FUEL and ln.get("fuel"):
                ws.cell(row=r, column=L_FUEL,
                        value=ln["fuel"]).number_format = ACC_FMT
            ws.cell(row=r, column=L_DRAW,
                    value=draw_name.get(ln["draw"], ln["draw"])).font = Font(size=SZ)
            for col in (L_QTY, L_RATE, L_AMT, L_TAX, L_FUEL):
                if col:
                    ws.cell(row=r, column=col).font = Font(size=SZ)
            # Re-apply the estimator's row mark from the previous version —
            # green = the PM confirmed this bill (the user 2026-07-31). The
            # exact color is preserved, so other conventions survive too.
            #
            # INVARIANT (binding): the script must NEVER write a direct cell
            # fill on a dated bill row — that is what makes every fill on
            # such a row human-owned, regardless of color. Script coloring
            # on data rows goes through CONDITIONAL FORMATTING instead (a
            # separate xlsx layer the mark readback cannot see). Breaking
            # this turns script decoration into phantom "PM confirmations".
            _mk = (marks or {}).get(
                (str(ln["doc"] or ln["txn_id"] or "(no #)").strip(),
                 str(ln["date"]).strip(), str(ln["vendor"]).strip(),
                 round(float(ln["amount"]), 2)))
            if _mk:
                _mfill = PatternFill("solid", fgColor=_mk)
                for col in range(1, L_DESC + 1):
                    ws.cell(row=r, column=col).fill = _mfill
                n_marks_kept[0] += 1
            ws.row_dimensions[r].height = ROW_H
            r += 1
    # Deliver fitted: every column sized to its longest line, header rows to
    # their line count. DESCRIPTION is excluded — it spills right by design.
    # Measured rows: the scoreboard table, the yards strip, and the ledger —
    # never the spilling note lines.
    _autofit(ws, 2, L_DESC - 1, [hdr, r_ledger_hdr],
             [(hdr, tot), (r_ledger_hdr, ws.max_row)])
    if n_marks_kept[0] or marks:
        kept, had = n_marks_kept[0], len(marks or {})
        ui_event(f"{kind}: {kept} PM-confirmed row mark(s) preserved"
                 + (f"  ⚑ {had - kept} mark(s) no longer match a row"
                    if had > kept else ""),
                 icon="·", color=(_YEL if had > kept else ""))
    _setup_print(ws, L_DESC)
    return ws.title


def _yards_strip(ws, row: int, col0: int, codes: List[str],
                 code_costs: Dict[str, dict], budget: Dict[str, float],
                 yards: Tuple[float, Dict[str, float], str], sz: int) -> None:
    """Concrete's yards / $-per-yd comparison, parked TOP-RIGHT beside the
    title (the user's own arrangement, 2026-08-04) — labels at `row`, values
    beneath, the lump-bill note under those. The takeoff's implied rate
    (concrete $ ÷ CONCRETE YARDS total) vs what QBO actually paid per yard,
    ex-tax; lump bills with no yardage are excluded from the rate (they'd
    fake the $/yd) and remain fully visible in the ledger."""
    bud_yards, _rows, _src = yards
    lines = [ln for c in codes for ln in (code_costs.get(c) or {}).get("lines", [])]
    yard_lines = [ln for ln in lines if _is_yard_line(ln)]
    lump = [ln for ln in lines
            if not ln["is_tax"] and not ln.get("is_fuel") and not _is_yard_line(ln)]
    act_yards = round(sum(ln["qty"] for ln in yard_lines), 2)
    act_dollars = round(sum(ln["amount"] for ln in yard_lines), 2)
    bud_dollars = round(sum(budget.get(c, 0.0) for c in codes), 2)
    rate_b = round(bud_dollars / bud_yards, 2) if bud_yards else 0.0
    rate_a = round(act_dollars / act_yards, 2) if act_yards else 0.0

    cells = [("BUDGET YARDS", bud_yards, "#,##0.00"),
             ("BUDGET $/YD (implied)", rate_b, ACC_FMT),
             ("ACTUAL YARDS", act_yards, "#,##0.00"),
             ("ACTUAL $/YD paid", rate_a, ACC_FMT),
             ("VAR $/YD", round(rate_b - rate_a, 2), ACC_FMT),
             ("VAR YARDS", round(act_yards - bud_yards, 2), "#,##0.00")]
    for i, (label, _v, _f) in enumerate(cells):
        col = col0 + i
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(bold=True, size=sz - 1, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="44546A")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = THIN_BORDER
        L = get_column_letter(col)
        if (ws.column_dimensions[L].width or 8.43) < 15:
            ws.column_dimensions[L].width = 15
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 32)
    for i, (_l, value, fmt) in enumerate(cells):
        c = ws.cell(row=row + 1, column=col0 + i, value=value)
        c.number_format = fmt
        c.font = Font(bold=True, size=sz)
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    if lump:
        w = ws.cell(row=row + 2, column=col0, value=(
            f"⚑ {len(lump)} bill line(s) carry no yardage "
            f"(${sum(l['amount'] for l in lump):,.2f}) — in every total, "
            f"excluded from $/yd; see them in the ledger below"))
        w.font = Font(italic=True, size=sz - 1, color="9C5700")


def build_sheet_budget_vs_actual(wb, proj, cust_info, wip_info,
                                 budget: Dict[str, float],
                                 actuals: Dict[str, dict],
                                 as_of: str, co_flag: bool = False,
                                 budget_source: str = "", realm: str = "") -> Optional[str]:
    """BUDGET (takeoff cost codes) vs ACTUAL (QBO cost-code totals) per line,
    with EVERY actual transaction listed under its code for audit — name /
    date / amount / QBO link (the user 2026-07-17). CP budget = the takeoff's
    Cost Code sheet; RP budget = the takeoff's last sheet. Jobs WITH change
    orders get a warning banner (CO costs aren't in the budget). Returns the
    sheet name, or None when no budget."""
    if not budget:
        return None
    ws = wb.create_sheet("Budget vs Actual")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.sheet_properties.outlinePr.summaryBelow = False   # code row sits ABOVE its txns
    for col, w in (("A", 48), ("B", 16), ("C", 16), ("D", 16), ("E", 10)):
        ws.column_dimensions[col].width = w
    known = _project_name_words(cust_info.get("name", ""))
    exp_cls = _expected_class(proj)
    n_class_flags = sum(1 for g in actuals.values() for t in g.get("txns", [])
                        if not _class_ok(t.get("class"), exp_cls))

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    if budget_source:
        src = ws.cell(row=r, column=1, value=f"Budget source: {budget_source}")
        src.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
        r += 1
    if n_class_flags:
        cf = ws.cell(row=r, column=1, value=(
            f"⚑ CLASS CHECK: {n_class_flags} transaction(s) are not classed "
            f"'{exp_cls}' — flagged red in the last column; fix the Class in QBO"))
        cf.font = Font(bold=True, size=BASE_SIZE - 1, color="C00000")
        for cc in range(1, 6):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="FCE4D6")
        r += 1
    if co_flag:
        warn = ws.cell(row=r, column=1, value=(
            "⚠ MAY BE INACCURATE — this job has change orders; CO costs are "
            "NOT in the budget (CO-template cost line pending), so actuals "
            "can exceed budget for CO work"))
        warn.font = Font(bold=True, size=BASE_SIZE - 1, color="9C5700")
        for cc in range(1, 6):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="FFF2CC")
        r += 1
    tip = ws.cell(row=r, column=1, value=(
        "Each cost code expands to its bills — click the ± in the margin; the "
        "QBO ↗ button opens the bill."))
    tip.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    hdr = ws.cell(row=r, column=1, value="BUDGET vs ACTUAL — by cost code")
    hdr.font = Font(bold=True, size=BASE_SIZE, color="FFFFFF")
    for cc in range(1, 6):
        ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    for c, h in ((1, "Cost Code  /  transaction"), (2, "Budget  /  date"),
                 (3, "Actual (QBO)  /  amount"), (4, "Variance  /  link"),
                 (5, "Used %")):
        hc = ws.cell(row=r, column=c, value=h)
        hc.font = Font(bold=True, size=BASE_SIZE - 1, color=NAVY)
        hc.border = BOTTOM_BORDER
        if c > 1:
            hc.alignment = Alignment(horizontal="right" if c < 5 else "center")
    r += 1

    # Row set = union; codes in Cost-Code-Sheet order, non-code actuals last.
    codes = sorted(set(budget) | set(actuals), key=_cost_code_sort_key)
    code_rows = []                       # only these feed the TOTAL (no double-count)
    for code in codes:
        b = budget.get(code)
        grp = actuals.get(code) or {}
        a = grp.get("total", 0.0)
        code_row = r
        code_rows.append(code_row)
        lc = ws.cell(row=r, column=1, value=_cost_code_label(code))
        lc.font = Font(bold=True, size=BASE_SIZE - 1,
                       color="000000" if b is not None else "9C5700")
        if b is not None:
            bc = ws.cell(row=r, column=2, value=round(b, 2))
            bc.number_format = CURR_FMT
            bc.font = Font(bold=True, size=BASE_SIZE - 1)
        else:
            nb = ws.cell(row=r, column=2, value="not budgeted")
            nb.font = Font(italic=True, size=BASE_SIZE - 2, color="9C5700")
            nb.alignment = Alignment(horizontal="right")
        ac = ws.cell(row=r, column=3, value=round(a, 2))
        ac.number_format = CURR_FMT
        ac.font = Font(bold=True, size=BASE_SIZE - 1)
        vc = ws.cell(row=r, column=4, value=f"=B{r}-C{r}" if b is not None else None)
        vc.number_format = CURR_FMT
        vc.font = Font(bold=True, size=BASE_SIZE - 1)
        pc = ws.cell(row=r, column=5,
                     value=(f'=IF(B{r}=0,"",C{r}/B{r})' if b is not None else None))
        pc.number_format = "0%"
        pc.font = Font(bold=True, size=BASE_SIZE - 1)
        pc.alignment = Alignment(horizontal="center")
        # job-type band across the code row so trades group visually (the user
        # 2026-07-17). Transaction rows below stay white for contrast.
        _band = _cost_band_fill(code)
        for cc in range(1, 6):
            ws.cell(row=code_row, column=cc).fill = _band
        r += 1
        # ── transactions under the code: A name · B date · C amount · D link ──
        for t in sorted(grp.get("txns", []),
                        key=lambda x: (_parse_date(x.get("date", "")) or dt.date.min,
                                       abs(float(x.get("amount", 0) or 0))),
                        reverse=True):
            nm = _clean_cost_text(t.get("desc") or t.get("memo") or "", known)
            label = f"    #{t.get('ref', '')}  {t.get('vendor', '')}"
            if nm:
                label += f" — {nm}"
            nc = ws.cell(row=r, column=1, value=label)
            nc.font = Font(size=BASE_SIZE - 2, color="404040")
            dv = _parse_date(t.get("date", ""))
            dc = ws.cell(row=r, column=2, value=dv or t.get("date", ""))
            if dv:
                dc.number_format = "mm/dd/yyyy"
            dc.font = Font(size=BASE_SIZE - 2, color="404040")
            dc.alignment = Alignment(horizontal="right")
            amc = ws.cell(row=r, column=3, value=round(float(t.get("amount", 0) or 0), 2))
            amc.number_format = CURR_FMT
            amc.font = Font(size=BASE_SIZE - 2, color="404040")
            url = _qbo_txn_url(t.get("tx_type", ""), t.get("txn_id", ""), realm)
            lk = ws.cell(row=r, column=4, value="QBO ↗" if url else "")
            if url:
                lk.hyperlink = url
                lk.font = Font(size=BASE_SIZE - 2, color=LINK, underline="single", bold=True)
            lk.alignment = Alignment(horizontal="center")
            # class reconciliation: flag a line whose class ≠ the division class
            _cls = t.get("class")
            if not _class_ok(_cls, exp_cls):
                fc = ws.cell(row=r, column=5,
                             value=f"⚑ {_cls}" if _cls else "⚑ no class")
                fc.font = Font(size=BASE_SIZE - 2, bold=True,
                               color="C00000" if _cls else "9C5700")
                # tint the whole transaction row so it's easy to spot
                for cc in range(1, 6):
                    if not ws.cell(row=r, column=cc).fill.patternType:
                        ws.cell(row=r, column=cc).fill = PatternFill(
                            "solid", fgColor="FDECEA")
            ws.row_dimensions[r].outline_level = 1     # collapsible under the code
            r += 1
    tl = ws.cell(row=r, column=1, value="TOTAL")
    tl.font = Font(bold=True, size=BASE_SIZE - 1)
    tl.border = TOP_BORDER
    # SUM (not a + chain) over the explicit code-row cells: it skips the
    # transaction rows (their col B holds a DATE, col C an amount) AND ignores
    # the "not budgeted" TEXT in some B cells — a + chain #VALUE!-errors on that
    # text (the user 2026-07-17).
    _sum_b = ("=SUM(" + ",".join(f"B{cr}" for cr in code_rows) + ")") if code_rows else "=0"
    _sum_c = ("=SUM(" + ",".join(f"C{cr}" for cr in code_rows) + ")") if code_rows else "=0"
    for c, f, fmt in ((2, _sum_b, CURR_FMT), (3, _sum_c, CURR_FMT),
                      (4, f"=B{r}-C{r}", CURR_FMT),
                      (5, f'=IF(B{r}=0,"",C{r}/B{r})', "0%")):
        tc = ws.cell(row=r, column=c, value=f)
        tc.number_format = fmt
        tc.font = Font(bold=True, size=BASE_SIZE - 1)
        tc.border = TOP_BORDER
        if c == 5:
            tc.alignment = Alignment(horizontal="center")
    # over-budget flag: Used % turns red past 100% (only the code rows carry E)
    if code_rows:
        ws.conditional_formatting.add(
            f"E{code_rows[0]}:E{r}",
            CellIsRule(operator="greaterThan", formula=["1"],
                       font=Font(bold=True, color="C00000")))
    _setup_print(ws, 5)
    return ws.title


def build_sheet_cashflow(wb, proj, cust_info, wip_info, events, as_of, realm=""):
    """CASH FLOW timeline (the user 2026-06-26) — the FUNDING lens, not the profit lens.
    One row per ACTUAL payment: AP cash-out (when we paid suppliers/subs) and AR
    cash-in (when the GC's money landed), chronological, with a running balance.
    The lowest running balance is the PEAK CASH REQUIREMENT — how far in the hole
    the job went before it paid itself back. (Profit is accrual; this is cash.)"""
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    for col, w in (("A", 13), ("B", 14), ("C", 28), ("D", 14),
                   ("E", 15), ("F", 15), ("G", 17)):
        ws.column_dimensions[col].width = w

    def wc(row, col, val, *, bold=False, color="000000", fmt=None, fill=None,
           indent=0, link=None, size=None):
        v = ("    " * indent + val) if isinstance(val, str) else val
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(bold=bold, size=size or (BASE_SIZE - 1),
                      color=LINK if link else color, underline="single" if link else None)
        if fmt:
            c.number_format = fmt
        if fill is not None:
            c.fill = fill
        if link:
            c.hyperlink = link
        return c

    def wdate(row, col, ds, *, bold=False):
        d = _parse_date(ds) if isinstance(ds, str) else (ds if isinstance(ds, dt.date) else None)
        c = ws.cell(row=row, column=col, value=d or ds)
        if d:
            c.number_format = "mm/dd/yyyy"
        c.font = Font(size=BASE_SIZE - 1, bold=bold)

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    wc(r, 1, f"CASH FLOW — actual payment timeline  ·  {proj}", bold=True,
       color=NAVY, size=BASE_SIZE + 3)
    r += 1
    wc(r, 1, "Dates are when cash MOVED (bill/invoice PAYMENT dates), not bill or "
             "invoice dates. This is the funding lens — profit is separate.",
       color="595959").font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    if not events:
        wc(r, 1, "No AP/AR payments recorded in QBO for this project window.",
           color="595959", bold=True)
        _setup_print(ws, 7)
        return ws

    trough = min(e["running"] for e in events)
    trough_ev = next(e for e in events if e["running"] == trough)
    total_out = round(sum(e["amount"] for e in events if e["kind"] == "out"), 2)
    total_in = round(sum(e["amount"] for e in events if e["kind"] == "in"), 2)
    final = events[-1]["running"]

    # ── summary box ──
    sx = r
    wc(r, 1, "SUMMARY", bold=True, color="FFFFFF", fill=HDR_FILL)
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = HDR_FILL
    r += 1

    def s(label, val, color="000000", datev=None):
        nonlocal r
        wc(r, 1, label, color="595959")
        if datev is not None:
            wdate(r, 5, datev, bold=True)
        wc(r, 7, val, fmt=CURR_FMT, bold=True, color=color)
        r += 1

    s("Peak cash needed (deepest in the hole)", trough,
      color=RED, datev=trough_ev["date"])
    s("Total paid out (AP)", -total_out, color="C55A11")
    s("Total received (AR)", total_in, color=GREEN)
    s("Net cash position now", final, color=(GREEN if final >= 0 else RED))
    for rr in range(sx, r):
        for cc in range(1, 8):
            cur = ws.cell(row=rr, column=cc).border
            ws.cell(row=rr, column=cc).border = Border(
                left=_THICK if cc == 1 else cur.left,
                right=_THICK if cc == 7 else cur.right,
                top=_THICK if rr == sx else cur.top,
                bottom=_THICK if rr == r - 1 else cur.bottom)
    r += 2

    # ── event table ──
    for c, h in ((1, "Date"), (2, "Cash event"), (3, "Party"), (4, "Ref"),
                 (5, "Paid out (AP)"), (6, "Received (AR)"), (7, "Running balance")):
        wc(r, c, h, bold=True, color=NAVY).border = BOTTOM_BORDER
    r += 1
    TROUGH_FILL = PatternFill("solid", fgColor="FCE4D6")   # light salmon
    for e in events:
        is_trough = (e is trough_ev)
        wdate(r, 1, e["date"])
        if e["kind"] == "out":
            wc(r, 2, "Paid AP", color="C55A11", bold=True)
            wc(r, 5, e["amount"], fmt=CURR_FMT, color="C55A11")
        else:
            wc(r, 2, "Received AR", color=GREEN, bold=True)
            wc(r, 6, e["amount"], fmt=CURR_FMT, color=GREEN)
        wc(r, 3, e["party"])
        wc(r, 4, str(e["ref"]) or "—",
           link=_qbo_txn_url(e["page"], e["doc_id"], realm))
        run = e["running"]
        wc(r, 7, run, fmt=CURR_FMT, bold=True, color=(RED if run < 0 else GREEN))
        # the deepest point: salmon-highlight the whole row + a marker in the Party
        # cell prefix (event/amount/running columns stay intact).
        if is_trough:
            for cc in range(1, 8):
                ws.cell(row=r, column=cc).fill = TROUGH_FILL
            pc = ws.cell(row=r, column=3)
            pc.value = f"◀ PEAK  ·  {e['party']}"
            pc.font = Font(bold=True, size=BASE_SIZE - 1, color=RED)
        r += 1

    _setup_print(ws, 7)
    return ws


def build_sheet_pos(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    unused: list, used: list, as_of: str, realm: str = "",
) -> None:
    """
    POs sheet — every PurchaseOrder for the project matched to its Bill(s).
    UNUSED (open / not yet billed) on TOP — that total is the BUFFER for the
    real-world gap where bills get entered after the report is pulled. USED
    (billed) below. PO info + matched Bill ref#/date/amount only — no account
    detail (the user 2026-06-09). Conditional color: amber for the buffer, red when
    a bill overran its PO.
    """
    ws = wb.create_sheet("POs")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    for col, w in (("A", 16), ("B", 13), ("C", 36), ("D", 16),
                   ("E", 16), ("F", 13), ("G", 16), ("H", 18)):
        ws.column_dimensions[col].width = w

    BUFFER_FILL = PatternFill("solid", fgColor="FFF2CC")  # amber = buffer

    def _po_order(recs):
        """Vendor A→Z, then PO date NEWEST→oldest (the user 2026-06-09)."""
        return sorted(recs, key=lambda x: (
            (x.get("vendor") or "").lower(),
            -((_parse_date(x.get("po_date")) or dt.date.min).toordinal())))

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    intro = ws.cell(row=r, column=1, value=(
        "Purchase Orders matched to their Bills. UNUSED (open) POs are a "
        "BUFFER — committed costs not yet billed, for bills entered after the "
        "report was pulled. USED POs are already billed."))
    intro.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    headers = ["PO #", "PO Date", "Vendor", "PO Amount",
               "Bill Ref #", "Bill Date", "Bill Amount", "Status"]

    def hdr_row():
        nonlocal r
        for ci, h in enumerate(headers, start=1):
            c = _write_cell(ws, r, ci, h)
            c.font = SUBHDR_FONT
            c.border = BOTTOM_BORDER
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        r += 1

    def date_cell(col, date_s, *, bold=False):
        dv = _parse_date(date_s)
        cell = ws.cell(row=r, column=col, value=dv or date_s)
        if dv:
            cell.number_format = "mm/dd/yyyy"
        cell.font = Font(bold=bold, size=BASE_SIZE - 1)
        cell.alignment = Alignment(horizontal="center")

    def po_row(rec, *, fill=None, status="", status_color="000000"):
        """One row per PO; if it has matched bills, one row per (PO, bill)."""
        nonlocal r
        bills = rec.get("bills") or [None]
        for i, b in enumerate(bills):
            # PO cols only on the first line of a multi-bill PO
            po_num = rec["po_num"] if i == 0 else ""
            c = _write_cell(ws, r, 1, po_num)
            c.font = Font(bold=True, size=BASE_SIZE - 1)
            _url = _qbo_txn_url("purchaseorder", rec.get("id", ""), realm) if i == 0 else None
            if _url:
                c.hyperlink = _url
                c.font = Font(bold=True, size=BASE_SIZE - 1, color=LINK, underline="single")
            if i == 0:
                date_cell(2, rec["po_date"])
                _write_cell(ws, r, 3, rec["vendor"]).font = Font(size=BASE_SIZE - 1)
                amt = _write_cell(ws, r, 4, rec["po_amt"]); amt.number_format = CURR_FMT
                amt.font = Font(bold=True, size=BASE_SIZE - 1)
            if b is not None:
                bref = _write_cell(ws, r, 5, b["ref"])
                _burl = _qbo_txn_url(b.get("tx_type", "Bill"), b.get("txn_id", ""), realm)
                if _burl:                            # Bill Ref # links to the QBO bill
                    bref.hyperlink = _burl
                    bref.font = Font(size=BASE_SIZE - 1, color=LINK, underline="single")
                else:
                    bref.font = Font(size=BASE_SIZE - 1)
                date_cell(6, b["date"])
                ba = _write_cell(ws, r, 7, b["amt"]); ba.number_format = CURR_FMT
                ba.font = Font(size=BASE_SIZE - 1)
                # red if this bill overran the PO amount
                over = rec["po_amt"] and b["amt"] > rec["po_amt"] + 0.005
                if over:
                    ba.font = Font(size=BASE_SIZE - 1, color=RED)
            st = _write_cell(ws, r, 8, status if i == 0 else "")
            st.font = Font(bold=True, size=BASE_SIZE - 1, color=status_color)
            if fill is not None:
                for ci in range(1, 9):
                    ws.cell(row=r, column=ci).fill = fill
            r += 1

    # ── UNUSED (buffer) on TOP ──
    ws.cell(row=r, column=1, value="UNUSED POs — BUFFER (open, not yet billed)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=BASE_SIZE + 1, color=NAVY)
    r += 1
    hdr_row()
    if unused:
        ustart = r
        for rec in _po_order(unused):
            po_row(rec, fill=BUFFER_FILL, status="OPEN — buffer",
                   status_color="BF8F00")
        uend = r - 1
        tc = _write_cell(ws, r, 3, "Total unused PO buffer")
        tc.font = Font(bold=True, size=BASE_SIZE)
        tot = ws.cell(row=r, column=4,
                      value=f"=SUM(D{ustart}:D{uend})")
        tot.number_format = CURR_FMT
        tot.font = Font(bold=True, size=BASE_SIZE, color="BF8F00")
        tot.fill = BUFFER_FILL
        tc.fill = BUFFER_FILL
        r += 1
    else:
        ws.cell(row=r, column=1, value="(none)").font = Font(
            italic=True, size=BASE_SIZE - 1, color="808080")
        r += 1
    r += 1

    # ── USED (billed) BELOW ──
    ws.cell(row=r, column=1, value="USED POs — billed")
    ws.cell(row=r, column=1).font = Font(bold=True, size=BASE_SIZE + 1, color=NAVY)
    r += 1
    hdr_row()
    if used:
        for rec in _po_order(used):
            po_row(rec, status="BILLED", status_color=GREEN)
    else:
        ws.cell(row=r, column=1, value="(none)").font = Font(
            italic=True, size=BASE_SIZE - 1, color="808080")
        r += 1
    _setup_print(ws, 8)


_FNAME_SAFE_RE = re.compile(r"[^A-Za-z0-9._ -]+")
_attachable_index_cache: Optional[Dict[Tuple[str, str], List[dict]]] = None
# The Attachable sweep walks EVERY scan AP ever uploaded (~10 min of paging),
# so the index persists to disk for a week (logs dir per the repo rule — never
# inside the repo). TempDownloadUri expires in minutes and is NOT cached; a
# download re-reads its attachable by id for a fresh link. Attachments
# uploaded since the cache was built are invisible until it expires — delete
# the file (or wait out the TTL) to pick them up sooner.
_ATT_CACHE_TTL_S = 7 * 24 * 3600


def _att_cache_file(company_id: str) -> Path:
    d = Path.home() / "Library/Logs/Proficient/project-pnl"
    d.mkdir(parents=True, exist_ok=True)
    return d / f"attachable_index_{company_id}.json"


def _load_attachable_index(access: str, company_id: str
                           ) -> Dict[Tuple[str, str], List[dict]]:
    """(entity type, txn id) → [attachable {Id, FileName}] — from the disk
    cache when fresh, else a full sweep (then cached)."""
    global _attachable_index_cache
    if _attachable_index_cache is not None:
        return _attachable_index_cache
    cache = _att_cache_file(company_id)
    try:
        if cache.exists() and time.time() - cache.stat().st_mtime < _ATT_CACHE_TTL_S:
            raw = json.loads(cache.read_text())
            by_key: Dict[Tuple[str, str], List[dict]] = {}
            for item in raw["items"]:
                for etype, evalue in item["refs"]:
                    by_key.setdefault((etype, evalue), []).append(
                        {"Id": item["id"], "FileName": item["file"]})
            _attachable_index_cache = by_key
            ui_event(f"attachment index: cached "
                     f"({dt.datetime.fromtimestamp(cache.stat().st_mtime):%m/%d %H:%M}"
                     f", {len(raw['items'])} files) — delete the cache file to re-sweep",
                     icon="·")
            return by_key
    except Exception:
        pass                                   # unreadable cache → re-sweep
    by_key = {}
    items = []
    for a in query_all(access, company_id, "Attachable"):
        if not a.get("FileName"):
            continue                           # a bare note, not a file
        refs = [((r.get("EntityRef") or {}).get("type"),
                 (r.get("EntityRef") or {}).get("value"))
                for r in a.get("AttachableRef") or []]
        refs = [x for x in refs if x[0] and x[1]]
        if not refs:
            continue
        items.append({"id": a["Id"], "file": a["FileName"], "refs": refs})
        for key in refs:
            by_key.setdefault(key, []).append(
                {"Id": a["Id"], "FileName": a["FileName"]})
    try:
        cache.write_text(json.dumps({"fetched": dt.datetime.now().isoformat(),
                                     "items": items}))
    except OSError:
        pass
    _attachable_index_cache = by_key
    return by_key


def fetch_txn_attachments(access: str, company_id: str,
                          txns: List[Tuple[str, str, str]],
                          dest: Path) -> Dict[str, str]:
    """Download each transaction's QBO ATTACHMENT (the uploaded bill scan) into
    `dest` and return {txn_id: relative link}. The user 2026-07-31: the ledger
    link should open the uploaded file itself, not the QBO bill page — but QBO
    only serves attachments through TempDownloadUri links that EXPIRE in
    minutes, so a workbook can't link them directly. We pull the file at
    export time instead; the cell links the local copy (works offline, no QBO
    login), and bills with no attachment keep the QBO bill link.

    txns: (txn_id, tx_type 'Bill'/'Expense', doc#). Idempotent — a file whose
    name is already in dest is not re-downloaded. The company-wide Attachable
    sweep (every scan AP ever uploaded) is paged and slow, so it runs ONCE per
    process and is reused across a batch run's projects."""
    try:
        by_key = _load_attachable_index(access, company_id)
    except Exception as e:
        ui_warn(f"attachment sweep failed ({e}) — ledger keeps QBO bill links")
        return {}
    out: Dict[str, dict] = {}
    n_dl = 0
    for txn_id, tx_type, doc in txns:
        etype = "Purchase" if tx_type == "Expense" else "Bill"
        attachables = by_key.get((etype, txn_id), [])
        multi = len(attachables) > 1
        prefix = _FNAME_SAFE_RE.sub("_", str(doc or txn_id)).strip()
        # A bill with SEVERAL scans gets its OWN subfolder (the user
        # 2026-07-31 — the folder link must show that bill's files, not the
        # whole attachments library); a single scan stays flat and the cell
        # opens it directly.
        tdir = dest / prefix if multi else dest
        got: List[str] = []
        for a in attachables:
            fname = _FNAME_SAFE_RE.sub("_", a["FileName"]).strip() or "attachment"
            name = (fname if multi
                    else (f"{prefix}_{fname}" if not fname.startswith(prefix)
                          else fname))
            path = tdir / name
            if not path.exists():
                legacy = dest / (f"{prefix}_{fname}"
                                 if not fname.startswith(prefix) else fname)
                if multi and legacy.exists():
                    # already downloaded flat by the earlier layout — move it
                    tdir.mkdir(parents=True, exist_ok=True)
                    legacy.rename(path)
                else:
                    # TempDownloadUri expires in minutes — always fetch a
                    # fresh one by re-reading the attachable (one GET each).
                    try:
                        fresh = _api_get(f"/v3/company/{company_id}/attachable/"
                                         f"{a['Id']}", access)
                        uri = (fresh.get("Attachable") or {}).get("TempDownloadUri")
                        if not uri:
                            continue
                        resp = requests.get(uri, timeout=60)
                        resp.raise_for_status()
                        tdir.mkdir(parents=True, exist_ok=True)
                        path.write_bytes(resp.content)
                        n_dl += 1
                    except Exception:
                        continue           # this one keeps its QBO bill link
            got.append(name)
        if got:
            out[txn_id] = {
                "link": (f"{dest.name}/{prefix}" if multi
                         else f"{dest.name}/{got[0]}"),
                "n": len(got)}
    if out:
        multi = sum(1 for v in out.values() if v["n"] > 1)
        ui_event(f"bill attachments: {len(out)} transaction(s) linked "
                 f"({n_dl} newly downloaded"
                 + (f", {multi} with multiple scans → folder link" if multi else "")
                 + f") → {dest.name}/")
    return out


def _qbo_txn_url(tx_type: str, txn_id: str, realm: str) -> Optional[str]:
    """Deep link to a QBO transaction so the user can click straight to the
    bill/expense to review it (the user 2026-06-19). Uses the login deep-link form
    QBO itself returns, so it works regardless of the user's session."""
    if not txn_id or not realm:
        return None
    from urllib.parse import quote
    t = (tx_type or "").lower()
    page = {"bill": "bill", "invoice": "invoice",
            "purchaseorder": "purchaseorder", "po": "purchaseorder"}.get(t, "expense")
    return (f"https://qbo.intuit.com/app/login?pagereq="
            f"{quote(f'{page}?txnId={txn_id}')}&deeplinkcompanyid={realm}")


def _qbo_customer_url(customer_id: str, realm: str) -> Optional[str]:
    """Deep link to the QBO customer page for the project — every invoice and
    payment on one screen (the user 2026-07-13: Billed totals must be
    click-to-verify). Same login deep-link form as _qbo_txn_url so it works
    regardless of the user's session."""
    if not customer_id or not realm:
        return None
    from urllib.parse import quote
    return (f"https://qbo.intuit.com/app/login?pagereq="
            f"{quote(f'customerdetail?nameId={customer_id}')}"
            f"&deeplinkcompanyid={realm}")


def _qbo_project_pl_url(customer_id: str, realm: str) -> Optional[str]:
    """Deep link to the QBO P&L report filtered to the project customer —
    billed AND costs in one drillable report (the user 2026-07-13: Costs
    totals must be click-to-verify). QBO web occasionally ignores the URL
    filter params and opens the report unfiltered — still lands on the P&L."""
    if not customer_id or not realm:
        return None
    from urllib.parse import quote
    return (f"https://qbo.intuit.com/app/login?pagereq="
            f"{quote(f'report?rptId=PANDL&customer={customer_id}&date_macro=alldates')}"
            f"&deeplinkcompanyid={realm}")


def _setup_print(ws, last_col: int, header_rows: int = 2) -> None:
    """Printer-friendly: landscape, scaled to ONE page wide (height flows to as
    many pages as needed), tight margins, print area = used cols, title rows
    repeat on every page (the user 2026-06-19)."""
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.2)
    ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"
    if header_rows:
        ws.print_title_rows = f"1:{header_rows}"


def build_sheet_job_rp(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    invoices: List[dict], job_groups: dict, job_total: float,
    billed_total: float, as_of: str, overhead_pct: float = 10.0,
    realm: str = "",
) -> None:
    """
    RESIDENTIAL (RP) — the main "Job P&L" sheet (the user 2026-06-19). Most important
    info first: JOB PROFIT box pinned at the TOP (Billed, Costs, Gross Profit,
    Margin %, Markup %, Overhead, TRUE NET PROFIT — hero). Then the INVOICE, then
    the JOB COSTS transaction detail — ALWAYS shown here, broken out JOB TYPE →
    COST NAME → VENDOR → TRANSACTION, collapsed to the job-type/cost totals.
    Wreck labor billed after the invoice is tagged (wreck) green; open POs dated
    before the invoice (bill pending) are tagged and highlighted YELLOW. Budget
    vs Actual lives on its OWN sheet, not here. Big (16pt) text, thick boxes.
    """
    SZ = BASE_SIZE - 1            # body text — smaller, compact (the user 2026-06-26)
    HSZ = BASE_SIZE + 1           # section headers / hero
    ws = wb.create_sheet("Job P&L")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.sheet_properties.outlinePr.summaryBelow = False
    for col, w in (("A", 40), ("B", 14), ("C", 44), ("D", 15)):
        ws.column_dimensions[col].width = w

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    # Header shortcut (the user 2026-08-06): "Open Project in QBO" → project HOME
    # page, in the free header cell to the right of the 4-col RP layout.
    _home_url = _qbo_customer_url(cust_info.get("id", ""), realm)
    if _home_url:
        # Column 9, NOT 5: the meta block merges A2:H2 (subtitle), and writing
        # into a merged cell raises 'MergedCell.value is read-only' — the CP/MFD
        # template already dodges this by using I2 (fixed 2026-08-06).
        _hc = ws.cell(row=2, column=9, value="Open Project in QBO  ↗")
        _hc.hyperlink = _home_url
        _hc.font = Font(size=BASE_SIZE, color="0563C1", underline="single")
    r += 1

    HERO = PatternFill("solid", fgColor="1F3A5F")
    SECT = PatternFill("solid", fgColor="DDEBF7")
    INV_FILL = PatternFill("solid", fgColor="C6E0B4")    # green — invoice header
    COST_FILL = PatternFill("solid", fgColor="EFF5FC")   # light cost-name band

    def _box(r0, r1, c0, c1, side):
        for gr in range(r0, r1 + 1):
            for cc in range(c0, c1 + 1):
                b = ws.cell(row=gr, column=cc).border
                ws.cell(row=gr, column=cc).border = Border(
                    left=side if cc == c0 else b.left,
                    right=side if cc == c1 else b.right,
                    top=side if gr == r0 else b.top,
                    bottom=side if gr == r1 else b.bottom)

    def thick_box(r0, r1, c0, c1):
        _box(r0, r1, c0, c1, _THICK)

    def thin_box(r0, r1, c0, c1):
        _box(r0, r1, c0, c1, _THINB)

    last_col = 4
    pcol = 2                       # JOB PROFIT values live in col B (the user 2026-06-19)

    # ════════════ JOB PROFIT — compact card pinned at the very top (A:B) ═══
    ph = ws.cell(row=r, column=1, value="JOB PROFIT")
    ph.font = Font(bold=True, size=HSZ, color="FFFFFF")
    for cc in range(1, pcol + 1):
        ws.cell(row=r, column=cc).fill = HERO
    ws.row_dimensions[r].height = 26
    prof_top = r
    r += 1

    def profit_line(label, *, hero=False, color="000000", fill=None):
        """Color-graded like the Draw P&L (the user 2026-06-26): green income, tan
        costs, yellow gross profit, gray overhead, navy hero for True Net Profit."""
        nonlocal r
        lc = _write_cell(ws, r, 1, label)
        lc.font = Font(bold=True, size=(HSZ if hero else SZ),
                       color=("FFFFFF" if hero else color))
        vc = ws.cell(row=r, column=pcol)       # value next to label in col B
        vc.font = Font(bold=True, size=(HSZ if hero else SZ),
                       color=("C6EFCE" if hero else color))
        use_fill = HERO if hero else fill
        if use_fill is not None:
            for cc in range(1, pcol + 1):
                ws.cell(row=r, column=cc).fill = use_fill
        if hero:
            ws.row_dimensions[r].height = 26
        used = r
        r += 1
        return vc, used

    billed_vc, billed_pr = profit_line("Billed", color="375623", fill=INCOME_FILL)
    costs_vc, costs_pr = profit_line("Job Costs", color="C55A11", fill=COGS_FILL)
    gp_vc, gp_pr = profit_line("Gross Profit", fill=GP_FILL)
    margin_vc, _ = profit_line("Margin % (of billed)", fill=GP_FILL)
    markup_vc, _ = profit_line("Markup % (of cost)", fill=GP_FILL)
    ohx_vc, ohx_pr = profit_line(f"less: Overhead ({overhead_pct:.1f}% of billed)",
                                 color="C0504D", fill=SECT_FILL)
    tnp_vc, tnp_pr = profit_line("TRUE NET PROFIT", hero=True)
    tnppct_vc, _ = profit_line("True Net Profit %", hero=True)
    # QBO deep links (the user 2026-07-13): Billed → the customer page (all
    # invoices on one screen); Job Costs → the project-filtered P&L report.
    for _vc, _u in ((billed_vc, _qbo_customer_url(cust_info.get("id", ""), realm)),
                    (costs_vc, _qbo_project_pl_url(cust_info.get("id", ""), realm))):
        if _u:
            _vc.hyperlink = _u
            _f = _vc.font
            _vc.font = Font(bold=_f.bold, size=_f.size, color=_f.color,
                            underline="single")
    thick_box(prof_top, r - 1, 1, pcol)
    r += 1

    # ════════════ WIP / PROJECTION (the user 2026-06-26) ════════════
    #  Bid Proposal (the residential "contract") + ETC as yellow inputs; the rest
    #  is computed. Same idea as the Draw P&L ① WIP block.
    YEL = PatternFill("solid", fgColor="FFE699")

    def _wnum(*keys):
        for k in keys:
            v = wip_info.get(k)
            if v not in (None, ""):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    pass
        return None

    wh = ws.cell(row=r, column=1, value="WIP / PROJECTION  (yellow = your input)")
    wh.font = Font(bold=True, size=HSZ, color="FFFFFF"); wh.fill = HERO
    for cc in range(2, pcol + 1):
        ws.cell(row=r, column=cc).fill = HERO
    wip_top = r
    r += 1
    bid_c = _write_cell(ws, r, 1, "Bid Proposal (contract)"); bid_c.font = Font(bold=True, size=SZ)
    bv = ws.cell(row=r, column=pcol,
                 value=_wnum("contract_saved", "revised_contract", "original_contract", "contract"))
    bv.number_format = CURR_FMT; bv.font = Font(bold=True, size=SZ); bv.fill = YEL
    bid_row = r; r += 1
    etc_c = _write_cell(ws, r, 1, "Estimated Total Cost (ETC)"); etc_c.font = Font(bold=True, size=SZ)
    ev = ws.cell(row=r, column=pcol,
                 value=_wnum("etc_saved", "revised_etc", "original_etc", "etc"))
    ev.number_format = CURR_FMT; ev.font = Font(bold=True, size=SZ); ev.fill = YEL
    etc_row = r; r += 1

    def _wrow(label, formula, fmt=CURR_FMT, color="000000"):
        nonlocal r
        _write_cell(ws, r, 1, label).font = Font(size=SZ, color="595959")
        c = ws.cell(row=r, column=pcol, value=formula)
        c.number_format = fmt; c.font = Font(bold=True, size=SZ, color=color)
        r += 1

    _wrow("% Complete (cost ÷ ETC)", f'=IF(B{etc_row}=0,"",{job_total}/B{etc_row})',
          fmt="0%", color=NAVY)
    pct_row = r - 1
    _wrow("Earned Revenue (bid × %)", f"=B{bid_row}*B{pct_row}")
    _wrow("Cost to date", f"={job_total}", color="C55A11")
    _wrow("Projected Profit (bid − ETC)", f"=B{bid_row}-B{etc_row}", color=GREEN)
    _wrow("Projected Margin %", f'=IF(B{bid_row}=0,"",(B{bid_row}-B{etc_row})/B{bid_row})',
          fmt="0%", color=GREEN)
    thick_box(wip_top, r - 1, 1, pcol)
    r += 1

    # ════════════ INVOICE ════════════
    ih = ws.cell(row=r, column=1, value="INVOICE")
    ih.font = Font(bold=True, size=HSZ, color="375623")
    ih.fill = INV_FILL
    for cc in range(2, last_col + 1):
        ws.cell(row=r, column=cc).fill = INV_FILL
    inv_top = r
    r += 1
    inv_rows = []
    for inv in invoices or []:
        lc = _write_cell(ws, r, 1, f"  Inv #{inv.get('doc_num', '')}")
        lc.font = Font(size=SZ)
        dv = _parse_date(inv.get("date", ""))
        dc = ws.cell(row=r, column=2, value=dv or inv.get("date", ""))
        if dv:
            dc.number_format = "mm/dd/yyyy"
        dc.font = Font(size=SZ)
        mc2 = _write_cell(ws, r, 3, inv.get("memo", "")); mc2.font = Font(size=SZ)
        vc = ws.cell(row=r, column=last_col, value=float(inv.get("amount", 0) or 0))
        vc.number_format = CURR_FMT; vc.font = Font(size=SZ)
        inv_rows.append(r)
        r += 1
    blc = _write_cell(ws, r, 1, "Total billed"); blc.font = Font(bold=True, size=SZ)
    if inv_rows:
        bcell = ws.cell(row=r, column=last_col,
                        value="=" + "+".join(f"{get_column_letter(last_col)}{ir}"
                                             for ir in inv_rows))
    else:
        bcell = ws.cell(row=r, column=last_col, value=0)
    bcell.number_format = CURR_FMT
    bcell.font = Font(bold=True, size=SZ, color="375623")
    blc.border = TOP_BORDER; bcell.border = TOP_BORDER  # close the total, no box
    billed_row = r
    r += 2

    # ════════════ JOB COSTS — BY ACCOUNT → VENDOR (the user 2026-06-26) ════════════
    #  Show the ACCOUNT (resolved from items, readable — not the cost code), then
    #  the VENDORS under it, then the bills (collapsed). A parent with ONE sub
    #  collapses to the sub name ("Subcontractor Expense: Labor" → "Labor"); a
    #  parent with MANY subs shows the parent total, then each sub.
    PEND_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber — bill pending
    dh = ws.cell(row=r, column=1, value="JOB COSTS  —  account → vendor")
    dh.font = Font(bold=True, size=HSZ, color=NAVY); dh.fill = SECT
    for cc in range(2, last_col + 1):
        ws.cell(row=r, column=cc).fill = SECT
    det_top = r
    r += 1
    acct_rows = []

    def _rp_vendors(vendors, vlevel, vindent):
        """vendor (total, visible) → bills (collapsed). Bills link to QBO; bill-
        pending POs shown amber, wreck labor green."""
        nonlocal r
        for vend in sorted(vendors, key=lambda v: -vendors[v]["total"]):
            vg = vendors[vend]
            vc = _write_cell(ws, r, 1, vindent + vend)
            vc.font = Font(bold=True, size=SZ)
            vt = ws.cell(row=r, column=last_col, value=round(vg["total"], 2))
            vt.number_format = CURR_FMT; vt.font = Font(bold=True, size=SZ)
            ws.row_dimensions[r].outline_level = vlevel
            ws.row_dimensions[r].collapsed = True
            r += 1
            for t in sorted(vg["txns"], key=lambda t: _parse_date(t.get("date", ""))
                            or dt.date.min):
                pend = t.get("po_pending")
                tag = (" (bill pending)" if pend else " (wreck)" if t.get("wreck") else "")
                clr = "BF8F00" if pend else GREEN if t.get("wreck") else "000000"
                dc2 = _write_cell(ws, r, 1, vindent + "    " + str(t["doc_num"]) + tag)
                dc2.font = Font(size=SZ, color=clr)
                u = _qbo_txn_url(t.get("tx_type", ""), t.get("txn_id", ""), realm)
                if u:
                    dc2.hyperlink = u
                    dc2.font = Font(size=SZ, color=LINK, underline="single")
                dv = _parse_date(t.get("date", ""))
                d2 = ws.cell(row=r, column=2, value=dv or t.get("date", ""))
                if dv:
                    d2.number_format = "mm/dd/yyyy"
                d2.font = Font(size=SZ)
                _write_cell(ws, r, 3, t.get("desc", "")).font = Font(size=SZ)
                a2 = ws.cell(row=r, column=last_col, value=round(t["amount"], 2))
                a2.number_format = CURR_FMT; a2.font = Font(size=SZ)
                if pend:
                    for cc in range(1, last_col + 1):
                        ws.cell(row=r, column=cc).fill = PEND_FILL
                ws.row_dimensions[r].outline_level = vlevel + 1
                ws.row_dimensions[r].hidden = True
                r += 1

    if not job_groups:
        nc = _write_cell(ws, r, 1, "  (no job costs found)")
        nc.font = Font(italic=True, size=SZ, color="808080")
        r += 1
    for parent in sorted(job_groups, key=lambda p: -job_groups[p]["total"]):
        pg = job_groups[parent]
        subs = pg.get("subs", {})
        if len(subs) <= 1:                       # collapse parent+single sub → sub name
            leaf = next(iter(subs), parent) or parent
            lc = _write_cell(ws, r, 1, leaf)
            lc.font = Font(bold=True, size=SZ, color=NAVY)
            tc = ws.cell(row=r, column=last_col, value=round(pg["total"], 2))
            tc.number_format = CURR_FMT; tc.font = Font(bold=True, size=SZ, color="C55A11")
            for cc in range(1, last_col + 1):
                ws.cell(row=r, column=cc).fill = COST_FILL
            acct_rows.append(r)
            r += 1
            if subs:
                _rp_vendors(subs[leaf]["vendors"], 1, "    ")
        else:                                    # parent total, then its sub-accounts
            pc = _write_cell(ws, r, 1, parent)
            pc.font = Font(bold=True, size=SZ, color=NAVY)
            pt = ws.cell(row=r, column=last_col, value=round(pg["total"], 2))
            pt.number_format = CURR_FMT; pt.font = Font(bold=True, size=SZ, color=NAVY)
            for cc in range(1, last_col + 1):
                ws.cell(row=r, column=cc).fill = ACCENT_FILL
            acct_rows.append(r)
            r += 1
            for leaf in sorted(subs, key=lambda l: -subs[l]["total"]):
                sc = _write_cell(ws, r, 1, "    " + leaf); sc.font = Font(bold=True, size=SZ)
                st = ws.cell(row=r, column=last_col, value=round(subs[leaf]["total"], 2))
                st.number_format = CURR_FMT; st.font = Font(size=SZ, color="C55A11")
                for cc in range(1, last_col + 1):
                    ws.cell(row=r, column=cc).fill = COST_FILL
                ws.row_dimensions[r].outline_level = 1
                r += 1
                _rp_vendors(subs[leaf]["vendors"], 2, "        ")
    tlc = _write_cell(ws, r, 1, "Total job costs"); tlc.font = Font(bold=True, size=SZ)
    tlc.border = TOP_BORDER
    cf = ("=" + "+".join(f"D{jr}" for jr in acct_rows)) if acct_rows else "=0"
    cc = ws.cell(row=r, column=4, value=cf)
    cc.number_format = CURR_FMT
    cc.font = Font(bold=True, size=SZ, color="C55A11")
    cc.border = TOP_BORDER
    costs_row = r
    costs_col = "D"
    r += 1

    # ── fill the JOB PROFIT card (col B) now that the totals exist ──
    #    Billed/Costs mirror the section totals (col D); the rest is computed
    #    within the card so the box is self-contained.
    oh = overhead_pct / 100.0
    underwater = (billed_total - job_total) < 0
    billed_vc.value = f"=D{billed_row}"
    billed_vc.number_format = CURR_FMT
    costs_vc.value = f"={costs_col}{costs_row}"
    costs_vc.number_format = CURR_FMT
    gp_vc.value = f"=B{billed_pr}-B{costs_pr}"
    gp_vc.number_format = CURR_FMT
    if underwater:
        gp_vc.font = Font(bold=True, size=SZ, color=RED)
    margin_vc.value = f'=IF(B{billed_pr}=0,"",B{gp_pr}/B{billed_pr})'
    margin_vc.number_format = "0.0%"
    markup_vc.value = f'=IF(B{costs_pr}=0,"",B{gp_pr}/B{costs_pr})'
    markup_vc.number_format = "0.0%"
    ohx_vc.value = f"=-B{billed_pr}*{oh}"
    ohx_vc.number_format = CURR_FMT
    tnp_vc.value = f"=B{gp_pr}+B{ohx_pr}"
    tnp_vc.number_format = CURR_FMT
    tnppct_vc.value = f'=IF(B{billed_pr}=0,"",B{tnp_pr}/B{billed_pr})'
    tnppct_vc.number_format = "0.0%"

    _setup_print(ws, last_col)   # printer-friendly (landscape, fit 1 page wide)


def build_sheet_pending_rp(
    wb: Workbook, proj: str, cust_info: dict, wip_info: dict,
    pending: List[dict], dup_flags: List[dict], as_of: str,
    company_id: Optional[str] = None,
) -> None:
    """
    RESIDENTIAL (RP) Pending Review — the data-quality net (the user 2026-06-09).
    EXCLUDED from job costs: bills after the invoice that aren't wreck labor,
    zero/negative lines, and uncategorized lines. Plus an informational
    DUPLICATE-REF list (those ARE still counted in costs — verify them).
    Each row shows the line DESCRIPTION and an "Open in QBO" link straight to
    the bill (the user 2026-06-19).
    """
    SZ = BASE_SIZE - 1
    ws = wb.create_sheet("Pending Review")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    for col, w in (("A", 18), ("B", 13), ("C", 30), ("D", 42),
                   ("E", 14), ("F", 34), ("G", 16)):
        ws.column_dimensions[col].width = w

    r = _write_meta_block(ws, proj, cust_info, wip_info, as_of)
    leg = ws.cell(row=r, column=1, value=(
        "Items NOT counted in job costs — click 'Open in QBO' to review the bill, "
        "fix it, then re-run. Duplicate refs (bottom) ARE counted; verify they "
        "aren't double-entered."))
    leg.font = Font(italic=True, size=BASE_SIZE - 2, color="595959")
    r += 2

    def hdr():
        nonlocal r
        for ci, h in enumerate(["Ref #", "Date", "Vendor", "Description",
                                "Amount", "Reason", "Link"], start=1):
            c = _write_cell(ws, r, ci, h)
            c.font = SUBHDR_FONT
            c.border = BOTTOM_BORDER
        r += 1

    def row_for(rec, *, amt_color="000000"):
        nonlocal r
        url = _qbo_txn_url(rec.get("tx_type"), rec.get("txn_id"), company_id)
        ref = _write_cell(ws, r, 1, str(rec.get("doc_num", "")))
        if url:                                  # ref # itself is clickable
            ref.hyperlink = url
            ref.font = Font(size=SZ, color=LINK, underline="single")
        else:
            ref.font = Font(size=SZ)
        dv = _parse_date(rec.get("date", ""))
        dc = ws.cell(row=r, column=2, value=dv or rec.get("date", ""))
        if dv:
            dc.number_format = "mm/dd/yyyy"
        dc.font = Font(size=SZ)
        _write_cell(ws, r, 3, rec.get("vendor", "")).font = Font(size=SZ)
        _write_cell(ws, r, 4, rec.get("desc", "")).font = Font(size=SZ)
        a = _write_cell(ws, r, 5, float(rec.get("amount", 0) or 0))
        a.number_format = CURR_FMT; a.font = Font(size=SZ, color=amt_color)
        _write_cell(ws, r, 6, rec.get("reason", "")).font = Font(size=SZ, color=RED)
        lk = _write_cell(ws, r, 7, "Open in QBO ▸")
        if url:
            lk.hyperlink = url
            lk.font = Font(size=SZ, color=LINK, underline="single")
        else:
            lk.value = ""
        r += 1

    def grouped(items, amt_color):
        """Vendor header (count) → its rows; group headers leave col E blank so a
        SUM over the range totals only bill rows (the user 2026-06-26 — group by vendor)."""
        nonlocal r
        byv = {}
        for rec in items:
            byv.setdefault(rec.get("vendor") or "(no vendor)", []).append(rec)
        for vend in sorted(byv, key=lambda v: -sum(float(x.get("amount", 0) or 0)
                                                   for x in byv[v])):
            vh = _write_cell(ws, r, 1, f"{vend}  ({len(byv[vend])})")
            vh.font = Font(bold=True, size=SZ, color=NAVY)
            r += 1
            for rec in sorted(byv[vend], key=lambda x: -float(x.get("amount", 0) or 0)):
                row_for(rec, amt_color=amt_color)

    h = ws.cell(row=r, column=1, value="PENDING REVIEW — excluded from job costs")
    h.font = Font(bold=True, size=BASE_SIZE + 1, color=RED)
    r += 1
    if pending:
        hdr()
        start = r
        grouped(pending, RED)
        _write_cell(ws, r, 4, "Total pending").font = Font(bold=True, size=SZ)
        tc = ws.cell(row=r, column=5, value=f"=SUM(E{start}:E{r-1})")
        tc.number_format = CURR_FMT; tc.font = Font(bold=True, size=SZ, color=RED)
        tc.border = TOP_BORDER
        r += 2
    else:
        ws.cell(row=r, column=1, value="(none — clean)").font = Font(
            italic=True, size=SZ, color=GREEN)
        r += 2

    if dup_flags:
        h2 = ws.cell(row=r, column=1, value="POSSIBLE DUPLICATES — still in costs, verify")
        h2.font = Font(bold=True, size=BASE_SIZE + 1, color="BF8F00")
        r += 1
        hdr()
        grouped(dup_flags, "BF8F00")
    _setup_print(ws, 7)


# ────────────────────────── orchestration ──────────────────────────

def generate_project_pnl(
    access: str,
    company_id: str,
    proj: str,
    cust_info: dict,
    wip_info: dict,
    start_date: str,
    end_date: str,
    out_dir: Path,
    as_of: str,
    dry_run: bool = False,
    overhead_pct: float = 10.0,
    interactive: bool = False,
    infer_periods: bool = False,
    simple: bool = False,
) -> Optional[Path]:
    ui_proj(proj, f"{cust_info['name']}  ·  id {cust_info['id']}")

    # Residential uses a different template: no draws, expenses → invoice →
    # profit, wreck-labor rule + Pending Review (the user 2026-06-09).
    if _project_division(proj).startswith("Residential"):
        return generate_project_pnl_rp(
            access, company_id, proj, cust_info, wip_info,
            start_date, end_date, out_dir, as_of,
            dry_run=dry_run, overhead_pct=overhead_pct,
        )

    # Each project gets ONE home folder named by its number (the user 2026-06-25):
    #   <out>/MFD192/                    ← the P&L workbook lands here
    #   <out>/MFD192/rd-reports/         ← PM drops draw-cost reports here
    # Created on the first P&L run, skipped if it already exists.
    # rd-reports is the MFD draw-cost-report cross-check workflow ONLY — Commercial
    # doesn't use it, so don't litter CP folders with it (the user 2026-07-02). rd_dir
    # stays defined so index_pm_reports() just finds nothing for non-MFD.
    is_mfd = proj.upper().startswith("MFD")
    # Dual overhead view (MFD 9% on costs vs Company 10% on revenue) is MFD-only —
    # MFD is a different player. CP (and any non-MFD draw job) shows ONLY the company
    # overhead; keep MFD out of it (the user 2026-07-02). None => single company view.
    _alt_oh = 9.0 if is_mfd else None
    # CP drops into the awarded-project folder on the Common drive; MFD stays in the
    # OneDrive tree (the user 2026-07-02).
    proj_dir, _cp_note = _resolve_project_out_dir(proj, out_dir)
    if _cp_note:
        ui_event(_cp_note, icon="⚑", color=_YEL)
    elif proj.upper().startswith("CP"):
        ui_event(f"CP output → {proj_dir}", icon="→", color=_CYAN)
    rd_dir = proj_dir / DRAW_REPORTS_SUBDIR
    if not dry_run:
        proj_dir.mkdir(parents=True, exist_ok=True)
        if is_mfd:
            rd_dir.mkdir(parents=True, exist_ok=True)

    invoices = fetch_customer_invoices(access, company_id, cust_info["id"])
    # An older job invoices on the PARENT customer, not the project (the user
    # 2026-08-24) — without this its billed-to-date reads as zero. Pull the
    # parent's too and let the memo decide which ones are this job's.
    if _LEGACY_MATCH is not None and cust_info.get("parent_id"):
        _seen = {i["Id"] for i in invoices}
        _extra = [i for i in fetch_customer_invoices(
                      access, company_id, cust_info["parent_id"])
                  if i["Id"] not in _seen and _LEGACY_MATCH.invoice_belongs(i)]
        if _extra:
            invoices += _extra
            ui_event(f"legacy: +{len(_extra)} invoice(s) billed on the parent "
                     f"customer", icon="⚑", color=_YEL)
    if infer_periods:
        _shape = learn_period_shape([(i.get("PrivateNote") or "")
                                     for i in invoices])
        if not _shape:
            ui_warn("--infer-periods: no invoice carries a (Period:…) tag, so "
                    "there is no shape to learn from — left as-is")
        else:
            _n = 0
            for _inv in invoices:
                _tag = infer_period_tag(_inv.get("PrivateNote") or "",
                                        _parse_date(_inv.get("TxnDate", "")),
                                        _shape)
                if _tag:
                    _inv["PrivateNote"] = ((_inv.get("PrivateNote") or "").rstrip()
                                           + "\n" + _tag)
                    _n += 1
            ui_event(f"periods inferred for {_n} invoice(s) — window learned "
                     f"from {_shape['n']} tagged: day {_shape['start_day']} of "
                     f"the prior month → day {_shape['end_day']}",
                     icon="⚑", color=_YEL)
    income_groups = group_invoices_by_draw(invoices, interactive=interactive)
    # AR/AP payment state (the user 2026-08-05): invoice/bill Balance == 0 is
    # PAID. Purchases (checks/CC) are paid by nature.
    # Exclude ALL special buckets (__untagged, __retainage, …) — only real
    # draw groups have a "period".
    tagged = [l for l in income_groups if not l.startswith("__")]
    tagged.sort(key=lambda l: income_groups[l]["period"][0])
    ui_event(f"{len(invoices)} invoices · {len(tagged)} draw periods")

    draw_periods = [
        (l, income_groups[l]["period"][0], income_groups[l]["period"][1])
        for l in tagged
    ]

    # P&L cuts off at the last draw period end; costs after that date are
    # the ACCUMULATING COSTS block (what's stacking toward the next draw).
    last_end = draw_periods[-1][2] if draw_periods else None
    pl_end = end_date
    pl_cutoff = None
    if last_end and last_end.isoformat() < end_date:
        pl_end = last_end.isoformat()
        pl_cutoff = last_end.strftime("%m/%d/%y")

    # QBO's own project P&L report is keyed to the project customer, so on a
    # legacy job it cannot see the line-text / bill-memo costs OR the invoices
    # billed on the parent — it would report a job millions short. In legacy
    # mode the totals are synthesized from the SAME attributed lines the rest
    # of the workbook is built from (below, once the bills are pulled).
    if _LEGACY_MATCH is not None:
        pl_data, pl_totals = {}, {}
    else:
        pl_data = fetch_project_pl(access, company_id, cust_info["id"],
                                   start_date, pl_end)
        pl_totals = extract_pl_totals(pl_data)
        ui_event(f"P&L through {pl_end}  ·  income ${pl_totals['income']:,.0f} · "
                 f"COGS ${pl_totals['cogs']:,.0f} · "
                 f"GP ${pl_totals['gross_profit']:,.0f}")

    accounts = query_all(access, company_id, "Account")
    parent_map = build_account_parent_map(accounts)
    account_names = {a.get("Id"): a.get("Name")
                     for a in accounts if a.get("Id")}
    # Fully-qualified names for the Transactions/P&L account labels (the user
    # 2026-06-22: "Job Materials: Rebar", not just "Rebar").
    account_fqn = {a.get("Id"): ((a.get("FullyQualifiedName") or a.get("Name") or "")
                                 .replace(":", ": "))
                   for a in accounts if a.get("Id")}
    acct_type = build_account_type_map(accounts)

    items = query_all(access, company_id, "Item")
    item_account = {it.get("Id"): (it.get("ExpenseAccountRef") or {}).get("value")
                    for it in items if it.get("Id")}
    ui_event(f"{len(accounts)} accounts · {len(items)} items")

    retainage_bs = fetch_retainage_held(
        access, company_id, cust_info["id"], end_date, accounts=accounts,
    )

    # Narrow the company-wide Bill/Purchase pull to this project's actual
    # life — earliest draw start or invoice date, minus a 180-day buffer for
    # pre-mobilization costs — instead of scanning from start_date (default
    # 2020). QBO can't filter these by line-level customer, so a tighter
    # TxnDate window is the cheapest way to cut the rows we download+filter.
    activity = [income_groups[l]["period"][0] for l in tagged]
    for inv in invoices:
        d = _parse_date(inv.get("TxnDate", ""))
        if d:
            activity.append(d)
    bill_start = start_date
    if activity:
        cand = (min(activity) - dt.timedelta(days=_BILL_LOOKBACK_DAYS)).isoformat()
        if cand > start_date:
            bill_start = cand
    bills, purchases = fetch_customer_bills_and_purchases(
        access, company_id, cust_info["id"], bill_start, end_date,
    )
    # {txn id: (balance, total)} — the balance is what makes PARTIAL possible.
    paid_map = {b.get("Id"): (float(b.get("Balance", 0) or 0),
                              float(b.get("TotalAmt", 0) or 0)) for b in bills}
    paid_map.update({pch.get("Id"): (0.0, 0.0) for pch in purchases})
    if _LEGACY_MATCH is not None:
        pl_totals = _synth_pl_totals(bills, purchases, income_groups,
                                     cust_info["id"], acct_type, item_account,
                                     pl_end)
        ui_event(f"P&L through {pl_end}  ·  income ${pl_totals['income']:,.0f} · "
                 f"COGS ${pl_totals['cogs']:,.0f} · "
                 f"GP ${pl_totals['gross_profit']:,.0f}  "
                 f"{_DIM}(synthesized — legacy attribution){_RESET}")
    ui_event(f"{len(bills)} bills · {len(purchases)} purchases  "
             f"{_DIM}(from {bill_start}){_RESET}")
    # If the OLDEST attributed cost sits within a month of the window start,
    # the window may still be clipping older cost — say so rather than quietly
    # under-report the job.
    if bill_start > start_date:
        _d = [t.get("TxnDate") or "" for t in list(bills) + list(purchases)
              if _txn_touches_job(t, cust_info["id"])]
        _d = [x for x in _d if x]
        if _d:
            _edge = (dt.date.fromisoformat(bill_start)
                     + dt.timedelta(days=30)).isoformat()
            if min(_d) <= _edge:
                ui_event(f"oldest attributed cost {min(_d)} sits at the edge of "
                         f"the pull window ({bill_start}) — re-run with "
                         f"--start-date to confirm nothing older was clipped",
                         icon="⚑", color=_YEL)

    pos = fetch_customer_purchase_orders(
        access, company_id, cust_info["id"], bill_start, end_date,
    )
    po_unused, po_used = match_pos_to_bills(pos, bills, cust_info["id"])
    buf = sum(p["po_amt"] for p in po_unused)
    ui_event(f"{len(pos)} POs  ·  {len(po_unused)} open ${buf:,.0f} · "
             f"{len(po_used)} billed")

    # Cash flow (the user 2026-06-26): ACTUAL payment dates — AP out + AR in.
    bill_pmts = fetch_bill_payments(access, company_id, bill_start, end_date)
    cust_pmts = fetch_customer_payments(access, company_id, bill_start, end_date)
    cash_events = build_cashflow_events(bills, invoices, bill_pmts, cust_pmts,
                                        cust_info["id"])
    if cash_events:
        _trough = min(e["running"] for e in cash_events)
        ui_event(f"cash flow: {len(cash_events)} payments · "
                 f"peak ${_trough:,.0f} · now ${cash_events[-1]['running']:,.0f}",
                 icon="$", color=_CYAN)

    draw_costs = bucket_costs_by_draw_window(
        bills, purchases, cust_info["id"], draw_periods, parent_map,
        account_names=account_names,
    )

    # Accumulating costs for the next draw = EXACTLY the costs that fell
    # outside every draw window (the Draws "Costs Outside Draw Windows"
    # block). Using the same set makes the P&L total tie to the Draws total
    # and the ➜ link land on the matching detail (the user 2026-06-09).
    accum = None
    outside_acc = draw_costs.get("__outside")
    if outside_acc and (outside_acc.get("total") or outside_acc.get("groups")):
        end_d = _parse_date(end_date) or dt.date.today()
        acc_groups: Dict[str, float] = {}
        for pg in (outside_acc.get("groups") or {}).values():
            for leaf, lg in pg.get("subs", {}).items():
                acc_groups[leaf] = acc_groups.get(leaf, 0.0) + lg["total"]
        accum = {
            "through": end_d.strftime("%m/%d/%y"),
            "groups": acc_groups,
            "total": outside_acc.get("total", 0.0),
        }
        print(f"      accumulating (outside draw windows) = {accum['total']:,.2f}")
    _dis = draw_costs.get("__disregarded")
    if _dis:
        _anch = _dis.get("anchor") or ""
        ui_event(f"pre-period history disregarded from draw views: "
                 f"{_dis['count']} cost line(s) ${_dis['total']:,.0f} before {_anch} "
                 f"(still in P&L totals/Transactions)", icon="⚑", color=_YEL)

    # .get — special buckets like __retainage have no "net_billed"
    net_billed = sum(g.get("net_billed", 0.0) for g in income_groups.values())

    # Draw label comes from the invoice MEMO number first (the user 2026-06-09).
    # If a memo has NO 'Draw N', DON'T fabricate a number — label that draw
    # by its month/year ("Draw – April 2026") and sort by date (tagged is
    # already chronological). Tell the operator in the terminal.
    # Retainage-not-billed total (excluded from income per the user 2026-06-09)
    retainage_nb = (income_groups.get("__retainage") or {}).get("total", 0.0)
    # Total retainage CALCULATED from the invoice data (the user 2026-06-09 — the
    # left-side retainage must sum the draws' retainage + the other retainage
    # invoices, not just the not-billed piece the Balance Sheet showed):
    #   withheld on draws (held) + retainage billed back by GC + not yet billed.
    ret_withheld_total = sum(g.get("retainage_held", 0.0)
                             for g in income_groups.values())
    ret_billed_total = sum(g.get("retainage_billed", 0.0)
                           for g in income_groups.values())
    total_retainage = ret_withheld_total + ret_billed_total + retainage_nb
    ui_event(f"retainage ${total_retainage:,.0f}  "
             f"{_DIM}(withheld ${ret_withheld_total:,.0f} · billed ${ret_billed_total:,.0f} "
             f"· not-billed ${retainage_nb:,.0f} · BS ${retainage_bs:,.0f}){_RESET}")

    # draw_rows = (name, period, net_billed, costs, retainage_held, retainage_billed)
    # retainage_held  = GC holding back (positive)        — Retained col, black
    # retainage_billed = GC paying retainage back (positive) — Retained col, green
    # A draw that is ONLY retainage (no work) is NAMED "Retainage" (the user 2026-06-09).
    draw_rows: List[Tuple[str, str, float, float, float, float]] = []
    for lbl in tagged:
        net = income_groups[lbl]["net_billed"]
        held = income_groups[lbl].get("retainage_held", 0.0)
        billed = income_groups[lbl].get("retainage_billed", 0.0)
        work = income_groups[lbl].get("gross_income", 0.0)
        costs = (draw_costs.get(lbl) or {}).get("total", 0.0)
        draw_no = None
        for inv in income_groups[lbl].get("invoices") or []:
            draw_no = extract_draw_number(inv.get("memo", ""))
            if draw_no is not None:
                break
        if work <= 0.005 and billed > 0.005:
            name = "Retainage"            # pure retainage billing, no new work
        elif draw_no is not None:
            name = f"Draw {draw_no}"
        else:
            month = income_groups[lbl]["period"][1].strftime("%B %Y")  # period END
            name = f"Draw – {month}"
            ui_warn(f"no draw # in memo for {lbl} — labeled '{name}' by date")
        draw_rows.append((name, lbl, net, costs, held, billed))
        cov_s = f"{net / costs * 100:.0f}%" if costs else "—"
        ui_event(f"{name}  {_DIM}{lbl}{_RESET}  billed ${net:,.0f} · "
                 f"costs ${costs:,.0f} · coverage {cov_s}", icon="•", color=_CYAN)

    # PM draw-report CAPTURE INDEX (the user 2026-06-26): pool every report for this
    # project; a bill is 'captured' if it's on ANY report (the draw period from QBO
    # is authoritative — report periods get readjusted). Per-draw underbilling and
    # the orphan list are computed from this index, not from report periods.
    report_index, parsed_reports = index_pm_reports(rd_dir, proj)

    # ORPHANS (the user 2026-06-26 — reconcile = final catcher): report lines whose
    # bill#+amount match NO QBO bill anywhere on the project (PM typo / wrong amount
    # / cost not in QBO). Prior-period lines excluded (informational).
    _qbo_keys = set()
    _qbo_by_num: Dict[str, list] = {}     # bill # → QBO lines with that # (any amount)
    _qbo_lines: list = []                 # candidate pool for the difference-finder
    for _dc in draw_costs.values():
        for _b in _draw_flat_bills(_dc):
            num = str(_b["num"]).strip()
            amt = round(float(_b["amount"]), 2)
            _qbo_keys.add((num, amt))
            _qbo_by_num.setdefault(num, []).append(_b)
            _qbo_lines.append(_b)
    # A report line that doesn't EXACT-match QBO is either an AMOUNT MISMATCH (its
    # bill # IS in QBO, different amount — the user 2026-06-26) or a true ORPHAN (bill #
    # not in QBO at all). Splitting them stops "in QBO, wrong amount" from being
    # mislabeled "not in QuickBooks".
    orphans, mismatches, _seen = [], [], set()
    for _fname, _rep in parsed_reports:
        for _ln in _rep["lines"]:
            if _ln.get("prior"):
                continue
            num = str(_ln["num"]).strip()
            amt = round(float(_ln["amount"]), 2)
            _k = (num, amt)
            if _k in _qbo_keys or (_fname, _k) in _seen:
                continue
            _seen.add((_fname, _k))
            base = {"num": _ln["num"], "date": _ln["date"], "vendor": _ln["vendor"],
                    "desc": _ln["desc"], "amount": _ln["amount"], "source": _fname}
            if num in _qbo_by_num:                 # bill # exists in QBO → $ mismatch
                cand = min(_qbo_by_num[num],
                           key=lambda c: abs(float(c["amount"]) - amt))
                base["qbo_amount"] = round(float(cand["amount"]), 2)
                base["qbo_txn_id"] = cand.get("txn_id", "")
                base["qbo_tx_type"] = cand.get("tx_type", "")
                base["diff"] = round(amt - base["qbo_amount"], 2)
                mismatches.append(base)
            else:
                orphans.append(base)

    if dry_run:
        ui_event(f"[dry run] would write {proj_dir}/Project_PnL_{proj}.xlsx", color=_YEL)
        return None

    wb = Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    # Transactions sheet FIRST — its subtotal cells are the source the P&L's
    # Income/Retainage/COGS SUM-link to (the user 2026-06-22). Build it before the
    # P&L so the cell refs exist.
    tx = gather_transactions(income_groups, bills, purchases, cust_info["id"],
                             parent_map, account_names=account_fqn,
                             acct_type=acct_type, item_account=item_account)
    tx_refs = build_sheet_transactions(wb, proj, cust_info, wip_info, tx, as_of,
                                       paid_map=paid_map,
                                       realm=company_id)
    # Second view of the same lines, pivoted account → vendor, so a P&L figure
    # can be clicked straight through to what makes it up (the user 2026-08-27).
    acct_anchors = build_sheet_by_account(wb, proj, cust_info, wip_info, tx,
                                          as_of, realm=company_id,
                                          paid_map=paid_map)
    qbo_exp = pl_totals.get("gross_profit", 0.0) - pl_totals.get("net_ordinary_income", 0.0)
    # Candidate pool for the difference-finder: every project cost line + invoice
    # facet, so a gap of $X can be traced to a single transaction equaling it.
    diff_candidates = [{"ref": str(b.get("num", "")), "amount": round(float(b["amount"]), 2),
                        "party": b.get("vendor", ""), "txn_id": b.get("txn_id", ""),
                        "tx_type": b.get("tx_type", "Bill")} for b in _qbo_lines]
    for _g in income_groups.values():
        for _inv in _g.get("invoices") or []:
            for _f in ("amount", "gross", "retainage", "retainage_billed"):
                _v = _inv.get(_f)
                if _v:
                    diff_candidates.append({"ref": str(_inv.get("doc_num", "")),
                                            "amount": round(float(_v), 2),
                                            "party": cust_info["name"],
                                            "txn_id": _inv.get("id", ""), "tx_type": "invoice"})
    build_sheet_reconciliations(
        wb, proj, cust_info, wip_info, pl_totals.get("income", 0.0),
        pl_totals.get("cogs", 0.0), qbo_exp, tx_refs, as_of, has_retainage=True,
        orphans=orphans, reports_relpath=DRAW_REPORTS_SUBDIR,
        mismatches=mismatches, candidates=diff_candidates, tx_totals=tx["tot"],
        realm=company_id)

    # ONE SHEET PER DRAW (the user 2026-06-26). Newest draw first. Each draw's name
    # becomes its own sheet; draw_anchors maps name → sheet so the P&L coverage
    # table links to it. Match each draw to the PM-report cross-check whose period
    # best overlaps the draw window.
    used_names: set = set()

    def _sheet_name(base: str) -> str:
        s = re.sub(r"[:\\/?*\[\]]", "-", base).strip()[:31] or "Draw"
        cand, i = s, 2
        while cand in used_names or cand in wb.sheetnames:
            suf = f" ({i})"
            cand = s[:31 - len(suf)] + suf
            i += 1
        used_names.add(cand)
        return cand

    # Where each QBO bill actually lives (for the PM-only "→ in QBO <draw>" note).
    _lbl_to_name = {lbl: nm for nm, lbl, *_ in draw_rows}
    qbo_loc: Dict[tuple, str] = {}
    for _lbl, _dc in draw_costs.items():
        _loc = ("Next Draw" if str(_lbl).startswith("__")
                else _lbl_to_name.get(_lbl, _lbl))
        for _b in _draw_flat_bills(_dc):
            qbo_loc.setdefault((str(_b["num"]).strip(), round(float(_b["amount"]), 2)), _loc)

    def _match_report(prd):
        """The PM report whose period best overlaps this draw window (its 'version'
        of the draw). Period overlap, not filename — robust to readjusted periods."""
        if not prd:
            return None
        ds0, de0 = prd
        best, best_ov = None, 0
        for fname, rep in parsed_reports:
            rp = rep.get("period")
            if not rp:
                continue
            ov = (min(rp[1], de0) - max(rp[0], ds0)).days
            if ov > best_ov:
                best, best_ov = (fname, rep), ov
        return best

    draw_anchors: Dict[str, str] = {}
    draw_sheet_order: List[str] = []
    underbill_total, underbill_count = 0.0, 0
    for nm, lbl, net, costs, held, billed in (
            [] if simple else reversed(draw_rows)):
        sn = _sheet_name(nm)
        _bottom, m_tot, m_cnt = build_sheet_one_draw(
            wb, sn, proj, cust_info, wip_info, nm, lbl, net, costs, held, billed,
            (income_groups.get(lbl) or {}).get("invoices") or [],
            draw_costs.get(lbl) or {},
            _match_report(income_groups[lbl]["period"]), report_index, qbo_loc,
            income_groups[lbl]["period"], as_of, overhead_pct=overhead_pct,
            realm=company_id, alt_overhead_pct=_alt_oh,
            reports_relpath=DRAW_REPORTS_SUBDIR, paid_map=paid_map)
        underbill_total += m_tot
        underbill_count += m_cnt
        draw_anchors[nm] = sn
        draw_sheet_order.append(sn)
    underbill_total = round(underbill_total, 2)
    if parsed_reports:
        msg = (f"PM cross-check: {underbill_count} bill(s) on no report "
               f"(${underbill_total:,.0f} underbilled) · {len(orphans)} orphan line(s)")
        if underbill_count or orphans:
            ui_event(msg, icon="⚑", color=_YEL)
        else:
            ui_done(msg)

    leftover = None if simple else build_sheet_next_draw_retainage(
        wb, proj, cust_info, wip_info, income_groups, draw_costs, as_of,
        realm=company_id)
    if leftover is not None:
        draw_anchors["__outside"] = leftover.title

    build_sheet_pos(wb, proj, cust_info, wip_info, po_unused, po_used, as_of, realm=company_id)
    build_sheet_cashflow(wb, proj, cust_info, wip_info, cash_events, as_of, realm=company_id)
    # Budget vs Actual (the user 2026-07-16): CP budget = the takeoff's Cost
    # Code sheet; actuals = QBO cost-code totals. MFD has no takeoff → skipped.
    if not is_mfd:
        _bud, _bud_src = load_cp_budget(proj)
        if _bud:
            _cof = False
            try:
                _cof = bool(float(wip_info.get("change_orders") or 0))
            except (TypeError, ValueError):
                pass
            _acts = costs_by_code(bills, purchases, cust_info["id"], parent_map,
                                  account_names=account_names)
            build_sheet_budget_vs_actual(
                wb, proj, cust_info, wip_info, _bud, _acts, as_of,
                co_flag=_cof, budget_source=_bud_src, realm=company_id)
            _ncf = sum(1 for g in _acts.values() for t in g.get("txns", [])
                       if not _class_ok(t.get("class"), _expected_class(proj)))
            ui_event(f"Budget vs Actual: {len(_bud)} budgeted codes "
                     f"(${sum(_bud.values()):,.0f}) from {_bud_src}"
                     + ("  ⚑ CO flag" if _cof else "")
                     + (f"  ⚑ {_ncf} class flag(s)" if _ncf else ""))
        else:
            ui_event("no takeoff cost-code budget found — Budget vs Actual "
                     "skipped", icon="⚑", color=_YEL)

        # LABOR + CONCRETE, per draw — the PM/ops manager's main view (the user
        # 2026-07-29). Materials come as packages; these two are what gets
        # tracked line by line. NO BUDGET → NO SHEET: a scoreboard whose every
        # BUDGET cell is $0 (takeoff unreadable, or the Common drive not
        # mounted) reads as "wildly over budget", which is worse than absent.
        if not _bud:
            ui_warn("Labor/Concrete sheets skipped — no takeoff budget "
                    "(is the Common drive mounted?)")
        _cc = code_costs_by_draw(bills, purchases, cust_info["id"],
                                 draw_periods, account_names=account_names)
        _yards = load_cp_concrete_yards(proj)
        # Columns are headed by the DRAW NUMBER, not the period dates (the
        # user 2026-07-29 — the takeoff template's WEEK columns become draws).
        _dnames = {lbl: nm for nm, lbl, *_rest in draw_rows}
        _dcols = [(l, f"{_dnames.get(l, l)}\n"
                      f"{_s.strftime('%m/%d/%y')}–{_e.strftime('%m/%d/%y')}")
                  for l, _s, _e in draw_periods]
        _atts: Dict[str, str] = {}
        if _bud and not dry_run:
            _seen = set()
            _txns = []
            for _g in _cc.values():
                for _l in _g["lines"]:
                    if _l["txn_id"] and _l["txn_id"] not in _seen:
                        _seen.add(_l["txn_id"])
                        _txns.append((_l["txn_id"], _l["tx_type"], _l["doc"]))
            _atts = fetch_txn_attachments(access, company_id, _txns,
                                          proj_dir / "attachments")
        _marks = (read_back_ledger_marks(proj_dir / f"Project_PnL_{proj}.xlsx")
                  if _bud else {})
        for _kind in ("Labor", "Concrete") if _bud else ():
            _nm = build_sheet_labor_concrete(
                wb, _kind, proj, cust_info, wip_info, _bud, _cc,
                _dcols, as_of, yards=_yards,
                budget_source=_bud_src, realm=company_id, att_links=_atts,
                marks=_marks.get(_kind))
            if _nm:
                _n = _FOCUS_NUM[_kind]
                _b = sum(v for k, v in _bud.items() if _split_code(k)[1] == _n)
                _a = sum(sum(g["draws"].values())
                         for k, g in _cc.items() if _split_code(k)[1] == _n)
                _t = sum(sum(g["tax"].values())
                         for k, g in _cc.items() if _split_code(k)[1] == _n)
                ui_event(f"{_kind}: budget ${_b:,.0f} · actual ${_a:,.0f} "
                         f"({_a - _b:+,.0f})"
                         + (f" · tax ${_t:,.0f} shown separately" if _t else ""),
                         icon="▪", color=(_YEL if _a > _b else _CYAN))

    # Preserve Contract Price / ETC / CO Costs typed into the PRIOR sheet
    # across syncs (CO cost is manual until the CO template has a cost line).
    _saved = read_back_inputs(proj_dir / f"Project_PnL_{proj}.xlsx")
    if _saved.get("contract") is not None:
        wip_info["contract_saved"] = _saved["contract"]
    if _saved.get("etc") is not None:
        wip_info["etc_saved"] = _saved["etc"]
    if _saved.get("co_cost") is not None:
        wip_info["co_cost_saved"] = _saved["co_cost"]
    # CONTRACT PRICE + APPROVED COs come from the SIGNED G702 pay application,
    # not from the draw invoices and not from a hand-typed cell (the user
    # 2026-07-29) — the G702 is the document the GC certifies. The WIP master
    # stays the fallback for jobs with no pay app on the drive yet.
    if not is_mfd:
        _g702 = load_g702(proj)
        if _g702.get("error"):
            ui_warn(f"G702 {_g702.get('source')}: {_g702['error']} — "
                    f"contract falls back to the WIP master")
        elif _g702.get("original_contract") is not None:
            _typed = wip_info.pop("contract_saved", None)
            wip_info["original_contract"] = _g702["original_contract"]
            wip_info["change_orders"] = _g702.get("co_net") or 0.0
            if _g702.get("contract_to_date") is not None:
                wip_info["revised_contract"] = _g702["contract_to_date"]
            wip_info["contract_g702"] = f"G702 pay app, Draw #{_g702['draw_no']}"
            if (_typed is not None
                    and abs(float(_typed) - _g702["original_contract"]) > 1.0):
                wip_info["contract_g702_typed"] = float(_typed)
            ui_event(f"contract per G702 (Draw #{_g702['draw_no']}): "
                     f"${_g702['original_contract']:,.0f} + COs "
                     f"${_g702.get('co_net') or 0:,.0f} = "
                     f"${_g702.get('contract_to_date') or 0:,.0f}"
                     + (f"  ⚑ overrode typed ${_typed:,.0f}"
                        if wip_info.get("contract_g702_typed") else ""),
                     icon="§", color=_CYAN)
    build_sheet_pl(
        wb, proj, cust_info, wip_info, pl_data, pl_totals,
        net_billed, ret_withheld_total, as_of, overhead_pct=overhead_pct,
        pl_cutoff=pl_cutoff, accum=accum, draw_rows=draw_rows,
        draw_anchors=draw_anchors, retainage_nb=retainage_nb,
        retainage_billed_total=ret_billed_total, tx_refs=tx_refs,
        alt_overhead_pct=_alt_oh, underbill_total=underbill_total,
        underbill_count=underbill_count, income_rows=tx.get("income"),
        simple=simple, acct_anchors=acct_anchors, realm=company_id,
    )
    # Order (the user 2026-07-16; Labor/Concrete first among the analysis tabs
    # 2026-07-29 — they're the PM/ops manager's main view): P&L, Transactions,
    # Labor, Concrete, their detail, Budget vs Actual, Next Draw, the draw
    # sheets, POs, Reconciliations; Cash Flow trails.
    _order_sheets(wb, ["P&L", "Transactions", "By Account",
                       "Labor", "Concrete", "Budget vs Actual",
                       *(["Next Draw"] if leftover is not None else []),
                       *draw_sheet_order,
                       "POs", "Reconciliations", "Cash Flow"])

    # Color-code the tabs for navigation (the user 2026-06-26).
    _tabcolors = {"P&L": "1F3A5F", "By Account": "375623", "Cash Flow": "C55A11",
                  "Next Draw": "808080",
                  "Labor": "7030A0", "Concrete": "7030A0",
                  "Budget vs Actual": "BF8F00",
                  "Transactions": "548235", "POs": "808080",
                  "Reconciliations": "808080"}
    for _sn, _col in _tabcolors.items():
        if _sn in wb.sheetnames:
            wb[_sn].sheet_properties.tabColor = _col
    for _sn in draw_sheet_order:                      # all draw tabs = blue
        wb[_sn].sheet_properties.tabColor = "2E75B6"

    proj_dir.mkdir(parents=True, exist_ok=True)
    out_path = proj_dir / f"Project_PnL_{proj}.xlsx"
    saved = safe_save(wb, out_path)
    if saved:
        ui_done(f"wrote {saved.parent.name}/{saved.name}  ·  "
                f"{len(draw_sheet_order)} draw sheet(s)")
    return saved


def generate_project_pnl_rp(
    access: str,
    company_id: str,
    proj: str,
    cust_info: dict,
    wip_info: dict,
    start_date: str,
    end_date: str,
    out_dir: Path,
    as_of: str,
    dry_run: bool = False,
    overhead_pct: float = 10.0,
) -> Optional[Path]:
    """
    RESIDENTIAL "Job P&L" template — no draws, no retainage (the user 2026-06-09).
    Sheets: Job P&L (JOB PROFIT → WIP/Projection → INVOICE → JOB COSTS account →
    vendor → bills), Cash Flow, Transactions, Pending Review, POs, Reconciliations.
    """
    ui_event("Residential Job P&L template", color=_DIM)
    invoices = fetch_customer_invoices(access, company_id, cust_info["id"])
    inv_info = []
    inv_dates = []
    rp_not_billed = []          # retainage-not-billed JE docs — NOT income (the user 2026-07-02)
    for inv in invoices:
        rec = {
            "doc_num": _xml_clean(inv.get("DocNumber", "") or inv.get("Id", "")),
            "id": inv.get("Id", ""),
            "date": inv.get("TxnDate", ""),
            "memo": _xml_clean(inv.get("PrivateNote", "") or ""),
            "amount": float(inv.get("TotalAmt", 0) or 0),
        }
        if _is_retainage_not_billed(inv):
            rp_not_billed.append(rec)          # moved to retainage receivable, not revenue
            continue
        d = _parse_date(inv.get("TxnDate", ""))
        if d:
            inv_dates.append(d)
        inv_info.append(rec)
    billed_total = sum(i["amount"] for i in inv_info)
    invoice_date = max(inv_dates) if inv_dates else None
    if rp_not_billed:
        ui_event(f"{len(rp_not_billed)} retainage-not-billed doc(s) "
                 f"${sum(i['amount'] for i in rp_not_billed):,.0f} EXCLUDED from income "
                 f"(moved to retainage receivable)", icon="⚑", color=_YEL)
    ui_event(f"{len(invoices)} invoice(s) · billed ${billed_total:,.0f} · "
             f"invoice date {invoice_date}")

    accounts = query_all(access, company_id, "Account")
    parent_map = build_account_parent_map(accounts)
    account_names = {a.get("Id"): a.get("Name") for a in accounts if a.get("Id")}
    account_fqn = {a.get("Id"): ((a.get("FullyQualifiedName") or a.get("Name") or "")
                                 .replace(":", ": "))
                   for a in accounts if a.get("Id")}
    acct_type = build_account_type_map(accounts)

    items = query_all(access, company_id, "Item")
    item_account = {it.get("Id"): (it.get("ExpenseAccountRef") or {}).get("value")
                    for it in items if it.get("Id")}

    pl_data = fetch_project_pl(access, company_id, cust_info["id"], start_date, end_date)
    pl_totals = extract_pl_totals(pl_data)
    ui_event(f"{len(accounts)} accounts · {len(items)} items · "
             f"P&L income ${pl_totals['income']:,.0f}")

    # Budget vs Actual was DROPPED (the user 2026-06-19): QBO's Project cost-estimate
    # feature isn't exposed by the Accounting API, so there's no cost-code budget
    # source. The dormant budget code was removed in the 2026-06-26 cleanup.

    # Narrow the bill/purchase pull to project life (earliest invoice − 180d).
    bill_start = start_date
    if inv_dates:
        cand = (min(inv_dates) - dt.timedelta(days=180)).isoformat()
        if cand > start_date:
            bill_start = cand
    bills, purchases = fetch_customer_bills_and_purchases(
        access, company_id, cust_info["id"], bill_start, end_date,
    )
    # {txn id: (balance, total)} — the balance is what makes PARTIAL possible.
    paid_map = {b.get("Id"): (float(b.get("Balance", 0) or 0),
                              float(b.get("TotalAmt", 0) or 0)) for b in bills}
    paid_map.update({pch.get("Id"): (0.0, 0.0) for pch in purchases})
    ui_event(f"{len(bills)} bills · {len(purchases)} purchases  "
             f"{_DIM}(from {bill_start}){_RESET}")

    pos = fetch_customer_purchase_orders(
        access, company_id, cust_info["id"], bill_start, end_date,
    )
    po_unused, po_used = match_pos_to_bills(pos, bills, cust_info["id"])

    job_groups, job_total, pending, dup_flags = gather_rp_costs(
        bills, purchases, cust_info["id"], invoice_date, parent_map,
        account_names=account_names, account_fqn=account_fqn,
        item_account=item_account,
    )
    # Open POs dated ON/BEFORE the invoice = committed cost, bill pending →
    # add to the P&L (yellow). After-invoice opens stay buffer (PO sheet only).
    unused_ids = {rec["id"] for rec in po_unused if rec.get("id")}
    pend_total, pend_n = inject_pending_pos(
        job_groups, pos, unused_ids, cust_info["id"], invoice_date,
        parent_map, account_names=account_names, account_fqn=account_fqn,
        item_account=item_account,
    )
    job_total += pend_total
    ui_event(f"job costs ${job_total:,.0f}  {_DIM}({pend_n} bill-pending POs "
             f"${pend_total:,.0f} · {len(pending)} pending-review · "
             f"{len(dup_flags)} dup flags){_RESET}")

    # Cash flow — same as the Draw template (the user 2026-06-26): actual payment dates.
    bill_pmts = fetch_bill_payments(access, company_id, bill_start, end_date)
    cust_pmts = fetch_customer_payments(access, company_id, bill_start, end_date)
    cash_events = build_cashflow_events(bills, invoices, bill_pmts, cust_pmts,
                                        cust_info["id"])
    if cash_events:
        _trough = min(e["running"] for e in cash_events)
        ui_event(f"cash flow: {len(cash_events)} payments · peak ${_trough:,.0f} · "
                 f"now ${cash_events[-1]['running']:,.0f}", icon="$", color=_CYAN)

    if dry_run:
        ui_event(f"[dry run] would write {proj}/Project_PnL_{proj}.xlsx", color=_YEL)
        return None

    wb = Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    build_sheet_job_rp(wb, proj, cust_info, wip_info, inv_info, job_groups,
                       job_total, billed_total, as_of, overhead_pct=overhead_pct,
                       realm=company_id)
    # Transactions sheet — same traceability as the Draw template (the user 2026-06-22).
    # RP has no draws/retainage, so income is the invoices billed; bills split
    # COGS vs Expense by account type.
    rp_income_groups = {"__rp": {"invoices": [
        {"doc_num": i["doc_num"], "id": i.get("id", ""), "date": i["date"], "memo": i["memo"],
         "gross": i["amount"], "retainage": 0.0, "retainage_billed": 0.0}
        for i in inv_info]}}
    if rp_not_billed:           # show them on the Transactions sheet, excluded from income
        rp_income_groups["__retainage"] = {
            "invoices": [{"doc_num": i["doc_num"], "id": i.get("id", ""), "date": i["date"],
                          "memo": i["memo"], "amount": i["amount"],
                          "lines": [{"desc": "Retainage (not billed)", "amt": i["amount"]}]}
                         for i in rp_not_billed],
            "total": sum(i["amount"] for i in rp_not_billed)}
    tx = gather_transactions(rp_income_groups, bills, purchases, cust_info["id"],
                             parent_map, account_names=account_fqn,
                             acct_type=acct_type, item_account=item_account)
    tx_refs = build_sheet_transactions(wb, proj, cust_info, wip_info, tx, as_of,
                                       paid_map=paid_map,
                                       realm=company_id)
    qbo_exp = pl_totals.get("gross_profit", 0.0) - pl_totals.get("net_ordinary_income", 0.0)
    # Difference-finder candidates (same as Draw template): cost lines + invoices.
    diff_candidates = []
    for vg in {**tx.get("cogs", {}), **tx.get("exp", {})}.values():
        for ln in vg:
            diff_candidates.append({"ref": str(ln.get("ref", "")),
                                    "amount": round(float(ln.get("amount", 0) or 0), 2),
                                    "party": "", "txn_id": ln.get("txn_id", ""),
                                    "tx_type": ln.get("tx_type", "Bill")})
    for i in inv_info:
        diff_candidates.append({"ref": str(i.get("doc_num", "")),
                                "amount": round(float(i.get("amount", 0) or 0), 2),
                                "party": cust_info["name"], "txn_id": i.get("id", ""),
                                "tx_type": "invoice"})
    build_sheet_reconciliations(
        wb, proj, cust_info, wip_info, pl_totals.get("income", 0.0),
        pl_totals.get("cogs", 0.0), qbo_exp, tx_refs, as_of, has_retainage=False,
        candidates=diff_candidates, tx_totals=tx["tot"], realm=company_id)
    build_sheet_pending_rp(wb, proj, cust_info, wip_info, pending, dup_flags, as_of,
                           company_id=company_id)
    build_sheet_pos(wb, proj, cust_info, wip_info, po_unused, po_used, as_of, realm=company_id)
    build_sheet_cashflow(wb, proj, cust_info, wip_info, cash_events, as_of, realm=company_id)
    # Budget vs Actual (the user 2026-07-16): RP budget = the takeoff's Cost
    # Gral sheet (FW codes → the -FTW project; SL/PR → the slab project).
    _bud, _bud_src = load_rp_budget(proj)
    if _bud:
        _cof = False
        try:
            _cof = bool(float(wip_info.get("change_orders") or 0))
        except (TypeError, ValueError):
            pass
        _acts = costs_by_code(bills, purchases, cust_info["id"], parent_map,
                              account_names=account_names)
        build_sheet_budget_vs_actual(
            wb, proj, cust_info, wip_info, _bud, _acts, as_of,
            co_flag=_cof, budget_source=_bud_src, realm=company_id)
        _ncf = sum(1 for g in _acts.values() for t in g.get("txns", [])
                   if not _class_ok(t.get("class"), _expected_class(proj)))
        ui_event(f"Budget vs Actual: {len(_bud)} budgeted codes "
                 f"(${sum(_bud.values()):,.0f}) from {_bud_src}"
                 + ("  ⚑ CO flag" if _cof else "")
                 + (f"  ⚑ {_ncf} class flag(s)" if _ncf else ""))
    else:
        ui_event("no takeoff Cost Gral budget found — Budget vs Actual "
                 "skipped", icon="⚑", color=_YEL)
    _order_sheets(wb, ["Job P&L", "Transactions", "Budget vs Actual",
                       "Pending Review", "POs", "Reconciliations", "Cash Flow"])

    # Color-code the tabs to match the Draw template (the user 2026-06-26).
    for _sn, _col in {"Job P&L": "1F3A5F", "Cash Flow": "C55A11",
                      "Transactions": "548235", "Budget vs Actual": "BF8F00",
                      "Pending Review": "808080", "POs": "808080",
                      "Reconciliations": "808080"}.items():
        if _sn in wb.sheetnames:
            wb[_sn].sheet_properties.tabColor = _col

    # Folder + file named "RP#### - Customer" (the user 2026-06-26). Customer = the top
    # builder (first segment of the fully-qualified name, e.g. "Grand Homes").
    _fqn = cust_info.get("fully_qualified_name") or cust_info.get("name") or ""
    _client = (_fqn.split(":")[0].strip() if ":" in _fqn else (cust_info.get("name") or ""))
    _client = re.sub(r'[:\\/?*\[\]<>|"]', "-", _client).strip()
    label = f"{proj} - {_client}" if _client else proj
    proj_dir = out_dir / label                    # one home folder per project
    proj_dir.mkdir(parents=True, exist_ok=True)
    out_path = proj_dir / f"{label}.xlsx"
    saved = safe_save(wb, out_path)
    if saved:
        ui_done(f"wrote {label}/{saved.name}  ·  Residential Job P&L")
    return saved


# A "prior draw" marker the PM writes in a note column to the RIGHT of the amount
# ("INCLUDED IN APRIL REPORT", "IN APRIL", "already billed") — means the cost was
# pulled in an earlier draw, so it must NOT be flagged as missed this draw.
_PRIOR_MARK_RE = re.compile(
    r"includ|already|\bprior\b|\bprev|^\s*in\s+"
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", re.I)

_MONTHS = ("january february march april may june july august september "
           "october november december").split()


def _parse_report_period(text: str):
    """Parse a QB report date-range header into (start, end). Handles
    'March 1-April 28, 2026' (cross-month), 'February 2-26, 2026' (same month),
    and 'March 1-30, 2026'. Returns None if no range found."""
    m = re.search(
        r"([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*(?:([A-Za-z]+)\s+)?(\d{1,2}),?\s*(\d{4})",
        text)
    if not m:
        return None
    mon1, d1, mon2, d2, yr = m.groups()
    mon2 = mon2 or mon1
    try:
        s = dt.datetime.strptime(f"{mon1} {d1} {yr}", "%B %d %Y").date()
        e = dt.datetime.strptime(f"{mon2} {d2} {yr}", "%B %d %Y").date()
        return (s, e)
    except ValueError:
        return None


def parse_pm_report(path: Path) -> dict:
    """Parse a PM 'Cost By Vendor' / project-cost draw report (the user 2026-06-23).
    Returns {project, period:(start,end), lines:[{vendor,num,date,account,desc,
    amount,prior}]}.

    QB exports these in MANY column layouts (6/7/9/15/16 cols; Amount in col F/G/H;
    optional 'Transaction type' column; sidecar tables to the right). So instead of
    hardcoding columns, DETECT them from the report's own header row — the row that
    contains 'Date', 'Num', and 'Amount' (the user 2026-06-25, after the parser silently
    failed on every layout but MAYHILL and flagged real bills as 'missed')."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    maxc = ws.max_column

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    proj = None
    for r in range(1, 14):
        for c in range(1, 6):
            m = re.search(r"\b(MFD|CP|RP)\s*\d+\b", str(cell(r, c) or ""), re.I)
            if m:
                proj = m.group(0).replace(" ", "").upper()
                break
        if proj:
            break

    period = None
    for r in range(1, 7):
        period = _parse_report_period(str(cell(r, 1) or ""))
        if period:
            break

    # Locate the column-header row (has Date + Num + Amount) and map columns.
    hdr_row, cols = None, {}
    for r in range(1, 18):
        labels = {c: str(cell(r, c)).strip().lower()
                  for c in range(1, maxc + 1) if cell(r, c) not in (None, "")}
        low = set(labels.values())
        if "date" in low and "num" in low and "amount" in low:
            def pick(*cands):
                for cand in cands:                       # exact first, then contains
                    for c, v in labels.items():
                        if v == cand:
                            return c
                    for c, v in labels.items():
                        if cand in v:
                            return c
                return None
            hdr_row = r
            cols = {"date": pick("date"), "num": pick("num"),
                    "amount": pick("amount"),
                    "desc": pick("description", "memo/description", "memo"),
                    "acct": pick("distribution account", "account")}
            break

    lines, vendor = [], None
    if hdr_row and cols["date"] and cols["num"] and cols["amount"]:
        dcol, ncol, acol = cols["date"], cols["num"], cols["amount"]
        for r in range(hdr_row + 1, ws.max_row + 1):
            dval = cell(r, dcol)
            dparsed = (dval.date() if hasattr(dval, "date")
                       else _parse_date(str(dval)) if dval else None)
            if dparsed:                                  # ── transaction row ──
                amt, amt_at = None, None
                for cc in (acol, acol + 1):              # MAYHILL shifts amount +1
                    v = cell(r, cc)
                    if isinstance(v, (int, float)):
                        amt, amt_at = float(v), cc
                        break
                if not amt:                              # skip blank/0 desc-only lines
                    continue
                prior = any(
                    isinstance(cell(r, cc), str) and _PRIOR_MARK_RE.search(cell(r, cc))
                    for cc in range(amt_at + 1, maxc + 1))
                lines.append({
                    "vendor": vendor or "", "num": str(cell(r, ncol) or "").strip(),
                    "date": dparsed,
                    "account": str(cell(r, cols["acct"]) or "").strip() if cols["acct"] else "",
                    "desc": str(cell(r, cols["desc"]) or "").strip() if cols["desc"] else "",
                    "amount": round(amt, 2), "prior": prior})
            else:                                        # ── label row ──
                A = cell(r, 1)
                if A and isinstance(A, str):
                    s = A.strip()
                    if re.match(r"(?i)^total\b", s):
                        continue                         # vendor/section/grand total
                    if re.search(r"\b(MFD|CP|RP)\s*\d+\b", s, re.I):
                        continue                         # project subheader under vendor
                    vendor = s                           # otherwise: a new vendor

    if period is None and lines:                         # fall back to min/max dates
        ds = [l["date"] for l in lines if isinstance(l["date"], dt.date)]
        if ds:
            period = (min(ds), max(ds))
    return {"project": proj, "period": period, "lines": lines}


def compare_report_to_qbo(rep_lines: list, qbo_lines: list):
    """Match by Bill#(num) + round(amount,2), MULTISET. Returns (missing, extra):
    missing = QBO lines not in the report (PM didn't pull → underbilled); extra =
    THIS-DRAW report lines with no QBO match (typo / wrong amount).

    ALL report lines — including prior-period ones (PM marks "INCLUDED IN <prior>
    REPORT") — are in the match set so a bill the PM already pulled in an earlier
    draw is NOT flagged as missed. But prior lines are EXCLUDED from `extra`: their
    bills are dated before this window, so "no in-window QBO match" is expected,
    not an error."""
    from collections import Counter
    rep_ct = Counter((l["num"], l["amount"]) for l in rep_lines)
    missing = []
    for q in qbo_lines:
        k = (q["num"], q["amount"])
        if rep_ct.get(k, 0) > 0:
            rep_ct[k] -= 1
        else:
            missing.append(q)
    extra, used = [], Counter()
    for l in rep_lines:
        if l.get("prior"):           # prior-draw line — not expected in this window
            continue
        k = (l["num"], l["amount"])
        if rep_ct.get(k, 0) - used.get(k, 0) > 0:
            extra.append(l); used[k] += 1
    return missing, extra


def _qbo_lines_from_fetched(bills, purchases, customer_id, anames, item_acct,
                            start=None, end=None) -> list:
    """Build cost LINES {num,amount,vendor,account,date,desc,tx_type,txn_id} from
    already-fetched bills/purchases, optionally date-windowed (so the orchestration
    can reuse its single fetch instead of re-pulling per report)."""
    out = []

    def in_win(ds):
        if not (start or end):
            return True
        d = _parse_date(ds)
        if not d:
            return True
        return not ((start and d < start) or (end and d > end))

    def take(txn, tx_type, vfield):
        if not in_win(txn.get("TxnDate", "")):
            return
        vendor = _xml_clean(((txn.get(vfield) or {}).get("name") or "").strip())
        num = _xml_clean(str(txn.get("DocNumber") or ""))
        tid = txn.get("Id", "")
        date = txn.get("TxnDate", "")
        for ln in txn.get("Line") or []:
            det = (ln.get("AccountBasedExpenseLineDetail")
                   or ln.get("ItemBasedExpenseLineDetail") or {})
            if not (det and _line_belongs(det, ln, txn, customer_id)):
                continue
            amt = float(ln.get("Amount", 0) or 0)
            if amt == 0:
                continue
            aid = ((det.get("AccountRef") or {}).get("value")
                   or item_acct.get((det.get("ItemRef") or {}).get("value")))
            out.append({"num": num, "amount": round(amt, 2), "vendor": vendor,
                        "account": _xml_clean(anames.get(aid) or ""), "date": date,
                        "desc": _xml_clean((ln.get("Description") or "").strip()),
                        "tx_type": tx_type, "txn_id": tid})

    for b in bills:
        take(b, "Bill", "VendorRef")
    for p in purchases:
        take(p, "Expense", "EntityRef")
    return out


def _qbo_project_cost_lines(access, company_id, customer_id, start, end) -> list:
    """Fetch + build cost lines for the standalone cross-check."""
    bills, purchases = fetch_customer_bills_and_purchases(
        access, company_id, customer_id, start, end)
    # {txn id: (balance, total)} — the balance is what makes PARTIAL possible.
    paid_map = {b.get("Id"): (float(b.get("Balance", 0) or 0),
                              float(b.get("TotalAmt", 0) or 0)) for b in bills}
    paid_map.update({pch.get("Id"): (0.0, 0.0) for pch in purchases})
    accounts = query_all(access, company_id, "Account")
    anames = {a.get("Id"): a.get("Name") for a in accounts if a.get("Id")}
    items = query_all(access, company_id, "Item")
    item_acct = {it.get("Id"): (it.get("ExpenseAccountRef") or {}).get("value")
                 for it in items if it.get("Id")}
    return _qbo_lines_from_fetched(bills, purchases, customer_id, anames, item_acct)


def cross_check_draw_report(access, company_id, report_path: Path,
                            cust_map: dict, out_dir: Path) -> Optional[Path]:
    """Cross-check a PM draw-cost report against QBO. Flags MISSING (in QBO, not
    in the report — costs the PM didn't pull) and EXTRA (report lines with no QBO
    match — possible typo / wrong amount). Writes a standalone workbook so the
    main P&L is never touched (the user 2026-06-23). Match key = Bill# + line amount."""
    print(f"\n  Cross-checking {report_path.name} ...")
    rep = parse_pm_report(report_path)
    proj = rep["project"]
    if not proj:
        print("    ✗ couldn't read a project # from the report"); return None
    if proj not in cust_map:
        print(f"    ✗ {proj} not found in QBO customers"); return None
    if not rep["period"]:
        print("    ✗ couldn't read the report period"); return None
    start, end = rep["period"]
    print(f"    {proj}  period {start}→{end}  ({len(rep['lines'])} report lines)")
    cust = cust_map[proj]
    qbo = _qbo_project_cost_lines(access, company_id, cust["id"],
                                  start.isoformat(), end.isoformat())
    print(f"    {len(qbo)} QBO cost lines in window")

    missing, extra = compare_report_to_qbo(rep["lines"], qbo)

    wb = Workbook()
    realm = company_id
    NB, RED, GRN = NAVY, "C00000", "375623"

    def sect(ws, r, txt, color=NB):
        c = _write_cell(ws, r, 1, txt)
        c.font = Font(bold=True, size=BASE_SIZE + 1, color=color)
        return r + 1

    def table(ws, r, headers, rows, link_col=None):
        for ci, h in enumerate(headers, start=1):
            hc = _write_cell(ws, r, ci, h)
            hc.font = Font(bold=True, color=NB); hc.border = BOTTOM_BORDER
        r += 1
        for row in rows:
            for ci, v in enumerate(row, start=1):
                cc = (ws.cell(row=r, column=ci, value=v) if isinstance(v, (int, float))
                      else _write_cell(ws, r, ci, v))
                cc.font = Font(size=BASE_SIZE - 1)
                if isinstance(v, (int, float)):
                    cc.number_format = CURR_FMT
            r += 1
        return r

    # ── Summary ──
    ws = wb.active; ws.title = "Cross-Check"
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 22), ("B", 12), ("C", 22), ("D", 40), ("E", 14), ("F", 14)):
        ws.column_dimensions[col].width = w
    t = _write_cell(ws, 1, 1, f"DRAW CROSS-CHECK — {proj}")
    t.font = Font(bold=True, size=BASE_SIZE + 3, color=NB)
    _write_cell(ws, 2, 1, f"Report: {report_path.name}   ·   Period {start}–{end}"
                ).font = Font(italic=True, size=BASE_SIZE - 1, color="595959")
    miss_tot = sum(m["amount"] for m in missing)
    extra_tot = sum(e["amount"] for e in extra)
    _write_cell(ws, 3, 1, f"MISSING (in QBO, not in report): {len(missing)} lines  "
                f"${miss_tot:,.2f}        EXTRA (no QBO match): {len(extra)} lines  "
                f"${extra_tot:,.2f}").font = Font(bold=True, size=BASE_SIZE, color=RED)
    r = 5
    r = sect(ws, r, "MISSING IN DRAW  —  in QBO but NOT pulled by the PM", RED)
    r = table(ws, r, ["Bill #", "Date", "Vendor", "Description", "Account", "Amount"],
              [[m["num"], m["date"], m["vendor"], m["desc"], m["account"], m["amount"]]
               for m in sorted(missing, key=lambda x: -x["amount"])])
    if missing:
        _write_cell(ws, r, 4, "TOTAL MISSING").font = Font(bold=True)
        mt = ws.cell(row=r, column=6, value=miss_tot); mt.number_format = CURR_FMT
        mt.font = Font(bold=True, color=RED); r += 1
        # hyperlink the Bill # cells
        for i, m in enumerate(sorted(missing, key=lambda x: -x["amount"])):
            url = _qbo_txn_url(m["tx_type"], m["txn_id"], realm)
            if url:
                cc = ws.cell(row=r - len(missing) - 1 + i, column=1)
                cc.hyperlink = url; cc.font = Font(size=BASE_SIZE - 1, color="0563C1", underline="single")
    r += 2
    r = sect(ws, r, "EXTRA IN REPORT  —  report lines with NO QBO match (check typo/amount)", "C55A11")
    r = table(ws, r, ["Bill #", "Date", "Vendor", "Description", "Account", "Amount"],
              [[e["num"], e["date"], e["vendor"], e["desc"], e["account"], e["amount"]]
               for e in sorted(extra, key=lambda x: -x["amount"])])

    # ── PM Report (parsed, memorized) ──
    ws2 = wb.create_sheet("PM Report (parsed)")
    ws2.sheet_view.showGridLines = False
    for col, w in (("A", 22), ("B", 12), ("C", 22), ("D", 40), ("E", 14), ("F", 12)):
        ws2.column_dimensions[col].width = w
    table(ws2, 1, ["Bill #", "Date", "Vendor", "Description", "Account", "Amount", "In this draw?"],
          [[l["num"], l["date"], l["vendor"], l["desc"], l["account"], l["amount"],
            ("prior draw" if l["prior"] else "THIS DRAW")] for l in rep["lines"]])

    for w in (ws, ws2):
        _setup_print(w, 6)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{report_path.stem} — CROSS-CHECK.xlsx"
    saved = safe_save(wb, out_path)
    if saved:
        print(f"    ✓ {saved.name}  ({len(missing)} missing ${miss_tot:,.0f}, "
              f"{len(extra)} extra ${extra_tot:,.0f})")
    return saved


def index_pm_reports(reports_dir, proj: str):
    """Pool ALL PM reports for this project into one capture index (the user 2026-06-26).
    The report's own header period is IGNORED — the director readjusts draw periods,
    so a bill is 'captured' if it appears on ANY report in the folder. Returns:
      index  : {(bill#, amount): {report filenames that list it}}
      parsed : [(filename, parsed-report)] for relevance/orphan lookups."""
    index, parsed = {}, []
    if not reports_dir:
        return index, parsed
    d = Path(reports_dir).expanduser()
    if not d.exists():
        return index, parsed
    for f in sorted(d.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            rep = parse_pm_report(f)
        except Exception:
            continue
        if (rep.get("project") or "").upper() != (proj or "").upper():
            continue
        parsed.append((f.name, rep))
        for ln in rep["lines"]:
            index.setdefault((str(ln["num"]).strip(), round(float(ln["amount"]), 2)),
                             set()).add(f.name)
    if parsed:
        print(f"      indexed {len(parsed)} PM report(s) → {len(index)} bill keys for {proj}")
    return index, parsed


def expand_active_projects(tokens: List[str],
                           wip_master: Dict[str, dict]) -> Tuple[List[str], bool]:
    """`ACTIVE [CP|RP|MFD ...]` → every Active project of those divisions from
    the WIP master (Test-Master STATUS), so ONE run refreshes a whole division
    (the user 2026-07-16: `project-pnl active cp`). No division token = all
    three. Non-keyword tokens still pass through, so `active cp MFD177` works.
    Returns (projects, expanded?)."""
    # Split on the separators people actually type: `active-cp`, `active,cp`
    # and `active cp` all mean the same batch. Without this, `active-cp` fell
    # through as a literal project number, QBO had no customer called
    # ACTIVE-CP, and the run did nothing but say "skipped" (the user
    # 2026-08-25: "it's not loading").
    toks = [part for t in tokens
            for part in re.split(r"[-,/]", t.strip().upper()) if part]
    if "ACTIVE" not in toks:
        return [t.strip().upper() for t in tokens], False
    divs = {t for t in toks if t in ("CP", "RP", "MFD")} or {"CP", "RP", "MFD"}
    extras = [t for t in toks if t not in ("ACTIVE", "CP", "RP", "MFD")]
    matched = sorted(
        p for p, info in wip_master.items()
        if str(info.get("status") or "").strip().lower() == "active"
        and any(p.startswith(d) for d in divs))
    return matched + [e for e in extras if e not in matched], True


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate per-project P&L workbooks from QBO")
    ap.add_argument("projects", nargs="+",
                    help="Project numbers (e.g. MFD177 CP672), or `active` plus "
                         "divisions (`active cp`, `active rp mfd`) to batch every "
                         "Active project from the WIP master — OR drag/drop a "
                         "PM draw-cost .xlsx report to cross-check it vs QBO")
    ap.add_argument("--out", default=str(DEFAULT_OUT),
                    help=f"Output folder (default: {DEFAULT_OUT})")
    ap.add_argument("--start-date", default="2020-01-01")
    ap.add_argument("--end-date", default=dt.date.today().isoformat())
    ap.add_argument("--wip-master", default=str(DEFAULT_WIP_MASTER),
                    help="Path to WIP master .xlsx for Contract/ETC lookup")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--overhead-pct", type=float, default=10.0,
                    help="Company overhead as %% of revenue (default 10.0, the user 2026-07-16). "
                         "Drives Overhead Allocation + True Net Profit rows.")
    ap.add_argument("--legacy", action="store_true",
                    help="LEGACY JOB attribution for jobs that predate "
                         "consistent project coding: a cost line counts when "
                         "the line's customer is the project, OR the line text "
                         "names the job, OR the bill memo names it and names "
                         "exactly one job. Also pulls invoices billed on the "
                         "PARENT customer. Off by default — see "
                         "shared/job_lines.py.")
    ap.add_argument("--alias", action="append", default=[],
                    help="Street name the job goes by (repeatable), e.g. "
                         "--alias 'BONDS RANCH'. Only used with --legacy.")
    ap.add_argument("--job-class", default="",
                    help="The job's OWN class branch, e.g. "
                         "'MULTI FAMILY:MARKER LAPIZ'. Matched as a PREFIX so "
                         "the live parent and its deleted per-job leaf both "
                         "count. A bare division name is refused. Only used "
                         "with --legacy.")
    ap.add_argument("--class", dest="use_class", action="store_true",
                    help="Add the job's CLASS-coded lines to the P&L. The class "
                         "is found automatically from the job number, active or "
                         "not - you never type its name. Shorthand: put +class "
                         "anywhere on the command line.")
    ap.add_argument("--simple", action="store_true",
                    help="STRIPPED-BACK P&L for a COMPLETED job: no draw "
                         "sheets, no Next Draw sheet, and no draw-coverage or "
                         "accumulating-costs blocks on the P&L. Shorthand: "
                         "+simple anywhere on the command line.")
    ap.add_argument("--class-project", action="store_true",
                    help="CLASS/PROJECT LOOKUP (the user 2026-08-25): the job's "
                         "cost is exactly its CLASS lines plus its PROJECT "
                         "lines, with the line-text and bill-memo rules OFF. "
                         "For a job that ran across the class→project coding "
                         "switchover. Implies --legacy; needs --job-class.")
    ap.add_argument("--infer-periods", action="store_true",
                    help="Learn the draw-window shape from the invoices that DO "
                         "carry a (Period:…) tag and apply it to the ones that "
                         "don't, instead of falling back to the calendar month. "
                         "The draw's MONTH still comes from the memo's own "
                         "wording. Retainage invoices are left untagged.")
    ap.add_argument("--no-prompt", action="store_true",
                    help="Don't pause to ask about mistyped invoice period "
                         "dates (skip them and warn instead). Use for "
                         "unattended/scheduled runs.")
    args = ap.parse_args()

    # Drag/drop of a .xlsx report → cross-check mode; bare names → P&L export.
    def _is_report(a):
        return a.lower().endswith(".xlsx") and Path(a).expanduser().exists()
    report_files = [Path(a).expanduser() for a in args.projects if _is_report(a)]
    # `+class` reads naturally mid-command ("project-pnl MFD228 +class") and
    # argparse hands it through as a positional, so pull it out here rather
    # than making anyone type a class name they should never have to know
    # (the user 2026-08-25: "i will never remember that huge line of text").
    if any(a.strip().lower() in ("+class", "+classes") for a in args.projects):
        args.use_class = True
    if any(a.strip().lower() in ("+simple", "+basic") for a in args.projects):
        args.simple = True
    projects = [a.strip().upper() for a in args.projects
                if not _is_report(a) and not a.strip().startswith("+")]
    out_dir = Path(args.out).expanduser()
    # Prompt to fix mistyped period dates only when attached to a terminal.
    interactive = (not args.no_prompt) and sys.stdin.isatty()

    # WIP master loads BEFORE auth: `active <division>` expands from it, and
    # it's a local file read — no Touch ID needed to know the batch.
    wip_master = load_wip_master(Path(args.wip_master).expanduser())
    projects, was_expanded = expand_active_projects(projects, wip_master)
    if was_expanded and not projects:
        print(f"  ✗ `active` found no Active projects in the WIP master "
              f"({args.wip_master}) — is the Test-Master tab current?")
        return 1

    # --alias is GLOBAL to the run, so a batch would apply EVERY job's street
    # name to EVERY job. That is exactly how MFD228 picked up MFD172's Bonds
    # Ranch costs and reported 4.2M instead of 880K (2026-08-27). Aliases are
    # per-job by nature — refuse the combination rather than produce a wrong
    # number that looks plausible.
    if args.alias and len(projects) > 1:
        print(f"  ✗ --alias applies to EVERY project in the run, so it cannot be "
              f"used with {len(projects)} projects at once "
              f"({', '.join(projects[:4])}{'…' if len(projects) > 4 else ''}).\n"
              f"    Run each job on its own with its own --alias, or drop --alias "
              f"(+class alone is per-job and safe in a batch).")
        return 1

    ui_banner("Project P&L Export")
    if was_expanded:
        ui_cfg("Active batch", f"{len(projects)} project(s) from the WIP master")
    ui_cfg("Projects", ", ".join(projects) or "(none)")
    if report_files:
        ui_cfg("Cross-check", ", ".join(f.name for f in report_files))
    ui_cfg("Dates", f"{args.start_date} → {args.end_date}")
    ui_cfg("Output", f"{out_dir}")
    ui_cfg("", f"{_DIM}one folder per project: <out>/<PROJ>/{_RESET}")
    if args.dry_run:
        ui_cfg("Mode", f"{_YEL}DRY RUN — no files written{_RESET}")
    ui_close()

    print(f"\n  {_DIM}Authenticating to QBO (Touch ID)…{_RESET}")
    access, company_id = load_credentials()
    ui_step("Connected to QBO")
    cust_map = build_project_customer_map(access, company_id)
    ui_step("Project → customer map", f"{len(cust_map)} projects")
    ui_step("WIP master loaded", f"{len(wip_master)} rows")

    _ALL_CLASSES = [None]        # lazy: pulled once per run, only if +class
    as_of = dt.datetime.now().strftime("%Y-%m-%d %I:%M %p")  # 12-hour + AM/PM
    generated: List[Path] = []
    not_found: List[str] = []

    for proj in projects:
        if proj not in cust_map:
            ui_proj(proj, "not found in QBO customers")
            ui_warn("skipped — no matching customer")
            not_found.append(proj)
            continue
        # Per project, so a batch can never carry one job's aliases into
        # the next; clears itself when --legacy is off.
        # +class / --class: find the job's OWN class from the job number,
        # active or not, and key on its ID so a QBO reactivate-rename can't
        # break it (the user 2026-08-25 - QBO renames on reactivate).
        _cls = {}
        if args.use_class or args.class_project:
            if _ALL_CLASSES[0] is None:              # one pull, reused by a batch
                _ALL_CLASSES[0] = (query_all(access, company_id, "Class")
                                   + query_all(access, company_id, "Class",
                                               "Active = false"))
            _cls = discover_job_classes(_ALL_CLASSES[0], proj)
            if not _cls:
                ui_warn(f"{proj}: no class names this job — continuing without "
                        f"the class rule")
        try:
            _set_legacy_matcher(
                proj, cust_map[proj]["id"],
                args.legacy or args.class_project or args.use_class,
                args.alias, args.job_class,
                # +class alone = project ∪ class ("the whole P&L, plus the
                # class lines"). Adding --legacy/--alias turns the line-text
                # and bill-memo rules back on as well.
                text_rules=not (args.class_project
                                or (args.use_class and not args.legacy)),
                class_ids=list(_cls.keys()))
        except ValueError as e:
            ui_fail(f"{proj}: {e}")
            return 1
        if args.class_project and not (args.job_class or _cls):
            ui_fail(f"{proj}: no class found for this job — the class/project "
                    f"lookup would be project-only.")
            return 1
        if args.legacy or args.class_project or args.use_class:
            _cn = ", ".join(sorted(_cls.values())) or args.job_class
            ui_event(("CLASS/PROJECT lookup ON for " if args.class_project
                      else "legacy attribution ON for ") + proj
                     + (f"  · class: {_cn}" if _cn else "")
                     + (f"  · aliases: {', '.join(args.alias)}"
                        if args.alias else ""),
                     icon="⚑", color=_YEL)
        try:
            path = generate_project_pnl(
                access, company_id, proj,
                infer_periods=args.infer_periods,
                cust_info=cust_map[proj],
                wip_info=wip_master.get(proj, {}),
                start_date=args.start_date,
                end_date=args.end_date,
                out_dir=out_dir,
                as_of=as_of,
                dry_run=args.dry_run,
                overhead_pct=args.overhead_pct,
                interactive=interactive,
                simple=args.simple,
            )
            if path:
                generated.append(path)
        except Exception as e:
            ui_fail(f"{proj}: {e}")
            if os.getenv("ACB_DEBUG"):
                import traceback
                traceback.print_exc()

    for rf in report_files:
        try:
            p = cross_check_draw_report(access, company_id, rf, cust_map, out_dir)
            if p:
                generated.append(p)
        except Exception as e:
            ui_fail(f"cross-check {rf.name}: {e}")

    ui_banner(f"Done — {len(generated)} workbook(s)")
    if not_found:
        ui_cfg("Not in QBO", f"{_YEL}{', '.join(not_found)}{_RESET}")
    for p in generated:
        print(f"  {_GREEN}→{_RESET} {_BOLD}{p.name}{_RESET}")
        print(f"    {_DIM}{p.parent}{_RESET}")
    ui_close()
    return 0 if not not_found else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(130)
