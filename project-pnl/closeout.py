#!/usr/bin/env python3
"""
closeout.py — the FINAL closeout report: one permanent page per finished job.

The owner asked for "a final closeout report that says FINAL, to have as
permanent to the job folder, to review jobs based on type, client etc"
(2026-09-02). Two deliverables:

  <job folder>/<JOB> FINAL Closeout.xlsx   one page, stamped, per job
  <division>/Closeout Index.xlsx           every FINAL issued, one row each

FINAL MEANS IMMUTABLE, and that is the whole point of the file.
A live P&L is regenerated whenever someone asks and its numbers move with QBO;
that is right for a job in flight and wrong for a record. A closeout is the
number the job ENDED on, so this refuses to overwrite a FINAL that already
exists. `--reissue` overrides, and then the sheet says so on its face - it
carries the original issue date and the reason, because a document labelled
FINAL whose figures quietly changed is worse than no document at all.

REVIEW BY TYPE AND CLIENT is what the index is for: division, client/GC, and
type (RP Tract/Custom from the WIP master; slab vs -FTW from the project #) sit
beside the final figures so the whole book sorts and filters on them.

Reads the generated `Project_PnL_<job>.xlsx` - the same reader the overview
uses, whose numbers are proven line-level against QBO. No QBO pull, no
credentials, runs in a second.

USAGE
  python3 project-pnl/closeout.py MFD133 MFD160
  python3 project-pnl/closeout.py --division mfd --all-completed
  python3 project-pnl/closeout.py MFD133 --reissue "restated after bill recode"
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import sqlite3
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import pnl_paths                                    # noqa: E402
from shared.xlsx_verify import assert_clean                     # noqa: E402

import completed_pnl as C                                       # same tool folder

FINAL_NAME = "{job} FINAL Closeout.xlsx"
INDEX_NAME = "Closeout Index.xlsx"
LEDGER_DB = Path(os.environ.get(
    "ACB_LEDGER_DB",
    Path.home() / "Library" / "Application Support" / "Proficient" / "ledger.sqlite3"))

GOLD = "8A6D1F"          # the FINAL stamp - the one colour this sheet adds


def job_meta(job: str) -> dict:
    """Division, client/GC and type for the index. RP carries builder and
    Tract/Custom on the WIP master; CP and MFD carry neither, so the client
    comes from the P&L's own customer path instead."""
    out = {"division": "", "client": "", "type": "", "category": ""}
    ju = job.upper()
    out["division"] = ("MFD" if ju.startswith("MFD") else
                       "CP" if ju.startswith("CP") else
                       "RP" if ju.startswith("RP") else "")
    out["type"] = "Flatwork" if ju.endswith("-FTW") else ""
    try:
        con = sqlite3.connect(f"file:{LEDGER_DB}?mode=ro", uri=True)
    except sqlite3.Error:
        return out
    try:
        row = con.execute("SELECT builder_or_gc, type, rp_category FROM project "
                          "WHERE project_no = ?", (ju,)).fetchone()
        if row:
            out["client"] = row[0] or ""
            out["type"] = row[1] or out["type"]
            out["category"] = row[2] or ""
    except sqlite3.Error:
        pass
    finally:
        con.close()
    return out


def _client_from_title(title: str) -> str:
    """The GC out of `PROJECT P&L — Parent:Project`, for the divisions the WIP
    master does not carry a builder for."""
    fqn = (title or "").replace("PROJECT P&L — ", "").strip()
    return fqn.split(":")[0].strip() if ":" in fqn else fqn


def _span(src: dict) -> tuple:
    """(first activity, last activity) across every cost line and invoice."""
    ds = [i["date"] for i in src["invoices"] if i["date"]]
    ds += [ln["date"] for s in src["sections"] for a in s["accounts"]
           for v in a["vendors"] for ln in v["lines"] if ln["date"]]
    ds = [d for d in ds if d]
    return (min(ds), max(ds)) if ds else (None, None)


def build_final(job: str, src: dict, t: dict, out: Path, meta: dict,
                reissue: str = "", first_issued: str = "") -> None:
    """One page. Identity, the result, where the money went, what was billed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Closeout"
    ws.sheet_view.showGridLines = False
    today = dt.datetime.now()
    first, last = _span(src)

    def t_(r, c, v, **kw):
        return C._t(ws, r, c, v, **kw)

    # ── the stamp ──
    t_(1, 1, "FINAL", size=C.SZ_TITLE + 6, bold=True, color=GOLD)
    t_(1, C.C0 + 1, C.job_label(job, src.get("title", "")),
       size=C.SZ_TITLE, bold=True, color=C.NAVY)
    ws.merge_cells(start_row=1, start_column=C.C0 + 1, end_row=1, end_column=C.C0 + 5)
    ws.row_dimensions[1].height = 34
    t_(2, C.C0 + 1, f"Closeout report · issued {today:%m/%d/%Y}"
                    + (f" · REISSUED, supersedes the {first_issued} issue"
                       if reissue else ""),
       size=C.SZ_SMALL, color=C.GREY)
    if reissue:
        t_(3, C.C0 + 1, f"reason: {reissue}", size=C.SZ_SMALL, color=C.RED)
    for c in range(1, C.C0 + 7):
        ws.cell(row=3 if not reissue else 4, column=c).border = Border(
            bottom=Side(style="medium", color=GOLD))

    r = (5 if reissue else 4) + 1

    # ── identity: what this job WAS ──
    t_(r, C.C0, "THE JOB", size=C.SZ + 1, bold=True, color=C.NAVY)
    r += 1
    ident = [("Project #", job),
             ("Division", meta.get("division", "")),
             ("Client / GC", meta.get("client") or _client_from_title(src.get("title", ""))),
             ("Type", meta.get("type") or "Slab"),
             ("QBO customer", (src.get("title", "")
                               .replace("PROJECT P&L — ", "").strip())),
             ("First activity", first.strftime("%m/%d/%Y") if first else "-"),
             ("Last activity", last.strftime("%m/%d/%Y") if last else "-"),
             ("Invoices", f"{len(src['invoices'])}"),
             ("Cost lines", f"{sum(len(v['lines']) for s in src['sections'] for a in s['accounts'] for v in a['vendors'])}")]
    for label, val in ident:
        t_(r, C.C0, label, size=C.SZ_SMALL, color=C.GREY)
        t_(r, C.C0 + 2, val, size=C.SZ_SMALL, color=C.INK)
        ws.merge_cells(start_row=r, start_column=C.C0 + 2, end_row=r, end_column=C.C0 + 5)
        r += 1
    r += 1

    # ── the result ──
    t_(r, C.C0, "THE RESULT", size=C.SZ + 1, bold=True, color=C.NAVY)
    r += 1
    res = [("Billed (incl. retainage)", t["billed"], C.MONEY, C.INK),
           ("Job cost", t["cost"], C.MONEY, C.INK),
           ("Gross profit", t["gp"], C.MONEY, C.GREEN if t["gp"] >= 0 else C.RED),
           ("Gross margin", t["gpm"], C.PCT, C.GREEN if t["gp"] >= 0 else C.RED),
           (f"less overhead ({C.OVERHEAD_PCT:.0%} of revenue)", -t["oh"], C.MONEY, C.GREY),
           ("NET PROFIT", t["net"], C.MONEY, C.GREEN if t["net"] >= 0 else C.RED),
           ("Net margin", t["netm"], C.PCT, C.GREEN if t["net"] >= 0 else C.RED)]
    if meta.get("division") == "MFD":
        res += [(f"MFD view: overhead ({C.MFD_OVERHEAD_PCT:.0%} of cost)", -t["moh"], C.MONEY, C.GREY),
                ("MFD view: net profit", t["mnet"], C.MONEY,
                 C.GREEN if t["mnet"] >= 0 else C.RED)]
    hero = r + 5
    for label, val, fmt, col in res:
        big = label.startswith("NET PROFIT")
        t_(r, C.C0, label, size=C.SZ + (2 if big else 0), bold=big,
           color=C.INK if not big else C.NAVY)
        t_(r, C.C0 + 3, val, size=C.SZ + (4 if big else 0), bold=True, fmt=fmt,
           color=col, align="right")
        ws.merge_cells(start_row=r, start_column=C.C0, end_row=r, end_column=C.C0 + 2)
        ws.row_dimensions[r].height = 26 if big else 20
        r += 1
    C._thick_box(ws, hero, hero, C.C0, C.C0 + 3)
    r += 1

    # ── where the money went ──
    t_(r, C.C0, "WHERE THE MONEY WENT", size=C.SZ + 1, bold=True, color=C.NAVY)
    t_(r, C.C0 + 3, t["cost"], size=C.SZ + 1, bold=True, color=C.NAVY,
       fmt=C.MONEY, align="right")
    r += 1
    for c, h in ((C.C0, "Account"), (C.C0 + 3, "Amount"), (C.C0 + 4, "% of cost")):
        t_(r, c, h, size=C.SZ_SMALL - 1, bold=True, color="FFFFFF", fill=C.F_HDR,
           align="right" if c != C.C0 else "left", indent=1 if c == C.C0 else 0)
    for c in range(C.C0, C.C0 + 5):
        ws.cell(row=r, column=c).fill = C.F_HDR
    r += 1
    for sec in src["sections"]:
        # Name the section. Without it the list looks like one ranking that
        # inexplicably restarts - which on MFD133 pushed its LARGEST account
        # (a deleted expense account carrying 56% of the job) far down the page.
        if len(src["sections"]) > 1:
            t_(r, C.C0, sec["name"], size=C.SZ_SMALL, bold=True, color=C.NAVY)
            t_(r, C.C0 + 3, sec["total"], size=C.SZ_SMALL, bold=True,
               color=C.NAVY, fmt=C.MONEY, align="right")
            for c in range(C.C0, C.C0 + 5):
                ws.cell(row=r, column=c).border = Border(
                    bottom=Side(style="thin", color=C.NAVY))
            r += 1
        for acct in sorted(sec["accounts"], key=lambda a: -a["total"]):
            t_(r, C.C0, acct["name"], size=C.SZ_SMALL, color=C.INK)
            ws.merge_cells(start_row=r, start_column=C.C0, end_row=r, end_column=C.C0 + 2)
            t_(r, C.C0 + 3, acct["total"], size=C.SZ_SMALL, fmt=C.MONEY, align="right")
            t_(r, C.C0 + 4, (acct["total"] / t["cost"]) if t["cost"] else 0,
               size=C.SZ_SMALL, fmt=C.PCT, color=C.GREY, align="right")
            if r % 2 == 0:
                for c in range(C.C0, C.C0 + 5):
                    ws.cell(row=r, column=c).fill = C.F_BAND
            r += 1
    if src.get("payroll"):
        t_(r, C.C0, f"excludes payroll of {src['payroll']:,.0f} - carried in the "
                    f"overhead %, not charged to the job",
           size=C.SZ_SMALL - 1, color=C.GREY)
        r += 1
    r += 1

    # ── what was billed ──
    t_(r, C.C0, "WHAT WAS BILLED", size=C.SZ + 1, bold=True, color=C.NAVY)
    t_(r, C.C0 + 3, t["billed"], size=C.SZ + 1, bold=True, color=C.NAVY,
       fmt=C.MONEY, align="right")
    r += 1
    for c, h in ((C.C0, "Invoice"), (C.C0 + 1, "Date"), (C.C0 + 2, "What for"),
                 (C.C0 + 3, "Amount"), (C.C0 + 4, "Paid?")):
        t_(r, c, h, size=C.SZ_SMALL - 1, bold=True, color="FFFFFF", fill=C.F_HDR,
           align="right" if c == C.C0 + 3 else
           ("center" if c == C.C0 + 4 else "left"), indent=1 if c == C.C0 else 0)
    r += 1
    unpaid = 0.0
    for inv in sorted(src["invoices"], key=lambda i: str(i["date"])):
        cell = t_(r, C.C0, inv["doc"], size=C.SZ_SMALL, align="left")
        cell.number_format = "0"
        if inv["url"]:
            cell.hyperlink = inv["url"]
            cell.font = Font(size=C.SZ_SMALL, color=C.LINK, underline="single")
        dc = t_(r, C.C0 + 1, inv["date"], size=C.SZ_SMALL, align="left")
        dc.number_format = "mm/dd/yyyy"
        t_(r, C.C0 + 2, str(inv["memo"])[:70], size=C.SZ_SMALL)
        t_(r, C.C0 + 3, inv["gross"] + inv["ret_billed"], size=C.SZ_SMALL,
           fmt=C.MONEY, align="right")
        paid = str(inv["paid"])
        if not paid.startswith("PAID"):
            unpaid += inv["gross"] + inv["ret_billed"]
        t_(r, C.C0 + 4, paid, size=C.SZ_SMALL, align="center",
           color=C.GREY if paid.startswith("PAID") else C.RED)
        if r % 2 == 0:
            for c in range(C.C0, C.C0 + 5):
                ws.cell(row=r, column=c).fill = C.F_BAND
        r += 1
    if src["not_billed"]:
        t_(r, C.C0, "(journal entry)", size=C.SZ_SMALL, color=C.GREY)
        t_(r, C.C0 + 2, "retainage moved by journal entry", size=C.SZ_SMALL, color=C.GREY)
        t_(r, C.C0 + 3, src["not_billed"], size=C.SZ_SMALL, fmt=C.MONEY,
           align="right", color=C.GREY)
        r += 1
    r += 1
    if unpaid > 0.5:
        t_(r, C.C0, f"⚠ {unpaid:,.0f} still OPEN on this job at closeout",
           size=C.SZ, bold=True, color=C.RED)
        r += 2

    # ── the provenance, so the number can be defended later ──
    t_(r, C.C0, "Figures are FINAL as at the issue date above, taken line-level "
                "from QuickBooks Online.", size=C.SZ_SMALL - 1, color=C.GREY)
    r += 1
    t_(r, C.C0, "This page is a record, not a live report - it is not "
                "regenerated when QBO changes.", size=C.SZ_SMALL - 1, color=C.GREY)

    ws.column_dimensions["A"].width = C.GUTTER_W
    for col, w in zip("BCDEFG", (30, 14, 30, 18, 16, 14)):
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    tmp.replace(out)


def _issued_on(path: Path) -> str:
    """The issue date already inside a FINAL, so a reissue can name it."""
    try:
        wb = load_workbook(str(path))
        v = str(wb["Closeout"].cell(2, C.C0 + 1).value or "")
        wb.close()
        for tok in v.split():
            if tok.count("/") == 2:
                return tok
    except Exception:
        pass
    return "an earlier"


def build_index(rows: List[dict], out: Path) -> None:
    """Every FINAL issued, sortable by division, client and type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Closeouts"
    ws.sheet_view.showGridLines = False
    C._t(ws, 1, 1, "CLOSEOUT INDEX", size=C.SZ_TITLE, bold=True, color=C.NAVY)
    C._t(ws, 2, C.C0, f"{len(rows)} job(s) closed out · sort or filter on "
                      f"division, client or type", size=C.SZ_SMALL, color=C.GREY)
    C._t(ws, 1, C.C0 + 9, f"Generated {dt.datetime.now():%m/%d/%Y %I:%M %p}",
         size=C.SZ_SMALL, color=C.GREY, align="right")
    heads = [("Job", 30), ("Division", 11), ("Client / GC", 26), ("Type", 12),
             ("Closed", 12), ("Billed", 16), ("Cost", 16), ("Gross profit", 16),
             ("GP %", 9), ("Net", 16)]
    r = 4
    for i, (h, _w) in enumerate(heads):
        C._t(ws, r, C.C0 + i, h, size=C.SZ_SMALL - 1, bold=True, color="FFFFFF",
             fill=C.F_HDR, align="left" if i < 5 else "right",
             indent=1 if i == 0 else 0)
    ws.row_dimensions[r].height = 22
    r += 1
    for d in sorted(rows, key=lambda x: (x["division"], x["client"], x["job"])):
        vals = [(d["label"], None), (d["division"], None), (d["client"], None),
                (d["type"], None), (d["closed"], "mm/dd/yyyy"),
                (d["billed"], C.MONEY), (d["cost"], C.MONEY),
                (d["gp"], C.MONEY), (d["gpm"], C.PCT), (d["net"], C.MONEY)]
        for i, (v, fmt) in enumerate(vals):
            col = C.GREEN if (i in (7, 8, 9) and isinstance(v, (int, float)) and v >= 0) else \
                  C.RED if (i in (7, 8, 9) and isinstance(v, (int, float))) else C.INK
            cell = C._t(ws, r, C.C0 + i, v, size=C.SZ_SMALL, fmt=fmt,
                        align="left" if i < 5 else "right", color=col,
                        bold=(i == 7))
            if i == 0 and d.get("rel"):
                cell.hyperlink = d["rel"]
                cell.font = Font(size=C.SZ_SMALL, color=C.LINK, underline="single")
        if r % 2 == 0:
            for i in range(len(heads)):
                ws.cell(row=r, column=C.C0 + i).fill = C.F_BAND
        r += 1
    ws.auto_filter.ref = (f"{get_column_letter(C.C0)}4:"
                          f"{get_column_letter(C.C0 + len(heads) - 1)}{r - 1}")
    ws.freeze_panes = f"A{5}"
    ws.column_dimensions["A"].width = C.GUTTER_W
    for i, (_h, w) in enumerate(heads):
        ws.column_dimensions[get_column_letter(C.C0 + i)].width = w
    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    tmp.replace(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="FINAL closeout report per job")
    ap.add_argument("jobs", nargs="*", help="e.g. MFD133 MFD160")
    ap.add_argument("--division", choices=sorted(C.DIVISIONS), default="mfd")
    ap.add_argument("--all-completed", action="store_true",
                    help="every job filed under the division's archive folder")
    ap.add_argument("--reissue", default="",
                    help="overwrite an existing FINAL, stating why (it is "
                         "recorded on the sheet)")
    ap.add_argument("--index-only", action="store_true",
                    help="rebuild the index from the FINALs already on disk")
    a = ap.parse_args()

    div = C.DIVISIONS[a.division]
    div_dir = pnl_paths.division_dir(div["prefix"])
    found = C._iter_jobs(div_dir, div["prefix"])
    if a.jobs:
        want = {j.upper() for j in a.jobs}
        found = [f for f in found if f[0].upper() in want]
    elif a.all_completed:
        found = [f for f in found if f[2] == "Completed"]
    elif not a.index_only:
        print("✗  name a job, or pass --all-completed")
        return 1

    rows: List[dict] = []
    for job, src_path, _status in found:
        folder = src_path.parent
        final = folder / FINAL_NAME.format(job=job)
        src = C.read_source(src_path)
        if not src["sections"]:
            print(f"  ⚠ {job}: no cost detail in {src_path.name}")
            continue
        t = C._totals(src)
        meta = job_meta(job)
        meta["client"] = meta["client"] or _client_from_title(src.get("title", ""))
        _f, last = _span(src)
        if final.exists() and not (a.reissue or a.index_only):
            print(f"  · {job}: FINAL already issued {_issued_on(final)} - left "
                  f"alone (use --reissue \"<reason>\" to supersede it)")
        elif not a.index_only:
            build_final(job, src, t, final, meta,
                        reissue=a.reissue,
                        first_issued=_issued_on(final) if final.exists() else "")
            print(f"  ✓ {job}  →  {final.name}"
                  + ("  (REISSUED)" if a.reissue and final.exists() else ""))
        if final.exists():
            rows.append({
                "job": job, "label": C.job_label(job, src.get("title", "")),
                "division": meta["division"], "client": meta["client"],
                "type": meta["type"] or "Slab", "closed": last,
                "billed": t["billed"], "cost": t["cost"], "gp": t["gp"],
                "gpm": t["gpm"], "net": t["net"],
                "rel": os.path.relpath(final, div_dir).replace(os.sep, "/")})
    if rows:
        idx = div_dir / INDEX_NAME
        build_index(rows, idx)
        print(f"\n  index: {len(rows)} closeout(s)  →  {idx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
