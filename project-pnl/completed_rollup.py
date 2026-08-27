#!/usr/bin/env python3
"""
completed_rollup.py — one P&L over every COMPLETED job in a folder.

The per-project workbooks answer "how did THIS job do". This answers "how did
the finished work do, together" — one row per job, a portfolio total, and a
link on each row that opens that job's own workbook.

READS THE WORKBOOKS, NOT QBO. Every figure is lifted from each project P&L's
Transactions sheet (the same raw invoice rows and cost lines the P&L itself
sums), so the rollup can never disagree with the file it links to, and it runs
offline in seconds with no credential unlock.

Why not read the P&L sheet's totals? They are live Excel FORMULAS - openpyxl
reads the formula text, not a value, and `data_only=True` returns None unless
Excel has opened and cached the file. The Transactions sheet holds real
numbers, so it is the honest source.

Links are STORED RELATIVE targets (`MFD133/Project_PnL_MFD133.xlsx`), the same
mechanism the ledgers use for bill scans - proven on Mac and Windows, and it
survives an Excel save. It needs the rollup to sit BESIDE the job folders.

USAGE
  python3 project-pnl/completed_rollup.py
  python3 project-pnl/completed_rollup.py --folder "<path>" --out "<file.xlsx>"
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import pnl_paths
from shared.xlsx_verify import assert_clean

DEFAULT_SUBDIR = "completed mfd project p&l"
OUT_NAME = "Completed MFD P&L.xlsx"

SZ = 12
NAVY = "1F3A5F"
LINK = "0563C1"
GREEN = "008000"
RED = "C00000"
ACC = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
PCT = "0.00%"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
TOT_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Border(top=Side(style="thin", color=NAVY))


def read_project_pnl(path: Path) -> Optional[dict]:
    """Billed / cost / contract / ETC out of one project P&L workbook."""
    try:
        wb = load_workbook(str(path), data_only=False)
    except Exception as e:
        print(f"    ⚠ {path.name}: unreadable ({e})")
        return None
    try:
        if "Transactions" not in wb.sheetnames:
            return None
        ws = wb["Transactions"]
        hdr = next((r for r in range(1, 60)
                    if str(ws.cell(r, 1).value or "").strip() == "Inv #"), None)
        if hdr is None:
            return None
        end = next((r for r in range(hdr + 1, ws.max_row + 1)
                    if str(ws.cell(r, 1).value or "").startswith("TOTAL")), None)
        if end is None:
            return None

        def num(r, c):
            v = ws.cell(r, c).value
            return float(v) if isinstance(v, (int, float)) else 0.0

        gross = sum(num(r, 4) for r in range(hdr + 1, end))
        ret_billed = sum(num(r, 7) for r in range(hdr + 1, end))
        # A row below the income total with an ACCOUNT in col D is a cost line;
        # without one it is the retainage-not-billed block, which is income.
        cost = sum(num(r, 5) for r in range(end + 1, ws.max_row + 1)
                   if ws.cell(r, 4).value)
        not_billed = sum(num(r, 5) for r in range(end + 1, ws.max_row + 1)
                         if not ws.cell(r, 4).value)
        contract = etc = None
        if "P&L" in wb.sheetnames:
            pl = wb["P&L"]
            for r in range(1, 40):
                lbl = str(pl.cell(r, 1).value or "").strip().lower()
                v = pl.cell(r, 2).value
                if not isinstance(v, (int, float)):
                    continue
                if lbl.startswith("original contract price"):
                    contract = float(v)
                elif lbl.startswith("original etc"):
                    etc = float(v)
        return {"billed": round(gross + ret_billed + not_billed, 2),
                "cost": round(cost, 2),
                "contract": contract, "etc": etc,
                "invoices": end - hdr - 1}
    finally:
        wb.close()


def collect(folder: Path) -> List[dict]:
    out = []
    for sub in sorted(folder.iterdir()):
        if not sub.is_dir():
            continue
        f = sub / f"Project_PnL_{sub.name}.xlsx"
        if not f.exists():
            hits = list(sub.glob("Project_PnL_*.xlsx"))
            if not hits:
                continue
            f = hits[0]
        d = read_project_pnl(f)
        if not d:
            continue
        d["job"] = sub.name
        d["rel"] = f"{sub.name}/{f.name}"
        out.append(d)
    return out


def build(rows: List[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed MFD"
    ws.sheet_view.showGridLines = False

    t = ws.cell(row=1, column=1, value="COMPLETED MFD PROJECTS — COMBINED P&L")
    t.font = Font(bold=True, size=SZ + 4, color=NAVY)
    s = ws.cell(row=2, column=1,
                value=(f"{len(rows)} finished job(s) · every figure read from that "
                       f"job's own workbook · click OPEN to jump to it"))
    s.font = Font(italic=True, size=SZ - 2, color="595959")

    heads = ["JOB", "CONTRACT", "ETC", "BILLED", "COST",
             "GROSS PROFIT", "GP %", "vs ETC", "OPEN"]
    hr = 4
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(bold=True, size=SZ, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    r = hr + 1
    first = r
    for d in sorted(rows, key=lambda x: -x["billed"]):
        ws.cell(row=r, column=1, value=d["job"]).font = Font(bold=True, size=SZ)
        for c, v in ((2, d["contract"]), (3, d["etc"]),
                     (4, d["billed"]), (5, d["cost"])):
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = ACC
            cell.font = Font(size=SZ)
        gp = ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        gp.number_format = ACC
        gp.font = Font(bold=True, size=SZ)
        pc = ws.cell(row=r, column=7, value=f'=IF(D{r}=0,"",F{r}/D{r})')
        pc.number_format = PCT
        pc.font = Font(bold=True, size=SZ)
        # Cost against the budget the job was sold on. Blank when no ETC was
        # ever entered — a zero there would read as "on budget".
        ov = ws.cell(row=r, column=8, value=f'=IF(C{r}=0,"",E{r}-C{r})')
        ov.number_format = ACC
        ov.font = Font(size=SZ)
        lk = ws.cell(row=r, column=9, value="OPEN  ↗")
        lk.hyperlink = d["rel"]
        lk.font = Font(size=SZ, color=LINK, underline="single", bold=True)
        lk.alignment = Alignment(horizontal="center")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    last = r - 1

    ws.cell(row=r, column=1, value=f"TOTAL — {len(rows)} JOBS").font = Font(bold=True, size=SZ)
    for c in (2, 3, 4, 5):
        L = get_column_letter(c)
        cell = ws.cell(row=r, column=c, value=f"=SUM({L}{first}:{L}{last})")
        cell.number_format = ACC
        cell.font = Font(bold=True, size=SZ)
    for c, f in ((6, f"=D{r}-E{r}"), (7, f'=IF(D{r}=0,"",F{r}/D{r})'),
                 (8, f'=IF(C{r}=0,"",E{r}-C{r})')):
        cell = ws.cell(row=r, column=c, value=f)
        cell.number_format = PCT if c == 7 else ACC
        cell.font = Font(bold=True, size=SZ)
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = TOT_FILL
        ws.cell(row=r, column=c).border = TOP

    # red = lost money / over the budget it was sold on
    from openpyxl.formatting.rule import CellIsRule
    for col in ("F", "G", "H"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{r}",
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(bold=True, color=RED)))
    ws.conditional_formatting.add(
        f"H{first}:H{r}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(bold=True, color=RED)))

    for col, w in zip("ABCDEFGHI", (14, 17, 17, 17, 17, 17, 10, 17, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hr + 1}"

    tmp = out_path.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)                    # never publish a file Excel would repair
    tmp.replace(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Combined P&L over completed jobs")
    ap.add_argument("--folder", default=None,
                    help=f"folder holding the job subfolders "
                         f"(default: <PROJECT P&Ls>/{DEFAULT_SUBDIR})")
    ap.add_argument("--out", default=None, help=f"output file (default: {OUT_NAME})")
    a = ap.parse_args()

    folder = (Path(a.folder).expanduser() if a.folder
              else pnl_paths.pnl_out_dir() / DEFAULT_SUBDIR)
    if not folder.is_dir():
        print(f"✗  no such folder: {folder}")
        return 1
    rows = collect(folder)
    if not rows:
        print(f"✗  no project P&L workbooks under {folder}")
        return 1
    out = Path(a.out).expanduser() if a.out else folder / OUT_NAME
    build(rows, out)
    tb = sum(d["billed"] for d in rows)
    tc = sum(d["cost"] for d in rows)
    print(f"  {len(rows)} job(s)   billed ${tb:,.2f}   cost ${tc:,.2f}   "
          f"GP ${tb - tc:,.2f}  ({(tb - tc) / tb * 100 if tb else 0:.2f}%)")
    print(f"  → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
