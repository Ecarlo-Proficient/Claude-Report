#!/usr/bin/env python3
"""
completed_pnl.py — the division OVERVIEW, assembled from the P&L workbooks.

ONE workbook per division: `<DIV> Overview.xlsx`, an overview sheet with a row
per job over a sheet per job (its P&L and grouped transactions). Every link is
internal, so it survives being emailed. Live and finished jobs sit in the same
book, split into their own sections.

READS THE GENERATED WORKBOOKS, NOT QBO. It re-shapes each
`Project_PnL_<job>.xlsx`, whose numbers have already been proven line-level
against QBO by `one-offs/pnl_line_level_audit.py`. So it cannot introduce an
attribution bug, needs no credentials, and runs in seconds — which is why
`project_pnl_export.py` rebuilds the Overview at the END of every run rather
than leaving it to go stale (the user 2026-09-03).

RETIRED 2026-09-03 (the user: "why is there a job result excel? shouldn't this
be merged with the P&L? i feel that we are confused and all over the place").
The per-job `<JOB> Job Result.xlsx` is gone. It re-derived, in a second shape
and a second file, what `project_pnl_export.py --simple` already produces: the
stripped-back P&L for a finished job, with no draw sheets or coverage blocks.
One job = one P&L. Sibling retirements: `closeout.py` (FINAL Closeout + Closeout
Index) and `completed_rollup.py` (never once run).

USAGE
  python3 project-pnl/completed_pnl.py --division mfd
  python3 project-pnl/completed_pnl.py --division rp --year 2026
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import pnl_paths
from shared.xlsx_verify import assert_clean

# One overview per DIVISION, written INTO that division's folder - the folder
# whose OneDrive link the owner shares with the PM who runs it
# (shared/pnl_paths.division_dir). CP and RP default to the current year only:
# "need one for CP and RP just for this year projects, don't go further back"
# (the user 2026-08-31). MFD is small enough to carry every job.
DIVISIONS = {
    "mfd": {"prefix": "MFD", "label": "MFD", "title": "MFD OVERVIEW — ALL JOBS",
            "default_year": None,
            "alt": {"label": "MFD — 9% of cost", "short": "NET  (MFD 9%)"}},
    "cp": {"prefix": "CP", "label": "CP", "title": "COMMERCIAL OVERVIEW",
           "default_year": dt.date.today().year, "alt": None},
    "rp": {"prefix": "RP", "label": "RP", "title": "RESIDENTIAL OVERVIEW",
           "default_year": dt.date.today().year, "alt": None},
}
# Payroll is carried in the overhead %, so charging it to the job as well
# double-counts it (the user 2026-08-27). Excluded from job cost and reported
# on the sheet as a named exclusion — never silently dropped.
PAYROLL_RE = re.compile(r"payroll|wages|salar|employee benefit|workers.?comp", re.I)
OVERHEAD_PCT = 0.10          # company view: 10% of REVENUE
MFD_OVERHEAD_PCT = 0.09      # MFD's own view: 9% of COSTS (CLAUDE.md)

# ── one accent, lots of white space ────────────────────────────────────
# The first cut used solid navy on the tile headers AND the section headers AND
# the table headers, borders on every cell, row banding, and red on every
# negative — four treatments competing on one page, which is what read as
# amateurish (the user 2026-08-27). Now: navy anchors ONE band per table, rules
# do the separating instead of boxes, and red is reserved for profit lines.
SZ = 14                      # body
SZ_SMALL = 13                # detail rows — was 12 and read too small
SZ_TITLE = 20
NAVY = "1F3A5F"
INK = "1F2937"               # near-black body text, softer than pure black
GREEN = "1E6B3A"
RED = "B00020"
LINK = "0563C1"
GREY = "6B7280"
MONEY = '#,##0;[Red](#,##0)'          # birds-eye: no cents
MONEY_C = '#,##0.00;[Red](#,##0.00)'  # detail: cents
PCT = "0.0%"

# Column A is a narrow GUTTER on the overview sheets and content starts in B
# (the user 2026-08-31: "goal: have ability to move info away from left side").
# The big title is the one thing left in A, hanging into the gutter.
GUTTER_W = 3
C0 = 2                       # first content column

F_HDR = PatternFill("solid", fgColor=NAVY)      # the ONE navy band per table
F_BAND = PatternFill("solid", fgColor="F4F6F9")  # barely-there row banding
F_KPI = PatternFill("none")                      # tiles sit on white
HAIR = Side(style="thin", color="E3E8EF")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
RULE = Side(style="thin", color=NAVY)
UNDER = Border(bottom=Side(style="medium", color=NAVY))
THICK = Side(style="medium", color=NAVY)


def _thick_box(ws, r0: int, r1: int, c0: int, c1: int) -> None:
    """Heavy rule around a block so it reads as its own thing."""
    for rr in range(r0, r1 + 1):
        for cc in range(c0, c1 + 1):
            cur = ws.cell(row=rr, column=cc).border
            ws.cell(row=rr, column=cc).border = Border(
                left=THICK if cc == c0 else cur.left,
                right=THICK if cc == c1 else cur.right,
                top=THICK if rr == r0 else cur.top,
                bottom=THICK if rr == r1 else cur.bottom)


# ───────────────────────────── read the source ─────────────────────────────

# A vendor row in the Transactions cost block carries its line count:
# "Estrada Ready Mix  (28)". That suffix is what separates it from a category
# banner ("CONCRETE", "LABOR"), which carries no amount and no count.
_VENDOR_RE = re.compile(r"^(.*?)\s*\((\d+)\)\s*$")


def _first_col(ws) -> int:
    """Where the Transactions sheet's labels actually start.

    P&Ls generated from 2026-08-31 carry a LEFT GUTTER: column A is empty and
    every label sits in B. Reading column A regardless is not a wrong number,
    it is a SILENT ZERO - the job reports 0 billed / 0 cost and drops out of
    the overview totals looking like a job with no activity. So find the
    'Inv #' header and take its column; pre-gutter workbooks answer 1."""
    for r in range(1, 60):
        for c in range(1, 5):
            if str(ws.cell(r, c).value or "").strip() == "Inv #":
                return c
    return 1


def _header_map(ws, row: int, c0: int) -> Dict[str, int]:
    """{header text -> column} for one header row, read from c0 rightwards."""
    return {str(ws.cell(row, c).value or "").strip().lower(): c
            for c in range(c0, c0 + 10)}


def _read_invoices(ws, c0: int) -> tuple:
    """Every invoice, and the retainage that moved by journal entry."""
    inv: List[dict] = []
    hdr = next((r for r in range(1, 60)
                if str(ws.cell(r, c0).value or "").strip() == "Inv #"), None)
    if hdr is None:
        return inv, 0.0
    col = _header_map(ws, hdr, c0)
    DOC = col.get("inv #", c0)
    DATE = col.get("date", c0 + 1)
    MEMO = col.get("memo", c0 + 2)
    GROSS = col.get("gross income", c0 + 3)
    WHELD = col.get("retainage withheld", c0 + 4)
    RETB = col.get("retainage billed", c0 + 6)
    PAID = col.get("paid?", c0 + 7)
    end = next((r for r in range(hdr + 1, ws.max_row + 1)
                if str(ws.cell(r, c0).value or "").startswith("TOTAL")), ws.max_row)

    def f(rr, cc):
        v = ws.cell(rr, cc).value
        return float(v) if isinstance(v, (int, float)) else 0.0

    for r in range(hdr + 1, end):
        inv.append({"doc": ws.cell(r, DOC).value, "date": ws.cell(r, DATE).value,
                    "memo": ws.cell(r, MEMO).value or "",
                    "gross": f(r, GROSS), "withheld": f(r, WHELD),
                    "ret_billed": f(r, RETB),
                    "paid": ws.cell(r, PAID).value or "",
                    "url": (ws.cell(r, DOC).hyperlink.target
                            if ws.cell(r, DOC).hyperlink else None)})
    # Rows between the income total and the first cost block, with no account:
    # retainage the bookkeeper moved by journal entry instead of invoicing.
    stop = next((r for r in range(end + 1, ws.max_row + 1)
                 if str(ws.cell(r, c0).value or "").strip()
                 .startswith(("COGS", "EXPENSES"))), ws.max_row + 1)
    not_billed = sum(f(r, WHELD) for r in range(end + 1, stop)
                     if not ws.cell(r, GROSS).value)
    return inv, not_billed


def _read_costs(ws, c0: int) -> List[dict]:
    """COST OF GOODS SOLD / OPERATING EXPENSES → account → vendor → line.

    Read from the **Transactions** sheet, which every template writes the same
    way — MFD, CP and RP alike. The old reader used the `By Account` sheet,
    which only the current CP/MFD template has: CP672 and every RP workbook
    have no such sheet, so a division overview could never have been built from
    it. Verified equal to the cent against `By Account` on all 14 MFD jobs
    before the switch (2026-08-31).

    The block is written vendor-major; the account each line belongs to rides
    on the line itself, so the account tree is regrouped here.
    """
    sections: List[dict] = []
    cur: Optional[dict] = None
    cols: Dict[str, int] = {}
    vendor = ""
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, c0).value or "").strip()
        lvl = ws.row_dimensions[r].outline_level or 0
        if label.startswith("COGS"):
            cur = {"name": "COST OF GOODS SOLD", "rows": []}
            sections.append(cur); cols, vendor = {}, ""
            continue
        if label.startswith("EXPENSES"):
            cur = {"name": "OPERATING EXPENSES (non-COGS)", "rows": []}
            sections.append(cur); cols, vendor = {}, ""
            continue
        if label.startswith("INCOME"):
            cur = None
            continue
        if cur is None:
            continue
        if label == "Ref #":
            cols = _header_map(ws, r, c0)
            continue
        if label.startswith("TOTAL"):
            cur = None
            continue
        if not cols:
            continue
        if lvl == 0:
            m = _VENDOR_RE.match(label)
            vendor = m.group(1).strip() if m else ""   # else: a category banner
            continue
        amt = ws.cell(r, cols.get("amount", c0 + 4)).value
        if not isinstance(amt, (int, float)):
            continue
        doc_c = cols.get("ref #", c0)
        desc_c = cols.get("memo", cols.get("description", c0 + 2))
        paid_c = cols.get("paid?")
        cur["rows"].append({
            "acct": str(ws.cell(r, cols.get("account", c0 + 3)).value or "(unclassified)"),
            "vendor": vendor or "(no vendor)",
            "date": ws.cell(r, cols.get("date", c0 + 1)).value,
            "doc": ws.cell(r, doc_c).value,
            "desc": ws.cell(r, desc_c).value or "",
            "amt": float(amt),
            "paid": (ws.cell(r, paid_c).value or "") if paid_c else "",
            "url": (ws.cell(r, doc_c).hyperlink.target
                    if ws.cell(r, doc_c).hyperlink else None)})

    out: List[dict] = []
    for sec in sections:
        accounts: List[dict] = []
        by_acct: Dict[str, dict] = {}
        for ln in sec["rows"]:
            a = by_acct.get(ln["acct"])
            if a is None:
                a = by_acct[ln["acct"]] = {"name": ln["acct"], "total": 0.0,
                                           "vendors": [], "_v": {}}
                accounts.append(a)
            v = a["_v"].get(ln["vendor"])
            if v is None:
                v = a["_v"][ln["vendor"]] = {"name": ln["vendor"], "total": 0.0,
                                             "lines": []}
                a["vendors"].append(v)
            v["lines"].append(ln)
            v["total"] = round(v["total"] + ln["amt"], 2)
            a["total"] = round(a["total"] + ln["amt"], 2)
        for a in accounts:
            a.pop("_v", None)
        accounts.sort(key=lambda a: -a["total"])
        for a in accounts:
            a["vendors"].sort(key=lambda v: -v["total"])
        if accounts:
            out.append({"name": sec["name"],
                        "total": round(sum(a["total"] for a in accounts), 2),
                        "accounts": accounts})
    return out


def read_source(path: Path) -> dict:
    """Invoices + account/vendor/line detail out of a generated project P&L.
    Template-agnostic: MFD, CP and RP workbooks all read the same way."""
    wb = load_workbook(str(path), data_only=False)
    try:
        ws = wb["Transactions"]
        c0 = _first_col(ws)          # 1 pre-gutter, 2 from 2026-08-31 on
        inv, not_billed = _read_invoices(ws, c0)
        sections = _read_costs(ws, c0)

        # Payroll out of job cost (it lives in overhead), recorded so the
        # sheet can say what it left out.
        payroll = 0.0
        for sec in sections:
            keep = []
            for acct in sec["accounts"]:
                if PAYROLL_RE.search(acct["name"]):
                    payroll += acct["total"]
                else:
                    keep.append(acct)
            sec["accounts"] = keep
            sec["total"] = round(sum(a["total"] for a in keep), 2)
        sections = [s for s in sections if s["accounts"]]

        # One row per BILL, not per line: a bill split over several lines of
        # the same account is one document and reads as one (the user
        # 2026-08-27). The line count rides along so nothing is hidden.
        for sec in sections:
            for acct in sec["accounts"]:
                for v in acct["vendors"]:
                    merged: Dict[str, dict] = {}
                    for ln in v["lines"]:
                        key = str(ln["doc"] or id(ln))
                        if key in merged:
                            m = merged[key]
                            m["amt"] += ln["amt"]
                            m["n"] += 1
                        else:
                            merged[key] = dict(ln, n=1)
                    for m in merged.values():
                        if m["n"] > 1:
                            m["desc"] = f"{m['n']} lines · {m['desc']}"
                    v["lines"] = sorted(merged.values(),
                                        key=lambda x: str(x["date"]), reverse=True)

        pl = next((n for n in ("P&L", "Job P&L") if n in wb.sheetnames), None)
        title = str(wb[pl].cell(1, 1).value or "") if pl else ""
        return {"invoices": inv, "not_billed": not_billed, "payroll": round(payroll, 2),
                "sections": sections, "title": title,
                "status": "Completed", "rel": None}
    finally:
        wb.close()


# ───────────────────────────── write the report ────────────────────────────

def _t(ws, r, c, v, *, size=SZ, bold=False, color="000000", fmt=None,
       fill=None, align=None, wrap=False, border=None, indent=0):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(size=size, bold=bold, color=color)
    if fmt:
        cell.number_format = fmt
    if fill is not None:
        cell.fill = fill
    if align or wrap or indent:
        cell.alignment = Alignment(horizontal=align, wrap_text=wrap,
                                   vertical="center", indent=indent)
    if border is not None:
        cell.border = border
    return cell


def job_label(job: str, title: str) -> str:
    """'MFD133 — JLB Builders' — the number and the name in ONE cell, never a
    separate column (the user 2026-08-27).

    The name comes from the QBO customer path `<GC>:<project>`. Some projects
    carry the site in their own name ('MFD295 - Rock Creek Apartments'); most
    are bare ('MFD133'), so the GC is the useful name. Trailing LLC/INC is
    dropped and a SHOUTED name is title-cased, because a report reads better
    than a database does."""
    fqn = (title or "").replace("PROJECT P&L — ", "").strip()
    if not fqn:
        return job
    parts = [x.strip() for x in fqn.split(":") if x.strip()]
    leaf = parts[-1] if parts else ""
    name = re.sub(r"^\s*" + re.escape(job) + r"\s*[-–—:]*\s*", "", leaf,
                  flags=re.I).strip()
    if not name:
        # Some GCs nest a development under themselves
        # ("JPI Construction, LLC:SPCA:MFD160"), so take the TOP-level
        # customer — the GC — not the intermediate grouping, which is usually
        # an internal abbreviation nobody outside recognises.
        name = parts[0] if parts else ""
    # A GC's legal name often carries the development after a dash
    # ("Embrey Builders LLC-Champions Way DFW LP") — the entity is the part
    # that identifies the job on a report, so cut there.
    head = re.split(r"\s*[-–]\s*", name, 1)[0]
    if len(head) >= 6:
        name = head
    name = re.sub(r"\b(L\.?L\.?C\.?|INC\.?|LTD\.?|CORP\.?|L\.?P\.?)\b", "", name,
                  flags=re.I)
    name = re.sub(r"\s{2,}", " ", name).strip(" ,-")
    if name and name == name.upper():
        name = name.title()
    if len(name) > 26:                      # keep the cell one clean line
        cut = name[:26].rsplit(" ", 1)[0]
        name = cut if len(cut) >= 10 else name[:26]
    return f"{job} — {name}" if name else job


def _totals(src: dict) -> dict:
    billed = sum(i["gross"] + i["ret_billed"] for i in src["invoices"]) + src["not_billed"]
    cogs = next((x["total"] for x in src["sections"] if x["name"].startswith("COST")), 0.0)
    opex = next((x["total"] for x in src["sections"] if x["name"].startswith("OPERATING")), 0.0)
    cost = cogs + opex
    gp = billed - cost
    oh = billed * OVERHEAD_PCT
    moh = cost * MFD_OVERHEAD_PCT
    return {"billed": billed, "cogs": cogs, "opex": opex, "cost": cost,
            "gp": gp, "oh": oh, "net": gp - oh,
            "gpm": gp / billed if billed else 0.0,
            "netm": (gp - oh) / billed if billed else 0.0,
            "moh": moh, "mnet": gp - moh,
            "mnetm": (gp - moh) / billed if billed else 0.0}


def _tiles(ws, r: int, items, spans) -> None:
    """Metric tiles on MERGED spans of their own.

    They used to borrow the data table's column widths, so a label like
    "Company — 10% of revenue" was clipped by an 18-wide column while its
    value landed at the far edge of a 74-wide description column, nowhere
    near it. A tile owns its span, and the label and value are centred in it
    (the user 2026-08-27: "it just looks amateurish")."""
    for (label, val, fmt, color), (c0, c1) in zip(items, spans):
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
        ws.merge_cells(start_row=r + 1, start_column=c0, end_row=r + 1, end_column=c1)
        _t(ws, r, c0, label, size=SZ_SMALL - 1, bold=True, color=GREY,
           align="left")
        _t(ws, r + 1, c0, val, size=SZ + 9, bold=True, color=color, fmt=fmt,
           align="left")
        # a hairline under the label is all the separation a tile needs
        for cc in range(c0, c1 + 1):
            ws.cell(row=r, column=cc).border = Border(bottom=RULE)
    ws.row_dimensions[r].height = 18
    ws.row_dimensions[r + 1].height = 40


def _kpi_strip(ws, r: int, t: dict, spans, alt=None) -> int:
    """Metrics ACROSS, exactly 4 cells wide so the strip lines up with the
    table beneath it instead of running off to column N (the user 2026-08-27).
    The OVERHEAD views get their own boxed block below. MFD carries TWO - its
    own 9%-of-cost view alongside the company's 10%-of-revenue - and the MFD one
    must not get lost among the others; CP and RP have only the company view."""
    _tiles(ws, r, [
        ("BILLED", t["billed"], MONEY, NAVY),
        ("COST", t["cost"], MONEY, NAVY),
        ("GROSS PROFIT", t["gp"], MONEY, GREEN if t["gp"] >= 0 else RED),
        ("GROSS MARGIN", t["gpm"], PCT, GREEN if t["gp"] >= 0 else RED)], spans)

    r2 = r + 3
    lab, oh_c, net_c, pct_c = spans[0][0], spans[1], spans[2], spans[3]
    ws.merge_cells(start_row=r2, start_column=lab, end_row=r2, end_column=spans[0][1])
    _t(ws, r2, lab, "AFTER OVERHEAD", size=SZ_SMALL - 1, bold=True, color=GREY,
       indent=1)
    for span, txt in ((oh_c, "OVERHEAD"), (net_c, "NET PROFIT"), (pct_c, "NET MARGIN")):
        ws.merge_cells(start_row=r2, start_column=span[0], end_row=r2, end_column=span[1])
        _t(ws, r2, span[0], txt, size=SZ_SMALL - 1, bold=True, color=GREY,
           align="right")
    for cc in range(spans[0][0], spans[3][1] + 1):
        ws.cell(row=r2, column=cc).border = Border(bottom=HAIR)
    views = [(f"Company — {OVERHEAD_PCT:.0%} of revenue", t["oh"], t["net"], t["netm"])]
    if alt:
        views.append((alt["label"], t["moh"], t["mnet"], t["mnetm"]))
    for i, (lbl, oh, net, netm) in enumerate(views):
        rr = r2 + 1 + i
        ws.merge_cells(start_row=rr, start_column=lab, end_row=rr, end_column=spans[0][1])
        _t(ws, rr, lab, lbl, size=SZ, bold=(i == len(views) - 1 and len(views) > 1),
           color=INK, indent=1)
        for span, val, fmt, col in (
                (oh_c, -oh, MONEY, GREY),
                (net_c, net, MONEY, GREEN if net >= 0 else RED),
                (pct_c, netm, PCT, GREEN if net >= 0 else RED)):
            ws.merge_cells(start_row=rr, start_column=span[0], end_row=rr,
                           end_column=span[1])
            _t(ws, rr, span[0], val, size=SZ + 2, bold=True, fmt=fmt,
               color=col, align="right")
        ws.row_dimensions[rr].height = 26
    _thick_box(ws, r2, r2 + len(views), spans[0][0], spans[3][1])
    return r2 + len(views) + 2


def _spans(c0: int, last: int, n: int = 4):
    """Split the content width into `n` merged tile spans, so the metric strip
    is exactly as wide as the table under it whatever the division's column
    count is."""
    width = last - c0 + 1
    base, extra = divmod(width, n)
    out, c = [], c0
    for i in range(n):
        w = base + (1 if i < extra else 0)
        out.append((c, c + w - 1))
        c += w
    return out


def build_bundle(jobs: List[tuple], out: Path, div: dict) -> None:
    """<DIV> Overview: the overview sheet, then a sheet per job carrying its
    P&L and its transactions grouped account → vendor → line.

    EVERY link is INTERNAL. The per-job reports link to files beside them,
    which is right on the share and broken the moment the file is emailed —
    this one has to survive being sent (the user 2026-08-27).

    COLUMN A IS A NARROW GUTTER and every table starts in B (the user
    2026-08-31): content welded to the window edge has nowhere to breathe.
    The one thing that stays in A is the big title, which hangs into the
    gutter and spills across the empty cells beside it, so the eye still
    starts at the corner. Nothing else is written to column A - which is why
    `lint_layout` is told where the content actually begins.
    """
    alt = div.get("alt")
    cols = [("BILLED", "billed", MONEY, 18), ("COST", "cost", MONEY, 18),
            ("GROSS PROFIT", "gp", MONEY, 18), ("GP %", "gpm", PCT, 11),
            (f"NET  (company {OVERHEAD_PCT:.0%})", "net", MONEY, 20)]
    if alt:
        cols.append((alt["short"], "mnet", MONEY, 20))

    wb = Workbook()
    sm = wb.active
    sm.title = "Summary"
    sm.sheet_view.showGridLines = False

    LAST = C0 + len(cols) + 1          # job col + data cols + the file link
    _n_act = sum(1 for _, src, _x in jobs if src.get("status") == "Active")
    _t(sm, 1, 1, div["title"], size=SZ_TITLE, bold=True, color=NAVY)
    _pay = sum(src.get("payroll", 0.0) for _, src, _t2 in jobs)
    _note = (f" · excludes payroll of {_pay:,.0f} (carried in overhead)"
             if _pay else "")
    _scope = f" · {div['scope']}" if div.get("scope") else ""
    _t(sm, 2, C0, f"{_n_act} active · {len(jobs) - _n_act} completed{_scope} · click a "
                  f"job for its detail sheet, or 'open workbook' for the full file"
                  f"{_note}",
       size=SZ_SMALL, color=GREY)
    # Generated stamp with the TIME — these are re-run through the day and a
    # date alone cannot tell you which pull you are looking at (the user
    # 2026-08-31). Its own cell so it keeps its own format.
    _t(sm, 1, LAST, f"Generated {dt.datetime.now():%m/%d/%Y %I:%M %p}",
       size=SZ_SMALL, color=GREY, align="right")
    # the rule runs from the gutter so the hanging title sits ON it
    for c in range(1, LAST + 1):
        sm.cell(row=2, column=c).border = Border(bottom=HAIR)
    sm.row_dimensions[1].height = 30
    sm.row_dimensions[3].height = 8

    def _sum(sel):
        d = {k: sum(t[k] for _, _, t in sel)
             for k in ("billed", "cost", "gp", "oh", "net", "moh", "mnet")}
        for a, b in (("gpm", "gp"), ("netm", "net"), ("mnetm", "mnet")):
            d[a] = d[b] / d["billed"] if d["billed"] else 0
        return d

    tot = _sum(jobs)
    r = _kpi_strip(sm, 4, tot, _spans(C0, LAST), alt)

    _t(sm, r, C0, "JOB", size=SZ_SMALL - 1, bold=True, color="FFFFFF",
       fill=F_HDR, align="left", wrap=True, indent=1)
    for i, (h, _k, _f, _w) in enumerate(cols):
        _t(sm, r, C0 + 1 + i, h, size=SZ_SMALL - 1, bold=True, color="FFFFFF",
           fill=F_HDR, align="right", wrap=True)
    sm.row_dimensions[r].height = 30
    _t(sm, r, LAST, "FULL DETAIL", size=SZ_SMALL - 1, bold=True, color="FFFFFF",
       fill=F_HDR, align="center")
    r += 1

    def _row_figures(rr, t, bold_keys=("gp", "mnet")):
        for i, (_h, k, fmt, _w) in enumerate(cols):
            pos = k in ("gp", "gpm", "net", "mnet")
            _t(sm, rr, C0 + 1 + i, t[k], size=SZ, fmt=fmt, align="right",
               bold=k in bold_keys,
               color=(GREEN if t[k] >= 0 else RED) if pos else INK)

    def _job_row(job, _src, t):
        nonlocal r
        cell = _t(sm, r, C0, job_label(job, _src.get("title", "")), size=SZ, bold=True)
        cell.hyperlink = f"#'{job}'!A1"
        cell.font = Font(size=SZ, bold=True, color=LINK, underline="single")
        _row_figures(r, t)
        # "see the actual project excel for details" — the job name jumps to
        # its sheet INSIDE this file (survives being emailed); this opens the
        # real workbook on the share.
        if _src.get("rel"):
            lk = _t(sm, r, LAST, "open workbook  ↗", size=SZ_SMALL, color=LINK)
            lk.hyperlink = _src["rel"]
            lk.font = Font(size=SZ_SMALL, color=LINK, underline="single")
        if r % 2 == 0:
            for c in range(C0, LAST + 1):
                sm.cell(row=r, column=c).fill = F_BAND
        sm.row_dimensions[r].height = 22
        r += 1

    def _section(title, sel, note=""):
        nonlocal r
        if not sel:
            return
        _t(sm, r, C0, title, size=SZ, bold=True, color="FFFFFF", fill=F_HDR)
        if note:
            _t(sm, r, C0 + 1, note, size=SZ_SMALL - 1, color="FFFFFF", fill=F_HDR)
        for c in range(C0, LAST + 1):
            sm.cell(row=r, column=c).fill = F_HDR
        sm.row_dimensions[r].height = 22
        r += 1
        for _j, _sc, _tt in sorted(sel, key=lambda x: -x[2]["billed"]):
            _job_row(_j, _sc, _tt)
        st = _sum(sel)
        _t(sm, r, C0, f"subtotal — {len(sel)} job(s)", size=SZ, bold=True, color=NAVY)
        _row_figures(r, st, bold_keys=tuple(k for _h, k, _f, _w in cols))
        for c in range(C0, LAST + 1):
            sm.cell(row=r, column=c).border = Border(top=HAIR)
        r += 2

    # Kept apart on purpose: an ACTIVE job's cost is only what has landed so
    # far, so folding it into one total with finished work reads as a margin
    # nobody has earned yet.
    _active = [j for j in jobs if j[1].get("status") == "Active"]
    _done = [j for j in jobs if j[1].get("status") != "Active"]
    _section("ACTIVE — in progress", _active,
             "costs to date only — not finished")
    _section("COMPLETED", _done)
    _t(sm, r, C0, f"ALL {div['label']} — {len(jobs)} JOBS", size=SZ, bold=True,
       color=NAVY)
    _row_figures(r, tot, bold_keys=tuple(k for _h, k, _f, _w in cols))
    for c in range(C0, LAST + 1):
        sm.cell(row=r, column=c).border = Border(top=RULE)
    sm.row_dimensions[r].height = 24
    sm.column_dimensions["A"].width = GUTTER_W
    sm.column_dimensions[get_column_letter(C0)].width = 34
    for i, (_h, _k, _f, w) in enumerate(cols):
        sm.column_dimensions[get_column_letter(C0 + 1 + i)].width = w
    sm.column_dimensions[get_column_letter(LAST)].width = 19

    for job, src, t in sorted(jobs, key=lambda x: -x[2]["billed"]):
        ws = wb.create_sheet(job[:31])
        ws.sheet_view.showGridLines = False
        JLAST = C0 + 6                 # rightmost content column on a job sheet
        _t(ws, 1, 1, job_label(job, src.get("title", "")), size=SZ_TITLE - 2,
           bold=True, color=NAVY)
        back = _t(ws, 1, JLAST, "← back to Summary", size=SZ_SMALL, align="right")
        back.hyperlink = "#'Summary'!A1"
        back.font = Font(size=SZ_SMALL, color=LINK, underline="single")
        _t(ws, 2, C0, f"{src.get('status', 'Completed').lower()} job · "
                      f"{src['title'].replace('PROJECT P&L — ', '')}",
           size=SZ_SMALL, color=GREY)
        _t(ws, 2, JLAST, f"Generated {dt.datetime.now():%m/%d/%Y %I:%M %p}",
           size=SZ_SMALL, color=GREY, align="right")
        for c in range(1, JLAST + 1):
            ws.cell(row=2, column=c).border = Border(bottom=HAIR)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 8
        r = _kpi_strip(ws, 4, t, _spans(C0, JLAST), alt)

        # ONE grid for both blocks: B label/doc · C date · D what-for ·
        # G amount · H paid. No column is ever left blank in a data row —
        # the empty column A down the cost detail is what made the first cut
        # read as sloppy (the user 2026-08-27).
        _t(ws, r, C0, "INVOICED", size=SZ + 2, bold=True, color=NAVY)
        _t(ws, r, C0 + 5, t["billed"], size=SZ + 2, bold=True, color=NAVY,
           fmt=MONEY, align="right")
        r += 1
        for c, h in ((C0, "Invoice"), (C0 + 1, "Date"), (C0 + 2, "What for"),
                     (C0 + 5, "Amount"), (C0 + 6, "Paid?")):
            _t(ws, r, c, h, size=SZ_SMALL - 1, bold=True, color="FFFFFF",
               fill=F_HDR, align="right" if c == C0 + 5 else
               ("center" if c == C0 + 6 else "left"), indent=1 if c == C0 else 0)
        for c in range(C0, JLAST + 1):
            ws.cell(row=r, column=c).fill = F_HDR
        ws.row_dimensions[r].height = 22
        r += 1
        for inv in sorted(src["invoices"], key=lambda i: str(i["date"]), reverse=True):
            # An invoice number is an IDENTIFIER, so it reads left. Numeric
            # right-alignment parked it at the far edge of a wide column,
            # disconnected from its own header.
            c1 = _t(ws, r, C0, inv["doc"], size=SZ_SMALL, align="left")
            c1.number_format = "0"
            if inv["url"]:                      # → the invoice in QBO
                c1.hyperlink = inv["url"]
                c1.font = Font(size=SZ_SMALL, color=LINK, underline="single")
            dc = _t(ws, r, C0 + 1, inv["date"], size=SZ_SMALL, align="left")
            dc.number_format = "mm/dd/yyyy"
            _t(ws, r, C0 + 2, str(inv["memo"])[:110], size=SZ_SMALL)   # spills E:F
            _t(ws, r, C0 + 5, inv["gross"] + inv["ret_billed"], size=SZ_SMALL,
               fmt=MONEY, align="right")
            _t(ws, r, C0 + 6, str(inv["paid"]), size=SZ_SMALL, align="center",
               color=GREY if str(inv["paid"]).startswith("PAID") else RED)
            if r % 2 == 0:
                for c in range(C0, JLAST + 1):
                    ws.cell(row=r, column=c).fill = F_BAND
            ws.row_dimensions[r].height = 20
            r += 1
        if src["not_billed"]:
            _t(ws, r, C0, "(journal entry)", size=SZ_SMALL, color=GREY, align="left")
            _t(ws, r, C0 + 2, "retainage moved by journal entry", size=SZ_SMALL,
               color=GREY)
            _t(ws, r, C0 + 5, src["not_billed"], size=SZ_SMALL, fmt=MONEY,
               align="right", color=GREY)
            r += 1
        r += 1

        _t(ws, r, C0, "COSTS — account, then vendor, then every line",
           size=SZ + 2, bold=True, color=NAVY)
        _t(ws, r, C0 + 5, t["cost"], size=SZ + 2, bold=True, color=NAVY,
           fmt=MONEY, align="right")
        r += 1
        if src.get("payroll"):
            _t(ws, r, C0, f"excludes payroll of {src['payroll']:,.0f} — carried "
                          f"in the overhead %, not charged to the job",
               size=SZ_SMALL - 1, color=GREY)
            r += 1
        for c, h in ((C0, "Account / Vendor / Doc #"), (C0 + 1, "Date"),
                     (C0 + 2, "Description"), (C0 + 5, "Amount"), (C0 + 6, "Paid?")):
            _t(ws, r, c, h, size=SZ_SMALL - 1, bold=True, color="FFFFFF",
               fill=F_HDR, align="right" if c == C0 + 5 else
               ("center" if c == C0 + 6 else "left"), indent=1 if c == C0 else 0)
        for c in range(C0, JLAST + 1):
            ws.cell(row=r, column=c).fill = F_HDR
        ws.row_dimensions[r].height = 22
        r += 1
        for sec in src["sections"]:
            _t(ws, r, C0, sec["name"], size=SZ + 1, bold=True, color=NAVY)
            _t(ws, r, C0 + 5, sec["total"], size=SZ + 1, bold=True, color=NAVY,
               fmt=MONEY, align="right")
            for c in range(C0, JLAST + 1):
                ws.cell(row=r, column=c).border = Border(bottom=RULE)
            ws.row_dimensions[r].height = 26
            r += 1
            for acct in sec["accounts"]:
                _t(ws, r, C0, acct["name"], size=SZ, bold=True, color=INK, fill=F_BAND)
                _t(ws, r, C0 + 5, acct["total"], size=SZ, bold=True, color=INK,
                   fill=F_BAND, fmt=MONEY, align="right")
                for c in (C0 + 1, C0 + 2, C0 + 3, C0 + 4, C0 + 6):
                    ws.cell(row=r, column=c).fill = F_BAND
                ws.row_dimensions[r].height = 21
                r += 1
                for v in acct["vendors"]:
                    _t(ws, r, C0, v["name"], size=SZ_SMALL, bold=True, indent=1)
                    _t(ws, r, C0 + 5, v["total"], size=SZ_SMALL, bold=True,
                       fmt=MONEY, align="right")
                    ws.row_dimensions[r].outline_level = 1
                    ws.row_dimensions[r].hidden = True
                    r += 1
                    for ln in v["lines"]:
                        dc1 = _t(ws, r, C0, ln["doc"], size=SZ_SMALL, indent=2,
                                 align="left")
                        if ln["url"]:            # → the bill in QBO
                            dc1.hyperlink = ln["url"]
                            dc1.font = Font(size=SZ_SMALL, color=LINK,
                                            underline="single")
                        dc = _t(ws, r, C0 + 1, ln["date"], size=SZ_SMALL, align="left")
                        dc.number_format = "mm/dd/yyyy"
                        _t(ws, r, C0 + 2, str(ln["desc"])[:110], size=SZ_SMALL)
                        _t(ws, r, C0 + 5, ln["amt"], size=SZ_SMALL, fmt=MONEY_C,
                           align="right")
                        _t(ws, r, C0 + 6, ln["paid"], size=SZ_SMALL, color=GREY,
                           align="center")
                        ws.row_dimensions[r].outline_level = 2
                        ws.row_dimensions[r].hidden = True
                        r += 1
        # D carries the description and SPILLS across E:F (empty on data rows;
        # the metric tiles above are what keep those columns from reading as
        # gutters). Amount and Paid sit at the right edge, always aligned.
        ws.column_dimensions["A"].width = GUTTER_W
        for col, w in zip("BCDEFGH", (28, 14, 25, 25, 25, 21, 15)):
            ws.column_dimensions[col].width = w
        ws.sheet_properties.outlinePr.summaryBelow = False

    for ws in wb.worksheets:                 # lands on one page wide as a PDF
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    for msg in lint_layout(wb, first_col=C0):
        print(f"    ⚑ layout: {msg}")
    tmp.replace(out)


def lint_layout(wb, first_col: int = 1) -> List[str]:
    """Catch the things that make a sheet READ badly but verify fine — the
    class of defect that shipped twice before this existed (the user
    2026-08-27: "why don't you inspect it after? it looks sloppy").

    Checks: a column left empty inside the used range (a gutter), and data
    rows whose first column is blank (a ragged left edge). Cell-level checks
    catch corruption; these catch ugly.

    `first_col` is where the content actually begins - anything left of it is
    a DELIBERATE gutter, so neither check may fire on it."""
    from openpyxl.utils import get_column_letter
    out: List[str] = []
    for ws in wb.worksheets:
        last = max((c.column for row in ws.iter_rows() for c in row
                    if c.value is not None), default=0)
        if not last:
            continue
        used = set()
        for row in ws.iter_rows(min_row=1, max_col=last):
            for c in row:
                if c.value is not None:
                    used.add(c.column)
        # A column covered by a MERGED range is not a gutter — only the
        # top-left cell of a merge carries the value.
        for mr in ws.merged_cells.ranges:
            used.update(range(mr.min_col, mr.max_col + 1))
        gutters = [get_column_letter(i) for i in range(first_col, last + 1)
                   if i not in used]
        if gutters:
            out.append(f"{ws.title}: empty column(s) {gutters} inside A..{get_column_letter(last)}")
        blank_a = sum(1 for row in ws.iter_rows(min_row=6, max_col=last)
                      if row[first_col - 1].value is None
                      and any(c.value is not None for c in row[first_col:]))
        if blank_a > 5:
            out.append(f"{ws.title}: {blank_a} rows have data but a blank "
                       f"column {get_column_letter(first_col)}")
    return out


def _link_target(target: Path, base: Path) -> str:
    """A workbook link that resolves on BOTH Windows and Mac (the user
    2026-08-31).

    RELATIVE, always. An absolute path is either /Users/... or C:\\... and is
    wrong on the other machine the moment the file is opened there - and it
    breaks again the instant the folder is re-shared under a different OneDrive
    root, which is exactly what the division-folder sharing rule causes. Excel
    resolves a relative target against the workbook's own folder on both
    platforms, so the overview and the P&Ls only have to keep their positions
    relative to EACH OTHER, wherever the tree is mounted. Separators are forced
    to '/' - Windows Excel accepts them, Mac Excel needs them.
    """
    return Path(os.path.relpath(target, base)).as_posix()


# A job folder is named for its project, sometimes with the client appended
# ("RP7437 - OPA CONSTRUCTION"). The project # is the leading token.
_JOB_RE = re.compile(r"^(MFD|CP|RP)\s*-?\s*\d+(?:-FTW)?", re.I)


def _find_workbook(folder: Path, proj: str) -> Optional[Path]:
    """The generated P&L inside a job folder. CP/MFD write
    `Project_PnL_<proj>.xlsx`; RP writes `<proj> - <client>.xlsx`."""
    for cand in (folder / f"Project_PnL_{proj}.xlsx",
                 folder / f"{folder.name}.xlsx"):
        if cand.exists():
            return cand
    loose = [x for x in sorted(folder.glob("*.xlsx"))
             if not x.name.startswith("~$") and "Job Result" not in x.name]
    return loose[0] if len(loose) == 1 else None


def _iter_jobs(div_dir: Path, prefix: str):
    """(project #, workbook, status) for every job under a division.

    Live jobs are direct children of the division folder; finished ones are
    filed inside an archive subfolder ("completed mfd project p&l")."""
    out = []
    try:
        children = sorted(div_dir.iterdir())
    except OSError:
        return out
    for child in children:
        if not child.is_dir():
            continue
        if child.name.lower().startswith(pnl_paths.ARCHIVE_PREFIXES):
            for sub in sorted(child.iterdir()):
                if sub.is_dir():
                    out.append((sub, "Completed"))
            continue
        out.append((child, "Active"))
    jobs = []
    for folder, status in out:
        m = _JOB_RE.match(folder.name)
        if not m:
            continue
        proj = re.sub(r"[\s-]+", "", m.group(0).upper()).replace("FTW", "-FTW")
        if not proj.upper().startswith(prefix):
            continue
        wb = _find_workbook(folder, proj)
        if wb is None:
            print(f"  ⚠ {proj}: no P&L workbook in {folder.name}")
            continue
        jobs.append((proj, wb, status))
    return jobs


def _year_of(v) -> Optional[int]:
    return getattr(v, "year", None)


def _touched_year(src: dict, year: int) -> bool:
    """Did this job bill or spend anything in `year`? That is the test for
    "this year's projects, don't go further back" (the user 2026-08-31) - it
    keeps a job that is still running from a prior year and drops one that
    finished before the year started."""
    if any(_year_of(i["date"]) == year for i in src["invoices"]):
        return True
    return any(_year_of(ln["date"]) == year
               for s in src["sections"] for a in s["accounts"]
               for v in a["vendors"] for ln in v["lines"])


_DEFAULT_YEAR = object()


def resolve_year(division: dict, year=_DEFAULT_YEAR):
    """The year filter for a division: `_DEFAULT_YEAR` → its own default (CP/RP
    the current year, MFD all), 'all'/None → no filter, else that year."""
    if year is _DEFAULT_YEAR:
        year = division.get("default_year")
    return None if str(year).lower() in ("all", "none", "") else int(year)


def load_division(found, div_dir: Path, year):
    """(loaded, skipped) for the jobs in `found` (from `_iter_jobs`, already
    filtered by the caller).

    `loaded` is [(job, src, totals, src_path)] in `found` order. The ONE reader
    both the CLI and the post-run Overview rebuild go through, so the workbook a
    run produces can never be assembled two different ways."""
    loaded, skipped = [], []
    for job, src_path, status in found:
        src = read_source(src_path)
        if not src["sections"]:
            print(f"  ⚠ {job}: no cost detail in {src_path.name} — regenerate it")
            continue
        if year and not _touched_year(src, year):
            skipped.append(job)
            continue
        src["status"] = status
        src["rel"] = _link_target(src_path, div_dir)
        loaded.append((job, src, _totals(src), src_path))
    return loaded, skipped


def rebuild_overview(division: str, div_dir: "Path | None" = None,
                     year=_DEFAULT_YEAR) -> "dict | None":
    """Rebuild `<DIV> Overview.xlsx` from the workbooks on disk. Returns
    {path, jobs, billed, cost} or None when there is nothing to bundle.

    Called by the CLI's --bundle AND by project_pnl_export at the end of a run
    (the user 2026-09-03: "make sure now if we update any mfd p&l it will get
    updated on the overview"). Reads workbooks only - no QBO, no credentials -
    so it is cheap enough to run every time and the Overview can never be left
    describing a P&L that has since moved."""
    div = DIVISIONS[division]
    div_dir = (Path(div_dir).expanduser() if div_dir
               else pnl_paths.division_dir(div["prefix"]))
    year = resolve_year(div, year)
    loaded, _ = load_division(_iter_jobs(div_dir, div["prefix"]), div_dir, year)
    if not loaded:
        return None
    out = div_dir / f"{div['label']} Overview.xlsx"
    build_bundle([(j, s, t) for j, s, t, _ in loaded], out,
                 dict(div, scope=(f"{year} only" if year else "")))
    return {"path": out, "jobs": len(loaded),
            "billed": sum(t["billed"] for _, _, t, _ in loaded),
            "cost": sum(t["cost"] for _, _, t, _ in loaded)}


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Build <DIV> Overview.xlsx from the division's P&L workbooks")
    ap.add_argument("jobs", nargs="*",
                    help="limit the overview to these jobs (default: all of them)")
    ap.add_argument("--division", choices=sorted(DIVISIONS), default="mfd",
                    help="which division's folder to read and write (default mfd)")
    ap.add_argument("--year", default=None,
                    help="keep only jobs with an invoice or a cost dated in this "
                         "year (e.g. 2026); 'all' for no filter. Defaults to the "
                         "current year for CP and RP, all for MFD.")
    ap.add_argument("--folder", default=None,
                    help="override the division folder (rarely needed)")
    a = ap.parse_args()
    div = DIVISIONS[a.division]
    # The overview lands in the DIVISION folder, not the archive inside it: it
    # covers live and finished jobs alike, so filing it under "completed" put
    # it somewhere it did not belong (the user 2026-08-31).
    div_dir = (Path(a.folder).expanduser() if a.folder
               else pnl_paths.division_dir(div["prefix"]))
    year = resolve_year(div, a.year if a.year is not None else _DEFAULT_YEAR)

    found = _iter_jobs(div_dir, div["prefix"])
    if a.jobs:
        want = {j.upper() for j in a.jobs}
        found = [f for f in found if f[0].upper() in want]
    if not found:
        print(f"✗  no {div['label']} job folders under {div_dir}")
        return 1

    loaded, skipped = load_division(found, div_dir, year)
    if skipped:
        print(f"  · {len(skipped)} job(s) with nothing in {year}: "
              f"{', '.join(sorted(skipped))}")
    if not loaded:
        print("✗  nothing to build")
        return 1
    out = div_dir / f"{div['label']} Overview.xlsx"
    build_bundle([(j, s, t) for j, s, t, _ in loaded], out,
                 dict(div, scope=(f"{year} only" if year else "")))
    tb = sum(t["billed"] for _, _, t, _ in loaded)
    tc = sum(t["cost"] for _, _, t, _ in loaded)
    print(f"  {len(loaded)} jobs   billed ${tb:,.0f}   cost ${tc:,.0f}   "
          f"GP ${tb - tc:,.0f} ({(tb - tc) / tb * 100 if tb else 0:.2f}%)")
    print(f"  → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
