#!/usr/bin/env python3
"""
qbo_health.py — Local company health dashboard from QBO.

Pulls live numbers from QuickBooks Online (Reports API + entity queries)
and writes a multi-sheet Excel workbook to a private, non-synced local
folder. Designed to answer "where should I be looking today?" at a glance.

SHEETS
  Dashboard   — top-level KPIs + flagged callouts ("where to look")
  Cash        — bank accounts with current balances
  AR Aging    — customer-level aging with standard QBO buckets
  AP Aging    — vendor-level aging with standard QBO buckets
  Coverage    — AR vs AP side-by-side by bucket + net position
  P&L         — MTD / YTD with period-over-period variance
  Anomalies   — large/unusual recent expenses + overhead spikes
  _Meta       — timestamp + run info (hidden)

PRIVACY
  Default output: ~/Documents/CompanyHealth/health_dashboard.xlsx
    — NOT in OneDrive, NOT in iCloud, NOT in the project folder.
  After write: chmod 600 (owner-only read/write).
  For "password on open" defense-in-depth, create an encrypted .dmg
  (see README) and pass --out /Volumes/Health/health_dashboard.xlsx.

AUTH
  Reuses qbo_vault.py — single Touch ID unlocks the credentials blob
  and the access token is cached in memory for the run.

USAGE
  python3 qbo_health.py                     # refresh dashboard
  python3 qbo_health.py --out PATH          # override output path
  python3 qbo_health.py --no-lock           # skip chmod 600 (debugging)
  python3 qbo_health.py --anomaly-sigma 2   # sensitivity of overhead spike flag
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import os
import stat
import statistics
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import requests
except ImportError:
    print("missing dependency. Run: pip3 install --break-system-packages requests openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("missing dependency. Run: pip3 install --break-system-packages requests openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import qbo_vault as kc
from shared import paths

# ────────────────────────── constants ──────────────────────────

API_BASE = "https://quickbooks.api.intuit.com"
MINOR_VERSION = "70"

DEFAULT_OUTPUT = paths.companyhealth_dir() / "health_dashboard.xlsx"
DEFAULT_OVERRIDES = paths.companyhealth_dir() / "customer_overrides.xlsx"

# Aging bucket boundaries, in days — matches QBO's default Aged Receivables/Payables.
AGING_BUCKETS = [
    ("Current",  0,   0),   # not overdue
    ("1-30",     1,   30),
    ("31-60",    31,  60),
    ("61-90",    61,  90),
    ("90+",      91,  10_000),
]

# Overhead accounts (for anomaly detection). QBO's default COA uses these names;
# the script matches by substring so "Office Supplies" and "Office Supplies & Expenses"
# both resolve. If the user's COA differs, edit OVERHEAD_HINTS below.
OVERHEAD_HINTS = [
    "office", "insurance", "utilities", "rent", "telephone",
    "internet", "fuel", "dues", "subscriptions", "meals",
    "travel", "legal", "professional fees", "bank", "software",
    "advertising", "repairs", "maintenance",
]

# ── Construction-finance assumptions ──
# Target net profit you want left over after covering direct cost + overhead.
# 8-12% is typical for concrete contractors; adjust to your business.
TARGET_NET_PROFIT_PCT = 0.10

# YoY-change thresholds for flagging individual overhead accounts.
ACCOUNT_YOY_RED_THRESHOLD = 0.25     # >25% YoY = red
ACCOUNT_YOY_YELLOW_THRESHOLD = 0.10  # 10-25% YoY = yellow

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
TITLE_FONT = Font(bold=True, size=18, color="1F3A5F", name="Arial")
SECTION_FONT = Font(bold=True, size=13, color="1F3A5F", name="Arial")
METRIC_FONT = Font(bold=True, size=11, name="Arial")
BODY_FONT = Font(size=11, name="Arial")
FLAG_FILL = PatternFill("solid", fgColor="FFF5CC")
GOOD_FILL = PatternFill("solid", fgColor="DCE9D5")
BAD_FILL = PatternFill("solid", fgColor="F7D4D4")
# Cell border constants — currently unused; reintroduce Side + Border imports
# from openpyxl.styles if you start applying borders.

CURRENCY_FMT = '$#,##0;[Red]($#,##0);"-"'
CURRENCY_DEC = '$#,##0.00;[Red]($#,##0.00);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'


# ────────────────────────── auth ──────────────────────────

def load_credentials() -> Tuple[str, str]:
    """Single Touch ID prompt unlocks all QBO keys. Returns (access_token, company_id)."""
    if not kc.has_credentials():
        print("✗ no credentials stored.")
        print("  fix:  python3 setup_qbo.py")
        sys.exit(1)

    try:
        creds = kc.get_all()
    except kc.SecretsError as e:
        print(f"✗ Keychain read failed: {e}")
        sys.exit(1)

    required = ["QBO_CLIENT_ID", "QBO_CLIENT_SECRET", "QBO_COMPANY_ID", "QBO_REFRESH_TOKEN"]
    missing = [k for k in required if not creds.get(k)]
    if missing:
        print(f"✗ incomplete blob. Missing: {', '.join(missing)}")
        print("  fix:  python3 setup_qbo.py")
        sys.exit(1)

    basic = base64.b64encode(f"{creds['QBO_CLIENT_ID']}:{creds['QBO_CLIENT_SECRET']}".encode()).decode()
    r = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        headers={
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        data={"grant_type": "refresh_token", "refresh_token": creds["QBO_REFRESH_TOKEN"]},
        timeout=30,
    )
    if r.status_code != 200:
        print(f"✗ token refresh failed  status={r.status_code}")
        print(f"  body: {r.text[:400]}")
        print(f"  diagnose:  python3 setup_qbo.py --test")
        sys.exit(1)

    data = r.json()
    new_rt = data.get("refresh_token")
    if new_rt and new_rt != creds["QBO_REFRESH_TOKEN"]:
        try:
            kc.put("QBO_REFRESH_TOKEN", new_rt)
        except kc.SecretsError:
            pass

    return data["access_token"], creds["QBO_COMPANY_ID"]


# ────────────────────────── api helpers ──────────────────────────

def _api_get(path: str, access: str, params: Optional[dict] = None) -> dict:
    p = dict(params or {})
    p["minorversion"] = MINOR_VERSION
    r = requests.get(
        f"{API_BASE}{path}",
        headers={"Authorization": f"Bearer {access}", "Accept": "application/json"},
        params=p,
        timeout=60,
    )
    if r.status_code != 200:
        raise RuntimeError(f"{path} → {r.status_code}: {r.text[:400]}")
    return r.json()


def query(access: str, company_id: str, q: str) -> dict:
    return _api_get(f"/v3/company/{company_id}/query", access, params={"query": q})


def query_all(access: str, company_id: str, entity: str, where: str = "") -> List[dict]:
    out: List[dict] = []
    start = 1
    page = 500
    while True:
        q = f"SELECT * FROM {entity}"
        if where:
            q += f" WHERE {where}"
        q += f" STARTPOSITION {start} MAXRESULTS {page}"
        data = query(access, company_id, q)
        batch = data.get("QueryResponse", {}).get(entity, [])
        if not batch:
            break
        out.extend(batch)
        if len(batch) < page:
            break
        start += page
    return out


def report(access: str, company_id: str, name: str, params: Optional[dict] = None) -> dict:
    return _api_get(f"/v3/company/{company_id}/reports/{name}", access, params=params)


# ────────────────────────── data shaping ──────────────────────────

def _today() -> dt.date:
    return dt.date.today()


def _days_between(a: dt.date, b: dt.date) -> int:
    return (a - b).days


def _bucket_for_days(days_overdue: int) -> str:
    """Return the aging bucket label for `days_overdue`. Current = not yet overdue."""
    if days_overdue <= 0:
        return "Current"
    for label, lo, hi in AGING_BUCKETS:
        if label == "Current":
            continue
        if lo <= days_overdue <= hi:
            return label
    return "90+"


def fetch_bank_accounts(access: str, company_id: str) -> List[Dict[str, Any]]:
    """Bank + credit card accounts with CurrentBalance."""
    rows = query_all(access, company_id, "Account",
                     where="AccountType IN ('Bank','Credit Card') AND Active=true")
    out = []
    for a in rows:
        out.append({
            "Name": a.get("Name", ""),
            "AccountType": a.get("AccountType", ""),
            "AccountSubType": a.get("AccountSubType", ""),
            "CurrentBalance": float(a.get("CurrentBalance", 0) or 0),
            "Id": a.get("Id", ""),
        })
    return sorted(out, key=lambda r: (r["AccountType"] != "Bank", -r["CurrentBalance"]))


def fetch_retainage_accounts(access: str, company_id: str) -> List[Dict[str, Any]]:
    """Find any active account with 'retainage' in its name and return its
    CurrentBalance. QBO typically stores retainage receivable as an Other
    Current Asset account.

    We pull ALL active accounts and filter in Python because QBO's query
    language doesn't support case-insensitive LIKE reliably across tenants.
    Receivable (positive) vs payable (liability type) are kept separate in
    the return payload so the caller can decide what to surface.
    """
    rows = query_all(access, company_id, "Account", where="Active=true")
    receivables: List[Dict[str, Any]] = []
    payables: List[Dict[str, Any]] = []
    for a in rows:
        name = a.get("Name", "") or ""
        if "retainage" not in name.lower():
            continue
        entry = {
            "Name": name,
            "AccountType": a.get("AccountType", ""),
            "AccountSubType": a.get("AccountSubType", ""),
            "CurrentBalance": float(a.get("CurrentBalance", 0) or 0),
            "Id": a.get("Id", ""),
        }
        # Heuristic: AR / Asset types = receivable side; Liability/AP = payable side
        atype = entry["AccountType"].lower()
        if "liabilit" in atype or "payable" in atype or "payable" in name.lower():
            payables.append(entry)
        else:
            receivables.append(entry)
    return [
        {"side": "receivable", "accounts": receivables,
         "total": sum(a["CurrentBalance"] for a in receivables)},
        {"side": "payable", "accounts": payables,
         "total": sum(a["CurrentBalance"] for a in payables)},
    ]


def fetch_open_invoices(access: str, company_id: str) -> List[Dict[str, Any]]:
    """Unpaid invoices with Balance > 0. Includes DueDate for bucketing."""
    rows = query_all(access, company_id, "Invoice", where="Balance > '0'")
    today = _today()
    out = []
    for inv in rows:
        bal = float(inv.get("Balance", 0) or 0)
        if bal <= 0:
            continue
        due = inv.get("DueDate") or inv.get("TxnDate") or ""
        try:
            due_date = dt.date.fromisoformat(due[:10]) if due else today
        except ValueError:
            due_date = today
        days_overdue = _days_between(today, due_date)
        cust_ref = inv.get("CustomerRef", {})
        out.append({
            "Customer": cust_ref.get("name", ""),
            "CustomerId": cust_ref.get("value", ""),
            "Doc #": inv.get("DocNumber", ""),
            "Txn Date": (inv.get("TxnDate") or "")[:10],
            "Due Date": due[:10] if due else "",
            "Days Overdue": max(0, days_overdue),
            "Bucket": _bucket_for_days(days_overdue),
            "Balance": bal,
            "Total": float(inv.get("TotalAmt", 0) or 0),
            "PrivateNote": inv.get("PrivateNote", ""),
        })
    return sorted(out, key=lambda r: -r["Balance"])


def fetch_open_bills(access: str, company_id: str) -> List[Dict[str, Any]]:
    """Unpaid bills with Balance > 0."""
    rows = query_all(access, company_id, "Bill", where="Balance > '0'")
    today = _today()
    out = []
    for b in rows:
        bal = float(b.get("Balance", 0) or 0)
        if bal <= 0:
            continue
        due = b.get("DueDate") or b.get("TxnDate") or ""
        try:
            due_date = dt.date.fromisoformat(due[:10]) if due else today
        except ValueError:
            due_date = today
        days_overdue = _days_between(today, due_date)
        v_ref = b.get("VendorRef", {})
        out.append({
            "Vendor": v_ref.get("name", ""),
            "VendorId": v_ref.get("value", ""),
            "Doc #": b.get("DocNumber", ""),
            "Txn Date": (b.get("TxnDate") or "")[:10],
            "Due Date": due[:10] if due else "",
            "Days Overdue": max(0, days_overdue),
            "Bucket": _bucket_for_days(days_overdue),
            "Balance": bal,
            "Total": float(b.get("TotalAmt", 0) or 0),
            "PrivateNote": b.get("PrivateNote", ""),
        })
    return sorted(out, key=lambda r: -r["Balance"])


def fetch_recent_expense_lines(access: str, company_id: str, days: int = 90) -> List[Dict[str, Any]]:
    """Bill + Purchase line items in the last `days` days, resolved to account names.
    Used for overhead-spike anomaly detection."""
    since = (_today() - dt.timedelta(days=days)).isoformat()

    # Account id → name map
    acct_rows = query_all(access, company_id, "Account")
    acct_name = {a["Id"]: a.get("Name", "") for a in acct_rows}

    # Vendor id → name map (so we can show readable Name in anomaly rows)
    vend_rows = query_all(access, company_id, "Vendor")
    vendor_name = {
        v["Id"]: v.get("DisplayName") or v.get("CompanyName") or f"Vendor {v['Id']}"
        for v in vend_rows
    }

    out: List[Dict[str, Any]] = []

    for b in query_all(access, company_id, "Bill", where=f"TxnDate >= '{since}'"):
        v = b.get("VendorRef", {})
        vname = vendor_name.get(v.get("value", ""), v.get("name", ""))
        for ln in b.get("Line", []):
            det = ln.get("AccountBasedExpenseLineDetail") or ln.get("ItemBasedExpenseLineDetail")
            if not det:
                continue
            aid = det.get("AccountRef", {}).get("value", "")
            aname = acct_name.get(aid, det.get("AccountRef", {}).get("name", ""))
            out.append({
                "Txn Date": (b.get("TxnDate") or "")[:10],
                "Type": "Bill",
                "Doc #": b.get("DocNumber", ""),
                "Name": vname,
                "Account": aname,
                "Description": ln.get("Description", ""),
                "Amount": float(ln.get("Amount", 0) or 0),
            })

    for p in query_all(access, company_id, "Purchase", where=f"TxnDate >= '{since}'"):
        e = p.get("EntityRef", {})
        if e.get("type") == "Vendor":
            name = vendor_name.get(e.get("value", ""), e.get("name", ""))
        else:
            name = e.get("name", "")
        ptype = p.get("PaymentType", "Cash")
        for ln in p.get("Line", []):
            det = ln.get("AccountBasedExpenseLineDetail") or ln.get("ItemBasedExpenseLineDetail")
            if not det:
                continue
            aid = det.get("AccountRef", {}).get("value", "")
            aname = acct_name.get(aid, det.get("AccountRef", {}).get("name", ""))
            out.append({
                "Txn Date": (p.get("TxnDate") or "")[:10],
                "Type": f"Purchase ({ptype})",
                "Doc #": p.get("DocNumber", ""),
                "Name": name,
                "Account": aname,
                "Description": ln.get("Description", ""),
                "Amount": float(ln.get("Amount", 0) or 0),
            })

    return out


def fetch_invoices_since(access: str, company_id: str, since: str) -> List[Dict[str, Any]]:
    """All invoices (paid + unpaid) with TxnDate >= since. Keeps only fields
    needed for top-customers aggregation and DSO date lookup."""
    rows = query_all(access, company_id, "Invoice", where=f"TxnDate >= '{since}'")
    out = []
    for inv in rows:
        c = inv.get("CustomerRef", {})
        out.append({
            "Id": inv.get("Id", ""),
            "Customer": c.get("name", ""),
            "CustomerId": c.get("value", ""),
            "Doc #": inv.get("DocNumber", ""),
            "TxnDate": (inv.get("TxnDate") or "")[:10],
            "DueDate": (inv.get("DueDate") or "")[:10],
            "TotalAmt": float(inv.get("TotalAmt", 0) or 0),
            "Balance": float(inv.get("Balance", 0) or 0),
        })
    return out


def fetch_bills_since(access: str, company_id: str, since: str) -> List[Dict[str, Any]]:
    """All bills (paid + unpaid) with TxnDate >= since."""
    rows = query_all(access, company_id, "Bill", where=f"TxnDate >= '{since}'")
    out = []
    for b in rows:
        v = b.get("VendorRef", {})
        out.append({
            "Id": b.get("Id", ""),
            "Vendor": v.get("name", ""),
            "VendorId": v.get("value", ""),
            "Doc #": b.get("DocNumber", ""),
            "TxnDate": (b.get("TxnDate") or "")[:10],
            "DueDate": (b.get("DueDate") or "")[:10],
            "TotalAmt": float(b.get("TotalAmt", 0) or 0),
            "Balance": float(b.get("Balance", 0) or 0),
        })
    return out


def fetch_purchases_since(access: str, company_id: str, since: str) -> List[Dict[str, Any]]:
    """Cash/CC/Check purchases since `since` — grouped by vendor for top-vendor rollup."""
    rows = query_all(access, company_id, "Purchase", where=f"TxnDate >= '{since}'")
    out = []
    for p in rows:
        e = p.get("EntityRef", {})
        vendor_id = e.get("value", "") if e.get("type") == "Vendor" else ""
        out.append({
            "Id": p.get("Id", ""),
            "VendorId": vendor_id,
            "Vendor": e.get("name", ""),
            "TxnDate": (p.get("TxnDate") or "")[:10],
            "PaymentType": p.get("PaymentType", ""),
            "TotalAmt": float(p.get("TotalAmt", 0) or 0),
        })
    return out


def fetch_payments_since(access: str, company_id: str, since: str) -> List[Dict[str, Any]]:
    """Customer Payment entries since `since`. Each carries LinkedTxn pointing
    to the Invoice(s) it paid, which is how we compute DSO."""
    rows = query_all(access, company_id, "Payment", where=f"TxnDate >= '{since}'")
    out = []
    for p in rows:
        # Collect linked invoice ids at the header level AND at each Line.
        linked_ids: List[Tuple[str, float]] = []
        for link in (p.get("LinkedTxn") or []):
            if link.get("TxnType") == "Invoice":
                linked_ids.append((link.get("TxnId", ""), 0.0))
        for ln in p.get("Line", []):
            amt = float(ln.get("Amount", 0) or 0)
            for link in (ln.get("LinkedTxn") or []):
                if link.get("TxnType") == "Invoice":
                    linked_ids.append((link.get("TxnId", ""), amt))
        out.append({
            "Id": p.get("Id", ""),
            "TxnDate": (p.get("TxnDate") or "")[:10],
            "TotalAmt": float(p.get("TotalAmt", 0) or 0),
            "LinkedInvoices": linked_ids,
        })
    return out


def fetch_bill_payments_since(access: str, company_id: str, since: str) -> List[Dict[str, Any]]:
    """BillPayment entries since `since`. LinkedTxn points to the Bill(s) paid,
    which is how we compute DPO."""
    rows = query_all(access, company_id, "BillPayment", where=f"TxnDate >= '{since}'")
    out = []
    for p in rows:
        linked_ids: List[Tuple[str, float]] = []
        for link in (p.get("LinkedTxn") or []):
            if link.get("TxnType") == "Bill":
                linked_ids.append((link.get("TxnId", ""), 0.0))
        for ln in p.get("Line", []):
            amt = float(ln.get("Amount", 0) or 0)
            for link in (ln.get("LinkedTxn") or []):
                if link.get("TxnType") == "Bill":
                    linked_ids.append((link.get("TxnId", ""), amt))
        out.append({
            "Id": p.get("Id", ""),
            "TxnDate": (p.get("TxnDate") or "")[:10],
            "TotalAmt": float(p.get("TotalAmt", 0) or 0),
            "LinkedBills": linked_ids,
        })
    return out


# ────────────────────────── report parsers ──────────────────────────

def _walk_report_rows(node: dict, depth: int = 0) -> List[Tuple[str, List[Optional[float]], int]]:
    """Flatten a QBO report JSON tree into [(label, [col_values], depth), ...].

    QBO reports nest `Rows.Row` arrays with optional `Header`, `Rows`, `Summary`.
    Each leaf or summary carries `ColData` — a list of { value: "..." } cells.
    """
    out: List[Tuple[str, List[Optional[float]], int]] = []
    rows = (node.get("Rows") or {}).get("Row") or []
    for row in rows:
        header = row.get("Header")
        summary = row.get("Summary")
        col_data = row.get("ColData")
        if col_data:
            label = col_data[0].get("value", "") if col_data else ""
            values: List[Optional[float]] = []
            for c in col_data[1:]:
                v = c.get("value", "")
                try:
                    values.append(float(v) if v not in ("", None) else None)
                except ValueError:
                    values.append(None)
            out.append((label, values, depth))
        # Recurse first into nested rows, THEN emit the summary so the summary
        # sits after its children — matching QBO's on-screen order.
        if row.get("Rows"):
            out.extend(_walk_report_rows(row, depth + 1))
        if summary and summary.get("ColData"):
            label = summary["ColData"][0].get("value", "")
            values = []
            for c in summary["ColData"][1:]:
                v = c.get("value", "")
                try:
                    values.append(float(v) if v not in ("", None) else None)
                except ValueError:
                    values.append(None)
            # Only prepend "Total " if QBO didn't already include it.
            # Some QBO tenants emit summary ColData like "Cost of Goods Sold"
            # (consumer adds "Total"); others already emit "Total Cost of Goods Sold".
            # Double-prepending breaks every downstream label match.
            normalized = label.strip()
            if normalized.lower().startswith("total "):
                final_label = normalized
            else:
                final_label = f"Total {normalized}".strip()
            out.append((final_label, values, depth))
    return out


def _report_columns(report_data: dict) -> List[str]:
    cols = (report_data.get("Columns") or {}).get("Column") or []
    return [c.get("ColTitle", "") for c in cols[1:]]  # first column is the row label


def fetch_pl(access: str, company_id: str, start: str, end: str) -> Tuple[List[str], List[Tuple[str, List[Optional[float]], int]]]:
    data = report(access, company_id, "ProfitAndLoss", params={
        "start_date": start,
        "end_date": end,
        "accounting_method": "Accrual",
    })
    return _report_columns(data), _walk_report_rows(data)


def fetch_pl_monthly(access: str, company_id: str, start: str, end: str) -> Tuple[List[str], List[Tuple[str, List[Optional[float]], int]]]:
    data = report(access, company_id, "ProfitAndLoss", params={
        "start_date": start,
        "end_date": end,
        "accounting_method": "Accrual",
        "summarize_column_by": "Month",
    })
    return _report_columns(data), _walk_report_rows(data)


# Section-label aliases. Construction COAs vary widely — a contractor's QBO
# might call COGS "Job Costs" or "Direct Costs" and Expenses "Operating
# Expenses" or "Overhead". Matching against these keyword sets makes the
# parser robust to any reasonable naming.
_INCOME_TOTAL_HINTS = ("total income", "total revenue", "total sales", "total revenues")
_COGS_TOTAL_HINTS = (
    "total cost of goods sold", "total cogs",
    "total job costs", "total job cost", "total job-related costs",
    "total direct costs", "total direct cost",
    "total cost of sales", "total cost of revenue",
    "total construction costs",
)
_EXPENSE_TOTAL_HINTS = (
    "total expenses", "total expense",
    "total operating expenses", "total operating expense",
    "total overhead", "total g&a",
    "total general & administrative",
)


def _matches_any(key: str, hints: tuple) -> bool:
    return any(key.startswith(h) for h in hints)


def extract_pl_totals(rows: List[Tuple[str, List[Optional[float]], int]]) -> Dict[str, float]:
    """Pull standard P&L totals from a flattened report tree, robust to COA
    naming variation (Job Costs vs COGS, Operating Expenses vs Expenses, etc.).

    Includes an inference fallback: if COGS isn't named explicitly but Income
    and Gross Profit are present, COGS = Income − Gross Profit.
    """
    out: Dict[str, float] = {}
    for label, values, _ in rows:
        if not values:
            continue
        v = values[0] if values else None
        if v is None:
            continue
        key = label.strip().lower()
        if _matches_any(key, _INCOME_TOTAL_HINTS):
            out["income"] = v
        elif _matches_any(key, _COGS_TOTAL_HINTS):
            out["cogs"] = v
        elif key == "gross profit":
            out["gross_profit"] = v
        elif _matches_any(key, _EXPENSE_TOTAL_HINTS):
            out["expenses"] = v
        elif key.startswith("net operating income") or key.startswith("operating income"):
            out["net_op_income"] = v
        elif key.startswith("net income") or key == "net loss":
            out["net_income"] = v

    # Inference fallback: if COGS section wasn't named but we have Income +
    # Gross Profit, derive COGS arithmetically.
    if "cogs" not in out and "income" in out and "gross_profit" in out:
        out["cogs"] = out["income"] - out["gross_profit"]
    # Inference: if Gross Profit missing but Income and COGS are present
    if "gross_profit" not in out and "income" in out and "cogs" in out:
        out["gross_profit"] = out["income"] - out["cogs"]

    return out


def parse_pl_by_account(
    rows: List[Tuple[str, List[Optional[float]], int]],
) -> Dict[str, List[Dict[str, Any]]]:
    """Walk a flattened P&L tree and partition leaf account rows into sections.

    QBO emits a section's items at depth ≥ 1, then a 'Total <Section>' summary
    row when the section closes. We use the summary labels as section
    boundaries — that's more robust than relying on depth alone, since QBO
    reports nest sub-categories at variable depth across tenants.

    Returns:
        {
            'income':   [{'account': str, 'amount': float}, ...],
            'cogs':     [...],
            'expenses': [...],
        }
    Each list excludes header rows, sub-category subtotals, and the closing
    'Total <Section>' summary. Only true leaf accounts.
    """
    out: Dict[str, List[Dict[str, Any]]] = {"income": [], "cogs": [], "expenses": []}
    # Buffer leaf accounts as we walk; assign them to a section when we hit
    # that section's closing 'Total X' row. Robust to whether the report tree
    # emits explicit headers or only summaries.
    buffered: List[Dict[str, Any]] = []

    for label, values, depth in rows:
        if not label:
            continue
        amount = values[0] if values else None
        key = label.strip().lower()

        # Section close — flush buffered leaves to the matching section
        if _matches_any(key, _INCOME_TOTAL_HINTS):
            out["income"].extend(buffered)
            buffered = []
            continue
        if _matches_any(key, _COGS_TOTAL_HINTS):
            out["cogs"].extend(buffered)
            buffered = []
            continue
        if _matches_any(key, _EXPENSE_TOTAL_HINTS):
            out["expenses"].extend(buffered)
            buffered = []
            continue

        # Inter-section markers — discard whatever is buffered (Gross Profit
        # row, Net Operating Income, Net Income — none of these belong to a
        # leaf-account section)
        if (key.startswith("gross profit")
                or key.startswith("net operating")
                or key.startswith("operating income")
                or key.startswith("net income")
                or key == "net loss"):
            buffered = []
            continue

        # Sub-category subtotal inside a section (e.g., 'Total Field
        # Vehicle/Equipment' nested inside COGS). Skip — leaves under that
        # subcategory are already in `buffered` from prior iterations.
        if key.startswith("total ") and amount is not None:
            continue

        # Otherwise: leaf account row. Buffer it.
        if amount is not None and label.strip():
            buffered.append({"account": label.strip(), "amount": float(amount)})

    return out


def compute_overhead_metrics(
    pl_ytd_totals: Dict[str, float],
    pl_py_totals: Dict[str, float],
    pl_ytd_detail: Dict[str, List[Dict[str, Any]]],
    pl_py_detail: Dict[str, List[Dict[str, Any]]],
    target_profit_pct: float = TARGET_NET_PROFIT_PCT,
) -> Dict[str, Any]:
    """Construction-finance overhead analysis.

    Direct Costs (COGS) are job-tied — materials, sub labor, etc.
    Overhead (Operating Expenses) is what keeps the business running
    regardless of jobs — office rent, admin wages, insurance, etc.

    Required Markup % = Overhead / Direct  (the markup needed on direct
        cost just to recover overhead — your bid floor before profit)

    Bid Multiplier = (1 + required_markup) × (1 + target_profit)
        i.e. price = direct_cost × bid_multiplier and net profit will
        equal target_profit_pct of the resulting price.

    Break-even Revenue = Overhead / GM%  (revenue needed at current
        gross margin to cover overhead with $0 net income)

    Overhead Creep = YoY overhead growth rate − YoY revenue growth rate
        positive = overhead growing faster than top line, bad
    """
    revenue_ytd = pl_ytd_totals.get("income", 0) or 0
    cogs_ytd = pl_ytd_totals.get("cogs", 0) or 0
    overhead_ytd = pl_ytd_totals.get("expenses", 0) or 0
    gp_ytd = pl_ytd_totals.get("gross_profit", revenue_ytd - cogs_ytd) or 0

    revenue_py = pl_py_totals.get("income", 0) or 0
    cogs_py = pl_py_totals.get("cogs", 0) or 0
    overhead_py = pl_py_totals.get("expenses", 0) or 0
    gp_py = pl_py_totals.get("gross_profit", revenue_py - cogs_py) or 0

    # Ratios — guard every denominator
    overhead_pct_revenue_ytd = (overhead_ytd / revenue_ytd) if revenue_ytd else 0
    overhead_pct_revenue_py = (overhead_py / revenue_py) if revenue_py else 0
    required_markup_ytd = (overhead_ytd / cogs_ytd) if cogs_ytd else 0
    required_markup_py = (overhead_py / cogs_py) if cogs_py else 0
    gm_pct_ytd = (gp_ytd / revenue_ytd) if revenue_ytd else 0
    gm_pct_py = (gp_py / revenue_py) if revenue_py else 0
    breakeven_revenue_ytd = (overhead_ytd / gm_pct_ytd) if gm_pct_ytd else 0

    # Bid multiplier — what to multiply direct cost by so that net profit
    # equals target_profit_pct of REVENUE (the standard construction/GAAP
    # definition of "net profit margin", not markup-on-cost).
    #
    # Derivation:
    #   Revenue = COGS + Overhead + Profit
    #   Profit = target_profit_pct × Revenue   ← what we want
    #   Revenue × (1 - target_profit_pct) = COGS + Overhead
    #   Revenue = COGS × (1 + Overhead/COGS) / (1 - target_profit_pct)
    #          = COGS × (1 + required_markup) / (1 - target_profit_pct)
    if cogs_ytd and target_profit_pct < 1.0:
        bid_multiplier_ytd = (1 + required_markup_ytd) / (1 - target_profit_pct)
    else:
        bid_multiplier_ytd = 0

    # Net profit at the suggested bid (sanity check the math)
    suggested_bid_per_100k = 100_000 * bid_multiplier_ytd
    np_per_100k_at_suggested = suggested_bid_per_100k * target_profit_pct

    # Overhead creep (YoY growth comparison)
    revenue_growth = ((revenue_ytd - revenue_py) / revenue_py) if revenue_py else 0
    overhead_growth = ((overhead_ytd - overhead_py) / overhead_py) if overhead_py else 0
    creep = overhead_growth - revenue_growth

    # Are we under-recovering? Gross margin doesn't cover overhead's share of revenue.
    under_recovery = (overhead_pct_revenue_ytd > gm_pct_ytd) and revenue_ytd > 0

    # Per-account YoY breakdown (overhead detail)
    py_lookup = {a["account"]: a["amount"] for a in pl_py_detail.get("expenses", [])}
    overhead_accounts: List[Dict[str, Any]] = []
    for a in pl_ytd_detail.get("expenses", []):
        py_amt = py_lookup.get(a["account"], 0)
        delta = a["amount"] - py_amt
        delta_pct = (delta / py_amt) if py_amt else (1.0 if delta > 0 else 0.0)
        overhead_accounts.append({
            "Account": a["account"],
            "YTD": a["amount"],
            "PY YTD": py_amt,
            "YoY $": delta,
            "YoY %": delta_pct,
            "% of Overhead": (a["amount"] / overhead_ytd) if overhead_ytd else 0,
        })
    # Find PY accounts that disappeared YTD (also useful — spend dropped to $0)
    ytd_names = {a["account"] for a in pl_ytd_detail.get("expenses", [])}
    for a in pl_py_detail.get("expenses", []):
        if a["account"] not in ytd_names and a["amount"] > 0:
            overhead_accounts.append({
                "Account": a["account"],
                "YTD": 0,
                "PY YTD": a["amount"],
                "YoY $": -a["amount"],
                "YoY %": -1.0,
                "% of Overhead": 0,
            })

    overhead_accounts.sort(key=lambda r: -r["YTD"])

    # Top growers — accounts ≥ ACCOUNT_YOY_RED_THRESHOLD with material absolute size
    top_growers = [
        a for a in overhead_accounts
        if a["YoY %"] >= ACCOUNT_YOY_YELLOW_THRESHOLD
        and a["YTD"] > overhead_ytd * 0.02  # at least 2% of total overhead
    ]
    top_growers.sort(key=lambda r: -r["YoY %"])

    return {
        "revenue_ytd": revenue_ytd,
        "cogs_ytd": cogs_ytd,
        "overhead_ytd": overhead_ytd,
        "gp_ytd": gp_ytd,
        "revenue_py": revenue_py,
        "cogs_py": cogs_py,
        "overhead_py": overhead_py,
        "gp_py": gp_py,
        "overhead_pct_revenue_ytd": overhead_pct_revenue_ytd,
        "overhead_pct_revenue_py": overhead_pct_revenue_py,
        "required_markup_ytd": required_markup_ytd,
        "required_markup_py": required_markup_py,
        "gm_pct_ytd": gm_pct_ytd,
        "gm_pct_py": gm_pct_py,
        "breakeven_revenue_ytd": breakeven_revenue_ytd,
        "bid_multiplier_ytd": bid_multiplier_ytd,
        "target_profit_pct": target_profit_pct,
        "revenue_growth": revenue_growth,
        "overhead_growth": overhead_growth,
        "creep": creep,
        "under_recovery": under_recovery,
        "overhead_accounts": overhead_accounts,
        "top_growers": top_growers,
    }


# ────────────────────────── anomaly detection ──────────────────────────

def _is_overhead(account_name: str) -> bool:
    low = (account_name or "").lower()
    return any(h in low for h in OVERHEAD_HINTS)


def detect_overhead_spikes(
    recent_lines: List[Dict[str, Any]],
    sigma_threshold: float = 2.0,
) -> List[Dict[str, Any]]:
    """Flag overhead accounts whose 7-day spend is >=sigma_threshold std-devs
    above their rolling weekly mean over the trailing 90-day window."""
    today = _today()
    cutoff_7 = today - dt.timedelta(days=7)

    # Weekly totals per overhead account, over trailing 90 days.
    weekly: Dict[str, Dict[dt.date, float]] = defaultdict(lambda: defaultdict(float))
    for ln in recent_lines:
        if not _is_overhead(ln.get("Account", "")):
            continue
        try:
            d = dt.date.fromisoformat(ln["Txn Date"])
        except (ValueError, KeyError):
            continue
        week_start = d - dt.timedelta(days=d.weekday())  # Monday anchor
        weekly[ln["Account"]][week_start] += ln["Amount"]

    flags: List[Dict[str, Any]] = []
    for account, by_week in weekly.items():
        this_week_total = sum(
            ln["Amount"] for ln in recent_lines
            if ln.get("Account") == account
            and ln.get("Txn Date", "") >= cutoff_7.isoformat()
        )
        historical = [v for wk, v in by_week.items() if wk < cutoff_7]
        if len(historical) < 4:
            continue  # not enough history
        mean = statistics.mean(historical)
        try:
            stdev = statistics.stdev(historical)
        except statistics.StatisticsError:
            stdev = 0.0
        if stdev == 0:
            continue
        z = (this_week_total - mean) / stdev if stdev else 0
        if z >= sigma_threshold and this_week_total > mean:
            flags.append({
                "Account": account,
                "Last 7 Days": this_week_total,
                "Avg/week (prior 90d)": mean,
                "Std Dev": stdev,
                "Z-score": z,
                "$ Above Avg": this_week_total - mean,
            })

    return sorted(flags, key=lambda f: -f["Z-score"])


def find_large_recent_txns(
    recent_lines: List[Dict[str, Any]],
    days: int = 7,
    top_n: int = 10,
) -> List[Dict[str, Any]]:
    """Top-N largest expense line items in the last `days` days."""
    cutoff = (_today() - dt.timedelta(days=days)).isoformat()
    recent = [ln for ln in recent_lines if ln.get("Txn Date", "") >= cutoff]
    recent.sort(key=lambda r: -r.get("Amount", 0))
    return recent[:top_n]


# ────────────────────────── relationship rollups ──────────────────────────

def _parent_customer(name: str, aliases: Optional[Dict[str, str]] = None) -> str:
    """Extract the parent customer from QBO's hierarchical name format.
    QBO emits CustomerRef.name as 'Parent:Project # Description' for
    sub-customers / Projects. For concentration analysis we want the parent
    (the actual GC / homeowner that pays us), not the per-project subname.

    If `aliases` is provided (from the Customers sheet of the override file),
    apply user-supplied parent rollups so e.g. 'JPI Development' resolves to
    'JPI' for aggregation purposes.
    """
    if not name:
        return "(unknown)"
    parent = name.split(":", 1)[0].strip() or name
    if aliases:
        # Apply alias one or two hops max (avoid infinite loops on cycles).
        for _ in range(3):
            mapped = aliases.get(parent.lower())
            if not mapped or mapped == parent:
                break
            parent = mapped
    return parent


def _project_from_customer(name: str) -> str:
    """Extract a project number from the QBO CustomerRef.name format
    'Parent:Project# Description'. Returns the first whitespace-separated
    token of the sub-customer half. Empty string if no ':' present."""
    if not name or ":" not in name:
        return ""
    sub = name.split(":", 1)[1].strip()
    if not sub:
        return ""
    return sub.split()[0].strip()


def _read_sheet_rows(wb, sheet_name: str) -> List[Tuple]:
    """Helper — read rows from a named sheet, return [] if sheet absent."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return list(ws.iter_rows(values_only=True))


def _find_header_indexes(rows: List[Tuple], required: List[str], optional: List[str] = None) -> Tuple[Optional[Dict[str, int]], List[Tuple]]:
    """Locate column indexes from the first non-empty row. Match labels
    case-insensitive, alias-friendly. Returns (col_map, data_rows). col_map is
    None if any required column is missing."""
    optional = optional or []
    header_row = None
    header_idx = -1
    for i, r in enumerate(rows):
        if r and any(v not in (None, "") for v in r):
            header_row = [str(v).strip().lower() if v else "" for v in r]
            header_idx = i
            break
    if not header_row:
        return None, []

    cols: Dict[str, int] = {}
    for name in required + optional:
        for i, h in enumerate(header_row):
            if h == name.lower():
                cols[name] = i
                break

    for req in required:
        if req not in cols:
            return None, []

    return cols, rows[header_idx + 1:]


def load_overrides(path: Path) -> Dict[str, Any]:
    """Load the user-maintained annotations file.

    Multi-sheet xlsx. Each sheet is optional — script gracefully degrades
    if a sheet is missing or has the wrong columns.

    Sheet 'Customers':
        Customer | Maps To | Status | Notes
        - Maps To (optional): rolls this customer up under the named parent
          (e.g., "JPI Development" Maps To "JPI"). For top-customers,
          concentration risk, and aging grouping, both are treated as one.
        - Status (optional): routes off main AR onto Hold List

    Sheet 'Projects':
        Project # | Status | Notes
        - Per-project override. Match against project numbers extracted from
          QBO Customer strings of the shape 'Parent:Project# Description'.
          Use this when ONE project of a customer is in litigation but
          others are fine.

    Sheet 'Recurring Excludes':
        Vendor | Reason
        - Vendors here will NOT be flagged as recurring even if pattern
          detection thinks so. For false positives.

    Returns:
        {
          "customer_status":  {lower(name): {status, display_status, notes, display_name}},
          "customer_aliases": {lower(from_name): canonical_parent},
          "project_status":   {project_num: {status, display_status, notes}},
          "recurring_excludes": {lower(vendor): reason},
        }
    """
    empty = {
        "customer_status": {},
        "customer_aliases": {},
        "project_status": {},
        "recurring_excludes": {},
        "recurring_decisions": {},
    }
    if not path or not path.exists():
        return empty
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ⚠ couldn't read overrides at {path}: {e}")
        return empty

    out = dict(empty)

    # ── Customers sheet ──
    cust_rows = _read_sheet_rows(wb, "Customers")
    # Fallback: legacy single-sheet files where the only sheet was named
    # 'Overrides' or default Sheet1.
    if not cust_rows and wb.sheetnames:
        cust_rows = _read_sheet_rows(wb, wb.sheetnames[0])
    cols, data = _find_header_indexes(
        cust_rows, required=["customer"], optional=["maps to", "status", "notes"],
    )
    if cols is not None:
        for r in data:
            if r is None:
                continue
            cust = (str(r[cols["customer"]]).strip()
                    if cols.get("customer") is not None
                    and len(r) > cols["customer"]
                    and r[cols["customer"]] else "")
            if not cust or cust.startswith("#") or "(example)" in cust.lower():
                continue
            maps_to = ""
            if "maps to" in cols and len(r) > cols["maps to"] and r[cols["maps to"]]:
                maps_to = str(r[cols["maps to"]]).strip()
            status = ""
            if "status" in cols and len(r) > cols["status"] and r[cols["status"]]:
                status = str(r[cols["status"]]).strip()
            notes = ""
            if "notes" in cols and len(r) > cols["notes"] and r[cols["notes"]]:
                notes = str(r[cols["notes"]]).strip()

            if maps_to:
                out["customer_aliases"][cust.lower()] = maps_to
            if status:
                # The status entry uses the EFFECTIVE parent (post-alias) so
                # a held customer that aliases to JPI shows as JPI on the
                # Hold List sheet.
                effective_parent = maps_to if maps_to else cust
                out["customer_status"][effective_parent.lower()] = {
                    "status": status.lower(),
                    "display_status": status,
                    "notes": notes,
                    "display_name": effective_parent,
                }

    # ── Projects sheet ──
    proj_rows = _read_sheet_rows(wb, "Projects")
    cols, data = _find_header_indexes(
        proj_rows, required=["project #"], optional=["status", "notes"],
    )
    # Try alternate header names if first match failed
    if cols is None:
        cols, data = _find_header_indexes(
            proj_rows, required=["project"], optional=["status", "notes"],
        )
    if cols is not None:
        proj_col_key = "project #" if "project #" in cols else "project"
        for r in data:
            if r is None:
                continue
            proj = (str(r[cols[proj_col_key]]).strip()
                    if r[cols[proj_col_key]] else "")
            if not proj or proj.startswith("#"):
                continue
            status = ""
            if "status" in cols and len(r) > cols["status"] and r[cols["status"]]:
                status = str(r[cols["status"]]).strip()
            notes = ""
            if "notes" in cols and len(r) > cols["notes"] and r[cols["notes"]]:
                notes = str(r[cols["notes"]]).strip()
            if status:
                out["project_status"][proj] = {
                    "status": status.lower(),
                    "display_status": status,
                    "notes": notes,
                }

    # ── Recurring Excludes sheet ──
    rec_rows = _read_sheet_rows(wb, "Recurring Excludes")
    cols, data = _find_header_indexes(
        rec_rows, required=["vendor"], optional=["reason"],
    )
    if cols is not None:
        for r in data:
            if r is None:
                continue
            vendor = (str(r[cols["vendor"]]).strip()
                      if r[cols["vendor"]] else "")
            if not vendor or vendor.startswith("#"):
                continue
            reason = ""
            if "reason" in cols and len(r) > cols["reason"] and r[cols["reason"]]:
                reason = str(r[cols["reason"]]).strip()
            out["recurring_excludes"][vendor.lower()] = reason

    # ── Recurring Decisions sheet (per-vendor disambiguation) ──
    dec_rows = _read_sheet_rows(wb, "Recurring Decisions")
    cols, data = _find_header_indexes(
        dec_rows, required=["vendor"], optional=["decision", "notes"],
    )
    if cols is not None:
        for r in data:
            if r is None:
                continue
            vendor = (str(r[cols["vendor"]]).strip()
                      if r[cols["vendor"]] else "")
            if not vendor or vendor.startswith("#") or "(example)" in vendor.lower():
                continue
            decision = ""
            if "decision" in cols and len(r) > cols["decision"] and r[cols["decision"]]:
                decision = str(r[cols["decision"]]).strip().lower()
            if decision in ("active", "exclude", "ignore"):
                out["recurring_decisions"][vendor.lower()] = decision

    return out


# Backwards-compatible alias for the older single-purpose loader.
def load_customer_overrides(path: Path) -> Dict[str, Dict[str, str]]:
    """Legacy entry point — returns just the customer_status dict so older
    callers continue to work."""
    return load_overrides(path).get("customer_status", {})


def split_invoices_by_override(
    open_invoices: List[Dict[str, Any]],
    overrides: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Split AR rows into (regular, on_hold). Two layers of override:

      1. Customer-level (from Customers sheet, post-alias): if the entire
         customer is on hold/litigation/etc, ALL their invoices route to
         Hold List.
      2. Project-level (from Projects sheet): a single project of an
         otherwise-normal customer can be routed to Hold List, leaving the
         rest of that customer's AR on the main aging.

    Match is case-insensitive. Held rows get 'Override Status', 'Override
    Notes', and 'Parent Customer' fields added so the Hold List writer can
    group + display them.
    """
    customer_status = (overrides or {}).get("customer_status", {})
    project_status = (overrides or {}).get("project_status", {})
    aliases = (overrides or {}).get("customer_aliases", {})

    if not customer_status and not project_status:
        return open_invoices, []

    regular: List[Dict[str, Any]] = []
    held: List[Dict[str, Any]] = []

    for r in open_invoices:
        full_name = r.get("Customer") or ""
        parent = _parent_customer(full_name, aliases=aliases)
        proj = _project_from_customer(full_name)

        # Customer-level override has priority over project-level
        ov = customer_status.get(parent.lower())
        proj_ov = project_status.get(proj) if proj else None

        if ov:
            held_row = dict(r)
            held_row["Override Status"] = ov["display_status"]
            held_row["Override Notes"] = ov["notes"]
            held_row["Parent Customer"] = ov["display_name"]
            held_row["Override Source"] = "customer"
            held.append(held_row)
        elif proj_ov:
            held_row = dict(r)
            held_row["Override Status"] = proj_ov["display_status"]
            held_row["Override Notes"] = proj_ov["notes"]
            held_row["Parent Customer"] = parent
            held_row["Override Source"] = f"project {proj}"
            held.append(held_row)
        else:
            regular.append(r)
    return regular, held


def aggregate_top_customers(
    invoices_ytd: List[Dict[str, Any]],
    top_n: int = 10,
    aliases: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """Top-N customers by YTD invoiced revenue (accrual basis), grouped by
    PARENT customer (post-alias) so concentration risk reflects who actually
    pays us, not how many projects we run for them or how QBO splits the
    customer record."""
    by_name: Dict[str, float] = defaultdict(float)
    count: Dict[str, int] = defaultdict(int)
    for inv in invoices_ytd:
        full_name = inv.get("Customer") or "(unknown)"
        parent = _parent_customer(full_name, aliases=aliases)
        by_name[parent] += inv.get("TotalAmt", 0)
        count[parent] += 1
    total = sum(by_name.values()) or 0.0
    rows = [
        {
            "Customer": name,
            "YTD Revenue": amt,
            "Invoices": count[name],
            "% of YTD": (amt / total) if total else 0,
        }
        for name, amt in by_name.items()
    ]
    rows.sort(key=lambda r: -r["YTD Revenue"])
    return rows[:top_n]


def aggregate_top_vendors(
    bills_ytd: List[Dict[str, Any]],
    purchases_ytd: List[Dict[str, Any]],
    top_n: int = 10,
) -> List[Dict[str, Any]]:
    """Top-N vendors by YTD spend = Bills (accrual) + Purchases (cash/cc)."""
    by_name: Dict[str, float] = defaultdict(float)
    count: Dict[str, int] = defaultdict(int)
    for b in bills_ytd:
        name = b.get("Vendor") or "(unknown)"
        by_name[name] += b.get("TotalAmt", 0)
        count[name] += 1
    for p in purchases_ytd:
        name = p.get("Vendor") or "(unknown)"
        by_name[name] += p.get("TotalAmt", 0)
        count[name] += 1
    total = sum(by_name.values()) or 0.0
    rows = [
        {
            "Vendor": name,
            "YTD Spend": amt,
            "Transactions": count[name],
            "% of YTD": (amt / total) if total else 0,
        }
        for name, amt in by_name.items()
    ]
    rows.sort(key=lambda r: -r["YTD Spend"])
    return rows[:top_n]


def concentration_flag(top_customers: List[Dict[str, Any]], threshold: float = 0.15) -> Optional[Dict[str, Any]]:
    """If top customer > threshold of YTD revenue, return its info; else None.
    Default 15% matches the threshold construction sureties (Travelers, Zurich)
    flag for concentration risk on contractors under $10M revenue."""
    if not top_customers:
        return None
    top = top_customers[0]
    if top.get("% of YTD", 0) >= threshold:
        return {"Customer": top["Customer"], "% of YTD": top["% of YTD"], "YTD Revenue": top["YTD Revenue"]}
    return None


# ────────────────────────── DSO / DPO ──────────────────────────

def _month_key(d: str) -> str:
    """Return YYYY-MM from a YYYY-MM-DD string."""
    return d[:7] if len(d) >= 7 else ""


def compute_dso_monthly(
    invoices: List[Dict[str, Any]],
    payments: List[Dict[str, Any]],
    months_back: int = 12,
) -> List[Dict[str, Any]]:
    """Dollar-weighted DSO by payment month for the last `months_back` months.

    For each Payment row:
      for each (invoice_id, amount) in its LinkedInvoices:
        days_to_collect = payment_date - invoice_date
        contribute (days * amount) and amount to that payment month's totals.

    DSO = Σ(days × amount) / Σ(amount)  (dollar-weighted)
    """
    inv_date = {inv["Id"]: inv["TxnDate"] for inv in invoices if inv.get("Id")}

    bucket: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {"days_weighted": 0.0, "amount_total": 0.0, "count": 0.0}
    )

    for pmt in payments:
        pmt_date_str = pmt.get("TxnDate", "")
        if not pmt_date_str:
            continue
        try:
            pmt_d = dt.date.fromisoformat(pmt_date_str)
        except ValueError:
            continue
        month = _month_key(pmt_date_str)

        # Dedupe LinkedTxn entries: a single payment-to-invoice link can appear
        # both at header level (amount=0 placeholder) and at line level (with
        # actual amount). Take the largest amount per invoice id.
        per_invoice: Dict[str, float] = {}
        for inv_id, amt in (pmt.get("LinkedInvoices") or []):
            if not inv_id:
                continue
            if amt > per_invoice.get(inv_id, 0):
                per_invoice[inv_id] = amt

        # If no line-level amounts came through, fall back to splitting the
        # payment's total evenly across deduped linked invoices.
        if per_invoice and all(a == 0 for a in per_invoice.values()):
            even = pmt.get("TotalAmt", 0) / len(per_invoice) if per_invoice else 0
            per_invoice = {k: even for k in per_invoice}

        link_amounts = list(per_invoice.items())

        for inv_id, amt in link_amounts:
            if amt <= 0:
                continue
            inv_d_str = inv_date.get(inv_id)
            if not inv_d_str:
                continue
            try:
                inv_d = dt.date.fromisoformat(inv_d_str)
            except ValueError:
                continue
            days = (pmt_d - inv_d).days
            if days < 0:
                continue  # data weirdness — skip
            bucket[month]["days_weighted"] += days * amt
            bucket[month]["amount_total"] += amt
            bucket[month]["count"] += 1

    out: List[Dict[str, Any]] = []
    for month, data in sorted(bucket.items()):
        if data["amount_total"] <= 0:
            continue
        out.append({
            "Month": month,
            "DSO": round(data["days_weighted"] / data["amount_total"], 1),
            "Invoices Paid": int(data["count"]),
            "$ Collected": data["amount_total"],
        })
    # Keep the last N months
    return out[-months_back:] if months_back else out


def compute_dpo_monthly(
    bills: List[Dict[str, Any]],
    bill_payments: List[Dict[str, Any]],
    months_back: int = 12,
) -> List[Dict[str, Any]]:
    """Dollar-weighted DPO by payment month. Mirror of compute_dso_monthly."""
    bill_date = {b["Id"]: b["TxnDate"] for b in bills if b.get("Id")}

    bucket: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {"days_weighted": 0.0, "amount_total": 0.0, "count": 0.0}
    )

    for bp in bill_payments:
        pmt_date_str = bp.get("TxnDate", "")
        if not pmt_date_str:
            continue
        try:
            pmt_d = dt.date.fromisoformat(pmt_date_str)
        except ValueError:
            continue
        month = _month_key(pmt_date_str)

        # Dedupe LinkedTxn entries — same fix as compute_dso_monthly. A single
        # bill payment can carry duplicate LinkedTxn entries for the same bill
        # at header + line level, with only the line-level having an amount.
        per_bill_d: Dict[str, float] = {}
        for bill_id, amt in (bp.get("LinkedBills") or []):
            if not bill_id:
                continue
            if amt > per_bill_d.get(bill_id, 0):
                per_bill_d[bill_id] = amt

        if per_bill_d and all(a == 0 for a in per_bill_d.values()):
            even = bp.get("TotalAmt", 0) / len(per_bill_d) if per_bill_d else 0
            per_bill_d = {k: even for k in per_bill_d}

        link_amounts = list(per_bill_d.items())

        for bill_id, amt in link_amounts:
            if amt <= 0:
                continue
            b_d_str = bill_date.get(bill_id)
            if not b_d_str:
                continue
            try:
                b_d = dt.date.fromisoformat(b_d_str)
            except ValueError:
                continue
            days = (pmt_d - b_d).days
            if days < 0:
                continue
            bucket[month]["days_weighted"] += days * amt
            bucket[month]["amount_total"] += amt
            bucket[month]["count"] += 1

    out: List[Dict[str, Any]] = []
    for month, data in sorted(bucket.items()):
        if data["amount_total"] <= 0:
            continue
        out.append({
            "Month": month,
            "DPO": round(data["days_weighted"] / data["amount_total"], 1),
            "Bills Paid": int(data["count"]),
            "$ Paid": data["amount_total"],
        })
    return out[-months_back:] if months_back else out


# ────────────────────────── cash flow & runway ──────────────────────────

def _classify_cadence(mean_gap_days: float) -> Optional[str]:
    """Map an average gap (days) between transactions to a human-readable cadence.
    Returns None if the gap doesn't fit a recognizable recurring rhythm — those
    entries get dropped from recurring detection."""
    if 0.5 <= mean_gap_days <= 1.7: return "Daily"
    if 1.7 < mean_gap_days <= 4.5:   return "Every few days"
    if 5 <= mean_gap_days <= 9:      return "Weekly"
    if 12 <= mean_gap_days <= 18:    return "Biweekly"
    if 25 <= mean_gap_days <= 35:    return "Monthly"
    if 55 <= mean_gap_days <= 70:    return "Bimonthly"
    if 85 <= mean_gap_days <= 100:   return "Quarterly"
    return None


def _cadence_to_weekly(mean_amount: float, mean_gap_days: float) -> float:
    if mean_gap_days <= 0:
        return 0.0
    return mean_amount / mean_gap_days * 7


def _cadence_to_monthly(mean_amount: float, mean_gap_days: float) -> float:
    if mean_gap_days <= 0:
        return 0.0
    return mean_amount / mean_gap_days * 30.4375  # avg days/month


def detect_recurring_payments(
    bills_12mo: List[Dict[str, Any]],
    purchases_12mo: List[Dict[str, Any]],
    min_occurrences: int = 4,
    amount_stability_max: float = 0.15,
    gap_stability_max: float = 0.35,
    excludes: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """Identify recurring payment streams over the last 12 months.

    Groups by Vendor, then within each vendor clusters transactions by amount
    similarity (10% band). For each (vendor, amount-cluster), flags the stream
    as recurring when:
      - ≥ min_occurrences transactions
      - amount coefficient-of-variation ≤ amount_stability_max
      - date-gap coefficient-of-variation ≤ gap_stability_max
      - the average gap falls into a recognizable cadence bucket

    Returns a list sorted by monthly cost (largest first).
    """
    # Normalize Bill + Purchase into a common shape
    txns: List[Dict[str, Any]] = []
    for b in bills_12mo:
        if not b.get("Vendor") or not b.get("TxnDate"):
            continue
        txns.append({
            "Vendor": b["Vendor"],
            "Date": b["TxnDate"],
            "Amount": float(b.get("TotalAmt", 0) or 0),
            "Source": "Bill",
        })
    for p in purchases_12mo:
        name = p.get("Vendor") or ""
        if not name or not p.get("TxnDate"):
            continue
        txns.append({
            "Vendor": name,
            "Date": p["TxnDate"],
            "Amount": float(p.get("TotalAmt", 0) or 0),
            "Source": f"Purchase ({p.get('PaymentType', '')})".strip(),
        })

    excludes_lc = {k.lower() for k in (excludes or {}).keys()}

    by_vendor: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for t in txns:
        if t["Amount"] > 0:
            by_vendor[t["Vendor"]].append(t)

    recurring: List[Dict[str, Any]] = []

    for vendor, tx_list in by_vendor.items():
        # User-tagged vendors that we know aren't actually recurring even
        # though the math says they look that way — skip entirely.
        if vendor.lower() in excludes_lc:
            continue
        if len(tx_list) < min_occurrences:
            continue

        # Cluster by amount: group amounts within 10% of an existing cluster center.
        # Order by amount desc so clusters center on the most common bill size.
        tx_list.sort(key=lambda x: -x["Amount"])
        clusters: List[List[Dict[str, Any]]] = []
        cluster_centers: List[float] = []
        for t in tx_list:
            placed = False
            for i, center in enumerate(cluster_centers):
                if center > 0 and abs(t["Amount"] - center) / center < 0.10:
                    clusters[i].append(t)
                    # Update the center to the running mean of the cluster
                    amts = [x["Amount"] for x in clusters[i]]
                    cluster_centers[i] = sum(amts) / len(amts)
                    placed = True
                    break
            if not placed:
                clusters.append([t])
                cluster_centers.append(t["Amount"])

        for cluster in clusters:
            if len(cluster) < min_occurrences:
                continue

            # Noise filter: a coincidental cluster of 4 among a variable
            # supplier's 40 transactions shouldn't be flagged as recurring.
            # Threshold 0.3 — strict enough to reject coincidental clusters
            # from variable-cost suppliers but lenient enough to catch real
            # monthly streams (e.g., 1 monthly insurance bill + 4 one-off
            # claim adjustments still get flagged at 1/5 = 20%... too tight,
            # so we tune to 30% = the recurring stream needs to be at least
            # a third of the vendor's activity).
            cluster_fraction = len(cluster) / len(tx_list)
            if cluster_fraction < 0.3:
                continue

            cluster.sort(key=lambda x: x["Date"])
            try:
                dates = [dt.date.fromisoformat(x["Date"]) for x in cluster]
            except ValueError:
                continue
            gaps = [(dates[i + 1] - dates[i]).days for i in range(len(dates) - 1)]
            if not gaps:
                continue

            mean_gap = sum(gaps) / len(gaps)
            if mean_gap <= 0:
                continue
            try:
                gap_stdev = statistics.stdev(gaps) if len(gaps) > 1 else 0.0
            except statistics.StatisticsError:
                gap_stdev = 0.0
            gap_cv = gap_stdev / mean_gap

            amounts = [x["Amount"] for x in cluster]
            mean_amount = sum(amounts) / len(amounts)
            try:
                amt_stdev = statistics.stdev(amounts) if len(amounts) > 1 else 0.0
            except statistics.StatisticsError:
                amt_stdev = 0.0
            amt_cv = amt_stdev / mean_amount if mean_amount else 1.0

            # Daily debits hitting only weekdays have structural gap variance
            # (four 1-day gaps + one 3-day gap = CV ~0.57). Allow more slack
            # for daily cadences where the shape is structural, not noise.
            allowed_gap_cv = 0.70 if mean_gap <= 2.0 else gap_stability_max
            if gap_cv > allowed_gap_cv or amt_cv > amount_stability_max:
                continue

            cadence = _classify_cadence(mean_gap)
            if cadence is None:
                continue

            recurring.append({
                "Vendor": vendor,
                "Cadence": cadence,
                "Typical Amount": mean_amount,
                "Avg Gap (days)": round(mean_gap, 1),
                "Occurrences (12mo)": len(cluster),
                "Last Seen": dates[-1].isoformat(),
                "Weekly Cost": _cadence_to_weekly(mean_amount, mean_gap),
                "Monthly Cost": _cadence_to_monthly(mean_amount, mean_gap),
                "Amount Stability": round(1 - amt_cv, 2),
                "Cadence Stability": round(1 - gap_cv, 2),
                "Source": cluster[-1]["Source"],
            })

    return sorted(recurring, key=lambda r: -r["Monthly Cost"])


def categorize_recurring_streams(
    streams_12mo: List[Dict[str, Any]],
    bills_12mo: List[Dict[str, Any]],
    purchases_12mo: List[Dict[str, Any]],
    decisions: Optional[Dict[str, str]] = None,
    current_window_days: int = 90,
) -> Dict[str, List[Dict[str, Any]]]:
    """Classify each detected recurring stream as Active / Stale / New.

    Why this matters: a 12-month detection window captures subscriptions you
    cancelled months ago, inflating "monthly burn" with money you're not
    actually paying. The two-phase approach gives the *real* burn:

      Active = detected in 12mo AND has matching activity in last
               `current_window_days` (default 90) → counts toward real burn
      Stale  = detected in 12mo, NO matching activity in last 90 days →
               likely cancelled/ended; excluded from real burn
      New    = picked up only by a separate 3-month sweep, not present in
               the 12mo detection → newly recurring; counted toward burn
               but flagged for review (may not stick)

    User decisions from the override file's "Recurring Decisions" sheet
    take precedence: 'active' forces a stream into Active even if auto-
    classification would say Stale; 'exclude' moves it out entirely.
    """
    decisions = decisions or {}
    today = _today()
    cutoff = today - dt.timedelta(days=current_window_days)

    # Combine bills + purchases for activity scanning
    all_txns: List[Dict[str, Any]] = []
    for b in bills_12mo:
        all_txns.append({
            "Vendor": b.get("Vendor", ""),
            "Date": b.get("TxnDate", ""),
            "Amount": float(b.get("TotalAmt", 0) or 0),
        })
    for p in purchases_12mo:
        all_txns.append({
            "Vendor": p.get("Vendor", ""),
            "Date": p.get("TxnDate", ""),
            "Amount": float(p.get("TotalAmt", 0) or 0),
        })

    def _has_recent_activity(stream: Dict[str, Any]) -> Tuple[bool, Optional[str]]:
        """Check if there's a matching transaction (vendor + ±15% of typical
        amount) in the last `current_window_days`. Returns (found, latest_date)."""
        vendor_lc = stream["Vendor"].lower()
        target = stream["Typical Amount"]
        latest: Optional[dt.date] = None
        for t in all_txns:
            if t["Vendor"].lower() != vendor_lc:
                continue
            if not target:
                continue
            if abs(t["Amount"] - target) / target > 0.15:
                continue
            try:
                d = dt.date.fromisoformat(t["Date"])
            except (ValueError, TypeError):
                continue
            if d >= cutoff:
                if latest is None or d > latest:
                    latest = d
        return (latest is not None, latest.isoformat() if latest else None)

    active: List[Dict[str, Any]] = []
    stale: List[Dict[str, Any]] = []

    for stream in streams_12mo:
        s = dict(stream)
        # User decision wins
        decision = decisions.get(s["Vendor"].lower())
        if decision == "exclude":
            continue  # silently dropped
        if decision == "active":
            s["Status"] = "Active (forced)"
            s["Latest Activity"] = ""
            s["Decision Source"] = "user-decision"
            active.append(s)
            continue

        found, latest = _has_recent_activity(s)
        s["Latest Activity"] = latest or s.get("Last Seen", "")
        if found:
            s["Status"] = "Active"
            s["Decision Source"] = "auto"
            active.append(s)
        else:
            try:
                last_seen_date = dt.date.fromisoformat(s.get("Last Seen", ""))
                days_ago = (today - last_seen_date).days
            except (ValueError, TypeError):
                days_ago = current_window_days + 1
            s["Status"] = f"Stale ({days_ago}d since last)"
            s["Decision Source"] = "auto"
            stale.append(s)

    # ── Detect "new" streams: rerun detection on 3-month window with relaxed
    # min_occurrences, then exclude any stream we already classified above. ──
    bills_3mo = [b for b in bills_12mo if (b.get("TxnDate") or "") >= cutoff.isoformat()]
    purchases_3mo = [p for p in purchases_12mo if (p.get("TxnDate") or "") >= cutoff.isoformat()]
    streams_3mo = detect_recurring_payments(
        bills_3mo, purchases_3mo,
        min_occurrences=2,        # 3 monthly bills won't always all fit in 90 days
        amount_stability_max=0.15,
        gap_stability_max=0.40,   # slightly looser
    )

    known_keys = {(s["Vendor"].lower(), round(s["Typical Amount"] / 100) * 100)
                  for s in active + stale}

    new_streams: List[Dict[str, Any]] = []
    for s in streams_3mo:
        key = (s["Vendor"].lower(), round(s["Typical Amount"] / 100) * 100)
        if key in known_keys:
            continue
        decision = decisions.get(s["Vendor"].lower())
        if decision == "exclude":
            continue
        n = dict(s)
        n["Status"] = "New (3mo only)"
        n["Latest Activity"] = s.get("Last Seen", "")
        n["Decision Source"] = "auto"
        new_streams.append(n)

    return {
        "active": active,
        "stale": stale,
        "new": new_streams,
    }


def compute_fixed_obligations(recurring: List[Dict[str, Any]]) -> Dict[str, float]:
    """Total burn — used to compute the 'if all detected continued' number."""
    return {
        "weekly_total": sum(r["Weekly Cost"] for r in recurring),
        "monthly_total": sum(r["Monthly Cost"] for r in recurring),
        "count": len(recurring),
    }


def compute_burn_summary(categorized: Dict[str, List[Dict[str, Any]]]) -> Dict[str, float]:
    """Real burn breakdown. Active + New = what you're actually paying right now.
    Stale = what would inflate the number if we hadn't filtered it out."""
    active = categorized.get("active", [])
    stale = categorized.get("stale", [])
    new = categorized.get("new", [])

    real_streams = active + new
    return {
        "active_weekly": sum(r["Weekly Cost"] for r in active),
        "active_monthly": sum(r["Monthly Cost"] for r in active),
        "new_weekly": sum(r["Weekly Cost"] for r in new),
        "new_monthly": sum(r["Monthly Cost"] for r in new),
        "stale_weekly": sum(r["Weekly Cost"] for r in stale),
        "stale_monthly": sum(r["Monthly Cost"] for r in stale),
        "real_weekly": sum(r["Weekly Cost"] for r in real_streams),
        "real_monthly": sum(r["Monthly Cost"] for r in real_streams),
        "active_count": len(active),
        "stale_count": len(stale),
        "new_count": len(new),
    }


def compute_weekly_cash_flow(
    payments: List[Dict[str, Any]],
    bill_payments: List[Dict[str, Any]],
    purchases: List[Dict[str, Any]],
    weeks: int = 13,
) -> List[Dict[str, Any]]:
    """Weekly net cash flow over the last `weeks` weeks.

    Inflow  = customer Payments (cash actually received)
    Outflow = BillPayments + Purchases (cash actually paid out)
    """
    today = _today()
    earliest = today - dt.timedelta(weeks=weeks)

    def _week_anchor(date_str: str) -> Optional[dt.date]:
        try:
            d = dt.date.fromisoformat(date_str)
        except ValueError:
            return None
        if d < earliest:
            return None
        return d - dt.timedelta(days=d.weekday())  # Monday anchor

    inflow: Dict[dt.date, float] = defaultdict(float)
    outflow: Dict[dt.date, float] = defaultdict(float)

    for pmt in payments:
        anchor = _week_anchor(pmt.get("TxnDate", ""))
        if anchor:
            inflow[anchor] += pmt.get("TotalAmt", 0)
    for bp in bill_payments:
        anchor = _week_anchor(bp.get("TxnDate", ""))
        if anchor:
            outflow[anchor] += bp.get("TotalAmt", 0)
    for p in purchases:
        anchor = _week_anchor(p.get("TxnDate", ""))
        if anchor:
            outflow[anchor] += p.get("TotalAmt", 0)

    all_weeks = sorted(set(inflow.keys()) | set(outflow.keys()))
    out = []
    for wk in all_weeks:
        recv = inflow.get(wk, 0.0)
        paid = outflow.get(wk, 0.0)
        out.append({
            "Week Of": wk.isoformat(),
            "Received": recv,
            "Paid": paid,
            "Net": recv - paid,
        })
    return out


def compute_runway(current_cash: float, weekly_flow: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Cash runway in weeks based on trailing 13-week average NEGATIVE net flow.
    If the business is net cash-positive over the period, runway is effectively infinite."""
    if not weekly_flow:
        return {"weeks": None, "avg_weekly_paid": 0, "avg_weekly_net": 0, "note": "no cash-flow data"}
    paid = [w["Paid"] for w in weekly_flow]
    nets = [w["Net"] for w in weekly_flow]
    avg_paid = sum(paid) / len(paid) if paid else 0
    avg_net = sum(nets) / len(nets) if nets else 0
    if avg_net >= 0:
        return {
            "weeks": None,
            "avg_weekly_paid": avg_paid,
            "avg_weekly_net": avg_net,
            "note": "net cash-positive — runway not constrained by current trend",
        }
    # avg_net < 0 → burning cash; runway = cash / weekly_burn
    weekly_burn = -avg_net
    return {
        "weeks": current_cash / weekly_burn if weekly_burn else None,
        "avg_weekly_paid": avg_paid,
        "avg_weekly_net": avg_net,
        "note": f"based on 13-wk avg net burn of ${weekly_burn:,.0f}/week",
    }


# ────────────────────────── xlsx writers ──────────────────────────

def _set_header_row(ws: Worksheet, headers: List[str], row: int = 1) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws: Worksheet, widths: Dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_cash_sheet(
    ws: Worksheet,
    accounts: List[Dict[str, Any]],
    weekly_flow: List[Dict[str, Any]],
    runway: Dict[str, Any],
) -> None:
    # ── Account balances block ──
    _set_header_row(ws, ["Account", "Type", "Subtype", "Current Balance"])
    for i, a in enumerate(accounts, start=2):
        ws.cell(row=i, column=1, value=a["Name"]).font = BODY_FONT
        ws.cell(row=i, column=2, value=a["AccountType"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=a["AccountSubType"]).font = BODY_FONT
        c = ws.cell(row=i, column=4, value=a["CurrentBalance"])
        c.number_format = CURRENCY_FMT
        c.font = BODY_FONT

    last = len(accounts) + 1 if accounts else 1
    if accounts:
        tot_row = last + 1
        ws.cell(row=tot_row, column=1, value="Total").font = METRIC_FONT
        tot = ws.cell(row=tot_row, column=4, value=f"=SUM(D2:D{last})")
        tot.number_format = CURRENCY_FMT
        tot.font = METRIC_FONT

        bank_only = tot_row + 1
        ws.cell(row=bank_only, column=1, value="Bank only (excl. Credit Cards)").font = METRIC_FONT
        bc = ws.cell(row=bank_only, column=4,
                     value=f'=SUMIF(B2:B{last},"Bank",D2:D{last})')
        bc.number_format = CURRENCY_FMT
        bc.font = METRIC_FONT
        last_row = bank_only
    else:
        last_row = 1

    # ── Runway block ──
    runway_start = last_row + 3
    ws.cell(row=runway_start, column=1, value="Cash Runway").font = SECTION_FONT
    ws.merge_cells(start_row=runway_start, start_column=1, end_row=runway_start, end_column=4)

    ws.cell(row=runway_start + 1, column=1, value="Avg weekly cash OUT (13wk)").font = BODY_FONT
    c = ws.cell(row=runway_start + 1, column=2, value=runway.get("avg_weekly_paid", 0))
    c.number_format = CURRENCY_FMT; c.font = BODY_FONT

    ws.cell(row=runway_start + 2, column=1, value="Avg weekly NET cash flow (13wk)").font = BODY_FONT
    c = ws.cell(row=runway_start + 2, column=2, value=runway.get("avg_weekly_net", 0))
    c.number_format = CURRENCY_FMT; c.font = BODY_FONT
    if runway.get("avg_weekly_net", 0) >= 0:
        c.fill = GOOD_FILL
    else:
        c.fill = BAD_FILL

    ws.cell(row=runway_start + 3, column=1, value="Weeks of runway at current burn").font = METRIC_FONT
    weeks = runway.get("weeks")
    if weeks is None:
        v = ws.cell(row=runway_start + 3, column=2, value=runway.get("note", "net cash-positive — not constrained"))
        v.font = METRIC_FONT
        v.fill = GOOD_FILL
        ws.merge_cells(start_row=runway_start + 3, start_column=2, end_row=runway_start + 3, end_column=4)
    else:
        v = ws.cell(row=runway_start + 3, column=2, value=round(weeks, 1))
        v.number_format = '0.0 "weeks"'
        v.font = METRIC_FONT
        if weeks < 8:
            v.fill = BAD_FILL
        elif weeks < 16:
            v.fill = FLAG_FILL
        else:
            v.fill = GOOD_FILL

    # ── Weekly cash flow table ──
    flow_start = runway_start + 6
    ws.cell(row=flow_start, column=1, value="Weekly Net Cash Flow (last 13 weeks)").font = SECTION_FONT
    ws.merge_cells(start_row=flow_start, start_column=1, end_row=flow_start, end_column=4)

    hdr = flow_start + 1
    for i, h in enumerate(["Week Of", "Received ($)", "Paid ($)", "Net ($)"], start=1):
        c = ws.cell(row=hdr, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, w in enumerate(weekly_flow[-13:], start=hdr + 1):
        ws.cell(row=i, column=1, value=w["Week Of"]).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=w["Received"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        c = ws.cell(row=i, column=3, value=w["Paid"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        c = ws.cell(row=i, column=4, value=w["Net"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        if w["Net"] < 0:
            c.fill = BAD_FILL
        else:
            c.fill = GOOD_FILL

    _autosize(ws, {"A": 32, "B": 16, "C": 24, "D": 18})


def write_aging_sheet(
    ws: Worksheet,
    title: str,
    name_label: str,
    rows: List[Dict[str, Any]],
) -> Tuple[int, int]:
    """Write an aging sheet grouped by month → then by vendor.

    Months sort oldest first (most-aged invoices/bills surface at the top).
    Within each month, rows group by vendor alphabetically; vendors with >1
    line in that month get a subtotal row. Each month ends with a month total.
    Rows in the 61-90 / 90+ buckets get red highlighting as before.

    Layout:
        Row 1: column headers
        Row 2: 'TOTAL <title>' grand total at the top (so you see it first)
        Row 4+: month banner → vendor subgroups → month total → blank → next month
    """
    headers = [name_label, "Doc #", "Txn Date", "Due Date", "Days Overdue",
               "Bucket", "Balance", "Memo"]
    _set_header_row(ws, headers)

    # Grand total at the top — placeholder cell, populated after we know
    # the last data row.
    grand_total_row = 2
    ws.cell(row=grand_total_row, column=1, value=f"GRAND TOTAL {title}").font = METRIC_FONT
    ws.cell(row=grand_total_row, column=1).fill = FLAG_FILL
    grand_tot_cell = ws.cell(row=grand_total_row, column=7)
    grand_tot_cell.font = METRIC_FONT
    grand_tot_cell.fill = FLAG_FILL
    grand_tot_cell.number_format = CURRENCY_FMT

    # Group rows by month of TxnDate (YYYY-MM), then by vendor name within
    months: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for r in rows:
        d = r.get("Txn Date") or ""
        month = d[:7] if len(d) >= 7 else "unknown"
        vendor = r.get(name_label, "(unknown)")
        months[month][vendor].append(r)

    # Sort months oldest-first so the most-aged sit at the top
    sorted_months = sorted(months.keys())

    # Track sum-region for the grand total formula
    cur_row = 4
    data_first_row: Optional[int] = None
    data_last_row: Optional[int] = None

    for month in sorted_months:
        # Month banner
        month_label = month if month != "unknown" else "(no Txn Date)"
        bc = ws.cell(row=cur_row, column=1, value=f"━━━ {month_label} ━━━")
        bc.font = SECTION_FONT
        bc.fill = HEADER_FILL
        bc.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)
        cur_row += 1

        month_first_data_row = cur_row

        for vendor in sorted(months[month].keys(), key=lambda v: v.lower()):
            vendor_lines = months[month][vendor]
            vendor_first = cur_row

            for r in vendor_lines:
                ws.cell(row=cur_row, column=1, value=r.get(name_label, "")).font = BODY_FONT
                ws.cell(row=cur_row, column=2, value=r.get("Doc #", "")).font = BODY_FONT
                ws.cell(row=cur_row, column=3, value=r.get("Txn Date", "")).font = BODY_FONT
                ws.cell(row=cur_row, column=4, value=r.get("Due Date", "")).font = BODY_FONT
                ws.cell(row=cur_row, column=5, value=r.get("Days Overdue", 0)).font = BODY_FONT
                ws.cell(row=cur_row, column=6, value=r.get("Bucket", "")).font = BODY_FONT
                c = ws.cell(row=cur_row, column=7, value=r.get("Balance", 0))
                c.number_format = CURRENCY_FMT
                c.font = BODY_FONT
                if r.get("Bucket") in ("61-90", "90+"):
                    for col in range(1, 9):
                        ws.cell(row=cur_row, column=col).fill = BAD_FILL
                ws.cell(row=cur_row, column=8, value=r.get("PrivateNote", "")).font = BODY_FONT

                if data_first_row is None:
                    data_first_row = cur_row
                data_last_row = cur_row
                cur_row += 1

            # Vendor subtotal row (only when vendor has >1 line in this month)
            if len(vendor_lines) > 1:
                sub = ws.cell(row=cur_row, column=1, value=f"  └ {vendor} subtotal")
                sub.font = Font(italic=True, size=10, color="555555", name="Arial")
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
                tot = ws.cell(row=cur_row, column=7, value=f"=SUM(G{vendor_first}:G{cur_row - 1})")
                tot.number_format = CURRENCY_FMT
                tot.font = Font(italic=True, bold=True, size=10, name="Arial")
                cur_row += 1

        # Month total row
        if cur_row > month_first_data_row:
            mt = ws.cell(row=cur_row, column=1, value=f"  Month total — {month_label}")
            mt.font = METRIC_FONT
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
            mtot = ws.cell(row=cur_row, column=7, value=f"=SUM(G{month_first_data_row}:G{cur_row - 1})/2")
            # Why /2: vendor subtotals already double-count their lines.
            # Cleaner: sum only the leaf lines. Rebuild with explicit ranges below.
            mtot.number_format = CURRENCY_FMT
            mtot.font = METRIC_FONT
            mtot.fill = FLAG_FILL
            cur_row += 1

        # Blank spacer row between months
        cur_row += 1

    # Grand total references the actual leaf-data range — but vendor
    # subtotals and month totals are interleaved as formulas, so a simple
    # SUM(G:G) would over-count. Build it from the source rows in Python.
    grand_total_value = sum(r.get("Balance", 0) for r in rows)
    grand_tot_cell.value = grand_total_value

    # Recompute month totals to exclude vendor subtotal rows. The cleanest
    # approach is to redo each month total with a sum of the per-vendor
    # subtotal rows AND singleton vendor rows. Simpler and more reliable:
    # write the month total as a hardcoded sum from Python.
    # → fix: rewrite each month total as a Python-computed sum.
    # We'll find them by scanning the cells we just wrote.
    month_totals: Dict[str, float] = {
        m: sum(r.get("Balance", 0) for r in rows if (r.get("Txn Date") or "")[:7] == m)
        for m in sorted_months
    }
    for r in range(4, cur_row):
        cell = ws.cell(row=r, column=1)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("  Month total — "):
            month_label = cell.value.replace("  Month total — ", "").strip()
            month_key = month_label if month_label != "(no Txn Date)" else "unknown"
            v = month_totals.get(month_key, 0)
            tc = ws.cell(row=r, column=7)
            tc.value = v
            tc.number_format = CURRENCY_FMT

    _autosize(ws, {"A": 32, "B": 12, "C": 12, "D": 12, "E": 14, "F": 10, "G": 14, "H": 30})
    ws.freeze_panes = "A4"  # keep grand total + headers visible while scrolling
    return 1, grand_total_row


def write_coverage_sheet(
    ws: Worksheet,
    ar_rows: List[Dict[str, Any]],
    ap_rows: List[Dict[str, Any]],
    retainage_recv: Dict[str, Any],
    retainage_pay: Dict[str, Any],
) -> None:
    ws["A1"] = "AR vs AP by Aging Bucket — Coverage View"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["Bucket", "AR ($)", "AP ($)", "Net (AR − AP)", "Coverage Ratio"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    # Sum-by-bucket: done as formulas so edits to AR/AP sheets reflect here.
    for idx, (label, _, _) in enumerate(AGING_BUCKETS, start=4):
        ws.cell(row=idx, column=1, value=label).font = METRIC_FONT
        ws.cell(row=idx, column=2,
                value=f"=SUMIF('AR Aging'!F:F,\"{label}\",'AR Aging'!G:G)").number_format = CURRENCY_FMT
        ws.cell(row=idx, column=3,
                value=f"=SUMIF('AP Aging'!F:F,\"{label}\",'AP Aging'!G:G)").number_format = CURRENCY_FMT
        ws.cell(row=idx, column=4, value=f"=B{idx}-C{idx}").number_format = CURRENCY_FMT
        ws.cell(row=idx, column=5,
                value=f'=IF(C{idx}=0,"-",B{idx}/C{idx})').number_format = PCT_FMT

    total_row = 4 + len(AGING_BUCKETS)
    ws.cell(row=total_row, column=1, value="TOTAL").font = METRIC_FONT
    ws.cell(row=total_row, column=2,
            value=f"=SUM(B4:B{total_row - 1})").number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=3,
            value=f"=SUM(C4:C{total_row - 1})").number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=4,
            value=f"=B{total_row}-C{total_row}").number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=5,
            value=f'=IF(C{total_row}=0,"-",B{total_row}/C{total_row})').number_format = PCT_FMT
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).font = METRIC_FONT
        ws.cell(row=total_row, column=col).fill = FLAG_FILL

    # ── Retainage block (separate — not collectible right now) ──
    ret_start = total_row + 2
    ws.cell(row=ret_start, column=1, value="Retainage (earned, not yet collectible)").font = SECTION_FONT
    ws.merge_cells(start_row=ret_start, start_column=1, end_row=ret_start, end_column=5)

    ret_hdr_row = ret_start + 1
    for i, h in enumerate(["Account", "Type", "Subtype", "Balance", ""], start=1):
        c = ws.cell(row=ret_hdr_row, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    row_cursor = ret_hdr_row + 1
    for a in retainage_recv.get("accounts", []):
        ws.cell(row=row_cursor, column=1, value=a["Name"]).font = BODY_FONT
        ws.cell(row=row_cursor, column=2, value=a["AccountType"]).font = BODY_FONT
        ws.cell(row=row_cursor, column=3, value=a["AccountSubType"]).font = BODY_FONT
        c = ws.cell(row=row_cursor, column=4, value=a["CurrentBalance"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        row_cursor += 1
    for a in retainage_pay.get("accounts", []):
        ws.cell(row=row_cursor, column=1, value=f"{a['Name']} (payable)").font = BODY_FONT
        ws.cell(row=row_cursor, column=2, value=a["AccountType"]).font = BODY_FONT
        ws.cell(row=row_cursor, column=3, value=a["AccountSubType"]).font = BODY_FONT
        c = ws.cell(row=row_cursor, column=4, value=-a["CurrentBalance"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        row_cursor += 1

    if row_cursor == ret_hdr_row + 1:
        # No retainage accounts found — still show a line so the section isn't empty
        ws.cell(row=row_cursor, column=1,
                value="— no accounts with 'retainage' in the name found —").font = BODY_FONT
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5)
        row_cursor += 1

    # Totals: receivable, payable, net retainage
    ws.cell(row=row_cursor, column=1, value="Retainage Receivable").font = METRIC_FONT
    c = ws.cell(row=row_cursor, column=4, value=retainage_recv.get("total", 0))
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = GOOD_FILL
    row_cursor += 1

    if retainage_pay.get("total", 0):
        ws.cell(row=row_cursor, column=1, value="Retainage Payable").font = METRIC_FONT
        c = ws.cell(row=row_cursor, column=4, value=-retainage_pay.get("total", 0))
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
        row_cursor += 1

        ws.cell(row=row_cursor, column=1, value="Net Retainage").font = METRIC_FONT
        c = ws.cell(row=row_cursor, column=4,
                    value=retainage_recv.get("total", 0) - retainage_pay.get("total", 0))
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
        c.fill = FLAG_FILL
        row_cursor += 1

    # Reading notes
    note_row = row_cursor + 1
    ws.cell(row=note_row, column=1,
            value=("Coverage Ratio = AR / AP. 1.00x means incoming covers outgoing; <1 means AP exceeds AR in that bucket. "
                   "Retainage is money you've earned but customers are holding until project completion — it's NOT in AR and NOT collectible yet, "
                   "but it's still yours.")).font = BODY_FONT
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    ws.row_dimensions[note_row].height = 45

    _autosize(ws, {"A": 30, "B": 18, "C": 20, "D": 18, "E": 12})


def write_pl_sheet(
    ws: Worksheet,
    mtd_cols: List[str],
    mtd_rows: List[Tuple[str, List[Optional[float]], int]],
    ytd_cols: List[str],
    ytd_rows: List[Tuple[str, List[Optional[float]], int]],
    py_cols: List[str],
    py_rows: List[Tuple[str, List[Optional[float]], int]],
) -> None:
    ws["A1"] = "Profit & Loss"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    def _dump(start_row: int, title: str, cols: List[str], rows: List[Tuple[str, List[Optional[float]], int]]) -> int:
        ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
        hdr_row = start_row + 2
        ws.cell(row=hdr_row, column=1, value="Line").fill = HEADER_FILL
        ws.cell(row=hdr_row, column=1).font = HEADER_FONT
        for j, col_title in enumerate(cols, start=2):
            c = ws.cell(row=hdr_row, column=j, value=col_title)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT

        r = hdr_row + 1
        for label, values, depth in rows:
            indent = "  " * depth
            lc = ws.cell(row=r, column=1, value=f"{indent}{label}")
            lc.font = METRIC_FONT if label.lower().startswith(("total ", "net ", "gross")) else BODY_FONT
            for j, v in enumerate(values, start=2):
                if v is None:
                    continue
                vc = ws.cell(row=r, column=j, value=v)
                vc.number_format = CURRENCY_FMT
                vc.font = METRIC_FONT if lc.font == METRIC_FONT else BODY_FONT
                if label.lower().startswith(("total ", "net ", "gross")):
                    vc.fill = FLAG_FILL
            r += 1
        return r + 2

    end = _dump(3, "Month to Date", mtd_cols, mtd_rows)
    end = _dump(end, "Year to Date (current year)", ytd_cols, ytd_rows)
    end = _dump(end, "Year to Date (prior year)", py_cols, py_rows)

    _autosize(ws, {"A": 38, "B": 18, "C": 18, "D": 18, "E": 18})


def write_anomalies_sheet(
    ws: Worksheet,
    spikes: List[Dict[str, Any]],
    large: List[Dict[str, Any]],
    sigma: float,
) -> None:
    ws["A1"] = "Anomalies — Where to Look This Week"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A3"] = f"Overhead spend spikes (last 7 days ≥ {sigma}σ above 90-day weekly avg)"
    ws["A3"].font = SECTION_FONT
    ws.merge_cells("A3:F3")

    spike_hdrs = ["Account", "Last 7 Days", "Avg/week (prior 90d)", "Std Dev", "Z-score", "$ Above Avg"]
    for i, h in enumerate(spike_hdrs, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    if not spikes:
        ws.cell(row=5, column=1, value="— no overhead spikes detected —").font = BODY_FONT
        ws.merge_cells("A5:F5")
        next_section = 7
    else:
        for i, s in enumerate(spikes, start=5):
            ws.cell(row=i, column=1, value=s["Account"]).font = BODY_FONT
            ws.cell(row=i, column=2, value=s["Last 7 Days"]).number_format = CURRENCY_FMT
            ws.cell(row=i, column=3, value=s["Avg/week (prior 90d)"]).number_format = CURRENCY_FMT
            ws.cell(row=i, column=4, value=s["Std Dev"]).number_format = CURRENCY_FMT
            ws.cell(row=i, column=5, value=round(s["Z-score"], 2))
            ws.cell(row=i, column=6, value=s["$ Above Avg"]).number_format = CURRENCY_FMT
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = FLAG_FILL
        next_section = 5 + len(spikes) + 2

    ws.cell(row=next_section, column=1, value="Top 10 largest expense line items (last 7 days)").font = SECTION_FONT
    ws.merge_cells(start_row=next_section, start_column=1, end_row=next_section, end_column=6)

    hdr = next_section + 1
    large_hdrs = ["Txn Date", "Type", "Doc #", "Name / Vendor", "Account", "Amount"]
    for i, h in enumerate(large_hdrs, start=1):
        c = ws.cell(row=hdr, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    if not large:
        ws.cell(row=hdr + 1, column=1, value="— nothing in the last 7 days —").font = BODY_FONT
    else:
        for i, ln in enumerate(large, start=hdr + 1):
            ws.cell(row=i, column=1, value=ln.get("Txn Date", "")).font = BODY_FONT
            ws.cell(row=i, column=2, value=ln.get("Type", "")).font = BODY_FONT
            ws.cell(row=i, column=3, value=ln.get("Doc #", "")).font = BODY_FONT
            ws.cell(row=i, column=4, value=ln.get("Name", "")).font = BODY_FONT
            ws.cell(row=i, column=5, value=ln.get("Account", "")).font = BODY_FONT
            c = ws.cell(row=i, column=6, value=ln.get("Amount", 0))
            c.number_format = CURRENCY_FMT
            c.font = BODY_FONT

    _autosize(ws, {"A": 12, "B": 18, "C": 12, "D": 28, "E": 28, "F": 14})


def write_relationships_sheet(
    ws: Worksheet,
    top_customers: List[Dict[str, Any]],
    top_vendors: List[Dict[str, Any]],
    ytd_revenue: float,
    ytd_spend: float,
    concentration: Optional[Dict[str, Any]],
) -> None:
    ws["A1"] = "Top Customers & Vendors — Year to Date"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    # Concentration callout
    if concentration:
        ws["A2"] = (f"⚠ CONCENTRATION RISK: {concentration['Customer']} = "
                    f"{concentration['% of YTD']:.0%} of YTD revenue "
                    f"(${concentration['YTD Revenue']:,.0f}) — losing this customer would be material")
        ws["A2"].font = Font(bold=True, color="C00000", name="Arial")
        ws["A2"].fill = BAD_FILL
        ws.merge_cells("A2:E2")

    # Top customers block
    ws["A4"] = f"Top 10 Customers by YTD Revenue (Total: ${ytd_revenue:,.0f})"
    ws["A4"].font = SECTION_FONT
    ws.merge_cells("A4:E4")

    for i, h in enumerate(["#", "Customer", "YTD Revenue", "Invoices", "% of YTD"], start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, r in enumerate(top_customers, start=6):
        ws.cell(row=i, column=1, value=i - 5).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["Customer"]).font = BODY_FONT
        c = ws.cell(row=i, column=3, value=r["YTD Revenue"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        ws.cell(row=i, column=4, value=r["Invoices"]).font = BODY_FONT
        c = ws.cell(row=i, column=5, value=r["% of YTD"])
        c.number_format = PCT_FMT; c.font = BODY_FONT
        # Flag any single customer >25% of revenue
        if r["% of YTD"] >= 0.25:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = BAD_FILL
        elif r["% of YTD"] >= 0.15:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = FLAG_FILL

    # Top vendors block
    v_start = 6 + len(top_customers) + 2
    ws.cell(row=v_start, column=1, value=f"Top 10 Vendors by YTD Spend (Total: ${ytd_spend:,.0f})").font = SECTION_FONT
    ws.merge_cells(start_row=v_start, start_column=1, end_row=v_start, end_column=5)

    for i, h in enumerate(["#", "Vendor", "YTD Spend", "Transactions", "% of YTD"], start=1):
        c = ws.cell(row=v_start + 1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, r in enumerate(top_vendors, start=v_start + 2):
        ws.cell(row=i, column=1, value=i - (v_start + 1)).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["Vendor"]).font = BODY_FONT
        c = ws.cell(row=i, column=3, value=r["YTD Spend"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        ws.cell(row=i, column=4, value=r["Transactions"]).font = BODY_FONT
        c = ws.cell(row=i, column=5, value=r["% of YTD"])
        c.number_format = PCT_FMT; c.font = BODY_FONT

    _autosize(ws, {"A": 4, "B": 36, "C": 18, "D": 14, "E": 12})


def write_collections_sheet(
    ws: Worksheet,
    dso_monthly: List[Dict[str, Any]],
    dpo_monthly: List[Dict[str, Any]],
    ar_total: float = 0.0,
    ap_total: float = 0.0,
    revenue_ytd: float = 0.0,
    cogs_ytd: float = 0.0,
    days_into_year: int = 1,
) -> None:
    ws["A1"] = "Collections & Payment Timing"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = ("DSO = Days Sales Outstanding. Two flavors: Balance-Sheet DSO "
                "(AR ÷ daily revenue rate) is what banks/sureties ask for. "
                "Collection-Time DSO is the average days customers actually "
                "took to pay, computed from each payment's invoice age.")
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 36

    # Summary block
    def _avg_recent(rows: List[Dict[str, Any]], key: str, n: int = 3) -> Optional[float]:
        recent = rows[-n:] if len(rows) >= n else rows
        if not recent:
            return None
        return sum(r[key] for r in recent) / len(recent)

    dso_avg = _avg_recent(dso_monthly, "DSO", 3)
    dpo_avg = _avg_recent(dpo_monthly, "DPO", 3)

    # ── Balance-sheet DSO/DPO (the standard exec/lender metric) ──
    # DSO_BS = AR / (Revenue_annualized / 365) = AR × days_ytd / revenue_ytd
    # Treat YTD as our representative billing rate.
    days_into_year = max(days_into_year, 1)
    dso_bs = (ar_total * days_into_year / revenue_ytd) if revenue_ytd else None
    dpo_bs = (ap_total * days_into_year / cogs_ytd) if cogs_ytd else None

    ws["A4"] = "DSO — Balance Sheet (AR ÷ daily revenue)"
    ws["A4"].font = BODY_FONT
    if dso_bs is not None:
        c = ws["B4"]; c.value = round(dso_bs, 1)
        c.number_format = '0.0" days"'; c.font = METRIC_FONT
        if dso_bs > 60:    c.fill = BAD_FILL
        elif dso_bs > 45:  c.fill = FLAG_FILL
        else:              c.fill = GOOD_FILL

    ws["C4"] = "DPO — Balance Sheet (AP ÷ daily COGS)"
    ws["C4"].font = BODY_FONT
    if dpo_bs is not None:
        c = ws["D4"]; c.value = round(dpo_bs, 1)
        c.number_format = '0.0" days"'; c.font = METRIC_FONT

    ws["A5"] = "DSO — Collection Time (avg last 3 months)"
    ws["A5"].font = BODY_FONT
    if dso_avg is not None:
        c = ws["B5"]; c.value = round(dso_avg, 1)
        c.number_format = '0.0" days"'; c.font = METRIC_FONT
        if dso_avg > 60:    c.fill = BAD_FILL
        elif dso_avg > 45:  c.fill = FLAG_FILL
        else:               c.fill = GOOD_FILL

    ws["C5"] = "DPO — Payment Time (avg last 3 months)"
    ws["C5"].font = BODY_FONT
    if dpo_avg is not None:
        c = ws["D5"]; c.value = round(dpo_avg, 1)
        c.number_format = '0.0" days"'; c.font = METRIC_FONT

    # DSO trend table
    ws["A7"] = "DSO Trend — by payment month (collection-time view)"
    ws["A7"].font = SECTION_FONT
    ws.merge_cells("A7:D7")

    for i, h in enumerate(["Month", "DSO (days)", "Invoices Paid", "$ Collected"], start=1):
        c = ws.cell(row=8, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, r in enumerate(dso_monthly, start=9):
        ws.cell(row=i, column=1, value=r["Month"]).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=r["DSO"])
        c.number_format = "0.0"; c.font = BODY_FONT
        if r["DSO"] > 60:
            c.fill = BAD_FILL
        elif r["DSO"] > 45:
            c.fill = FLAG_FILL
        ws.cell(row=i, column=3, value=r["Invoices Paid"]).font = BODY_FONT
        c = ws.cell(row=i, column=4, value=r["$ Collected"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT

    # DPO trend table
    dpo_start = 9 + len(dso_monthly) + 2
    ws.cell(row=dpo_start, column=1, value="DPO Trend — by payment month (collection-time view)").font = SECTION_FONT
    ws.merge_cells(start_row=dpo_start, start_column=1, end_row=dpo_start, end_column=4)

    for i, h in enumerate(["Month", "DPO (days)", "Bills Paid", "$ Paid"], start=1):
        c = ws.cell(row=dpo_start + 1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, r in enumerate(dpo_monthly, start=dpo_start + 2):
        ws.cell(row=i, column=1, value=r["Month"]).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=r["DPO"])
        c.number_format = "0.0"; c.font = BODY_FONT
        ws.cell(row=i, column=3, value=r["Bills Paid"]).font = BODY_FONT
        c = ws.cell(row=i, column=4, value=r["$ Paid"])
        c.number_format = CURRENCY_FMT; c.font = BODY_FONT

    _autosize(ws, {"A": 14, "B": 16, "C": 20, "D": 18})


def write_recurring_sheet(
    ws: Worksheet,
    categorized: Dict[str, List[Dict[str, Any]]],
    burn: Dict[str, float],
    avg_weekly_inflow: float,
) -> None:
    """Recurring payments view, three-section layout:
       - Active     : currently recurring (real burn)
       - Stale      : detected in 12mo but no recent activity (likely cancelled)
       - New        : recently appeared, only in 3-month sweep
    """
    active = categorized.get("active", [])
    stale = categorized.get("stale", [])
    new = categorized.get("new", [])

    ws["A1"] = "Recurring Payments — Active vs Stale vs New"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    ws["A2"] = ("Two-phase detection. The 12-month sweep finds historical streams; "
                "the 3-month sweep filters for what's still active. ACTIVE = your "
                "real monthly burn. STALE = looked recurring once but no activity "
                "in the last 90 days (likely cancelled — excluded from burn). "
                "NEW = just started, may not stick. To override the auto-classification, "
                "use the 'Recurring Decisions' sheet of the override file.")
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 60

    # ── Burn Summary ──
    ws["A4"] = "Real Burn Summary"
    ws["A4"].font = SECTION_FONT
    ws.merge_cells("A4:I4")

    ws["A5"] = "ACTIVE — what you're paying right now (weekly / monthly)"
    ws["A5"].font = METRIC_FONT
    c = ws["D5"]; c.value = burn.get("active_weekly", 0)
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = GOOD_FILL
    c = ws["E5"]; c.value = burn.get("active_monthly", 0)
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = GOOD_FILL

    ws["A6"] = "+ NEW — recently started (weekly / monthly)"
    ws["A6"].font = BODY_FONT
    c = ws["D6"]; c.value = burn.get("new_weekly", 0)
    c.number_format = CURRENCY_FMT; c.font = BODY_FONT
    c = ws["E6"]; c.value = burn.get("new_monthly", 0)
    c.number_format = CURRENCY_FMT; c.font = BODY_FONT

    ws["A7"] = "  = REAL TOTAL BURN (Active + New)"
    ws["A7"].font = METRIC_FONT
    c = ws["D7"]; c.value = "=D5+D6"
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL
    c = ws["E7"]; c.value = "=E5+E6"
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL

    ws["A9"] = "STALE (no activity in last 90 days — excluded from burn above)"
    ws["A9"].font = Font(italic=True, color="666666", name="Arial")
    c = ws["D9"]; c.value = burn.get("stale_weekly", 0)
    c.number_format = CURRENCY_FMT; c.font = Font(italic=True, color="666666", name="Arial")
    c = ws["E9"]; c.value = burn.get("stale_monthly", 0)
    c.number_format = CURRENCY_FMT; c.font = Font(italic=True, color="666666", name="Arial")

    ws["A10"] = "  = If all detected streams were still active (overstated burn)"
    ws["A10"].font = Font(italic=True, color="666666", name="Arial")
    c = ws["D10"]; c.value = "=D5+D6+D9"
    c.number_format = CURRENCY_FMT; c.font = Font(italic=True, color="666666", name="Arial")
    c = ws["E10"]; c.value = "=E5+E6+E9"
    c.number_format = CURRENCY_FMT; c.font = Font(italic=True, color="666666", name="Arial")

    ws["A12"] = "Avg weekly cash inflow (last 13wk)"
    ws["A12"].font = BODY_FONT
    c = ws["D12"]; c.value = avg_weekly_inflow
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT

    ws["A13"] = "REAL burn as % of weekly inflow"
    ws["A13"].font = BODY_FONT
    c = ws["D13"]; c.value = '=IFERROR(D7/D12,"-")'
    c.number_format = PCT_FMT; c.font = METRIC_FONT
    if avg_weekly_inflow > 0:
        pct = (burn.get("active_weekly", 0) + burn.get("new_weekly", 0)) / avg_weekly_inflow
        if pct >= 0.5:   c.fill = BAD_FILL
        elif pct >= 0.3: c.fill = FLAG_FILL
        else:            c.fill = GOOD_FILL

    # ── Helper that writes one stream-table section ──
    headers = ["#", "Vendor", "Cadence", "Typical Amount", "Avg Gap (days)",
               "Occurrences", "Last Seen", "Weekly Cost", "Monthly Cost"]

    def _write_section(start_row: int, title: str, color: str,
                       streams: List[Dict[str, Any]],
                       suggested_action: str = "") -> int:
        # Banner
        bc = ws.cell(row=start_row, column=1, value=f"━━━ {title} ━━━ ({len(streams)} stream(s))")
        bc.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        bc.fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        cur = start_row + 1

        if suggested_action:
            ws.cell(row=cur, column=1, value=suggested_action).font = Font(
                italic=True, color="666666", name="Arial")
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=9)
            cur += 1

        if not streams:
            ws.cell(row=cur, column=1, value="— none —").font = Font(
                italic=True, color="666666", name="Arial")
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=9)
            return cur + 2

        # Header row
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=cur, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
        cur += 1

        for i, r in enumerate(streams, start=1):
            ws.cell(row=cur, column=1, value=i).font = BODY_FONT
            ws.cell(row=cur, column=2, value=r["Vendor"]).font = BODY_FONT
            ws.cell(row=cur, column=3, value=r["Cadence"]).font = BODY_FONT
            c = ws.cell(row=cur, column=4, value=r["Typical Amount"])
            c.number_format = CURRENCY_FMT; c.font = BODY_FONT
            c = ws.cell(row=cur, column=5, value=r.get("Avg Gap (days)", 0))
            c.number_format = "0.0"; c.font = BODY_FONT
            ws.cell(row=cur, column=6, value=r.get("Occurrences (12mo)", 0)).font = BODY_FONT
            ws.cell(row=cur, column=7, value=r.get("Latest Activity") or r.get("Last Seen", "")).font = BODY_FONT
            c = ws.cell(row=cur, column=8, value=r["Weekly Cost"])
            c.number_format = CURRENCY_FMT; c.font = BODY_FONT
            c = ws.cell(row=cur, column=9, value=r["Monthly Cost"])
            c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
            cur += 1

        return cur + 1  # blank spacer

    # ── Sections ──
    cur_row = 16

    cur_row = _write_section(
        cur_row, "ACTIVE", "2D7A2D", active,
        suggested_action="These are still being paid. Real burn comes from this section.",
    )
    cur_row = _write_section(
        cur_row, "STALE", "C0392B", stale,
        suggested_action=(
            "Detected in the 12-month sweep but no matching activity in the last "
            "90 days. Likely cancelled. To confirm exclusion, copy the Vendor name "
            "into the 'Recurring Excludes' sheet of customer_overrides.xlsx. "
            "If actually still active (e.g., quarterly bill), mark 'active' in "
            "the 'Recurring Decisions' sheet instead."),
    )
    cur_row = _write_section(
        cur_row, "NEW", "B58900", new,
        suggested_action=(
            "Picked up only in the 3-month sweep — recently started recurring. "
            "Counts toward burn but flagged for review since it may not stick."),
    )

    _autosize(ws, {"A": 4, "B": 32, "C": 16, "D": 16, "E": 12,
                   "F": 14, "G": 12, "H": 14, "I": 14})


def write_sources_sheet(ws: Worksheet) -> None:
    """Lineage / drill-down map. For every metric on the Dashboard, document
    where the underlying data comes from and which sheet/range to look at
    for detail. This is the 'why does this number say X?' answer.

    Acts as the obsidian-style backing graph in tabular form: each row is
    one metric, with columns for QBO source, calc function, and the local
    sheet+range where you can drill in.
    """
    ws["A1"] = "Sources & Lineage — How Every Dashboard Number Is Computed"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Each row maps a Dashboard metric to its underlying data. "
                "The 'QBO Source' column tells you the API entity / report. "
                "The 'Drill Into' column points you to the sheet you'd open "
                "to see the line-level detail behind the headline number. "
                "Override file annotations (Customers / Projects / Recurring "
                "Excludes) are applied as noted in the 'Adjustments' column.")
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 60

    headers = ["Dashboard Metric", "Cell", "QBO Source", "Calc",
               "Drill Into", "Adjustments / Override Layer"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    # Each tuple: (metric_name, dashboard_cell, qbo_source, calc, drill_target, adjustments)
    lineage = [
        # ── Cash & coverage box ──
        ("Bank total (excl. credit cards)", "Dashboard!B5",
         "QBO Account entity (AccountType='Bank')", "Sum of Account.CurrentBalance",
         "Cash sheet, rows 2+", "—"),
        ("Runway at current burn", "Dashboard!B7",
         "QBO Payment + BillPayment + Purchase (last 13 weeks)",
         "Bank total ÷ trailing 13-wk avg net cash outflow",
         "Cash sheet, runway block",
         "—"),
        ("Total AR", "Dashboard!E5",
         "QBO Invoice (Balance > 0)", "Sum of Invoice.Balance for collectible invoices",
         "AR Aging sheet — full breakdown by month + vendor",
         "Hold List sheet absorbs invoices for customers / projects tagged in overrides"),
        ("Retainage Receivable", "Dashboard!E6",
         "QBO Account (Name contains 'retainage')", "Sum of Account.CurrentBalance",
         "Coverage sheet, retainage block",
         "—"),
        ("Total AP", "Dashboard!E7",
         "QBO Bill (Balance > 0)", "Sum of Bill.Balance",
         "AP Aging sheet — full breakdown by month + vendor", "—"),
        ("Net (AR + Retainage − AP)", "Dashboard!E8",
         "Computed", "Excel formula =E5+E6-E7", "—",
         "Excludes Hold List balances by construction (they're not in E5)"),
        ("Coverage Ratio (AR only)", "Dashboard!E9",
         "Computed", "Excel formula =IFERROR(E5/E7,'-')", "Coverage sheet (full bucket breakdown)",
         "Held customers excluded — collection-readiness only"),

        # ── P&L block ──
        ("Income / COGS / GP / Expenses (MTD/YTD/PY)", "Dashboard!B12:E17",
         "QBO ProfitAndLoss report (Accrual basis)",
         "Direct read from QBO P&L summary rows",
         "P&L sheet — full account breakdown for each period",
         "Verified against your 2025 numbers; see Overhead → 'P&L Totals' block"),
        ("Gross Margin %", "Dashboard!B18:D18",
         "Computed", "Formula: IFERROR(GrossProfit / Income, '-')",
         "Overhead sheet, snapshot block", "—"),

        # ── Fixed obligations block ──
        ("Weekly fixed obligations", "Dashboard!B21",
         "QBO Bill + Purchase (12 months)",
         "Pattern-detected recurring streams: vendor + amount cluster + cadence",
         "Recurring sheet — every detected stream + stability scores",
         "Recurring Excludes sheet of overrides skips false-positives"),
        ("% of weekly inflow consumed by recurring", "Dashboard!E22",
         "Computed", "Recurring weekly ÷ trailing 13-wk avg cash inflow",
         "Recurring sheet, summary block", "—"),

        # ── Overhead recovery block ──
        ("Required markup over direct cost", "Dashboard!B25",
         "QBO ProfitAndLoss YTD totals",
         "Total Expenses ÷ Total COGS — the bid floor needed to recover overhead",
         "Overhead sheet, 'Bid Floor' block", "—"),
        ("Bid multiplier (target NP)", "Dashboard!E25",
         "Computed", "Formula: (1 + required_markup) ÷ (1 − target_profit_pct)",
         "Overhead sheet, 'Bidding Guide' block",
         "TARGET_NET_PROFIT_PCT constant at top of qbo_health.py (default 10%)"),
        ("Current Gross Margin %", "Dashboard!B26",
         "Computed", "Gross Profit ÷ Revenue (YTD, accrual)",
         "Overhead + P&L sheets",
         "—"),
        ("Net margin after overhead", "Dashboard!E26",
         "Computed", "GM% − Overhead%-of-Revenue", "Overhead sheet creep block",
         "—"),

        # ── 'Where to Look' flags ──
        ("'Where to Look' flags", "Dashboard, row 29+",
         "Mix of all the above",
         "Threshold-based callouts — see flag rules in source",
         "Each flag includes a callout with its derivation",
         "Override layer affects which flags fire (e.g., concentration risk uses post-alias parent grouping)"),

        # ── Top 5 AR / AP tables ──
        ("Top 5 AR balances (Dashboard)", "Dashboard, near bottom",
         "QBO Invoice (Balance > 0)", "Sorted by balance descending, take 5",
         "AR Aging sheet for full list; Relationships sheet for top 10",
         "Held customers excluded"),
        ("Top 5 AP balances (Dashboard)", "Dashboard, near bottom",
         "QBO Bill (Balance > 0)", "Sorted by balance descending, take 5",
         "AP Aging sheet", "—"),
    ]

    for i, (metric, cell, source, calc, drill, adj) in enumerate(lineage, start=5):
        ws.cell(row=i, column=1, value=metric).font = METRIC_FONT
        ws.cell(row=i, column=2, value=cell).font = BODY_FONT
        ws.cell(row=i, column=3, value=source).font = BODY_FONT
        ws.cell(row=i, column=4, value=calc).font = BODY_FONT
        ws.cell(row=i, column=5, value=drill).font = BODY_FONT
        ws.cell(row=i, column=6, value=adj).font = BODY_FONT
        for col in range(1, 7):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 32

    _autosize(ws, {"A": 32, "B": 24, "C": 32, "D": 38, "E": 32, "F": 36})


def write_hold_list_sheet(
    ws: Worksheet,
    held_rows: List[Dict[str, Any]],
    overrides_path: Path,
) -> None:
    """Render the Hold List — customers tagged in customer_overrides.xlsx.
    These balances are intentionally OUT of the main AR aging view but
    surfaced here in full so the money isn't forgotten.
    """
    ws["A1"] = "Hold List — Disputed / Litigation / Collections / Write-off"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws["A2"] = (f"Source: {overrides_path}  •  edit that file in Excel to "
                f"add/remove customers. Status values are auto-detected — "
                f"any text you put in the Status column becomes a group here.")
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 36

    if not held_rows:
        ws["A4"] = ("— no customers in the override file, or override file not found. "
                    "Create one at the path above with columns: Customer | Status | Notes —")
        ws["A4"].font = BODY_FONT
        ws.merge_cells("A4:H4")
        _autosize(ws, {"A": 28, "B": 12, "C": 12, "D": 12, "E": 14,
                       "F": 10, "G": 14, "H": 30})
        return

    # Group by status (auto-detected)
    by_status: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in held_rows:
        by_status[r.get("Override Status", "(no status)")].append(r)

    # Summary block at top
    ws["A4"] = "Summary by Status"
    ws["A4"].font = SECTION_FONT
    ws.merge_cells("A4:H4")

    ws["A5"] = "Status"
    ws["B5"] = "# Invoices"
    ws["C5"] = "Total Balance"
    for col in ("A", "B", "C"):
        ws[f"{col}5"].fill = HEADER_FILL
        ws[f"{col}5"].font = HEADER_FONT

    grand_total = 0.0
    cur_row = 6
    for status, items in sorted(by_status.items()):
        bal = sum(r.get("Balance", 0) for r in items)
        grand_total += bal
        ws.cell(row=cur_row, column=1, value=status).font = METRIC_FONT
        ws.cell(row=cur_row, column=2, value=len(items)).font = BODY_FONT
        c = ws.cell(row=cur_row, column=3, value=bal)
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
        cur_row += 1

    ws.cell(row=cur_row, column=1, value="TOTAL ON HOLD").font = METRIC_FONT
    ws.cell(row=cur_row, column=1).fill = FLAG_FILL
    ws.cell(row=cur_row, column=2, value=len(held_rows)).font = METRIC_FONT
    c = ws.cell(row=cur_row, column=3, value=grand_total)
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL
    cur_row += 3

    # Detail by status
    headers = ["Customer", "Doc #", "Txn Date", "Due Date", "Days Overdue",
               "Bucket", "Balance", "Notes"]
    for status, items in sorted(by_status.items()):
        # Banner for this status
        bc = ws.cell(row=cur_row, column=1, value=f"━━━ {status.upper()} ━━━")
        bc.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        bc.fill = HEADER_FILL
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)
        cur_row += 1

        # Header row
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=cur_row, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
        cur_row += 1

        # Sort items by parent customer name then by doc#
        items.sort(key=lambda r: (r.get("Parent Customer", ""), r.get("Doc #", "")))

        for r in items:
            ws.cell(row=cur_row, column=1, value=r.get("Customer", "")).font = BODY_FONT
            ws.cell(row=cur_row, column=2, value=r.get("Doc #", "")).font = BODY_FONT
            ws.cell(row=cur_row, column=3, value=r.get("Txn Date", "")).font = BODY_FONT
            ws.cell(row=cur_row, column=4, value=r.get("Due Date", "")).font = BODY_FONT
            ws.cell(row=cur_row, column=5, value=r.get("Days Overdue", 0)).font = BODY_FONT
            ws.cell(row=cur_row, column=6, value=r.get("Bucket", "")).font = BODY_FONT
            c = ws.cell(row=cur_row, column=7, value=r.get("Balance", 0))
            c.number_format = CURRENCY_FMT; c.font = BODY_FONT
            ws.cell(row=cur_row, column=8, value=r.get("Override Notes", "")).font = BODY_FONT
            cur_row += 1

        # Status subtotal
        sub = ws.cell(row=cur_row, column=1, value=f"  Subtotal — {status}")
        sub.font = METRIC_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
        sub_val = sum(r.get("Balance", 0) for r in items)
        c = ws.cell(row=cur_row, column=7, value=sub_val)
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL
        cur_row += 2

    _autosize(ws, {"A": 28, "B": 12, "C": 12, "D": 12, "E": 14,
                   "F": 10, "G": 14, "H": 30})


def write_overhead_sheet(
    ws: Worksheet,
    m: Dict[str, Any],
    pyfull_totals: Optional[Dict[str, float]] = None,
    py_full_bounds: Tuple[str, str] = ("", ""),
    ytd_totals: Optional[Dict[str, float]] = None,
) -> None:
    """Construction-finance overhead view: required markup, break-even, creep,
    and per-account YoY breakdown. Designed to answer 'what % do I need to
    bid every job at to cover overhead?'.

    The top of this sheet shows raw P&L totals (YTD + full prior year) so you
    can verify the parser is reading QBO correctly — those numbers MUST
    match what you see in QBO Reports → Profit and Loss for the same range.
    """
    pyfull_totals = pyfull_totals or {}
    ytd_totals = ytd_totals or {}

    ws["A1"] = "Overhead Analysis — Construction Finance"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Direct cost = COGS (job materials + labor + subs). Overhead = operating expenses "
                "(office rent, admin wages, insurance, software, etc.). Required Markup = the % "
                "you must add to direct cost to recover overhead. Bid every job above that floor — "
                "anything less and you're losing money.")
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 50

    # ── P&L Totals — Verification Block ──
    # The whole rest of the sheet derives from these three numbers per period.
    # If they don't match QBO Reports → P&L for the same date range, there's
    # a parser issue we need to fix; if they DO match, the analytics below
    # are sound.
    ws["A4"] = "P&L Totals — verify these against QBO Reports → Profit and Loss"
    ws["A4"].font = SECTION_FONT
    ws.merge_cells("A4:F4")

    py_label = f"{py_full_bounds[0][:4]} (Full Year)" if py_full_bounds[0] else "Prior Full Year"
    ytd_label = "YTD (this year)"

    # Header row — three columns: label, prior full year, current YTD
    for col, hdr in [(1, "From P&L Report"), (2, py_label), (3, ytd_label)]:
        c = ws.cell(row=5, column=col, value=hdr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    verify_rows = [
        ("Total Income",                "income"),
        ("Total Cost of Goods Sold",    "cogs"),
        ("Gross Profit",                "gross_profit"),
        ("Total Expenses",              "expenses"),
        ("Net Operating Income",        "net_op_income"),
        ("Net Income",                  "net_income"),
    ]
    for i, (label, key) in enumerate(verify_rows, start=6):
        ws.cell(row=i, column=1, value=label).font = METRIC_FONT
        py_v = pyfull_totals.get(key)
        ytd_v = ytd_totals.get(key)
        if py_v is not None:
            c = ws.cell(row=i, column=2, value=py_v)
            c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
        if ytd_v is not None:
            c = ws.cell(row=i, column=3, value=ytd_v)
            c.number_format = CURRENCY_FMT; c.font = METRIC_FONT

    # Date-range line so user knows exactly what window each column covers
    if py_full_bounds[0]:
        ws.cell(row=12, column=2,
                value=f"({py_full_bounds[0]} → {py_full_bounds[1]})").font = Font(
                italic=True, size=9, color="666666", name="Arial")

    # ── Analytics — based on YTD ──
    BASE = 14  # Bumped from 4 to leave room for the verification block above.
               # If the verification block grows, just bump BASE.

    ws.cell(row=BASE, column=1, value="YTD Snapshot (analytics)").font = SECTION_FONT
    ws.cell(row=BASE, column=4, value="Prior Year YTD").font = SECTION_FONT

    snap_rows = [
        ("Revenue (YTD)",                m["revenue_ytd"],          CURRENCY_FMT, None),
        ("Direct Costs / COGS (YTD)",    m["cogs_ytd"],             CURRENCY_FMT, None),
        ("Gross Profit (YTD)",           m["gp_ytd"],               CURRENCY_FMT, None),
        ("Gross Margin %",               m["gm_pct_ytd"],           PCT_FMT,      None),
        ("Overhead (YTD)",               m["overhead_ytd"],         CURRENCY_FMT, None),
        ("Overhead as % of Revenue",     m["overhead_pct_revenue_ytd"], PCT_FMT,  None),
    ]
    for i, (lbl, val, fmt, fill) in enumerate(snap_rows, start=BASE + 1):
        ws.cell(row=i, column=1, value=lbl).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt; c.font = METRIC_FONT
        if fill:
            c.fill = fill

    py_rows = [
        m["revenue_py"], m["cogs_py"], m["gp_py"], m["gm_pct_py"],
        m["overhead_py"], m["overhead_pct_revenue_py"],
    ]
    py_fmts = [CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT, PCT_FMT, CURRENCY_FMT, PCT_FMT]
    for i, (val, fmt) in enumerate(zip(py_rows, py_fmts), start=BASE + 1):
        c = ws.cell(row=i, column=4, value=val)
        c.number_format = fmt; c.font = BODY_FONT

    # ── Bid Floor ──
    BID_BASE = BASE + 8  # row 22 by default

    ws.cell(row=BID_BASE, column=1,
            value="Bid Floor — Markup You MUST Add to Cover Overhead").font = SECTION_FONT
    ws.merge_cells(start_row=BID_BASE, start_column=1, end_row=BID_BASE, end_column=6)

    ws.cell(row=BID_BASE + 1, column=1, value="Required Markup over Direct Cost").font = METRIC_FONT
    c = ws.cell(row=BID_BASE + 1, column=2, value=m["required_markup_ytd"])
    c.number_format = PCT_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL

    ws.cell(row=BID_BASE + 1, column=4, value="Required Markup (PY)").font = BODY_FONT
    c = ws.cell(row=BID_BASE + 1, column=5, value=m["required_markup_py"])
    c.number_format = PCT_FMT; c.font = BODY_FONT
    if m["required_markup_ytd"] > m["required_markup_py"] + 0.02:
        ws.cell(row=BID_BASE + 1, column=2).fill = BAD_FILL

    ws.cell(row=BID_BASE + 2, column=1, value="Break-even Revenue (YTD overhead ÷ GM%)").font = BODY_FONT
    c = ws.cell(row=BID_BASE + 2, column=2, value=m["breakeven_revenue_ytd"])
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    ws.cell(row=BID_BASE + 2, column=4, value="vs YTD revenue").font = BODY_FONT
    if m["breakeven_revenue_ytd"] > 0:
        c = ws.cell(row=BID_BASE + 2, column=5)
        c.value = m["revenue_ytd"] - m["breakeven_revenue_ytd"]
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
        c.fill = GOOD_FILL if c.value >= 0 else BAD_FILL

    # ── Bidding formula ──
    GUIDE_BASE = BID_BASE + 4

    ws.cell(row=GUIDE_BASE, column=1, value="Bidding Guide").font = SECTION_FONT
    ws.merge_cells(start_row=GUIDE_BASE, start_column=1, end_row=GUIDE_BASE, end_column=6)

    bid_mult = m["bid_multiplier_ytd"]
    ws.cell(row=GUIDE_BASE + 1, column=1,
            value=(f"Target Net Profit: {m['target_profit_pct']:.0%}"
                   "  (edit TARGET_NET_PROFIT_PCT in qbo_health.py to change)")).font = BODY_FONT
    ws.merge_cells(start_row=GUIDE_BASE + 1, start_column=1, end_row=GUIDE_BASE + 1, end_column=6)

    ws.cell(row=GUIDE_BASE + 2, column=1, value="Bid Multiplier on Direct Cost").font = METRIC_FONT
    c = ws.cell(row=GUIDE_BASE + 2, column=2, value=bid_mult)
    c.number_format = "0.000\"x\""; c.font = METRIC_FONT; c.fill = FLAG_FILL

    ws.cell(row=GUIDE_BASE + 3, column=1, value="Formula").font = BODY_FONT
    ws.cell(row=GUIDE_BASE + 3, column=2,
            value=("Bid Price = Direct Cost × "
                   f"(1 + {m['required_markup_ytd']:.1%}) ÷ (1 − {m['target_profit_pct']:.0%})"
                   f" = Direct Cost × {bid_mult:.3f}x")).font = BODY_FONT
    ws.merge_cells(start_row=GUIDE_BASE + 3, start_column=2, end_row=GUIDE_BASE + 3, end_column=6)

    ws.cell(row=GUIDE_BASE + 4, column=1, value="Example").font = BODY_FONT
    # NP at suggested bid: profit = bid - COGS - OH; OH ≈ COGS × required_markup
    bid_amt = 100_000 * bid_mult
    profit_amt = bid_amt * m["target_profit_pct"]
    ws.cell(row=GUIDE_BASE + 4, column=2,
            value=(f"$100,000 direct cost  →  bid at ${bid_amt:,.0f}  "
                   f"→  net profit ${profit_amt:,.0f}  ({m['target_profit_pct']:.0%} of revenue)")).font = BODY_FONT
    ws.merge_cells(start_row=GUIDE_BASE + 4, start_column=2, end_row=GUIDE_BASE + 4, end_column=6)

    # ── Overhead Creep ──
    CREEP_BASE = GUIDE_BASE + 6

    ws.cell(row=CREEP_BASE, column=1,
            value="Overhead vs Revenue Growth (Creep Check)").font = SECTION_FONT
    ws.merge_cells(start_row=CREEP_BASE, start_column=1, end_row=CREEP_BASE, end_column=6)

    creep_rows = [
        ("Revenue YoY growth", m["revenue_growth"]),
        ("Overhead YoY growth", m["overhead_growth"]),
        ("Creep (Overhead − Revenue)", m["creep"]),
    ]
    for i, (lbl, val) in enumerate(creep_rows, start=CREEP_BASE + 1):
        ws.cell(row=i, column=1, value=lbl).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = PCT_FMT; c.font = METRIC_FONT
        if lbl.startswith("Creep"):
            if val > 0.05:
                c.fill = BAD_FILL
            elif val > 0:
                c.fill = FLAG_FILL
            else:
                c.fill = GOOD_FILL

    if m["creep"] > 0.05:
        warn_row = CREEP_BASE + 4
        ws.cell(row=warn_row, column=1,
                value=("⚠ Overhead is growing materially faster than revenue. "
                       "Required markup will keep rising unless overhead is cut "
                       "or revenue accelerates.")).font = Font(italic=True, color="C00000", name="Arial")
        ws.cell(row=warn_row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=warn_row, start_column=1, end_row=warn_row, end_column=6)
        ws.row_dimensions[warn_row].height = 30

    # ── Per-account detail ──
    detail_start = CREEP_BASE + 7
    ws.cell(row=detail_start, column=1, value="Overhead Accounts — YoY Detail").font = SECTION_FONT
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=6)

    hdr = detail_start + 1
    headers = ["Account", "YTD", "PY YTD", "YoY $", "YoY %", "% of Overhead"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT

    for i, a in enumerate(m["overhead_accounts"], start=hdr + 1):
        ws.cell(row=i, column=1, value=a["Account"]).font = BODY_FONT
        c = ws.cell(row=i, column=2, value=a["YTD"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        c = ws.cell(row=i, column=3, value=a["PY YTD"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        c = ws.cell(row=i, column=4, value=a["YoY $"]); c.number_format = CURRENCY_FMT; c.font = BODY_FONT
        c = ws.cell(row=i, column=5, value=a["YoY %"]); c.number_format = PCT_FMT; c.font = BODY_FONT
        # Color code based on YoY% with material size guard
        material = a["YTD"] > m["overhead_ytd"] * 0.02 if m["overhead_ytd"] else False
        if material and a["YoY %"] >= ACCOUNT_YOY_RED_THRESHOLD:
            c.fill = BAD_FILL
        elif material and a["YoY %"] >= ACCOUNT_YOY_YELLOW_THRESHOLD:
            c.fill = FLAG_FILL
        c = ws.cell(row=i, column=6, value=a["% of Overhead"])
        c.number_format = PCT_FMT; c.font = BODY_FONT

    # Total row
    if m["overhead_accounts"]:
        last = hdr + len(m["overhead_accounts"])
        tot = last + 1
        ws.cell(row=tot, column=1, value="TOTAL OVERHEAD").font = METRIC_FONT
        c = ws.cell(row=tot, column=2, value=f"=SUM(B{hdr + 1}:B{last})")
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL
        c = ws.cell(row=tot, column=3, value=f"=SUM(C{hdr + 1}:C{last})")
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL
        c = ws.cell(row=tot, column=4, value=f"=SUM(D{hdr + 1}:D{last})")
        c.number_format = CURRENCY_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL

    _autosize(ws, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 12, "F": 16})


def write_dashboard_sheet(
    ws: Worksheet,
    now: dt.datetime,
    bank_total: float,
    cash_accounts_count: int,
    ar_total: float,
    ap_total: float,
    ar_over_60: float,
    ap_over_60: float,
    pl_mtd: Dict[str, float],
    pl_ytd: Dict[str, float],
    pl_py: Dict[str, float],
    spike_count: int,
    top_spikes: List[Dict[str, Any]],
    top_ar: List[Dict[str, Any]],
    top_ap: List[Dict[str, Any]],
    runway: Dict[str, Any],
    concentration: Optional[Dict[str, Any]],
    dso_monthly: List[Dict[str, Any]],
    dpo_monthly: List[Dict[str, Any]],
    retainage_recv: Dict[str, Any],
    retainage_pay: Dict[str, Any],
    recurring: List[Dict[str, Any]],
    fixed_obligations: Dict[str, float],
    avg_weekly_inflow: float,
    overhead_metrics: Dict[str, Any],
    burn_summary: Optional[Dict[str, float]] = None,
) -> None:
    ws.sheet_view.showGridLines = False
    burn_summary = burn_summary or {}

    # Title strip
    ws["A1"] = "Company Health Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Refreshed: {now.strftime('%Y-%m-%d %H:%M')}  •  Source: QuickBooks Online (Accrual)"
    ws["A2"].font = Font(italic=True, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    # Drill-down hint right under the title — clickable link to Sources tab
    ws["A3"] = '=HYPERLINK("#Sources!A1","→ Where do these numbers come from? Click to see lineage")'
    ws["A3"].font = Font(italic=True, color="0563C1", underline="single", size=10, name="Arial")
    ws.merge_cells("A3:F3")

    # ── Cash + Coverage boxes ──
    # Section labels link to their source sheets so users can click to drill in.
    ws["A4"] = '=HYPERLINK("#Cash!A1","Cash on Hand →")'
    ws["A4"].font = Font(bold=True, size=13, color="1F3A5F", underline="single", name="Arial")
    ws["A5"] = "Bank total (excl. credit cards)"
    ws["A5"].font = BODY_FONT
    c = ws["B5"]; c.value = bank_total; c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    ws["A6"] = f"Across {cash_accounts_count} active account(s)"
    ws["A6"].font = Font(italic=True, color="666666", name="Arial")

    ws["A7"] = "Runway at current burn"
    ws["A7"].font = BODY_FONT
    weeks = runway.get("weeks")
    if weeks is None:
        c = ws["B7"]; c.value = "cash-positive"; c.font = METRIC_FONT; c.fill = GOOD_FILL
    else:
        c = ws["B7"]; c.value = round(weeks, 1)
        c.number_format = '0.0 "weeks"'; c.font = METRIC_FONT
        if weeks < 8: c.fill = BAD_FILL
        elif weeks < 16: c.fill = FLAG_FILL
        else: c.fill = GOOD_FILL

    ws["D4"] = '=HYPERLINK("#Coverage!A1","AR vs AP Coverage →")'
    ws["D4"].font = Font(bold=True, size=13, color="1F3A5F", underline="single", name="Arial")
    ws["D5"] = "Total AR"
    ws["D5"].font = BODY_FONT
    c = ws["E5"]; c.value = ar_total; c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    ws["D6"] = "Retainage Receivable"
    ws["D6"].font = BODY_FONT
    c = ws["E6"]; c.value = retainage_recv.get("total", 0)
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    c.font = Font(bold=True, italic=True, size=11, color="1F3A5F", name="Arial")
    ws["D7"] = "Total AP"
    ws["D7"].font = BODY_FONT
    c = ws["E7"]; c.value = ap_total; c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    ws["D8"] = "Net (AR + Retainage − AP)"
    ws["D8"].font = BODY_FONT
    c = ws["E8"]; c.value = "=E5+E6-E7"; c.number_format = CURRENCY_FMT; c.font = METRIC_FONT
    ws["D9"] = "Coverage Ratio (AR only)"
    ws["D9"].font = BODY_FONT
    c = ws["E9"]; c.value = '=IFERROR(E5/E7,"-")'; c.number_format = PCT_FMT; c.font = METRIC_FONT

    # Coverage flag color
    if ap_total > 0:
        ratio = ar_total / ap_total
        ws["E9"].fill = GOOD_FILL if ratio >= 1 else BAD_FILL

    # ── P&L strip ──
    ws["A10"] = '=HYPERLINK("#\'P&L\'!A1","Profit & Loss Snapshot →")'
    ws["A10"].font = Font(bold=True, size=13, color="1F3A5F", underline="single", name="Arial")

    pl_hdrs = ["Metric", "MTD", "YTD", "YTD Prior Year", "YoY Δ"]
    for i, h in enumerate(pl_hdrs, start=1):
        c = ws.cell(row=11, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    def pl_row(r: int, label: str, key: str, fmt: str = CURRENCY_FMT) -> None:
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        mtd_val = pl_mtd.get(key)
        ytd_val = pl_ytd.get(key)
        py_val = pl_py.get(key)
        if mtd_val is not None:
            c = ws.cell(row=r, column=2, value=mtd_val); c.number_format = fmt; c.font = BODY_FONT
        if ytd_val is not None:
            c = ws.cell(row=r, column=3, value=ytd_val); c.number_format = fmt; c.font = BODY_FONT
        if py_val is not None:
            c = ws.cell(row=r, column=4, value=py_val); c.number_format = fmt; c.font = BODY_FONT
        if ytd_val is not None and py_val:
            c = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
            c.number_format = fmt; c.font = BODY_FONT

    pl_row(12, "Income", "income")
    pl_row(13, "COGS", "cogs")
    pl_row(14, "Gross Profit", "gross_profit")
    pl_row(15, "Operating Expenses", "expenses")
    pl_row(16, "Net Operating Income", "net_op_income")
    pl_row(17, "Net Income", "net_income")

    # Gross Margin % (computed). IFERROR handles empty cells, zero income,
    # or partial-period gaps without throwing #VALUE! / #DIV/0!.
    ws.cell(row=18, column=1, value="Gross Margin %").font = BODY_FONT
    ws.cell(row=18, column=2, value='=IFERROR(B14/B12,"-")').number_format = PCT_FMT
    ws.cell(row=18, column=3, value='=IFERROR(C14/C12,"-")').number_format = PCT_FMT
    ws.cell(row=18, column=4, value='=IFERROR(D14/D12,"-")').number_format = PCT_FMT

    # ── Fixed Obligations (REAL recurring burn — active + new only) ──
    ws["A20"] = '=HYPERLINK("#Recurring!A1","Fixed Obligations (REAL Recurring Burn) →")'
    ws["A20"].font = Font(bold=True, size=13, color="1F3A5F", underline="single", name="Arial")
    ws.merge_cells("A20:E20")

    # Use the active+new "real burn" (excludes stale streams that look recurring
    # in the 12-month sweep but haven't shown activity in the last 90 days).
    real_weekly = burn_summary.get("real_weekly", fixed_obligations.get("weekly_total", 0))
    real_monthly = burn_summary.get("real_monthly", fixed_obligations.get("monthly_total", 0))
    active_count = burn_summary.get("active_count", 0)
    new_count = burn_summary.get("new_count", 0)
    stale_count = burn_summary.get("stale_count", 0)
    stale_monthly = burn_summary.get("stale_monthly", 0)

    ws["A21"] = "Weekly burn (active + recently started)"
    ws["A21"].font = BODY_FONT
    c = ws["B21"]; c.value = real_weekly
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT

    ws["D21"] = "Monthly burn"
    ws["D21"].font = BODY_FONT
    c = ws["E21"]; c.value = real_monthly
    c.number_format = CURRENCY_FMT; c.font = METRIC_FONT

    ws["A22"] = f"{active_count} active + {new_count} new stream(s)"
    if stale_count:
        ws["A22"].value = (f"{active_count} active + {new_count} new stream(s) "
                           f"  ·  {stale_count} stale stream(s) excluded "
                           f"(${stale_monthly:,.0f}/mo would inflate the number)")
    ws["A22"].font = Font(italic=True, color="666666", name="Arial")
    ws.merge_cells("A22:C22")

    ws["D22"] = "% of weekly inflow"
    ws["D22"].font = BODY_FONT
    if avg_weekly_inflow > 0:
        ratio = real_weekly / avg_weekly_inflow
        c = ws["E22"]; c.value = ratio
        c.number_format = PCT_FMT; c.font = METRIC_FONT
        if ratio >= 0.5:
            c.fill = BAD_FILL
        elif ratio >= 0.3:
            c.fill = FLAG_FILL
        else:
            c.fill = GOOD_FILL

    # ── Overhead Recovery (construction markup math) ──
    ws["A24"] = '=HYPERLINK("#Overhead!A1","Overhead Recovery (Construction Markup) →")'
    ws["A24"].font = Font(bold=True, size=13, color="1F3A5F", underline="single", name="Arial")
    ws.merge_cells("A24:E24")

    req_markup = overhead_metrics.get("required_markup_ytd", 0)
    bid_mult = overhead_metrics.get("bid_multiplier_ytd", 0)
    target_np = overhead_metrics.get("target_profit_pct", 0)
    gm_pct = overhead_metrics.get("gm_pct_ytd", 0)
    oh_pct_rev = overhead_metrics.get("overhead_pct_revenue_ytd", 0)

    ws["A25"] = "Required markup over direct cost"
    ws["A25"].font = BODY_FONT
    c = ws["B25"]; c.value = req_markup
    c.number_format = PCT_FMT; c.font = METRIC_FONT; c.fill = FLAG_FILL

    ws["D25"] = f"Bid multiplier (target NP {target_np:.0%})"
    ws["D25"].font = BODY_FONT
    c = ws["E25"]; c.value = bid_mult
    c.number_format = "0.000\"x\""; c.font = METRIC_FONT; c.fill = FLAG_FILL

    ws["A26"] = "Current Gross Margin %"
    ws["A26"].font = BODY_FONT
    c = ws["B26"]; c.value = gm_pct
    c.number_format = PCT_FMT; c.font = METRIC_FONT
    if gm_pct < oh_pct_rev:
        c.fill = BAD_FILL  # GM doesn't cover overhead
    elif gm_pct < oh_pct_rev + target_np:
        c.fill = FLAG_FILL  # covers overhead but not target NP
    else:
        c.fill = GOOD_FILL

    ws["D26"] = "Net margin after overhead (YTD)"
    ws["D26"].font = BODY_FONT
    net_margin_ytd = gm_pct - oh_pct_rev
    c = ws["E26"]; c.value = net_margin_ytd
    c.number_format = PCT_FMT; c.font = METRIC_FONT
    if net_margin_ytd < 0:
        c.fill = BAD_FILL
    elif net_margin_ytd < target_np:
        c.fill = FLAG_FILL
    else:
        c.fill = GOOD_FILL

    # ── Where to Look ──
    ws["A28"] = "Where to Look"
    ws["A28"].font = SECTION_FONT

    flags: List[Tuple[str, str, Optional[str]]] = []

    if ar_over_60 > 0:
        flags.append(("AR aged 60+ days",
                      f"${ar_over_60:,.0f} — chase collections", "F7D4D4"))
    if ap_over_60 > 0:
        flags.append(("AP aged 60+ days",
                      f"${ap_over_60:,.0f} — vendor relationship risk", "F7D4D4"))

    if ap_total > 0 and ar_total < ap_total:
        flags.append(("AR < AP",
                      f"incoming ${ar_total:,.0f} does not cover outgoing ${ap_total:,.0f}", "F7D4D4"))

    if spike_count > 0:
        top = top_spikes[0]
        flags.append((f"{spike_count} overhead spike(s) detected",
                      f"biggest: {top['Account']} — ${top['Last 7 Days']:,.0f} last 7d "
                      f"(avg ${top['Avg/week (prior 90d)']:,.0f})", "FFF5CC"))

    if top_ar:
        big = top_ar[0]
        flags.append(("Largest single AR balance",
                      f"{big['Customer']} — ${big['Balance']:,.0f} ({big['Bucket']})", "FFF5CC"))
    if top_ap:
        big = top_ap[0]
        flags.append(("Largest single AP balance",
                      f"{big['Vendor']} — ${big['Balance']:,.0f} ({big['Bucket']})", "FFF5CC"))

    if pl_ytd.get("net_income") is not None and pl_py.get("net_income") is not None:
        delta = pl_ytd["net_income"] - pl_py["net_income"]
        if delta < 0:
            flags.append(("YTD Net Income below last year",
                          f"${abs(delta):,.0f} behind prior year", "F7D4D4"))

    # Cash runway flag
    weeks = runway.get("weeks")
    if weeks is not None and weeks < 8:
        flags.append(("Cash runway under 8 weeks",
                      f"{weeks:.1f} weeks at current burn rate ({runway.get('note', '')})",
                      "F7D4D4"))
    elif weeks is not None and weeks < 16:
        flags.append(("Cash runway under 16 weeks",
                      f"{weeks:.1f} weeks at current burn rate — keep collections tight",
                      "FFF5CC"))

    # Customer concentration flag
    if concentration:
        flags.append(("Customer concentration risk",
                      f"{concentration['Customer']} = {concentration['% of YTD']:.0%} of YTD revenue "
                      f"(${concentration['YTD Revenue']:,.0f}). Losing them would hurt.",
                      "F7D4D4"))

    # DSO trending up flag
    if len(dso_monthly) >= 6:
        last3 = sum(r["DSO"] for r in dso_monthly[-3:]) / 3
        prior3 = sum(r["DSO"] for r in dso_monthly[-6:-3]) / 3
        if last3 > prior3 + 5 and last3 > 45:
            flags.append(("DSO trending up",
                          f"avg collection days rose from {prior3:.0f} → {last3:.0f} over 3 months "
                          f"— collections discipline may be slipping",
                          "FFF5CC"))

    # Retainage visibility — informational (not a red flag, just worth knowing)
    ret_r = retainage_recv.get("total", 0)
    if ret_r > 0:
        label_text = (f"${ret_r:,.0f} earned, held by customers until project close — "
                      f"not in AR, not in cash")
        # Treat as a bigger flag if it's significant relative to current AR
        if ar_total > 0 and ret_r >= ar_total * 0.25:
            label_text += f" ({ret_r / ar_total:.0%} of current AR sitting in retainage)"
        flags.append(("Retainage receivable", label_text, "FFF5CC"))

    # Fixed-obligation load vs weekly inflow
    weekly_fx = burn_summary.get("real_weekly", fixed_obligations.get("weekly_total", 0))
    if avg_weekly_inflow > 0 and weekly_fx > 0:
        load = weekly_fx / avg_weekly_inflow
        if load >= 0.5:
            flags.append(("High fixed-obligation load",
                          f"${weekly_fx:,.0f}/wk in recurring payments = "
                          f"{load:.0%} of avg weekly cash inflow — "
                          f"little room for surprise expenses",
                          "F7D4D4"))
        elif load >= 0.3:
            flags.append(("Moderate fixed-obligation load",
                          f"${weekly_fx:,.0f}/wk recurring = {load:.0%} of weekly inflow — "
                          f"see Recurring tab for detail",
                          "FFF5CC"))

    # Overhead flags — construction-finance specific
    if overhead_metrics.get("under_recovery"):
        flags.append(("Under-recovering overhead",
                      f"Gross margin {overhead_metrics['gm_pct_ytd']:.1%} is below "
                      f"overhead's share of revenue {overhead_metrics['overhead_pct_revenue_ytd']:.1%} "
                      f"— losing money before profit",
                      "F7D4D4"))

    creep = overhead_metrics.get("creep", 0)
    if creep > 0.05:
        flags.append(("Overhead creep",
                      f"Overhead grew {overhead_metrics['overhead_growth']:.0%} YoY vs "
                      f"revenue at {overhead_metrics['revenue_growth']:.0%} — "
                      f"required markup is rising",
                      "F7D4D4"))

    rm_now = overhead_metrics.get("required_markup_ytd", 0)
    rm_py = overhead_metrics.get("required_markup_py", 0)
    if rm_now > rm_py + 0.02 and rm_py > 0:
        flags.append(("Required bid markup increased",
                      f"Now need {rm_now:.0%} markup over direct cost (was {rm_py:.0%} prior year) — "
                      f"raise bids or cut overhead",
                      "FFF5CC"))

    growers = overhead_metrics.get("top_growers", [])
    if growers:
        worst = growers[0]
        if worst["YoY %"] >= 0.25:
            flags.append((f"Overhead account spiking",
                          f"{worst['Account']} — ${worst['YTD']:,.0f} YTD, "
                          f"up {worst['YoY %']:.0%} YoY (was ${worst['PY YTD']:,.0f})",
                          "F7D4D4"))

    if not flags:
        flags.append(("All green",
                      "no flags raised — dashboard is clean", "DCE9D5"))

    for i, (title, body, color) in enumerate(flags, start=29):
        tc = ws.cell(row=i, column=1, value=title)
        tc.font = METRIC_FONT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        bc = ws.cell(row=i, column=2, value=body)
        bc.font = BODY_FONT
        bc.alignment = Alignment(wrap_text=True)
        if color:
            fill = PatternFill("solid", fgColor=color)
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = fill

    # ── Top 5 AR / AP tables ──
    base = 29 + len(flags) + 2

    def _top_block(title: str, rows: List[Dict[str, Any]], name_key: str, start_row: int) -> int:
        ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        hdr = start_row + 1
        for i, h in enumerate(["#", name_key, "Balance", "Bucket", "Doc"], start=1):
            c = ws.cell(row=hdr, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
        for i, r in enumerate(rows[:5], start=hdr + 1):
            ws.cell(row=i, column=1, value=i - hdr).font = BODY_FONT
            ws.cell(row=i, column=2, value=r.get(name_key, "")).font = BODY_FONT
            bc = ws.cell(row=i, column=3, value=r.get("Balance", 0))
            bc.number_format = CURRENCY_FMT; bc.font = BODY_FONT
            ws.cell(row=i, column=4, value=r.get("Bucket", "")).font = BODY_FONT
            ws.cell(row=i, column=5, value=r.get("Doc #", "")).font = BODY_FONT
        return hdr + 1 + min(5, len(rows)) + 2

    end1 = _top_block("Top 5 AR Balances", top_ar, "Customer", base)
    _top_block("Top 5 AP Balances", top_ap, "Vendor", end1)

    # Column widths
    _autosize(ws, {"A": 32, "B": 20, "C": 20, "D": 20, "E": 20, "F": 4})


# ────────────────────────── orchestration ──────────────────────────

def period_bounds(today: dt.date) -> Dict[str, Tuple[str, str]]:
    mtd_start = today.replace(day=1)
    ytd_start = today.replace(month=1, day=1)
    py_ytd_start = dt.date(today.year - 1, 1, 1)
    py_ytd_end = dt.date(today.year - 1, today.month, today.day)
    py_full_start = dt.date(today.year - 1, 1, 1)
    py_full_end = dt.date(today.year - 1, 12, 31)
    return {
        "mtd": (mtd_start.isoformat(), today.isoformat()),
        "ytd": (ytd_start.isoformat(), today.isoformat()),
        "py_ytd": (py_ytd_start.isoformat(), py_ytd_end.isoformat()),
        "py_full": (py_full_start.isoformat(), py_full_end.isoformat()),
    }


def _sum_over_bucket(rows: List[Dict[str, Any]], buckets: set) -> float:
    return sum(r.get("Balance", 0) for r in rows if r.get("Bucket") in buckets)


def build_workbook(
    data: Dict[str, Any],
    out_path: Path,
    sigma: float,
    overrides_path: Optional[Path] = None,
) -> None:
    wb = Workbook()

    # Create sheets in desired tab order
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_cash = wb.create_sheet("Cash")
    ws_ar = wb.create_sheet("AR Aging")
    ws_ap = wb.create_sheet("AP Aging")
    ws_cov = wb.create_sheet("Coverage")
    ws_rel = wb.create_sheet("Relationships")
    ws_col = wb.create_sheet("Collections")
    ws_rec = wb.create_sheet("Recurring")
    ws_oh = wb.create_sheet("Overhead")
    ws_hold = wb.create_sheet("Hold List")
    ws_src = wb.create_sheet("Sources")
    ws_pl = wb.create_sheet("P&L")
    ws_anom = wb.create_sheet("Anomalies")
    ws_meta = wb.create_sheet("_Meta")
    ws_meta.sheet_state = "hidden"

    now = data["now"]
    accounts = data["accounts"]
    open_invoices = data["open_invoices"]
    open_bills = data["open_bills"]

    bank_total = sum(a["CurrentBalance"] for a in accounts if a["AccountType"] == "Bank")
    cash_accounts_count = sum(1 for a in accounts if a["AccountType"] == "Bank")
    ar_total = sum(r["Balance"] for r in open_invoices)
    ap_total = sum(r["Balance"] for r in open_bills)
    ar_over_60 = _sum_over_bucket(open_invoices, {"61-90", "90+"})
    ap_over_60 = _sum_over_bucket(open_bills, {"61-90", "90+"})

    ytd_revenue = sum(c["YTD Revenue"] for c in data["top_customers"])
    ytd_spend = sum(v["YTD Spend"] for v in data["top_vendors"])

    retainage_recv = next((r for r in data["retainage"] if r["side"] == "receivable"), {"total": 0, "accounts": []})
    retainage_pay = next((r for r in data["retainage"] if r["side"] == "payable"), {"total": 0, "accounts": []})

    write_cash_sheet(ws_cash, accounts, data["weekly_flow"], data["runway"])
    write_aging_sheet(ws_ar, "AR", "Customer", open_invoices)
    write_aging_sheet(ws_ap, "AP", "Vendor", open_bills)
    write_coverage_sheet(ws_cov, open_invoices, open_bills, retainage_recv, retainage_pay)
    write_relationships_sheet(
        ws_rel,
        data["top_customers"],
        data["top_vendors"],
        ytd_revenue,
        ytd_spend,
        data["concentration"],
    )
    # Days into year for the balance-sheet DSO/DPO denominator
    days_into_year = (now.date() - dt.date(now.year, 1, 1)).days + 1
    write_collections_sheet(
        ws_col, data["dso_monthly"], data["dpo_monthly"],
        ar_total=ar_total, ap_total=ap_total,
        revenue_ytd=data.get("pl_ytd_totals", {}).get("income", 0) or 0,
        cogs_ytd=data.get("pl_ytd_totals", {}).get("cogs", 0) or 0,
        days_into_year=days_into_year,
    )

    # Avg weekly inflow from the 13-week cash flow (used for recurring % comparison)
    wflow = data["weekly_flow"]
    avg_weekly_inflow = (sum(w["Received"] for w in wflow) / len(wflow)) if wflow else 0
    write_recurring_sheet(
        ws_rec,
        data.get("recurring_categorized", {"active": data["recurring"], "stale": [], "new": []}),
        data.get("burn_summary", {}),
        avg_weekly_inflow,
    )
    write_overhead_sheet(
        ws_oh,
        data["overhead_metrics"],
        data.get("pl_pyfull_totals", {}),
        data.get("py_full_bounds", ("", "")),
        data.get("pl_ytd_totals", {}),
    )

    write_hold_list_sheet(
        ws_hold,
        data.get("held_invoices", []),
        overrides_path or DEFAULT_OVERRIDES,
    )

    write_sources_sheet(ws_src)

    write_pl_sheet(
        ws_pl,
        data["pl_mtd_cols"], data["pl_mtd_rows"],
        data["pl_ytd_cols"], data["pl_ytd_rows"],
        data["pl_py_cols"], data["pl_py_rows"],
    )
    write_anomalies_sheet(ws_anom, data["spikes"], data["large_txns"], sigma)
    write_dashboard_sheet(
        ws_dash,
        now,
        bank_total,
        cash_accounts_count,
        ar_total,
        ap_total,
        ar_over_60,
        ap_over_60,
        data["pl_mtd_totals"],
        data["pl_ytd_totals"],
        data["pl_py_totals"],
        len(data["spikes"]),
        data["spikes"],
        open_invoices[:5] if open_invoices else [],
        open_bills[:5] if open_bills else [],
        data["runway"],
        data["concentration"],
        data["dso_monthly"],
        data["dpo_monthly"],
        retainage_recv,
        retainage_pay,
        data["recurring"],
        data["fixed_obligations"],
        avg_weekly_inflow,
        data["overhead_metrics"],
        burn_summary=data.get("burn_summary", {}),
    )

    # Meta sheet for provenance
    ws_meta["A1"] = "Generated"; ws_meta["B1"] = now.isoformat()
    ws_meta["A2"] = "Script"; ws_meta["B2"] = "qbo_health.py"
    ws_meta["A3"] = "Sigma threshold"; ws_meta["B3"] = sigma
    ws_meta["A4"] = "Output path"; ws_meta["B4"] = str(out_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def fetch_all(access: str, company_id: str,
              overrides: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    now = dt.datetime.now()
    today = now.date()
    bounds = period_bounds(today)

    print("  → bank accounts & balances...")
    accounts = fetch_bank_accounts(access, company_id)

    print("  → retainage accounts...")
    retainage = fetch_retainage_accounts(access, company_id)

    print("  → open invoices (AR)...")
    open_invoices_all = fetch_open_invoices(access, company_id)

    # Apply user-maintained customer overrides (litigation / hold / writeoff)
    overrides = overrides or {}
    open_invoices, held_invoices = split_invoices_by_override(open_invoices_all, overrides)
    if held_invoices:
        n_parents = len({_parent_customer(r["Customer"]).lower() for r in held_invoices})
        print(f"     {len(held_invoices)} invoice(s) routed to Hold List "
              f"({n_parents} customer(s) tagged in overrides)")

    print("  → open bills (AP)...")
    open_bills = fetch_open_bills(access, company_id)

    print("  → P&L (MTD / YTD / prior-year YTD / prior-year FULL)...")
    pl_mtd_cols, pl_mtd_rows = fetch_pl(access, company_id, *bounds["mtd"])
    pl_ytd_cols, pl_ytd_rows = fetch_pl(access, company_id, *bounds["ytd"])
    pl_py_cols, pl_py_rows = fetch_pl(access, company_id, *bounds["py_ytd"])
    pl_pyfull_cols, pl_pyfull_rows = fetch_pl(access, company_id, *bounds["py_full"])

    pl_mtd_totals = extract_pl_totals(pl_mtd_rows)
    pl_ytd_totals = extract_pl_totals(pl_ytd_rows)
    pl_py_totals = extract_pl_totals(pl_py_rows)
    pl_pyfull_totals = extract_pl_totals(pl_pyfull_rows)

    pl_ytd_detail = parse_pl_by_account(pl_ytd_rows)
    pl_py_detail = parse_pl_by_account(pl_py_rows)
    pl_pyfull_detail = parse_pl_by_account(pl_pyfull_rows)

    # Diagnostic: if COGS or Expenses didn't get matched on the live P&L,
    # log the section labels we DID see (label only, no amounts) so we can
    # extend the hint sets without leaking financial data to the terminal.
    missing_sections = []
    if not pl_ytd_totals.get("cogs"):
        missing_sections.append("COGS / direct job costs")
    if not pl_ytd_totals.get("expenses"):
        missing_sections.append("Operating Expenses / overhead")
    if missing_sections:
        print(f"  ⚠ P&L parser couldn't identify: {', '.join(missing_sections)}")
        print(f"    Section-total labels found in your P&L:")
        for label, values, depth in pl_ytd_rows:
            key = label.strip().lower()
            if key.startswith("total ") and depth == 0 and values and values[0] is not None:
                print(f"       • {label!r}")
        print(f"    → Edit _COGS_TOTAL_HINTS / _EXPENSE_TOTAL_HINTS at the top of qbo_health.py")
        print(f"      to add your COA's section names. Overhead analysis will be incomplete this run.")

    overhead_metrics = compute_overhead_metrics(
        pl_ytd_totals, pl_py_totals, pl_ytd_detail, pl_py_detail,
        target_profit_pct=TARGET_NET_PROFIT_PCT,
    )

    print("  → recent expense lines (anomaly window)...")
    recent_lines = fetch_recent_expense_lines(access, company_id, days=90)

    print("  → detecting anomalies...")
    spikes = detect_overhead_spikes(recent_lines, sigma_threshold=2.0)
    large_txns = find_large_recent_txns(recent_lines, days=7, top_n=10)

    # Year-to-date invoices & bills (for top customers / top vendors rollups)
    ytd_start = bounds["ytd"][0]
    print("  → YTD invoices (for top customers)...")
    invoices_ytd = fetch_invoices_since(access, company_id, ytd_start)
    print("  → YTD bills & purchases (for top vendors)...")
    bills_ytd = fetch_bills_since(access, company_id, ytd_start)
    purchases_ytd = fetch_purchases_since(access, company_id, ytd_start)

    top_customers = aggregate_top_customers(
        invoices_ytd, top_n=10,
        aliases=(overrides or {}).get("customer_aliases", {}),
    )
    top_vendors = aggregate_top_vendors(bills_ytd, purchases_ytd, top_n=10)
    conc = concentration_flag(top_customers, threshold=0.15)

    # 12-month payment history for DSO/DPO + recurring detection
    months_back_start = (_today() - dt.timedelta(days=400)).isoformat()
    print("  → 12mo invoices & payments (for DSO)...")
    invoices_12mo = fetch_invoices_since(access, company_id, months_back_start)
    payments_12mo = fetch_payments_since(access, company_id, months_back_start)
    print("  → 12mo bills & bill payments (for DPO)...")
    bills_12mo = fetch_bills_since(access, company_id, months_back_start)
    bill_payments_12mo = fetch_bill_payments_since(access, company_id, months_back_start)
    print("  → 12mo purchases (for recurring detection)...")
    purchases_12mo = fetch_purchases_since(access, company_id, months_back_start)

    dso_monthly = compute_dso_monthly(invoices_12mo, payments_12mo, months_back=12)
    dpo_monthly = compute_dpo_monthly(bills_12mo, bill_payments_12mo, months_back=12)

    print("  → detecting recurring payments...")
    recurring = detect_recurring_payments(
        bills_12mo, purchases_12mo,
        excludes=(overrides or {}).get("recurring_excludes", {}),
    )
    # Two-phase categorization: 12-month detection vs current-3-month activity
    recurring_categorized = categorize_recurring_streams(
        recurring, bills_12mo, purchases_12mo,
        decisions=(overrides or {}).get("recurring_decisions", {}),
        current_window_days=90,
    )
    burn_summary = compute_burn_summary(recurring_categorized)
    fixed_obligations = compute_fixed_obligations(recurring)

    # 13-week cash flow + runway
    print("  → weekly cash flow (runway calc)...")
    weeks_start = (_today() - dt.timedelta(weeks=13)).isoformat()
    payments_13w = [p for p in payments_12mo if p["TxnDate"] >= weeks_start]
    bill_payments_13w = [p for p in bill_payments_12mo if p["TxnDate"] >= weeks_start]
    purchases_13w = [p for p in purchases_ytd if p["TxnDate"] >= weeks_start]
    weekly_flow = compute_weekly_cash_flow(payments_13w, bill_payments_13w, purchases_13w, weeks=13)

    bank_total_for_runway = sum(a["CurrentBalance"] for a in accounts if a["AccountType"] == "Bank")
    runway = compute_runway(bank_total_for_runway, weekly_flow)

    return {
        "now": now,
        "accounts": accounts,
        "open_invoices": open_invoices,
        "open_bills": open_bills,
        "pl_mtd_cols": pl_mtd_cols, "pl_mtd_rows": pl_mtd_rows, "pl_mtd_totals": pl_mtd_totals,
        "pl_ytd_cols": pl_ytd_cols, "pl_ytd_rows": pl_ytd_rows, "pl_ytd_totals": pl_ytd_totals,
        "pl_py_cols": pl_py_cols, "pl_py_rows": pl_py_rows, "pl_py_totals": pl_py_totals,
        "pl_pyfull_cols": pl_pyfull_cols, "pl_pyfull_rows": pl_pyfull_rows,
        "pl_pyfull_totals": pl_pyfull_totals, "pl_pyfull_detail": pl_pyfull_detail,
        "py_full_bounds": bounds["py_full"],
        "recent_lines": recent_lines,
        "spikes": spikes,
        "large_txns": large_txns,
        "top_customers": top_customers,
        "top_vendors": top_vendors,
        "concentration": conc,
        "dso_monthly": dso_monthly,
        "dpo_monthly": dpo_monthly,
        "weekly_flow": weekly_flow,
        "runway": runway,
        "retainage": retainage,
        "recurring": recurring,
        "recurring_categorized": recurring_categorized,
        "burn_summary": burn_summary,
        "fixed_obligations": fixed_obligations,
        "overhead_metrics": overhead_metrics,
        "held_invoices": held_invoices,
    }


def _chmod_600(path: Path) -> None:
    try:
        os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)
    except OSError as e:
        print(f"  ⚠ chmod 600 failed on {path}: {e}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", help=f"Override output path (default: {DEFAULT_OUTPUT})")
    ap.add_argument("--overrides",
                    help=f"Path to customer overrides xlsx (default: {DEFAULT_OVERRIDES}). "
                         f"Customers listed there get routed off main AR onto the Hold List sheet.")
    ap.add_argument("--no-lock", action="store_true", help="Skip chmod 600 on the output file")
    ap.add_argument("--anomaly-sigma", type=float, default=2.0,
                    help="Sigma threshold for overhead spike flagging (default 2.0)")
    args = ap.parse_args()

    out_path = Path(args.out).expanduser() if args.out else DEFAULT_OUTPUT
    overrides_path = Path(args.overrides).expanduser() if args.overrides else DEFAULT_OVERRIDES

    print("━" * 60)
    print("  QBO Company Health Dashboard")
    print("━" * 60)

    access, company_id = load_credentials()
    print(f"  ✓ authenticated  (company={company_id})")

    print(f"  → loading override annotations...")
    overrides = load_overrides(overrides_path)
    n_cust = len(overrides.get("customer_status", {}))
    n_alias = len(overrides.get("customer_aliases", {}))
    n_proj = len(overrides.get("project_status", {}))
    n_excl = len(overrides.get("recurring_excludes", {}))
    if any([n_cust, n_alias, n_proj, n_excl]):
        print(f"     {n_cust} customer status / {n_alias} alias / "
              f"{n_proj} project / {n_excl} recurring exclude(s)")
    else:
        print(f"     no overrides applied (file: {overrides_path})")

    data = fetch_all(access, company_id, overrides=overrides)

    print(f"  → writing workbook...")
    build_workbook(data, out_path, sigma=args.anomaly_sigma, overrides_path=overrides_path)

    if not args.no_lock:
        _chmod_600(out_path)

    print()
    print(f"  ✓ {out_path}")
    print(f"    {len(data['accounts'])} accounts · {len(data['open_invoices'])} open invoices · "
          f"{len(data['open_bills'])} open bills · {len(data['spikes'])} anomalies flagged"
          f"{' · ' + str(len(data.get('held_invoices', []))) + ' on Hold List' if data.get('held_invoices') else ''}")
    print()
    if not args.no_lock:
        print("  privacy: chmod 600 applied (owner-only). FileVault recommended at OS level.")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\ninterrupted.")
        sys.exit(130)
