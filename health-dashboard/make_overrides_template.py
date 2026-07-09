#!/usr/bin/env python3
"""
Generate a starter customer_overrides.xlsx template with three sheets:

  Customers           — alias map (Maps To) + status overrides for whole customers
  Projects            — per-project status overrides (one project of a customer
                        can be tagged separately from the rest)
  Recurring Excludes  — vendors that detection thinks are recurring but aren't

Run once to create the template at the path the dashboard reads from. After
that, edit the file directly in Excel. The dashboard auto-loads it on every
run; no code changes needed.

Usage:
  python3 make_overrides_template.py            # default path
  python3 make_overrides_template.py --out PATH # custom path
"""
import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import paths

DEFAULT_PATH = paths.companyhealth_dir() / "customer_overrides.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ITALIC_GREY = Font(italic=True, color="666666", name="Arial")


def _write_headers(ws, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT


def _write_examples(ws, examples):
    for r_idx, row in enumerate(examples, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if str(val).startswith("#") or "(example)" in str(val).lower():
                cell.font = ITALIC_GREY


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", help=f"Output path (default: {DEFAULT_PATH})")
    args = ap.parse_args()

    out = Path(args.out).expanduser() if args.out else DEFAULT_PATH

    if out.exists():
        print(f"⚠ {out} already exists — refusing to overwrite.")
        print(f"  Edit it directly in Excel, or delete it first if you want a fresh template.")
        return 1

    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Customers ──
    ws = wb.active
    ws.title = "Customers"
    _write_headers(ws, ["Customer", "Maps To", "Status", "Notes"])
    cust_examples = [
        ("# Use this sheet for customer-level overrides.", "", "", ""),
        ("# 'Maps To' (optional): roll this customer up under another parent for", "", "", ""),
        ("#   concentration / aggregation purposes. Useful when QBO has separate", "", "", ""),
        ("#   customer entries for what's really one payer (e.g. JPI Construction", "", "", ""),
        ("#   and JPI Development are both JPI). Leave blank to keep customer as itself.", "", "", ""),
        ("# 'Status' (optional): routes ALL of this customer's invoices off the main", "", "", ""),
        ("#   AR aging and onto the Hold List sheet (litigation, hold, collections,", "", "", ""),
        ("#   writeoff, or any free-form label).", "", "", ""),
        ("# Match is by parent name, case-insensitive.", "", "", ""),
        ("", "", "", ""),
        ("JPI Development (example)", "JPI", "", "alias only — group with JPI parent"),
        ("JPI Construction (example)", "JPI", "", "alias only — group with JPI parent"),
        ("Acme Construction (example)", "", "litigation", "filed 2026-01-15 — Smith & Jones"),
        ("Slow Pay LLC (example)", "", "collections", "with attorney 2026-03-04"),
        ("Old Customer Inc (example)", "", "writeoff", "approved for write-off 2025-12"),
    ]
    _write_examples(ws, cust_examples)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 50
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 2: Projects ──
    ws = wb.create_sheet("Projects")
    _write_headers(ws, ["Project #", "Status", "Notes"])
    proj_examples = [
        ("# Use this sheet when ONE project of a customer needs special handling but", "", ""),
        ("#   the customer's other projects are normal. Match by project number that", "", ""),
        ("#   appears in QBO's CustomerRef.name format 'Parent:RP123 Description'.", "", ""),
        ("# Customer-level overrides take priority over project-level if both apply.", "", ""),
        ("", "", ""),
        ("RP742 (example)", "hold", "scope dispute — pause AR until resolved"),
        ("CP301 (example)", "litigation", "in dispute, filed 2026-02"),
    ]
    _write_examples(ws, proj_examples)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 3: Recurring Excludes ──
    ws = wb.create_sheet("Recurring Excludes")
    _write_headers(ws, ["Vendor", "Reason"])
    rec_examples = [
        ("# Use this sheet to silence false-positive 'recurring' detections.", "", ),
        ("# Sometimes a vendor's bills look recurring statistically but you know", ""),
        ("#   they're not (e.g., coincidental same-amount one-offs, or a vendor", ""),
        ("#   that bills monthly only because of a temporary contract).", ""),
        ("# Vendor name match is case-insensitive.", ""),
        ("", ""),
        ("Wells Fargo Bank (example)", "transfers, not recurring expense"),
        ("Bob's Concrete (example)", "happens to have similar amounts but project-driven"),
    ]
    _write_examples(ws, rec_examples)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 4: Recurring Decisions ──
    ws = wb.create_sheet("Recurring Decisions")
    _write_headers(ws, ["Vendor", "Decision", "Notes"])
    dec_examples = [
        ("# Per-vendor disambiguation for cases the auto-classifier got wrong.", "", ""),
        ("# Decision values:", "", ""),
        ("#   active   = treat as currently recurring even if no recent activity", "", ""),
        ("#              (e.g., quarterly bill that happens to be outside the 90-day window)", "", ""),
        ("#   exclude  = remove from recurring view entirely (same effect as", "", ""),
        ("#              putting on Recurring Excludes sheet)", "", ""),
        ("#   ignore   = no opinion, let auto-classification stand", "", ""),
        ("# Vendor match is case-insensitive.", "", ""),
        ("", "", ""),
        ("State Comp Insurance (example)", "active", "quarterly — out of 90-day window but real"),
        ("Old Software (example)", "exclude", "cancelled subscription, don't show"),
    ]
    _write_examples(ws, dec_examples)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(out)
    print(f"✓ wrote template: {out}")
    print(f"")
    print(f"  Open in Excel. Replace (example) rows with your real entries, save, and")
    print(f"  re-run the dashboard. Changes apply immediately on the next refresh.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
