#!/usr/bin/env python3
"""
company_tracker.py — the ONE consolidated workbook (the user 2026-07-17).

Reads every tracker workbook (the data layer) and folds them into a single
`Company Tracker.xlsx` organised MONEY IN / MONEY OUT / POSITION, then renders
the `Company Dashboard.html` that breaks the same model down. One command →
one workbook + one dashboard, both from ONE metric model (company_dashboard.
build_sections) so they can never disagree.

  Summary tab   — the whole story: hero numbers + every metric, grouped In/Out/
                  Position, colour-coded (AR/backlog green, money-out amber/red,
                  position flags red when bad).
  Money In / Money Out / Position tabs — each section on its own tab with its
                  metric table + aging / LOC-by-division bars (Excel data bars).

No QBO calls, no Touch ID, offline. Regenerate the source trackers first.
OUTPUT  ~/Documents/CompanyHealth/Company Tracker.xlsx  (chmod 600)
        ~/Documents/CompanyHealth/Company Dashboard.html (chmod 600)

USAGE
  python3 health-dashboard/company_tracker.py
  python3 health-dashboard/company_tracker.py --open
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import stat
import sys
import webbrowser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import schedule as sched
from shared import recurring as rec
import json
import company_dashboard as cd    # same tool (health-dashboard/) — shared readers + model

CH = paths.companyhealth_dir()
XLSX_OUT = CH / "Company Tracker.xlsx"
HTML_OUT = CH / "Company Dashboard.html"

# ── styling (matches the other trackers / the HTML tones) ──
_NAVY = PatternFill("solid", fgColor="1F3864")
_ZEBRA = PatternFill("solid", fgColor="EEF3FA")
_WHITE = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D7E5")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
TONE = {                                   # (band fill, value font color)
    "be":  ("7030A0", "1C2430"),
    "in":  ("1F6B4C", "1F6B4C"),
    "out": ("C00000", "9C0006"),
    "pos": ("1F3864", "1C2430"),
}
VCOLOR = {"g": "1F6B4C", "r": "9C0006", "a": "9C6500", "n": "1C2430"}
CUR = '#,##0'


def _band(ws, ncols, text, color):
    ws.append([text] + [""] * (ncols - 1))
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22


def _heroes(ws, ncols, heroes, tone):
    """A row of big hero numbers under a band."""
    ws.append([f"{v}   {lbl}" for lbl, v, _cls in heroes] + [""] * (ncols - len(heroes)))
    r = ws.max_row
    for i, (_lbl, _v, cls) in enumerate(heroes):
        cell = ws.cell(r, i + 1)
        cell.font = Font(bold=True, size=13, color=VCOLOR.get(cls, "1C2430"))
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 26


def _metric_table(ws, rows):
    """metric | value | detail rows, value coloured by class."""
    hdr = ws.max_row + 1
    ws.append(["METRIC", "VALUE", "DETAIL"])
    for c in range(1, 4):
        cell = ws.cell(hdr, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
    for i, (metric, val, detail, cls) in enumerate(rows):
        ws.append([metric, val, detail])
        r = ws.max_row
        for c in range(1, 4):
            ws.cell(r, c).border = _BORDER
            if i % 2:
                ws.cell(r, c).fill = _ZEBRA
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).font = Font(bold=True, color=VCOLOR.get(cls, "1C2430"))
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        ws.cell(r, 3).font = Font(color="6B7280", size=10)


def _bar_block(ws, header, items):
    """LABEL | VALUE(number) rows with a data bar on the value column."""
    ws.append([header])
    ws.cell(ws.max_row, 1).font = Font(bold=True, italic=True, color="6B7280")
    first = ws.max_row + 1
    for label, val, _color, sub in items:
        ws.append([f"{label}" + (f"  ({sub})" if sub else ""), round(val)])
        r = ws.max_row
        ws.cell(r, 2).number_format = CUR
        for c in (1, 2):
            ws.cell(r, c).border = _BORDER
    last = ws.max_row
    if last >= first:
        ws.conditional_formatting.add(
            f"B{first}:B{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color="8EAADB", showValue=True))


def _section_to_sheet(ws, sec, standalone):
    ncols = 3
    band_color = TONE[sec["tone"]][0]
    _band(ws, ncols, sec["title"], band_color)
    _heroes(ws, ncols, sec["heroes"], sec["tone"])
    ws.append([])
    _metric_table(ws, sec["rows"])
    for header, items in sec.get("bars", []):
        if not items:
            continue
        ws.append([])
        _bar_block(ws, header, items)
    if standalone:
        for col, w in zip("ABC", (44, 16, 60)):
            ws.column_dimensions[col].width = w
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = band_color


def _write_schedule_tab(wb, model) -> None:
    """Weekly stage Gantt as a tab: job × day, cell coloured by stage."""
    if not model or not model.get("jobs"):
        return
    ws = wb.create_sheet("Weekly Schedule")
    ws.sheet_view.showGridLines = False
    dates = model["dates"]
    hdr = ["PROJECT #", "ADDRESS", "BUILDER", "CONTRACT $", "CURRENT STAGE"] + \
          [d.strftime("%a %m/%d") for d in dates]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([11, 30, 26, 13, 14] + [12] * len(dates)):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "F2"
    for i, j in enumerate(model["jobs"]):
        row = [j["proj"], j["address"], j["builder"], j["contract"],
               sched.stage_cat(j["current"])[0] if j["current"] else ""]
        for d in dates:
            row.append(sched.stage_cat(j["days"][d])[0] if d in j["days"] else "")
        ws.append(row)
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(r, c).border = _BORDER
            if i % 2:
                ws.cell(r, c).fill = _ZEBRA
        ws.cell(r, 4).number_format = CUR
        for k, d in enumerate(dates):
            cell = ws.cell(r, 6 + k)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if d in j["days"]:
                cell.fill = PatternFill("solid", fgColor=sched.stage_cat(j["days"][d])[1])
                cell.font = Font(bold=True, color="FFFFFF", size=9)
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    # legend runs DOWN a column to the right of the grid — laid across the
    # narrow day columns it truncated the labels (the user 2026-08-03)
    lc = 6 + len(dates) + 1
    ws.column_dimensions[get_column_letter(lc)].width = 18
    ws.cell(1, lc, "LEGEND").font = Font(bold=True)
    for i, (name, color) in enumerate(sched.legend(), start=2):
        cc = ws.cell(i, lc, name)
        cc.fill = PatternFill("solid", fgColor=color)
        cc.font = Font(bold=True, color="FFFFFF", size=9)
        cc.alignment = Alignment(horizontal="left", indent=1)
        cc.border = _BORDER
    ws.sheet_properties.tabColor = "305496"


NOTES_FILE = paths.companyhealth_sources_dir() / "recurring_notes.json"


def _load_notes() -> dict:
    """The owner's annotations, kept in a sidecar so regenerating the workbook
    never wipes them (the user 2026-08-03: "I let you know what really happened
    and you update")."""
    try:
        return json.loads(NOTES_FILE.read_text())
    except Exception:
        return {}


def _write_recurring_tab(wb, model) -> None:
    """Every recurring obligation in front of the owner: what we count, when it
    was last paid, what CHANGED / STOPPED, and their own note."""
    if not model:
        return
    notes = _load_notes()
    ws = wb.create_sheet("Recurring & Debt")
    ws.sheet_view.showGridLines = False
    hdr = ["STATUS", "OBLIGATION", "TYPE", "PER MONTH $", "PRIOR $",
           "LAST PAID", "BALANCE $", "WHAT CHANGED", "YOUR NOTE (edit me)"]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([11, 36, 10, 14, 13, 12, 14, 38, 34]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    SC = {"CHANGED": "FFEB9C", "STOPPED": "FFC7CE", "NEW": "C6EFCE"}

    def block(label, rows, color):
        if not rows:
            return
        ws.append([f"  {label}   —   {len(rows)} item(s)   —   "
                   f"${sum(r['last_amount'] for r in rows):,.0f}/month"] + [""] * 8)
        br = ws.max_row
        ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=9)
        ws.cell(br, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(br, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.row_dimensions[br].height = 20
        for i, r in enumerate(rows):
            ws.append([r["status"], r["name"], r.get("kind", "debt"),
                       r["last_amount"], r.get("prior_amount", 0),
                       r.get("last_month", ""), r.get("balance_now", ""),
                       r.get("note", "") + (" · fees already in overhead"
                                            if r.get("fees_booked") else ""),
                       notes.get(r["name"], "")])
            rr = ws.max_row
            for c in range(1, 10):
                ws.cell(rr, c).border = _BORDER
                if i % 2:
                    ws.cell(rr, c).fill = _ZEBRA
            for c in (4, 5, 7):
                ws.cell(rr, c).number_format = CUR
            if r["status"] in SC:
                ws.cell(rr, 1).fill = PatternFill("solid", fgColor=SC[r["status"]])
                ws.cell(rr, 1).font = Font(bold=True)
            ws.cell(rr, 9).fill = PatternFill("solid", fgColor="FFFDE7")

    alerts = [r for r in model["alerts"]]
    refi = model.get("refinancing", [])
    block("⚠ NEEDS YOUR REVIEW — changed, stopped, or new", alerts, "C00000")
    block("FIXED OVERHEAD — steady", [r for r in model["overhead"]
                                      if r["kind"] == "fixed" and r not in alerts], "1F6B4C")
    block("DEBT SERVICE — steady", [r for r in model["debt"]
                                    if r not in alerts and r["status"] != "STOPPED"], "305496")
    block("MCA / REFINANCING — balance moves by JE, NOT a cash payment "
          "(fees already in overhead) — tell us the real ACH amount", refi, "BF8F00")
    block("VARIABLE / one-off — excluded from the fixed nut",
          [r for r in model["overhead"] if r["kind"] == "variable" and r not in alerts], "7F7F7F")
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    ws.append([])
    ws.append([f"Fixed overhead ${model['fixed_overhead_month']:,.0f}/mo  +  "
               f"debt service ${model['debt_service_month']:,.0f}/mo  =  "
               f"${model['total_monthly_obligation']:,.0f}/mo total obligation. "
               f"Debt payment = the month-over-month drop in the liability balance "
               f"(interest isn't booked until the CPA's year-end JE, so the whole "
               f"payment lands there). MCA/refinancing rows are EXCLUDED — a new "
               f"advance pays off the prior one via the CPA's JE, so the balance "
               f"delta isn't cash. Type your corrections in the last column — "
               f"they are preserved across runs."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="808080")
    ws.sheet_properties.tabColor = "C55A11"


def _write_breakeven_audit(ws, audit) -> None:
    """Append the full audit trail under the Break-Even metrics so every figure
    can be traced back to its QBO source and re-derived by hand."""
    if not audit:
        return
    ws.append([])
    ws.append(["AUDIT TRAIL — where every number comes from"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="7030A0")
    ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=11)
    ws.row_dimensions[r].height = 20
    for item, val, where in audit:
        ws.append([item, val, where])
        rr = ws.max_row
        if item.startswith("──"):
            for c in range(1, 4):
                ws.cell(rr, c).fill = PatternFill("solid", fgColor="EDE7F6")
            ws.cell(rr, 1).font = Font(bold=True, color="4A3AA7")
        else:
            ws.cell(rr, 1).font = Font(bold=True)
            ws.cell(rr, 2).alignment = Alignment(horizontal="right")
            ws.cell(rr, 2).font = Font(bold=True)
            ws.cell(rr, 3).font = Font(color="6B7280", size=9)
        for c in range(1, 4):
            ws.cell(rr, c).border = _BORDER
            ws.cell(rr, c).alignment = Alignment(
                horizontal="right" if c == 2 else "left",
                vertical="top", wrap_text=(c == 3))


def _write_rp_billing_tab(wb, rows) -> None:
    """RP Billing Status detail — poured & unbilled vs backlog (the user
    2026-07-28). Computed by money_bleeds (needs QBO); folded in here so it
    lives in the ONE workbook rather than its own file."""
    if not rows:
        return
    ws = wb.create_sheet("RP Billing Status")
    ws.sheet_view.showGridLines = False
    hdr = ["PROJECT #", "SCOPE", "ADDRESS", "BUILDER", "BID $", "BILLED $",
           "DUE $", "DAYS SINCE POUR"]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([11, 7, 30, 24, 13, 13, 13, 15]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    for status, label, color in (
            ("INVOICE NOW", "POURED — NOT BILLED · INVOICE NOW", "C00000"),
            ("draw-based", "DRAW / PHASE-BILLED — bill by draw, not pour", "BF8F00"),
            ("backlog", "NOT POURED — BACKLOG (not owed)", "7F7F7F")):
        items = [r for r in rows if r["status"] == status]
        if not items:
            continue
        ws.append([f"  {label}   —   {len(items)} scope(s)   —   "
                   f"${sum(r['gap'] for r in items):,.0f}"] + [""] * 7)
        br = ws.max_row
        ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=8)
        ws.cell(br, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(br, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.row_dimensions[br].height = 20
        first = ws.max_row + 1
        for i, r in enumerate(items):
            ws.append([r["proj"], r["scope"], r["addr"], r["builder"],
                       r["bid"], r["billed"], r["gap"],
                       r["days"] if r["days"] is not None else ""])
            rr = ws.max_row
            for c in range(1, 9):
                ws.cell(rr, c).border = _BORDER
                if i % 2:
                    ws.cell(rr, c).fill = _ZEBRA
            for c in (5, 6, 7):
                ws.cell(rr, c).number_format = CUR
            if status == "INVOICE NOW":
                ws.cell(rr, 7).font = Font(bold=True, color="9C0006")
            ws.row_dimensions[rr].outlineLevel = 1
        last = ws.max_row
        if last >= first:
            ws.conditional_formatting.add(
                f"G{first}:G{last}",
                DataBarRule(start_type="num", start_value=0, end_type="max",
                            color="F8696B" if status == "INVOICE NOW" else "BFBFBF",
                            showValue=True))
        ws.append([""] * 4 + ["", "subtotal", sum(r["gap"] for r in items), ""])
        sr = ws.max_row
        for c in range(1, 9):
            ws.cell(sr, c).fill = PatternFill("solid", fgColor="D9E1F2")
            ws.cell(sr, c).border = _BORDER
            ws.cell(sr, c).font = Font(bold=True)
        ws.cell(sr, 7).number_format = CUR
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    ws.append([])
    ws.append(["Poured = the crew schedule shows this scope reaching pour/wreck/"
               "stress/punch. Slab vs flatwork from the stage text; billing per "
               "scope (RP#### vs RP####-FTW). Small gaps (builder fee) ignored."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="808080")
    ws.sheet_properties.tabColor = "7030A0"


def build_workbook(path: Path, sections, health, schedule=None, rp_rows=None,
                   recurring_model=None) -> None:
    wb = Workbook()

    # Summary tab — the whole story
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    for col, w in zip("ABC", (44, 16, 60)):
        s.column_dimensions[col].width = w
    s.append(["COMPANY TRACKER — Money In · Money Out · Position"])
    s.merge_cells("A1:C1")
    s.cell(1, 1).font = Font(bold=True, size=15, color="FFFFFF")
    for c in (1, 2, 3):
        s.cell(1, c).fill = _NAVY
    s.row_dimensions[1].height = 32
    gen = health.get("generated")
    s.append([f"Generated {dt.datetime.now():%Y-%m-%d %H:%M}  ·  cash / AR-AP / margins "
              f"as of {gen:%Y-%m-%d}" if gen else "Generated now"])
    s.cell(2, 1).font = Font(italic=True, color="6B7280")
    s.merge_cells("A2:C2")
    for sec in sections:
        s.append([])
        _section_to_sheet(s, sec, standalone=False)

    # one tab per section
    for sec, tab in zip(sections, ("Money In", "Money Out", "Position", "Break-Even")):
        ws = wb.create_sheet(tab)
        _section_to_sheet(ws, sec, standalone=True)
        if sec.get("audit"):
            _write_breakeven_audit(ws, sec["audit"])

    _write_recurring_tab(wb, recurring_model)
    _write_rp_billing_tab(wb, rp_rows or [])
    _write_schedule_tab(wb, schedule)

    s.sheet_properties.tabColor = "1F3864"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)


def main() -> int:
    ap = argparse.ArgumentParser(description="One consolidated company tracker workbook + HTML")
    ap.add_argument("--xlsx", type=Path, default=XLSX_OUT)
    ap.add_argument("--html", type=Path, default=HTML_OUT)
    ap.add_argument("--open", action="store_true")
    args = ap.parse_args()

    print("\n  COMPANY TRACKER — one workbook, In / Out / Position")
    print("  " + "─" * 50)
    cards = cd.read_money_bleeds(cd.MB_PATH)
    loc = cd.read_sub_loc(cd.LOC_PATH)
    mor = cd.read_money_out(cd.MOR_PATH)
    health = cd.read_health(cd.HEALTH_PATH)
    wip = cd.read_wip(cd.WIP_PATH)
    print(f"  sources: MB {len(cards)} · LOC div {len(loc['divisions'])} · "
          f"WIP active {wip['count']} · cash {cd.hk(health, 'Bank total')}")

    sections = cd.build_sections(cards, loc, mor, health, wip)   # ONE model
    schedule = sched.build_model()          # week grid → Excel Gantt tab
    rp_stages = sched.build_rp_stages()     # all-RP stage view → HTML
    print(f"  schedule: {len(schedule['jobs'])} jobs this week · "
          f"{len(rp_stages['rows'])} RP projects staged" if schedule["jobs"]
          else "  schedule: none (volume not mounted / no files) — tab skipped")
    rp_rows = cd.read_rp_billing(cd.MB_PATH)
    print(f"  RP billing rows: {len(rp_rows)}")
    recurring_model = None
    try:
        from shared.qbo_cache import QboCache
        recurring_model = rec.build(QboCache())
        print(f"  recurring: {len(recurring_model['overhead'])} overhead + "
              f"{len(recurring_model['debt'])} debt · "
              f"{len(recurring_model['alerts'])} need review")
    except Exception as e:
        print(f"  recurring: skipped ({type(e).__name__}) — needs QBO")
    build_workbook(args.xlsx, sections, health, schedule, rp_rows, recurring_model)
    print(f"  ✓ {args.xlsx}")

    sources = [cd.MB_PATH, cd.LOC_PATH, cd.MOR_PATH, cd.HEALTH_PATH, cd.WIP_PATH, cd.BILL_PATH]
    args.html.write_text(cd.render(sections, health, sources, rp_stages), encoding="utf-8")
    os.chmod(args.html, stat.S_IRUSR | stat.S_IWUSR)
    print(f"  ✓ {args.html}")
    if args.open:
        webbrowser.open(f"file://{args.xlsx}")
        webbrowser.open(f"file://{args.html}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
