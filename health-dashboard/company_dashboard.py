#!/usr/bin/env python3
"""
company_dashboard.py — one consolidated company-health view (HTML).

Reads the tracker workbooks that the other tools generate (the data layer) and
renders ONE self-contained HTML page, organised as MONEY IN / MONEY OUT /
POSITION (the user 2026-07-17) — grouped tables with a few hero numbers, not a
wall of identical boxes. No QBO calls, no Touch ID, offline. Regenerate the
trackers first, then run this.

Colour logic is semantic: money owed TO us (AR, backlog, retainage) is green;
money OUT (AP, POs, checks, LOC) is amber/red; position flags turn red when bad
(runway < 8 wks, coverage < 1, AR < AP).

Sources: Money Bleeds.xlsx · Sub LOC Report.xlsx · Money Out Register.xlsx ·
health_dashboard.xlsx (cash/AR/AP/margins) · WIP master Test-Master (backlog +
over/under-billing). Each carries its freshness; stale ones are flagged.

OUTPUT  ~/Documents/CompanyHealth/Company Dashboard.html  (chmod 600)

USAGE
  python3 health-dashboard/company_dashboard.py --open
"""
from __future__ import annotations

import argparse
import datetime as dt
import html
import os
import re
import stat
import sys
import webbrowser
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths

CH = paths.companyhealth_dir()
MB_PATH = CH / "Money Bleeds.xlsx"
LOC_PATH = CH / "Sub LOC Report.xlsx"
MOR_PATH = CH / "Money Out Register.xlsx"
HEALTH_PATH = CH / "health_dashboard.xlsx"
BILL_PATH = CH / "Bill Tracker.xlsx"
WIP_PATH = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx")
OUT_PATH = CH / "Company Dashboard.html"
STALE_DAYS = 7


# ────────────────────────── helpers ──────────────────────────

def _mtime(p: Path) -> Optional[dt.datetime]:
    try:
        return dt.datetime.fromtimestamp(p.stat().st_mtime)
    except OSError:
        return None


def _num(v) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def _money(v) -> str:
    n = _num(v)
    return f"${n:,.0f}" if n is not None else (html.escape(str(v)) if v else "—")


# ────────────────────────── source readers ──────────────────────────

def read_money_bleeds(path: Path) -> Dict[str, dict]:
    """Money Bleeds 'Dashboard' cards keyed by a normalised label →
    {count, amount}."""
    out: Dict[str, dict] = {}
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Dashboard" not in wb.sheetnames:
        wb.close()
        return out
    for row in wb["Dashboard"].iter_rows(min_row=2, values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        if not b:
            continue
        label = str(b).strip()
        if label.startswith(("Generated", "RP data")):
            continue
        m = re.search(r"^(.*?):\s*(\d+)\s*$", label)
        if m:
            out[m.group(1).strip().lower()] = {"count": int(m.group(2)),
                                               "amount": _num(c)}
    wb.close()
    return out


def mb(cards: dict, needle: str) -> dict:
    for k, v in cards.items():
        if needle.lower() in k:
            return v
    return {"count": 0, "amount": None}


def read_sub_loc(path: Path) -> dict:
    out: Dict[str, Any] = {"summary": {}, "divisions": []}
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(min_row=3, values_only=True):
            if len(row) > 2 and row[1] and row[2] not in (None, ""):
                out["summary"][str(row[1]).strip()] = row[2]
    if "By Division" in wb.sheetnames:
        for row in wb["By Division"].iter_rows(min_row=2, values_only=True):
            nm = str(row[0]).strip() if row and row[0] else ""
            if not nm or nm in ("TOTAL",) or nm.startswith("Note"):
                continue
            out["divisions"].append({"name": nm, "peak": _num(row[1]) or 0,
                                     "outstanding": _num(row[6]) or 0,
                                     "float": _num(row[7]) or 0})
    wb.close()
    return out


def loc_val(loc: dict, needle: str):
    for k, v in loc["summary"].items():
        if needle.lower() in k.lower():
            return v
    return None


def read_money_out(path: Path) -> dict:
    out: Dict[str, Any] = {}
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(min_row=3, values_only=True):
            lbl = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            val = row[2] if len(row) > 2 else None
            if "aged > 30" in lbl and "$" in lbl:
                out["aged_total"] = _num(val)
            elif "aged > 30" in lbl:
                out["aged_count"] = _num(val)
            elif lbl.startswith("unmarked checks"):
                out["unmarked_count"] = _num(val)
            elif lbl.startswith("unmarked $"):
                out["unmarked_total"] = _num(val)
    wb.close()
    return out


AGING = ["Current", "1-30", "31-60", "61-90", "90+"]


def read_health(path: Path) -> dict:
    """Pre-computed KPIs from health_dashboard's Dashboard grid + AR/AP aging
    buckets + top-customer concentration."""
    out: Dict[str, Any] = {"generated": None, "kpis": {},
                           "ar_buckets": {}, "ap_buckets": {},
                           "concentration": None, "top_customer": None}
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    if "_Meta" in wb.sheetnames:
        for row in wb["_Meta"].iter_rows(values_only=True):
            if row and str(row[0]).strip() == "Generated" and len(row) > 1:
                try:
                    out["generated"] = dt.datetime.fromisoformat(str(row[1]))
                except ValueError:
                    pass
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        for r in range(1, ws.max_row + 1):
            for lc, vc in ((1, 2), (4, 5)):
                lbl, val = ws.cell(r, lc).value, ws.cell(r, vc).value
                if isinstance(lbl, str) and isinstance(val, (int, float)):
                    out["kpis"][lbl.strip()] = val
    for sheet, key in (("AR Aging", "ar_buckets"), ("AP Aging", "ap_buckets")):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        hi = {n: i for i, n in enumerate(hdr)}
        bi, vi = hi.get("Bucket"), hi.get("Balance")
        ni = hi.get("Customer", hi.get("Vendor"))
        b: Dict[str, float] = {}
        for row in it:
            nm = str(row[ni] or "") if ni is not None else ""
            if nm.startswith("━") or "GRAND TOTAL" in nm:
                continue
            bk = row[bi] if bi is not None else None
            vv = _num(row[vi]) if vi is not None else None
            if bk and vv:
                b[str(bk)] = b.get(str(bk), 0) + vv
        out[key] = b
    if "Relationships" in wb.sheetnames:
        ws = wb["Relationships"]
        for r in range(1, min(ws.max_row + 1, 12)):
            if str(ws.cell(r, 1).value).strip() == "1":       # first ranked customer
                out["top_customer"] = ws.cell(r, 2).value
                out["concentration"] = _num(ws.cell(r, 5).value)
                break
    wb.close()
    return out


def hk(health: dict, needle: str):
    for k, v in health["kpis"].items():
        if needle.lower() in k.lower():
            return v
    return None


def read_wip(path: Path) -> dict:
    """Active-project totals from the WIP master Test-Master tab (header row
    detected). Backlog = LEFT TO BILL; over/under-billing net; job borrow."""
    out: Dict[str, Any] = {"count": 0, "contract": 0.0, "billed": 0.0,
                           "left_to_bill": 0.0, "over": 0.0, "under": 0.0,
                           "job_borrow": 0.0}
    if not path.exists():
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    if "Test-Master" not in wb.sheetnames:
        wb.close()
        return out
    ws = wb["Test-Master"]
    hdr_row, H = None, {}
    for r in range(1, 6):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any("CONTRACT" in str(v).upper() for v in vals if v):
            hdr_row = r
            H = {str(v).strip().upper(): i + 1 for i, v in enumerate(vals) if v}
            break
    if not hdr_row:
        wb.close()
        return out

    def col(r, name):
        c = H.get(name)
        v = ws.cell(r, c).value if c else None
        return v if isinstance(v, (int, float)) else 0

    for r in range(hdr_row + 1, ws.max_row + 1):
        proj = ws.cell(r, H.get("PROJECT #", 2)).value
        if not proj:
            continue
        if str(ws.cell(r, H.get("STATUS", 7)).value or "").strip().lower() != "active":
            continue
        out["count"] += 1
        out["contract"] += col(r, "TOTAL CONTRACT PRICE")
        out["billed"] += col(r, "BILLED TO DATE")
        out["left_to_bill"] += col(r, "LEFT TO BILL")
        out["over"] += col(r, "OVERBILLINGS")
        out["under"] += col(r, "UNDERBILLINGS")
        out["job_borrow"] += col(r, "PURE JOB BORROW")
    wb.close()
    return out


# ────────────────────────── HTML render ──────────────────────────

def _fresh_badge(p: Path) -> str:
    mt = _mtime(p)
    if not mt:
        return f'<span class="badge missing">{html.escape(p.name)} · not generated</span>'
    age = (dt.datetime.now() - mt).days
    cls = "stale" if age > STALE_DAYS else "fresh"
    return f'<span class="badge {cls}">{html.escape(p.name)} · {mt:%Y-%m-%d} ({age}d)</span>'


def _pct(v) -> str:
    n = _num(v)
    return f"{n*100:.1f}%" if n is not None else "—"


def _bar(buckets: dict) -> str:
    if not buckets:
        return ""
    mx = max(buckets.values(), default=1) or 1
    shade = {"Current": "#1F6B4C", "1-30": "#7F9F3F", "31-60": "#BF8F00",
             "61-90": "#D2691E", "90+": "#C00000"}
    out = ""
    for b in AGING:
        v = buckets.get(b, 0)
        if not v:
            continue
        w = max(3, round(100 * v / mx))
        out += (f'<div class="agerow"><span class="agelbl">{b}</span>'
                f'<span class="agetrack"><span class="agefill" '
                f'style="width:{w}%;background:{shade.get(b, "#888")}">'
                f'${v:,.0f}</span></span></div>')
    return out


def _section(title, tone, heroes, rows, extra="") -> str:
    """tone: in | out | pos. heroes: [(label,value,cls)]. rows: [(metric,value,
    detail,valcls)]."""
    hero_html = "".join(
        f'<div class="hero {cls}"><div class="hv">{val}</div>'
        f'<div class="hl">{html.escape(lbl)}</div></div>'
        for lbl, val, cls in heroes)
    row_html = "".join(
        f'<tr><td class="m">{html.escape(m)}</td>'
        f'<td class="v {vcls}">{val}</td>'
        f'<td class="d">{det}</td></tr>'
        for m, val, det, vcls in rows)
    return f'''<section class="sec {tone}">
      <h2>{html.escape(title)}</h2>
      <div class="heroes">{hero_html}</div>
      <table class="metrics"><tbody>{row_html}</tbody></table>
      {extra}
    </section>'''


def render(cards, loc, mor, health, wip, sources) -> str:
    now = dt.datetime.now()
    gen = health.get("generated")
    hage = (now - gen).days if gen else None
    hstale = hage is not None and hage > STALE_DAYS
    asof = (f'<span class="asof {"bad" if hstale else ""}">cash / AR-AP / margins '
            f'as of {gen:%Y-%m-%d}{" — STALE, run qbo_health.py" if hstale else ""}'
            f'</span>' if gen else "")

    ar = hk(health, "Total AR")
    ap = hk(health, "Total AP")
    retain = hk(health, "Retainage Receivable")
    cash = hk(health, "Bank total")
    runway = hk(health, "Runway at current burn")
    gm = hk(health, "Current Gross Margin")
    nm = hk(health, "Net margin after overhead")
    burn = hk(health, "Weekly burn")
    cov = (ar / ap) if (_num(ar) and _num(ap)) else None

    # ---- MONEY IN ----
    lien_past = mb(cards, "past deadline")
    cp_draw = mb(cards, "under-invoiced")
    rp_wait = mb(cards, "waiting on punch")
    money_in = _section(
        "Money In — owed to us / to bill", "in",
        [(f"{_money(ar)}", _money(ar), "g"),
         (f"Unbilled backlog {_money(wip['left_to_bill'])}", _money(wip["left_to_bill"]), "g"),
         (f"Retainage {_money(retain)}", _money(retain), "g")],
        [("Accounts Receivable", _money(ar), f"aged 60+ {_money(hk(health,'AR aged 60+') or '')}"[:60] if hk(health,'AR aged 60+') else "money owed to us", "g"),
         ("Unbilled backlog (WIP active)", _money(wip["left_to_bill"]), f"{wip['count']} active jobs · contract {_money(wip['contract'])}", "g"),
         ("Underbilled / job borrow (WIP)", _money(wip["under"]), "earned but not yet billed", "g"),
         ("Retainage receivable", _money(retain), "held by clients until close", "g"),
         ("Lien deadline PAST", _money(lien_past["amount"]), f"{lien_past['count']} invoices — money with no lien backup", "r" if lien_past["count"] else "g"),
         ("CP draws billed, not invoiced", _money(cp_draw["amount"]), f"{cp_draw['count']} project(s)", "r" if cp_draw["count"] else "g"),
         ("RP slabs 100% waiting to bill", str(rp_wait["count"]), "wrap up to get paid", "a" if rp_wait["count"] else "g")],
        extra=f'<div class="aging"><div class="agehd">AR aging</div>{_bar(health.get("ar_buckets", {}))}</div>')

    # ---- MONEY OUT ----
    ap_pay = mb(cards, "client hasn't paid")
    ap_noinv = mb(cards, "no invoice issued")
    ap_now = mb(cards, "lien coming up")
    unused_po = mb(cards, "unused pos")
    loc_peak = loc_val(loc, "LOC you truly need")
    div_bars = "".join(
        f'<div class="agerow"><span class="agelbl">{html.escape(d["name"])}</span>'
        f'<span class="agetrack"><span class="agefill" style="width:'
        f'{max(3, round(100*d["peak"]/(max((x["peak"] for x in loc["divisions"]), default=1) or 1)))}%;'
        f'background:{ {"MFD":"#305496","CP":"#1F6B4C","RP":"#7030A0"}.get(d["name"],"#888") }">'
        f'${d["peak"]:,.0f} · {d["float"]:.0f}d</span></span></div>'
        for d in sorted(loc["divisions"], key=lambda x: -x["peak"]))
    money_out = _section(
        "Money Out — we owe / committed", "out",
        [(_money(ap), _money(ap), "r"),
         (_money(loc_peak), _money(loc_peak), "a"),
         (_money(unused_po["amount"]), _money(unused_po["amount"]), "a")],
        [("Accounts Payable", _money(ap), f"aged 60+ {_money(hk(health,'AP aged 60+') or '')}"[:60] if hk(health,'AP aged 60+') else "money we owe", "r"),
         ("Bills to pay NOW (paid by client + lien due)", _money(ap_now["amount"]), f"{ap_now['count']} bills — money in, sub can lien us", "r" if ap_now["count"] else "g"),
         ("Bills — client hasn't paid us", _money(ap_pay["amount"]), f"{ap_pay['count']} bills, collect from GC first", "a"),
         ("Bills — no GC invoice issued yet", _money(ap_noinv["amount"]), f"{ap_noinv['count']} bills, bill the GC", "a"),
         ("Unused POs ≥30 days", _money(unused_po["amount"]), f"{unused_po['count']} open POs, no bill (ready-mix blankets may be intentional)", "a"),
         ("Unreconciled checks >30 days", _money(mor.get("aged_total")), f"{int(mor.get('aged_count') or 0)} checks NOT marked cleared (QBO can't confirm cashed — mark them in the register)", "a" if (mor.get("aged_count") or 0) else "g"),
         ("Sub LOC peak needed", _money(loc_peak), "high-water to float sub payments", "a"),
         ("Weekly sub/vendor burn", _money(burn), "recurring outflow", "a")],
        extra=(f'<div class="aging"><div class="agehd">AP aging</div>{_bar(health.get("ap_buckets", {}))}</div>'
               f'<div class="aging"><div class="agehd">Sub LOC peak by division</div>{div_bars}</div>'))

    # ---- POSITION ----
    over_under = (wip["over"] or 0) - (wip["under"] or 0)
    runway_bad = _num(runway) is not None and runway < 8
    cov_bad = _num(cov) is not None and cov < 1
    position = _section(
        "Position — where we stand", "pos",
        [(_money(cash), _money(cash), "r" if (_num(cash) or 0) < 0 else "n"),
         (f"{_num(runway):.1f} wk" if _num(runway) is not None else "—",
          f"{_num(runway):.1f} wk" if _num(runway) is not None else "—", "r" if runway_bad else "n"),
         (_pct(gm), _pct(gm), "n")],
        [("Cash (bank, excl. credit cards)", _money(cash), "spendable now", "r" if (_num(cash) or 0) < 0 else "n"),
         ("Runway at current burn", f"{_num(runway):.1f} weeks" if _num(runway) is not None else "—", "weeks of cash left" + (" — TIGHT" if runway_bad else ""), "r" if runway_bad else "n"),
         ("Coverage — AR vs AP", f"{_num(cov):.2f}" if _num(cov) is not None else "—", "AR ÷ AP; <1 = inflows don't cover outflows", "r" if cov_bad else "n"),
         ("Over / under-billing net (WIP)", _money(over_under), ("overbilled — liability" if over_under > 0 else "underbilled — bill it"), "a" if over_under > 0 else "g"),
         ("Gross margin %", _pct(gm), "revenue − direct cost", "n"),
         ("Net margin after overhead (YTD)", _pct(nm), "the real bottom line", "r" if (_num(nm) or 0) < 0.02 else "n"),
         ("Top-customer concentration", _pct(health.get("concentration")), html.escape(str(health.get("top_customer") or "")), "a" if (_num(health.get("concentration")) or 0) > 0.25 else "n")])

    badges = " ".join(_fresh_badge(p) for p in sources)
    return f'''<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Company Dashboard</title>
<style>
  :root {{ --bg:#eef1f6; --card:#fff; --ink:#1c2430; --muted:#6b7280;
    --navy:#1F3864; --green:#1F6B4C; --amber:#BF8F00; --red:#C00000;
    --line:rgba(120,130,150,.18); }}
  @media (prefers-color-scheme: dark) {{
    :root {{ --bg:#0f1420; --card:#171e2b; --ink:#e6eaf2; --muted:#98a2b3;
      --line:rgba(150,160,180,.16); }}
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--bg); color:var(--ink);
    font:14.5px/1.45 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; }}
  header {{ background:var(--navy); color:#fff; padding:18px 30px; }}
  header h1 {{ margin:0; font-size:21px; }}
  header .sub {{ opacity:.82; font-size:12.5px; margin-top:3px; }}
  .asof {{ margin-left:0; }} .asof.bad {{ color:#ffb3b3; font-weight:600; }}
  main {{ max-width:1220px; margin:0 auto; padding:22px 30px 20px;
    display:grid; grid-template-columns:1fr 1fr; gap:22px; }}
  .sec {{ background:var(--card); border-radius:12px; padding:16px 18px 18px;
    box-shadow:0 1px 4px rgba(0,0,0,.07); border-top:4px solid var(--muted); }}
  .sec.in {{ border-top-color:var(--green); }}
  .sec.out {{ border-top-color:var(--red); }}
  .sec.pos {{ border-top-color:var(--navy); grid-column:1 / -1; }}
  .sec h2 {{ margin:2px 0 12px; font-size:14px; letter-spacing:.04em;
    text-transform:uppercase; color:var(--muted); }}
  .heroes {{ display:flex; gap:26px; flex-wrap:wrap; margin-bottom:12px;
    padding-bottom:12px; border-bottom:1px solid var(--line); }}
  .hero .hv {{ font-size:27px; font-weight:750; letter-spacing:-.5px; }}
  .hero .hl {{ font-size:11px; color:var(--muted); margin-top:1px; }}
  .hero.g .hv {{ color:var(--green); }} .hero.r .hv {{ color:var(--red); }}
  .hero.a .hv {{ color:var(--amber); }} .hero.n .hv {{ color:var(--ink); }}
  table.metrics {{ width:100%; border-collapse:collapse; }}
  table.metrics td {{ padding:7px 6px; border-bottom:1px solid var(--line);
    vertical-align:top; }}
  td.m {{ font-weight:600; width:44%; }}
  td.v {{ text-align:right; font-weight:750; white-space:nowrap; width:20%;
    font-variant-numeric:tabular-nums; }}
  td.v.g {{ color:var(--green); }} td.v.r {{ color:var(--red); }}
  td.v.a {{ color:var(--amber); }} td.v.n {{ color:var(--ink); }}
  td.d {{ color:var(--muted); font-size:12px; }}
  .aging {{ margin-top:12px; }}
  .agehd {{ font-size:11px; text-transform:uppercase; letter-spacing:.05em;
    color:var(--muted); margin-bottom:5px; }}
  .agerow {{ display:grid; grid-template-columns:56px 1fr; align-items:center;
    gap:8px; margin:3px 0; }}
  .agelbl {{ font-size:12px; color:var(--muted); }}
  .agetrack {{ background:var(--line); border-radius:5px; height:20px; }}
  .agefill {{ height:20px; border-radius:5px; color:#fff; font-size:11px;
    font-weight:600; display:flex; align-items:center; padding-left:7px;
    white-space:nowrap; min-width:40px; }}
  footer {{ max-width:1220px; margin:0 auto; padding:6px 30px 40px; }}
  footer h3 {{ font-size:12px; text-transform:uppercase; color:var(--muted);
    letter-spacing:.05em; }}
  .badge {{ display:inline-block; font-size:11px; padding:4px 9px;
    border-radius:20px; margin:3px 4px 0 0; background:var(--line);
    color:var(--muted); }}
  .badge.stale {{ background:rgba(192,0,0,.15); color:var(--red); font-weight:600; }}
  .badge.fresh {{ background:rgba(31,107,76,.16); color:var(--green); }}
  .badge.missing {{ background:rgba(191,143,0,.18); color:var(--amber); }}
  @media (max-width:780px) {{ main {{ grid-template-columns:1fr; }}
    .sec.pos {{ grid-column:auto; }} }}
</style></head><body>
<header>
  <h1>Company Dashboard</h1>
  <div class="sub">Generated {now:%Y-%m-%d %H:%M} · {asof}</div>
</header>
<main>
  {money_in}
  {money_out}
  {position}
</main>
<footer>
  <h3>Sources &amp; freshness</h3>
  {badges}
</footer>
</body></html>'''


def main() -> int:
    ap = argparse.ArgumentParser(description="Consolidated company dashboard (HTML)")
    ap.add_argument("--out", type=Path, default=OUT_PATH)
    ap.add_argument("--open", action="store_true")
    args = ap.parse_args()

    print("\n  COMPANY DASHBOARD — Money In / Out / Position")
    print("  " + "─" * 48)
    cards = read_money_bleeds(MB_PATH)
    loc = read_sub_loc(LOC_PATH)
    mor = read_money_out(MOR_PATH)
    health = read_health(HEALTH_PATH)
    wip = read_wip(WIP_PATH)
    print(f"  MB cards {len(cards)} · LOC div {len(loc['divisions'])} · "
          f"WIP active {wip['count']} · cash {hk(health, 'Bank total')}")

    htmlout = render(cards, loc, mor, health, wip,
                     [MB_PATH, LOC_PATH, MOR_PATH, HEALTH_PATH, WIP_PATH, BILL_PATH])
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(htmlout, encoding="utf-8")
    os.chmod(args.out, stat.S_IRUSR | stat.S_IWUSR)
    print(f"\n  ✓ {args.out}  (chmod 600)")
    if args.open:
        webbrowser.open(f"file://{args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
