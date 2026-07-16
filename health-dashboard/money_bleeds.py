#!/usr/bin/env python3
"""
money_bleeds.py — the company-health exceptions report ("Money Bleeds").

Replaces the old KPI dashboard idea with a short list of things that are
provably wrong and cost money (the user 2026-07-16):

  1. DRAWS WITH NO INVOICE (MFD + CP only — RP never has a G702)
       MFD: active projects come from the 'WIP Master' tab; each project's
            DRAWS folder on the Multi Family volume is checked — LATEST
            numbered draw only ("we are already billing draws; this is for
            the future case — if the last draw has an invoice, move on").
       CP:  latest draw's G702 earned-less-retainage vs cumulative QBO
            invoiced for the project — a shortfall = a draw never invoiced.

  2. LIEN CLOCK on every open invoice (Texas Prop. Code Ch. 53)
       CP/MFD (commercial):  notice due the 15th of the 3rd month after the
                             WORK month.
       RP (residential):     15th of the 2nd month.
       Work month = INVOICE month (the user 2026-07-16: the RP invoice date
       is the last day on the job / finished 100%, and draws are billed in
       their work month — always run the clock from the invoice month).
       Deadlines roll BACKWARD off weekends (never forward). Retainage is a
       separate statutory track (completion-based, § 53.057) — retainage
       invoices are listed apart, not run through the monthly clock.
       Equipment-lease / note-payment invoices to subs are NOT construction
       income — excluded from the clock, listed on their own sheet.

  3. RP WRAP-UP — SLAB lines 100% complete in the General List but not
       fully billed: waiting on punch, but it must be chased to get paid.
       FTW lines are ignored — the General List's 100% column is slab-only
       (the FTW completion section isn't used). Read from the WIP workbook's
       'Test - RP' tab (rp_wip_reader output — run the WIP readers first).

OUTPUT
  ~/Documents/CompanyHealth/Money Bleeds.xlsx  (chmod 600 after write)
  Plain white/black formatting per repo rules. READ-ONLY everywhere else:
  QBO reads, WIP workbook reads, volume folder scans — no writes.

USAGE
  python3 health-dashboard/money_bleeds.py
  python3 health-dashboard/money_bleeds.py --out /path/x.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import stat
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import qbo_api
from shared import draws

# ────────────────────────── config / paths ──────────────────────────

DEFAULT_OUTPUT = paths.companyhealth_dir() / "Money Bleeds.xlsx"

WIP_EXCEL_PATH = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx",
)
MFD_ROOT = paths.get_path(
    "MFD_PROJECTS_DIR",
    "/Volumes/Multi Family/MULTY FAMILY DIVISION PROJECTS",
)
CP_ACTIVE_DIR = paths.get_path(
    "CP_ACTIVE_DIR",
    "/Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects",
)

MASTER_SHEET = "WIP Master"          # header row 3, data 4+ (col A proj, B name)
RP_TEST_SHEET = "Test - RP"          # rp_wip_reader output (header row 1)

# 'MFD177 - JPI MERRITT PARK' → MFD177 ; \b keeps MFD1770 from matching MFD177
_MFD_FOLDER_RE = re.compile(r"^(MFD\d{3,4})\b", re.IGNORECASE)
_CP_FOLDER_RE = re.compile(r"^(CP\d{3,4})\b", re.IGNORECASE)
# Draw folder naming varies per project: '13- JULY 2026 DRAW' (MFD177) vs
# '6-JULY DRAW 2026' / '2-FEB DRAW 2026' (MFD192/325). Parse tolerantly:
# leading sequence number, then a month name + 4-digit year anywhere.
_MFD_DRAW_NUM_RE = re.compile(r"^\s*(\d{1,2})\s*-")
_YEAR_RE = re.compile(r"\b(20\d{2})\b")
_MONTHS = {m.upper(): i + 1 for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"])}
_MONTHS.update({m.upper(): i + 1 for i, m in enumerate(
    ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
     "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"])})


def _month_from_token(token: str) -> Optional[int]:
    """'JUNE'/'FEB'/'SEPT'/'MARZO' → month number. Abbreviations are any
    ≥3-char prefix of a full month name (English or Spanish)."""
    t = token.strip().upper()
    if len(t) < 3:
        return None
    for name, num in _MONTHS.items():
        if name.startswith(t):
            return num
    return None


def _parse_mfd_draw_name(name: str) -> Optional[Tuple[int, int, int]]:
    """'6-JULY DRAW 2026' / '12- JUNE 2026 DRAW' → (num, month, year)."""
    m = _MFD_DRAW_NUM_RE.match(name)
    if not m:
        return None
    y = _YEAR_RE.search(name)
    if not y:
        return None
    for token in re.findall(r"[A-Za-z]+", name):
        month = _month_from_token(token)
        if month:
            return int(m.group(1)), month, int(y.group(1))
    return None

_SKIP_DIRS = {"#recycle", "@eadir", "completed jobs", "z - non active or bidding"}

# Lien clock (Texas Prop. Code Ch. 53, first-tier subcontractor):
#   commercial notice = 15th of the 3rd month after the work month
#   residential       = 15th of the 2nd month
NOTICE_MONTHS = {"MFD": 3, "CP": 3, "RP": 2}
URGENT_DAYS = 15
WATCH_DAYS = 45

_RETAINAGE_RE = re.compile(r"retainage|retention|retenci[oó]n", re.IGNORECASE)

# Equipment leases / note payments invoiced to subs are NOT construction
# income (the user 2026-07-16) — no lien rights ride on them. Matched on
# line items + descriptions ('Other Charges:Equipment Lease', 'Note
# Principal Payment', 'Interest fee', 'Pump #14', 'Tractor #179', …).
_NON_CONSTRUCTION_RE = re.compile(
    r"equipment lease|monthly equipment|note principal|principal payment"
    r"|interest (charge|fee)|lease payment", re.IGNORECASE)
_EQUIP_ITEM_RE = re.compile(
    r"^(pump|tractor|truck|dump truck|trailer|excavator|other charges)\b",
    re.IGNORECASE)

CP_TOLERANCE = 1.0   # $ tolerance G702 earned vs QBO invoiced


# ────────────────────────── small helpers ──────────────────────────

def _today() -> dt.date:
    return dt.date.today()


def _add_months(year: int, month: int, n: int) -> Tuple[int, int]:
    m = month - 1 + n
    return year + m // 12, m % 12 + 1


def _roll_back_weekend(d: dt.date) -> dt.date:
    """Statute rule: deadlines roll BACKWARD to the prior business day,
    never forward. (Holidays not modeled — weekend-only.)"""
    while d.weekday() >= 5:          # 5=Sat, 6=Sun
        d -= dt.timedelta(days=1)
    return d


def notice_deadline(work_year: int, work_month: int, division: str) -> dt.date:
    n = NOTICE_MONTHS.get(division, NOTICE_MONTHS["RP"])   # unknown → shorter clock
    y, m = _add_months(work_year, work_month, n)
    return _roll_back_weekend(dt.date(y, m, 15))


def _division(customer_name: str) -> str:
    """Division from the invoice's customer name. MFD = parent customer
    'Multi Family' (never the Class field); else project-# prefix."""
    name = customer_name or ""
    if name.strip().lower().startswith("multi family"):
        return "MFD"
    proj = qbo_api.extract_proj(name)
    if proj:
        return "MFD" if proj.startswith("MFD") else ("CP" if proj.startswith("CP") else "RP")
    return "?"


def _parent(customer_name: str) -> str:
    """'Multi Family:MFD177 JPI Merritt Park' → 'Multi Family'; top-level
    customers have no parent."""
    return customer_name.split(":")[0].strip() if ":" in (customer_name or "") else ""


def _is_retainage_invoice(inv: dict) -> bool:
    texts = [inv.get("PrivateNote") or "",
             (inv.get("CustomerMemo") or {}).get("value") or ""]
    for line in inv.get("Line", []) or []:
        texts.append(line.get("Description") or "")
    return any(_RETAINAGE_RE.search(t) for t in texts if t)


def _is_lease_invoice(inv: dict) -> bool:
    """Equipment lease / note-payment invoices to subs — not construction
    income. Text match first; else every itemized line is an equipment or
    'Other Charges' item."""
    items, texts = [], []
    for line in inv.get("Line", []) or []:
        det = line.get("SalesItemLineDetail") or {}
        item = (det.get("ItemRef") or {}).get("name") or ""
        if item:
            items.append(item)
        if line.get("Description"):
            texts.append(line["Description"])
    if any(_NON_CONSTRUCTION_RE.search(t) for t in items + texts):
        return True
    return bool(items) and all(_EQUIP_ITEM_RE.match(i) for i in items)


def _parse_date(s: str) -> Optional[dt.date]:
    try:
        return dt.date.fromisoformat(s)
    except (TypeError, ValueError):
        return None


# ────────────────── check 1a — MFD draws vs invoices ────────────────

def read_active_mfd(wip_path: Path) -> List[Tuple[str, str]]:
    """MFD#### rows from the 'WIP Master' tab — the master sheet IS the MFD
    source for active jobs (same rule as wip/master_wip_test.py)."""
    wb = load_workbook(wip_path, data_only=True, read_only=True)
    ws = wb[MASTER_SHEET]
    out = []
    for r in range(4, ws.max_row + 1):
        proj = ws.cell(r, 1).value
        if proj and str(proj).strip().upper().startswith("MFD"):
            out.append((str(proj).strip().upper(),
                        str(ws.cell(r, 2).value or "").strip()))
    wb.close()
    return out


def find_mfd_project_folder(proj: str) -> Optional[Path]:
    """<MFD root>/<CLIENT>/<MFD### - NAME>. Scans client dirs (skipping
    recycle/completed/non-active); word-boundary match so MFD177 never
    matches MFD1770."""
    want = re.compile(rf"^{re.escape(proj)}\b", re.IGNORECASE)
    try:
        clients = [d for d in MFD_ROOT.iterdir() if d.is_dir()
                   and d.name.strip().lower() not in _SKIP_DIRS]
    except OSError:
        return None
    for client in clients:
        try:
            for sub in client.iterdir():
                if sub.is_dir() and want.match(sub.name.strip()):
                    return sub
        except OSError:
            continue
    return None


def find_mfd_draws_folder(project_folder: Path) -> Optional[Path]:
    """'PM MISC/DRAWS' first (the standard layout), else any 'Draw(s)' dir
    within two levels."""
    direct = project_folder / "PM MISC" / "DRAWS"
    if direct.is_dir():
        return direct
    for depth1 in project_folder.iterdir():
        if not depth1.is_dir():
            continue
        if draws.DRAWS_FOLDER_RE.match(depth1.name.strip()):
            return depth1
        try:
            for depth2 in depth1.iterdir():
                if depth2.is_dir() and draws.DRAWS_FOLDER_RE.match(depth2.name.strip()):
                    return depth2
        except OSError:
            continue
    return None


def latest_mfd_draw(draws_dir: Path) -> Optional[Tuple[int, dt.date, str]]:
    """Highest-numbered non-empty 'N- MONTH YEAR DRAW' folder →
    (draw_num, first-of-draw-month, folder name)."""
    best = None
    try:
        entries = list(draws_dir.iterdir())
    except OSError:
        return None
    for entry in entries:
        if not entry.is_dir():
            continue
        parsed = _parse_mfd_draw_name(entry.name)
        if not parsed:
            continue
        num, month, year = parsed
        try:                                     # empty folder = draw not built yet
            if not any(entry.iterdir()):
                continue
        except OSError:
            continue
        if best is None or num > best[0]:
            best = (num, dt.date(year, month, 1), entry.name.strip())
    return best


def check_mfd_draws(access: str, company_id: str,
                    proj_map: Dict[str, dict]) -> List[Dict[str, Any]]:
    """One row per active MFD project: latest draw folder vs latest QBO
    invoice date. PASS = latest invoice lands in/after the draw month
    (we're already billing draws — this is the future-case tripwire)."""
    results = []
    today = _today()
    for proj, name in read_active_mfd(WIP_EXCEL_PATH):
        row = {"project": proj, "name": name, "draw": "", "draw_month": None,
               "last_invoice": None, "verdict": "", "detail": ""}
        results.append(row)

        folder = find_mfd_project_folder(proj)
        if folder is None:
            row["verdict"] = "REVIEW"
            row["detail"] = "No project folder found on the Multi Family volume"
            continue
        draws_dir = find_mfd_draws_folder(folder)
        if draws_dir is None:
            row["verdict"] = "REVIEW"
            row["detail"] = f"No DRAWS folder under {folder.name}"
            continue
        latest = latest_mfd_draw(draws_dir)
        if latest is None:
            row["verdict"] = "REVIEW"
            row["detail"] = "DRAWS folder has no numbered draw subfolders"
            continue
        draw_num, draw_month, draw_label = latest
        row["draw"] = draw_label
        row["draw_month"] = draw_month

        cust = proj_map.get(proj)
        if cust is None:
            row["verdict"] = "RED"
            row["detail"] = "No QBO project customer matches this project #"
            continue
        invs = qbo_api.fetch_customer_invoices(access, company_id, cust["id"])
        last = max((_parse_date(i.get("TxnDate")) for i in invs
                    if _parse_date(i.get("TxnDate"))), default=None)
        row["last_invoice"] = last

        if last is not None and last >= draw_month:
            row["verdict"] = "PASS"
            row["detail"] = f"Draw {draw_num} invoiced ({last.isoformat()})"
        elif (draw_month.year, draw_month.month) == (today.year, today.month):
            row["verdict"] = "PENDING"
            row["detail"] = f"Draw {draw_num} is current-month — invoice not in QBO yet"
        else:
            row["verdict"] = "RED"
            row["detail"] = (f"Latest draw ({draw_label}) has NO QBO invoice — "
                             f"last invoice {last.isoformat() if last else 'NEVER'}")
    return results


# ────────────────── check 1b — CP draws vs invoices ────────────────

def check_cp_draws(access: str, company_id: str,
                   proj_map: Dict[str, dict]) -> List[Dict[str, Any]]:
    """One row per active CP project folder that has a draw: G702
    earned-less-retainage (cumulative, what should be invoiced) vs the sum
    of ALL QBO invoices for the project. Shortfall = a draw never invoiced."""
    results = []
    try:
        folders = sorted(d for d in CP_ACTIVE_DIR.iterdir() if d.is_dir())
    except OSError as e:
        raise SystemExit(f"✗ CP active dir unreadable: {CP_ACTIVE_DIR} ({e})")
    for folder in folders:
        m = _CP_FOLDER_RE.match(folder.name.strip())
        if not m:
            continue
        proj = m.group(1).upper()
        found = draws.find_latest_draw(folder)
        if not found:
            continue                              # no draw yet — nothing to verify
        draw_num, draw_file = found
        row = {"project": proj, "name": folder.name, "draw_num": draw_num,
               "earned": None, "invoiced": None, "verdict": "", "detail": ""}
        results.append(row)

        data, flags = draws.read_draw_g702(draw_file)
        earned = (data or {}).get("earned_less_retainage")
        if earned is None:
            earned = (data or {}).get("billed")    # fallback: gross billed
        if earned is None:
            row["verdict"] = "REVIEW"
            row["detail"] = "; ".join(flags) or f"Draw #{draw_num} G702 unreadable"
            continue
        row["earned"] = earned

        cust = proj_map.get(proj)
        if cust is None:
            row["verdict"] = "RED"
            row["detail"] = "No QBO project customer matches this project #"
            continue
        invs = qbo_api.fetch_customer_invoices(access, company_id, cust["id"])
        invoiced = sum(float(i.get("TotalAmt") or 0) for i in invs)
        row["invoiced"] = invoiced

        gap = earned - invoiced
        if gap > CP_TOLERANCE:
            row["verdict"] = "RED"
            row["detail"] = (f"Draw #{draw_num} earned ${earned:,.2f} but QBO "
                             f"invoiced ${invoiced:,.2f} — ${gap:,.2f} never invoiced")
        else:
            row["verdict"] = "PASS"
            row["detail"] = f"QBO invoiced covers draw #{draw_num}"
    return results


# ────────────────── check 2 — lien clock on open invoices ───────────

def check_lien_clock(access: str, company_id: str,
                     proj_map: Dict[str, dict]) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Every open invoice → notice deadline + days left. Work month =
    invoice month (the user 2026-07-16: RP invoices go out the day the job
    finishes; draws bill their work month). Returns (monthly_rows,
    retainage_rows, lease_rows) — retainage runs its own statutory track;
    equipment-lease/note invoices aren't construction income at all."""
    today = _today()
    monthly, retainage, leases = [], [], []
    for inv in qbo_api.query_all(access, company_id, "Invoice", "Balance > '0'"):
        txn = _parse_date(inv.get("TxnDate"))
        if txn is None:
            continue
        cust_name = (inv.get("CustomerRef") or {}).get("name") or ""
        div = _division(cust_name)
        proj = qbo_api.extract_proj(cust_name) or ""
        # Invoice CustomerRef.name is the bare display name — the parent
        # chain comes from the customer map's FullyQualifiedName.
        parent = _parent(cust_name)
        if not parent and proj and proj in proj_map:
            fqn = proj_map[proj].get("fully_qualified_name", "")
            parent = ":".join(fqn.split(":")[:-1])
        row = {
            "division": div if div != "?" else "? (review)",
            "project": proj,
            "parent": parent,
            "customer": cust_name,
            "doc": inv.get("DocNumber") or "",
            "txn_date": txn,
            "balance": float(inv.get("Balance") or 0),
        }
        if _is_lease_invoice(inv):
            row["note"] = "Equipment lease / note payment — not construction income"
            leases.append(row)
            continue
        if _is_retainage_invoice(inv):
            row["note"] = ("Retainage — separate track (§ 53.057, keyed to "
                           "completion, not the monthly clock)")
            retainage.append(row)
            continue
        deadline = notice_deadline(txn.year, txn.month, div if div != "?" else "RP")
        days = (deadline - today).days
        row["work_month"] = dt.date(txn.year, txn.month, 1)
        row["deadline"] = deadline
        row["days_left"] = days
        row["status"] = ("PAST" if days < 0 else
                         "URGENT" if days <= URGENT_DAYS else
                         "WATCH" if days <= WATCH_DAYS else "OK")
        monthly.append(row)
    monthly.sort(key=lambda r: r["days_left"])
    retainage.sort(key=lambda r: r["txn_date"])
    leases.sort(key=lambda r: -r["balance"])
    return monthly, retainage, leases


# ────────────────── check 3 — RP wrap-up (waiting on punch) ─────────

def check_rp_wrapup(wip_path: Path) -> Tuple[List[Dict[str, Any]], str]:
    """'Test - RP' SLAB rows 100% complete in the General List but not fully
    billed — waiting on punch, chase to get paid. FTW lines are skipped: the
    General List's 100% column is slab-only; the FTW completion section
    isn't used (the user 2026-07-16). Returns (rows, synced)."""
    wb = load_workbook(wip_path, data_only=True, read_only=True)
    ws = wb[RP_TEST_SHEET]
    hdr = {str(ws.cell(1, c).value or "").strip().upper(): c
           for c in range(1, ws.max_column + 1)}

    def col(r, label):
        c = hdr.get(label)
        return ws.cell(r, c).value if c else None

    rows, synced = [], ""
    for r in range(2, ws.max_row + 1):
        proj = col(r, "PROJECT #")
        if not proj:
            continue
        synced = str(col(r, "LAST SYNCED") or synced)
        proj = str(proj).strip().upper()
        if proj.endswith("-FTW"):
            continue                      # 100% is slab-only — FTW ignored
        notes = str(col(r, "NOTES") or "")
        status = str(col(r, "STATUS") or "").strip()
        if "100% complete (list)" not in notes or status.lower() != "active":
            continue
        contract = float(col(r, "TOTAL CONTRACT PRICE") or 0)
        billed = float(col(r, "BILLED TO DATE") or 0)
        left = float(col(r, "LEFT TO BILL") or (contract - billed))
        if left <= 0:
            continue
        # WHY: the classification segment, minus the RED echo + the 100% tag
        why = "; ".join(
            s.strip() for s in notes.split(";")
            if s.strip() and not s.strip().startswith("RED:")
            and s.strip() != "100% complete (list)") or notes
        rows.append({"project": proj, "name": str(col(r, "PROJECT NAME") or ""),
                     "client": str(col(r, "CLIENT") or ""),
                     "contract": contract, "billed": billed, "left": left,
                     "note": why})
    wb.close()
    rows.sort(key=lambda x: -x["left"])
    return rows, synced


# ────────────────────────── Excel output ─────────────────────────────
# Color + framing per the user's explicit ask (2026-07-16) — this workbook
# is exempt from the plain-white/black repo rule. Label + amount stay on
# the same row; separate sheets over crowding one.

CUR = '#,##0.00'
_BOLD = Font(bold=True)
_HDR_FILL = PatternFill("solid", fgColor="1F3864")           # dark navy
_HDR_FONT = Font(bold=True, color="FFFFFF")
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT = Font(bold=True, color="9C0006")
_AMB_FILL = PatternFill("solid", fgColor="FFEB9C")
_AMB_FONT = Font(color="9C6500")
_GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
_GRN_FONT = Font(color="006100")
_ROW_RED = PatternFill("solid", fgColor="FDECEA")            # light row wash
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_STYLE = {
    "PASS": (_GRN_FILL, _GRN_FONT), "OK": (_GRN_FILL, _GRN_FONT),
    "PENDING": (_AMB_FILL, _AMB_FONT), "WATCH": (_AMB_FILL, _AMB_FONT),
    "REVIEW": (_AMB_FILL, _AMB_FONT),
    "RED": (_RED_FILL, _RED_FONT), "PAST": (_RED_FILL, _RED_FONT),
    "URGENT": (_RED_FILL, _RED_FONT),
}


def _sheet(wb: Workbook, title: str, headers: List[str],
           widths: List[int]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A2"
    return ws


def _frame_row(ws, ncols: int, status: str = "", status_col: int = 0) -> None:
    """Borders on the just-appended row; status cell colored by verdict;
    red verdicts wash the whole row."""
    r = ws.max_row
    fill_font = _STATUS_STYLE.get(status)
    for c in range(1, ncols + 1):
        ws.cell(r, c).border = _BORDER
        if fill_font and fill_font[0] is _RED_FILL and c != status_col:
            ws.cell(r, c).fill = _ROW_RED
    if fill_font and status_col:
        ws.cell(r, status_col).fill = fill_font[0]
        ws.cell(r, status_col).font = fill_font[1]


def _finish(ws, ncols: int) -> None:
    ws.auto_filter.ref = f"A1:{chr(64 + ncols)}{ws.max_row}"


def write_workbook(out: Path,
                   mfd: List[dict], cp: List[dict],
                   lien: List[dict], reten: List[dict], leases: List[dict],
                   wrapup: List[dict], rp_synced: str) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"

    # ── bleed sheets first (dashboard summarizes them) ──
    ws = _sheet(wb, "Draws MFD", ["PROJECT #", "PROJECT NAME", "LATEST DRAW",
                                  "LAST QBO INVOICE", "VERDICT", "DETAIL"],
                [11, 30, 26, 17, 10, 70])
    for r in mfd:
        ws.append([r["project"], r["name"], r["draw"],
                   r["last_invoice"].isoformat() if r["last_invoice"] else "—",
                   r["verdict"], r["detail"]])
        _frame_row(ws, 6, r["verdict"], 5)
    _finish(ws, 6)

    ws = _sheet(wb, "Draws CP", ["PROJECT #", "FOLDER", "DRAW #",
                                 "G702 EARNED (LESS RET.) $", "QBO INVOICED $",
                                 "VERDICT", "DETAIL"],
                [11, 34, 8, 22, 16, 10, 70])
    for r in cp:
        ws.append([r["project"], r["name"], r["draw_num"], r["earned"],
                   r["invoiced"], r["verdict"], r["detail"]])
        ws.cell(ws.max_row, 4).number_format = CUR
        ws.cell(ws.max_row, 5).number_format = CUR
        _frame_row(ws, 7, r["verdict"], 6)
    _finish(ws, 7)

    ws = _sheet(wb, "Lien Clock", ["STATUS", "DAYS LEFT", "NOTICE DEADLINE",
                                   "DIVISION", "PARENT", "PROJECT #",
                                   "INVOICE #", "INVOICE DATE", "WORK MONTH",
                                   "OPEN BALANCE $", "CUSTOMER"],
                [9, 10, 16, 10, 20, 11, 11, 12, 12, 15, 45])
    for r in lien:
        ws.append([r["status"], r["days_left"], r["deadline"].isoformat(),
                   r["division"], r["parent"], r["project"], r["doc"],
                   r["txn_date"].isoformat(), r["work_month"].strftime("%b %Y"),
                   r["balance"], r["customer"]])
        ws.cell(ws.max_row, 10).number_format = CUR
        _frame_row(ws, 11, r["status"], 1)
    _finish(ws, 11)

    ws = _sheet(wb, "Lien Retainage", ["DIVISION", "PARENT", "PROJECT #",
                                       "INVOICE #", "INVOICE DATE",
                                       "OPEN BALANCE $", "CUSTOMER", "NOTE"],
                [10, 20, 11, 11, 12, 15, 40, 55])
    for r in reten:
        ws.append([r["division"], r["parent"], r["project"], r["doc"],
                   r["txn_date"].isoformat(), r["balance"], r["customer"],
                   r["note"]])
        ws.cell(ws.max_row, 6).number_format = CUR
        _frame_row(ws, 8)
    _finish(ws, 8)

    ws = _sheet(wb, "Leases (excluded)", ["PARENT", "CUSTOMER", "INVOICE #",
                                          "INVOICE DATE", "OPEN BALANCE $",
                                          "WHY EXCLUDED"],
                [20, 40, 11, 12, 15, 55])
    for r in leases:
        ws.append([r["parent"], r["customer"], r["doc"],
                   r["txn_date"].isoformat(), r["balance"], r["note"]])
        ws.cell(ws.max_row, 5).number_format = CUR
        _frame_row(ws, 6)
    _finish(ws, 6)

    ws = _sheet(wb, "RP Wrap-Up", ["PROJECT #", "PROJECT NAME", "CLIENT",
                                   "CONTRACT $", "BILLED $", "LEFT TO BILL $",
                                   "WHY"],
                [12, 26, 24, 14, 14, 15, 60])
    for r in wrapup:
        ws.append([r["project"], r["name"], r["client"], r["contract"],
                   r["billed"], r["left"], r["note"]])
        for c in (4, 5, 6):
            ws.cell(ws.max_row, c).number_format = CUR
        _frame_row(ws, 7)
        ws.cell(ws.max_row, 6).font = _RED_FONT
    _finish(ws, 7)

    # ── dashboard ──
    mfd_red = [r for r in mfd if r["verdict"] == "RED"]
    mfd_rev = [r for r in mfd if r["verdict"] in ("REVIEW", "PENDING")]
    cp_red = [r for r in cp if r["verdict"] == "RED"]
    cp_gap = sum((r["earned"] or 0) - (r["invoiced"] or 0) for r in cp_red)
    past = [r for r in lien if r["status"] == "PAST"]
    urgent = [r for r in lien if r["status"] == "URGENT"]
    watch = [r for r in lien if r["status"] == "WATCH"]
    wrap_total = sum(r["left"] for r in wrapup)

    def _row(label, amount=None, detail="", bad=False):
        dash.append([label, amount, detail])
        r = dash.max_row
        for c in (1, 2, 3):
            dash.cell(r, c).border = _BORDER
            dash.cell(r, c).fill = _ROW_RED if bad else _GRN_FILL
        dash.cell(r, 1).font = _RED_FONT if bad else _GRN_FONT
        if amount is not None:
            dash.cell(r, 2).number_format = CUR
            dash.cell(r, 2).font = _RED_FONT if bad else _GRN_FONT

    dash.append(["MONEY BLEEDS — company health exceptions"])
    for c in (1, 2, 3):
        dash.cell(1, c).fill = _HDR_FILL
    dash.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    dash.append([f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} · lien clock runs "
                 f"from the invoice month (RP invoices = job finished; draws bill "
                 f"their work month)"])
    dash.append([f"RP data from '{RP_TEST_SHEET}' tab (slab lines only), "
                 f"last synced {rp_synced or 'unknown'}"])
    dash.append([])
    _row(f"MFD draws with NO invoice: {len(mfd_red)}", None,
         "; ".join(r["project"] for r in mfd_red) or "all clear",
         bad=bool(mfd_red))
    _row(f"MFD needs review / pending: {len(mfd_rev)}", None,
         "; ".join(r["project"] for r in mfd_rev) or "—",
         bad=bool(mfd_rev))
    _row(f"CP draws under-invoiced: {len(cp_red)}", cp_gap or None,
         "; ".join(r["project"] for r in cp_red) or "all clear",
         bad=bool(cp_red))
    dash.append([])
    _row(f"Lien notice PAST deadline: {len(past)}",
         sum(r["balance"] for r in past) or None,
         "money with no lien backup — see Lien Clock", bad=bool(past))
    _row(f"URGENT (≤{URGENT_DAYS} days): {len(urgent)}",
         sum(r["balance"] for r in urgent) or None,
         "notice must go out now", bad=bool(urgent))
    _row(f"WATCH (≤{WATCH_DAYS} days): {len(watch)}",
         sum(r["balance"] for r in watch) or None, "", bad=False)
    _row(f"Retainage invoices (own track): {len(reten)}",
         sum(r["balance"] for r in reten) or None,
         "completion-based deadline — not the monthly clock", bad=False)
    _row(f"Lease/note invoices excluded: {len(leases)}",
         sum(r["balance"] for r in leases) or None,
         "equipment leases to subs — not construction income", bad=False)
    dash.append([])
    _row(f"RP slabs 100% waiting on punch: {len(wrapup)}", wrap_total or None,
         "wrap up to get paid — see RP Wrap-Up", bad=bool(wrapup))
    dash.column_dimensions["A"].width = 44
    dash.column_dimensions["B"].width = 16
    dash.column_dimensions["C"].width = 70

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    os.chmod(out, stat.S_IRUSR | stat.S_IWUSR)   # 600 — owner-only


# ────────────────────────── main ─────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="Money Bleeds — company health exceptions report")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    args = ap.parse_args()

    print("\n  MONEY BLEEDS — draws↔invoices · lien clock · RP wrap-up")
    print("  " + "─" * 60)

    # Hard-fail preflight — mounts must be present, same pattern as Synology.
    problems = []
    if not WIP_EXCEL_PATH.exists():
        problems.append(f"WIP workbook not found: {WIP_EXCEL_PATH} (OneDrive synced?)")
    if not MFD_ROOT.is_dir():
        problems.append(f"Multi Family volume not mounted: {MFD_ROOT}")
    if not CP_ACTIVE_DIR.is_dir():
        problems.append(f"Common volume not mounted: {CP_ACTIVE_DIR}")
    if problems:
        for p in problems:
            print(f"  ✗ {p}")
        return 2

    access, company_id = qbo_api.load_credentials()
    print("  QBO project-customer map …")
    proj_map = qbo_api.build_project_customer_map(access, company_id)

    print("  1a. MFD draws vs invoices …")
    mfd = check_mfd_draws(access, company_id, proj_map)
    print(f"      {len(mfd)} active MFD project(s), "
          f"{sum(1 for r in mfd if r['verdict'] == 'RED')} RED")

    print("  1b. CP draws vs invoices …")
    cp = check_cp_draws(access, company_id, proj_map)
    print(f"      {len(cp)} CP project(s) with draws, "
          f"{sum(1 for r in cp if r['verdict'] == 'RED')} RED")

    print("  2.  Lien clock (work month = invoice month) …")
    lien, reten, leases = check_lien_clock(access, company_id, proj_map)
    past = sum(1 for r in lien if r["status"] == "PAST")
    urgent = sum(1 for r in lien if r["status"] == "URGENT")
    print(f"      {len(lien)} open invoice(s) on the clock — "
          f"{past} PAST, {urgent} URGENT · {len(reten)} retainage (own track) "
          f"· {len(leases)} lease/note excluded")

    print("  3.  RP wrap-up (slabs 100%, waiting on punch) …")
    wrapup, rp_synced = check_rp_wrapup(WIP_EXCEL_PATH)
    print(f"      {len(wrapup)} slab(s), ${sum(r['left'] for r in wrapup):,.0f} left to bill")

    write_workbook(args.out, mfd, cp, lien, reten, leases, wrapup, rp_synced)
    print(f"\n  ✓ {args.out}  (chmod 600)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
