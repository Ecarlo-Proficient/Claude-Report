#!/usr/bin/env python3
"""
load_bill_tracker.py — land Bill Tracker's vendor bills into the ledger (AP + liens).

WHAT IT DOES
Reads the line-level display sheets (Bills, Inventory) of "Bill Tracker.xlsx" and
fills `ap_bill_line` — vendor, project, account, amount, open balance, pay status,
and the Texas lien clock per bill line.

WHAT IT IS *NOT*
Not the cost ledger. Bill Tracker's display sheets EXCLUDE subs, and for a sub-based
labor company subs are most of the cost — so this understates job cost by design
(measured 25–98% short vs the QBO WIP truth). Job cost stays in wip_snapshot; the
complete cost_line (incl subs + true SL/PV cost codes) comes later from qbo-export.
What this uniquely adds is AP pay status + lien deadlines the WIP snapshot lacks.

SAFETY
    * The workbook is opened READ-ONLY — this tool never writes the sheet.
    * Idempotent: each run FULL-REPLACES source='bill_tracker' rows, mirroring the file.
    * --dry-run parses and reports; writes nothing.

USAGE
    python3 ledger/load_bill_tracker.py --dry-run --show 8
    python3 ledger/load_bill_tracker.py
    python3 ledger/load_bill_tracker.py --db /path/ledger.sqlite3 --excel "/path/Bill Tracker.xlsx"
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths  # noqa: E402

HERE = Path(__file__).resolve().parent
SCHEMA_SQL = HERE / "schema.sql"

DEFAULT_DB = paths.get_path(
    "ACB_LEDGER_DB",
    Path.home() / "Library" / "Application Support" / "Proficient" / "ledger.sqlite3",
)
DEFAULT_EXCEL = paths.get_path(
    "ACB_BILL_TRACKER_XLSX",
    paths.onedrive_base() / "Automations-/Bill Tracker.xlsx",
)

SHEETS = ("Bills", "Inventory")   # the line-level display sheets
HEADER_ROW = 2                    # row 1 is the grouped banner; real headers are row 2

# The "Open" column carries `=HYPERLINK("…/app/bill?txnId=<id>","↗")` — a QBO deep
# link to the bill. data_only reads the cached "↗" glyph, so we re-open for formulas.
QBO_BILL_RE = re.compile(r"app/bill\?txnId=(\d+)", re.I)
QBO_BILL_URL = "https://qbo.intuit.com/app/bill?txnId={}"

# a leading "[TAG] " on the Matched Invoice cell (see excel_bill_sync._invoice_cell)
MATCH_TAG_RE = re.compile(r"^\s*\[([^\]]*)\]\s*(.*)$", re.S)

# canonical field -> Bill Tracker header label
FIELD_HEADERS = {
    "vendor":       "Vendor",
    "bill_ref":     "Bill #",
    "bill_date":    "Bill Date",
    "project_no":   "Project #",
    "division":     "Division",
    "account":      "Account",
    "description":  "Line Description",
    "line_amount":  "Line Amount",
    "bill_total":   "Bill Total",
    "open_balance": "Bill Open Bal",
    "pay_status":   "Pay Status",
    "approved":     "Approved",
    "lien_status":  "Lien",
    "matched_invoice": "Matched Invoice",
    "invoice_status":  "Invoice Status",
    "invoice_no":      "Invoice #",
    "gc_paid_date":    "GC Paid Date",
    "pay_date":        "Pay Date",
    "bt_key":          "_Key",
}

# lien states that put a bill on the action watchlist, most-urgent first
LIEN_RANK = {
    "Notice PAST due": 0, "Notice due in ≤7d": 1, "Notice due in ≤15d": 2,
    "Notice due in ≤30d": 3, "Notice Sent": 4, "Lien Filed": 5,
}


def _num(v):
    if isinstance(v, bool):
        return None
    return float(v) if isinstance(v, (int, float)) else None


def _text(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    return s or None


def _header_index(ws):
    idx = {}
    for j, cell in enumerate(ws[HEADER_ROW]):
        label = _text(cell.value)
        if label and label not in idx:
            idx[label] = j
    return idx


def read_sheet(ws):
    """Yield (excel_row, record) for each real bill line on a display sheet."""
    hidx = _header_index(ws)
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1
    ):
        def cell(header):
            j = hidx.get(header)
            return row[j] if (j is not None and j < len(row)) else None

        rec = {}
        for field, header in FIELD_HEADERS.items():
            raw = cell(header)
            rec[field] = _num(raw) if field in ("line_amount", "bill_total", "open_balance") else _text(raw)
        # a real line needs a vendor and an amount (skip spacers / totals)
        if not rec["vendor"] and rec["line_amount"] is None:
            continue
        # the tracker prefixes Matched Invoice with a "[...]" tag on special matches
        # ([DRAW] / [FULLY BILLED] / [PUSHED from Draw #3]); keep the tag beside the
        # bare "invoice — memo" so every bill on one draw shares the SAME key
        rec["match_tag"] = None
        m = MATCH_TAG_RE.match(rec.get("matched_invoice") or "")
        if m:
            rec["match_tag"], rec["matched_invoice"] = m.group(1).strip(), (m.group(2).strip() or None)
        rec["source_sheet"] = ws.title
        yield excel_row, rec


def read_bill_links(ws):
    """Map line_uid -> QBO bill deep link, read from each row's =HYPERLINK() formula.

    Needs a workbook opened data_only=False (the cached value is only the "↗" glyph).
    Scans the whole row for the app/bill?txnId= pattern so it never depends on the
    "Open" column staying in a fixed position.
    """
    links = {}
    for excel_row, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1), start=HEADER_ROW + 1):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "app/bill?txnId=" in v:
                m = QBO_BILL_RE.search(v)
                if m:
                    links[f"bt:{ws.title}:{excel_row}"] = QBO_BILL_URL.format(m.group(1))
                break
    return links


def load(excel_path: Path, db_path: Path, dry_run: bool, show: int):
    if not excel_path.exists():
        sys.exit(f"ERROR: Bill Tracker not found: {excel_path}\n"
                 f"Set ACB_BILL_TRACKER_XLSX or pass --excel.")
    print(f"Reading (read-only): {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    records = []
    for name in SHEETS:
        if name not in wb.sheetnames:
            print(f"  ! sheet not found, skipping: {name}")
            continue
        n = 0
        for excel_row, rec in read_sheet(wb[name]):
            rec["line_uid"] = f"bt:{name}:{excel_row}"
            records.append(rec)
            n += 1
        print(f"  {name:<10} -> {n:>5} bill lines")
    wb.close()

    # second pass (formulas only): attach each bill's QBO deep link from the "Open"
    # ↗ =HYPERLINK() cell — the cached value in the pass above is just the glyph.
    wbf = load_workbook(excel_path, read_only=True, data_only=False)
    links = {}
    for name in SHEETS:
        if name in wbf.sheetnames:
            links.update(read_bill_links(wbf[name]))
    wbf.close()
    for r in records:
        r["qbo_link"] = links.get(r["line_uid"])
    print(f"  QBO bill links matched: {sum(1 for r in records if r.get('qbo_link'))}/{len(records)}")

    open_total = sum(r["open_balance"] or 0 for r in records)
    watch = [r for r in records if (r["lien_status"] in LIEN_RANK)]
    with_proj = sum(1 for r in records if r["project_no"])
    print(f"\nTotal lines: {len(records)}   with project#: {with_proj}   "
          f"open AP: ${open_total:,.0f}   lien-watch: {len(watch)}")

    if dry_run:
        print("\n--dry-run: nothing written.")
        _print_watch(sorted(watch, key=lambda r: LIEN_RANK[r['lien_status']]), show)
        return

    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    con.execute("PRAGMA foreign_keys = ON;")
    con.executescript(SCHEMA_SQL.read_text(encoding="utf-8"))
    # migrate ap_bill_line if it predates the draw/invoice columns — its rows are
    # reloaded from Excel in THIS run, so a drop + recreate is lossless.
    have = {r[1] for r in con.execute("PRAGMA table_info(ap_bill_line)")}
    if "matched_invoice" not in have or "qbo_link" not in have:
        con.execute("DROP TABLE ap_bill_line")
        con.executescript(SCHEMA_SQL.read_text(encoding="utf-8"))
    elif "match_tag" not in have:
        con.execute("ALTER TABLE ap_bill_line ADD COLUMN match_tag TEXT")
    now = dt.datetime.now().isoformat(timespec="seconds")

    # full replace: this feed always mirrors the current file
    con.execute("DELETE FROM ap_bill_line WHERE source = 'bill_tracker'")
    cols = ["line_uid", "project_no", "division", "vendor", "bill_ref", "bill_date",
            "account", "description", "line_amount", "bill_total", "open_balance",
            "pay_status", "approved", "lien_status", "matched_invoice", "match_tag",
            "invoice_status", "invoice_no", "gc_paid_date", "pay_date", "bt_key", "qbo_link",
            "source_sheet", "loaded_at"]
    ph = ", ".join(f":{c}" for c in cols)
    for r in records:
        r["loaded_at"] = now
        con.execute(f"INSERT INTO ap_bill_line ({', '.join(cols)}) VALUES ({ph})",
                    {c: r.get(c) for c in cols})
    con.commit()
    print(f"\nWrote {len(records)} bill lines -> {db_path}")

    # how many AP projects aren't in the WIP-derived project table (informational)
    known = {r[0] for r in con.execute("SELECT project_no FROM project")}
    ap_projs = {r["project_no"] for r in records if r["project_no"]}
    off_wip = sorted(ap_projs - known)
    if off_wip:
        print(f"note: {len(off_wip)} AP project#s not in the WIP project table "
              f"(off-WIP / closed) — kept anyway: {', '.join(off_wip[:8])}"
              f"{' …' if len(off_wip) > 8 else ''}")
    _print_watch_db(con, show)
    con.close()


def _print_watch(watch, show):
    if show <= 0:
        return
    print(f"\nLien watch (top {show}):")
    for r in watch[:show]:
        print(f"  {(r['lien_status'] or ''):<20} {(r['project_no'] or '—'):<10} "
              f"{(r['vendor'] or '')[:26]:<26} bill {r['bill_ref'] or '—':<10} "
              f"open ${r['open_balance'] or 0:,.0f}")


def _print_watch_db(con, show):
    if show <= 0:
        return
    rows = con.execute(
        "SELECT lien_status, project_no, vendor, bill_ref, open_balance "
        "FROM ap_bill_line WHERE lien_status IS NOT NULL").fetchall()
    rows = [r for r in rows if r[0] in LIEN_RANK]
    rows.sort(key=lambda r: (LIEN_RANK[r[0]], -(r[4] or 0)))
    print(f"\nLien watch (top {show}):")
    for r in rows[:show]:
        print(f"  {r[0]:<20} {(r[1] or '—'):<10} {(r[2] or '')[:26]:<26} "
              f"bill {r[3] or '—':<10} open ${r[4] or 0:,.0f}")


def main():
    ap = argparse.ArgumentParser(description="Load Bill Tracker vendor bills (AP + liens) into the ledger.")
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Bill Tracker workbook (read-only).")
    ap.add_argument("--db", type=Path, default=DEFAULT_DB, help="SQLite ledger to write.")
    ap.add_argument("--dry-run", action="store_true", help="Parse and report; write nothing.")
    ap.add_argument("--show", type=int, default=0, help="Print N lien-watch rows.")
    args = ap.parse_args()
    load(args.excel, args.db, args.dry_run, args.show)


if __name__ == "__main__":
    main()
