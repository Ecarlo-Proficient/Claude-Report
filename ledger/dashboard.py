#!/usr/bin/env python3
"""
dashboard.py — a local web dashboard over the project ledger.

Reads the SQLite ledger (READ-ONLY) and serves a single-page dashboard at
http://127.0.0.1:<port>. No terminal, no SQL: portfolio KPIs, per-division
rollups, a searchable / sortable projects table, click-into-a-job detail, and
one-click copy + CSV export. Appearance — theme, font, size, density, content
width, which widgets and which columns show — is customizable in the UI and
saved in the browser (localStorage), per person.

SAFETY
  * Reads open the database READ-ONLY (SQLite mode=ro). The ONLY writes are the owner's
    own marks, each to its own tiny overlay table (never to the mirrored source tables):
    waiver (draw waivers) and bill_mark (lien tags → mirrored to the workbook on sync-ap).
  * The server binds to 127.0.0.1 only — it is not exposed on the network.

USAGE
  python3 ledger/dashboard.py                       # start + open your browser
  python3 ledger/dashboard.py --port 8787           # pick the port
  python3 ledger/dashboard.py --no-open             # don't auto-open a browser
  python3 ledger/dashboard.py --db /path/ledger.sqlite3
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import os
import platform
import re
import signal
import sqlite3
import subprocess
import sys
import threading
import time
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths, pnl_paths, bill_marks, lien_clock  # noqa: E402

import registry_view  # noqa: E402  (local: parses the vault's process registry for the Systems tab)
import vault_graph    # noqa: E402  (local: vault [[link]] graph + docs/ARCHITECTURE.md diagrams for the Graph tab)

HERE = Path(__file__).resolve().parent
STATIC = HERE / "static"

# Dashboard build version - shown in the top bar so the owner can confirm which build is
# live. Bump on every user-visible release. 1.0.0 = Open Invoices tab + lien columns;
# 1.0.1 = Open Invoices client shows the parent GC (not the project-level name);
# 1.1.0 = Systems tab - the process registry rendered live from the vault.
LEDGER_VERSION = "1.1.0"

DEFAULT_DB = paths.get_path(
    "ACB_LEDGER_DB",
    Path.home() / "Library" / "Application Support" / "Proficient" / "ledger.sqlite3",
)

CONTENT_TYPES = {
    ".html": "text/html; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".svg": "image/svg+xml",
}

# ── P&L link (project-pnl) ──────────────────────────────────────────────────
# The dashboard can OPEN an existing per-project P&L and, on an explicit owner
# confirm, RUN project-pnl to (re)generate it. Generation shells out to the tool's
# own CLI (run_pnl.sh) — a subprocess, never an import (tools never import tools).
# QBO stays read-only inside project-pnl; the ONE data write (the .xlsx) is gated
# behind the UI confirm + a confirm flag on the request. Logs land under
# ~/Library/Logs/Proficient/ (never inside the repo).
_PNL_DIR = PROJECT_ROOT / "project-pnl"
_PNL_RUN = _PNL_DIR / "run_pnl.sh"
_PNL_LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient" / "ledger-pnl"
_PROJ_RE = re.compile(r"^(MFD|CP|RP)\d+(-FTW)?$", re.IGNORECASE)
_PNL_JOBS: dict = {}                 # proj -> {state, started, log, detail, proc, file}
_PNL_LOCK = threading.Lock()
# Where the lien notice / affidavit PDFs are filed (Synology). Per-machine override via
# machine.env LIEN_FOLDER; default is the Accounting share's Vendor Liens folder.
_LIEN_FOLDER = paths.get_path("LIEN_FOLDER",
                              "/Volumes/Accounting/LIENS & MONTHLY NOTICES/Vendor Liens/2026")


def _os_open(path: str):
    """Open a file/folder in the host OS file manager. Cross-platform so the same
    dashboard works on Mac OR Windows — the LOCAL server opens it with the native
    command, so the browser never has to handle smb:// or \\\\server paths. Returns
    None on success, an error string otherwise. Open-only — never executes."""
    system = platform.system()
    try:
        if system == "Darwin":
            subprocess.Popen(["open", path])           # Finder (Mac)
        elif system == "Windows":
            os.startfile(path)                         # Explorer (Windows)  # noqa: B606
        else:
            subprocess.Popen(["xdg-open", path])       # Linux desktops
    except OSError as e:
        return f"open failed: {e}"
    return None


def _ensure_lien_vendor_dir(vendor: str):
    """`_LIEN_FOLDER/<vendor>` for a vendor's lien docs (owner: organize by vendor, auto-create on
    mark). Creates ONLY the vendor subfolder and ONLY if the base already exists - so we never
    build a fake tree that shadows an unmounted share. Sanitized (no path traversal). Path or None."""
    safe = re.sub(r"[\\/]+", " ", (vendor or "")).strip().strip(".")
    base = Path(_LIEN_FOLDER)
    if not safe or not base.is_dir():
        return None
    d = base / safe
    try:
        d.mkdir(exist_ok=True)
        return d
    except OSError:
        return None


# ── live P&L compute (folds project-pnl's numbers INTO the dashboard) ────────
# Conventions match project-pnl/project_pnl_export.py so the two reconcile:
# Earned Revenue = contract × %complete; costs = cost_line (QBO truth, incl subs);
# Overhead = 10% of revenue (MFD alt = 9% of costs, the user 2026-07-16); net =
# revenue − costs − overhead. Billed (AR) is shown alongside as the realized view.
_OVERHEAD_REV = 0.10          # company: 10% of earned revenue
_OVERHEAD_MFD_COST = 0.09     # MFD alt view: 9% of costs


def _project_pnl(con, proj: str) -> dict:
    """A live per-project P&L assembled from the ledger spine — no Excel needed."""
    proj = (proj or "").strip().upper()
    row = con.execute(
        "SELECT project_no, division, total_contract_price tcp, percent_complete pc, "
        "costs_to_date ctd, estimated_total_costs etc FROM v_wip_latest WHERE project_no = ?",
        (proj,)).fetchone()
    div = (row["division"] if row else None) or ("Multi Family" if proj.startswith("MFD")
           else "Commercial" if proj.startswith("CP") else "Residential")
    is_mfd = proj.startswith("MFD") or (div or "").lower().startswith("multi")
    contract = (row["tcp"] if row else 0) or 0
    pct = (row["pc"] if row else 0) or 0
    earned = round(contract * pct, 2)
    # costs from the QBO-complete cost_line (incl subs), itemized by cost code
    cost = con.execute("SELECT COALESCE(SUM(amount),0) c FROM cost_line WHERE project_no = ?", (proj,)).fetchone()["c"] or 0
    by_code = [dict(r) for r in con.execute(
        "SELECT COALESCE(cost_code,'(uncoded)') code, COALESCE(SUM(amount),0) amount, "
        "COUNT(*) lines, MAX(is_sub) is_sub FROM cost_line WHERE project_no = ? "
        "GROUP BY COALESCE(cost_code,'(uncoded)') ORDER BY amount DESC", (proj,))]
    billed = con.execute("SELECT COALESCE(SUM(amount),0) a FROM billing_event WHERE project_no = ?", (proj,)).fetchone()["a"] or 0
    # Every AR invoice (draw) this project has billed - the make-up of billed-to-date, oldest first.
    invoices = [dict(r) for r in con.execute(
        "SELECT doc_number, qbo_txn_id, amount, balance, txn_date, status, paid_date, due_date, memo "
        "FROM billing_event WHERE project_no = ? ORDER BY txn_date, doc_number", (proj,))]
    overhead = round((_OVERHEAD_MFD_COST * cost) if is_mfd else (_OVERHEAD_REV * earned), 2)
    net = round(earned - cost - overhead, 2)
    return {
        "proj": proj, "division": div,
        "contract": contract, "pct_complete": pct, "earned": earned, "billed": billed,
        "cost": cost, "overhead": overhead,
        "overhead_basis": "9% of costs (MFD)" if is_mfd else "10% of revenue",
        "net": net, "net_pct": (net / earned) if earned else None,
        "by_code": by_code, "invoices": invoices,
        "has_wip": row is not None,
    }


def _portfolio_pnl(con) -> dict:
    """Company P&L: every ACTIVE job's live P&L + division and company totals.
    Active = WIP status 'Active' OR NULL (MFD — Test-Master carries no STATUS column,
    so its jobs are active by construction); Closed/Complete are excluded. Same
    per-project math as _project_pnl, batched into 3 aggregate reads."""
    wip = list(con.execute(
        "SELECT project_no, division, status, total_contract_price tcp, percent_complete pc, "
        "builder_or_gc, project_name FROM v_wip_latest"))
    costs = {r["project_no"]: r["c"] for r in con.execute(
        "SELECT project_no, COALESCE(SUM(amount),0) c FROM cost_line GROUP BY project_no")}
    billed = {r["project_no"]: r["a"] for r in con.execute(
        "SELECT project_no, COALESCE(SUM(amount),0) a FROM billing_event "
        "WHERE project_no IS NOT NULL GROUP BY project_no")}
    client_of = _project_customer_map(con)   # the shared client resolver (same as Bills/Liens/projects)
    try:  # project → QBO customer id (CustomerRef.value) for the project# deep link; absent-safe
        cust_of = {r["project_no"]: r["customer_id"] for r in con.execute(
            "SELECT project_no, customer_id FROM cost_line "
            "WHERE customer_id IS NOT NULL AND customer_id <> '' GROUP BY project_no")}
    except sqlite3.OperationalError:
        cust_of = {}
    rows, div = [], {}
    comp = {"earned": 0.0, "cost": 0.0, "overhead": 0.0, "net": 0.0, "billed": 0.0, "n": 0}
    for w in wip:
        st_raw = (w["status"] or "").strip()
        active = st_raw.lower() in ("", "active")    # blank = MFD (active by construction)
        p = w["project_no"]
        division = w["division"] or ("Multi Family" if p.startswith("MFD")
                   else "Commercial" if p.startswith("CP") else "Residential")
        is_mfd = p.startswith("MFD") or division.lower().startswith("multi")
        contract = w["tcp"] or 0
        pc = w["pc"] or 0
        earned = round(contract * pc, 2)
        cost = costs.get(p, 0) or 0
        oh = round((_OVERHEAD_MFD_COST * cost) if is_mfd else (_OVERHEAD_REV * earned), 2)
        net = round(earned - cost - oh, 2)
        b = billed.get(p, 0) or 0
        try:                                             # ~4 stats/project (no glob) - cheap, cached client-side
            mtime = pnl_paths.find_pnl(p).get("mtime")
        except Exception:  # noqa: BLE001 - a path hiccup must never break the P&L
            mtime = None
        rows.append({"proj": p, "division": division, "contract": contract,
                     "pct_complete": pc, "earned": earned, "cost": cost,
                     "overhead": oh, "net": net,
                     "net_pct": (net / earned) if earned else None, "billed": b,
                     "name": w["project_name"],                          # the job name / address
                     "client": client_of.get(p) or w["builder_or_gc"] or None,
                     "cust_id": cust_of.get(p), "pnl_mtime": mtime,
                     "status": st_raw or "Active", "active": active})
        if not active:                               # Closed/Complete: shown + filterable, but OFF the totals
            continue
        d = div.setdefault(division, {"division": division, "earned": 0.0, "cost": 0.0,
                                      "overhead": 0.0, "net": 0.0, "billed": 0.0, "n": 0})
        for k, v in (("earned", earned), ("cost", cost), ("overhead", oh), ("net", net), ("billed", b)):
            d[k] += v
            comp[k] += v
        d["n"] += 1
        comp["n"] += 1
    for d in list(div.values()) + [comp]:
        d["net_pct"] = (d["net"] / d["earned"]) if d["earned"] else None
    rows.sort(key=lambda r: r["net"])                # worst margin first — the ones to watch
    by_div = sorted(div.values(), key=lambda d: -d["earned"])
    return {"rows": rows, "by_division": by_div, "company": comp}


_ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


def _pnl_last_line(path) -> str:
    """The last MEANINGFUL line of a generation log - the tool's own progress text ('where it's
    at'), not the whole log. Reads only the tail, strips ANSI + box-drawing, skips blank/border
    lines. Empty string if the log isn't readable yet."""
    if not path:
        return ""
    try:
        with open(path, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            f.seek(max(0, size - 4096))
            data = f.read().decode("utf-8", "replace")
    except OSError:
        return ""
    for ln in reversed(data.splitlines()):
        s = _ANSI_RE.sub("", ln).strip().strip("─╭╮╰╯│▌ ").strip()
        if s and re.search(r"[A-Za-z0-9]", s):
            return s[:130]
    return ""


def _pnl_wait(proj: str) -> None:
    """Reap a generation subprocess and record its outcome (running → done/error)."""
    j = _PNL_JOBS.get(proj)
    if not j:
        return
    rc = j["proc"].wait()
    try:
        j["file"].close()
    except OSError:
        pass
    with _PNL_LOCK:
        j["state"] = "done" if rc == 0 else "error"
        j["rc"] = rc
        if rc != 0:
            j["detail"] = f"exit {rc} — see {j['log']}"


# ── the ledger's CONTROL PLANE: a pipeline registry ────────────────────────
# Each pipeline is an ordered list of steps run as subprocesses (tools never IMPORT
# tools - repo rule). `script` is repo-relative; `side: True` marks a PRODUCER with
# real writes (QBO/Notion/Teams/Excel) vs a read-only loader. The default Resync
# ("reload") runs the LOADERS ONLY - read the current sources into the ledger, fast +
# read-only. "all" (Full refresh) also runs the producers. The WIP DRAFT generators
# (the readers write Test tabs for PM review, then it's implemented) are a SEPARATE,
# confirm-gated action, never in a refresh. 'QBO costs' pulls the last 90 days
# incrementally (Touch ID). WIP loads first (it creates the project table).
_SYNC_LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient" / "ledger-sync"
_SYNC_COST_WINDOW_DAYS = 90


def _pipelines():
    """Built fresh each call so the cost --since window stays current."""
    since = (_dt.date.today() - _dt.timedelta(days=_SYNC_COST_WINDOW_DAYS)).isoformat()
    return [
        {"key": "wip", "label": "WIP master", "steps": [
            {"label": "Load current WIP -> ledger", "script": "ledger/load_wip_master.py", "args": []},
        ], "draft": {"label": "Generate DRAFT WIP (Test tabs, for PM review)", "steps": [
            {"label": "Draft CP WIP", "script": "wip/cp_wip_reader.py", "args": [], "side": True},
            {"label": "Draft RP WIP", "script": "wip/rp_wip_reader.py", "args": [], "side": True},
        ]}},
        # Hidden as its own Console card (owner 2026-08-19 - costs belong in a future "Company P&L"
        # view). It STAYS in the reload/all chains so Resync keeps costs fresh for the Project P&L
        # (which reads cost_line); it just isn't a standalone button any more.
        {"key": "costs", "label": "Costs (QBO, 90d)", "hidden": True, "steps": [
            {"label": "Pull costs (90d, Touch ID)", "script": "ledger/load_costs.py", "args": ["--active", "--since", since]},
        ]},
        {"key": "ap", "label": "AP - bills + liens", "steps": [
            {"label": "Sync bills (QBO -> Bill Tracker.xlsx)", "script": "bill-tracker/excel_bill_sync.py", "args": [], "side": True},
            {"label": "Load bills -> ledger", "script": "ledger/load_bill_tracker.py", "args": []},
        ]},
        {"key": "ar", "label": "AR - invoices / draws", "steps": [
            {"label": "Sync invoices (QBO -> Notion + Teams)", "script": "invoice-sync/run_invoice_sync.py", "args": [], "side": True},
            {"label": "Load invoices -> ledger", "script": "ledger/load_invoices.py", "args": []},
        ]},
        {"key": "payments", "label": "Payments (QBO)", "steps": [
            # Payments are money IN and drive the Payments tab. Was wired into NO sync at all
            # (owner 2026-08-27: "payments section is not showing recent payments"), so it now
            # rides the reload chain like every other loader. load_payments DELETE+reloads its
            # window, so the window IS the history depth: a rolling 12-month year (its default)
            # keeps the Payments tab's history while catching the newest payments each run.
            {"label": "Pull payments (QBO, rolling year)", "script": "ledger/load_payments.py", "args": ["--months", "12"]},
        ]},
        {"key": "crm", "label": "CRM - customers", "steps": [
            {"label": "Pull customers (Notion)", "script": "ledger/load_customers.py", "args": []},
        ]},
        # Workbook GENERATOR, not a loader: its steps live under "actions" so it
        # can never be swept into the reload/all chains - a Full refresh must not
        # kick off 138 project P&L runs (same reason the WIP draft sits outside
        # "steps"). One action per division; each is its own explicit click.
        {"key": "pnl", "label": "Project P&L workbooks", "steps": [], "actions": [
            {"key": "pnl-cp", "label": "Active CP"},
            {"key": "pnl-rp", "label": "Active RP"},
            {"key": "pnl-mfd", "label": "Active MFD"},
        ]},
        {"key": "subloc", "label": "Sub LOC (QBO float)", "steps": [
            {"label": "Load sub LOC float (Touch ID)", "script": "ledger/load_sub_loc.py", "args": []},
        ]},
    ]


def _resolve_steps(pipeline_key):
    """Ordered steps for: a pipeline key (its full chain) - 'reload' (every loader, the
    safe default Resync) - 'all' (every full chain incl producers) - 'wip-draft'."""
    pls = _pipelines()
    if pipeline_key == "reload":
        return [s for p in pls for s in p["steps"] if not s.get("side")]
    if pipeline_key == "all":
        return [s for p in pls for s in p["steps"]]
    if pipeline_key == "wip-draft":
        return next((p["draft"]["steps"] for p in pls if p["key"] == "wip" and p.get("draft")), [])
    if pipeline_key.startswith("pnl-"):
        div = pipeline_key[4:].upper()
        if div not in ("CP", "RP", "MFD"):
            return []
        return [{"label": f"Regenerate Active {div} P&L workbooks",
                 "script": "project-pnl/project_pnl_export.py",
                 "args": ["active", div.lower(), "--no-prompt"], "side": True}]
    return next((p["steps"] for p in pls if p["key"] == pipeline_key), [])


# ── WIP Review: the accept/merge flow over the three Test tabs ────────────────
# The ledger shows the pending WIP update as a per-job before/after diff, the owner
# approves/disapproves each change, and approved values are written to Test - CP
# (CP), Test - RP (RP), and Test-Master (all three). The dashboard ONLY orchestrates
# subprocesses and reads/writes JSON - it never imports the wip tools (repo rule).
# Each tool's --emit-review dumps its tab's diff; --apply-review applies the owner's
# decisions and writes. JSON lives OUTSIDE the repo (Claude-visible/synced folder).
_WIP_REVIEW_DIR = Path(os.environ.get(
    "ACB_WIP_REVIEW_DIR",
    Path.home() / "Library" / "Application Support" / "Proficient" / "wip-review"))
# division -> emit file. Order is the display order (CP, RP, then the MFD-on-Master).
_WIP_REVIEW_FILES = [("Commercial", "cp.json"), ("Residential", "rp.json"),
                     ("Multi-Family", "master.json")]


def _rp_wip_file() -> Path:
    """The owner's verified RP WIP workbook - the binding RP source, resolved the
    same way rp_wip_reader does so Master and Test - RP read the identical file."""
    return paths.get_path("RP_WIP_FILE", paths.onedrive_base() / "RP WIP TO FIX_Final.xlsx")


def _wip_review_steps(mode: str):
    """mode 'emit' -> the three diff runs (no tab write); 'apply' -> the three
    guarded writes (Test - CP / Test - RP / Test-Master), each honouring the SAME
    decisions.json. Master needs --rp-from-file so its RP rows match Test - RP."""
    _WIP_REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    cp = str(_WIP_REVIEW_DIR / "cp.json")
    rp = str(_WIP_REVIEW_DIR / "rp.json")
    mst = str(_WIP_REVIEW_DIR / "master.json")
    dec = str(_WIP_REVIEW_DIR / "decisions.json")
    if mode == "emit":
        return [
            {"label": "Review CP (Test - CP, Touch ID)", "script": "wip/cp_wip_reader.py", "args": ["--emit-review", cp]},
            {"label": "Review RP (Test - RP, Touch ID)", "script": "wip/rp_wip_reader.py", "args": ["--emit-review", rp]},
            {"label": "Review MFD (Test-Master, Touch ID)", "script": "wip/master_wip_test.py", "args": ["--emit-review", mst]},
        ]
    return [
        {"label": "Write CP → Test - CP", "script": "wip/cp_wip_reader.py", "args": ["--apply-review", dec], "side": True},
        {"label": "Write RP → Test - RP", "script": "wip/rp_wip_reader.py", "args": ["--apply-review", dec], "side": True},
        {"label": "Write all → Test-Master", "script": "wip/master_wip_test.py",
         "args": ["--apply-review", dec, "--rp-from-file", str(_rp_wip_file())], "side": True},
    ]


_SYNC = {"state": "idle", "current": -1, "started": 0.0, "log": None, "pipeline": None, "steps": []}
_SYNC_LOCK = threading.Lock()


def _run_sync(steps) -> None:
    """Run the resolved steps in order, recording per-step state for the progress bar.
    State was claimed under the lock by _sync_start; this just executes + finalizes."""
    _SYNC_LOG_DIR.mkdir(parents=True, exist_ok=True)
    logpath = _SYNC_LOG_DIR / "sync.log"
    with _SYNC_LOCK:
        _SYNC["log"] = str(logpath)
    ok = True
    with open(logpath, "w", encoding="utf-8") as logf:
        for i, step in enumerate(steps):
            with _SYNC_LOCK:
                _SYNC["current"] = i
                _SYNC["steps"][i]["state"] = "running"
            logf.write(f"\n===== {step['label']} ({step['script']}) =====\n")
            logf.flush()
            try:
                rc = subprocess.call([sys.executable, str(PROJECT_ROOT / step["script"])] + step.get("args", []),
                                     cwd=str(PROJECT_ROOT), stdout=logf, stderr=subprocess.STDOUT)
            except OSError as e:
                logf.write(f"launch failed: {e}\n")
                rc = 1
            with _SYNC_LOCK:
                _SYNC["steps"][i]["state"] = "done" if rc == 0 else "error"
            if rc != 0:
                ok = False
                break
    with _SYNC_LOCK:
        _SYNC["state"] = "done" if ok else "error"
        _SYNC["current"] = -1


# lien states that put a bill on the action watchlist, most-urgent first
LIEN_RANK = {
    "Notice PAST due": 0, "Notice due in ≤7d": 1, "Notice due in ≤15d": 2,
    "Notice due in ≤30d": 3, "Notice Sent": 4, "Lien Filed": 5,
}


_PROJ_SHAPED = re.compile(r"^(RP|CP|MFD)\d", re.I)


def _project_customer_map(con) -> dict:
    """{project_no -> client (the GC)} - the ONE client resolver, shared by the AP/Liens view, the
    projects list, and the portfolio P&L so they never disagree. PRIMARY: project_customer (the QBO
    Customer:Project hierarchy reversed by load_payments, covers EVERY project); then payments'
    resolved GC, then a non-project-shaped billing_event customer, fill any gap."""
    out = {}
    try:
        for r in con.execute("SELECT project_no pn, client FROM project_customer WHERE COALESCE(client,'') <> ''"):
            out[r["pn"]] = r["client"]
    except sqlite3.OperationalError:
        pass
    try:
        for r in con.execute(
                "SELECT pa.project_no pn, p.parent_customer gc, COUNT(*) n FROM payment_application pa "
                "JOIN payment p ON p.qbo_txn_id = pa.payment_txn_id "
                "WHERE pa.project_no IS NOT NULL AND COALESCE(p.parent_customer,'') <> '' "
                "GROUP BY pa.project_no, p.parent_customer ORDER BY pa.project_no, n DESC"):
            out.setdefault(r["pn"], r["gc"])
    except sqlite3.OperationalError:
        pass
    try:
        for r in con.execute("SELECT project_no pn, customer c FROM billing_event "
                             "WHERE project_no IS NOT NULL AND COALESCE(customer,'') <> ''"):
            if r["pn"] not in out and not _PROJ_SHAPED.match(r["c"] or ""):
                out.setdefault(r["pn"], r["c"])
    except sqlite3.OperationalError:
        pass
    return out


def _fetch_ap(con) -> dict:
    """AP + lien view from ap_bill_line; empty (not an error) if the table is absent."""
    ap = {"summary": {"open_balance": 0, "open_lines": 0, "watch_count": 0},
          "lien_watch": [], "by_project": {}, "bills": []}
    try:
        rows = con.execute(
            "SELECT project_no, division, vendor, bill_ref, open_balance, lien_status, "
            "matched_invoice, invoice_no, qbo_link FROM ap_bill_line").fetchall()
    except sqlite3.OperationalError:
        return ap
    # Site lien marks (Notice Sent / Lien Filed / Released) the owner set on the dashboard.
    # They win over the computed/loaded lien_status INSTANTLY, before the next sync-ap mirrors
    # them into the workbook. Keyed by QBO bill id (= the workbook's _Key). Absent-safe.
    marks = bill_marks.read_lien_marks()
    def _eff_lien(r):
        bid = bill_marks.bill_id_from_link(r["qbo_link"])
        tag = marks.get(bid) if bid else None
        return tag if tag in bill_marks.LIEN_STATES else r["lien_status"]
    open_bal, open_lines = 0.0, 0
    for r in rows:
        ob = r["open_balance"] or 0
        if ob > 0:
            open_bal += ob
            open_lines += 1
    by_project = {}
    for r in con.execute("SELECT project_no, open_lines, open_balance FROM v_ap_by_project "
                         "WHERE COALESCE(open_balance,0) > 0"):
        if r["project_no"]:
            by_project[r["project_no"]] = {"open_lines": r["open_lines"], "open_balance": r["open_balance"]}
    ap["by_project"] = by_project
    # Full bill list for the Bill Tracker tab. Every row is one bill (open_balance is
    # per-bill, safe to sum). Description is omitted to keep the payload lean; account
    # carries the QBO category. Newest first (bill_date is ISO). Each bill is enriched
    # with its AR invoice (from billing_event, joined on Invoice #) so the Bills tab can
    # show the real invoice pay status and deep-link to the invoice in QuickBooks.
    bmap = {}
    try:
        for be in con.execute("SELECT doc_number, qbo_txn_id, amount, balance, status, txn_date, customer "
                              "FROM billing_event WHERE doc_number IS NOT NULL"):
            bmap[str(be["doc_number"])] = dict(be)
    except sqlite3.OperationalError:
        pass
    proj_customer = _project_customer_map(con)   # project -> client (the GC), the shared resolver
    pay = bill_marks.read_pay_marks()             # current check run: {bill_id -> {amount}}
    bills = []
    for r in con.execute(
        "SELECT project_no, division, vendor, bill_ref, bill_date, account, "
        "line_amount, open_balance, pay_status, approved, invoice_status, lien_status, "
        "matched_invoice, invoice_no, gc_paid_date, pay_date, qbo_link "
        "FROM ap_bill_line ORDER BY bill_date DESC"):
        b = dict(r)
        b["client"] = proj_customer.get(b.get("project_no"))
        bid = bill_marks.bill_id_from_link(b.get("qbo_link"))
        b["bill_id"] = bid                           # QBO bill id = workbook _Key; None → not markable
        pm = pay.get(bid) if bid else None           # is this bill in the current pay run?
        b["pay_selected"] = bool(pm)
        b["pay_amount"] = (pm.get("amount") if pm and pm.get("amount") is not None else None)  # None → full open bal
        b["lien_marked"] = bool(bid and marks.get(bid))   # currently a site override (vs computed)
        if bid and marks.get(bid) in bill_marks.LIEN_STATES:
            b["lien_status"] = marks[bid]            # site mark wins until the next sync-ap
        inv = bmap.get(str(b.get("invoice_no") or ""))
        if inv:
            b["inv_qbo_id"] = inv["qbo_txn_id"]      # QBO Invoice Id → company-scoped deep link
            b["inv_ar_status"] = inv["status"]       # actual AR status: Paid | Partially Paid | Unpaid
            b["inv_amount"] = inv["amount"]
            b["inv_balance"] = inv["balance"]
            b["inv_date"] = inv["txn_date"]
            b["inv_customer"] = inv["customer"]
        bills.append(b)
    ap["bills"] = bills
    # The lien WATCHLIST = the enriched bills that are on the clock OR marked (any LIEN_RANK status),
    # so the Liens page carries the client, bill date, the AR invoice + its pay status - not just the
    # project #. Most-urgent first. It's ALL divisions, not CP-only.
    watch = [b for b in bills if (b.get("lien_status") or "") in LIEN_RANK]
    watch.sort(key=lambda b: (LIEN_RANK[b["lien_status"]], -(b.get("open_balance") or 0)))
    ap["lien_watch"] = watch[:500]
    # The Liens PAGE = bills where a notice was actually SENT or a lien FILED (the owner's marks),
    # NOT the deadline clock (owner 2026-08-20). Filed first, then by open $. Small list, no cap.
    _SENT = {"Lien Filed": 0, "Notice Sent": 1}
    liens = [b for b in bills if (b.get("lien_status") or "") in _SENT]
    liens.sort(key=lambda b: (_SENT[b["lien_status"]], -(b.get("open_balance") or 0)))
    ap["liens"] = liens
    ap["summary"] = {"open_balance": open_bal, "open_lines": open_lines, "watch_count": len(watch)}
    return ap


_AGING_BUCKETS = ["Current", "1-30", "31-60", "61-90", "90+"]


def _aging_bucket(days_past_due) -> int:
    """Index into _AGING_BUCKETS by signed days-past-due - the SAME thresholds as
    invoice-sync/aging_sheet.py, so the tab ages exactly like the AR Aging workbook.
    None / not-yet-due (<= 0) sits in Current."""
    if days_past_due is None or days_past_due <= 0:
        return 0
    if days_past_due <= 30:
        return 1
    if days_past_due <= 60:
        return 2
    if days_past_due <= 90:
        return 3
    return 4


def _client_pay_speed(con) -> dict:
    """How long each client takes to pay: avg days from INVOICE date to PAID date over their
    PAID invoices - the input to a future-cash-in forecast (owner 2026-08-25). Returns
    {by_client: {client_lower: {avg_days, n}}, all_avg: <portfolio avg or None>}. Absent-safe."""
    try:
        rows = con.execute(
            "SELECT customer, txn_date, paid_date FROM billing_event "
            "WHERE COALESCE(balance,0) <= 0.005 AND COALESCE(paid_date,'') <> '' AND COALESCE(txn_date,'') <> ''"
        ).fetchall()
    except sqlite3.OperationalError:
        return {"by_client": {}, "all_avg": None}
    agg, alld = {}, []
    for r in rows:
        try:
            d = (_dt.date.fromisoformat(r["paid_date"][:10]) - _dt.date.fromisoformat(r["txn_date"][:10])).days
        except (ValueError, TypeError):
            continue
        if d < 0 or d > 400:                      # guard against typo'd dates
            continue
        agg.setdefault((r["customer"] or "").strip().lower(), []).append(d)
        alld.append(d)
    by_client = {c: {"avg_days": round(sum(v) / len(v)), "n": len(v)} for c, v in agg.items() if v}
    return {"by_client": by_client, "all_avg": (round(sum(alld) / len(alld)) if alld else None)}


def _fetch_open_invoices(con) -> dict:
    """Open AR invoices (the draws the GC still owes you) from billing_event, aged by DUE
    DATE into the same Current/1-30/31-60/61-90/90+ buckets as the Invoice Tracker's AR
    Aging tab, each carrying its related Lien Tracker status. Empty (not an error) if
    billing_event predates the AR columns or is unloaded."""
    out = {"as_of": _dt.date.today().isoformat(), "buckets": _AGING_BUCKETS, "invoices": []}
    try:
        rows = con.execute(
            "SELECT doc_number, qbo_txn_id, project_no, division, customer, memo, amount, "
            "balance, txn_date, due_date, net_terms, aging_bucket, status, litigation, "
            "lien_status, lien_notice, paid_date, draw_period FROM billing_event "
            "WHERE COALESCE(balance,0) > 0.005").fetchall()
    except sqlite3.OperationalError:
        return out
    # Project → QBO customer id (CustomerRef.value from cost_line) for the project# deep link. Absent-safe.
    try:
        cust_of = {r["project_no"]: r["customer_id"] for r in con.execute(
            "SELECT project_no, customer_id FROM cost_line "
            "WHERE customer_id IS NOT NULL AND customer_id <> '' GROUP BY project_no")}
    except sqlite3.OperationalError:
        cust_of = {}
    today = _dt.date.today()
    invs = []
    for r in rows:
        d = dict(r)
        days, due = None, (d.get("due_date") or "")[:10]
        if due:
            try:
                days = (today - _dt.date.fromisoformat(due)).days
            except ValueError:
                days = None
        bi = _aging_bucket(days)
        # Only fall back to Notion's stored bucket when there's no due date to age by live.
        if days is None and d.get("aging_bucket") in _AGING_BUCKETS:
            bi = _AGING_BUCKETS.index(d["aging_bucket"])
        d["days_past_due"] = days
        d["bucket"] = _AGING_BUCKETS[bi]
        d["bucket_index"] = bi
        # Computed Texas lien-notice CLOCK ("when the lien is due") - the SAME shared/lien_clock
        # the AR Aging Excel uses, so the site and the workbook agree. Division from the project #.
        proj = d.get("project_no") or ""
        div_code = "MFD" if proj.startswith("MFD") else "CP" if proj.startswith("CP") else "RP"
        inv_date = None
        tx = (d.get("txn_date") or "")[:10]
        if tx:
            try:
                inv_date = _dt.date.fromisoformat(tx)
            except ValueError:
                inv_date = None
        # The Notion Lien Tracker status advances the clock: once the notice is Mailed, it moves
        # from the notice deadline to the lien-AFFIDAVIT deadline (the real cutoff); a filed/paid
        # status ends it. Same call the AR Aging Excel makes, so the site and workbook agree.
        ls = lien_clock.lien_state(div_code, inv_date, today, memo=(d.get("memo") or ""),
                                   lien_status=d.get("lien_status"))
        d["lien_due_label"] = ls.label or None
        d["lien_due_state"] = ls.state
        d["cust_id"] = cust_of.get(proj)             # project# → QBO customerdetail deep link
        invs.append(d)
    invs.sort(key=lambda x: ((x.get("customer") or "~").lower(),
                             x.get("due_date") or "9999", x.get("doc_number") or ""))
    out["invoices"] = invs
    out["pay_speed"] = _client_pay_speed(con)   # avg days-to-pay per client → cash-in forecast
    return out


def _fetch_payments(con) -> dict:
    """Received customer payments, each as ONE transaction (money IN) with the
    invoice(s) it paid grouped beneath it. Reads the payment / payment_application
    tables (QBO Payment objects via load_payments.py). Empty (not an error) when
    that loader hasn't run yet."""
    out = {"payments": [], "total_received": 0.0, "count": 0, "invoices_paid": 0, "loaded_at": None}
    try:
        prows = con.execute(
            "SELECT qbo_txn_id, txn_date, customer, customer_id, parent_customer, parent_customer_id, "
            "total_amt, unapplied_amt, method, ref_no, loaded_at FROM payment").fetchall()
    except sqlite3.OperationalError:
        return out
    if not prows:
        return out
    try:            # project -> QBO customer id for the deep link (same source as the P&L/invoices)
        cust_of = {r["project_no"]: r["customer_id"] for r in con.execute(
            "SELECT project_no, customer_id FROM cost_line "
            "WHERE customer_id IS NOT NULL AND customer_id <> '' GROUP BY project_no")}
    except sqlite3.OperationalError:
        cust_of = {}
    apps_by_pay = {}
    try:
        for a in con.execute("SELECT payment_txn_id, invoice_txn_id, invoice_no, project_no, "
                             "division, amount, invoice_open FROM payment_application"):
            d = dict(a)
            d["cust_id"] = cust_of.get(d["project_no"])       # project deep link
            apps_by_pay.setdefault(a["payment_txn_id"], []).append(d)
    except sqlite3.OperationalError:
        pass
    pays, recv_tot, links = [], 0.0, 0
    for r in prows:
        p = dict(r)
        apps = apps_by_pay.get(p["qbo_txn_id"], [])
        apps.sort(key=lambda x: (x.get("project_no") or "", x.get("invoice_no") or ""))
        divs = {a.get("division") for a in apps if a.get("division")}
        p["applications"] = apps
        p["invoice_count"] = len(apps)
        p["applied_total"] = round(sum(a.get("amount") or 0 for a in apps), 2)
        p["division"] = next(iter(divs)) if len(divs) == 1 else ("Mixed" if divs else None)
        recv_tot += p.get("total_amt") or 0
        links += len(apps)
        pays.append(p)
    pays.sort(key=lambda x: (x.get("txn_date") or ""), reverse=True)      # most recent payment first
    out["payments"] = pays
    out["total_received"] = round(recv_tot, 2)
    out["count"] = len(pays)
    out["invoices_paid"] = links
    out["loaded_at"] = pays[0].get("loaded_at") if pays else None
    return out


def _fetch_costs(con) -> dict:
    """QBO cost rollups from cost_line; empty (not an error) if unloaded."""
    out = {"by_code": [], "by_project_code": {}, "by_project": {}, "loaded_total": 0}
    try:
        bp = {r["project_no"]: {"costs_loaded": r["costs_loaded"], "sub_costs": r["sub_costs"],
                                "lines": r["lines"]}
              for r in con.execute("SELECT project_no, costs_loaded, sub_costs, lines "
                                   "FROM v_cost_by_project")}
    except sqlite3.OperationalError:
        return out
    out["by_project"] = bp
    out["loaded_total"] = sum((v["costs_loaded"] or 0) for v in bp.values())
    # portfolio, by cost code (join the friendly description)
    for r in con.execute(
            "SELECT vc.code, vc.cost_code, cc.description, SUM(vc.actual) actual, SUM(vc.lines) lines "
            "FROM v_cost_by_code vc LEFT JOIN cost_code cc ON cc.code = vc.cost_code "
            "GROUP BY vc.code, vc.cost_code, cc.description ORDER BY actual DESC"):
        out["by_code"].append({"code": r["code"], "cost_code": r["cost_code"],
                               "description": r["description"], "actual": r["actual"], "lines": r["lines"]})
    # per-project, by cost code (for the job detail drill)
    pc: dict = {}
    for r in con.execute("SELECT project_no, code, cost_code, actual, lines "
                         "FROM v_cost_by_code ORDER BY actual DESC"):
        pc.setdefault(r["project_no"], []).append(
            {"code": r["code"], "cost_code": r["cost_code"], "actual": r["actual"], "lines": r["lines"]})
    out["by_project_code"] = pc

    # grouped: cost TYPE = parent, job TYPE = sub — the JobTread model (material
    # links to ONE cost-type parent; the job-type sub shows cost-to-budget).
    from shared.qbo_costs import cost_code_meta, job_type_name
    groups: dict = {}
    for c in out["by_code"]:
        if c["cost_code"]:
            m = cost_code_meta(c["cost_code"])
            parent = m["description"] or c["cost_code"]
            sub = job_type_name(m["prefix"]) or (m["prefix"] or "—")
        else:                                   # account-based line: no job-type split
            parent = c["code"]
            sub = "(account)"
        g = groups.setdefault(parent, {"parent": parent, "actual": 0, "lines": 0, "subs": {}})
        g["actual"] += c["actual"] or 0
        g["lines"] += c["lines"] or 0
        s = g["subs"].setdefault(sub, {"sub": sub, "code": c["cost_code"], "actual": 0, "lines": 0})
        s["actual"] += c["actual"] or 0
        s["lines"] += c["lines"] or 0
    by_type = []
    for g in groups.values():
        g["subs"] = sorted(g["subs"].values(), key=lambda s: -(s["actual"] or 0))
        by_type.append(g)
    by_type.sort(key=lambda g: -(g["actual"] or 0))
    out["by_cost_type"] = by_type

    # by vendor — the Vendors page (who we pay the most, subs vs suppliers)
    vend = []
    for r in con.execute(
            "SELECT vendor, SUM(amount) spend, COUNT(*) lines, "
            "COUNT(DISTINCT project_no) jobs, "
            "SUM(CASE WHEN is_sub=1 THEN amount ELSE 0 END) sub_spend "
            "FROM cost_line WHERE vendor IS NOT NULL AND vendor <> '' "
            "GROUP BY vendor ORDER BY spend DESC"):
        vend.append({"vendor": r["vendor"], "spend": r["spend"], "lines": r["lines"],
                     "jobs": r["jobs"], "sub_spend": r["sub_spend"]})
    # vendor TYPE — Sub (labor) vs Supplier: <material> — from each vendor's cost mix.
    mix: dict = {}
    for r in con.execute("SELECT vendor, cost_code, account, SUM(amount) amt FROM cost_line "
                         "WHERE vendor IS NOT NULL AND vendor <> '' GROUP BY vendor, cost_code, account"):
        parent = cost_code_meta(r["cost_code"])["description"] if r["cost_code"] else None
        parent = parent or r["account"] or "Materials"
        mix.setdefault(r["vendor"], {})
        mix[r["vendor"]][parent] = mix[r["vendor"]].get(parent, 0) + (r["amt"] or 0)
    for v in vend:
        spend = v["spend"] or 0
        sub_share = (v["sub_spend"] or 0) / spend if spend else 0
        vm = mix.get(v["vendor"], {})
        top = max(vm, key=vm.get) if vm else None
        v["vtype"] = "Sub" if sub_share >= 0.5 else (f"Supplier: {top}" if top else "Supplier")
    out["by_vendor"] = vend
    return out


def _freshness(con) -> dict:
    """When each feed last landed (ledger loaded_at) + when each SOURCE sync wrote
    its file (mtime) — the owner's "is my data current?" strip."""
    import os
    out = {"ledger": {}, "sources": {}}
    for tbl, key in (("wip_snapshot", "WIP"), ("ap_bill_line", "AP (Bill Tracker)"),
                     ("cost_line", "Costs (QBO)"), ("billing_event", "AR (invoices)"),
                     ("payment", "Payments"),
                     ("customer", "CRM (customers)"), ("sub_loc_run", "Sub LOC")):
        try:
            r = con.execute(f"SELECT MAX(loaded_at) FROM {tbl}").fetchone()
            out["ledger"][key] = r[0] if r and r[0] else None
        except sqlite3.OperationalError:
            out["ledger"][key] = None

    def mtime(p):
        try:
            return _dt.datetime.fromtimestamp(os.path.getmtime(str(p))).isoformat(timespec="minutes")
        except OSError:
            return None

    ob = paths.onedrive_base()
    bt = paths.get_path("ACB_BILL_TRACKER_XLSX", ob / "Automations-/Bill Tracker.xlsx")
    wm = paths.get_path("WIP_EXCEL_PATH", ob / "Company Files - WIP Report/WIP - MASTER new.xlsx")
    out["sources"]["sync-ap"] = mtime(bt)
    out["sources"]["WIP master"] = mtime(wm)
    # AR mirror: the owner's live file is OneDrive Collections/Invoice Tracker.xlsx (2026-08-18;
    # the old Open_Invoices.xlsx is now backup_dont_use). Resolve that first (configurable), then
    # the invoice-sync default (INVOICE_EXPORT_PATH / repo root) and legacy OneDrive names, so the
    # AR card shows a real "last ran" - not blank - which was the "AP showed, AR didn't" bug.
    for cand in (paths.get_path("ACB_INVOICE_TRACKER_XLSX", ob / "Collections/Invoice Tracker.xlsx"),
                 paths.get_path("INVOICE_EXPORT_PATH", PROJECT_ROOT / "Open_Invoices.xlsx"),
                 ob / "Collections/Open_Invoices.xlsx", ob / "Automations-/Open_Invoices.xlsx",
                 ob / "Open_Invoices.xlsx"):
        m = mtime(cand)
        if m:
            out["sources"]["sync-ar"] = m
            break
    return out


def _fetch_actions(con) -> dict:
    """Map action_key → {url, status} from the `action` table (Notion mirror).
    Empty (not an error) if sync_actions hasn't run."""
    out: dict = {}
    try:
        for r in con.execute("SELECT action_key, status, notion_url FROM action"):
            out[r["action_key"]] = {"status": r["status"], "url": r["notion_url"]}
    except sqlite3.OperationalError:
        pass
    return out


def _waiver_key(mi, vendor, bill_ref) -> str:
    """Deterministic key for a bill's waiver — survives ap_bill_line reloads."""
    raw = f"{mi or ''}|{vendor or ''}|{bill_ref or ''}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


# ── Accounting fixes: the Bill Tracker audits, folded into the ledger ─────────
_AUDIT_SHEETS = ("Audit - Coding", "Audit - PO", "Audit - Bills")


def _audit_div(project: str) -> str:
    p = (project or "").upper()
    if p.startswith("MFD"):
        return "Multi Family"
    if p.startswith("CP"):
        return "Commercial"
    if p.startswith("RP"):
        return "Residential"
    return ""


def _fetch_accounting_audits() -> dict:
    """The three Bill Tracker audit sheets (the fixes to make) as ONE filterable list.
    Read live from Bill Tracker.xlsx: bill-tracker already computes the findings with the
    FULL bill data (incl subs + cost codes) that ap_bill_line doesn't carry, so the ledger
    reads the sheets rather than recomputing. Read-only; a missing file is reported, not raised."""
    from openpyxl import load_workbook
    bt = paths.get_path("ACB_BILL_TRACKER_XLSX", paths.onedrive_base() / "Automations-/Bill Tracker.xlsx")
    out = {"ok": False, "source": str(bt), "findings": [], "counts": {}}
    if not bt.exists():
        out["error"] = f"Bill Tracker.xlsx not found ({bt}). Run the AP sync first."
        return out
    try:
        # data_only=False: the Open column is an =HYPERLINK("url","↗") FORMULA, so the cached
        # value is just "↗" - we read the formula and pull the URL out. Every other audit cell
        # is a plain stored value, so it reads the same either way.
        wb = load_workbook(bt)
    except Exception as e:                          # noqa: BLE001
        out["error"] = f"could not open Bill Tracker.xlsx: {e}"
        return out
    findings = []
    # line memo (Line Description) lives only on the Bills/Inventory sheets, keyed per line
    # by (Bill #, Line Amount) - join it on so each finding shows what the clerk actually
    # wrote on the bill line, not just why it was flagged. Subs aren't on those display
    # sheets, so sub findings carry no memo; a single-line bill falls back to its one desc.
    memo_by_line, descs_by_bill = {}, {}
    for msheet in ("Bills", "Inventory"):
        if msheet not in wb.sheetnames:
            continue
        ms = wb[msheet]
        mhdr, midx = None, {}
        for r in range(1, 8):
            vals = [ms.cell(r, c).value for c in range(1, (ms.max_column or 0) + 1)]
            if "Line Description" in vals and "Bill #" in vals:
                mhdr, midx = r, {v: c for c, v in enumerate(vals, 1) if v}
                break
        if not mhdr:
            continue
        bcol, acol, dcol = midx.get("Bill #"), midx.get("Line Amount"), midx.get("Line Description")
        if not (bcol and dcol):
            continue
        for r in range(mhdr + 1, (ms.max_row or 0) + 1):
            bno, desc = ms.cell(r, bcol).value, ms.cell(r, dcol).value
            if not bno or desc in (None, ""):
                continue
            bno, desc = str(bno).strip(), str(desc).strip()
            amt = ms.cell(r, acol).value if acol else None
            akey = round(float(amt), 2) if isinstance(amt, (int, float)) else None
            memo_by_line.setdefault((bno, akey), desc)
            descs_by_bill.setdefault(bno, []).append(desc)
    for sheet in _AUDIT_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr, idx = None, {}
        for r in range(1, 8):
            vals = [ws.cell(r, c).value for c in range(1, (ws.max_column or 0) + 1)]
            if "Issue" in vals:
                hdr, idx = r, {v: c for c, v in enumerate(vals, 1) if v}
                break
        if not hdr:
            continue
        ocol = idx.get("Open")

        def cell(r, *names):
            for n in names:
                c = idx.get(n)
                if c and ws.cell(r, c).value not in (None, ""):
                    return ws.cell(r, c).value
            return None

        for r in range(hdr + 1, (ws.max_row or 0) + 1):
            iv = idx.get("Issue") and ws.cell(r, idx["Issue"]).value
            if not iv or not str(iv).strip():
                continue
            url = None
            if ocol:
                oc = ws.cell(r, ocol)
                if oc.hyperlink:
                    url = oc.hyperlink.target
                elif isinstance(oc.value, str) and "HYPERLINK(" in oc.value:
                    m = re.search(r'HYPERLINK\("([^"]+)"', oc.value)
                    url = m.group(1) if m else None
                elif isinstance(oc.value, str) and oc.value.startswith("http"):
                    url = oc.value
            proj = str(cell(r, "Project", "Project/Job") or "").strip()
            dv = cell(r, "Bill Date", "Date")
            amt = cell(r, "Amount")
            bno = str(cell(r, "Bill #", "PO/Bill #", "Bill/Ref #") or "").strip()
            akey = round(float(amt), 2) if isinstance(amt, (int, float)) else None
            memo = memo_by_line.get((bno, akey)) if bno else None
            if not memo and bno in descs_by_bill:      # single-line bill: unambiguous fallback
                uniq = set(descs_by_bill[bno])
                if len(uniq) == 1:
                    memo = next(iter(uniq))
            findings.append({
                "issue": str(iv).strip(),
                "vendor": str(cell(r, "Vendor") or "").strip(),
                "bill_no": bno,
                "date": (dv.isoformat()[:10] if hasattr(dv, "isoformat") else (str(dv)[:10] if dv else None)),
                "project": proj, "division": _audit_div(proj),
                "cost_code": str(cell(r, "Cost Code") or "").strip(),
                "amount": (float(amt) if isinstance(amt, (int, float)) else None),
                "memo": memo or "",
                "detail": str(cell(r, "Detail") or "").strip(),
                "url": url, "group": sheet.replace("Audit - ", ""),
            })
    wb.close()
    # attachment count per finding, from the shared attachable index (disk cache, NO QBO
    # call). The UI shows a scan link only when att>0 and fetches fresh links on click, so
    # no-scan bills never trigger a lookup. Missing cache -> everything reads 0 (no links).
    try:
        from shared import qbo_attachments
        idx = qbo_attachments.index_from_cache()
        for f in findings:
            m = re.search(r"txnId=(\d+)", f.get("url") or "")
            f["att"] = (len(idx.get(("Bill", m.group(1)), [])) if (idx and m) else 0)
    except Exception:                                  # noqa: BLE001 - never let counts break the audit
        for f in findings:
            f["att"] = 0
    counts = {}
    for f in findings:
        counts[f["issue"]] = counts.get(f["issue"], 0) + 1
    out.update(ok=True, findings=findings, counts=counts)
    return out


# race-through stages, in worklist priority — Ready-to-turn-in first (all bills
# paid → turn it in to unlock the next draw), then pay vendors, then awaiting GC.
# A draw is "done" (green) the moment every bill is PAID; unconditional waivers are
# tracked per bill (the checkboxes) for the owner's records but no longer gate green.
# Vendors that must NOT hold a draw in "Pay vendors" (the owner 2026-08-28).
# The pump companies invoice on their own cycle and are paid outside the draw,
# so an open pump bill was parking otherwise-finished draws in the wrong stage.
# Their bills STILL SHOW on the draw and still count in the money — they just
# don't gate it. Matched tightly: "JD Core Construction" and "CORE SUPPLY EAST"
# are different vendors and must not be caught.
_NON_GATING_VENDOR_RE = re.compile(
    r"^\s*(MCP\s+CONCRETE\s+PUMPING|CORE\s+CONCRETE\s+PUMPING)\b", re.I)


def _gates_stage(vendor: str) -> bool:
    return not _NON_GATING_VENDOR_RE.match(vendor or "")


_STAGE_ORDER = {"Ready to turn in": 0, "Fund in — pay vendors": 1,
                "Awaiting GC funding": 2, "All paid": 3}


def _fetch_draws(con, limit: int = 100) -> dict:
    """Roll AP bills up BY DRAW (matched invoice) → the race-through pipeline,
    joined to the owner's waiver marks. Empty (not an error) if not loaded."""
    try:
        rows = con.execute(
            "SELECT matched_invoice mi, project_no, division, vendor, bill_ref, "
            "MAX(bill_total) amount, MAX(open_balance) open_bal, pay_status, invoice_status, "
            "MAX(gc_paid_date) gc, MAX(pay_date) pd, MAX(bill_date) bd, MAX(invoice_no) invoice_no, "
            "MAX(qbo_link) qbo_link "
            "FROM ap_bill_line WHERE matched_invoice IS NOT NULL AND matched_invoice <> '' "
            # RP isn't draws — RP bills at completion / milestones, not formal draws (owner).
            "AND COALESCE(project_no,'') NOT LIKE 'RP%' AND matched_invoice NOT LIKE '%— RP%' "
            "GROUP BY matched_invoice, vendor, bill_ref").fetchall()
    except sqlite3.OperationalError:
        return {"draws": [], "total": 0}
    wmap = {w["waiver_key"]: w["received"] for w in con.execute("SELECT waiver_key, received FROM waiver")}
    # AR side (money IN) from the Invoice Tracker load — joined by Invoice #.
    bmap: dict = {}
    try:
        for b in con.execute("SELECT doc_number, qbo_txn_id, project_no, division, customer, memo, "
                             "amount, balance, status, txn_date, due_date, net_terms, paid_date, "
                             "draw_period, lien_status, lien_notice "
                             "FROM billing_event WHERE doc_number IS NOT NULL"):
            bmap[str(b["doc_number"])] = dict(b)
    except sqlite3.OperationalError:
        pass
    draws: dict = {}
    for r in rows:
        d = draws.setdefault(r["mi"], {"matched_invoice": r["mi"], "project_no": r["project_no"],
                                       "division": r["division"], "invoice_no": None, "bills": []})
        if not d["invoice_no"] and r["invoice_no"]:
            d["invoice_no"] = r["invoice_no"]
        wk = _waiver_key(r["mi"], r["vendor"], r["bill_ref"])
        d["bills"].append({
            "vendor": r["vendor"], "bill_ref": r["bill_ref"], "amount": r["amount"] or 0,
            "open": r["open_bal"] or 0, "pay_status": r["pay_status"], "invoice_status": r["invoice_status"],
            "gc_paid": r["gc"], "pay_date": r["pd"], "bill_date": r["bd"], "qbo_link": r["qbo_link"],
            "waiver_key": wk, "waiver": bool(wmap.get(wk, 0)),
            "gates": _gates_stage(r["vendor"]),
        })
    out = []
    for mi, d in draws.items():
        bills = d["bills"]
        n = len(bills)
        paid = sum(1 for b in bills if b["pay_date"])
        # Stage is decided on the GATING bills only. If every bill on the draw
        # is a pump bill there is nothing else to judge by, so fall back to all
        # of them rather than calling the draw done on no evidence.
        gate = [b for b in bills if b["gates"]] or bills
        n_gate = len(gate)
        paid_gate = sum(1 for b in gate if b["pay_date"])
        funded = any(b["gc_paid"] for b in bills)
        waivers = sum(1 for b in bills if b["waiver"])
        ar = bmap.get(str(d.get("invoice_no") or ""))
        gc_paid_in = bool(ar and (ar.get("status") == "Paid" or (ar.get("balance") or 0) <= 0.005))
        if not funded:
            stage = "Awaiting GC funding"
        elif paid_gate < n_gate:
            stage = "Fund in — pay vendors"
        elif gc_paid_in:                   # vendors paid AND the GC has paid our AR = fully settled
            stage = "All paid"
        else:                              # vendors paid, GC AR still open (we fronted it - collect)
            stage = "Ready to turn in"
        d.update({
            "label": (mi or "").split("\n")[0].strip(), "n": n, "paid": paid, "funded": funded,
            # how many bills actually decide the stage, and how many open bills
            # are parked because they are non-gating vendors
            "n_gate": n_gate, "paid_gate": paid_gate,
            "parked": sum(1 for b in bills if not b["gates"] and not b["pay_date"]),
            "waivers": waivers, "total": sum(b["amount"] for b in bills),
            # what we actually pay = gating bills only (MCP/CORE concrete pumping excluded - not paid by us)
            "total_gate": sum(b["amount"] for b in bills if b["gates"]), "stage": stage,
            "recency": max([(b["gc_paid"] or b["pay_date"] or b["bill_date"] or "") for b in bills] or [""]),
            # money IN (billed to GC) — from the Invoice Tracker, by Invoice #
            "billed": (ar["amount"] if ar else None),          # net billed to the GC
            "ar_open": (ar["balance"] if ar else None),        # GC still owes this much
            "ar_status": (ar["status"] if ar else None),       # Paid | Partially Paid | Unpaid
            "ar_date": (ar["txn_date"] if ar else None),       # invoice date
            "ar_qbo_id": (ar["qbo_txn_id"] if ar else None),   # QBO Invoice Id → deep link
            "customer": (ar["customer"] if ar else None),
            # The whole invoice row (memo + fields) so the Draws sidebar can show the same
            # invoice detail the Invoices tab does - read a draw without opening QBO.
            "inv": ar,
            "gc_paid_in": gc_paid_in,
        })
        out.append(d)
    out.sort(key=lambda d: d["recency"], reverse=True)
    out.sort(key=lambda d: _STAGE_ORDER.get(d["stage"], 9))
    return {"draws": out[:limit], "total": len(out)}


# sales pipeline stages in funnel order
_SALES_ORDER = {"Lead": 0, "Follow up": 1, "Contacted": 2, "Interested": 3,
                "No response": 4, "Closed - Won": 5, "Closed - Lost": 6, "(none)": 7}


# Non-sales accounts to keep OUT of the sales-rep view: Notion integration bots
# (they arrive as a bare UUID) plus any account named in ACB_SALES_AUTOMATION_REPS
# (machine.env, gitignored — so real names never enter the repo). These create /
# import records but do no outreach, so crediting them as a "rep" is misleading.
_UUID_RE = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")
_AUTOMATION_REPS = {s.strip() for s in (paths.get("ACB_SALES_AUTOMATION_REPS", "") or "").split(",") if s.strip()}


def _is_automation(name) -> bool:
    return bool(name) and (_UUID_RE.fullmatch(name) is not None or name in _AUTOMATION_REPS)


def _rep_label(name):
    """Display label for an editor: automation → 'Automation'; real people/emails kept."""
    if not name:
        return "(unknown)"
    return "Automation" if _is_automation(name) else name


def _fetch_sales(con) -> dict:
    """CRM / sales pipeline from customer + sales_touch; empty (not an error) if
    load_customers.py hasn't run. Read-only, like every other feed."""
    out = {"pipeline": [], "by_rep": [], "warm": [], "customers": [], "touch_log": [],
           "totals": {"customers": 0, "touches": 0, "interested": 0}}
    try:
        pipe = [dict(r) for r in con.execute(
            "SELECT sales_status, customers, touches FROM v_sales_pipeline")]
    except sqlite3.OperationalError:
        return out
    pipe.sort(key=lambda r: _SALES_ORDER.get(r["sales_status"], 9))
    out["pipeline"] = pipe
    out["by_rep"] = []
    for r in con.execute("SELECT rep, worked, contacted, interested, won FROM v_sales_by_rep ORDER BY worked DESC"):
        if _is_automation(r["rep"]):   # keep automation / import accounts out of the sales scoreboard
            continue
        out["by_rep"].append(dict(r))
    # touch logs grouped by customer (for the warm-account drill)
    touches: dict = {}
    for r in con.execute("SELECT customer_key, touch_date, note FROM sales_touch ORDER BY customer_key, seq"):
        touches.setdefault(r["customer_key"], []).append({"date": r["touch_date"], "note": r["note"]})
    for r in con.execute(
            "SELECT customer_key, name, division, last_contacted, last_edited_by, notion_url "
            "FROM customer WHERE sales_status = 'Interested' ORDER BY last_contacted DESC"):
        d = dict(r)
        d["last_edited_by"] = _rep_label(d["last_edited_by"])
        d["touches"] = touches.get(r["customer_key"], [])
        out["warm"].append(d)
    out["customers"] = []
    for r in con.execute("SELECT name, division, sales_status, main_status, last_contacted, "
                         "follow_up_date, last_edited_by, n_touches, notion_url "
                         "FROM customer ORDER BY last_contacted DESC"):
        d = dict(r); d["last_edited_by"] = _rep_label(d["last_edited_by"]); out["customers"].append(d)
    # full dated touch log with rep attribution (feeds the per-rep activity drill:
    # weekly/daily timeline + recent-touch list). Rep = the customer's last editor.
    for r in con.execute(
            "SELECT t.touch_date date, t.note, c.name customer, c.division, "
            "c.sales_status stage, c.last_edited_by "
            "FROM sales_touch t JOIN customer c ON c.customer_key = t.customer_key "
            "WHERE t.touch_date IS NOT NULL ORDER BY t.touch_date"):
        d = dict(r); d["rep"] = _rep_label(d.pop("last_edited_by")); out["touch_log"].append(d)
    tot = con.execute(
        "SELECT COUNT(*) c, COALESCE(SUM(n_touches),0) t, "
        "SUM(CASE WHEN sales_status='Interested' THEN 1 ELSE 0 END) i FROM customer").fetchone()
    out["totals"] = {"customers": tot["c"], "touches": tot["t"], "interested": tot["i"]}
    return out


def _fetch_sub_loc(con) -> dict:
    """Subcontractor LOC float from sub_loc_run / sub_loc_event; empty (not an error) if the
    loader hasn't run. summary.outstanding = fronted-but-uncollected NOW; peak = the LOC to size."""
    out = {"summary": None, "divisions": {}, "projects": [], "open_by_project": {}, "events": []}
    try:
        run = con.execute("SELECT * FROM sub_loc_run WHERE id=1").fetchone()
    except sqlite3.OperationalError:
        return out
    if not run:
        return out
    cols = run.keys()
    out["summary"] = {k: run[k] for k in cols if k not in ("divisions", "projects", "open_by_project", "id")}
    try:
        out["divisions"] = json.loads(run["divisions"] or "{}")
        out["projects"] = json.loads(run["projects"] or "[]")
        if "open_by_project" in cols:
            out["open_by_project"] = json.loads(run["open_by_project"] or "{}")
    except (ValueError, TypeError):
        pass
    try:
        # `settled` may be absent on a ledger written before this column existed - select it
        # defensively so an old DB degrades instead of throwing (the loader adds it on next run).
        ecols = {r[1] for r in con.execute("PRAGMA table_info(sub_loc_event)")}
        settled_col = ", settled" if "settled" in ecols else ""
        rows = con.execute("SELECT event_date, type, project, division, party, out_amt, in_amt, "
                           f"lag_days, balance, note, invoice, reimb{settled_col} FROM sub_loc_event ORDER BY seq")
        for r in rows:
            d = dict(r)
            for col in ("reimb", "settled"):
                if d.get(col):
                    try:
                        d[col] = json.loads(d[col])
                    except (ValueError, TypeError):
                        d[col] = []
            out["events"].append(d)
    except sqlite3.OperationalError:   # degrade to summary-only, never break the whole dashboard
        pass
    return out


def _connect(db_path: Path) -> sqlite3.Connection:
    """Open the ledger READ-ONLY. New connection per request (SQLite + threads)."""
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    return con


def fetch_data(db_path: Path, scope: str = "full") -> dict:
    """Return the ledger payload, or {error} if the db isn't ready.

    scope splits the load so the app is interactive on first paint instead of parsing the
    whole ~5 MB blob up front:
      * 'full'  - everything (used by manual + the 90s auto-refresh).
      * 'light' - the Overview core (projects/costs/draws/liens); the heavy tab blobs
                  (bills, sub_loc, payments, sales) are DEFERRED. Sent on first page load.
      * 'heavy' - only those deferred blobs, fetched in the background right after 'light'.
    """
    if not db_path.exists():
        return {"error": f"No ledger database at {db_path}. "
                         f"Run:  python3 ledger/load_wip_master.py"}
    try:
        con = _connect(db_path)
    except sqlite3.OperationalError as e:
        return {"error": f"Could not open {db_path}: {e}"}
    if scope == "heavy":                    # only the deferred, tab-specific blobs
        try:
            out = {"ap_bills": _fetch_ap(con).get("bills", []),
                   "sub_loc": _fetch_sub_loc(con),
                   "payments": _fetch_payments(con),
                   "sales": _fetch_sales(con)}
        except sqlite3.OperationalError as e:
            con.close()
            return {"error": f"Ledger schema not found ({e}). Run the loader first."}
        con.close()
        return out
    try:
        rows = [dict(r) for r in con.execute("SELECT * FROM v_wip_latest")]
        report_date = None
        pcount = len(rows)
        if rows:
            report_date = max((r.get("report_date") or "") for r in rows) or None
        loaded_at = None
        cur = con.execute("SELECT MAX(loaded_at) FROM wip_snapshot")
        got = cur.fetchone()
        if got:
            loaded_at = got[0]
        ap = _fetch_ap(con)
        proj_client = _project_customer_map(con)   # project -> client, to show on the projects list too
        costs = _fetch_costs(con)
        draws = _fetch_draws(con)
        open_invoices = _fetch_open_invoices(con)
        actions = _fetch_actions(con)
        freshness = _freshness(con)
        sub_loc = sales = payments = None          # heavy blobs: only on a full load (deferred otherwise)
        if scope == "full":
            sales = _fetch_sales(con)
            sub_loc = _fetch_sub_loc(con)
            payments = _fetch_payments(con)
        try:                                # realm for company-scoped QBO deep links (never logged)
            _mr = con.execute("SELECT value FROM meta WHERE key='qbo_realm'").fetchone()
            qbo_realm = _mr[0] if _mr else None
        except sqlite3.OperationalError:
            qbo_realm = None
    except sqlite3.OperationalError as e:
        con.close()
        return {"error": f"Ledger schema not found ({e}). Run the loader first."}
    con.close()
    for r in rows:  # attach QBO cost rollup + the client onto each project row
        cp = costs["by_project"].get(r["project_no"])
        r["costs_loaded"] = cp["costs_loaded"] if cp else None
        r["sub_costs"] = cp["sub_costs"] if cp else None
        r["client"] = proj_client.get(r["project_no"])
    for d in draws.get("draws", []):  # attach the Notion action link (if tracked)
        inv = (d.get("matched_invoice") or "").split("—")[0].strip()
        d["action"] = actions.get(f"draw:{inv}")
    if scope == "light":
        ap = {**ap, "bills": []}      # defer the ~2.7 MB bill list to the heavy fetch
    out = {
        "meta": {
            "db_path": str(db_path),
            "version": LEDGER_VERSION,
            "report_date": report_date,
            "loaded_at": loaded_at,
            "project_count": pcount,
            "freshness": freshness,
            "qbo_realm": qbo_realm,   # company-scopes the QBO deep links; None -> bare fallback
        },
        "projects": rows,
        "ap": ap,
        "cost": costs,
        "draws": draws,
        "open_invoices": open_invoices,
    }
    if scope == "full":               # the heavy tab blobs ride along on a full refresh
        out["sales"] = sales
        out["sub_loc"] = sub_loc
        out["payments"] = payments
    return out


class Handler(BaseHTTPRequestHandler):
    db_path: Path = DEFAULT_DB

    def log_message(self, *args):  # quiet; no request spam in the terminal
        pass

    def _send(self, code: int, body: bytes, ctype: str):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _json(self, obj, code: int = 200):
        self._send(code, json.dumps(obj, default=str).encode("utf-8"),
                   "application/json; charset=utf-8")

    def _static(self, name: str):
        # Serve only files that actually live in static/ (no traversal).
        target = (STATIC / name).resolve()
        if STATIC.resolve() not in target.parents or not target.is_file():
            self._send(404, b"not found", "text/plain; charset=utf-8")
            return
        ctype = CONTENT_TYPES.get(target.suffix, "application/octet-stream")
        self._send(200, target.read_bytes(), ctype)

    def _query(self) -> dict:
        return {k: v[0] for k, v in parse_qs(urlparse(self.path).query).items()}

    def _attachment(self, q: dict) -> None:
        """Resolve a bill's QBO scan link(s) on demand - subprocess the loader (never an
        import), which prints JSON {ok, files:[{name,url}]} with fresh TempDownloadUris."""
        bill = (q.get("bill") or "").strip()
        if not bill.isdigit():
            return self._json({"ok": False, "error": "bad bill id"}, 400)
        tx_type = q.get("type") or "Bill"
        try:
            out = subprocess.check_output(
                [sys.executable, str(PROJECT_ROOT / "ledger" / "attachments.py"), bill, "--type", tx_type],
                cwd=str(PROJECT_ROOT), stderr=subprocess.DEVNULL, timeout=60)
            lines = [ln for ln in out.decode("utf-8", "replace").splitlines() if ln.strip()]
            self._json(json.loads(lines[-1]) if lines else {"ok": False, "error": "no output"})
        except subprocess.TimeoutExpired:
            self._json({"ok": False, "error": "attachment lookup timed out"})
        except Exception as e:                          # noqa: BLE001
            self._json({"ok": False, "error": f"attachment lookup failed: {e}"})

    def do_GET(self):
        path = self.path.split("?", 1)[0]
        if path == "/":
            self._static("index.html")
        elif path == "/api/data":
            q = self._query()
            scope = "light" if q.get("light") else ("heavy" if q.get("heavy") else "full")
            self._json(fetch_data(self.db_path, scope))
        elif path == "/api/health":
            self._json({"ok": True})
        elif path == "/api/pnl":
            self._pnl_find(self._query().get("proj", ""))
        elif path == "/api/pnl/pl":       # live computed P&L (numbers folded into the UI)
            self._pnl_pl(self._query().get("proj", ""))
        elif path == "/api/pnl/portfolio":  # company P&L: every active job + totals
            self._pnl_portfolio()
        elif path == "/api/pnl/status":
            self._pnl_status(self._query().get("proj", ""))
        elif path == "/api/sync/status":   # progress for the in-app Resync / Console runs
            self._sync_status()
        elif path == "/api/pipelines":     # the Console registry (pipelines + their steps)
            self._pipelines_list()
        elif path == "/api/processes":     # the Systems tab (vault process registry, live)
            self._processes()
        elif path == "/api/graph":         # the Graph tab (vault link-graph + system diagrams, live)
            self._graph()
        elif path == "/api/wip/review":    # the WIP Review tab (pending before/after diff, merged)
            self._wip_review_get()
        elif path == "/api/accounting":    # the Audit tab (Bill Tracker audits, live)
            try:
                self._json(_fetch_accounting_audits())
            except Exception as e:         # noqa: BLE001
                self._json({"ok": False, "findings": [], "error": f"audit read failed: {e}"})
        elif path == "/api/attachment":    # a bill's scan link(s), resolved fresh from QBO on click
            self._attachment(self._query())
        elif path.startswith("/static/"):
            self._static(path[len("/static/"):])
        else:
            self._send(404, b"not found", "text/plain; charset=utf-8")

    def do_POST(self):
        p = urlparse(self.path).path
        if p == "/api/waiver":            # a ledger write - the owner's waiver marks
            self._set_waiver()
        elif p == "/api/bill-mark":       # a ledger write - the owner's lien mark (mirrors to the workbook on sync-ap)
            self._set_bill_mark()
        elif p == "/api/pay-run":         # a ledger write - the owner's check-run worksheet (LOCAL only, never pays QBO)
            self._save_pay_run()
        elif p == "/api/pay-run/clear":   # empty the whole pay run (after the check run is done)
            self._clear_pay_run()
        elif p == "/api/pnl/open":        # open the P&L workbook (or ?folder=1 → its folder), cross-platform
            self._pnl_open(self._query().get("proj", ""), folder=self._query().get("folder") == "1")
        elif p == "/api/job/open":        # open the SOURCE job folder (Synology CP/RP, OneDrive MFD)
            self._job_open(self._query().get("proj", ""))
        elif p == "/api/lien/folder":     # open the Synology Vendor Liens folder (where notice/lien PDFs live)
            self._lien_folder(self._query().get("vendor", ""))
        elif p == "/api/pnl/generate":    # run project-pnl (gated by an explicit confirm)
            self._pnl_generate()
        elif p == "/api/sync":            # in-app Resync: run the ledger loaders (gated)
            self._sync_start()
        elif p == "/api/wip/review":      # WIP Review: compute the pending diff (no writes)
            self._wip_review_run()
        elif p == "/api/wip/merge":       # WIP Review: write approved changes to the 3 Test tabs (gated)
            self._wip_merge()
        else:
            self._send(404, b"not found", "text/plain; charset=utf-8")

    # ── P&L link handlers ───────────────────────────────────────────────────
    def _pnl_pl(self, proj: str):
        """Live computed P&L for the job detail — the numbers, not just a link."""
        proj = (proj or "").strip().upper()
        if not _PROJ_RE.match(proj):
            return self._json({"error": "bad or missing project"}, 400)
        con = _connect(self.db_path)
        try:
            self._json(_project_pnl(con, proj))
        finally:
            con.close()

    def _pnl_portfolio(self):
        con = _connect(self.db_path)
        try:
            self._json(_portfolio_pnl(con))
        finally:
            con.close()

    # ── in-app Console: run any pipeline (or the safe loaders-only reload) ────
    def _processes(self):
        """The systems & process registry, parsed fresh from the vault markdown.

        Read-only and uncached on purpose: the vault files are the source of
        truth, so editing them and reloading the tab is the whole update loop.
        A missing vault is reported in the payload, never raised - the rest of
        the dashboard must keep working on a machine with no vault checkout.
        """
        try:
            self._json(registry_view.load_registry())
        except Exception as e:                      # noqa: BLE001
            self._json({"ok": False, "domains": [], "rows": [],
                        "error": f"registry parse failed: {e}"})

    def _graph(self):
        """The org map (vault notes + [[wikilinks]]) and the imported system
        diagrams (mermaid in docs/ARCHITECTURE.md). Same contract as _processes:
        parsed live, never cached, read-only, and a missing vault is reported in
        the payload rather than raised so the rest of the dashboard keeps working.
        """
        try:
            self._json(vault_graph.load_all())
        except Exception as e:                      # noqa: BLE001
            self._json({"org": {"ok": False, "nodes": [], "links": [],
                                "error": f"graph parse failed: {e}"},
                        "diagrams": {"ok": False, "diagrams": []}})

    def _pipelines_list(self):
        out = []
        for p in _pipelines():
            if p.get("hidden"):        # kept in reload/all, just no standalone Console card
                continue
            out.append({
                "key": p["key"], "label": p["label"],
                "steps": [{"label": s["label"], "side": bool(s.get("side"))} for s in p["steps"]],
                "has_producer": any(s.get("side") for s in p["steps"]),
                "draft": ({"label": p["draft"]["label"]} if p.get("draft") else None),
                "actions": p.get("actions") or [],
            })
        self._json({"pipelines": out})

    def _sync_status(self):
        with _SYNC_LOCK:
            out = {"state": _SYNC["state"], "current": _SYNC["current"], "pipeline": _SYNC["pipeline"],
                   "steps": [dict(s) for s in _SYNC["steps"]]}
            if _SYNC["state"] == "running":
                out["elapsed"] = int(time.time() - _SYNC["started"])
        self._json(out)

    def _sync_start(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        if not body.get("confirm"):                   # gate: no confirm, no run
            return self._json({"error": "confirm required"}, 400)
        pipeline = str(body.get("pipeline") or "reload")   # default = safe loaders-only reload
        steps = _resolve_steps(pipeline)
        if not steps:
            return self._json({"error": f"unknown pipeline '{pipeline}'"}, 400)
        return self._launch(steps, pipeline)

    def _launch(self, steps, pipeline):
        """Claim the single run-lock and spawn the steps thread. Shared by the
        Console sync and the WIP Review emit/apply runs so only ONE can hold the
        lock at a time (concurrent QBO pulls + tab writes would corrupt)."""
        with _SYNC_LOCK:
            if _SYNC["state"] == "running":
                return self._json({"error": "a sync is already running", "running": True}, 409)
            # claim it INSIDE the lock so a second POST can't also see "idle" and launch
            # a second run that writes at once (TOCTOU). Single authoritative claim.
            _SYNC["state"] = "running"
            _SYNC["started"] = time.time()
            _SYNC["current"] = -1
            _SYNC["pipeline"] = pipeline
            _SYNC["steps"] = [{"label": s["label"], "state": "pending"} for s in steps]
        threading.Thread(target=_run_sync, args=(steps,), daemon=True).start()
        return self._json({"ok": True, "pipeline": pipeline, "steps": [s["label"] for s in steps]})

    # ── WIP Review endpoints (emit the diff · read it · write approved) ──────
    def _wip_review_run(self):
        """Start the three emit runs (CP/RP/MFD) - compute the pending update and
        diff each Test tab, no writes. Confirm-gated; each run does a QBO pull."""
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        if not body.get("confirm"):
            return self._json({"error": "confirm required"}, 400)
        return self._launch(_wip_review_steps("emit"), "wip-review")

    def _wip_review_get(self):
        """Merge the three emit JSONs into one review payload for the UI."""
        out = {"ok": True, "generated": {}, "records": [], "ready": False}
        for div, fn in _WIP_REVIEW_FILES:
            p = _WIP_REVIEW_DIR / fn
            if not p.exists():
                continue
            try:
                d = json.loads(p.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                continue
            out["ready"] = True
            out["records"].extend(d.get("records", []))
            out["generated"][div] = {"tab": d.get("tab"), "count": d.get("count"),
                                     "changed": d.get("changed"),
                                     "at": _dt.datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds")}
        dec = _WIP_REVIEW_DIR / "decisions.json"
        try:
            out["decisions"] = json.loads(dec.read_text(encoding="utf-8")) if dec.exists() else {}
        except (OSError, ValueError):
            out["decisions"] = {}
        changed = [r for r in out["records"] if r["status"] != "SAME"]
        out["counts"] = {"jobs": len(out["records"]), "changed": len(changed),
                         "added": sum(r["status"] == "ADDED" for r in out["records"]),
                         "removed": sum(r["status"] == "REMOVED" for r in out["records"])}
        return self._json(out)

    def _wip_merge(self):
        """Save the owner's decisions, then start the three guarded writes."""
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        if not body.get("confirm"):
            return self._json({"error": "confirm required"}, 400)
        decisions = body.get("decisions") or {}
        _WIP_REVIEW_DIR.mkdir(parents=True, exist_ok=True)
        (_WIP_REVIEW_DIR / "decisions.json").write_text(
            json.dumps(decisions, indent=2), encoding="utf-8")
        return self._launch(_wip_review_steps("apply"), "wip-merge")

    def _job_open(self, proj: str):
        """Open the SOURCE job folder on the file server (docs/takeoffs/photos) —
        CP → the Synology awarded folder, RP → the builder folder, MFD → its OneDrive
        folder as a fallback. Cross-platform via _os_open."""
        proj = (proj or "").strip().upper()
        if not _PROJ_RE.match(proj):
            return self._json({"error": "bad project"}, 400)
        con = _connect(self.db_path)
        try:
            r = con.execute("SELECT builder_or_gc FROM project WHERE project_no = ?", (proj,)).fetchone()
        finally:
            con.close()
        folder, note = pnl_paths.job_folder(proj, r["builder_or_gc"] if r else None)
        if folder is None:                            # MFD / unresolved → the P&L's folder (OneDrive)
            info = pnl_paths.find_pnl(proj)
            if info.get("exists"):
                folder = Path(info["path"]).parent
            else:
                return self._json({"error": note or "no folder found"}, 404)
        err = _os_open(str(folder))
        if err:
            return self._json({"error": err}, 500)
        self._json({"ok": True, "path": str(folder), "note": note})

    def _lien_folder(self, vendor: str = ""):
        """Open the Synology Vendor Liens folder. With ?vendor=, drill to (and create, if missing)
        that vendor's subfolder - organize by vendor, auto-create (owner 2026-08-20); without it,
        open the base. Only ever under _LIEN_FOLDER (no path traversal). Cross-platform."""
        target = Path(_LIEN_FOLDER)
        if (vendor or "").strip():
            d = _ensure_lien_vendor_dir(vendor)
            if d is not None:
                target = d                       # else fall back to the base folder
        if not target.exists():
            return self._json({"error": "folder not found (is the Accounting share mounted?)"}, 404)
        err = _os_open(str(target))
        if err:
            return self._json({"error": err}, 500)
        self._json({"ok": True, "path": str(target)})

    def _pnl_find(self, proj: str):
        proj = (proj or "").strip().upper()
        if not _PROJ_RE.match(proj):
            return self._json({"error": "bad or missing project"}, 400)
        info = pnl_paths.find_pnl(proj)
        j = _PNL_JOBS.get(proj)
        info["job"] = j["state"] if j else "idle"
        self._json(info)

    def _pnl_status(self, proj: str):
        proj = (proj or "").strip().upper()
        j = _PNL_JOBS.get(proj)
        if not j:
            return self._json({"state": "idle"})
        out = {"state": j["state"]}
        if j["state"] == "running":
            out["elapsed"] = int(time.time() - j["started"])
            out["status"] = _pnl_last_line(j.get("log"))   # one live line: where the run is at
        if j.get("detail"):
            out["detail"] = j["detail"]
        self._json(out)

    def _pnl_open(self, proj: str, folder: bool = False):
        """Open the project's P&L workbook (or its containing folder when folder=True)
        in the host OS file manager. The LOCAL server opens it with the native command
        so the same dashboard works on Mac or Windows — CP resolves onto the Synology
        Common drive, RP/MFD onto OneDrive, per pnl_paths (the owner's convention)."""
        proj = (proj or "").strip().upper()
        if not _PROJ_RE.match(proj):
            return self._json({"error": "bad project"}, 400)
        info = pnl_paths.find_pnl(proj)
        if not info.get("exists"):
            return self._json({"error": "no P&L generated yet"}, 404)
        path = Path(info["path"])
        if path.name != f"Project_PnL_{proj}.xlsx":   # only ever open the resolved workbook / its folder
            return self._json({"error": "unexpected file"}, 400)
        target = path.parent if folder else path
        err = _os_open(str(target))
        if err:
            return self._json({"error": err}, 500)
        self._json({"ok": True, "path": str(target)})

    def _pnl_generate(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        proj = str(body.get("proj", "")).strip().upper()
        if not _PROJ_RE.match(proj):
            return self._json({"error": "bad or missing project"}, 400)
        if not body.get("confirm"):                   # gate: no confirm, no run
            return self._json({"error": "confirm required"}, 400)
        if not _PNL_RUN.exists():
            return self._json({"error": "project-pnl runner not found"}, 500)
        with _PNL_LOCK:
            j = _PNL_JOBS.get(proj)
            if j and j["state"] == "running":
                return self._json({"state": "running", "proj": proj, "already": True})
            try:
                _PNL_LOG_DIR.mkdir(parents=True, exist_ok=True)
                f = open(_PNL_LOG_DIR / f"{proj}.log", "w")
                proc = subprocess.Popen(
                    ["/bin/bash", str(_PNL_RUN), proj],
                    cwd=str(_PNL_DIR), stdout=f, stderr=subprocess.STDOUT,
                    env={**os.environ, "PYTHONUNBUFFERED": "1"})   # flush progress live so the status line updates
            except OSError as e:
                return self._json({"error": f"spawn failed: {e}"}, 500)
            _PNL_JOBS[proj] = {"state": "running", "started": time.time(),
                               "log": str(_PNL_LOG_DIR / f"{proj}.log"),
                               "proc": proc, "file": f}
        threading.Thread(target=_pnl_wait, args=(proj,), daemon=True).start()
        self._json({"state": "running", "proj": proj})

    def _set_waiver(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        mi, vendor, bill = body.get("matched_invoice"), body.get("vendor"), body.get("bill_ref")
        received = 1 if body.get("received") else 0
        if not mi:
            return self._json({"error": "matched_invoice required"}, 400)
        key = _waiver_key(mi, vendor, bill)
        now = _dt.datetime.now().isoformat(timespec="seconds")
        try:
            con = sqlite3.connect(self.db_path)          # WRITABLE (the one write surface)
            con.execute("CREATE TABLE IF NOT EXISTS waiver (waiver_key TEXT PRIMARY KEY, "
                        "matched_invoice TEXT, vendor TEXT, bill_ref TEXT, received INTEGER NOT NULL "
                        "DEFAULT 0, received_date TEXT, note TEXT, updated_at TEXT NOT NULL)")
            con.execute(
                "INSERT INTO waiver (waiver_key, matched_invoice, vendor, bill_ref, received, "
                "received_date, updated_at) VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(waiver_key) DO UPDATE SET received=excluded.received, "
                "received_date=excluded.received_date, updated_at=excluded.updated_at",
                (key, mi, vendor, bill, received, now if received else None, now))
            con.commit(); con.close()
        except sqlite3.OperationalError as e:
            return self._json({"error": f"write failed: {e}"}, 500)
        self._json({"ok": True, "received": bool(received), "waiver_key": key})

    def _set_bill_mark(self):
        """Set / clear a bill's lien tag (Notice Sent / Lien Filed / Released). Persists in the
        ledger overlay instantly; the next sync-ap mirrors it into the workbook's Lien cell."""
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        bill_id = str(body.get("bill_id") or "").strip()
        lien = (body.get("lien") or "").strip()        # "" clears the mark
        if not bill_id:
            return self._json({"error": "bill_id required (this bill has no QBO bill link)"}, 400)
        if lien and lien not in bill_marks.LIEN_STATES:
            return self._json({"error": f"lien must be one of {bill_marks.LIEN_STATES} or empty"}, 400)
        try:
            bill_marks.set_lien_mark(bill_id, lien, _dt.datetime.now().isoformat(timespec="seconds"))
        except sqlite3.OperationalError as e:
            return self._json({"error": f"write failed: {e}"}, 500)
        # Auto-create the vendor's lien folder on an ACTUAL notice/lien (owner 2026-08-20) so the
        # PDF has a home. Only Notice Sent / Lien Filed - a cleared or Released mark makes no folder.
        folder = None
        vendor = (body.get("vendor") or "").strip()
        if lien in ("Notice Sent", "Lien Filed") and vendor:
            d = _ensure_lien_vendor_dir(vendor)
            folder = str(d) if d is not None else None
        self._json({"ok": True, "bill_id": bill_id, "lien": lien, "folder": folder})

    def _save_pay_run(self):
        """Persist the check-run worksheet: which bills to pay + a partial amount override.
        LOCAL intent only - this NEVER pays QBO or moves money; the owner records the real
        payment in QBO and the next sync-ap pulls the true status back."""
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self._json({"error": "bad request"}, 400)
        items = body.get("items")
        if not isinstance(items, list):
            return self._json({"error": "items (a list) required"}, 400)
        try:
            kept = bill_marks.set_pay_marks(items, _dt.datetime.now().isoformat(timespec="seconds"))
        except (sqlite3.OperationalError, ValueError, TypeError) as e:
            return self._json({"error": f"write failed: {e}"}, 500)
        self._json({"ok": True, "count": kept})

    def _clear_pay_run(self):
        """Empty the whole pay run (after the check run is done)."""
        try:
            n = bill_marks.clear_pay_marks()
        except sqlite3.OperationalError as e:
            return self._json({"error": f"write failed: {e}"}, 500)
        self._json({"ok": True, "cleared": n})


def _daemonize():
    """Double-fork + setsid so the server outlives whatever launched it. A GUI app's
    `do shell script` (Project Ledger.app) reaps ordinary backgrounded children when it
    returns; a process in its OWN session survives. stdout/stderr stay on whatever the
    caller redirected (the launcher points them at a log)."""
    if os.fork() > 0:
        os._exit(0)                       # first parent exits → caller returns immediately
    os.setsid()                           # new session, detached from the controlling group
    if os.fork() > 0:
        os._exit(0)                       # second parent exits → can't reacquire a terminal
    devnull = os.open(os.devnull, os.O_RDWR)
    os.dup2(devnull, 0)                   # stdin ← /dev/null (nothing keeps us tethered)


def main():
    ap = argparse.ArgumentParser(description="Local web dashboard over the project ledger.")
    ap.add_argument("--db", type=Path, default=DEFAULT_DB, help="SQLite ledger to read.")
    ap.add_argument("--port", type=int, default=8787, help="Port (default 8787).")
    ap.add_argument("--no-open", action="store_true", help="Don't auto-open a browser.")
    ap.add_argument("--background", action="store_true",
                    help="Detach into a new session (daemonize) and serve in the background — "
                         "so a GUI launcher (Project Ledger.app) can't reap it.")
    args = ap.parse_args()

    if args.background:                   # detach BEFORE binding, so the grandchild owns the socket
        _daemonize()

    Handler.db_path = args.db
    url = f"http://127.0.0.1:{args.port}"
    httpd = ThreadingHTTPServer(("127.0.0.1", args.port), Handler)

    ready = fetch_data(args.db)
    if "error" in ready:
        print(f"⚠  {ready['error']}")
    else:
        m = ready["meta"]
        print(f"Ledger: {m['project_count']} projects · report {m['report_date']} · {m['db_path']}")
    print(f"Dashboard: {url}   (Ctrl-C to stop)")

    # Quit cleanly on SIGTERM too — when the server IS the app process (Project
    # Ledger.app runs it in the foreground), Cmd-Q / Dock-Quit / logout sends SIGTERM.
    # default_int_handler raises KeyboardInterrupt, same as Ctrl-C, so the handler below runs.
    signal.signal(signal.SIGTERM, signal.default_int_handler)

    if not args.no_open:
        threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        httpd.shutdown()


if __name__ == "__main__":
    main()
