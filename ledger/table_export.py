"""table_export.py - a filtered ledger table as an Excel report (owner 2026-09-02: "export to Excel
that way we can keep the conditional formatting ... ask me how i would like it grouped").

POST /api/export/xlsx  {name, title, subtitle, columns:[{label,type}], rows:[[...]], group_by: <col idx|null>,
                        fmt:[{row idx, col idx, cls}]}  ->  ~/Downloads/<name> <stamp>.xlsx, revealed in Finder.

The workbook is a REPORT of what the person had on screen: title block, one header row (frozen,
autofiltered), rows grouped under a band per group with a subtotal, Excel outline groups so a band
collapses, a grand total. Colour ENCODES STATE ONLY (paid green / open red / lien or approval
amber) - the owner asked for exactly this on this export; every other workbook stays plain (repo
rule 5). Last step is always shared/xlsx_verify.assert_clean (rule 5b).
"""
from __future__ import annotations

import datetime as dt
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook                                  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter                   # noqa: E402

from shared import xlsx_verify                                 # noqa: E402

_MONEY = '"$"#,##0.00_);[Red]("$"#,##0.00)'
_FILL_HDR = PatternFill("solid", fgColor="D9D9D9")
_FILL_BAND = PatternFill("solid", fgColor="F2F2F2")
_FILLS = {"pos": PatternFill("solid", fgColor="E2F0D9"),    # paid / settled
          "neg": PatternFill("solid", fgColor="F8D7DA"),    # open / owed
          "warn": PatternFill("solid", fgColor="FFF2CC")}   # lien risk / not approved
_FONTS = {"pos": Font(color="1F7A4D"), "neg": Font(color="B4341E", bold=True), "warn": Font(color="8A6508")}
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(bottom=_THIN)
_SAFE = re.compile(r"[^A-Za-z0-9 _.()#-]+")


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build(spec: dict, dest_dir: Path) -> Path:
    cols = spec.get("columns") or []
    rows = spec.get("rows") or []
    gcol = spec.get("group_by")
    gcol = int(gcol) if gcol is not None and str(gcol) != "" else None
    fmt = {(int(f["r"]), int(f["c"])): f.get("cls") for f in (spec.get("fmt") or []) if f.get("cls") in _FILLS}
    money_cols = {i for i, c in enumerate(cols) if (c.get("type") == "money")}
    name = _SAFE.sub("", str(spec.get("name") or "ledger export")).strip() or "ledger export"
    stamp = dt.datetime.now().strftime("%m-%d-%Y %H%M")
    dest_dir.mkdir(parents=True, exist_ok=True)
    path = dest_dir / f"{name} {stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = (_SAFE.sub("", str(spec.get("sheet") or name))[:31] or "Report")
    ws["A1"] = str(spec.get("title") or name); ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = str(spec.get("subtitle") or f"as of {dt.date.today().strftime('%m/%d/%Y')}"); ws["A2"].font = Font(color="595959")
    hdr_row = 4
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=hdr_row, column=j, value=str(c.get("label") or ""))
        cell.font = Font(bold=True); cell.fill = _FILL_HDR; cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # order: keep the on-screen order inside each group; groups in first-seen order
    order: list = []
    groups: dict = {}
    for ri, r in enumerate(rows):
        key = (str(r[gcol]) if gcol is not None and gcol < len(r) and r[gcol] not in (None, "") else "(none)") if gcol is not None else "__all__"
        if key not in groups:
            groups[key] = []; order.append(key)
        groups[key].append(ri)

    r = hdr_row + 1
    grand = {j: 0.0 for j in money_cols}
    for key in order:
        idxs = groups[key]
        band_row = None
        if gcol is not None:
            band_row = r
            ws.cell(row=r, column=1, value=f"{key}  ({len(idxs)})").font = Font(bold=True)
            for j in range(1, len(cols) + 1):
                ws.cell(row=r, column=j).fill = _FILL_BAND
            r += 1
        first = r
        sub = {j: 0.0 for j in money_cols}
        for ri in idxs:
            row = rows[ri]
            for j, c in enumerate(cols):
                v = row[j] if j < len(row) else None
                cell = ws.cell(row=r, column=j + 1)
                if j in money_cols:
                    n = _num(v); cell.value = n; cell.number_format = _MONEY
                    if n is not None:
                        sub[j] += n; grand[j] += n
                else:
                    cell.value = "" if v is None else v
                cls = fmt.get((ri, j))
                if cls:
                    cell.fill = _FILLS[cls]; cell.font = _FONTS[cls]
            r += 1
        if gcol is not None:
            for j in money_cols:
                cell = ws.cell(row=band_row, column=j + 1, value=round(sub[j], 2)); cell.number_format = _MONEY; cell.font = Font(bold=True)
            if r - 1 >= first:
                ws.row_dimensions.group(first, r - 1, outline_level=1, hidden=False)
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for j in money_cols:
        cell = ws.cell(row=r, column=j + 1, value=round(grand[j], 2)); cell.number_format = _MONEY; cell.font = Font(bold=True)
    for j in range(1, len(cols) + 1):
        ws.cell(row=r, column=j).border = Border(top=Side(style="medium"))
    # widths from content, autofilter on the header block
    for j, c in enumerate(cols, start=1):
        longest = max([len(str(c.get("label") or ""))] + [len(str(rw[j - 1])) for rw in rows if j - 1 < len(rw) and rw[j - 1] is not None][:400])
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(9, longest + 2))
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(cols))}{max(r - 1, hdr_row)}"
    ws.sheet_properties.outlinePr.summaryBelow = False
    wb.save(path)
    xlsx_verify.assert_clean(path)      # rule 5b - the LAST step before anyone sees the file
    return path


def _selftest() -> None:
    import tempfile
    spec = {"name": "selftest bills", "title": "Bill Tracker - CP800 to 07/25/2026", "subtitle": "44 bills",
            "columns": [{"label": "Vendor"}, {"label": "Bill #"}, {"label": "Amount", "type": "money"}, {"label": "Open", "type": "money"}, {"label": "Paid"}],
            "rows": [["VENDOR A", "1001", 100.5, 0, "Paid 08/01/2026"], ["VENDOR A", "1002", 50, 50, "Open"], ["VENDOR B", "77", 20, 20, "Open"]],
            "group_by": 0, "fmt": [{"r": 0, "c": 4, "cls": "pos"}, {"r": 1, "c": 3, "cls": "neg"}, {"r": 2, "c": 3, "cls": "neg"}]}
    with tempfile.TemporaryDirectory() as d:
        p = build(spec, Path(d))
        from openpyxl import load_workbook
        ws = load_workbook(p).active
        assert ws["A4"].value == "Vendor" and ws["A5"].value.startswith("VENDOR A") and ws["C5"].value == 150.5
        assert ws["A8"].value.startswith("VENDOR B") and ws["A10"].value == "TOTAL" and ws["D10"].value == 70
        assert ws.freeze_panes == "A5" and xlsx_verify.verify_xlsx(p) == []
    print("table_export selftest OK - bands + subtotals + total, money formats, state fills, clean xlsx")


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        _selftest()
    else:
        print("usage: table_export.py --selftest   (served by dashboard.py as POST /api/export/xlsx)")
