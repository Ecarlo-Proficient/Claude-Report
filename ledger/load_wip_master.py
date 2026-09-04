#!/usr/bin/env python3
"""
load_wip_master.py — land the FINAL WIP master sheet into the project ledger.

WHAT IT DOES
Reads the three Test tabs of "WIP - MASTER new.xlsx" (the finalized WIP report)
and populates two tables of the canonical ledger:
    project        one row per real job   (identity: division, type, builder, ...)
    wip_snapshot   one row per (project, report_date)  (the computed WIP position)

It DOES NOT generate anything from QBO — the master sheet is the source of truth
for this load. The granular cost_code / budget_line / cost_line / billing_event
tables in schema.sql are filled later by the QBO connectors, not here.

EACH PROJECT IS READ FROM ITS RICHEST SOURCE, exactly once:
    Test - CP     -> division Commercial   (has retainage, notes, over/under, earned)
    Test - RP     -> division Residential  (has builder, Tract/Custom, category, marks)
    Test-Master   -> division Multi Family (MFD rows only; MFD has no own Test tab)

SAFETY
    * The Excel workbook is opened READ-ONLY — this tool never writes the sheet.
    * Writes go to a local SQLite file (default) or any DB you point --db at.
    * Upserts are idempotent: re-running replaces the same (project) / (project,
      report_date) rows, never duplicates them.
    * --dry-run parses and reports counts WITHOUT touching the database.

USAGE
    python3 ledger/load_wip_master.py --dry-run          # preview, write nothing
    python3 ledger/load_wip_master.py                    # load into the default SQLite db
    python3 ledger/load_wip_master.py --show 8           # load, then print 8 sample rows
    python3 ledger/load_wip_master.py --db /path/to.sqlite3 --excel "/path/WIP - MASTER new.xlsx"

Postgres: this loader targets SQLite for the zero-install spike. The SAME
schema.sql deploys to Postgres unchanged; pointing the load at Postgres is a
driver swap (psycopg) with the identical INSERT ... ON CONFLICT statements.
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths  # noqa: E402

HERE = Path(__file__).resolve().parent
SCHEMA_SQL = HERE / "schema.sql"

# Default local database — outside the repo (never committed), created on demand.
DEFAULT_DB = (
    Path.home() / "Library" / "Application Support" / "Proficient" / "ledger.sqlite3"
)

# Same default the WIP writer uses, so this reads the exact file the readers write.
DEFAULT_EXCEL = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx",
)

HEADER_ROW = 3               # the Test tabs put column headers on row 3
DATE_ROW = 2                 # "REPORT DATE: AUG 07, 2026" lives on row 2
PROJECT_RE = re.compile(r"^(MFD|CP|RP)\d+(-FTW)?$", re.IGNORECASE)

# canonical field  ->  the header label that carries it on the Test tabs.
# ('TYPE' is deliberately NOT here: on Test - RP it means Tract/Custom, but on
#  Test-Master it means the division/home-type string. Handled per-tab below.)
FIELD_HEADERS = {
    "project_no":               "PROJECT #",
    "name":                     "PROJECT NAME",
    "status":                   "STATUS",
    "original_contract":        "ORIGINAL CONTRACT",
    "approved_cos":             "APPROVED COs",
    "total_contract_price":     "TOTAL CONTRACT PRICE",
    "original_estimated_cost":  "ORIGINAL ESTIMATED COST",
    "co_costs":                 "CO COSTS",
    "estimated_total_costs":    "ESTIMATED TOTAL COSTS",
    "original_profit":          "ORIGINAL PROFIT",
    "gross_profit_pct":         "GROSS PROFIT %",
    "costs_to_date":            "COSTS TO DATE",
    "cost_to_complete":         "COST TO COMPLETE",
    "percent_complete":         "PERCENT COMPLETE",
    "revenues_earned_to_date":  "REVENUES EARNED TO DATE",
    "profit_earned_to_date":    "PROFIT EARNED TO DATE",
    "billed_to_date":           "BILLED TO DATE",
    "overbillings":             "OVERBILLINGS",
    "underbillings":            "UNDERBILLINGS",
    "retainage_held":           "RETAINAGE HELD",
    "left_to_bill":             "LEFT TO BILL",
    "future_profit_to_earn":    "FUTURE PROFIT TO EARN",
    "pure_job_borrow":          "PURE JOB BORROW",
    "notes":                    "NOTES",
}

# per-tab load plan: which division, plus the tab-specific columns.
TAB_PLAN = [
    {
        "tab": "Test - CP",
        "division": "Commercial",
        "mfd_only": False,
    },
    {
        "tab": "Test - RP",
        "division": "Residential",
        "mfd_only": False,
        "type_header": "TYPE",          # Tract / Custom
        "builder_header": "BUILDER",
        "category_header": "CATEGORY",
        "marks": {
            "mark_schedule": "SCHEDULE",
            "mark_general_list": "GENERAL LIST",
            "mark_jobtread": "JOBTREAD",
        },
    },
    {
        "tab": "Test-Master",
        "division": "Multi Family",
        "mfd_only": True,               # keep only rows whose TYPE starts 'Multi-Family'
        "bonded_header": "BONDED",
    },
]

SNAPSHOT_FIELDS = [
    "status", "original_contract", "approved_cos", "total_contract_price",
    "original_estimated_cost", "co_costs", "estimated_total_costs",
    "original_profit", "gross_profit_pct", "costs_to_date", "cost_to_complete",
    "percent_complete", "revenues_earned_to_date", "profit_earned_to_date",
    "billed_to_date", "overbillings", "underbillings", "retainage_held",
    "left_to_bill", "future_profit_to_earn", "pure_job_borrow",
    "mark_schedule", "mark_general_list", "mark_jobtread", "notes",
]


def _num(v):
    """Return a float if the cell is numeric, else None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _bonded(v):
    s = _text(v)
    if s is None:
        return None
    return 1 if s.upper().startswith("Y") else 0


def _report_date(ws) -> str:
    """Parse 'REPORT DATE: AUG 07, 2026' from row 2 -> '2026-08-07'."""
    for cell in ws[DATE_ROW]:
        s = _text(cell.value)
        if s and "REPORT DATE" in s.upper():
            raw = s.split(":", 1)[1].strip()
            for fmt in ("%b %d, %Y", "%B %d, %Y"):
                try:
                    return dt.datetime.strptime(raw, fmt).date().isoformat()
                except ValueError:
                    continue
    return None


def _header_index(ws):
    """Map header label -> 0-based column index from the header row."""
    idx = {}
    for j, cell in enumerate(ws[HEADER_ROW]):
        label = _text(cell.value)
        if label and label not in idx:
            idx[label] = j
    return idx


def read_tab(ws, plan):
    """Yield (project_dict, snapshot_dict) for every real project row on a tab."""
    hidx = _header_index(ws)
    report_date = _report_date(ws)

    def cell(row, header):
        j = hidx.get(header)
        return row[j] if (j is not None and j < len(row)) else None

    pno_col = hidx.get("PROJECT #")
    type_col = hidx.get(plan.get("type_header")) if plan.get("type_header") else None
    mfd_type_col = hidx.get("TYPE") if plan["mfd_only"] else None

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        pno = _text(row[pno_col]) if (pno_col is not None and pno_col < len(row)) else None
        if not pno or not PROJECT_RE.match(pno.upper()):
            continue  # skips every legend / total / section-break row
        if plan["mfd_only"]:
            tv = _text(row[mfd_type_col]) if mfd_type_col is not None else None
            if not (tv and tv.startswith("Multi-Family")):
                continue

        proj = {
            "project_no": pno,
            "division": plan["division"],
            "is_ftw": 1 if pno.upper().endswith("-FTW") else 0,
            "name": _text(cell(row, "PROJECT NAME")),
            "type": _text(row[type_col]) if type_col is not None else None,
            "builder_or_gc": _text(cell(row, plan["builder_header"])) if plan.get("builder_header") else None,
            "bonded": _bonded(cell(row, plan["bonded_header"])) if plan.get("bonded_header") else None,
            "rp_category": _text(cell(row, plan["category_header"])) if plan.get("category_header") else None,
        }

        snap = {"project_no": pno, "report_date": report_date, "source_tab": ws.title}
        # numeric / text WIP figures shared across tabs (absent header -> None)
        for field, header in FIELD_HEADERS.items():
            if field in ("project_no", "name"):
                continue
            raw = cell(row, header)
            snap[field] = _text(raw) if field in ("status", "notes") else _num(raw)
        # RP cross-check marks
        for mfield, header in plan.get("marks", {}).items():
            snap[mfield] = _text(cell(row, header))

        _fill_derived(snap)
        yield proj, snap


def _fill_derived(snap):
    """Fill the WIP derived columns from the inputs when they came back blank.

    The Test tabs carry the derived figures (TOTAL CONTRACT PRICE, ESTIMATED TOTAL
    COSTS, PERCENT COMPLETE, ...) as Excel FORMULAS. openpyxl reads a formula cell
    as None whenever the workbook's cached values were stripped - which every script
    write to the tabs does - so a routine load leaves those columns NULL. Without
    contract or % complete the ledger's Project P&L reads every job as a total loss
    (owner 2026-08-27: "the P&L is not functioning ... due to it needing data").

    So compute them here from the input columns, using the SAME formulas the sheet
    uses (wip_writer's column guide): only ever FILLS a blank - a real typed value
    (e.g. MFD's contract off the WIP Master tab) is left untouched."""
    def n(k):
        v = snap.get(k)
        return v if isinstance(v, (int, float)) else None

    def put(k, v):
        if snap.get(k) in (None, "") and v is not None:
            snap[k] = round(v, 4)

    oc, co = n("original_contract"), (n("approved_cos") or 0.0)
    oec, coc = n("original_estimated_cost"), (n("co_costs") or 0.0)
    i_cost, billed = n("costs_to_date"), n("billed_to_date")
    tcp = n("total_contract_price") or ((oc + co) if oc is not None else None)          # C = A + B
    etc = n("estimated_total_costs") or ((oec + coc) if oec is not None else None)      # F = D + E
    put("total_contract_price", tcp)
    put("estimated_total_costs", etc)
    gp = (tcp - etc) if (tcp is not None and etc is not None) else None                 # G = C - F
    put("original_profit", gp)
    put("gross_profit_pct", (gp / tcp) if (gp is not None and tcp) else None)           # H = G / C
    put("cost_to_complete", (etc - i_cost) if (etc is not None and i_cost is not None) else None)  # J = F - I
    pct = (i_cost / etc) if (i_cost is not None and etc) else None                      # K = I / F
    put("percent_complete", pct)
    earned = (tcp * pct) if (tcp is not None and pct is not None) else None             # L = C * K
    put("revenues_earned_to_date", earned)
    put("profit_earned_to_date", (gp * pct) if (gp is not None and pct is not None) else None)  # M = G * K
    if earned is not None and billed is not None:
        put("overbillings", max(billed - earned, 0.0))                                 # O = MAX(N-L,0)
        put("underbillings", max(earned - billed, 0.0))                                # P = MAX(L-N,0)
    if tcp is not None and billed is not None:
        put("left_to_bill", tcp - billed)                                              # R = C - N
    if gp is not None and earned is not None:
        pe = gp * pct if pct is not None else None
        put("future_profit_to_earn", (gp - pe) if pe is not None else None)            # S = G - M
    jtc, ltb = snap.get("cost_to_complete"), snap.get("left_to_bill")
    if isinstance(jtc, (int, float)) and isinstance(ltb, (int, float)):
        put("pure_job_borrow", max(jtc - ltb, 0.0))                                    # T = MAX(J-R,0)


def load(excel_path: Path, db_path: Path, dry_run: bool, show: int):
    if not excel_path.exists():
        sys.exit(f"ERROR: WIP master not found: {excel_path}\n"
                 f"Set WIP_EXCEL_PATH or pass --excel.")
    print(f"Reading (read-only): {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    projects, snapshots = {}, []
    per_div = {}
    report_date = None
    for plan in TAB_PLAN:
        if plan["tab"] not in wb.sheetnames:
            print(f"  ! tab not found, skipping: {plan['tab']}")
            continue
        ws = wb[plan["tab"]]
        n = 0
        for proj, snap in read_tab(ws, plan):
            projects[proj["project_no"]] = proj
            snapshots.append(snap)
            report_date = report_date or snap["report_date"]
            n += 1
        per_div[plan["division"]] = per_div.get(plan["division"], 0) + n
        print(f"  {plan['tab']:<12} -> {plan['division']:<13} {n:>4} projects")
    wb.close()

    print(f"\nReport date: {report_date}")
    print("By division: " + ", ".join(f"{d} {c}" for d, c in per_div.items()))
    print(f"Total projects: {len(projects)}   snapshots: {len(snapshots)}")

    if dry_run:
        print("\n--dry-run: nothing written.")
        _print_sample_from_memory(projects, snapshots, show)
        return

    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    con.execute("PRAGMA foreign_keys = ON;")
    con.executescript(SCHEMA_SQL.read_text(encoding="utf-8"))
    now = dt.datetime.now().isoformat(timespec="seconds")

    for p in projects.values():
        con.execute(
            """INSERT INTO project
                 (project_no, division, is_ftw, name, type, builder_or_gc,
                  bonded, rp_category, updated_at)
               VALUES (:project_no, :division, :is_ftw, :name, :type,
                       :builder_or_gc, :bonded, :rp_category, :updated_at)
               ON CONFLICT(project_no) DO UPDATE SET
                 division=excluded.division, is_ftw=excluded.is_ftw,
                 name=excluded.name, type=excluded.type,
                 builder_or_gc=excluded.builder_or_gc, bonded=excluded.bonded,
                 rp_category=excluded.rp_category, updated_at=excluded.updated_at""",
            {**p, "updated_at": now},
        )

    cols = ["project_no", "report_date"] + SNAPSHOT_FIELDS + ["source_tab", "loaded_at"]
    placeholders = ", ".join(f":{c}" for c in cols)
    updates = ", ".join(f"{c}=excluded.{c}" for c in cols
                        if c not in ("project_no", "report_date"))
    for s in snapshots:
        row = {c: s.get(c) for c in cols}
        row["loaded_at"] = now
        con.execute(
            f"INSERT INTO wip_snapshot ({', '.join(cols)}) VALUES ({placeholders}) "
            f"ON CONFLICT(project_no, report_date) DO UPDATE SET {updates}",
            row,
        )

    con.commit()
    print(f"\nWrote {len(projects)} projects + {len(snapshots)} snapshots -> {db_path}")
    _print_sample_from_db(con, show)
    con.close()


def _fmt(v):
    if isinstance(v, float):
        return f"{v:,.0f}"
    return "" if v is None else str(v)


def _print_sample_from_memory(projects, snapshots, show):
    if show <= 0:
        return
    print(f"\nSample (first {show}):")
    for s in snapshots[:show]:
        p = projects[s["project_no"]]
        print(f"  {s['project_no']:<12} {p['division']:<12} "
              f"contract {_fmt(s.get('total_contract_price')):>12} "
              f"costs {_fmt(s.get('costs_to_date')):>12} "
              f"billed {_fmt(s.get('billed_to_date')):>12}  {p['name']}")


def _print_sample_from_db(con, show):
    if show <= 0:
        return
    print(f"\nv_wip_latest (top {show} by contract):")
    q = ("SELECT project_no, division, total_contract_price, costs_to_date, "
         "billed_to_date, project_name FROM v_wip_latest "
         "ORDER BY total_contract_price DESC NULLS LAST LIMIT ?")
    try:
        rows = con.execute(q, (show,)).fetchall()
    except sqlite3.OperationalError:
        rows = con.execute(q.replace(" NULLS LAST", ""), (show,)).fetchall()
    for r in rows:
        print(f"  {r[0]:<12} {r[1]:<12} contract {_fmt(r[2]):>12} "
              f"costs {_fmt(r[3]):>12} billed {_fmt(r[4]):>12}  {r[5]}")


def main():
    ap = argparse.ArgumentParser(description="Load the final WIP master sheet into the project ledger.")
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="WIP master workbook (read-only).")
    ap.add_argument("--db", type=Path, default=DEFAULT_DB, help="SQLite ledger file to write.")
    ap.add_argument("--dry-run", action="store_true", help="Parse and report; write nothing.")
    ap.add_argument("--show", type=int, default=0, help="Print N sample rows after loading.")
    args = ap.parse_args()
    load(args.excel, args.db, args.dry_run, args.show)


if __name__ == "__main__":
    main()
