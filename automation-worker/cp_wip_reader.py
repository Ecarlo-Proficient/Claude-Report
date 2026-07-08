#!/usr/bin/env python3
"""
cp_wip_reader.py — CP division WIP row builder.

Reads Commercial takeoffs from Synology, joins each project with its QBO
Billed/Costs, and writes results to the "Test - CP" tab of the SharePoint
WIP file. Live tabs (WIP - CP, WIP Master) are never touched — the guard
in wip_excel_guard.py enforces this at the code level.

Per-project extraction rules (locked 2026-06-30):
  Contract Price = cell immediately right of "GRAND TOTAL" on sheet
                   "Commercial Proposal"
  ETC            = cell AP1961 on sheet "Bid"
  Change Orders  = flag-only in v1 (CO sheets detected but not summed;
                   Change Orders/ sub-folder → flagged and skipped)

Failure modes never crash the run — they surface as a Status column value
so Ted can triage from the Test - CP tab.

Data sources:
  - Synology (READ-ONLY):
      Active:    /Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects
      Completed: same path + /Completed Projects
  - QBO (READ-ONLY GET): via project-pnl/project_pnl_export.py reused
    helpers (load_credentials, build_project_customer_map, fetch_project_pl,
    extract_pl_totals).
  - OneDrive local mirror of SharePoint (WRITE, Test - CP only):
      ~/Library/CloudStorage/OneDrive-ProficientConcrete,LLC/
        Company Files - WIP Report/WIP - MASTER new.xlsx

Usage:
  python3 cp_wip_reader.py --dry-run                     # preview, no write
  python3 cp_wip_reader.py --project CP672               # one project only
  python3 cp_wip_reader.py                                # live run, all CP
  python3 cp_wip_reader.py --no-qbo                       # skip QBO join
                                                          #   (fast local test)

Overridable via env or flags:
  CP_ACTIVE_DIR, CP_COMPLETED_DIR, WIP_EXCEL_PATH
"""
from __future__ import annotations

import argparse
import datetime as dt
import importlib.util
import logging
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

# Silence known-benign noise BEFORE importing openpyxl / requests.
# 1. openpyxl warns on cross-sheet INDIRECT() print areas (harmless — we're
#    only reading data, print settings don't matter).
# 2. macOS system Python 3.9 ships with LibreSSL; urllib3 v2 prefers OpenSSL.
#    HTTPS still works fine over LibreSSL; the warning is cosmetic.
warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name.*")
warnings.filterwarnings("ignore", message=".*LibreSSL.*")
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.NotOpenSSLWarning)  # type: ignore[attr-defined]
except Exception:
    pass

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from wip_excel_guard import (
    ALLOWED_WRITE_SHEETS,
    WipWriteDenied,
    assert_write_allowed,
    open_wip_workbook_for_write,
)


# ─────────────────────── pretty terminal output ───────────────────
class _Term:
    """Small ANSI helper — degrades gracefully when stdout isn't a TTY."""
    ENABLED = sys.stdout.isatty()

    RESET = "\033[0m"     if ENABLED else ""
    BOLD  = "\033[1m"     if ENABLED else ""
    DIM   = "\033[2m"     if ENABLED else ""
    CYAN  = "\033[36m"    if ENABLED else ""
    GREEN = "\033[32m"    if ENABLED else ""
    AMBER = "\033[33m"    if ENABLED else ""
    RED   = "\033[31m"    if ENABLED else ""
    GRAY  = "\033[90m"    if ENABLED else ""

    @staticmethod
    def color(code: str, text: str) -> str:
        return f"{code}{text}{_Term.RESET}"


_ANSI_RE = re.compile(r"\033\[[0-9;]*m")

def _vlen(s: str) -> int:
    """Visible length (strips ANSI escape codes so padding math works)."""
    return len(_ANSI_RE.sub("", s))

def _rpad(s: str, width: int) -> str:
    """Right-pad to `width` visible chars (for left-aligned columns)."""
    return s + " " * max(0, width - _vlen(s))

def _lpad(s: str, width: int) -> str:
    """Left-pad to `width` visible chars (for right-aligned columns)."""
    return " " * max(0, width - _vlen(s)) + s


def _section(title: str) -> None:
    """Print a section header."""
    print()
    print(_Term.color(_Term.BOLD + _Term.CYAN, f"▸ {title}"))
    print(_Term.color(_Term.DIM, "─" * (len(title) + 2)))


def _kv(label: str, value, indent: int = 2) -> None:
    """Print a key/value pair aligned."""
    print(f"{' ' * indent}{label:<18} {value}")


def _fmt_money(v: Optional[float]) -> str:
    """Money as a plain (uncolored) string. Coloring happens at pad time."""
    if v is None:
        return "—"
    return f"${v:,.2f}"


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "—"
    return f"{v*100:.1f}%"


def _dim_if_dash(s: str) -> str:
    """Dim the em-dash placeholder so real numbers pop; leave money plain."""
    return _Term.color(_Term.DIM, s) if s == "—" else s


# ─────────────────────── config / paths ────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent

CP_ACTIVE_DIR = Path(os.getenv(
    "CP_ACTIVE_DIR",
    "/Volumes/Common/CURRENT PROJECTS/Awarded Projects Commercial projects",
))
CP_COMPLETED_DIR = Path(os.getenv(
    "CP_COMPLETED_DIR",
    str(CP_ACTIVE_DIR / "Completed Projects"),
))
WIP_EXCEL_PATH = Path(os.getenv(
    "WIP_EXCEL_PATH",
    str(Path.home()
        / "Library/CloudStorage/OneDrive-ProficientConcrete,LLC"
        / "Company Files - WIP Report/WIP - MASTER new.xlsx"),
))

TEST_TAB = "Test - CP"  # write target — must be in ALLOWED_WRITE_SHEETS

# Cell anchors on the takeoff (locked)
PROPOSAL_SHEET = "Commercial Proposal"
GRAND_TOTAL_LABEL = "GRAND TOTAL"
BID_SHEET = "Bid"
ETC_CELL = "AP1961"

# CO sheet detection (case-insensitive)
_CO_SHEET_RE = re.compile(r"^(CO|CHANGE\s*ORDER)[\s#-]*\d+$", re.IGNORECASE)
_CO_FOLDER = "Change Orders"

# Project # from folder name — e.g. "CP672 - FIRESTONE RED OAK" → "CP672"
_CP_FOLDER_RE = re.compile(r"^(CP\d{3,4})\b", re.IGNORECASE)


log = logging.getLogger("cp_wip_reader")


# ─────────────────────── project-pnl reuse ─────────────────────────
def _load_project_pnl_module():
    """Load project-pnl/project_pnl_export.py despite its hyphenated folder
    name (not a valid Python package). Uses importlib to bypass sys.path
    identifier restrictions."""
    path = PROJECT_ROOT / "project-pnl" / "project_pnl_export.py"
    if not path.exists():
        raise FileNotFoundError(
            f"project-pnl module not at {path}. "
            f"CP reader depends on project_pnl_export.py for QBO helpers."
        )
    spec = importlib.util.spec_from_file_location("project_pnl_export", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ─────────────────────── data models ───────────────────────────────
@dataclass
class CpRow:
    """One row per CP project, ready to write to Test - CP."""
    project_num: str
    project_name: str                # from folder name after " - "
    is_completed: bool               # from Completed Projects/ subfolder
    base_contract: Optional[float]   # from Commercial Proposal Grand Total (audit)
    co_revenue: Optional[float]      # sum of Change Orders/ xlsx TOTAL: cells
    base_etc: Optional[float]        # from Bid!AP1961 (audit — pre-CO)
    billed_to_date: Optional[float]  # QBO P&L income = GROSS billed (incl retainage)
    costs_to_date: Optional[float]   # from QBO P&L COGS + Expenses
    retainage_held: Optional[float] = None  # gross billed − net collectible (retainage receivable)
    status_flags: List[str] = field(default_factory=list)  # multiple can stack
    takeoff_path: Optional[Path] = None   # audit trail: first included takeoff (hyperlink anchor)
    co_folder_path: Optional[Path] = None # audit trail: Change Orders/ folder
    co_details: List[str] = field(default_factory=list)  # per-CO audit list
    included_takeoffs: List[Path] = field(default_factory=list)  # takeoff(s) summed into this row
    folder_path: Optional[Path] = None   # explicit folder for the project-name link
                                          # (RP sets this; CP falls back to takeoff_path.parent)

    @property
    def contract_price(self) -> Optional[float]:
        """Total Contract Price = base contract + summed COs. Returns None when
        the base is undetermined (e.g. Missing Grand Total / contract not yet
        decided) — a CO-only total would be misleading, so the cell stays blank
        and the flag explains why."""
        if self.base_contract is None:
            return None
        return self.base_contract + (self.co_revenue or 0.0)

    @property
    def co_cost_estimate(self) -> Optional[float]:
        """CO Cost is None until estimators add a real cost cell to the CO
        template. No proxy — Ted's rule (2026-07-01): "no false numbers,
        don't populate if there is no source. If no CO costs, don't put
        it, just flag it."

        Property name kept as `co_cost_estimate` for column-mapping
        stability; semantics are now "CO cost sourced from template" and
        the value is None until sourcing exists.

        This will start returning a real value when parse_change_orders_folder
        gains a cost-cell extractor (blocked on template change)."""
        return None

    @property
    def etc(self) -> Optional[float]:
        """Revised ETC.
        - No COs → equals Base ETC (revised == base is truthful here).
        - COs present but no CO Cost data → defaults to Base ETC and the
          row is flagged provisional (Ted 2026-07-02). This WIP is a
          MONITORING view, fully script-driven from takeoffs — no manual
          entry. Defaulting to Original ETC keeps the row alive and
          surfaces the over-budget signal (Costs > Original ETC) instead
          of going dark. The bounded error (excludes the CO's cost) is
          disclosed by the `% based on Original ETC — excludes CO cost`
          flag. We still never FABRICATE a CO cost — the base ETC is a
          real number, just scoped pre-CO. See [[project-cp-wip-takeoff-extraction]]."""
        if self.base_etc is None:
            return None
        if not self.co_revenue:
            return self.base_etc                # no COs → Revised = Base
        if self.co_cost_estimate is None:
            return self.base_etc                # CO Rev but no CO Cost → provisional (Original ETC)
        return self.base_etc + self.co_cost_estimate

    @property
    def has_missing_co_cost(self) -> bool:
        """True when the project has CO Revenue but no CO Cost data — a
        signal that Revised ETC (and the % / Earned / Over-Under derived
        from it) is PROVISIONAL: computed off Original ETC, which excludes
        the CO's cost. Drives the provisional flag."""
        return bool(self.co_revenue) and self.co_cost_estimate is None

    @property
    def status(self) -> str:
        if not self.status_flags:
            return "OK"
        return "; ".join(self.status_flags)


# ─────────────────────── takeoff parsing ───────────────────────────
def _find_label_cells(ws, label: str) -> List[tuple]:
    """Return ALL (row, col) matches for the given label (case-insensitive,
    trailing colon/whitespace stripped).

    Returns a list because CO templates have TWO 'TOTAL:' cells: a header
    label at the top of the description table AND the summary total at
    the bottom. Caller must try each until one yields a numeric value to
    the right (see _read_number_to_right)."""
    label_norm = label.strip().upper().rstrip(":").strip()
    results: List[tuple] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip().upper().rstrip(":").strip()
            if s == label_norm:
                results.append((cell.row, cell.column))
    return results


def _find_number_after_label(ws_data, ws_formula, label: str) -> Optional[float]:
    """Find `label` on the sheet (case-insensitive), then scan rightward
    from each match for a numeric value. Handles both:
      - Multiple label occurrences (header + summary): tries each until
        one yields a number.
      - Merged label cells + separate currency-symbol cells: the right-
        scan skips empty/currency cells (see _read_number_to_right).
    Returns None if no match yields a number."""
    for (r, c) in _find_label_cells(ws_data, label):
        v = _read_number_to_right(ws_data, ws_formula, r, c)
        if v is not None:
            return v
    return None


def _resolve_sheet_name(wb, target: str) -> Optional[str]:
    """Case-insensitive, whitespace-tolerant sheet name lookup. Returns the
    ACTUAL sheet name as it appears in the workbook, or None if not found.
    Handles: 'BID' vs 'Bid' vs 'bid ' vs '  Commercial Proposal '."""
    target_norm = target.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target_norm:
            return name
    return None


def _read_number_to_right(ws_data, ws_formula, row: int, start_col: int,
                          max_scan: int = 12) -> Optional[float]:
    """Scan rightward from `start_col` looking for the first numeric value.

    Skips empty cells (None), cells that are part of a merged region but
    aren't the top-left anchor, and cells that hold just a currency
    symbol like '$'. This handles two real-world CP takeoff patterns:

      1. TOTAL: label lives in a merged cell — the value sits past the
         merge's right edge, not one column right of the label anchor.
      2. Currency symbol is a separate cell from the number:
             [ ...merged label... ] [ $ ] [ 3,840.00 ]

    Returns the first cell that resolves to a number via
    `_read_number_smart` (which itself handles formula fallback). None
    if nothing found within `max_scan` columns."""
    for offset in range(1, max_scan + 1):
        col = start_col + offset
        try:
            ref = f"{get_column_letter(col)}{row}"
        except ValueError:
            return None
        # Try the smart reader first — handles cached values + formula refs.
        v = _read_number_smart(ws_data, ws_formula, ref)
        if v is not None:
            return v
        # If not numeric, check if the raw string is a currency marker; if
        # so, keep scanning. Any other non-empty non-numeric value → treat
        # as a stop signal (we've passed the value region).
        raw = ws_data[ref].value
        if raw is None or raw == "":
            continue
        s = str(raw).strip()
        if s in ("$", "USD", "€", "£", "-", "—"):
            continue
        # Non-numeric text (like a next-row label) — bail.
        return None
    return None


def _coerce_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ─────────────────────── formula-safe cell reader ────────────────
# openpyxl(data_only=True) returns cached formula results. If the takeoff was
# saved without recalc (common when files are edited via Google Sheets export,
# LibreOffice, or a Python tool), the cache is empty and formula cells come
# back as None. In that case we open a SECOND view of the workbook (data_only
# =False) to see the formula string, then evaluate it dynamically against the
# cached view — so we NEVER hardcode the specific cell-relationship (the
# estimators can change AP1961's formula without breaking us).
_CELL_REF_RE = re.compile(r"^[A-Z]+\d+$")
_TOKEN_RE    = re.compile(r"[A-Z]+\d+|[+\-]")


def _read_number_smart(ws_data, ws_formula, cell_ref: str, depth: int = 0) -> Optional[float]:
    """Read a numeric value from `cell_ref`. Order of attempts:
       1. Cached value from data_only workbook.
       2. If the cell holds a formula, evaluate simple +/- expressions of
          cell references (with recursive fallback for each ref).

    Recursion depth is bounded to prevent runaway on circular refs."""
    if depth > 4:
        return None
    try:
        cached = ws_data[cell_ref].value
    except (KeyError, ValueError):
        return None
    if cached is not None:
        coerced = _coerce_float(cached)
        if coerced is not None:
            return coerced

    # Cached is missing / non-numeric — try the formula view.
    try:
        raw = ws_formula[cell_ref].value
    except (KeyError, ValueError):
        return None
    if not isinstance(raw, str) or not raw.strip().startswith("="):
        return None

    expr = raw.strip().lstrip("=").replace(" ", "").upper()
    # Support only simple +/- chains of cell refs (e.g. "AM1948+AN1948+AO1948").
    # Anything more complex (SUM, ranges, other sheets) falls back to None +
    # the caller will flag the row for manual review.
    if not re.fullmatch(r"[A-Z]+\d+([+\-][A-Z]+\d+)*", expr):
        return None
    tokens = _TOKEN_RE.findall(expr)
    if not tokens:
        return None

    total = 0.0
    sign = 1.0
    expect_ref = True
    for tok in tokens:
        if tok in ("+", "-"):
            if expect_ref:
                return None  # malformed
            sign = 1.0 if tok == "+" else -1.0
            expect_ref = True
        else:
            if not expect_ref or not _CELL_REF_RE.match(tok):
                return None
            v = _read_number_smart(ws_data, ws_formula, tok, depth + 1)
            if v is None:
                return None
            total += sign * v
            expect_ref = False
            sign = 1.0
    return total


_WIP_TAG_RE = re.compile(r"\bwip\b", re.IGNORECASE)
# Auxiliary xlsx that live alongside takeoffs in a project folder and are NOT
# takeoffs (used only as a fallback when no file is named 'takeoff').
_AUX_XLSX_RE = re.compile(r"cost\s*code|explanation", re.IGNORECASE)


def _find_contract_total(ws_data, ws_formula):
    """Read the contract total off a proposal sheet. Templates are inconsistent
    (Ted 2026-07-02) — the overall total is labeled 'GRAND TOTAL', 'SUB TOTAL',
    or just 'TOTAL'. Match is EXACT (normalized, colon-stripped) so 'TOTAL SQFT',
    'TOTAL YARDS', etc. are NOT caught — only cells that are exactly one of the
    three dollar-total labels. Collect every value from all three and take the
    **largest** = the overall contract total (an overall is ≥ any section
    subtotal). Returns (value_or_None, label_used_or_None)."""
    best = None
    best_label = None
    for label in ("GRAND TOTAL", "SUB TOTAL", "TOTAL"):
        for (r, c) in _find_label_cells(ws_data, label):
            v = _read_number_to_right(ws_data, ws_formula, r, c)
            if v is not None and (best is None or v > best):
                best = v
                best_label = label
    return best, best_label


def _find_co_total(ws_data, ws_formula):
    """Read a Change Order's total off a CO sheet/file. Templates label it
    'CHANGE ORDER TOTAL' or plain 'TOTAL' (Ted 2026-07-02) — prefer
    'CHANGE ORDER TOTAL' (avoids grabbing a line-item 'TOTAL:' column header),
    fall back to 'TOTAL'. Returns the value or None."""
    for label in ("CHANGE ORDER TOTAL", "TOTAL"):
        v = _find_number_after_label(ws_data, ws_formula, label)
        if v is not None:
            return v
    return None


def _select_proposal_sheet(wb):
    """Pick which proposal tab to read the contract GRAND TOTAL from, when a
    takeoff has multiple proposal sheets (Ted 2026-07-02):
      - a tab with 'final' in its name → the final proposal, use it;
      - else if exactly ONE proposal tab → use it;
      - else (multiple proposals, none marked FINAL) → don't guess, flag it.
    'Proposal' tabs = any sheet whose name contains 'proposal' (Commercial /
    Residential / Alternative / '11.14.25 Proposal' …). Non-proposal tabs
    (Bid, JMP Subcontract, Change Order#N, Cost Codes) are ignored here.
    Returns (sheet_name_or_None, flag_or_None)."""
    proposals = [s for s in wb.sheetnames if "proposal" in s.lower()]
    if not proposals:
        return None, "Missing Proposal Sheet"
    finals = [s for s in proposals if "final" in s.lower()]
    if len(finals) == 1:
        return finals[0], None
    if len(finals) > 1:
        return finals[0], f"Multiple FINAL proposals — used '{finals[0]}'"
    if len(proposals) == 1:
        return proposals[0], None
    return None, (f"Multiple proposals ({len(proposals)}), none marked FINAL — "
                  f"mark the final one: {', '.join(proposals[:4])}"
                  f"{'...' if len(proposals) > 4 else ''}")


def _parse_one_takeoff(tk: Path):
    """Read Contract Price (final proposal Grand Total) + ETC (Bid!AP1961)
    + any in-takeoff Change Order sheets from ONE takeoff file.
    Returns (contract, etc, co_total, flags). Never raises."""
    flags: List[str] = []
    # Two views: cached values (data_only) fast path + formulas (fallback for
    # cells saved without recalc). Random-access needed → not read_only.
    try:
        wb_data    = load_workbook(tk, data_only=True)
        wb_formula = load_workbook(tk, data_only=False)
    except Exception as e:
        return None, None, None, [f"Takeoff Read Failed: {type(e).__name__}"]

    contract = etc = None
    co_total = None
    try:
        prop_sheet, prop_flag = _select_proposal_sheet(wb_data)
        if prop_flag:
            flags.append(prop_flag)
        if prop_sheet is not None:
            cp, label_used = _find_contract_total(wb_data[prop_sheet],
                                                  wb_formula[prop_sheet])
            if cp is None:
                has_label = (_find_label_cells(wb_data[prop_sheet], "GRAND TOTAL")
                             or _find_label_cells(wb_data[prop_sheet], "SUB TOTAL"))
                flags.append("Bad Contract Price" if has_label
                             else "Missing Grand/Sub Total")
            else:
                contract = cp

        bid_sheet = _resolve_sheet_name(wb_data, BID_SHEET)
        if not bid_sheet:
            flags.append("Missing Bid Sheet")
        else:
            e = _read_number_smart(wb_data[bid_sheet], wb_formula[bid_sheet], ETC_CELL)
            if e is None:
                flags.append(f"Bad ETC ({ETC_CELL})")
            else:
                etc = e

        # Change Order sheets INSIDE the takeoff (Scenario A) — parse + SUM
        # each sheet's 'TOTAL:' cell (same pattern as standalone CO files).
        co_sheets = [s for s in wb_data.sheetnames if _CO_SHEET_RE.match(s.strip())]
        if co_sheets:
            parsed = []
            running = 0.0
            for cs in co_sheets:
                amt = _find_co_total(wb_data[cs], wb_formula[cs])
                if amt is not None:
                    running += amt
                    parsed.append(f"{cs} ${amt:,.2f}")
                else:
                    flags.append(f"CO sheet '{cs}': no TOTAL cell found — manual review")
            if parsed:
                co_total = running
                flags.append(f"CO sheets in takeoff summed ({len(parsed)}): "
                             f"{', '.join(parsed)}")
    finally:
        wb_data.close()
        wb_formula.close()
    return contract, etc, co_total, flags


def parse_takeoff(folder: Path, row: CpRow) -> None:
    """Extract Contract Price + ETC into `row`, identifying takeoff files by
    'takeoff' in the filename (Ted 2026-07-02):
      - Project folders also hold AUXILIARY xlsx (Cost Codes.xlsx, Explanation
        OH.xlsx, etc.). Only files whose name contains 'takeoff' are treated as
        takeoffs; the rest are ignored so they don't trip the multi-file logic.
      - ONE takeoff → use it (no tag needed).
      - MULTIPLE takeoffs → include ONLY the 'WIP'-tagged one(s), and SUM them
        (a project can have multiple scopes, e.g. FDT + PAVING, both tagged).
      - Multiple takeoffs, none tagged 'WIP' → flag + leave blank (never guess).
    Appends to row.status_flags on any failure; never raises."""
    xlsx_files = sorted([p for p in folder.iterdir()
                         if p.suffix.lower() in (".xlsx", ".xlsm")
                         and not p.name.startswith("~$")])
    if len(xlsx_files) == 0:
        row.status_flags.append("No Takeoff")
        return

    # Takeoffs are named with 'takeoff'; everything else (Cost Codes,
    # Explanation OH, …) is auxiliary and ignored.
    takeoffs = [p for p in xlsx_files if "takeoff" in p.name.lower()]
    if not takeoffs:
        # No file named 'takeoff' — fall back to the non-auxiliary xlsx.
        non_aux = [p for p in xlsx_files if not _AUX_XLSX_RE.search(p.name)]
        if len(non_aux) == 1:
            takeoffs = non_aux
        else:
            row.status_flags.append(
                f"No takeoff file identified ({len(xlsx_files)} xlsx, none named "
                f"'takeoff') — rename the takeoff to include 'takeoff'")
            return

    if len(takeoffs) == 1:
        included = takeoffs                          # single takeoff — tag not required
    else:
        included = [p for p in takeoffs if _WIP_TAG_RE.search(p.name)]
        if not included:
            row.status_flags.append(
                f"Multiple takeoffs ({len(takeoffs)}) — none tagged 'WIP'; "
                f"estimator must tag the one(s) to include")
            return

    row.included_takeoffs = included
    row.takeoff_path = included[0]                  # hyperlink anchor

    contract_total = 0.0
    etc_total = 0.0
    co_total = 0.0
    got_contract = False
    got_etc = False
    got_co = False
    multi = len(included) > 1
    for tk in included:
        c, e, co, fflags = _parse_one_takeoff(tk)
        prefix = f"{tk.name}: " if multi else ""
        for f in fflags:
            row.status_flags.append(prefix + f)
        if c is not None:
            contract_total += c
            got_contract = True
        if e is not None:
            etc_total += e
            got_etc = True
        if co is not None:
            co_total += co
            got_co = True

    row.base_contract = contract_total if got_contract else None
    row.base_etc = etc_total if got_etc else None
    # In-takeoff CO sheets (Scenario A) → add to co_revenue (Scenario B folder
    # COs, if any, are added on top by parse_change_orders_folder).
    if got_co:
        row.co_revenue = (row.co_revenue or 0.0) + co_total

    if multi:
        row.status_flags.append(
            f"WIP takeoffs summed ({len(included)}): "
            f"{', '.join(p.name for p in included)}")


# ─────────────────────── Scenario B: sum CO xlsx files ────────────
def parse_change_orders_folder(co_folder: Path, row: CpRow) -> None:
    """Sum all standalone CO xlsx files in the Change Orders/ sub-folder.

    Each CO file has a plain 'TOTAL:' label with the CO amount immediately
    one cell to the right (sample locked 2026-07-01 from CP672).

    Populates row.co_revenue with the sum, row.co_details with a per-CO
    audit list, and flags any per-file issues (missing TOTAL, unreadable
    file, etc.) without aborting the sum."""
    row.co_folder_path = co_folder

    co_files = sorted([
        p for p in co_folder.iterdir()
        if p.is_file()
        and p.suffix.lower() in (".xlsx", ".xlsm")
        and not p.name.startswith("~$")   # Excel lockfiles
    ])
    if not co_files:
        row.status_flags.append("CO Folder Empty")
        return

    total = 0.0
    per_file_issues: List[str] = []
    for co_file in co_files:
        try:
            wb_d = load_workbook(co_file, data_only=True)
            wb_f = load_workbook(co_file, data_only=False)
        except Exception as e:
            per_file_issues.append(f"{co_file.name}: read failed ({type(e).__name__})")
            continue

        try:
            # CO templates typically have TWO "TOTAL:" cells: one header
            # label above the description table, and the summary at the
            # bottom right. Also, the file may have multiple sheets (the
            # actual CO sheet + a "BidScreen XL Drawing Data" metadata
            # sheet). Scan every sheet, try every TOTAL: match, take the
            # first that yields a number to the right.
            amt = None
            found_sheet = None
            found_any_label = False
            for sheet_name in wb_d.sheetnames:
                ws_d = wb_d[sheet_name]
                ws_f = wb_f[sheet_name]
                if (_find_label_cells(ws_d, "CHANGE ORDER TOTAL")
                        or _find_label_cells(ws_d, "TOTAL")):
                    found_any_label = True
                v = _find_co_total(ws_d, ws_f)   # CHANGE ORDER TOTAL, else TOTAL
                if v is not None:
                    amt = v
                    found_sheet = sheet_name
                    break

            if amt is None:
                if found_any_label:
                    per_file_issues.append(f"{co_file.name}: TOTAL label found but no numeric value to the right")
                else:
                    per_file_issues.append(f"{co_file.name}: no TOTAL cell")
                continue

            total += amt
            row.co_details.append(f"{co_file.name}: ${amt:,.2f}")
        finally:
            wb_d.close()
            wb_f.close()

    # Additive: preserve any in-takeoff CO sheets (Scenario A) already summed
    # into co_revenue by parse_takeoff.
    if row.co_details:
        row.co_revenue = (row.co_revenue or 0.0) + total

    n_ok = len(row.co_details)
    if n_ok:
        row.status_flags.append(
            f"CO Summed ({n_ok} file(s), ${total:,.2f})"
        )
        # Ted 2026-07-02 (monitoring view): don't guess CO cost, but don't
        # blank the row either. CO Cost stays empty; Revised ETC defaults to
        # Original ETC so % / Earned / Over-Under still compute — flagged
        # provisional so the CO-cost exclusion is disclosed. Real fix is
        # upstream (estimators add a cost line to the CO template).
        row.status_flags.append(
            "% based on Original ETC — excludes CO cost (provisional; add CO cost line to template)"
        )
    if per_file_issues:
        row.status_flags.append(
            f"CO Read Issues: {'; '.join(per_file_issues[:3])}"
            + ("..." if len(per_file_issues) > 3 else "")
        )


# ─────────────────────── folder scan ───────────────────────────────
def _project_num_from_folder(folder: Path) -> Optional[str]:
    m = _CP_FOLDER_RE.match(folder.name)
    return m.group(1).upper() if m else None


def _project_name_from_folder(folder: Path) -> str:
    """'CP672 - FIRESTONE RED OAK' → 'FIRESTONE RED OAK'"""
    if " - " in folder.name:
        return folder.name.split(" - ", 1)[1].strip()
    return folder.name


def scan_cp_folders(root: Path, is_completed: bool) -> List[CpRow]:
    """Iterate CP#### folders under root. Returns one CpRow per project.
    Skips non-CP folders silently (folder naming discipline is on the
    estimators). Handles Change Orders/ sub-folder = flag-and-skip."""
    rows: List[CpRow] = []
    if not root.exists():
        log.warning("CP root does not exist: %s (Synology unmounted?)", root)
        return rows

    for entry in sorted(root.iterdir()):
        if not entry.is_dir():
            continue
        if entry.name == "Completed Projects":  # handled in a separate pass
            continue

        proj_num = _project_num_from_folder(entry)
        if not proj_num:
            continue  # not a CP folder — skip silently

        row = CpRow(
            project_num=proj_num,
            project_name=_project_name_from_folder(entry),
            is_completed=is_completed,
            base_contract=None,
            co_revenue=None,
            base_etc=None,
            billed_to_date=None,
            costs_to_date=None,
        )

        # Read the base takeoff (Contract Price + ETC).
        parse_takeoff(entry, row)

        # Scenario B: Change Orders/ sub-folder → sum standalone CO xlsx
        # files and add to Contract Price (via the co_revenue property).
        # ETC still comes from the main takeoff only — current CO template
        # has no cost cell, so ETC growth from COs is deferred (see
        # [[project-cp-wip-takeoff-extraction]] v2 options).
        co_folder = entry / _CO_FOLDER
        if co_folder.is_dir():
            parse_change_orders_folder(co_folder, row)

        rows.append(row)

    return rows


# ─────────────────────── QBO join ──────────────────────────────────
# Retainage detection — case-insensitive match on "Retainage Not Billed"
# anywhere in the invoice's PrivateNote (memo field). Pattern harvested
# from legacy wip/wip_sync.py; see [[reference-wip-qbo-aggregation-patterns]].
_RETAINAGE_NOT_BILLED_RE = re.compile(r"retainage\s+not\s+billed", re.IGNORECASE)


def _linked_txn_types(inv: dict) -> List[str]:
    """Normalized TxnType strings of every transaction linked to `inv`
    (payments, journal entries, credits) from QBO's LinkedTxn array.
    Spaces stripped + lowercased so 'JournalEntry' / 'Journal Entry' both
    normalize to 'journalentry'."""
    out = []
    for lt in (inv.get("LinkedTxn") or []):
        t = str(lt.get("TxnType", "")).replace(" ", "").lower()
        if t:
            out.append(t)
    return out


def _cleared_by_journal_entry(inv: dict) -> bool:
    """True if a Journal Entry is linked to this invoice — the reclass that
    moves the balance into the Retainage Receivable account (i.e. it was NOT
    settled by a real customer cash payment)."""
    return "journalentry" in _linked_txn_types(inv)


def _is_retainage_receivable_invoice(inv: dict) -> bool:
    """True if this is a standalone retainage-receivable invoice → its amount
    is retainage HELD (not collectible cash). It is excluded from Billed's
    net-collectible component and surfaces in the RETAINAGE HELD column.

    Keyed on the 'Retainage Not Billed' memo — Ted's explicit tag for
    retainage moved to (or awaiting) the Retainage Receivable account.

    CORRECTED 2026-07-02 (live run): an earlier version ALSO required a
    Journal Entry on the invoice's LinkedTxn. That left retainage held
    understated ($2,179.50 instead of $31,094.80 on CP672) because the JE
    that reclasses AR → Retainage Receivable is applied to the invoice via a
    PAYMENT (the "Payment on 12/31/25" that carries the JE credit) — so the
    JE never appears directly in Invoice.LinkedTxn and the check missed it.
    The memo is the reliable signal and correctly handles BOTH methods:
      * standalone 'Retainage Not Billed' invoice → matched here → held.
      * 2026 draw-with-retainage-LINE → memo is 'Draw #N' (not matched); its
        withheld portion is captured via gross income − net TotalAmt.
    (`_cleared_by_journal_entry` is retained for audit logging only.)"""
    memo = inv.get("PrivateNote") or ""
    return bool(_RETAINAGE_NOT_BILLED_RE.search(memo))


def _fetch_billed_ex_retainage(pnl, access: str, company_id: str,
                               customer_id: str) -> float:
    """Sum TotalAmt of all invoices for a customer, EXCLUDING standalone
    retainage-receivable invoices (memo 'Retainage Not Billed'). This is the
    NET-COLLECTIBLE billed; the caller derives Retainage Held = gross − net.
    Returns 0.0 if the query fails or there are no invoices.

    Every retainage-memo invoice is logged with its LinkedTxn types + whether
    a JE is directly attached, so the retainage reclass path stays auditable."""
    try:
        invoices = pnl.fetch_customer_invoices(access, company_id, customer_id)
    except Exception as e:
        log.warning("Invoice fetch failed for customer %s: %s — falling back to 0",
                    customer_id, e)
        return 0.0

    total = 0.0
    skipped = 0
    for inv in invoices:
        if _is_retainage_receivable_invoice(inv):
            skipped += 1
            doc = inv.get("DocNumber", "?")
            amt = inv.get("TotalAmt", "?")
            types = _linked_txn_types(inv) or ["<none>"]
            je = "direct" if _cleared_by_journal_entry(inv) else "via-payment/none"
            log.info("  EXCL retainage-receivable inv #%s ($%s) — memo-tagged "
                     "(LinkedTxn=%s, JE=%s) → counts as Retainage Held",
                     doc, amt, ",".join(types), je)
            continue
        total += float(inv.get("TotalAmt", 0) or 0)

    if skipped:
        log.info("Customer %s: %d retainage-receivable invoice(s) → Retainage Held",
                 customer_id, skipped)
    return total


def enrich_with_qbo(rows: List[CpRow]) -> None:
    """Fetch QBO Billed/Costs per project and populate rows in-place.
    All-time window (start_date = 2019-01-01, end_date = today) — CP is
    slow-turn commercial work, worth the extra API cost. Never raises."""
    pnl = _load_project_pnl_module()
    try:
        access, company_id = pnl.load_credentials()
    except SystemExit:
        log.error("QBO auth failed — leaving Billed/Costs blank on all rows")
        for r in rows:
            r.status_flags.append("QBO Auth Failed")
        return
    except Exception as e:
        log.error("QBO auth error: %s", e)
        for r in rows:
            r.status_flags.append(f"QBO Auth Error: {type(e).__name__}")
        return

    try:
        proj_map = pnl.build_project_customer_map(access, company_id)
    except Exception as e:
        log.error("QBO customer map fetch failed: %s", e)
        for r in rows:
            r.status_flags.append(f"QBO Customer Map Failed: {type(e).__name__}")
        return

    start_date = "2019-01-01"
    end_date = dt.date.today().isoformat()

    for row in rows:
        cust = proj_map.get(row.project_num)
        if not cust:
            row.status_flags.append("QBO Not Found")
            continue

        try:
            report_data = pnl.fetch_project_pl(
                access, company_id, cust["id"], start_date, end_date
            )
            totals = pnl.extract_pl_totals(report_data)
            # Billed to Date = GROSS billed, INCLUDING retainage — this is
            # the standard WIP basis (Marcum/CFMA: job-to-date billing is
            # "the total requisitioned by the customer, including retainage
            # held"). Verified 2026-07-02. QBO P&L Total Income = gross
            # billed (retainage invoices post to income; negative retainage
            # lines post to Retainage Receivable on the balance sheet, so
            # income is not reduced). We ALSO compute the net-collectible
            # billed (retainage-receivable invoices excluded, via the
            # JE-clearing check) and derive Retainage Held = gross − net.
            gross_billed = float(totals.get("income", 0.0) or 0.0)
            net_collectible = _fetch_billed_ex_retainage(
                pnl, access, company_id, cust["id"]
            )
            row.billed_to_date = gross_billed
            row.retainage_held = max(gross_billed - net_collectible, 0.0)
            log.info("  %s billed(gross)=%.2f net-collectible=%.2f retainage-held=%.2f",
                     row.project_num, gross_billed, net_collectible, row.retainage_held)
            # Costs = COGS + Expenses. QBO Projects UI sums both; per GAAP
            # job costing, direct project spending is a project cost
            # regardless of which account category it lands in. Coding
            # errors in QBO (bills posted to Expenses instead of COGS)
            # should NOT hide costs from the WIP — the WIP should surface
            # true total spend so the coding error becomes visible.
            row.costs_to_date = (
                (totals.get("cogs", 0.0) or 0.0)
                + (totals.get("expenses", 0.0) or 0.0)
            )
            # Loud over-budget signal (Ted 2026-07-02: "need signal loud
            # and clear"). Costs exceeding the ETC shown = job is over its
            # estimate → surface as an explicit flag, not just an implied
            # >100% in the % column. % is left UNCAPPED on purpose so the
            # overage stays visible. On provisional rows (CO w/o cost) the
            # ETC is Original ETC, so this can fire on CO spend too — that's
            # acceptable; the provisional flag sits right beside it.
            if (row.etc not in (None, 0)
                    and row.costs_to_date is not None
                    and row.costs_to_date > row.etc):
                row.status_flags.append("⚠ OVER BUDGET — Costs exceed ETC")
        except Exception as e:
            row.status_flags.append(f"QBO P&L Failed: {type(e).__name__}")


# ─────────────────────── Excel write (Test - CP only) ──────────────
# Clean layout matching the team's WIP: light-gray bold header, thin grid
# borders on every cell, light-yellow ONLY on sourced inputs, white calcs,
# no green rows (Ted 2026-07-02 "make it clean and easy to read").
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")        # light gray header
HDR_FONT = Font(bold=True, color="000000", size=10)
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")       # flag cell
FLAG_FONT = Font(italic=True, color="7F6000", size=9)
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")      # light yellow — SOURCED inputs
LINK_FONT = Font(color="0563C1", underline="single")     # Excel link (project name only)
_SIDE = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
CURRENCY_FMT = '"$"#,##0.00_);[Red]("$"#,##0.00)'

# Which fields render as currency vs percent (drives number_format in the write
# loop). Everything money-valued is currency; the two ratios are percent.
_MONEY_FIELDS = frozenset({
    "base_contract", "co_revenue", "contract_price", "base_etc",
    "co_cost_estimate", "etc", "original_profit", "costs_to_date",
    "cost_to_complete", "_earned_revenue", "profit_earned", "future_profit",
    "billed_to_date", "retainage_held", "left_to_bill", "overbillings",
    "underbillings", "job_borrow",
})
_PCT_FIELDS = frozenset({"_pct_complete", "gross_profit_pct"})

# SOURCED cells (raw numbers straight from the takeoff or QBO) get a yellow
# fill so they're visually distinct from the white CALCULATED cells (Excel
# formulas). Ted 2026-07-02: "yellow = metrics that have sources, calculations
# leave white." Identifier/metadata columns (project #, name, status, flags,
# last synced) stay white — they're not metrics.
_SOURCE_FIELDS = frozenset({
    "contract_price",    # TOTAL CONTRACT PRICE (Original + COs)
    "etc",               # ESTIMATED TOTAL COSTS (Revised ETC)
    "billed_to_date",    # QBO (gross income)
    "costs_to_date",     # QBO
    # trailing cross-check columns (also sourced → yellow)
    "co_revenue",        # APPROVED COs
    "retainage_held",    # QBO (gross − net collectible)
})


def _sheet_fragment(sheet_name: str, cell_ref: str = "A1") -> str:
    """Build the '#SheetName!CellRef' fragment appended to a file:// URI so
    Excel opens the workbook AND jumps to the specified cell.

    Sheet names with spaces or special chars need single-quote wrapping in
    the URL fragment (Excel convention). Cell ref defaults to A1."""
    return f"#'{sheet_name}'!{cell_ref}"


def _apply_hyperlink(cell, target: Optional[Path], fragment: str = "") -> None:
    """Attach a file:// hyperlink to a cell and apply link styling.

    `target` may be a Path to a file OR a directory. If None, this is a
    no-op so callers can safely pass None when the source is missing.
    `fragment` is optional (e.g. \"#'Bid'!AP1961\" to jump to a cell).

    **Never attaches a hyperlink to a cell whose value is None.** openpyxl's
    behavior in that case is to display the hyperlink URL as the cell text,
    which produces confusing "why is there a URL in this blank cell?"
    output. If the cell is empty, there's nothing to click anyway."""
    if target is None:
        return
    if cell.value is None or cell.value == "":
        return
    try:
        uri = target.as_uri() + fragment
    except (ValueError, OSError):
        # Path can't be converted to URI (unusual on macOS/Linux) — skip
        # rather than crash the whole run.
        return
    cell.hyperlink = uri
    cell.font = LINK_FONT

# Full standard construction-WIP column set. Formulas verified 2026-07-02
# against Marcum LLP / Construction Executive, Wouch Maloney CPAs, CFMA/
# AICPA-referenced guidance. Billed to Date is GROSS (incl retainage); a
# separate RETAINAGE HELD column keeps the cash-collection timeline visible.
# Column order mirrors the team's WIP schedule — reads left→right as a story:
# the deal (contract, cost) → progress (billed, spent) → projection (cost to
# complete, profit) → recognition (% , earned) → billing position (over/under,
# left to bill) → profitability (GP%, future profit, job borrow). The four
# yellow inputs (contract, cost, billed, costs) + retainage lead; everything
# after is derived. Original/CO breakdown lives in the FLAGS ("CO Summed …"),
# not as separate columns (kept clean per Ted 2026-07-02).
COLS = [
    # ── Identifiers ──
    ("PROJECT #",                                       12, "project_num"),
    ("PROJECT NAME",                                    30, "project_name"),
    ("STATUS",                                          10, "_active_status"),
    # ── Inputs (yellow) — the 4 core WIP inputs, matching the team sheet ──
    ("TOTAL CONTRACT PRICE",                            16, "contract_price"),   # Original + COs
    ("ESTIMATED TOTAL COSTS",                           16, "etc"),              # Revised ETC
    ("BILLED TO DATE",                                  15, "billed_to_date"),   # gross (incl retainage)
    ("COSTS TO DATE",                                   15, "costs_to_date"),
    # ── Derived story (white) ──
    ("COST TO COMPLETE",                                14, "cost_to_complete"),
    ("ORIGINAL PROFIT",                                 14, "original_profit"),
    ("PERCENT COMPLETE",                                11, "_pct_complete"),
    ("REVENUES EARNED TO DATE",                         16, "_earned_revenue"),
    ("PROFIT EARNED TO DATE",                           15, "profit_earned"),
    ("BILLINGS IN EXCESS OF EARNED REV. (OVERBILLINGS)", 20, "overbillings"),
    ("EARN. REV. IN EXCESS OF BILLINGS (UNDERBILLINGS)", 20, "underbillings"),
    ("LEFT TO BILL",                                    14, "left_to_bill"),
    ("GROSS PROFIT %",                                  12, "gross_profit_pct"),
    ("FUTURE PROFIT TO EARN",                           15, "future_profit"),
    ("PURE JOB BORROW",                                 14, "job_borrow"),
    # ── Trailing cross-check / reference (yellow, sourced) ──
    ("APPROVED COs",                                    14, "co_revenue"),
    ("RETAINAGE HELD",                                  14, "retainage_held"),
    ("LAST SYNCED",                                     18, "_last_synced"),
    ("FLAGS",                                           50, "status"),   # after Sync (Ted 2026-07-02)
]


# Fields written as Excel FORMULAS (not static values) so any input change
# auto-recalculates and the formula is visible in the formula bar for audit.
# Mirrors the project-pnl workbook's formula-driven design.
FORMULA_FIELDS: frozenset = frozenset({
    # Total Contract Price (contract_price) + Estimated Total Costs (etc) are
    # now written as VALUES (the yellow inputs, = Original+COs / Revised ETC via
    # the CpRow properties); the derived columns below reference those cells.
    "original_profit",   # ORIGINAL PROFIT   = Total Contract − Estimated Costs
    "gross_profit_pct",  # GROSS PROFIT %    = Original Profit / Revised Contract
    "cost_to_complete",  # COST TO COMPLETE  = Revised ETC − Costs to Date
    "_pct_complete",     # % COMPLETE        = Costs / Revised ETC
    "_earned_revenue",   # REVENUES EARNED   = Revised Contract × (Costs / Revised ETC)
    "profit_earned",     # PROFIT EARNED     = Earned Revenue − Costs to Date
    "future_profit",     # FUTURE PROFIT     = Original Profit − Profit Earned
    "left_to_bill",      # LEFT TO BILL      = Revised Contract − Billed
    "overbillings",      # OVERBILLINGS      = MAX(Billed − Earned, 0)
    "underbillings",     # UNDERBILLINGS     = MAX(Earned − Billed, 0)
    "job_borrow",        # PURE JOB BORROW   = MAX(Cost to Complete − Left to Bill, 0)
})


def _build_formula(field_name: str, row_num: int,
                   col_letter_by_field: Dict[str, str]) -> str:
    """Build the Excel formula string for a derived-field cell in row `row_num`.

    Every reference is resolved by FIELD NAME → column letter (dynamic), so
    the formulas survive any column reordering. Blank inputs propagate to a
    blank output (never a false zero or #DIV/0), matching the Python
    None-cascades."""
    def ref(f: str) -> str:
        return col_letter_by_field[f] + str(row_num)

    F  = ref("contract_price")     # Total Contract Price (value)
    I  = ref("etc")                # Estimated Total Costs (value)
    J  = ref("billed_to_date")     # Billed (gross)
    K  = ref("costs_to_date")      # Costs to Date
    OP = ref("original_profit")    # Original Profit
    CTC = ref("cost_to_complete")  # Cost to Complete
    ER = ref("_earned_revenue")    # Revenues Earned to Date
    PE = ref("profit_earned")      # Profit Earned to Date
    LTB = ref("left_to_bill")      # Left to Bill

    if field_name == "original_profit":
        return f'=IF(OR({F}="",{I}=""),"",{F}-{I})'
    if field_name == "gross_profit_pct":
        return f'=IF(OR({OP}="",{F}="",{F}=0),"",{OP}/{F})'
    if field_name == "cost_to_complete":
        return f'=IF(OR({I}="",{K}=""),"",{I}-{K})'
    if field_name == "_pct_complete":
        return f'=IF(OR({K}="",{I}="",{I}=0),"",{K}/{I})'
    if field_name == "_earned_revenue":
        return f'=IF(OR({F}="",{K}="",{I}="",{I}=0),"",{F}*{K}/{I})'
    if field_name == "profit_earned":
        return f'=IF(OR({ER}="",{K}=""),"",{ER}-{K})'
    if field_name == "future_profit":
        return f'=IF(OR({OP}="",{PE}=""),"",{OP}-{PE})'
    if field_name == "left_to_bill":
        return f'=IF(OR({F}="",{J}=""),"",{F}-{J})'
    if field_name == "overbillings":
        # BIE — billings in excess of earned revenue (liability). 0 if underbilled.
        return f'=IF(OR({J}="",{ER}=""),"",MAX({J}-{ER},0))'
    if field_name == "underbillings":
        # CIE — earned revenue in excess of billings (asset). 0 if overbilled.
        return f'=IF(OR({ER}="",{J}=""),"",MAX({ER}-{J},0))'
    if field_name == "job_borrow":
        # Pure Job Borrow — remaining costs in excess of remaining billing
        # capacity (Wouch Maloney CPAs): costs-to-complete beyond what's left
        # to bill = cash this job must borrow from others. 0 if none.
        return f'=IF(OR({CTC}="",{LTB}=""),"",MAX({CTC}-{LTB},0))'
    raise ValueError(f"No formula defined for field {field_name!r}")


def _row_display_value(row: CpRow, field_name: str, sync_ts: str):
    """Return the value to write into the cell for a NON-formula `field_name`
    (formula fields are written via _build_formula, never here)."""
    if field_name == "_active_status":
        return "Closed" if row.is_completed else "Active"
    if field_name == "_last_synced":
        return sync_ts
    return getattr(row, field_name, None)


def write_test_cp(rows: List[CpRow], wip_path: Path, dry_run: bool = False,
                  tab_name: str = TEST_TAB) -> bool:
    """Write rows to the given WIP tab (default 'Test - CP'; RP passes
    'Test - RP'). Same structure/formatting for every division. Guarded by
    wip_excel_guard. Returns True if written, False if skipped (dry-run, or the
    file is open in Excel)."""
    assert_write_allowed(tab_name)  # tripwire before we even open the workbook

    if dry_run:
        _print_rows_table(rows, wip_path, tab_name)
        return False

    if not wip_path.exists():
        raise FileNotFoundError(
            f"WIP Excel not found at {wip_path}. "
            f"Check WIP_EXCEL_PATH env or verify OneDrive is synced."
        )

    # Never clobber a workbook that's open in Excel. Excel drops a
    # ~$<name> owner-lock file next to an open workbook (project-pnl
    # safe_save pattern, Ted 2026-06-23). If present, skip with a clear
    # message rather than writing underneath the open file — a
    # last-writer-wins save would silently lose the sync or the user's edits.
    lock = wip_path.with_name("~$" + wip_path.name)
    if lock.exists():
        print(_Term.color(_Term.AMBER,
              f"  ⚠ {wip_path.name} looks OPEN in Excel — skipped the write to "
              f"avoid overwriting it. Close the file and re-run."))
        return False

    wb = open_wip_workbook_for_write(wip_path)
    try:
        if tab_name not in wb.sheetnames:
            wb.create_sheet(tab_name)
        ws = wb[tab_name]
        assert_write_allowed(ws.title)  # belt + suspenders

        # Full replace: clear existing rows AND explicitly wipe every cell
        # attribute (value, hyperlink, number_format, fill, font) up to the
        # prior extent so leftovers can't leak into new rows.
        # openpyxl's delete_rows sometimes leaves formatting or hyperlinks
        # behind on cells that were populated before — belt-and-suspenders
        # clear here to guarantee a clean slate.
        prior_max_row = ws.max_row or 0
        prior_max_col = ws.max_column or 0
        for r in range(1, max(prior_max_row, len(rows) + 1) + 1):
            for c in range(1, max(prior_max_col, len(COLS)) + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.hyperlink = None
                cell.number_format = "General"
        if prior_max_row > 0:
            ws.delete_rows(1, prior_max_row)

        # Header row — gray, bold, centered + wrapped, bordered.
        for c, (label, width, _key) in enumerate(COLS, start=1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = CELL_BORDER
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        sync_ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

        # Column indices for hyperlink attachment (1-based like openpyxl)
        # + column LETTERS for building cross-cell formulas.
        col_idx = {field: i + 1 for i, (_, _, field) in enumerate(COLS)}
        col_letter_by_field = {
            field: get_column_letter(i + 1)
            for i, (_, _, field) in enumerate(COLS)
        }

        for i, row in enumerate(rows, start=2):
            # Invariant guard: if CO Cost is populated but CO Revenue is
            # not, refuse to write the CO Cost. That's either a bug or a
            # corrupted state, and quietly writing a made-up cost would
            # be worse than surfacing the issue.
            if row.co_revenue is None and row.co_cost_estimate is not None:
                log.warning(
                    "%s: CO Cost populated ($%s) with no CO Revenue — "
                    "refusing to write CO Cost. Investigate parse logic.",
                    row.project_num, row.co_cost_estimate
                )
                row.status_flags.append("Data integrity: CO Cost without CO Rev — dropped")

            for c, (_label, _width, field_name) in enumerate(COLS, start=1):
                if field_name in FORMULA_FIELDS:
                    # Derived cell — write an Excel formula referencing
                    # the input cells in the same row. Excel evaluates on
                    # open/edit so any input change auto-recalculates.
                    val = _build_formula(field_name, i, col_letter_by_field)
                else:
                    val = _row_display_value(row, field_name, sync_ts)
                cell = ws.cell(row=i, column=c, value=val)
                cell.border = CELL_BORDER
                if field_name in _MONEY_FIELDS:
                    cell.number_format = CURRENCY_FMT
                elif field_name in _PCT_FIELDS:
                    cell.number_format = "0.0%"
                # Yellow = sourced input (raw from takeoff / QBO); white = calc.
                if field_name in _SOURCE_FIELDS:
                    cell.fill = INPUT_FILL
                elif field_name == "status" and row.status_flags:
                    cell.fill = FLAG_FILL
                    cell.font = FLAG_FONT

            # PROJECT NAME → project (Awarded Project) folder. Ted 2026-07-02
            # wants this kept even though on Windows/OneDrive Excel rewrites the
            # macOS file:// link into a (broken) SharePoint URL — it works on his
            # Mac and he needs the trace point. (Number-cell links stay off.)
            # Link target: explicit folder_path (RP) else the takeoff's parent (CP).
            link_folder = (row.folder_path if row.folder_path is not None
                           else (row.takeoff_path.parent if row.takeoff_path else None))
            if link_folder is not None:
                _apply_hyperlink(ws.cell(row=i, column=col_idx["project_name"]),
                                 link_folder)

        # Wrap the range in an Excel Table — gives filter/sort dropdowns and a
        # structured, clean look. Explicit gray header + borders above override
        # the table style, so it stays clean (no row stripes).
        last_row = len(rows) + 1
        last_col = get_column_letter(len(COLS))
        for tname in list(ws.tables):
            del ws.tables[tname]          # drop any prior run's table first
        tbl_name = re.sub(r"[^A-Za-z0-9]", "", tab_name) or "WIP"  # unique per tab
        table = Table(displayName=tbl_name, ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        ws.add_table(table)

        # Atomic write — save to a temp file then os.replace() so a crash
        # or interruption can't leave a half-written WIP (safe_save pattern).
        tmp = wip_path.with_name(wip_path.name + ".tmp")
        wb.save(str(tmp))
        try:
            os.replace(str(tmp), str(wip_path))
        except OSError as e:
            # Keep the good data in the .tmp rather than losing it.
            log.error("Could not replace %s (%s); new data saved to %s",
                      wip_path.name, e, tmp.name)
            raise
        log.info("Wrote %d rows to %s in %s", len(rows), tab_name, wip_path)
    finally:
        wb.close()
    return True


# ─────────────────────── pretty run report ────────────────────────
def _shorten(path: Path, max_len: int = 70) -> str:
    """Shorten long paths by inserting an ellipsis in the middle."""
    s = str(path)
    if len(s) <= max_len:
        return s
    head = s[: max_len // 2 - 2]
    tail = s[-(max_len // 2 - 2):]
    return f"{head}…{tail}"


def _wip_metrics(r: CpRow) -> Dict[str, Optional[float]]:
    """Compute the derived WIP numbers in Python for TERMINAL DISPLAY only —
    mirrors the Excel formulas exactly (same blank-propagation), so the
    dry-run preview matches what the workbook will show after recalc."""
    F = r.contract_price          # Revised Contract
    I = r.etc                     # Revised ETC
    K = r.costs_to_date
    J = r.billed_to_date          # Billed (gross)

    def sub(a, b):
        return (a - b) if (a is not None and b is not None) else None

    orig_profit = sub(F, I)
    gp_pct = (orig_profit / F) if (orig_profit is not None and F not in (None, 0)) else None
    ctc = sub(I, K)
    pct = (K / I) if (K is not None and I not in (None, 0)) else None
    earned = (F * K / I) if (F is not None and K is not None and I not in (None, 0)) else None
    profit_earned = sub(earned, K)
    future_profit = sub(orig_profit, profit_earned)
    left_to_bill = sub(F, J)
    overbill = max(J - earned, 0.0) if (J is not None and earned is not None) else None
    underbill = max(earned - J, 0.0) if (earned is not None and J is not None) else None
    job_borrow = max(ctc - left_to_bill, 0.0) if (ctc is not None and left_to_bill is not None) else None
    return {
        "orig_profit": orig_profit, "gp_pct": gp_pct, "ctc": ctc, "pct": pct,
        "earned": earned, "profit_earned": profit_earned,
        "future_profit": future_profit, "left_to_bill": left_to_bill,
        "overbill": overbill, "underbill": underbill, "job_borrow": job_borrow,
    }


def _print_rows_table(rows: List[CpRow], wip_path: Path, tab_name: str = TEST_TAB) -> None:
    """Dry-run preview. Vertical per-project detail for small runs (the
    single-project verify case), compact table for bulk runs. Values match
    the Excel formulas via _wip_metrics."""
    _section(f"DRY RUN — would write {len(rows)} row(s) to {tab_name!r}")
    print(f"  target: {_Term.color(_Term.DIM, _shorten(wip_path))}")

    clean = sum(1 for r in rows if not r.status_flags)
    flagged = len(rows) - clean

    if len(rows) <= 3:
        # Full vertical WIP card per project — every column, easy to verify.
        for r in rows:
            m = _wip_metrics(r)
            status = "Closed" if r.is_completed else "Active"
            print()
            print(_Term.color(_Term.BOLD + _Term.CYAN,
                  f"  ▌ {r.project_num}  {r.project_name}") +
                  _Term.color(_Term.DIM, f"   [{status}]"))
            rows_out = [
                ("Original Contract",   _fmt_money(r.base_contract)),
                ("Approved COs",        _fmt_money(r.co_revenue)),
                ("Revised Contract",    _fmt_money(r.contract_price)),
                ("Original ETC",        _fmt_money(r.base_etc)),
                ("CO Cost",             _fmt_money(r.co_cost_estimate)),
                ("Revised ETC",         _fmt_money(r.etc)),
                ("Original Profit",     _fmt_money(m["orig_profit"])),
                ("Gross Profit %",      _fmt_pct(m["gp_pct"])),
                ("Costs to Date",       _fmt_money(r.costs_to_date)),
                ("Cost to Complete",    _fmt_money(m["ctc"])),
                ("% Complete",          _fmt_pct(m["pct"])),
                ("Revenues Earned",     _fmt_money(m["earned"])),
                ("Profit Earned",       _fmt_money(m["profit_earned"])),
                ("Future Profit",       _fmt_money(m["future_profit"])),
                ("Billed (gross)",      _fmt_money(r.billed_to_date)),
                ("Retainage Held",      _fmt_money(r.retainage_held)),
                ("Left to Bill",        _fmt_money(m["left_to_bill"])),
                ("Overbillings",        _fmt_money(m["overbill"])),
                ("Underbillings",       _fmt_money(m["underbill"])),
                ("Pure Job Borrow",     _fmt_money(m["job_borrow"])),
            ]
            for label, val in rows_out:
                print(f"    {label:<20} {_lpad(_dim_if_dash(val), 16)}")
            if r.status_flags:
                print(_Term.color(_Term.AMBER, "    ⚑ " + "; ".join(r.status_flags)))
            else:
                print(_Term.color(_Term.GREEN, "    ✓ clean"))
    else:
        # Compact table — the headline columns for scanning many jobs.
        W_P, W_N, W_M, W_PC = 8, 24, 13, 6
        SEP = _Term.color(_Term.DIM, " │ ")
        hdr = [
            _rpad("PROJECT", W_P), _rpad("NAME", W_N),
            _lpad("CONTRACT", W_M), _lpad("CO $", W_M), _lpad("ETC", W_M),
            _lpad("COSTS", W_M), _lpad("%", W_PC), _lpad("BILLED", W_M),
            _lpad("RETAIN", W_M), _lpad("OVER", W_M), _lpad("UNDER", W_M),
            _lpad("BORROW", W_M), "FLAGS",
        ]
        print()
        print(_Term.color(_Term.BOLD, "  " + SEP.join(hdr)))
        print(_Term.color(_Term.DIM, "  " + "─" * 150))
        for r in rows:
            m = _wip_metrics(r)
            flag = (_Term.color(_Term.AMBER, "⚑ " + "; ".join(r.status_flags))
                    if r.status_flags else _Term.color(_Term.GREEN, "✓"))
            name = _rpad(r.project_name[:W_N], W_N)
            if r.is_completed:
                name = _Term.color(_Term.GREEN, name)
            cells = [
                _rpad(r.project_num, W_P), name,
                _lpad(_dim_if_dash(_fmt_money(r.contract_price)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.co_revenue)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.etc)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.costs_to_date)), W_M),
                _lpad(_dim_if_dash(_fmt_pct(m["pct"])), W_PC),
                _lpad(_dim_if_dash(_fmt_money(r.billed_to_date)), W_M),
                _lpad(_dim_if_dash(_fmt_money(r.retainage_held)), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["overbill"])), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["underbill"])), W_M),
                _lpad(_dim_if_dash(_fmt_money(m["job_borrow"])), W_M),
                flag,
            ]
            print("  " + SEP.join(cells))

    print()
    print(f"  {_Term.color(_Term.BOLD, 'Summary')}:  "
          f"{len(rows)} total  ·  "
          f"{_Term.color(_Term.GREEN, str(clean) + ' clean')}  ·  "
          f"{_Term.color(_Term.AMBER, str(flagged) + ' flagged')}")

    # Audit trail — file paths the reader touched. Shown when small enough
    # to be useful (≤ 5 projects), so single-project runs always get it.
    if len(rows) <= 5:
        print()
        print(_Term.color(_Term.DIM, "  Audit trail:"))
        for r in rows:
            if len(r.included_takeoffs) > 1:
                print(_Term.color(_Term.DIM, f"    {r.project_num}  takeoffs (summed, {len(r.included_takeoffs)}):"))
                for tk in r.included_takeoffs:
                    print(_Term.color(_Term.DIM, f"    {r.project_num}    · {tk.name}"))
            elif r.takeoff_path:
                print(_Term.color(_Term.DIM, f"    {r.project_num}  takeoff:      {r.takeoff_path}"))
            if r.co_folder_path:
                print(_Term.color(_Term.DIM, f"    {r.project_num}  CO folder:    {r.co_folder_path}"))
            for detail in r.co_details:
                print(_Term.color(_Term.DIM, f"    {r.project_num}    · {detail}"))


# ─────────────────────── orchestration ─────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse takeoffs + fetch QBO but don't write to Excel.")
    ap.add_argument("--project",
                    help="Filter to one project # (e.g. CP672). Case-insensitive.")
    ap.add_argument("--no-qbo", action="store_true",
                    help="Skip QBO join (fast local test of takeoff parsing).")
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="Show debug logs (timestamps + module names).")
    args = ap.parse_args()

    # Interactive mode uses pretty terminal output (no log prefixes).
    # --verbose reverts to the standard timestamped log format for debugging.
    # Route logging to stdout (not stderr) so pretty print order is preserved
    # when stdout/stderr are merged.
    if args.verbose:
        logging.basicConfig(
            level=logging.DEBUG,
            format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
            stream=sys.stdout,
        )
    else:
        logging.basicConfig(
            level=logging.WARNING,
            format="  %(levelname)s: %(message)s",
            stream=sys.stdout,
        )

    # ── Header ──
    print()
    print(_Term.color(_Term.BOLD + _Term.CYAN, "  CP WIP Reader"))
    print(_Term.color(_Term.DIM, "  " + dt.datetime.now().strftime("%Y-%m-%d %H:%M")))

    _section("Configuration")
    _kv("Active dir",    _shorten(CP_ACTIVE_DIR))
    _kv("Completed dir", _shorten(CP_COMPLETED_DIR))
    _kv("WIP target",    _shorten(WIP_EXCEL_PATH))
    _kv("Write target",  f"{TEST_TAB!r} tab (guard allow-list: "
                         f"{', '.join(sorted(ALLOWED_WRITE_SHEETS))})")

    # ── Scan ──
    _section("Scanning CP folders on Synology")
    active_rows = scan_cp_folders(CP_ACTIVE_DIR, is_completed=False)
    completed_rows = scan_cp_folders(CP_COMPLETED_DIR, is_completed=True)
    rows = active_rows + completed_rows
    _kv("Active",    f"{len(active_rows)} folder(s)")
    _kv("Completed", f"{len(completed_rows)} folder(s)")
    _kv("Total",     f"{len(rows)} folder(s) scanned")

    if args.project:
        pf = args.project.upper()
        rows = [r for r in rows if r.project_num == pf]
        _kv(f"--project {pf}", f"{len(rows)} row(s) after filter")

    if not rows:
        print()
        print(_Term.color(_Term.AMBER, "  No CP projects to process — exiting"))
        return 0

    # ── QBO enrichment ──
    if not args.no_qbo:
        _section("Enriching with QBO Billed/Costs")
        print(f"  {_Term.color(_Term.DIM, 'Fetching...')}")
        t0 = dt.datetime.now()
        enrich_with_qbo(rows)
        elapsed = (dt.datetime.now() - t0).total_seconds()
        print(_Term.color(_Term.GREEN, f"  ✓ {len(rows)} project(s) enriched") +
              _Term.color(_Term.DIM, f" ({elapsed:.1f}s)"))
    else:
        _section("Skipping QBO join (--no-qbo)")

    # ── Write / dry-run report ──
    try:
        wrote = write_test_cp(rows, WIP_EXCEL_PATH, dry_run=args.dry_run)
    except WipWriteDenied as e:
        print(_Term.color(_Term.RED, f"  ✗ Guard blocked write: {e}"))
        return 2
    except FileNotFoundError as e:
        print(_Term.color(_Term.RED, f"  ✗ {e}"))
        return 3

    if not args.dry_run and wrote:
        _section("Done")
        print(_Term.color(_Term.GREEN, f"  ✓ Wrote {len(rows)} row(s) to {TEST_TAB!r}"))
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
