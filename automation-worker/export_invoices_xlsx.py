"""
Export Notion Invoice Trackers to a single Excel file.

One-direction sync: Notion → Excel. The file is a read-only artifact for
Eduardo's ad-hoc summing (select rows in Excel → see Sum in status bar) and
copy-paste-to-email workflow. Eduardo NEVER edits the Excel file — every
edit-worthy field (Quick Status, Next Follow-Up, Assignee, etc.) lives in
Notion. This file is regenerated every sync run, so any edits made directly
would be clobbered.

Why this exists:
  Notion can't sum arbitrary row selection (only filtered totals). Eduardo's
  workflow includes ad-hoc questions like "what do customer X's last 3
  invoices total?" — for that, Excel is the right tool. Notion is still the
  source of truth and the working surface for all human notes / status.

Run via run_invoice_sync.py after the main sync completes. Errors here
don't affect the QBO→Notion sync (caught and logged separately).
"""
from __future__ import annotations

import datetime as dt
import logging
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from notion_client import NotionClient

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import paths


log = logging.getLogger("automation_worker.export_xlsx")

# Default file location — overrideable via INVOICE_EXPORT_PATH (env or
# machine.env) so you can point this at Synology, OneDrive, Desktop, etc.
# without code change. Fallback = repo root (same location as before).
DEFAULT_EXPORT_PATH = paths.get_path(
    "INVOICE_EXPORT_PATH",
    Path(__file__).resolve().parent.parent / "Open_Invoices.xlsx",
)


# Column order matches Eduardo's preferred layout (he sorts/scans left-to-right
# in this order: Division/Project # first for quick triage, then customer/invoice
# detail, then $ columns, then status). PM, Quick Status, and QBO Link live only
# in Notion — this Excel is a read-only summing sheet for the collections workflow.
COLUMNS = [
    ("Division", 10),
    ("Project #", 12),
    ("Client", 28),
    ("Date", 12),
    ("Invoice #", 12),
    ("Net Terms", 14),
    ("Due Date", 12),
    ("Past Due", 10),
    ("Memo", 40),
    ("Total Amount", 14),
    ("Open Balance", 14),
    ("Status", 14),
    ("Aging Bucket", 12),
]


def _text(prop: dict) -> str:
    """Pull plain text from a Notion rich_text or title property."""
    if not prop:
        return ""
    arr = prop.get("rich_text") or prop.get("title") or []
    return "".join(t.get("plain_text", "") for t in arr).strip()


def _select_name(prop: dict) -> str:
    sel = (prop or {}).get("select") or {}
    return sel.get("name", "")


def _multi_select_names(prop: dict) -> str:
    items = (prop or {}).get("multi_select") or []
    return ", ".join(i.get("name", "") for i in items)


def _people_names(prop: dict) -> str:
    """Notion people property → comma-separated names."""
    items = (prop or {}).get("people") or []
    return ", ".join(p.get("name", "") for p in items)


def _date_value(prop: dict) -> Optional[dt.date]:
    d = (prop or {}).get("date") or {}
    start = d.get("start")
    if not start:
        return None
    try:
        return dt.date.fromisoformat(start[:10])
    except ValueError:
        return None


def _number(prop: dict) -> Optional[float]:
    return (prop or {}).get("number")


def _url(prop: dict) -> str:
    return (prop or {}).get("url") or ""


def _relation_first_title(prop: dict, customer_title_cache: Dict[str, str]) -> str:
    """For a relation property, return the first related page's title via cache."""
    rels = (prop or {}).get("relation") or []
    if not rels:
        return ""
    page_id = rels[0].get("id")
    return customer_title_cache.get(page_id, "")


def _build_customer_title_cache(
    notion: NotionClient, customer_ds_id: str, title_prop: str = "Client"
) -> Dict[str, str]:
    """{page_id → display name} for all customers in a customer DB."""
    cache: Dict[str, str] = {}
    for page in notion.query_data_source(customer_ds_id, page_size=100):
        page_id = page.get("id")
        if not page_id:
            continue
        title_prop_data = (page.get("properties") or {}).get(title_prop) or {}
        title_arr = title_prop_data.get("title") or []
        name = "".join(t.get("plain_text", "") for t in title_arr).strip()
        if name:
            cache[page_id] = name
    return cache


def _row_for(
    page: dict,
    division_label: str,
    customer_title_cache: Dict[str, str],
    today: dt.date,
) -> List[Any]:
    """Build one Excel row from a Notion invoice page."""
    props = page.get("properties") or {}

    # Customer — prefer the relation's resolved name, fall back to Customer (raw).
    customer_relation_name = _relation_first_title(
        props.get("Customer"), customer_title_cache
    )
    customer = customer_relation_name or _text(props.get("Customer (raw)"))

    invoice_num = _text(props.get("Invoice #"))
    project_num = _text(props.get("Project #"))
    memo = _text(props.get("Memo"))

    invoice_date = _date_value(props.get("Date"))
    due_date = _date_value(props.get("Due Date"))
    # Past Due — signed days. Positive = actually past due. Negative = not yet due
    # (e.g., -2 = "due in 2 days"). Zero = due today. "" only when no due date set.
    # Number format on the cell (see below) adds the "+" prefix on positives.
    past_due = (today - due_date).days if due_date else ""

    total_amt = _number(props.get("Total Amount"))
    open_bal = _number(props.get("Open balance"))

    status = _select_name(props.get("Status"))
    aging = _select_name(props.get("Aging Bucket"))
    net_terms = _select_name(props.get("Net Terms"))
    division = _select_name(props.get("Division")) or division_label

    return [
        division,                              # Division
        project_num,                           # Project #
        customer,                              # Client
        invoice_date,                          # Date
        invoice_num,                           # Invoice #
        net_terms,                             # Net Terms
        due_date,                              # Due Date
        past_due,                              # Past Due
        memo,                                  # Memo
        total_amt,                             # Total Amount
        open_bal,                              # Open Balance
        status,                                # Status
        aging,                                 # Aging Bucket
    ]


def _open_invoices(notion: NotionClient, ds_id: str) -> List[dict]:
    """Pull non-Paid invoices from a tracker."""
    filter_body = {
        "or": [
            {"property": "Status", "select": {"equals": "Unpaid"}},
            {"property": "Status", "select": {"equals": "Partially Paid"}},
        ]
    }
    return list(notion.query_data_source(ds_id, filter_body=filter_body, page_size=100))


def export_open_invoices_xlsx(
    *,
    notion: NotionClient,
    res_com_ds_id: str,
    mfd_ds_id: str,
    customer_list_ds_id: str,
    mfd_client_list_ds_id: str,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Pull open invoices from both trackers, write a single Excel file
    formatted for Eduardo's collections workflow. Returns the output path.
    """
    target = output_path or Path(
        os.getenv("INVOICE_EXPORT_PATH", str(DEFAULT_EXPORT_PATH))
    )

    # Office lock check — Excel creates a hidden '~$Open_Invoices.xlsx' file in
    # the same folder while someone has the workbook open (Eduardo on his side,
    # Ted on his, doesn't matter — OneDrive syncs the lock to both machines).
    # If we overwrite while Excel is open, his AutoSave will silently write the
    # stale buffer back on top of our update, losing the sync. Skip cleanly and
    # tell the operator to re-run once it's closed.
    #
    # Edge case: if Excel crashed and left a stale '~$' file behind, manually
    # delete it from the OneDrive folder.
    lock_file = target.parent / f"~${target.name}"
    if lock_file.exists():
        log.warning(
            "Excel file is OPEN (%s present in OneDrive). "
            "Skipped Excel export this run — overwriting now would collide with "
            "the open Excel session and lose the update. "
            "Close the file and re-run `sync-ar` to refresh the OneDrive mirror.",
            lock_file.name,
        )
        return target

    log.info("Loading Notion customer caches for export…")
    res_com_titles = _build_customer_title_cache(notion, customer_list_ds_id)
    mfd_titles = _build_customer_title_cache(notion, mfd_client_list_ds_id)

    log.info("Pulling open invoices from Notion…")
    res_com_pages = _open_invoices(notion, res_com_ds_id)
    mfd_pages = _open_invoices(notion, mfd_ds_id)
    log.info(
        "Export: %d Res/Com + %d MFD open invoices",
        len(res_com_pages), len(mfd_pages),
    )

    today = dt.date.today()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Invoices"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        col = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    row_num = 2
    for page in res_com_pages:
        row = _row_for(page, "RP/CP", res_com_titles, today)
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1
    for page in mfd_pages:
        row = _row_for(page, "MFD", mfd_titles, today)
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

    # Number / currency / date formatting (column indices match COLUMNS order)
    for r in range(2, row_num):
        ws.cell(row=r, column=4).number_format = "mm/dd/yyyy"   # Date
        ws.cell(row=r, column=7).number_format = "mm/dd/yyyy"   # Due Date
        ws.cell(row=r, column=10).number_format = '"$"#,##0.00' # Total Amount
        ws.cell(row=r, column=11).number_format = '"$"#,##0.00' # Open Balance
        # Past Due (col 8) — "+12" for past due, "-2" for not yet due, "0" for due today.
        # Custom format: positive;negative;zero (Excel applies "+" prefix on positives).
        past_due_cell = ws.cell(row=r, column=8)
        past_due_cell.number_format = '"+"0;-0;0'
        # Bold red when > 30 days overdue
        if isinstance(past_due_cell.value, int) and past_due_cell.value > 30:
            past_due_cell.font = Font(bold=True, color="C00000")

    # Auto-filter on the header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row_num - 1}"

    # Write
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    log.info("Exported %d invoices → %s", row_num - 2, target)
    return target
