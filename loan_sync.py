#!/usr/bin/env python3
"""
loan_sync.py — One-click QBO → Equipment Debt Schedule sync.

WHAT IT DOES
  Reads loan-payment activity from QuickBooks (Proficient Concrete, LLC) and
  updates the automated equipment debt-schedule workbook:

    1. Current Balance (Master Debt Schedule col G) ← QBO liability account
       *actual* balance (CurrentBalance).  No principal/interest split is
       computed or required — whatever QBO says is owed is what we show.
       Only rows tagged with the active --company (default Proficient) are
       synced; L&A Holdings is a separate QBO file.
    2. Payment transactions hitting each loan's liability account are pulled
       and written to a "QBO Payment Ledger" sheet (append-only, idempotent).
    3. "As-of Date" (col F) stamped with the sync date.

  The amortization tabs and all their formulas are left untouched.  This script
  only writes value cells (E, F) on the Master and maintains two helper sheets
  ("QBO Setup", "QBO Payment Ledger").

MATCHING ANCHOR (set up once)
  Each loan is matched to QBO by its **liability Account Id** (stable, survives
  rename).  You confirm that mapping ONCE in the "QBO Setup" sheet — that is the
  "enter the terms at the beginning" step.  `--discover` pre-fills it with a
  best-guess match by lender/asset name so you only correct, never type IDs.

USAGE
  python3 loan_sync.py --discover     # 1st time: pull QBO liability accounts,
                                       #   write/refresh the QBO Setup mapping.
                                       #   Then open the workbook & confirm it.
  python3 loan_sync.py --dry-run      # safe preview: writes *_PREVIEW.xlsx,
                                       #   changes nothing live. Prints a diff.
  python3 loan_sync.py                 # the one click: backup → sync → save.
  python3 loan_sync.py --since 2025-01-01   # widen the transaction window.

  Auth = single Touch ID (reads the QBO blob from Keychain, same as the other
  scripts).  A valid refresh token must already be set up (setup_qbo.py).

SAFETY
  * The live workbook is ALWAYS copied to a timestamped backup before any write.
  * --dry-run never touches the live file.
  * Re-running is safe: payments already in the ledger (by QBO Txn Id + account)
    are never duplicated.

FIRST-RUN VERIFICATION (flagged in output — confirm once, see README_loan_sync):
  [V1] Liability CurrentBalance sign (we display abs()).
  [V2] Loan payments are Check/Expense (Purchase) vs Bill vs JournalEntry.
  [V3] The discover auto-match is correct for every loan.
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import difflib
import logging
import re
import shutil
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("missing dependency. Run: pip3 install --break-system-packages openpyxl requests")
    sys.exit(1)

import qbo_vault as kc

# ───────────────────────────  config  ───────────────────────────

API_BASE = "https://quickbooks.api.intuit.com"

# The automated debt schedule (single source of truth for terms).
WORKBOOK = Path(__file__).resolve().parent / "debt-schedule" / "Equipment_Debt_Schedule_v2.xlsx"

# Which company this QBO login belongs to. Only rows tagged with this company are
# synced (L&A Holdings is a SEPARATE QuickBooks file — run with --company "L&A Holdings"
# and its own credentials to sync those).
DEFAULT_COMPANY = "Proficient"

# Logs live OUTSIDE the project folder (house rule).
LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient"
BACKUP_DIR = Path.home() / "Library" / "Application Support" / "proficient-automation" / "debt-schedule-backups"

MASTER_SHEET = "Master Debt Schedule"
SETUP_SHEET = "QBO Setup"
LEDGER_SHEET = "QBO Payment Ledger"

# Master layout (v2 workbook — Company + Account # added).
MASTER_HEADER_ROW = 3
MASTER_FIRST_DATA_ROW = 4
COL_COMPANY = 1                    # A
COL_EQUIP = 2                      # B
COL_DESC = 3                       # C
COL_LENDER = 4                     # D
COL_ACCT_NO = 5                    # E  ← lender account number (QBO match helper)
COL_ORIG = 6                       # F
COL_CURRENT_BAL = 7               # G  ← QBO actual balance lands here
COL_ASOF = 8                       # H  ← sync date
COL_TERM = 10                      # J
COL_STATUS = 14                    # N
COL_AMORT_TAB = 18                 # R

# Transaction entities that can carry a loan-payment line.
TXN_ENTITIES = ("Purchase", "Bill", "JournalEntry")

DEFAULT_SINCE_MONTHS = 24

# ───────────────────────────  logging  ───────────────────────────

def _setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("loan_sync")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        fh = logging.FileHandler(LOG_DIR / "loan_sync.log")
        fh.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-7s %(message)s"))
        logger.addHandler(fh)
    return logger

log = _setup_logging()

def out(msg: str) -> None:
    """Console + log."""
    print(msg)
    log.info(msg)

# ───────────────────────────  auth  ───────────────────────────

def load_credentials() -> Tuple[str, str]:
    """Single Touch ID prompt unlocks all QBO keys. Returns (access_token, company_id)."""
    if not kc.has_credentials():
        out("✗ no credentials stored.  fix:  python3 setup_qbo.py")
        sys.exit(1)
    try:
        creds = kc.get_all()
    except kc.SecretsError as e:
        out(f"✗ Keychain read failed: {e}")
        sys.exit(1)

    required = ["QBO_CLIENT_ID", "QBO_CLIENT_SECRET", "QBO_COMPANY_ID", "QBO_REFRESH_TOKEN"]
    missing = [k for k in required if not creds.get(k)]
    if missing:
        out(f"✗ incomplete blob. Missing: {', '.join(missing)}  fix:  python3 setup_qbo.py")
        sys.exit(1)

    basic = base64.b64encode(
        f"{creds['QBO_CLIENT_ID']}:{creds['QBO_CLIENT_SECRET']}".encode()
    ).decode()
    r = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        headers={
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        data={"grant_type": "refresh_token", "refresh_token": creds["QBO_REFRESH_TOKEN"]},
        timeout=30,
    )
    if r.status_code != 200:
        out(f"✗ token refresh failed  status={r.status_code}  body={r.text[:300]}")
        out("  diagnose:  python3 setup_qbo.py --test")
        sys.exit(1)

    data = r.json()
    new_rt = data.get("refresh_token")
    if new_rt and new_rt != creds["QBO_REFRESH_TOKEN"]:
        try:
            kc.put("QBO_REFRESH_TOKEN", new_rt)
        except kc.SecretsError:
            pass
    return data["access_token"], creds["QBO_COMPANY_ID"]

# ───────────────────────────  api (with retry)  ───────────────────────────

def _api_get(path: str, access: str, params: Optional[dict] = None) -> dict:
    p = dict(params or {})
    p["minorversion"] = "70"
    delay = 1
    for attempt in range(6):
        r = requests.get(
            f"{API_BASE}{path}",
            headers={"Authorization": f"Bearer {access}", "Accept": "application/json"},
            params=p,
            timeout=45,
        )
        if r.status_code == 200:
            return r.json()
        # transient: 429 rate limit, 5xx server hiccup → backoff & retry
        if r.status_code == 429 or 500 <= r.status_code < 600:
            if attempt < 5:
                log.warning(f"{path} → {r.status_code}; retry in {delay}s")
                time.sleep(delay)
                delay *= 2
                continue
        raise RuntimeError(f"{path} → {r.status_code}: {r.text[:300]}")
    raise RuntimeError(f"{path} → exhausted retries")

def query(access: str, company_id: str, q: str) -> dict:
    return _api_get(f"/v3/company/{company_id}/query", access, params={"query": q})

def query_all(access: str, company_id: str, entity: str, where: str = "") -> List[dict]:
    """Paginated query — every row of `entity`, optional WHERE (no OR; QBO AND-only)."""
    out_rows: List[dict] = []
    start, page = 1, 500
    while True:
        q = f"SELECT * FROM {entity}"
        if where:
            q += f" WHERE {where}"
        q += f" STARTPOSITION {start} MAXRESULTS {page}"
        data = query(access, company_id, q)
        batch = data.get("QueryResponse", {}).get(entity, [])
        if not batch:
            break
        out_rows.extend(batch)
        if len(batch) < page:
            break
        start += page
    return out_rows

# ───────────────────────────  QBO helpers  ───────────────────────────

def fetch_liability_accounts(access: str, company_id: str) -> List[dict]:
    """All liability accounts with balances. Classification == 'Liability'."""
    rows = query_all(access, company_id, "Account", where="Classification = 'Liability'")
    if not rows:  # fallback if Classification filter unsupported on a tier
        rows = [a for a in query_all(access, company_id, "Account")
                if a.get("Classification") == "Liability"]
    return rows

def account_balance(acct: dict) -> float:
    """Display balance = amount owed. We show abs() — see [V1]."""
    return abs(float(acct.get("CurrentBalance", 0) or 0))

def _line_hits(line: dict, acct_id: str) -> Optional[float]:
    """If this line posts to acct_id, return its amount; else None.
    Covers Purchase/Bill (AccountBasedExpenseLineDetail) and JournalEntry."""
    det = line.get("AccountBasedExpenseLineDetail")
    if det and det.get("AccountRef", {}).get("value") == acct_id:
        return float(line.get("Amount", 0) or 0)
    jed = line.get("JournalEntryLineDetail")
    if jed and jed.get("AccountRef", {}).get("value") == acct_id:
        amt = float(line.get("Amount", 0) or 0)
        # Debit to a liability = paydown (positive); Credit = draw (negative).
        return amt if jed.get("PostingType") == "Debit" else -amt
    return None

def _txn_name(txn: dict) -> str:
    for k in ("EntityRef", "VendorRef", "CustomerRef"):
        if txn.get(k):
            return txn[k].get("name", "")
    return ""

def extract_payments(txns: List[dict], entity: str, acct_id: str) -> List[dict]:
    """Keep only txns with a line hitting acct_id; one ledger row per hit."""
    rows = []
    for t in txns:
        for ln in t.get("Line", []):
            amt = _line_hits(ln, acct_id)
            if amt is None:
                continue
            rows.append({
                "date": t.get("TxnDate", ""),
                "type": f"{entity}" + (f" ({t.get('PaymentType')})" if entity == "Purchase" and t.get("PaymentType") else ""),
                "doc": t.get("DocNumber", ""),
                "name": _txn_name(t),
                "acct_amount": amt,                       # portion hitting the loan account
                "total_amount": float(t.get("TotalAmt", 0) or 0),
                "memo": (t.get("PrivateNote") or ln.get("Description") or "")[:200],
                "txn_id": t.get("Id", ""),
                "synctoken": t.get("SyncToken", ""),
            })
    return rows

# ───────────────────────────  fuzzy matching  ───────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9 ]", " ", (s or "").lower())

def _asset_numbers(text: str) -> List[str]:
    """Extract 2-4 digit asset/unit numbers e.g. '#013' -> '013','13'."""
    nums = re.findall(r"#?\s*(\d{2,4})", text or "")
    extra = [n.lstrip("0") or "0" for n in nums]
    return list({*nums, *extra})

def _digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")

def match_account(loan: dict, accounts: List[dict]) -> Tuple[Optional[dict], float]:
    """Score each liability account against a loan row. Returns (best, score 0-1).
    Strongest signal: the lender account number matching the QBO account's AcctNum
    or appearing in its name."""
    lender = _norm(loan["lender"])
    desc = _norm(loan["desc"])
    equip = loan["equip"]
    acct_no = loan.get("acct_no", "")
    acct_digits = _digits(acct_no)
    loan_nums = _asset_numbers(equip) + _asset_numbers(desc)
    best, best_score = None, 0.0
    for a in accounts:
        raw_name = a.get("Name", "") + " " + (a.get("FullyQualifiedName", "") or "")
        name = _norm(raw_name)
        toks_loan = set((lender + " " + desc).split())
        toks_acct = set(name.split())
        overlap = len(toks_loan & toks_acct) / max(1, len(toks_loan))
        seq = difflib.SequenceMatcher(None, lender, name).ratio()
        num_hit = 1.0 if any(n in name.split() for n in loan_nums) else 0.0
        # lender account-number match (very strong): QBO AcctNum or digits-in-name
        acct_hit = 0.0
        if acct_digits and len(acct_digits) >= 4:
            qbo_acctnum_digits = _digits(a.get("AcctNum", ""))
            if qbo_acctnum_digits and acct_digits in qbo_acctnum_digits:
                acct_hit = 1.0
            elif acct_digits in _digits(raw_name):
                acct_hit = 1.0
        if acct_hit:
            score = 0.80 + 0.20 * overlap          # near-certain
        else:
            score = 0.45 * overlap + 0.25 * seq + 0.30 * num_hit
        if score > best_score:
            best, best_score = a, score
    return best, round(best_score, 3)

# ───────────────────────────  workbook IO  ───────────────────────────

def _read_loans(wb) -> List[dict]:
    """Read financed-asset rows: those with a Company tag + amort tab, up to TOTALS."""
    ws = wb[MASTER_SHEET]
    loans = []
    for r in range(MASTER_FIRST_DATA_ROW, ws.max_row + 1):
        company = ws.cell(r, COL_COMPANY).value
        if str(company or "").strip().upper() == "TOTALS":
            break
        equip = ws.cell(r, COL_EQUIP).value
        amort = ws.cell(r, COL_AMORT_TAB).value
        if not equip or not amort:           # skips receivable-only / blank rows
            continue
        loans.append({
            "row": r,
            "company": str(company or "").strip(),
            "equip": str(equip),
            "desc": str(ws.cell(r, COL_DESC).value or ""),
            "lender": str(ws.cell(r, COL_LENDER).value or ""),
            "acct_no": str(ws.cell(r, COL_ACCT_NO).value or "").strip(),
            "amort_tab": str(amort),
            "status": str(ws.cell(r, COL_STATUS).value or ""),
        })
    return loans

def _read_setup(wb) -> Dict[str, dict]:
    """equip# -> {acct_id, acct_name, acctnum} from the QBO Setup sheet, if present."""
    if SETUP_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SETUP_SHEET]
    mapping = {}
    for r in range(2, ws.max_row + 1):
        equip = ws.cell(r, 2).value            # col B = Equipment #
        acct_id = ws.cell(r, 6).value          # col F = QBO Acct ID (the anchor)
        confirm = str(ws.cell(r, 10).value or "").strip().upper()  # col J = CONFIRM
        if equip and acct_id and confirm == "Y":   # only CONFIRMED rows ever sync
            mapping[str(equip)] = {
                "acct_id": str(acct_id).strip(),
                "acct_name": str(ws.cell(r, 5).value or ""),
                "acctnum": str(ws.cell(r, 7).value or ""),
            }
    return mapping

# styling helpers (match the workbook's existing palette)
HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name="Arial", color="FFFFFF", bold=True)
SUB_FONT = Font(name="Arial", color="000000")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
CUR_FMT = '_-"$"* #,##0.00_-;[Red]_-"$"* (#,##0.00)_-;_-"$"* "-"??_-;_-@_-'

def write_setup_sheet(wb, loans: List[dict], accounts: List[dict], logger_lines: List[str],
                      company: str = DEFAULT_COMPANY) -> None:
    """Create/refresh QBO Setup with one row per loan + best-guess account match.
    Rows whose Company != `company` belong to a different QBO file and are not matched."""
    if SETUP_SHEET in wb.sheetnames:
        existing = _read_setup(wb)
        del wb[SETUP_SHEET]
    else:
        existing = {}
    ws = wb.create_sheet(SETUP_SHEET, index=1)
    headers = ["Company", "Equipment #", "Lender", "Lender Acct #",
               "QBO Account (matched)", "QBO Acct ID", "QBO AcctNum",
               "QBO Balance", "Match Confidence", "CONFIRM (Y/N)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    acct_names = [a.get("Name", "") for a in accounts]
    by_id = {a["Id"]: a for a in accounts}

    for i, loan in enumerate(loans, start=2):
        is_other_company = loan.get("company") != company
        prior = existing.get(loan["equip"])
        if prior and prior.get("acct_id") in by_id:
            a, conf, confirmed = by_id[prior["acct_id"]], 1.0, "Y"
        elif is_other_company:
            a, conf, confirmed = None, 0.0, "N"   # different QBO file — don't match here
        else:
            a, conf = match_account(loan, accounts)
            confirmed = "N"
        ws.cell(i, 1, loan.get("company", ""))
        ws.cell(i, 2, loan["equip"])
        ws.cell(i, 3, loan["lender"])
        ws.cell(i, 4, loan.get("acct_no", ""))
        ws.cell(i, 5, a.get("Name", "") if a else "")
        ws.cell(i, 6, a.get("Id", "") if a else "")
        ws.cell(i, 7, a.get("AcctNum", "") if a else "")
        ws.cell(i, 8, account_balance(a) if a else 0).number_format = CUR_FMT
        ws.cell(i, 9, conf)
        ws.cell(i, 10, "—" if is_other_company else confirmed)
        if is_other_company:
            for c in range(1, 11):
                ws.cell(i, c).fill = PatternFill("solid", start_color="E2EFDA")  # L&A / other co.
        elif conf < 0.55 and confirmed != "Y":
            for c in range(1, 11):
                ws.cell(i, c).fill = WARN_FILL
        tag = "  [other QBO file — skip here]" if is_other_company else ""
        logger_lines.append(
            f"  {loan['equip']:<22} → {a.get('Name','(no match)') if a else '(no match)':<30} "
            f"conf={conf}  bal={account_balance(a) if a else 0:,.2f}{tag}"
        )
    for c, w in zip(range(1, 11), (14, 22, 24, 18, 30, 12, 12, 16, 14, 14)):
        ws.column_dimensions[get_column_letter(c)].width = w
    if acct_names:
        joined = ",".join(n.replace(",", " ") for n in acct_names if n)[:255]
        dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"E2:E{len(loans)+1}")
    note = ws.cell(len(loans) + 3, 1,
                   "Confirm each row: set CONFIRM=Y (col J). Anchor = QBO Acct ID (col F). "
                   "Fix a wrong match: pick the right account name in col E, then put its Id in "
                   "col F (re-run --discover to auto-fill the Id). Green rows = a different company's "
                   "QBO file (e.g. L&A Holdings) — they sync only when you run with that company's login.")
    note.font = Font(name="Arial", italic=True, color="666666")

def ensure_ledger_sheet(wb):
    if LEDGER_SHEET in wb.sheetnames:
        return wb[LEDGER_SHEET]
    ws = wb.create_sheet(LEDGER_SHEET)
    headers = ["Date", "Equipment #", "QBO Account", "Type", "Doc #", "Name",
               "Amount to Loan Acct", "Total Payment", "Memo", "QBO Txn ID",
               "SyncToken", "Acct ID"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for c, w in zip(range(1, 13), (12, 18, 26, 18, 12, 24, 18, 16, 40, 14, 12, 10)):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    return ws

# ledger column index of the QBO Account Id (the dedup anchor)
LEDGER_TXN_COL = 10
LEDGER_ACCTID_COL = 12

def existing_ledger_keys(ws) -> set:
    """Dedup identity = (QBO Txn Id, QBO Account Id) — stable across renames."""
    keys = set()
    for r in range(2, ws.max_row + 1):
        txn = ws.cell(r, LEDGER_TXN_COL).value
        acct_id = ws.cell(r, LEDGER_ACCTID_COL).value
        if txn:
            keys.add((str(txn), str(acct_id or "")))
    return keys

# ───────────────────────────  core sync  ───────────────────────────

def _since_default() -> str:
    today = dt.date.today()
    y, m = today.year, today.month - DEFAULT_SINCE_MONTHS
    while m <= 0:
        m += 12; y -= 1
    return f"{y:04d}-{m:02d}-01"

def run_sync(since: str, dry_run: bool, company: str = DEFAULT_COMPANY) -> int:
    if not WORKBOOK.exists():
        out(f"✗ workbook not found: {WORKBOOK}")
        return 1

    access, company_id = load_credentials()
    out(f"✓ authenticated  (QBO realm {company_id})  syncing company: {company}")

    wb = openpyxl.load_workbook(WORKBOOK)
    all_loans = _read_loans(wb)
    loans = [l for l in all_loans if l.get("company") == company]
    skipped = [l for l in all_loans if l.get("company") != company]
    if skipped:
        out(f"  (skipping {len(skipped)} row(s) for other companies — separate QBO file: "
            f"{', '.join(sorted({l.get('company','?') for l in skipped}))})")
    setup = _read_setup(wb)
    if not setup:
        out("✗ no confirmed mappings in 'QBO Setup'. Run:  python3 loan_sync.py --discover")
        out("  then open the workbook, confirm each row (CONFIRM=Y), save, and re-run.")
        return 2

    # 1) refresh balances from the live liability accounts
    accounts = {a["Id"]: a for a in fetch_liability_accounts(access, company_id)}
    ws_master = wb[MASTER_SHEET]
    today_str = dt.date.today().strftime("%Y-%m-%d")
    bal_changes = []
    mapped_acct_ids = {}
    for loan in loans:
        m = setup.get(loan["equip"])
        if not m:
            continue
        mapped_acct_ids[loan["equip"]] = m["acct_id"]
        acct = accounts.get(m["acct_id"])
        if not acct:
            out(f"  ! {loan['equip']}: mapped acct {m['acct_id']} not found in QBO liabilities")
            continue
        new_bal = account_balance(acct)
        old_bal = ws_master.cell(loan["row"], COL_CURRENT_BAL).value
        bal_changes.append((loan["equip"], old_bal, new_bal))
        ws_master.cell(loan["row"], COL_CURRENT_BAL, new_bal)
        ws_master.cell(loan["row"], COL_ASOF, today_str)

    # 2) pull payment transactions hitting each mapped account
    where = f"TxnDate >= '{since}'"
    all_new_rows = []
    ledger = ensure_ledger_sheet(wb)
    seen = existing_ledger_keys(ledger)
    acct_name_by_id = {aid: accounts.get(aid, {}).get("Name", "") for aid in set(mapped_acct_ids.values())}
    equip_by_acct = {}
    for eq, aid in mapped_acct_ids.items():
        equip_by_acct.setdefault(aid, eq)

    for entity in TXN_ENTITIES:
        txns = query_all(access, company_id, entity, where=where)
        for aid in set(mapped_acct_ids.values()):
            for row in extract_payments(txns, entity, aid):
                key = (row["txn_id"], aid)
                if key in seen:
                    continue
                seen.add(key)
                row["equip"] = equip_by_acct.get(aid, "")
                row["acct_name"] = acct_name_by_id.get(aid, "")
                row["acct_id"] = aid
                all_new_rows.append(row)

    # write new ledger rows (append-only)
    all_new_rows.sort(key=lambda r: (r["equip"], r["date"]))
    start_row = ledger.max_row + 1
    for i, r in enumerate(all_new_rows):
        rr = start_row + i
        ledger.cell(rr, 1, r["date"])
        ledger.cell(rr, 2, r["equip"])
        ledger.cell(rr, 3, r["acct_name"])
        ledger.cell(rr, 4, r["type"])
        ledger.cell(rr, 5, r["doc"])
        ledger.cell(rr, 6, r["name"])
        ledger.cell(rr, 7, r["acct_amount"]).number_format = CUR_FMT
        ledger.cell(rr, 8, r["total_amount"]).number_format = CUR_FMT
        ledger.cell(rr, 9, r["memo"])
        ledger.cell(rr, 10, r["txn_id"])
        ledger.cell(rr, 11, r["synctoken"])
        ledger.cell(rr, 12, r["acct_id"])

    # ── report ──
    out("\n── Balance refresh (QBO actual) ──")
    for eq, old, new in bal_changes:
        flag = "" if (isinstance(old, (int, float)) and abs((old or 0) - new) < 0.01) else "  *changed*"
        out(f"  {eq:<22} {(old or 0):>14,.2f} → {new:>14,.2f}{flag}")
    out(f"\n── New payments added to ledger: {len(all_new_rows)} ──")
    for r in all_new_rows[:40]:
        out(f"  {r['date']}  {r['equip']:<20} {r['acct_amount']:>12,.2f}  {r['type']}  {r['txn_id']}")
    if len(all_new_rows) > 40:
        out(f"  …and {len(all_new_rows) - 40} more")

    # ── save ──
    if dry_run:
        preview = WORKBOOK.with_name(WORKBOOK.stem + "_PREVIEW.xlsx")
        wb.save(preview)
        out(f"\n[DRY RUN] nothing live was changed.")
        out(f"   preview written → {preview}")
    else:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = BACKUP_DIR / f"{WORKBOOK.stem}__{stamp}.xlsx"
        shutil.copy2(WORKBOOK, backup)
        wb.save(WORKBOOK)
        out(f"\n✓ saved.  backup → {backup}")
    out("\nReminder [V1/V2/V3]: confirm balance sign, txn types, and mappings on first run.")
    return 0

def run_discover(company: str = DEFAULT_COMPANY) -> int:
    if not WORKBOOK.exists():
        out(f"✗ workbook not found: {WORKBOOK}")
        return 1
    access, company_id = load_credentials()
    out(f"✓ authenticated  (QBO realm {company_id})  mapping company: {company}")
    accounts = fetch_liability_accounts(access, company_id)
    out(f"✓ pulled {len(accounts)} liability accounts from QBO")

    wb = openpyxl.load_workbook(WORKBOOK)
    loans = _read_loans(wb)
    lines: List[str] = []
    write_setup_sheet(wb, loans, accounts, lines, company=company)
    out("\n── Proposed loan → QBO account matches (confirm in the workbook) ──")
    for ln in lines:
        out(ln)

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"{WORKBOOK.stem}__predisc_{stamp}.xlsx"
    shutil.copy2(WORKBOOK, backup)
    wb.save(WORKBOOK)
    out(f"\n✓ 'QBO Setup' sheet written.  backup → {backup}")
    out("  NEXT: open the workbook → QBO Setup → fix any low-confidence (yellow) rows,")
    out("        set CONFIRM=Y on every loan, save, then run:  python3 loan_sync.py --dry-run")
    return 0

# ───────────────────────────  cli  ───────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="One-click QBO → equipment debt-schedule sync.")
    ap.add_argument("--discover", action="store_true",
                    help="Pull QBO liability accounts and write/refresh the QBO Setup mapping.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Preview only: write *_PREVIEW.xlsx, change nothing live.")
    ap.add_argument("--since", default=None,
                    help="Transaction window start (YYYY-MM-DD). Default: last 24 months.")
    ap.add_argument("--company", default=DEFAULT_COMPANY,
                    help="Which company's rows to sync (matches the Company column / your QBO "
                         "login). Default: Proficient. Use 'L&A Holdings' with L&A credentials.")
    args = ap.parse_args()

    try:
        if args.discover:
            return run_discover(company=args.company)
        since = args.since or _since_default()
        out(f"▶ loan_sync  since={since}  company={args.company}  {'(dry run)' if args.dry_run else ''}")
        return run_sync(since, dry_run=args.dry_run, company=args.company)
    except RuntimeError as e:
        out(f"✗ QBO error: {e}")
        return 1
    except Exception as e:  # noqa
        log.exception("unexpected failure")
        out(f"✗ unexpected: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main())
