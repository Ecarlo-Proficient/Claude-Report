"""Build PC Equipment Debt Schedule (master-driven amortization workbook)."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

# ---------- Styling ----------
FONT_NAME = "Arial"
BLUE_INPUT = Font(name=FONT_NAME, color="0000FF", bold=False)
BLACK_FORMULA = Font(name=FONT_NAME, color="000000")
GREEN_LINK = Font(name=FONT_NAME, color="006100")  # link from other sheet
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
SUBHEADER_FONT = Font(name=FONT_NAME, color="000000", bold=True, size=11)
TITLE_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=14)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")  # dark blue
SUB_FILL = PatternFill("solid", start_color="D9E1F2")  # light blue
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")  # needs input
GREY_FILL = PatternFill("solid", start_color="F2F2F2")
TOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

CURRENCY_FMT = '_-"$"* #,##0.00_-;[Red]_-"$"* (#,##0.00)_-;_-"$"* "-"??_-;_-@_-'
NUMBER_FMT = '#,##0.00;[Red](#,##0.00);"-"'
PCT_FMT = "0.0000%"
DATE_FMT = "m/d/yyyy"
INT_FMT = "0"

# ---------- Loan seed data ----------
# Parsed from existing PC_Equipment_Debt_Schedule - March 2026 FINAL.xlsx
# rate is annual %. None means compute from RATE() in Excel.
LOANS = [
    # equip_no, description, lender, original, current, payment, term, rate, notes_original
    ("PUMP #013", "2020 MACK Granite", "Alliance Funding Group", 190000.00, 33927.21, 3769.69, 60, None, "60 months"),
    ("TRUCK #089", "2019 Ram 3500", "Ally Financial-Truck", 52863.14, 25637.87, 888.66, 75, 7.64, "7.64% Interest (75 months)"),
    ("TRUCK #098", "2024 Toyota Sequoia", "Ally Financial-Truck", 111266.64, 65351.35, 1545.37, 72, 8.14, "8.14% Interest (72 months)"),
    ("PUMP #018", "2019 PUTZMEISTER", "First Citizen Bank", 140000.00, 84691.61, 3467.07, 48, None, "48 months (Rate Factor: 0.0247648)"),
    ("PUMP #014", "2023 FRHT", "Frost Leasing", 422200.00, 128138.06, 8432.76, 60, 7.3689, "7.3689% Interest (60 months)"),
    ("PUMP #015", "2023 MACK GR85-F", "Frost Leasing", 620000.00, 287716.63, 12501.75, 60, 7.7677, "7.7677% Interest (60 months)"),
    ("PUMP #017", "2023 MACK TE86", "Frost Leasing", 863327.54, 631222.60, 13744.69, 84, 8.6731, "8.6731% Interest (84 months)"),
    ("SOMERO #219", "Somero 2024 S15R", "Frost Leasing", 266414.42, 172524.37, 5413.53, 60, 8.0957, "8.0957% (60 months)"),
    ("SOMERO #205,#206", "Somero Laser system", "Frost Leasing", 313585.58, 88535.86, 7719.29, 48, 8.4377, "8.4377% Interest (48 months)"),
    ("SAW #207", "SAW - HUSQ", "Leaf Capital (Lease Services)", 37978.43, 1235.06, 1235.06, 36, None, "36 months"),
    ("Truck #096", "2021 RAM 350", "Leaf Capital (Lease Services)", 51357.03, 13369.28, 1671.16, 36, None, "36 months"),
    ("Truck #097", "2015 RAM 3500", "Leaf Capital (Lease Services)", 36425.86, 14695.68, 1224.64, 36, None, "36 months"),
    ("Equipment #211-#217", "Equipment #211-#217", "Leaf Capital (Lease Services)", 141133.60, 104128.40, 3062.60, 60, None, "60 months"),
    ("Pump #019", "2021 FREIGHTLINER", "M2 Equipment Finance", 415000.00, 399501.96, 7682.73, 72, 9.97, "72 months (9.97%)"),
    ("Tractor #201", "2022 Bobcat T66", "Wells Fargo", 79503.03, 16877.23, 1428.79, 60, 2.90, "2.90% Interest (60 months)"),
    ("Tractor #204", "2022 Bobcat E42 T4", "Wells Fargo", 70942.68, 29572.99, 1106.68, 72, 3.90, "3.90% Interest (72 months)"),
    ("Tractors #208,#209, #210", "2023 Bobcat T770, T66, & 30C Auger", "Wells Fargo", 177866.49, 34585.22, 4940.74, 36, 0.00, "0.00% (36 months)"),
    ("Tractor #218", "2024 Bobcat T66", "Wells Fargo", 66206.72, 33103.28, 1379.31, 48, 0.00, "0.00% (48 months)"),
    ("Tractor #220/221", "2024 Bobcat T66 & sweeper", "Wells Fargo", 80395.90, 54440.45, 1405.64, 60, 1.90, "1.90% ( 60 months)"),
    ("Tractor #222", "2024 Bobcat T66", "Wells Fargo", 64014.29, 40008.95, 1333.63, 48, 0.00, "0.00% (48 months)"),
]

# Receivables (preserved from existing) — (equip#, customer, orig, current, payment)
RECEIVABLES = {
    "PUMP #013": ("MCP - Pump #013", 84886.83, 84886.83, 4364.11),
    "TRUCK #089": ("Escobar Concrete #089", 51563.78, 51563.78, 954.88),
    "PUMP #014": ("Core Concrete #014 | MCP - Pump #014 -", 216456.84, 216456.84, 12740.12),
    "PUMP #015": ("MCP - Pump #015", 337547.25, 337547.25, 12501.75),
    "PUMP #017": ("Core Concrete #017", 640000.00, 640000.00, 10877.27),
    "SOMERO #219": ("JCP Concrete SOMERO #219-C/15", 340000.00, 340000.00, 11056.08),
    "Pump #019": ("MCP - Pump #019", 288168.32, 288168.32, 7926.98),
}

# Receivable-only rows (no loan side)
RECEIVABLE_ONLY = [
    ("Equipment ##010", "Erick Martinez-Pump #010-C/15", 59703.65, 59703.65, 2323.67),
    ("Equipment ##011", "Core Concrete #011", 60000.00, 60000.00, 1837.19),
    ("Equipment ##090", "Escobar Concrete #090", 58289.48, 58289.48, 1079.43),
    ("Equipment ##146", "Escobar Conctete #146", 3148.25, 3148.25, 524.71),
    ("Equipment ##157", "Escobar Concrete #157", 3148.25, 3148.25, 524.71),
    ("Equipment ##159", "Escobar Concrete #159", 3148.25, 3148.25, 524.71),
    ("Equipment ##160", "Escobar Concrete #160", 3148.25, 3148.25, 524.71),
    ("Equipment ##179", "Escobar Concrete #179", 28023.77, 28023.77, 518.96),
    ("Equipment ##183", "Escobar Concrete #183", 28023.77, 28023.77, 518.96),
    ("Equipment ##088", "Carlos Alvarez-Dump Truck #088", 80000.00, 80000.00, 1246.90),
    ("Equipment ##003", "Javier Alvarez-Dump Truck #003", 40000.00, 40000.00, 976.52),
]


def sanitize_sheet_name(name: str) -> str:
    """Excel sheet name: <=31 chars, no []:*?/\\, no leading/trailing apostrophe."""
    clean = re.sub(r"[\[\]\:\*\?/\\]", "", name)
    clean = clean.replace("#", "").replace(",", "_").replace(" ", "_").replace("/", "_")
    clean = re.sub(r"_+", "_", clean).strip("_")
    return clean[:31]


def fmt_cell(cell, font=None, fill=None, fmt=None, align=None, border=BORDER):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def build_master(wb, as_of_date_str):
    ws = wb.create_sheet("Master Debt Schedule")
    # Column widths
    widths = {
        "A": 26, "B": 32, "C": 28, "D": 16, "E": 16, "F": 13, "G": 13,
        "H": 9, "I": 11, "J": 13, "K": 6, "L": 10, "M": 14, "N": 14,
        "O": 38, "P": 22, "Q": 3, "R": 32, "S": 18, "T": 18, "U": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title row
    ws.merge_cells("A1:U1")
    ws["A1"] = "EQUIPMENT DEBT & RECEIVABLE SCHEDULE — Master (single source of truth)"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 26

    # Section banners
    ws.merge_cells("A2:B2"); ws["A2"] = "Equipment"
    ws.merge_cells("C2:O2"); ws["C2"] = "Loan Information"
    ws.merge_cells("R2:U2"); ws["R2"] = "Receivable Information"
    for c in ("A2", "C2", "R2"):
        fmt_cell(ws[c], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    # Header row 3
    headers = {
        "A": "Equipment #",
        "B": "Description",
        "C": "Lender",
        "D": "Original Loan Balance",
        "E": "Current Balance",
        "F": "As-of Date",
        "G": "Monthly Payment",
        "H": "Term (mo)",
        "I": "Annual Rate %",
        "J": "Start Date",
        "K": "Pmt Day",
        "L": "Status",
        "M": "YTD Principal",
        "N": "YTD Interest",
        "O": "Notes (original Terms text)",
        "P": "Amort Tab",
        "R": "Customer",
        "S": "Original Receivable",
        "T": "Current Receivable",
        "U": "Receivable Payment",
    }
    for col, text in headers.items():
        c = ws[f"{col}3"]
        c.value = text
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"

    # Data rows
    row = 4
    sheet_names = []
    for (equip, desc, lender, orig, curr, pmt, term, rate, notes) in LOANS:
        ws[f"A{row}"] = equip
        ws[f"B{row}"] = desc
        ws[f"C{row}"] = lender
        ws[f"D{row}"] = orig
        ws[f"E{row}"] = curr
        ws[f"F{row}"] = as_of_date_str  # date string; will format below
        ws[f"G{row}"] = pmt
        ws[f"H{row}"] = term
        # Rate: hardcode if known; else compute fully-amortizing rate via RATE() (FV=0)
        if rate is None:
            # Solve for rate that amortizes Original over Term months to $0
            ws[f"I{row}"] = f"=IFERROR(RATE(H{row},-G{row},D{row})*12,0)"
        else:
            ws[f"I{row}"] = rate / 100.0  # store as decimal, format as %
        # Start Date, Pmt Day — leave blank for user input
        ws[f"J{row}"] = None
        ws[f"K{row}"] = 1
        ws[f"L{row}"] = "Active"
        # YTD Principal/Interest — link to amort tab fixed cells (B16, B17)
        tab = sanitize_sheet_name(equip)
        sheet_names.append((equip, tab))
        ws[f"M{row}"] = f"=IFERROR(INDIRECT(\"'\"&P{row}&\"'!P5\"),0)"
        ws[f"N{row}"] = f"=IFERROR(INDIRECT(\"'\"&P{row}&\"'!Q5\"),0)"
        ws[f"O{row}"] = notes
        ws[f"P{row}"] = tab

        # Receivable (if exists)
        if equip in RECEIVABLES:
            cust, r_orig, r_curr, r_pmt = RECEIVABLES[equip]
            ws[f"R{row}"] = cust
            ws[f"S{row}"] = r_orig
            ws[f"T{row}"] = r_curr
            ws[f"U{row}"] = r_pmt

        # Formatting per cell
        fmt_cell(ws[f"A{row}"], font=BLACK_FORMULA, align=LEFT)
        fmt_cell(ws[f"B{row}"], font=BLUE_INPUT, align=LEFT)
        fmt_cell(ws[f"C{row}"], font=BLUE_INPUT, align=LEFT)
        fmt_cell(ws[f"D{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"E{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"F{row}"], font=BLUE_INPUT, fmt=DATE_FMT, align=CENTER, fill=YELLOW_FILL)
        fmt_cell(ws[f"G{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"H{row}"], font=BLUE_INPUT, fmt=INT_FMT, align=CENTER)
        if rate is None:
            fmt_cell(ws[f"I{row}"], font=BLACK_FORMULA, fmt=PCT_FMT, align=CENTER, fill=GREY_FILL)
        else:
            fmt_cell(ws[f"I{row}"], font=BLUE_INPUT, fmt=PCT_FMT, align=CENTER)
        fmt_cell(ws[f"J{row}"], font=BLUE_INPUT, fmt=DATE_FMT, align=CENTER, fill=YELLOW_FILL)
        fmt_cell(ws[f"K{row}"], font=BLUE_INPUT, fmt=INT_FMT, align=CENTER)
        fmt_cell(ws[f"L{row}"], font=BLUE_INPUT, align=CENTER)
        fmt_cell(ws[f"M{row}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"N{row}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"O{row}"], font=BLACK_FORMULA, align=LEFT)
        fmt_cell(ws[f"P{row}"], font=BLACK_FORMULA, align=LEFT, fill=GREY_FILL)
        # Receivable cols
        for col in ("R", "S", "T", "U"):
            cell = ws[f"{col}{row}"]
            f = BLUE_INPUT
            fmt = None
            if col in ("S", "T", "U"):
                fmt = CURRENCY_FMT
            fmt_cell(cell, font=f, fmt=fmt, align=RIGHT if col != "R" else LEFT)
        row += 1

    # Receivable-only rows
    for (equip, cust, r_orig, r_curr, r_pmt) in RECEIVABLE_ONLY:
        ws[f"A{row}"] = equip
        ws[f"B{row}"] = ""
        ws[f"L{row}"] = "Receivable Only"
        ws[f"R{row}"] = cust
        ws[f"S{row}"] = r_orig
        ws[f"T{row}"] = r_curr
        ws[f"U{row}"] = r_pmt
        fmt_cell(ws[f"A{row}"], font=BLACK_FORMULA, align=LEFT)
        for col in "BCDEFGHIJKMNOP":
            fmt_cell(ws[f"{col}{row}"], font=BLACK_FORMULA, fill=GREY_FILL)
        fmt_cell(ws[f"L{row}"], font=BLACK_FORMULA, align=CENTER, fill=GREY_FILL)
        fmt_cell(ws[f"R{row}"], font=BLUE_INPUT, align=LEFT)
        for col in ("S", "T", "U"):
            fmt_cell(ws[f"{col}{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        row += 1

    # Totals row
    total_row = row
    ws[f"A{total_row}"] = "TOTALS"
    fmt_cell(ws[f"A{total_row}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for col, sumcol in (("D", "D"), ("E", "E"), ("G", "G"), ("M", "M"), ("N", "N"),
                        ("S", "S"), ("T", "T"), ("U", "U")):
        ws[f"{col}{total_row}"] = f"=SUM({sumcol}4:{sumcol}{total_row-1})"
        fmt_cell(ws[f"{col}{total_row}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, fmt=CURRENCY_FMT, align=RIGHT)
    for col in "BCFHIJKLOPRQ":
        if col != "Q":
            fmt_cell(ws[f"{col}{total_row}"], fill=TOTAL_FILL)

    return sheet_names, total_row


def build_amort_tab(wb, equip, desc, lender, orig, curr, pmt, term, rate_pct,
                     master_row, tab_name, max_term=120):
    """Build one amort tab. References Master row for live inputs.
    Schedule is origination-based: starts at (Effective) Start Date with Original Balance,
    runs Term months. If user leaves Start Date blank on Master, an implied start is
    computed by NPER from current balance back to origination.
    """
    ws = wb.create_sheet(tab_name)

    # Column widths — compromise between horizontal header section and schedule section.
    # Description/Lender cells in row 5 will wrap text since the row is tall.
    widths = {
        "A": 13, "B": 22, "C": 22, "D": 16, "E": 16, "F": 14, "G": 16,
        "H": 9,  "I": 11, "J": 14, "K": 11, "L": 12, "M": 11, "N": 13,
        "O": 9,  "P": 15, "Q": 15,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    M = master_row
    MASTER = "'Master Debt Schedule'"

    # ---------- ROW 1: Title (merged A:Q) ----------
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"AMORTIZATION SCHEDULE — {equip}"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 24

    # ---------- ROW 3: Loan info banner (merged) ----------
    ws.merge_cells("A3:Q3")
    ws["A3"] = "LOAN INFORMATION & KPIs — pulled live from Master. YTD cols (P/Q) reflect current calendar year."
    fmt_cell(ws["A3"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    # ---------- ROW 4: Column headers (horizontal) ----------
    headers = [
        "Equipment #", "Description", "Lender", "Original Balance",
        "Current Balance", "As-of Date", "Monthly Payment", "Term (mo)",
        "Annual Rate %", "Start Date", "Pmt Day", "Monthly Rate",
        "Implied Mo Made", "Effective Start", "Current Year",
        "YTD Principal", "YTD Interest",
    ]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}4"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[4].height = 32

    # ---------- ROW 5: Values ----------
    ws["A5"] = f"={MASTER}!A{M}"
    ws["B5"] = f"={MASTER}!B{M}"
    ws["C5"] = f"={MASTER}!C{M}"
    ws["D5"] = f"={MASTER}!D{M}"
    ws["E5"] = f"={MASTER}!E{M}"
    ws["F5"] = f"={MASTER}!F{M}"
    ws["G5"] = f"={MASTER}!G{M}"
    ws["H5"] = f"={MASTER}!H{M}"
    ws["I5"] = f"={MASTER}!I{M}"
    # Start Date — blank-safe so empty Master cell doesn't show 1/0/1900
    ws["J5"] = f'=IF({MASTER}!J{M}="","",{MASTER}!J{M})'
    ws["K5"] = f"={MASTER}!K{M}"
    ws["L5"] = "=I5/12"
    ws["M5"] = "=IFERROR(ROUND(IF(I5=0,IF(G5=0,0,(D5-E5)/G5),NPER(L5,-G5,D5,-E5)),0),0)"
    ws["N5"] = '=IFERROR(IF(AND(ISNUMBER(J5),J5>0),J5,EDATE(F5,-M5)),"")'
    ws["O5"] = "=YEAR(TODAY())"
    # P5 / Q5 filled after tracker

    fmt_cell(ws["A5"], font=GREEN_LINK, align=LEFT)
    fmt_cell(ws["B5"], font=GREEN_LINK,
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    fmt_cell(ws["C5"], font=GREEN_LINK,
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    fmt_cell(ws["D5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["E5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["F5"], font=GREEN_LINK, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["G5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["H5"], font=GREEN_LINK, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["I5"], font=GREEN_LINK, fmt=PCT_FMT, align=CENTER)
    fmt_cell(ws["J5"], font=GREEN_LINK, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["K5"], font=GREEN_LINK, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["L5"], font=BLACK_FORMULA, fmt=PCT_FMT, align=CENTER)
    fmt_cell(ws["M5"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["N5"], font=BLACK_FORMULA, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["O5"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["P5"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=TOTAL_FILL)
    fmt_cell(ws["Q5"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=TOTAL_FILL)
    ws.row_dimensions[5].height = 36

    # ---------- PAYMENT TRACKER (rows 7-21) ----------
    tracker_banner = 7
    tracker_header_row = 8
    tracker_start = 9
    tracker_years = 12
    tracker_end = tracker_start + tracker_years - 1  # 20
    tracker_totals = tracker_end + 1  # 21

    ws.merge_cells(f"A{tracker_banner}:P{tracker_banner}")
    ws[f"A{tracker_banner}"] = "PAYMENT TRACKER — enter $ paid each month when bank drafts"
    fmt_cell(ws[f"A{tracker_banner}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    tracker_headers = ["Year"] + ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"] + \
                       ["Total Paid", "Principal", "Interest"]
    for i, h in enumerate(tracker_headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}{tracker_header_row}"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    # Layout constants for schedule (defined here so tracker auto-pop formulas can ref them)
    sched_banner = tracker_totals + 2  # 23
    sched_header = sched_banner + 1     # 24
    sched_start = sched_header + 1      # 25
    sched_end = sched_start + max_term - 1  # 144

    # Tracker rows: Year col + auto-populate month cells + Total col
    # Month cells AUTO-FILL with scheduled payment when DATE(year, month, Pmt Day) <= TODAY().
    # SUMIFS pulls Sched Pmt (col D) from the amort schedule row matching this year+month.
    # To override, just type a number — it replaces the formula.
    for i in range(tracker_years):
        r = tracker_start + i
        ws[f"A{r}"] = f'=IFERROR(YEAR($N$5)+{i},"")'
        fmt_cell(ws[f"A{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
        for mc in range(12):
            col = get_column_letter(2 + mc)
            month_idx = mc + 1  # 1=Jan, ..., 12=Dec
            cell = ws[f"{col}{r}"]
            cell.value = (
                f'=IFERROR(IF(AND(ISNUMBER($A{r}),'
                f'DATE($A{r},{month_idx},$K$5)<=TODAY()),'
                f'SUMIFS(D{sched_start}:D{sched_end},'
                f'H{sched_start}:H{sched_end},$A{r},'
                f'I{sched_start}:I{sched_end},{month_idx}),'
                f'0),0)'
            )
            # Blue font signals "you can type here to override the auto-fill"
            fmt_cell(cell, font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        ws[f"N{r}"] = f"=SUM(B{r}:M{r})"
        fmt_cell(ws[f"N{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- AMORT SCHEDULE ----------

    ws.merge_cells(f"A{sched_banner}:L{sched_banner}")
    ws[f"A{sched_banner}"] = "AMORT SCHEDULE (origination-based, runs full Term)"
    fmt_cell(ws[f"A{sched_banner}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    sched_headers = ["Pmt #", "Date", "Beg Balance", "Scheduled Pmt",
                     "Sched Principal", "Sched Interest", "End Balance",
                     "Year", "Month", "Actual Paid", "Eff Principal", "Eff Interest"]
    for i, h in enumerate(sched_headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}{sched_header}"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    n_rows = max_term
    for i in range(n_rows):
        r = sched_start + i
        pmt_n = i + 1
        ws[f"A{r}"] = pmt_n
        ws[f"B{r}"] = (f'=IFERROR(IF(OR(A{r}>$H$5,$N$5="",$N$5=0),"",'
                       f'DATE(YEAR(EDATE($N$5,A{r}-1)),'
                       f'MONTH(EDATE($N$5,A{r}-1)),'
                       f'MIN($K$5,DAY(EOMONTH(EDATE($N$5,A{r}-1),0))))),"")')
        if i == 0:
            ws[f"C{r}"] = f'=IF(B{r}="","",$D$5)'
        else:
            ws[f"C{r}"] = f'=IF(B{r}="","",IFERROR(G{r-1},0))'
        ws[f"F{r}"] = f'=IF(B{r}="","",IFERROR(C{r}*$L$5,0))'
        ws[f"D{r}"] = f'=IF(B{r}="","",IFERROR(MIN($G$5,C{r}+F{r}),0))'
        ws[f"E{r}"] = f'=IF(B{r}="","",IFERROR(D{r}-F{r},0))'
        ws[f"G{r}"] = f'=IF(B{r}="","",IFERROR(MAX(0,C{r}-E{r}),0))'
        ws[f"H{r}"] = f'=IF(B{r}="","",YEAR(B{r}))'
        ws[f"I{r}"] = f'=IF(B{r}="","",MONTH(B{r}))'
        ws[f"J{r}"] = (f'=IFERROR(IF(B{r}="","",'
                       f'INDEX($B${tracker_start}:$M${tracker_end},'
                       f'MATCH(H{r},$A${tracker_start}:$A${tracker_end},0),'
                       f'I{r})),0)')
        ws[f"K{r}"] = f'=IF(B{r}="","",IF(J{r}>0,E{r},0))'
        ws[f"L{r}"] = f'=IF(B{r}="","",IF(J{r}>0,F{r},0))'

        fmt_cell(ws[f"A{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
        fmt_cell(ws[f"B{r}"], font=BLACK_FORMULA, fmt=DATE_FMT, align=CENTER)
        for col in ("C", "D", "E", "F", "G"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)
        for col in ("H", "I"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER, fill=GREY_FILL)
        fmt_cell(ws[f"J{r}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT, fill=GREY_FILL)
        for col in ("K", "L"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=GREY_FILL)

    # sched_end already defined above; just compute totals row
    sched_totals = sched_end + 1
    ws[f"A{sched_totals}"] = "TOTALS"
    fmt_cell(ws[f"A{sched_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for col in ("D", "E", "F", "J", "K", "L"):
        ws[f"{col}{sched_totals}"] = f"=SUM({col}{sched_start}:{col}{sched_end})"
        fmt_cell(ws[f"{col}{sched_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- Fill Tracker Principal / Interest ----------
    for i in range(tracker_years):
        r = tracker_start + i
        ws[f"O{r}"] = (f'=SUMIFS(K{sched_start}:K{sched_end},'
                       f'H{sched_start}:H{sched_end},A{r})')
        ws[f"P{r}"] = (f'=SUMIFS(L{sched_start}:L{sched_end},'
                       f'H{sched_start}:H{sched_end},A{r})')
        fmt_cell(ws[f"O{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"P{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)

    # Tracker totals row
    ws[f"A{tracker_totals}"] = "TOTALS"
    fmt_cell(ws[f"A{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for mc in range(12):
        col = get_column_letter(2 + mc)
        ws[f"{col}{tracker_totals}"] = f"=SUM({col}{tracker_start}:{col}{tracker_end})"
        fmt_cell(ws[f"{col}{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)
    for col in ("N", "O", "P"):
        ws[f"{col}{tracker_totals}"] = f"=SUM({col}{tracker_start}:{col}{tracker_end})"
        fmt_cell(ws[f"{col}{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- Populate P5 / Q5 (current-year YTD lookups) ----------
    ws["P5"] = (f'=IFERROR(VLOOKUP($O$5,A{tracker_start}:P{tracker_end},15,FALSE),0)')
    ws["Q5"] = (f'=IFERROR(VLOOKUP($O$5,A{tracker_start}:P{tracker_end},16,FALSE),0)')

    # Freeze the top header section (rows 1-5) so loan info stays visible while scrolling
    ws.freeze_panes = "A6"


def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    ws.merge_cells("A1:B1")
    ws["A1"] = "PC Equipment Debt Schedule — How this workbook works"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 28

    rows = [
        ("PURPOSE", "Single workbook to manage all equipment loans and their amortizations. "
                    "Master tab is the only place you enter loan info; each amort tab pulls from Master."),
        ("STRUCTURE", "Master Debt Schedule = inputs.  One amort tab per loan, named after Equipment #.  "
                      "_TEMPLATE_AMORT = copy this to add a new loan.  Archive = move paid-off tabs here."),
        ("INPUT CELLS", "Blue text = you type a value.  Yellow fill = required input (Start Date, As-of Date).  "
                        "Black text = formulas — do not type over.  Green text = pulled from another sheet."),
        ("ADDING A LOAN", "1) Add a row at bottom of Master with Equipment #, Description, Lender, balances, payment, term, rate.  "
                          "2) Right-click _TEMPLATE_AMORT tab → Move or Copy → check 'Create a copy'.  "
                          "3) Rename the new tab to match the Master 'Amort Tab' value for that row.  "
                          "4) On the new tab, change the Master row reference (the row number in cells B2:B11)."),
        ("MONTHLY WORKFLOW", "When a payment drafts: open that loan's amort tab → find the year row in PAYMENT TRACKER → "
                             "type the $ amount in that month's cell.  Workbook auto-splits to Principal & Interest "
                             "using the scheduled amort row matching that calendar month."),
        ("YEAR-END", "On each amort tab, the YTD Principal / YTD Interest cells (top, B16/B17) reflect the current calendar year "
                     "automatically.  Tracker has 12 years pre-built.  For prior-year totals, look at the corresponding row in Payment Tracker."),
        ("PAYOFF / ARCHIVE", "When a loan pays off: change Status to 'Paid Off' on Master, right-click the amort tab → Move or Copy → "
                              "to Archive.  Master totals will still include it unless you delete the row; or filter Master by Status."),
        ("RATE COMPUTATION", "Loans where the original Terms/Interest text did not include a rate are computed via Excel's RATE() function "
                              "on Master (cell shows in grey).  Verify these against your loan paperwork."),
        ("ASSUMPTIONS", "Schedule starts from CURRENT balance (as of the As-of Date), not from origination.  "
                        "Remaining months are computed from current balance + payment + rate.  "
                        "Each amort row uses standard PPMT/IPMT math: Interest = Beg Balance × monthly rate; Principal = Payment − Interest."),
        ("WHAT TO CHECK FIRST", "1) Update As-of Date on Master row to the current month.  2) Fill in Start Date and Pmt Day (yellow cells) for accurate scheduling.  "
                                  "3) Spot-check 2-3 loans: open amort tab, verify scheduled monthly P&I split looks right vs. lender paperwork."),
    ]
    for i, (label, body) in enumerate(rows):
        r = 3 + i * 2
        ws[f"A{r}"] = label
        ws[f"B{r}"] = body
        fmt_cell(ws[f"A{r}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        fmt_cell(ws[f"B{r}"], font=BLACK_FORMULA, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[r].height = 45


def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    AS_OF = "2026-03-31"  # last day of month per existing 'March 2026 FINAL' file

    # 1) README first (becomes index 0)
    build_readme(wb)

    # 2) Master
    sheet_names, _total_row = build_master(wb, AS_OF)

    # 3) Amort tabs
    for i, (equip, desc, lender, orig, curr, pmt, term, rate, notes) in enumerate(LOANS):
        master_row = 4 + i
        tab = sheet_names[i][1]
        build_amort_tab(wb, equip, desc, lender, orig, curr, pmt, term, rate,
                        master_row, tab)

    # 4) _TEMPLATE_AMORT (clone of one with a clear marker)
    # We'll build a generic template that references Master row 4 — user copies and edits row ref
    template_ws = wb.create_sheet("_TEMPLATE_AMORT")
    template_ws["A1"] = "TEMPLATE — copy this tab to add a new loan. Update Master row refs in B2:B11."
    fmt_cell(template_ws["A1"], font=TITLE_FONT, fill=PatternFill("solid", start_color="C00000"), align=CENTER)
    template_ws.column_dimensions["A"].width = 80

    # 5) Archive
    arch = wb.create_sheet("Archive")
    arch["A1"] = "PAID-OFF / RETIRED LOANS — move tabs here when paid off"
    fmt_cell(arch["A1"], font=TITLE_FONT, fill=PatternFill("solid", start_color="595959"), align=CENTER)
    arch.column_dimensions["A"].width = 80

    out_path = "/sessions/gifted-zealous-ptolemy/mnt/Automate Concrete Business/debt-schedule/PC_Equipment_Debt_Schedule_Automated.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
