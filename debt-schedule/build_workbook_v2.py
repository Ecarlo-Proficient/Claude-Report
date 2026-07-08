"""Build PC Equipment Debt Schedule (master-driven amortization workbook)."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

# ---------- Styling ----------
FONT_NAME = "Arial"
BLUE_INPUT = Font(name=FONT_NAME, color="0000FF", bold=False)
BLACK_FORMULA = Font(name=FONT_NAME, color="000000")
GREEN_LINK = Font(name=FONT_NAME, color="006100")  # link from other sheet
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
SUBHEADER_FONT = Font(name=FONT_NAME, color="000000", bold=True, size=11)
TITLE_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=14)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")  # dark blue
SUB_FILL = PatternFill("solid", start_color="D9E1F2")  # light blue
YELLOW_FILL = PatternFill("solid", start_color="FFFF00")  # needs input (terms)
ORANGE_FILL = PatternFill("solid", start_color="FCE4D6")  # missing account # (QBO lookup)
LA_FILL = PatternFill("solid", start_color="E2EFDA")       # L&A Holdings company tag
GREY_FILL = PatternFill("solid", start_color="F2F2F2")
TOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

CURRENCY_FMT = '_-"$"* #,##0.00_-;[Red]_-"$"* (#,##0.00)_-;_-"$"* "-"??_-;_-@_-'
NUMBER_FMT = '#,##0.00;[Red](#,##0.00);"-"'
PCT_FMT = "0.0000%"
DATE_FMT = "m/d/yyyy"
INT_FMT = "0"

# ---------- Loan seed data ----------
# Merged 2026-06-23 from 'Copy of Monthly payments.xlsx' (current balances, payments,
# lender account numbers, terms) joined to prior seed (original amounts + known rates).
# Fields: company, equip, desc, lender, acct_no, original, current, payment, term, rate, status, notes
#   original=None  -> not known, user must enter (highlighted yellow)
#   rate=None      -> computed via Excel RATE()
#   acct_no=""     -> missing, flagged orange (needed to look up QBO transactions)
LOANS = [
    # ── Proficient Concrete ──
    ("Proficient", "PUMP #013", "2020 MACK Granite", "Alliance Funding Group", "22-15181", 190000.00, 30408.83, 3769.69, 60, None, "Active", "Yes, 60 months, Pump #013"),
    ("Proficient", "PUMP #018", "2019 PUTZMEISTER", "First Citizen Bank", "097-0210991-000", 140000.00, 76506.68, 3467.07, 48, None, "Active", "Yes, 48 Months - Pump #018"),
    ("Proficient", "PUMP #014", "2023 FRHT", "Frost Leasing", "776061089-005", 422200.00, 118058.64, 8432.76, 60, 7.3689, "Active", "Yes, 60 months - Pump #014"),
    ("Proficient", "PUMP #015", "2023 MACK GR85-F", "Frost Leasing", "776061089-007", 620000.00, 287540.56, 12501.75, 60, 7.7677, "Active", "Yes, 60 months - Pump #015"),
    ("Proficient", "PUMP #017", "2023 MACK TE86", "Frost Leasing", "776061089-008", 863327.54, 450986.84, 13744.69, 84, 8.6731, "Active", "Yes, 84 months-Pump #017"),
    ("Proficient", "SOMERO #205,#206", "Somero Laser system", "Frost Leasing", "776061089-006", 313585.58, 77192.90, 7719.29, 48, 8.4377, "Active", "Yes, 48 months - #205,#206"),
    ("Proficient", "SOMERO #219", "Somero 2024 S15R", "Frost Leasing", "776061089-009", 266414.42, 184060.02, 5413.53, 60, 8.0957, "Active", "Yes, 60 months- Somero #219"),
    ("Proficient", "Pump #019", "2021 FREIGHTLINER", "M2 Equipment Finance", "25925-1", 415000.00, 391819.23, 7682.73, 72, 9.97, "Active", "Yes, 72 months, Pump #019"),
    ("Proficient", "Excavator #196", "", "Wells Fargo", "9843925-017", None, 3.40, 636.48, 60, None, "Active", "Yes, (60 months) Excavator #196 — NEW: enter original"),
    ("Proficient", "Tractor #201", "2022 Bobcat T66", "Wells Fargo", "9843925-022", 79503.03, 14287.90, 1428.79, 60, 2.90, "Active", "Yes, 60 months, Tractor #201"),
    ("Proficient", "Tractor #204", "2022 Bobcat E42 T4", "Wells Fargo", "9843925-024", 70942.68, 28773.68, 1106.68, 72, 3.90, "Active", "Yes, 72 months, Tractor #204 (a.k.a. Excavator #204)"),
    ("Proficient", "Tractors #208,#209, #210", "2023 Bobcat T770, T66, & 30C Auger", "Wells Fargo", "9843925-025", 177866.49, 24703.70, 4940.74, 36, 0.00, "Active", "Yes, 36 months, Tractors #208, #209 #210 Auger"),
    ("Proficient", "Tractor #218", "2024 Bobcat T66", "Wells Fargo", "9843925-026", 66206.72, 30344.66, 1379.31, 48, 0.00, "Active", "Tractor #218"),
    ("Proficient", "Tractor #220/221", "2024 Bobcat T66 & sweeper", "Wells Fargo", "9843925-027", 80395.90, 52008.68, 1405.64, 60, 1.90, "Active", "Yes, 60 months"),
    ("Proficient", "Tractor #222", "2024 Bobcat T66", "Wells Fargo", "9843925-028", 64014.29, 38675.32, 1333.63, 48, 0.00, "Active", "Yes, 48 months"),
    ("Proficient", "Tractor #223", "Bobcat (M&M Concrete)", "Wells Fargo", "9843925-029", None, 52855.32, 1258.46, 60, None, "Active", "Yes, 60 months, Tractor #223 — NEW: enter original"),
    ("Proficient", "Tractor #203", "", "Leaf Capital (Lease Services)", "100-4018008-004", None, 0.00, 930.34, 48, None, "Active", "Yes, 48 months, Tractor #203 — NEW: enter original"),
    ("Proficient", "SAW #207", "SAW - HUSQ", "Leaf Capital (Lease Services)", "100-4018008-006", 37978.43, 0.00, 1235.06, 36, None, "Active", "Yes, 36 months"),
    ("Proficient", "Truck #090", "", "Leaf Capital (Lease Services)", "100-4018008-005", None, 0.00, 1623.34, 48, None, "Paid", "Yes, 48 months, Truck #090 (PAID)"),
    ("Proficient", "Truck #093", "", "Leaf Capital (Lease Services)", "100-4018008-007", None, 0.00, 1581.41, 36, None, "Paid", "Yes, 36 months (PAID)"),
    ("Proficient", "Truck #096", "2021 RAM 350", "Leaf Capital (Lease Services)", "100-4018008-008", 51357.03, 6684.64, 1671.16, 36, None, "Active", "Yes, 36 months"),
    ("Proficient", "Truck #097", "2015 RAM 3500", "Leaf Capital (Lease Services)", "100-4018008-010", 36425.86, 13471.04, 1224.64, 36, None, "Active", "Yes, 36 months"),
    ("Proficient", "Equipment #211-#217", "Equipment #211-#217", "Leaf Capital (Lease Services)", "100-4018008-009", 141133.60, 91449.32, 3062.60, 60, None, "Active", "Yes, 60 months - Equipment #211 - #217"),
    ("Proficient", "TRUCK #089", "2019 Ram 3500", "Ally Financial", "228-1220-75707", 52863.14, 2575.52, 888.66, 75, 7.64, "Paid", "Yes, 75 months, Truck #089 (PAID)"),
    ("Proficient", "TRUCK #098", "2024 Toyota Sequoia", "Ally Financial", "228-3765-12198", 111266.64, 74177.76, 1545.37, 72, 8.14, "Active", "Yes, 72 months, Truck #098"),
    ("Proficient", "Marine Loan", "Francisco Escobar — Marine", "Bank of the West (BMO)", "354182825", None, 54961.93, 572.07, None, None, "Active", "Francisco Escobar-Marine Loan — NEW: enter original, term, rate"),
    # ── L&A Holdings, LP (separate QBO file) ──
    ("L&A Holdings", "Property — CF Hawn Office", "10702-10704 CF Hawn Fwy", "American National Bank Texas", "88156922", None, 37885.00, 2440.15, None, None, "Active", "Contract date: October 20, 2014 — NEW: enter original, term, rate"),
    ("L&A Holdings", "Property — Balch Springs", "3116 Balch Springs Rd", "American National Bank Texas", "88232129", None, 72728.80, 1959.18, None, None, "Active", "3116 Balch Springs Rd. — NEW: enter original, term, rate"),
]

# Receivables (preserved from existing) — (equip#, customer, orig, current, payment)
RECEIVABLES = {
    "PUMP #013": ("MCP - Pump #013", 84886.83, 84886.83, 4364.11),
    "TRUCK #089": ("Escobar Concrete #089", 51563.78, 51563.78, 954.88),
    "PUMP #014": ("Core Concrete #014 | MCP - Pump #014 -", 216456.84, 216456.84, 12740.12),
    "PUMP #015": ("MCP - Pump #015", 337547.25, 337547.25, 12501.75),
    "PUMP #017": ("Core Concrete #017", 640000.00, 640000.00, 10877.27),
    "SOMERO #219": ("JCP Concrete SOMERO #219-C/15", 340000.00, 340000.00, 11056.08),
    "Pump #019": ("MCP - Pump #019", 288168.32, 288168.32, 7926.98),
}

# Receivable-only rows (no loan side)
RECEIVABLE_ONLY = [
    ("Equipment ##010", "Erick Martinez-Pump #010-C/15", 59703.65, 59703.65, 2323.67),
    ("Equipment ##011", "Core Concrete #011", 60000.00, 60000.00, 1837.19),
    ("Equipment ##090", "Escobar Concrete #090", 58289.48, 58289.48, 1079.43),
    ("Equipment ##146", "Escobar Conctete #146", 3148.25, 3148.25, 524.71),
    ("Equipment ##157", "Escobar Concrete #157", 3148.25, 3148.25, 524.71),
    ("Equipment ##159", "Escobar Concrete #159", 3148.25, 3148.25, 524.71),
    ("Equipment ##160", "Escobar Concrete #160", 3148.25, 3148.25, 524.71),
    ("Equipment ##179", "Escobar Concrete #179", 28023.77, 28023.77, 518.96),
    ("Equipment ##183", "Escobar Concrete #183", 28023.77, 28023.77, 518.96),
    ("Equipment ##088", "Carlos Alvarez-Dump Truck #088", 80000.00, 80000.00, 1246.90),
    ("Equipment ##003", "Javier Alvarez-Dump Truck #003", 40000.00, 40000.00, 976.52),
]

# Non-loan monthly trackers (combined tab, sortable by Type). Merged from
# 'Copy of Monthly payments.xlsx' 2026-06-23. No amortization — these are
# month-to-month expenses / receivables.
# Fields: type, company, name, memo/acct#, balance, payment, months[12 Jan-Dec], beginning_balance, notes
MONTHLY = [
    # ── Utilities ──
    ("Utilities", "Proficient", "Dallas Water", "Water bill", None, 250.0, [723.33, 1173.83, 241.89, None, None, None, None, None, None, None, None, None], 0.0, "No CONTRACT, month to month payment Water bill for office/ yard"),
    ("Utilities", "Proficient", "AT&T Mobility", "", 0.0, 1500.0, [2381.42, 2372.22, 2455.22, 2818.51, None, None, None, None, None, None, None, None], 0.0, "Proficient phones"),
    ("Utilities", "Proficient", "AT&T (U-VERSE)", "Account # : 140347133-5", 0.0, 86.02, [112.9, 112.9, 116.12, 116.12, 116.12, None, None, None, None, None, None, None], 0.0, "No CONTRACT, month to month payment Interent"),
    ("Utilities", "Proficient", "ATMOS ENERGY", "Account #: 4009102149", 0.0, 100.0, [133.64, 210.78, 109.13, 112.89, None, None, None, None, None, None, None, None], 0.0, "No CONTRACT, month to month payment Office gas"),
    ("Utilities", "Proficient", "TXU Energy", "", 0.0, 450.0, [744.4, 816.68, 846.8, 699.98, 699.96, None, None, None, None, None, None, None], 0.0, "Office electricity"),
    ("Utilities", "L&A Holdings", "L & A Holdings, LP", "Rent for 10702 CF HAWN", None, 3000.0, [3000.0, 3000.0, 3000.0, 3000.0, None, None, None, None, None, None, None, None], None, "Rent for 10702 CF Hawn"),
    ("Utilities", "Proficient", "Hernandez Cleaning- office and shop", "Cleaning Office and shop", None, 1141.0, [5705.0, 4564.0, 4636.0, 4652.0, None, None, None, None, None, None, None, None], None, ""),
    # ── Insurance ──
    ("Insurance", "Proficient", "Texas Mutual Insurance Company", "Policy# 0001222475", 0.0, None, [None, None, None, None, None, None, None, None, None, None, None, None], None, "Workers Comp. monthly payment"),
    ("Insurance", "Proficient", "Farmers Insurance", "Auto Policy NEW 2025-2026", None, None, [None, None, None, 1701.11, None, None, None, None, None, None, None, None], 21572.0, "Policy #: 60709-88-62"),
    ("Insurance", "Proficient", "Farmers Insurance", "Equipment Policy", 0.0, None, [None, None, None, None, None, None, None, None, None, None, None, None], None, ""),
    ("Insurance", "Proficient", "AFCO-BARNARD DONEGAN INS", "GL-Pollution Policy", -702.7, 712.7, [722.7, 712.7, 712.7, 712.7, 712.7, None, None, None, None, None, None, None], 4276.2, ""),
    ("Insurance", "Proficient", "AFCO-BARNARD DONEGAN INS", "GL & UMBRELLA", None, 17639.97, [17717.72, 17717.72, 17717.72, 17717.72, 17717.72, None, None, None, None, None, None, None], 158759.73, ""),
    ("Insurance", "Proficient", "First Insurance", "Property Insurance | 106995228", None, None, [None, None, None, None, None, 448.78, None, None, None, None, None, None], None, ""),
    # ── Life Insurance ──
    ("Life Insurance", "Proficient", "FNWL", "", None, 73.51, [73.51, 73.51, 73.51, 73.51, None, None, None, None, None, None, None, None], None, "Term insurance for Araceli Perez"),
    ("Life Insurance", "Proficient", "Transamerica Life Insurance", "Policy # 42872171", None, 900.0, [900.0, None, None, None, None, None, None, None, None, None, None, None], None, "Term Life Insurance Luis Perez"),
    ("Life Insurance", "Proficient", "Ohio National Insurance. (Luis Perez)", "Life Insurance", None, 176.89, [176.89, 176.89, 176.89, 176.89, None, None, None, None, None, None, None, None], None, ""),
    # ── Other Notes Payable (add rows here: legal settlements, misc notes payable) ──
    ("Other Notes Payable", "Proficient", "", "", None, None, [None]*12, None, ""),
    # ── Receivable ──
    ("Receivable", "Proficient", "JCP Concrete SOMERO #219-C/15", "", 541747.92, 11056.08, [11056.08, 11056.08, None, None, None, None, None, None, None, None, None, None], 563860.08, ""),
    ("Receivable", "Proficient", "Erick Martinez-Pump #010-C/15", "", 50007.65, 2323.67, [2424.0, 2424.0, 2424.0, 2424.0, None, None, None, None, None, None, None, None], 59703.65, ""),
    ("Receivable", "L&A Holdings", "Core Concrete #011 w/ L&A", "", 60000.0, 1837.19, [None, None, None, None, None, None, None, None, None, None, None, None], 60000.0, ""),
    ("Receivable", "Proficient", "Core Concrete #014", "", 140000.0, 4307.36, [None, None, None, None, None, None, None, None, None, None, None, None], 140000.0, ""),
    ("Receivable", "Proficient", "Core Concrete #017", "", 640000.0, 10877.27, [None, None, None, None, None, None, None, None, None, None, None, None], 640000.0, ""),
    ("Receivable", "Proficient", "Core Concrete #015", "", None, 5080.6, [None, None, None, None, None, None, None, None, None, None, None, None], 300000.0, ""),
    ("Receivable", "Proficient", "Core Concrete #019", "", None, 7041.24, [None, None, None, None, None, None, None, None, None, None, None, None], 415000.0, ""),
    ("Receivable", "Proficient", "Escobar Concrete  #089", "", 47744.26, 954.88, [954.88, 954.88, 954.88, 954.88, None, None, None, None, None, None, None, None], 51563.78, ""),
    ("Receivable", "Proficient", "Escobar Concrete  #090", "", 55051.19, 1079.43, [1079.43, 1079.43, 1079.43, None, None, None, None, None, None, None, None, None], 58289.48, ""),
    ("Receivable", "Proficient", "Escobar Conctete #146", "", 1049.41, 524.71, [524.71, 524.71, 524.71, 524.71, None, None, None, None, None, None, None, None], 3148.25, ""),
    ("Receivable", "Proficient", "Escobar Concrete #157", "", 1049.41, 524.71, [524.71, 524.71, 524.71, 524.71, None, None, None, None, None, None, None, None], 3148.25, ""),
    ("Receivable", "Proficient", "Escobar Concrete #159", "", 1049.41, 524.71, [524.71, 524.71, 524.71, 524.71, None, None, None, None, None, None, None, None], 3148.25, ""),
    ("Receivable", "Proficient", "Escobar Concrete #160", "", 1049.41, 524.71, [524.71, 524.71, 524.71, 524.71, None, None, None, None, None, None, None, None], 3148.25, ""),
    ("Receivable", "Proficient", "Escobar Concrete #179", "", 25947.93, 518.96, [518.96, 518.96, 518.96, 518.96, None, None, None, None, None, None, None, None], 28023.77, ""),
    ("Receivable", "Proficient", "Escobar Concrete #183", "", 25947.93, 518.96, [518.96, 518.96, 518.96, 518.96, None, None, None, None, None, None, None, None], 28023.77, ""),
    ("Receivable", "Proficient", "Carlos Alvarez-Dump Truck #088", "", 75006.2, 1246.9, [None, 1250.0, 1250.0, 1246.9, 1246.9, None, None, None, None, None, None, None], 80000.0, ""),
    ("Receivable", "Proficient", "Javier Alvarez-Dump Truck #003", "", 38046.96, 976.52, [None, 976.52, 976.52, None, None, None, None, None, None, None, None, None], 40000.0, ""),
    ("Receivable", "Proficient", "MCP - Pump #013 - STOPPED BILLING PER LUIS 5/4", "", 67430.39, 4364.11, [4364.11, 4364.11, 4364.11, 4364.11, None, None, None, None, None, None, None, None], 84886.83, ""),
    ("Receivable", "Proficient", "MCP - Pump #014 - SOLD TO CORE", "", 59591.32, 8432.76, [8432.76, 8432.76, None, None, None, None, None, None, None, None, None, None], 76456.84, ""),
    ("Receivable", "Proficient", "MCP - Pump #015 SOLD TO CORE", "", 287540.25, 12501.75, [12501.75, 12501.75, 12501.75, 12501.75, None, None, None, None, None, None, None, None], 337547.25, ""),
    ("Receivable", "Proficient", "MCP - Pump #019 SOLD TO CORE", "", 256460.4, 7926.98, [7926.98, 7926.98, 7926.98, 7926.98, None, None, None, None, None, None, None, None], 288168.32, ""),
]


def sanitize_sheet_name(name: str) -> str:
    """Excel sheet name: <=31 chars, no []:*?/\\, no leading/trailing apostrophe."""
    clean = re.sub(r"[\[\]\:\*\?/\\]", "", name)
    clean = clean.replace("—", "-").replace("–", "-")  # em/en dash -> hyphen
    clean = clean.replace("#", "").replace(",", "_").replace(" ", "_").replace("/", "_")
    clean = re.sub(r"_+", "_", clean).strip("_")
    return clean[:31]


def fmt_cell(cell, font=None, fill=None, fmt=None, align=None, border=BORDER):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def build_master(wb, as_of_date_str):
    ws = wb.create_sheet("Master Debt Schedule")
    # Column widths  (A=Company .. R=Amort Tab, S=spacer, T..W=Receivables)
    widths = {
        "A": 14, "B": 22, "C": 26, "D": 26, "E": 18, "F": 16, "G": 16,
        "H": 12, "I": 14, "J": 9, "K": 11, "L": 12, "M": 7, "N": 11,
        "O": 14, "P": 14, "Q": 38, "R": 22, "S": 2, "T": 30, "U": 16,
        "V": 16, "W": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title row
    ws.merge_cells("A1:W1")
    ws["A1"] = "EQUIPMENT DEBT & RECEIVABLE SCHEDULE — Master (single source of truth)"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 26

    # Section banners
    ws.merge_cells("A2:B2"); ws["A2"] = "Company / Equipment"
    ws.merge_cells("C2:R2"); ws["C2"] = "Loan Information"
    ws.merge_cells("T2:W2"); ws["T2"] = "Receivable Information"
    for c in ("A2", "C2", "T2"):
        fmt_cell(ws[c], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    # Header row 3
    headers = {
        "A": "Company",
        "B": "Equipment #",
        "C": "Description",
        "D": "Lender",
        "E": "Account #",
        "F": "Original Loan Balance",
        "G": "Current Balance",
        "H": "As-of Date",
        "I": "Monthly Payment",
        "J": "Term (mo)",
        "K": "Annual Rate %",
        "L": "Start Date",
        "M": "Pmt Day",
        "N": "Status",
        "O": "YTD Principal",
        "P": "YTD Interest",
        "Q": "Notes (original Terms text)",
        "R": "Amort Tab",
        "T": "Customer",
        "U": "Original Receivable",
        "V": "Current Receivable",
        "W": "Receivable Payment",
    }
    for col, text in headers.items():
        c = ws[f"{col}3"]
        c.value = text
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "C4"

    # Data rows
    row = 4
    sheet_names = []
    for (company, equip, desc, lender, acct_no, orig, curr, pmt, term, rate, status, notes) in LOANS:
        ws[f"A{row}"] = company
        ws[f"B{row}"] = equip
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = lender
        ws[f"E{row}"] = acct_no
        ws[f"F{row}"] = orig                 # None -> blank (input needed)
        ws[f"G{row}"] = curr
        ws[f"H{row}"] = as_of_date_str
        ws[f"I{row}"] = pmt
        ws[f"J{row}"] = term                 # None -> blank (input needed)
        # Rate: hardcode if known; else compute fully-amortizing rate via RATE() (FV=0)
        if rate is None:
            ws[f"K{row}"] = f"=IFERROR(RATE(J{row},-I{row},F{row})*12,0)"
        else:
            ws[f"K{row}"] = rate / 100.0
        ws[f"L{row}"] = None                 # Start Date — user input (yellow)
        ws[f"M{row}"] = 1                     # Pmt Day default
        ws[f"N{row}"] = status
        tab = sanitize_sheet_name(equip)
        sheet_names.append((equip, tab))
        ws[f"O{row}"] = f"=IFERROR(INDIRECT(\"'\"&R{row}&\"'!P5\"),0)"
        ws[f"P{row}"] = f"=IFERROR(INDIRECT(\"'\"&R{row}&\"'!Q5\"),0)"
        ws[f"Q{row}"] = notes
        ws[f"R{row}"] = tab

        # Receivable (if exists)
        if equip in RECEIVABLES:
            cust, r_orig, r_curr, r_pmt = RECEIVABLES[equip]
            ws[f"T{row}"] = cust
            ws[f"U{row}"] = r_orig
            ws[f"V{row}"] = r_curr
            ws[f"W{row}"] = r_pmt

        # ── Formatting ──
        company_fill = LA_FILL if company == "L&A Holdings" else None
        fmt_cell(ws[f"A{row}"], font=SUBHEADER_FONT, align=CENTER, fill=company_fill)
        fmt_cell(ws[f"B{row}"], font=BLACK_FORMULA, align=LEFT)
        fmt_cell(ws[f"C{row}"], font=BLUE_INPUT, align=LEFT)
        fmt_cell(ws[f"D{row}"], font=BLUE_INPUT, align=LEFT)
        # Account # — orange if missing (cannot look up QBO transactions without it)
        fmt_cell(ws[f"E{row}"], font=BLUE_INPUT, align=LEFT,
                 fill=(ORANGE_FILL if not acct_no else None))
        # Original — yellow if missing (needed for amortization)
        fmt_cell(ws[f"F{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT,
                 fill=(YELLOW_FILL if orig is None else None))
        fmt_cell(ws[f"G{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"H{row}"], font=BLUE_INPUT, fmt=DATE_FMT, align=CENTER, fill=YELLOW_FILL)
        fmt_cell(ws[f"I{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        # Term — yellow if missing (needed for amortization)
        fmt_cell(ws[f"J{row}"], font=BLUE_INPUT, fmt=INT_FMT, align=CENTER,
                 fill=(YELLOW_FILL if term is None else None))
        if rate is None:
            fmt_cell(ws[f"K{row}"], font=BLACK_FORMULA, fmt=PCT_FMT, align=CENTER, fill=GREY_FILL)
        else:
            fmt_cell(ws[f"K{row}"], font=BLUE_INPUT, fmt=PCT_FMT, align=CENTER)
        fmt_cell(ws[f"L{row}"], font=BLUE_INPUT, fmt=DATE_FMT, align=CENTER, fill=YELLOW_FILL)
        fmt_cell(ws[f"M{row}"], font=BLUE_INPUT, fmt=INT_FMT, align=CENTER)
        fmt_cell(ws[f"N{row}"], font=BLUE_INPUT, align=CENTER,
                 fill=(GREY_FILL if status == "Paid" else None))
        fmt_cell(ws[f"O{row}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"P{row}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"Q{row}"], font=BLACK_FORMULA, align=LEFT)
        fmt_cell(ws[f"R{row}"], font=BLACK_FORMULA, align=LEFT, fill=GREY_FILL)
        for col in ("T", "U", "V", "W"):
            fmt = CURRENCY_FMT if col != "T" else None
            fmt_cell(ws[f"{col}{row}"], font=BLUE_INPUT, fmt=fmt,
                     align=RIGHT if col != "T" else LEFT)
        row += 1

    # Receivable-only rows
    for (equip, cust, r_orig, r_curr, r_pmt) in RECEIVABLE_ONLY:
        ws[f"A{row}"] = "Proficient"
        ws[f"B{row}"] = equip
        ws[f"N{row}"] = "Receivable Only"
        ws[f"T{row}"] = cust
        ws[f"U{row}"] = r_orig
        ws[f"V{row}"] = r_curr
        ws[f"W{row}"] = r_pmt
        fmt_cell(ws[f"A{row}"], font=BLACK_FORMULA, align=CENTER, fill=GREY_FILL)
        fmt_cell(ws[f"B{row}"], font=BLACK_FORMULA, align=LEFT, fill=GREY_FILL)
        for col in "CDEFGHIJKLMOPQR":
            fmt_cell(ws[f"{col}{row}"], font=BLACK_FORMULA, fill=GREY_FILL)
        fmt_cell(ws[f"N{row}"], font=BLACK_FORMULA, align=CENTER, fill=GREY_FILL)
        fmt_cell(ws[f"T{row}"], font=BLUE_INPUT, align=LEFT)
        for col in ("U", "V", "W"):
            fmt_cell(ws[f"{col}{row}"], font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        row += 1

    # Totals row
    total_row = row
    ws[f"A{total_row}"] = "TOTALS"
    fmt_cell(ws[f"A{total_row}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for col in ("F", "G", "I", "O", "P", "U", "V", "W"):
        ws[f"{col}{total_row}"] = f"=SUM({col}4:{col}{total_row-1})"
        fmt_cell(ws[f"{col}{total_row}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, fmt=CURRENCY_FMT, align=RIGHT)
    for col in "BCDEHJKLMNQRT":
        fmt_cell(ws[f"{col}{total_row}"], fill=TOTAL_FILL)

    return sheet_names, total_row


def build_amort_tab(wb, equip, desc, lender, orig, curr, pmt, term, rate_pct,
                     master_row, tab_name, max_term=120):
    """Build one amort tab. References Master row for live inputs.
    Schedule is origination-based: starts at (Effective) Start Date with Original Balance,
    runs Term months. If user leaves Start Date blank on Master, an implied start is
    computed by NPER from current balance back to origination.
    """
    ws = wb.create_sheet(tab_name)

    # Column widths — compromise between horizontal header section and schedule section.
    # Description/Lender cells in row 5 will wrap text since the row is tall.
    widths = {
        "A": 13, "B": 22, "C": 22, "D": 16, "E": 16, "F": 14, "G": 16,
        "H": 9,  "I": 11, "J": 14, "K": 11, "L": 12, "M": 11, "N": 13,
        "O": 9,  "P": 15, "Q": 15,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    M = master_row
    MASTER = "'Master Debt Schedule'"

    # ---------- ROW 1: Title (merged A:Q) ----------
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"AMORTIZATION SCHEDULE — {equip}"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 24

    # ---------- ROW 3: Loan info banner (merged) ----------
    ws.merge_cells("A3:Q3")
    ws["A3"] = "LOAN INFORMATION & KPIs — pulled live from Master. YTD cols (P/Q) reflect current calendar year."
    fmt_cell(ws["A3"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    # ---------- ROW 4: Column headers (horizontal) ----------
    headers = [
        "Equipment #", "Description", "Lender", "Original Balance",
        "Current Balance", "As-of Date", "Monthly Payment", "Term (mo)",
        "Annual Rate %", "Start Date", "Pmt Day", "Monthly Rate",
        "Implied Mo Made", "Effective Start", "Current Year",
        "YTD Principal", "YTD Interest",
    ]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}4"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[4].height = 32

    # ---------- ROW 5: Values (pulled from Master — column map updated for v2 layout) ----------
    # Master v2 cols: B=Equip C=Desc D=Lender F=Original G=Current H=As-of
    #                 I=Payment J=Term K=Rate L=Start M=Pmt Day
    ws["A5"] = f"={MASTER}!B{M}"   # Equipment #
    ws["B5"] = f"={MASTER}!C{M}"   # Description
    ws["C5"] = f"={MASTER}!D{M}"   # Lender
    ws["D5"] = f"={MASTER}!F{M}"   # Original Balance
    ws["E5"] = f"={MASTER}!G{M}"   # Current Balance
    ws["F5"] = f"={MASTER}!H{M}"   # As-of Date
    ws["G5"] = f"={MASTER}!I{M}"   # Monthly Payment
    ws["H5"] = f"={MASTER}!J{M}"   # Term
    ws["I5"] = f"={MASTER}!K{M}"   # Annual Rate %
    # Start Date — blank-safe so empty Master cell doesn't show 1/0/1900
    ws["J5"] = f'=IF({MASTER}!L{M}="","",{MASTER}!L{M})'   # Start Date
    ws["K5"] = f"={MASTER}!M{M}"   # Pmt Day
    ws["L5"] = "=I5/12"
    ws["M5"] = "=IFERROR(ROUND(IF(I5=0,IF(G5=0,0,(D5-E5)/G5),NPER(L5,-G5,D5,-E5)),0),0)"
    ws["N5"] = '=IFERROR(IF(AND(ISNUMBER(J5),J5>0),J5,EDATE(F5,-M5)),"")'
    ws["O5"] = "=YEAR(TODAY())"
    # P5 / Q5 filled after tracker

    fmt_cell(ws["A5"], font=GREEN_LINK, align=LEFT)
    fmt_cell(ws["B5"], font=GREEN_LINK,
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    fmt_cell(ws["C5"], font=GREEN_LINK,
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    fmt_cell(ws["D5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["E5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["F5"], font=GREEN_LINK, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["G5"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT)
    fmt_cell(ws["H5"], font=GREEN_LINK, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["I5"], font=GREEN_LINK, fmt=PCT_FMT, align=CENTER)
    fmt_cell(ws["J5"], font=GREEN_LINK, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["K5"], font=GREEN_LINK, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["L5"], font=BLACK_FORMULA, fmt=PCT_FMT, align=CENTER)
    fmt_cell(ws["M5"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["N5"], font=BLACK_FORMULA, fmt=DATE_FMT, align=CENTER)
    fmt_cell(ws["O5"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
    fmt_cell(ws["P5"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=TOTAL_FILL)
    fmt_cell(ws["Q5"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=TOTAL_FILL)
    ws.row_dimensions[5].height = 36

    # ---------- PAYMENT TRACKER (rows 7-21) ----------
    tracker_banner = 7
    tracker_header_row = 8
    tracker_start = 9
    tracker_years = 12
    tracker_end = tracker_start + tracker_years - 1  # 20
    tracker_totals = tracker_end + 1  # 21

    ws.merge_cells(f"A{tracker_banner}:P{tracker_banner}")
    ws[f"A{tracker_banner}"] = "PAYMENT TRACKER — enter $ paid each month when bank drafts"
    fmt_cell(ws[f"A{tracker_banner}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    tracker_headers = ["Year"] + ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"] + \
                       ["Total Paid", "Principal", "Interest"]
    for i, h in enumerate(tracker_headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}{tracker_header_row}"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    # Layout constants for schedule (defined here so tracker auto-pop formulas can ref them)
    sched_banner = tracker_totals + 2  # 23
    sched_header = sched_banner + 1     # 24
    sched_start = sched_header + 1      # 25
    sched_end = sched_start + max_term - 1  # 144

    # Tracker rows: Year col + auto-populate month cells + Total col
    # Month cells AUTO-FILL with scheduled payment when DATE(year, month, Pmt Day) <= TODAY().
    # SUMIFS pulls Sched Pmt (col D) from the amort schedule row matching this year+month.
    # To override, just type a number — it replaces the formula.
    for i in range(tracker_years):
        r = tracker_start + i
        ws[f"A{r}"] = f'=IFERROR(YEAR($N$5)+{i},"")'
        fmt_cell(ws[f"A{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
        for mc in range(12):
            col = get_column_letter(2 + mc)
            month_idx = mc + 1  # 1=Jan, ..., 12=Dec
            cell = ws[f"{col}{r}"]
            cell.value = (
                f'=IFERROR(IF(AND(ISNUMBER($A{r}),'
                f'DATE($A{r},{month_idx},$K$5)<=TODAY()),'
                f'SUMIFS(D{sched_start}:D{sched_end},'
                f'H{sched_start}:H{sched_end},$A{r},'
                f'I{sched_start}:I{sched_end},{month_idx}),'
                f'0),0)'
            )
            # Blue font signals "you can type here to override the auto-fill"
            fmt_cell(cell, font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        ws[f"N{r}"] = f"=SUM(B{r}:M{r})"
        fmt_cell(ws[f"N{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- AMORT SCHEDULE ----------

    ws.merge_cells(f"A{sched_banner}:L{sched_banner}")
    ws[f"A{sched_banner}"] = "AMORT SCHEDULE (origination-based, runs full Term)"
    fmt_cell(ws[f"A{sched_banner}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=CENTER)

    sched_headers = ["Pmt #", "Date", "Beg Balance", "Scheduled Pmt",
                     "Sched Principal", "Sched Interest", "End Balance",
                     "Year", "Month", "Actual Paid", "Eff Principal", "Eff Interest"]
    for i, h in enumerate(sched_headers):
        col = get_column_letter(i + 1)
        c = ws[f"{col}{sched_header}"]
        c.value = h
        fmt_cell(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    n_rows = max_term
    for i in range(n_rows):
        r = sched_start + i
        pmt_n = i + 1
        ws[f"A{r}"] = pmt_n
        ws[f"B{r}"] = (f'=IFERROR(IF(OR(A{r}>$H$5,$N$5="",$N$5=0),"",'
                       f'DATE(YEAR(EDATE($N$5,A{r}-1)),'
                       f'MONTH(EDATE($N$5,A{r}-1)),'
                       f'MIN($K$5,DAY(EOMONTH(EDATE($N$5,A{r}-1),0))))),"")')
        if i == 0:
            ws[f"C{r}"] = f'=IF(B{r}="","",$D$5)'
        else:
            ws[f"C{r}"] = f'=IF(B{r}="","",IFERROR(G{r-1},0))'
        ws[f"F{r}"] = f'=IF(B{r}="","",IFERROR(C{r}*$L$5,0))'
        ws[f"D{r}"] = f'=IF(B{r}="","",IFERROR(MIN($G$5,C{r}+F{r}),0))'
        ws[f"E{r}"] = f'=IF(B{r}="","",IFERROR(D{r}-F{r},0))'
        ws[f"G{r}"] = f'=IF(B{r}="","",IFERROR(MAX(0,C{r}-E{r}),0))'
        ws[f"H{r}"] = f'=IF(B{r}="","",YEAR(B{r}))'
        ws[f"I{r}"] = f'=IF(B{r}="","",MONTH(B{r}))'
        ws[f"J{r}"] = (f'=IFERROR(IF(B{r}="","",'
                       f'INDEX($B${tracker_start}:$M${tracker_end},'
                       f'MATCH(H{r},$A${tracker_start}:$A${tracker_end},0),'
                       f'I{r})),0)')
        ws[f"K{r}"] = f'=IF(B{r}="","",IF(J{r}>0,E{r},0))'
        ws[f"L{r}"] = f'=IF(B{r}="","",IF(J{r}>0,F{r},0))'

        fmt_cell(ws[f"A{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER)
        fmt_cell(ws[f"B{r}"], font=BLACK_FORMULA, fmt=DATE_FMT, align=CENTER)
        for col in ("C", "D", "E", "F", "G"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)
        for col in ("H", "I"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=INT_FMT, align=CENTER, fill=GREY_FILL)
        fmt_cell(ws[f"J{r}"], font=GREEN_LINK, fmt=CURRENCY_FMT, align=RIGHT, fill=GREY_FILL)
        for col in ("K", "L"):
            fmt_cell(ws[f"{col}{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=GREY_FILL)

    # sched_end already defined above; just compute totals row
    sched_totals = sched_end + 1
    ws[f"A{sched_totals}"] = "TOTALS"
    fmt_cell(ws[f"A{sched_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for col in ("D", "E", "F", "J", "K", "L"):
        ws[f"{col}{sched_totals}"] = f"=SUM({col}{sched_start}:{col}{sched_end})"
        fmt_cell(ws[f"{col}{sched_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- Fill Tracker Principal / Interest ----------
    for i in range(tracker_years):
        r = tracker_start + i
        ws[f"O{r}"] = (f'=SUMIFS(K{sched_start}:K{sched_end},'
                       f'H{sched_start}:H{sched_end},A{r})')
        ws[f"P{r}"] = (f'=SUMIFS(L{sched_start}:L{sched_end},'
                       f'H{sched_start}:H{sched_end},A{r})')
        fmt_cell(ws[f"O{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws[f"P{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)

    # Tracker totals row
    ws[f"A{tracker_totals}"] = "TOTALS"
    fmt_cell(ws[f"A{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for mc in range(12):
        col = get_column_letter(2 + mc)
        ws[f"{col}{tracker_totals}"] = f"=SUM({col}{tracker_start}:{col}{tracker_end})"
        fmt_cell(ws[f"{col}{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)
    for col in ("N", "O", "P"):
        ws[f"{col}{tracker_totals}"] = f"=SUM({col}{tracker_start}:{col}{tracker_end})"
        fmt_cell(ws[f"{col}{tracker_totals}"], font=SUBHEADER_FONT, fill=TOTAL_FILL,
                 fmt=CURRENCY_FMT, align=RIGHT)

    # ---------- Populate P5 / Q5 (current-year YTD lookups) ----------
    ws["P5"] = (f'=IFERROR(VLOOKUP($O$5,A{tracker_start}:P{tracker_end},15,FALSE),0)')
    ws["Q5"] = (f'=IFERROR(VLOOKUP($O$5,A{tracker_start}:P{tracker_end},16,FALSE),0)')

    # Freeze the top header section (rows 1-5) so loan info stays visible while scrolling
    ws.freeze_panes = "A6"


def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    ws.merge_cells("A1:B1")
    ws["A1"] = "Equipment Debt Schedule — How this workbook works"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 28

    rows = [
        ("PURPOSE", "Single workbook to manage all equipment loans and their amortizations. "
                    "Master tab is the only place you enter loan info; each amort tab pulls from Master."),
        ("STRUCTURE", "Index = clickable map of every sheet.  Master Debt Schedule = where you enter loan info.  "
                      "Monthly Payments = utilities/insurance/receivables tracker.  One amort tab per loan, named after "
                      "Equipment #.  _TEMPLATE_AMORT = copy this to add a new loan.  Archive = move paid-off tabs here."),
        ("CELL COLORS", "YELLOW fill = an input — you type here.  BLUE text = automated (a formula, a value pulled from "
                        "another sheet, or a balance filled by the QBO sync) — do not type over it.  "
                        "Light-green Company cell = L&A Holdings (separate QuickBooks file)."),
        ("COMPANY COLUMN", "Column A tags each loan Proficient or L&A Holdings.  L&A loans (American National Bank — "
                           "CF Hawn office & Balch Springs) live in a SEPARATE QuickBooks file, so the QBO sync skips them "
                           "by default — keep their balances current manually, or add L&A credentials to sync them too."),
        ("WHERE TERMS GO (amortization)", "On the Master row, fill these for a correct amortization: F=Original Loan Balance, "
                                          "J=Term (months), K=Annual Rate % (leave blank/grey to auto-compute via RATE), "
                                          "L=Start Date, M=Pmt Day.  Current Balance (G) + As-of (H) are kept current by the QBO sync.  "
                                          "Any yellow cell on a row is a term the amortization needs."),
        ("ACCOUNT # (QBO lookup)", "Column E holds each loan's lender account number (from your monthly file).  The QBO sync "
                                   "uses it to help match the loan to its QuickBooks liability account.  Rows with an orange "
                                   "Account # cell need a number entered before they can be looked up."),
        ("ADD A NEW SCHEDULE", "1) On the MASTER tab, add a row at the bottom of the loan list (above TOTALS): fill Company, "
                               "Equipment # (col B), Description, Lender, Account #, Original (F), Payment (I), Term (J), Rate (K, or "
                               "leave blank to auto-compute), Start Date (L), Pmt Day (M).  In 'Amort Tab' (col R) type the new tab name.  "
                               "2) Right-click the _TEMPLATE_AMORT tab → Move or Copy → ✔ 'Create a copy'.  3) On the copy, type that "
                               "same Equipment # into the YELLOW cell A5 — the whole schedule fills in automatically.  4) Rename the copied "
                               "tab to match col R.  No formula editing.  (TIP: the Index tab won't list it until the file is rebuilt, but the "
                               "Master 'Amort Tab' link will work.)"),
        ("MONTHLY WORKFLOW", "When a payment drafts: open that loan's amort tab → PAYMENT TRACKER → type the $ amount in that month's "
                             "YELLOW cell.  It auto-splits to Principal & Interest using the scheduled amort row for that month.  "
                             "(Or let the QBO sync pull actual payments + balances for you.)"),
        ("CLOSE / PAY OFF A SCHEDULE", "When a loan is paid off: 1) On the Master, set Status (col N) to 'Paid'.  The row's amort tab keeps "
                                       "its history.  2) To declutter, move the tab to the end: right-click the amort tab → Move or Copy → "
                                       "move it BEFORE 'Archive' (or just leave it — the Index groups everything for you).  3) The loan stays in "
                                       "Master TOTALS at $0 balance; to drop it from totals, delete the Master row AND its tab."),
        ("THE EXTRA / LEFTOVER TAB", "_TEMPLATE_AMORT is the master copy — NEVER delete it; you copy FROM it each time.  After you copy it "
                                     "to make a new loan, the copy becomes that loan's real tab (rename it).  A paid-off tab you no longer want "
                                     "to see: move it next to 'Archive' (keeps history) or delete it if you've recorded the payoff elsewhere.  "
                                     "Archive is just a parking spot — anything to the right of it is 'retired'."),
        ("RATE COMPUTATION", "Loans where the original Terms/Interest text did not include a rate are computed via Excel's RATE() function "
                              "on Master (cell shows in grey).  Verify these against your loan paperwork."),
        ("ASSUMPTIONS", "Schedule starts from CURRENT balance (as of the As-of Date), not from origination.  "
                        "Remaining months are computed from current balance + payment + rate.  "
                        "Each amort row uses standard PPMT/IPMT math: Interest = Beg Balance × monthly rate; Principal = Payment − Interest."),
        ("WHAT TO CHECK FIRST", "1) Update As-of Date on Master row to the current month.  2) Fill in Start Date and Pmt Day (yellow cells) for accurate scheduling.  "
                                  "3) Spot-check 2-3 loans: open amort tab, verify scheduled monthly P&I split looks right vs. lender paperwork."),
    ]
    for i, (label, body) in enumerate(rows):
        r = 3 + i * 2
        ws[f"A{r}"] = label
        ws[f"B{r}"] = body
        fmt_cell(ws[f"A{r}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        fmt_cell(ws[f"B{r}"], font=BLACK_FORMULA, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[r].height = 45


LINK_FONT = Font(name=FONT_NAME, color="0563C1", underline="single")
TYPE_FILLS = {
    "Utilities": PatternFill("solid", start_color="DDEBF7"),
    "Insurance": PatternFill("solid", start_color="FCE4D6"),
    "Life Insurance": PatternFill("solid", start_color="E2EFDA"),
    "Other Notes Payable": PatternFill("solid", start_color="FFF2CC"),
    "Receivable": PatternFill("solid", start_color="EDEDED"),
}


def _link(target_sheet, label):
    """Excel in-workbook hyperlink formula."""
    safe = str(label).replace('"', "'")
    return f'=HYPERLINK("#\'{target_sheet}\'!A1","{safe}")'


def build_index(wb, loan_tabs, extra_tabs):
    """Navigation hub: clickable links to every sheet, grouped by Company."""
    ws = wb.create_sheet("Index", index=1)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.merge_cells("A1:C1")
    ws["A1"] = "INDEX — click any name to jump to its sheet"
    fmt_cell(ws["A1"], font=TITLE_FONT, fill=HEADER_FILL, align=CENTER, border=None)
    ws.row_dimensions[1].height = 24

    r = 3
    ws[f"A{r}"] = "Key sheets"
    fmt_cell(ws[f"A{r}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=LEFT)
    r += 1
    for sheet, label in extra_tabs:
        ws[f"A{r}"] = _link(sheet, label)
        fmt_cell(ws[f"A{r}"], font=LINK_FONT, align=LEFT)
        r += 1

    # Loans grouped by company
    by_company = {}
    for company, equip, tab in loan_tabs:
        by_company.setdefault(company, []).append((equip, tab))
    for company in sorted(by_company):
        r += 1
        ws[f"A{r}"] = f"{company} — loans ({len(by_company[company])})"
        fmt_cell(ws[f"A{r}"], font=SUBHEADER_FONT, fill=SUB_FILL, align=LEFT)
        fmt_cell(ws[f"B{r}"], fill=SUB_FILL)
        fmt_cell(ws[f"C{r}"], fill=SUB_FILL)
        r += 1
        for equip, tab in by_company[company]:
            ws[f"A{r}"] = _link(tab, equip)
            fmt_cell(ws[f"A{r}"], font=LINK_FONT, align=LEFT)
            # live current balance pulled from that tab
            ws[f"B{r}"] = f"=IFERROR('{tab}'!E5,\"\")"
            fmt_cell(ws[f"B{r}"], font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT)
            ws[f"C{r}"] = "Current balance"
            fmt_cell(ws[f"C{r}"], font=Font(name=FONT_NAME, color="808080", size=9), align=LEFT)
            r += 1
    ws.freeze_panes = "A2"


def build_monthly_tracker(wb):
    """One combined, sortable tracker for all non-loan monthly payments."""
    ws = wb.create_sheet("Monthly Payments", index=3)
    headers = ["Type", "Company", "Name", "Memo / Acct #", "Balance", "Payment",
               "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Total Paid", "Beginning Balance", "Notes"]
    widths = [18, 13, 34, 26, 13, 12] + [10] * 12 + [13, 16, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells("A1:U1")
    ws["A1"] = ('MONTHLY PAYMENTS — utilities, insurance, life insurance, receivables, other notes payable.  '
                'Use the Type filter ▾ to group.  ← back to Index')
    fmt_cell(ws["A1"], font=Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11),
             fill=HEADER_FILL, align=LEFT, border=None)
    ws["V1"] = _link("Index", "← Index")
    fmt_cell(ws["V1"], font=LINK_FONT, align=LEFT)
    ws.row_dimensions[1].height = 22

    hr = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        fmt_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[hr].height = 20

    row = hr + 1
    for (typ, company, name, memo, bal, pmt, months, beg, notes) in MONTHLY:
        ws.cell(row, 1, typ)
        ws.cell(row, 2, company)
        ws.cell(row, 3, name)
        ws.cell(row, 4, memo)
        ws.cell(row, 5, bal)
        ws.cell(row, 6, pmt)
        for mi in range(12):
            ws.cell(row, 7 + mi, months[mi])
        ws.cell(row, 19, f"=SUM(G{row}:R{row})")   # Total Paid YTD
        ws.cell(row, 20, beg)
        ws.cell(row, 21, notes)
        # formatting
        fmt_cell(ws.cell(row, 1), font=SUBHEADER_FONT, align=LEFT, fill=TYPE_FILLS.get(typ))
        fmt_cell(ws.cell(row, 2), font=BLACK_FORMULA, align=CENTER,
                 fill=(LA_FILL if company == "L&A Holdings" else None))
        fmt_cell(ws.cell(row, 3), font=BLACK_FORMULA, align=LEFT)
        fmt_cell(ws.cell(row, 4), font=BLUE_INPUT, align=LEFT)
        for c in [5, 6] + list(range(7, 21)):
            fmt_cell(ws.cell(row, c), font=BLUE_INPUT, fmt=CURRENCY_FMT, align=RIGHT)
        fmt_cell(ws.cell(row, 19), font=BLACK_FORMULA, fmt=CURRENCY_FMT, align=RIGHT, fill=GREY_FILL)
        fmt_cell(ws.cell(row, 21), font=BLACK_FORMULA, align=LEFT)
        row += 1

    last = row - 1
    # Grand total row
    ws.cell(row, 3, "TOTAL")
    fmt_cell(ws.cell(row, 3), font=SUBHEADER_FONT, fill=TOTAL_FILL, align=LEFT)
    for c in [5, 6] + list(range(7, 21)):
        ws.cell(row, c, f"=SUM({get_column_letter(c)}{hr+1}:{get_column_letter(c)}{last})")
        fmt_cell(ws.cell(row, c), font=SUBHEADER_FONT, fill=TOTAL_FILL, fmt=CURRENCY_FMT, align=RIGHT)
    for c in (1, 2, 4, 21):
        fmt_cell(ws.cell(row, c), fill=TOTAL_FILL)
    # AutoFilter over the data (sort/group by Type)
    ws.auto_filter.ref = f"A{hr}:U{last}"
    ws.freeze_panes = "C3"


def build_template_real(wb):
    """A REAL, copyable amort tab. Type an Equipment # in A5 — everything else
    pulls from the matching Master row via INDEX/MATCH. Copy the tab, type the
    new Equipment # in A5, rename the tab. No row-reference editing."""
    # Build a normal amort tab first (references Master row 4 as a scaffold)…
    build_amort_tab(wb, "TYPE EQUIP # IN A5", "", "", 0, 0, 0, 0, None, 4, "_TEMPLATE_AMORT")
    ws = wb["_TEMPLATE_AMORT"]
    MASTER = "'Master Debt Schedule'"
    mrow = f"MATCH($A$5,{MASTER}!$B:$B,0)"   # find this equipment's Master row
    # A5 is the ONLY input now; the rest look it up.
    ws["A5"] = "TYPE EQUIP # HERE"
    fmt_cell(ws["A5"], font=BLUE_INPUT, fill=YELLOW_FILL, align=LEFT)
    ws["B5"] = f"=IFERROR(INDEX({MASTER}!$C:$C,{mrow}),\"\")"
    ws["C5"] = f"=IFERROR(INDEX({MASTER}!$D:$D,{mrow}),\"\")"
    ws["D5"] = f"=IFERROR(INDEX({MASTER}!$F:$F,{mrow}),0)"
    ws["E5"] = f"=IFERROR(INDEX({MASTER}!$G:$G,{mrow}),0)"
    ws["F5"] = f"=IFERROR(INDEX({MASTER}!$H:$H,{mrow}),\"\")"
    ws["G5"] = f"=IFERROR(INDEX({MASTER}!$I:$I,{mrow}),0)"
    ws["H5"] = f"=IFERROR(INDEX({MASTER}!$J:$J,{mrow}),0)"
    ws["I5"] = f"=IFERROR(INDEX({MASTER}!$K:$K,{mrow}),0)"
    ws["J5"] = f'=IFERROR(IF(INDEX({MASTER}!$L:$L,{mrow})="","",INDEX({MASTER}!$L:$L,{mrow})),"")'
    ws["K5"] = f"=IFERROR(INDEX({MASTER}!$M:$M,{mrow}),1)"
    # instructions banner
    ws.merge_cells("A2:Q2")
    ws["A2"] = ("HOW TO USE: right-click this tab → Move or Copy → ✔ Create a copy.  On the copy, type the "
                "Equipment # (exactly as on Master col B) into the yellow A5 cell, then rename the tab.  Done.")
    fmt_cell(ws["A2"], font=Font(name=FONT_NAME, color="C00000", bold=True), fill=YELLOW_FILL, align=LEFT)


BLUE_AUTO = Font(name=FONT_NAME, color="0000FF")          # automated / pulled / synced
INPUT_FONT = Font(name=FONT_NAME, color="000000")          # you type here (on yellow)


def _yellow(cell):
    cell.fill = YELLOW_FILL
    cell.font = INPUT_FONT

def _blue(cell):
    cell.font = BLUE_AUTO
    if cell.fill and getattr(cell.fill, "fgColor", None) and cell.fill.fgColor.rgb in ("00FFFF00", "FFFFFF00"):
        cell.fill = PatternFill()  # clear yellow off automated cells


def apply_color_convention(wb, loan_tab_titles):
    """Workbook-wide: YELLOW fill = input (you type) · BLUE text = automated."""
    # ---- Master Debt Schedule ----
    ms = wb["Master Debt Schedule"]
    total_row = None
    for r in range(4, ms.max_row + 1):
        if str(ms.cell(r, 1).value or "").strip().upper() == "TOTALS":
            total_row = r
            break
    last = (total_row or ms.max_row + 1) - 1
    INPUT_COLS = [2, 3, 4, 5, 6, 9, 10, 12, 13, 14, 20, 21, 22, 23]  # B..N inputs + receivables
    AUTO_COLS = [7, 8, 15, 16, 18]                                   # G,H,O,P,R
    for r in range(4, last + 1):
        is_loan = bool(ms.cell(r, 18).value)        # has amort tab
        for c in INPUT_COLS:
            cell = ms.cell(r, c)
            # receivable-only rows: only the receivable cols are inputs
            if not is_loan and c < 20:
                continue
            if c >= 20 and not is_loan and not (ms.cell(r, 20).value or ms.cell(r, 21).value
                                                or ms.cell(r, 22).value or ms.cell(r, 23).value):
                continue
            _yellow(cell)
        if is_loan:
            for c in AUTO_COLS:
                _blue(ms.cell(r, c))
            # Rate (K=11): formula -> auto/blue; literal -> input/yellow
            kv = ms.cell(r, 11).value
            if isinstance(kv, str) and kv.startswith("="):
                _blue(ms.cell(r, 11))
            else:
                _yellow(ms.cell(r, 11))

    # ---- Monthly Payments ----
    mp = wb["Monthly Payments"]
    mp_total = None
    for r in range(3, mp.max_row + 1):
        if str(mp.cell(r, 3).value or "").strip().upper() == "TOTAL":
            mp_total = r
            break
    mlast = (mp_total or mp.max_row + 1) - 1
    for r in range(3, mlast + 1):
        for c in [3, 4, 5, 6] + list(range(7, 19)) + [20, 21]:  # Name,Memo,Bal,Pmt,Jan-Dec,Beg,Notes
            _yellow(mp.cell(r, c))
        _blue(mp.cell(r, 19))                                   # Total Paid = automated

    # ---- Amort tabs (+ template): tracker inputs yellow, row-5 KPIs blue ----
    for title in list(loan_tab_titles) + ["_TEMPLATE_AMORT"]:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        for r in range(9, 21):              # PAYMENT TRACKER years
            for c in range(2, 14):          # Jan..Dec (B..M) — type actual $ here
                _yellow(ws.cell(r, c))
        for c in range(1, 12):              # row 5 values pulled from Master
            _blue(ws.cell(5, c))
        _blue(ws["P5"]); _blue(ws["Q5"])    # YTD principal / interest


def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    AS_OF = "2026-03-31"  # last day of month per existing 'March 2026 FINAL' file

    # 1) README first (becomes index 0)
    build_readme(wb)

    # 2) Master
    sheet_names, _total_row = build_master(wb, AS_OF)

    # 3) Amort tabs (collect company per tab for the Index) + back-link to Index
    loan_tabs = []
    for i, (company, equip, desc, lender, acct_no, orig, curr, pmt, term, rate, status, notes) in enumerate(LOANS):
        master_row = 4 + i
        tab = sheet_names[i][1]
        build_amort_tab(wb, equip, desc, lender, orig, curr, pmt, term, rate,
                        master_row, tab)
        loan_tabs.append((company, equip, tab))
        # back-link to Index in the empty row-2 area
        aws = wb[tab]
        aws["Q2"] = _link("Index", "← Index")
        fmt_cell(aws["Q2"], font=LINK_FONT, align=RIGHT)

    # 4) Make Master 'Amort Tab' column (R) clickable WITHOUT changing its text
    #    (the YTD O/P formulas read R as the plain tab name via INDIRECT — keep it).
    mws = wb["Master Debt Schedule"]
    for i in range(len(LOANS)):
        r = 4 + i
        tab = sheet_names[i][1]
        cell = mws[f"R{r}"]            # value already = tab name (set in build_master)
        cell.hyperlink = f"#'{tab}'!A1"
        fmt_cell(cell, font=LINK_FONT, align=LEFT, fill=GREY_FILL)
    mws["A2"] = _link("Index", "← Index")
    fmt_cell(mws["A2"], font=LINK_FONT, align=CENTER, fill=SUB_FILL)

    # 5) Monthly Payments tracker (non-loan), 6) Index hub, 7) real template
    build_monthly_tracker(wb)
    extra_tabs = [("Master Debt Schedule", "Master Debt Schedule (loans + amortization)"),
                  ("Monthly Payments", "Monthly Payments (utilities, insurance, receivables…)"),
                  ("README", "README — how it works"),
                  ("_TEMPLATE_AMORT", "Template — add a new loan"),
                  ("Archive", "Archive — paid-off loans")]
    build_index(wb, loan_tabs, extra_tabs)
    build_template_real(wb)

    # Archive
    arch = wb.create_sheet("Archive")
    arch["A1"] = "PAID-OFF / RETIRED LOANS — move tabs here when paid off"
    fmt_cell(arch["A1"], font=TITLE_FONT, fill=PatternFill("solid", start_color="595959"), align=CENTER)
    arch.column_dimensions["A"].width = 80

    # 7b) Color convention: yellow = input, blue text = automated
    apply_color_convention(wb, [t for (_c, _e, t) in loan_tabs])

    # 8) Final sheet order: README, Index, Master, Monthly Payments, loans…, template, archive
    desired = ["README", "Index", "Master Debt Schedule", "Monthly Payments"]
    desired += [t for (_c, _e, t) in loan_tabs]
    desired += ["_TEMPLATE_AMORT", "Archive"]
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 999)

    import os
    # Recalc all formulas when the file is opened (Excel + LibreOffice honor this).
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "Equipment_Debt_Schedule_v2.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
