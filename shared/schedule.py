"""
schedule.py — weekly crew-schedule stage model (shared).

The daily crew schedules live at
`…/OPERATIONS/SCHEDULE/<yr>/<month>/Schedule M-D-YY.xlsx` ('Daily Schedule'
tab): crew/foreman names in col A above their jobs, each job row carrying an
address (col B), city (C), builder (D) and the STAGE that day (E, 'Description'
— Pour / Wreck / Forms / …).

This module reads a Mon–Fri week into a job × day × stage model, matches pricing
best-effort from the WIP master by address, and classifies each stage into a
coloured category. Extracted from one-offs/schedule_report.py (2026-07-24) the
moment company-tracker also needed it — repo rule: tools never import tools,
shared code lives here.
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from shared import paths

SCHEDULE_DIR = paths.get_path("RP_SCHEDULE_DIR", "/Volumes/Common/OPERATIONS/SCHEDULE")
WIP_PATH = paths.get_path(
    "WIP_EXCEL_PATH",
    paths.onedrive_base() / "Company Files - WIP Report/WIP - MASTER new.xlsx")
GL_PATH = paths.get_path(
    "RP_ALPHA_PATH",
    "/Volumes/Common/OPERATIONS/GENERAL LIST/LISTA GENERAL AÑO 2026.xlsx")

# General List column map (1-based → 0-based for values_only): job C, house D,
# street E, slab-bid AI(35), flat-bid AK(37). Header row 4, data row 6+.
_GL_JOB, _GL_HOUSE, _GL_STREET, _GL_SLAB, _GL_FLAT = 2, 3, 4, 34, 36
_GL_PRICED_SHEETS = ("General list - Alpha order", "Small Jobs")
_JOB_RE = re.compile(r"^(RP\d{4}(?:-[A-Za-z]{2,6})?|CP\d{3,4})\b", re.IGNORECASE)

_FILE_RE = re.compile(r"Schedule (\d{1,2})-(\d{1,2})-(\d{2})\.xlsx$", re.IGNORECASE)

# stage → (category, hex color). First keyword hit wins.
_STAGE_RULES = [
    ("pour", ("Pour", "C00000")),
    ("wreck", ("Wreck", "305496")),
    ("form", ("Forms", "BF8F00")),
    ("cable", ("Cables", "7030A0")), ("tension", ("Cables", "7030A0")),
    ("stress", ("Stress", "1F6B4C")),
    ("plumb", ("Plumbing", "2E75B6")), ("trench", ("Plumbing", "2E75B6")),
    ("grade", ("Grade", "548235")), ("back", ("Grade", "548235")),
    ("punch", ("Punch", "7F7F7F")),
    ("set up", ("Set up", "9E480E")), ("setup", ("Set up", "9E480E")),
    ("pad", ("Pads", "9E480E")), ("patio", ("Pads", "9E480E")),
    # observed in the Main Schedule (the user 2026-07-28) — 'make up' is the
    # ready-to-build queue; the rest are demo/repair scopes
    ("make up", ("Make-up", "C55A11")), ("makeup", ("Make-up", "C55A11")),
    ("pier", ("Piers", "2F5597")), ("footing", ("Piers", "2F5597")),
    ("demo", ("Demo/Repair", "833C00")), ("saw", ("Demo/Repair", "833C00")),
    ("cut joint", ("Demo/Repair", "833C00")), ("patch", ("Demo/Repair", "833C00")),
]
_DEFAULT_STAGE = ("Other", "404040")

# The Main Schedule groups jobs under SECTION headers, and those sections ARE
# the pipeline stage the office works from (the user 2026-08-03: jobs under
# "Grade and Backout / Trenched" were missing from the stage list). The
# DESCRIPTION column is the specific task within that stage, so classifying by
# description alone scattered one section across four invented buckets.
_SECTION_RULES = [
    ("wreck", ("Wreck & Clean", "305496")),
    ("flatwork", ("Flatwork", "BF8F00")),
    ("grade", ("Grade & Backout", "548235")), ("backout", ("Grade & Backout", "548235")),
    ("form", ("Form Set", "2F5597")),
    ("punch", ("Punch List", "7F7F7F")),
    ("pour", ("Pour", "C00000")),
]


def section_cat(section: str) -> Tuple[str, str]:
    """Section header → (stage label, colour). Falls back to the raw section
    name so a NEW section the office adds shows up as itself, never dropped."""
    lo = (section or "").lower()
    for kw, cat in _SECTION_RULES:
        if kw in lo:
            return cat
    clean = (section or "").strip()
    if clean:
        return (clean[:28].title(), "404040")
    return _DEFAULT_STAGE


def stage_cat(desc: str) -> Tuple[str, str]:
    lo = (desc or "").lower()
    for kw, cat in _STAGE_RULES:
        if kw in lo:
            return cat
    return _DEFAULT_STAGE


def legend() -> List[Tuple[str, str]]:
    seen: Dict[str, str] = {}
    for _, cat in _STAGE_RULES:
        seen.setdefault(cat[0], cat[1])
    seen[_DEFAULT_STAGE[0]] = _DEFAULT_STAGE[1]
    return list(seen.items())


# ────────────────────────── discovery + parse ──────────────

def _monday(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def find_week_files(target: Optional[dt.date] = None) -> List[Tuple[dt.date, Path]]:
    """Mon–Fri Schedule files for the target week (or the latest week with
    files). Empty list if the schedule volume isn't mounted / none found."""
    found: Dict[dt.date, Path] = {}
    if not SCHEDULE_DIR.is_dir():
        return []
    for p in SCHEDULE_DIR.rglob("Schedule *.xlsx"):
        m = _FILE_RE.search(p.name)
        if not m or p.name.startswith("~$"):
            continue
        try:
            d = dt.date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            continue
        found[d] = p
    if not found:
        return []
    if target is None:
        # the week containing TODAY (not the newest file — next week's schedule
        # is often pre-loaded); fall back to the latest week that has files
        mon = _monday(dt.date.today())
        if not any((mon + dt.timedelta(days=i)) in found for i in range(5)):
            mon = _monday(max(found))
    else:
        mon = _monday(target)
    return [(mon + dt.timedelta(days=i), found[mon + dt.timedelta(days=i)])
            for i in range(5) if (mon + dt.timedelta(days=i)) in found]


def parse_daily(path: Path) -> List[dict]:
    """'Daily Schedule' tab → [{crew, address, city, builder, stage}]."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == "daily schedule"),
                 wb.sheetnames[0])
    ws = wb[sheet]
    crew = None
    out = []
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip()
        b = str(ws.cell(r, 2).value or "").strip()
        c = str(ws.cell(r, 3).value or "").strip()
        d = str(ws.cell(r, 4).value or "").strip()
        e = str(ws.cell(r, 5).value or "").strip()
        if a and not b and a.lower() != "name":
            crew = a
        elif b and b.lower() != "address":
            out.append({"crew": crew, "address": b, "city": c,
                        "builder": d, "stage": e})
    wb.close()
    return out


# ────────────────────────── pricing (best-effort) ──────────

_SUFFIX = {"DRIVE": "DR", "STREET": "ST", "AVENUE": "AVE", "LANE": "LN",
           "ROAD": "RD", "COURT": "CT", "BOULEVARD": "BLVD", "CIRCLE": "CIR",
           "PLACE": "PL", "TRAIL": "TRL", "PARKWAY": "PKWY", "HIGHWAY": "HWY",
           "TERRACE": "TER", "COVE": "CV", "RANCH": "RCH"}
_DIR = {"NORTH": "N", "SOUTH": "S", "EAST": "E", "WEST": "W"}


def _norm_addr(s: str) -> str:
    s = re.sub(r"[^A-Z0-9 ]", " ", (s or "").upper())
    toks = []
    for t in s.split():
        t = _SUFFIX.get(t, _DIR.get(t, t))
        toks.append(t)
    return " ".join(toks).strip()


def _money(v):
    try:
        f = float(v)
        return f if f else 0.0
    except (TypeError, ValueError):
        return 0.0


def _read_general_list(path: Path) -> Tuple[Dict[str, dict], Dict[str, float]]:
    """(by_addr, by_proj) from the General List: every RP/CP row across all
    sheets gives address→project# (broad coverage); the priced Alpha/Small
    sheets add the slab+flat bid as the contract."""
    by_addr: Dict[str, dict] = {}
    by_proj: Dict[str, float] = {}
    if not path.exists():
        return by_addr, by_proj
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return by_addr, by_proj
    for sn in wb.sheetnames:
        priced = sn in _GL_PRICED_SHEETS
        ws = wb[sn]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if len(row) <= _GL_STREET:
                continue
            job = row[_GL_JOB]
            m = _JOB_RE.match(str(job).strip()) if job else None
            if not m:
                continue
            proj = m.group(1).upper()
            house = row[_GL_HOUSE] if len(row) > _GL_HOUSE else None
            street = row[_GL_STREET] if len(row) > _GL_STREET else None
            if not street:
                continue
            key = _norm_addr(f"{house or ''} {street}")
            contract = None
            if priced and len(row) > _GL_FLAT:
                c = _money(row[_GL_SLAB]) + _money(row[_GL_FLAT])
                contract = c or None
            if contract and proj not in by_proj:
                by_proj[proj] = contract
            if key:
                # don't let a later, price-less row clobber a priced one
                if key not in by_addr or (contract and by_addr[key].get("contract") is None):
                    by_addr[key] = {"proj": proj, "contract": contract, "name": street}
    wb.close()
    return by_addr, by_proj


def _read_wip(path: Path) -> Tuple[Dict[str, dict], Dict[str, float]]:
    """(by_addr, by_proj) from the WIP master Test-Master — clean active
    contracts (PROJECT NAME is the address for RP)."""
    by_addr: Dict[str, dict] = {}
    by_proj: Dict[str, float] = {}
    if not path.exists():
        return by_addr, by_proj
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return by_addr, by_proj
    if "Test-Master" not in wb.sheetnames:
        wb.close()
        return by_addr, by_proj
    ws = wb["Test-Master"]
    hdr_row, H = None, {}
    for r in range(1, 6):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any("CONTRACT" in str(v).upper() for v in vals if v):
            hdr_row = r
            H = {str(v).strip().upper(): i + 1 for i, v in enumerate(vals) if v}
            break
    if not hdr_row:
        wb.close()
        return by_addr, by_proj
    pc, nc, cc = H.get("PROJECT #"), H.get("PROJECT NAME"), H.get("TOTAL CONTRACT PRICE")
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(r, nc).value if nc else None
        proj = str(ws.cell(r, pc).value or "").upper() if pc else ""
        contract = ws.cell(r, cc).value if cc else None
        contract = contract if isinstance(contract, (int, float)) else None
        if proj and contract and proj not in by_proj:
            by_proj[proj] = contract
        if name:
            key = _norm_addr(str(name))
            if key and key not in by_addr:
                by_addr[key] = {"proj": proj, "contract": contract, "name": str(name)}
    wb.close()
    return by_addr, by_proj


def build_price_map(path: Path = WIP_PATH, gl_path: Path = GL_PATH) -> dict:
    """Address→pricing lookup, broadened: the General List gives address→
    project# across ALL its jobs plus bid prices on the priced sheets; the WIP
    master overlays clean active contracts. Returns {'by_addr':…, 'by_proj':…}.
    match_price cross-looks the project# in by_proj when an address has no price
    of its own."""
    gl_addr, gl_proj = _read_general_list(gl_path)
    wip_addr, wip_proj = _read_wip(path)
    by_addr = dict(gl_addr)
    by_addr.update(wip_addr)                       # WIP contract wins on collision
    by_proj = dict(gl_proj)
    by_proj.update(wip_proj)                       # WIP contract wins
    return {"by_addr": by_addr, "by_proj": by_proj}


def _with_price(hit: dict, by_proj: Dict[str, float]) -> dict:
    if hit and hit.get("contract") is None and hit.get("proj") in by_proj:
        return {**hit, "contract": by_proj[hit["proj"]]}
    return hit


# Street-type words are typed inconsistently between the schedule and the
# General List (the same job appears as "1234 OAKWOOD PARKWAY" and "1234
# OAKWOOD PKWY"), so they are DROPPED from the comparison rather than normalised — the
# house number + street NAME is the reliable key (the user 2026-07-28).
_STREET_TYPES = {"DR", "ST", "AVE", "LN", "RD", "CT", "BLVD", "CIR", "PL",
                 "TRL", "PKWY", "HWY", "TER", "CV", "RCH", "WAY", "LOOP",
                 "PASS", "PATH", "RUN", "ROW", "BEND", "CROSSING", "XING"}


def _addr_core(key: str) -> Tuple[str, frozenset]:
    """'1234 OAKWOOD RD' → ('1234', {'OAKWOOD'}). Drops street types and any
    trailing descriptors so '1234 OAKWOOD PKWY GYM ADDITION' still matches
    '1234 OAKWOOD PARKWAY'."""
    toks = key.split()
    if not toks:
        return "", frozenset()
    house = toks[0].split("-")[0]                 # '420-422' → '420'
    words = {t for t in toks[1:] if t not in _STREET_TYPES and not t.isdigit()}
    return house, frozenset(words)


def match_price(address: str, pmap: dict) -> dict:
    """Exact → house# + street-name core (street types ignored) → best
    token-overlap, then a project#→contract cross-lookup so a matched job shows
    a price even if its own address row had none."""
    by_addr = pmap.get("by_addr", {})
    by_proj = pmap.get("by_proj", {})
    miss = {"proj": "", "contract": None, "name": ""}
    key = _norm_addr(address)
    if not key:
        return miss
    if key in by_addr:
        return _with_price(by_addr[key], by_proj)
    house, words = _addr_core(key)
    if not house or not words:
        return miss
    best, best_score = None, 0.0
    for k, v in by_addr.items():
        kh, kw = _addr_core(k)
        if kh != house or not kw:
            continue
        # the shorter side drives the score, so extra descriptors on either
        # side ("GYM ADDITION") don't sink an otherwise exact street match
        overlap = len(words & kw) / min(len(words), len(kw))
        if overlap > best_score:
            best, best_score = v, overlap
    if best and best_score >= 0.75:
        return _with_price(best, by_proj)
    return miss


# ────────────────────────── model ──────────────────────────

_MAIN_SHEET = "Main Schedule"
# Main Schedule layout (1-based): A super, B PROJECT #, C address, D city,
# E builder, F description/stage, G date, H crew. Section bands ('WRECK AND
# CLEAN', 'FLATWORK', '… PUNCH LIST') and repeated header rows are skipped.
_MS_SUPER, _MS_PROJ, _MS_ADDR, _MS_CITY, _MS_BUILDER, _MS_STAGE, _MS_DATE, _MS_CREW = range(8)


def parse_main_schedule(path: Path) -> List[dict]:
    """'Main Schedule' tab → the full job pipeline.

    Columns are located dynamically, because the layout CHANGED mid-2026: older
    files run NAME | ADDRESS | CITY | BUILDER | DESCRIPTION with **no project #
    column**, while newer ones insert PROJECT # as column B. Parsing the old
    files by address (proj left blank, resolved later from newer sightings) is
    what lets "in progress" reach back past the column change — otherwise every
    job looks like it started the week the column appeared (the user 2026-07-28).
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == _MAIN_SHEET.lower()),
                 None)
    if sheet is None:
        wb.close()
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    # locate a header row and map the columns we need by name
    cols: Dict[str, int] = {}
    for row in rows[:12]:
        names = {str(c).strip().upper(): i for i, c in enumerate(row) if c}
        if "ADDRESS" in names and "DESCRIPTION" in names:
            cols = names
            break
    if not cols:
        return []
    c_addr = cols["ADDRESS"]
    c_stage = cols["DESCRIPTION"]
    c_city = cols.get("CITY")
    c_bldr = cols.get("BUILDER")
    c_crew = cols.get("CREW")
    c_name = cols.get("NAME", 0)

    out = []
    section = ""
    for row in rows:
        if not row or len(row) <= max(c_addr, c_stage):
            continue
        addr = str(row[c_addr] or "").strip()
        first = str(row[c_name] or "").strip() if len(row) > c_name else ""
        if not addr or addr.upper() == "ADDRESS":
            # a lone label in the first column is a SECTION header
            if first and first.upper() not in ("NAME", "SUPERINTENDENT"):
                section = first
            continue
        # project # lives in its own column on newer files; find it anywhere
        proj = ""
        for c in row:
            m = _JOB_RE.match(str(c).strip()) if c else None
            if m:
                proj = m.group(1).upper()
                break
        out.append({
            "proj": proj,
            "section": section,
            "address": addr,
            "city": str(row[c_city] or "").strip() if c_city is not None and len(row) > c_city else "",
            "builder": str(row[c_bldr] or "").strip() if c_bldr is not None and len(row) > c_bldr else "",
            "stage": str(row[c_stage] or "").strip(),
            "crew": str(row[c_crew] or "").strip() if c_crew is not None and len(row) > c_crew else "",
            "super": str(row[c_name] or "").strip() if len(row) > c_name else "",
        })
    return out


def recent_files(weeks_back: int = 6) -> List[Tuple[dt.date, Path]]:
    """Every schedule file within the last `weeks_back` weeks, oldest first."""
    cutoff = dt.date.today() - dt.timedelta(weeks=weeks_back)
    found: List[Tuple[dt.date, Path]] = []
    if not SCHEDULE_DIR.is_dir():
        return found
    for p in SCHEDULE_DIR.rglob("Schedule *.xlsx"):
        m = _FILE_RE.search(p.name)
        if not m or p.name.startswith("~$"):
            continue
        try:
            d = dt.date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            continue
        if d >= cutoff:
            found.append((d, p))
    return sorted(found)


def build_rp_stages(weeks_back: int = 10,
                    pmap: Optional[Dict[str, dict]] = None) -> dict:
    """Every RP project and the stage it is currently in (the user 2026-07-28:
    "just show me the stages of all RP projects").

    Source is the Schedule file's **Main Schedule** tab, which carries the
    project # alongside the address and stage — NOT the General List, which
    misses jobs. Walks the schedule files oldest→newest across `weeks_back`
    weeks so each later day UPDATES the job's stage; the newest sighting wins.
    Pricing is the only thing still looked up externally (General List / WIP).
    """
    if pmap is None:
        pmap = build_price_map()
    files = recent_files(weeks_back)
    parsed = [(d, parse_main_schedule(p)) for d, p in files]   # oldest → newest

    # Older files carry no project # column, so build address → project # from
    # every sighting that HAS one, then use it to attribute the older,
    # address-only rows. Without this, "in progress" restarts at the date the
    # project-# column was introduced.
    addr_to_proj: Dict[str, str] = {}
    for _d, rws in parsed:
        for r in rws:
            if r["proj"]:
                addr_to_proj.setdefault(_norm_addr(r["address"]), r["proj"])

    latest: Dict[str, dict] = {}                 # PROJECT # → newest sighting
    first_seen: Dict[str, dt.date] = {}          # PROJECT # → oldest sighting
    for d, rws in parsed:                        # oldest → newest, later wins
        for row in rws:
            if not row["stage"]:
                continue
            proj = row["proj"] or addr_to_proj.get(_norm_addr(row["address"]), "")
            if not proj:
                continue
            first_seen.setdefault(proj, d)
            if row["proj"]:                      # only trust fully-keyed rows
                latest[proj] = {**row, "date": d}
    earliest = files[0][0] if files else None
    today = dt.date.today()
    rows = []
    for proj, s in latest.items():
        if not proj.upper().startswith("RP"):
            continue                              # RP projects only
        # group by the schedule SECTION (the real pipeline stage); the
        # description stays as the task detail (the user 2026-08-03)
        cat, color = section_cat(s.get("section", "")) if s.get("section") \
            else stage_cat(s["stage"])
        pr = match_price(s["address"], pmap)
        contract = pr["contract"]
        if contract is None:                      # fall back to the project # itself
            contract = pmap.get("by_proj", {}).get(proj)
        fs = first_seen.get(proj)
        # "in progress" = days since the job FIRST showed up on a schedule in
        # the scanned window. If that first sighting is the oldest file we read,
        # the job predates the window, so the figure is a floor (shown as "45+").
        in_progress = (today - fs).days if fs else None
        rows.append({
            "proj": proj, "address": s["address"], "builder": s["builder"],
            "city": s["city"], "crew": s["crew"], "super": s.get("super", ""),
            "contract": contract,
            "stage": s["stage"], "section": s.get("section", ""),
            "stage_cat": cat, "stage_color": color,
            "last_seen": s["date"], "days_ago": (today - s["date"]).days,
            "first_seen": fs, "in_progress": in_progress,
            "in_progress_capped": bool(fs and earliest and fs <= earliest),
        })
    rows.sort(key=lambda r: (r["days_ago"], r["proj"]))
    return {"rows": rows, "weeks_back": weeks_back, "files": len(files),
            "generated": dt.datetime.now()}


def build_model(week: Optional[List[Tuple[dt.date, Path]]] = None,
                pmap: Optional[Dict[str, dict]] = None) -> dict:
    """job (address) → {proj, contract, builder, current, days:{date:stage}}.
    Returns {'dates': [...], 'jobs': [...]} (empty if no schedule files)."""
    if week is None:
        week = find_week_files()
    if not week:
        return {"dates": [], "jobs": []}
    if pmap is None:
        pmap = build_price_map()
    jobs: Dict[str, dict] = {}
    for d, path in week:
        for row in parse_daily(path):
            addr = row["address"]
            j = jobs.setdefault(addr, {"address": addr, "builder": row["builder"],
                                       "city": row["city"], "days": {},
                                       "crew": row["crew"]})
            if row["stage"]:
                j["days"][d] = row["stage"]
            if row["builder"] and not j["builder"]:
                j["builder"] = row["builder"]
    for addr, j in jobs.items():
        pr = match_price(addr, pmap)
        j["proj"], j["contract"] = pr["proj"], pr["contract"]
        last = max((d for d in j["days"]), default=None)
        j["current"] = j["days"].get(last, "") if last else ""
    dates = [d for d, _ in week]
    ordered = sorted(jobs.values(), key=lambda j: (-len(j["days"]), j["address"]))
    return {"dates": dates, "jobs": ordered}

# ── choosing WHICH schedule file to read ────────────────────────────
_SCHED_NAME_RE = re.compile(r"Schedule (\d{1,2})-(\d{1,2})-(\d{2})\.xlsx$",
                            re.IGNORECASE)


def all_schedule_files(sched_dir: Path = None):
    """[(date, path)] for every 'Schedule M-D-YY.xlsx', oldest → newest."""
    root = Path(sched_dir or SCHEDULE_DIR)
    out = []
    if not root.exists():
        return out
    for year_dir in root.iterdir():
        if not (year_dir.is_dir() and year_dir.name.isdigit()):
            continue
        for month_dir in year_dir.iterdir():
            if not month_dir.is_dir():
                continue
            for f in month_dir.iterdir():
                m = _SCHED_NAME_RE.search(f.name)
                if m:
                    mo, dy, yy = (int(g) for g in m.groups())
                    try:
                        out.append((dt.date(2000 + yy, mo, dy), f))
                    except ValueError:
                        pass
    return sorted(out)


def schedule_on_or_before(as_of: Optional[dt.date] = None,
                          sched_dir: Path = None):
    """The newest schedule file dated ON OR BEFORE `as_of` (default: today).

    NEVER returns a FUTURE-dated file. The team pre-loads tomorrow's board, so
    a plain "highest filename wins" pick silently jumps to tomorrow the moment
    that file lands — which is how a 7-29 run started reading 7-30 and would
    have written tomorrow's dates into JobTread (the user 2026-07-29).

    Returns (date, path) or None.
    """
    cap = as_of or dt.date.today()
    files = [(d, p) for d, p in all_schedule_files(sched_dir) if d <= cap]
    return files[-1] if files else None
