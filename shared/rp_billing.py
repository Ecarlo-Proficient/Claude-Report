"""
rp_billing.py — RP billing status: what's poured and unbilled vs still backlog.

The question (the user 2026-07-28): "find which jobs were completed and find out
if we billed for them — that's our underbilling for RP."

The trap this module exists to avoid: the General List's completion % certifies
the SLAB only, and each job carries TWO priced scopes (slab + flatwork) billed
to two QBO projects (`RP####` and `RP####-FTW`). Comparing slab completion
against slab+flatwork bid overstates underbilling by counting flatwork that was
never poured — measured 2026-07-28: nearly all of the apparent RP underbilling
was unbuilt flatwork plus builder-fee rounding, not real slab underbilling.

So "done" is not taken from a completion column at all. A scope counts as
POURED only when the crew schedule shows it reaching pour/wreck/stress/punch,
with the scope read from the stage text ("POUR SLAB" vs "POUR FLATWORK"/"SET UP
FTW"). Slabs legitimately sit unbilled for a while during punch work, so the
schedule is the only evidence we have that the concrete is actually down.

  poured  + unbilled → INVOICE NOW (real underbilling)
  unpoured + unbilled → BACKLOG (work to come, not money owed)
"""
from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from pathlib import Path
from typing import Callable, Dict, List, Optional

from openpyxl import load_workbook

from shared import schedule as sched

# A scope is flatwork when the stage text says so; everything else is slab.
_FTW_RE = re.compile(r"flatwork|\bftw\b|patio|driveway|sidewalk|approach", re.I)
# Stages that only happen once concrete is down (or is being finished).
_POURED_RE = re.compile(r"\bpour|wreck|stress|punch|strip", re.I)
_JOB_RE = re.compile(r"^(RP\d{4})\b", re.I)

# Builder fees (e.g. WRH's 1.5%) leave small residuals that are NOT underbilling
# — the user's standing materiality rule. Anything under this is "billed out".
FEE_TOLERANCE = 600.0
LOOKBACK_WEEKS = 14

# A few RP jobs are billed by DRAW / phase (piers → foundation → flatwork, with
# their own invoices and change orders) instead of one invoice at completion —
# RP6533 is one (the user 2026-07-28). For these, "poured but unbilled" is the
# WRONG test: billing follows the draw schedule, so a poured slab can be
# correctly unbilled. They're reported in their own bucket, never as
# underbilling. Add project numbers here as they come up.
DRAW_BASED = {"RP6533"}

# Typical RP slab bids run ~$25–50K. A much larger contract is usually a
# phase-billed job, so anything above this that ISN'T already in DRAW_BASED is
# surfaced for review rather than silently counted as underbilling.
DRAW_BASED_REVIEW_OVER = 120_000.0


def _f(v) -> float:
    try:
        return float(v) or 0.0
    except (TypeError, ValueError):
        return 0.0


def read_general_list_bids(path: Optional[Path] = None) -> Dict[str, dict]:
    """{RP####: {addr, slab, flat}} — the priced scopes per job."""
    p = Path(path or sched.GL_PATH)
    out: Dict[str, dict] = {}
    if not p.exists():
        return out
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return out
    for sn in ("General list - Alpha order", "Small Jobs"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(6, ws.max_row + 1):
            job = ws.cell(r, 3).value
            m = _JOB_RE.match(str(job).strip()) if job else None
            if not m:
                continue
            out.setdefault(m.group(1).upper(), {
                "addr": f"{ws.cell(r, 4).value or ''} "
                        f"{ws.cell(r, 5).value or ''}".strip(),
                "builder": ws.cell(r, 2).value or "",
                "slab": _f(ws.cell(r, 35).value),
                "flat": _f(ws.cell(r, 37).value),
            })
    wb.close()
    return out


def pour_history(weeks_back: int = LOOKBACK_WEEKS) -> Dict[str, dict]:
    """{RP####: {slab_poured, ftw_poured, slab_date, ftw_date}} from the crew
    schedules. Older files carry no project # column, so address→project is
    learned from the newer ones and applied backwards."""
    files = sched.recent_files(weeks_back)
    parsed = [(d, sched.parse_main_schedule(p)) for d, p in files]
    addr2proj: Dict[str, str] = {}
    for _d, rows in parsed:
        for r in rows:
            if r["proj"]:
                addr2proj.setdefault(sched._norm_addr(r["address"]), r["proj"])

    hist: Dict[str, dict] = defaultdict(
        lambda: {"slab_poured": False, "ftw_poured": False,
                 "slab_date": None, "ftw_date": None})
    for d, rows in parsed:
        for r in rows:
            stage = r["stage"]
            if not stage:
                continue
            proj = r["proj"] or addr2proj.get(sched._norm_addr(r["address"]), "")
            m = _JOB_RE.match(proj) if proj else None
            if not m:
                continue
            base = m.group(1).upper()
            if not _POURED_RE.search(stage):
                continue
            h = hist[base]
            if _FTW_RE.search(stage):
                h["ftw_poured"] = True
                h["ftw_date"] = max(h["ftw_date"] or d, d)
            else:
                h["slab_poured"] = True
                h["slab_date"] = max(h["slab_date"] or d, d)
    return dict(hist)


def build(billed_for: Callable[[str], Optional[float]],
          weeks_back: int = LOOKBACK_WEEKS) -> dict:
    """Join priced scopes × pour evidence × QBO billing.

    `billed_for(project_key)` returns cumulative invoiced for e.g. 'RP7470' or
    'RP7470-FTW', or None when no QBO customer exists. Returns
    {invoice_now: [...], backlog: [...], weeks_back, generated} with each row
    carrying proj, scope, addr, bid, billed, gap, poured, last_pour, days.
    """
    bids = read_general_list_bids()
    hist = pour_history(weeks_back)
    today = dt.date.today()
    invoice_now: List[dict] = []
    backlog: List[dict] = []
    draw_based: List[dict] = []

    for proj, v in sorted(bids.items()):
        h = hist.get(proj, {})
        for scope, bid, key, poured, last in (
            ("SLAB", v["slab"], proj, h.get("slab_poured", False), h.get("slab_date")),
            ("FTW", v["flat"], f"{proj}-FTW", h.get("ftw_poured", False), h.get("ftw_date")),
        ):
            if bid <= 0:
                continue
            billed = billed_for(key)
            gap = bid - (billed or 0.0)
            if gap <= FEE_TOLERANCE:          # billed out (fee/rounding residue)
                continue
            row = {
                "proj": proj, "scope": scope, "addr": v["addr"],
                "builder": str(v.get("builder") or ""),
                "bid": bid, "billed": billed or 0.0, "gap": gap,
                "poured": poured, "last_pour": last,
                "days": (today - last).days if last else None,
                "in_qbo": billed is not None,
            }
            # draw/phase-billed jobs are judged by their draw schedule, not by
            # pour evidence — keep them out of the underbilling number
            if proj in DRAW_BASED:
                row["why"] = "draw/phase-billed — bill by draw, not pour"
                draw_based.append(row)
            elif poured and bid >= DRAW_BASED_REVIEW_OVER:
                row["why"] = (f"contract over ${DRAW_BASED_REVIEW_OVER:,.0f} — "
                              "confirm whether this job bills by draw")
                draw_based.append(row)
            else:
                (invoice_now if poured else backlog).append(row)

    for lst in (invoice_now, backlog, draw_based):
        lst.sort(key=lambda r: -r["gap"])
    return {"invoice_now": invoice_now, "backlog": backlog,
            "draw_based": draw_based,
            "weeks_back": weeks_back, "generated": dt.datetime.now(),
            "due_total": sum(r["gap"] for r in invoice_now),
            "backlog_total": sum(r["gap"] for r in backlog),
            "draw_total": sum(r["gap"] for r in draw_based)}
