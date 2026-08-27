"""
draws.py — CP/commercial draw (AIA G702/G703) discovery + parsing.

Extracted from wip/cp_wip_reader.py (2026-07-16) the moment a second tool
(health-dashboard) needed it — repo rule: tools never import tools; common
code lives in shared/.

Draw folder detection (the user 2026-07-09): the WIP update for a CP project
comes from the LATEST draw (AIA G702/G703 payment application), not the
takeoff.
  • Container folder is named 'Draw' or 'Draws' (inclusive), never 'Drawings'.
  • Draw # is the SEQUENCE — the highest draw # wins, read from the filename
    or the numbered 'Draw #N' subfolder.
  • If no draw folder / no draw yet → callers fall back to the takeoff
    proposal (Original Contract Price); once Draw #1 lands, use the draw.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

DRAWS_FOLDER_RE = re.compile(r"^draws?\b", re.IGNORECASE)   # 'Draw' / 'Draws', not 'Drawings'
DRAW_NUM_RE = re.compile(r"draw\s*#?\s*(\d+)", re.IGNORECASE)  # 'Draw #4', 'DRAW#4', 'Draw 4'
G702_SHEET = "G702"


def coerce_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_draws_folder(project_folder: Path) -> Optional[Path]:
    """Return the project's draw container ('Draw'/'Draws', case-insensitive),
    or None. 'Drawings' is excluded by the word-boundary in DRAWS_FOLDER_RE."""
    try:
        for entry in project_folder.iterdir():
            if entry.is_dir() and DRAWS_FOLDER_RE.match(entry.name.strip()):
                return entry
    except OSError:
        pass
    return None


def draw_num_from_name(name: str) -> Optional[int]:
    m = DRAW_NUM_RE.search(name)
    return int(m.group(1)) if m else None


def has_g702(xlsx: Path) -> bool:
    """True if the workbook has a G702 sheet (payment-application front page).
    read_only so it's a cheap header read, not a full parse."""
    try:
        wb = load_workbook(xlsx, read_only=True)
        try:
            return any(s.strip().lower() == G702_SHEET.lower() for s in wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return False


# Subfolders inside a draw that must never supply the workbook — the team
# parks superseded paperwork in them.
_DEAD_DIR_RE = re.compile(
    r"DO ?NOT ?USE|SUPERSEDED|\bOLD\b|ARCHIVE|VOID|BACKUP", re.I)


def _draw_workbooks(draw_folder: Path, max_depth: int = 3):
    """Every candidate draw workbook inside a 'Draw #N' folder.

    Walks DOWN, because the workbook is not always at the top of the draw
    folder: the team often files it under a per-company subfolder (ours and the
    GC's sit side by side). CP861's draws #4 and #5 did exactly that, so a
    single-level scan found neither and the reader silently fell back to
    draw #3 — costing $52,576 of change orders and $194,906 of billing on that
    job (2026-08-04). 'DO NOT USE' style folders are skipped, and the caller
    still requires `has_g702()`, so the GC's own spreadsheets can't win.
    """
    out = []

    def walk(folder: Path, depth: int) -> None:
        if depth > max_depth:
            return
        try:
            entries = list(folder.iterdir())
        except OSError:
            return
        for f in entries:
            try:
                if f.is_file():
                    if (f.suffix.lower() in (".xlsx", ".xlsm")
                            and not f.name.startswith("~$")):
                        out.append(f)
                elif f.is_dir() and not _DEAD_DIR_RE.search(f.name):
                    walk(f, depth + 1)
            except OSError:
                continue

    walk(draw_folder, 1)
    # Shallowest first: a workbook filed at the top of the draw folder outranks
    # one buried in a subfolder.
    out.sort(key=lambda p: len(p.relative_to(draw_folder).parts))
    return out


# A draw can be re-cut without changing its number — the revision sits in the
# same 'Draw #N' folder beside the original it replaces, and the ORIGINAL is
# dead paper the moment it does. Caught on CP765 draw #4 (the user 2026-08-04):
# 'Revised LP Draw Excel #4' backed a $13,552 change order out of the original,
# and reading the superseded file invented a $13,552 shortfall against a QBO
# invoice that was in fact correct to the dollar.
_REVISED_RE = re.compile(r"revis", re.IGNORECASE)          # Revised / Revision


def _supersedes(path: Path) -> Tuple[int, float]:
    """Tie-break key WITHIN one draw number: a revised workbook beats the
    original, and among equals the newest file wins. Without this the winner
    is filesystem iteration order — i.e. arbitrary."""
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return (1 if _REVISED_RE.search(path.name) else 0), mtime


def find_latest_draw(project_folder: Path) -> Optional[Tuple[int, Path]]:
    """Locate the LATEST draw workbook for a project (the user 2026-07-09: draw # is
    the sequence — highest wins). Draw workbooks live in the 'Draws' container,
    either directly or inside a numbered 'Draw #N' subfolder; if there's no
    container, numbered draws directly under the project folder are also
    accepted. Only numbered draw subfolders are descended into — the Supplier
    Release folders (full of PDFs) are never opened. Returns (draw_num,
    draw_file) or None (→ caller falls back to the takeoff proposal)."""
    scan_root = find_draws_folder(project_folder) or project_folder
    candidates = []   # (draw_num, xlsx_path)
    try:
        entries = list(scan_root.iterdir())
    except OSError:
        return None
    for entry in entries:
        try:
            if entry.is_file():
                if entry.suffix.lower() in (".xlsx", ".xlsm") \
                        and not entry.name.startswith("~$"):
                    n = draw_num_from_name(entry.name)
                    if n is not None:
                        candidates.append((n, entry))
            elif entry.is_dir():
                n = draw_num_from_name(entry.name)      # numbered 'Draw #N' subfolder
                if n is not None:
                    for f in _draw_workbooks(entry):
                        fn = draw_num_from_name(f.name)
                        candidates.append((fn if fn is not None else n, f))
        except OSError:
            continue
    if not candidates:
        return None
    # Highest draw # wins; require a real G702 so a stray xlsx can't win.
    # Within one draw #, _supersedes() breaks the tie — see CP765 draw #4.
    for n, f in sorted(candidates,
                       key=lambda x: (x[0],) + _supersedes(x[1]), reverse=True):
        if has_g702(f):
            return n, f
    return None


def _g702_value(ws, label_sub: str, max_scan: int = 14) -> Optional[float]:
    """Find `label_sub` (case-insensitive substring) in the G702 label column
    and read the first numeric value to its right (skipping '$'/blank cells).
    G702 line labels sit in column A with the amount a few columns right."""
    needle = label_sub.upper()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and needle in str(v).upper():
                for off in range(1, max_scan + 1):
                    raw = ws.cell(r, c + off).value
                    num = coerce_float(raw)
                    if num is not None:
                        return num
                    if raw is None or str(raw).strip() in ("", "$", "USD", "-", "—"):
                        continue
                    break   # hit other text — stop scanning this label row
    return None


def read_draw_g702(draw_file: Path):
    """Read the WIP inputs off a draw's G702 (AIA payment application).
    Mapping (verified 2026-07-09 against CP585 Draws #1–#4):
      Contract Price  = Line 3  Contract Sum to Date (= Line 1 + Line 2)
      Approved COs    = Line 2  Net change by Change Orders
      Billed (gross)  = Line 4  Total Completed & Stored to Date
      Retainage       = Line 4 − Line 6  (Total Earned Less Retainage)
    RETAINAGE IS NOT read from the labeled 'Total Retainage' cell — that cell
    is unreliable across draws (0 / mismatched); Line 4 − Line 6 ties to the
    10% on Line 5a every time. Returns (dict, flags); never raises."""
    flags: List[str] = []
    try:
        wb = load_workbook(draw_file, data_only=True)
    except Exception as e:
        return None, [f"Draw read failed: {type(e).__name__}"]
    try:
        sheet = next((s for s in wb.sheetnames
                      if s.strip().lower() == G702_SHEET.lower()), None)
        if sheet is None:
            return None, ["Draw has no G702 sheet"]
        ws = wb[sheet]
        orig   = _g702_value(ws, "ORIGINAL CONTRACT SUM")          # Line 1
        net_co = _g702_value(ws, "NET CHANGE BY CHANGE ORDERS")    # Line 2
        c2d    = _g702_value(ws, "CONTRACT SUM TO DATE")           # Line 3
        billed = _g702_value(ws, "TOTAL COMPLETED")                # Line 4
        earned = _g702_value(ws, "TOTAL EARNED LESS RETAINAGE")    # Line 6
    finally:
        wb.close()

    # Contract: prefer Line 3; else reconstruct Line 1 + Line 2.
    if c2d is None and orig is not None:
        c2d = orig + (net_co or 0.0)
    # Original contract (base) so contract_price property = base + CO = Line 3.
    if orig is None and c2d is not None:
        orig = c2d - (net_co or 0.0)
    retainage = (billed - earned) if (billed is not None and earned is not None) else None

    data = {
        "orig_contract": orig, "net_co": net_co, "contract_to_date": c2d,
        "billed": billed, "earned_less_retainage": earned, "retainage": retainage,
    }
    missing = [k for k in ("contract_to_date", "billed") if data[k] is None]
    if missing:
        flags.append(f"Draw G702 missing {', '.join(missing)} — open & save the "
                     f"draw in Excel to refresh cached values")
    return data, flags


# ── legacy .xls pay applications (the user 2026-07-29) ───────────────────────
# The readers above expect a modern draw workbook with a literal 'G702' sheet.
# Plenty of jobs still bill on the old AIA template saved as legacy .xls, whose
# sheets are named 'A' (G702 front page) and 'B' (G703 continuation) — e.g.
# CP745's 'Proficient Concrete - 3 - Pay App.xls'. project-pnl takes the CP
# CONTRACT PRICE and APPROVED COs from that document rather than from the draw
# invoices, so the reader lives here where any tool can reach it.
PAY_APP_RE = re.compile(r"pay\s*app", re.IGNORECASE)
# (key, label fragment) — the first row carrying the fragment wins, and the
# value is the RIGHTMOST numeric cell on that row (the G702 puts its dollars in
# the far-right column, with the label and a lone '$' to its left).
PAY_APP_FIELDS = (
    ("original_contract", "original contract sum"),
    ("co_net",            "net change by change order"),
    ("contract_to_date",  "contract sum to date"),
    ("completed_to_date", "total completed"),
    ("retainage",         "retainage"),
    ("earned_to_date",    "total earned"),
    ("previous_apps",     "less: previous application"),
    ("current_due",       "current payment due"),
    ("balance_to_finish", "balance to finish"),
)


def sheet_rows(path: Path, index: int) -> List[list]:
    """One sheet as plain value lists — reads legacy .xls (xlrd) and .xlsx
    (openpyxl) the same way. Empty list when unreadable or absent."""
    if path.suffix.lower() == ".xls":
        try:
            import xlrd                                   # legacy BIFF only
        except ImportError:
            return []
        try:
            book = xlrd.open_workbook(str(path))
            if index >= book.nsheets:
                return []
            sh = book.sheet_by_index(index)
            return [[sh.cell_value(r, c) for c in range(sh.ncols)]
                    for r in range(sh.nrows)]
        except Exception:
            return []
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        names = wb.sheetnames
        if index >= len(names):
            wb.close()
            return []
        rows = [list(r) for r in wb[names[index]].iter_rows(values_only=True)]
        wb.close()
        return rows
    except Exception:
        return []


def _row_label(cells: list) -> str:
    parts = [str(c) for c in cells if isinstance(c, str) and c.strip()]
    return re.sub(r"[.\s]+", " ", " ".join(parts)).strip().lower()


def _rightmost_number(cells: list) -> Optional[float]:
    for v in reversed(cells):
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v)
        if isinstance(v, str):
            n = coerce_float(v)
            if n is not None:
                return n
    return None


def parse_g703_rows(rows: List[list]) -> List[dict]:
    """G703 continuation sheet → schedule-of-values lines. A data row carries a
    numeric ITEM in col A and a description in col B; the 'TOTAL' row ends it.
    Cols: A item · B description · C scheduled · D previous · E this app ·
    F stored · G completed-to-date · H % · I balance · J retainage."""
    out: List[dict] = []
    for cells in rows:
        if len(cells) < 3:
            continue
        a, b = cells[0], cells[1]
        if isinstance(b, str) and b.strip().upper() == "TOTAL":
            break
        if not isinstance(a, (int, float)) or isinstance(a, bool):
            continue
        desc = str(b or "").strip()
        if not desc:
            continue

        def num(i):
            v = cells[i] if i < len(cells) else None
            return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else 0.0

        out.append({
            "item": int(a), "desc": desc,
            "scheduled": round(num(2), 2), "previous": round(num(3), 2),
            "this_app": round(num(4), 2), "stored": round(num(5), 2),
            "completed": round(num(6), 2), "balance": round(num(8), 2),
            "retainage": round(num(9), 2),
        })
    return out


def find_pay_app(project_folder: Path) -> Optional[Tuple[int, Path]]:
    """(draw #, path) of the HIGHEST-numbered '*Pay App*' workbook under the
    project's draw folder, or None."""
    draws = find_draws_folder(project_folder)
    if draws is None:
        return None
    cands: List[Tuple[int, float, Path]] = []
    try:
        for d in draws.iterdir():
            files, n = [], 0
            if d.is_dir():
                n = draw_num_from_name(d.name) or 0
                try:
                    files = [f for f in d.iterdir() if f.is_file()]
                except OSError:
                    continue
            elif d.is_file():
                files = [d]
                n = draw_num_from_name(d.name) or 0
            for f in files:
                if (f.suffix.lower() in (".xls", ".xlsx", ".xlsm")
                        and PAY_APP_RE.search(f.name)
                        and not f.name.startswith("~$")):
                    cands.append((draw_num_from_name(f.name) or n,
                                  f.stat().st_mtime, f))
    except OSError:
        return None
    if not cands:
        return None
    n, _mt, path = max(cands)
    return n, path


def read_pay_app(project_folder: Path) -> dict:
    """CONTRACT PRICE + APPROVED COs (and the G703 schedule of values) from the
    project's latest signed pay application. Returns {} when there is no pay
    app, and {"error": …} when one exists but can't be read — so callers can
    say so instead of silently falling back to another source."""
    found = find_pay_app(project_folder)
    if not found:
        return {}
    draw_no, path = found
    rows = sheet_rows(path, 0)
    if not rows:
        return {"error": "pay app unreadable (legacy .xls needs xlrd installed)",
                "source": path.name, "draw_no": draw_no}
    out: dict = {"source": path.name, "draw_no": draw_no, "path": str(path)}
    for cells in rows:
        text = _row_label(cells)
        if not text:
            continue
        for key, frag in PAY_APP_FIELDS:
            if key in out or frag not in text:
                continue
            val = _rightmost_number(cells)
            if val is not None:
                out[key] = round(val, 2)
    out["sov"] = parse_g703_rows(sheet_rows(path, 1))
    return out


# ───────────────── retroactive draw periods (the user 2026-08-25) ─────────────────
#
# On an older job the recent invoices carry an explicit "(Period:MM/DD/YYYY -
# MM/DD/YYYY)" tag and the earlier ones do not. Without a tag the P&L falls back
# to the CALENDAR month, which is wrong whenever the GC's draw window straddles
# month end - MFD295 bills the 21st through the 20th, so a calendar-month
# fallback pushes three weeks of cost into the wrong draw.
#
# These learn the window SHAPE from the invoices that ARE tagged and apply it
# backwards. Nothing is guessed about which month a draw belongs to: that comes
# from the memo's own "December Draw 2024" / "February 2025 Draw" wording, and
# only falls back to the invoice date when the memo names no month.

_MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}
_MONTH_YEAR_RE = re.compile(
    r"\b(" + "|".join(_MONTHS) + r")\b[^0-9]{0,14}(\d{4})?", re.IGNORECASE)
_PERIOD_TAG_RE = re.compile(
    r"\(\s*Period\s*:\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*\)",
    re.IGNORECASE)
_RETAINAGE_ONLY_RE = re.compile(r"retainage", re.IGNORECASE)


def _mode(vals, default):
    """Most common value; ties break toward the LATER one (a window that starts
    on the 21st and once on the 20th is a 21st window with one typo)."""
    if not vals:
        return default
    best, n = default, 0
    for v in sorted(set(vals), reverse=True):
        c = vals.count(v)
        if c > n:
            best, n = v, c
    return best


def _add_months(y: int, m: int, delta: int):
    i = (y * 12 + (m - 1)) + delta
    return i // 12, i % 12 + 1


def _clamp_day(y: int, m: int, day: int) -> "_dt.date":
    import calendar
    import datetime as _d
    return _d.date(y, m, min(day, calendar.monthrange(y, m)[1]))


def learn_period_shape(memos) -> Optional[dict]:
    """Window shape from the memos that DO carry a Period tag:
    {end_day, start_day, span} where `span` is how many months back the window
    starts. Returns None when fewer than one tagged memo is available."""
    import datetime as _d
    ends, starts, spans = [], [], []
    for memo in memos or []:
        m = _PERIOD_TAG_RE.search(memo or "")
        if not m:
            continue
        try:
            s = _d.datetime.strptime(m.group(1), "%m/%d/%Y").date()
            e = _d.datetime.strptime(m.group(2), "%m/%d/%Y").date()
        except ValueError:
            continue
        if s > e:
            continue
        ends.append(e.day)
        starts.append(s.day)
        spans.append((e.year * 12 + e.month) - (s.year * 12 + s.month))
    if not ends:
        return None
    return {"end_day": _mode(ends, 20), "start_day": _mode(starts, 21),
            "span": _mode(spans, 1), "n": len(ends)}


def draw_month_from_memo(memo: str, fallback_date):
    """(year, month) of the draw this invoice bills, from the memo's own month
    wording; the invoice date only when the memo names no month."""
    for m in _MONTH_YEAR_RE.finditer(memo or ""):
        mon = _MONTHS[m.group(1).lower()]
        yr = int(m.group(2)) if m.group(2) else None
        if yr is None and fallback_date is not None:
            # No year beside the month name: take the invoice's, stepping back
            # a year if that would put the draw in the future (a January draw
            # billed in January is this year; billed in December it is next).
            yr = fallback_date.year
            if mon - fallback_date.month > 6:
                yr -= 1
        if yr:
            return yr, mon
    if fallback_date is None:
        return None
    return fallback_date.year, fallback_date.month


def infer_period_tag(memo: str, txn_date, shape: dict) -> Optional[str]:
    """The `(Period:… - …)` tag this invoice would have carried, or None when
    it should not get one (already tagged, or a pure retainage invoice, which
    bills no work window and is handled by the retainage blocks)."""
    if not shape or _PERIOD_TAG_RE.search(memo or ""):
        return None
    ym = draw_month_from_memo(memo or "", txn_date)
    if not ym:
        return None
    if _RETAINAGE_ONLY_RE.search(memo or "") and not re.search(r"draw", memo or "", re.I):
        return None
    y, mo = ym
    end = _clamp_day(y, mo, shape["end_day"])
    sy, sm = _add_months(y, mo, -abs(shape["span"]))
    start = _clamp_day(sy, sm, shape["start_day"])
    if start > end:
        return None
    return (f"(Period:{start.strftime('%m/%d/%Y')} - "
            f"{end.strftime('%m/%d/%Y')})")
