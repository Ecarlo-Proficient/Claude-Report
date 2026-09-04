#!/usr/bin/env python3
"""
wip_contracts.py - the ONE lookup for a job's CONTRACT and ETC, with its SOURCE.

Overhead is charged at a percentage of the CONTRACT, not of what has been
billed (the owner 2026-09-03: "it's contract 10%"). A LIVE job carries its
contract on the WIP master. A FINISHED job is off the master entirely, so the
only record is the WIP reports - and those must be quoted, never guessed: the
owner 2026-09-04, "sources sources sources always sources so i can see what
you are grabbing these numbers from". Every number this returns comes back
with the file, the sheet and the ROW it was read from, and the P&L prints it.

SOURCE ORDER, and why:
  1. the WIP master's own row - current, and the only source for a LIVE job
  2. `WIP History/WIP 3-31-26.xlsx`, then `WIP History/WIP 12-31-25.xlsx` -
     THE reports, the ones the owner pointed at (2026-09-04). They carry the
     figures that went to the bank; `WIP History/Backups/…- SENT 08-25-26.xlsx`
     are frozen copies of the same rows (verified row-for-row on all 11
     finished MFD jobs), so read the live pair and leave the backups alone.
  3. `MFD WIP Report 12-2025 & 3-2026.xlsx` - the MFD division's own report.
     It covers jobs the bank reports never carried (MFD228) and jobs whose
     bank row is not the job (MFD172, below).

ADDENDUM ROWS ARE SKIPPED. A report can carry a second row for extra scope
sold later - "TDC-BONDS RANCH - ADDED SITE…", 15,077 against MFD172's real
4,976,536. Reading that row as the contract flips MFD172 from a real loss to a
fictitious profit, so any row whose PROJECT NAME says ADDED is not a contract
row. This is why MFD172 resolves to the MFD report while every other job
resolves to the bank's.

Coverage is genuinely incomplete and that is reported, never papered over:
6 of the 11 finished MFD jobs (MFD133, 160, 166, 182, 186, 281) appear in NO
report. `contract_for` returns (None, None, "") for those and the caller drops
the projection block rather than inventing one.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Optional, Tuple

from shared import paths

WIP_DIR_DEFAULT = "Company Files - WIP Report"
_CONTRACT_HDR = "TOTAL CONTRACT PRICE"
_ETC_HDR = "ESTIMATED TOTAL COSTS"
_PROJ_HDR = "PROJECT #"
_NAME_HDR = "PROJECT NAME"
# extra scope sold later, carried as its own row - never the job's contract
_ADDENDUM_RE = re.compile(r"\bADDED\b|\bADDENDUM\b", re.I)

# Highest priority LAST is NOT how this reads - the list is in priority order,
# best first, and the first hit for a job wins.
_SOURCES = (
    ("WIP History/WIP 3-31-26.xlsx", "the 3-31-26 WIP report"),
    ("WIP History/WIP 12-31-25.xlsx", "the 12-31-25 WIP report"),
    ("MFD WIP Report 12-2025 & 3-2026.xlsx", "MFD division report"),
)


def _wip_dir() -> Path:
    return paths.onedrive_base() / WIP_DIR_DEFAULT


def _scan(path: Path) -> Dict[str, Tuple[float, Optional[float], str]]:
    """{project -> (contract, etc, 'sheet <name> row <n>')} from one report.

    The FIRST usable row per project wins; addendum rows never count. Sheets
    are read in workbook order, which puts the newest period first in the MFD
    report ('3-31-26' before '12-31-25')."""
    from openpyxl import load_workbook
    out: Dict[str, Tuple[float, Optional[float], str]] = {}
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return out
    try:
        for sn in wb.sheetnames:
            rows = list(wb[sn].iter_rows(max_col=24, values_only=True))
            hdr = next((i for i, r in enumerate(rows[:10])
                        if any(isinstance(v, str) and _CONTRACT_HDR in v.upper()
                               for v in r if v)), None)
            if hdr is None:
                continue
            H = {str(v).strip().upper(): j for j, v in enumerate(rows[hdr]) if v}
            pc, cc = H.get(_PROJ_HDR), H.get(_CONTRACT_HDR)
            ec, nc = H.get(_ETC_HDR), H.get(_NAME_HDR)
            if pc is None or cc is None:
                continue
            for i, r in enumerate(rows[hdr + 1:], start=hdr + 2):   # 1-based row #
                if pc >= len(r) or cc >= len(r):
                    continue
                p = str(r[pc] or "").strip().upper().replace(" ", "")
                v = r[cc]
                if not p or p in out or not isinstance(v, (int, float)) or v <= 0:
                    continue
                name = str(r[nc] or "") if nc is not None and nc < len(r) else ""
                if _ADDENDUM_RE.search(name):
                    continue                      # extra scope, not the contract
                etc = r[ec] if ec is not None and ec < len(r) else None
                out[p] = (float(v),
                          float(etc) if isinstance(etc, (int, float)) and etc > 0 else None,
                          f"sheet '{sn}' row {i}")
    finally:
        wb.close()
    return out


_cache: Optional[Dict[str, Tuple[float, Optional[float], str]]] = None


def contract_index() -> Dict[str, Tuple[float, Optional[float], str]]:
    """{project -> (contract, etc, source)} across the reports, best source
    first. `source` names the file, why it is trusted, and the row."""
    global _cache
    if _cache is not None:
        return _cache
    idx: Dict[str, Tuple[float, Optional[float], str]] = {}
    d = _wip_dir()
    for rel, why in _SOURCES:
        p = d / rel
        if not p.exists():
            continue
        for job, (ctr, etc, where) in _scan(p).items():
            if job not in idx:                       # first source wins
                idx[job] = (ctr, etc, f"{Path(rel).name} ({why}) · {where}")
    _cache = idx
    return idx


def contract_for(job: str) -> Tuple[Optional[float], Optional[float], str]:
    """(contract, etc, source) for one job; (None, None, "") when no report
    names it - the caller must then say so rather than substitute a number."""
    return contract_index().get(job.upper(), (None, None, ""))
