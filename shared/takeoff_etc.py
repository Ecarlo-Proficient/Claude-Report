"""Takeoff → ETC (budget) extractor — the ONE copy, importable by any tool.

The verified logic used to live in `one-offs/rp_schedule_wip_preview.py`; it was
moved here (2026-08-07) the moment a second tool — the WIP reader — needed it, so
`find_takeoff_etc` is never cross-imported between tool folders. The one-off now
imports these names from here; behaviour is unchanged (verified on 64 active RP
takeoffs — every count, LF, cut-list row and cost tied exactly).

WHAT IT DOES: given a project FOLDER, pick the best takeoff workbook for the scope
and read the BUDGET from the cost sheet's OWN subtotal cells (not Σ items):
  · slab scope → SL + PR bands of 'JobTread Cost Gral' (or a commercial 'BID'
    sheet AP1948/AP1961 when a CP PM helped on the RP job)
  · ftw scope  → the FW band
Returns (path, budget, note, fragment): `note` names the sheet + cells, `fragment`
is a jump-link into them. budget is None when no cost sheet is present.
"""
import re
from pathlib import Path

from openpyxl import load_workbook

# Scope words that mark a proposal/takeoff as a SIDE scope (not the base slab)
_SIDE_TOKENS = {"POOL", "CASITA", "CABANA", "FENCE", "CAPS", "FLATWORK",
                "RETAINING", "COURTYARD", "WALL", "PATIO", "DRIVEWAY",
                "FOOTINGS", "FOOTER", "PAVING"}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").upper().strip())


def _desc_tokens(desc: str):
    return {w for w in re.split(r"[^A-Z]+", desc or "")
            if len(w) > 3 and w not in ("POUR", "READY", "MAKE")}


def _score_name(name_u: str, scope: str, desc: str) -> float:
    """Rank a proposal/takeoff filename for the (scope, description) asked."""
    toks = _desc_tokens(desc)
    side_in_name = {t for t in _SIDE_TOKENS if t in name_u}
    score = 0.0
    score += 2.0 * len({t for t in toks if t in name_u})       # desc words hit
    if scope == "ftw":
        score += 3.0 if ("FLATWORK" in name_u or "FTW" in name_u) else 0.0
    else:
        # Slab wants the BASE proposal — side-scope names only help when the
        # schedule description itself asked for them.
        stray = side_in_name - toks
        score -= 2.0 * len(stray)
    if "REVISED" in name_u or "UPDATED" in name_u:
        score += 1.0
    return score


# ─────────────────── takeoff last sheet → ETC ──────────────────────
def _cost_sheet_totals(ws):
    """Parse 'JobTread Cost Gral' in BOTH layouts:
      A) code rows (SL#/PR#/FW# in col A, qty col C, unit cost col D)
      B) banded rows (SLAB/PIERS/FLATWORK in col A, items col B)

    THE NUMBER THAT COUNTS is the template's OWN subtotal cell at the foot
    of each band — NOT Σ(qty × cost) of the visible items. RP5542 FLATWORK:
    the visible items sum to barely half of D33 (the blue subtotal the clerk
    keys into the General List) — its formula pulls from the Flatwork takeoff
    sheet beyond these rows (the user 2026-07-17). Items are kept only as
    the fallback when the subtotal cell errors (#N/A) + a mismatch check.

    Returns {'SL'|'PR'|'FW': {'sub': float|None, 'cell': 'D33'|None,
                              'items': float}}."""
    bands = {k: {"sub": None, "cell": None, "items": 0.0}
             for k in ("SL", "PR", "FW")}
    band = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
        a = _norm(row[0].value if len(row) > 0 else "")
        b = _norm(row[1].value if len(row) > 1 else "")
        qty = row[2].value if len(row) > 2 else None
        cost_cell = row[3] if len(row) > 3 else None
        cost = cost_cell.value if cost_cell is not None else None
        m = re.match(r"^(SL|PR|FW)\d", a)
        if m:                                           # layout A code row
            band = m.group(1)
            try:
                bands[band]["items"] += float(qty or 0) * float(cost or 0)
            except (TypeError, ValueError):
                pass
            continue
        if a.startswith("SLAB"):
            band = "SL"
        elif a.startswith("PIER"):
            band = "PR"
        elif a.startswith("FLATWORK"):
            band = "FW"
        if band and b and b not in ("DESCRIPTION",):    # layout B item row
            try:
                bands[band]["items"] += float(qty or 0) * float(cost or 0)
            except (TypeError, ValueError):
                pass
            continue
        # Band-foot subtotal: no code, no item label, a value in the COST
        # column — the template's own total (first one after the band wins).
        if band and not a and not b and cost is not None \
                and bands[band]["sub"] is None and bands[band]["cell"] is None:
            bands[band]["cell"] = cost_cell.coordinate
            if isinstance(cost, (int, float)):
                bands[band]["sub"] = float(cost)
            # strings like '#N/A' leave sub=None → items fallback
    return bands


def find_takeoff_etc(folder: Path, job: str, scope: str, desc: str):
    """Best takeoff for the scope → BUDGET from its cost sheet's own
    subtotal cells. Returns (path, budget, note, fragment) — note names the
    sheet + cells the number came from; fragment jump-links there.

    COMMERCIAL TAKEOFF WINS for slab scope (the user 2026-07-21, RP6586):
    when the CP PM helps on an RP job he uses the CP template — a workbook
    with a 'BID' sheet, budget in AP1948 (AP1961 in some revisions). Those
    files may not carry the RP# in the name ('Peninsula Takeoff …'), so any
    *takeoff*-named workbook is a candidate, and a found BID sheet beats the
    residential 'JobTread Cost Gral' (which holds partial garbage on those
    jobs)."""
    cands = []
    try:
        for f in folder.iterdir():
            if f.suffix.lower() not in (".xlsm", ".xlsx"):
                continue
            if f.name.startswith("~$"):
                continue
            n = _norm(f.name)
            if n.startswith(job) or job in n or "TAKEOFF" in n:
                cands.append(f)
    except OSError:
        return None, None, "folder unreadable", None
    if not cands:
        return None, None, "No budget takeoff in this folder — add it", None
    cands.sort(key=lambda f: (_score_name(_norm(f.name), scope, desc),
                              f.stat().st_mtime), reverse=True)

    if scope != "ftw":
        for f in cands[:6]:
            try:
                wb = load_workbook(f, data_only=True, read_only=True)
            except Exception:
                continue
            bid = next((s for s in wb.sheetnames if _norm(s) == "BID"), None)
            if bid is None:
                wb.close()
                continue
            ws = wb[bid]
            for ref in ("AP1948", "AP1961"):
                v = ws[ref].value
                if isinstance(v, (int, float)) and v:
                    wb.close()
                    return (f, round(float(v), 2),
                            f"Commercial Takeoff '{bid}' {ref}",
                            f"#'{bid}'!{ref}")
            wb.close()

    found_takeoff = False
    for f in cands[:4]:
        try:
            wb = load_workbook(f, data_only=True, read_only=True)
        except Exception:
            continue
        found_takeoff = True
        sheet = next((s for s in wb.sheetnames if "COST GRAL" in _norm(s)), None)
        if sheet is None:
            wb.close()
            continue
        bands = _cost_sheet_totals(wb[sheet])
        wb.close()
        # A SIDE-SCOPE takeoff (its name matches the schedule description —
        # fence, pool house, caps…) is a whole little job of its own: its
        # entire cost sheet IS the scope's ETC. Only the BASE takeoff splits
        # slab (SL+PR) vs flatwork (FW) bands.
        if _desc_tokens(desc) & set(_norm(f.name).split()):
            keys = ("SL", "PR", "FW")
        else:
            keys = ("FW",) if scope == "ftw" else ("SL", "PR")
        etc, cells, notes = 0.0, [], []
        for k in keys:
            b = bands[k]
            if b["sub"] is not None:
                etc += b["sub"]
                cells.append(b["cell"])
            elif b["items"]:
                etc += b["items"]
                cells.append(f"{k} items (subtotal cell "
                             f"{b['cell'] or 'missing'} unreadable)")
            if (b["sub"] is not None and b["items"]
                    and abs(b["sub"] - b["items"]) > 1):
                notes.append(f"{k} subtotal ${b['sub']:,.0f} ≠ item rows "
                             f"${b['items']:,.0f} — template pulls extra scope")
        if etc:
            src = f"'{sheet}' {' + '.join(cells)}"
            note = "; ".join([src] + notes)
            frag = f"#'{sheet}'!{next((c for c in cells if re.match(r'^[A-Z]+[0-9]+$', c)), 'A1')}"
            return f, round(etc, 2), note, frag
    return (cands[0], None,
            ("Missing 'JobTread Cost Gral' sheet — add the budget sheet to "
             "the takeoff" if found_takeoff
             else "No budget takeoff in this folder — add it"), None)
