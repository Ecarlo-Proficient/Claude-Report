"""
breakeven.py — how much we must SELL and COLLECT to cover overhead.

The question (the user 2026-07-29): "I have X in contracts won — how much sales
do I need to cover this month's overhead, weekly and monthly, and how much do I
need to collect?"

The business term is **break-even analysis**. The core identity:

    break-even revenue = fixed overhead ÷ gross margin %

Gross margin % is the *contribution margin ratio* — the share of each revenue
dollar left after direct job costs to pay overhead. Its inverse is the revenue
needed per dollar of overhead (at 17.9% GM, $5.59 of revenue per $1).

Two numbers, not one — the caveat that matters in construction:
  • SELL    (accrual break-even) — revenue that must be BOOKED/billed.
  • COLLECT (cash break-even)    — cash that must actually ARRIVE. Same target,
    but lagged by DSO, and retainage is earned yet uncollectible until close.
    A company can clear accrual break-even and still miss payroll.

And the "contracts won" angle: **backlog coverage** = unbilled backlog ÷ monthly
break-even = months of overhead already sold. High coverage means the job is to
BILL AND COLLECT what's won, not to sell more.

Source: the health_dashboard P&L blocks (MTD / YTD current / YTD prior). Read
only — no QBO.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

# P&L block headers written by qbo_health
BLOCK_MTD = "month to date"
BLOCK_YTD = "year to date (current year)"
BLOCK_PRIOR = "year to date (prior year)"

_TOTALS = {
    "income": "total income",
    "cogs": "total cost of goods sold",
    "gross_profit": "total gross profit",
    "overhead": "total expenses",
    "net_operating": "total net operating income",
}
DAYS_PER_MONTH = 30.4375


def read_pl_blocks(path: Path) -> Dict[str, dict]:
    """{'mtd'|'ytd'|'prior': {income, cogs, gross_profit, overhead,
    net_operating}} from the health_dashboard P&L sheet.

    Blocks are delimited by their header rows, so a total is only ever credited
    to the period it sits under — the sheet repeats the same labels three times.
    """
    out: Dict[str, dict] = {}
    if not path.exists():
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    if "P&L" not in wb.sheetnames:
        wb.close()
        return out
    ws = wb["P&L"]
    current: Optional[str] = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = str(row[0]).strip() if row[0] else ""
        val = row[1] if len(row) > 1 else None
        low = label.lower()
        if low.startswith(BLOCK_MTD):
            current = "mtd"
            out.setdefault(current, {})
            continue
        if low.startswith(BLOCK_YTD):
            current = "ytd"
            out.setdefault(current, {})
            continue
        if low.startswith(BLOCK_PRIOR):
            current = "prior"
            out.setdefault(current, {})
            continue
        if current is None:
            continue
        for key, want in _TOTALS.items():
            if low == want and key not in out[current]:
                try:
                    out[current][key] = float(val)
                except (TypeError, ValueError):
                    pass
    wb.close()
    return out


def build(health_path: Path, as_of: Optional[dt.datetime] = None,
          backlog: float = 0.0, ar: float = 0.0, retainage: float = 0.0,
          dso_days: Optional[float] = None) -> dict:
    """Break-even model from the YTD block, annualised over the elapsed period.

    `as_of` is the health_dashboard generation time (the YTD cut-off). Overhead
    is spread evenly across the elapsed period — see `caveat` in the result.
    """
    blocks = read_pl_blocks(health_path)
    ytd = blocks.get("ytd", {})
    prior = blocks.get("prior", {})
    income = ytd.get("income") or 0.0
    gp = ytd.get("gross_profit") or 0.0
    overhead = ytd.get("overhead") or 0.0
    if not income or not overhead:
        return {"ok": False, "reason": "P&L YTD block not found or empty"}

    as_of = as_of or dt.datetime.now()
    days = max((as_of.date() - dt.date(as_of.year, 1, 1)).days, 1)
    months = days / DAYS_PER_MONTH
    weeks = days / 7.0

    gm = gp / income if income else 0.0
    oh_month = overhead / months
    oh_week = overhead / weeks
    be_month = oh_month / gm if gm else 0.0
    be_week = oh_week / gm if gm else 0.0
    rev_run_month = income / months
    margin_of_safety = ((rev_run_month - be_month) / rev_run_month
                        if rev_run_month else 0.0)

    prior_gm = ((prior.get("gross_profit") or 0.0) / prior["income"]
                if prior.get("income") else None)

    return {
        "ok": True,
        "as_of": as_of, "days": days, "months": months, "weeks": weeks,
        "income_ytd": income, "gross_profit_ytd": gp, "overhead_ytd": overhead,
        "net_operating_ytd": ytd.get("net_operating"),
        "gm": gm,
        "overhead_month": oh_month, "overhead_week": oh_week,
        "breakeven_month": be_month, "breakeven_week": be_week,
        "revenue_per_overhead_dollar": (1 / gm) if gm else 0.0,
        "run_rate_month": rev_run_month,
        "margin_of_safety": margin_of_safety,
        "coverage_ratio": (rev_run_month / be_month) if be_month else 0.0,
        # contracts won → months of overhead already sold
        "backlog": backlog,
        "backlog_coverage_months": (backlog / be_month) if be_month else 0.0,
        "backlog_gross_profit": backlog * gm,
        # the cash side
        "ar": ar, "retainage": retainage, "dso_days": dso_days,
        "ar_coverage_months": (ar / be_month) if be_month else 0.0,
        # prior-year comparison
        "prior_gm": prior_gm,
        "prior_overhead": prior.get("overhead"),
        "prior_net_operating": prior.get("net_operating"),
        "caveat": ("Overhead is the YTD total spread evenly over the elapsed "
                   "period; a one-off (fee spike, insurance credit) shifts the "
                   "monthly figure. A 12-month trailing average is firmer."),
    }


def audit_rows(m: dict, source: Optional[Path] = None) -> list:
    """[(item, value, where_it_came_from)] — the full trail behind every
    break-even figure, so the numbers can be checked against QBO by hand
    (the user 2026-08-03: "I need to audit the numbers but can't find them").
    """
    if not m.get("ok"):
        return []
    src = (f"{source.parent.name}/{source.name}" if source else "health_dashboard.xlsx")
    pl = f"{src} → 'P&L' sheet → 'Year to Date (current year)' block"
    gm_pct = m["gm"] * 100

    def money(v):
        return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"

    return [
        ("── INPUTS (all from QBO via qbo_health) ──", "", ""),
        ("Revenue YTD", money(m["income_ytd"]), pl + " → 'Total Income'"),
        ("Direct job costs YTD (COGS)", money(m["income_ytd"] - m["gross_profit_ytd"]),
         pl + " → 'Total Cost of Goods Sold'"),
        ("Gross profit YTD", money(m["gross_profit_ytd"]),
         pl + " → 'Total Gross Profit'"),
        ("Overhead YTD (operating expenses)", money(m["overhead_ytd"]),
         pl + " → 'Total Expenses'"),
        ("Net operating income YTD",
         money(m["net_operating_ytd"] or 0), pl + " → 'Total Net Operating Income'"),
        ("Period covered", f"Jan 1 → {m['as_of']:%b %d, %Y}",
         f"{m['days']} days = {m['months']:.2f} months = {m['weeks']:.1f} weeks"),
        ("── DERIVED (the formulas) ──", "", ""),
        ("Gross margin %", f"{gm_pct:.2f}%",
         f"{money(m['gross_profit_ytd'])} gross profit ÷ {money(m['income_ytd'])} revenue"),
        ("Overhead per month", money(m["overhead_month"]),
         f"{money(m['overhead_ytd'])} ÷ {m['months']:.2f} months"),
        ("Overhead per week", money(m["overhead_week"]),
         f"{money(m['overhead_ytd'])} ÷ {m['weeks']:.1f} weeks"),
        ("BREAK-EVEN sales / month", money(m["breakeven_month"]),
         f"{money(m['overhead_month'])} overhead ÷ {gm_pct:.2f}% gross margin"),
        ("BREAK-EVEN sales / week", money(m["breakeven_week"]),
         f"{money(m['overhead_week'])} overhead ÷ {gm_pct:.2f}% gross margin"),
        ("Revenue per $1 of overhead",
         f"${m['revenue_per_overhead_dollar']:,.2f}", f"1 ÷ {gm_pct:.2f}%"),
        ("Actual run rate / month", money(m["run_rate_month"]),
         f"{money(m['income_ytd'])} ÷ {m['months']:.2f} months"),
        ("Coverage ratio", f"{m['coverage_ratio']:.2f}×",
         f"{money(m['run_rate_month'])} run rate ÷ {money(m['breakeven_month'])} break-even"),
        ("Margin of safety", f"{m['margin_of_safety']*100:.0f}%",
         "(run rate − break-even) ÷ run rate"),
        ("Backlog coverage", f"{m['backlog_coverage_months']:.1f} months",
         f"{money(m['backlog'])} backlog ÷ {money(m['breakeven_month'])} monthly break-even"),
        ("  ↳ backlog source", "", "WIP master 'Test-Master' tab → LEFT TO BILL, "
                                   "Active rows (computed: contract − billed)"),
        ("AR coverage", f"{m['ar_coverage_months']:.1f} months",
         f"{money(m['ar'])} AR ÷ monthly break-even  ·  AR from the AR Aging sheet"),
        ("── PRIOR YEAR (same YTD window) ──", "", ""),
        ("Prior-year gross margin",
         f"{m['prior_gm']*100:.2f}%" if m.get("prior_gm") else "—",
         pl.replace("current year", "prior year")),
        ("Prior-year net operating income",
         money(m["prior_net_operating"]) if m.get("prior_net_operating") is not None else "—",
         "same block → 'Total Net Operating Income'"),
        ("── CAVEAT ──", "", m["caveat"]),
    ]


def rows_for_display(m: dict) -> list:
    """[(metric, value_str, detail, class)] — same shape the tracker/dashboard
    metric tables use. class: g good · r bad · a warn · n neutral."""
    if not m.get("ok"):
        return []

    def money(v):
        return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"

    mos = m["margin_of_safety"]
    cov = m["backlog_coverage_months"]
    return [
        ("Break-even SALES — per month", money(m["breakeven_month"]),
         f"revenue needed to cover ${m['overhead_month']:,.0f}/mo overhead at "
         f"{m['gm']*100:.1f}% GM", "n"),
        ("Break-even SALES — per week", money(m["breakeven_week"]),
         f"${m['overhead_week']:,.0f}/wk overhead ÷ gross margin", "n"),
        ("Break-even COLLECTIONS — per month", money(m["breakeven_month"]),
         (f"cash that must ARRIVE; work billed ~{m['dso_days']:.0f} days earlier"
          if m.get("dso_days") else
          "cash that must ARRIVE — lagged by DSO, retainage excluded"), "a"),
        ("Revenue needed per $1 of overhead",
         f"${m['revenue_per_overhead_dollar']:,.2f}",
         "the inverse of gross margin", "n"),
        ("Actual run rate", money(m["run_rate_month"]) + "/mo",
         f"{m['coverage_ratio']:.2f}× break-even", "g" if m["coverage_ratio"] >= 1 else "r"),
        ("Margin of safety", f"{mos*100:.0f}%",
         "revenue could fall this much before a loss",
         "g" if mos >= 0.25 else "a" if mos > 0 else "r"),
        ("Backlog coverage", f"{cov:.1f} months",
         f"{money(m['backlog'])} won & unbilled = "
         f"{money(m['backlog_gross_profit'])} of gross profit",
         "g" if cov >= 3 else "a" if cov >= 1 else "r"),
        ("AR coverage", f"{m['ar_coverage_months']:.1f} months",
         f"{money(m['ar'])} receivable ÷ monthly break-even", "n"),
        ("Retainage (earned, not collectible)", money(m["retainage"]),
         "counts as profit but pays no bills until job close", "a"),
    ]
