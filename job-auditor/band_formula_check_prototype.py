#!/usr/bin/env python3
"""Is the Cost Gral sheet CUMULATIVE by design, or is =SUM(D10:D17) a mistake?

The estimator says the range is intentional. There is a reading where he is
right: if the band-foot rows are RUNNING totals — slab, then slab+piers, then
slab+piers+flatwork — then =SUM(D10:D17) is correct and OUR ETC (SL_sub +
PR_sub) is what double-counts.

The discriminator is the FLATWORK subtotal:
    cumulative design → its range STARTS at the piers subtotal row
    band-only design  → its range STARTS at the first FW item row

Reports every band-foot formula and classifies the sheet. Read-only.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_R = str(Path(__file__).resolve().parent.parent)   # repo root — self-locating
for _p in (_R, _R + "/one-offs", _R + "/wip"):
    sys.path.insert(0, _p)

import rp_wip_reader as RP            # noqa: E402
import rp_schedule_wip_preview as P   # noqa: E402

# 2 known-bad + 1 known-good + 5 from the unclassifiable population
TARGETS = ["RP7482", "RP7529", "RP7470",
           "RP7083", "RP7518", "RP7490", "RP7553", "RP7455"]

RANGE = re.compile(r"SUM\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)", re.I)


def analyse(tk):
    from openpyxl import load_workbook as _lw
    try:
        v = _lw(tk, data_only=True)
        f = _lw(tk, data_only=False)
    except Exception as e:
        return None, f"unreadable ({type(e).__name__})"
    try:
        name = next((n for n in v.sheetnames if "cost gral" in n.lower()), None)
        if not name:
            return None, "no Cost Gral sheet"
        wv, wf = v[name], f[name]
        rows = []                      # (row, band, code, value, formula)
        band = None
        first_item_row = {}
        for r in range(1, min(wv.max_row, 140) + 1):
            a = wv.cell(r, 1).value
            d = wv.cell(r, 4).value
            fd = wf.cell(r, 4).value
            an = str(a).strip().upper() if a is not None else ""
            m = re.match(r"^(SL|PR|FW)\d", an)
            if m:
                band = m.group(1)
                first_item_row.setdefault(band, r)
                rows.append((r, band, an, d, fd))
                continue
            if an.startswith("SLAB"):
                band = "SL"
            elif an.startswith("PIER"):
                band = "PR"
            elif an.startswith("FLATWORK"):
                band = "FW"
            if a is None and (d is not None or isinstance(fd, str)):
                rows.append((r, band, "<SUBTOTAL>", d, fd))
        return (rows, first_item_row), ""
    finally:
        v.close()
        f.close()


def main():
    (_k, sp) = P.latest_schedule(RP.SCHEDULE_DIR)
    sched = {s["job"]: s for s in P.read_main_schedule(sp)}
    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)

    verdicts = []
    for job in TARGETS:
        s = sched.get(job)
        print("\n" + "=" * 76)
        print(f"  {job}  {s['address'] if s else '?'}")
        print("=" * 76)
        if not s:
            print("  not on the current schedule")
            continue
        fs = sorted(rp_to_folders.get(job, ()), key=lambda f: (f.parent.name, f.name))
        folder = fs[0] if fs else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            folder = RP.match_by_address({"house": parts[0] if parts else "",
                                          "street": parts[1] if len(parts) > 1
                                          else s["address"]}, addr_folders)
        if folder is None:
            print("  no folder")
            continue
        tk, _e, _n, _f = P.find_takeoff_etc(folder, job, "slab", s["desc"])
        if tk is None:
            print("  no takeoff")
            continue
        got, err = analyse(tk)
        if got is None:
            print(f"  {err}")
            continue
        rows, first = got
        print(f"  takeoff: {Path(tk).name}")
        subs = [(r, b, d, fd) for (r, b, c, d, fd) in rows if c == "<SUBTOTAL>"]
        for r, b, d, fd in subs:
            dv = f"{d:,.2f}" if isinstance(d, (int, float)) else str(d)
            print(f"     r{r:<4} band={b or '?':<3} value={dv:<15} formula={fd}")
        # classify using the FW subtotal (and PR as corroboration)
        verdict = "unknown"
        pr_sub = next((x for x in subs if x[1] == "PR"), None)
        fw_sub = next((x for x in subs if x[1] == "FW"), None)
        sl_sub = next((x for x in subs if x[1] == "SL"), None)
        def start_of(x):
            if not x or not isinstance(x[3], str):
                return None
            m = RANGE.search(x[3])
            return int(m.group(2)) if m else None
        pr_start, fw_start = start_of(pr_sub), start_of(fw_sub)
        sl_row = sl_sub[0] if sl_sub else None
        pr_row = pr_sub[0] if pr_sub else None
        notes = []
        if pr_start is not None and sl_row is not None:
            notes.append("PR range starts AT the SL subtotal row"
                         if pr_start == sl_row else
                         f"PR range starts at r{pr_start} (band items begin r{first.get('PR')})")
        if fw_start is not None and pr_row is not None:
            if fw_start == pr_row:
                verdict = "CUMULATIVE (chains off the previous subtotal)"
            elif fw_start == first.get("FW"):
                verdict = "BAND-ONLY (FW starts at its own first item)"
            notes.append(f"FW range starts at r{fw_start} "
                         f"(PR subtotal r{pr_row}, FW items begin r{first.get('FW')})")
        for n in notes:
            print(f"     · {n}")
        print(f"     ⇒ {verdict}")
        verdicts.append((job, verdict))

    print("\n" + "=" * 76)
    print("  VERDICTS")
    for j, v in verdicts:
        print(f"    {j:<9} {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
