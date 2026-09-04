#!/usr/bin/env python3
"""Verify the claim: the Piers takeoff sheet carries NON-pier line items
(Tie wire / Scrape lot / Reset forms) that land in PR3 / PR6.

Checks, per job:
  1. dump 'Piers takeoff' rows around the subtotal cells F30 / F33
  2. show which line items each subtotal range actually covers
  3. confirm the genuine pier lines are zero on a no-piers job
  4. search 'Slab takeoff' for scrape lot / reset forms / tie wire — if the slab
     sheet already carries them, the cost would be DOUBLE counted, not misfiled
Read-only.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_R = str(Path(__file__).resolve().parent.parent)   # repo root — self-locating
for _p in (_R, _R + "/one-offs", _R + "/wip"):
    sys.path.insert(0, _p)

import rp_wip_reader as RP            # noqa: E402
import rp_schedule_wip_preview as P   # noqa: E402

JOBS = ["RP7518", "RP7491", "RP7492", "RP7553", "RP7591"]
WANT = re.compile(r"(?i)scrape|reset\s*form|tie\s*wire|stake|drill|labor|bell|void|jack|"
                  r"rebar|concrete|yard|pier")


def main():
    (_k, sp) = P.latest_schedule(RP.SCHEDULE_DIR)
    sched = {s["job"]: s for s in P.read_main_schedule(sp)}
    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)
    from openpyxl import load_workbook as lw

    for job in JOBS:
        s = sched.get(job)
        print("\n" + "=" * 78)
        print(f"  {job}  {s['address'] if s else '?'}")
        print("=" * 78)
        if not s:
            continue
        fs = sorted(rp_to_folders.get(job, ()), key=lambda f: (f.parent.name, f.name))
        folder = fs[0] if fs else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            folder = RP.match_by_address({"house": parts[0] if parts else "",
                                          "street": parts[1] if len(parts) > 1
                                          else s["address"]}, addr_folders)
        tk, _e, _n, _f = P.find_takeoff_etc(folder, job, "slab", s["desc"]) if folder else (None,)*4
        if tk is None:
            print("  no takeoff")
            continue
        try:
            v = lw(tk, data_only=True)
            f = lw(tk, data_only=False)
        except Exception as e:
            print(f"  unreadable: {e}")
            continue
        try:
            pn = next((n for n in v.sheetnames if "piers takeoff" in n.lower()), None)
            sn = next((n for n in v.sheetnames if "slab takeoff" in n.lower()), None)
            if pn is None:
                print("  no 'Piers takeoff' sheet")
                continue
            pv, pf = v[pn], f[pn]
            print(f"  takeoff: {Path(tk).name}")
            print(f"\n  ── '{pn}' rows 20..36  (col D label · col E amount) ──")
            for r in range(20, 37):
                lab = pv.cell(r, 4).value
                if lab is None:
                    lab = pv.cell(r, 3).value
                if lab is None:
                    lab = pv.cell(r, 2).value
                amt = pv.cell(r, 5).value
                famt = pf.cell(r, 6).value
                fcol = pf.cell(r, 6).value
                av = f"{amt:,.2f}" if isinstance(amt, (int, float)) else str(amt)
                ff = f"   F{r}={fcol}" if isinstance(fcol, str) and fcol.startswith("=") else ""
                if lab is None and amt is None and not ff:
                    continue
                print(f"     r{r:<3} {str(lab)[:40]:<40} E={av:<12}{ff}")
            # subtotal formulas
            print("\n  ── subtotal cells ──")
            for cell in ("F25", "F26", "F30", "F33"):
                print(f"     {cell}: value={pv[cell].value!r:<22} formula={pf[cell].value!r}")
            # slab sheet check
            if sn:
                print(f"\n  ── '{sn}': does it carry scrape lot / reset forms / tie wire? ──")
                sv = v[sn]
                found = []
                for row in sv.iter_rows(min_row=1, max_row=min(sv.max_row, 90)):
                    for c in row:
                        if isinstance(c.value, str) and re.search(
                                r"(?i)scrape|reset\s*form|tie\s*wire", c.value):
                            amt = sv.cell(c.row, 5).value
                            found.append((c.row, c.value[:34],
                                          f"{amt:,.2f}" if isinstance(amt, (int, float)) else amt))
                if found:
                    for r, lab, amt in found:
                        print(f"     r{r:<3} {lab:<36} E={amt}")
                else:
                    print("     none — slab sheet has no scrape lot / reset forms / tie wire")
        finally:
            v.close()
            f.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
