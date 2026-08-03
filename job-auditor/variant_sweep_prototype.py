#!/usr/bin/env python3
"""Which Cost Gral VARIANT is each takeoff, and what does our ETC reader get wrong?

The estimator's design (confirmed): the sheet holds TWO scopes, not three bands.
  FOUNDATION = slab + piers  → its SUB TOTAL is the piers-band foot (r18),
                               deliberately =SUM(<slab subtotal>:<last PR row>)
  FLATWORK   = its own scope → subtotal starts at the first FW item (r19)
No piers on the job ⇒ the PR item rows read 0 and you read the slab subtotal.

So there are two variants in circulation:
  CUMULATIVE  r18 = slab + piers   → correct ETC is r18 ALONE
  BAND_ONLY   r18 = piers only     → correct ETC is r10 + r18

Our reader always does r10 + r18, which is right for BAND_ONLY and
double-counts the slab on CUMULATIVE. Read-only; writes nothing.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_R = str(Path(__file__).resolve().parent.parent)   # repo root — self-locating
for _p in (_R, _R + "/one-offs", _R + "/wip"):
    sys.path.insert(0, _p)

import rp_wip_reader as RP            # noqa: E402
import rp_schedule_wip_preview as P   # noqa: E402

RANGE = re.compile(r"SUM\(\s*\$?[A-Z]+\$?(\d+)\s*:\s*\$?[A-Z]+\$?(\d+)\s*\)", re.I)


def read(tk):
    from openpyxl import load_workbook as _lw
    try:
        v = _lw(tk, data_only=True)
        f = _lw(tk, data_only=False)
    except Exception:
        return None
    try:
        nm = next((n for n in v.sheetnames if "cost gral" in n.lower()), None)
        if not nm:
            return None
        wv, wf = v[nm], f[nm]
        band, first = None, {}
        sl = pr = None          # (row, value, formula)
        pr_items = 0.0
        for r in range(1, min(wv.max_row, 140) + 1):
            a = wv.cell(r, 1).value
            d = wv.cell(r, 4).value
            fd = wf.cell(r, 4).value
            an = str(a).strip().upper() if a is not None else ""
            m = re.match(r"^(SL|PR|FW)\d", an)
            if m:
                band = m.group(1)
                first.setdefault(band, r)
                if band == "PR" and isinstance(d, (int, float)):
                    pr_items += float(d)
                continue
            if a is None and (d is not None or isinstance(fd, str)):
                if band == "SL" and sl is None:
                    sl = (r, d, fd)
                elif band == "PR" and pr is None:
                    pr = (r, d, fd)
        return sl, pr, pr_items, first
    finally:
        v.close()
        f.close()


def num(x):
    return float(x) if isinstance(x, (int, float)) else None


def main():
    (_k, sp) = P.latest_schedule(RP.SCHEDULE_DIR)
    sched = [s for s in P.read_main_schedule(sp) if not s["job"].startswith("CP")]
    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)

    cum, band, unk, skip = [], [], [], []
    seen = set()
    for s in sched:
        job = s["job"]
        if job in seen:
            continue
        seen.add(job)
        fs = sorted(rp_to_folders.get(job, ()), key=lambda f: (f.parent.name, f.name))
        folder = fs[0] if fs else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            folder = RP.match_by_address({"house": parts[0] if parts else "",
                                          "street": parts[1] if len(parts) > 1
                                          else s["address"]}, addr_folders)
        if folder is None:
            skip.append((job, s["address"], "no folder"))
            continue
        tk, _e, _n, _f = P.find_takeoff_etc(folder, job, "slab", s["desc"])
        if tk is None:
            skip.append((job, s["address"], "no takeoff"))
            continue
        got = read(tk)
        if got is None:
            skip.append((job, s["address"], "unreadable / no Cost Gral"))
            continue
        sl, pr, pr_items, first = got
        if not sl or not pr or not isinstance(pr[2], str):
            unk.append((job, s["address"], "no readable PR subtotal formula"))
            continue
        m = RANGE.search(pr[2])
        if not m:
            unk.append((job, s["address"], f"unparsed formula {str(pr[2])[:30]}"))
            continue
        start = int(m.group(1))
        slv, prv = num(sl[1]), num(pr[1])
        rec = dict(job=job, addr=s["address"], sl_row=sl[0], pr_row=pr[0],
                   start=start, sl=slv, pr=prv, items=pr_items,
                   na=(prv is None))
        if start == sl[0]:
            cum.append(rec)
        elif start == first.get("PR"):
            band.append(rec)
        else:
            unk.append((job, s["address"], f"range starts r{start} (SL sub r{sl[0]}, "
                                           f"PR items r{first.get('PR')})"))

    print("\n  COST GRAL VARIANT SWEEP — which sheets does our ETC reader mis-read?")
    print("  " + "─" * 72)
    print(f"    CUMULATIVE (r18 = slab+piers, per the estimator's design) : {len(cum)}")
    print(f"    BAND_ONLY  (r18 = piers only)                             : {len(band)}")
    print(f"    unclassified                                              : {len(unk)}")
    print(f"    skipped                                                   : {len(skip)}")

    over = [r for r in cum if not r["na"]]
    hidden = [r for r in cum if r["na"]]
    print(f"\n  ── CUMULATIVE sheets, subtotal READABLE — our ETC OVERSTATES by one slab ──")
    print(f"    {'JOB':<9} {'ADDRESS':<30} {'ETC ours':>11} {'ETC correct':>12} {'OVER':>11}")
    tot = 0.0
    for r in sorted(over, key=lambda r: -(r["sl"] or 0)):
        ours = (r["sl"] or 0) + (r["pr"] or 0)
        good = r["pr"] or 0
        tot += ours - good
        print(f"    {r['job']:<9} {r['addr'][:30]:<30} {ours:>11,.0f} {good:>12,.0f} "
              f"{ours - good:>11,.0f}")
    print(f"    {'':<40} {'TOTAL OVERSTATED':>24} {tot:>11,.0f}")

    print(f"\n  ── CUMULATIVE sheets with an #N/A subtotal — ETC UNDERSTATES (band dropped) ──")
    tot2 = 0.0
    for r in sorted(hidden, key=lambda r: -(r["sl"] or 0)):
        # reader drops the #N/A band ⇒ ETC = slab only; correct = slab + pier items
        ours = r["sl"] or 0
        good = (r["sl"] or 0) + r["items"]
        tot2 += good - ours
        print(f"    {r['job']:<9} {r['addr'][:30]:<30} {ours:>11,.0f} {good:>12,.0f} "
              f"{good - ours:>11,.0f}")
    print(f"    {'':<40} {'TOTAL UNDERSTATED':>24} {tot2:>11,.0f}")

    if band:
        print(f"\n  ── BAND_ONLY sheets — our reader is CORRECT on these ──")
        for r in band[:14]:
            print(f"    {r['job']:<9} {r['addr'][:30]:<30} "
                  f"ETC {(r['sl'] or 0) + (r['pr'] or 0):>11,.0f}")
    if unk:
        print(f"\n  ── unclassified ──")
        for j, a, w in unk:
            print(f"    {j:<9} {a[:30]:<30} {w}")
    if skip:
        print(f"\n  ── skipped ──")
        for j, a, w in skip:
            print(f"    {j:<9} {a[:30]:<30} {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
