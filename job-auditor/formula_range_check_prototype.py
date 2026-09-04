#!/usr/bin/env python3
"""BLAST RADIUS: how many active takeoffs have the PIERS-subtotal formula bug?

The bug: the PIERS band subtotal is `=SUM(D10:D17)` where D10 is the SLAB
band's own subtotal, not a pier row. So

    PR_subtotal = SL_subtotal + Σ(PR items)

and because the ETC is SL_subtotal + PR_subtotal, the SLAB IS COUNTED TWICE.
ETC is overstated by exactly one slab.

Numeric test (no formula parsing needed, so it survives template variants):
    BUG      when |PR_sub - (SL_sub + Σ PR items)| < $1
    CORRECT  when |PR_sub -            Σ PR items|  < $1
Anything else is UNKNOWN and reported as such, never assumed clean.
"""
from __future__ import annotations

import sys
from pathlib import Path

_R = str(Path(__file__).resolve().parent.parent)   # repo root — self-locating
for _p in (_R, _R + "/one-offs", _R + "/wip"):
    sys.path.insert(0, _p)

import rp_wip_reader as RP            # noqa: E402
import rp_schedule_wip_preview as P   # noqa: E402


def read_bands(tk):
    """→ (sl_sub, pr_sub, pr_items, pr_formula) straight off the sheet."""
    from openpyxl import load_workbook as _lw
    try:
        v = _lw(tk, data_only=True)
        f = _lw(tk, data_only=False)
    except Exception:
        return None
    try:
        name = next((n for n in v.sheetnames if "cost gral" in n.lower()), None)
        if not name:
            return None
        wv, wf = v[name], f[name]
        sl_sub = pr_sub = pr_formula = None
        pr_items = 0.0
        band = None
        for r in range(1, min(wv.max_row, 120) + 1):
            a = wv.cell(r, 1).value
            d = wv.cell(r, 4).value
            an = str(a).strip().upper() if a is not None else ""
            if an[:2] in ("SL", "PR", "FW") and any(ch.isdigit() for ch in an[2:4]):
                band = an[:2]
                if band == "PR" and isinstance(d, (int, float)):
                    pr_items += float(d)
                continue
            if a is None and isinstance(d, (int, float)):
                if band == "SL" and sl_sub is None:
                    sl_sub = float(d)
                elif band == "PR" and pr_sub is None:
                    pr_sub = float(d)
                    fd = wf.cell(r, 4).value
                    pr_formula = fd if isinstance(fd, str) else None
        return sl_sub, pr_sub, pr_items, pr_formula
    finally:
        v.close()
        f.close()


def main():
    print("\n  PIERS-SUBTOTAL FORMULA BUG — blast radius")
    print("  " + "─" * 72)
    (_k, sp) = P.latest_schedule(RP.SCHEDULE_DIR)
    sched = [s for s in P.read_main_schedule(sp) if not s["job"].startswith("CP")]
    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)

    bug, ok, unknown, skipped = [], [], [], []
    seen = set()
    for s in sched:
        job = s["job"]
        if job in seen:
            continue
        seen.add(job)
        fs = sorted(rp_to_folders.get(job, ()), key=lambda f: (f.parent.name, f.name))
        folder = fs[0] if fs else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            folder = RP.match_by_address({"house": parts[0] if parts else "",
                                          "street": parts[1] if len(parts) > 1
                                          else s["address"]}, addr_folders)
        if folder is None:
            skipped.append((job, s["address"], "no folder"))
            continue
        tk, _e, _n, _f = P.find_takeoff_etc(folder, job, "slab", s["desc"])
        if tk is None:
            skipped.append((job, s["address"], "no takeoff"))
            continue
        got = read_bands(tk)
        if got is None:
            skipped.append((job, s["address"], "unreadable / no Cost Gral"))
            continue
        sl, pr, items, formula = got
        if pr is None or sl is None:
            unknown.append((job, s["address"], sl, pr, items, formula))
            continue
        if abs(pr - (sl + items)) < 1.0:
            bug.append((job, s["address"], sl, pr, items, formula))
        elif abs(pr - items) < 1.0:
            ok.append((job, s["address"], sl, pr, items, formula))
        else:
            unknown.append((job, s["address"], sl, pr, items, formula))

    tot = len(bug) + len(ok) + len(unknown)
    print(f"\n  takeoffs tested : {tot}")
    print(f"  ✗ BUG           : {len(bug)}   (PR subtotal = SL subtotal + PR items)")
    print(f"  ✓ correct       : {len(ok)}")
    print(f"  ? unknown       : {len(unknown)}")
    print(f"  – skipped       : {len(skipped)}")

    if bug:
        over = sum(b[2] for b in bug)
        print(f"\n  ── ETC OVERSTATED (by exactly one slab each) — total ${over:,.0f} ──")
        print(f"    {'JOB':<9} {'ADDRESS':<30} {'SLAB $':>11} {'ETC now $':>12} "
              f"{'ETC true $':>12} {'OVER $':>11}")
        for job, addr, sl, pr, items, _f in sorted(bug, key=lambda b: -b[2]):
            print(f"    {job:<9} {addr[:30]:<30} {sl:>11,.0f} {sl + pr:>12,.0f} "
                  f"{sl + items:>12,.0f} {sl:>11,.0f}")
    if ok:
        print("\n  ── correct takeoffs ──")
        for job, addr, sl, pr, items, _f in ok[:10]:
            print(f"    {job:<9} {addr[:30]:<30} PR sub {pr:>10,.0f} = items {items:>10,.0f}")
    if unknown:
        print("\n  ── UNKNOWN (do not assume clean) ──")
        for job, addr, sl, pr, items, f in unknown:
            print(f"    {job:<9} {addr[:26]:<26} SL={sl} PR={pr} items={items:,.0f}  {str(f)[:26]}")
    if skipped:
        print("\n  ── skipped ──")
        for job, addr, why in skipped:
            print(f"    {job:<9} {addr[:30]:<30} {why}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
