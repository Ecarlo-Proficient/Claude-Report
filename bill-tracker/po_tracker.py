"""
po_tracker.py — read the office PO tracker workbook (READ-ONLY) and reconcile it
against QBO purchase orders for the bill-tracker's "Unused PO" audit.

The office keeps a manual PO log on OneDrive. QBO carries the real PO
transactions (POStatus Open/Closed, bills linked). This module is the "two tools,
one story" join: match tracker `P.O. #` to QBO `DocNumber` and surface POs that
need AP attention.

Flags (the user 2026-08-25):
  1. "Open, no bill"        — QBO PO is Open with no bill linked (issued, nothing billed).
  2. "Stale >Nd"            — an Open+unbilled QBO PO older than STALE_DAYS (default 60).
  3. "On tracker, not in QBO" — a recent tracker PO # with no matching QBO PO
                                (logged/requested but never issued). Bounded to
                                the last TRACKER_MISSING_WINDOW_DAYS so ancient
                                tracker rows QBO no longer retains don't flood it.

The tracker is manual and lags — its data currently runs months behind. We stamp
the tracker's freshness (max data date) on the sheet so nobody trusts a stale
join blindly. Reconciliation is pure (no QBO/IO) so it unit-tests offline; only
`load_po_tracker` touches the workbook.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from shared import paths  # noqa: E402

# Tracker of record (the user 2026-08-25). Override with ACB_PO_TRACKER_XLSX.
DEFAULT_TRACKER = paths.onedrive_base() / "Purchase Orders" / "Copy 05 dic.xlsx"
TRACKER_SHEET = "Orders"

STALE_DAYS = 60                      # Open+unbilled older than this → "Stale"
TRACKER_MISSING_WINDOW_DAYS = 365    # only flag recent tracker POs as missing-from-QBO


def tracker_path() -> Path:
    return paths.get_path("ACB_PO_TRACKER_XLSX", DEFAULT_TRACKER)


def _norm_po(v) -> str:
    """Normalize a PO number to a comparable key. Numbers read as int/float
    collapse to a plain integer string; text is trimmed + uppercased. '' if blank."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip().upper()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _as_date(v) -> Optional[dt.date]:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _as_amount(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def days_between(d: Optional[dt.date], today: dt.date) -> Optional[int]:
    d = _as_date(d)
    return (today - d).days if d is not None else None


def _find_header(ws) -> Tuple[int, Dict[str, int]]:
    """Locate the header row (the one carrying 'P.O. #') within the first rows and
    map lower-cased header name → 0-based column index."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        if any(c is not None and str(c).strip() == "P.O. #" for c in row):
            colmap = {str(c).strip().lower(): j for j, c in enumerate(row)
                      if c not in (None, "")}
            return i, colmap
    raise ValueError("po_tracker: could not find the 'P.O. #' header row")


def load_po_tracker(path: Optional[Path] = None) -> Tuple[Dict[str, dict], dict]:
    """Read the tracker's Orders sheet → per-PO aggregate records.

    Returns (by_po, meta):
      by_po[norm_po] = {po, date(max), job, customer, project, vendor,
                        amount(sum), bill_no(joined), qb_date, rows}
      meta = {path, max_date, po_count, row_count}
    Multiple rows for one PO are aggregated (a PO is "billed on the tracker" if
    ANY of its rows carries a Bill #).
    """
    from openpyxl import load_workbook

    p = Path(path) if path else tracker_path()
    if not p.exists():
        return {}, {"path": str(p), "max_date": None, "po_count": 0, "row_count": 0,
                    "error": "file not found"}

    wb = load_workbook(p, read_only=True, data_only=True)
    if TRACKER_SHEET not in wb.sheetnames:
        return {}, {"path": str(p), "max_date": None, "po_count": 0, "row_count": 0,
                    "error": f"no '{TRACKER_SHEET}' sheet"}
    ws = wb[TRACKER_SHEET]
    hdr_row, cm = _find_header(ws)

    def col(name: str) -> Optional[int]:
        return cm.get(name.lower())

    iPO, iDate, iJob, iCust, iProj, iVend, iAmt, iBill, iQB = (
        col("P.O. #"), col("Date"), col("Job No."), col("Customer"),
        col("Project / Address"), col("Vendor"), col("Amount Due"),
        col("Bill #"), col("QB Date"))

    def get(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    by_po: Dict[str, dict] = {}
    row_count = 0
    max_date: Optional[dt.date] = None
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        po = _norm_po(get(row, iPO))
        d = _as_date(get(row, iDate))
        amt = _as_amount(get(row, iAmt))
        if not po and d is None and amt is None:
            continue  # blank/spacer row
        if not po:
            continue  # no PO number → can't reconcile
        row_count += 1
        if d and (max_date is None or d > max_date):
            max_date = d
        bill = get(row, iBill)
        bill = str(bill).strip() if bill not in (None, "") else ""
        rec = by_po.get(po)
        if rec is None:
            rec = {"po": po, "date": d, "job": "", "customer": "", "project": "",
                   "vendor": "", "amount": 0.0, "_amt_seen": False,
                   "bill_no": "", "qb_date": "", "rows": 0}
            by_po[po] = rec
        rec["rows"] += 1
        if d and (rec["date"] is None or d > rec["date"]):
            rec["date"] = d
        for key, i in (("job", iJob), ("customer", iCust), ("project", iProj),
                       ("vendor", iVend), ("qb_date", iQB)):
            if not rec[key]:
                v = get(row, i)
                if v not in (None, ""):
                    rec[key] = str(v).strip()
        if amt is not None:
            rec["amount"] += amt
            rec["_amt_seen"] = True
        if bill:
            rec["bill_no"] = (rec["bill_no"] + ", " + bill).strip(", ") if rec["bill_no"] else bill

    for rec in by_po.values():
        if not rec.pop("_amt_seen"):
            rec["amount"] = None

    meta = {"path": str(p), "max_date": max_date, "po_count": len(by_po),
            "row_count": row_count}
    return by_po, meta


def index_by_doc(po_index: Dict[str, dict]) -> Dict[str, dict]:
    """Re-key a {po_id: rec} QBO PO index by normalized DocNumber, for the
    reconcile join. Later dup docs win (rare)."""
    out: Dict[str, dict] = {}
    for rec in po_index.values():
        d = _norm_po(rec.get("doc"))
        if d:
            out[d] = rec
    return out


def reconcile_unused_pos(
    po_by_doc: Dict[str, dict],
    tracker_by_po: Dict[str, dict],
    today: dt.date,
    stale_days: int = STALE_DAYS,
    tracker_window_days: int = TRACKER_MISSING_WINDOW_DAYS,
) -> List[dict]:
    """Join QBO POs (by normalized DocNumber) with tracker POs (by normalized #)
    into one flagged-row-per-PO list. Pure — no IO. Only POs with ≥1 reason are
    returned. See module docstring for the flag definitions."""
    out: List[dict] = []
    for key in set(po_by_doc) | set(tracker_by_po):
        q = po_by_doc.get(key)
        t = tracker_by_po.get(key)
        reasons: List[str] = []

        if q is not None:
            status = (q.get("status") or "").strip()
            if status.lower() == "open" and not q.get("has_bill"):
                reasons.append("Open, no bill")
                age = days_between(q.get("date"), today)
                if age is not None and age > stale_days:
                    reasons.append(f"Stale >{stale_days}d")
        else:
            # tracker-only → missing from QBO. Only flag the truly-dropped ones:
            # recent enough to still matter AND unbilled on the tracker (a tracker
            # PO that already carries a Bill # got processed — not "unused").
            age = days_between(t.get("date") if t else None, today)
            if (age is not None and age <= tracker_window_days
                    and not (t.get("bill_no") or "").strip()):
                reasons.append("On tracker, not in QBO")

        if not reasons:
            continue

        po_date = (q.get("date") if q else None) or (t.get("date") if t else None)
        amount = None
        if q and q.get("total"):
            amount = q.get("total")
        elif t:
            amount = t.get("amount")

        out.append({
            "po": key,
            "vendor": (q.get("vendor") if q else "") or (t.get("vendor") if t else ""),
            "po_date": _as_date(po_date),
            "days_open": days_between(po_date, today),
            "job": (q.get("job") if q else "") or (t.get("job") if t else ""),
            "amount": amount,
            "qbo_status": (q.get("status") if q else "") or "— not in QBO",
            "qbo_bill": "Yes" if (q and q.get("has_bill")) else ("No" if q else "—"),
            "tracker_bill": (t.get("bill_no") if t else "") or "",
            "tracker_qb": (t.get("qb_date") if t else "") or "",
            "reason": " · ".join(reasons),
            "po_id": (q.get("id") if q else "") or "",
        })

    def _sort_key(r: dict):
        # not-in-QBO first, then stale, then plain open; oldest within each.
        if "not in QBO" in r["reason"]:
            pr = 0
        elif "Stale" in r["reason"]:
            pr = 1
        else:
            pr = 2
        return (pr, r["po_date"] or dt.date.max)

    out.sort(key=_sort_key)
    return out
