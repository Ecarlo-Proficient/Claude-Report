#!/usr/bin/env python3
"""
excel_bill_sync.py — daily QBO → xlsx AP bill tracker.

Output: ~/Documents/CompanyHealth/Bill Tracker.xlsx  (chmod 600)

Five sheets, every one an Excel Table with AutoFilter on every column.
Same column set across all five, sorted differently per sheet:

  1. Pay List        — Pay = x AND not yet paid; vendor → date
  2. Open Bills      — open bills only (Balance > 0); vendor → date → line
  3. Project Ledger  — open + paid since paid_cutoff; project → vendor → date
  4. Client Ledger   — open + paid since paid_cutoff; client → project → date
  5. By Division     — open + paid since paid_cutoff; div → proj → inv → bill

Editable cells: Lien (Notice Sent / Lien Filed), Notes (free text). Both are
preserved across syncs by _Key (bill_id). Pay marking was removed 2026-06-18 —
the owner now copies the bills to pay into a separate workbook, so Bill
Tracker.xlsx is script-owned and the clerk treats it as read-only except for
the Lien tag and Notes.

Lien tag (escalation, two steps): "Notice Sent" = a supplier/sub sent a
preliminary lien notice on an unpaid bill; "Lien Filed" = a real lien was
recorded. Both persist across syncs (a lien doesn't auto-clear on payment).

Color: RED row tint = NOT APPROVED. Lien cell tint: amber = Notice Sent,
red = Lien Filed.

USAGE
  python3 bill-tracker/excel_bill_sync.py
  python3 bill-tracker/excel_bill_sync.py --dry-run   # build rows, don't write
  python3 bill-tracker/excel_bill_sync.py --limit 50  # smoke test
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import stat
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple


# ─────────────────────── path bootstraps ───────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent
BILL_TRACKER_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))           # qbo_vault.py
sys.path.insert(0, str(BILL_TRACKER_DIR))       # qbo_bill_tracker.py

from shared import paths


# ─────────────────────── deps ───────────────────────

try:
    import requests  # noqa: F401  (used transitively by QBO extraction)
except ImportError:
    print("✗ pip3 install --break-system-packages requests")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Color
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.filters import AutoFilter, FilterColumn, Filters
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import Rule
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:
    print("✗ pip3 install --break-system-packages openpyxl")
    sys.exit(1)

from qbo_bill_tracker import (
    load_credentials, query_all, is_sub_bill, is_excluded_invoice,
    parse_date,
    STATUS_OK_TO_PAY, STATUS_AWAITING_PAYMENT, STATUS_AWAITING_INVOICE,
    STATUS_PAID, STATUS_NO_PROJECT, STATUS_PARTIAL_PAID,
    STATUS_PARTIALLY_PAID_REMAINDER,
    MATCH_BASIS_DRAW, MATCH_BASIS_FINAL, MATCH_BASIS_PUSHED,
)
from bill_rows import (
    build_account_maps, build_po_index, build_payment_map, build_rows,
    approved_text,
    collapse_rows, multi_project_bill_ids, MULTI_MARKER,
)
from po_tracker import load_po_tracker, reconcile_unused_pos, index_by_doc, _norm_po
from general_list import load_contracts
import cost_code_history as cchist
from shared.cost_code_audit import (
    classify_vendors, flag_lines, load_override, code_families, po_origin, TYPE_LABEL)


# ─────────────────────── constants ───────────────────────

OUTPUT_PATH = paths.get_path(
    "ACB_BILL_TRACKER_XLSX",
    paths.onedrive_base() / "Automations-/Bill Tracker.xlsx",
)
BACKUP_RETENTION_DAYS = 14

# Paid bills lookback. Per the user 2026-05-27: trailing 12mo was too much; use
# fixed YTD start date instead. Adjust here when crossing a fiscal year.
PAID_CUTOFF_DATE = "2026-01-01"

# Invoice query cutoff. Used to filter the Invoice pull — invoices older
# than this are skipped on the assumption they can't match any open bill.
# Pushed back to 2024-01-01 to comfortably cover the oldest open bills
# (the December 2024 MCP residual etc.) without dragging in years of
# historic invoices that quintuple the QBO query load.
INVOICE_CUTOFF_DATE = "2024-01-01"

QBO_BILL_URL_TEMPLATE = "https://qbo.intuit.com/app/bill?txnId={bill_id}"
QBO_INVOICE_URL_TEMPLATE = "https://qbo.intuit.com/app/invoice?txnId={inv_id}"
QBO_BILLPAYMENT_URL_TEMPLATE = "https://qbo.intuit.com/app/billpayment?txnId={pay_id}"
QBO_PO_URL_TEMPLATE = "https://qbo.intuit.com/app/purchaseorder?txnId={po_id}"

# Muted palette — chrome is calm so the two CF signals can read at a glance.
HEADER_FILL = PatternFill("solid", start_color="6E7E94")
DIVIDER_FILL = PatternFill("solid", start_color="2F3B4D")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Supra-header band colors — one per section. Semantic mapping:
#   BILL    = navy   (cool, identification)
#   STATUS  = amber  (warm, action needed)
#   INVOICE = green  (stable, verified data)
# The user 2026-06-04: swapped INVOICE and STATUS so green sits on Invoice.
SUPRA_BILL_FILL    = PatternFill("solid", start_color="1F3864")  # deep navy
SUPRA_STATUS_FILL  = PatternFill("solid", start_color="BF8F00")  # warm amber
SUPRA_INVOICE_FILL = PatternFill("solid", start_color="375623")  # deep green
SUPRA_PAY_FILL     = PatternFill("solid", start_color="31849B")  # teal (AP cash-out)
SUPRA_FONT         = Font(name="Calibri", size=12, bold=True,
                          color="FFFFFF", italic=False)
SUPRA_HEIGHT       = 22

MONEY_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
DATE_FMT = "m/d/yyyy"

# Status × Approved row colors. Pay-cell green stacks on top.
# The user 2026-06-04: softened "hold" (red was too aggressive) and bumped
# "partial" to a warmer amber-yellow so it reads as actionable.
CF_GREEN_READY = "C6EFCE"        # Invoice paid + approved → pay the vendor
CF_ORANGE_URGENT = "FFC299"      # Invoice paid + not approved → approve now
CF_RED_HOLD = "FCE4E4"           # Awaiting + not approved → held (soft pink)
CF_YELLOW_AUDIT = "FFEB9C"       # Bill paid + not approved → audit
CF_NEUTRAL = "FFEB9C"            # Excel "Neutral" tan — partial/awaiting GC pay
CF_PURPLE_NEEDS_PROJECT = "E4D7F5"  # No project # → needs project
CF_PEACH_PARTIAL = "FFCC66"      # Partial paid (multi-project) → warm amber
CF_LIEN_NOTICE = "FFC000"        # amber — lien NOTICE sent (escalation step 1)
CF_LIEN_FILED  = "FF0000"        # red — real LIEN filed (escalation step 2)
CF_LIEN_RELEASED = "92D050"      # green — lien RELEASED / satisfied (resolved)

# Lien-NOTICE timer (the user 2026-06-18): an UNPAID, UNTAGGED bill whose supplier
# lien-notice date (15th of the 2nd month after the bill date — TX practice)
# is approaching. A yellow→orange countdown. It NEVER goes red: red is reserved
# for an actual FILED lien (a real legal event), not a "this is coming" warning.
# 2026-07-10 (the user): old ramp was too soft — ≤30d was near-invisible and ≤15d
# blended into the Notice Sent amber. Each step now one notch hotter.
# 2026-07-10 (the user): these cells are SCRIPT-written (the countdown), so they read
# as soft/tentative — PALE tints + grey italic text — vs the human lien tags,
# which stay saturated + bold black. The tier still escalates gold→coral→rose;
# the bucket text carries the exact window. PAST is a distinct rose (NOT red —
# red stays reserved for an actual FILED lien).
CF_TIMER_YELLOW = "FFF2CC"       # ≤ 30 days out — pale gold
CF_TIMER_ORANGE = "FCE4D6"       # ≤ 15 days out — pale peach
CF_TIMER_HOT    = "F8CBAD"       # ≤ 7 days out — pale coral
CF_TIMER_PAST   = "E6B8B8"       # notice deadline passed — pale rose
CF_TIMER_TEXT   = "808080"       # grey italic — signals "script-written, not a tag"

# The Lien cell itself shows the countdown bucket as text (the user 2026-07-10) so a
# clerk reads the state without cross-referencing the key. These label strings
# are the single source of truth: written into the cell AND matched by the CF
# rules that color it, so text and color can never disagree.
TIMER_LABEL_30   = "Notice due in ≤30d"
TIMER_LABEL_15   = "Notice due in ≤15d"
TIMER_LABEL_7    = "Notice due in ≤7d"
TIMER_LABEL_PAST = "Notice PAST due"

# Lien tag values (escalation order). "Outstanding" = NOTICE or FILED; a bill
# being paid does NOT resolve a lien — only RELEASED does.
LIEN_NOTICE = "Notice Sent"
LIEN_FILED = "Lien Filed"
LIEN_RELEASED = "✓ Released"
LIEN_OUTSTANDING = (LIEN_NOTICE, LIEN_FILED)
LIEN_ALL = (LIEN_NOTICE, LIEN_FILED, LIEN_RELEASED)

# Aging bucket boundaries — strictly informational column on every sheet
AGING_BUCKETS: List[Tuple[str, int, int]] = [
    ("0–30",  0,  30),
    ("31–60", 31, 60),
    ("61–90", 61, 90),
    ("90+",   91, 10**9),
]


# ─────────────────────── unified bill-row column spec ───────────────────────
#
# Every visible sheet emits rows in this exact column order. Hidden columns
# (_Key) are present for the merge logic but hidden from view. Divider columns
# (│) are 1.5-wide dark stripes splitting bill / invoice / status sections.

BILL_ROW_COLS: List[Tuple[str, str]] = [
    # (header, kind) — kind controls number_format / alignment / font
    # 2026-06-03: dropped Days Open / Bucket / Bill Type. Moved Open to col A.
    # 2026-06-04: full reorder — BILL → STATUS → INVOICE.
    # 2026-06-04: Line Amount restored — hidden on Bills (where at bill grain it
    #   equals Bill Total) but VISIBLE on Inventory (where it's the per-line
    #   dollar amount, the whole point of that sheet).
    # 2026-07-13: reordered to the bill life-cycle read (the user). One Status split
    # into Pay Status (AP, "what's open") + Invoice Status (AR, "client paid?").
    # Payment Date renamed GC Paid Date.
    # 2026-08-12: Client moved into the AR section, right BEFORE Matched Invoice,
    #   and Invoice # moved to right AFTER Matched Invoice (the user) — so the AR
    #   read is Invoice Status · Client · Matched Invoice · Invoice #.
    ("Open",               "link"),      # 1  — tiny QBO bill link
    # ── THE BILL — what's open ──
    ("Vendor",             "text"),      # 2
    ("Bill #",             "text"),      # 3
    ("Bill Date",          "date"),      # 4
    ("Project #",          "text"),      # 5
    ("Division",           "text"),      # 6
    ("Account",            "text"),      # 7
    ("Line Description",   "text"),      # 8
    ("Line Amount",        "money"),     # 9  — hidden on Bills, shown on Inventory
    ("Bill Total",         "money"),     # 10 — original bill amount
    ("Bill Open Bal",      "money"),     # 11 — what's still owed the vendor
    ("Pay Status",         "text"),      # 12 — AP: Unpaid / Partial paid / Bill paid
    ("│",                  "div"),       # 13 — divider bill | handling
    # ── OUR HANDLING ──
    ("Approved",           "text"),      # 14 — "approved" / "not approved"
    ("Lien",               "text"),      # 15 — Notice Sent / Lien Filed; hidden on Inventory
    ("Notes",              "text"),      # 16
    ("┃",                  "div"),       # 17 — divider handling | client-payment
    # ── CLIENT PAYMENT · AR ──
    ("Invoice Status",     "text"),      # 18 — AR: Awaiting Invoice / Awaiting Payment / Invoice paid / No project #
    ("Client",             "text"),      # 19 — GC/parent customer (sits by the match to eyeball it)
    ("Matched Invoice",    "text"),      # 20 — # + memo (scope), for eyeballing the match
    ("Invoice #",          "text"),      # 21 — just the # → QBO link
    ("Invoice Date",       "date"),      # 22
    ("Invoice Open Bal",   "money"),     # 23 — what the GC still owes on the invoice
    ("Invoice Total",      "money"),     # 24
    ("GC Paid Date",       "date"),      # 25 — when the GC paid the invoice (money IN)
    # ── HOW WE PAID (AP cash-OUT; band derived from Pay Ref #) ──
    ("Pay Ref #",          "text"),      # 26 — check # we paid with (blank for CC)
    ("Pay Date",           "date"),      # 27 — when we paid the vendor
    ("Pay Method",         "text"),      # 28 — Check / CC / (multiple)
    ("_Key",               "text"),      # 29 — hidden merge join key
]
LINE_AMT_COL_INDEX = next(i + 1 for i, (h, _) in enumerate(BILL_ROW_COLS) if h == "Line Amount")
HEADERS = [h for h, _ in BILL_ROW_COLS]
KINDS   = [k for _, k in BILL_ROW_COLS]
# Divider columns are marked by kind=="div"; their distinct headers (│ vs ┃)
# render visually similar but satisfy Excel Table's unique-column-name rule.
DIVIDER_COL_INDEXES = {i + 1 for i, (_, k) in enumerate(BILL_ROW_COLS) if k == "div"}
KEY_COL_INDEX = HEADERS.index("_Key") + 1
LIEN_COL_INDEX = HEADERS.index("Lien") + 1
NOTES_COL_INDEX = HEADERS.index("Notes") + 1
BILL_OPEN_BAL_COL_INDEX = HEADERS.index("Bill Open Bal") + 1
PAY_STATUS_COL_INDEX = HEADERS.index("Pay Status") + 1
INVOICE_STATUS_COL_INDEX = HEADERS.index("Invoice Status") + 1
APPROVED_COL_INDEX = HEADERS.index("Approved") + 1
OPEN_COL_INDEX = HEADERS.index("Open") + 1
ACCOUNT_COL_INDEX = HEADERS.index("Account") + 1
LINE_DESC_COL_INDEX = HEADERS.index("Line Description") + 1

# Per-sheet column widths — matches BILL → STATUS → INVOICE.
COL_WIDTHS: Dict[int, float] = {
    1: 5,                                            # Open
    # THE BILL — what's open
    2: 28,                                           # Vendor
    3: 12, 4: 11,                                    # Bill # / Bill Date
    5: 11, 6: 8,                                     # Project # / Division
    7: 22,                                           # Account
    8: 35,                                           # Line Description
    9: 13,                                           # Line Amount (Inventory only)
    10: 13, 11: 13,                                  # Bill Total / Bill Open Bal
    12: 13,                                          # Pay Status
    13: 1.5,                                         # divider │
    # OUR HANDLING
    14: 14, 15: 15, 16: 28,                          # Approved / Lien / Notes
    17: 1.5,                                         # divider ┃
    # CLIENT PAYMENT · AR
    18: 16,                                          # Invoice Status
    19: 26,                                          # Client (sits by the match)
    20: 60, 21: 10, 22: 11,                          # Matched Inv / Invoice # / Inv Date
    23: 13, 24: 12, 25: 12,                          # Invoice Open Bal / Inv Total / GC Paid Date
    # HOW WE PAID (AP cash-out)
    26: 14, 27: 11, 28: 12,                          # Pay Ref # / Pay Date / Pay Method
    29: 14,                                          # _Key (hidden)
}


# ─────────────────────── small helpers ───────────────────────

def _col_letter(idx_1based: int) -> str:
    return get_column_letter(idx_1based)


def _bucket_for_age(days: Optional[int]) -> str:
    if days is None:
        return ""
    for label, lo, hi in AGING_BUCKETS:
        if lo <= days <= hi:
            return label
    return ""


def _qbo_link(bill_id: str) -> str:
    """Render the QBO bill link as a single-glyph 'button'. Column is ~5 wide
    so we keep the label tiny; the cell is centered and underlined by
    _format_data_cell's link branch."""
    if not bill_id:
        return ""
    url = QBO_BILL_URL_TEMPLATE.format(bill_id=bill_id)
    return f'=HYPERLINK("{url}","↗")'


def _qbo_po_link(po_id: str) -> str:
    """QBO purchase-order link as a single-glyph 'button'. Blank when the PO has
    no QBO id (tracker-only rows)."""
    if not po_id:
        return ""
    url = QBO_PO_URL_TEMPLATE.format(po_id=po_id)
    return f'=HYPERLINK("{url}","↗")'


def _invoice_cell(inv_doc: str, inv_memo: str, match_basis: str = "",
                  match_note: str = "") -> str:
    """Format Invoice # together with the invoice's PrivateNote (draw memo)
    so the AP team can eyeball whether the QBO match is correct. No truncation
    — column is wide enough, and clipping the period notation defeats the
    purpose of showing the memo at all.

    `match_basis` prefixes a per-row signal when General List RP draw semantics
    applied: [DRAW] = the next draw authorizes this bill; [FULLY BILLED] = job
    is 100% billed, matched to the last draw. [PUSHED from Draw #N] = a CP/MFD
    bill the supplier agreed to carry into this later draw (shared/draw_moves);
    the ledger loader splits that leading tag off so the draw key stays clean."""
    if not inv_doc:
        return ""
    prefix = ""
    if match_basis == MATCH_BASIS_DRAW:
        prefix = "[DRAW] "
    elif match_basis == MATCH_BASIS_FINAL:
        prefix = "[FULLY BILLED] "
    elif match_basis == MATCH_BASIS_PUSHED:
        prefix = "[" + ((match_note or "pushed").replace("pushed", "PUSHED", 1)) + "] "
    memo = (inv_memo or "").strip()
    body = inv_doc if not memo else f"{inv_doc} — {memo}"
    return prefix + body


def _esc_col(name: str) -> str:
    """Escape a column name for use inside Excel `tbl[Col]` structured refs.
    `#`, `'`, `[`, `]` are reserved and must be prefixed with `'`.
    LibreOffice accepts unescaped; Excel silently strips formulas without it."""
    return (name.replace("'", "''")
                .replace("#", "'#")
                .replace("[", "'[")
                .replace("]", "']"))


def _format_data_cell(c, kind: str) -> None:
    c.font = BODY_FONT
    if kind == "money":
        c.number_format = MONEY_FMT
        c.alignment = ALIGN_RIGHT
    elif kind == "date":
        c.number_format = DATE_FMT
        c.alignment = ALIGN_CENTER
    elif kind == "flag":
        c.number_format = "0"
        c.alignment = ALIGN_CENTER
    elif kind == "link":
        c.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        c.alignment = ALIGN_CENTER
    elif kind == "div":
        # Divider cells get the dark fill; no value, no alignment
        c.fill = DIVIDER_FILL
    else:
        c.alignment = ALIGN_LEFT


def _cf_fill_rule(formula: str, hex_color: str, bold: bool = False,
                  stop_if_true: bool = False, font_color: str = "",
                  italic: bool = False) -> Rule:
    """Build a properly-formed conditional-formatting Rule.

    openpyxl's PatternFill("solid", start_color=…) shortcut emits a dxf that
    Excel rejects (missing patternType). We construct DifferentialStyle with
    explicit fg/bg colors; the post_process_xlsx step further patches the XML
    if openpyxl strips patternType anyway.

    `stop_if_true=True` halts CF evaluation when this rule matches — used for
    the NOT APPROVED red rule so it wins over OK-TO-PAY green when a row could
    match both (NOT APPROVED bill that's also OK TO PAY = still don't pay).

    `font_color` sets the dxf font color; `italic`/`bold` set the weight/style
    (grey italic marks script-written text; bold black marks a human tag).
    """
    fill = PatternFill(
        patternType="solid",
        fgColor=Color(rgb=f"FF{hex_color}"),
        bgColor=Color(rgb=f"FF{hex_color}"),
    )
    if bold or font_color or italic:
        dxf = DifferentialStyle(
            fill=fill,
            font=Font(bold=bold, italic=italic, color=(font_color or None)))
    else:
        dxf = DifferentialStyle(fill=fill)
    return Rule(type="expression", formula=[formula], dxf=dxf, stopIfTrue=stop_if_true)


# ─────────────────────── post-process Excel quirk fixes ───────────────────────

def post_process_xlsx(path: Path) -> None:
    """Patch openpyxl's two known Excel-rejection quirks via XML regex.

    1. dxf <patternFill> elements emitted by openpyxl drop `patternType="solid"`.
       Excel needs it; LibreOffice tolerates absence. Inject it.
    2. type="list" data validations get serialized with `operator="between"`
       and `<formula2>0</formula2>`. Excel rejects on list validations.
       Strip both.

    Both issues produce "We found a problem with some content" warnings on open.
    """
    import zipfile
    import shutil

    tmp = path.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r'<patternFill(?![^>]*patternType)([^>/]*)>',
                    r'<patternFill patternType="solid"\1>',
                    text,
                )
                data = text.encode("utf-8")
            elif item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                def _fix_dv(m: "re.Match") -> str:
                    block = m.group(0)
                    if 'type="list"' not in block:
                        return block
                    block = re.sub(r'\s+operator="[^"]*"', "", block)
                    block = re.sub(r'<formula2>[^<]*</formula2>', "", block)
                    return block
                text = re.sub(
                    r"<dataValidation\b[^>]*>.*?</dataValidation>",
                    _fix_dv, text, flags=re.S,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


# ─────────────────────── xlsx validator (Excel-strict preflight) ───────────────────────

def validate_xlsx(path: Path) -> List[str]:
    """Inspect the built xlsx for patterns Excel rejects but LibreOffice tolerates.

    Returns a list of human-readable failure messages. Empty list = clean.
    Run AFTER post_process_xlsx so we catch only un-patched issues.

    Patterns checked (each one has burned us before):
      A. dxf <patternFill> missing patternType="solid"
      B. dataValidation type="list" with operator="between" or <formula2>
      C. Excel Table with duplicate column names
      D. Unescaped # / [ / ] inside tbl[Col] structured references
      E. Excel 365 dynamic-array functions (FILTER/SORT/UNIQUE/etc) in cells
         without t="array" on the <f> tag
    """
    import zipfile
    DYNAMIC_ARRAY_FNS = ("FILTER", "SORT", "SORTBY", "UNIQUE", "SEQUENCE",
                         "RANDARRAY", "TOROW", "TOCOL", "WRAPROWS", "WRAPCOLS")
    failures: List[str] = []

    with zipfile.ZipFile(path, "r") as z:
        names = z.namelist()

        # A. dxf patternType
        if "xl/styles.xml" in names:
            text = z.read("xl/styles.xml").decode("utf-8")
            m = re.search(r"<dxfs[^>]*>(.*?)</dxfs>", text, re.S)
            if m:
                content = m.group(1)
                total = len(re.findall(r"<patternFill\b", content))
                ok = len(re.findall(r'<patternFill[^>]*patternType=', content))
                if total != ok:
                    failures.append(
                        f"A. styles.xml dxf: {total - ok} <patternFill> missing patternType "
                        f"(Excel rejects — openpyxl drops it in DifferentialStyle output)"
                    )

        # Per-sheet checks
        for fn in names:
            if not (fn.startswith("xl/worksheets/sheet") and fn.endswith(".xml")):
                continue
            sheet_text = z.read(fn).decode("utf-8")
            sheet_label = fn.rsplit("/", 1)[-1]

            # B. dataValidation list with stray operator / formula2
            for dv in re.findall(r"<dataValidation\b[^>]*>.*?</dataValidation>", sheet_text, re.S):
                if 'type="list"' in dv:
                    if 'operator="between"' in dv:
                        failures.append(f"B. {sheet_label}: dataValidation type=list has operator=between")
                    if "<formula2>" in dv:
                        failures.append(f"B. {sheet_label}: dataValidation type=list has stray <formula2>")

            # D. Unescaped # / [ / ] inside tbl[Col]. Match `tblName[Foo#Bar]` style.
            for ref in re.findall(r"\btbl[A-Za-z][A-Za-z0-9_]*\[([^]]*)\]", sheet_text):
                if "#" in ref and "'#" not in ref:
                    fixed = ref.replace("#", "'#")
                    failures.append(
                        f"D. {sheet_label}: structured ref contains unescaped # → "
                        f"tbl[{ref}] (should be tbl[{fixed}])"
                    )

            # E. dynamic-array functions written without t="array"
            array_formula_refs: List[str] = []  # collect for F check below
            for fn_match in re.finditer(r"<c\b[^>]*>(.*?)</c>", sheet_text, re.S):
                cell_xml = fn_match.group(0)
                f_match = re.search(r"<f\b([^>]*)>([^<]*)</f>", cell_xml)
                if not f_match:
                    continue
                f_attrs, f_text = f_match.group(1), f_match.group(2)
                if 't="array"' in f_attrs:
                    ref_m = re.search(r'ref="([^"]+)"', f_attrs)
                    if ref_m:
                        array_formula_refs.append(ref_m.group(1))
                    continue
                # Check if the formula uses any dynamic-array function
                f_text_upper = f_text.upper()
                for fn_name in DYNAMIC_ARRAY_FNS:
                    if re.search(rf"\b{fn_name}\s*\(", f_text_upper):
                        failures.append(
                            f"E. {sheet_label}: dynamic-array fn {fn_name}() in cell formula "
                            f"without t=\"array\" attribute → Excel rejects on open. "
                            f"Use openpyxl's ArrayFormula(ref=…, text=…) instead."
                        )
                        break

            # F. ArrayFormula spill range contains other formatted cells.
            # `<c r="A3" s="2"/>` (style-only, no value) inside the declared
            # spill envelope blocks Excel's array spill → "Removed Records".
            # Detect by checking for any `<c s="…"/>` cells within the ref range.
            for spill_ref in array_formula_refs:
                # parse range like "A2:Y1001"
                m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", spill_ref)
                if not m:
                    continue
                start_row, end_row = int(m.group(2)), int(m.group(4))
                # find any <c> within the row range that has s= but no value
                # (the FILTER's anchor cell at start_row is exempt — it's the array)
                blockers = 0
                for cell_m in re.finditer(r'<c r="([A-Z]+)(\d+)"([^>]*)/>', sheet_text):
                    cell_row = int(cell_m.group(2))
                    if cell_row <= start_row or cell_row > end_row:
                        continue
                    attrs = cell_m.group(3)
                    if "s=" in attrs:  # style-attached, blocks spill
                        blockers += 1
                if blockers > 0:
                    failures.append(
                        f"F. {sheet_label}: ArrayFormula spill range {spill_ref} contains "
                        f"{blockers} style-attached empty cells — Excel will refuse spill. "
                        f"Remove per-cell formatting inside the spill range; use column "
                        f"defaults (ws.column_dimensions[X].number_format) instead."
                    )

        # C. Tables with duplicate column names
        for fn in names:
            if not (fn.startswith("xl/tables/table") and fn.endswith(".xml")):
                continue
            tbl_text = z.read(fn).decode("utf-8")
            tbl_name_m = re.search(r'displayName="([^"]+)"', tbl_text)
            tbl_name = tbl_name_m.group(1) if tbl_name_m else fn
            cols = re.findall(r'<tableColumn[^/]*name="([^"]+)"', tbl_text)
            dupes = sorted({c for c in cols if cols.count(c) > 1})
            if dupes:
                failures.append(
                    f"C. Table {tbl_name}: duplicate column names {dupes} "
                    f"(Excel Tables require unique headers)"
                )

    return failures


# ─────────────────────── backup rotation ───────────────────────

def rotate_backup(workbook_path: Path, retention_days: int = BACKUP_RETENTION_DAYS) -> None:
    """Snapshot the existing workbook before overwriting. One file per day;
    prune anything older than `retention_days`.
    """
    import shutil
    if not workbook_path.exists():
        return
    backups_dir = workbook_path.parent / "_backups"
    backups_dir.mkdir(parents=True, exist_ok=True)
    stem = workbook_path.stem
    today_iso = dt.date.today().isoformat()
    target = backups_dir / f"{stem} — {today_iso}{workbook_path.suffix}"
    try:
        shutil.copy2(workbook_path, target)
    except Exception as e:
        print(f"  ⚠ backup copy failed: {e}")
        return

    cutoff = dt.date.today() - dt.timedelta(days=retention_days)
    pruned = 0
    for f in backups_dir.glob(f"{stem} — *.xlsx"):
        try:
            date_str = f.stem.rsplit(" — ", 1)[1]
            file_date = dt.date.fromisoformat(date_str)
        except (IndexError, ValueError):
            continue
        if file_date < cutoff:
            try:
                f.unlink()
                pruned += 1
            except Exception as e:
                print(f"  ⚠ prune failed for {f.name}: {e}")
    print(f"  ✓ backup → _backups/{target.name}"
          f"{f' (pruned {pruned} older than {retention_days}d)' if pruned else ''}")


# ─────────────────────── Lien/Notes preservation (single-sheet) ───────────────────────
#
# Bills is the only editable sheet (division + Audit sheets are read-only,
# rebuilt each sync). Preservation reads the existing Bills sheet and restores
# the two human-edited columns — Lien and Notes — by _Key (bill_id) on the
# next build. Pay was removed 2026-06-18 (owner pastes the pay list into a
# separate workbook), so it is no longer carried forward.

SHEET_NAMES = ["Bills"]
BILLS_TABLE = "tblBills"


def _read_sheet_edits(ws) -> Dict[str, Dict[str, str]]:
    """Return {_Key → {Lien, Notes}} from a sheet that follows BILL_ROW_COLS.

    Auto-detects whether the header row is row 1 (legacy single-header) or
    row 2 (current layout with supra-header at row 1). Requires "_Key" and
    "Notes"; "Lien" is optional (absent in pre-2026-06-18 workbooks → blank).
    """
    out: Dict[str, Dict[str, str]] = {}
    if ws.max_row < 2:
        return out

    # Detect header row — try row 1 first (legacy), fall back to row 2 (current)
    header_row_idx = None
    for candidate in (1, 2):
        try:
            cells = [c.value for c in ws[candidate]]
        except Exception:
            continue
        if "_Key" in cells and "Notes" in cells:
            header_row_idx = candidate
            break
    if header_row_idx is None:
        return out

    header_row = [c.value for c in ws[header_row_idx]]
    key_i = header_row.index("_Key")
    notes_i = header_row.index("Notes")
    lien_i = header_row.index("Lien") if "Lien" in header_row else None

    data_start = header_row_idx + 1
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if key_i >= len(row):
            continue
        key = row[key_i]
        if not key:
            continue
        lien_val = ""
        if lien_i is not None and lien_i < len(row) and row[lien_i] is not None:
            # Only real tags are preserved — the notice-countdown text the sheet
            # writes into blank Lien cells must NOT be carried forward as an edit
            # (it's recomputed each build).
            if row[lien_i] in LIEN_ALL:
                lien_val = row[lien_i]
        out[str(key)] = {
            "Lien": lien_val,
            "Notes": (row[notes_i] if notes_i < len(row) and row[notes_i] is not None else ""),
        }
    return out


def preserve_edits(path: Path) -> Dict[str, Dict[str, str]]:
    """Read existing workbook's Bills sheet and return {_Key → {Lien, Notes}}.

    Bills is the only editable sheet — division sheets and the Audit sheet are
    read-only views, rebuilt each sync. Both Lien and Notes are ALWAYS
    preserved (never auto-cleared — a lien persists until manually released,
    and Notes are commentary).

    Legacy migration: _Key was per-line (`bill_id-line_id`) before the Bills
    sheet collapsed to bill-grain rows. Any line-level keys found are collapsed
    to bill-level so edits carry over. First non-empty value wins per bill.

    Returns empty dict on first run (no file) or unreadable workbook —
    safe to feed into `_render_rows` either way.
    """
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ⚠ couldn't open existing workbook for preservation: {e}")
        return {}
    # Fall back through legacy sheet names so an upgrade from older layouts
    # (Open Bills, Pay List, AP Workspace) doesn't lose pending edits.
    raw: Dict[str, Dict[str, str]] = {}
    for name in ("Bills", "Open Bills", "Pay List", "AP Workspace"):
        if name in wb.sheetnames:
            raw = _read_sheet_edits(wb[name])
            break
    if not raw:
        return {}

    # Migrate line-level keys (e.g. "12345-0") to bill-level ("12345").
    # Detection: a key with a trailing "-<digits>" segment is line-level.
    migrated: Dict[str, Dict[str, str]] = {}
    line_keys_seen = 0
    for k, v in raw.items():
        m = re.match(r"^(.+?)-(\d+)$", k)
        bill_key = m.group(1) if m else k
        if m:
            line_keys_seen += 1
        slot = migrated.setdefault(bill_key, {"Lien": "", "Notes": ""})
        # First non-empty value wins per bill (Lien and Notes alike).
        if (v.get("Lien") or "").strip() and not slot["Lien"]:
            slot["Lien"] = v["Lien"]
        if (v.get("Notes") or "").strip() and not slot["Notes"]:
            slot["Notes"] = v["Notes"]
    if line_keys_seen:
        print(f"  ↻ migrated {line_keys_seen} line-level _Keys → "
              f"{len(migrated)} bill-level _Keys for collapsed view")
    return migrated


def preserve_audit_marks(path: Path) -> Dict[str, str]:
    """Read the existing `Audit - Coding` sheet's user-entered `Status` marks,
    keyed by the hidden `_Key` (bill_id). These persist across runs and mirror into
    the Bills Notes so the audit and Bills stay consistent (the user 2026-08-25).
    First non-empty mark per bill wins. Empty/first-run/unreadable → {}."""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ⚠ couldn't read audit marks: {e}")
        return {}
    if "Audit - Coding" not in wb.sheetnames:
        return {}
    ws = wb["Audit - Coding"]
    hdr = [(c.value or "") for c in ws[1]]
    if "_Key" not in hdr or "Status" not in hdr:
        return {}
    ki, si = hdr.index("_Key"), hdr.index("Status")
    marks: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ki >= len(row) or si >= len(row):
            continue
        bid, st = row[ki], row[si]
        if bid and st and str(st).strip():
            marks.setdefault(str(bid), str(st).strip())
    return marks


# ─────────────────────── sheet rendering ───────────────────────

def _apply_header(ws, hide_cols: Optional[List[int]] = None) -> None:
    """Write a two-row header: supra-header bands at row 1 (BILL / INVOICE /
    STATUS & ACTION) + column-name headers at row 2 (these are the Table
    headers with AutoFilter dropdowns).

    The supra-header sits OUTSIDE the Table — Excel only allows one header
    row per Table, so the Table's ref begins at row 2.
    """
    # Row 1 — supra-header section bands.
    # The three sections are defined by the divider column positions:
    #   BILL    = cols 1 .. divider1-1
    #   INVOICE = cols divider1+1 .. divider2-1
    #   STATUS  = cols divider2+1 .. end (excluding hidden _Key)
    dividers = sorted(DIVIDER_COL_INDEXES)  # [11, 16] for current layout
    d1, d2 = dividers[0], dividers[1]
    # The PAYMENT band (AP cash-out) sits at the end. It has no divider column
    # of its own (that would break the 2-divider assumptions elsewhere) — the
    # band-color change alone separates it from INVOICE INFO.
    pay_from = HEADERS.index("Pay Ref #") + 1
    sections = [
        ("THE BILL — what's open", 1,     d1 - 1,                     SUPRA_BILL_FILL),
        ("OUR HANDLING",       d1 + 1,   d2 - 1,                     SUPRA_STATUS_FILL),
        ("CLIENT PAYMENT (AR)", d2 + 1,  pay_from - 1,               SUPRA_INVOICE_FILL),
        ("HOW WE PAID (AP)",   pay_from, KEY_COL_INDEX - 1,          SUPRA_PAY_FILL),
    ]
    for label, c_from, c_to, fill in sections:
        ws.cell(row=1, column=c_from, value=label)
        ws.merge_cells(start_row=1, start_column=c_from, end_row=1, end_column=c_to)
        # Apply font + fill to every cell in the merged range
        # (Excel needs all cells styled or the merge can de-style on edit)
        for ci in range(c_from, c_to + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = fill
            cell.font = SUPRA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Divider columns in row 1 also get the dark fill so the section bands
    # are visually separated by the same divider as the data rows.
    for c_i in DIVIDER_COL_INDEXES:
        ws.cell(row=1, column=c_i).fill = DIVIDER_FILL
    # _Key column in row 1 stays plain (it will be hidden anyway)

    ws.row_dimensions[1].height = SUPRA_HEIGHT

    # Row 2 — column headers (this row is the Table's header).
    for col_i, name in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col_i, value=name)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.fill = DIVIDER_FILL if col_i in DIVIDER_COL_INDEXES else HEADER_FILL
    ws.row_dimensions[2].height = 22

    # Freeze both header rows so they stay visible as user scrolls
    ws.freeze_panes = "A3"

    for col_i, width in COL_WIDTHS.items():
        ws.column_dimensions[_col_letter(col_i)].width = width
    # hide _Key always, plus any sheet-specific hidden columns (Line Amount
    # on Bills, Pay on Inventory).
    ws.column_dimensions[_col_letter(KEY_COL_INDEX)].hidden = True
    for col_i in (hide_cols or []):
        ws.column_dimensions[_col_letter(col_i)].hidden = True


def _notice_date(bill_date) -> Optional[dt.date]:
    """Supplier lien-notice deadline: 15th of the 2nd month after the bill date
    (TX practice). None if the bill date is missing/invalid."""
    if not isinstance(bill_date, dt.date):
        return None
    total = bill_date.year * 12 + (bill_date.month - 1) + 2
    y, m0 = divmod(total, 12)
    return dt.date(y, m0 + 1, 15)


def _notice_timer_label(bill_date, balance, lien_val: str,
                        today: dt.date) -> Optional[str]:
    """Bucket label for an UNPAID, UNTAGGED bill whose lien-notice deadline is
    near or past. None when tagged, paid, no bill date, or >30 days out."""
    if lien_val:
        return None                       # a real tag wins over the countdown
    if not (isinstance(balance, (int, float)) and balance > 0):
        return None                       # paid → no notice risk
    nd = _notice_date(bill_date)
    if nd is None:
        return None
    days = (nd - today).days
    if days < 0:
        return TIMER_LABEL_PAST
    if days <= 7:
        return TIMER_LABEL_7
    if days <= 15:
        return TIMER_LABEL_15
    if days <= 30:
        return TIMER_LABEL_30
    return None


def _pay_status_display(base: str, gc_paid_date, our_pay_date) -> str:
    """Tag a paid bill with the cash-flow timing: 'Bill paid (fronted)' when we
    paid the vendor before the GC paid us (we floated our own cash — including
    when the GC hasn't paid at all yet), or 'Bill paid (collected)' when the GC
    funded us first. Falls back to the plain 'Bill paid' when we can't tell (no
    recorded pay date). Non-paid statuses pass through unchanged. Prefix stays
    'Bill paid' so the CF rules + the hide-paid filter still match on it."""
    if base != "Bill paid":
        return base
    if not isinstance(gc_paid_date, dt.date):
        return "Bill paid (fronted)"          # GC hasn't paid us → still floating
    if not isinstance(our_pay_date, dt.date):
        return base                           # unknown when we paid → leave plain
    return ("Bill paid (fronted)" if our_pay_date < gc_paid_date
            else "Bill paid (collected)")


def _write_bill_row(ws, r_i: int, r: dict, edits: Dict[str, Dict[str, str]],
                    today: dt.date) -> None:
    """Render one bill-line row using the unified BILL_ROW_COLS layout.

    Lien and Notes both reflect the preserved value (neither auto-clears — a
    lien persists until manually released, Notes are commentary).

    2026-06-03: dropped Days Open / Bucket / Bill Type. Open link moved to
    col A. Status split into pipeline + Approved columns.
    2026-06-18: Pay column replaced by Lien (Notice Sent / Lien Filed).
    """
    key = r.get("key", "") or ""
    pres = edits.get(key, {"Lien": "", "Notes": ""})
    approved = bool(r.get("approved"))
    lien_val = pres.get("Lien", "") or ""
    # AP cash-out: the payment(s) that paid this bill (set in main() for the run).
    _pm = _BILL_PAY_MAP.get(r.get("bill_id", "") or "", {})
    # A paid bill tags whether we FRONTED (paid the vendor before the GC paid us
    # → floated our own cash) or paid after we COLLECTED.
    pay_status_val = _pay_status_display(
        r.get("pay_status", ""), r.get("payment_date"), _pm.get("date"))
    # Untagged + unpaid + near the notice deadline → the cell shows the countdown
    # bucket as text (colored by the matching CF rule). A real tag always wins;
    # this computed text is never preserved (see _read_sheet_edits).
    lien_display = lien_val or (
        _notice_timer_label(r.get("bill_date"), r.get("bill_balance"),
                            lien_val, today) or "")

    values: List[Any] = [
        _qbo_link(r.get("bill_id", "")),                         # 1  Open
        # ── THE BILL — what's open ──
        r.get("vendor", ""),                                     # 2
        r.get("bill_doc", ""),                                   # 3  Bill #
        r.get("bill_date"),                                      # 4
        r.get("project_num", ""),                                # 5
        r.get("division", ""),                                   # 6
        r.get("account", ""),                                    # 7  Account
        r.get("line_desc", ""),                                  # 8  Line Description
        r.get("line_amount"),                                    # 9  Line Amount
        r.get("bill_total"),                                     # 10 Bill Total
        r.get("bill_balance"),                                   # 11 Bill Open Bal
        pay_status_val,                                          # 12 Pay Status (AP)
        None,                                                    # 13 divider │
        # ── OUR HANDLING ──
        approved_text(approved),                                 # 14 Approved
        lien_display,                                            # 15 Lien (tag or countdown)
        pres.get("Notes", "") or "",                             # 16 Notes
        None,                                                    # 17 divider ┃
        # ── CLIENT PAYMENT · AR ──
        r.get("invoice_status", ""),                             # 18 Invoice Status (AR)
        r.get("gc_name", ""),                                    # 19 Client (GC/parent, by the match)
        _invoice_cell(r.get("inv_doc", ""), r.get("inv_memo", ""), r.get("match_basis", ""),
                      r.get("match_note", "")),                  # 20 Matched Invoice
        r.get("inv_doc", ""),                                    # 21 Invoice # (→ link)
        r.get("inv_date"),                                       # 22
        r.get("inv_balance"),                                    # 23 Invoice Open Bal (GC still owes)
        r.get("inv_total"),                                      # 24 Invoice Total
        r.get("payment_date"),                                   # 25 GC Paid Date (money IN)
        # ── HOW WE PAID (AP cash-out) ──
        _pm.get("ref", ""),                                      # 26 Pay Ref #
        _pm.get("date"),                                         # 27 Pay Date
        _pm.get("method", ""),                                   # 28 Pay Method
        key,                                                     # 29 _Key
    ]
    for c_i, (val, kind) in enumerate(zip(values, KINDS), start=1):
        c = ws.cell(row=r_i, column=c_i, value=val)
        _format_data_cell(c, kind)

    # Hyperlink the dedicated "Invoice #" cell → the QBO invoice (just the # is
    # the link; the "Matched Invoice" column keeps the # + memo for reading).
    inv_id = r.get("inv_id")
    if inv_id:
        mi = ws.cell(row=r_i, column=HEADERS.index("Invoice #") + 1)
        if mi.value:
            mi.hyperlink = QBO_INVOICE_URL_TEMPLATE.format(inv_id=inv_id)
            mi.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    # Hyperlink the Pay Ref # cell → the QBO bill payment (latest, for a bill
    # paid by several). Blank for CC (no ref text to click); '(multiple)' links
    # to the most recent payment.
    if _pm.get("pay_id") and (_pm.get("ref") or "").strip():
        pr = ws.cell(row=r_i, column=HEADERS.index("Pay Ref #") + 1)
        pr.hyperlink = QBO_BILLPAYMENT_URL_TEMPLATE.format(pay_id=_pm["pay_id"])
        pr.font = Font(name="Calibri", size=11, color="0563C1", underline="single")


def _lien_legend(ws, start_col: int) -> None:
    """HORIZONTAL key for the Lien column, on the frozen header row (row 1) to
    the right of the table. A single row so AutoFilter (which hides data rows)
    can never fragment it (the user 2026-06-18). Each cell is a colored swatch with
    its label inside. Human TAGS (bold black on a saturated fill) come first,
    then the SCRIPT countdown (grey italic on a pale tint) — the same styling
    the data cells get, so the key reads exactly like the column."""
    # (hex fill or None, label, font hex, italic) — italic marks script-written.
    items = [
        (None,             "LIEN KEY →",        "1F3864", False),
        (CF_LIEN_NOTICE,   LIEN_NOTICE,         "000000", False),
        (CF_LIEN_FILED,    LIEN_FILED,          "000000", False),
        (CF_LIEN_RELEASED, LIEN_RELEASED,       "000000", False),
        (CF_TIMER_YELLOW,  TIMER_LABEL_30,      CF_TIMER_TEXT, True),
        (CF_TIMER_ORANGE,  TIMER_LABEL_15,      CF_TIMER_TEXT, True),
        (CF_TIMER_HOT,     TIMER_LABEL_7,       CF_TIMER_TEXT, True),
        (CF_TIMER_PAST,    TIMER_LABEL_PAST,    CF_TIMER_TEXT, True),
    ]
    for i, (hex_color, text, font_hex, italic) in enumerate(items):
        col = start_col + i
        ws.column_dimensions[_col_letter(col)].width = 20
        c = ws.cell(row=1, column=col, value=text)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if hex_color:
            c.fill = PatternFill(patternType="solid",
                                 fgColor=Color(rgb=f"FF{hex_color}"),
                                 bgColor=Color(rgb=f"FF{hex_color}"))
            # Tags read bold; script countdown reads grey italic (not bold).
            c.font = Font(bold=not italic, italic=italic, size=10, color=font_hex)
        else:
            c.font = Font(bold=True, size=11, color=font_hex)


def _rowcolor_legend(ws, start_col: int) -> None:
    """Horizontal key for the reconciliation status-cell colors, on frozen ROW 2 to the
    right of the table (row 1 holds the lien key). It MUST live on a header row:
    AutoFilter hides DATA rows, so a legend placed in the body would vanish the
    moment the user filters. Each cell is a colored swatch with its meaning
    inside, mirroring the lien key one row above."""
    # (hex fill or None, label, font hex)
    items = [
        (None,                    "STATUS KEY →",           "1F3864"),
        (CF_GREEN_READY,          "Invoice paid",           "000000"),
        (CF_NEUTRAL,              "Awaiting / partial pay", "000000"),
        (CF_RED_HOLD,             "Fronted: not inv'd",     "000000"),
        (CF_PEACH_PARTIAL,        "Partly funded (multi)",  "000000"),
        (CF_PURPLE_NEEDS_PROJECT, "No project #",           "000000"),
        (None,                    "Pipeline: awaiting GC",  "1F3864"),
    ]
    for i, (hex_color, text, font_hex) in enumerate(items):
        col = start_col + i
        ws.column_dimensions[_col_letter(col)].width = 20
        c = ws.cell(row=2, column=col, value=text)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=10, color=font_hex)
        if hex_color:
            c.fill = PatternFill(patternType="solid",
                                 fgColor=Color(rgb=f"FF{hex_color}"),
                                 bgColor=Color(rgb=f"FF{hex_color}"))


def _finalize_sheet(ws, table_name: str, last_row: int,
                    hide_paid_default: bool = False,
                    lien_editable: bool = False) -> None:
    """Wrap rendered range in an Excel Table (gives AutoFilter on every col)
    and apply universal conditional formatting.

    Tables require ≥1 data row, so we ensure at least row 3 exists even when
    empty (data starts at row 3 since row 1 = supra-header, row 2 = headers).

    `hide_paid_default=True` pre-filters the Status column to hide "Bill paid"
    rows on open — used for the Bills sheet so paid bills don't clutter the
    active view. User clears the filter (Status dropdown → check all) to see
    paid bills when needed.
    """
    HEADER_ROW = 2   # column-name headers (Table header row)
    DATA_START = 3   # data rows begin here

    if last_row < DATA_START:
        # Empty sheet — drop a single blank data row so the Table has body
        for c_i in range(1, len(HEADERS) + 1):
            ws.cell(row=DATA_START, column=c_i, value=None)
        last_row = DATA_START

    table_ref = f"A{HEADER_ROW}:{_col_letter(len(HEADERS))}{last_row}"
    tbl = Table(displayName=table_name, ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight15",
        showFirstColumn=False, showLastColumn=False,
        # Row stripes OFF: the table's cosmetic grey/white banding collided with
        # grey = "Done" in the reconciliation colors. White by default now means
        # "pipeline" (matches the legend); grey appears only on real done rows.
        showRowStripes=False, showColumnStripes=False,
    )

    # Default view = the "what do we still owe" working set: Pay Status Unpaid /
    # Partial paid (hide Bill paid). Set the AutoFilter funnel AND physically
    # hide paid rows to match — Excel trusts each row's hidden flag on open and
    # ignores the filter criteria until the user toggles it. A Bill-paid row with
    # an OUTSTANDING lien stays visible (a lien outlives payment). To review paid
    # / fronted bills, check "Bill paid" in the Pay Status filter — the row color
    # + Invoice Status then flag the fronted ones.
    if hide_paid_default:
        tbl.autoFilter = AutoFilter(ref=table_ref)
        tbl.autoFilter.filterColumn = [
            FilterColumn(colId=PAY_STATUS_COL_INDEX - 1,
                         filters=Filters(filter=["Unpaid", "Partial paid"]))
        ]
        for r in range(DATA_START, last_row + 1):
            # Pay Status is "Bill paid" or "Bill paid (fronted)/(collected)".
            if not str(ws.cell(row=r, column=PAY_STATUS_COL_INDEX).value or "").startswith("Bill paid"):
                continue
            if ws.cell(row=r, column=LIEN_COL_INDEX).value in LIEN_OUTSTANDING:
                continue
            ws.row_dimensions[r].hidden = True
    ws.add_table(tbl)

    # Universal CF — color by RECONCILIATION (Pay Status × Invoice Status):
    # fronted / funded / done / pipeline. The color paints ONLY the Invoice
    # Status cell now, never the whole row (the user 2026-08-12) — Lien and
    # Approved likewise tint just their own cell. No cell is colored by another
    # cell's status, so the sheet reads cell-by-cell instead of as banded rows.
    pay_letter = _col_letter(PAY_STATUS_COL_INDEX)
    inv_letter = _col_letter(INVOICE_STATUS_COL_INDEX)
    approved_letter = _col_letter(APPROVED_COL_INDEX)
    lien_letter = _col_letter(LIEN_COL_INDEX)

    # CF formula row anchor — use DATA_START (3) so absolute-row references
    # in $-prefixed formulas point at the first DATA row, not the header.
    fr = DATA_START

    # ── Lien column CF (its own cell only) ──
    # A lien tag or an active supplier-notice timer paints just the Lien cell.
    # No row band overlaps it any more (colors are cell-scoped, the user
    # 2026-08-12), so the Lien cell stays uncolored when neither a tag nor a
    # timer applies.
    if lien_editable:
        lien_range = f"{lien_letter}{DATA_START}:{lien_letter}{last_row}"
        ob_letter = _col_letter(BILL_OPEN_BAL_COL_INDEX)
        unpaid = f"${ob_letter}{fr}>0"
        # Outstanding liens stay colored even when paid (lien outlives payment
        # until formally released).
        ws.conditional_formatting.add(lien_range,
            _cf_fill_rule(f'${lien_letter}{fr}="{LIEN_NOTICE}"', CF_LIEN_NOTICE, bold=True))
        ws.conditional_formatting.add(lien_range,
            _cf_fill_rule(f'${lien_letter}{fr}="{LIEN_FILED}"', CF_LIEN_FILED, bold=True))
        # Released tints only while UNPAID → paid+released archives clean.
        ws.conditional_formatting.add(lien_range,
            _cf_fill_rule(f'AND(${lien_letter}{fr}="{LIEN_RELEASED}",{unpaid})',
                          CF_LIEN_RELEASED, bold=True))
        # Supplier-notice timer: the Lien cell CARRIES the countdown bucket as
        # text (written in _write_bill_row), so color it by matching that text.
        # Script-written → pale tint + GREY ITALIC (vs the human tags above,
        # which are saturated + bold black). stop-if-true so the band below
        # doesn't repaint them.
        for label, color in (
            (TIMER_LABEL_PAST, CF_TIMER_PAST),
            (TIMER_LABEL_7,    CF_TIMER_HOT),
            (TIMER_LABEL_15,   CF_TIMER_ORANGE),
            (TIMER_LABEL_30,   CF_TIMER_YELLOW),
        ):
            ws.conditional_formatting.add(lien_range,
                _cf_fill_rule(f'${lien_letter}{fr}="{label}"', color,
                              italic=True, font_color=CF_TIMER_TEXT,
                              stop_if_true=True))

    # ── Approved-cell tint (its own cell only, never the row) ──
    # "approved" → green, "not approved" → red — both painted on the Approved
    # cell alone (the user 2026-08-12). Exact-match so "approved" doesn't also
    # fire on "not approved".
    approved_range = f"{approved_letter}{DATA_START}:{approved_letter}{last_row}"
    ws.conditional_formatting.add(approved_range,
        _cf_fill_rule(f'${approved_letter}{fr}="not approved"', CF_RED_HOLD, bold=True))
    ws.conditional_formatting.add(approved_range,
        _cf_fill_rule(f'${approved_letter}{fr}="approved"', CF_GREEN_READY))

    # ── Invoice-Status cell: color by the AR state (Good / Neutral / Bad) ──
    # Paints ONLY the Invoice Status cell (never the row). "Invoice paid" is
    # ALWAYS green now (the user 2026-08-12) — the old gray "done" tint is gone;
    # "Partially Paid/Awaiting Remainder" is neutral tan. Awaiting Payment /
    # Awaiting Invoice tint only once we've FRONTED the vendor (bill paid), so an
    # unpaid pipeline bill stays uncolored. States are mutually exclusive, so
    # rule order doesn't affect correctness.
    inv_range = f"{inv_letter}{DATA_START}:{inv_letter}{last_row}"
    # Pay Status carries a (fronted)/(collected) suffix on paid bills, so match
    # on the "Bill paid" PREFIX rather than the exact string.
    paid = f'LEFT(${pay_letter}{fr},9)="Bill paid"'
    cf_rules: List[Tuple[str, str]] = [
        # No project # (data issue) wins regardless of the other axis.
        (f'${inv_letter}{fr}="No project #"',                            CF_PURPLE_NEEDS_PROJECT),
        # Invoice paid in full → green, whether or not we've paid the vendor.
        (f'${inv_letter}{fr}="Invoice paid"',                            CF_GREEN_READY),
        # GC paid part of the invoice → neutral tan, remainder still due.
        (f'${inv_letter}{fr}="{STATUS_PARTIALLY_PAID_REMAINDER}"',       CF_NEUTRAL),
        # Fronted — we paid the vendor, GC hasn't reimbursed us.
        (f'AND({paid},${inv_letter}{fr}="Awaiting Payment")',           CF_YELLOW_AUDIT),
        (f'AND({paid},${inv_letter}{fr}="Awaiting Invoice")',           CF_RED_HOLD),
        # Partly funded across a multi-project bill → decide float or wait.
        (f'${inv_letter}{fr}="Partial paid"',                            CF_PEACH_PARTIAL),
        # Everything else (bill open + awaiting GC) → no tint, normal pipeline.
    ]
    for formula, color in cf_rules:
        ws.conditional_formatting.add(inv_range, _cf_fill_rule(formula, color))

    # Lien legend (horizontal, frozen row 1) + dropdown. The lien CF rules
    # themselves were added ABOVE the row-status loop so they out-prioritize
    # the row band on the Lien cell.
    if lien_editable:
        _lien_legend(ws, len(HEADERS) + 2)
        _rowcolor_legend(ws, len(HEADERS) + 2)
        # Dropdown so the tag is foolproof (no typos that would break the CF
        # match). Inline list — no formula2, so the Excel-strict validator
        # passes. allow_blank lets a bill carry no lien tag.
        lien_dv = DataValidation(
            type="list",
            formula1=f'"{LIEN_NOTICE},{LIEN_FILED},{LIEN_RELEASED}"',
            allow_blank=True,
        )
        ws.add_data_validation(lien_dv)
        lien_dv.add(f"{lien_letter}{DATA_START}:{lien_letter}{last_row}")

    # Repaint divider columns dark on every data row (Tables zebra-stripe
    # overrides per-cell fills otherwise; setting on each row is reliable).
    # Status CF is cell-scoped and never touches a divider column, so this fill
    # survives on every row.
    for r_i in range(DATA_START, last_row + 1):
        for c_i in DIVIDER_COL_INDEXES:
            ws.cell(row=r_i, column=c_i).fill = DIVIDER_FILL


def _render_rows(
    ws,
    rows: List[dict],
    table_name: str,
    edits: Dict[str, Dict[str, str]],
    row_filter: Optional[Callable[[dict], bool]] = None,
    sort_key: Optional[Callable[[dict], Tuple]] = None,
    hide_paid_default: bool = False,
    hide_cols: Optional[List[int]] = None,
    inv_anchor_map: Optional[Dict[str, int]] = None,
    lien_editable: bool = False,
) -> Tuple[int, Dict[str, int]]:
    """Apply filter + sort, write rows, finalize as Excel Table.

    Returns (row_count, bill_id_to_row_map).

    Layout: row 1 = supra-header bands, row 2 = column headers (Table
    header row), row 3+ = data. Table.ref starts at row 2.

    `hide_cols` — extra columns to hide (in addition to _Key). Use for
    Bills (hide Line Amount, redundant with Bill Total at bill grain) and
    Inventory (hide Pay — Bills-only decision).

    `inv_anchor_map` — {bill_id → row_on_Inventory}. If provided and a
    row's `is_multi_project` is True, an internal hyperlink is attached to
    the Project # cell so clicking "(multiple)" jumps to that bill's first
    line on the Inventory sheet.
    """
    _apply_header(ws, hide_cols=hide_cols)

    items = [r for r in rows if (row_filter is None or row_filter(r))]
    if sort_key is not None:
        items = sorted(items, key=sort_key)

    today = dt.date.today()
    bill_id_to_row: Dict[str, int] = {}
    proj_col_letter = _col_letter(HEADERS.index("Project #") + 1)
    for r_i, r in enumerate(items, start=3):
        _write_bill_row(ws, r_i, r, edits, today)
        bid = r.get("bill_id")
        if bid and bid not in bill_id_to_row:
            bill_id_to_row[bid] = r_i
        # Hyperlink the "(multiple)" cell on multi-project bills →
        # Inventory anchor row. Cell value stays as the visible text.
        if inv_anchor_map and r.get("is_multi_project") and bid in inv_anchor_map:
            target_row = inv_anchor_map[bid]
            cell = ws[f"{proj_col_letter}{r_i}"]
            cell.hyperlink = f"#Inventory!A{target_row}"
            cell.font = Font(
                name="Calibri", size=11, color="0563C1", underline="single",
            )

    last_row = max(len(items) + 2, 3)
    _finalize_sheet(ws, table_name, last_row, hide_paid_default=hide_paid_default,
                    lien_editable=lien_editable)
    return len(items), bill_id_to_row


# ─────────────────────── sort key ───────────────────────

def _vendor_key(r: dict) -> Tuple:
    """Default sort on the Bills sheet: Vendor → Bill Date → Bill # → Line."""
    return ((r.get("vendor") or "").upper(),
            r.get("bill_date") or dt.date(1900, 1, 1),
            r.get("bill_doc", ""),
            int(r.get("line_id") or 0) if str(r.get("line_id") or "").isdigit() else 0)


# ─────────────────────── Pay List (static, rebuilt each sync) ───────────────────────

# build_pay_list_sheet — removed 2026-05-29.
# Workflow: manager filters Bills sheet by Pay=x (Status filter already hides
# paid bills by default), selects visible rows, copy-pastes to a new workbook
# for AP. Simpler than maintaining a static-snapshot sheet that confused
# people about freshness.


# ─────────────────────── Liens sheet (live register of liened bills) ───────────────────────

# A LIVE view, not a sync-time snapshot. Each row pulls from the Bills sheet
# with SMALL/IF/INDEX formulas, so the instant the clerk tags a lien on Bills
# the row appears here — no waiting for the next sync (like a Notion filtered
# view). Read-only: liens are edited on Bills.
#
# Why not FILTER(): openpyxl can't write a true dynamic-array spill — Excel
# imports it as a single-cell legacy array and nothing spills (verified — the
# sheet came up blank). SMALL/IF/INDEX is the version-proof equivalent: a
# hidden helper column finds the k-th Bills row whose Lien is set (CSE array),
# and each visible cell INDEXes that row out of the Bills sheet. Because it
# reads the Bills rows directly, it also shows liened bills that are paid
# (those rows are kept visible on Bills).

LIENS_MAX_ROWS = 200          # pre-filled live-formula slots (liens are rare)
LIENS_SCAN_END = 5000         # Bills row range each helper scans

# (Liens header, Bills source header). INDEX pulls each column live from Bills.
LIENS_COLS = [
    ("Vendor",        "Vendor"),
    ("Bill #",        "Bill #"),
    ("Bill Date",     "Bill Date"),
    ("Project #",     "Project #"),
    ("Division",      "Division"),
    ("Client",        "Client"),
    ("Bill Total",    "Bill Total"),
    ("Bill Open Bal", "Bill Open Bal"),
    ("Pay Status",    "Pay Status"),
    ("Lien",          "Lien"),
    ("Notes",         "Notes"),
]
_LIENS_KIND = {"Bill Date": "date", "Bill Total": "money", "Bill Open Bal": "money"}
_LIENS_WIDTH = {"Vendor": 26, "Bill #": 12, "Bill Date": 11, "Project #": 12,
                "Division": 9, "Client": 26, "Bill Total": 13, "Bill Open Bal": 13,
                "Pay Status": 13, "Lien": 14, "Notes": 34}


def build_liens_sheet(ws) -> None:
    """Live register of every bill whose Lien tag is set (SMALL/IF/INDEX pulls
    from Bills). Updates on edit, no sync required."""
    n = len(LIENS_COLS)
    help_letter = _col_letter(n + 1)                 # hidden helper column
    bills_lien = _col_letter(LIEN_COL_INDEX)         # Lien column letter on Bills
    lien_out_letter = _col_letter(
        next(i + 1 for i, (h, _) in enumerate(LIENS_COLS) if h == "Lien"))

    # Header + widths.
    for c_i, (h, _) in enumerate(LIENS_COLS, 1):
        c = ws.cell(row=1, column=c_i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        ws.column_dimensions[_col_letter(c_i)].width = _LIENS_WIDTH.get(h, 14)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    ws.column_dimensions[help_letter].hidden = True

    last = LIENS_MAX_ROWS + 1
    scan = (f"Bills!${bills_lien}$3:${bills_lien}${LIENS_SCAN_END}")
    for r in range(2, last + 1):
        # Helper (CSE array): worksheet-row of the k-th Bills bill with a lien,
        # where k = ROW()-1. Blank once we run past the last lien.
        ws.cell(row=r, column=n + 1).value = ArrayFormula(
            ref=f"{help_letter}{r}",
            text=(f'=IFERROR(SMALL(IF({scan}<>"",ROW({scan})),ROW()-1),"")'),
        )
        for c_i, (h, src) in enumerate(LIENS_COLS, 1):
            src_letter = _col_letter(HEADERS.index(src) + 1)
            kind = _LIENS_KIND.get(h)
            idx = f'INDEX(Bills!{src_letter}:{src_letter},${help_letter}{r})'
            # Text columns: append &"" so an empty source cell shows blank, not
            # the 0 that a bare INDEX returns. Money/date stay numeric so a real
            # 0 / date isn't turned into text.
            expr = f'{idx}&""' if kind is None else idx
            cell = ws.cell(
                row=r, column=c_i,
                value=f'=IF(${help_letter}{r}="","",{expr})',
            )
            cell.font = BODY_FONT
            if kind == "money":
                cell.number_format = MONEY_FMT
                cell.alignment = ALIGN_RIGHT
            elif kind == "date":
                cell.number_format = DATE_FMT
                cell.alignment = ALIGN_CENTER

    # AutoFilter so AP can slice by Division / Lien within the register.
    ws.auto_filter.ref = f"A1:{_col_letter(n)}{last}"

    # Lien cell tint (amber / red / green) over the slot range.
    lien_range = f"{lien_out_letter}2:{lien_out_letter}{last}"
    for val, color in ((LIEN_NOTICE, CF_LIEN_NOTICE),
                       (LIEN_FILED, CF_LIEN_FILED),
                       (LIEN_RELEASED, CF_LIEN_RELEASED)):
        ws.conditional_formatting.add(
            lien_range,
            _cf_fill_rule(f'${lien_out_letter}2="{val}"', color, bold=True),
        )


# ─────────────────────── Inventory sheet (multi-project drill-down) ───────────────────────

# Multi-project bills (Martin Marietta-style inventory tickets that distribute
# across jobs) appear on the master Bills sheet as ONE summary row with
# Project# = "(multiple)". The per-line / per-project detail lives here so AP
# can see which jobs ate which dollars + each line's individual GC-invoice
# status (so the bill-level Status on the Bills sheet can roll up to
# "Partial paid" when some lines are funded and others aren't).
#
# 2026-06-04: Inventory now mirrors the Bills sheet layout exactly (same 22
# columns, same supra-headers, same CF). Pay column is hidden — that decision
# lives on the Bills sheet. Line Amount is visible here (the whole point —
# per-line dollar attribution).

INVENTORY_TABLE = "tblInventory"


def build_inventory_sheet(ws, line_rows: List[dict]
                          ) -> Tuple[int, Dict[str, int]]:
    """Render multi-project bills' line-level rows using the same layout as
    the master Bills sheet, with the Pay column hidden.

    Returns (line_count, {bill_id → first_row_on_inventory}) so the Bills
    renderer can attach hyperlinks from each "(multiple)" cell back to the
    matching anchor row here.

    Sort: Vendor → Bill Date → Bill # → Project # — keeps every bill's lines
    grouped together so the hyperlink lands you at the FIRST line of the
    bill and the rest sit immediately below.
    """
    multi_ids = multi_project_bill_ids(line_rows)
    items = [r for r in line_rows if r.get("bill_id") in multi_ids]

    sort_key = lambda r: (
        (r.get("vendor") or "").upper(),
        r.get("bill_date") or dt.date(1900, 1, 1),
        r.get("bill_doc", ""),
        (r.get("project_num") or ""),
    )

    # Inventory hides Pay (col 15) — Bills sheet is the source of truth for
    # the pay decision. _Key is always hidden by _apply_header.
    n, bill_id_to_row = _render_rows(
        ws,
        items,
        INVENTORY_TABLE,
        edits={},                       # no Lien/Notes preservation here
        sort_key=sort_key,
        hide_paid_default=True,         # same default as Bills
        hide_cols=[LIEN_COL_INDEX],
        inv_anchor_map=None,            # this IS the inventory sheet
    )
    return n, bill_id_to_row


# ─────────────────────── Bill payments (AP cash-out) ───────────────────────
# Rather than a separate sheet, the payment that paid each bill is shown as
# Pay Ref # / Pay Date / Pay Method columns on the Bills sheet, so the clerk can
# just filter by check # or pay date (the user 2026-07-13). No out-of-window bill
# fetch and no second sheet — we only annotate bills already loaded. The map is
# built once per run and read by _write_bill_row via this module global.
_BILL_PAY_MAP: Dict[str, dict] = {}


def _bp_method(bp: dict) -> str:
    """Normalize a BillPayment's PayType to a short label the clerk reads as
    'which statement do I open'."""
    pt = (bp.get("PayType") or "").strip().lower()
    if pt == "check":
        return "Check"
    if pt in ("creditcard", "cc"):
        return "CC"
    if pt == "cash":
        return "Cash"
    return (bp.get("PayType") or "?")


def _bp_linked_bills(bp: dict) -> List[Tuple[str, Optional[float]]]:
    """Return [(bill_id, amount_applied)] for a BillPayment. QBO puts the link
    either on each Line (with the applied Amount) or at the top level (no
    amount). Line-level wins; top-level fills any gap with amount None."""
    seen: Dict[str, Optional[float]] = {}
    for ln in bp.get("Line") or []:
        amt = float(ln.get("Amount") or 0)
        for lk in ln.get("LinkedTxn") or []:
            if lk.get("TxnType") == "Bill" and lk.get("TxnId"):
                seen[lk["TxnId"]] = amt
    for lk in bp.get("LinkedTxn") or []:
        if lk.get("TxnType") == "Bill" and lk.get("TxnId") and lk["TxnId"] not in seen:
            seen[lk["TxnId"]] = None
    return list(seen.items())


def build_bill_payment_map(bill_payments: List[dict]) -> Dict[str, dict]:
    """Map each paid bill_id -> {ref, date, method} from its BillPayment(s).
    Checks show the check # (DocNumber); CC refs are left blank (a generic
    label is no lookup key -- pay date + vendor + amount locate it). A bill paid
    by more than one payment shows the latest date and '(multiple)'."""
    acc: Dict[str, List[Tuple[Optional[dt.date], str, str, str]]] = defaultdict(list)
    for bp in bill_payments:
        pay_date = parse_date(bp.get("TxnDate"))
        method = _bp_method(bp)
        ref = (bp.get("DocNumber") or "").strip() if method == "Check" else ""
        pay_id = bp.get("Id", "") or ""
        for bid, _amt in _bp_linked_bills(bp):
            acc[bid].append((pay_date, method, ref, pay_id))
    out: Dict[str, dict] = {}
    for bid, lst in acc.items():
        lst.sort(key=lambda t: t[0] or dt.date.min)
        if len(lst) == 1:
            d, m, rf, pid = lst[0]
            out[bid] = {"date": d, "method": m, "ref": rf, "pay_id": pid}
        else:
            methods = {t[1] for t in lst}
            out[bid] = {
                "date": lst[-1][0],
                "method": methods.pop() if len(methods) == 1 else "(multiple)",
                "ref": "(multiple)",
                "pay_id": lst[-1][3],   # link the latest payment
            }
    return out

# ─────────────────────── QBO Audit sheet ───────────────────────

# QBO Class field uses spelled-out division names, not the RP/CP/MFD codes.
# Normalize before comparing in the audit. Anything not in this map gets the
# raw class string echoed back (so the audit message says exactly what's in
# QBO when there's a real mismatch — easier to spot bad data entry).
CLASS_TO_DIVISION = {
    "RESIDENTIAL": "RP",
    "RP":          "RP",
    "RES":         "RP",
    "COMMERCIAL":  "CP",
    "CP":          "CP",
    "COM":         "CP",
    "MULTI FAMILY":  "MFD",
    "MULTIFAMILY":   "MFD",
    "MULTI-FAMILY":  "MFD",
    "MFD":           "MFD",
    "MF":            "MFD",
}


def _normalize_class(class_name: str) -> Optional[str]:
    """Map a QBO Class value to a division code (RP/CP/MFD), or None if
    unrecognized. Case- and whitespace-insensitive."""
    if not class_name:
        return None
    key = " ".join(class_name.strip().upper().split())  # collapse internal spaces
    return CLASS_TO_DIVISION.get(key)


AUDIT_HEADERS = [
    "Bill #", "Vendor", "Bill Date", "Customer/Project (QBO)",
    "Project # (parsed)", "Division (expected)", "Class field (QBO)",
    "Memo (PrivateNote)", "Line Description",
    "Mismatch", "Open",
]

# 2026-06-04: rolling buffer for the stale-not-approved audit. NOT APPROVED
# bills older than this many days surface in their own audit section so the
# clerk can chase approvals before they age further. 30 days mirrors the
# old end-of-month batch cadence and self-corrects past month boundaries
# (a bill from May 30 doesn't become urgent until ~June 29, not June 1).
NOT_APPROVED_BUFFER_DAYS = 30

# Audit section banner styling — distinct color per section so the two
# concern types (approval pipeline vs data entry) read as separate work.
AUDIT_SECTION_STALE_FILL = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFFFCC66"),
                                       bgColor=Color(rgb="FFFFCC66"))
AUDIT_SECTION_DATA_FILL  = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFD9E1F2"),
                                       bgColor=Color(rgb="FFD9E1F2"))
AUDIT_SECTION_FONT       = Font(name="Calibri", size=12, bold=True, color="1F3864")
AUDIT_SECTION_HEIGHT     = 24

# Section 3 — uncoded JOB-COST lines (a real cost about to be left off a
# project). Strong orange so it reads as the most actionable section. Overhead-
# only uncoded lines are deliberately NOT shown (the user 2026-06-18: they'd flood
# the audit and are legitimately projectless). Sub bills are handled by their own
# section (6) — this section runs on the non-sub display population only.
AUDIT_SECTION_UNCODED_FILL = PatternFill(patternType="solid",
                                         fgColor=Color(rgb="FFF4B183"),
                                         bgColor=Color(rgb="FFF4B183"))
AUDIT_UNCODED_CELL_FILL    = PatternFill(patternType="solid",
                                         fgColor=Color(rgb="FFFCE4D6"),
                                         bgColor=Color(rgb="FFFCE4D6"))

# Age escalation on the Bill Date cell in the MISSING-project section. A blank
# project # is tolerable while the job is still being identified, but a real
# job cost shouldn't sit uncoded for long — the project is knowable within ~2
# weeks. >15 days old → yellow, >30 days → red. (the user 2026-07-10.)
AUDIT_MISSING_AGE_YELLOW_DAYS = 15
AUDIT_MISSING_AGE_RED_DAYS    = 30
AUDIT_AGE_YELLOW_FILL = PatternFill(patternType="solid",
                                    fgColor=Color(rgb="FFFFE699"),
                                    bgColor=Color(rgb="FFFFE699"))
AUDIT_AGE_RED_FILL    = PatternFill(patternType="solid",
                                    fgColor=Color(rgb="FFFF9999"),
                                    bgColor=Color(rgb="FFFF9999"))

# Project-code pattern for the uncoded job-cost check (own constant so the
# existing _audit_row_checks stays untouched).
_UNCODED_PROJ_RE = re.compile(r"\b(MFD|CP|RP)\d+(?:-FTW)?\b", re.IGNORECASE)

# Aging bucket sub-banner inside the stale section. Lighter amber than the
# main section banner — reads as a sub-heading, not a peer break.
AUDIT_BUCKET_FILL        = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFFFE5B0"),
                                       bgColor=Color(rgb="FFFFE5B0"))
AUDIT_BUCKET_FONT        = Font(name="Calibri", size=11, bold=True, color="1F3864")
AUDIT_BUCKET_HEIGHT      = 20

# Section 4 — duplicate bill # within a vendor tree (double-entry / double-pay
# risk). Purple family so it reads as its own concern, distinct from the amber
# approval section and blue data-entry section. Always rendered — even at zero,
# the clerk sees the check ran and found nothing.
AUDIT_SECTION_DUP_FILL   = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFB1A0C7"),
                                       bgColor=Color(rgb="FFB1A0C7"))
AUDIT_DUP_GROUP_FILL     = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFCCC0DA"),
                                       bgColor=Color(rgb="FFCCC0DA"))
AUDIT_DUP_CELL_FILL      = PatternFill(patternType="solid",
                                       fgColor=Color(rgb="FFE4DFEC"),
                                       bgColor=Color(rgb="FFE4DFEC"))

# Section 5 — FW (flatwork) cost code on a job where it doesn't belong: any CP
# or MFD job, or a base RP#### slab (FW is legitimate ONLY on the -FTW project).
# Red family — a coding error to fix at the source. Fed by the FULL bill
# population incl. subs (the user 2026-08-06).
AUDIT_SECTION_FW_FILL = PatternFill(patternType="solid",
                                    fgColor=Color(rgb="FFE59090"),
                                    bgColor=Color(rgb="FFE59090"))
AUDIT_FW_CELL_FILL    = PatternFill(patternType="solid",
                                    fgColor=Color(rgb="FFFCE4E4"),
                                    bgColor=Color(rgb="FFFCE4E4"))

# Section 6 — SUB bills missing a project # (the sub_bill_audit check, folded in
# 2026-08-06). Subs are off the display sheets, but a sub line with no project #
# is a real coding gap. Teal family so it reads as its own concern.
AUDIT_SECTION_SUB_FILL = PatternFill(patternType="solid",
                                     fgColor=Color(rgb="FF8FBFBF"),
                                     bgColor=Color(rgb="FF8FBFBF"))
AUDIT_SUB_CELL_FILL    = PatternFill(patternType="solid",
                                     fgColor=Color(rgb="FFE2F0F0"),
                                     bgColor=Color(rgb="FFE2F0F0"))

# FW (flatwork) cost-code prefix: item name whose leaf starts with FW + a digit
# (FW1, FW-2, FW 6, FW52 …). Cost codes live in the QBO Item name, not the account.
_FW_CODE_RE = re.compile(r"^FW\s*-?\s*\d", re.IGNORECASE)

# Shared cell-kind list for every audit data row (all sections use the same
# 11-column AUDIT_HEADERS layout).
AUDIT_ROW_KINDS = ["text", "text", "date", "text", "text", "text", "text",
                   "text", "text", "text", "link"]

# Aging buckets — most overdue first, so the clerk's eye lands on the worst
# offenders. (label, lower_inclusive, upper_exclusive)
AGING_BUCKETS: List[Tuple[str, int, int]] = [
    ("90+ days",   90, 10**9),
    ("60–90 days", 60, 90),
    ("30–60 days", 30, 60),
]


def _aging_bucket_for(days_old: int) -> str:
    """Map a days-old value to its aging-bucket label."""
    for label, lo, hi in AGING_BUCKETS:
        if lo <= days_old < hi:
            return label
    return AGING_BUCKETS[0][0]  # 90+ catches anything past the top


def _stale_not_approved_bills(
    line_rows: List[dict],
    today: dt.date,
    threshold_days: int = NOT_APPROVED_BUFFER_DAYS,
) -> List[dict]:
    """Return bills that are NOT APPROVED and aged past `threshold_days`.

    Deduped by bill_id — approval is bill-level, so one entry per bill even
    when the bill has many lines. Sorted oldest-first so the most overdue
    bills sit at the top of the audit section.

    `today` is passed in by the caller so the date stays consistent across
    the entire audit pass.
    """
    by_bill: Dict[str, dict] = {}
    for r in line_rows:
        bid = r.get("bill_id") or ""
        if not bid or bid in by_bill:
            continue
        if r.get("approved", True):
            continue  # bill IS approved — not in our scope
        bill_date = r.get("bill_date")
        if not isinstance(bill_date, dt.date):
            continue
        days_old = (today - bill_date).days
        if days_old < threshold_days:
            continue
        by_bill[bid] = {
            "bill_id": bid,
            "bill_doc": r.get("bill_doc") or "",
            "vendor": r.get("vendor") or "",
            "bill_date": bill_date,
            "customer_name": r.get("customer_name") or "",
            "project_num": r.get("project_num") or "",
            "division": r.get("division") or "",
            "class_name": r.get("class_name") or "",
            "days_old": days_old,
        }
    return sorted(by_bill.values(), key=lambda x: -x["days_old"])


def _audit_section_banner(ws, row_idx: int, text: str,
                          fill: PatternFill, n_cols: int) -> None:
    """Render a merged section banner row on the audit sheet."""
    ws.cell(row=row_idx, column=1, value=text)
    ws.merge_cells(start_row=row_idx, start_column=1,
                   end_row=row_idx, end_column=n_cols)
    for c_i in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c_i)
        cell.fill = fill
        cell.font = AUDIT_SECTION_FONT
    ws.cell(row=row_idx, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.row_dimensions[row_idx].height = AUDIT_SECTION_HEIGHT


def _audit_row_checks(r: dict) -> List[str]:
    """Run all audit checks for a single bill-line row. Return list of
    mismatch labels (empty list = no problems)."""
    issues: List[str] = []
    project_num = (r.get("project_num") or "").upper()
    expected_div = (r.get("division") or "").upper()
    actual_class = (r.get("class_name") or "").strip()

    # 1) Class field empty — only meaningful if project IS assigned
    if expected_div and not actual_class:
        issues.append("Class not set")

    # 2) Class set but normalizes to wrong division.
    #    QBO Class is spelled-out ("Residential" / "Commercial" / "Multi Family"),
    #    not the RP/CP/MFD codes — normalize via CLASS_TO_DIVISION first.
    if expected_div and actual_class:
        normalized = _normalize_class(actual_class)
        if normalized is None:
            # Class set but unrecognized — clerk typed something off-script
            issues.append(f"Class={actual_class!r} (not a recognized division)")
        elif normalized != expected_div:
            issues.append(f"Class={actual_class!r} ({normalized}) but project is {expected_div}")

    # 3) Line description has a different project # than Customer/Project
    line_desc = r.get("line_desc", "") or ""
    desc_match = re.search(r"\b(MFD|CP|RP)\d+(?:-FTW)?\b", line_desc, re.IGNORECASE)
    if desc_match:
        desc_proj = desc_match.group(0).upper()
        if project_num and desc_proj != project_num:
            issues.append(f"Line desc says {desc_proj} but project is {project_num}")
    return issues


def _uncoded_job_cost(r: dict) -> Optional[Tuple[str, str]]:
    """Guardrail: a line with NO project that shows job-cost signals is a real
    miss (a cost about to be left off a project). Return (reason, division) or
    None.

    Signals (high-confidence job work):
      1. item / cost-code line (bill_type == "COGS"), or
      2. Class set to a TOP-PARENT division class (_normalize_class → RP/CP/MFD;
         sub-classes / off-script classes don't count), or
      3. line description carries a project code (MFD/CP/RP####).

    Lines with no signal are treated as legitimate overhead and intentionally
    NOT flagged — that's the "don't flood the audit" guardrail (the user 2026-06-18).
    """
    if (r.get("project_num") or "").strip():
        return None  # has a project — covered by _audit_row_checks, not here

    is_item   = r.get("bill_type") == "COGS"
    class_div = _normalize_class(r.get("class_name") or "")   # top-parent division only
    desc      = r.get("line_desc") or ""
    code_m    = _UNCODED_PROJ_RE.search(desc)
    if not (is_item or class_div or code_m):
        return None  # no job-cost signal → overhead, skip

    cust = (r.get("customer_name") or "").strip()
    reason = (f"Missing project # (parent only: {cust})" if cust
              else "Missing Customer/Project")
    hints = []
    if is_item:
        hints.append("item/cost-code line")
    if class_div:
        hints.append(f"class={class_div}")
    if code_m:
        hints.append(f"desc says {code_m.group(0).upper()}")
    if hints:
        reason += " — " + ", ".join(hints)
    return reason, (class_div or "(none)")


def _fw_cost_code(cost_code: str) -> bool:
    """True if a cost code (raw QBO Item name) is an FW / flatwork code. Takes
    the leaf after the last ':' so a hierarchical item name still matches."""
    leaf = (cost_code or "").split(":")[-1].strip()
    return bool(_FW_CODE_RE.match(leaf))


def _fw_misplaced(r: dict) -> Optional[str]:
    """FW (flatwork) cost code on a job where it doesn't belong. FW is legitimate
    ONLY on -FTW projects; it is a miscode on any CP job, any MFD job, or a base
    RP#### slab (the user 2026-08-06). Division/slab are read from the project #,
    never the Class field. Returns a reason string, or None.

    Lines with no project # are out of scope here — they surface in the
    missing-project sections instead."""
    if not _fw_cost_code(r.get("cost_code") or ""):
        return None
    proj = (r.get("project_num") or "").strip().upper()
    div = (r.get("division") or "").strip().upper()
    if not div:
        return None
    code = (r.get("cost_code") or "").split(":")[-1].strip().upper()
    if div in ("CP", "MFD"):
        return f"FW code {code} on a {div} job ({proj}) — flatwork belongs to RP"
    if div == "RP" and not proj.endswith("-FTW"):
        return f"FW code {code} on RP slab {proj} — FW belongs on the -FTW project"
    return None


def _norm_ref(doc: str) -> str:
    """Normalize a bill ref # for duplicate matching: trim + uppercase."""
    return (doc or "").strip().upper()


# Credit-card-fee bills reuse a generic label ("CC", "CC FEE", "MONTHLY CC FEE")
# as their ref #, so the same label recurs across many unrelated bills and dates
# — not real duplicates. A CC-marker ref only counts as a duplicate when the
# copies also land on the SAME DAY (and same vendor), i.e. a genuine same-day
# double entry. The user 2026-07-10.
_CC_REF_RE = re.compile(r"\bCC\b")


def _is_cc_ref(ref_key: str) -> bool:
    """True if a normalized ref # is a credit-card-fee marker (has 'CC' as a
    standalone token)."""
    return bool(_CC_REF_RE.search(ref_key))


def _build_vendor_root(vendors: List[dict]) -> Dict[str, str]:
    """Map each vendor Id to its top-most ancestor Id by walking ParentRef
    (cycle-guarded). Root + every sub-vendor collapse to one tree key, so a
    duplicate ref # booked once to a parent and once to a sub-vendor is still
    caught. A vendor with no parent maps to itself."""
    parent: Dict[str, str] = {}
    ids = set()
    for v in vendors:
        vid = v.get("Id")
        if not vid:
            continue
        ids.add(vid)
        pid = (v.get("ParentRef") or {}).get("value")
        if pid and pid != vid:
            parent[vid] = pid

    def root_of(vid: str) -> str:
        seen = set()
        cur = vid
        while cur in parent and cur not in seen:
            seen.add(cur)
            cur = parent[cur]
        return cur

    return {vid: root_of(vid) for vid in ids}


def _duplicate_bill_groups(
    rows: List[dict],
    vendor_root: Optional[Dict[str, str]] = None,
    vendor_map: Optional[Dict[str, str]] = None,
) -> List[List[dict]]:
    """Group bills (deduped by bill_id) by (vendor tree root, normalized ref #)
    and return only the groups with 2+ distinct bills — the same ref # entered
    more than once under one vendor tree. Blank ref #s are skipped (nothing to
    compare). Groups sorted by tree name then ref #; bills within a group by
    date."""
    vendor_root = vendor_root or {}
    vendor_map = vendor_map or {}
    bills: Dict[str, dict] = {}
    for r in rows:
        bid = r.get("bill_id") or ""
        if not bid or bid in bills:
            continue
        doc_key = _norm_ref(r.get("bill_doc"))
        if not doc_key:
            continue
        vid = r.get("vendor_id") or ""
        root_id = vendor_root.get(vid, vid)
        bills[bid] = {
            "bill_id": bid,
            "bill_doc": r.get("bill_doc") or "",
            "vendor": r.get("vendor") or "",
            "root_id": root_id,
            "root_name": vendor_map.get(root_id, "") or (r.get("vendor") or ""),
            "bill_date": r.get("bill_date"),
            "bill_total": float(r.get("bill_total") or 0),
            "customer_name": r.get("customer_name") or "",
            "doc_key": doc_key,
        }
    groups: Dict[Tuple[str, str, str], List[dict]] = defaultdict(list)
    for b in bills.values():
        # Fall back to vendor name when we have no id (defensive — a row should
        # always carry vendor_id, but never merge two vendors on an empty key).
        tree_key = b["root_id"] or ("NAME:" + b["vendor"].upper())
        # CC-fee markers only group within the same day (same-day double entry);
        # a real ref # groups across all dates (date_part stays empty).
        date_part = ""
        if _is_cc_ref(b["doc_key"]):
            date_part = b["bill_date"].isoformat() if b["bill_date"] else "NODATE"
        groups[(tree_key, b["doc_key"], date_part)].append(b)
    out = [g for g in groups.values() if len(g) >= 2]
    for g in out:
        g.sort(key=lambda x: (x["bill_date"] or dt.date.min, x["bill_doc"]))
    out.sort(key=lambda g: ((g[0]["root_name"] or g[0]["vendor"] or "").upper(),
                            g[0]["doc_key"]))
    return out


def _audit_write_row(ws, row_idx: int, values: List, tint_fill: Optional[PatternFill],
                     level: int) -> None:
    """Write one 11-column audit data row, tint the Mismatch cell (col 10), and
    place it at the given outline level so the section is collapsible."""
    for c_i, (val, kind) in enumerate(zip(values, AUDIT_ROW_KINDS), start=1):
        c = ws.cell(row=row_idx, column=c_i, value=val)
        _format_data_cell(c, kind)
    if tint_fill is not None:
        ws.cell(row=row_idx, column=10).fill = tint_fill
    ws.row_dimensions[row_idx].outline_level = level


def _audit_sub_banner(ws, row_idx: int, text: str, fill: PatternFill,
                      n_cols: int, level: int) -> None:
    """Render a nested sub-group banner (aging bucket, duplicate group) at the
    given outline level."""
    ws.cell(row=row_idx, column=1, value=text)
    ws.merge_cells(start_row=row_idx, start_column=1,
                   end_row=row_idx, end_column=n_cols)
    for c_i in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c_i)
        cell.fill = fill
        cell.font = AUDIT_BUCKET_FONT
    ws.cell(row=row_idx, column=1).alignment = Alignment(
        horizontal="left", vertical="center", indent=2
    )
    ws.row_dimensions[row_idx].height = AUDIT_BUCKET_HEIGHT
    ws.row_dimensions[row_idx].outline_level = level


def _audit_none_row(ws, row_idx: int) -> None:
    """Placeholder row shown when a section has zero findings — the header still
    renders so the clerk sees at a glance that the check ran and was clean."""
    c = ws.cell(row=row_idx, column=1, value="✓ none")
    c.font = Font(italic=True, color="808080")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].outline_level = 1


def _missing_project_age_fill(bill_date, today: dt.date) -> Optional[PatternFill]:
    """Escalate the Bill Date cell of a MISSING-project line by age: yellow past
    AUDIT_MISSING_AGE_YELLOW_DAYS, red past AUDIT_MISSING_AGE_RED_DAYS. A recent
    bill (project still being identified) stays uncolored. None = no fill."""
    if not isinstance(bill_date, dt.date):
        return None
    days_old = (today - bill_date).days
    if days_old > AUDIT_MISSING_AGE_RED_DAYS:
        return AUDIT_AGE_RED_FILL
    if days_old > AUDIT_MISSING_AGE_YELLOW_DAYS:
        return AUDIT_AGE_YELLOW_FILL
    return None


def _audit_table_sheet(wb, sheet_name: str, table_name: str,
                       headers: List[str], col_kinds: List[str],
                       data_rows: List[list], widths: List[int]) -> int:
    """Create one audit section as its OWN sheet, wrapped in a proper Excel Table
    (AutoFilter + sort on every column). `col_kinds` formats each column
    (text/date/money/flag/link); a 'link' column's value is a bill_id, rendered as
    a QBO ↗ hyperlink button. An empty section still yields a valid one-row table
    ('✓ none found'). Table headers must be unique per table (validate_xlsx check C).
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    for c_i, name in enumerate(headers, 1):
        hc = ws.cell(row=1, column=c_i, value=name)
        hc.font = HEADER_FONT
        hc.fill = HEADER_FILL
        hc.alignment = ALIGN_CENTER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for c_i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(c_i)].width = w

    if not data_rows:
        nc = ws.cell(row=2, column=1, value="✓ none found")
        nc.font = Font(name="Calibri", size=11, italic=True, color="808080")
        last_row = 2
    else:
        r = 2
        for vals in data_rows:
            for c_i, (val, kind) in enumerate(zip(vals, col_kinds), 1):
                if kind == "link":
                    link = _qbo_link(val)
                    c = ws.cell(row=r, column=c_i, value=link or None)
                    _format_data_cell(c, "link")
                elif kind == "polink":
                    link = _qbo_po_link(val)
                    c = ws.cell(row=r, column=c_i, value=link or None)
                    _format_data_cell(c, "link")
                elif kind == "url":
                    c = ws.cell(row=r, column=c_i,
                                value=(f'=HYPERLINK("{val}","↗")' if val else None))
                    _format_data_cell(c, "link")
                else:
                    c = ws.cell(row=r, column=c_i, value=val)
                    _format_data_cell(c, kind)
            r += 1
        last_row = r - 1

    ref = f"A1:{_col_letter(len(headers))}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    return len(data_rows)


def _bill_url(bid: str) -> str:
    return QBO_BILL_URL_TEMPLATE.format(bill_id=bid) if bid else ""


def _po_url(pid: str) -> str:
    return QBO_PO_URL_TEMPLATE.format(po_id=pid) if pid else ""


def _load_audit_exclusions() -> Dict[str, Dict[str, list]]:
    """EXCEL-ONLY audit suppressions from <companyhealth>/audit_exclusions.json
    (the user 2026-08-25 — the ledger is untouched). Known-legit no-project vendors
    (equipment rental, overhead, insurance) and classes that should never trip the
    Missing Project check. Shape: {"missing_project": {"vendors": [...],
    "classes": [...]}}. Missing/broken file → no exclusions."""
    p = paths.companyhealth_dir() / "audit_exclusions.json"
    try:
        data = json.loads(p.read_text()) if p.exists() else {}
    except (OSError, json.JSONDecodeError):
        data = {}
    out: Dict[str, Dict[str, list]] = {}
    for check, cfg in data.items():
        if check.startswith("_") or not isinstance(cfg, dict):
            continue
        out[check] = {"vendors": [str(x).strip().upper() for x in cfg.get("vendors", [])],
                      "classes": [str(x).strip().upper() for x in cfg.get("classes", [])]}
    return out


def _excluded(r: dict, excl: Dict[str, list]) -> bool:
    """True if the row's vendor or QBO class matches an exclusion (substring, ci)."""
    vend = (r.get("vendor") or "").upper()
    cls = (r.get("class_name") or "").upper()
    return (any(v and v in vend for v in excl.get("vendors", ()))
            or any(c and c in cls for c in excl.get("classes", ())))


def _missing_po_bills(all_rows: List[dict], today: dt.date, days: int = 90) -> List[dict]:
    """Bills with NO PO that are real COGS purchases: not a sub, has ≥1 item/COGS
    line (so not an expense-only bill), within the last `days` (the user 2026-08-25).
    One row per bill. The mirror of the Unused-PO audit."""
    cutoff = today - dt.timedelta(days=days)
    by_bill: Dict[str, dict] = {}
    for r in all_rows:
        bid = r.get("bill_id")
        if not bid:
            continue
        b = by_bill.setdefault(bid, {"row": r, "cogs": False, "projects": set()})
        if r.get("bill_type") == "COGS":
            b["cogs"] = True
        if r.get("project_num"):
            b["projects"].add(r["project_num"])
    out = []
    for bid, b in by_bill.items():
        r = b["row"]
        d = r.get("bill_date")
        if (r.get("is_sub") or (r.get("po_num") or "").strip() or not b["cogs"]
                or not (isinstance(d, dt.date) and d >= cutoff)):
            continue
        out.append({"bill_id": bid, "bill_doc": r.get("bill_doc", ""),
                    "vendor": r.get("vendor", ""), "bill_date": d,
                    "project": ", ".join(sorted(b["projects"]))
                    or (r.get("customer_name", "") or ""),
                    "amount": r.get("bill_total") or 0.0})
    out.sort(key=lambda x: ((x["vendor"] or "").upper(), x["bill_date"]))
    return out


def _cost_code_findings(all_rows: List[dict], po_index: Optional[Dict[str, dict]],
                        tracker_by_po: Optional[Dict[str, dict]]
                        ) -> Tuple[List[list], Dict[str, str], List[dict]]:
    """Cost-code family miscodes → themed Coding rows, each with its PO origin.
    Returns (rows, vtype, flags). `flags` are the raw miscode dicts (vendor,
    bill_id, bill_doc, cost_code, project, amount, date, reason, vtype) that feed
    the persistent history log. Same logic as the old Audit - Cost Code sheet."""
    recs = []
    for r in all_rows:
        raw = r.get("cost_code", "") or ""
        number, cost_name = code_families(raw)
        recs.append({"vendor": r.get("vendor", "") or "", "number": number,
                     "cost_code": raw.split(":")[-1].strip(), "cost_name": cost_name,
                     "desc": r.get("line_desc", "") or "", "account": r.get("account", "") or "",
                     "bill_id": r.get("bill_id", ""), "bill_doc": r.get("bill_doc", ""),
                     "date": r.get("bill_date"), "project": r.get("project_num", "") or "",
                     "amount": r.get("line_amount") or 0.0, "po_num": r.get("po_num", "") or ""})
    override = load_override(paths.companyhealth_dir() / "concrete_suppliers.json")
    _agg, vtype = classify_vendors(recs, override=override)
    flags = flag_lines(recs, vtype)

    po_by_doc = index_by_doc(po_index) if po_index else {}
    bill_to_po: Dict[str, str] = {}
    for po, rec in (tracker_by_po or {}).items():
        for b in (rec.get("bill_no") or "").split(","):
            b = b.strip()
            if b:
                bill_to_po.setdefault(b, po)
    rows = []
    for f in flags:
        doc = _norm_po(f.get("po_num")) or _norm_po(
            bill_to_po.get((f.get("bill_doc") or "").strip(), ""))
        rec = po_by_doc.get(doc) if doc else None
        origin = po_origin(f.get("number"), rec.get("numbers") if rec else None, bool(doc))
        detail = f"[{TYPE_LABEL.get(f['vtype'], '')}] {f['reason']} · {origin}"
        rows.append(["Cost Code", f["vendor"], f["bill_doc"], f["date"], f["project"],
                     f["cost_code"], round(float(f["amount"]), 2), detail,
                     _bill_url(f["bill_id"]), f["bill_id"]])
    return rows, vtype, flags


def build_audits(wb, all_rows: List[dict],
                 po_index: Optional[Dict[str, dict]] = None,
                 tracker_by_po: Optional[Dict[str, dict]] = None,
                 tracker_meta: Optional[dict] = None,
                 vendor_root: Optional[Dict[str, str]] = None,
                 vendor_map: Optional[Dict[str, str]] = None,
                 audit_marks: Optional[Dict[str, str]] = None,
                 history_path: Optional[Path] = None) -> int:
    """THREE themed audit sheets (the user 2026-08-25 — de-bloat from 9 tabs). Each
    is one filterable Excel Table with an 'Issue' column so a single sheet covers a
    family of checks:
      Audit - Coding : Data Entry · Missing Project · FW Misplaced · Sub No Project · Cost Code
      Audit - PO     : Unused PO · Missing PO
      Audit - Bills  : Not Approved · Duplicates
    All the finding logic is unchanged — only the rendering is consolidated."""
    today = dt.date.today()
    display_rows = [r for r in all_rows if not r.get("is_sub")]
    sub_rows = [r for r in all_rows if r.get("is_sub")]
    mp_excl = _load_audit_exclusions().get("missing_project", {})
    audit_marks = audit_marks or {}

    # ── CODING ──────────────────────────────────────────────────────────
    # Each row ends with the raw bill_id (last element) so the Status mark can be
    # keyed to the bill and preserved across runs.
    coding: List[list] = []
    for r in display_rows:
        issues = _audit_row_checks(r)
        if issues:
            coding.append(["Data Entry", r.get("vendor", ""), r.get("bill_doc", ""),
                           r.get("bill_date"),
                           r.get("project_num", "") or (r.get("customer_name", "") or ""),
                           "", r.get("line_amount") or 0.0,
                           f"Class {r.get('class_name', '') or '(empty)'} · "
                           + " · ".join(issues), _bill_url(r.get("bill_id", "")),
                           r.get("bill_id", "")])
        uc = _uncoded_job_cost(r)
        if uc and not _excluded(r, mp_excl):   # skip known-legit no-project vendors/classes
            coding.append(["Missing Project", r.get("vendor", ""), r.get("bill_doc", ""),
                           r.get("bill_date"), r.get("customer_name", "") or "(none)",
                           "", r.get("line_amount") or 0.0, uc[0],
                           _bill_url(r.get("bill_id", "")), r.get("bill_id", "")])
    for r in all_rows:
        reason = _fw_misplaced(r)
        if reason:
            coding.append(["FW Misplaced", r.get("vendor", ""), r.get("bill_doc", ""),
                           r.get("bill_date"), r.get("project_num", "") or "(none)",
                           (r.get("cost_code", "") or "").split(":")[-1].strip(),
                           r.get("line_amount") or 0.0,
                           ("SUB · " if r.get("is_sub") else "") + reason,
                           _bill_url(r.get("bill_id", "")), r.get("bill_id", "")])
    for r in sub_rows:
        if not (r.get("project_num") or "").strip() and r.get("bill_type") == "COGS":
            coding.append(["Sub No Project", r.get("vendor", ""), r.get("bill_doc", ""),
                           r.get("bill_date"), "(none)",
                           (r.get("cost_code", "") or "").split(":")[-1].strip(),
                           r.get("line_amount") or 0.0, r.get("line_desc", "") or "",
                           _bill_url(r.get("bill_id", "")), r.get("bill_id", "")])
    cc_rows, vtype, cc_flags = _cost_code_findings(all_rows, po_index, tracker_by_po)
    coding.extend(cc_rows)
    coding.sort(key=lambda x: (x[0], (x[1] or "").upper(), str(x[2])))
    # Insert the editable, preserved Status (from the mark keyed by bill_id) after
    # Issue; carry the bill_id as a hidden _Key so the mark survives the next run.
    coding_data = [[r[0], audit_marks.get(r[9], ""), r[1], r[2], r[3], r[4], r[5],
                    r[6], r[7], r[8], r[9]] for r in coding]
    _audit_table_sheet(
        wb, "Audit - Coding", "tblAuditCoding",
        ["Issue", "Status", "Vendor", "Bill #", "Bill Date", "Project", "Cost Code",
         "Amount", "Detail", "Open", "_Key"],
        ["text", "text", "text", "text", "date", "text", "text", "money", "text",
         "url", "text"],
        coding_data, [16, 20, 24, 12, 11, 12, 11, 12, 46, 6, 12])
    ws = wb["Audit - Coding"]
    ws.column_dimensions[get_column_letter(11)].hidden = True   # hide _Key
    cnt = {t: sum(1 for v in vtype.values() if v == t)
           for t in ("concrete", "material", "both", "hauler", "review")}
    ws.cell(row=1, column=13,
            value=(f"Mark a Status (e.g. KEEP - reason) to acknowledge & keep an item; "
                   f"it persists and mirrors to the Bills Notes.  |  Cost-code vendor "
                   f"types: {cnt['concrete']} concrete · {cnt['material']} material · "
                   f"{cnt['both']} both · {cnt['hauler']} hauler · "
                   f"{cnt['review']} review")).font = Font(
        name="Calibri", size=10, italic=True, color="808080")

    # ── HISTORY ─────────────────────────────────────────────────────────
    # Persistent cost-code miscode log (the owner 2026-09-01): how often the
    # bill clerk miscodes over time + what got FIXED between refreshes. State is
    # a JSON OUTSIDE the repo; only a real run reaches here (dry-run bails first).
    hp = history_path or (paths.companyhealth_dir() / "cost_code_history.json")
    try:
        hist = cchist.load(hp)
        recap = cchist.update(hist, cc_flags, today)
        cchist.save(hp, hist)
        hist_rows = cchist.to_rows(hist, today)
        hist_data = [row[:-1] + [_bill_url(row[-1])] for row in hist_rows]
        _audit_table_sheet(
            wb, "Audit - History", "tblAuditHistory",
            ["Status", "Vendor", "Bill #", "Cost Code", "Reason", "Project",
             "First Seen", "Last Seen", "Times", "Fixed On", "QBO"],
            ["text", "text", "text", "text", "text", "text",
             "date", "date", "flag", "date", "url"],
            hist_data, [10, 24, 12, 11, 44, 11, 11, 11, 7, 11, 6])
        wb["Audit - History"].cell(row=1, column=13,
                                   value=cchist.recap_note(recap, today)).font = Font(
            name="Calibri", size=10, italic=True, color="808080")
        print(f"  cost-code history: {recap['open']} open · {recap['new']} new · "
              f"{recap['fixed']} fixed this run · {recap['rate']} new/run "
              f"({recap['runs']} runs) → {hp.name}")
    except Exception as e:                          # never let the log kill a run
        print(f"  ⚠ cost-code history skipped: {e}")

    # ── PO ──────────────────────────────────────────────────────────────
    po: List[list] = []
    if po_index is not None:
        po_by_doc = index_by_doc(po_index)
        for f in reconcile_unused_pos(po_by_doc, tracker_by_po or {}, today):
            detail = (f"QBO {f['qbo_status']} · bill?={f['qbo_bill']} · tracker "
                      f"bill {f['tracker_bill'] or '-'} · "
                      f"{f['days_open'] if f['days_open'] is not None else '?'}d open")
            po.append([f["reason"], f["vendor"], f["po"], f["po_date"], f["job"],
                       f["amount"] or 0.0, detail, _po_url(f["po_id"])])
    for m in _missing_po_bills(all_rows, today, 90):
        po.append(["Missing PO (COGS, no PO)", m["vendor"], m["bill_doc"],
                   m["bill_date"], m["project"], m["amount"] or 0.0,
                   "COGS bill with no PO (last 90 days)", _bill_url(m["bill_id"])])
    po.sort(key=lambda x: (x[0], (x[1] or "").upper()))
    _audit_table_sheet(
        wb, "Audit - PO", "tblAuditPO",
        ["Issue", "Vendor", "PO/Bill #", "Date", "Project/Job", "Amount",
         "Detail", "Open"],
        ["text", "text", "text", "date", "text", "money", "text", "url"],
        po, [26, 26, 12, 11, 12, 12, 46, 6])
    if tracker_meta:
        md = tracker_meta.get("max_date")
        cap = (f"PO tracker unavailable ({tracker_meta['error']})"
               if tracker_meta.get("error") else
               f"PO tracker: {Path(tracker_meta.get('path', '')).name} · through {md}")
        wb["Audit - PO"].cell(row=1, column=10, value=cap).font = Font(
            name="Calibri", size=10, italic=True, color="808080")

    # ── BILLS ───────────────────────────────────────────────────────────
    bills: List[list] = []
    for s in _stale_not_approved_bills(display_rows, today):
        bills.append(["Not Approved", s["vendor"], s["bill_doc"], s["bill_date"],
                      s["project_num"] or (s["customer_name"] or ""), "",
                      f"{_aging_bucket_for(s['days_old'])} · {s['days_old']}d old",
                      _bill_url(s["bill_id"])])
    for g in _duplicate_bill_groups(all_rows, vendor_root, vendor_map):
        same_amt = len({round(b["bill_total"], 2) for b in g}) == 1
        flag = "same $" if same_amt else "amounts differ"
        for b in g:
            bills.append(["Duplicate", b["vendor"], b["bill_doc"], b["bill_date"],
                          b["customer_name"], round(float(b["bill_total"]), 2),
                          f"vendor tree {b['root_name'] or b['vendor']} · {flag}",
                          _bill_url(b["bill_id"])])
    bills.sort(key=lambda x: (x[0], (x[1] or "").upper()))
    _audit_table_sheet(
        wb, "Audit - Bills", "tblAuditBills",
        ["Issue", "Vendor", "Bill/Ref #", "Date", "Project", "Amount", "Detail", "Open"],
        ["text", "text", "text", "date", "text", "money", "text", "url"],
        bills, [16, 26, 12, 11, 12, 12, 42, 6])

    print(f"  audits: Coding {len(coding)} · PO {len(po)} · Bills {len(bills)}")
    return len(coding) + len(po) + len(bills)


# ─────────────────────── main ───────────────────────

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="build rows but do not write xlsx")
    ap.add_argument("--limit", type=int, default=0, help="cap row count (smoke test)")
    args = ap.parse_args()

    started = dt.datetime.now()
    print(f"→ {started:%Y-%m-%d %H:%M:%S}  excel bill sync starting")
    print(f"   output: {OUTPUT_PATH}")

    print("→ authenticating to QBO (Touch ID) …")
    qbo_access, qbo_cid = load_credentials()
    print("  ok.")   # never echo the company_id / realm (consistent with sync-ar)

    print("→ fetching vendors …")
    vendors = query_all(qbo_access, qbo_cid, "Vendor")
    vendor_map = {
        v["Id"]: v.get("DisplayName") or v.get("CompanyName") or f"Vendor {v['Id']}"
        for v in vendors
    }
    vendor_root = _build_vendor_root(vendors)   # id → top-most parent id, for the dup audit
    print(f"  {len(vendor_map)} vendors")

    print("→ building account + item maps …")
    account_map, item_map = build_account_maps(qbo_access, qbo_cid)
    print(f"  {len(account_map)} accounts, {len(item_map)} items")

    print("→ building PO index …")
    po_index = build_po_index(qbo_access, qbo_cid, vendor_map)
    po_map = {pid: rec["doc"] for pid, rec in po_index.items()}
    print(f"  {len(po_index)} purchase orders")

    print("→ building invoice → payment date map …")
    payment_map = build_payment_map(qbo_access, qbo_cid)
    print(f"  {len(payment_map)} invoices with payment dates")

    # Open bills — any date, Balance > 0. FULL pull incl. sub bills: subs are
    # filtered out of the Bills/Inventory display sheets below but flow to the
    # QBO Audit sheet (the user 2026-08-06 — the tracker now pulls every bill;
    # sub findings surface in the audit, never on the display sheets).
    print("→ fetching OPEN bills (Balance > 0) …")
    open_bills = query_all(qbo_access, qbo_cid, "Bill", where="Balance > '0'")
    print(f"  {len(open_bills)} open bills")

    # Paid bills since cutoff — for ledger sheets
    print(f"→ fetching PAID bills since {PAID_CUTOFF_DATE} (Balance = 0) …")
    paid_bills = query_all(
        qbo_access, qbo_cid, "Bill",
        where=f"Balance = '0' AND TxnDate >= '{PAID_CUTOFF_DATE}'",
    )
    print(f"  {len(paid_bills)} paid bills")

    # Sub-bill ids (memo contains 'sub') — marked per row so subs reach the
    # audit only. Bills/Inventory exclude them; the audit sheet includes them.
    sub_ids = {b.get("Id", "") for b in (open_bills + paid_bills) if is_sub_bill(b)}
    print(f"  {len(sub_ids)} sub bills → audit only (excluded from Bills/Inventory)")

    print(f"→ fetching invoices since {INVOICE_CUTOFF_DATE} …")
    invoices_raw = query_all(
        qbo_access, qbo_cid, "Invoice",
        where=f"TxnDate >= '{INVOICE_CUTOFF_DATE}'",
    )
    invoices = [inv for inv in invoices_raw if not is_excluded_invoice(inv)]
    print(f"  {len(invoices_raw)} → {len(invoices)} after retainage exclusion")

    invoices_by_customer: Dict[str, List[dict]] = defaultdict(list)
    for inv in invoices:
        cv = (inv.get("CustomerRef") or {}).get("value", "")
        if cv:
            invoices_by_customer[cv].append(inv)

    # General List (READ-ONLY) → RP draw semantics. None (share unmounted / file
    # unreadable) degrades RP matching to today's amount-cover-only behavior.
    print("→ loading General List (read-only) …")
    gl_contracts = load_contracts()
    if gl_contracts is None:
        print("  ⚠ General List unavailable — RP matching degraded to amount-cover only")
    else:
        print(f"  {len(gl_contracts)} RP contract entries")

    print("→ building rows …")
    open_rows = build_rows(
        open_bills, invoices_by_customer, vendor_map, account_map, item_map, po_map,
        payment_map=payment_map, gl_contracts=gl_contracts,
    )
    paid_rows = build_rows(
        paid_bills, invoices_by_customer, vendor_map, account_map, item_map, po_map,
        payment_map=payment_map, gl_contracts=gl_contracts,
    )
    all_rows = open_rows + paid_rows
    # Mark sub-bill lines. all_rows (incl. subs) → QBO Audit sheet;
    # display_rows (subs removed) → Bills / Inventory / Liens, unchanged.
    for r in all_rows:
        r["is_sub"] = r.get("bill_id", "") in sub_ids
    display_rows = [r for r in all_rows if not r["is_sub"]]
    n_sub_rows = len(all_rows) - len(display_rows)
    print(f"  {len(open_rows)} open + {len(paid_rows)} paid = {len(all_rows)} line rows "
          f"({len(display_rows)} display · {n_sub_rows} sub lines audit-only)")

    # status / division breakdowns — over the display population (subs were
    # never counted on the display sheets, so these stay comparable run-to-run).
    breakdown: Dict[str, int] = defaultdict(int)
    by_division: Dict[str, int] = defaultdict(int)
    for r in display_rows:
        breakdown[r["auto_status"]] += 1
        by_division[r.get("division") or "(none)"] += 1
    print("→ status breakdown (display rows):")
    for s, n in sorted(breakdown.items(), key=lambda x: -x[1]):
        print(f"    {s}: {n}")
    print("→ by division:")
    for d, n in sorted(by_division.items()):
        print(f"    {d}: {n}")
    n_draw = sum(1 for r in display_rows if r.get("match_basis") == MATCH_BASIS_DRAW)
    n_final = sum(1 for r in display_rows if r.get("match_basis") == MATCH_BASIS_FINAL)
    print(f"→ RP draw semantics: {n_draw} draw-matched · {n_final} fully-billed")
    n_pushed = sum(1 for r in display_rows if r.get("match_basis") == MATCH_BASIS_PUSHED)
    if n_pushed:
        print(f"→ pushed to a later draw by agreement (shared/draw_moves): {n_pushed} rows")

    if args.limit > 0:
        print(f"→ --limit {args.limit}: capping row sets")
        all_rows = all_rows[: args.limit]
        display_rows = [r for r in all_rows if not r["is_sub"]]

    if args.dry_run:
        elapsed = (dt.datetime.now() - started).total_seconds()
        print(f"\n✓ dry run complete in {elapsed:.1f}s — workbook NOT written")
        return 0

    # Read existing workbook's Bills sheet for Lien/Notes preservation
    print("→ reading existing workbook for Lien/Notes preservation …")
    edits = preserve_edits(OUTPUT_PATH)
    if edits:
        nonblank = sum(1 for v in edits.values()
                       if (v.get("Lien") or "") or (v.get("Notes") or ""))
        print(f"  {nonblank} _Keys with Lien/Notes edits carried forward")
    else:
        print("  no existing workbook — first run, nothing to preserve")

    # Audit Status marks (Audit - Coding) — persist across runs AND push into the
    # Bills Notes so the two stay consistent (the audit is the entry point).
    audit_marks = preserve_audit_marks(OUTPUT_PATH)
    for _bid, _mark in audit_marks.items():
        _mark = (_mark or "").strip()
        if not _mark:
            continue
        _slot = edits.setdefault(_bid, {"Lien": "", "Notes": ""})
        _note = _slot.get("Notes") or ""
        if _mark not in _note:                      # idempotent append (no dup)
            _slot["Notes"] = f"{_note} · {_mark}".strip(" ·") if _note else _mark
    if audit_marks:
        print(f"  {len(audit_marks)} audit Status mark(s) carried forward → Bills Notes")

    print("→ rotating backup …")
    rotate_backup(OUTPUT_PATH)

    # Two derived row-sets from the line-level data:
    #   • bills_view_rows  — bill-grain (one row per bill). Multi-project
    #                         bills show Project# = "(multiple)".
    #   • all_rows         — line-level. Inventory + Audit sheets use this.
    # (The per-division MFD/RP/CP grain was dropped 2026-07-13 — the Project #
    # already lives on the Bills sheet and the division sheets went unused.)
    print("→ collapsing rows for display …")
    bills_view_rows = collapse_rows(display_rows, grain="bill")
    n_multi = len(multi_project_bill_ids(display_rows))
    print(f"  {len(bills_view_rows)} bills · "
          f"{n_multi} multi-project bills routed to Inventory")

    # Bill payments (AP cash-out) → Pay Ref #/Date/Method columns on the Bills
    # sheet (no separate sheet, no out-of-window bill fetch — we only annotate
    # bills already loaded). One BillPayment fetch since the cutoff, then a
    # bill_id → payment map that _write_bill_row reads via the module global.
    global _BILL_PAY_MAP
    print(f"→ fetching bill payments since {PAID_CUTOFF_DATE} …")
    bill_payments = query_all(qbo_access, qbo_cid, "BillPayment",
                              where=f"TxnDate >= '{PAID_CUTOFF_DATE}'")
    _BILL_PAY_MAP = build_bill_payment_map(bill_payments)
    print(f"  {len(bill_payments)} bill payments → {len(_BILL_PAY_MAP)} bills annotated")

    print("→ building workbook …")
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    ws_bills = wb.create_sheet("Bills")
    ws_liens = wb.create_sheet("Liens")
    ws_inv   = wb.create_sheet("Inventory")
    # The audit is now THREE themed Excel Tables (build_audits): Coding · PO ·
    # Bills — created after the display sheets, below.
    # MFD/RP/CP division sheets removed 2026-07-13 (unused; Project # is on Bills).
    # Bill payments show as Pay columns on Bills, not a separate sheet (2026-07-13).

    # Render order matters: build Inventory FIRST so we have the
    # {bill_id → row} anchor map to wire the Bills sheet hyperlinks.
    # Display sheets use display_rows (subs excluded); the audit uses all_rows.
    n_inv, inv_anchor_map = build_inventory_sheet(ws_inv, display_rows)

    n_bills, _ = _render_rows(
        ws_bills, bills_view_rows, BILLS_TABLE, edits,
        sort_key=_vendor_key,
        hide_paid_default=True,
        # Bills hides Line Amount (= Bill Total at bill grain) plus Account (H)
        # and Line Description (I) — detail lives on Inventory; the bill-grain
        # summary doesn't need them by default.
        hide_cols=[ACCOUNT_COL_INDEX, LINE_DESC_COL_INDEX, LINE_AMT_COL_INDEX],
        inv_anchor_map=inv_anchor_map,    # hyperlink "(multiple)" → Inventory
        lien_editable=True,               # Bills is the only sheet with the Lien tag + dropdown
    )
    build_liens_sheet(ws_liens)       # live FILTER view of tblBills (Lien set)

    # PO tracker (read-only) feeds the Unused-PO + Cost-Code PO-origin checks.
    print("→ loading PO tracker (read-only) …")
    tracker_by_po, tracker_meta = load_po_tracker()
    if tracker_meta.get("error"):
        print(f"  ⚠ PO tracker unavailable: {tracker_meta['error']} — PO checks QBO-only")
    else:
        print(f"  tracker: {tracker_meta['po_count']} POs, data through {tracker_meta['max_date']}")

    # THREE themed audit sheets (Coding · PO · Bills) — de-bloat from 9 tabs.
    n_audit = build_audits(wb, all_rows, po_index=po_index,
                           tracker_by_po=tracker_by_po, tracker_meta=tracker_meta,
                           vendor_root=vendor_root, vendor_map=vendor_map,
                           audit_marks=audit_marks)

    print(f"  Bills: {n_bills} bills (open + paid since {PAID_CUTOFF_DATE})")
    print(f"  Liens: live view  ·  Inventory: {n_inv} lines  ·  Audit: {n_audit} rows "
          f"across 3 themed sheets + cost-code History log")

    wb.save(OUTPUT_PATH)
    post_process_xlsx(OUTPUT_PATH)

    # Excel-strict preflight — bail loud if any known-bad pattern survived.
    # Better to fail here than ship a workbook that Excel will mangle on open.
    print("→ validating xlsx (Excel-strict preflight) …")
    failures = validate_xlsx(OUTPUT_PATH)
    if failures:
        print("✗ validation FAILED — Excel would warn on open:")
        for f in failures:
            print(f"    {f}")
        print(f"\n  workbook saved anyway at: {OUTPUT_PATH}")
        print("  but you should NOT trust it until these are fixed.")
        return 2
    print(f"  ✓ saved {OUTPUT_PATH}")

    try:
        os.chmod(OUTPUT_PATH, stat.S_IRUSR | stat.S_IWUSR)
    except Exception as e:
        print(f"  ⚠ chmod 600 failed: {e}")

    elapsed = (dt.datetime.now() - started).total_seconds()
    print(f"\n✓ done in {elapsed:.1f}s")
    print(f"   open: {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
