#!/usr/bin/env python3
"""
general_list.py — READ-ONLY reader for the company General List (LISTA GENERAL).

Pulls each RP job's contract value (slab + flatwork bids) so the bill tracker can
apply RP draw semantics: a job with a signed contract that QBO shows only
partially billed authorizes early AP bills off its existing invoices (draws).

The General List is the estimators' LIVE entry sheet on Synology — this module
NEVER writes to it (opened read_only, closed immediately). When the share is
unmounted or the file can't be read, load_contracts() returns None and the
caller degrades to today's amount-cover-only matching.

USAGE (self-check, read-only):
  python3 bill-tracker/general_list.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Dict, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # repo root → shared
from shared import paths

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip3 install --break-system-packages openpyxl")
    sys.exit(1)

# READ-ONLY. Never write. Per-machine override: ACB_GENERAL_LIST_XLSX.
GENERAL_LIST_XLSX = paths.get_path(
    "ACB_GENERAL_LIST_XLSX",
    "/Volumes/Common/OPERATIONS/GENERAL LIST/LISTA GENERAL AÑO 2026.xlsx",
)

_SHEETS = ("General list - Alpha order", "Small Jobs")
_DATA_START_ROW = 6
_COL_JOB = 3     # C  — job number
_COL_SLAB = 35   # AI — slab / post-tension bid
_COL_FLAT = 37   # AK — flatwork bid
_RP_RE = re.compile(r"^\s*(RP\d{4})\b", re.IGNORECASE)


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_contracts(path: Optional[Path] = None) -> Optional[Dict[str, float]]:
    """{RP job base or base-FTW → contract $} from the General List, or None if
    the file is missing/unreadable (caller degrades). Read-only; never saves.

    Slab bid (col AI) keys the base job (e.g. RP7186); flatwork bid (col AK)
    keys the -FTW variant. Alpha sheet wins over Small Jobs via setdefault.
    """
    p = Path(path) if path else GENERAL_LIST_XLSX
    if not p.exists():
        return None
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception:
        return None
    contracts: Dict[str, float] = {}
    try:
        for sheet in _SHEETS:
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(min_row=_DATA_START_ROW, values_only=True):
                if len(row) < _COL_JOB:
                    continue
                m = _RP_RE.match(str(row[_COL_JOB - 1] or ""))
                if not m:
                    continue
                base = m.group(1).upper()
                slab = _num(row[_COL_SLAB - 1]) if len(row) >= _COL_SLAB else 0.0
                flat = _num(row[_COL_FLAT - 1]) if len(row) >= _COL_FLAT else 0.0
                if slab > 0:
                    contracts.setdefault(base, slab)
                if flat > 0:
                    contracts.setdefault(base + "-FTW", flat)
    finally:
        wb.close()
    return contracts


def _self_check() -> int:
    print(f"General List (READ-ONLY): {GENERAL_LIST_XLSX}")
    print(f"  exists: {GENERAL_LIST_XLSX.exists()}")
    contracts = load_contracts()
    if contracts is None:
        print("  load_contracts() → None (missing/unreadable — caller degrades)")
        return 1
    print(f"  {len(contracts)} RP contract entries")
    for k in sorted(contracts)[:10]:
        print(f"    {k}: {contracts[k]:,.2f}")
    return 0


if __name__ == "__main__":
    sys.exit(_self_check())
