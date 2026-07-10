#!/usr/bin/env python3
"""
duplicate_bill_audit.py — same Bill Ref # entered twice under one vendor tree.

Finds AP Bills that share a DocNumber (Bill Ref #) within the SAME vendor
"tree" — a top-level vendor plus any of its QBO sub-vendors. A vendor's
invoice number is unique to that vendor, so the same ref # on two bills is a
double-entry / double-payment risk. Comparing across the whole tree (root +
sub-vendors) catches the case where the same paper invoice was booked once to
"ABC Supply" and once to "ABC Supply:Yard 2".

Matching is case-insensitive and whitespace-trimmed on the ref #. Bills with a
blank ref # are skipped (nothing to compare). Read-only — reports only, never
edits QBO. Reuses qbo_bill_tracker (auth/paging/parsing).

USAGE
  python3 duplicate_bill_audit.py                 # all bills, all time
  python3 duplicate_bill_audit.py 2026-01-01      # bills dated on/after
  python3 duplicate_bill_audit.py 2026-01-01 2026-06-30
  python3 duplicate_bill_audit.py --dry-run       # fetch + report, write nothing
"""
from __future__ import annotations
import argparse
import datetime as dt
import re
import sys
from pathlib import Path

from qbo_bill_tracker import load_credentials, query_all, parse_date
from job_coding_audit import txn_link
import paths

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip3 install --break-system-packages openpyxl")
    sys.exit(1)

OUTDIR = paths.get_path(
    "ACB_AUDIT_OUT_DIR",
    paths.onedrive_base() / "Works In Progress" / "QBO Audits",
)

COLUMNS = [
    ("Group", 7), ("Vendor Tree (root)", 28), ("Bill Date", 12),
    ("Vendor (as entered)", 28), ("Bill Ref #", 16), ("Bill Total", 13),
    ("Bill Open Bal", 14), ("Copies in tree", 14), ("Same amount?", 13),
    ("Memo", 40), ("Open in QBO", 12),
]


def fetch_vendor_tree(access, cid):
    """Return (id_to_name, id_to_root). id_to_root resolves each vendor Id to
    the Id of its top-most ancestor by walking ParentRef (cycle-guarded), so
    root + every sub-vendor collapse to a single tree key."""
    vendors = query_all(access, cid, "Vendor", verbose=False)
    parent = {}
    id_to_name = {}
    for v in vendors:
        vid = v.get("Id")
        if not vid:
            continue
        id_to_name[vid] = v.get("DisplayName") or v.get("CompanyName") or vid
        pref = v.get("ParentRef") or {}
        pid = pref.get("value")
        if pid and pid != vid:
            parent[vid] = pid

    def root_of(vid):
        seen = set()
        cur = vid
        while cur in parent and cur not in seen:
            seen.add(cur)
            cur = parent[cur]
        return cur

    id_to_root = {vid: root_of(vid) for vid in id_to_name}
    return id_to_name, id_to_root


def _norm_ref(doc):
    return (doc or "").strip().upper()


# Credit-card-fee bills reuse a generic label ("CC", "CC FEE", "MONTHLY CC FEE")
# as their ref #, so the same label recurs across many unrelated bills and dates
# — not real duplicates. A CC-marker ref only counts as a duplicate when the
# copies also land on the SAME DAY (and same vendor), i.e. a genuine same-day
# double entry. Ted 2026-07-10.
_CC_REF_RE = re.compile(r"\bCC\b")


def _is_cc_ref(ref_key):
    """True if a normalized ref # is a credit-card-fee marker (has 'CC' as a
    standalone token)."""
    return bool(_CC_REF_RE.search(ref_key))


def build_rows(bills, id_to_name, id_to_root):
    """Group bills by (vendor tree root, normalized ref #); keep only groups
    with 2+ distinct bills. Returns display rows, group-numbered, sorted so a
    tree's duplicate sets sit together.

    CC-fee markers ('CC', 'CC FEE', …) are the exception: they only group
    within the same day, so a reused CC label across dates isn't flagged."""
    groups = {}  # (root_id, ref_key, date_part) -> list of bill dicts
    for b in bills:
        ref_key = _norm_ref(b.get("DocNumber"))
        if not ref_key:
            continue
        vref = b.get("VendorRef") or {}
        vid = vref.get("value") or ""
        root_id = id_to_root.get(vid, vid)
        # CC-fee markers only group within the same day; a real ref # groups
        # across all dates (date_part stays empty).
        date_part = ""
        if _is_cc_ref(ref_key):
            bd = parse_date(b.get("TxnDate"))
            date_part = bd.isoformat() if bd else "NODATE"
        groups.setdefault((root_id, ref_key, date_part), []).append(b)

    dup_groups = []
    for (root_id, ref_key, _date_part), bs in groups.items():
        # De-dupe by Bill Id in case the same record surfaced twice.
        by_id = {b.get("Id"): b for b in bs}
        if len(by_id) < 2:
            continue
        root_name = id_to_name.get(root_id, "")
        dup_groups.append((root_name, ref_key, list(by_id.values())))

    # Sort groups by tree name, then ref #, so duplicate sets read top-to-bottom.
    dup_groups.sort(key=lambda g: (g[0].lower(), g[1]))

    rows = []
    for gi, (root_name, ref_key, bs) in enumerate(dup_groups, 1):
        bs.sort(key=lambda b: (parse_date(b.get("TxnDate")) or dt.date.min))
        totals = {round(float(b.get("TotalAmt") or 0), 2) for b in bs}
        same_amt = "YES" if len(totals) == 1 else "no"
        for b in bs:
            vref = b.get("VendorRef") or {}
            vname = id_to_name.get(vref.get("value") or "",
                                   vref.get("name") or "")
            rows.append({
                "group": gi,
                "tree": root_name,
                "bill_date": parse_date(b.get("TxnDate")),
                "vendor": vname,
                "doc": (b.get("DocNumber") or "").strip(),
                "total": float(b.get("TotalAmt") or 0),
                "bal": float(b.get("Balance") or 0),
                "copies": len(bs),
                "same_amt": same_amt,
                "memo": (b.get("PrivateNote") or "").strip(),
                "link": txn_link("Bill", b),
            })
    return rows


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate Bill Refs"
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top")
    for ci, (name, w) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = bold
        c.alignment = left
        ws.column_dimensions[get_column_letter(ci)].width = w
    if not rows:
        ws.cell(row=2, column=1,
                value="No duplicate bill ref #s within a vendor tree in this "
                      "window.").alignment = left
        ws.freeze_panes = "A2"
    else:
        r = 2
        for row in rows:
            ws.cell(row=r, column=1, value=row["group"]).alignment = left
            ws.cell(row=r, column=2, value=row["tree"]).alignment = left
            d = row["bill_date"]
            ws.cell(row=r, column=3, value=d.strftime("%m/%d/%Y") if d else "").alignment = left
            ws.cell(row=r, column=4, value=row["vendor"]).alignment = left
            ws.cell(row=r, column=5, value=row["doc"]).alignment = left
            for col, key in ((6, "total"), (7, "bal")):
                c = ws.cell(row=r, column=col, value=row[key])
                c.number_format = '#,##0.00'
                c.alignment = right
            ws.cell(row=r, column=8, value=row["copies"]).alignment = left
            ws.cell(row=r, column=9, value=row["same_amt"]).alignment = left
            ws.cell(row=r, column=10, value=row["memo"]).alignment = left
            lk = ws.cell(row=r, column=11, value="open")
            lk.hyperlink = row["link"]
            lk.alignment = left
            r += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:" + get_column_letter(len(COLUMNS)) + str(r - 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run(a):
    since = parse_date(a.since) if a.since else None
    if a.since and not since:
        print("bad start date: " + repr(a.since) + " (YYYY-MM-DD)")
        return 2
    until = parse_date(a.until) if a.until else None
    if a.until and not until:
        print("bad end date: " + repr(a.until) + " (YYYY-MM-DD)")
        return 2
    where = ""
    if since:
        where = "TxnDate >= '" + since.isoformat() + "'"
    if until:
        where = (where + " AND " if where else "") + "TxnDate <= '" + until.isoformat() + "'"
    scope = (str(since) if since else "all time") + " -> " + (str(until) if until else "today")
    print("Duplicate bill ref #s within a vendor tree  (" + scope + ")")
    if not where:
        print("  (no date filter - pulling all bills, this can take a bit)")
    access, cid = load_credentials()
    print("  fetching Vendors ...", flush=True)
    id_to_name, id_to_root = fetch_vendor_tree(access, cid)
    print("    " + str(len(id_to_name)) + " vendors")
    print("  fetching Bills ...", flush=True)
    bills = query_all(access, cid, "Bill", where=where, verbose=True)
    print("    " + str(len(bills)) + " bills")
    rows = build_rows(bills, id_to_name, id_to_root)
    n_groups = len({r["group"] for r in rows})
    print("  " + str(n_groups) + " duplicate group(s), "
          + str(len(rows)) + " bill(s) flagged")
    if a.dry_run:
        print("(dry run - no file written)")
        return 0
    suffix = ("_" + str(since.year) if since else "")
    out = (Path(a.out).expanduser() if a.out
           else OUTDIR / ("Duplicate_Bill_Refs" + suffix + ".xlsx"))
    write_xlsx(out, rows)
    print("wrote " + str(out))
    return 0


def main():
    p = argparse.ArgumentParser(
        description="Audit AP bills for a duplicate ref # within a vendor tree.")
    p.add_argument("since", nargs="?", default="",
                   help="Start date YYYY-MM-DD (default all time).")
    p.add_argument("until", nargs="?", default="",
                   help="End date YYYY-MM-DD (optional).")
    p.add_argument("--out", default="",
                   help="Output .xlsx (default OneDrive QBO Audits).")
    p.add_argument("--dry-run", action="store_true")
    return run(p.parse_args())


if __name__ == "__main__":
    sys.exit(main())
