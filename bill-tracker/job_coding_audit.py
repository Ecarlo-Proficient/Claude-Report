#!/usr/bin/env python3
"""
job_coding_audit.py — Reusable QBO coding-mistake catcher for ANY job code.

WHY THIS EXISTS
  Jobs were tracked as CLASSES in QuickBooks Desktop. QBO doesn't like that —
  the job belongs in Customer/Project, and Class should only carry the DIVISION
  (Residential / Commercial / Multi Family). This script finds the leftovers:
  lines that mention a job in their text but aren't coded to that job's Project,
  and lines on the job that are missing a Customer/Project or still carry a
  non-parent (project-named) class.

  It is NOT hard-wired to one job. You pass the job code at run time, so the
  same tool works for MFD281 today and CP142 tomorrow.

WHAT IT CHECKS  (run one or both)
  CHECK A — "desc"  (description mismatch)
      Every expense line whose DESCRIPTION contains the job code but whose
      Customer/Project does NOT resolve to that exact job code. These are lines
      that talk about the job but are coded to the wrong project (or none).

  CHECK B — "memo"  (memo scan)
      Every transaction whose MEMO (PrivateNote) contains the job code. For
      each expense line on those transactions, flag the line if EITHER:
        (a) it has no Customer/Project assigned, OR
        (b) its Class is not exactly a division parent
            (Residential / Commercial / Multi Family). A blank class, or a
            sub/project-named class like "Multi Family:MFD281", is flagged.

TRANSACTION TYPES SCANNED  (choose with --types)
  bill          AP Bills
  check         Checks            (QBO Purchase, PaymentType=Check)
  cc            Credit-card / cash expenses (QBO Purchase, CreditCard/Cash)
  vendorcredit  Vendor Credits
  je            Journal Entries   (opt-in; off by default)
  invoice       Invoices          (opt-in; off by default)
  Default --types = bill,check,cc,vendorcredit  (all vendor-cost txns).

USAGE
  python3 job_coding_audit.py --job MFD281
  python3 job_coding_audit.py --job CP142 --checks desc
  python3 job_coding_audit.py --job MFD281 --types bill,check --since 2025-01-01
  python3 job_coding_audit.py --job MFD281 --out ~/Desktop/MFD281_audit.xlsx
  python3 job_coding_audit.py --job MFD281 --dry-run     # fetch + report, no file

OUTPUT
  A plain Excel workbook (black on white, no fills) with up to two sheets:
    "Desc Mismatch"  — Check A rows
    "Memo Scan"      — Check B rows
  Each row carries a ↗ hyperlink that opens the exact transaction in QBO so
  you can fix the coding in place.

NOTE
  Read-only against QBO. It never edits a transaction — it only reports.
  Runs on the Mac that holds the QBO Keychain blob (Touch ID once per run).
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Reuse the existing QBO plumbing — same auth, paging, parsing the bill
# tracker uses. Keeps one source of truth for credentials and project parsing.
from qbo_bill_tracker import (
    load_credentials,
    query_all,
    get_line_customer_ref,
    get_project_num,
    parse_date,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ pip3 install --break-system-packages openpyxl")
    sys.exit(1)


# ───────────────────────── configuration ─────────────────────────

# The ONLY acceptable Class values are the division parents. QBO stores them
# exactly like this (verified) — note "Multi Family" with a space, not
# "Multifamily". Override with --parents if your file differs.
DEFAULT_PARENT_CLASSES = ["Residential", "Commercial", "Multi Family"]

# Only these line detail types carry a Customer/Project + Class. Everything
# else on a transaction (subtotals, description-only, the source bank line on
# a Purchase, etc.) is skipped.
EXPENSE_DETAIL_TYPES = ("AccountBasedExpenseLineDetail", "ItemBasedExpenseLineDetail")

# Transaction-type registry. Each entry: QBO entity name + how to build the
# deep link + how to read the vendor. PaymentType splits Purchase into
# check vs cc/cash so the link lands on the right QBO screen.
QBO_BASE = "https://app.qbo.intuit.com/app"

# txn-type token → (QBO entity, default DocNumber-less link path)
TYPE_ENTITY = {
    "bill": "Bill",
    "check": "Purchase",
    "cc": "Purchase",
    "vendorcredit": "VendorCredit",
    "je": "JournalEntry",
    "invoice": "Invoice",
}
ALL_TYPES = list(TYPE_ENTITY.keys())
DEFAULT_TYPES = ["bill", "check", "cc", "vendorcredit"]


# ───────────────────────── link + field helpers ─────────────────────────

def _purchase_link_path(payment_type: str) -> str:
    """Purchase rows are Checks, CC charges, or cash expenses. QBO opens each
    on a different screen."""
    pt = (payment_type or "").strip().lower()
    if pt == "check":
        return "check"
    # CreditCard and Cash both open on the Expense screen in QBO.
    return "expense"


def txn_link(entity: str, txn: dict) -> str:
    """Build a click-through QBO URL for a transaction object."""
    txn_id = txn.get("Id", "")
    if entity == "Bill":
        path = "bill"
    elif entity == "VendorCredit":
        path = "vendorcredit"
    elif entity == "JournalEntry":
        path = "journal"
    elif entity == "Invoice":
        path = "invoice"
    elif entity == "Purchase":
        path = _purchase_link_path(txn.get("PaymentType", ""))
    else:
        path = "homepage"
    return f"{QBO_BASE}/{path}?txnId={txn_id}"


def txn_type_label(entity: str, txn: dict) -> str:
    """Human label for the Txn Type column."""
    if entity == "Purchase":
        pt = (txn.get("PaymentType") or "").strip()
        if pt.lower() == "check":
            return "Check"
        if pt.lower() == "creditcard":
            return "CC Charge"
        if pt.lower() == "cash":
            return "Cash Expense"
        return "Expense"
    return {
        "Bill": "Bill",
        "VendorCredit": "Vendor Credit",
        "JournalEntry": "Journal Entry",
        "Invoice": "Invoice",
    }.get(entity, entity)


def txn_vendor(entity: str, txn: dict, vendor_map: Dict[str, str]) -> str:
    """Pull the vendor/payee name regardless of entity shape."""
    ref = {}
    if entity in ("Bill", "VendorCredit"):
        ref = txn.get("VendorRef") or {}
    elif entity == "Purchase":
        ref = txn.get("EntityRef") or {}      # payee (vendor) on a Purchase
    elif entity == "Invoice":
        ref = txn.get("CustomerRef") or {}
    # JournalEntry has no header vendor.
    rid = ref.get("value", "")
    return vendor_map.get(rid, ref.get("name", "") or "")


def txn_doc(txn: dict) -> str:
    return (txn.get("DocNumber") or "").strip()


def get_line_class(line: dict, txn: dict) -> str:
    """Class for a line. Prefer the line-level ClassRef (a txn can split across
    divisions); fall back to the header ClassRef on older entries."""
    dt_type = line.get("DetailType", "")
    detail = line.get(dt_type) or {}
    line_class = (detail.get("ClassRef") or {}).get("name", "")
    if line_class:
        return line_class
    return (txn.get("ClassRef") or {}).get("name", "") or ""


def get_je_line_customer_ref(line: dict) -> dict:
    """JournalEntry lines hold Entity/Customer differently."""
    jed = line.get("JournalEntryLineDetail") or {}
    ent = jed.get("Entity") or {}
    if (ent.get("Type") or "").lower() == "customer":
        return ent.get("EntityRef") or {}
    return {}


def get_invoice_line_customer_ref(txn: dict) -> dict:
    """Invoices code the customer at the header, not per line."""
    return txn.get("CustomerRef") or {}


def iter_expense_lines(entity: str, txn: dict):
    """Yield (line, customer_ref, class_name) for each codeable line on a txn,
    normalizing the differences between entity shapes."""
    for line in (txn.get("Line") or []):
        dt_type = line.get("DetailType", "")
        if entity in ("Bill", "Purchase", "VendorCredit"):
            if dt_type not in EXPENSE_DETAIL_TYPES:
                continue
            cust = get_line_customer_ref(line)
            cls = get_line_class(line, txn)
        elif entity == "JournalEntry":
            if dt_type != "JournalEntryLineDetail":
                continue
            cust = get_je_line_customer_ref(line)
            jed = line.get("JournalEntryLineDetail") or {}
            cls = (jed.get("ClassRef") or {}).get("name", "") or \
                  (txn.get("ClassRef") or {}).get("name", "") or ""
        elif entity == "Invoice":
            if dt_type not in ("SalesItemLineDetail", "ItemBasedExpenseLineDetail"):
                continue
            cust = get_invoice_line_customer_ref(txn)
            sid = line.get("SalesItemLineDetail") or {}
            cls = (sid.get("ClassRef") or {}).get("name", "") or \
                  (txn.get("ClassRef") or {}).get("name", "") or ""
        else:
            continue
        yield line, cust, cls


# ───────────────────────── the two checks ─────────────────────────

def build_term_re(term: str) -> re.Pattern:
    """Case-insensitive search for the job/term as a whole token. The trailing
    boundary stops MFD281 from matching MFD2810, but still matches the MFD281
    inside 'MFD281-FTW' (a hyphen is a boundary) so those surface for review."""
    return re.compile(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", re.IGNORECASE)


def check_desc_mismatch(
    entity: str, txn: dict, job: str, desc_re: re.Pattern,
    vendor_map: Dict[str, str],
) -> List[dict]:
    """CHECK A. Lines whose description contains the job but whose
    Customer/Project is not that exact job."""
    rows: List[dict] = []
    job_u = job.upper()
    for line, cust, cls in iter_expense_lines(entity, txn):
        desc = line.get("Description", "") or ""
        if not desc_re.search(desc):
            continue
        cust_name = cust.get("name", "") or ""
        detected = get_project_num(cust_name)            # strict, respects -FTW
        if detected and detected.upper() == job_u:
            continue                                     # correctly coded → skip
        rows.append(_row(
            entity, txn, line, cust_name, detected, cls, vendor_map,
            issue=("No Customer/Project — desc says " + job
                   if not detected
                   else f"Coded to {detected}, not {job}"),
        ))
    return rows


def check_memo_scan(
    entity: str, txn: dict, parents_lower: set,
    vendor_map: Dict[str, str],
) -> List[dict]:
    """CHECK B. For a txn whose memo mentions the job, flag any line missing a
    Customer/Project or carrying a non-parent class."""
    rows: List[dict] = []
    for line, cust, cls in iter_expense_lines(entity, txn):
        cust_name = cust.get("name", "") or ""
        detected = get_project_num(cust_name)
        has_project = bool(cust_name.strip())
        class_ok = cls.strip().lower() in parents_lower

        if has_project and class_ok:
            continue                                     # both fine → skip

        problems: List[str] = []
        if not has_project:
            problems.append("No Customer/Project")
        if not class_ok:
            problems.append(
                f"Class '{cls}' not a division parent" if cls.strip()
                else "Class is blank"
            )
        rows.append(_row(
            entity, txn, line, cust_name, detected, cls, vendor_map,
            issue="; ".join(problems),
        ))
    return rows


def _row(
    entity: str, txn: dict, line: dict, cust_name: str,
    detected: Optional[str], cls: str, vendor_map: Dict[str, str],
    issue: str,
) -> dict:
    """Assemble one output row."""
    return {
        "txn_type": txn_type_label(entity, txn),
        "txn_date": parse_date(txn.get("TxnDate")),
        "vendor": txn_vendor(entity, txn, vendor_map),
        "doc": txn_doc(txn),
        "line_no": line.get("LineNum") or line.get("Id") or "",
        "line_desc": line.get("Description", "") or "",
        "line_amount": float(line.get("Amount") or 0),
        "customer_project": cust_name,
        "detected_project": detected or "",
        "class_name": cls,
        "issue": issue,
        "link": txn_link(entity, txn),
    }


# ───────────────────────── excel output ─────────────────────────

# Plain black-on-white, no fills (binding house style). Label + amount share a
# row, no hidden rows. Header is bold only.
SHEET_COLUMNS = [
    ("Txn Type", 14),
    ("Txn Date", 12),
    ("Vendor / Payee", 28),
    ("Ref / Doc #", 14),
    ("Line #", 8),
    ("Line Description", 40),
    ("Line Amount", 14),
    ("Customer/Project (as coded)", 34),
    ("Detected Project", 16),
    ("Class (as coded)", 22),
    ("Issue", 40),
    ("Open in QBO", 12),
]


def _write_sheet(ws, rows: List[dict], empty_msg: str) -> None:
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top")

    for ci, (name, width) in enumerate(SHEET_COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = bold
        c.alignment = left
        ws.column_dimensions[get_column_letter(ci)].width = width

    if not rows:
        ws.cell(row=2, column=1, value=empty_msg).alignment = left
        ws.freeze_panes = "A2"
        return

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["txn_type"]).alignment = left
        d = row["txn_date"]
        ws.cell(row=r, column=2,
                value=(d.strftime("%m/%d/%Y") if d else "")).alignment = left
        ws.cell(row=r, column=3, value=row["vendor"]).alignment = left
        ws.cell(row=r, column=4, value=row["doc"]).alignment = left
        ws.cell(row=r, column=5, value=str(row["line_no"])).alignment = left
        ws.cell(row=r, column=6, value=row["line_desc"]).alignment = left
        amt = ws.cell(row=r, column=7, value=row["line_amount"])
        amt.number_format = '#,##0.00'
        amt.alignment = right
        ws.cell(row=r, column=8, value=row["customer_project"]).alignment = left
        ws.cell(row=r, column=9, value=row["detected_project"]).alignment = left
        ws.cell(row=r, column=10, value=row["class_name"]).alignment = left
        ws.cell(row=r, column=11, value=row["issue"]).alignment = left
        link = ws.cell(row=r, column=12, value="↗ open")
        link.hyperlink = row["link"]
        link.alignment = left
        r += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET_COLUMNS))}{r - 1}"


def write_workbook(
    path: Path, desc_rows: List[dict], memo_rows: List[dict],
    run_checks: List[str], job: str,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    if "desc" in run_checks:
        _write_sheet(
            wb.create_sheet("Desc Mismatch"), desc_rows,
            f"No lines found whose description mentions {job} but is coded to a "
            f"different / missing project. ✓",
        )
    if "memo" in run_checks:
        _write_sheet(
            wb.create_sheet("Memo Scan"), memo_rows,
            f"Every line on transactions whose memo mentions {job} already has "
            f"a Customer/Project and a division-parent class. ✓",
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ───────────────────────── orchestration ─────────────────────────

def fetch_vendor_map(access: str, cid: str) -> Dict[str, str]:
    vm: Dict[str, str] = {}
    for v in query_all(access, cid, "Vendor", verbose=False):
        vm[v["Id"]] = v.get("DisplayName") or v.get("CompanyName") or v.get("Id")
    return vm


def memo_of(txn: dict) -> str:
    return txn.get("PrivateNote", "") or ""


def run(args) -> int:
    job = args.job.strip()
    if not job:
        print("✗ --job is required")
        return 2

    types = [t.strip().lower() for t in args.types.split(",") if t.strip()]
    bad = [t for t in types if t not in TYPE_ENTITY]
    if bad:
        print(f"✗ unknown --types: {', '.join(bad)} (valid: {', '.join(ALL_TYPES)})")
        return 2

    run_checks = [c.strip().lower() for c in args.checks.split(",") if c.strip()]
    bad_c = [c for c in run_checks if c not in ("desc", "memo")]
    if bad_c:
        print(f"✗ unknown --checks: {', '.join(bad_c)} (valid: desc, memo)")
        return 2

    parents = [p.strip() for p in args.parents.split(",") if p.strip()]
    parents_lower = {p.lower() for p in parents}

    desc_term = (args.desc_term or job).strip()
    memo_term = (args.memo_term or job).strip()
    desc_re = build_term_re(desc_term)
    memo_re = build_term_re(memo_term)

    since = parse_date(args.since) if args.since else None
    until = parse_date(args.until) if args.until else None
    if args.since and not since:
        print(f"✗ bad --since date: {args.since!r} (use YYYY-MM-DD)")
        return 2
    if args.until and not until:
        print(f"✗ bad --until date: {args.until!r} (use YYYY-MM-DD)")
        return 2

    # Server-side date filter. QBO can ONLY filter on top-level fields (TxnDate
    # here) — it cannot search line descriptions, line Customer/Project, line
    # Class, or memo text, so those are scanned client-side. Pushing the date
    # window into the WHERE clause is the one lever that cuts the row count at
    # the source (e.g. 95k bills → just the job's active window).
    where_clauses: List[str] = []
    if since:
        where_clauses.append(f"TxnDate >= '{since.isoformat()}'")
    if until:
        where_clauses.append(f"TxnDate <= '{until.isoformat()}'")
    where = " AND ".join(where_clauses)

    print(f"▶ Job coding audit for {job}")
    print(f"  types   : {', '.join(types)}")
    print(f"  checks  : {', '.join(run_checks)}")
    print(f"  parents : {', '.join(parents)}")
    print(f"  window  : {since.isoformat() if since else 'beginning'} → "
          f"{until.isoformat() if until else 'today'}")
    if not where:
        print("  ⚠ no --since given → pulling FULL history (slow). Tip: pass "
              "--since YYYY-MM-DD for when this job started.")

    access, cid = load_credentials()
    vendor_map = fetch_vendor_map(access, cid)

    # De-dupe entities (check + cc both → Purchase) so we fetch each once.
    entities_needed = sorted({TYPE_ENTITY[t] for t in types})
    txns_by_entity: Dict[str, List[dict]] = {}
    for ent in entities_needed:
        print(f"  fetching {ent}"
              f"{(' since ' + since.isoformat()) if since else ''} …", flush=True)
        rows = query_all(access, cid, ent, where=where, verbose=False)
        txns_by_entity[ent] = rows
        print(f"    {len(rows)} {ent}")

    # For Purchase we may want only checks OR only cc — filter by PaymentType
    # so the user's --types choice is honored.
    want_check = "check" in types
    want_cc = "cc" in types

    def purchase_wanted(txn: dict) -> bool:
        pt = (txn.get("PaymentType") or "").strip().lower()
        if pt == "check":
            return want_check
        # CreditCard / Cash / anything else → the "cc" bucket
        return want_cc

    desc_rows: List[dict] = []
    memo_rows: List[dict] = []

    for ent in entities_needed:
        for txn in txns_by_entity[ent]:
            if ent == "Purchase" and not purchase_wanted(txn):
                continue
            if "desc" in run_checks:
                desc_rows.extend(
                    check_desc_mismatch(ent, txn, job, desc_re, vendor_map)
                )
            if "memo" in run_checks:
                if memo_re.search(memo_of(txn)):
                    memo_rows.extend(
                        check_memo_scan(ent, txn, parents_lower, vendor_map)
                    )

    # Stable sort: newest first, then vendor.
    def _sort_key(r: dict):
        return (-(r["txn_date"].toordinal() if r["txn_date"] else 0), r["vendor"])
    desc_rows.sort(key=_sort_key)
    memo_rows.sort(key=_sort_key)

    print()
    if "desc" in run_checks:
        print(f"  CHECK A (desc mismatch): {len(desc_rows)} line(s) flagged")
    if "memo" in run_checks:
        print(f"  CHECK B (memo scan)    : {len(memo_rows)} line(s) flagged")

    if args.dry_run:
        print("\n(dry run — no file written)")
        return 0

    out = Path(args.out).expanduser() if args.out else \
        Path(__file__).resolve().parent / f"Job_Coding_Audit_{job}.xlsx"
    write_workbook(out, desc_rows, memo_rows, run_checks, job)
    print(f"\n✓ wrote {out}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Catch QBO coding mistakes for a given job code.")
    p.add_argument("--job", required=True,
                   help="Job code to audit, e.g. MFD281, CP142, RP7234")
    p.add_argument("--types", default=",".join(DEFAULT_TYPES),
                   help=f"Comma list from {ALL_TYPES}. Default vendor-cost txns.")
    p.add_argument("--checks", default="desc,memo",
                   help="Which checks to run: desc, memo, or both (default).")
    p.add_argument("--parents", default=",".join(DEFAULT_PARENT_CLASSES),
                   help="Acceptable division-parent class names (comma list).")
    p.add_argument("--desc-term", default="",
                   help="Override the text searched in line descriptions "
                        "(default = the job code).")
    p.add_argument("--memo-term", default="",
                   help="Override the text searched in memos "
                        "(default = the job code).")
    p.add_argument("--since", default="",
                   help="Only scan txns on/after this date (YYYY-MM-DD). "
                        "Filtered server-side — the main speed lever.")
    p.add_argument("--until", default="",
                   help="Only scan txns on/before this date (YYYY-MM-DD).")
    p.add_argument("--out", default="",
                   help="Output .xlsx path (default beside this script).")
    p.add_argument("--dry-run", action="store_true",
                   help="Fetch + report counts, but write no file.")
    return p


def main() -> int:
    args = build_parser().parse_args()
    try:
        return run(args)
    except KeyboardInterrupt:
        print("\n✗ interrupted")
        return 130


if __name__ == "__main__":
    sys.exit(main())
