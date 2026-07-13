#!/usr/bin/env python3
"""
sub_bill_audit.py — SUB bills missing a Customer/Project #.

Finds AP Bills whose memo (PrivateNote) STARTS with "Sub" and flags every
expense line that has no project # in its Customer/Project. Default window is
the current calendar year. Reuses qbo_bill_tracker (auth/paging/parsing) and
job_coding_audit (line iteration, QBO links, vendor map). Read-only.

USAGE
  python3 sub_bill_audit.py                      # this year to date
  python3 sub_bill_audit.py 2026-01-01           # since a date
  python3 sub_bill_audit.py 2026-01-01 2026-06-30
  python3 sub_bill_audit.py --dry-run
"""
from __future__ import annotations
import argparse
import datetime as dt
import re
import sys
from pathlib import Path

# Repo root on sys.path for shared/ (explicit — do NOT rely on
# qbo_bill_tracker's insert having run first).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from qbo_bill_tracker import load_credentials, query_all, parse_date, get_project_num
from job_coding_audit import iter_expense_lines, txn_link, txn_vendor, fetch_vendor_map
from shared import paths

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip3 install --break-system-packages openpyxl")
    sys.exit(1)

SUB_RE = re.compile(r"^\s*sub\b", re.IGNORECASE)

OUTDIR = paths.get_path(
    "ACB_AUDIT_OUT_DIR",
    paths.onedrive_base() / "Works In Progress" / "QBO Audits",
)

COLUMNS = [
    ("Bill Date", 12), ("Vendor", 28), ("Bill Ref #", 14),
    ("Bill Total", 13), ("Bill Open Bal", 14), ("Memo", 42),
    ("Line Description", 40), ("Line Amount", 13),
    ("Customer/Project (as coded)", 32), ("Issue", 34), ("Open in QBO", 12),
]


def is_sub(bill: dict) -> bool:
    return bool(SUB_RE.search(bill.get("PrivateNote") or ""))


def build_rows(bills, vendor_map):
    rows = []
    for b in bills:
        if not is_sub(b):
            continue
        memo = (b.get("PrivateNote") or "").strip()
        bill_date = parse_date(b.get("TxnDate"))
        vendor = txn_vendor("Bill", b, vendor_map)
        doc = (b.get("DocNumber") or "").strip()
        total = float(b.get("TotalAmt") or 0)
        bal = float(b.get("Balance") or 0)
        link = txn_link("Bill", b)
        for line, cust, cls in iter_expense_lines("Bill", b):
            name = (cust.get("name") or "").strip()
            if get_project_num(name):
                continue
            issue = ("No Customer/Project" if not name
                     else "No project # (parent only: " + name + ")")
            rows.append({
                "bill_date": bill_date, "vendor": vendor, "doc": doc,
                "total": total, "bal": bal, "memo": memo,
                "line_desc": (line.get("Description") or ""),
                "line_amount": float(line.get("Amount") or 0),
                "customer": name, "issue": issue, "link": link,
            })
    rows.sort(key=lambda r: (-(r["bill_date"].toordinal() if r["bill_date"] else 0),
                             r["vendor"]))
    return rows


def write_xlsx(path, rows, since, until):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sub Bills - No Project"
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top")
    for ci, (name, w) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = bold
        c.alignment = left
        ws.column_dimensions[get_column_letter(ci)].width = w
    if not rows:
        ws.cell(row=2, column=1,
                value="No sub bills missing a project # in this window.").alignment = left
        ws.freeze_panes = "A2"
    else:
        r = 2
        for row in rows:
            d = row["bill_date"]
            ws.cell(row=r, column=1, value=d.strftime("%m/%d/%Y") if d else "").alignment = left
            ws.cell(row=r, column=2, value=row["vendor"]).alignment = left
            ws.cell(row=r, column=3, value=row["doc"]).alignment = left
            for col, key in ((4, "total"), (5, "bal"), (8, "line_amount")):
                c = ws.cell(row=r, column=col, value=row[key])
                c.number_format = '#,##0.00'
                c.alignment = right
            ws.cell(row=r, column=6, value=row["memo"]).alignment = left
            ws.cell(row=r, column=7, value=row["line_desc"]).alignment = left
            ws.cell(row=r, column=9, value=row["customer"]).alignment = left
            ws.cell(row=r, column=10, value=row["issue"]).alignment = left
            lk = ws.cell(row=r, column=11, value="open")
            lk.hyperlink = row["link"]
            lk.alignment = left
            r += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:" + get_column_letter(len(COLUMNS)) + str(r - 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run(a) -> int:
    today = dt.date.today()
    since = parse_date(a.since) if a.since else today.replace(month=1, day=1)
    if a.since and not since:
        print("bad start date: " + repr(a.since) + " (YYYY-MM-DD)")
        return 2
    until = parse_date(a.until) if a.until else None
    if a.until and not until:
        print("bad end date: " + repr(a.until) + " (YYYY-MM-DD)")
        return 2
    where = "TxnDate >= '" + since.isoformat() + "'"
    if until:
        where += " AND TxnDate <= '" + until.isoformat() + "'"
    print("Sub bills missing a project #  (" + str(since) + " -> "
          + (str(until) if until else "today") + ")")
    access, cid = load_credentials()
    vendor_map = fetch_vendor_map(access, cid)
    print("  fetching Bills ...", flush=True)
    bills = query_all(access, cid, "Bill", where=where, verbose=False)
    print("    " + str(len(bills)) + " bills in window")
    rows = build_rows(bills, vendor_map)
    n_bills = len({(r["doc"], r["vendor"]) for r in rows})
    print("  " + str(len(rows)) + " uncoded line(s) across "
          + str(n_bills) + " sub bill(s) flagged")
    if a.dry_run:
        print("(dry run - no file written)")
        return 0
    out = (Path(a.out).expanduser() if a.out
           else OUTDIR / ("Sub_Bills_Missing_Project_" + str(since.year) + ".xlsx"))
    write_xlsx(out, rows, since, until)
    print("wrote " + str(out))
    return 0


def main() -> int:
    p = argparse.ArgumentParser(
        description="Audit sub bills (memo starts 'Sub') missing a project #.")
    p.add_argument("since", nargs="?", default="",
                   help="Start date YYYY-MM-DD (default Jan 1 this year).")
    p.add_argument("until", nargs="?", default="",
                   help="End date YYYY-MM-DD (optional).")
    p.add_argument("--out", default="", help="Output .xlsx (default OneDrive QBO Audits).")
    p.add_argument("--dry-run", action="store_true")
    return run(p.parse_args())


if __name__ == "__main__":
    sys.exit(main())
