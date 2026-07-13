#!/usr/bin/env python3
"""
item_no_project_audit.py — Bill ITEM lines (not Category/Account) missing a
project #. Item lines are job costs and must carry a Customer/Project.

Scans all Bills (optionally a date window), flags every item-based line whose
Customer/Project has no project #. Read-only. Reuses qbo_bill_tracker +
job_coding_audit helpers.

USAGE
  python3 item_no_project_audit.py                 # all bills, all time
  python3 item_no_project_audit.py 2026-01-01      # since a date
  python3 item_no_project_audit.py 2026-01-01 2026-06-30
  python3 item_no_project_audit.py --dry-run
"""
from __future__ import annotations
import argparse
import datetime as dt
import sys
from pathlib import Path

# Repo root on sys.path for shared/ (explicit — do NOT rely on
# qbo_bill_tracker's insert having run first).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from qbo_bill_tracker import load_credentials, query_all, parse_date, get_project_num
from job_coding_audit import iter_expense_lines, txn_link, txn_vendor, fetch_vendor_map
from shared import paths

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip3 install --break-system-packages openpyxl")
    sys.exit(1)

OUTDIR = paths.get_path(
    "ACB_AUDIT_OUT_DIR",
    paths.onedrive_base() / "Works In Progress" / "QBO Audits",
)

COLUMNS = [
    ("Bill Date", 12), ("Vendor", 28), ("Bill Ref #", 14), ("Bill Total", 13),
    ("Bill Open Bal", 14), ("Item", 26), ("Line Description", 38),
    ("Line Amount", 13), ("Customer/Project (as coded)", 32),
    ("Issue", 30), ("Open in QBO", 12),
]


def item_name(line):
    d = line.get("ItemBasedExpenseLineDetail") or {}
    return (d.get("ItemRef") or {}).get("name", "") or ""


def build_rows(bills, vmap):
    rows = []
    for b in bills:
        bd = parse_date(b.get("TxnDate"))
        vendor = txn_vendor("Bill", b, vmap)
        doc = (b.get("DocNumber") or "").strip()
        total = float(b.get("TotalAmt") or 0)
        bal = float(b.get("Balance") or 0)
        link = txn_link("Bill", b)
        for line, cust, cls in iter_expense_lines("Bill", b):
            if line.get("DetailType") != "ItemBasedExpenseLineDetail":
                continue
            name = (cust.get("name") or "").strip()
            if get_project_num(name):
                continue
            issue = ("No Customer/Project" if not name
                     else "No project # (parent only: " + name + ")")
            rows.append({
                "bill_date": bd, "vendor": vendor, "doc": doc, "total": total,
                "bal": bal, "item": item_name(line),
                "line_desc": (line.get("Description") or ""),
                "line_amount": float(line.get("Amount") or 0),
                "customer": name, "issue": issue, "link": link,
            })
    rows.sort(key=lambda r: (-(r["bill_date"].toordinal() if r["bill_date"] else 0),
                             r["vendor"]))
    return rows


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Lines - No Project"
    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top")
    for ci, (n, w) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=n)
        c.font = bold
        c.alignment = left
        ws.column_dimensions[get_column_letter(ci)].width = w
    if not rows:
        ws.cell(row=2, column=1,
                value="No item lines missing a project # in this window.").alignment = left
        ws.freeze_panes = "A2"
    else:
        r = 2
        for row in rows:
            d = row["bill_date"]
            ws.cell(row=r, column=1, value=d.strftime("%m/%d/%Y") if d else "").alignment = left
            ws.cell(row=r, column=2, value=row["vendor"]).alignment = left
            ws.cell(row=r, column=3, value=row["doc"]).alignment = left
            for col, key in ((4, "total"), (5, "bal"), (8, "line_amount")):
                c = ws.cell(row=r, column=col, value=row[key])
                c.number_format = '#,##0.00'
                c.alignment = right
            ws.cell(row=r, column=6, value=row["item"]).alignment = left
            ws.cell(row=r, column=7, value=row["line_desc"]).alignment = left
            ws.cell(row=r, column=9, value=row["customer"]).alignment = left
            ws.cell(row=r, column=10, value=row["issue"]).alignment = left
            lk = ws.cell(row=r, column=11, value="open")
            lk.hyperlink = row["link"]
            lk.alignment = left
            r += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:" + get_column_letter(len(COLUMNS)) + str(r - 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run(a):
    since = parse_date(a.since) if a.since else None
    if a.since and not since:
        print("bad start date: " + repr(a.since))
        return 2
    until = parse_date(a.until) if a.until else None
    if a.until and not until:
        print("bad end date: " + repr(a.until))
        return 2
    where = ""
    if since:
        where = "TxnDate >= '" + since.isoformat() + "'"
    if until:
        where = (where + " AND " if where else "") + "TxnDate <= '" + until.isoformat() + "'"
    scope = (str(since) if since else "all time") + " -> " + (str(until) if until else "today")
    print("Item lines missing a project #  (" + scope + ")")
    if not where:
        print("  (no date filter - pulling all bills, this can take a bit)")
    access, cid = load_credentials()
    vmap = fetch_vendor_map(access, cid)
    print("  fetching Bills ...", flush=True)
    bills = query_all(access, cid, "Bill", where=where, verbose=True)
    print("    " + str(len(bills)) + " bills")
    rows = build_rows(bills, vmap)
    nb = len({(r["doc"], r["vendor"]) for r in rows})
    print("  " + str(len(rows)) + " item line(s) across " + str(nb)
          + " bill(s) missing a project #")
    if a.dry_run:
        print("(dry run - no file written)")
        return 0
    suffix = ("_" + str(since.year) if since else "")
    out = (Path(a.out).expanduser() if a.out
           else OUTDIR / ("Item_Lines_Missing_Project" + suffix + ".xlsx"))
    write_xlsx(out, rows)
    print("wrote " + str(out))
    return 0


def main():
    p = argparse.ArgumentParser(
        description="Audit Bill ITEM lines missing a project #.")
    p.add_argument("since", nargs="?", default="", help="Start date YYYY-MM-DD (default all time).")
    p.add_argument("until", nargs="?", default="", help="End date YYYY-MM-DD (optional).")
    p.add_argument("--out", default="")
    p.add_argument("--dry-run", action="store_true")
    return run(p.parse_args())


if __name__ == "__main__":
    sys.exit(main())
