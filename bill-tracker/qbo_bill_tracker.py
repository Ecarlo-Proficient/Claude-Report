#!/usr/bin/env python3
"""
qbo_bill_tracker.py — Bill Payment Tracker for Proficient Concrete

Pulls open AP bills from QBO, matches each line to the corresponding GC
invoice, classifies status by division (RP/MFD/CP), and writes a 4-sheet
Excel report.  Notes and status overrides are preserved across runs by
stable QBO key (Bill.Id + Line.Id), so they survive bill state changes
(e.g., paid → bounced check → unpaid).

USAGE
  python3 qbo_bill_tracker.py                    # standard run
  python3 qbo_bill_tracker.py --dry-run          # build but do not write
  python3 qbo_bill_tracker.py --out /path/x.xlsx # override output path

OUTPUT
  Bill_Payment_Tracker.xlsx with four sheets:
    1. Bill Tracker  — open bills (Balance > 0), excluding subs, editable
    2. By Vendor     — same data grouped, locked
    3. By Project    — same data grouped, locked
    4. Archive       — paid bills with notes/overrides preserved, locked
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import re
import sys
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Allow import of qbo_vault from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import requests
except ImportError:
    print("✗ pip3 install --break-system-packages requests")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("✗ pip3 install --break-system-packages openpyxl")
    sys.exit(1)

from shared import qbo_vault as kc
from shared import paths

# ───────────────────────── constants ─────────────────────────

API_BASE = "https://quickbooks.api.intuit.com"
MINOR_VERSION = "70"

PROJECT_NUM_RE = re.compile(r"\b((?:MFD|CP|RP)\d+(?:-FTW)?)\b", re.IGNORECASE)
DRAW_PERIOD_RE = re.compile(
    r"\(\s*Period\s*:\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*\)",
    re.IGNORECASE,
)
RETAINAGE_NOT_BILLED_RE = re.compile(r"retainage\s+not\s+billed", re.IGNORECASE)
RETAINAGE_RE = re.compile(r"\bretainage\b", re.IGNORECASE)
SUB_RE = re.compile(r"\bsub\b", re.IGNORECASE)

# Status values — plain-English pipeline state. Ted 2026-06-03: re-split
# AWAITING into PAYMENT vs INVOICE because lumping them lost the signal:
#   - "Awaiting Invoice" = we owe the bill, GC hasn't been billed yet
#   - "Awaiting Payment" = we owe the bill, GC was billed but hasn't paid us
# Approval status now lives in its own column (Approved), not suffixed onto
# Status. display_status() returns the pipeline label only.
STATUS_OK_TO_PAY = "Invoice paid"
STATUS_AWAITING_PAYMENT = "Awaiting Payment"
STATUS_AWAITING_INVOICE = "Awaiting Invoice"
STATUS_NO_PROJECT = "No project #"
STATUS_PAID = "Bill paid"
# 2026-06-04: aggregated bill-level status only emitted by collapse_rows
# when a multi-project bill has SOME lines invoice-paid + others not. Each
# line on a multi-project bill matches its own project's invoice, so the
# "is it fully funded" question is line-aggregated. Partial paid signals
# "some lines funded, AP can decide to float the remainder or wait."
STATUS_PARTIAL_PAID = "Partial paid"

# 2026-07-13: the single Status split into two axes (Ted). PAY STATUS is the AP
# side (did WE pay the vendor) from the bill balance; INVOICE STATUS is the AR
# side (did the GC fund us) computed INDEPENDENTLY of payment, so a paid bill
# still shows whether we've been reimbursed ("fronted").
STATUS_UNPAID = "Unpaid"          # pay status: bill balance == full total


def compute_pay_status(bill: dict) -> str:
    """AP side — did we pay the vendor. From the bill balance alone."""
    total = float(bill.get("TotalAmt") or 0)
    bal = float(bill.get("Balance") or 0)
    if bal == 0:
        return STATUS_PAID              # "Bill paid"
    if 0 < bal < total:
        return STATUS_PARTIAL_PAID      # "Partial paid"
    return STATUS_UNPAID                # "Unpaid"


def compute_invoice_status(matched: Optional[dict], division: Optional[str]) -> str:
    """AR side — has the GC funded this draw. Independent of the bill balance,
    so it stays visible even after we've paid the vendor."""
    if not division:
        return STATUS_NO_PROJECT
    if matched is None:
        return STATUS_AWAITING_INVOICE
    if float(matched.get("Balance") or 0) == 0:
        return STATUS_OK_TO_PAY         # "Invoice paid"
    return STATUS_AWAITING_PAYMENT
OVERRIDE_VALUES = [STATUS_OK_TO_PAY, "ON HOLD", "REVIEW", "DISPUTED"]

# Color fills
COLOR_OK_TO_PAY = "C6EFCE"           # green
COLOR_AWAITING_PAYMENT = "FFEB9C"    # yellow
COLOR_AWAITING_INVOICE = "FFF8DC"    # beige
COLOR_NO_PROJECT = "FFC7CE"          # red
COLOR_PAID = "D9D9D9"                # grey
COLOR_ON_HOLD = "FFD580"             # orange
COLOR_REVIEW = "FFCCCC"              # light red
COLOR_DISPUTED = "F8696B"            # bright red
COLOR_HEADER = "1F4E78"              # dark blue
COLOR_HEADER_FONT = "FFFFFF"         # white
COLOR_SUBTOTAL = "E7E6E6"            # light grey

STATUS_FILL_MAP = {
    STATUS_OK_TO_PAY: COLOR_OK_TO_PAY,
    STATUS_AWAITING_PAYMENT: COLOR_AWAITING_PAYMENT,
    STATUS_AWAITING_INVOICE: COLOR_AWAITING_INVOICE,
    STATUS_NO_PROJECT: COLOR_NO_PROJECT,
    STATUS_PAID: COLOR_PAID,
    "ON HOLD": COLOR_ON_HOLD,
    "REVIEW": COLOR_REVIEW,
    "DISPUTED": COLOR_DISPUTED,
}

# This module is now a LIBRARY (excel_bill_sync.py imports its functions). The
# standalone path below is legacy. DEFAULT_OUTPUT points OUTSIDE the project
# folder so a stray standalone run can't drop real AP data into this
# AI-visible/synced folder.
DEFAULT_OUTPUT = paths.companyhealth_dir() / "_legacy_bill_payment_tracker.xlsx"

# Column headers in canonical order. (header, width)
COLUMNS = [
    ("Bill Date", 12),
    ("Vendor", 28),
    ("Bill Ref #", 14),
    ("Bill Total", 14),
    ("Bill Open Bal", 14),
    ("Customer/Project", 36),
    ("Division", 9),
    ("Line Amount", 14),
    ("Line Description", 36),
    ("Matched Invoice #", 16),
    ("Invoice Date", 12),
    ("Draw Period", 22),
    ("Invoice Total", 14),
    ("Invoice Open Bal", 16),
    ("Status (Auto)", 22),
    ("Status Override", 22),
    ("Status (Final)", 22),
    ("Notes", 40),
    ("_Key", 16),
]
COL_IDX = {name: i + 1 for i, (name, _) in enumerate(COLUMNS)}  # 1-indexed

KEY_COL = "_Key"
OVERRIDE_COL = "Status Override"
NOTES_COL = "Notes"
AUTO_COL = "Status (Auto)"
FINAL_COL = "Status (Final)"

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ───────────────────────── auth ─────────────────────────

def load_credentials() -> Tuple[str, str]:
    if not kc.has_credentials():
        print("✗ no credentials. Run: python3 setup_qbo.py")
        sys.exit(1)
    creds = kc.get_all()
    required = ["QBO_CLIENT_ID", "QBO_CLIENT_SECRET", "QBO_COMPANY_ID", "QBO_REFRESH_TOKEN"]
    if any(not creds.get(k) for k in required):
        print("✗ blob incomplete. Run: python3 setup_qbo.py")
        sys.exit(1)
    basic = base64.b64encode(
        f"{creds['QBO_CLIENT_ID']}:{creds['QBO_CLIENT_SECRET']}".encode()
    ).decode()
    r = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        headers={
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        data={"grant_type": "refresh_token", "refresh_token": creds["QBO_REFRESH_TOKEN"]},
        timeout=30,
    )
    if r.status_code != 200:
        print(f"✗ token refresh status={r.status_code} body={r.text[:300]}")
        sys.exit(1)
    body = r.json()
    new_rt = body.get("refresh_token")
    if new_rt and new_rt != creds["QBO_REFRESH_TOKEN"]:
        try:
            kc.put("QBO_REFRESH_TOKEN", new_rt)
        except kc.SecretsError:
            pass
    return body["access_token"], creds["QBO_COMPANY_ID"]


# ───────────────────────── api ─────────────────────────

def _api_get(access: str, path: str, params: Optional[dict] = None) -> dict:
    """GET with exponential-backoff retry on transient QBO errors.

    QBO is moody on big queries — `SystemFailureError` (500) and 429 rate
    limits come and go without any change on our side. Five retries with
    1s/2s/4s/8s/16s backoff catches almost all of these. 4xx (real client
    errors) and anything still failing after retries propagate as before.
    """
    import time
    p = dict(params or {})
    p["minorversion"] = MINOR_VERSION
    delays = [1, 2, 4, 8, 16]   # 5 retries, ~31s total worst case
    last_exc_text = ""
    for attempt in range(len(delays) + 1):
        try:
            r = requests.get(
                f"{API_BASE}{path}",
                headers={"Authorization": f"Bearer {access}", "Accept": "application/json"},
                params=p,
                timeout=60,
            )
        except requests.RequestException as e:
            # Network blip — retryable
            last_exc_text = f"network error: {e}"
            if attempt < len(delays):
                time.sleep(delays[attempt])
                print(f"  ⟳ retry {attempt+1}/{len(delays)} ({last_exc_text}) …", flush=True)
                continue
            raise RuntimeError(f"GET {path} → {last_exc_text} (after {len(delays)} retries)")
        if r.status_code == 200:
            return r.json()
        # 5xx and 429 are retryable; 4xx are not
        is_retryable = r.status_code >= 500 or r.status_code == 429
        if not is_retryable or attempt >= len(delays):
            raise RuntimeError(f"GET {path} → {r.status_code}: {r.text[:300]}")
        wait = delays[attempt]
        print(f"  ⟳ retry {attempt+1}/{len(delays)} after {wait}s "
              f"({r.status_code}: {r.text[:120]}) …", flush=True)
        time.sleep(wait)
    # Defensive fallthrough
    raise RuntimeError(f"GET {path} → exhausted retries")


def query(access: str, cid: str, q: str) -> dict:
    return _api_get(access, f"/v3/company/{cid}/query", {"query": q})


def query_all(access: str, cid: str, entity: str, where: str = "",
              verbose: bool = True) -> List[dict]:
    """Paginated SELECT * FROM <entity> [WHERE ...]. Logs page-by-page
    progress so the user can see the sync isn't stuck during long pulls."""
    out: List[dict] = []
    start = 1
    page = 500
    page_num = 0
    while True:
        page_num += 1
        q = f"SELECT * FROM {entity}"
        if where:
            q += f" WHERE {where}"
        q += f" STARTPOSITION {start} MAXRESULTS {page}"
        data = query(access, cid, q)
        batch = data.get("QueryResponse", {}).get(entity, [])
        if not batch:
            break
        out.extend(batch)
        if verbose and (page_num % 5 == 0 or len(batch) < page):
            # log every 5th page + the final partial page
            print(f"    page {page_num}: {len(out)} {entity} so far", flush=True)
        if len(batch) < page:
            break
        start += page
    return out


def fetch_bills_by_ids(access: str, cid: str, ids: set) -> List[dict]:
    """Fetch specific bills by Id list. Used to re-fetch newly-archived bills."""
    if not ids:
        return []
    BATCH = 30
    out: List[dict] = []
    ids_list = sorted(ids)
    for i in range(0, len(ids_list), BATCH):
        batch = ids_list[i: i + BATCH]
        in_clause = ",".join(f"'{x}'" for x in batch)
        try:
            data = query(access, cid, f"SELECT * FROM Bill WHERE Id IN ({in_clause})")
            out.extend(data.get("QueryResponse", {}).get("Bill", []))
        except Exception as e:
            print(f"  warn: failed to fetch bill batch {i // BATCH}: {e}")
    return out


# ───────────────────────── parsing & classification ─────────────────────────

def parse_date(s: Any) -> Optional[dt.date]:
    """Parse various date inputs (str, datetime, date) to date or None."""
    if s is None or s == "":
        return None
    if isinstance(s, dt.datetime):
        return s.date()
    if isinstance(s, dt.date):
        return s
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def get_project_num(name: str) -> Optional[str]:
    """Extract MFD###, CP###, or RP#### (with optional -FTW) from anywhere in a
    customer name. In QBO the Customer/Project field on a line is formatted
    'Parent Customer: Project # Project Name' — so we search the whole string,
    not just the prefix."""
    if not name:
        return None
    m = PROJECT_NUM_RE.search(name)
    return m.group(1).upper() if m else None


def get_division(project_num: Optional[str]) -> Optional[str]:
    if not project_num:
        return None
    p = project_num.upper()
    if p.startswith("MFD"):
        return "MFD"
    if p.startswith("CP"):
        return "CP"
    if p.startswith("RP"):
        return "RP"
    return None


def parse_draw_period(invoice: dict) -> Optional[Tuple[dt.date, dt.date]]:
    """Pull (start, end) from invoice PrivateNote like '(Period: MM/DD/YY - MM/DD/YY)'."""
    pn = invoice.get("PrivateNote", "") or ""
    m = DRAW_PERIOD_RE.search(pn)
    if not m:
        return None
    s = parse_date(m.group(1))
    e = parse_date(m.group(2))
    if s and e and s <= e:
        return (s, e)
    return None


def is_sub_bill(bill: dict) -> bool:
    """Bill memo (PrivateNote) contains the word 'sub' → exclude entirely."""
    memo = (bill.get("PrivateNote") or "")
    return SUB_RE.search(memo) is not None


def is_excluded_invoice(invoice: dict) -> bool:
    """
    True if invoice should be excluded from the matching pool.

    Rules:
      1) PrivateNote contains 'retainage not billed' (historical mistakes)
      2) Every non-empty line description contains 'retainage' (standalone
         retainage release / full retainage draw)
    """
    memo = (invoice.get("PrivateNote") or "")
    if RETAINAGE_NOT_BILLED_RE.search(memo):
        return True

    lines = invoice.get("Line") or []
    descs: List[str] = []
    for ln in lines:
        if ln.get("DetailType") in ("SubTotalLineDetail", "DiscountLineDetail"):
            continue
        d = (ln.get("Description") or "").strip()
        if d:
            descs.append(d)
    if descs and all(RETAINAGE_RE.search(d) for d in descs):
        return True
    return False


def get_line_customer_ref(line: dict) -> dict:
    """Return CustomerRef sub-object from a bill line (regardless of detail type)."""
    dt_type = line.get("DetailType", "")
    if dt_type == "AccountBasedExpenseLineDetail":
        return (line.get("AccountBasedExpenseLineDetail") or {}).get("CustomerRef") or {}
    if dt_type == "ItemBasedExpenseLineDetail":
        return (line.get("ItemBasedExpenseLineDetail") or {}).get("CustomerRef") or {}
    return {}


# ───────────────────────── matching ─────────────────────────

_PUMP_RE = re.compile(r"\bpump", re.IGNORECASE)

# RP amount-aware matching (Ted 2026-06-18). A concrete bill is our COST; the
# invoice is what we billed the GC for that scope, so the bill must not exceed
# the invoice. Among same-project invoices within this forward window, keep
# only those that COVER the bill amount and take the LARGEST; if none cover,
# leave unmatched (→ Awaiting Invoice) rather than pin it to a too-small one.
_RP_MATCH_FWD_DAYS = 60          # invoice TxnDate must be within 60 days AFTER the bill
_RP_COVER_TOLERANCE = 0.01       # 1-cent slack so an exact-equal invoice still covers


def _invoice_scope(inv: dict) -> str:
    """Return the last ' - '-delimited segment of an invoice's PrivateNote.
    RP invoices follow `PROJECT# - ADDRESS - SCOPE` (e.g. `RP7342 - 2112 Ten
    Mile Creek Court - Pump Charges`). Used to detect pump-truck invoices.
    """
    memo = (inv.get("PrivateNote") or "").strip()
    if not memo:
        return ""
    parts = memo.rsplit(" - ", 1)
    return parts[1].strip() if len(parts) == 2 else memo


def find_matching_invoice(
    bill_date: dt.date,
    division: str,
    customer_id: str,
    invoices_by_customer: Dict[str, List[dict]],
    bill_text: str = "",
    bill_amount: float = 0.0,
) -> Optional[dict]:
    """Apply division-specific matching rules. Return matched invoice or None.

    `bill_text` — bill description + account name concatenated, used by the
    RP branch to keep pump-truck bills from being matched to non-pump
    invoices and vice versa. RP projects can have several scoped invoices
    on the same house (Pump Charges, Foundation, Driveway, …); matching
    purely by date proximity puts pump bills on foundation invoices etc.
    Ted 2026-06-04: only the pump case has burned us; broader scope
    matching is deferred.

    `bill_amount` — the cost being attributed to this project (the line
    amount). RP matching is amount-aware (Ted 2026-06-18): the bill must not
    exceed the invoice it's matched to, so we only consider invoices that
    COVER the bill amount and pick the largest. See `_RP_MATCH_FWD_DAYS`.
    """
    candidates = invoices_by_customer.get(customer_id, [])
    if not candidates:
        return None

    if division == "RP":
        # Two-way pump filter: pump bill ↔ pump invoice; non-pump bill ↔
        # non-pump invoice. Only applied when there's any pump invoice for
        # this customer — otherwise we'd over-filter ourselves to no match.
        bill_is_pump = bool(_PUMP_RE.search(bill_text or ""))
        any_pump_inv = any(_PUMP_RE.search(_invoice_scope(inv)) for inv in candidates)
        if any_pump_inv:
            candidates = [
                inv for inv in candidates
                if bool(_PUMP_RE.search(_invoice_scope(inv))) == bill_is_pump
            ]
            if not candidates:
                return None

        # Date window: invoice on/after the bill, within the forward window.
        window_end = bill_date + dt.timedelta(days=_RP_MATCH_FWD_DAYS)
        eligible = []
        for inv in candidates:
            d = parse_date(inv.get("TxnDate"))
            if d is not None and bill_date <= d <= window_end:
                eligible.append(inv)
        if not eligible:
            return None

        # Amount-aware: a bill (our cost) must not exceed the invoice (what we
        # billed the GC for that scope). Keep only invoices that COVER the bill
        # amount; among those take the LARGEST. If none cover, leave unmatched
        # so it surfaces as Awaiting Invoice instead of pinning to a too-small
        # invoice (e.g. $2,703 concrete cost wrongly stuck on a $246 invoice).
        covering = [
            inv for inv in eligible
            if float(inv.get("TotalAmt") or 0) >= bill_amount - _RP_COVER_TOLERANCE
        ]
        if not covering:
            return None
        covering.sort(key=lambda inv: (
            -float(inv.get("TotalAmt") or 0),                             # largest first
            abs(((parse_date(inv.get("TxnDate")) or bill_date) - bill_date).days),  # then closest date
        ))
        return covering[0]

    if division in ("MFD", "CP"):
        for inv in candidates:
            period = parse_draw_period(inv)
            if not period:
                continue
            start, end = period
            if start <= bill_date <= end:
                return inv
        return None

    return None


def compute_status(bill: dict, matched: Optional[dict], division: Optional[str]) -> str:
    bill_balance = float(bill.get("Balance") or 0)
    if bill_balance == 0:
        return STATUS_PAID
    if not division:
        return STATUS_NO_PROJECT
    if matched is None:
        return STATUS_AWAITING_INVOICE
    inv_balance = float(matched.get("Balance") or 0)
    if inv_balance == 0:
        return STATUS_OK_TO_PAY
    return STATUS_AWAITING_PAYMENT


# ───────────────────────── existing-file IO ─────────────────────────

def read_user_edits(path: Path) -> Dict[str, Tuple[str, str]]:
    """Build {key → (override, notes)} from existing Bill Tracker + Archive sheets."""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(filename=path, data_only=True)
    except Exception as e:
        print(f"  ⚠ could not read existing file: {e}")
        return {}

    db: Dict[str, Tuple[str, str]] = {}
    for sheet in ("Bill Tracker", "Archive"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = {str(c.value): c.column for c in ws[1] if c.value}
        kc_idx = headers.get(KEY_COL)
        ov_idx = headers.get(OVERRIDE_COL)
        nt_idx = headers.get(NOTES_COL)
        if not kc_idx:
            continue
        for row in ws.iter_rows(min_row=2, values_only=False):
            kv = str(row[kc_idx - 1].value or "").strip()
            if not kv or kv.startswith("SUBTOTAL"):
                continue
            ov = str(row[ov_idx - 1].value or "").strip() if ov_idx else ""
            nt = str(row[nt_idx - 1].value or "").strip() if nt_idx else ""
            db[kv] = (ov, nt)
    return db


# ───────────────────────── row building ─────────────────────────

def build_rows(
    bills: List[dict],
    invoices_by_customer: Dict[str, List[dict]],
    vendor_map: Dict[str, str],
    notes_db: Dict[str, Tuple[str, str]],
) -> List[dict]:
    """Build canonical row dicts (one per bill line)."""
    rows: List[dict] = []
    for bill in bills:
        bill_id = bill.get("Id", "")
        bill_date = parse_date(bill.get("TxnDate")) or dt.date.today()
        bill_total = float(bill.get("TotalAmt") or 0)
        bill_balance = float(bill.get("Balance") or 0)
        bill_doc = bill.get("DocNumber", "") or ""
        v_ref = bill.get("VendorRef") or {}
        vendor_name = vendor_map.get(v_ref.get("value", ""), v_ref.get("name", "?"))

        for line in (bill.get("Line") or []):
            # Only Product/Service lines are job costs (each item carries a
            # cost code). Account/Category lines are overhead — utilities,
            # office supplies, software, etc. — and don't belong here.
            if line.get("DetailType") != "ItemBasedExpenseLineDetail":
                continue
            line_id = line.get("Id", "")
            key = f"{bill_id}-{line_id}"
            line_amt = float(line.get("Amount") or 0)
            line_desc = line.get("Description", "") or ""

            cust = get_line_customer_ref(line)
            cust_id = cust.get("value", "")
            cust_name = cust.get("name", "")
            project_num = get_project_num(cust_name)
            division = get_division(project_num)

            # Skip non-job-cost lines entirely (office expenses, utilities, etc.)
            # — they don't belong in a GC-invoice-matching report.
            if not division:
                continue

            matched: Optional[dict] = None
            if cust_id and division:
                matched = find_matching_invoice(
                    bill_date, division, cust_id, invoices_by_customer,
                    bill_text=line_desc,
                )

            inv_doc = ""
            inv_date: Optional[dt.date] = None
            inv_total = 0.0
            inv_balance = 0.0
            draw_period_str = ""
            if matched:
                inv_doc = matched.get("DocNumber", "") or ""
                inv_date = parse_date(matched.get("TxnDate"))
                inv_total = float(matched.get("TotalAmt") or 0)
                inv_balance = float(matched.get("Balance") or 0)
                period = parse_draw_period(matched)
                if period:
                    draw_period_str = f"{period[0].strftime('%m/%d/%y')} – {period[1].strftime('%m/%d/%y')}"

            auto_status = compute_status(bill, matched, division)
            ov, nt = notes_db.get(key, ("", ""))

            rows.append({
                "key": key,
                "bill_date": bill_date,
                "vendor": vendor_name,
                "bill_doc": bill_doc,
                "bill_total": bill_total,
                "bill_balance": bill_balance,
                "customer": cust_name,
                "division": division or "",
                "project_num": project_num or "",
                "line_amount": line_amt,
                "line_desc": line_desc,
                "inv_doc": inv_doc,
                "inv_date": inv_date,
                "draw_period": draw_period_str,
                "inv_total": inv_total,
                "inv_balance": inv_balance,
                "auto_status": auto_status,
                "override": ov,
                "notes": nt,
            })
    return rows


# ───────────────────────── excel writing ─────────────────────────

def _row_values_for_write(row: dict) -> List[Any]:
    """Convert row dict to ordered list matching COLUMNS, with None placeholder for the formula column."""
    return [
        row["bill_date"],
        row["vendor"],
        row["bill_doc"],
        row["bill_total"],
        row["bill_balance"],
        row["customer"],
        row["division"],
        row["line_amount"],
        row["line_desc"],
        row["inv_doc"],
        row["inv_date"],
        row["draw_period"],
        row["inv_total"],
        row["inv_balance"],
        row["auto_status"],
        row["override"],
        None,                  # Status (Final) — set as formula
        row["notes"],
        row["key"],
    ]


def _write_header(ws):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (name, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = fill
        c.font = font
        c.alignment = align
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36
    ws.column_dimensions[get_column_letter(COL_IDX[KEY_COL])].hidden = True


def _format_cell(c, col_name: str):
    """Apply alignment and number format based on column name."""
    c.border = THIN_BORDER
    if col_name in ("Bill Total", "Bill Open Bal", "Line Amount", "Invoice Total", "Invoice Open Bal"):
        c.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        c.alignment = Alignment(horizontal="right", vertical="center")
    elif col_name in ("Bill Date", "Invoice Date"):
        c.number_format = "mm/dd/yy"
        c.alignment = Alignment(horizontal="center", vertical="center")
    elif col_name in ("Division", "Bill Ref #", "Matched Invoice #",
                      "Status (Auto)", "Status Override", "Status (Final)"):
        c.alignment = Alignment(horizontal="center", vertical="center")
    elif col_name in ("Line Description", "Notes", "Customer/Project", "Vendor", "Draw Period"):
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    else:
        c.alignment = Alignment(horizontal="left", vertical="center")


def _write_data_row(ws, row_num: int, row: dict, editable: bool):
    """Write one data row with formula + color coding. `editable` unlocks Override + Notes cells."""
    values = _row_values_for_write(row)
    for i, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=i, value=val)
        col_name = COLUMNS[i - 1][0]
        _format_cell(c, col_name)
        if editable and col_name in (OVERRIDE_COL, NOTES_COL):
            c.protection = Protection(locked=False)

    # Status (Final) formula
    ovr_letter = get_column_letter(COL_IDX[OVERRIDE_COL])
    auto_letter = get_column_letter(COL_IDX[AUTO_COL])
    final_cell = ws.cell(row=row_num, column=COL_IDX[FINAL_COL])
    final_cell.value = f'=IF({ovr_letter}{row_num}="",{auto_letter}{row_num},{ovr_letter}{row_num})'
    _format_cell(final_cell, FINAL_COL)
    final_cell.font = Font(bold=True)

    # Color the three status cells based on resolved final value
    final_status = row["override"] if row["override"] else row["auto_status"]
    fill_color = STATUS_FILL_MAP.get(final_status, "")
    if fill_color:
        fill = PatternFill("solid", fgColor=fill_color)
        for col_name in (AUTO_COL, OVERRIDE_COL, FINAL_COL):
            ws.cell(row=row_num, column=COL_IDX[col_name]).fill = fill


def _add_override_dropdown(ws, last_row: int):
    if last_row < 2:
        return
    letter = get_column_letter(COL_IDX[OVERRIDE_COL])
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(OVERRIDE_VALUES)}"',
        allow_blank=True,
        showErrorMessage=False,  # let user type free text too
    )
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{last_row}")


def write_main_sheet(wb: Workbook, rows: List[dict]):
    ws = wb.create_sheet("Bill Tracker", 0)
    _write_header(ws)
    rows_sorted = sorted(rows, key=lambda r: (r["bill_date"], r["vendor"]), reverse=True)
    for i, row in enumerate(rows_sorted, start=2):
        _write_data_row(ws, i, row, editable=True)
    _add_override_dropdown(ws, len(rows_sorted) + 1)
    # Soft sheet protection: locked by default; Override + Notes cells unlocked above.
    # NOTE: do NOT set ws.protection.password = None — openpyxl's password setter
    # tries to iterate over the value to hash it, which TypeErrors on None.
    ws.protection.sheet = True


def _write_subtotal_row(ws, row_num: int, label: str, line_total: float, bal_total: float):
    fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
    bold = Font(bold=True)
    for ci in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=row_num, column=ci)
        c.fill = fill
        c.border = THIN_BORDER
    ws.cell(row=row_num, column=1, value=f"Subtotal: {label}").font = bold
    c = ws.cell(row=row_num, column=COL_IDX["Line Amount"], value=line_total)
    c.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    c.font = bold
    c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=row_num, column=COL_IDX["Bill Open Bal"], value=bal_total)
    c.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    c.font = bold
    c.alignment = Alignment(horizontal="right")
    # sentinel key so re-reads skip this row
    ws.cell(row=row_num, column=COL_IDX[KEY_COL], value=f"SUBTOTAL:{label}")


def write_grouped_sheet(
    wb: Workbook,
    rows: List[dict],
    title: str,
    group_key_fn,
    sort_fn,
):
    ws = wb.create_sheet(title)
    _write_header(ws)
    rows_sorted = sorted(rows, key=sort_fn)

    if not rows_sorted:
        ws.protection.sheet = True
        return

    cur = 2
    last_key = None
    group_buf: List[dict] = []

    def flush(key: Optional[str], buf: List[dict], start: int) -> int:
        if not buf:
            return start
        for r in buf:
            _write_data_row(ws, start, r, editable=False)
            ws.row_dimensions[start].outline_level = 1
            start += 1
        line_tot = sum(r["line_amount"] for r in buf)
        bal_tot = sum(r["bill_balance"] for r in buf)
        _write_subtotal_row(ws, start, str(key) if key is not None else "", line_tot, bal_tot)
        return start + 1

    for r in rows_sorted:
        gk = group_key_fn(r)
        if gk != last_key and group_buf:
            cur = flush(last_key, group_buf, cur)
            group_buf = []
        last_key = gk
        group_buf.append(r)
    cur = flush(last_key, group_buf, cur)

    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.protection.sheet = True


def write_archive_sheet(wb: Workbook, rows: List[dict]):
    ws = wb.create_sheet("Archive")
    _write_header(ws)
    rows_sorted = sorted(rows, key=lambda r: (r["bill_date"], r["vendor"]), reverse=True)
    for i, row in enumerate(rows_sorted, start=2):
        _write_data_row(ws, i, row, editable=False)
    ws.protection.sheet = True


def write_workbook(path: Path, open_rows: List[dict], archive_rows: List[dict]) -> Path:
    """Write workbook. Falls back to PENDING file if canonical path is locked."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_main_sheet(wb, open_rows)
    write_grouped_sheet(
        wb, open_rows, "By Vendor",
        group_key_fn=lambda r: r["vendor"],
        sort_fn=lambda r: (r["vendor"].lower(), -r["bill_date"].toordinal()),
    )
    write_grouped_sheet(
        wb, open_rows, "By Project",
        group_key_fn=lambda r: f"{r['division']} {r['project_num']}".strip(),
        sort_fn=lambda r: (r["division"], r["project_num"], -r["bill_date"].toordinal()),
    )
    write_archive_sheet(wb, archive_rows)

    try:
        wb.save(path)
        return path
    except (PermissionError, OSError) as e:
        pending = path.with_name(path.stem + ".PENDING.xlsx")
        wb.save(pending)
        print(f"  ⚠ could not write canonical file ({e}); wrote {pending} instead")
        return pending


# ───────────────────────── main ─────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", type=Path, default=DEFAULT_OUTPUT,
                    help=f"output file path (default: {DEFAULT_OUTPUT})")
    ap.add_argument("--dry-run", action="store_true",
                    help="build everything but do not write the file")
    args = ap.parse_args()

    started = dt.datetime.now()
    print(f"→ {started:%Y-%m-%d %H:%M:%S}  bill tracker run starting")
    print(f"  output target: {args.out}")

    # 1) read existing edits
    print("→ reading existing file for preserved Notes / Status Override …")
    notes_db = read_user_edits(args.out)
    print(f"  {len(notes_db)} preserved entries")

    # 2) auth
    print("→ authenticating to QBO (Touch ID) …")
    access, cid = load_credentials()
    print(f"  ok. company_id={cid}")

    # 3) reference data
    print("→ fetching vendors …")
    vendors = query_all(access, cid, "Vendor")
    vendor_map = {
        v["Id"]: v.get("DisplayName") or v.get("CompanyName") or f"Vendor {v['Id']}"
        for v in vendors
    }
    print(f"  {len(vendors)} vendors")

    # 4) open bills (we'll filter subs in Python)
    print("→ fetching open bills (Balance > 0) …")
    bills = query_all(access, cid, "Bill", where="Balance > '0'")
    print(f"  {len(bills)} bills with open balance")

    n_pre = len(bills)
    bills = [b for b in bills if not is_sub_bill(b)]
    print(f"  excluding subs: {n_pre - len(bills)} dropped, {len(bills)} kept")

    # 5) invoices (excluding retainage)
    print("→ fetching invoices …")
    invoices_raw = query_all(access, cid, "Invoice")
    print(f"  {len(invoices_raw)} invoices total")
    n_pre = len(invoices_raw)
    invoices = [inv for inv in invoices_raw if not is_excluded_invoice(inv)]
    print(f"  excluding retainage: {n_pre - len(invoices)} dropped, {len(invoices)} kept")

    invoices_by_customer: Dict[str, List[dict]] = defaultdict(list)
    for inv in invoices:
        cid_v = (inv.get("CustomerRef") or {}).get("value", "")
        if cid_v:
            invoices_by_customer[cid_v].append(inv)

    # 6) build current open rows
    print("→ building rows for open bills …")
    open_rows = build_rows(bills, invoices_by_customer, vendor_map, notes_db)
    print(f"  {len(open_rows)} bill lines")

    # 7) archive: bills tracked previously but no longer open
    open_keys = {r["key"] for r in open_rows}
    prev_keys = set(notes_db.keys())
    archived_keys = prev_keys - open_keys
    archive_rows: List[dict] = []
    if archived_keys:
        archived_bill_ids = {k.split("-")[0] for k in archived_keys}
        print(f"→ fetching {len(archived_bill_ids)} previously-tracked bills now closed …")
        arch_bills = fetch_bills_by_ids(access, cid, archived_bill_ids)
        # filter sub bills (in case classification changed)
        arch_bills = [b for b in arch_bills if not is_sub_bill(b)]
        archive_rows = build_rows(arch_bills, invoices_by_customer, vendor_map, notes_db)
        # only keep keys that are actually archived (a single bill could have lines in both states)
        archive_rows = [r for r in archive_rows if r["key"] in archived_keys]
        print(f"  {len(archive_rows)} archived bill lines")

    # 8) status breakdown (using effective / final status)
    breakdown: Dict[str, int] = defaultdict(int)
    for r in open_rows:
        eff = r["override"] or r["auto_status"]
        breakdown[eff] += 1
    print("→ status breakdown (open bills):")
    for s, n in sorted(breakdown.items(), key=lambda x: -x[1]):
        print(f"    {s}: {n}")

    # 9) write
    if args.dry_run:
        print("→ --dry-run: not writing file")
        return 0

    print("→ writing workbook …")
    actual = write_workbook(args.out, open_rows, archive_rows)
    elapsed = (dt.datetime.now() - started).total_seconds()
    print(f"✓ done in {elapsed:.1f}s. file: {actual}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n✗ interrupted")
        sys.exit(130)
    except Exception as e:
        print(f"\n✗ FAILED: {type(e).__name__}: {e}")
        traceback.print_exc()
        sys.exit(1)
