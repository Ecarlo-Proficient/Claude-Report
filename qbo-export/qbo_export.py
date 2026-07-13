#!/usr/bin/env python3
"""
qbo_export.py — Export QBO transactions to a flat xlsx table.

Pulls: Bill, Purchase (cc/cash expense), JournalEntry, Invoice.
Writes ONE ROW PER LINE ITEM, so a bill with 3 expense lines becomes
3 rows. Columns:

  Txn Date | Type | Doc # | Name | Project # | Account (Category) |
  Description | Amount | Memo | Txn ID

Output:  <OneDrive inbox>/qbexp_transactions_<YYYYMMDD_HHMMSS>.xlsx

Usage:
  python3 qbo_export.py                     # last 24 months
  python3 qbo_export.py --since 2024-01-01  # custom start
  python3 qbo_export.py --all               # everything, from the beginning
  python3 qbo_export.py --out /path/out.xlsx  # override output path

Auth: reads the QBO credentials blob from Keychain ONCE (single Touch ID
prompt), caches in memory for the rest of the run.
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("missing dependency. Run: pip3 install --break-system-packages openpyxl requests")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import qbo_vault as kc
from shared import paths

# Production only — no env selector.
API_BASE = "https://quickbooks.api.intuit.com"

DEFAULT_OUTPUT_DIR = paths.get_path(
    "ACB_EXPORT_INBOX_DIR",
    paths.onedrive_base() / "-Inbox- Project Report Exports",
)


# ──────────────────────────  auth  ──────────────────────────

def load_credentials() -> Tuple[str, str]:
    """Single Touch ID prompt unlocks all QBO keys. Returns (access_token, company_id)."""
    if not kc.has_credentials():
        print("✗ no credentials stored.")
        print("  fix:  python3 setup_qbo.py")
        sys.exit(1)

    try:
        creds = kc.get_all()   # ← single Touch ID
    except kc.SecretsError as e:
        print(f"✗ Keychain read failed: {e}")
        sys.exit(1)

    required = ["QBO_CLIENT_ID", "QBO_CLIENT_SECRET", "QBO_COMPANY_ID", "QBO_REFRESH_TOKEN"]
    missing = [k for k in required if not creds.get(k)]
    if missing:
        print(f"✗ incomplete blob. Missing: {', '.join(missing)}")
        print("  fix:  python3 setup_qbo.py")
        sys.exit(1)

    basic = base64.b64encode(f"{creds['QBO_CLIENT_ID']}:{creds['QBO_CLIENT_SECRET']}".encode()).decode()
    r = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        headers={
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        data={"grant_type": "refresh_token", "refresh_token": creds["QBO_REFRESH_TOKEN"]},
        timeout=30,
    )
    if r.status_code != 200:
        print(f"✗ token refresh failed  status={r.status_code}")
        print(f"  body: {r.text[:400]}")
        print(f"  diagnose:  python3 setup_qbo.py --test")
        sys.exit(1)

    data = r.json()
    new_rt = data.get("refresh_token")
    if new_rt and new_rt != creds["QBO_REFRESH_TOKEN"]:
        try:
            kc.put("QBO_REFRESH_TOKEN", new_rt)
        except kc.SecretsError:
            pass  # non-fatal; will work next run

    return data["access_token"], creds["QBO_COMPANY_ID"]


# ──────────────────────────  api  ──────────────────────────

def _api_get(path: str, access: str, params: Optional[dict] = None) -> dict:
    p = dict(params or {})
    p["minorversion"] = "70"
    r = requests.get(
        f"{API_BASE}{path}",
        headers={"Authorization": f"Bearer {access}", "Accept": "application/json"},
        params=p,
        timeout=45,
    )
    if r.status_code != 200:
        raise RuntimeError(f"{path} → {r.status_code}: {r.text[:400]}")
    return r.json()


def query(access: str, company_id: str, q: str) -> dict:
    return _api_get(f"/v3/company/{company_id}/query", access, params={"query": q})


def query_all(access: str, company_id: str, entity: str, where: str = "") -> List[dict]:
    """Paginated query — pulls every row of `entity`, optional WHERE clause."""
    out: List[dict] = []
    start = 1
    page = 500
    while True:
        q = f"SELECT * FROM {entity}"
        if where:
            q += f" WHERE {where}"
        q += f" STARTPOSITION {start} MAXRESULTS {page}"
        data = query(access, company_id, q)
        batch = data.get("QueryResponse", {}).get(entity, [])
        if not batch:
            break
        out.extend(batch)
        if len(batch) < page:
            break
        start += page
    return out


# ──────────────────────────  name lookups  ──────────────────────────

def _build_account_map(access: str, company_id: str) -> Dict[str, str]:
    rows = query_all(access, company_id, "Account")
    return {a["Id"]: a.get("Name", f"Account {a['Id']}") for a in rows}


def _build_customer_map(access: str, company_id: str) -> Dict[str, dict]:
    rows = query_all(access, company_id, "Customer")
    return {c["Id"]: c for c in rows}


def _build_vendor_map(access: str, company_id: str) -> Dict[str, str]:
    rows = query_all(access, company_id, "Vendor")
    return {
        v["Id"]: v.get("DisplayName") or v.get("CompanyName") or f"Vendor {v['Id']}"
        for v in rows
    }


def _build_item_map(access: str, company_id: str, accounts: Dict[str, str]) -> Dict[str, str]:
    """Map Item.Id → account name (income for sales items, expense for purchase)."""
    rows = query_all(access, company_id, "Item")
    out: Dict[str, str] = {}
    for it in rows:
        acct_id = None
        if "IncomeAccountRef" in it:
            acct_id = it["IncomeAccountRef"].get("value")
        elif "ExpenseAccountRef" in it:
            acct_id = it["ExpenseAccountRef"].get("value")
        out[it["Id"]] = accounts.get(acct_id, it.get("Name", f"Item {it['Id']}"))
    return out


def _customer_display(cust: dict) -> Tuple[str, str]:
    """Returns (display_name, project_number).

    QBO stores project/job info in various places across versions. The
    DisplayName is the most reliable canonical identifier — it matches how
    Ted names projects in QBO."""
    name = cust.get("DisplayName") or cust.get("CompanyName") or f"Customer {cust.get('Id')}"
    return name, name


# ──────────────────────────  line extractors  ──────────────────────────

def _date_str(v: Optional[str]) -> str:
    return (v or "")[:10]


def _bill_rows(bills, accounts, customers, vendors):
    rows = []
    for b in bills:
        v_ref = b.get("VendorRef", {})
        vendor = vendors.get(v_ref.get("value", ""), v_ref.get("name", ""))
        for ln in b.get("Line", []):
            det = ln.get("AccountBasedExpenseLineDetail") or ln.get("ItemBasedExpenseLineDetail")
            if not det:
                continue
            if "AccountRef" in det:
                acct_id = det["AccountRef"].get("value", "")
                acct_name = accounts.get(acct_id, det["AccountRef"].get("name", ""))
            else:
                acct_name = det.get("ItemRef", {}).get("name", "")

            cust_id = det.get("CustomerRef", {}).get("value", "")
            cust = customers.get(cust_id, {})
            cust_name, proj = _customer_display(cust) if cust else ("", "")

            rows.append({
                "Txn Date": _date_str(b.get("TxnDate")),
                "Type": "Bill",
                "Doc #": b.get("DocNumber", ""),
                "Name": vendor,
                "Project #": proj,
                "Account (Category)": acct_name,
                "Description": ln.get("Description", ""),
                "Amount": float(ln.get("Amount", 0) or 0),
                "Memo": b.get("PrivateNote", ""),
                "Txn ID": b.get("Id", ""),
            })
    return rows


def _purchase_rows(purchases, accounts, customers, vendors):
    rows = []
    for p in purchases:
        entity = p.get("EntityRef", {})
        if entity.get("type") == "Vendor":
            name = vendors.get(entity.get("value", ""), entity.get("name", ""))
        else:
            name = entity.get("name", "")

        ptype = p.get("PaymentType", "Cash")  # Cash / Check / CreditCard
        type_label = f"Purchase ({ptype})"

        for ln in p.get("Line", []):
            det = ln.get("AccountBasedExpenseLineDetail") or ln.get("ItemBasedExpenseLineDetail")
            if not det:
                continue
            if "AccountRef" in det:
                acct_id = det["AccountRef"].get("value", "")
                acct_name = accounts.get(acct_id, det["AccountRef"].get("name", ""))
            else:
                acct_name = det.get("ItemRef", {}).get("name", "")

            cust_id = det.get("CustomerRef", {}).get("value", "")
            cust = customers.get(cust_id, {})
            cust_name, proj = _customer_display(cust) if cust else ("", "")

            rows.append({
                "Txn Date": _date_str(p.get("TxnDate")),
                "Type": type_label,
                "Doc #": p.get("DocNumber", ""),
                "Name": name,
                "Project #": proj,
                "Account (Category)": acct_name,
                "Description": ln.get("Description", ""),
                "Amount": float(ln.get("Amount", 0) or 0),
                "Memo": p.get("PrivateNote", ""),
                "Txn ID": p.get("Id", ""),
            })
    return rows


def _journal_rows(journals, accounts, customers):
    rows = []
    for j in journals:
        for ln in j.get("Line", []):
            det = ln.get("JournalEntryLineDetail")
            if not det:
                continue
            acct_id = det.get("AccountRef", {}).get("value", "")
            acct_name = accounts.get(acct_id, det.get("AccountRef", {}).get("name", ""))

            amt = float(ln.get("Amount", 0) or 0)
            if det.get("PostingType") == "Credit":
                amt = -amt

            # Customer / Project tag on the JE line, if present
            entity = det.get("Entity") or {}
            cust_id = ""
            if entity.get("Type") == "Customer":
                cust_id = entity.get("EntityRef", {}).get("value", "")
            cust = customers.get(cust_id, {}) if cust_id else {}
            cust_name, proj = _customer_display(cust) if cust else ("", "")

            rows.append({
                "Txn Date": _date_str(j.get("TxnDate")),
                "Type": "Journal Entry",
                "Doc #": j.get("DocNumber", ""),
                "Name": cust_name,
                "Project #": proj,
                "Account (Category)": acct_name,
                "Description": ln.get("Description", ""),
                "Amount": amt,
                "Memo": j.get("PrivateNote", ""),
                "Txn ID": j.get("Id", ""),
            })
    return rows


_RETAINAGE_RE = re.compile(r"retainage\s+not\s+billed", re.IGNORECASE)

def _is_retainage_line(ln: dict) -> bool:
    """True for retainage holdback lines — excluded from billed totals."""
    desc      = ln.get("Description") or ""
    item_name = (ln.get("SalesItemLineDetail") or {}).get("ItemRef", {}).get("name") or ""
    return bool(_RETAINAGE_RE.search(desc) or _RETAINAGE_RE.search(item_name))

def _invoice_rows(invoices, items, customers):
    rows = []
    for inv in invoices:
        c_ref = inv.get("CustomerRef", {})
        cust_id = c_ref.get("value", "")
        cust = customers.get(cust_id, {})
        if cust:
            cust_name, proj = _customer_display(cust)
        else:
            cust_name = c_ref.get("name", "")
            proj = cust_name

        for ln in inv.get("Line", []):
            det = ln.get("SalesItemLineDetail")
            if not det:
                continue
            if _is_retainage_line(ln):
                continue                       # exclude retainage holdback lines
            item_id = det.get("ItemRef", {}).get("value", "")
            acct_name = items.get(item_id, det.get("ItemRef", {}).get("name", ""))

            rows.append({
                "Txn Date": _date_str(inv.get("TxnDate")),
                "Type": "Invoice",
                "Doc #": inv.get("DocNumber", ""),
                "Name": cust_name,
                "Project #": proj,
                "Account (Category)": acct_name,
                "Description": ln.get("Description", ""),
                "Amount": float(ln.get("Amount", 0) or 0),
                "Memo": inv.get("PrivateNote", ""),
                "Txn ID": inv.get("Id", ""),
            })
    return rows


# ──────────────────────────  xlsx writer  ──────────────────────────

HEADERS = [
    "Txn Date", "Type", "Doc #", "Name", "Project #",
    "Account (Category)", "Description", "Amount", "Memo", "Txn ID",
]


def write_xlsx(rows: List[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    rows_sorted = sorted(rows, key=lambda r: r.get("Txn Date", ""))
    for r_idx, row in enumerate(rows_sorted, start=2):
        for c_idx, h in enumerate(HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))

    amt_col = HEADERS.index("Amount") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=amt_col).number_format = '#,##0.00;[Red](#,##0.00)'

    widths = {
        "Txn Date": 11, "Type": 18, "Doc #": 12, "Name": 28, "Project #": 24,
        "Account (Category)": 32, "Description": 42, "Amount": 14,
        "Memo": 28, "Txn ID": 10,
    }
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ──────────────────────────  main  ──────────────────────────

def _default_since() -> str:
    today = dt.date.today()
    y = today.year - 2
    return dt.date(y, today.month, 1).isoformat()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", help="Start date YYYY-MM-DD (default: 24 months back)")
    ap.add_argument("--all", action="store_true", help="No date filter — export everything")
    ap.add_argument("--out", help="Override output file path")
    args = ap.parse_args()

    print("━" * 60)
    print("  QBO Transaction Export")
    print("━" * 60)

    access, company_id = load_credentials()
    print(f"  ✓ authenticated  (company={company_id})")

    if args.all:
        print(f"  → pulling ALL transactions (no date filter)")
        where = ""
    else:
        since = args.since or _default_since()
        print(f"  → pulling transactions since {since}")
        where = f"TxnDate >= '{since}'"

    print(f"  → loading name maps (Account, Customer, Vendor, Item)...")
    accounts = _build_account_map(access, company_id)
    customers = _build_customer_map(access, company_id)
    vendors = _build_vendor_map(access, company_id)
    items = _build_item_map(access, company_id, accounts)
    print(f"    accounts: {len(accounts)}   customers: {len(customers)}"
          f"   vendors: {len(vendors)}   items: {len(items)}")

    all_rows: List[dict] = []

    print(f"  → Bills...")
    bills = query_all(access, company_id, "Bill", where=where)
    br = _bill_rows(bills, accounts, customers, vendors)
    print(f"    {len(bills)} bills → {len(br)} lines")
    all_rows.extend(br)

    print(f"  → Purchases (cash/cc/check)...")
    purchases = query_all(access, company_id, "Purchase", where=where)
    pr = _purchase_rows(purchases, accounts, customers, vendors)
    print(f"    {len(purchases)} purchases → {len(pr)} lines")
    all_rows.extend(pr)

    print(f"  → Journal Entries...")
    journals = query_all(access, company_id, "JournalEntry", where=where)
    jr = _journal_rows(journals, accounts, customers)
    print(f"    {len(journals)} journal entries → {len(jr)} lines")
    all_rows.extend(jr)

    print(f"  → Invoices...")
    invoices = query_all(access, company_id, "Invoice", where=where)
    ir = _invoice_rows(invoices, items, customers)
    print(f"    {len(invoices)} invoices → {len(ir)} lines")
    all_rows.extend(ir)

    if args.out:
        out_path = Path(args.out)
    else:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = DEFAULT_OUTPUT_DIR / f"qbexp_transactions_{ts}.xlsx"

    write_xlsx(all_rows, out_path)
    print()
    print(f"  ✓ wrote {len(all_rows)} rows → {out_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\ninterrupted.")
        sys.exit(130)
