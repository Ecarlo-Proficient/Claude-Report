"""
Export Notion Invoice Trackers to a single Excel file.

One-direction sync: Notion → Excel. The file is a read-only artifact for
The developer's ad-hoc summing (select rows in Excel → see Sum in status bar) and
copy-paste-to-email workflow. The developer NEVER edits the Excel file — every
edit-worthy field (Quick Status, Next Follow-Up, Assignee, etc.) lives in
Notion. This file is regenerated every sync run, so any edits made directly
would be clobbered.

Why this exists:
  Notion can't sum arbitrary row selection (only filtered totals). The developer's
  workflow includes ad-hoc questions like "what do customer X's last 3
  invoices total?" — for that, Excel is the right tool. Notion is still the
  source of truth and the working surface for all human notes / status.

Tabs:
  1. "Open Invoices" — the flat row-per-invoice list (the original sheet).
  2. "CP Aging" / "MFD Aging" / "RP Aging" — QBO-style aging buckets rolled up
     by parent client, one tab per division (the user 2026-08-10), each with the
     lien-notice clock and the collections clerk's notes. CP and MFD also carry
     the previous-draw block; RP drops it (RP doesn't bill in draws). There is
     no Division column — the tab IS the division. See aging_sheet.py.

Run via run_invoice_sync.py after the main sync completes. Errors here
don't affect the QBO→Notion sync (caught and logged separately).
"""
from __future__ import annotations

import datetime as dt
import logging
from collections import defaultdict
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from notion_client import NotionClient
from draw_chain import DrawChains
from notes_preserve import read_notes, reapply_notes, absorb_into_records, PreservedNotes
from cash_flow import build_cash_flow_sheets
from aging_sheet import (
    DIVISION_TABS,
    DRAW_DIVISIONS,
    RP_DROP_COLUMNS,
    load_vendor_bill_map,
    vendor_cells,
    build_aging_sheet,
)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import lien_clock


log = logging.getLogger("automation_worker.export_xlsx")

# Default file location — overrideable via INVOICE_EXPORT_PATH (env or
# machine.env) so you can point this at Synology, OneDrive, Desktop, etc.
# without code change. Fallback = repo root (same location as before).
DEFAULT_EXPORT_PATH = paths.get_path(
    "INVOICE_EXPORT_PATH",
    Path(__file__).resolve().parent.parent / "Open_Invoices.xlsx",
)


# Column order matches the developer's preferred layout (he sorts/scans left-to-right
# in this order: Division/Project # first for quick triage, then customer/invoice
# detail, then $ columns, then status). PM, Quick Status, and QBO Link live only
# in Notion — this Excel is a read-only summing sheet for the collections workflow.
COLUMNS = [
    ("Division", 10),
    ("Project #", 12),
    ("Client", 28),
    ("Date", 12),
    ("Invoice #", 12),
    ("Net Terms", 14),
    ("Due Date", 12),
    ("Past Due", 10),
    ("Memo", 40),
    ("Total Amount", 14),
    ("Open Balance", 14),
    ("Status", 14),
    ("Aging Bucket", 12),
]


def _text(prop: dict) -> str:
    """Pull plain text from a Notion rich_text or title property."""
    if not prop:
        return ""
    arr = prop.get("rich_text") or prop.get("title") or []
    return "".join(t.get("plain_text", "") for t in arr).strip()


def _select_name(prop: dict) -> str:
    sel = (prop or {}).get("select") or {}
    return sel.get("name", "")


def _multi_select_names(prop: dict) -> str:
    items = (prop or {}).get("multi_select") or []
    return ", ".join(i.get("name", "") for i in items)


def _people_names(prop: dict) -> str:
    """Notion people property → comma-separated names."""
    items = (prop or {}).get("people") or []
    return ", ".join(p.get("name", "") for p in items)


def _date_value(prop: dict) -> Optional[dt.date]:
    d = (prop or {}).get("date") or {}
    start = d.get("start")
    if not start:
        return None
    try:
        return dt.date.fromisoformat(start[:10])
    except ValueError:
        return None


def _number(prop: dict) -> Optional[float]:
    return (prop or {}).get("number")


def _url(prop: dict) -> str:
    return (prop or {}).get("url") or ""


def _relation_first_title(prop: dict, customer_title_cache: Dict[str, str]) -> str:
    """For a relation property, return the first related page's title via cache."""
    rels = (prop or {}).get("relation") or []
    if not rels:
        return ""
    page_id = rels[0].get("id")
    return customer_title_cache.get(page_id, "")


def _build_customer_title_cache(
    notion: NotionClient, customer_ds_id: str, title_prop: str = "Client"
) -> Dict[str, str]:
    """{page_id → display name} for all customers in a customer DB."""
    cache: Dict[str, str] = {}
    for page in notion.query_data_source(customer_ds_id, page_size=100):
        page_id = page.get("id")
        if not page_id:
            continue
        title_prop_data = (page.get("properties") or {}).get(title_prop) or {}
        title_arr = title_prop_data.get("title") or []
        name = "".join(t.get("plain_text", "") for t in title_arr).strip()
        if name:
            cache[page_id] = name
    return cache


def _row_for(
    page: dict,
    division_label: str,
    customer_title_cache: Dict[str, str],
    today: dt.date,
) -> List[Any]:
    """Build one Excel row from a Notion invoice page."""
    props = page.get("properties") or {}

    # Customer — prefer the relation's resolved name, fall back to Customer (raw).
    customer_relation_name = _relation_first_title(
        props.get("Customer"), customer_title_cache
    )
    customer = customer_relation_name or _text(props.get("Customer (raw)"))

    invoice_num = _text(props.get("Invoice #"))
    project_num = _text(props.get("Project #"))
    memo = _text(props.get("Memo"))

    invoice_date = _date_value(props.get("Date"))
    due_date = _date_value(props.get("Due Date"))
    # Past Due — signed days. Positive = actually past due. Negative = not yet due
    # (e.g., -2 = "due in 2 days"). Zero = due today. "" only when no due date set.
    # Number format on the cell (see below) adds the "+" prefix on positives.
    past_due = (today - due_date).days if due_date else ""

    total_amt = _number(props.get("Total Amount"))
    open_bal = _number(props.get("Open balance"))

    status = _select_name(props.get("Status"))
    aging = _select_name(props.get("Aging Bucket"))
    net_terms = _select_name(props.get("Net Terms"))
    division = _select_name(props.get("Division")) or division_label

    return [
        division,                              # Division
        project_num,                           # Project #
        customer,                              # Client
        invoice_date,                          # Date
        invoice_num,                           # Invoice #
        net_terms,                             # Net Terms
        due_date,                              # Due Date
        past_due,                              # Past Due
        memo,                                  # Memo
        total_amt,                             # Total Amount
        open_bal,                              # Open Balance
        status,                                # Status
        aging,                                 # Aging Bucket
    ]


def _aging_record(
    page: dict,
    division_label: str,
    customer_title_cache: Dict[str, str],
    today: dt.date,
    vendor_map: Optional[Dict[str, Any]],
    chains: Optional[Any] = None,
) -> dict:
    """Flatten one Notion invoice page into the dict the aging tab consumes.

    Kept separate from `_row_for` on purpose: the Open Invoices tab is a
    faithful mirror of the tracker, while the aging tab is a derived view with
    its own fields (parent client, computed days-past-due, vendor status).
    """
    props = page.get("properties") or {}

    due_date = _date_value(props.get("Due Date"))
    division = _select_name(props.get("Division")) or division_label
    invoice_num = _text(props.get("Invoice #"))
    prev_draw, vendor_status, vendor_bills, vendor_amount, this_amount = vendor_cells(
        division, invoice_num, vendor_map, chains
    )

    return {
        # Page id so an absorbed note can be mirrored up to this invoice's Notion
        # Quick Status without a second lookup.
        "page_id": page.get("id"),
        # The relation resolves to the PARENT customer (e.g. the GC), while
        # "Customer (raw)" is the project-level child ("MFD177 - MERRITT PARK").
        # Grouping needs the parent, so the raw name is only a last resort.
        "parent": _relation_first_title(props.get("Customer"), customer_title_cache)
        or _text(props.get("Customer (raw)")),
        "division": division,
        "project_num": _text(props.get("Project #")),
        "invoice_num": invoice_num,
        "invoice_date": _date_value(props.get("Date")),
        "due_date": due_date,
        "days_past_due": (today - due_date).days if due_date else None,
        "open_balance": _number(props.get("Open balance")) or 0.0,
        # Original invoice total — the Open Balance data bar on the aging tab is
        # scaled to this (bar fill = open ÷ total). Fall back to the open balance
        # so the bar is a full cell rather than dividing by a missing/zero total.
        "total_amount": _number(props.get("Total Amount"))
        or (_number(props.get("Open balance")) or 0.0),
        "memo": _text(props.get("Memo")),
        # "Quick Status" is where the collections clerk writes what's actually
        # happening on the invoice ("1st reminder", "Lien Notice Sent",
        # "Waiting vendor unconditional"). That's the Notes the owner wants.
        # Texas lien-notice clock, shared with money_bleeds via shared/lien_clock.
        # Work month = invoice month (settled 2026-07-16). Retainage and an
        # already-sent notice are read off the memo / clerk's note.
        "lien": lien_clock.lien_state(
            division, _date_value(props.get("Date")), today,
            memo=_text(props.get("Memo")),
            note=_text(props.get("Quick Status")),
        ),
        "notes": _text(props.get("Quick Status")),
        "last_action": _date_value(props.get("Last Action Date")),
        # Deep link to the invoice in QBO, written by the sync. The aging tab
        # hangs it off the invoice number so the row is one click from source.
        "qbo_link": _url(props.get("QBO Link")),
        "prev_draw": prev_draw,
        "vendor_status": vendor_status,
        "vendor_bills": vendor_bills,
        "vendor_amount": vendor_amount,
        "this_draw_amount": this_amount,
    }


def _is_litigation(page: dict) -> bool:
    return bool(((page.get("properties") or {}).get("Litigation") or {}).get("checkbox"))


OPEN_STATUSES = ("Unpaid", "Partially Paid")


def _status_name(page: dict) -> str:
    return _select_name((page.get("properties") or {}).get("Status"))


def _all_invoices(notion: NotionClient, ds_id: str) -> List[dict]:
    """Every invoice in a tracker, any status.

    The AR Aging tab needs the PAID ones too: an open draw's predecessor is
    almost always already paid, and a draw chain built from open invoices alone
    would be blind exactly where it has to see (see draw_chain.py). The Open
    Invoices tab still shows only the open ones — filtered below.
    """
    return list(notion.query_data_source(ds_id, page_size=100))


def _open_invoices(notion: NotionClient, ds_id: str) -> List[dict]:
    """Pull non-Paid invoices from a tracker."""
    filter_body = {
        "or": [
            {"property": "Status", "select": {"equals": s}} for s in OPEN_STATUSES
        ]
    }
    return list(notion.query_data_source(ds_id, filter_body=filter_body, page_size=100))


def _push_absorbed_notes(notion: NotionClient, plans: List[dict], today: dt.date) -> List[dict]:
    """Mirror absorbed notes to Notion: archive the prior Quick Status to the page
    body (the documented 'Collection Log'), then set Quick Status to the note text.

    Returns the list of plans whose push FAILED, so the caller can keep those cell
    Notes rather than dropping them (a failed push must never lose a note). Order of
    ops per invoice matters: set Quick Status FIRST, then append the archive line,
    so a mid-way failure can't leave the old status archived while Quick Status
    still holds it (which a re-run would then double-archive).

    Idempotent across runs: absorb only emits a plan when the note text differs from
    the current Quick Status, so a re-run finds them equal and plans nothing.
    """
    pushed = 0
    failed: List[dict] = []
    for p in plans:
        page_id = p.get("page_id")
        if not page_id:
            log.warning("Absorb push: %s has no Notion page id, kept as a cell Note.", p["invoice_num"])
            failed.append(p)
            continue
        try:
            notion.update_page(
                page_id,
                {"Quick Status": {"rich_text": [{"text": {"content": p["new_text"][:2000]}}]}},
            )
            old = (p.get("old_quick_status") or "").strip()
            if old:
                # Prior status to the Collection Log (page body), dated, so the
                # page reads as a chronological status history.
                notion.append_paragraph(page_id, f"{today.month}/{today.day}: {old}")
            pushed += 1
        except Exception as e:  # noqa: BLE001 — one bad page shouldn't sink the run
            failed.append(p)
            log.warning("Absorb push failed for invoice %s (cell Note kept): %s", p["invoice_num"], e)
    log.info("Absorb push: %d Quick Status updated in Notion (%d kept as cell Notes).", pushed, len(failed))
    return failed


def export_open_invoices_xlsx(
    *,
    notion: NotionClient,
    res_com_ds_id: str,
    mfd_ds_id: str,
    customer_list_ds_id: str,
    mfd_client_list_ds_id: str,
    output_path: Optional[Path] = None,
    absorb_notes: bool = False,
    apply_notion: bool = False,
) -> Path:
    """
    Pull open invoices from both trackers, write a single Excel file
    formatted for the developer's collections workflow. Returns the output path.

    Notes handling (the clerk's Excel Notes on the aging tabs):
      * absorb_notes=False (default) — PRESERVE: re-attach every cell Note to its
        row so nothing typed in Excel is lost on the rebuild. Quick Status (from
        Notion) still drives the Notes column. Safe: no Notion writes.
      * absorb_notes=True — ABSORB (the user 2026-08-11): a cell Note IS the new
        status. Its text (stamped `— Name, M/D`) replaces the Notes column and the
        cell Note is dropped (not re-attached). A per-client Note lands on every
        open invoice of that client; a per-invoice Note wins its own row.
      * apply_notion — only meaningful with absorb_notes. False = dry-run: log
        exactly what WOULD be written to Notion Quick Status, write nothing. True =
        push the absorbed text up to Notion. Absorb without a Notion push would
        lose the note on the next pull, so the live flow must pair them; the
        preview runs absorb + dry-run against a throwaway file (live copy safe).
    """
    # DEFAULT_EXPORT_PATH already resolved INVOICE_EXPORT_PATH via paths.get_path
    # (env var > machine.env > repo default) — don't re-resolve it here.
    target = output_path or DEFAULT_EXPORT_PATH

    # Office lock check — Excel creates a hidden '~$Open_Invoices.xlsx' file in
    # the same folder while someone has the workbook open (the developer on his side,
    # The user on his, doesn't matter — OneDrive syncs the lock to both machines).
    # If we overwrite while Excel is open, his AutoSave will silently write the
    # stale buffer back on top of our update, losing the sync. Skip cleanly and
    # tell the operator to re-run once it's closed.
    #
    # Edge case: if Excel crashed and left a stale '~$' file behind, manually
    # delete it from the OneDrive folder.
    lock_file = target.parent / f"~${target.name}"
    if lock_file.exists():
        log.warning(
            "Excel file is OPEN (%s present in OneDrive). "
            "Skipped Excel export this run — overwriting now would collide with "
            "the open Excel session and lose the update. "
            "Close the file and re-run `sync-ar` to refresh the OneDrive mirror.",
            lock_file.name,
        )
        return target

    # Read the clerk's Excel Notes off the current file BEFORE we rebuild it, so
    # they survive the overwrite (the user 2026-08-11). The file isn't locked
    # here (checked above), so this is a clean read of the last-synced workbook.
    preserved = read_notes(target)

    log.info("Loading Notion customer caches for export…")
    res_com_titles = _build_customer_title_cache(notion, customer_list_ds_id)
    mfd_titles = _build_customer_title_cache(notion, mfd_client_list_ds_id)

    log.info("Pulling invoices from Notion…")
    # One pull per tracker covering EVERY status: the open subset feeds both
    # tabs, the full set builds the draw chains (paid predecessors included).
    res_com_all = _all_invoices(notion, res_com_ds_id)
    mfd_all = _all_invoices(notion, mfd_ds_id)
    res_com_pages = [p for p in res_com_all if _status_name(p) in OPEN_STATUSES]
    mfd_pages = [p for p in mfd_all if _status_name(p) in OPEN_STATUSES]
    log.info(
        "Export: %d Res/Com + %d MFD open (of %d + %d total)",
        len(res_com_pages), len(mfd_pages), len(res_com_all), len(mfd_all),
    )

    today = dt.date.today()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Invoices"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        col = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    row_num = 2
    for page in res_com_pages:
        row = _row_for(page, "RP/CP", res_com_titles, today)
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1
    for page in mfd_pages:
        row = _row_for(page, "MFD", mfd_titles, today)
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

    # Number / currency / date formatting (column indices match COLUMNS order)
    for r in range(2, row_num):
        ws.cell(row=r, column=4).number_format = "mm/dd/yyyy"   # Date
        ws.cell(row=r, column=7).number_format = "mm/dd/yyyy"   # Due Date
        ws.cell(row=r, column=10).number_format = '"$"#,##0.00' # Total Amount
        ws.cell(row=r, column=11).number_format = '"$"#,##0.00' # Open Balance
        # Past Due (col 8) — "+12" for past due, "-2" for not yet due, "0" for due today.
        # Custom format: positive;negative;zero (Excel applies "+" prefix on positives).
        past_due_cell = ws.cell(row=r, column=8)
        past_due_cell.number_format = '"+"0;-0;0'
        # Bold red when > 30 days overdue
        if isinstance(past_due_cell.value, int) and past_due_cell.value > 30:
            past_due_cell.font = Font(bold=True, color="C00000")

    # Auto-filter on the header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row_num - 1}"

    # ── Tab 2: AR Aging ──
    # Built from the SAME page lists as tab 1 — no second Notion round-trip.
    vendor_map, vendor_as_of = load_vendor_bill_map()

    # Draw chains come from EVERY invoice — litigation ones included. They are
    # excluded from the aging rows, but an excluded invoice is still a real link
    # in its project's draw sequence and dropping it would silently make the
    # draw before it look like the immediate predecessor.
    chains = DrawChains()
    for page in res_com_all + mfd_all:
        props = page.get("properties") or {}
        chains.add(
            invoice_num=_text(props.get("Invoice #")),
            project_num=_text(props.get("Project #")),
            memo=_text(props.get("Memo")),
            invoice_date=_date_value(props.get("Date")),
            is_paid=_select_name(props.get("Status")) == "Paid",
        )
    chains.finalize()

    aging_records: List[dict] = []
    litigation_excluded: Dict[str, int] = defaultdict(int)
    for pages, label, cache in (
        (res_com_pages, "RP/CP", res_com_titles),
        (mfd_pages, "MFD", mfd_titles),
    ):
        for page in pages:
            # Litigation invoices are legal work, not collections work — leaving
            # them in would inflate every aging bucket the owner reads. (the user)
            if _is_litigation(page):
                litigation_excluded[
                    _select_name((page.get("properties") or {}).get("Division")) or label
                ] += 1
                continue
            aging_records.append(
                _aging_record(page, label, cache, today, vendor_map, chains)
            )

    # ── Notes: absorb the clerk's cell Notes as the new status, or preserve them ──
    # Absorb rewrites the Notes column HERE (before the sheets are built) for the
    # invoices whose Notion push succeeded; a failed push instead keeps the cell
    # Note (collected in `keep`) so it is never lost. In preserve mode `keep` is
    # unused and every cell Note is re-attached after the build.
    keep = PreservedNotes({}, {})
    if absorb_notes and preserved.total():
        note_plans = absorb_into_records(preserved, aging_records, today)
        log.info(
            "Notes ABSORB: %d invoice row(s) take a cell Note as their status.",
            len(note_plans),
        )
        for p in note_plans:
            verb = "PUSH" if apply_notion else "would push (dry-run)"
            log.info(
                "  %s %s [%s] Quick Status: %r -> %r",
                verb, p["invoice_num"], p["scope"], p["old_quick_status"], p["new_text"],
            )
        # Dry-run (preview) never writes Notion, so treat every plan as applied for
        # the throwaway file; a real run only applies the ones that actually pushed.
        failed = _push_absorbed_notes(notion, note_plans, today) if apply_notion else []
        failed_plan_ids = {id(p) for p in failed}
        for p in note_plans:
            if id(p) in failed_plan_ids:
                if p["scope"] == "invoice":
                    keep.per_invoice[(p["sheet"], p["invoice_num"])] = p["saved"]
                else:
                    keep.per_client[(p["sheet"], p["client"])] = p["saved"]
            else:
                # Absorbed: the Note becomes this row's Notes column value.
                p["rec"]["notes"] = p["new_text"]

    # One tab per division (the user 2026-08-10 — "keep cp and mfd separated").
    # There is no Division column any more: the tab IS the division, so a column
    # repeating it on every row earns nothing.
    counts = []
    for sheet_name, division, sheet_title in DIVISION_TABS:
        rows = [r for r in aging_records if r["division"] == division]
        build_aging_sheet(
            wb.create_sheet(sheet_name),
            rows,
            today=today,
            litigation_excluded=litigation_excluded.get(division, 0),
            vendor_as_of=vendor_as_of,
            # RP doesn't bill in draws, so the previous-draw block is dropped
            # there rather than rendered as a column of grey "n/a".
            drop_columns=RP_DROP_COLUMNS if division not in DRAW_DIVISIONS else (),
            title=sheet_title,
        )
        counts.append(f"{sheet_name} {len(rows)}")
    log.info(
        "Aging tabs: %s (%d litigation excluded)",
        " · ".join(counts), sum(litigation_excluded.values()),
    )

    # ── Cash-flow forecast (the user 2026-08-12) ── built from the SAME notes the
    # aging tabs carry, so absorb/preserve both feed it. Two tabs: a weekly list
    # and a month calendar grid; only clear dated promises land on them.
    build_cash_flow_sheets(wb, aging_records, today)

    # Re-attach cell Notes:
    #   ABSORB  → only those whose Notion push FAILED (kept so they're not lost);
    #             absorbed Notes have become the Notes column value and are dropped.
    #   PRESERVE → every cell Note, so nothing typed in Excel is lost on regen.
    if absorb_notes:
        if keep.total():
            log.warning(
                "%d note(s) could not push to Notion this run and were kept as cell "
                "Notes (not lost); they'll retry next sync.", keep.total(),
            )
            reapply_notes(wb, keep)
    elif preserved.total():
        reapply_notes(wb, preserved)

    # Write
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    log.info("Exported %d invoices → %s", row_num - 2, target)
    return target
