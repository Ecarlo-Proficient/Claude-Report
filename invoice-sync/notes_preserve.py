"""
notes_preserve.py — carry the collections clerk's Excel Notes across a rebuild.

Why this exists (the user 2026-08-11):
    The owner writes running collections notes directly on the aging tabs of
    Open_Invoices.xlsx as Excel **Notes** (the legacy yellow sticky — `Insert ▸
    Note`, author attached; NOT the newer threaded *Comments*, which the file
    format stores separately and openpyxl can't read). Those Notes live only in
    the workbook. `export_invoices_xlsx` rebuilds the whole file from Notion every
    run, so without this module every Note is wiped on the next `sync-ar`.

    Two scopes, both anchored on the Invoice # column (column C on every aging
    tab):
      * **Per-invoice** — a Note on a detail row's invoice number → that invoice.
      * **Per-client** — a Note on a client summary row's "N inv" cell → covers
        ALL of that client's open invoices (the "sum of invoices" the owner
        pointed at).

    This module only round-trips the Notes IN EXCEL: read them off the existing
    file before it is overwritten, re-attach them to the matching cells in the
    freshly built workbook. Verbatim — same text, same author — so the file the
    owner reopens looks exactly as they left it. Pushing the text up to Notion is
    a separate, gated step (design D) and lives elsewhere; nothing here writes to
    Notion.

Legacy Notes are read via openpyxl's `cell.comment` (openpyxl models legacy
notes and threaded-comment fallbacks under the same `Comment` object).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.workbook import Workbook

log = logging.getLogger("automation_worker.notes_preserve")

# The aging tabs, with their Invoice # column (1-based) and client-label column.
# Column C (=3) is the invoice number on every aging tab; column A (=1) is the
# client name on summary rows. The "Open Invoices" flat tab carries the invoice
# number in column E (=5) and has no client-summary rows.
_AGING_SHEETS = ("CP Aging", "MFD Aging", "RP Aging")
_SHEET_INV_COL = {"CP Aging": 3, "MFD Aging": 3, "RP Aging": 3, "Open Invoices": 5}
_LABEL_COL = 1

# A client summary row's Invoice # cell reads like "4 inv" (see aging_sheet).
_N_INV_RE = re.compile(r"^\s*\d+\s*inv\s*$", re.IGNORECASE)
# A real invoice number is all digits (QBO DocNumber), possibly with a suffix.
_INV_NUM_RE = re.compile(r"^[0-9][0-9A-Za-z\-]*$")


@dataclass
class SavedNote:
    """One preserved Excel Note: verbatim text + author + original box size."""
    text: str
    author: str
    width: Optional[str] = None
    height: Optional[str] = None

    def to_comment(self) -> Comment:
        c = Comment(self.text, self.author or "")
        if self.width:
            c.width = self.width
        if self.height:
            c.height = self.height
        return c


@dataclass
class PreservedNotes:
    # keyed by (sheet_name, invoice_num) and (sheet_name, client_name)
    per_invoice: Dict[Tuple[str, str], SavedNote]
    per_client: Dict[Tuple[str, str], SavedNote]

    def total(self) -> int:
        return len(self.per_invoice) + len(self.per_client)


def _classify(sheet_name: str, inv_val, label_val) -> Optional[Tuple[str, str]]:
    """Return ('invoice', num) or ('client', name) for a row, or None.

    Keyed off the Invoice # cell so it doesn't matter which cell the Note is
    physically pinned to — a Note on the client name or on the "N inv" cell both
    resolve to the same client via the row.
    """
    inv_txt = "" if inv_val is None else str(inv_val).strip()
    if sheet_name in _AGING_SHEETS and _N_INV_RE.match(inv_txt):
        name = "" if label_val is None else str(label_val).strip()
        return ("client", name) if name else None
    if inv_txt and _INV_NUM_RE.match(inv_txt):
        return ("invoice", inv_txt)
    return None


def read_notes(path: Path) -> PreservedNotes:
    """Read every Note off the existing workbook, keyed for re-attachment.

    Returns empty sets if the file doesn't exist yet (first run) or can't be
    read — never raises into the export, which must still produce a fresh file.
    """
    per_invoice: Dict[Tuple[str, str], SavedNote] = {}
    per_client: Dict[Tuple[str, str], SavedNote] = {}
    if not path.exists():
        return PreservedNotes(per_invoice, per_client)
    try:
        # NOT read_only — read_only workbooks drop comments entirely.
        wb = load_workbook(path, data_only=False)
    except Exception as e:  # noqa: BLE001 — a bad existing file must not block export
        log.warning("Could not read existing notes from %s (%s) — none carried over.", path, e)
        return PreservedNotes(per_invoice, per_client)

    try:
        for sheet_name in wb.sheetnames:
            inv_col = _SHEET_INV_COL.get(sheet_name)
            if inv_col is None:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment is None:
                        continue
                    r = cell.row
                    inv_val = ws.cell(row=r, column=inv_col).value
                    label_val = ws.cell(row=r, column=_LABEL_COL).value
                    key = _classify(sheet_name, inv_val, label_val)
                    if key is None:
                        continue
                    saved = SavedNote(
                        text=cell.comment.text or "",
                        author=cell.comment.author or "",
                        width=getattr(cell.comment, "width", None),
                        height=getattr(cell.comment, "height", None),
                    )
                    if key[0] == "invoice":
                        per_invoice[(sheet_name, key[1])] = saved
                    else:
                        per_client[(sheet_name, key[1])] = saved
    finally:
        wb.close()

    log.info(
        "Preserved notes read: %d per-invoice, %d per-client",
        len(per_invoice), len(per_client),
    )
    return PreservedNotes(per_invoice, per_client)


_DIVISION_TO_SHEET = {"CP": "CP Aging", "MFD": "MFD Aging", "RP": "RP Aging"}


def first_name(author: str) -> str:
    """'Sebastian' from 'Sebastian Perez' — the attribution the owner asked for."""
    a = (author or "").strip()
    return a.split()[0] if a else ""


def clean_text(saved: SavedNote) -> str:
    """The note body, minus the 'Author:' header Excel auto-inserts, on one line.

    Legacy Notes store the author both as a `<author>` attribute and as a bold
    first line ('Sebastian Perez:') inside the text. Drop that first line if it
    just repeats the author, and collapse the rest to a single spaced line for the
    Notes column / Quick Status.
    """
    text = saved.text or ""
    author = (saved.author or "").strip()
    lines = text.split("\n")
    if lines and author and lines[0].strip().rstrip(":").strip() == author:
        lines = lines[1:]
    return " ".join(" ".join(lines).split())


def stamped_text(saved: SavedNote, today) -> str:
    """'<note> – <Name>, <M>/<D>' (numeric date, no year; the user 2026-08-11).

    Separator is an EN dash (medium), never an em dash — the owner's standing
    rule is short '-' or medium '–' only.
    """
    body = clean_text(saved)
    fn = first_name(saved.author)
    stamp = f"{fn}, {today.month}/{today.day}" if fn else f"{today.month}/{today.day}"
    return f"{body} – {stamp}" if body else stamp


def absorb_into_records(preserved: PreservedNotes, aging_records: List[dict], today) -> List[dict]:
    """Plan folding each cell Note into its invoice(s) as the new status ("absorb").

    A per-invoice Note replaces that invoice's `notes` (the Notes column, which
    mirrors Notion `Quick Status`); a per-client Note replaces it on EVERY open
    invoice of that client, but a per-invoice Note wins for its own row.

    Does NOT mutate the records here: it returns one plan per affected invoice, and
    the caller applies the override (`rec['notes'] = new_text`) ONLY for invoices
    whose Notion push succeeded, so a failed push keeps its cell Note instead of
    silently dropping it. Each plan carries `rec` and `saved` refs for exactly that
    decision, plus the old status (captured before any override) for the archive.
    """
    plans: List[dict] = []
    for rec in aging_records:
        sheet = _DIVISION_TO_SHEET.get(rec.get("division"))
        if not sheet:
            continue
        saved = preserved.per_invoice.get((sheet, str(rec.get("invoice_num"))))
        scope = "invoice"
        if saved is None:
            saved = preserved.per_client.get((sheet, rec.get("parent")))
            scope = "client"
        if saved is None:
            continue
        new_text = stamped_text(saved, today)
        old_qs = rec.get("notes") or ""
        if new_text == old_qs:
            continue  # already reflects the note — nothing to move or push
        plans.append({
            "sheet": sheet,
            "invoice_num": str(rec.get("invoice_num")),
            "page_id": rec.get("page_id"),
            "scope": scope,
            "author": saved.author,
            "old_quick_status": old_qs,
            "new_text": new_text,
            "rec": rec,
            "saved": saved,
            "client": rec.get("parent") if scope == "client" else None,
        })
    return plans


def reapply_notes(wb: Workbook, notes: PreservedNotes) -> Tuple[int, List[str]]:
    """Re-attach preserved Notes to the freshly built workbook.

    Matches per-invoice notes by invoice number and per-client notes by client
    name, on the same aging tab they came from. Returns (reapplied count, orphan
    descriptions) — an orphan is a Note whose invoice/client is no longer on the
    sheet (paid off, removed). Orphans are reported so the operator knows; the
    text is still in the pre-run backup regardless.
    """
    reapplied = 0
    matched_inv: set = set()
    matched_client: set = set()

    for sheet_name in wb.sheetnames:
        inv_col = _SHEET_INV_COL.get(sheet_name)
        if inv_col is None:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            r = row[0].row
            inv_val = ws.cell(row=r, column=inv_col).value
            label_val = ws.cell(row=r, column=_LABEL_COL).value
            key = _classify(sheet_name, inv_val, label_val)
            if key is None:
                continue
            if key[0] == "invoice":
                saved = notes.per_invoice.get((sheet_name, key[1]))
                if saved is not None:
                    ws.cell(row=r, column=inv_col).comment = saved.to_comment()
                    matched_inv.add((sheet_name, key[1]))
                    reapplied += 1
            else:
                saved = notes.per_client.get((sheet_name, key[1]))
                if saved is not None:
                    ws.cell(row=r, column=inv_col).comment = saved.to_comment()
                    matched_client.add((sheet_name, key[1]))
                    reapplied += 1

    orphans: List[str] = []
    for (sheet_name, inv), _ in notes.per_invoice.items():
        if (sheet_name, inv) not in matched_inv:
            orphans.append(f"{sheet_name}: invoice {inv}")
    for (sheet_name, client), _ in notes.per_client.items():
        if (sheet_name, client) not in matched_client:
            orphans.append(f"{sheet_name}: client '{client}'")

    if orphans:
        log.warning(
            "%d preserved note(s) had no matching row this run (invoice paid/removed "
            "or client cleared): %s. Text remains in the pre-run backup.",
            len(orphans), "; ".join(orphans),
        )
    log.info("Notes re-applied: %d", reapplied)
    return reapplied, orphans
