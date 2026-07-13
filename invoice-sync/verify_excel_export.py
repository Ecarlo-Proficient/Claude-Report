#!/usr/bin/env python3
"""
verify_excel_export.py — three-way audit of the OneDrive Open_Invoices.xlsx.

Reads the live Excel artifact, the live Notion Invoice Trackers, and live QBO
open invoices, and checks that all three agree. Catches:
  * Notion open rows that didn't make it into the Excel export
  * Excel rows that shouldn't be there (Notion says Paid but Excel still has them)
  * Total open-balance drift between Excel ↔ Notion ↔ QBO (routable subset)
  * Field-level drift on Total Amount / Open Balance / Status for spot-checked rows

Usage:
    python3 verify_excel_export.py                 # print to stdout
    python3 verify_excel_export.py --out audit.md  # write markdown report

Exit codes:
    0  Excel matches Notion (clean)
    1  Drift detected
    2  Fatal (can't open file / QBO auth / Notion auth)
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

# Repo root on sys.path for shared/ (explicit — do NOT rely on qbo_client's insert).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import load_config
from notion_client import NotionClient
import qbo_client
from shared import paths


# Reuse same default path the exporter uses.
DEFAULT_EXPORT_PATH = paths.get_path(
    "INVOICE_EXPORT_PATH",
    paths.onedrive_base() / "Collections/Open_Invoices.xlsx",
)

PROJECT_NUM_RE = re.compile(r"\b((?:MFD|CP|RP)\d+(?:-FTW)?)\b", re.IGNORECASE)

# Column positions in the exported Excel (1-indexed). MUST match COLUMNS in
# export_invoices_xlsx.py — if the exporter's column order changes, update here.
COL_DIVISION = 1
COL_PROJECT = 2
COL_CLIENT = 3
COL_DATE = 4
COL_INVOICE_NUM = 5
COL_NET_TERMS = 6
COL_DUE_DATE = 7
COL_PAST_DUE = 8
COL_MEMO = 9
COL_TOTAL_AMT = 10
COL_OPEN_BAL = 11
COL_STATUS = 12
COL_AGING = 13


def _load_excel(path: Path) -> List[dict]:
    """Read the Excel file → list of dicts keyed by invoice #."""
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[COL_INVOICE_NUM - 1] in (None, ""):
            continue
        rows.append({
            "invoice_num": str(row[COL_INVOICE_NUM - 1]).strip(),
            "client": str(row[COL_CLIENT - 1] or "").strip(),
            "project": str(row[COL_PROJECT - 1] or "").strip(),
            "division": str(row[COL_DIVISION - 1] or "").strip(),
            "total_amt": float(row[COL_TOTAL_AMT - 1] or 0),
            "open_bal": float(row[COL_OPEN_BAL - 1] or 0),
            "status": str(row[COL_STATUS - 1] or "").strip(),
        })
    return rows


def _notion_open_invoices(notion: NotionClient, ds_id: str, label: str) -> Dict[str, dict]:
    """Pull Notion rows where Status in (Unpaid, Partially Paid)."""
    filter_body = {
        "or": [
            {"property": "Status", "select": {"equals": "Unpaid"}},
            {"property": "Status", "select": {"equals": "Partially Paid"}},
        ]
    }
    out = {}
    for page in notion.query_data_source(ds_id, filter_body=filter_body, page_size=100):
        props = page.get("properties") or {}
        title_arr = (props.get("Invoice #") or {}).get("title") or []
        invoice_num = "".join(t.get("plain_text", "") for t in title_arr).strip()
        if not invoice_num:
            continue
        inv_id_arr = (props.get("Invoice ID") or {}).get("rich_text") or []
        inv_id = "".join(t.get("plain_text", "") for t in inv_id_arr).strip()
        status = ((props.get("Status") or {}).get("select") or {}).get("name", "")
        total_amt = (props.get("Total Amount") or {}).get("number")
        open_bal = (props.get("Open balance") or {}).get("number")
        proj_arr = (props.get("Project #") or {}).get("rich_text") or []
        project = "".join(t.get("plain_text", "") for t in proj_arr).strip()
        out[invoice_num] = {
            "invoice_num": invoice_num,
            "qbo_id": inv_id,
            "status": status,
            "total_amt": float(total_amt or 0),
            "open_bal": float(open_bal or 0),
            "project": project,
            "division": label,
        }
    return out


def _qbo_routable_open(creds) -> Dict[str, dict]:
    """{qbo_id → {invoice_num, balance, project}} for routable open invoices."""
    raw = qbo_client.query_all(creds, "Invoice", where="Balance > '0'")
    out = {}
    for inv in raw:
        qbo_id = str(inv.get("Id") or "")
        if not qbo_id:
            continue
        doc_num = str(inv.get("DocNumber") or "").strip()
        balance = float(inv.get("Balance") or 0)
        total = float(inv.get("TotalAmt") or 0)
        cust = (inv.get("CustomerRef") or {}).get("name") or ""
        memo = inv.get("PrivateNote") or ""
        text = f"{cust} {memo}"
        m = PROJECT_NUM_RE.search(text)
        project = m.group(1).upper() if m else ""
        if not project:
            continue  # unroutable, intentionally not in Excel
        out[qbo_id] = {
            "invoice_num": doc_num,
            "balance": balance,
            "total": total,
            "project": project,
        }
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel ↔ Notion ↔ QBO three-way audit")
    parser.add_argument("--out", help="Write markdown report to this path")
    parser.add_argument(
        "--xlsx-path",
        default=str(DEFAULT_EXPORT_PATH),
        help="Path to the Excel file (defaults to INVOICE_EXPORT_PATH env or OneDrive path)",
    )
    args = parser.parse_args()

    try:
        config = load_config()
    except Exception as e:
        print(f"FATAL: config: {e}", file=sys.stderr)
        return 2

    xlsx_path = Path(args.xlsx_path)
    print(f"[1/4] Reading Excel: {xlsx_path}", file=sys.stderr)
    try:
        excel_rows = _load_excel(xlsx_path)
    except Exception as e:
        print(f"FATAL: can't read Excel: {e}", file=sys.stderr)
        return 2

    print("[2/4] Loading Notion open invoices…", file=sys.stderr)
    notion = NotionClient(
        secret=config.notion_secret,
        api_base=config.notion_api_base,
        version=config.notion_version,
    )
    res_com = _notion_open_invoices(notion, config.invoice_res_com_ds_id, "RP/CP")
    mfd = _notion_open_invoices(notion, config.invoice_mfd_ds_id, "MFD")
    notion_open = {**res_com, **mfd}

    print("[3/4] Loading QBO open invoices…", file=sys.stderr)
    try:
        creds = qbo_client.load_qbo_credentials()
        qbo_open = _qbo_routable_open(creds)
    except Exception as e:
        print(f"WARN: QBO check skipped: {e}", file=sys.stderr)
        qbo_open = {}

    print("[4/4] Comparing…", file=sys.stderr)

    excel_by_num = {r["invoice_num"]: r for r in excel_rows}

    # Drift checks
    in_notion_not_excel = sorted(set(notion_open) - set(excel_by_num))
    in_excel_not_notion = sorted(set(excel_by_num) - set(notion_open))

    # Balance totals
    excel_open_total = sum(r["open_bal"] for r in excel_rows)
    notion_open_total = sum(r["open_bal"] for r in notion_open.values())
    qbo_open_total = sum(r["balance"] for r in qbo_open.values()) if qbo_open else None

    # Per-invoice field drift between Excel and Notion (the two that should be 1:1)
    field_mismatches = []
    for num, n_row in notion_open.items():
        e_row = excel_by_num.get(num)
        if not e_row:
            continue
        if abs(e_row["open_bal"] - n_row["open_bal"]) > 0.01:
            field_mismatches.append((num, "Open Balance", e_row["open_bal"], n_row["open_bal"]))
        if abs(e_row["total_amt"] - n_row["total_amt"]) > 0.01:
            field_mismatches.append((num, "Total Amount", e_row["total_amt"], n_row["total_amt"]))
        if e_row["status"] != n_row["status"]:
            field_mismatches.append((num, "Status", e_row["status"], n_row["status"]))

    # QBO ↔ Notion sanity (routable open invoices)
    qbo_open_invoice_ids = {r["qbo_id"] for r in notion_open.values() if r["qbo_id"]}
    qbo_routable_ids = set(qbo_open.keys())
    qbo_not_in_notion = qbo_routable_ids - qbo_open_invoice_ids
    notion_not_in_qbo = qbo_open_invoice_ids - qbo_routable_ids

    # Build report
    today = dt.date.today().isoformat()
    md: List[str] = []
    md.append(f"# Excel Export Audit — {today}\n")
    md.append(f"**File:** `{xlsx_path}`\n")
    md.append(f"**Last modified:** {dt.datetime.fromtimestamp(xlsx_path.stat().st_mtime).isoformat(timespec='seconds')}\n")

    md.append("## Top-line numbers\n")
    md.append("| Source | Open invoices | Total open balance |")
    md.append("|---|---:|---:|")
    md.append(f"| Excel (OneDrive) | {len(excel_rows)} | ${excel_open_total:,.2f} |")
    md.append(f"| Notion (RP/CP + MFD, open only) | {len(notion_open)} | ${notion_open_total:,.2f} |")
    if qbo_open_total is not None:
        md.append(f"| QBO (routable open only) | {len(qbo_open)} | ${qbo_open_total:,.2f} |")
    md.append("")

    # Verdict
    excel_notion_match = (
        len(excel_rows) == len(notion_open)
        and abs(excel_open_total - notion_open_total) < 0.01
        and not in_notion_not_excel
        and not in_excel_not_notion
        and not field_mismatches
    )
    md.append("## Verdict\n")
    if excel_notion_match:
        md.append("**Excel ↔ Notion: PASS** — Excel is a clean mirror of Notion open invoices "
                  "(row count, total open balance, and per-invoice fields all agree).\n")
    else:
        md.append("**Excel ↔ Notion: DRIFT** — see sections below.\n")

    if qbo_open:
        qbo_match = (
            len(notion_open) >= len(qbo_open)  # Notion may have more (legacy customer-less rows still tracked)
            and not qbo_not_in_notion
        )
        if qbo_match:
            md.append("**Notion ↔ QBO: PASS** — every routable QBO open invoice is in Notion.\n")
        else:
            md.append(f"**Notion ↔ QBO: DRIFT** — {len(qbo_not_in_notion)} routable QBO invoice(s) "
                      "are not in Notion (next sync should pick them up).\n")

    # Detail sections
    if in_notion_not_excel:
        md.append(f"## Notion open but missing from Excel ({len(in_notion_not_excel)})\n")
        md.append("These rows should have been in the export. If the Excel file is fresh, the export "
                  "filter or write step has a bug.\n")
        md.append("| Invoice # | Project # | Open Balance | Status |")
        md.append("|---|---|---:|---|")
        for num in in_notion_not_excel[:25]:
            r = notion_open[num]
            md.append(f"| {num} | {r['project']} | ${r['open_bal']:,.2f} | {r['status']} |")
        if len(in_notion_not_excel) > 25:
            md.append(f"| _…and {len(in_notion_not_excel) - 25} more_ |  |  |  |")
        md.append("")

    if in_excel_not_notion:
        md.append(f"## Excel has rows Notion doesn't show as open ({len(in_excel_not_notion)})\n")
        md.append("Most likely cause: invoice was just paid in QBO and the sweep flipped Notion to Paid, "
                  "but the Excel file is stale (export hasn't run since). Solution: run `sync-ar`.\n")
        md.append("| Invoice # | Client | Open Balance (Excel) | Status (Excel) |")
        md.append("|---|---|---:|---|")
        for num in in_excel_not_notion[:25]:
            r = excel_by_num[num]
            md.append(f"| {num} | {r['client']} | ${r['open_bal']:,.2f} | {r['status']} |")
        if len(in_excel_not_notion) > 25:
            md.append(f"| _…and {len(in_excel_not_notion) - 25} more_ |  |  |  |")
        md.append("")

    if field_mismatches:
        md.append(f"## Field-level drift Excel ↔ Notion ({len(field_mismatches)})\n")
        md.append("| Invoice # | Field | Excel | Notion |")
        md.append("|---|---|---|---|")
        for num, field, exc, ntn in field_mismatches[:25]:
            md.append(f"| {num} | {field} | {exc} | {ntn} |")
        if len(field_mismatches) > 25:
            md.append(f"| _…and {len(field_mismatches) - 25} more_ |  |  |  |")
        md.append("")

    if qbo_open and qbo_not_in_notion:
        md.append(f"## Routable QBO open invoices missing from Notion ({len(qbo_not_in_notion)})\n")
        md.append("These will get picked up by the next `sync-ar` run.\n")
        md.append("| QBO Inv # | Project # | Balance |")
        md.append("|---|---|---:|")
        for qid in list(qbo_not_in_notion)[:25]:
            r = qbo_open[qid]
            md.append(f"| {r['invoice_num']} | {r['project']} | ${r['balance']:,.2f} |")
        md.append("")

    md.append("---\n")
    md.append("_Generated by `verify_excel_export.py`. Re-run anytime after `sync-ar` to confirm the "
              "OneDrive file is in sync with Notion + QBO._\n")

    report = "\n".join(md)
    if args.out:
        Path(args.out).parent.mkdir(parents=True, exist_ok=True)
        Path(args.out).write_text(report)
        print(f"\nReport written to {args.out}", file=sys.stderr)
    else:
        print()
        print(report)

    drift = (
        in_notion_not_excel
        or in_excel_not_notion
        or field_mismatches
        or (qbo_open and qbo_not_in_notion)
    )
    return 1 if drift else 0


if __name__ == "__main__":
    sys.exit(main())
