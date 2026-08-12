"""
aging_sheet.py — the per-division aging tabs of Open_Invoices.xlsx.

Why this exists (the user 2026-08-05):
    Notion is good for reading ONE invoice page. It is bad at the thing the
    owner actually does every week — scanning a hundred rows at once to see
    "who owes me money, how old is it, and what is holding it up." That is a
    QBO-style aging view: one row per invoice, the open balance dropped into
    a Current / 1-30 / 31-60 / 61-90 / 90+ column, all three divisions in one
    place, rolled up under the parent client.

What this tab shows that QBO's own aging does NOT:
    1. **Notes** — the collections clerk's running note on each invoice
       (Notion `Quick Status`) plus the date they last touched it.
    2. **The lien clock** — the `Lien` column (which replaced Days Past Due,
       the user 2026-08-10) gives the date a Ch. 53 notice must be MAILED by.
       Days-past-due was already legible from which bucket the money sits in;
       the lien deadline is not, and it is the one that EXPIRES. Dates come from
       `shared/lien_clock.py`, shared with money_bleeds so both tools can never
       drift apart on a statutory date.
    3. **Why the draw isn't funded yet** — for MFD and CP, the state of the
       PREVIOUS draw. The funding is a chain (the user 2026-08-05): the GC funds
       draw N, we pay draw N's vendor bills, those vendors issue unconditional
       waivers, and the GC needs the waivers before releasing draw N+1. So an
       unpaid draw is rarely about its own bills — it's about whether the one
       before it is cleared. `draw_chain.py` builds the sequence;
       `load_vendor_bill_map` reads what's still owed, from the bill-tracker's
       output file.

       The verdict splits the two very different holds: `PAY BILLS → unlock`
       (previous draw funded, our vendors still owed — **ours** to fix) versus
       `Waiting GC on prev` (previous draw not funded either — upstream of us).
       `This Draw $ Open` keeps the older same-draw figure alongside it.

Rules baked in:
    - **Litigation invoices are excluded** (the `Litigation` checkbox in both
      Notion trackers). They are not collections work anymore; they are legal
      work, and leaving them in the aging inflates every bucket.
    - Aging is by DUE DATE, matching QBO's default AR aging and the
      `Aging Bucket` select that invoice_sync already writes to Notion.
    - Parent-client groups are **collapsed by default** — the owner opens the
      client they care about instead of scrolling 60 invoice rows.

This module only builds the worksheet; `export_invoices_xlsx.py` owns the
workbook and the Notion pull.
"""
from __future__ import annotations

import datetime as dt
import logging
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import draw_chain

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import lien_clock


log = logging.getLogger("automation_worker.aging_sheet")

# One tab per division (the user 2026-08-10 — "keep cp and mfd separated").
# (sheet name, division, title, columns to omit) — RP drops the whole
# previous-draw block because RP doesn't bill in draws.
DIVISION_TABS = (
    ("CP Aging", "CP", "CP AGING"),
    ("MFD Aging", "MFD", "MFD AGING"),
    ("RP Aging", "RP", "RP AGING"),
)

# Type sizes (the user 2026-08-05 — "make the font 12 to see it a bit bigger",
# client rows 13). Client rows carry a point more because the sheet is read
# COLLAPSED: the client name is the line that has to land first.
BODY_PT = 12
CLIENT_PT = 13
TITLE_PT = 14
# Widths are in characters of the default 11pt font; _autofit scales for 12pt.
# The cap stops a long memo or note from creating a column you have to scroll.
MAX_COL_WIDTH = 62

# The bill-tracker's Excel output. Same env key the bill-tracker itself uses,
# so a machine.env override moves both together. We read the FILE, not the
# bill-tracker's code — tools never import tools (repo rule 3).
BILL_TRACKER_PATH = paths.get_path(
    "ACB_BILL_TRACKER_XLSX",
    paths.onedrive_base() / "Automations-" / "Bill Tracker.xlsx",
)

# Previous-draw verdicts. The funding chain (the user 2026-08-05): the GC funds
# draw N, we pay draw N's vendor bills, the vendors issue unconditional waivers,
# and the GC needs those waivers before funding draw N+1. So for an unpaid draw
# the question is never "are THIS draw's bills paid" — it's whether the PREVIOUS
# draw is cleared. These are a small fixed set so the column filters to two
# clicks; counts and dollars live in their own columns.
PREV_BLOCKED = "PAY BILLS → unlock"   # prev draw funded, our vendors still owed — WE are the blocker
PREV_WAITING_GC = "Waiting GC on prev"  # prev draw itself unpaid — blocker is upstream of us
PREV_CLEAR = "Clear"                  # prev draw funded and its vendors paid
PREV_FIRST = "First draw"             # nothing before it to gate funding
PREV_NOT_DRAW = "Not a draw"          # retainage / turnkey one-offs — no chain
PREV_MULTI = "Multi-contract"         # parallel contracts, bills unattributable (see draw_chain)
VENDOR_NA = "n/a"                     # RP — no draws at all
VENDOR_UNKNOWN = "?"                  # bill tracker file missing / unreadable

# Verdicts with no previous draw to report on. The count/$ cells under them are
# not "zero", they are unanswerable — so the whole block greys out. The verdict
# text itself survives (see _grey_out_vendor_block: it only fills EMPTY cells),
# which keeps "Multi-contract" and "First draw" readable while making clear
# there is no number beside them.
NO_CHAIN_VERDICTS = (
    VENDOR_NA, PREV_NOT_DRAW, PREV_MULTI, PREV_FIRST, VENDOR_UNKNOWN,
)

# The divisions that bill in draws at all. Bill lines are matched to the invoice
# authorising their payment via the DRAW PERIOD here (bill-tracker README, "How
# matching works"). RP matches on "earliest invoice on/after bill date" — not a
# draw period, and RP has no draw chain to walk — so the whole block is n/a.
DRAW_DIVISIONS = ("MFD", "CP")

# (header, width, number_format)
# No Division column: each tab IS one division (the user 2026-08-10), so it
# would repeat the tab name on every row.
COLUMNS: List[Tuple[str, int, Optional[str]]] = [
    ("Client / Invoice", 34, None),
    ("Project #",        13, None),
    ("Invoice #",        11, None),
    ("Date",             11, "mm/dd/yyyy"),
    ("Due Date",         11, "mm/dd/yyyy"),
    # Replaced Days Past Due (the user 2026-08-10). Days-past-due was already
    # legible from the bucket the money sits in; the lien deadline is not, and
    # it's the one that expires.
    ("Lien",             22, None),
    ("Current",          14, '"$"#,##0.00'),
    ("1-30",             14, '"$"#,##0.00'),
    ("31-60",            14, '"$"#,##0.00'),
    ("61-90",            14, '"$"#,##0.00'),
    ("90+",              14, '"$"#,##0.00'),
    # Open Balance first, then the invoice's original Total Amount, so the pair
    # reads open→total left-to-right (the user 2026-08-11). A data bar on Open
    # Balance is scaled to that row's Total Amount, so its fill = how much of the
    # invoice is still open (full bar = untouched, short bar = partly collected).
    ("Open Balance",     15, '"$"#,##0.00'),
    ("Total Amount",     15, '"$"#,##0.00'),
    ("Prev Draw",        11, None),
    ("Prev Draw Status", 19, None),
    ("Prev Bills Open",   9, "0"),
    ("Prev $ Open",      15, '"$"#,##0.00'),
    ("This Draw $ Open", 15, '"$"#,##0.00'),
    ("Notes",            46, None),
    ("Last Action",      12, "mm/dd/yyyy"),
]

# 0-based positions used when writing rows (kept in sync with COLUMNS above).
C_LABEL, C_PROJ, C_INV, C_DATE, C_DUE, C_LIEN = range(6)
C_CURRENT, C_1_30, C_31_60, C_61_90, C_90 = range(6, 11)
# C_TOTAL = Open Balance (kept name for the sum-of-buckets), C_INVTOTAL = the
# invoice's original Total Amount that the Open Balance bar is scaled against.
C_TOTAL, C_INVTOTAL, C_PREV, C_VSTATUS, C_VBILLS, C_VAMT, C_THIS, C_NOTES, C_ACTION = range(11, 20)

BUCKET_COLS = (C_CURRENT, C_1_30, C_31_60, C_61_90, C_90)

# The previous-draw block — greyed out wherever there is no chain to read.
VENDOR_COLS = (C_PREV, C_VSTATUS, C_VBILLS, C_VAMT)

# Columns the RP tab omits (spreadsheet N through R): the whole previous-draw
# block plus the same-draw figure. RP doesn't bill in draws, so on the combined
# tab these are 34 rows of grey "n/a". (the user 2026-08-05)
RP_DROP_COLUMNS = (C_PREV, C_VSTATUS, C_VBILLS, C_VAMT, C_THIS)

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")


# ─────────────────────── palette ───────────────────────
# Colour is here to answer "how bad is this" without reading a number, so it
# only ever encodes AGE (the five buckets) and STATE (vendors, dead cells).
# Nothing decorative gets a fill.

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")   # matches the Open Invoices tab
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Bucket headers, green → red as the money ages. White bold text on all five.
BUCKET_HEADER_FILLS = (
    PatternFill("solid", fgColor="2E7D32"),   # Current — not yet due
    PatternFill("solid", fgColor="7CB342"),   # 1-30
    PatternFill("solid", fgColor="D68910"),   # 31-60
    PatternFill("solid", fgColor="C0552B"),   # 61-90
    PatternFill("solid", fgColor="922B21"),   # 90+
)
# Light tint of the same hue for the ONE cell in each row that holds the amount.
# Scanning down the sheet, colour drifting rightward = money getting older.
BUCKET_CELL_FILLS = (
    PatternFill("solid", fgColor="E8F5E9"),
    PatternFill("solid", fgColor="F1F8E9"),
    PatternFill("solid", fgColor="FDF2E0"),
    PatternFill("solid", fgColor="FAE7E0"),
    PatternFill("solid", fgColor="F8DDDA"),
)

_GRAND_FILL = PatternFill("solid", fgColor="BFD3E6")    # the all-clients roll-up
_PARENT_FILL = PatternFill("solid", fgColor="DCE6F1")   # each client summary row

# Dead cells: the previous-draw block where there's nothing to read. Grey fill,
# darker grey text — "nothing to see here" without looking like missing data.
_NA_FILL = PatternFill("solid", fgColor="D9D9D9")
_NA_COLOR = "808080"

_BLOCKED_FILL = PatternFill("solid", fgColor="F8DDDA")  # the one row-level call to action
_LINK_COLOR = "0563C1"                                  # Excel's own hyperlink blue

# Lien deadlines are the only HARD expiry on this sheet — miss one and the right
# is gone, not merely late. Past due gets the only reversed-out cell here.
_LIEN_PAST_FILL = PatternFill("solid", fgColor="922B21")
_LIEN_URGENT_FILL = PatternFill("solid", fgColor="F8DDDA")
_LIEN_WATCH_FILL = PatternFill("solid", fgColor="FDF2E0")


# ─────────────────────── aging ───────────────────────

def bucket_index(days_past_due: Optional[int]) -> int:
    """Position in BUCKET_COLS for a signed days-past-due value.

    Positive = overdue. <= 0 (not yet due, or due today) = Current, which is how
    QBO ages AR and how invoice_sync assigns the Notion `Aging Bucket` select.
    An invoice with no due date can't be aged — it lands in Current rather than
    silently disappearing from the row total.
    """
    if days_past_due is None or days_past_due <= 0:
        return 0
    if days_past_due <= 30:
        return 1
    if days_past_due <= 60:
        return 2
    if days_past_due <= 90:
        return 3
    return 4


# ─────────────────── vendor bills (bill tracker) ───────────────────

def _humanize_age(stamp: dt.datetime) -> str:
    """'3 hours' / '2 days' — for telling the reader how old the vendor data is."""
    delta = dt.datetime.now() - stamp
    hours = delta.days * 24 + delta.seconds // 3600
    if hours < 1:
        return "under an hour"
    if hours < 48:
        return f"{hours} hour{'s' if hours != 1 else ''}"
    return f"{delta.days} days"

def load_vendor_bill_map(
    path: Path = BILL_TRACKER_PATH,
) -> Tuple[Optional[Dict[str, Tuple[float, int, int]]], Optional[dt.datetime]]:
    """{invoice # → (open $, bill count, vendor count)} from Bill Tracker.xlsx.

    The bill-tracker's `Bills` sheet is LINE-level: one row per bill line, with
    the bill's own `Bill Open Bal` repeated on every line of that bill. Summing
    the column directly would multiply a bill by its line count, so we dedupe on
    (vendor, bill #, bill date) per invoice before adding.

    Only MFD/CP rows carry a draw-period invoice match, and only lines with an
    open balance are owed — everything else is already paid and irrelevant here.

    Returns (None, None) if the file is missing or unreadable; the caller then
    shows "?" instead of claiming vendors are paid on stale/absent data.
    """
    if not path.exists():
        log.warning(
            "Bill Tracker not found at %s — Vendor Status will show '%s'. "
            "Run `sync-ap` to generate it.", path, VENDOR_UNKNOWN,
        )
        return None, None

    try:
        from openpyxl import load_workbook

        as_of = dt.datetime.fromtimestamp(path.stat().st_mtime)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["Bills"]
            rows = ws.iter_rows(min_row=2, values_only=True)  # row 1 = banner
            header = next(rows)
            idx = {name: i for i, name in enumerate(header) if name}
            need = ("Division", "Invoice #", "Bill Open Bal", "Vendor", "Bill #", "Bill Date")
            missing = [c for c in need if c not in idx]
            if missing:
                log.warning("Bill Tracker 'Bills' sheet missing columns %s", missing)
                return None, None

            seen: set = set()
            agg: Dict[str, List[Any]] = defaultdict(lambda: [0.0, 0, set()])
            for row in rows:
                if not any(row):
                    continue
                if row[idx["Division"]] not in DRAW_DIVISIONS:
                    continue
                invoice_num = str(row[idx["Invoice #"]] or "").strip()
                if not invoice_num:
                    continue
                try:
                    balance = float(row[idx["Bill Open Bal"]] or 0)
                except (TypeError, ValueError):
                    continue
                if balance <= 0:
                    continue
                vendor = row[idx["Vendor"]]
                bill_key = (invoice_num, vendor, row[idx["Bill #"]], row[idx["Bill Date"]])
                if bill_key in seen:
                    continue
                seen.add(bill_key)
                entry = agg[invoice_num]
                entry[0] += balance
                entry[1] += 1
                entry[2].add(vendor)
        finally:
            wb.close()

        result = {inv: (amt, bills, len(vendors)) for inv, (amt, bills, vendors) in agg.items()}
        log.info(
            "Vendor bills: %d MFD/CP invoices still carry unpaid bills (tracker as of %s)",
            len(result), as_of.strftime("%Y-%m-%d %H:%M"),
        )
        # AR runs after AP by design (`sync-all`). A tracker older than today
        # means that order was broken — say so at run time instead of letting a
        # day-old vendor column pass for current.
        if as_of.date() < dt.date.today():
            log.warning(
                "Bill Tracker is from %s, not today — the Vendor columns are "
                "%s old. Run `sync-ap` (or `sync-all`, which runs AP first) "
                "to refresh them.",
                as_of.strftime("%b %d"), _humanize_age(as_of),
            )
        return result, as_of
    except Exception as e:
        log.warning("Could not read Bill Tracker (%s): %s", path, e)
        return None, None


def _open_bills_for(
    invoice_num: str, vendor_map: Optional[Dict[str, Tuple[float, int, int]]]
) -> Tuple[Optional[int], Optional[float]]:
    """(open bill count, open $) still owed against one invoice's draw period."""
    if vendor_map is None:
        return None, None
    hit = vendor_map.get(str(invoice_num).strip())
    if not hit:
        return None, None
    amount, bills, _vendors = hit
    return bills, amount


def vendor_cells(
    division: str,
    invoice_num: str,
    vendor_map: Optional[Dict[str, Tuple[float, int, int]]],
    chains: Optional[Any] = None,
) -> Tuple[str, str, Optional[int], Optional[float], Optional[float]]:
    """The previous-draw block for one invoice.

    Returns (prev draw label, verdict, prev open bill count, prev open $,
    this draw's own open $).

    The verdict separates the two very different reasons a draw sits unpaid:

      * `PAY BILLS → unlock` — the previous draw WAS funded, but our vendors on
        it are still owed, so no unconditional waivers exist and the GC won't
        release this draw. **We** are the blocker, and the fix is ours.
      * `Waiting GC on prev` — the previous draw hasn't been funded either, so
        we couldn't have paid those vendors yet. The blocker is upstream.

    Collapsing those into one "unpaid bills" flag (as the first cut of this tab
    did) hides which of the two you're looking at, and they lead to opposite
    actions.
    """
    this_bills, this_amount = _open_bills_for(invoice_num, vendor_map)
    if division not in DRAW_DIVISIONS:
        return "", VENDOR_NA, None, None, None
    if chains is None:
        return "", VENDOR_UNKNOWN, None, None, this_amount

    outcome, prev = chains.previous_draw(invoice_num)
    if outcome == draw_chain.CHAIN_NOT_A_DRAW:
        return "", PREV_NOT_DRAW, None, None, this_amount
    if outcome == draw_chain.CHAIN_MULTI_CONTRACT:
        return "", PREV_MULTI, None, None, this_amount
    if outcome == draw_chain.CHAIN_FIRST_DRAW or prev is None:
        return "", PREV_FIRST, None, None, this_amount

    prev_num = prev["invoice_num"]
    if vendor_map is None:
        return prev_num, VENDOR_UNKNOWN, None, None, this_amount

    prev_bills, prev_amount = _open_bills_for(prev_num, vendor_map)
    if not prev["is_paid"]:
        # Show the bills anyway — they're what we'll owe once it does fund.
        return prev_num, PREV_WAITING_GC, prev_bills, prev_amount, this_amount
    if prev_bills:
        return prev_num, PREV_BLOCKED, prev_bills, prev_amount, this_amount
    return prev_num, PREV_CLEAR, None, None, this_amount


# ─────────────────────── sheet build ───────────────────────

class _Grid:
    """Maps logical column indices (the C_* constants) to physical ones.

    A tab can drop columns — the RP view hides the whole previous-draw block,
    which is meaningless there — but every row is still built at full width
    against the C_* constants. This does the projection in one place so no
    caller has to think about which physical column a field landed in.
    """

    def __init__(self, drop: Tuple[int, ...] = ()) -> None:
        self.visible = [i for i in range(len(COLUMNS)) if i not in drop]
        self._pos = {logical: n for n, logical in enumerate(self.visible)}

    def __contains__(self, logical: int) -> bool:
        return logical in self._pos

    def col(self, logical: int) -> int:
        """1-based physical column for a logical index."""
        return self._pos[logical] + 1

    def cell(self, ws: Worksheet, row: int, logical: int):
        """The cell for a logical column, or None when that column is hidden."""
        if logical not in self._pos:
            return None
        return ws.cell(row=row, column=self.col(logical))

    @property
    def width(self) -> int:
        return len(self.visible)


def _autofit(ws: Worksheet, grid: _Grid, last_row: int) -> None:
    """Size every column to its widest cell, in the sheet's own font.

    openpyxl has no real autofit — Excel computes widths at render time and
    openpyxl never renders. Measuring the strings we wrote is the honest
    approximation: character count scaled for the 12pt body (wider than the
    11pt default the width unit assumes), plus padding for the filter arrow.
    """
    for logical in grid.visible:
        name, floor_width, number_format = COLUMNS[logical]
        letter = get_column_letter(grid.col(logical))
        widest = len(name) + 3  # header text + room for the autofilter arrow
        for row in range(5, last_row + 1):
            value = ws.cell(row=row, column=grid.col(logical)).value
            if value is None:
                continue
            if isinstance(value, dt.date):
                rendered = 10                      # mm/dd/yyyy
            elif isinstance(value, float) and number_format and "$" in number_format:
                rendered = len(f"{value:,.2f}") + 2   # "$" plus separators
            else:
                rendered = len(str(value))
            widest = max(widest, rendered)
        width = max(floor_width, widest * BODY_PT / 11.0 + 1.5)
        ws.column_dimensions[letter].width = min(width, MAX_COL_WIDTH)


def _write_row(
    ws: Worksheet,
    grid: _Grid,
    row_num: int,
    values: List[Any],
    *,
    bold: bool,
    size: float = BODY_PT,
    fill: Optional[PatternFill] = None,
) -> None:
    for logical in grid.visible:
        cell = ws.cell(row=row_num, column=grid.col(logical), value=values[logical])
        number_format = COLUMNS[logical][2]
        if number_format:
            cell.number_format = number_format
        cell.font = Font(bold=bold, size=size)
        if fill:
            cell.fill = fill
    notes = grid.cell(ws, row_num, C_NOTES)
    if notes is not None:
        notes.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _grey_out_vendor_block(ws: Worksheet, grid: _Grid, row_num: int) -> None:
    """Mark the previous-draw cells as not-applicable on this row.

    Used wherever there is no draw chain to read — RP (no draws at all), non-draw
    invoices, first draws, and the multi-contract projects whose bills can't be
    attributed. A blank would read as "not looked up yet"; grey fill with darker
    grey text says the cell is intentionally dead. (the user)

    Only EMPTY cells get the "n/a" text, so a verdict already written there
    ("Multi-contract", "First draw") stays readable inside the grey block.
    """
    for logical in VENDOR_COLS:
        cell = grid.cell(ws, row_num, logical)
        if cell is None:
            continue
        if cell.value in (None, ""):
            cell.value = VENDOR_NA
        cell.number_format = "General"  # else "n/a" fights the $ / 0 formats
        cell.fill = _NA_FILL
        cell.font = Font(color=_NA_COLOR, italic=True, size=BODY_PT)
        cell.alignment = Alignment(horizontal="center")


def build_aging_sheet(
    ws: Worksheet,
    invoices: List[dict],
    *,
    today: dt.date,
    litigation_excluded: int,
    vendor_as_of: Optional[dt.datetime] = None,
    drop_columns: Tuple[int, ...] = (),
    title: str = "AR AGING",
    scope_note: str = "",
) -> None:
    """Write an aging tab.

    `invoices` are plain dicts (built by export_invoices_xlsx._aging_record) so
    this module stays free of Notion property plumbing.

    Layout: a header, an always-visible ALL CLIENTS total, then one bold summary
    row per parent client with its invoices grouped underneath and COLLAPSED
    (the owner's ask — Notion-style one-page scanning, drill in on demand).
    Summary-above-detail requires outlinePr.summaryBelow = False; without it
    Excel puts the collapse toggle on the row after the block and the grouping
    reads backwards.

    `drop_columns` hides logical columns entirely — the RP tab passes the whole
    previous-draw block, which has no meaning for a division that doesn't bill
    in draws.
    """
    grid = _Grid(drop_columns)
    shows_vendor_block = C_VSTATUS in grid

    # Group by parent client. Unresolved relations fall back to a stable label
    # rather than being dropped — an invoice with no parent is still money owed.
    by_parent: Dict[str, List[dict]] = defaultdict(list)
    for rec in invoices:
        by_parent[rec["parent"] or "(no client on file)"].append(rec)

    # ── title block ──
    ws.cell(
        row=1, column=1,
        value=f"{title} — as of {today.strftime('%b %d, %Y').upper()}",
    ).font = Font(bold=True, size=TITLE_PT)

    parts = [f"{len(invoices)} open invoices", f"litigation excluded ({litigation_excluded})"]
    if scope_note:
        parts.insert(0, scope_note)
    stale = vendor_as_of is None or vendor_as_of.date() < today
    if shows_vendor_block:
        # The vendor columns have a different clock from everything else on this
        # tab: they come from the AP tool's last run, not from this one. Say
        # which, in plain words, so nobody reads a day-old figure as current.
        if vendor_as_of is None:
            parts.append("Vendor bill status UNAVAILABLE — run `sync-ap`, then re-run this")
        elif vendor_as_of.date() < today:
            parts.append(
                f"⚠ VENDOR COLUMNS ARE {_humanize_age(vendor_as_of).upper()} OLD "
                f"(Bill Tracker last run {vendor_as_of.strftime('%b %d, %I:%M %p')}) — "
                f"run `sync-all`, which runs AP before AR"
            )
        else:
            parts.append(
                f"Vendor bill status current as of {vendor_as_of.strftime('%b %d, %I:%M %p')}"
            )
    subtitle_cell = ws.cell(row=2, column=1, value=" · ".join(parts))
    subtitle_cell.font = (
        Font(bold=True, color="C00000", size=BODY_PT)
        if (shows_vendor_block and stale)
        else Font(size=BODY_PT)
    )

    header_row = 4
    for logical in grid.visible:
        name = COLUMNS[logical][0]
        cell = ws.cell(row=header_row, column=grid.col(logical), value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=BODY_PT)
        # The five bucket headers run green→red so the severity scale is legible
        # from the header alone; everything else takes the standard blue.
        cell.fill = (
            BUCKET_HEADER_FILLS[BUCKET_COLS.index(logical)]
            if logical in BUCKET_COLS
            else _HEADER_FILL
        )
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_MEDIUM)
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── grand total (always visible, never grouped) ──
    # Roll the grand total up from the same per-invoice values the parent rows
    # use, so the top row always ties to the sum of the groups below it.
    grand = [0.0] * len(BUCKET_COLS)
    grand_vendor = 0.0
    grand_bills = 0
    grand_this = 0.0
    grand_invtotal = 0.0
    for rec in invoices:
        grand[bucket_index(rec["days_past_due"])] += rec["open_balance"] or 0.0
        grand_vendor += rec["vendor_amount"] or 0.0
        grand_bills += rec["vendor_bills"] or 0
        grand_this += rec["this_draw_amount"] or 0.0
        grand_invtotal += rec["total_amount"] or 0.0

    row_num = header_row + 1
    total_row: List[Any] = [""] * len(COLUMNS)
    total_row[C_LABEL] = f"ALL CLIENTS ({len(by_parent)})"
    for logical, amount in zip(BUCKET_COLS, grand):
        total_row[logical] = amount or None
    total_row[C_TOTAL] = sum(grand)
    total_row[C_INVTOTAL] = grand_invtotal or None
    total_row[C_VBILLS] = grand_bills or None
    total_row[C_VAMT] = grand_vendor or None
    total_row[C_THIS] = grand_this or None
    _write_row(ws, grid, row_num, total_row, bold=True, size=CLIENT_PT, fill=_GRAND_FILL)
    for logical in grid.visible:
        ws.cell(row=row_num, column=grid.col(logical)).border = Border(bottom=_THIN)
    row_num += 1

    # ── parent groups, biggest balance first ──
    def parent_total(records: List[dict]) -> float:
        return sum(r["open_balance"] or 0.0 for r in records)

    for parent in sorted(by_parent, key=lambda p: parent_total(by_parent[p]), reverse=True):
        records = sorted(
            by_parent[parent],
            key=lambda r: (r["due_date"] or dt.date.max, r["invoice_num"]),
        )
        buckets = [0.0] * len(BUCKET_COLS)
        vendor_sum = 0.0
        vendor_bills = 0
        this_sum = 0.0
        invtotal_sum = 0.0
        for rec in records:
            buckets[bucket_index(rec["days_past_due"])] += rec["open_balance"] or 0.0
            vendor_sum += rec["vendor_amount"] or 0.0
            vendor_bills += rec["vendor_bills"] or 0
            this_sum += rec["this_draw_amount"] or 0.0
            invtotal_sum += rec["total_amount"] or 0.0

        summary: List[Any] = [""] * len(COLUMNS)
        summary[C_LABEL] = parent
        summary[C_INV] = f"{len(records)} inv"
        for logical, amount in zip(BUCKET_COLS, buckets):
            summary[logical] = amount or None
        summary[C_TOTAL] = sum(buckets)
        summary[C_INVTOTAL] = invtotal_sum or None
        summary[C_VBILLS] = vendor_bills or None
        summary[C_VAMT] = vendor_sum or None
        summary[C_THIS] = this_sum or None
        # Client rows sit a point above the body — the owner reads this sheet
        # collapsed, so the client name is the line that has to carry.
        _write_row(ws, grid, row_num, summary, bold=True, size=CLIENT_PT, fill=_PARENT_FILL)
        for logical in grid.visible:
            ws.cell(row=row_num, column=grid.col(logical)).border = Border(top=_THIN)
        # A client with no MFD/CP work has no vendor answer either — grey the
        # block on the summary row too, or an all-RP client reads as "clear".
        if shows_vendor_block and not any(r["division"] in DRAW_DIVISIONS for r in records):
            _grey_out_vendor_block(ws, grid, row_num)
        row_num += 1

        for rec in records:
            detail: List[Any] = [""] * len(COLUMNS)
            # QBO memos carry hard line breaks ("… Draw 2026\n(Period: …)").
            # Left raw they render as one run-on line or a stray box, since this
            # column doesn't wrap — collapse to single-spaced text.
            label = " ".join((rec["memo"] or rec["project_num"] or "").split())
            detail[C_LABEL] = f"    {label}"[:120]
            detail[C_PROJ] = rec["project_num"]
            detail[C_INV] = rec["invoice_num"]
            detail[C_DATE] = rec["invoice_date"]
            detail[C_DUE] = rec["due_date"]
            detail[C_LIEN] = rec["lien"].label
            detail[BUCKET_COLS[bucket_index(rec["days_past_due"])]] = rec["open_balance"]
            detail[C_TOTAL] = rec["open_balance"]
            detail[C_INVTOTAL] = rec["total_amount"]
            detail[C_PREV] = rec["prev_draw"]
            detail[C_VSTATUS] = rec["vendor_status"]
            detail[C_VBILLS] = rec["vendor_bills"]
            detail[C_VAMT] = rec["vendor_amount"]
            detail[C_THIS] = rec["this_draw_amount"]
            detail[C_NOTES] = rec["notes"]
            detail[C_ACTION] = rec["last_action"]
            _write_row(ws, grid, row_num, detail, bold=False)

            # The invoice number opens the invoice in QBO. Putting the link on
            # the number rather than in its own column keeps the sheet narrow
            # and puts the click where the eye already is.
            if rec.get("qbo_link"):
                inv_cell = grid.cell(ws, row_num, C_INV)
                if inv_cell is not None:
                    inv_cell.hyperlink = rec["qbo_link"]
                    inv_cell.font = Font(color=_LINK_COLOR, underline="single", size=BODY_PT)

            # Tint only the ONE bucket cell carrying this invoice's balance, so
            # colour drifting rightward down the page = money getting older.
            slot = bucket_index(rec["days_past_due"])
            ws.cell(
                row=row_num, column=grid.col(BUCKET_COLS[slot])
            ).fill = BUCKET_CELL_FILLS[slot]

            # The lien cell carries the only hard expiry on this sheet, so it
            # gets the strongest cue: a missed or imminent notice deadline is a
            # right that disappears, not just money that is late.
            lien_cell = grid.cell(ws, row_num, C_LIEN)
            if lien_cell is not None:
                state = rec["lien"].state
                if state == lien_clock.STATE_PAST:
                    lien_cell.font = Font(bold=True, color="FFFFFF", size=BODY_PT)
                    lien_cell.fill = _LIEN_PAST_FILL
                elif state == lien_clock.STATE_URGENT:
                    lien_cell.font = Font(bold=True, color="922B21", size=BODY_PT)
                    lien_cell.fill = _LIEN_URGENT_FILL
                elif state == lien_clock.STATE_WATCH:
                    lien_cell.font = Font(color="8A5A00", size=BODY_PT)
                    lien_cell.fill = _LIEN_WATCH_FILL
                elif state == lien_clock.STATE_SENT:
                    lien_cell.font = Font(color="2E7D32", size=BODY_PT)
                elif state == lien_clock.STATE_RETAINAGE:
                    lien_cell.font = Font(color=_NA_COLOR, italic=True, size=BODY_PT)
                    lien_cell.fill = _NA_FILL

            if shows_vendor_block:
                verdict = rec["vendor_status"]
                if verdict in NO_CHAIN_VERDICTS:
                    _grey_out_vendor_block(ws, grid, row_num)
                else:
                    status_cell = grid.cell(ws, row_num, C_VSTATUS)
                    if verdict == PREV_BLOCKED:
                        # The one verdict that is ours to act on — pay these,
                        # get the waivers, unlock this draw.
                        status_cell.font = Font(bold=True, color="C00000", size=BODY_PT)
                        status_cell.fill = _BLOCKED_FILL
                    elif verdict == PREV_WAITING_GC:
                        status_cell.font = Font(color="B06000", size=BODY_PT)
                    elif verdict == PREV_CLEAR:
                        status_cell.font = Font(color="2E7D32", size=BODY_PT)

            ws.row_dimensions[row_num].outlineLevel = 1
            ws.row_dimensions[row_num].hidden = True  # collapsed by default
            row_num += 1

    last_data_row = row_num - 1

    # Open Balance data bar, scaled per row to that row's Total Amount (the user
    # 2026-08-11 — "a bar of how much open balance is to the total"). The bar max
    # is a FORMULA pointing at the Total Amount cell on the same row (column
    # absolute, row relative), so Excel adjusts it down every row: a fully-open
    # invoice fills the cell, a partly-collected one shows a short bar. Applies
    # to the grand row through the last invoice — summary rows scale to their
    # client's own open/total, so the bars stay coherent when collapsed.
    if C_INVTOTAL in grid and last_data_row >= header_row + 1:
        open_col = get_column_letter(grid.col(C_TOTAL))
        tot_col = get_column_letter(grid.col(C_INVTOTAL))
        first = header_row + 1
        bar_rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="formula", end_value=f"${tot_col}{first}",
            color="FF5B9BD5", showValue=True, minLength=0, maxLength=100,
        )
        ws.conditional_formatting.add(
            f"{open_col}{first}:{open_col}{last_data_row}", bar_rule
        )

    # Legend, below the data so it never pushes the numbers down. Colour that
    # needs explaining is colour that failed, but the grey block is a deliberate
    # "don't read this" and that one is worth spelling out.
    legend_row = row_num + 1
    ws.cell(row=legend_row, column=1, value="KEY").font = Font(bold=True, size=BODY_PT)
    for slot, logical in enumerate(BUCKET_COLS):
        cell = ws.cell(row=legend_row, column=grid.col(logical), value=COLUMNS[logical][0])
        cell.fill = BUCKET_CELL_FILLS[slot]
        cell.font = Font(size=BODY_PT)
        cell.alignment = Alignment(horizontal="center")

    lien_legend = grid.cell(ws, legend_row, C_LIEN)
    if lien_legend is not None:
        lien_legend.value = "PAST DUE"
        lien_legend.fill = _LIEN_PAST_FILL
        lien_legend.font = Font(bold=True, color="FFFFFF", size=BODY_PT)
        lien_legend.alignment = Alignment(horizontal="center")

    notes: List[str] = [
        "Aged by due date. Invoice # links to the invoice in QBO.",
        "Lien = the date a notice must be MAILED by (Tex. Prop. Code Ch. 53, first-tier sub): "
        "CP/MFD the 15th of the 3rd month after the work month, RP the 15th of the 2nd; "
        "rolled BACK off weekends. Work month = invoice month. Retainage runs its own track and is not dated here.",
        "The lien column is a deadline watchlist, not legal advice — confirm project type, parcel and owning entity before sending anything.",
    ]
    if shows_vendor_block:
        blocked_cell = grid.cell(ws, legend_row, C_VSTATUS)
        blocked_cell.value = PREV_BLOCKED
        blocked_cell.fill = _BLOCKED_FILL
        blocked_cell.font = Font(bold=True, color="C00000", size=BODY_PT)
        na_cell = grid.cell(ws, legend_row, C_VBILLS)
        na_cell.value = VENDOR_NA
        na_cell.fill = _NA_FILL
        na_cell.font = Font(color=_NA_COLOR, italic=True, size=BODY_PT)
        na_cell.alignment = Alignment(horizontal="center")
        notes += [
            f"'{PREV_BLOCKED}' = the previous draw was funded but our vendors on it are still owed — no unconditional waivers, so the GC won't release this draw. That one is ours to fix.",
            f"'{PREV_WAITING_GC}' = the previous draw hasn't been funded either, so the hold-up is upstream of us.  ·  '{PREV_CLEAR}' = previous draw funded and its vendors paid.",
            f"'{PREV_MULTI}' = the project runs parallel contracts and bills carry a project #, not a contract, so the previous draw can't be identified yet.  ·  'n/a' = RP, which doesn't bill in draws.",
        ]
    else:
        notes.append(
            "RP doesn't bill in draws, so the previous-draw and vendor-bill columns are omitted here. "
            "See the AR Aging tab for CP and MFD."
        )
    for offset, text in enumerate(notes):
        ws.cell(row=legend_row + 1 + offset, column=1, value=text).font = Font(
            italic=True, color=_NA_COLOR, size=BODY_PT
        )

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = False
    # Filter spans the header + data only — never the legend rows below it.
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(grid.width)}{last_data_row}"
    )
    _autofit(ws, grid, last_data_row)
    ws.sheet_view.showGridLines = True
