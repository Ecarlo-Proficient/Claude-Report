"""
cash_flow.py — a note-driven payment forecast for the collections file.

Why this exists (the user 2026-08-12):
    The collections notes already carry the real intel on WHEN money is coming
    ("payment promise 8/13", "paying this friday", "get check friday, deposit
    Monday"). That beats a generic QBO forecast. This turns those notes into a
    cash-flow forecast: two tabs in Open_Invoices.xlsx — a weekly list and a
    month calendar grid — plus a "no date / review" section for the vaguer signals.

What it is NOT:
    A parser of free text will never be perfect, so the rule is: only a CLEAR,
    unconditional payment date lands on the calendar; everything else with a
    payment promise but a fuzzy/absent date goes to the review section (nothing is
    silently dropped), and notes with no inflow promise at all (follow-ups,
    disputes, or "pay to <vendor>" OUTFLOWS) are excluded. Every row shows its
    source note so the owner can sanity-check the read.

Amounts are the invoice open balance — an EXPECTED inflow, never a guarantee.
"""
from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# The absorb/preserve stamp the note carries: " – Name, M/D". Strip it before
# parsing so the sync date (e.g. 8/12) is never read as a payment date.
_STAMP_RE = re.compile(r"\s+[-–]\s+\w+,\s*\d{1,2}/\d{1,2}\s*$")

_WEEKDAYS = {
    "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
    "friday": 4, "saturday": 5, "sunday": 6,
}

# A positive inflow promise (THEY will pay US). "payment" alone is deliberately
# NOT here — "asking for payment" / "send payment reminder" are chases, not promises.
_PROMISE_RE = re.compile(
    r"\b(promise\w*|paying|will\s+pay|pays?|paid|deposit\w*|wir(?:e|ed|ing)|remit\w*)\b",
    re.I,
)
# A payment CHECK as a noun (not "check PO" / "check if" / "checked", which are verbs).
_CHECK_NOUN_RE = re.compile(
    r"\b(?:courtesy|cashier'?s?|company|personal|the|a|pick\s?up|get|cut|send|mail|drop|receiv\w*|deposit)\s+check\b"
    r"|\bcheck\s*(?:#|friday|monday|tuesday|wednesday|thursday|ready|cut|by|on|for)",
    re.I,
)
# Money going OUT (draw-funding "pay Hope", pay a sub/vendor) — not an inflow.
_OUTFLOW_RE = re.compile(r"\bpay\s+(?:to\b|out\b|hope\b|the\s+sub\b|sub\b|vendor\b|\$)", re.I)
# Conditional / uncertain / stale — keeps a promise off the calendar, into review.
_UNCERTAIN_RE = re.compile(
    r"\b(if\s+not|if\s+no|might|maybe|possibl\w*|no\s+update|hope\s+to|should|tentativ\w*|dispute\w*|no\s+response)\b",
    re.I,
)

SCHEDULED = "scheduled"
REVIEW = "review"


@dataclass
class PaymentSignal:
    kind: str                     # SCHEDULED | REVIEW
    date: Optional[dt.date]       # the expected date (None when undated)
    reason: str                   # why it's in review ("" for scheduled)
    note: str                     # the cleaned note (stamp stripped), for display


def strip_stamp(note: str) -> str:
    return _STAMP_RE.sub("", note or "").strip()


def _next_weekday(today: dt.date, target: int, *, this_wk: bool, next_wk: bool) -> dt.date:
    delta = (target - today.weekday()) % 7
    if delta == 0:
        delta = 7  # "friday" on a Friday means the coming one, not today
    d = today + dt.timedelta(days=delta)
    if next_wk:
        d += dt.timedelta(days=7)
    return d


def _confident_date(low: str, today: dt.date) -> Optional[dt.date]:
    """The single expected date from a note, or None if there's no clear one.

    Collects every explicit M/D, weekday, and today/tomorrow, then picks: the
    deposit day if 'deposit' is named (that's when cash lands), else the earliest
    future date."""
    cand: List[dt.date] = []
    for m in re.finditer(r"\b(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?\b", low):
        mo, d = int(m.group(1)), int(m.group(2))
        try:
            dd = dt.date(today.year, mo, d)
        except ValueError:
            continue
        if (today - dd).days > 180:      # a date far in the past → next year
            try:
                dd = dt.date(today.year + 1, mo, d)
            except ValueError:
                pass
        cand.append(dd)
    for name, idx in _WEEKDAYS.items():
        if re.search(rf"\b{name}\b", low):
            cand.append(_next_weekday(
                today, idx,
                this_wk=f"this {name}" in low,
                next_wk=f"next {name}" in low,
            ))
    if re.search(r"\btomorrow\b", low):
        cand.append(today + dt.timedelta(days=1))
    # "today" is deliberately NOT a date candidate: in real notes it almost always
    # marks when an ACTION happened ("OM texted today", "texting today"), not when
    # payment lands, so it produced false same-day forecasts. A genuinely same-day
    # payment is imminent and known anyway; if undated it falls to review.
    if not cand:
        return None
    future = [c for c in cand if c >= today]
    pool = future or cand
    return max(pool) if "deposit" in low else min(pool)


def classify_note(note: str, today: dt.date) -> Optional[PaymentSignal]:
    """Classify one note into a payment signal, or None to exclude it.

    Excluded = no inflow promise (chase / follow-up / dispute / outflow).
    SCHEDULED = an inflow promise with a clear, unconditional, future date.
    REVIEW = an inflow promise but the date is conditional / stale / absent.
    """
    text = strip_stamp(note)
    low = text.lower()
    if not text:
        return None
    if _OUTFLOW_RE.search(low):
        return None  # "pay to Hope $158k" etc. — money out, not in
    has_promise = bool(_PROMISE_RE.search(low) or _CHECK_NOUN_RE.search(low)
                       or "payment promise" in low)
    if not has_promise:
        return None
    date = _confident_date(low, today)
    if _UNCERTAIN_RE.search(low):
        return PaymentSignal(REVIEW, date, "conditional / uncertain", text)
    if date is None:
        return PaymentSignal(REVIEW, None, "promised, no date", text)
    if date < today:
        return PaymentSignal(REVIEW, date, "date passed, no update", text)
    return PaymentSignal(SCHEDULED, date, "", text)


# ─────────────────────── sheet building ───────────────────────
# Styling mirrors the aging tabs (Font size 12 body; blue header banding).

_PT = 12
_TITLE_PT = 14
_MONEY = '"$"#,##0.00'
_MONEY0 = '"$"#,##0'

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=_PT)
_WEEK_FILL = PatternFill("solid", fgColor="DCE6F1")      # week banding (matches client rows)
_TOTAL_FILL = PatternFill("solid", fgColor="BFD3E6")     # grand/subtotal
_REVIEW_FILL = PatternFill("solid", fgColor="FDF2E0")    # amber for the review block
_PAY_FILL = PatternFill("solid", fgColor="E2EFDA")       # a day / row with money (soft green)
_WEEKEND_FILL = PatternFill("solid", fgColor="F2F2F2")   # calendar weekend
_TODAY_FILL = PatternFill("solid", fgColor="FFF2CC")     # calendar today
_GRID = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_GRID, right=_GRID, top=_GRID, bottom=_GRID)
_MEDIUM = Side(style="medium", color="000000")


def _friendly(d: dt.date) -> str:
    return d.strftime("%a %b %d").replace(" 0", " ")


def _split_signals(records: List[dict], today: dt.date):
    scheduled: List[Tuple[dict, PaymentSignal]] = []
    review: List[Tuple[dict, PaymentSignal]] = []
    for rec in records:
        s = classify_note(rec.get("notes", ""), today)
        if s is None:
            continue
        (scheduled if s.kind == SCHEDULED else review).append((rec, s))
    scheduled.sort(key=lambda x: (x[1].date, -(x[0].get("open_balance") or 0.0)))
    review.sort(key=lambda x: (x[1].date or dt.date.max, -(x[0].get("open_balance") or 0.0)))
    return scheduled, review


# ── Tab 1: weekly forecast list ──

_LIST_COLS = [
    ("Expected", 13, "mm/dd/yyyy"),
    ("Client", 30, None),
    ("Project #", 12, None),
    ("Invoice #", 11, None),
    ("Amount", 15, _MONEY),
    ("Cumulative", 15, _MONEY),
    ("Note", 60, None),
]


def _week_start(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())  # Monday


def build_cash_flow_list(ws: Worksheet, records: List[dict], today: dt.date) -> None:
    scheduled, review = _split_signals(records, today)
    sched_total = sum((r.get("open_balance") or 0.0) for r, _ in scheduled)
    review_total = sum((r.get("open_balance") or 0.0) for r, _ in review)

    ws.cell(row=1, column=1,
            value=f"CASH FLOW FORECAST — as of {today.strftime('%b %d, %Y').upper()}"
            ).font = Font(bold=True, size=_TITLE_PT)
    ws.cell(row=2, column=1, value=(
        f"Expected inflows read from collections notes — NOT guaranteed. "
        f"{len(scheduled)} scheduled (${sched_total:,.0f}) · "
        f"{len(review)} promised but need a date (${review_total:,.0f})."
    )).font = Font(size=_PT)

    header_row = 4
    for i, (name, width, _fmt) in enumerate(_LIST_COLS, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=_MEDIUM)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 24
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    def money(cell, val, fmt=_MONEY):
        cell.value = val
        cell.number_format = fmt

    row = header_row + 1
    running = 0.0
    by_week: Dict[dt.date, List[Tuple[dict, PaymentSignal]]] = defaultdict(list)
    for rec, s in scheduled:
        by_week[_week_start(s.date)].append((rec, s))

    for wk in sorted(by_week):
        wk_end = wk + dt.timedelta(days=6)
        wk_rows = by_week[wk]
        wk_total = sum((r.get("open_balance") or 0.0) for r, _ in wk_rows)
        hdr = ws.cell(row=row, column=1,
                      value=f"Week of {_friendly(wk)}  –  {_friendly(wk_end)}")
        hdr.font = Font(bold=True, size=_PT)
        for col in range(1, len(_LIST_COLS) + 1):
            ws.cell(row=row, column=col).fill = _WEEK_FILL
        money(ws.cell(row=row, column=5), wk_total)
        ws.cell(row=row, column=5).font = Font(bold=True, size=_PT)
        row += 1
        for rec, s in wk_rows:
            running += rec.get("open_balance") or 0.0
            ws.cell(row=row, column=1, value=s.date).number_format = "mm/dd/yyyy"
            ws.cell(row=row, column=2, value=rec.get("parent") or "")
            ws.cell(row=row, column=3, value=rec.get("project_num") or "")
            ws.cell(row=row, column=4, value=str(rec.get("invoice_num") or ""))
            money(ws.cell(row=row, column=5), rec.get("open_balance") or 0.0)
            money(ws.cell(row=row, column=6), running)
            note_cell = ws.cell(row=row, column=7, value=s.note)
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            for col in range(1, len(_LIST_COLS) + 1):
                ws.cell(row=row, column=col).font = Font(size=_PT)
            row += 1

    # Grand scheduled total
    if scheduled:
        ws.cell(row=row, column=2, value="SCHEDULED TOTAL").font = Font(bold=True, size=_PT)
        money(ws.cell(row=row, column=5), sched_total)
        money(ws.cell(row=row, column=6), running)
        for col in range(1, len(_LIST_COLS) + 1):
            ws.cell(row=row, column=col).fill = _TOTAL_FILL
            ws.cell(row=row, column=col).font = Font(bold=True, size=_PT)
        row += 1

    # ── review block ──
    row += 1
    rhdr = ws.cell(row=row, column=1,
                   value="PROMISED — NEEDS A DATE (conditional, stale, or no day given)")
    rhdr.font = Font(bold=True, color="8A5A00", size=_PT)
    for col in range(1, len(_LIST_COLS) + 1):
        ws.cell(row=row, column=col).fill = _REVIEW_FILL
    row += 1
    for rec, s in review:
        ws.cell(row=row, column=1,
                value=(s.date if s.date else None)).number_format = "mm/dd/yyyy"
        ws.cell(row=row, column=2, value=rec.get("parent") or "")
        ws.cell(row=row, column=3, value=rec.get("project_num") or "")
        ws.cell(row=row, column=4, value=str(rec.get("invoice_num") or ""))
        money(ws.cell(row=row, column=5), rec.get("open_balance") or 0.0)
        ws.cell(row=row, column=6, value=s.reason).font = Font(italic=True, color="8A5A00", size=_PT)
        nc = ws.cell(row=row, column=7, value=s.note)
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        for col in (1, 2, 3, 4, 5):
            ws.cell(row=row, column=col).font = Font(size=_PT)
        row += 1
    if review:
        ws.cell(row=row, column=2, value="REVIEW TOTAL").font = Font(bold=True, size=_PT)
        money(ws.cell(row=row, column=5), review_total)
        for col in range(1, len(_LIST_COLS) + 1):
            ws.cell(row=row, column=col).fill = _REVIEW_FILL
            if col in (2, 5):
                ws.cell(row=row, column=col).font = Font(bold=True, size=_PT)

    ws.sheet_view.showGridLines = False


# ── Tab 2: month calendar grid (rolling 6 weeks from this week's Monday) ──

_CAL_WEEKS = 6
_DAY_NAMES = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")


def build_cash_flow_calendar(ws: Worksheet, records: List[dict], today: dt.date) -> None:
    scheduled, _review = _split_signals(records, today)
    by_day: Dict[dt.date, List[Tuple[dict, PaymentSignal]]] = defaultdict(list)
    for rec, s in scheduled:
        by_day[s.date].append((rec, s))

    ws.cell(row=1, column=1,
            value=f"PAYMENT CALENDAR — {_CAL_WEEKS} weeks from {today.strftime('%b %d, %Y')}"
            ).font = Font(bold=True, size=_TITLE_PT)
    ws.cell(row=2, column=1, value=(
        "Scheduled inflows from notes, on the day the money is expected. "
        "Undated / conditional promises live on the Cash Flow tab, not here."
    )).font = Font(size=_PT)

    header_row = 4
    for i, name in enumerate(_DAY_NAMES, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = 20

    start = _week_start(today)
    r = header_row + 1
    for wk in range(_CAL_WEEKS):
        ws.row_dimensions[r].height = 54
        for col in range(7):
            day = start + dt.timedelta(days=wk * 7 + col)
            cell = ws.cell(row=r, column=col + 1)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            hits = by_day.get(day)
            head = day.strftime("%b %d").replace(" 0", " ")
            if hits:
                amt = sum((rec.get("open_balance") or 0.0) for rec, _ in hits)
                names = ", ".join(sorted({(rec.get("parent") or "")[:14] for rec, _ in hits}))
                cell.value = f"{head}\n${amt:,.0f}  ({len(hits)})\n{names}"
                cell.fill = _PAY_FILL
                cell.font = Font(size=_PT, bold=True)
            else:
                cell.value = head
                cell.font = Font(size=_PT, color="909090")
                if col >= 5:  # weekend
                    cell.fill = _WEEKEND_FILL
            if day == today:
                cell.fill = _TODAY_FILL
        r += 1

    # A small legend / total under the grid.
    total = sum((rec.get("open_balance") or 0.0) for rec, _ in scheduled)
    ws.cell(row=r + 1, column=1,
            value=f"Scheduled in view: ${total:,.0f} across {len(scheduled)} invoice(s).").font = (
        Font(italic=True, color="808080", size=_PT))
    ws.sheet_view.showGridLines = False


def build_cash_flow_sheets(wb, records: List[dict], today: dt.date) -> None:
    """Add the two cash-flow tabs to the workbook."""
    build_cash_flow_list(wb.create_sheet("Cash Flow"), records, today)
    build_cash_flow_calendar(wb.create_sheet("Pay Calendar"), records, today)
