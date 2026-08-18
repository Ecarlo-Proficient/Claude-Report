#!/usr/bin/env python3
"""
sub_loc_report.py - Subcontractor line-of-credit (LOC) float model.

The question (the user 2026-07-17): if every subcontractor payment were funded
by a line of credit and repaid when the client pays us for that work, how big
a LOC do we truly need, and how long is our money out before it comes back?

MODEL (all four choices are the user's, 2026-07-17):
  • Sub bill = a QBO Bill whose memo/PrivateNote contains "sub" (same rule as
    bill-tracker's is_sub_bill). Worked at the LINE level - each line carries
    its project on CustomerRef; one bill can span several projects.
  • DRAW (money out) = when the sub is actually PAID - the QBO BillPayment
    date - allocated across the bill's lines pro-rata by line amount.
  • REPAY (money in) = when the client pays us - the QBO customer Payment
    date, mapped to a project through the invoice it was applied to.
  • Matching = per-project FIFO: a client payment repays the oldest still-out
    sub draws on that project first, only up to what was drawn (the margin
    stays as profit and does NOT pay down the LOC).
  • Window = first Friday of the month 3 months back → today; balance starts 0.

Running LOC balance = cumulative draws − cumulative applied repayments over
time; its PEAK is the LOC you truly need. Averages: amount-weighted draw→repay
lag, average draw, average repayment.

READ-ONLY against QBO. Output: ~/Documents/CompanyHealth/Sub LOC Report.xlsx
(chmod 600). One Touch ID per run.

USAGE
  python3 one-offs/sub_loc_report.py
  python3 one-offs/sub_loc_report.py --months 3
  python3 one-offs/sub_loc_report.py --start 2026-04-03 --out /path/x.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import stat
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import qbo_api
from shared import paths

from shared import sub_loc as sl

DEFAULT_OUTPUT = paths.companyhealth_sources_dir() / "Sub LOC Report.xlsx"


# ────────────────────────── Excel ──────────────────────────

CUR = '#,##0'
_NAVY = PatternFill("solid", fgColor="1F3864")
_ZEBRA = PatternFill("solid", fgColor="EEF3FA")
_DRAW_FILL = PatternFill("solid", fgColor="FCE4D6")     # money out (peach)
_REPAY_FILL = PatternFill("solid", fgColor="E2EFDA")    # money in (green)
_PEAK_FILL = PatternFill("solid", fgColor="C00000")
_WHITE = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D7E5")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, headers, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center",
                                   wrap_text=True)
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"


_DIV_COLOR = {"MFD": "305496", "CP": "1F6B4C", "RP": "7030A0", "Other": "808080"}


def _write_division_kpi(wb, divisions: dict) -> None:
    """Per-division LOC KPI - judge the LOC need by division at a glance."""
    ws = wb.create_sheet("By Division")
    ws.sheet_view.showGridLines = False
    _hdr(ws, ["DIVISION", "PEAK LOC NEEDED $", "PEAK DATE", "DRAWN $",
              "PREFUNDED $", "REPAID $", "OUTSTANDING $", "AVG FLOAT (days)",
              "# DRAWS"],
         [11, 18, 13, 15, 14, 14, 15, 16, 9])
    order = sorted(divisions.items(), key=lambda kv: -kv[1]["peak"])
    for i, (name, d) in enumerate(order):
        ws.append([name, round(d["peak"]),
                   d["peak_date"].isoformat() if d["peak_date"] else "-",
                   round(d["drawn"]), round(d["prefunded"]), round(d["repaid"]),
                   round(d["outstanding"]), round(d["avg_lag"], 1), d["n_draws"]])
        r = ws.max_row
        for c in range(1, 10):
            ws.cell(r, c).border = _BORDER
            if i % 2:
                ws.cell(r, c).fill = _ZEBRA
        dc = ws.cell(r, 1)
        dc.fill = PatternFill("solid", fgColor=_DIV_COLOR.get(name, "808080"))
        dc.font = Font(bold=True, color="FFFFFF")
        for c in (2, 4, 5, 6, 7):
            ws.cell(r, c).number_format = CUR
        ws.cell(r, 2).font = Font(bold=True)
        if d["outstanding"] > 0:
            ws.cell(r, 7).font = Font(bold=True, color="9C0006")
    last = ws.max_row
    # total row
    if order:
        ws.append(["TOTAL", round(sum(d["peak"] for _, d in order)), "",
                   round(sum(d["drawn"] for _, d in order)),
                   round(sum(d["prefunded"] for _, d in order)),
                   round(sum(d["repaid"] for _, d in order)),
                   round(sum(d["outstanding"] for _, d in order)), "",
                   sum(d["n_draws"] for _, d in order)])
        tr = ws.max_row
        for c in range(1, 10):
            ws.cell(tr, c).fill = PatternFill("solid", fgColor="D9E1F2")
            ws.cell(tr, c).border = _BORDER
            ws.cell(tr, c).font = Font(bold=True)
        for c in (2, 4, 5, 6, 7):
            ws.cell(tr, c).number_format = CUR
        # note: the sum of per-division peaks ≥ the company peak (peaks can fall
        # on different days) - it's an upper bound if each division held its own
        # LOC. The company-wide peak is on the Summary sheet.
    for col in (2, 7):     # data bars on peak + outstanding
        ws.conditional_formatting.add(
            f"{get_column_letter(col)}2:{get_column_letter(col)}{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color="F8696B" if col == 7 else "8EAADB", showValue=True))
    ws.append([])
    ws.append(["Note: division peaks can fall on different days, so they sum to "
               "an upper bound (each division on its own LOC). The single "
               "company-wide LOC need is on the Summary sheet."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="808080")
    ws.sheet_properties.tabColor = "305496"


def write_workbook(out: Path, events, summary, projects, start, today):
    wb = Workbook()

    # ── Summary ──
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 3
    s.column_dimensions["B"].width = 42
    s.column_dimensions["C"].width = 20
    s.append(["", "SUBCONTRACTOR LOC - float model"])
    s.merge_cells("B1:C1")
    s.cell(1, 2).font = Font(bold=True, size=15, color="FFFFFF")
    for c in (1, 2, 3):
        s.cell(1, c).fill = _NAVY
    s.row_dimensions[1].height = 32
    s.append(["", f"Window {start:%Y-%m-%d} → {today:%Y-%m-%d}  ·  sub bills paid "
              f"from LOC, repaid when the client pays  ·  per-project FIFO"])
    s.cell(2, 2).font = Font(italic=True, color="595959")
    s.merge_cells("B2:C2")

    def line(label, value, money=True, big=False, bad=False):
        s.append(["", label, value])
        r = s.max_row
        s.cell(r, 2).font = Font(bold=big, size=12 if big else 11)
        vc = s.cell(r, 3)
        vc.font = Font(bold=True, size=13 if big else 11,
                       color="C00000" if bad else ("1F6B4C" if big else "000000"))
        if money:
            vc.number_format = CUR
        for c in (2, 3):
            s.cell(r, c).border = _BORDER
            if big:
                s.cell(r, c).fill = PatternFill("solid", fgColor="FCE4E4" if bad
                                                else "E7F2E7")

    s.append([])
    line("LOC you truly need (peak balance)", round(summary["peak"]), big=True, bad=True)
    pk = summary["peak_date"]
    line("…reached on", pk.isoformat() if pk else "-", money=False)
    s.append([])
    line("Total drawn (paid to subs)", round(summary["total_drawn"]))
    line("…of which prefunded by the client first", round(summary["prefunded"]))
    line("Total repaid (client → LOC)", round(summary["total_repaid"]))
    line("Still outstanding (not yet repaid)", round(summary["outstanding"]), bad=True)
    s.append([])
    line("Avg days our cash is out (draw→repay)", round(summary["avg_lag"], 1),
         money=False, big=True)
    line("Avg draw (paid to a sub)", round(summary["avg_draw"]))
    line("Avg repayment chunk", round(summary["avg_repay"]))
    line("# draws", summary["n_draws"], money=False)
    line("# repayment chunks", summary["n_repay_chunks"], money=False)
    s.sheet_properties.tabColor = "1F3864"

    # ── By Division (KPI) ──
    _write_division_kpi(wb, summary.get("divisions") or {})

    # ── Ledger (running balance) ──
    lg = wb.create_sheet("Ledger")
    _hdr(lg, ["DATE", "TYPE", "PROJECT #", "PARTY", "DRAW OUT $", "REPAY IN $",
              "RUNNING LOC $", "LAG (days)", "NOTE", "REIMBURSING INVOICE #",
              "INVOICE PAID DATE"],
         [12, 8, 12, 32, 14, 14, 16, 11, 16, 20, 15])
    first = 2
    peak_marked = False
    for i, e in enumerate(events):
        reimb = e.get("reimb") or []
        if e["type"] == "DRAW":
            invs = sorted({str(x[0]) for x in reimb if x[0]})
            inv_str = ", ".join(invs)
            paid = max((x[1] for x in reimb), default=None)
            paid_str = paid.isoformat() if paid else ("" if not e["out"] else "still out")
        else:
            inv_str = e.get("invoice", "")
            paid_str = e["date"].isoformat()
        lg.append([e["date"].isoformat(), e["type"], e["project"], e["party"],
                   round(e["out"]) if e["out"] else "",
                   round(e["inn"]) if e["inn"] else "",
                   round(e["balance"]), e["lag"] if e["lag"] is not None else "",
                   e["note"], inv_str, paid_str])
        r = lg.max_row
        for c in range(1, 12):
            lg.cell(r, c).border = _BORDER
            if i % 2:
                lg.cell(r, c).fill = _ZEBRA
        lg.cell(r, 2).fill = _DRAW_FILL if e["type"] == "DRAW" else _REPAY_FILL
        lg.cell(r, 2).font = Font(bold=True,
                                  color="9C6500" if e["type"] == "DRAW" else "1F6B4C")
        for c in (5, 6, 7):
            lg.cell(r, c).number_format = CUR
        if (not peak_marked and e["date"] == summary["peak_date"]
                and round(e["balance"]) == round(summary["peak"])):
            lg.cell(r, 7).fill = _PEAK_FILL
            lg.cell(r, 7).font = Font(bold=True, color="FFFFFF")
            peak_marked = True
    last = lg.max_row
    lg.conditional_formatting.add(
        f"G{first}:G{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="8EAADB", showValue=True))
    lg.auto_filter.ref = f"A1:K{last}"
    # legend
    lg.append([])
    lg.append(["Legend:  prefunded = this draw was covered by client cash already "
               "received for the same project+draw-month (no LOC needed).  "
               "surplus = the client paid more than we had drawn that period "
               "(our margin / advance toward the next draw).  REIMBURSING INVOICE "
               "= the client invoice(s) whose payment repaid this draw."])
    lg.cell(lg.max_row, 1).font = Font(italic=True, color="808080")
    lg.sheet_properties.tabColor = "C00000"

    # ── Per-Project ──
    pp = wb.create_sheet("Per-Project")
    _hdr(pp, ["PROJECT #", "DRAWN (paid subs) $", "REPAID $",
              "OUTSTANDING $", "AVG DAYS draw→repay"], [14, 20, 16, 16, 20])
    for i, p in enumerate(projects):
        pp.append([p["project"], round(p["drawn"]), round(p["repaid"]),
                   round(p["outstanding"]), round(p["avg_lag"], 1)])
        r = pp.max_row
        for c in range(1, 6):
            pp.cell(r, c).border = _BORDER
            if i % 2:
                pp.cell(r, c).fill = _ZEBRA
        for c in (2, 3, 4):
            pp.cell(r, c).number_format = CUR
        if p["outstanding"] > 0:
            pp.cell(r, 4).font = Font(bold=True, color="9C0006")
    last = pp.max_row
    for col in (2, 4):
        pp.conditional_formatting.add(
            f"{get_column_letter(col)}2:{get_column_letter(col)}{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color="F8696B" if col == 4 else "8EAADB", showValue=True))
    pp.auto_filter.ref = f"A1:E{last}"
    pp.sheet_properties.tabColor = "548235"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    os.chmod(out, stat.S_IRUSR | stat.S_IWUSR)


# ────────────────────────── main ──────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="Subcontractor LOC float model")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument("--months", type=int, default=3,
                    help="months back to the first Friday (default 3)")
    ap.add_argument("--start", type=str, help="override start date YYYY-MM-DD")
    args = ap.parse_args()

    today = dt.date.today()
    start = sl._parse(args.start) if args.start else sl.first_friday_months_back(today, args.months)
    lookback = sl.first_friday_months_back(today, sl.LOOKBACK_MONTHS).isoformat()

    print("\n  SUBCONTRACTOR LOC - float model")
    print("  " + "─" * 55)
    print(f"  window {start} → {today}")

    access, cid = qbo_api.load_credentials()
    print("  sub bill lines …")
    sub_lines = sl.build_sub_bill_lines(access, cid, lookback)
    print(f"    {len(sub_lines)} sub bill(s)")
    print("  invoice → project map …")
    inv_proj = sl.build_invoice_meta(access, cid, lookback)
    print("  draws (sub payments) …")
    draws = sl.collect_draws(access, cid, sub_lines, start)
    print(f"    {len(draws)} draw line-event(s), ${sum(d['amount'] for d in draws):,.0f}")
    print("  repayments (client payments) …")
    repays = sl.collect_repays(access, cid, inv_proj, start)
    print(f"    {len(repays)} repay event(s), ${sum(r['amount'] for r in repays):,.0f}")

    events, summary = sl.run_fifo(draws, repays)
    projects = sl.per_project(events)
    print(f"\n  PEAK LOC needed: ${summary['peak']:,.0f} on {summary['peak_date']}")
    print(f"  avg draw→repay: {summary['avg_lag']:.1f} days · "
          f"still out ${summary['outstanding']:,.0f}")

    write_workbook(args.out, events, summary, projects, start, today)
    print(f"\n  ✓ {args.out}  (chmod 600)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
