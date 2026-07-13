#!/usr/bin/env python3
"""
qbo_recode_review.py  —  Human-in-the-loop job-cost re-coder / consistency auditor for QBO
(Supersedes qbo_recode_jobcosts.py — you can delete that one.)

THE CHECK (per cost line item)
  A line is OK only if ALL THREE are true:
    1. the LINE DESCRIPTION contains a project #            (e.g. MFD1234 / RP7186 / CP585)
    2. the line's Customer:Project number == that project #
    3. the line's Class == the division the project # starts with
         MFD -> Multi Family,  RP -> Residential,  CP -> Commercial
  If any one is false  ->  status = AUDIT  (with audit_reason saying which check failed).

THREE STEPS
  1. EXPORT  (read-only)   python qbo_recode_review.py --export review.xlsx [--project MFD1234]
       Writes an Excel: every line marked OK or AUDIT. Locked columns identify the line;
       you edit the decision columns. Use the AutoFilter to show only status = AUDIT.
  2. AUDIT   (you, in Excel)
       On AUDIT rows you want fixed: set Approved=Y, pick Final_Customer / Final_Class from the
       dropdowns (real QBO names = exact spelling). Blank Approved = skip.
  3. APPLY   python qbo_recode_review.py --apply review.xlsx           # validate only (no writes)
             python qbo_recode_review.py --apply review.xlsx --commit  # write approved rows

SAFETY
  * EXPORT and "--apply without --commit" never write to QBO.
  * Only rows you mark Approved=Y are ever touched.
  * Final_Customer / Final_Class must match a real QBO name exactly, or the row is skipped+reported.
  * SyncToken captured at export; if the txn changed since, the row is skipped+reported.
  * Closed-period rows (<= QBO book-close date) skipped unless --include-closed.
  * Per-txn JSON backups on every apply.  Logs -> ~/Library/Logs/Proficient/.

AUTH: wire get_auth() to your qbo_vault (no secrets stored here).   DEPS: pip install requests openpyxl
"""

import os, re, sys, json, time, csv, argparse, logging
from datetime import datetime
from pathlib import Path

try:
    import requests
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("pip install requests openpyxl")

MINOR_VERSION = "70"
COST_ENTITIES = ["Bill", "Purchase", "VendorCredit"]
PROJECT_RE = re.compile(r"\b((?:MFD|RP|CP)\s?\d{2,5}(?:-FTW)?)\b", re.IGNORECASE)
PREFIX_TO_DIVISION = {"MFD": "Multi Family", "RP": "Residential", "CP": "Commercial"}
DIVISION_TO_CLASS = {"MULTI FAMILY": "Multi Family", "MULTIFAMILY": "Multi Family",
                     "RESIDENTIAL": "Residential", "COMMERCIAL": "Commercial"}
LINE_DETAILS = ("AccountBasedExpenseLineDetail", "ItemBasedExpenseLineDetail")

LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient"
BACKUP_ROOT = LOG_DIR / "recode_backups"
LOG_DIR.mkdir(parents=True, exist_ok=True); BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.FileHandler(LOG_DIR / "recode_review.log"), logging.StreamHandler()])
log = logging.getLogger("recode")

LOCKED = ["entity","txn_id","line_id","sync_token","txn_date","vendor","doc_number","account","amount",
          "line_description","desc_project","current_customer","customer_project",
          "class_current","class_expected","status","audit_reason","proposed_customer","proposed_class"]
EDITABLE = ["Approved(Y/N)","Final_Customer","Final_Class","Note"]
HEADERS = LOCKED + EDITABLE

def get_auth():
    # Replace with your qbo_vault, e.g.:
    #   from qbo_vault import get_qbo_session ; tok, realm = get_qbo_session()
    #   return tok, realm, "https://quickbooks.api.intuit.com"
    tok = os.environ.get("QBO_ACCESS_TOKEN"); realm = os.environ.get("QBO_REALM_ID")
    base = os.environ.get("QBO_BASE_URL", "https://quickbooks.api.intuit.com")
    if not tok or not realm:
        sys.exit("get_auth(): wire to qbo_vault (or set QBO_ACCESS_TOKEN / QBO_REALM_ID).")
    return tok, realm, base

class QBO:
    def __init__(self):
        self.tok, self.realm, self.base = get_auth()
        self.s = requests.Session()
        self.s.headers.update({"Authorization": f"Bearer {self.tok}",
                               "Accept": "application/json", "Content-Type": "application/json"})
    def _req(self, method, path, **kw):
        delay = 1
        for attempt in range(6):
            r = self.s.request(method, f"{self.base}/v3/company/{self.realm}/{path}",
                               params={**kw.pop("params", {}), "minorversion": MINOR_VERSION}, **kw)
            if r.status_code < 400: return r.json()
            if r.status_code in (429,500,502,503,504) and attempt < 5:
                log.warning("HTTP %s %s — retry %ss", r.status_code, path, delay); time.sleep(delay); delay *= 2; continue
            raise RuntimeError(f"QBO {r.status_code}: {r.text[:400]}")
        raise RuntimeError("retries exhausted")
    def query(self, stmt): return self._req("GET","query",params={"query":stmt}).get("QueryResponse",{})
    def query_all(self, entity, where=""):
        out, start = [], 1
        while True:
            rows = self.query(f"SELECT * FROM {entity} {where} STARTPOSITION {start} MAXRESULTS 1000".strip()).get(entity, [])
            out += rows
            if len(rows) < 1000: return out
            start += 1000
    def get(self, entity, _id): return self._req("GET", f"{entity.lower()}/{_id}").get(entity, {})
    def update(self, entity, body): return self._req("POST", entity.lower(), json=body).get(entity, {})
    def book_close_date(self):
        bcd = self._req("GET","preferences").get("Preferences",{}).get("AccountingInfoPrefs",{}).get("BookCloseDate")
        return datetime.strptime(bcd,"%Y-%m-%d").date() if bcd else None

def parse_proj(text):
    m = PROJECT_RE.search(str(text or ""))
    return m.group(1).upper().replace(" ", "") if m else None

def division_of(proj):
    if not proj: return ""
    pref = "MFD" if proj.startswith("MFD") else proj[:2]
    return PREFIX_TO_DIVISION.get(pref, "")

def normalize_division(name): return DIVISION_TO_CLASS.get((name or "").strip().upper())

def build_job_index(q):
    cust = q.query_all("Customer"); by_id = {c["Id"]: c for c in cust}
    def top_name(c):
        seen=set()
        while c.get("ParentRef") and c["Id"] not in seen:
            seen.add(c["Id"]); c = by_id.get(c["ParentRef"]["value"], {})
            if not c: break
        return c.get("DisplayName") or c.get("FullyQualifiedName","")
    idx={}
    for c in cust:
        m=parse_proj(" ".join(filter(None,[c.get("DisplayName"),c.get("FullyQualifiedName"),c.get("CompanyName")])))
        if not m: continue
        idx[m]={"customer_id":c["Id"],
                "customer_name":c.get("FullyQualifiedName") or c.get("DisplayName"),
                "class_name":normalize_division(top_name(c)) or division_of(m)}
    return idx

def build_class_index(q): return {c["Name"]: c["Id"] for c in q.query_all("Class")}

# ---------------------------------------------------------------- EXPORT (Pass 1)
def export(q, path, only_project, include_ok):
    jobs=build_job_index(q); classes=build_class_index(q); close=q.book_close_date()
    log.info("Indexed %d jobs, %d classes. Book-close: %s", len(jobs), len(classes), close or "none")
    rows=[]; n_ok=0
    for entity in COST_ENTITIES:
        for t in q.query_all(entity):
            txn_p = parse_proj(t.get("PrivateNote")) or parse_proj(t.get("DocNumber")) or parse_proj(t.get("Memo"))
            tdate=t.get("TxnDate",""); vendor=(t.get("VendorRef") or {}).get("name","")
            for line in t.get("Line",[]):
                dk=next((k for k in LINE_DETAILS if k in line),None)
                if not dk: continue
                det=line[dk]
                desc=line.get("Description","") or ""
                desc_p=parse_proj(desc)
                eff_p=desc_p or txn_p
                if only_project and eff_p!=only_project: continue
                cur_cust=(det.get("CustomerRef") or {}).get("name","")
                cust_p=parse_proj(cur_cust)
                cur_cls=(det.get("ClassRef") or {}).get("name","")
                exp_cls=division_of(eff_p)

                # THE CHECK — all three must be true for OK
                reasons=[]
                if not desc_p:
                    reasons.append("no project# in line description" + (f" (txn note has {txn_p})" if txn_p else ""))
                if eff_p and cust_p!=eff_p:
                    reasons.append("customer≠project#" if cust_p else "no customer on line")
                if eff_p and cur_cls!=exp_cls:
                    reasons.append(f"class '{cur_cls or '-'}'≠'{exp_cls}'")
                if eff_p and eff_p not in jobs:
                    reasons.append("project# not a QBO job")
                ok = bool(desc_p) and (cust_p==desc_p) and (cur_cls==division_of(desc_p))
                status = "OK" if ok else "AUDIT"
                if ok: n_ok+=1
                if ok and not include_ok:   # keep the workbook focused on what needs auditing
                    continue

                closed = bool(close and tdate and datetime.strptime(tdate,"%Y-%m-%d").date()<=close)
                if closed and status=="AUDIT": reasons.append("CLOSED period")
                prop_cust = jobs[eff_p]["customer_name"] if eff_p in jobs else ""
                rows.append([entity,t["Id"],line.get("Id",""),t.get("SyncToken",""),tdate,vendor,
                             t.get("DocNumber",""),(det.get("AccountRef") or {}).get("name",""),line.get("Amount",""),
                             desc[:160],desc_p or "",cur_cust,cust_p or "",cur_cls,exp_cls,status,
                             "; ".join(reasons),prop_cust,exp_cls,"","","",""])
    _write_xlsx(path, rows, sorted({j["customer_name"] for j in jobs.values()}|{r[11] for r in rows if r[11]}),
                sorted(classes.keys()))
    log.info("Wrote %d rows (%d OK lines %s) -> %s",
             len(rows), n_ok, "included" if include_ok else "hidden; use --include-ok to see", path)

def _write_xlsx(path, rows, customer_names, class_names):
    wb=Workbook(); ws=wb.active; ws.title="Review"; ws.append(HEADERS)
    for c in range(1,len(HEADERS)+1): ws.cell(1,c).font=ws.cell(1,c).font.copy(bold=True)
    for r in rows: ws.append(r)
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}{max(ws.max_row,1)}"
    ls=wb.create_sheet("_lists")
    ls["A1"]="Customers"
    for i,n in enumerate(customer_names,2): ls.cell(i,1,n)
    ls["B1"]="Classes"
    for i,n in enumerate(class_names,2): ls.cell(i,2,n)
    ls["C1"]="Approved"; ls["C2"]="Y"; ls["C3"]="N"
    nC=max(len(customer_names)+1,2); nK=max(len(class_names)+1,2)
    col={h:get_column_letter(i+1) for i,h in enumerate(HEADERS)}
    dvA=DataValidation(type="list",formula1="=_lists!$C$2:$C$3",allow_blank=True)
    dvU=DataValidation(type="list",formula1=f"=_lists!$A$2:$A${nC}",allow_blank=True)
    dvK=DataValidation(type="list",formula1=f"=_lists!$B$2:$B${nK}",allow_blank=True)
    for dv in (dvA,dvU,dvK): ws.add_data_validation(dv)
    last=max(ws.max_row,2)
    dvA.add(f"{col['Approved(Y/N)']}2:{col['Approved(Y/N)']}{last}")
    dvU.add(f"{col['Final_Customer']}2:{col['Final_Customer']}{last}")
    dvK.add(f"{col['Final_Class']}2:{col['Final_Class']}{last}")
    ins=wb.create_sheet("Instructions",0)
    for line in ["HOW TO USE","",
        "THE RULE — a line is OK only if all three are true:",
        "   1) the line description contains a project # (MFD####/RP####/CP####)",
        "   2) the Customer:Project number == that project #",
        "   3) the Class == the division the project # starts with (MFD=Multi Family, RP=Residential, CP=Commercial)",
        "Any one false => status AUDIT, with audit_reason naming the failed check(s).","",
        "1. On the Review tab, AutoFilter status = AUDIT to see only what needs work.",
        "2. Locked columns (left) identify the line — do not edit.",
        "3. To FIX a line: set Approved(Y/N)=Y, and pick Final_Customer / Final_Class from the dropdowns",
        "   (real QBO names). Leave them blank to use proposed_customer / proposed_class.",
        "4. Leave Approved blank to SKIP.",
        "5. Save, then:  python qbo_recode_review.py --apply <thisfile>        (validate, no writes)",
        "6. If the results look right:  add --commit  to write the approved rows.","",
        "Only Approved=Y rows are written. Exact-spelling + stale-txn + closed-period checks run on apply."]:
        ins.append([line])
    wb.save(path)

# ---------------------------------------------------------------- APPLY (Pass 3)
def apply(q, path, commit, include_closed):
    wb=load_workbook(path); ws=wb["Review"]
    hdr=[c.value for c in ws[1]]; ix={h:i for i,h in enumerate(hdr)}
    classes=build_class_index(q)
    cust_by_name={c.get("FullyQualifiedName") or c.get("DisplayName"):c["Id"] for c in q.query_all("Customer")}
    close=q.book_close_date()
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); run=BACKUP_ROOT/stamp; run.mkdir(parents=True,exist_ok=True)
    res=[("entity","txn_id","line_id","result","detail")]
    st={"approved":0,"applied":0,"skip_name":0,"skip_stale":0,"skip_closed":0,"already":0,"err":0}
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row or not row[ix["txn_id"]]: continue
        if str(row[ix["Approved(Y/N)"]] or "").strip().upper() not in ("Y","YES","TRUE","1"): continue
        st["approved"]+=1
        entity=row[ix["entity"]]; tid=str(row[ix["txn_id"]]); lid=str(row[ix["line_id"]] or "")
        tgt_cust=(row[ix["Final_Customer"]] or row[ix["proposed_customer"]] or "").strip()
        tgt_cls=(row[ix["Final_Class"]] or row[ix["proposed_class"]] or "").strip()
        cust_id=cust_by_name.get(tgt_cust); cls_id=classes.get(tgt_cls)
        if not cust_id or (tgt_cls and not cls_id):
            res.append((entity,tid,lid,"SKIP_NAME_NOT_FOUND",f"cust={tgt_cust!r} class={tgt_cls!r}")); st["skip_name"]+=1; continue
        try:
            fresh=q.get(entity,tid)
            if str(fresh.get("SyncToken",""))!=str(row[ix["sync_token"]] or ""):
                res.append((entity,tid,lid,"SKIP_STALE","txn changed since export — re-export")); st["skip_stale"]+=1; continue
            tdate=fresh.get("TxnDate","")
            if close and tdate and datetime.strptime(tdate,"%Y-%m-%d").date()<=close and not include_closed:
                res.append((entity,tid,lid,"SKIP_CLOSED",tdate)); st["skip_closed"]+=1; continue
            (run/f"{entity}_{tid}.json").write_text(json.dumps(fresh,indent=1))
            changed=False
            for line in fresh.get("Line",[]):
                if lid and str(line.get("Id",""))!=lid: continue
                dk=next((k for k in LINE_DETAILS if k in line),None)
                if not dk: continue
                det=line[dk]
                if (det.get("CustomerRef") or {}).get("value")!=cust_id:
                    det["CustomerRef"]={"value":cust_id,"name":tgt_cust}; det.setdefault("BillableStatus","NotBillable"); changed=True
                if cls_id and (det.get("ClassRef") or {}).get("value")!=cls_id:
                    det["ClassRef"]={"value":cls_id,"name":tgt_cls}; changed=True
            if not changed:
                res.append((entity,tid,lid,"ALREADY_OK","")); st["already"]+=1; continue
            if commit:
                fresh["sparse"]=False; q.update(entity,fresh)
                res.append((entity,tid,lid,"APPLIED",f"{tgt_cust} / {tgt_cls}")); st["applied"]+=1
                log.info("APPLIED %s %s line %s -> %s / %s",entity,tid,lid,tgt_cust,tgt_cls); time.sleep(0.25)
            else:
                res.append((entity,tid,lid,"WOULD_APPLY",f"{tgt_cust} / {tgt_cls}"))
        except Exception as e:
            res.append((entity,tid,lid,"ERROR",str(e)[:200])); st["err"]+=1; log.error("ERR %s %s: %s",entity,tid,e)
    out=run/"apply_results.csv"
    with open(out,"w",newline="") as f: csv.writer(f).writerows(res)
    log.info("%s — %s", "COMMIT" if commit else "VALIDATE-ONLY", json.dumps(st)); log.info("Results: %s", out)
    if not commit: log.info("No writes made. Re-run with --commit to apply the approved rows.")

def main():
    ap=argparse.ArgumentParser(description="Audit-gated QBO job-cost re-coder (export -> you edit -> apply).")
    ap.add_argument("--export",metavar="XLSX",help="Pass 1: write review workbook")
    ap.add_argument("--apply",metavar="XLSX",help="Pass 3: read your edited workbook")
    ap.add_argument("--project",help="limit export to one project #, e.g. MFD1234")
    ap.add_argument("--include-ok",action="store_true",help="export also includes OK lines (default: AUDIT only)")
    ap.add_argument("--commit",action="store_true",help="with --apply: actually write to QBO")
    ap.add_argument("--include-closed",action="store_true",help="DANGER: also write closed-period txns")
    a=ap.parse_args()
    q=QBO()
    if a.export: export(q,a.export,(a.project or "").upper().replace(" ","") or None, a.include_ok)
    elif a.apply: apply(q,a.apply,a.commit,a.include_closed)
    else: ap.print_help()

if __name__=="__main__":
    main()
