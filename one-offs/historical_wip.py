"""Historical WIP reconstruction — a point-in-time WIP as of a past date, built
SEPARATELY from the live WIP master (the user 2026-08-08). Two dates asked for:
12-31-2025 and 3-31-2026.

SOURCES (all confirmed present):
  · MFD + CP — the monthly .xlsb snapshot in 'Company Files - WIP Report/WIP
    History' already froze each division's WIP at month-end. Read the
    'WIP - CP' / 'WIP - MFD' tabs (pyxlsb) — that IS the truth as of the date,
    no QBO needed. Snapshot carries contract · COs · revised · billed
    (COMPLETED TO DATE) · % complete · retainage — NO cost column (it's a
    billing-based WIP).
  · RP — NOT in the old snapshots. Rebuilt from the schedule of that exact day
    → the RP jobs working then → each job's bid proposal vs. its invoice
    (match 100% ⇒ contract acquired, else FLAG) → QBO billed/costs dated
    on/before the report date. [stage B/C — added next]

Output: '~/Downloads/WIP as of <date>.xlsx', one tab per division. Touches
nothing else. Stage A here = MFD + CP from the snapshots.
"""
import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from pyxlsb import open_workbook as open_xlsb

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))
from shared import qbo_api                                    # noqa: E402
from shared.xlsx_verify import assert_clean, safe_table_name  # noqa: E402
import rp_wip_reader as RP                                    # noqa: E402
import rp_schedule_wip_preview as P                           # noqa: E402  (schedule + proposal)

WIP_HISTORY = Path(os.getenv(
    "WIP_HISTORY_DIR",
    str(Path.home() / "Library/CloudStorage/OneDrive-ProficientConcrete,LLC"
        / "Company Files - WIP Report" / "WIP History")))

# date key → (snapshot .xlsb, human label)
DATES = {
    "12-31-2025": ("WIP_12-31.25.xlsb", "December 31, 2025"),
    "3-31-2026":  ("WIP - 03-31-26.xlsb", "March 31, 2026"),
}
# date key → (RP schedule of that exact day, report-date ISO for the QBO cutoff)
SCHED_ROOT = Path(os.getenv("RP_SCHEDULE_ROOT", "/Volumes/Common/OPERATIONS/SCHEDULE"))
SCHEDULES = {
    "12-31-2025": (SCHED_ROOT / "2025" / "December 2025" / "Schedule 12-31-25.xlsx", "2025-12-31"),
    "3-31-2026":  (SCHED_ROOT / "2026" / "March 2026" / "Schedule 3-31-26.xlsx", "2026-03-31"),
}

# The snapshot columns we lift, by division tab. (header label → out label)
CP_COLS = [
    ("PROJECT", "PROJECT"), ("CUSTOMER", "CUSTOMER"), ("CONTRACT", "CONTRACT"),
    ("CHANGE ORDERS", "CHANGE ORDERS"), ("REV. CONTRACT", "REVISED CONTRACT"),
    ("COMPLETED TO DATE", "BILLED TO DATE"), ("% COMPLETE", "% COMPLETE"),
    ("BALANCE TO FINISH INCL'N RET", "BALANCE TO FINISH"),
    ("Total Retainage", "RETAINAGE"),
]
MFD_COLS = CP_COLS  # same header names on the MFD tab

HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
_thin = Side(style="thin", color="000000")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
MONEY = '"$"#,##0_);[Red]("$"#,##0)'
PCT = "0%"
FONT = "Tahoma"


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def read_division(xlsb_path: Path, tab: str, colspec):
    """Rows from a snapshot division tab. Excludes rows explicitly marked
    COMPLETED in col A; keeps every other job (its status travels along).
    Returns (rows, header_out_labels)."""
    with open_xlsb(str(xlsb_path)) as wb:
        with wb.get_sheet(tab) as sh:
            grid = [[c.v for c in row] for row in sh.rows()]
    hdr_i = next((i for i, r in enumerate(grid)
                  if any(str(v).strip() == "PROJECT" for v in r if v)), None)
    if hdr_i is None:
        return [], [o for _, o in colspec]
    hdr = grid[hdr_i]
    idx = {str(v).strip(): j for j, v in enumerate(hdr) if v not in (None, "")}
    pcol = idx.get("% COMPLETE")
    bcol = idx.get("BALANCE TO FINISH INCL'N RET")
    out = []
    for r in grid[hdr_i + 1:]:
        if not r:
            continue
        status = str(r[0]).strip() if r[0] not in (None, "") else ""
        proj = r[idx["PROJECT"]] if idx.get("PROJECT") is not None and len(r) > idx["PROJECT"] else None
        if not proj or not str(proj).strip():
            continue
        # ACTIVE = still billing as of the snapshot. The tab is a cumulative
        # ledger (mostly completed jobs kept for reference), and the col-A
        # COMPLETED tag is only on some — so key off % complete: < 100% billed
        # is active; fully-billed/retainage-only jobs live in the RETAINAGE tab.
        pct = _num(r[pcol]) if (pcol is not None and len(r) > pcol) else None
        bal = _num(r[bcol]) if (bcol is not None and len(r) > bcol) else None
        if "COMPLETED" in status.upper():
            continue
        active = (pct is not None and pct < 0.999) or (pct is None and bool(bal and bal > 1))
        if not active:
            continue
        rec = {"STATUS": status or "active"}
        for src, dst in colspec:
            j = idx.get(src)
            v = r[j] if (j is not None and len(r) > j) else None
            rec[dst] = _num(v) if dst not in ("PROJECT", "CUSTOMER") else (
                str(v).strip() if v not in (None, "") else "")
        out.append(rec)
    return out, [o for _, o in colspec]


def _write_tab(ws, title, subtitle, rows, out_labels, table_name):
    ws.cell(1, 1, title).font = Font(name=FONT, size=11, bold=True)
    ws.cell(2, 1, subtitle).font = Font(name=FONT, size=8)
    headers = ["STATUS"] + out_labels
    hdr = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(hdr, c, label)
        cell.fill = HDR_FILL
        cell.font = Font(name=FONT, size=8, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    widths = {"PROJECT": 26, "CUSTOMER": 22, "STATUS": 11}
    for c, label in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(hdr, c).column_letter].width = widths.get(label, 15)
    for i, rec in enumerate(rows, hdr + 1):
        for c, label in enumerate(headers, 1):
            cell = ws.cell(i, c, rec.get(label))
            cell.font = Font(name=FONT, size=8)
            cell.border = BORDER
            if label in ("CONTRACT", "CHANGE ORDERS", "REVISED CONTRACT",
                         "BILLED TO DATE", "RETAINAGE"):
                cell.number_format = MONEY
            elif label == "% COMPLETE":
                cell.number_format = PCT
    last = hdr + len(rows)
    if rows:
        from openpyxl.utils import get_column_letter
        ref = f"A{hdr}:{get_column_letter(len(headers))}{last}"
        t = Table(displayName=table_name, ref=ref)   # already valid + unique
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.tables.add(t)
    ws.freeze_panes = f"A{hdr+1}"


RP_COLS = ["PROJECT", "ADDRESS", "BUILDER", "SCOPE", "CONTRACT (proposal)",
           "BILLED (as of date)", "COSTS (as of date)", "NOTES"]

# schedule task-section title → the scope it implies. Punch-list sections are by
# superintendent name (people) — skipped: not tracked, and we store no names.
_SECT_SCOPE = [("FORM SET", "slab"), ("GRADE", "slab"),
               ("FLATWORK", "ftw"), ("WRECK", "wreck")]
_SCOPE_RANK = {"slab": 3, "ftw": 2, "wreck": 1}


def _read_schedule_jobs(path: Path):
    """The daily schedule (this format: NAME · ADDRESS · CITY · BUILDER ·
    DESCRIPTION, banded by task section) → unique jobs by ADDRESS. Person names
    in col 1 and the punch-list-by-name sections are ignored. When a job shows
    up in several sections, keep the most-advanced scope (slab > ftw > wreck)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.strip().lower() == "main schedule"), None)
    if ws is None:
        wb.close()
        return []
    scope, jobs = None, {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        v = [str(c.value).strip() if c.value is not None else "" for c in row[:6]]
        v += [""] * (6 - len(v))
        c1, addr, _city, builder, desc = v[0], v[1], v[2], v[3], v[4]
        if c1 and not addr and not desc:                 # a band title
            up = c1.upper()
            scope = next((sc for key, sc in _SECT_SCOPE if key in up), None)
            continue
        if c1.upper() == "NAME" or not addr or scope is None:
            continue                                     # header row / untracked band
        key = re.sub(r"\s+", " ", addr.upper()).strip()
        prev = jobs.get(key)
        if prev is None or _SCOPE_RANK.get(scope, 0) > _SCOPE_RANK.get(prev["scope"], 0):
            jobs[key] = {"address": addr, "builder": builder, "desc": desc, "scope": scope}
    wb.close()
    return list(jobs.values())


def _pl_totals(access, cid, cust_id, end_iso):
    """QBO billed (income) + costs (COGS+expenses) for a project, dated on/before
    end_iso. Returns (billed, costs), or (None, '<err>') on failure."""
    try:
        t = qbo_api.extract_pl_totals(
            qbo_api.fetch_project_pl(access, cid, cust_id, "2019-01-01", end_iso))
        return (float(t.get("income", 0.0) or 0.0),
                (t.get("cogs", 0.0) or 0.0) + (t.get("expenses", 0.0) or 0.0))
    except Exception as e:
        return None, f"QBO fail {type(e).__name__}"


def build_rp(date_key, access, cid, proj_map, folders_idx):
    """RP starting point (the user 2026-08-08: "pull the schedule as of, gather
    costs and billed from QBO, put that as the starting point"). The schedule of
    the day IS the active list — match each address to its project folder for the
    RP# + bid proposal, then pull QBO billed/costs dated on/before the date.
    Bounded to what's on the schedule; unmatched addresses are flagged, not guessed."""
    sched_path, end_iso = SCHEDULES[date_key]
    if not sched_path.exists():
        print(f"    ⚠ RP: schedule not found ({sched_path}) — RP tab left empty")
        return []
    jobs = _read_schedule_jobs(sched_path)
    _rp_to_folders, addr_folders = folders_idx
    rows = []
    for j in jobs:
        addr, scope = j["address"], j["scope"]
        parts = addr.split(None, 1)
        folder = RP.match_by_address(
            {"house": parts[0] if parts else "",
             "street": parts[1] if len(parts) > 1 else addr}, addr_folders)
        rp_num, contract, note = None, None, ""
        if folder is not None:
            # The RP# lives in the FILE names inside the address folder
            # (RP####_ADDRESS…), not the folder name itself.
            try:
                for x in folder.iterdir():
                    mm = re.search(r"(RP\d{3,4})", x.name.upper())
                    if mm:
                        rp_num = mm.group(1)
                        break
            except OSError:
                pass
            _pp, contract, note = P.find_proposal(
                folder, "slab" if scope == "wreck" else scope, j["desc"])
            note = note or ""
            if rp_num is None:
                note = (note + "; " if note else "") + "matched a folder but no RP# in its files"
        else:
            note = "unmatched address — verify job"
        line = f"{rp_num}-FTW" if (rp_num and scope == "ftw") else rp_num
        billed = costs = None
        cust = (proj_map.get(line) or proj_map.get(rp_num)) if rp_num else None
        if cust:
            billed, costs = _pl_totals(access, cid, cust["id"], end_iso)
            if isinstance(costs, str):
                note = (note + "; " if note else "") + costs
                billed = costs = None
        elif rp_num:
            note = (note + "; " if note else "") + "no QBO project"
        rows.append({
            "PROJECT": line or "(unmatched)", "ADDRESS": addr, "BUILDER": j["builder"],
            "SCOPE": scope, "CONTRACT (proposal)": contract,
            "BILLED (as of date)": billed, "COSTS (as of date)": costs, "NOTES": note,
        })
    matched = sum(1 for r in rows if r["PROJECT"] != "(unmatched)")
    print(f"    RP: {len(jobs)} scheduled job(s) · {matched} matched to a project · "
          f"{len(rows) - matched} unmatched → {sched_path.name}")
    return rows


def _write_rp_tab(ws, label, rows):
    ws.cell(1, 1, f"WIP as of {label} — RESIDENTIAL").font = Font(name=FONT, size=11, bold=True)
    ws.cell(2, 1, "STARTING POINT: active jobs = the schedule of the day, matched to the "
                  "project folder for RP# + bid proposal; billed/costs from QBO as of the date")\
        .font = Font(name=FONT, size=8)
    hdr = 4
    for c, label_ in enumerate(RP_COLS, 1):
        cell = ws.cell(hdr, c, label_)
        cell.fill = HDR_FILL
        cell.font = Font(name=FONT, size=8, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    widths = {"PROJECT": 15, "ADDRESS": 30, "BUILDER": 22, "SCOPE": 8, "NOTES": 40}
    for c, label_ in enumerate(RP_COLS, 1):
        ws.column_dimensions[ws.cell(hdr, c).column_letter].width = widths.get(label_, 16)
    for i, rec in enumerate(rows, hdr + 1):
        for c, label_ in enumerate(RP_COLS, 1):
            cell = ws.cell(i, c, rec.get(label_))
            cell.font = Font(name=FONT, size=8)
            cell.border = BORDER
            if label_ in ("CONTRACT (proposal)", "BILLED (as of date)", "COSTS (as of date)"):
                cell.number_format = MONEY
            if label_ in ("NOTES", "ADDRESS", "BUILDER"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    if rows:
        from openpyxl.utils import get_column_letter
        ref = f"A{hdr}:{get_column_letter(len(RP_COLS))}{hdr+len(rows)}"
        t = Table(displayName=safe_table_name("histRP", set()), ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.tables.add(t)
    ws.freeze_panes = f"A{hdr+1}"


def build(date_key: str, out_dir: Path, qbo=None, folders_idx=None) -> Path:
    snap, label = DATES[date_key]
    xlsb = WIP_HISTORY / snap
    if not xlsb.exists():
        raise FileNotFoundError(f"snapshot not found: {xlsb}")
    cp, cp_h = read_division(xlsb, "WIP - CP", CP_COLS)
    mfd, mfd_h = read_division(xlsb, "WIP - MFD", MFD_COLS)
    rp = []
    if qbo and folders_idx:
        rp = build_rp(date_key, qbo[0], qbo[1], qbo[2], folders_idx)
    seen = set()
    wb = Workbook()
    _write_tab(wb.active, f"WIP as of {label} — COMMERCIAL",
               f"from snapshot '{snap}' · WIP - CP tab · billing-based (no cost column in source)",
               cp, cp_h, safe_table_name("histCP", seen))
    wb.active.title = "CP"
    _write_tab(wb.create_sheet("MFD"), f"WIP as of {label} — MULTI-FAMILY",
               f"from snapshot '{snap}' · WIP - MFD tab", mfd, mfd_h,
               safe_table_name("histMFD", seen))
    rp_ws = wb.create_sheet("RP")
    if qbo:
        _write_rp_tab(rp_ws, label, rp)
    else:
        rp_ws.cell(1, 1, f"WIP as of {label} — RESIDENTIAL").font = Font(name=FONT, size=11, bold=True)
        rp_ws.cell(3, 1, "run with QBO (no --skip-rp) to build this tab").font = Font(name=FONT, size=9, italic=True)
    out = out_dir / f"WIP as of {date_key}.xlsx"
    wb.save(out)
    assert_clean(out)          # NEVER hand over a file that would trip Excel repair
    print(f"  ✓ {date_key}: CP {len(cp)} · MFD {len(mfd)} · RP {len(rp)} · "
          f"xlsx verified clean → {out}")
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Point-in-time WIP as of a past date (MFD/CP snapshot + RP rebuilt).")
    ap.add_argument("--date", choices=list(DATES) + ["all"], default="all")
    ap.add_argument("--out", default=str(Path.home() / "Downloads"))
    ap.add_argument("--skip-rp", action="store_true",
                    help="MFD/CP only — no QBO, no proposal reads (fast, no Touch ID)")
    args = ap.parse_args()
    out_dir = Path(args.out).expanduser()
    keys = list(DATES) if args.date == "all" else [args.date]
    print("\n  HISTORICAL WIP — MFD/CP (snapshots) + RP (schedule/proposal/QBO)")
    qbo = folders_idx = None
    if not args.skip_rp:
        access, cid = qbo_api.load_credentials()          # one Touch ID
        proj_map = qbo_api.build_project_customer_map(access, cid)
        qbo = (access, cid, proj_map)
        print("  indexing residential folders for proposals …")
        folders_idx = RP.index_residential(RP.RP_ROOT)
    for k in keys:
        build(k, out_dir, qbo, folders_idx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
