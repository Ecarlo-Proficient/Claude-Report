#!/usr/bin/env python3
"""
jobtread_migration_setup.py — "JobTread Migration, Setup for Success": the
JOBS-TO-ADD list (the user 2026-07-24). READ-ONLY against JobTread, Notion,
and the schedule; writes nothing but one Downloads Excel.

WHAT WE NEED TO ADD TO JOBTREAD — two clean tabs, one as-of:
  1. ACTIVE JOBS TO ADD   — jobs on the daily schedule (our active/production
     work) that JobTread does NOT have yet.  Most important.
  2. BIDDING TO ADD       — Notion RP bids still out (not won/lost) that aren't
     in JobTread yet — the pipeline to pre-load for going JobTread-only.

"Already in JobTread" is judged by exact job # against a full JobTread sweep,
so a job is on a list only if it genuinely needs creating.

SOURCES / AUTH (all read-only, one Touch ID for the JobTread key)
  • Schedule : /Volumes/.../SCHEDULE  (xlsx 'Main Schedule' tab) → active jobs
  • JobTread : JT_GRANT_KEY in the shared key library (Keychain automation-qbo)
  • Notion   : the invoice-sync integration token (Keychain
               proficient-automation-worker / notion, or $NOTION_SECRET).

Usage
  python3 jobtread_migration_setup.py                    # yesterday/latest schedule
  python3 jobtread_migration_setup.py --schedule <f.xlsx>
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from shared import qbo_vault
import rp_wip_reader as RP
import rp_schedule_wip_preview as P
import rp_jobtread_coverage as C   # pave(), ORG_ID

# ── Notion ──────────────────────────────────────────────────────────
NOTION_VER = "2025-09-03"
BIDLIST_DS = "19db24f7-5585-814a-a453-000b798447a9"   # Bid List data source
BUILDERS_DS = "19db24f7-5585-81af-a4e1-000bbe22e6cc"  # Builders data source
# Lead Status values that mean WON or dead — everything else is "still out".
TERMINAL = {"Sold", "Lost", "GC Not Awarded", "No Response", "No Opportunity"}
# stage order for display (lower = hotter); unknown/blank sorts last.
STAGE_RANK = {
    "Negotiating": 0, "Bid Pending Items": 1, "Revised Bid Sent": 2,
    "Bid Sent": 3, "Estimate Sent": 4, "Ready to Review": 5,
    "Waiting Plans/Info": 6, "On Hold": 7, "Preliminary": 8, "Not Started": 9,
}
_JOBNO_RE = re.compile(r"^(RP|CP|MFD)\d{3,4}(?:-FTW)?", re.IGNORECASE)


def notion_token() -> str:
    """invoice-sync's Notion integration token — Keychain first, env second.
    Read via the same service/key invoice-sync uses; we never import its code."""
    svc = os.getenv("KEYSTORE_SERVICE", "proficient-automation-worker")
    key = os.getenv("KEYSTORE_KEY_NOTION", "notion")
    try:
        import keyring
        tok = keyring.get_password(svc, key)
        if tok:
            return tok
    except Exception:
        pass
    tok = os.getenv("NOTION_SECRET")
    if tok:
        return tok
    raise SystemExit("No Notion token (Keychain proficient-automation-worker/"
                     "notion or $NOTION_SECRET).")


def notion_query(token: str, ds: str, body: dict) -> dict:
    req = urllib.request.Request(
        f"https://api.notion.com/v1/data_sources/{ds}/query",
        data=json.dumps(body).encode(),
        headers={"Authorization": f"Bearer {token}",
                 "Notion-Version": NOTION_VER,
                 "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def _prop(props: dict, name: str):
    """Scalar-ish value out of a Notion property, by its type."""
    p = props.get(name)
    if not p:
        return None
    t = p.get("type")
    if t == "title":
        return "".join(x["plain_text"] for x in p["title"]).strip() or None
    if t == "rich_text":
        return "".join(x["plain_text"] for x in p["rich_text"]).strip() or None
    if t == "select":
        return (p["select"] or {}).get("name")
    if t == "status":
        return (p["status"] or {}).get("name")
    if t == "number":
        return p["number"]
    if t == "date":
        return (p["date"] or {}).get("start")
    if t == "url":
        return p["url"]
    if t == "relation":
        return [x["id"] for x in p["relation"]]
    if t == "people":
        return [x.get("name") or x.get("id") for x in p["people"]]
    return None


def builder_map(token: str) -> dict:
    """id → builder/client name (the title of each Builders page)."""
    out, cur = {}, None
    while True:
        body = {"page_size": 100}
        if cur:
            body["start_cursor"] = cur
        r = notion_query(token, BUILDERS_DS, body)
        for pg in r["results"]:
            title = next((v for v in pg["properties"].values()
                          if v.get("type") == "title"), None)
            name = ("".join(x["plain_text"] for x in title["title"]).strip()
                    if title else "")
            out[pg["id"]] = name or "(unnamed builder)"
        if not r.get("has_more"):
            break
        cur = r["next_cursor"]
    return out


def open_rp_bids(token: str, bmap: dict) -> list:
    """RP (Residential) Bid List rows whose Lead Status is NOT won/lost."""
    rows, cur = [], None
    while True:
        body = {"page_size": 100,
                "filter": {"property": "Division",
                           "select": {"equals": "Residential"}}}
        if cur:
            body["start_cursor"] = cur
        r = notion_query(token, BIDLIST_DS, body)
        for pg in r["results"]:
            pr = pg["properties"]
            status = _prop(pr, "Lead Status")
            if status in TERMINAL:
                continue
            jobname = _prop(pr, "Job Name") or ""
            if not jobname.strip():          # blank/template row — skip
                continue
            m = _JOBNO_RE.match(jobname)
            bids = _prop(pr, "Builder") or []
            rows.append({
                "job": (m.group(0).upper() if m else ""),
                "address": _prop(pr, "Job Address") or "",
                "city": _prop(pr, "City") or "",
                "type": _prop(pr, "Job Type") or "",
                "builder": ", ".join(bmap.get(b, "?") for b in bids),
                "amount": _prop(pr, "Bid Amount"),
                "status": status or "(no status)",
                "bid_sent": _prop(pr, "Bid Sent Date") or "",
            })
        if not r.get("has_more"):
            break
        cur = r["next_cursor"]
    return rows


def _neg_date(d: str):
    """Sort dates newest-first as strings (YYYY-MM-DD)."""
    return tuple(-int(x) for x in d.split("-")) if d else (0,)


# ── JobTread full sweep ─────────────────────────────────────────────
def all_jt_jobs(key: str) -> list:
    out, page, guard = [], None, 0
    while True:
        dollar = {"size": 100}
        if page:
            dollar["page"] = page
        q = {"organization": {"$": {"id": C.ORG_ID}, "jobs": {
            "$": dollar, "nextPage": {},
            "nodes": {"number": {}, "name": {}, "status": {}}}}}
        r = C.pave(key, q)["organization"]["jobs"]
        out.extend(r["nodes"])
        page = r.get("nextPage")
        guard += 1
        if not page or guard > 60:
            break
    return out


# ── Excel — clean: one-line title, bold header w/ underline, autofilter,
#    no fills, no grid boxes ───────────────────────────────────────────
BOLD = Font(bold=True)
UNDER = Border(bottom=Side(style="thin", color="000000"))
CUR = '"$"#,##0.00'


def _add_sheet(wb, title, subtitle, header, rows, widths, money_cols=()):
    ws = wb.create_sheet(title)
    ws["A1"] = subtitle
    ws["A1"].font = BOLD                       # single line, no wrap
    ws.append(header)                          # row 2
    for c in range(1, len(header) + 1):
        cell = ws.cell(2, c)
        cell.font = BOLD
        cell.border = UNDER
        cell.alignment = Alignment(vertical="bottom")
    for row in rows:
        ws.append(row)
    for c in money_cols:
        for r in range(3, ws.max_row + 1):
            ws.cell(r, c).number_format = CUR
    last = chr(ord("A") + len(header) - 1)
    ws.auto_filter.ref = f"A2:{last}{max(2, ws.max_row)}"
    ws.freeze_panes = "A3"
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    return ws


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--schedule", help="schedule xlsx (default: yesterday/latest)")
    args = ap.parse_args()

    print("\n  JOBTREAD — Jobs to Add (read-only)")
    print("  " + "─" * 44)

    # 1) schedule → active jobs -------------------------------------
    if args.schedule:
        sched_path = Path(args.schedule)
        m = RP._SCHED_FILE_RE.search(sched_path.name)
        label = "-".join(m.groups()) if m else sched_path.stem
    else:
        best = P.latest_schedule(RP.SCHEDULE_DIR)
        if best is None:
            print("  ✗ no schedule file found")
            return 1
        _k, sched_path = best
        label = f"{_k[1]}-{_k[2]}-{_k[0] % 100:02d}"
    if not sched_path.exists():
        print(f"  ✗ schedule not found: {sched_path}")
        return 1
    print(f"  schedule: {sched_path.name}")
    sched = P.read_main_schedule(sched_path)
    jobs = {}
    for s in sched:
        jobs.setdefault(s["job"], s)
    print(f"  active jobs on the schedule: {len(jobs)}")

    # 2) JobTread sweep → what already exists -----------------------
    key = qbo_vault.get("JT_GRANT_KEY")
    print("  JobTread: full job sweep …")
    jt_all = all_jt_jobs(key)
    jt_numbers = {(j.get("number") or "").upper() for j in jt_all}
    print(f"    {len(jt_all)} JobTread jobs total")

    # tab 1 — active jobs NOT in JobTread
    active_add = [(job, s) for job, s in jobs.items() if job not in jt_numbers]
    active_add.sort(key=lambda t: (t[1]["builder"], t[0]))
    active_rows = [[job, s["address"], s["city"], s["builder"], s["section"]]
                   for job, s in active_add]
    print(f"    active jobs to add: {len(active_rows)}")

    # 3) Notion bids still out, NOT in JobTread ---------------------
    print("  Notion: Bid List RP (still out) …")
    tok = notion_token()
    bmap = builder_map(tok)
    bids = open_rp_bids(tok, bmap)
    bid_add = [b for b in bids if b["job"] not in jt_numbers]
    bid_add.sort(key=lambda x: (STAGE_RANK.get(x["status"], 99),
                                x["bid_sent"] == "", _neg_date(x["bid_sent"]),
                                x["job"]))
    bid_rows = [[b["job"], b["address"], b["city"], b["type"], b["builder"],
                 b["amount"], b["status"], b["bid_sent"]] for b in bid_add]
    bid_total = sum(b["amount"] or 0 for b in bid_add)
    print(f"    bidding jobs to add: {len(bid_rows)}  (Σ bid ${bid_total:,.0f})")

    # ── workbook ────────────────────────────────────────────────────
    out = Path(os.getenv(
        "JT_MIGRATION_XLSX",
        str(Path.home() / "Downloads" / "JobTread Migration Setup.xlsx")))
    lock = out.with_name("~$" + out.name)
    if lock.exists():
        raise SystemExit(f"{out.name} is open in Excel — close it first")

    wb = Workbook()
    wb.remove(wb.active)   # drop default sheet

    _add_sheet(
        wb, "Active Jobs to Add",
        f"ACTIVE JOBS TO ADD TO JOBTREAD — {len(active_rows)} jobs on the "
        f"{label} schedule not yet in JobTread (as of {label}).",
        ["JOB #", "ADDRESS", "CITY", "BUILDER", "SCHEDULE PHASE"],
        active_rows, (14, 32, 16, 28, 16))

    _add_sheet(
        wb, "Bidding to Add",
        f"BIDDING — JOBS TO ADD TO JOBTREAD — {len(bid_rows)} open RP bids "
        f"(not won/lost) not yet in JobTread. Σ bid ${bid_total:,.0f}.",
        ["JOB #", "ADDRESS", "CITY", "JOB TYPE", "BUILDER", "BID $",
         "LEAD STATUS", "BID SENT"],
        bid_rows, (14, 30, 16, 16, 28, 14, 16, 12), money_cols=(6,))

    wb.save(out)
    print(f"\n  ✓ Workbook → {out}")
    print(f"    Active to add {len(active_rows)} · Bidding to add "
          f"{len(bid_rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
