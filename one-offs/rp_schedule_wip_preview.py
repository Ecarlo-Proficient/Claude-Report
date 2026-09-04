#!/usr/bin/env python3
"""
rp_schedule_wip_preview.py — PREVIEW of the schedule-driven RP WIP method
(the user 2026-07-16). READ-ONLY everywhere; never touches the WIP master.

THE METHOD BEING PREVIEWED
  The daily schedule's 'Main Schedule' tab is the truth for which RP jobs are
  ACTIVE — the General List lags it (e.g. RP6440's fence scope is on the
  schedule but has no GL row). Under the new method each schedule job gets:
    • CONTRACT  = the bid proposal PDF in the project folder (last
      "SUB TOTAL: $…" in the text — the signed client number)
    • ETC       = the takeoff workbook's cost sheet ('JobTread Cost Gral',
      the last sheet): slab = SL+PR sections, flatwork = FW section
  Scope: FLATWORK schedule section → RP#-FTW line; wreck rows follow their
  description; everything else (trench/piers/forms/grade) → slab line.

OUTPUT — one Excel (never the WIP): ~/Downloads/RP WIP - Schedule Method
Preview.xlsx, grouped NEW (on schedule, missing from the current WIP) /
CHANGED (numbers differ from the General List) / MATCHES, with the source
file of every number.

Usage:
  python3 rp_schedule_wip_preview.py
  python3 rp_schedule_wip_preview.py --project RP6440
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))

from openpyxl import Workbook, load_workbook

import wip_writer as W  # shared report engine (was cp_wip_reader)
import rp_wip_reader as RP
# Takeoff→ETC extractor moved to shared/ (2026-08-07) once the WIP reader also
# needed it — re-exported here so P._norm / P._score_name / P.find_takeoff_etc
# stay valid for the job-auditor prototypes that call them through this module.
from shared.takeoff_etc import (          # noqa: F401  (re-export)
    _SIDE_TOKENS, _norm, _desc_tokens, _score_name,
    _cost_sheet_totals, find_takeoff_etc)

MAIN_SHEET = "Main Schedule"
# Section bands on the Main Schedule tab → scope of the WIP line they feed.
SECTION_SCOPE = [
    ("WRECK", "desc"),          # wreck slab → slab · wreck flatwork → ftw
    ("FLATWORK", "ftw"),
    ("GRADE AND BACKOUT", "slab"),
    ("TRENCH", "slab"),
    ("PIERS", "slab"),
    ("FORM SET", "slab"),
    ("TRACTOR", "slab"),
    ("CONCRETE CUTTING", "desc"),
]
_JOB_RE = re.compile(r"^(RP|CP)\d{3,4}$", re.IGNORECASE)
_SUBTOTAL_RE = re.compile(r"(?:SUB\s*)?TOTAL\s*:?\s*\$\s*([\d,]+(?:\.\d{1,2})?)",
                          re.IGNORECASE)
# _SIDE_TOKENS now lives in shared/takeoff_etc.py (imported above).

# Team-corrected source paths (RP WIP Fixes.xlsx → rp_source_overrides.json,
# 2026-07-22): WIP-LINE → {proposal, takeoff} Mac paths that OVERRIDE the
# folder guess. The team fixed wrong folders, wrong builders, and typos, so
# these win over any automatic file match.
_OVR_PATH = Path(__file__).resolve().parent / "rp_source_overrides.json"
try:
    OVERRIDES = json.loads(_OVR_PATH.read_text()) if _OVR_PATH.exists() else {}
except Exception:
    OVERRIDES = {}

# TRACT volume builders (the team 2026-07-22): NO bid proposal — contract +
# cost come from P.O.'s / a builder price list. Don't false-flag 'no proposal';
# take the General Lista price as the contract.
TRACT_TOKENS = ("CAMDEN", "GRAND HOMES", "HABITAT", "HABITART", "WILLIAM RYAN")


def _is_tract(builder: str) -> bool:
    b = _norm(builder)
    return any(t in b for t in TRACT_TOKENS)


# Jobs whose BUDGET comes from the General Lista, not the takeoff (per the
# user, case-by-case). RP7535 (2026-07-23).
GL_ETC_JOBS = {"RP7535"}


def latest_schedule(sched_dir: Path):
    best = None
    for year_dir in sched_dir.iterdir():
        if not (year_dir.is_dir() and year_dir.name.isdigit()):
            continue
        for month_dir in year_dir.iterdir():
            if not month_dir.is_dir():
                continue
            for f in month_dir.iterdir():
                m = RP._SCHED_FILE_RE.search(f.name)
                if m:
                    mo, dy, yy = (int(g) for g in m.groups())
                    key = (2000 + yy, mo, dy)
                    if best is None or key > best[0]:
                        best = (key, f)
    return best


def read_main_schedule(path: Path):
    """'Main Schedule' → [{job, scope, section, address, builder, desc}].
    Section bands tracked from the band-title rows; a data row is any row
    whose PROJECT column matches RP####/CP####."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames
                  if s.strip().lower() == MAIN_SHEET.lower()), None)
    if sheet is None:
        wb.close()
        raise SystemExit(f"No {MAIN_SHEET!r} tab in {path.name}")
    ws = wb[sheet]
    out, seen, section = [], {}, None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        vals = [str(c.value).strip() if c.value is not None else ""
                for c in row[:12]]
        joined = _norm(" ".join(vals))
        job = vals[1].upper() if len(vals) > 1 else ""
        if not _JOB_RE.match(job):
            for token, scope in SECTION_SCOPE:
                if token in joined and "NAME" not in joined \
                        and "SUPERINTENDENT" not in joined:
                    section = (token, scope)
                    break
            continue
        if section is None:
            continue
        desc = _norm(vals[5])
        token, scope = section
        if scope == "desc":
            scope = "ftw" if ("FLATWORK" in desc or "FTW" in desc
                              or "PAVING" in desc) else "slab"
        key = (job, scope)
        rec = seen.get(key)
        if rec:
            if desc and desc not in rec["desc"]:
                rec["desc"] += f" · {desc}"
            continue
        rec = {"job": job, "scope": scope, "section": token,
               "address": _norm(vals[2]), "city": _norm(vals[3]),
               "builder": _norm(vals[4]), "desc": desc}
        seen[key] = rec
        out.append(rec)
    wb.close()
    return out


def read_gl_jobs(path: Path):
    """ALL job numbers present in the General List (priced or not) — presence
    check for the lag report. RP.read_general_list drops unpriced rows, so
    presence needs its own pass."""
    wb = load_workbook(path, data_only=True, read_only=True)
    jobs = set()
    for sheet in (RP.ALPHA_SHEET, RP.SMALL_SHEET):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=6, min_col=3, max_col=3):
            v = row[0].value
            if v:
                m = RP._JOB_RE.match(str(v).strip())
                if m:
                    jobs.add(m.group(1).upper())
    wb.close()
    return jobs


# ─────────────────── proposal PDF → contract ───────────────────────
def find_proposal(folder: Path, scope: str, desc: str):
    """Best bid-proposal PDF + its extracted SUB TOTAL. Returns
    (path, amount, note) — amount None when no priced proposal reads.

    FTW scope only accepts PDFs whose NAME says flatwork (or matches the
    schedule description) — the base proposal is the SLAB contract, and
    silently returning it would grossly overstate the -FTW line (the whole
    house-slab bid vs the far smaller flatwork scope — seen on RP5542). No
    such PDF → caller falls back to the takeoff's bid sheets."""
    import pdfplumber
    cands = []
    try:
        for f in folder.iterdir():
            if f.suffix.lower() != ".pdf":
                continue
            n = _norm(f.name)
            if "INVOICE" in n or "DIAGRAM" in n or "PLAN" in n:
                continue
            if "BID" in n or "PROPOSAL" in n or _desc_tokens(desc) & set(n.split()):
                if scope == "ftw" and not (
                        "FLATWORK" in n or "FTW" in n
                        or _desc_tokens(desc) & set(n.split())):
                    continue
                cands.append(f)
    except OSError:
        return None, None, "folder unreadable"
    if not cands:
        return None, None, ("no flatwork proposal PDF" if scope == "ftw"
                            else "no proposal PDF in folder")
    cands.sort(key=lambda f: (_score_name(_norm(f.name), scope, desc),
                              f.stat().st_mtime), reverse=True)
    for f in cands[:4]:                    # best-ranked few; first priced wins
        try:
            with pdfplumber.open(f) as pdf:
                txt = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
            hits = _SUBTOTAL_RE.findall(txt)
            if hits:
                return f, float(hits[-1].replace(",", "")), None
        except Exception:
            continue
    return cands[0], None, "proposal found but no SUB TOTAL text (unpriced?)"


# ─────────────────────────── report ────────────────────────────────
# Implied gross-margin sanity band for the NEW numbers (the user 2026-07-17,
# RP5542-FTW: contract $91K vs takeoff budget $48K = 47.6% GP — "way too
# low"). RP jobs run ~10–25% GP; outside the band the pair is mismatched
# (wrong scope file, stale takeoff, or missing cost sections).
_GP_HI = 0.35     # above → budget too low vs the contract
_GP_LO = 0.05     # below → margin too thin (budget too high / contract too low)


def margin_flag(contract, budget):
    """(gp_pct, instructions) for a contract/budget pair — None when sane.
    Instructions are estimator to-dos (the user 2026-07-21), one per line;
    `_split_needs` routes them to the AR (proposal) or JR (budget) column."""
    if not contract or budget is None:
        return None, None
    gp = (contract - budget) / contract
    if gp > _GP_HI:
        return gp, (f"Budget way too low vs contract (implied GP "
                    f"{gp * 100:.0f}%; RP runs ~10–25%) — find the real "
                    f"budget takeoff sheet for this scope")
    if gp < 0:
        return gp, (f"Contract below budget (GP {gp * 100:.0f}%) — "
                    f"export/confirm the latest bid proposal PDF sent to "
                    f"the client\n"
                    f"Also confirm the budget takeoff isn't carrying "
                    f"extra scope")
    if gp < _GP_LO:
        return gp, (f"Margin too thin (GP {gp * 100:.0f}%) — verify the "
                    f"budget takeoff against the contract")
    return gp, None


def _split_needs(needs: str):
    """Route each NEEDS instruction to (AR text, JR text) — the user
    2026-07-21 wants owners in SEPARATE COLUMNS, not colored runs in one
    cell. AR (Accounts/estimator) = anything about the bid proposal PDF;
    JR = budget/takeoff/sheet/folder work; pure GL-status notes are dropped
    (the GROUP + 'IN GEN. LIST?' columns already say that).

    Separate single-font columns replace the old rich-text cell entirely —
    Mac Excel rejected openpyxl's multi-run inline strings ('String
    properties' repair), so the report now carries NO rich text at all."""
    ar, jr = [], []
    for p in (p.strip() for p in re.split(r"[;\n]", needs or "") if p.strip()):
        u = p.upper()
        if "GENERAL LIST" in u or "IN GL" in u or "UNPRICED" in u:
            continue
        (ar if "PROPOSAL" in u else jr).append(p)
    return "\n".join(ar), "\n".join(jr)


def _breadcrumb(path: Path) -> str:
    """Human-navigable path text starting at CURRENT PROJECTS (the user
    2026-07-21): the file:// links die on Windows/OneDrive, so estimators
    need a breadcrumb they can walk in Explorer — 'CURRENT PROJECTS >
    Residential > <client> > <address> > <file>'. Mac links stay attached."""
    parts = list(path.parts)
    for i, p in enumerate(parts):
        if p.upper() == "CURRENT PROJECTS":
            return " > ".join(parts[i:])
    return " > ".join(parts[-4:])


# ── extraction from a SPECIFIC file (used by team-corrected overrides) ──
def pdf_subtotal(path: Path):
    """Last 'SUB TOTAL: $…' in a specific proposal PDF, or None."""
    import pdfplumber
    try:
        with pdfplumber.open(path) as pdf:
            txt = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
    except Exception:
        return None
    hits = _SUBTOTAL_RE.findall(txt)
    return float(hits[-1].replace(",", "")) if hits else None


def takeoff_budget_from(f: Path, scope: str, desc: str):
    """Budget from a SPECIFIC takeoff file → (budget, note, fragment). Tries
    the commercial 'BID' sheet (AP1948/AP1961) for slab scope, then the
    residential 'JobTread Cost Gral' bands."""
    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception:
        return None, "takeoff unreadable", None
    if scope != "ftw":
        bid = next((s for s in wb.sheetnames if _norm(s) == "BID"), None)
        if bid is not None:
            ws = wb[bid]
            for ref in ("AP1948", "AP1961"):
                v = ws[ref].value
                if isinstance(v, (int, float)) and v:
                    wb.close()
                    return round(float(v), 2), f"'{bid}' {ref}", f"#'{bid}'!{ref}"
    sheet = next((s for s in wb.sheetnames if "COST GRAL" in _norm(s)), None)
    if sheet is None:
        wb.close()
        return (None, "Missing 'JobTread Cost Gral' sheet — add the budget "
                "sheet to the takeoff", None)
    bands = _cost_sheet_totals(wb[sheet])
    wb.close()
    if _desc_tokens(desc) & set(_norm(f.name).split()):
        keys = ("SL", "PR", "FW")
    else:
        keys = ("FW",) if scope == "ftw" else ("SL", "PR")
    budget, cells = 0.0, []
    for k in keys:
        b = bands[k]
        if b["sub"] is not None:
            budget += b["sub"]
            cells.append(b["cell"])
        elif b["items"]:
            budget += b["items"]
            cells.append(f"{k} items")
    if budget:
        frag = (f"#'{sheet}'!"
                + next((c for c in cells if re.match(r'^[A-Z]+[0-9]+$', c)), "A1"))
        return round(budget, 2), f"'{sheet}' {' + '.join(cells)}", frag
    return (None, "Missing 'JobTread Cost Gral' sheet — add the budget sheet "
            "to the takeoff", None)


def _blank_item(**kw):
    """Full item dict with defaults so cross-check rows (from the General
    Lista) carry every key the writer reads."""
    base = dict(group="", line="", section="", desc="", address="", builder="",
                in_gl=False, gl_contract=None, gl_etc=None, new_contract=None,
                new_etc=None, proposal=None, p_note="", takeoff=None, t_note="",
                needs="", folder=None, t_frag=None, gl_sheet=None, gl_row=None)
    base.update(kw)
    return base


def _link(cell, target, fragment: str = ""):
    """file:// hyperlink + blue underline; no-op when target is None.

    The sheet/cell jump goes in the hyperlink's `location` ATTRIBUTE, not
    appended to the URI (the user 2026-07-21): a fragment like
    #'JobTread Cost Gral'!D11 puts raw spaces in the .rels target URI —
    invalid XML that makes Excel demand a repair on open."""
    from openpyxl.styles import Font
    from openpyxl.worksheet.hyperlink import Hyperlink
    if target is None or cell.value in (None, ""):
        return
    try:
        uri = Path(target).as_uri()
    except (ValueError, OSError):
        return
    loc = fragment.lstrip("#") if fragment else None
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=uri,
                               location=loc or None)
    cell.font = Font(color="0563C1", underline="single")


def write_report(items, sched_label, out_path: Path) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    YELLOW = PatternFill("solid", fgColor="FFF2CC")   # General List numbers
    GREEN = PatternFill("solid", fgColor="E2EFDA")    # proposal/takeoff numbers
    AMBER = PatternFill("solid", fgColor="FCE4D6")    # significant deltas
    NEW_F = PatternFill("solid", fgColor="F8CBAD")    # NEW group band
    BAD = Font(color="9C0006", bold=True)
    GOOD = Font(color="006100", bold=True)
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CUR = '"$"#,##0.00_);[Red]("$"#,##0.00)'

    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        raise SystemExit(f"{out_path.name} is open in Excel — close it first")
    wb = Workbook()
    ws = wb.active
    ws.title = "PREVIEW"
    ws["A1"] = (f"SCHEDULE-DRIVEN RP WIP — PREVIEW ONLY (schedule "
                f"{sched_label}; the WIP master was NOT touched). "
                f"YELLOW = General Lista numbers · GREEN = bid proposal / "
                f"takeoff numbers · every number links to its source. "
                f"If we grabbed the WRONG file (or it says missing), delete "
                f"what's in the PROPOSAL PDF / TAKEOFF FILE box and paste the "
                f"correct full file path there.")
    ws["A1"].font = Font(bold=True)
    ws.append([])
    # Legend — two plain single-font cells (NO rich text anywhere in this
    # workbook: Mac Excel rejects openpyxl's multi-run inline strings with a
    # 'String properties' repair). AR and JR each read their own column.
    ORANGE_F = Font(color="ED7D31", bold=True, size=12)
    BLUE_F = Font(color="0070C0", bold=True, size=12)
    ws["A2"] = "ORANGE = AR — bid proposal / contract actions"
    ws["A2"].font = ORANGE_F
    ws["H2"] = "BLUE = JR — budget takeoff actions"
    ws["H2"].font = BLUE_F
    HDR = ["GROUP", "WIP LINE", "WORK DESC",
           "ADDRESS", "BUILDER", "IN GENERAL LISTA?",
           "General Lista CONTRACT $", "General Lista BUDGET $",
           "NEW CONTRACT $ (proposal)", "NEW BUDGET $ (takeoff)", "NEW GP %",
           "PROPOSAL PDF", "TAKEOFF FILE",
           "AR — PROPOSAL / CONTRACT", "JR — BUDGET / TAKEOFF"]
    ws.append(HDR)
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(3, c)
        cell.font = Font(bold=True)
        cell.fill = GRAY
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    # Header tint: General Lista columns yellow, source columns green, the
    # two owner columns carry their owner's color.
    for c in (7, 8):
        ws.cell(3, c).fill = YELLOW
    for c in (9, 10, 11):
        ws.cell(3, c).fill = GREEN
    ws.cell(3, 14).font = ORANGE_F
    ws.cell(3, 15).font = BLUE_F

    order = {"⚠ IN GENERAL LISTA, NOT ON SCHEDULE": 0, "NEW — not in WIP": 1,
             "CHANGED": 2, "MATCHES": 3, "FTW BACKLOG (GL, not scheduled)": 4,
             "FLATWORK TAKEN BY OTHER": 5}
    for it in sorted(items, key=lambda x: (order.get(x["group"], 6),
                                           x["line"])):
        gp, gp_flag = margin_flag(it["new_contract"], it["new_etc"])
        needs = "; ".join(x for x in (it["needs"], gp_flag) if x)
        ar_needs, jr_needs = _split_needs(needs)
        ws.append([it["group"], it["line"], it["desc"],
                   it["address"], it["builder"],
                   ("yes" if it["in_gl"] else "NO"),
                   it["gl_contract"], it["gl_etc"],
                   it["new_contract"], it["new_etc"], gp,
                   (_breadcrumb(it["proposal"]) if it["proposal"]
                    else it["p_note"]),
                   ((_breadcrumb(it["takeoff"])
                     + (f"\n[{it['t_note']}]" if it["t_note"] else ""))
                    if it["takeoff"] else it["t_note"]),
                   ar_needs, jr_needs])
        r = ws.max_row
        for cc in range(1, len(HDR) + 1):
            ws.cell(r, cc).border = BORDER
            ws.cell(r, cc).alignment = Alignment(
                vertical="top", wrap_text=(cc in (3, 12, 13, 14, 15)))
        for cc in (7, 8, 9, 10):
            ws.cell(r, cc).number_format = CUR
        ws.cell(r, 11).number_format = "0.0%"
        # Source-colored numbers: General Lista yellow · proposal/takeoff green.
        for cc in (7, 8):
            ws.cell(r, cc).fill = YELLOW
        for cc in (9, 10, 11):
            ws.cell(r, cc).fill = GREEN
        # Owner NEEDS columns carry the owner's font color.
        if ws.cell(r, 14).value:
            ws.cell(r, 14).font = Font(color="ED7D31", bold=True)
        if ws.cell(r, 15).value:
            ws.cell(r, 15).font = Font(color="0070C0", bold=True)
        # Links: line → folder · General Lista $ → its list row · new numbers
        # → the exact proposal PDF / takeoff cell; file columns → the folder
        # (they delete the path and paste the correct one in the same box).
        _link(ws.cell(r, 2), it.get("folder"))
        gl_frag = (f"#'{it['gl_sheet']}'!C{it['gl_row']}"
                   if it.get("gl_row") else "")
        for cc in (7, 8):
            _link(ws.cell(r, cc), RP.ALPHA_PATH if it.get("gl_row") else None,
                  gl_frag)
        _link(ws.cell(r, 9), it.get("proposal"))
        _link(ws.cell(r, 10), it.get("takeoff"), it.get("t_frag") or "")
        _link(ws.cell(r, 12),
              it["proposal"].parent if it.get("proposal") else None)
        _link(ws.cell(r, 13),
              it["takeoff"].parent if it.get("takeoff") else None)
        # GROUP color: RED-flag not-on-schedule = red fill; NEW banded orange;
        # CHANGED amber-text; MATCHES green; FTW backlog / OTHER neutral-band.
        RED_FILL = PatternFill("solid", fgColor="F4CCCC")
        if it["group"].startswith("⚠"):
            ws.cell(r, 1).fill = RED_FILL
            ws.cell(r, 1).font = BAD
        elif it["group"].startswith("NEW"):
            ws.cell(r, 1).fill = NEW_F
            ws.cell(r, 1).font = BAD
        elif it["group"] == "CHANGED":
            ws.cell(r, 1).font = Font(color="BF6000", bold=True)
        elif it["group"] == "MATCHES":
            ws.cell(r, 1).font = GOOD
        else:                      # FTW backlog / taken by OTHER
            ws.cell(r, 1).fill = GRAY
            ws.cell(r, 1).font = Font(bold=True)
        if gp is not None and gp_flag:
            ws.cell(r, 11).font = BAD
        if not it["in_gl"]:
            ws.cell(r, 6).font = BAD
    widths = (16, 12, 26, 22, 20, 11, 15, 15, 14, 13, 9, 34, 34, 32, 32)
    for col, w in zip("ABCDEFGHIJKLMNO", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # ── second sheet: budget-sheet CLEANUP CHECKLIST (the user 2026-07-22) ──
    # Every takeoff missing its 'JobTread Cost Gral' sheet (or with no takeoff
    # at all) — the #1 data gap the team surfaced. Add the sheet → the budget
    # extracts automatically next run.
    cw = wb.create_sheet("CLEANUP CHECKLIST")
    cw["A1"] = ("BUDGET-SHEET CLEANUP — takeoffs missing the 'JobTread Cost "
                "Gral' sheet (or no takeoff file). Add the sheet to the takeoff "
                "and the budget reads automatically.")
    cw["A1"].font = Font(bold=True)
    cw.append([])
    ch = ["JOB #", "BUILDER", "ISSUE", "TAKEOFF FILE / FOLDER"]
    cw.append(ch)
    for c in range(1, len(ch) + 1):
        cell = cw.cell(3, c)
        cell.font = Font(bold=True)
        cell.fill = GRAY
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    seen = set()
    for it in sorted(items, key=lambda x: (x["builder"], x["line"])):
        tn = it["t_note"] or ""
        # Genuine misses ONLY — a SUCCESSFUL extraction's note also names the
        # 'Cost Gral' sheet, so match the miss phrasing, not any mention.
        miss = ("Missing 'JobTread Cost Gral'" in tn
                or "No budget takeoff" in tn or "takeoff unreadable" in tn)
        if miss and it["line"] not in seen:
            seen.add(it["line"])
            where = (_breadcrumb(it["takeoff"]) if it["takeoff"]
                     else (_breadcrumb(it["folder"]) if it["folder"] else "—"))
            cw.append([it["line"], it["builder"], tn, where])
            rr = cw.max_row
            for cc in range(1, len(ch) + 1):
                cw.cell(rr, cc).border = BORDER
                cw.cell(rr, cc).alignment = Alignment(vertical="top",
                                                      wrap_text=True)
    for col, w in zip("ABCD", (14, 24, 50, 58)):
        cw.column_dimensions[col].width = w
    cw.freeze_panes = "A4"

    wb.save(out_path)
    print(f"  ✓ Preview → {out_path}")


# ─────────── commit READY → Test - RP · MISSING → Downloads ───────────
def _cprow_from_item(it):
    """Build a WIP CpRow from a schedule item. Contract/budget prefer the
    extracted numbers, falling back to the General Lista (backlog rows carry
    only GL numbers)."""
    contract = (it["new_contract"] if it["new_contract"] is not None
                else it["gl_contract"])
    etc = it["new_etc"] if it["new_etc"] is not None else it["gl_etc"]
    row = W.CpRow(it["line"], it["address"] or it["line"], False,
                   contract, None, etc, None, None)
    row.folder_path = it.get("folder")
    if it.get("takeoff"):
        row.takeoff_path = it["takeoff"]
    row.client = it["builder"] or None
    row.home_type = "Tract" if _is_tract(it["builder"]) else "Custom"
    return row


def commit_to_test_rp(ready, backlog, label):
    """Write the READY schedule jobs (contract + budget known) to the
    'Test - RP' tab as the new RP WIP report, QBO-enriched, with the General
    Lista FTW backlog as an appendix. Production write — guarded/QC'd by
    write_test_cp."""
    main_rows = [_cprow_from_item(it) for it in ready]
    bk_rows = [_cprow_from_item(it) for it in backlog]
    pairs = [(r, None, None) for r in main_rows + bk_rows]
    print("  Enriching READY + backlog with QBO Billed/Costs …")
    RP.enrich_with_qbo(pairs)
    for r in main_rows:
        RP._classify(r, None)
    for r in bk_rows:
        RP._classify(r, None)
        r.needs_review = False            # backlog = expected, not an error
    wrote = W.write_test_cp(
        main_rows, W.WIP_EXCEL_PATH, tab_name="Test - RP",
        appendix=("FTW BACKLOG — flatwork bid in the General Lista, not yet "
                  "on the schedule (expected wins)", bk_rows),
        cols=RP._rp_cols(), default_filter_active=True,
        title=f"RP WIP REPORT — schedule-driven, as of {label}", summary=True)
    # return the enriched rows keyed by PROJECT # so the audit view can show
    # billed/costs next to the contract/budget.
    return wrote, {r.project_num: r for r in main_rows}


def write_one_file(wip_audit, missing, not_scheduled, label, out_path):
    """ONE workbook (the user 2026-07-24), schedule-driven, two sheets:
      'WIP (AUDIT)'  jobs IN the WIP — original contract, budget, billed &
                     costs to date; the $ opens the file it came from, the
                     FROM column opens that file's folder and names it.
      'MISSING'      General Lista jobs in progress but NOT on the schedule
                     (red catch), then schedule jobs missing contract/budget.
    No rich text."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    GREEN = PatternFill("solid", fgColor="E2EFDA")
    RED = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CUR = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        raise SystemExit(f"{out_path.name} is open in Excel — close it first")
    wb = Workbook()

    # ── Sheet 1: WIP (AUDIT) ──
    ws = wb.active
    ws.title = "WIP (AUDIT)"
    ws["A1"] = (f"RP WIP - AUDIT (schedule {label}). Jobs in the WIP with the "
                f"numbers to check: original contract, budget, billed & costs "
                f"to date. Each $ opens the file it came from; the FROM column "
                f"opens that file's folder and names it.")
    ws["A1"].font = Font(bold=True)
    ws.append([])
    H = ["JOB #", "WORK DESC", "BUILDER", "ORIGINAL CONTRACT $", "BUDGET $",
         "BILLED TO DATE $", "COSTS TO DATE $", "CONTRACT FROM", "BUDGET FROM"]
    ws.append(H)
    for c in range(1, len(H) + 1):
        cell = ws.cell(3, c)
        cell.font = Font(bold=True); cell.fill = GRAY; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for it, row in sorted(wip_audit, key=lambda t: (t[0]["builder"], t[0]["line"])):
        billed = getattr(row, "billed_to_date", None) if row else None
        costs = getattr(row, "costs_to_date", None) if row else None
        prop, take = it.get("proposal"), it.get("takeoff")
        cfrom = _breadcrumb(prop) if prop else (it.get("p_note") or "-")
        bfrom = _breadcrumb(take) if take else (it.get("t_note") or "-")
        ws.append([it["line"], it["desc"], it["builder"], it["new_contract"],
                   it["new_etc"], billed, costs, cfrom, bfrom])
        r = ws.max_row
        for cc in range(1, len(H) + 1):
            ws.cell(r, cc).border = BORDER
            ws.cell(r, cc).alignment = Alignment(vertical="top",
                                                 wrap_text=(cc in (2, 8, 9)))
        for cc in (4, 5, 6, 7):
            ws.cell(r, cc).number_format = CUR
            ws.cell(r, cc).fill = GREEN
        _link(ws.cell(r, 4), prop)
        _link(ws.cell(r, 5), take, it.get("t_frag") or "")
        _link(ws.cell(r, 8), prop.parent if prop else None)
        _link(ws.cell(r, 9), take.parent if take else None)
    for col, w in zip("ABCDEFGHI", (12, 24, 22, 17, 14, 14, 14, 40, 40)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # ── Sheet 2: MISSING ──
    ms = wb.create_sheet("MISSING")
    ms["A1"] = (f"RP WIP - WHAT'S MISSING (schedule {label}). RED = in the "
                f"General Lista, in progress, but NOT on the schedule (must be "
                f"scheduled). Below = schedule jobs missing a contract/budget "
                f"(ORANGE=AR proposal, BLUE=JR budget).")
    ms["A1"].font = Font(bold=True)
    ms.append([])
    HM = ["JOB #", "WORK DESC", "ADDRESS", "BUILDER", "MISSING",
          "AR - PROPOSAL / CONTRACT", "JR - BUDGET / TAKEOFF",
          "PROPOSAL PDF", "TAKEOFF FILE"]
    ms.append(HM)
    for c in range(1, len(HM) + 1):
        cell = ms.cell(3, c)
        cell.font = Font(bold=True); cell.fill = GRAY; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def _emit(job, desc, addr, builder, gap, ar, jr, prop, take, red_row=False):
        ms.append([job, desc, addr, builder, gap, ar, jr, prop, take])
        r = ms.max_row
        for cc in range(1, len(HM) + 1):
            ms.cell(r, cc).border = BORDER
            ms.cell(r, cc).alignment = Alignment(vertical="top", wrap_text=True)
        ms.cell(r, 5).fill = RED
        ms.cell(r, 5).font = Font(color="9C0006", bold=True)
        if red_row:
            for cc in range(1, 5):
                ms.cell(r, cc).fill = RED
        if ms.cell(r, 6).value:
            ms.cell(r, 6).font = Font(color="ED7D31", bold=True)
        if ms.cell(r, 7).value:
            ms.cell(r, 7).font = Font(color="0070C0", bold=True)

    for it in sorted(not_scheduled, key=lambda x: (x["builder"], x["line"])):
        _emit(it["line"], it["desc"], it["address"], it["builder"],
              "NOT ON SCHEDULE", it["needs"], "", "", "", red_row=True)
    for it in sorted(missing, key=lambda x: (x["builder"], x["line"])):
        gap = []
        if it["new_contract"] is None:
            gap.append("CONTRACT")
        if it["new_etc"] is None:
            gap.append("BUDGET")
        ar, jr = _split_needs(it["needs"])
        _emit(it["line"], it["desc"], it["address"], it["builder"],
              " + ".join(gap), ar, jr,
              (_breadcrumb(it["proposal"]) if it["proposal"] else it["p_note"]),
              (_breadcrumb(it["takeoff"]) if it["takeoff"] else it["t_note"]))
    for col, w in zip("ABCDEFGHI", (12, 24, 22, 22, 16, 40, 40, 34, 34)):
        ms.column_dimensions[col].width = w
    ms.freeze_panes = "A4"

    wb.save(out_path)
    print(f"  ok RP WIP (one file) -> {out_path}  "
          f"(WIP {len(wip_audit)} - not-on-schedule {len(not_scheduled)} - "
          f"missing {len(missing)})")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--project", help="filter to one job #")
    ap.add_argument("--schedule", help="use this schedule file instead of the latest")
    ap.add_argument("--commit", action="store_true",
                    help="write READY jobs (contract+budget known) to the "
                         "'Test - RP' WIP tab; emit only the MISSING list")
    args = ap.parse_args()

    print("\n  RP WIP — schedule-driven method PREVIEW (read-only)")
    print("  " + "─" * 62)
    if args.schedule:
        sched_path = Path(args.schedule)
        if not sched_path.exists():
            print(f"  ✗ schedule not found: {sched_path}")
            return 1
        m = RP._SCHED_FILE_RE.search(sched_path.name)
        label = "-".join(m.groups()) if m else sched_path.stem
    else:
        best = latest_schedule(RP.SCHEDULE_DIR)
        if best is None:
            print("  ✗ no schedule file found")
            return 1
        (_k, sched_path) = best
        label = f"{_k[1]}-{_k[2]}-{_k[0] % 100:02d}"
    print(f"  schedule: {sched_path.name}")
    sched = read_main_schedule(sched_path)
    if args.project:
        sched = [s for s in sched if s["job"] == args.project.upper()]
    print(f"  active schedule lines: {len(sched)}")

    gl_jobs = read_gl_jobs(RP.ALPHA_PATH)
    records, _m = RP.read_general_list(RP.ALPHA_PATH)
    gl_by_job = {r["job"]: r for r in records}

    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)

    items = []
    for s in sched:
        job, scope = s["job"], s["scope"]
        line = job if (scope == "slab" or job.startswith("CP")) else f"{job}-FTW"
        rec = gl_by_job.get(job)
        if job.startswith("CP"):
            gl_k = ((rec["slab_bid"] or 0) + (rec["flat_bid"] or 0)) or None \
                if rec else None
            gl_e = ((rec["slab_cost"] or 0) + (rec["flat_cost"] or 0)) or None \
                if rec else None
        elif scope == "ftw":
            gl_k = rec and rec["flat_bid"]
            gl_e = rec and rec["flat_cost"]
            if rec and rec.get("flat_other"):
                gl_k = gl_e = None            # OTHER won it — no line today
        else:
            gl_k = rec and rec["slab_bid"]
            gl_e = rec and rec["slab_cost"]
        in_wip = bool(gl_k or gl_e)

        folders = sorted(rp_to_folders.get(job, ()),
                         key=lambda f: (f.parent.name, f.name))
        folder = folders[0] if folders else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            fake = {"house": parts[0] if parts else "",
                    "street": parts[1] if len(parts) > 1 else s["address"]}
            folder = RP.match_by_address(fake, addr_folders)

        prop = tkoff = None
        new_k = new_e = None
        p_note = t_note = ""
        t_frag = None
        ov = OVERRIDES.get(line, {})
        # Team-pinned files WIN (RP WIP Fixes.xlsx corrected wrong folders,
        # wrong builders and typos — trust the human over the folder guess).
        if ov.get("proposal"):
            prop = Path(ov["proposal"])
            new_k = pdf_subtotal(prop)
            p_note = ("team-pinned" if new_k is not None
                      else "team-pinned (price list — no SUB TOTAL)")
        if ov.get("takeoff"):
            tkoff = Path(ov["takeoff"])
            new_e, t_note, t_frag = takeoff_budget_from(tkoff, scope, s["desc"])
        # Folder search fills whatever the override didn't supply.
        if folder is not None:
            if prop is None:
                prop, new_k, p_note = find_proposal(folder, scope, s["desc"])
            if tkoff is None:
                tkoff, new_e, t_note, t_frag = find_takeoff_etc(
                    folder, job, scope, s["desc"])
        elif prop is None and tkoff is None:
            p_note = t_note = "no project folder found"

        # TRACT builders (Camden, Grand Homes, Habitat…): NO bid proposal —
        # contract comes from P.O.'s / a builder price list. Take the General
        # Lista price as the contract; never false-flag 'no proposal'.
        tract = _is_tract(s["builder"])
        if tract and new_k is None:
            p_note = "Tract — contract from P.O.'s / builder price list"
            if gl_k:
                new_k = gl_k
        # Per-job budget-from-General-Lista override (RP7535, the user 2026-07-23).
        if line in GL_ETC_JOBS and gl_e:
            new_e = gl_e
            t_note = "Budget from the General Lista (per the user)"

        needs = []
        if not in_wip and not tract and not rec:
            needs.append("not in General List" if job not in gl_jobs
                         else "in GL but unpriced")
        if new_k is None:
            needs.append("Tract — enter contract from P.O.'s / price list"
                         if tract else
                         "Export/find the bid proposal PDF that was sent "
                         "to the client → drop it in this project's folder")
        if new_e is None:
            needs.append(t_note or "no budget takeoff")
        if not in_wip:
            group = "NEW — not in WIP"
        elif ((new_k is not None and gl_k and abs(new_k - gl_k) > max(100, 0.01 * gl_k))
              or (new_e is not None and gl_e and abs(new_e - gl_e) > max(100, 0.01 * gl_e))):
            group = "CHANGED"
        else:
            group = "MATCHES"
        items.append(_blank_item(
            group=group, line=line, section=s["section"], desc=s["desc"],
            address=s["address"], builder=s["builder"],
            in_gl=(job in gl_jobs), gl_contract=gl_k or None, gl_etc=gl_e or None,
            new_contract=new_k, new_etc=new_e, proposal=prop, p_note=p_note,
            takeoff=tkoff, t_note=t_note, needs="; ".join(needs), folder=folder,
            t_frag=t_frag, gl_sheet=(rec["source"] if rec else None),
            gl_row=(rec["gl_row"] if rec else None)))

    # ── General Lista cross-checks (the user 2026-07-22) ──
    #   (1) A GL job in progress but NOT on the schedule → RED: every active
    #       job must be on the schedule.
    #   (2) Flatwork from the GL: AF=OTHER = won by another contractor
    #       (excluded); otherwise a priced -FTW not on today's schedule = the
    #       backlog the schedule can't show (expected, not scheduled yet).
    sched_bases = {s["job"] for s in sched}
    sched_ftw = {s["job"] for s in sched if s["scope"] == "ftw"}
    n_notsched = n_backlog = n_other = 0
    for rec in records:
        job = rec["job"]
        if job.startswith("CP"):
            continue
        comp = rec.get("completion")
        in_progress = comp is None or (isinstance(comp, (int, float)) and comp < 0.999)
        if (rec["slab_bid"] or rec["slab_cost"]) and in_progress \
                and job not in sched_bases:
            n_notsched += 1
            items.append(_blank_item(
                group="⚠ IN GENERAL LISTA, NOT ON SCHEDULE", line=job,
                section="—", desc="in progress on the list, missing from the schedule",
                builder=str(rec.get("builder") or ""), in_gl=True,
                gl_contract=rec["slab_bid"], gl_etc=rec["slab_cost"],
                needs="All active jobs must be on the schedule — add it, or "
                      "mark it complete if it's done",
                gl_sheet=rec["source"], gl_row=rec["gl_row"]))
        if rec["flat_bid"] or rec["flat_cost"]:
            if rec.get("flat_other"):
                n_other += 1
                items.append(_blank_item(
                    group="FLATWORK TAKEN BY OTHER", line=f"{job}-FTW",
                    section="OTHER",
                    desc="AF=OTHER — flatwork won by another contractor",
                    builder=str(rec.get("builder") or ""), in_gl=True,
                    gl_contract=rec["flat_bid"], gl_etc=rec["flat_cost"],
                    needs="Excluded — another contractor won the flatwork",
                    gl_sheet=rec["source"], gl_row=rec["gl_row"]))
            elif job not in sched_ftw:
                n_backlog += 1
                items.append(_blank_item(
                    group="FTW BACKLOG (GL, not scheduled)", line=f"{job}-FTW",
                    section="BACKLOG",
                    desc="flatwork bid, not yet on the schedule",
                    builder=str(rec.get("builder") or ""), in_gl=True,
                    gl_contract=rec["flat_bid"], gl_etc=rec["flat_cost"],
                    needs="Backlog — expected flatwork, not scheduled yet",
                    gl_sheet=rec["source"], gl_row=rec["gl_row"]))

    n_new = sum(1 for i in items if i["group"].startswith("NEW"))
    n_chg = sum(1 for i in items if i["group"] == "CHANGED")
    n_match = sum(1 for i in items if i["group"] == "MATCHES")
    print(f"  Schedule-active: NEW {n_new} · CHANGED {n_chg} · MATCHES {n_match}")
    print(f"  ⚠ GL active NOT on schedule: {n_notsched} · "
          f"FTW backlog (GL): {n_backlog} · FTW taken by OTHER: {n_other}")

    if args.commit:
        # Split the schedule-active jobs (exclude the GL cross-check rows):
        # READY = contract AND budget known → the RP WIP; MISSING = the gap.
        xcheck = {"⚠ IN GENERAL LISTA, NOT ON SCHEDULE",
                  "FTW BACKLOG (GL, not scheduled)", "FLATWORK TAKEN BY OTHER"}
        active = [i for i in items if i["group"] not in xcheck]
        backlog = [i for i in items
                   if i["group"] == "FTW BACKLOG (GL, not scheduled)"]
        ready = [i for i in active
                 if i["new_contract"] is not None and i["new_etc"] is not None]
        missing = [i for i in active
                   if i["new_contract"] is None or i["new_etc"] is None]
        not_sched = [i for i in items
                     if i["group"] == "⚠ IN GENERAL LISTA, NOT ON SCHEDULE"]
        print(f"  COMMIT: READY {len(ready)} → 'Test - RP' + audit · "
              f"MISSING {len(missing)} + not-on-schedule {len(not_sched)} · "
              f"backlog {len(backlog)}")
        try:
            wrote, rowmap = commit_to_test_rp(ready, backlog, label)
        except W.WipWriteDenied as e:
            print(f"  ✗ Guard blocked write: {e}")
            return 2
        if wrote:
            print("  ✓ READY jobs written to the 'Test - RP' WIP tab")
        # ONE file: WIP audit (with QBO billed/costs from the enriched rows)
        # + MISSING. Replaces the separate Missing/Justification files.
        wip_audit = [(it, rowmap.get(it["line"])) for it in ready]
        one_out = Path(os.getenv("RP_WIP_ONEFILE",
                   str(Path.home() / "Downloads" / "RP WIP.xlsx")))
        write_one_file(wip_audit, missing, not_sched, label, one_out)
        return 0

    out = Path(os.getenv("RP_SCHED_PREVIEW_XLSX",
               str(Path.home() / "Downloads"
                   / "RP WIP - Schedule Method Preview.xlsx")))
    write_report(items, label, out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
