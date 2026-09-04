#!/usr/bin/env python3
"""
schedule_report.py — weekly crew-schedule / stage Gantt (Excel + HTML).

Reads the daily crew schedules (`…/OPERATIONS/SCHEDULE/<yr>/<month>/Schedule
M-D-YY.xlsx`, 'Daily Schedule' tab) for the latest work-week (Mon–Fri) and
builds a Gantt-style view: one row per job (address + builder), one column per
day, each cell the STAGE that day (Pour / Wreck / Forms / Cables / …), coloured
by stage. Shows what stage each job is currently in and how it moved through the
week. Pricing is matched best-effort from the WIP master (RP project #/contract
by address).

READ-ONLY. Output (chmod 600):
  ~/Documents/CompanyHealth/Weekly Schedule.xlsx
  ~/Documents/CompanyHealth/Weekly Schedule.html

USAGE
  python3 one-offs/schedule_report.py --open
  python3 one-offs/schedule_report.py --week 2026-07-20   # any date in the week
"""
from __future__ import annotations

import argparse
import datetime as dt
import html
import os
import re
import stat
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import paths
from shared import schedule as sched   # model logic lives in shared/

SCHEDULE_DIR = sched.SCHEDULE_DIR
WIP_PATH = sched.WIP_PATH
stage_cat = sched.stage_cat
find_week_files = sched.find_week_files
build_price_map = sched.build_price_map
build_model = sched.build_model
_STAGE_RULES = sched._STAGE_RULES
_DEFAULT_STAGE = sched._DEFAULT_STAGE

CH = paths.companyhealth_dir()
XLSX_OUT = CH / "Weekly Schedule.xlsx"
HTML_OUT = CH / "Weekly Schedule.html"


# ────────────────────────── Excel ──────────────────────────

_NAVY = PatternFill("solid", fgColor="1F3864")
_ZEBRA = PatternFill("solid", fgColor="EEF3FA")
_WHITE = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D7E5")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CUR = '#,##0'


def _abbrev(stage: str) -> str:
    cat, _ = stage_cat(stage)
    return cat


def write_xlsx(path: Path, model: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Gantt"
    ws.sheet_view.showGridLines = False
    dates = model["dates"]
    hdr = ["PROJECT #", "ADDRESS", "BUILDER", "CONTRACT $", "CURRENT STAGE"] + \
          [d.strftime("%a %m/%d") for d in dates]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([11, 30, 26, 13, 14] + [13] * len(dates)):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "F2"

    for i, j in enumerate(model["jobs"]):
        row = [j["proj"], j["address"], j["builder"], j["contract"], j["current"]]
        for d in dates:
            row.append(_abbrev(j["days"][d]) if d in j["days"] else "")
        ws.append(row)
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(r, c).border = _BORDER
            if i % 2:
                ws.cell(r, c).fill = _ZEBRA
        ws.cell(r, 4).number_format = CUR
        for k, d in enumerate(dates):
            cell = ws.cell(r, 6 + k)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if d in j["days"]:
                _, color = stage_cat(j["days"][d])
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
    last = ws.max_row
    ws.auto_filter.ref = f"A1:E{last}"

    # legend
    ws.append([])
    ws.append(["Legend:"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    seen = {}
    for _, cat in _STAGE_RULES:
        seen[cat[0]] = cat[1]
    seen[_DEFAULT_STAGE[0]] = _DEFAULT_STAGE[1]
    lr = ws.max_row
    col = 2
    for name, color in seen.items():
        c = ws.cell(lr, col, name)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center")
        col += 1
    ws.sheet_properties.tabColor = "305496"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)


# ────────────────────────── HTML ──────────────────────────

def _money(v) -> str:
    return f"${v:,.0f}" if isinstance(v, (int, float)) else "—"


def write_html(path: Path, model: dict) -> None:
    dates = model["dates"]
    day_hdr = "".join(f"<th>{d.strftime('%a<br>%m/%d')}</th>" for d in dates)
    rows = ""
    for j in model["jobs"]:
        cells = ""
        for d in dates:
            if d in j["days"]:
                cat, color = stage_cat(j["days"][d])
                cells += (f'<td class="cell" style="background:#{color}" '
                          f'title="{html.escape(j["days"][d])}">{html.escape(cat)}</td>')
            else:
                cells += '<td class="cell empty"></td>'
        cur_cat, cur_color = stage_cat(j["current"])
        rows += (f'<tr><td class="proj">{html.escape(j["proj"])}</td>'
                 f'<td class="addr">{html.escape(j["address"])}'
                 f'<span class="bld">{html.escape(j["builder"])}</span></td>'
                 f'<td class="price">{_money(j["contract"])}</td>'
                 f'<td><span class="chip" style="background:#{cur_color}">{html.escape(cur_cat)}</span></td>'
                 f'{cells}</tr>')
    legend = ""
    seen = {}
    for _, cat in _STAGE_RULES:
        seen[cat[0]] = cat[1]
    seen[_DEFAULT_STAGE[0]] = _DEFAULT_STAGE[1]
    for name, color in seen.items():
        legend += f'<span class="chip" style="background:#{color}">{name}</span> '
    span = f"{dates[0]:%b %d} – {dates[-1]:%b %d, %Y}" if dates else "—"

    path.write_text(f'''<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Weekly Schedule</title>
<style>
  :root {{ --bg:#eef1f6; --card:#fff; --ink:#1c2430; --muted:#6b7280;
    --navy:#1F3864; --line:rgba(120,130,150,.2); }}
  @media (prefers-color-scheme: dark) {{ :root {{ --bg:#0f1420; --card:#171e2b;
    --ink:#e6eaf2; --muted:#98a2b3; --line:rgba(150,160,180,.16); }} }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--bg); color:var(--ink);
    font:14px/1.4 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; }}
  header {{ background:var(--navy); color:#fff; padding:16px 26px; }}
  header h1 {{ margin:0; font-size:20px; }}
  header .sub {{ opacity:.82; font-size:12.5px; margin-top:3px; }}
  main {{ max-width:1300px; margin:0 auto; padding:20px 26px 44px; }}
  .legend {{ margin:0 0 14px; }}
  .chip {{ display:inline-block; color:#fff; font-size:11px; font-weight:700;
    padding:3px 9px; border-radius:20px; margin:2px 3px 0 0; }}
  table {{ width:100%; border-collapse:separate; border-spacing:0; background:var(--card);
    border-radius:10px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,.07); }}
  th, td {{ padding:8px 9px; border-bottom:1px solid var(--line); text-align:left; }}
  thead th {{ background:var(--navy); color:#fff; font-size:12px; text-align:center;
    position:sticky; top:0; }}
  thead th:nth-child(-n+2) {{ text-align:left; }}
  td.proj {{ font-weight:700; color:var(--muted); white-space:nowrap; }}
  td.addr {{ font-weight:600; }}
  td.addr .bld {{ display:block; font-size:11px; color:var(--muted); font-weight:400; }}
  td.price {{ text-align:right; font-weight:700; font-variant-numeric:tabular-nums;
    white-space:nowrap; }}
  td.cell {{ text-align:center; color:#fff; font-size:11px; font-weight:700; }}
  td.cell.empty {{ background:transparent; }}
  tbody tr:nth-child(even) td:nth-child(-n+4) {{ background:rgba(120,130,150,.06); }}
</style></head><body>
<header><h1>Weekly Schedule — stage Gantt</h1>
  <div class="sub">Week of {span} · {len(model["jobs"])} jobs · generated
    {dt.datetime.now():%Y-%m-%d %H:%M} · stage = crew task that day, pricing matched
    from the WIP master by address</div></header>
<main>
  <div class="legend">{legend}</div>
  <table><thead><tr><th>Project #</th><th>Address / Builder</th><th>Contract</th>
    <th>Current stage</th>{day_hdr}</tr></thead><tbody>{rows}</tbody></table>
</main></body></html>''', encoding="utf-8")
    os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)


def main() -> int:
    ap = argparse.ArgumentParser(description="Weekly crew-schedule stage Gantt")
    ap.add_argument("--week", type=str, help="any date in the target week (YYYY-MM-DD)")
    ap.add_argument("--xlsx", type=Path, default=XLSX_OUT)
    ap.add_argument("--html", type=Path, default=HTML_OUT)
    ap.add_argument("--open", action="store_true")
    args = ap.parse_args()

    target = None
    if args.week:
        try:
            target = dt.date.fromisoformat(args.week)
        except ValueError:
            print(f"  ✗ bad --week date: {args.week}")
            return 2

    print("\n  WEEKLY SCHEDULE — stage Gantt")
    print("  " + "─" * 40)
    if not SCHEDULE_DIR.is_dir():
        print(f"  ✗ schedule volume not mounted: {SCHEDULE_DIR}")
        return 2
    week = find_week_files(target)
    if not week:
        print("  ✗ no schedule files found for the target week")
        return 3
    print(f"  week: {week[0][0]} → {week[-1][0]} ({len(week)} day file(s))")
    pmap = build_price_map()
    print(f"  price map: {len(pmap['by_addr'])} addresses "
          f"(General List + WIP master), {len(pmap['by_proj'])} priced projects")
    model = build_model(week, pmap)
    matched = sum(1 for j in model["jobs"] if j["contract"] is not None)
    print(f"  {len(model['jobs'])} jobs · {matched} price-matched")

    write_xlsx(args.xlsx, model)
    print(f"  ✓ {args.xlsx}")
    write_html(args.html, model)
    print(f"  ✓ {args.html}")
    if args.open:
        import webbrowser
        webbrowser.open(f"file://{args.xlsx}")
        webbrowser.open(f"file://{args.html}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
