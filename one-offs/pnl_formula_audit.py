#!/usr/bin/env python3
"""
pnl_formula_audit.py — do the P&L's cross-sheet formulas still point at the
columns they mean?

The failure this exists for (2026-09-03): the gutter post-pass shifts every
sheet one column right and rewrites the formulas to match, but its reference
regex required a ROW NUMBER - so WHOLE-COLUMN references were never shifted.
The P&L totals its COGS with

    =SUMIF(Transactions!$D:$D, "<account>", Transactions!$E:$E)

and those kept pointing one column left after Transactions moved. "Costs to
Date" read **0.00**, and every figure derived from it followed silently: %
complete, earned revenue, cost to complete, gross profit, net profit. Nothing
errored, nothing tripped the corruption gate - the workbook was structurally
perfect and arithmetically wrong.

So this checks MEANING, not structure: it reads the Transactions sheet's own
header row to find where Account and Amount actually are, then confirms every
SUMIF on the P&L agrees. Run it over delivered workbooks after any layout
change, the way `pnl_line_level_audit.py` re-proves the cost attribution.

USAGE
  python3 one-offs/pnl_formula_audit.py <workbook.xlsx> [more...]
  find "<P&L root>" -name "Project_PnL_*.xlsx" -exec python3 one-offs/pnl_formula_audit.py {} +
"""
import re, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def check(path):
    try:
        wb = load_workbook(path)
    except Exception as e:
        return [f"unreadable: {e}"]
    out = []
    pl = next((wb[n] for n in ("P&L", "Job P&L") if n in wb.sheetnames), None)
    if pl is None or "Transactions" not in wb.sheetnames:
        return out
    tx = wb["Transactions"]
    hdr = next((r for c0 in (1, 2) for r in range(1, 90)
                if str(tx.cell(r, c0).value or "").strip() == "Ref #"), None)
    if hdr is None:
        return out
    cols = {str(tx.cell(hdr, c).value or "").strip().lower(): c
            for c in range(1, 14)}
    acct = cols.get("account"); amt = cols.get("amount")
    if not acct or not amt:
        return out
    want_a, want_m = get_column_letter(acct), get_column_letter(amt)
    for row in pl.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=") and "SUMIF" in v):
                continue
            m = re.search(r"SUMIF\(\s*Transactions!\$?([A-Z]{1,3})\$?:", v)
            m2 = re.search(r",\s*Transactions!\$?([A-Z]{1,3})\$?:\$?[A-Z]{1,3}\$?\s*\)", v)
            if m and m.group(1) != want_a:
                out.append(f"{cell.coordinate}: SUMIF criteria column is "
                           f"{m.group(1)} but Transactions Account is {want_a}")
            if m2 and m2.group(1) != want_m:
                out.append(f"{cell.coordinate}: SUMIF sum column is "
                           f"{m2.group(1)} but Transactions Amount is {want_m}")
    return out[:3]

for p in sys.argv[1:]:
    iss = check(p)
    name = p.split("/")[-1]
    print(f"{'BROKEN ' if iss else 'ok     '}{name}")
    for i in iss:
        print("     -", i)
