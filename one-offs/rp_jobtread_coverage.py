#!/usr/bin/env python3
"""
rp_jobtread_coverage.py — the estimator's JobTread TO-DO list (the user
2026-07-22). READ-ONLY against JobTread; writes nothing anywhere but a
Downloads Excel.

WHY
  JobTread is the RP pricing source of record (decision 2026-07-22), but
  only a fraction of schedule-active jobs have an approved proposal there.
  This report is the gap as a work-list: for every job on the daily
  schedule, does JobTread have it, and does it carry an APPROVED proposal
  (the one thing the WIP needs)?  Close the list → close the coverage gap.

BUCKETS (per unique job # on the Main Schedule tab)
  • MISSING          — no job record in JobTread → create it + add proposal
  • NEEDS PROPOSAL   — job exists, no approved customerOrder → add proposal
  • COVERED          — approved proposal present (contract + budget shown)

Auth: JT_GRANT_KEY in the shared vault (one Touch ID). Nothing is pushed.

Usage:
  python3 rp_jobtread_coverage.py                       # latest schedule
  python3 rp_jobtread_coverage.py --schedule <file.xlsx>
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.request
import urllib.error
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))

from openpyxl import Workbook

from shared import qbo_vault
import rp_wip_reader as RP
import rp_schedule_wip_preview as P

ORG_ID = os.getenv("JT_ORG_ID", "22PFAfqHLF3a")   # Proficient Concrete LLC
API_URL = "https://api.jobtread.com/pave"


def pave(key: str, query: dict) -> dict:
    body = json.dumps({"query": {"$": {"grantKey": key}, **query}}).encode()
    req = urllib.request.Request(API_URL, data=body,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def job_coverage(key: str, number: str):
    """(status, contract, budget) for one job number. status ∈
    {'MISSING','NEEDS PROPOSAL','COVERED'}. contract/budget are the summed
    price/cost of APPROVED customerOrder docs (None unless COVERED)."""
    r = pave(key, {"organization": {"$": {"id": ORG_ID}, "jobs": {
        "$": {"size": 1, "where": {"and": [["number", "=", number]]}},
        "nodes": {"id": {}, "documents": {"$": {"size": 50}, "nodes": {
            "type": {}, "status": {}, "cost": {}, "price": {}}}}}}})
    nodes = r["organization"]["jobs"]["nodes"]
    if not nodes:
        return "MISSING", None, None
    approved = [d for d in nodes[0]["documents"]["nodes"]
                if d["type"] == "customerOrder" and d["status"] == "approved"]
    if not approved:
        return "NEEDS PROPOSAL", None, None
    contract = sum(float(d["price"] or 0) for d in approved)
    budget = sum(float(d["cost"] or 0) for d in approved)
    return "COVERED", round(contract, 2), round(budget, 2)


TODO = {
    "MISSING": "Create the job in JobTread, then add the approved proposal",
    "NEEDS PROPOSAL": "Add the approved proposal (the job already exists)",
    "COVERED": "Done — approved proposal is in JobTread",
}


def write_report(rows, sched_label, covered, needs, missing, out_path: Path):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    RED = PatternFill("solid", fgColor="F4CCCC")     # MISSING
    AMBER = PatternFill("solid", fgColor="FCE4D6")   # NEEDS PROPOSAL
    GREEN = PatternFill("solid", fgColor="D9EAD3")   # COVERED
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CUR = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    FILL = {"MISSING": RED, "NEEDS PROPOSAL": AMBER, "COVERED": GREEN}

    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        raise SystemExit(f"{out_path.name} is open in Excel — close it first")
    wb = Workbook()
    ws = wb.active
    ws.title = "JT COVERAGE"
    total = covered + needs + missing
    pct = (covered / total * 100) if total else 0
    ws["A1"] = (f"JOBTREAD COVERAGE — estimator TO-DO (schedule {sched_label}). "
                f"{total} active jobs · COVERED {covered} ({pct:.0f}%) · "
                f"NEEDS PROPOSAL {needs} · MISSING {missing}. "
                f"Read-only — nothing was pushed to JobTread.")
    ws["A1"].font = Font(bold=True)
    ws.append([])
    HDR = ["STATUS", "TO DO", "JOB #", "SCHEDULE PHASE", "ADDRESS", "BUILDER",
           "JT CONTRACT $", "JT BUDGET $"]
    ws.append(HDR)
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(3, c)
        cell.font = Font(bold=True)
        cell.fill = GRAY
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    for row in rows:
        ws.append([row["status"], TODO[row["status"]], row["job"],
                   row["phase"], row["address"], row["builder"],
                   row["contract"], row["budget"]])
        r = ws.max_row
        for cc in range(1, len(HDR) + 1):
            ws.cell(r, cc).border = BORDER
            ws.cell(r, cc).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, 1).fill = FILL[row["status"]]
        ws.cell(r, 1).font = Font(bold=True)
        for cc in (7, 8):
            ws.cell(r, cc).number_format = CUR
    for col, w in zip("ABCDEFGH", (16, 46, 12, 18, 26, 26, 15, 15)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(out_path)
    print(f"  ✓ Coverage → {out_path}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--schedule", help="use this schedule file instead of latest")
    args = ap.parse_args()

    print("\n  RP JobTread COVERAGE — estimator to-do (read-only)")
    print("  " + "─" * 56)
    if args.schedule:
        sched_path = Path(args.schedule)
        m = RP._SCHED_FILE_RE.search(sched_path.name)
        label = "-".join(m.groups()) if m else sched_path.stem
    else:
        best = P.latest_schedule(RP.SCHEDULE_DIR)
        if best is None:
            print("  ✗ no schedule file found")
            return 1
        _k, sched_path = best
        label = f"{_k[1]}-{_k[2]}-{_k[0] % 100:02d}"
    if not sched_path.exists():
        print(f"  ✗ schedule not found: {sched_path}")
        return 1
    print(f"  schedule: {sched_path.name}")

    sched = P.read_main_schedule(sched_path)
    # one entry per base job #; keep the first (earliest-listed) phase seen.
    jobs = {}
    for s in sched:
        jobs.setdefault(s["job"], s)
    print(f"  active jobs on the schedule: {len(jobs)}")

    key = qbo_vault.get("JT_GRANT_KEY")
    rows = []
    covered = needs = missing = 0
    for n, (job, s) in enumerate(sorted(jobs.items()), 1):
        try:
            status, contract, budget = job_coverage(key, job)
        except urllib.error.HTTPError as e:
            print(f"  ⚠ {job}: JobTread error {e.code} — skipped")
            continue
        rows.append({"status": status, "job": job, "phase": s["section"],
                     "address": s["address"], "builder": s["builder"],
                     "contract": contract, "budget": budget})
        covered += status == "COVERED"
        needs += status == "NEEDS PROPOSAL"
        missing += status == "MISSING"
        if n % 20 == 0:
            print(f"    …{n}/{len(jobs)} checked")

    # MISSING first, then NEEDS PROPOSAL, then COVERED; cluster by builder.
    rank = {"MISSING": 0, "NEEDS PROPOSAL": 1, "COVERED": 2}
    rows.sort(key=lambda x: (rank[x["status"]], x["builder"], x["job"]))
    total = covered + needs + missing
    pct = (covered / total * 100) if total else 0
    print(f"  COVERED {covered} ({pct:.0f}%) · NEEDS PROPOSAL {needs} · "
          f"MISSING {missing}")

    out = Path(os.getenv("RP_JT_COVERAGE_XLSX",
               str(Path.home() / "Downloads" / "RP JobTread Coverage.xlsx")))
    write_report(rows, label, covered, needs, missing, out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
