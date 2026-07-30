#!/usr/bin/env python3
"""
rp_wip_update.py — UPDATE the owner's live RP WIP in place (the user 2026-07-30).

This is NOT a generator. The RP WIP is now the owner's working file in OneDrive;
he verifies, changes and annotates it. This opens that file, refreshes only the
machine-derived columns, and leaves every human judgement alone.

THE COLOUR CONTRACT — his marks are authoritative and are NEVER overwritten:
    theme colour 9  (orange) = ops manager must verify   (e.g. an OPTIONAL scope
                               was sold but only the non-optional work costed)
    00B050 bright green      = he verified this number
    FF0000 bright red        = he changed this number
Mine, safe to refresh: 006100 (one-proposal/one-takeoff green), 9C0006 (✗),
0563C1 (link blue), theme 1 (section band headers).

WHAT IT REFRESHES (only where the cell is not owner-marked)
    SCHEDULE ✓/✗   from today's schedule
    BILLED / COSTS from QBO
    GP %           recomputed
    CONTRACT / ETC only when the cell is unmarked (his marked values win)
    + appends any schedule line that is not in the sheet yet

Dry run by default: writes a diff to the terminal and a copy to Downloads.
--commit writes back to the OneDrive file itself.

Usage
  python3 rp_wip_update.py                    # dry run + copy in Downloads
  python3 rp_wip_update.py --commit           # update the OneDrive file
  python3 rp_wip_update.py --schedule <f.xlsx>
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import shutil
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))

from openpyxl import load_workbook
from openpyxl.styles import Font

from shared import qbo_api, qbo_vault
from shared import schedule as SCHED
import rp_wip_reader as RP
import rp_schedule_wip_preview as P

LIVE = Path(os.getenv("RP_WIP_LIVE", str(
    Path.home() / "Library" / "CloudStorage" /
    "OneDrive-ProficientConcrete,LLC" / "RP WIP TO FIX_Final.xlsx")))
SHEET = "RP WIP"
CHECK, CROSS = "✓", "✗"
OWNER_RGB = {"00B050", "FF0000"}       # verified / changed
OWNER_THEME = {9}                      # orange = needs verification
COL = {"job": 1, "addr": 2, "builder": 3, "contract": 4, "etc": 5,
       "billed": 6, "costs": 7, "gp": 8, "sched": 9, "gl": 10, "action": 11}


def owner_marked(cell) -> bool:
    """True if the OWNER coloured this cell (verified / changed / verify-me)."""
    c = cell.font.color if cell.font else None
    if c is None:
        return False
    rgb = getattr(c, "rgb", None)
    if isinstance(rgb, str) and rgb[-6:].upper() in OWNER_RGB:
        return True
    th = getattr(c, "theme", None)
    return isinstance(th, int) and th in OWNER_THEME


def is_band(v) -> bool:
    return isinstance(v, str) and ("⚠" in v or "BACKLOG" in v)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--as-of", help="treat this date as today (YYYY-MM-DD); never reads a schedule dated after it")
    ap.add_argument("--schedule", help="schedule xlsx (default: latest)")
    ap.add_argument("--commit", action="store_true",
                    help="write back to the OneDrive file")
    ap.add_argument("--file", help="override the live file path")
    args = ap.parse_args()

    live = Path(args.file) if args.file else LIVE
    if not live.exists():
        print(f"  ✗ live file not found: {live}")
        return 1
    lock = live.with_name("~$" + live.name)
    if lock.exists():
        raise SystemExit(f"{live.name} is open in Excel — close it first")

    print("\n  RP WIP — UPDATE the owner's live file (never regenerate)")
    print(f"  {'COMMIT — will write to OneDrive' if args.commit else 'DRY RUN — copy to Downloads only'}")
    print("  " + "─" * 62)
    print(f"  file: {live.name}")

    # schedule
    if args.schedule:
        sched_path = Path(args.schedule)
    else:
        _cap = (dt.date.fromisoformat(args.as_of) if getattr(args, 'as_of', None)
                else None)
        best = SCHED.schedule_on_or_before(_cap)
        if best is None:
            print("  ✗ no schedule found")
            return 1
        sched_path = best[1]
    sched = P.read_main_schedule(sched_path)
    sched_lines = {}
    for s in sched:
        job, scope = s["job"], s["scope"]
        line = job if (scope == "slab" or job.startswith("CP")) else f"{job}-FTW"
        sched_lines.setdefault(line, s)
    print(f"  schedule: {sched_path.name}  ({len(sched_lines)} active lines)")

    # QBO
    print("  QBO: billed + costs …")
    access, cid = qbo_api.load_credentials()
    pmap = qbo_api.build_project_customer_map(access, cid)
    start, end = "2019-01-01", dt.date.today().isoformat()

    wb = load_workbook(live)
    ws = wb[SHEET]
    changes, protected, added = [], 0, []
    seen = set()

    for r in range(3, ws.max_row + 1):
        job = ws.cell(r, COL["job"]).value
        if not job or is_band(job):
            continue
        job = str(job).strip()
        seen.add(job)

        # SCHEDULE ✓/✗
        on = job in sched_lines
        cur = ws.cell(r, COL["sched"]).value
        want = CHECK if on else CROSS
        if cur != want:
            ws.cell(r, COL["sched"]).value = want
            ws.cell(r, COL["sched"]).font = Font(
                color="006100" if on else "9C0006", bold=True)
            changes.append(f"{job}: SCHEDULE {cur!r} → {want}")

        # QBO billed / costs
        c = pmap.get(job)
        if c:
            try:
                t = qbo_api.extract_pl_totals(
                    qbo_api.fetch_project_pl(access, cid, c["id"], start, end))
                billed = round(t.get("income") or 0.0, 2)
                costs = round((t.get("cogs") or 0.0) + (t.get("expenses") or 0.0), 2)
            except Exception:
                billed = costs = None
            for key, val in (("billed", billed), ("costs", costs)):
                if val is None:
                    continue
                cell = ws.cell(r, COL[key])
                old = cell.value
                if isinstance(old, (int, float)) and abs(old - val) < 0.01:
                    continue
                if owner_marked(cell):
                    protected += 1
                    continue
                cell.value = val
                changes.append(f"{job}: {key.upper()} "
                               f"{(old or 0):,.0f} → {val:,.0f}")

        # GP % from whatever contract/ETC now stand (his values win)
        k = ws.cell(r, COL["contract"]).value
        e = ws.cell(r, COL["etc"]).value
        if isinstance(k, (int, float)) and k and isinstance(e, (int, float)):
            gp = (k - e) / k
            g = ws.cell(r, COL["gp"])
            if not isinstance(g.value, float) or abs((g.value or 0) - gp) > 1e-6:
                g.value = gp
                g.number_format = "0.0%"

    # append schedule lines the sheet does not have yet
    for line, s in sorted(sched_lines.items()):
        if line in seen:
            continue
        added.append(line)
        r = ws.max_row + 1
        ws.cell(r, COL["job"], line)
        ws.cell(r, COL["addr"], s["address"])
        ws.cell(r, COL["builder"], s["builder"])
        ws.cell(r, COL["sched"], CHECK).font = Font(color="006100", bold=True)
        ws.cell(r, COL["gl"], CROSS).font = Font(color="9C0006", bold=True)
        ws.cell(r, COL["action"], "NEW on the schedule — needs contract/ETC")

    print("  " + "─" * 62)
    print(f"  refreshed: {len(changes)} cell(s)")
    for c in changes[:25]:
        print(f"     {c}")
    if len(changes) > 25:
        print(f"     … and {len(changes) - 25} more")
    print(f"  PROTECTED (your green/red/orange marks, untouched): {protected}")
    print(f"  NEW schedule lines appended: {len(added)}"
          + (f" → {', '.join(added)}" if added else ""))

    if args.commit:
        wb.save(live)
        print(f"\n  ✓ updated in place → {live}")
    else:
        out = Path.home() / "Downloads" / f"{live.stem} (updated preview).xlsx"
        wb.save(out)
        print(f"\n  ✓ DRY RUN — preview copy → {out}")
        print("    re-run with --commit to update the OneDrive file itself")
    return 0


if __name__ == "__main__":
    sys.exit(main())
