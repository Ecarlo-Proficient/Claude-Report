#!/usr/bin/env python3
"""
pnl_line_level_audit.py — prove a project P&L used LINE amounts, not bill totals.

THE FAILURE THIS EXISTS TO CATCH (the owner, standing instruction 2026-08-27):
"do not mix bill totals with line totals, that's where we can get a good job
made to look like a very bad job."

Most subcontractor bills are multi-line and span several jobs. If a job matches
ONE line on a bill and the code banks the bill's TOTAL, the job absorbs every
other job's money on that document. It does not error, it does not look wrong —
it just turns a profitable job into a loss. On MFD228 the two numbers are
879,732 (correct, line-level) and 1,516,919 (bill totals): a 72% overstatement
that would read as a catastrophic overrun.

WHAT IT DOES, per job:
  * re-derives cost from QBO INDEPENDENTLY of the workbook, at line level
  * sums the TOTALS of every bill it touched — the number a bill-total bug
    would have produced — and reports the gap
  * counts PARTIAL bills (only some lines are this job's), which is where the
    two diverge
  * reads the delivered workbook's own cost and asserts they MATCH

A MISMATCH means the workbook and QBO disagree. A large "bill totals" gap is
NOT a problem — it is the measure of how much this check is protecting you.

USAGE
  python3 one-offs/pnl_line_level_audit.py
  python3 one-offs/pnl_line_level_audit.py --folder "<archive folder>"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import pnl_paths, qbo_api
from shared.job_lines import JobMatcher, discover_job_classes

SINCE = "2021-01-01"
# Jobs whose P&L was built with the line-text / bill-memo rules ON, and the
# street name each was given. Everything else is project ∪ class only.
ALIASES: Dict[str, List[str]] = {"MFD172": ["BONDS RANCH"], "MFD228": ["LAPIZ"]}


def workbook_cost(path: Path):
    """The cost the DELIVERED workbook actually reports, from its Transactions
    sheet. A row below the income total with an account in col D is a cost
    line; without one it is the retainage block, which is income."""
    wb = load_workbook(str(path), data_only=False)
    try:
        ws = wb["Transactions"]
        hdr = next((r for r in range(1, 60)
                    if str(ws.cell(r, 1).value or "").strip() == "Inv #"), None)
        end = next((r for r in range(hdr + 1, ws.max_row + 1)
                    if str(ws.cell(r, 1).value or "").startswith("TOTAL")), None)
        if hdr is None or end is None:
            return None, 0
        tot = n = 0.0, 0
        s = 0.0
        n = 0
        for r in range(end + 1, ws.max_row + 1):
            v = ws.cell(r, 5).value
            if isinstance(v, (int, float)) and ws.cell(r, 4).value:
                s += float(v)
                n += 1
        return round(s, 2), n
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(description="Line-level vs bill-total audit")
    ap.add_argument("--folder", default=None)
    a = ap.parse_args()
    folder = (Path(a.folder).expanduser() if a.folder
              # inside the DIVISION folder - P&Ls are sorted by division so a
              # folder link can be shared with one PM (the user 2026-08-31)
              else pnl_paths.division_dir("MFD") / "completed mfd project p&l")
    jobs = sorted(d.name for d in folder.iterdir()
                  if d.is_dir() and (d / f"Project_PnL_{d.name}.xlsx").exists())
    if not jobs:
        print(f"✗  no project workbooks under {folder}")
        return 1

    access, cid = qbo_api.load_credentials()
    pmap = qbo_api.build_project_customer_map(access, cid)
    classes = (qbo_api.query_all(access, cid, "Class")
               + qbo_api.query_all(access, cid, "Class", "Active = false"))
    txns = ([(t, "Bill") for t in qbo_api.query_all(access, cid, "Bill", f"TxnDate >= '{SINCE}'")]
            + [(t, "Purchase") for t in qbo_api.query_all(access, cid, "Purchase", f"TxnDate >= '{SINCE}'")])
    print(f"pulled {len(txns):,} bills + purchases since {SINCE}\n")

    hdr = (f"{'JOB':<8} {'LINE-LEVEL (used)':>19} {'if BILL TOTALS':>18} "
           f"{'overstated by':>15} {'partial':>8}  workbook")
    print(hdr)
    print("-" * len(hdr))
    bad = []
    for j in jobs:
        info = pmap.get(j)
        if not info:
            print(f"{j:<8} not in QBO")
            continue
        cls = discover_job_classes(classes, j)
        al = ALIASES.get(j, [])
        m = JobMatcher(info["id"], j, al, legacy=True,
                       class_ids=list(cls.keys()), text_rules=bool(al))
        line_tot = 0.0
        bills: Dict[str, float] = {}
        partial = 0
        for t, _ent in txns:
            lines = [ln for ln in (t.get("Line") or [])
                     if (ln.get("AccountBasedExpenseLineDetail")
                         or ln.get("ItemBasedExpenseLineDetail"))]
            hits = [ln for ln in lines
                    if m((ln.get("AccountBasedExpenseLineDetail")
                          or ln.get("ItemBasedExpenseLineDetail")), ln, t)]
            if not hits:
                continue
            line_tot += sum(float(ln.get("Amount", 0) or 0) for ln in hits)
            bills[t["Id"]] = float(t.get("TotalAmt", 0) or 0)
            if len(hits) < len(lines):
                partial += 1
        bill_tot = round(sum(bills.values()), 2)
        line_tot = round(line_tot, 2)
        wb_cost, wb_n = workbook_cost(folder / j / f"Project_PnL_{j}.xlsx")
        ok = wb_cost is not None and abs(wb_cost - line_tot) < 1.0
        if not ok:
            bad.append((j, line_tot, wb_cost))
        infl = (bill_tot / line_tot - 1) * 100 if line_tot else 0
        print(f"{j:<8} {line_tot:>19,.2f} {bill_tot:>18,.2f} {infl:>14.1f}% "
              f"{partial:>8}  {'✓ matches' if ok else f'✗ workbook {wb_cost:,.2f}'}")
    print()
    if bad:
        print("✗ MISMATCH — the workbook disagrees with QBO on these jobs:")
        for j, q, w in bad:
            print(f"    {j}: QBO line-level {q:,.2f} vs workbook {w:,.2f} "
                  f"(diff {(w or 0) - q:,.2f})")
        return 1
    print("✓ every workbook's cost equals an independent LINE-LEVEL pull from QBO.")
    print("  The 'if BILL TOTALS' column is what the numbers would have been had")
    print("  the code banked whole bills — that gap is the damage this prevents.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
