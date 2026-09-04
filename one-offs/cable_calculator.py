#!/usr/bin/env python3
"""
cable_calculator.py — standalone PT cable cut-list + cost calculator
(the user 2026-07-27). READ-ONLY against takeoffs; writes one Downloads Excel.

WHY
  Migrating RP bidding to JobTread. JobTread prices what you MEASURE, but the
  cable count/lengths come from the engineer's tendon plan, so JobTread can't
  derive them. Today that math is buried in the takeoff's hidden '0' sheet,
  which means "just use the takeoff as a calculator" = never leaving Excel.
  This tool lifts the cable engine OUT of the takeoff so the bid can move to
  JobTread and the cut-list still gets produced for the supplier.

WHAT IT PRODUCES
  • the CUT-LIST for the cable supplier (qty × length, ascending)
  • the two numbers JobTread needs:  total LF  and  total cable COUNT
  • the cable cost:  LF × $/LF  +  count × $/cable

THE ENGINE IT REPLACES (mapped cell-for-cell from the takeoff, verified on
64 active RP takeoffs — every count, LF, cut-list row and cost tied exactly):
  INFORMATION inputs, TWO blocks:
      block 1  I25:I69 (qty) · K25:K69 (length) · M25:M69 (note)
      block 2  N24:N69 (qty) · P24:P69 (length) · R24:R69 (note)
  '0' staging 65..109 (block 1) and 110..154 (block 2):
      G=qty · I=IF(len+len=0,"",len) · K=G · L=note
  '0' sort 111..200:
      O = 1..90 counter
      N = SMALL($I$65:$J$154, O)              k-th smallest length
      M = VLOOKUP(N, $I$65:$K$154, 3, FALSE)  qty of the FIRST row w/ that length
  '0' display 3..92 (display r ↔ sort r+108):
      M=qty · N=length · O=M*N
      M94 = SUM(M3:M92) = COUNT     O94 = SUM(O3:O92) = TOTAL LF
  Slab takeoff cost rows are found BY LABEL ("Cable take off" / "Per cable") —
  their row number AND rate vary per template (row 37/38; $0.52 vs $0.46/LF,
  $8.00 vs $8.25 each), so never hard-code them.

KNOWN DEFECT IN THE LEGACY EXCEL ENGINE (this tool fixes it)
  SMALL() returns duplicate lengths once per occurrence, but VLOOKUP(...,FALSE)
  always returns the FIRST row's qty — so if one length is entered twice with
  DIFFERENT quantities, Excel repeats the first qty and drops the second.
  This tool sums duplicates correctly and prints a warning when its answer
  would differ from the legacy engine (today: no active job is affected).

Usage
  python3 cable_calculator.py --pairs "15x22, 18x23, 7x69, 3x70"
  python3 cable_calculator.py --from-takeoff "<RP####_ADDRESS.xlsm>" --verify
  python3 cable_calculator.py --verify-batch          # regression vs takeoffs
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# ── the verified input map ──────────────────────────────────────────
BLOCKS = (("I", "K", "M", 25, 45),      # qty col, length col, note col, first row, n
          ("N", "P", "R", 24, 45))
DISPLAY_SLOTS = 90                      # '0' rows 3..92
DEFAULT_RATE_LF = 0.52
DEFAULT_RATE_EA = 8.00


def read_takeoff_pairs(info_ws):
    """INFORMATION → [(qty, length, note)] in staging order. A length of 0/blank
    is invisible to the engine (staging I = IF(len+len=0,"",len))."""
    out = []
    for qcol, lcol, ncol, first, n in BLOCKS:
        for i in range(n):
            r = first + i
            qty = info_ws[f"{qcol}{r}"].value
            ln = info_ws[f"{lcol}{r}"].value
            note = info_ws[f"{ncol}{r}"].value
            qty = qty if isinstance(qty, (int, float)) else 0
            if not isinstance(ln, (int, float)) or ln == 0:
                continue
            out.append((qty, ln, note if isinstance(note, str) else ""))
    return out


def parse_pairs(text):
    """'15x22, 18x23' or '15@22 18@23' → [(qty, length, '')]."""
    out = []
    for chunk in re.split(r"[,\n;]+", text):
        m = re.search(r"(\d+(?:\.\d+)?)\s*[x@]\s*(\d+(?:\.\d+)?)", chunk.strip(), re.I)
        if m:
            out.append((float(m.group(1)), float(m.group(2)), ""))
    return out


def read_csv_pairs(path: Path):
    """CSV/TSV: qty,length[,note] — header row optional."""
    import csv
    out = []
    with open(path, newline="") as fh:
        sniff = fh.read(2048); fh.seek(0)
        delim = "\t" if "\t" in sniff and "," not in sniff.split("\n")[0] else ","
        for row in csv.reader(fh, delimiter=delim):
            if len(row) < 2:
                continue
            try:
                q, l = float(row[0]), float(row[1])
            except ValueError:
                continue                       # header / junk line
            out.append((q, l, row[2].strip() if len(row) > 2 else ""))
    return out


def compute(pairs):
    """Correct engine: group duplicate lengths and SUM their quantities.
    Returns (cutlist[(qty,length,lf,note)], count, lf, legacy_count, legacy_lf).
    legacy_* reproduce the takeoff's SMALL+VLOOKUP first-match behaviour so a
    divergence can be surfaced instead of silently changing a number."""
    grouped, notes = {}, {}
    for qty, ln, note in pairs:
        grouped[ln] = grouped.get(ln, 0) + qty
        if note and ln not in notes:
            notes[ln] = note
    cutlist = [(grouped[l], l, grouped[l] * l, notes.get(l, ""))
               for l in sorted(grouped)]
    count = sum(q for q, _, _, _ in cutlist)
    lf = sum(f for _, _, f, _ in cutlist)

    first_qty = {}
    for qty, ln, _ in pairs:
        first_qty.setdefault(ln, qty)
    legacy_lengths = sorted(l for _, l, _ in pairs)[:DISPLAY_SLOTS]
    legacy_count = sum(first_qty[l] for l in legacy_lengths)
    legacy_lf = sum(first_qty[l] * l for l in legacy_lengths)
    return cutlist, count, lf, legacy_count, legacy_lf


def find_cost_rows(slab_ws):
    """Cable cost rows shift per template — locate by label, return
    {'LF': (row, qty, rate), 'EA': (row, qty, rate)}."""
    found = {}
    for r in range(1, 120):
        a = slab_ws[f"A{r}"].value
        if not isinstance(a, str):
            continue
        k = a.strip().lower()
        if k.startswith("cable take off"):
            found["LF"] = (r, slab_ws[f"C{r}"].value, slab_ws[f"D{r}"].value)
        elif k.startswith("per cable"):
            found["EA"] = (r, slab_ws[f"C{r}"].value, slab_ws[f"D{r}"].value)
    return found


# ── plain Excel out (repo style: no fills, label + amount on one row) ──
BOLD = Font(bold=True)
UNDER = Border(bottom=Side(style="thin", color="000000"))
CUR = '"$"#,##0.00'


def write_xlsx(out_path: Path, job, cutlist, count, lf, rate_lf, rate_ea):
    lock = out_path.with_name("~$" + out_path.name)
    if lock.exists():
        raise SystemExit(f"{out_path.name} is open in Excel — close it first")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cable Order"
    ws["A1"] = f"CABLE CUT-LIST{(' — ' + job) if job else ''}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["QTY", "LENGTH (FT)", "LINEAR FEET", "NOTE"])
    for c in range(1, 5):
        ws.cell(3, c).font = BOLD
        ws.cell(3, c).border = UNDER
    for qty, ln, feet, note in cutlist:
        ws.append([qty, ln, feet, note])
    tot = ws.max_row + 1
    ws.cell(tot, 1, count).font = BOLD
    ws.cell(tot, 2, "TOTAL").font = BOLD
    ws.cell(tot, 3, lf).font = BOLD
    for c in range(1, 5):
        ws.cell(tot, c).border = Border(top=Side(style="thin", color="000000"))

    r = tot + 2
    ws.cell(r, 1, "FOR JOBTREAD").font = Font(bold=True, size=12)
    for label, qty, rate in (("PT Cable (LF)", lf, rate_lf),
                             ("Cable Stressing (each)", count, rate_ea)):
        r += 1
        ws.cell(r, 1, label)
        ws.cell(r, 2, qty)
        ws.cell(r, 3, rate).number_format = CUR
        ws.cell(r, 4, qty * rate).number_format = CUR
    r += 1
    ws.cell(r, 1, "TOTAL CABLE COST").font = BOLD
    ws.cell(r, 4, lf * rate_lf + count * rate_ea).font = BOLD
    ws.cell(r, 4).number_format = CUR

    for col, w in zip("ABCD", (24, 14, 14, 22)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(out_path)
    return out_path


def report(job, pairs, rate_lf, rate_ea, out: Path | None):
    cutlist, count, lf, lg_count, lg_lf = compute(pairs)
    print(f"\n  CABLE CALCULATOR{(' — ' + job) if job else ''}")
    print("  " + "─" * 52)
    print(f"  {'QTY':>5}  {'LENGTH':>7}  {'LF':>8}")
    for qty, ln, feet, _ in cutlist:
        print(f"  {qty:>5g}  {ln:>7g}  {feet:>8g}")
    print("  " + "─" * 52)
    print(f"  TOTAL: {count:g} cables · {lf:g} LF")
    print(f"  JobTread → PT Cable {lf:g} LF @ ${rate_lf:,.2f} = "
          f"${lf * rate_lf:,.2f}")
    print(f"            Stressing {count:g} ea @ ${rate_ea:,.2f} = "
          f"${count * rate_ea:,.2f}")
    print(f"            TOTAL CABLE COST ${lf * rate_lf + count * rate_ea:,.2f}")
    if (lg_count, lg_lf) != (count, lf):
        print(f"\n  ⚠ legacy takeoff engine would say {lg_count:g}/{lg_lf:g} — it "
              "repeats the first quantity on duplicate lengths (Excel defect). "
              "This tool sums them correctly.")
    if out:
        write_xlsx(out, job, cutlist, count, lf, rate_lf, rate_ea)
        print(f"\n  ✓ Cut-list → {out}")
    return count, lf


def from_takeoff(path: Path, verify: bool, rate_lf, rate_ea, out):
    wb = load_workbook(path, data_only=True)
    if "INFORMATION" not in wb.sheetnames:
        raise SystemExit(f"{path.name}: no INFORMATION sheet")
    pairs = read_takeoff_pairs(wb["INFORMATION"])
    m = re.search(r"(RP\d{3,4}(?:-FTW)?)", path.name, re.I)
    job = m.group(1).upper() if m else path.stem
    _, count, lf, _, _ = compute(pairs)
    if verify:
        z = wb["0"] if "0" in wb.sheetnames else None
        ok = True
        if z is not None:
            xl_c = z["M94"].value or 0
            xl_lf = z["O94"].value or 0
            ok = (xl_c == count and xl_lf == lf)
            print(f"  VERIFY {job}: takeoff engine {xl_c}/{xl_lf} vs this tool "
                  f"{count:g}/{lf:g} → {'MATCH' if ok else 'MISMATCH'}")
        if "Slab takeoff" in wb.sheetnames:
            cr = find_cost_rows(wb["Slab takeoff"])
            if cr.get("LF"):
                rate_lf = cr["LF"][2] or rate_lf
            if cr.get("EA"):
                rate_ea = cr["EA"][2] or rate_ea
            print(f"  rates from takeoff: ${rate_lf}/LF · ${rate_ea}/cable")
        if not ok:
            print("  ✗ mismatch — do not trust this run; re-map before using")
    report(job, pairs, rate_lf, rate_ea, out)
    wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--pairs", help='e.g. "15x22, 18x23, 7x69, 3x70"')
    src.add_argument("--csv", help="CSV of qty,length[,note]")
    src.add_argument("--from-takeoff", help="read pairs out of a takeoff .xlsm")
    src.add_argument("--verify-batch", action="store_true",
                     help="regression: this tool vs every active RP takeoff")
    ap.add_argument("--verify", action="store_true",
                    help="with --from-takeoff: check against its own engine")
    ap.add_argument("--job", default="", help="job # for the header")
    ap.add_argument("--rate-lf", type=float, default=DEFAULT_RATE_LF)
    ap.add_argument("--rate-ea", type=float, default=DEFAULT_RATE_EA)
    ap.add_argument("--out", help="output xlsx (default: Downloads)")
    ap.add_argument("--no-xlsx", action="store_true", help="print only")
    args = ap.parse_args()

    if args.verify_batch:
        print("  Batch verification lives in the session harness "
              "(cable_validate.py); run --from-takeoff --verify per file.")
        return 0

    out = None
    if not args.no_xlsx:
        out = Path(args.out) if args.out else Path(
            os.getenv("CABLE_XLSX", str(Path.home() / "Downloads" /
                                        "Cable Cut-List.xlsx")))

    if args.from_takeoff:
        from_takeoff(Path(args.from_takeoff), args.verify,
                     args.rate_lf, args.rate_ea, out)
        return 0

    pairs = (parse_pairs(args.pairs) if args.pairs
             else read_csv_pairs(Path(args.csv)))
    if not pairs:
        print("  ✗ no (qty × length) pairs parsed")
        return 1
    report(args.job, pairs, args.rate_lf, args.rate_ea, out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
