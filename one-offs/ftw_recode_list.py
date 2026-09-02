#!/usr/bin/env python3
"""
ftw_recode_list.py — the per-bill worklist for moving flatwork off the slabs.

`job_reality_audit.py` found FW (flatwork) cost sitting on jobs that are not
`-FTW` projects. This turns that into something a clerk can work: every
offending LINE, grouped by vendor, biggest vendor first, so one vendor gets
knocked out at a time instead of hopping between them.

THREE SHEETS, because they are three different actions:
  1. Recode by vendor  - the slab has a `-FTW` twin already, so the line just
                         moves job. Unambiguous; this is the work.
  2. Create -FTW first  - FW on an RP slab with NO twin. Somebody has to make
                         the job before anything can move.
  3. Wrong code         - FW on a CP or MFD job. Those divisions have no -FTW
                         twin at all, so this is a cost-code correction, not a
                         job move (the owner's rule, 2026-08-06).

EVERY SHEET LEADS WITH WHAT LANDED AFTER THE MEETING. The owner settled the
CP/MFD rule in a meeting on 2026-08-06, so a line coded that way SINCE is a
different problem from the historical backlog - the same mistake made after
everyone was told. Measured 2026-09-02: of 502,505 miscoded to CP/MFD, only
3,861 (4 lines) postdates the meeting and September is clean, so the meeting
worked and the rest is a cleanup of history. Keeping the two apart is the
point - burying four live lines inside 188 historical ones hides the only
part that is still happening.

Plain formatting on purpose (repo rule 5): this is a worklist, not a report.
Rows are grouped but NEVER hidden - the clerk can collapse a vendor with the
outline +/- if they want to.

Reads the cache `job_reality_audit.py` writes. No QBO pull of its own.

MERGING INTO THE CLERK'S WORKBOOK is the normal path (`--into`). The owner
already has one coding-errors workbook the clerk works in and does not want the
same information scattered over several files, so this writes ONE sheet into
that workbook, in that workbook's own column layout, and leaves every other
sheet untouched. A re-run replaces its own sheet rather than adding a second
copy. The file is backed up first and the corruption gate runs before it is put
back.

USAGE
  python3 one-offs/ftw_recode_list.py --into "/path/Code Code Errors 8.31.26.xlsx"
  python3 one-offs/ftw_recode_list.py --out ~/Downloads/"FTW Recodes.xlsx"
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared.xlsx_verify import assert_clean                    # noqa: E402

CACHE = Path(os.environ.get(
    "ACB_AUDIT_CACHE",
    Path(os.environ.get("TMPDIR", "/tmp")) / "job_reality_audit.json"))
FW_RE = re.compile(r"^FW\d+$", re.IGNORECASE)
MONEY = '#,##0.00;[Red](#,##0.00)'
THIN = Side(style="thin", color="000000")


def qbo_url(tx_type: str, txn_id: str, realm: str) -> str:
    """Deep link straight to the bill, so the clerk clicks instead of searching."""
    if not txn_id or not realm:
        return ""
    page = "bill" if (tx_type or "").lower() == "bill" else "expense"
    return (f"https://qbo.intuit.com/app/login?pagereq="
            f"{quote(f'{page}?txnId={txn_id}')}&deeplinkcompanyid={realm}")


def classify(costs: list) -> tuple:
    """Split the FW lines into the three actions."""
    jobs_with_cost = {c["proj"].upper() for c in costs}
    move, need_job, wrong_code = [], [], []
    for c in costs:
        leaf = (c.get("code") or "").split(":")[-1].strip()
        if not FW_RE.match(leaf):
            continue
        p = c["proj"].upper()
        if p.endswith("-FTW"):
            continue
        if p.startswith("RP"):
            twin = f"{p}-FTW"
            (move if twin in jobs_with_cost else need_job).append(dict(c, target=twin))
        else:
            wrong_code.append(dict(c, target=""))
    return move, need_job, wrong_code


def _t(ws, r, c, v, *, bold=False, fmt=None, align=None, wrap=False, size=11):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(size=size, bold=bold)
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    return cell


def sheet_by_vendor(wb, title: str, rows: list, realm: str, blurb: str,
                    show_target: bool, meeting: str = "") -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    total = sum(r["amt"] for r in rows)
    bills = len({(r.get("txn"), r.get("doc")) for r in rows})
    after = [r for r in rows if meeting and (r.get("date") or "") >= meeting]
    before = [r for r in rows if r not in after] if after else rows

    _t(ws, 1, 1, title.upper(), bold=True, size=14)
    _t(ws, 2, 1, blurb, size=10)
    _t(ws, 3, 1, f"{bills} bills · {len(rows)} lines · {total:,.2f} total · "
                 f"biggest vendor first", size=10)
    heads = ["Done", "Date", "Bill #", "Move FROM"] + \
            (["Move TO"] if show_target else []) + \
            ["Code", "Amount", "Description", "Open in QBO"]
    r = 5
    for i, h in enumerate(heads):
        c = _t(ws, r, 2 + i, h, bold=True,
               align="right" if h == "Amount" else "left")
        c.border = Border(bottom=THIN)
    r += 1
    for block, banner in ((after, f"CODED THIS WAY AFTER THE {_us(meeting)} MEETING "
                                  f"- {len(after)} line(s), {sum(x['amt'] for x in after):,.2f} "
                                  f"- raise these, they are still happening"),
                          (before, f"BEFORE THE MEETING - historical cleanup, "
                                   f"{len(before)} line(s), "
                                   f"{sum(x['amt'] for x in before):,.2f}")):
        if not block:
            continue
        if after:                      # only banner the split when there IS one
            _t(ws, r, 2, banner, bold=True)
            for i in range(len(heads)):
                ws.cell(row=r, column=2 + i).border = Border(top=THIN, bottom=THIN)
            r += 1
        r = _vendor_block(ws, r, block, heads, realm, show_target)
    _t(ws, r, 2, "TOTAL", bold=True)
    _t(ws, r, 2 + heads.index("Amount"), total, bold=True, fmt=MONEY, align="right")
    for i in range(len(heads)):
        ws.cell(row=r, column=2 + i).border = Border(top=THIN)

    ws.column_dimensions["A"].width = 3
    widths = [7, 11, 14, 13] + ([13] if show_target else []) + [8, 14, 52, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = f"A{6}"


def _us(iso: str) -> str:
    """2026-08-06 -> 08/06 - the owner never reads a date year-first."""
    p = (iso or "").split("-")
    return f"{p[1]}/{p[2]}" if len(p) == 3 else iso


def _vendor_block(ws, r: int, rows: list, heads: list, realm: str,
                  show_target: bool) -> int:
    by_vendor = defaultdict(list)
    for x in rows:
        by_vendor[x.get("vendor") or "(no vendor)"].append(x)
    order = sorted(by_vendor, key=lambda v: -sum(x["amt"] for x in by_vendor[v]))
    for vendor in order:
        lines = sorted(by_vendor[vendor], key=lambda x: (x.get("date") or ""))
        sub = sum(x["amt"] for x in lines)
        _t(ws, r, 2, f"{vendor}   ({len(lines)} line(s))", bold=True)
        _t(ws, r, 2 + heads.index("Amount"), sub, bold=True, fmt=MONEY, align="right")
        for i in range(len(heads)):
            ws.cell(row=r, column=2 + i).border = Border(top=THIN)
        r += 1
        for x in lines:
            vals = [None, x.get("date") or "", str(x.get("doc") or ""),
                    x["proj"]] + \
                   ([x.get("target") or ""] if show_target else []) + \
                   [(x.get("code") or "").split(":")[-1], x["amt"],
                    (x.get("desc") or x.get("memo") or "")[:70], None]
            for i, v in enumerate(vals):
                cell = _t(ws, r, 2 + i, v,
                          fmt=MONEY if heads[i] == "Amount" else None,
                          align="right" if heads[i] == "Amount" else None)
                if heads[i] == "Done":
                    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            link = qbo_url(x.get("type", ""), x.get("txn", ""), realm)
            if link:
                lc = ws.cell(row=r, column=2 + len(heads) - 1, value="open bill")
                lc.hyperlink = link
                lc.font = Font(size=11, color="0563C1", underline="single")
            # grouped, NOT hidden (repo rule 5: no hidden rows) - the clerk can
            # collapse a vendor with the outline +/- if they want to
            ws.row_dimensions[r].outline_level = 1
            r += 1
    return r


# The clerk's workbook speaks this column language - match it exactly rather
# than importing a second convention into the same file.
CLERK_COLS = ["Issue", "Vendor", "Bill #", "Date", "Project", "Class", "Cost",
              "Amount", "Line memo", "Why flagged", "Move to",
              "After 08/06 mtg", "Open in QBO"]
DIVISION = {"MFD": "Multi Family", "CP": "Commercial", "RP": "Residential"}


def _division(proj: str) -> str:
    """From the project #, never the QBO Class field (CLAUDE.md: do not trust
    Class for division)."""
    for pre, name in DIVISION.items():
        if proj.upper().startswith(pre):
            return name
    return ""


def merge_sheet(book: Path, move: list, need_job: list, wrong: list,
                realm: str, meeting: str, sheet_name: str) -> dict:
    """Write ONE sheet into the clerk's existing workbook, in their layout.

    Every other sheet is left exactly as it is - this only adds or replaces its
    own. Their workbook carries an Excel Table on the first sheet, and rule 5b
    is about a table ref left pointing past its data; nothing here touches that
    sheet's rows, so the ref stays valid, and `assert_clean` proves it before
    the file goes back."""
    lock = book.with_name("~$" + book.name)
    if lock.exists():
        raise SystemExit(f"✗  {book.name} is OPEN in Excel - close it and re-run "
                         f"(refusing to overwrite the clerk's work)")
    wb = load_workbook(str(book))
    before = {ws.title: ws.max_row for ws in wb.worksheets}
    if sheet_name in wb.sheetnames:          # re-run replaces, never duplicates
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    tagged = []
    for rows, issue, target in ((move, "FW - move to -FTW", True),
                                (need_job, "FW - create -FTW first", False),
                                (wrong, "FW - wrong code (CP/MFD)", False)):
        for x in rows:
            tagged.append(dict(x, _issue=issue, _target=x.get("target", "") if target else ""))
    # biggest vendor first, that vendor's rows together, and anything coded
    # since the meeting at the very top whatever vendor it is
    vend_total = defaultdict(float)
    for x in tagged:
        vend_total[x.get("vendor") or ""] += x["amt"]
    tagged.sort(key=lambda x: (
        0 if (x.get("date") or "") >= meeting else 1,
        -vend_total[x.get("vendor") or ""],
        x.get("vendor") or "",
        x.get("date") or ""))

    for i, h in enumerate(CLERK_COLS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(size=12, bold=True)
    for r, x in enumerate(tagged, start=2):
        after = (x.get("date") or "") >= meeting
        why = {"FW - move to -FTW":
               f"FW code on slab {x['proj']} - flatwork belongs on {x.get('_target')}",
               "FW - create -FTW first":
               f"FW code on slab {x['proj']} - no -FTW job exists yet, create it first",
               "FW - wrong code (CP/MFD)":
               f"FW code on a {_division(x['proj'])} job - flatwork belongs to RP"}[x["_issue"]]
        if after:
            why += f" - CODED AFTER THE {_us(meeting)} MEETING"
        # a REAL date, not the ISO text - the clerk's other sheets carry real
        # dates, and text sorts lexically and cannot be date-filtered
        try:
            when = dt.date.fromisoformat((x.get("date") or "")[:10])
        except ValueError:
            when = x.get("date") or ""
        vals = [x["_issue"], x.get("vendor") or "", str(x.get("doc") or ""),
                when, x["proj"], _division(x["proj"]),
                (x.get("code") or "").split(":")[-1], x["amt"],
                (x.get("desc") or x.get("memo") or "")[:80], why,
                x.get("_target") or "", "YES" if after else "", None]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(size=12, bold=bool(after and i == 12))
            if CLERK_COLS[i - 1] == "Amount":
                c.number_format = MONEY
            elif CLERK_COLS[i - 1] == "Date" and isinstance(v, dt.date):
                c.number_format = "mm/dd/yyyy"      # never year-first (owner)
        link = qbo_url(x.get("type", ""), x.get("txn", ""), realm)
        if link:
            lc = ws.cell(row=r, column=len(CLERK_COLS), value="open bill")
            lc.hyperlink = link
            lc.font = Font(size=12, color="0563C1", underline="single")
    ws.auto_filter.ref = (f"A1:{get_column_letter(len(CLERK_COLS))}"
                          f"{max(len(tagged) + 1, 2)}")
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJKLM",
                      (24, 30, 15, 12, 12, 14, 8, 13, 42, 62, 14, 15, 12)):
        ws.column_dimensions[col].width = w

    tmp = book.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    check = load_workbook(str(tmp))
    after_rows = {t: check[t].max_row for t in before}
    if after_rows != before:
        tmp.unlink()
        raise SystemExit(f"✗  refusing to write: an existing sheet changed size "
                         f"{before} -> {after_rows}")
    tmp.replace(book)
    return {"rows": len(tagged), "sheets_kept": len(before)}


def main() -> int:
    ap = argparse.ArgumentParser(description="Per-bill FTW recode worklist")
    ap.add_argument("--out", default=str(Path.home() / "Downloads" / "FTW Recodes.xlsx"))
    ap.add_argument("--into", default="",
                    help="merge ONE sheet into this existing workbook instead "
                         "of writing a separate file")
    ap.add_argument("--sheet", default="FW Misplaced",
                    help="sheet name used by --into (replaced on a re-run)")
    ap.add_argument("--cache", default=str(CACHE))
    ap.add_argument("--meeting", default="2026-08-06",
                    help="the date the rule was settled; anything coded this "
                         "way SINCE is listed first (default 2026-08-06)")
    a = ap.parse_args()
    cache = Path(a.cache).expanduser()
    if not cache.exists():
        print(f"✗  no audit cache at {cache} — run one-offs/job_reality_audit.py first")
        return 1
    data = json.loads(cache.read_text())
    realm = data.get("realm", "")
    move, need_job, wrong_code = classify(data["costs"])

    if a.into:
        book = Path(a.into).expanduser()
        if not book.exists():
            print(f"✗  no workbook at {book}")
            return 1
        bak = (Path(os.environ.get("TMPDIR", "/tmp")) /
               f"{book.stem}.backup-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
        bak.write_bytes(book.read_bytes())
        info = merge_sheet(book, move, need_job, wrong_code, realm,
                           a.meeting, a.sheet)
        after = sum(1 for r in move + need_job + wrong_code
                    if (r.get("date") or "") >= a.meeting)
        print(f"  sheet '{a.sheet}' written into {book.name}")
        print(f"    {info['rows']} rows · {after} coded after {_us(a.meeting)} "
              f"· {info['sheets_kept']} existing sheet(s) untouched")
        print(f"    backup: {bak}")
        return 0

    wb = Workbook()
    wb.remove(wb.active)
    sheet_by_vendor(
        wb, "Recode by vendor", move, realm, show_target=True, meeting=a.meeting,
        blurb="Move each line from the slab to its -FTW twin. The twin already "
              "exists, so nothing needs creating first.")
    if need_job:
        sheet_by_vendor(
            wb, "Create -FTW first", need_job, realm, show_target=False,
            meeting=a.meeting,
            blurb="FW cost on an RP slab with NO -FTW twin. The flatwork job has "
                  "to be created before these can move.")
    if wrong_code:
        sheet_by_vendor(
            wb, "Wrong code", wrong_code, realm, show_target=False,
            meeting=a.meeting,
            blurb="FW cost on a CP or MFD job. Those divisions have no -FTW "
                  "twin, so this is a cost-code correction, not a job move.")
    out = Path(a.out).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    tmp.replace(out)
    for label, rows in (("recode", move), ("create -FTW first", need_job),
                        ("wrong code", wrong_code)):
        print(f"  {label:20} {len(rows):5} lines  {sum(r['amt'] for r in rows):12,.2f}")
    print(f"\n  → {out}   (generated {dt.datetime.now():%m/%d/%Y %I:%M %p})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
