#!/usr/bin/env python3
"""
jobtread_close_jobs.py — audit-gated JobTread job closer (the user 2026-07-27).

WHY
  jobtread_bloat_report.py finds open JobTread jobs that are finished in real
  life (paid & idle, or stale shells). This closes them — but ONLY the ones the
  user approves, and it NEVER deletes anything.

HOW A JOB IS CLOSED (verified against the API 2026-07-27)
  A job's `status` is NOT directly settable. Closing = setting **`closedOn`** to
  a date; JobTread then reports status "closed". Clearing `closedOn` (null)
  restores the previous status. Proven round-trip on the CP000 placeholder:
      approved/None → (closedOn=2026-07-27) → closed → (closedOn=null) → approved/None
  So every close here is REVERSIBLE with --reopen. Nothing is ever deleted.

THE GATE (same pattern as qbo_recode_review.py)
  1. --export  → reads the bloat workbook, writes an APPROVE workbook with a
                 CLOSE? (Y/N) column pre-filled with the recommendation.
  2. the user edits it — Y = close, anything else = leave open.
  3. --apply   → DRY RUN: prints exactly what would change.
  4. --apply --commit → performs the closes. Every change logged with before/after.
  5. --reopen [--commit] → undoes rows marked Y (clears closedOn).

MFD is excluded by default (house rule: MFD closures are handled by hand).

Usage
  python3 jobtread_close_jobs.py --export
  python3 jobtread_close_jobs.py --apply
  python3 jobtread_close_jobs.py --apply --commit
  python3 jobtread_close_jobs.py --reopen --commit
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.datavalidation import DataValidation

from shared import qbo_vault

ORG_ID = os.getenv("JT_ORG_ID", "22PFAfqHLF3a")
API_URL = "https://api.jobtread.com/pave"
BLOAT_XLSX = Path(os.getenv(
    "JT_BLOAT_XLSX",
    str(Path.home() / "Downloads" / "JobTread Bloat - Close Candidates.xlsx")))
APPROVE_XLSX = Path(os.getenv(
    "JT_CLOSE_XLSX",
    str(Path.home() / "Downloads" / "JobTread Close - APPROVE.xlsx")))
LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient"

HDR = ["CLOSE? (Y/N)", "JOB #", "DIV", "VERDICT", "JT NAME", "QBO BILLED $",
       "AR BALANCE $", "LAST INVOICE", "DAYS IDLE", "CREATED"]


def pave(key: str, query: dict) -> dict:
    body = json.dumps({"query": {"$": {"grantKey": key}, **query}}).encode()
    req = urllib.request.Request(API_URL, data=body,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def job_by_number(key: str, number: str):
    """number → {id, number, status, closedOn} or None."""
    r = pave(key, {"organization": {"$": {"id": ORG_ID}, "jobs": {
        "$": {"size": 1, "where": {"and": [["number", "=", number]]}},
        "nodes": {"id": {}, "number": {}, "status": {}, "closedOn": {}}}}})
    nodes = r["organization"]["jobs"]["nodes"]
    return nodes[0] if nodes else None


def set_closed(key: str, job_id: str, when):
    """when = 'YYYY-MM-DD' to close, or None to reopen."""
    return pave(key, {"updateJob": {"$": {"id": job_id, "closedOn": when}}})


# ── export the approval workbook ────────────────────────────────────
BOLD = Font(bold=True)
UNDER = Border(bottom=Side(style="thin", color="000000"))
CUR = '"$"#,##0.00'


def do_export(include_mfd: bool) -> int:
    if not BLOAT_XLSX.exists():
        print(f"  ✗ {BLOAT_XLSX.name} not found — run jobtread_bloat_report.py first")
        return 1
    wb = load_workbook(BLOAT_XLSX)
    if "Close Candidates" not in wb.sheetnames:
        print("  ✗ 'Close Candidates' tab missing")
        return 1
    ws = wb["Close Candidates"]
    cols = {str(c.value).strip(): i for i, c in enumerate(ws[2], 1) if c.value}
    need = ["JOB #", "DIV", "VERDICT", "JT NAME", "QBO BILLED $",
            "AR BALANCE $", "LAST INVOICE", "DAYS IDLE", "CREATED"]
    missing = [n for n in need if n not in cols]
    if missing:
        print(f"  ✗ bloat workbook missing columns: {missing}")
        return 1

    rows, skipped_mfd = [], 0
    for r in range(3, ws.max_row + 1):
        get = lambda n: ws.cell(r, cols[n]).value
        job = get("JOB #")
        if not job:
            continue
        div = get("DIV") or ""
        if div == "Multi Family" and not include_mfd:
            skipped_mfd += 1
            continue
        verdict = get("VERDICT") or ""
        # pre-fill: confident bucket = Y, softer bucket = blank for review
        pre = "Y" if str(verdict).startswith("CLOSE —") else ""
        rows.append([pre, job, div, verdict, get("JT NAME"), get("QBO BILLED $"),
                     get("AR BALANCE $"), get("LAST INVOICE"), get("DAYS IDLE"),
                     get("CREATED")])
    wb.close()

    lock = APPROVE_XLSX.with_name("~$" + APPROVE_XLSX.name)
    if lock.exists():
        raise SystemExit(f"{APPROVE_XLSX.name} is open in Excel — close it first")
    out = Workbook()
    o = out.active
    o.title = "APPROVE"
    o["A1"] = (f"JOBTREAD CLOSE — APPROVAL LIST ({len(rows)} candidates). "
               "Put Y in column A to CLOSE that job; leave blank to keep it open. "
               "Pre-filled Y = paid & idle (high confidence); blank = no QBO "
               "activity (review). Closing is reversible (--reopen); nothing is deleted.")
    o["A1"].font = BOLD
    o.append(HDR)
    for c in range(1, len(HDR) + 1):
        o.cell(2, c).font = BOLD
        o.cell(2, c).border = UNDER
    for row in rows:
        o.append(row)
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    o.add_data_validation(dv)
    dv.add(f"A3:A{max(3, o.max_row)}")
    for c in (6, 7):
        for r in range(3, o.max_row + 1):
            o.cell(r, c).number_format = CUR
    o.auto_filter.ref = f"A2:J{max(2, o.max_row)}"
    o.freeze_panes = "A3"
    for col, w in zip("ABCDEFGHIJ", (14, 12, 14, 24, 30, 15, 15, 13, 10, 12)):
        o.column_dimensions[col].width = w
    out.save(APPROVE_XLSX)
    pre_y = sum(1 for r in rows if r[0] == "Y")
    print(f"  ✓ Approval list → {APPROVE_XLSX}")
    print(f"    {len(rows)} candidates · pre-filled Y: {pre_y} · blank (review): "
          f"{len(rows) - pre_y}" + (f" · MFD skipped: {skipped_mfd}" if skipped_mfd else ""))
    print("    Edit column A, then run:  --apply   (dry run)")
    return 0


def read_approved():
    if not APPROVE_XLSX.exists():
        raise SystemExit(f"{APPROVE_XLSX.name} not found — run --export first")
    wb = load_workbook(APPROVE_XLSX, data_only=True)
    ws = wb["APPROVE"]
    out = []
    for r in range(3, ws.max_row + 1):
        mark = str(ws.cell(r, 1).value or "").strip().upper()
        job = ws.cell(r, 2).value
        if job and mark == "Y":
            out.append({"job": str(job).strip(),
                        "verdict": ws.cell(r, 4).value or ""})
    wb.close()
    return out


def do_apply(commit: bool, reopen: bool, closed_date: str) -> int:
    approved = read_approved()
    if not approved:
        print("  nothing marked Y in the approval list — nothing to do")
        return 0
    action = "REOPEN" if reopen else "CLOSE"
    print(f"\n  {action} — {len(approved)} job(s) marked Y "
          f"{'(COMMIT)' if commit else '(DRY RUN — no changes)'}")
    print("  " + "─" * 60)
    key = qbo_vault.get("JT_GRANT_KEY")
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = LOG_DIR / f"jobtread_close_{stamp}.jsonl"
    done = skipped = failed = 0
    with open(log_path, "a") as log:
        for i, row in enumerate(approved, 1):
            num = row["job"]
            try:
                j = job_by_number(key, num)
            except urllib.error.HTTPError as e:
                print(f"   {num:10} ✗ lookup failed ({e.code})")
                failed += 1
                continue
            if not j:
                print(f"   {num:10} ✗ not found in JobTread")
                failed += 1
                continue
            is_closed = bool(j.get("closedOn")) or j.get("status") == "closed"
            if not reopen and is_closed:
                print(f"   {num:10} · already closed ({j.get('closedOn')}) — skip")
                skipped += 1
                continue
            if reopen and not is_closed:
                print(f"   {num:10} · already open — skip")
                skipped += 1
                continue
            target = None if reopen else closed_date
            if not commit:
                print(f"   {num:10} would {action.lower()}: status={j['status']} "
                      f"closedOn={j.get('closedOn')} → closedOn={target}")
                done += 1
                continue
            try:
                set_closed(key, j["id"], target)
                after = job_by_number(key, num)
                print(f"   {num:10} ✓ {j['status']}/{j.get('closedOn')} → "
                      f"{after['status']}/{after.get('closedOn')}")
                log.write(json.dumps({"ts": dt.datetime.now().isoformat(),
                                      "job": num, "id": j["id"], "action": action,
                                      "before": {"status": j["status"],
                                                 "closedOn": j.get("closedOn")},
                                      "after": {"status": after["status"],
                                                "closedOn": after.get("closedOn")}}) + "\n")
                done += 1
            except urllib.error.HTTPError as e:
                print(f"   {num:10} ✗ {action.lower()} failed ({e.code}) "
                      f"{e.read().decode()[:120]}")
                failed += 1
            if i % 25 == 0:
                print(f"    …{i}/{len(approved)}")
    print("  " + "─" * 60)
    verb = "would change" if not commit else "changed"
    print(f"  {verb}: {done} · skipped: {skipped} · failed: {failed}")
    if commit:
        print(f"  log → {log_path}")
        print("  undo any of these with:  --reopen --commit  "
              "(mark only those rows Y)")
    else:
        print("  DRY RUN — nothing was changed. Re-run with --commit to apply.")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--export", action="store_true",
                      help="build the approval workbook from the bloat report")
    mode.add_argument("--apply", action="store_true",
                      help="close jobs marked Y (dry run unless --commit)")
    mode.add_argument("--reopen", action="store_true",
                      help="reopen jobs marked Y (clears closedOn)")
    ap.add_argument("--commit", action="store_true", help="perform the writes")
    ap.add_argument("--closed-date", default=dt.date.today().isoformat(),
                    help="date to stamp as closedOn (default: today)")
    ap.add_argument("--include-mfd", action="store_true",
                    help="include Multi Family (excluded by default)")
    args = ap.parse_args()

    print("\n  JOBTREAD CLOSER — status change only, never deletes")
    print("  " + "─" * 60)
    if args.export:
        return do_export(args.include_mfd)
    return do_apply(args.commit, args.reopen, args.closed_date)


if __name__ == "__main__":
    sys.exit(main())
