#!/usr/bin/env python3
"""
job_vendor_report.py — every job cost in a date window, by vendor, complete.

Built to answer a PM's "Transaction List by Vendor" with the same shape but
nothing missing (the user 2026-08-27). On MFD325's July draw the PM's version
ran 05/30-07/01 instead of 06/02-07/01 AND dropped three subcontractors worth
~218k; this rebuilds the same window from QBO and, given the PM's file, shows
vendor-by-vendor what each side has.

LINE AMOUNTS ONLY, never a bill's TotalAmt — a sub bill spans jobs and banking
the whole document hands this job another job's money (see
one-offs/pnl_line_level_audit.py).

USAGE
  python3 one-offs/job_vendor_report.py --project MFD325 \
      --from 2026-05-30 --to 2026-07-01 --compare "<the PM's file.xlsx>"
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import qbo_api
from shared.job_lines import JobMatcher, discover_job_classes
from shared.xlsx_verify import assert_clean

SZ, SZ_S, SZ_T = 14, 13, 20
NAVY, INK, GREY = "1F3A5F", "1F2937", "6B7280"
GREEN, RED, LINK = "1E6B3A", "B00020", "0563C1"
MONEY = '#,##0.00;[Red](#,##0.00)'
F_HDR = PatternFill("solid", fgColor=NAVY)
F_BAND = PatternFill("solid", fgColor="F4F6F9")
RULE = Side(style="thin", color=NAVY)
HAIR = Side(style="thin", color="E3E8EF")


def _c(ws, r, col, v, *, size=SZ_S, bold=False, color=INK, fmt=None,
       fill=None, align=None, indent=0):
    cell = ws.cell(row=r, column=col, value=v)
    cell.font = Font(size=size, bold=bold, color=color)
    if fmt:
        cell.number_format = fmt
    if fill is not None:
        cell.fill = fill
    if align or indent:
        cell.alignment = Alignment(horizontal=align, indent=indent, vertical="center")
    return cell


def pull(project: str, d0: str, d1: str, aliases: List[str]) -> List[dict]:
    access, cid = qbo_api.load_credentials()
    pmap = qbo_api.build_project_customer_map(access, cid)
    info = pmap.get(project)
    if not info:
        print(f"✗  {project} not found in QBO")
        raise SystemExit(1)
    classes = (qbo_api.query_all(access, cid, "Class")
               + qbo_api.query_all(access, cid, "Class", "Active = false"))
    cls = discover_job_classes(classes, project)
    m = JobMatcher(info["id"], project, aliases, legacy=bool(aliases or cls),
                   class_ids=list(cls.keys()), text_rules=bool(aliases))
    where = f"TxnDate >= '{d0}' AND TxnDate <= '{d1}'"
    rows: List[dict] = []
    for ent, vf in (("Bill", "VendorRef"), ("Purchase", "EntityRef")):
        for t in qbo_api.query_all(access, cid, ent, where):
            vendor = ((t.get(vf) or {}).get("name") or "(no vendor)")
            note = t.get("PrivateNote") or ""
            bal = float(t.get("Balance", 0) or 0) if ent == "Bill" else 0.0
            for ln in t.get("Line") or []:
                det = (ln.get("AccountBasedExpenseLineDetail")
                       or ln.get("ItemBasedExpenseLineDetail"))
                if not det:
                    continue
                rule = m.rule(det, ln, t)
                if not rule:
                    continue
                rows.append({
                    "vendor": vendor, "date": t.get("TxnDate") or "",
                    "num": str(t.get("DocNumber") or ""),
                    "memo": (ln.get("Description") or note or "")[:90],
                    "amt": float(ln.get("Amount", 0) or 0),
                    "paid": ("Paid" if bal <= 0.005 else "Unpaid") if ent == "Bill" else "Paid",
                    "how": rule,
                    "acct": ((det.get("AccountRef") or {}).get("name")
                             or (det.get("ItemRef") or {}).get("name") or "")})
    return rows


def read_pm(path: Path) -> List[dict]:
    ws = load_workbook(str(path), data_only=True).worksheets[0]
    out, vend = [], None
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip()
        d, num, memo = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        amt = ws.cell(r, 5).value
        if a and not a.lower().startswith("total"):
            vend = a
            continue
        if a.lower().startswith("total"):
            continue
        if isinstance(amt, (int, float)) and d:
            out.append({"vendor": vend, "date": str(d)[:10],
                        "num": str(num or ""), "memo": str(memo or ""),
                        "amt": float(amt)})
    return out


def _norm(s: str) -> str:
    return "".join(ch for ch in (s or "").upper() if ch.isalnum())[:14]


def _doc(s: str) -> str:
    return "".join(ch for ch in (s or "").upper() if ch.isalnum())


def _match_pm(rows: List[dict], pm: List[dict]):
    """Tag each of OUR lines with whether the PM's report caught the BILL, and
    return the bills only he has.

    Matched per DOCUMENT, not per line. A bill that is one row on his report
    is often several lines on ours (a sub bills form/rebar/pour separately),
    so amount-for-amount matching wrongly called whole vendors "partly
    missing" when their totals agreed to the cent. A bill is a bill on both
    sides; the document number is the honest key."""
    theirs: Dict[tuple, float] = defaultdict(float)
    their_rows: Dict[tuple, list] = defaultdict(list)
    for x in pm:
        k = (_norm(x["vendor"]), _doc(x["num"]))
        theirs[k] += x["amt"]
        their_rows[k].append(x)
    ours: Dict[tuple, float] = defaultdict(float)
    for r in rows:
        ours[(_norm(r["vendor"]), _doc(r["num"]))] += r["amt"]

    caught = 0
    for r in rows:
        k = (_norm(r["vendor"]), _doc(r["num"]))
        if k in theirs:
            r["pm"] = "Yes"
            caught += 1
        else:
            r["pm"] = "No"
    extra = [x for k, lst in their_rows.items() if k not in ours for x in lst]
    return caught, extra


def build(project: str, d0: str, d1: str, rows: List[dict],
          pm: List[dict], out: Path, income: float = 0.0) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Report"
    ws.sheet_view.showGridLines = False
    total = sum(r["amt"] for r in rows)

    caught, extra = _match_pm(rows, pm) if pm else (0, [])
    missed = [x for x in rows if x.get("pm") == "No"]

    _c(ws, 1, 1, f"{project} — JOB COST BY VENDOR", size=SZ_T, bold=True, color=NAVY)
    _c(ws, 2, 1, f"{dt.date.fromisoformat(d0):%m/%d/%Y} – {dt.date.fromisoformat(d1):%m/%d/%Y}"
                 f"   ·   every bill and purchase coded to this job   ·   line amounts, "
                 f"never bill totals", size=SZ_S, color=GREY)
    for c in range(1, 8):
        ws.cell(row=2, column=c).border = Border(bottom=HAIR)

    # NET FIRST (the user 2026-08-27) — the number the draw turns on.
    net = income - total
    tiles = [("BILLED THIS DRAW", income, MONEY, NAVY),
             ("JOB COST", total, MONEY, NAVY),
             ("NET", net, MONEY, GREEN if net >= 0 else RED)]
    if pm:
        tiles.append((f"MISSED BY THE PM REPORT",
                      sum(x["amt"] for x in missed), MONEY, RED))
    for i, (lbl, val, fmt, col) in enumerate(tiles):
        c0 = 1 + i * 2
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 1)
        ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 1)
        _c(ws, 4, c0, lbl, size=SZ_S - 1, bold=True, color=GREY)
        _c(ws, 5, c0, val, size=SZ + 8, bold=True, color=col, fmt=fmt)
        for cc in (c0, c0 + 1):
            ws.cell(row=4, column=cc).border = Border(bottom=RULE)
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 40

    r = 7
    for c, h in ((1, "Vendor / Doc #"), (2, "Date"), (3, "Memo"),
                 (5, "Amount"), (6, "Paid?"), (7, "On PM report?")):
        _c(ws, r, c, h, size=SZ_S - 1, bold=True, color="FFFFFF", fill=F_HDR,
           align="right" if c == 5 else ("center" if c in (6, 7) else "left"),
           indent=1 if c == 1 else 0)
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = F_HDR
    ws.row_dimensions[r].height = 22
    r += 1

    by = defaultdict(list)
    for x in rows:
        by[x["vendor"]].append(x)
    for vendor in sorted(by, key=lambda v: -sum(x["amt"] for x in by[v])):
        lines = sorted(by[vendor], key=lambda x: x["date"])
        vmiss = sum(x["amt"] for x in lines if x.get("pm") == "No")
        _c(ws, r, 1, vendor, size=SZ, bold=True, color=INK, fill=F_BAND)
        _c(ws, r, 5, sum(x["amt"] for x in lines), size=SZ, bold=True,
           color=INK, fill=F_BAND, fmt=MONEY, align="right")
        if pm and vmiss:
            _c(ws, r, 7, "NOT ON PM REPORT" if vmiss == sum(x["amt"] for x in lines)
               else "partly missing", size=SZ_S - 1, bold=True, color=RED,
               fill=F_BAND, align="center")
        for c in (2, 3, 4, 6):
            ws.cell(row=r, column=c).fill = F_BAND
        if pm and not vmiss:
            ws.cell(row=r, column=7).fill = F_BAND
        ws.row_dimensions[r].height = 21
        r += 1
        for x in lines:
            _c(ws, r, 1, x["num"], indent=2, align="left")
            d = _c(ws, r, 2, dt.date.fromisoformat(x["date"]) if x["date"] else None)
            d.number_format = "mm/dd/yyyy"
            _c(ws, r, 3, x["memo"])
            _c(ws, r, 5, x["amt"], fmt=MONEY, align="right")
            _c(ws, r, 6, x["paid"], align="center",
               color=GREY if x["paid"] == "Paid" else RED)
            if pm:
                _c(ws, r, 7, x.get("pm", ""), align="center", bold=x.get("pm") == "No",
                   color=GREY if x.get("pm") == "Yes" else RED)
            r += 1

    if extra:
        r += 1
        _c(ws, r, 1, "ON THE PM REPORT BUT NOT CODED TO THIS JOB IN QBO",
           size=SZ, bold=True, color=NAVY)
        _c(ws, r, 5, sum(x["amt"] for x in extra), size=SZ, bold=True,
           color=NAVY, fmt=MONEY, align="right")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = Border(bottom=RULE)
        r += 1
        for x in sorted(extra, key=lambda y: -y["amt"]):
            _c(ws, r, 1, f"{x['vendor']}  {x['num']}".strip(), indent=1)
            _c(ws, r, 3, str(x["memo"])[:90])
            _c(ws, r, 5, x["amt"], fmt=MONEY, align="right")
            _c(ws, r, 7, "PM only", align="center", color=RED, bold=True)
            r += 1

    for col, w in zip("ABCDEFG", (34, 14, 58, 20, 20, 14, 20)):
        ws.column_dimensions[col].width = w

    if pm:
        cs = wb.create_sheet("vs PM Report")
        cs.sheet_view.showGridLines = False
        _c(cs, 1, 1, f"{project} — THIS REPORT vs THE PM'S", size=SZ_T, bold=True,
           color=NAVY)
        _c(cs, 2, 1, "Same window, vendor by vendor. A gap is money one side "
                     "has and the other does not.", size=SZ_S, color=GREY)
        ours, theirs, name = defaultdict(float), defaultdict(float), {}
        for x in rows:
            ours[_norm(x["vendor"])] += x["amt"]
            name[_norm(x["vendor"])] = x["vendor"]
        for x in pm:
            theirs[_norm(x["vendor"])] += x["amt"]
            name.setdefault(_norm(x["vendor"]), x["vendor"])
        rr = 4
        for c, h in ((1, "Vendor"), (2, "This report"), (3, "PM report"),
                     (4, "Difference"), (5, "What it is")):
            _c(cs, rr, c, h, size=SZ_S - 1, bold=True, color="FFFFFF", fill=F_HDR,
               align="right" if c in (2, 3, 4) else "left", indent=1 if c == 1 else 0)
        for c in range(1, 6):
            cs.cell(row=rr, column=c).fill = F_HDR
        rr += 1
        for k in sorted(set(ours) | set(theirs),
                        key=lambda k: -abs(ours.get(k, 0) - theirs.get(k, 0))):
            a, b = round(ours.get(k, 0.0), 2), round(theirs.get(k, 0.0), 2)
            note = ("" if abs(a - b) < 0.01 else
                    ("missing from the PM report" if b == 0 else
                     ("not coded to this job in QBO" if a == 0 else
                      "date range / partial")))
            _c(cs, rr, 1, name[k], bold=abs(a - b) > 0.01)
            _c(cs, rr, 2, a, fmt=MONEY, align="right")
            _c(cs, rr, 3, b, fmt=MONEY, align="right")
            _c(cs, rr, 4, a - b, fmt=MONEY, align="right", bold=True,
               color=INK if abs(a - b) < 0.01 else (GREEN if a > b else RED))
            _c(cs, rr, 5, note, color=GREY)
            if abs(a - b) < 0.01:
                for c in range(1, 6):
                    cs.cell(row=rr, column=c).fill = F_BAND
            rr += 1
        _c(cs, rr, 1, "TOTAL", size=SZ, bold=True, color=NAVY)
        for c, v in ((2, sum(ours.values())), (3, sum(theirs.values())),
                     (4, sum(ours.values()) - sum(theirs.values()))):
            _c(cs, rr, c, v, size=SZ, bold=True, fmt=MONEY, align="right", color=NAVY)
        for c in range(1, 6):
            cs.cell(row=rr, column=c).border = Border(top=RULE)
        for col, w in zip("ABCDE", (36, 20, 20, 20, 34)):
            cs.column_dimensions[col].width = w

    for sh in wb.worksheets:
        sh.page_setup.orientation = "landscape"
        sh.page_setup.fitToWidth = 1
        sh.page_setup.fitToHeight = 0
        sh.sheet_properties.pageSetUpPr.fitToPage = True
    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    assert_clean(tmp)
    tmp.replace(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="Complete job cost by vendor for a window")
    ap.add_argument("--project", required=True)
    ap.add_argument("--from", dest="d0", required=True)
    ap.add_argument("--to", dest="d1", required=True)
    ap.add_argument("--alias", action="append", default=[])
    ap.add_argument("--compare", default=None, help="the PM's xlsx, for a diff sheet")
    ap.add_argument("--income", type=float, default=0.0,
                    help="what the draw BILLED, so the sheet can lead with net. "
                         "Take it from the draw sheet's INCOME figure — it is the "
                         "invoice's own number, not something to infer from dates.")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()
    rows = pull(a.project, a.d0, a.d1, a.alias)
    pm = read_pm(Path(a.compare).expanduser()) if a.compare else []
    out = Path(a.out).expanduser() if a.out else (
        Path.home() / "Downloads" /
        f"{a.project} Vendor Report {a.d0} to {a.d1}.xlsx")
    build(a.project, a.d0, a.d1, rows, pm, out, income=a.income)
    print(f"  {len(rows)} line(s)   ${sum(r['amt'] for r in rows):,.2f}")
    if pm:
        print(f"  PM report: {len(pm)} line(s)   ${sum(r['amt'] for r in pm):,.2f}"
              f"   difference ${sum(r['amt'] for r in rows) - sum(r['amt'] for r in pm):,.2f}")
    print(f"  → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
