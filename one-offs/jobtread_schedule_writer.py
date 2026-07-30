#!/usr/bin/env python3
"""
jobtread_schedule_writer.py — push the daily Excel schedule into JobTread as
dated tasks (the user 2026-07-29). AUDIT-GATED: dry run by default.

WHY
  The schedule already says what stage each job is in and when. JobTread models
  exactly that as a TASK (name + startDate + endDate + taskType on a job), and
  the estimator has already built the stage vocabulary as task templates
  ('Residential Slab MASTER', '… Tella Firma MASTER', '… Piers & Slab MASTER',
  'Residential Flatwork'). So the schedule can drive JobTread instead of
  someone retyping it job by job.

WHAT IT WRITES
  For every scheduled line: one task named after the mapped stage, dated from
  the schedule (column G; if blank, the schedule file's own date), with the
  matching taskType. UPSERT by (job, task name):
      • task with that name already on the job → update its dates ONLY if they changed
      • no such task                           → create it
  It NEVER deletes a task and never touches a task it did not map (hand-added
  JobTread tasks are safe).

WHAT IT DOES NOT DO
  Proposals. `createDocument` is rejected by the API ("A job location name or
  address is required") even on jobs that have a full location — verified on
  CP000 and RP7552 — so pricing still goes in by hand. Tasks and job-close are
  the write paths that are proven to work.

Usage
  python3 jobtread_schedule_writer.py                     # DRY RUN (default)
  python3 jobtread_schedule_writer.py --job RP7552        # one job
  python3 jobtread_schedule_writer.py --commit            # write to JobTread
  python3 jobtread_schedule_writer.py --schedule <f.xlsx>
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))

from openpyxl import load_workbook

from shared import qbo_vault
from shared import schedule as SCHED
import rp_wip_reader as RP
import rp_schedule_wip_preview as P

ORG_ID = os.getenv("JT_ORG_ID", "22PFAfqHLF3a")
API_URL = "https://api.jobtread.com/pave"
LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient"
_JOB_RE = re.compile(r"^(RP|CP)\d{3,4}$", re.IGNORECASE)

# Schedule stage → the estimator's own JobTread task name + task type.
# Matched on the DESCRIPTION first (it is more specific than the band), then
# the section band. Anything unmapped keeps the schedule's own wording so
# nothing is silently dropped.
DESC_MAP = [
    (r"WRECK",                    "Wreck/Clean",    "Finish"),
    (r"TENSION|STRESS",           "Tension Cables", "Pour"),
    (r"\bPOUR\b",                 "Pour",           "Pour"),
    (r"DRILL|\bPIER",             "Drill Piers",    "Piers"),
    (r"TRENCH",                   "Trench",         "Site Prep"),
    (r"GRADE|BACKOUT",            "Grade/Backout",  "Slab Prep"),
    (r"INSPECT",                  "Inspection",     "Slab Prep"),
    (r"SET UP|FORM SET|SET FORM", "Set Forms",      "Site Prep"),
]
SECTION_MAP = {
    "WRECK":             ("Wreck/Clean",   "Finish"),
    "TRENCH":            ("Trench",        "Site Prep"),
    "PIERS":             ("Drill Piers",   "Piers"),
    "GRADE AND BACKOUT": ("Grade/Backout", "Slab Prep"),
    "FORM SET":          ("Set Forms",     "Site Prep"),
    "FLATWORK":          ("Set Forms",     "Flatwork"),
    "TRACTOR":           ("Grade/Backout", "Site Prep"),
    "CONCRETE CUTTING":  ("Concrete Cutting", "Site Prep"),
}


def pave(key, query):
    body = json.dumps({"query": {"$": {"grantKey": key}, **query}}).encode()
    req = urllib.request.Request(API_URL, data=body,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def stage_for(section, desc):
    d = (desc or "").upper()
    for pat, name, ttype in DESC_MAP:
        if re.search(pat, d):
            return name, ttype
    if section in SECTION_MAP:
        return SECTION_MAP[section]
    return (desc or section or "Scheduled work").strip()[:60], None


def read_schedule_dated(path: Path):
    """Main Schedule → [{job, line, section, desc, date, crew}]. Column G is the
    row's own date; blank means 'the board date' = the schedule file's date."""
    m = RP._SCHED_FILE_RE.search(path.name)
    board = (dt.date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
             if m else dt.date.today())
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames
                  if s.strip().lower() == "main schedule"), None)
    if sheet is None:
        wb.close()
        raise SystemExit(f"No 'Main Schedule' tab in {path.name}")
    ws = wb[sheet]
    out, section = [], None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        vals = [c.value for c in row[:12]]
        text = [str(v).strip() if v is not None else "" for v in vals]
        joined = P._norm(" ".join(text))
        job = text[1].upper() if len(text) > 1 else ""
        if not _JOB_RE.match(job):
            for token, _scope in P.SECTION_SCOPE:
                if token in joined and "NAME" not in joined \
                        and "SUPERINTENDENT" not in joined:
                    section = token
                    break
            continue
        if section is None:
            continue
        desc = P._norm(text[5])
        scope = "ftw" if ("FLATWORK" in desc or "FTW" in desc
                          or "PAVING" in desc) else "slab"
        if section == "FLATWORK":
            scope = "ftw"
        line = job if (scope == "slab" or job.startswith("CP")) else f"{job}-FTW"
        d = vals[6]
        when = d.date() if isinstance(d, dt.datetime) else (
            d if isinstance(d, dt.date) else board)
        out.append({"job": job, "line": line, "section": section, "desc": desc,
                    "date": when, "crew": text[7], "dated": bool(d)})
    wb.close()
    return out, board


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--as-of", help="treat this date as today (YYYY-MM-DD); never reads a schedule dated after it")
    ap.add_argument("--schedule", help="schedule xlsx (default: latest)")
    ap.add_argument("--job", help="limit to one job #")
    ap.add_argument("--commit", action="store_true",
                    help="actually write to JobTread (default: dry run)")
    args = ap.parse_args()

    print("\n  JOBTREAD SCHEDULE WRITER — Excel schedule → dated tasks")
    print(f"  {'COMMIT — writing to JobTread' if args.commit else 'DRY RUN — nothing will be written'}")
    print("  " + "─" * 62)

    if args.schedule:
        sched_path = Path(args.schedule)
    else:
        _cap = (dt.date.fromisoformat(args.as_of) if getattr(args, 'as_of', None)
                else None)
        best = SCHED.schedule_on_or_before(_cap)
        if best is None:
            print("  ✗ no schedule found")
            return 1
        sched_path = best[1]
    print(f"  schedule: {sched_path.name}")
    lines, board = read_schedule_dated(sched_path)
    if args.job:
        lines = [l for l in lines if l["job"] == args.job.upper()]
    print(f"  scheduled lines: {len(lines)} (board date {board})")

    key = qbo_vault.get("JT_GRANT_KEY")
    types = pave(key, {"organization": {"$": {"id": ORG_ID}, "taskTypes": {
        "$": {"size": 30}, "nodes": {"id": {}, "name": {}}}}})
    type_id = {t["name"]: t["id"]
               for t in types["organization"]["taskTypes"]["nodes"]}
    print(f"  task types: {', '.join(sorted(type_id))}")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = LOG_DIR / f"jobtread_schedule_{stamp}.jsonl"

    created = updated = same = missing = failed = 0
    log = open(log_path, "a")
    seen = set()
    for n, l in enumerate(lines, 1):
        key_ = (l["line"], stage_for(l["section"], l["desc"])[0])
        if key_ in seen:
            continue
        seen.add(key_)
        name, ttype = stage_for(l["section"], l["desc"])
        when = l["date"].isoformat()
        try:
            r = pave(key, {"organization": {"$": {"id": ORG_ID}, "jobs": {
                "$": {"size": 1, "where": {"and": [["number", "=", l["line"]]]}},
                "nodes": {"id": {}, "number": {}, "tasks": {
                    "$": {"size": 60},
                    "nodes": {"id": {}, "name": {}, "startDate": {},
                              "endDate": {}}}}}}})
            nodes = r["organization"]["jobs"]["nodes"]
        except urllib.error.HTTPError as e:
            print(f"   {l['line']:12} ✗ lookup failed {e.code}")
            failed += 1
            continue
        if not nodes:
            print(f"   {l['line']:12} · not in JobTread — skip "
                  f"({name} {when})")
            missing += 1
            continue
        job = nodes[0]
        existing = next((t for t in job["tasks"]["nodes"]
                         if (t["name"] or "").strip().lower() == name.lower()),
                        None)
        if existing:
            if existing.get("startDate") == when and existing.get("endDate") == when:
                same += 1
                continue
            if not args.commit:
                print(f"   {l['line']:12} would UPDATE {name!r} "
                      f"{existing.get('startDate')} → {when}")
                updated += 1
                continue
            try:
                pave(key, {"updateTask": {"$": {
                    "id": existing["id"], "startDate": when, "endDate": when}}})
                print(f"   {l['line']:12} ✓ updated {name!r} → {when}")
                log.write(json.dumps({"job": l["line"], "task": name,
                                      "action": "update",
                                      "from": existing.get("startDate"),
                                      "to": when}) + "\n")
                updated += 1
            except urllib.error.HTTPError as e:
                print(f"   {l['line']:12} ✗ update failed {e.code}")
                failed += 1
        else:
            if not args.commit:
                print(f"   {l['line']:12} would CREATE {name!r} [{ttype}] {when}")
                created += 1
                continue
            a = {"targetType": "job", "targetId": job["id"], "name": name,
                 "startDate": when, "endDate": when}
            if ttype and ttype in type_id:
                a["taskTypeId"] = type_id[ttype]
            try:
                pave(key, {"createTask": {"$": a, "createdTask": {"id": {}}}})
                print(f"   {l['line']:12} ✓ created {name!r} [{ttype}] {when}")
                log.write(json.dumps({"job": l["line"], "task": name,
                                      "action": "create", "date": when}) + "\n")
                created += 1
            except urllib.error.HTTPError as e:
                print(f"   {l['line']:12} ✗ create failed {e.code} "
                      f"{e.read().decode()[:110]}")
                failed += 1
        if n % 25 == 0:
            print(f"    …{n}/{len(lines)}")
    log.close()

    print("  " + "─" * 62)
    verb = "would create" if not args.commit else "created"
    verb2 = "would update" if not args.commit else "updated"
    print(f"  {verb}: {created} · {verb2}: {updated} · already correct: {same} "
          f"· job not in JobTread: {missing} · failed: {failed}")
    if args.commit:
        print(f"  log → {log_path}")
    else:
        print("  DRY RUN — nothing written. Re-run with --commit to apply.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
