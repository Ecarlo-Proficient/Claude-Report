#!/usr/bin/env python3
"""
money_out_register.py — check register for tracking uncashed (outstanding) checks.

QBO exposes every check we wrote (check #, payee, amount, bank, date) but NOT
whether it cleared the bank — there is no reconciled/cleared status in the API.
So this builds the money-OUT register and YOU own the cleared flag: mark a check
'Y' in the CLEARED? column once it clears, and everything still 'N'/blank is an
uncashed / outstanding check (the user 2026-07-17).

STATEFUL — your CLEARED?/CLEARED DATE marks are PRESERVED across runs (merged by
QBO txn id). New checks are appended; cleared checks older than the prune window
drop off so the register stays the outstanding list.

SOURCES (money out, check type only): BillPayment PayType=Check + Purchase
PaymentType=Check.

OUTPUT  ~/Documents/CompanyHealth/Money Out Register.xlsx  (chmod 600)
Read-only against QBO. One Touch ID per run.

USAGE
  python3 one-offs/money_out_register.py
  python3 one-offs/money_out_register.py --days 120     # pull window
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import stat
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import qbo_api
from shared import paths

DEFAULT_OUTPUT = paths.companyhealth_dir() / "Money Out Register.xlsx"
PULL_DAYS = 45           # how far back to pull checks (recent = plausibly uncashed)
PRUNE_CLEARED_DAYS = 45  # drop checks marked cleared older than this
SHEET = "Register"


def _today() -> dt.date:
    return dt.date.today()


def _parse(s) -> Optional[dt.date]:
    if isinstance(s, dt.datetime):
        return s.date()
    if isinstance(s, dt.date):
        return s
    try:
        return dt.date.fromisoformat(str(s))
    except (TypeError, ValueError):
        return None


# ────────────────────────── QBO pull ──────────────────────────

def pull_checks(access, cid, since: str) -> List[dict]:
    """All check-type money-out from QBO → register rows (no cleared status —
    that's ours to track)."""
    rows: List[dict] = []
    for bp in qbo_api.query_all(access, cid, "BillPayment", f"TxnDate >= '{since}'"):
        if bp.get("PayType") != "Check":
            continue
        chk = bp.get("CheckPayment") or {}
        rows.append({
            "key": f"BP:{bp['Id']}",
            "check": str(bp.get("DocNumber") or ""),
            "date": _parse(bp.get("TxnDate")),
            "payee": (bp.get("VendorRef") or {}).get("name") or "",
            "amount": float(bp.get("TotalAmt") or 0),
            "bank": (chk.get("BankAccountRef") or {}).get("name") or "",
            "type": "Bill Pmt",
            "memo": "",
        })
    for pu in qbo_api.query_all(access, cid, "Purchase", f"TxnDate >= '{since}'"):
        if pu.get("PaymentType") != "Check":
            continue
        rows.append({
            "key": f"PU:{pu['Id']}",
            "check": str(pu.get("DocNumber") or ""),
            "date": _parse(pu.get("TxnDate")),
            "payee": (pu.get("EntityRef") or {}).get("name") or "",
            "amount": float(pu.get("TotalAmt") or 0),
            "bank": (pu.get("AccountRef") or {}).get("name") or "",
            "type": "Check",
            "memo": (pu.get("PrivateNote") or "")[:60],
        })
    return rows


# ────────────────────── merge (preserve marks) ──────────────────────

def load_marks(path: Path) -> Dict[str, dict]:
    """Existing CLEARED?/CLEARED DATE keyed by _Key, so a refresh never wipes
    the user's reconciliation."""
    marks: Dict[str, dict] = {}
    if not path.exists():
        return marks
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return marks
    if SHEET not in wb.sheetnames:
        wb.close()
        return marks
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {n: i for i, n in enumerate(hdr)}
    kc, cc, dc = idx.get("_Key"), idx.get("CLEARED?"), idx.get("CLEARED DATE")
    for row in it:
        if kc is None or kc >= len(row) or not row[kc]:
            continue
        marks[str(row[kc])] = {
            "cleared": str(row[cc]).strip().upper() if cc is not None and row[cc] else "",
            "cleared_date": row[dc] if dc is not None and dc < len(row) else None,
        }
    wb.close()
    return marks


def merge(pulled: List[dict], marks: Dict[str, dict],
          existing_keys: Dict[str, dict]) -> List[dict]:
    """Union of pulled checks + any still-outstanding checks already in the
    register; carry forward marks; prune cleared checks past the window."""
    today = _today()
    by_key: Dict[str, dict] = {}
    # start from existing rows (so checks that aged out of the pull window but
    # are still uncashed stay on the register)
    for k, r in existing_keys.items():
        by_key[k] = dict(r)
    for r in pulled:
        by_key[r["key"]] = {**by_key.get(r["key"], {}), **r}
    out = []
    for k, r in by_key.items():
        m = marks.get(k, {})
        cleared = m.get("cleared", "")
        cdate = _parse(m.get("cleared_date"))
        d = _parse(r.get("date"))
        # prune long-cleared checks to keep the register the outstanding list
        if cleared == "Y" and cdate and (today - cdate).days > PRUNE_CLEARED_DAYS:
            continue
        if cleared == "Y" and not cdate and d and (today - d).days > PRUNE_CLEARED_DAYS:
            continue
        r["cleared"] = cleared
        r["cleared_date"] = cdate
        r["days_out"] = (today - d).days if d else None
        out.append(r)
    # outstanding first, oldest first; cleared sink to the bottom
    out.sort(key=lambda x: (x["cleared"] == "Y", -(x["days_out"] or 0)))
    return out


def load_existing_rows(path: Path) -> Dict[str, dict]:
    """Existing register rows keyed by _Key (to retain outstanding checks that
    fell outside the current pull window)."""
    rows: Dict[str, dict] = {}
    if not path.exists():
        return rows
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return rows
    if SHEET not in wb.sheetnames:
        wb.close()
        return rows
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {n: i for i, n in enumerate(hdr)}

    def g(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    for row in it:
        k = g(row, "_Key")
        if not k:
            continue
        rows[str(k)] = {
            "key": str(k), "check": str(g(row, "CHECK #") or ""),
            "date": _parse(g(row, "DATE")), "payee": str(g(row, "PAYEE") or ""),
            "amount": float(g(row, "AMOUNT $") or 0), "bank": str(g(row, "BANK") or ""),
            "type": str(g(row, "TYPE") or ""), "memo": str(g(row, "MEMO") or ""),
        }
    wb.close()
    return rows


# ────────────────────────── Excel ──────────────────────────

CUR = '#,##0.00'
_NAVY = PatternFill("solid", fgColor="1F3864")
_ZEBRA = PatternFill("solid", fgColor="EEF3FA")
_RED = PatternFill("solid", fgColor="FFC7CE")
_AMBER = PatternFill("solid", fgColor="FFEB9C")
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_WHITE = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D7E5")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
COLS = ["CHECK #", "DATE", "PAYEE", "AMOUNT $", "BANK", "TYPE", "MEMO",
        "DAYS OUT", "CLEARED?", "CLEARED DATE", "_Key"]
WIDTHS = [10, 12, 30, 14, 26, 10, 34, 10, 11, 13, 14]


def write_register(path: Path, rows: List[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(1, c)
        cell.font = _WHITE
        cell.fill = _NAVY
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows):
        ws.append([r["check"], r["date"].isoformat() if r["date"] else "",
                   r["payee"], r["amount"], r["bank"], r["type"], r["memo"],
                   r["days_out"] if r["days_out"] is not None else "",
                   r["cleared"], r["cleared_date"].isoformat() if r["cleared_date"] else "",
                   r["key"]])
        rr = ws.max_row
        for c in range(1, len(COLS) + 1):
            ws.cell(rr, c).border = _BORDER
            if i % 2:
                ws.cell(rr, c).fill = _ZEBRA
        ws.cell(rr, 4).number_format = CUR
        # colour the CLEARED? cell: cleared=green; else amber, red if aged >30d
        cc = ws.cell(rr, 9)
        if r["cleared"] == "Y":
            cc.fill = _GREEN
        elif (r["days_out"] or 0) > 30:
            cc.fill = _RED
            cc.value = cc.value or "N"
        else:
            cc.fill = _AMBER
    last = ws.max_row

    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)
    if last >= 2:
        dv.add(f"I2:I{last}")
    ws.auto_filter.ref = f"A1:J{last}"
    ws.column_dimensions["K"].hidden = True

    # Summary sheet
    s = wb.create_sheet("Summary", 0)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 3
    s.column_dimensions["B"].width = 40
    s.column_dimensions["C"].width = 18
    outstanding = [r for r in rows if r["cleared"] != "Y"]
    out_tot = sum(r["amount"] for r in outstanding)
    aged = [r for r in outstanding if (r["days_out"] or 0) > 30]
    s.append(["", "MONEY OUT — UNCASHED CHECK REGISTER"])
    s.merge_cells("B1:C1")
    s.cell(1, 2).font = Font(bold=True, size=14, color="FFFFFF")
    for c in (1, 2, 3):
        s.cell(1, c).fill = _NAVY
    s.row_dimensions[1].height = 30
    s.append(["", f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} · mark CLEARED? = Y "
              "on the Register tab as checks clear; the rest are outstanding"])
    s.cell(2, 2).font = Font(italic=True, color="595959")
    s.merge_cells("B2:C2")

    def kpi(label, value, money=True, bad=False):
        s.append(["", label, value])
        r = s.max_row
        s.cell(r, 2).font = Font(bold=True)
        v = s.cell(r, 3)
        v.font = Font(bold=True, size=12, color="9C0006" if bad else "1F6B4C")
        if money:
            v.number_format = CUR
        for c in (2, 3):
            s.cell(r, c).border = _BORDER
            s.cell(r, c).fill = PatternFill("solid", fgColor="FCE4E4" if bad else "E7F2E7")
    aged_tot = sum(r["amount"] for r in aged)
    s.append([])
    kpi("Unmarked checks (not yet reconciled)", len(outstanding), money=False)
    kpi("Unmarked $ (until you mark them cleared)", round(out_tot))
    s.append([])
    kpi("⚠ AGED > 30 DAYS & still unmarked", len(aged), money=False, bad=bool(aged))
    kpi("Aged > 30 days $ (the real chase list)", round(aged_tot), bad=bool(aged))
    s.sheet_properties.tabColor = "C00000"
    ws.sheet_properties.tabColor = "1F3864"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)


def main() -> int:
    ap = argparse.ArgumentParser(description="Money-out uncashed check register")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument("--days", type=int, default=PULL_DAYS)
    args = ap.parse_args()

    since = (_today() - dt.timedelta(days=args.days)).isoformat()
    print("\n  MONEY OUT — uncashed check register")
    print("  " + "─" * 45)
    marks = load_marks(args.out)
    existing = load_existing_rows(args.out)
    print(f"  existing register: {len(existing)} row(s), {len(marks)} mark(s) preserved")

    access, cid = qbo_api.load_credentials()
    pulled = pull_checks(access, cid, since)
    print(f"  QBO checks since {since}: {len(pulled)}")

    rows = merge(pulled, marks, existing)
    outstanding = [r for r in rows if r["cleared"] != "Y"]
    print(f"  register: {len(rows)} row(s) · outstanding {len(outstanding)} "
          f"= ${sum(r['amount'] for r in outstanding):,.0f}")

    try:
        write_register(args.out, rows)
    except PermissionError:
        print(f"  ✗ {args.out} is open in Excel — close it and re-run.")
        return 2
    print(f"\n  ✓ {args.out}  (chmod 600)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
