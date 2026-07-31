#!/usr/bin/env python3
"""
rp_wip_simple.py — the stripped-down RP WIP (the user 2026-07-27).

ALL RP WIP lines in ONE table on the 'Test - RP' tab. Reporting stripped to
the basics the user asked for:

    JOB # · ADDRESS · BUILDER
    CONTRACT $ · ETC $ · BILLED TO DATE $ · COSTS TO DATE $ · GP %
    SCHEDULE · GENERAL LISTA · ACTION

Rules (the user):
  • SCHEDULE / GENERAL LISTA — a check mark when the job is on that source.
  • GREEN numbers = the folder holds exactly ONE bid proposal and ONE takeoff,
    so the right number is unambiguous. Anything else stays black (judgment
    needed — several revisions/scopes to choose between).
  • CONTRACT / ETC link to the project FOLDER — they open the location in
    Finder, never the PDF or the workbook itself.
  • BILLED / COSTS link to that project in QBO.
  • ACTION says what the job still needs.

Sources: the daily schedule (active lines), the General Lista (presence +
fallback pricing), the project folders (proposal PDF → contract, takeoff cost
sheet → ETC), and QBO (billed/costs, read-only).

Production write is GATED: dry run writes a preview workbook to Downloads;
--commit writes the 'Test - RP' tab (guarded by wip_excel_guard).

Usage
  python3 rp_wip_simple.py                    # preview → Downloads
  python3 rp_wip_simple.py --commit           # write the Test - RP tab
  python3 rp_wip_simple.py --schedule <f.xlsx>
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from shared import qbo_api
from shared import schedule as SCHED
import rp_wip_reader as RP
import rp_schedule_wip_preview as P
import wip_excel_guard as GUARD
import cp_wip_reader as CP

TAB = "Test - RP"
CHECK = "✓"
CROSS = "✗"
OKF = Font(color="006100", bold=True)      # ✓ green
NOF = Font(color="9C0006", bold=True)      # ✗ red
GREEN = Font(color="006100", bold=True)
BLACK = Font(color="000000")
BOLD = Font(bold=True)
LINKF = Font(color="0563C1", underline="single")
LINKF_GREEN = Font(color="006100", bold=True, underline="single")
UNDER = Border(bottom=Side(style="thin", color="000000"))
CUR = '"$"#,##0.00'
PCT = '0.0%'

HDR = ["JOB #", "ADDRESS", "BUILDER", "CONTRACT $", "ETC $",
       "BILLED TO DATE $", "COSTS TO DATE $", "GP %",
       "SCHEDULE", "GENERAL LISTA", "ACTION", "CO $",
       "CONTRACT FILE (copy → Finder ⇧⌘G)", "ETC FILE (copy → Finder ⇧⌘G)",
       "CO FILE (copy → Finder ⇧⌘G)", "|"]
WIDTHS = (14, 30, 24, 15, 15, 16, 16, 9, 10, 14, 46, 13, 58, 58, 58, 3)
ITALIC = Font(italic=True)


# ── "is the number unambiguous?" ────────────────────────────────────
def _count_sources(folder: Path):
    """(#bid-proposal PDFs, #takeoff workbooks) in the project folder. Exactly
    one of each = the right number is obvious → the row goes green."""
    if folder is None:
        return 0, 0
    props = tkoffs = 0
    try:
        for f in folder.iterdir():
            n = P._norm(f.name)
            suf = f.suffix.lower()
            if f.name.startswith("~$"):
                continue
            if suf == ".pdf":
                if "INVOICE" in n or "DIAGRAM" in n or "PLAN" in n:
                    continue
                if "BID" in n or "PROPOSAL" in n:
                    props += 1
            elif suf in (".xlsm", ".xlsx"):
                if "TAKEOFF" in n or n.startswith("RP") or n.startswith("CP"):
                    tkoffs += 1
    except OSError:
        return 0, 0
    return props, tkoffs


def _sheet_subtotal(ws):
    """The 'SUB TOTAL:' figure on a bid sheet (value to the right of the label)."""
    if ws is None:
        return None
    for row in ws.iter_rows(min_row=1, max_row=60, max_col=12):
        for k, c in enumerate(row):
            if isinstance(c.value, str) and "SUB TOTAL" in c.value.upper():
                for v in row[k + 1:]:
                    if isinstance(v.value, (int, float)) and v.value:
                        return float(v.value)
    return None


def takeoff_etc(tk_path, scope):
    """ETC from the takeoff's cost sheet.

    THE RULE (the user 2026-07-30): **the PIERS line (row 18) IS the sub total.**
    Read it — but ALWAYS verify by reading its FORMULA first, because the
    template ships two conventions and they mean different things:

        D18 = SUM(D10:D17)  → range starts at the SLAB SUBTOTAL, so row 18 is
                              the FOUNDATION total (slab + piers) → ETC = D18
        D18 = SUM(D11:D17)  → range starts at PR1, so row 18 is piers ONLY
                              → ETC = D10 + D18

    Taking row 18 without reading the formula double-counts the slab on every
    takeoff using the first convention (verified: RP7482 combined,
    RP7470 piers-only). If row 18 is #N/A the pier cost is unknown — fall back
    to D10 + the numeric PR code rows and flag it.

    D10 is read as a VALUE, never recomputed: it sometimes carries a manual
    adder the estimator typed in (nuance cost with no cost code) that must be
    kept. FW is NEVER added to a non-FTW ETC — flatwork is its own line.

    Returns (etc, notes, detail)."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(tk_path, data_only=True)
        wbf = _lw(tk_path, data_only=False)
    except Exception:
        return None, ["takeoff unreadable"], {}
    try:
        sheet = next((n for n in wb.sheetnames if "cost gral" in n.lower()), None)
        if sheet is None:
            return None, ["no 'Cost Gral' sheet in the takeoff"], {}
        ws, wsf = wb[sheet], wbf[sheet]
        slab = ws["D10"].value
        flat = ws["D27"].value
        sub18 = ws["D18"].value                 # the PIERS line = the sub total
        f18 = str(wsf["D18"].value or "")       # ...verified by its formula
        pr_cells = [ws.cell(r, 4).value for r in range(11, 18)]
    finally:
        wb.close()
        wbf.close()

    notes = []
    pr_na = [c for c in pr_cells if isinstance(c, str) and c.strip()]
    piers_rows = sum(c for c in pr_cells if isinstance(c, (int, float)))

    # Which convention is row 18 using? Read the formula, don't assume.
    compact = f18.replace(" ", "").upper()
    if "D10:D17" in compact:
        conv = "combined"        # row 18 already includes the slab subtotal
    elif "D11:D17" in compact:
        conv = "piers-only"
    else:
        conv = "unknown"

    # Split the pier band: genuine pier work vs non-pier site work that merely
    # LIVES on the Piers takeoff sheet (tie wire / scrape lot / reset forms).
    # Those are REAL costs — they stay counted; the split only drives flags.
    genuine = nonpier = 0.0
    try:
        wb2 = _lw(tk_path, data_only=True)
        if "Piers takeoff" in wb2.sheetnames:
            pv = wb2["Piers takeoff"]
            for rr in range(18, 36):
                lab = str(pv.cell(rr, 1).value or "").upper()
                e = pv.cell(rr, 5).value
                if isinstance(e, (int, float)) and e > 0:
                    if any(k in lab for k in ("TIE WIRE", "SCRAPE", "RESET", "STAKE")):
                        nonpier += e
                    else:
                        genuine += e
        wb2.close()
    except Exception:
        pass
    detail = {"slab": slab, "piers": piers_rows, "flat": flat,
              "pr_na": len(pr_na), "sub18": sub18, "d18_formula": f18,
              "convention": conv, "pier_genuine": round(genuine, 2),
              "pier_nonpier": round(nonpier, 2)}

    if scope == "ftw":
        if not isinstance(flat, (int, float)):
            return None, ["⚠ flatwork subtotal is #N/A — fix the takeoff"], detail
        return round(flat, 2), notes, detail

    if not isinstance(slab, (int, float)):
        return None, ["⚠ slab subtotal is #N/A — fix the takeoff"], detail

    # Take the PIERS line as the sub total, per its verified formula.
    if isinstance(sub18, (int, float)) and conv != "unknown":
        etc = float(sub18) if conv == "combined" else float(slab) + float(sub18)
        # cross-check against the code rows; disagreement means the range was
        # edited or a cell was hand-typed
        expect = float(slab) + piers_rows
        if abs(etc - expect) > 1.0:
            notes.append(f"⚠ row 18 ({conv}, {f18}) gives ${etc:,.0f} but the code "
                         f"rows give ${expect:,.0f} — verify the takeoff")
    else:
        etc = float(slab) + piers_rows
        if isinstance(sub18, str) and sub18.strip():
            notes.append(f"⚠ piers sub total is {sub18} — used the code rows "
                         f"instead (${piers_rows:,.0f} of piers counted)")
        elif conv == "unknown" and f18:
            notes.append(f"⚠ unrecognised piers sub-total formula {f18} — "
                         "used the code rows instead")

    if pr_na:
        notes.append(f"⚠ {len(pr_na)} pier cost row(s) are #N/A — pier cost is "
                     f"INCOMPLETE — fix the takeoff")
    if nonpier:
        notes.append(f"note: ${nonpier:,.0f} of the pier band is site work "
                     "(tie wire / scrape lot / reset forms live on the Piers "
                     "takeoff sheet) — counted, as expected")
    return round(etc, 2), notes, detail


def read_proposal_lines(pdf_path):
    """Read the proposal's ITEMIZED LINES (scope section above SUB TOTAL) —
    never a whole-document keyword search (the user 2026-07-28, RP7490 fired on
    the terms-and-conditions boilerplate "…WHEN DIGGING PIERS…").

    Returns {'piers': bool, 'contingent': float, 'contingent_lines': [str]}
      piers      — a PRICED line item that quotes piers
      contingent — Σ of 'if needed / if required' lines: they carry a unit rate
                   but are DELIBERATELY excluded from SUB TOTAL (rock beams,
                   rock saw…). Never contract value; real exposure if hit.
    """
    out = {"piers": False, "contingent": 0.0, "contingent_lines": []}
    if not pdf_path:
        return out
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            txt = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception:
        return out
    up = txt.upper()
    cut = up.find("SUB TOTAL")
    scope = txt[:cut] if cut != -1 else txt      # terms/conditions live below
    for line in scope.splitlines():
        u = line.upper()
        if "PIER" in u and re.search(r"\d", u) and re.search(r"\$|\d{2,}", u):
            out["piers"] = True
        if re.search(r"IF NEEDED|IF REQUIRED", u):
            nums = re.findall(r"\$?([\d,]+\.?\d*)", line)
            if len(nums) >= 2:
                try:
                    q = float(nums[-2].replace(",", ""))
                    r_ = float(nums[-1].replace(",", ""))
                    if q > 10 and 0 < r_ < 10:      # qty × small unit rate
                        out["contingent"] += q * r_
                        out["contingent_lines"].append(line.strip()[:40])
                except ValueError:
                    pass
    out["contingent"] = round(out["contingent"], 2)
    return out


def find_change_orders(folder):
    """(total $, [paths]) for CHANGE ORDER / ADDENDUM docs in the project
    folder — the tool prices ONE proposal, so an approved CO is otherwise
    invisible and the contract is understated (the user 2026-07-28)."""
    if folder is None:
        return 0.0, []
    tot, paths = 0.0, []
    try:
        for f in sorted(folder.iterdir()):
            if f.name.startswith("~$") or f.suffix.lower() != ".pdf":
                continue
            if not re.search(r"CHANGE ORDER|\bC\.?O\.?\b|ADDENDUM", f.name, re.I):
                continue
            paths.append(f)
            amt = P.pdf_subtotal(f)
            if amt:
                tot += amt
    except OSError:
        return 0.0, []
    return round(tot, 2), paths


def proposal_has_piers(pdf_path):
    """Does the bid proposal actually QUOTE piers as a PRICED LINE ITEM?

    NOT a plain "PIER" text search — every proposal's terms-and-conditions
    boilerplate says "…WHEN DIGGING PIERS, GRADE BEAMS…", which made the naive
    check fire on jobs with no piers at all (RP7490, the user 2026-07-28).
    So: read only the scope section ABOVE 'SUB TOTAL', and only count a line
    that names a pier AND carries a number (qty or $)."""
    if not pdf_path:
        return False
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            txt = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception:
        return False
    up = txt.upper()
    cut = up.find("SUB TOTAL")
    scope = up[:cut] if cut != -1 else up      # terms/conditions live below it
    for line in scope.splitlines():
        if "PIER" not in line:
            continue
        if re.search(r"[\d]", line) and re.search(r"\$|\d{2,}", line):
            return True
    return False


def takeoff_bid(tk_path):
    """(bid, type) from a takeoff: INFORMATION!K16 'TYPE OF CONSTRUCTION' picks
    Bid Post Tension vs Bid Tella Firma. VERIFIED on 64 active takeoffs — the
    template prices BOTH sheets on 47 of them, so the type field is what decides;
    it resolves 81% outright and the rest are flagged rather than guessed."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(tk_path, data_only=True)
    except Exception:
        return None, "", "takeoff unreadable"
    try:
        info = wb["INFORMATION"] if "INFORMATION" in wb.sheetnames else None
        typ = str(info["K16"].value).strip() if info is not None and \
            info["K16"].value not in (None, "") else ""
        pt = _sheet_subtotal(wb["Bid Post Tension"]) if "Bid Post Tension" in wb.sheetnames else None
        tf = _sheet_subtotal(wb["Bid Tella Firma"]) if "Bid Tella Firma" in wb.sheetnames else None
    finally:
        wb.close()
    tu = typ.upper()
    if not typ:
        return (pt or tf), typ, "⚠ TYPE OF CONSTRUCTION blank — check the plans"
    if "POST" in tu:
        if pt:
            return pt, typ, ""
        return tf, typ, "⚠ type=POST TENSION but only Tella Firma is priced — check the plans"
    if "TELLA" in tu or "FIRMA" in tu:
        if tf:
            return tf, typ, ""
        return pt, typ, "⚠ type=TELLA FIRMA but only Post Tension is priced — check the plans"
    return (pt or tf), typ, f"⚠ unrecognised construction type {typ!r}"


def _folder_url(folder: Path):
    return f"file://{folder}" if folder else None


def all_schedules(sched_dir: Path):
    """[(key, path)] for every Schedule M-D-YY.xlsx, oldest → newest."""
    out = []
    for year_dir in sched_dir.iterdir():
        if not (year_dir.is_dir() and year_dir.name.isdigit()):
            continue
        for month_dir in year_dir.iterdir():
            if not month_dir.is_dir():
                continue
            for f in month_dir.iterdir():
                m = RP._SCHED_FILE_RE.search(f.name)
                if m:
                    mo, dy, yy = (int(g) for g in m.groups())
                    out.append(((2000 + yy, mo, dy), f))
    return sorted(out)


def prior_schedule(sched_dir: Path, current: Path):
    """The schedule file immediately before `current` (for the drop check)."""
    scheds = all_schedules(sched_dir)
    for i, (_k, p) in enumerate(scheds):
        if p == current and i > 0:
            return scheds[i - 1][1]
    return scheds[-2][1] if len(scheds) > 1 else None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--as-of", help="treat this date as today (YYYY-MM-DD); never reads a schedule dated after it")
    ap.add_argument("--schedule", help="schedule xlsx (default: latest)")
    ap.add_argument("--commit", action="store_true",
                    help="write the 'Test - RP' tab (production)")
    ap.add_argument("--out", help="preview xlsx path")
    args = ap.parse_args()

    print("\n  RP WIP — simplified (contract · ETC · billed · costs · GP%)")
    print("  " + "─" * 62)

    # 1) schedule ------------------------------------------------------
    if args.schedule:
        sched_path = Path(args.schedule)
        m = RP._SCHED_FILE_RE.search(sched_path.name)
        label = "-".join(m.groups()) if m else sched_path.stem
    else:
        _cap = (dt.date.fromisoformat(args.as_of) if getattr(args, 'as_of', None)
                else None)
        best = SCHED.schedule_on_or_before(_cap)
        if best is None:
            print("  ✗ no schedule found")
            return 1
        _d, sched_path = best
        label = f"{_d.month}-{_d.day}-{_d.year % 100:02d}"
    if not sched_path.exists():
        print(f"  ✗ schedule not found: {sched_path}")
        return 1
    print(f"  schedule: {sched_path.name}")
    sched = P.read_main_schedule(sched_path)
    print(f"  active schedule lines: {len(sched)}")

    # 2) General Lista + folders --------------------------------------
    gl_jobs = P.read_gl_jobs(RP.ALPHA_PATH)
    records, _ = RP.read_general_list(RP.ALPHA_PATH)
    gl_by_job = {r["job"]: r for r in records}
    rp_to_folders, addr_folders = RP.index_residential(RP.RP_ROOT)

    # 3) build one row per scheduled scope ----------------------------
    def build_row(s, on_sched=True):
        job, scope = s["job"], s["scope"]
        line = job if (scope == "slab" or job.startswith("CP")) else f"{job}-FTW"
        rec = gl_by_job.get(job)
        if job.startswith("CP"):
            gl_k = ((rec["slab_bid"] or 0) + (rec["flat_bid"] or 0)) or None if rec else None
            gl_e = ((rec["slab_cost"] or 0) + (rec["flat_cost"] or 0)) or None if rec else None
        elif scope == "ftw":
            gl_k = rec and rec["flat_bid"]
            gl_e = rec and rec["flat_cost"]
            if rec and rec.get("flat_other"):
                gl_k = gl_e = None            # another contractor won the flatwork
        else:
            gl_k = rec and rec["slab_bid"]
            gl_e = rec and rec["slab_cost"]

        folders = sorted(rp_to_folders.get(job, ()),
                         key=lambda f: (f.parent.name, f.name))
        folder = folders[0] if folders else None
        if folder is None and s["address"]:
            parts = s["address"].split(None, 1)
            folder = RP.match_by_address(
                {"house": parts[0] if parts else "",
                 "street": parts[1] if len(parts) > 1 else s["address"]},
                addr_folders)

        ov = P.OVERRIDES.get(line, {})
        contract = etc = None
        k_src = e_src = None          # the FILE each number came from
        etc_notes, bands = [], {}
        provisional = False           # tract, no P.O. yet → takeoff bid
        type_note = ""
        if ov.get("proposal"):
            contract = P.pdf_subtotal(Path(ov["proposal"]))
            k_src = Path(ov["proposal"])
        if ov.get("takeoff"):
            etc, _n, _f = P.takeoff_budget_from(Path(ov["takeoff"]), scope, s["desc"])
            e_src = Path(ov["takeoff"])
        tk_path = e_src
        if folder is not None:
            if contract is None:
                _p, contract, _n = P.find_proposal(folder, scope, s["desc"])
                if contract is not None:
                    k_src = _p
            if etc is None:
                _t, _etc_old, _n, _f = P.find_takeoff_etc(folder, job, scope, s["desc"])
                if _t is not None:
                    e_src = tk_path = _t
                    # recompute with OUR rules: no FW in a slab ETC, #N/A excluded
                    etc, etc_notes, bands = takeoff_etc(_t, scope)

        tract = P._is_tract(s["builder"])
        if tract and contract is None:
            # No signed P.O. yet. Use the TAKEOFF bid as the STARTING contract
            # (Post Tension vs Tella Firma chosen by INFORMATION!K16) and mark
            # it provisional — italic + a note (the user 2026-07-28).
            if tk_path is not None:
                tb, ttyp, tnote = takeoff_bid(tk_path)
                if tb:
                    contract, k_src, provisional = tb, tk_path, True
                    type_note = tnote
            if contract is None and gl_k:
                contract, provisional = gl_k, True
        if line in getattr(P, "GL_ETC_JOBS", ()) and gl_e:
            etc = gl_e
        if contract is None and gl_k:
            contract = gl_k
        if etc is None and gl_e:
            etc = gl_e

        n_prop, n_tko = _count_sources(folder)
        easy = (n_prop == 1 and n_tko == 1)
        co_amt, co_files = find_change_orders(folder)

        # A CP job on the RP schedule belongs to the CP WIP, not this one. Its
        # folder lives under AWARDED CP PROJECTS, so the Residential search
        # can't see it — never price it here, always flag it (the user
        # 2026-07-28, CP865 came through at ~$5 and read "OK").
        is_cp = job.startswith("CP")

        action = []
        if is_cp:
            action.append("⚠ CP JOB — belongs in the CP WIP, not RP"
                          + ("" if folder else
                             " · no folder under awarded CP projects — verify it exists"))
        elif contract is None:
            action.append("Tract — enter contract from P.O./price list" if tract
                          else "Find the bid proposal (contract)")
        # A contract that is present but implausibly small is NOT "OK" — it
        # means the wrong number was picked up.
        SUSPECT = 1000.0
        suspect_k = (isinstance(contract, (int, float)) and 0 < contract < SUSPECT)
        if not is_cp:
            if etc is None:
                action.append("Find the takeoff budget (ETC)")
            if suspect_k:
                action.append(f"⚠ contract looks wrong (${contract:,.2f}) — verify")
            for n in etc_notes:
                action.append(n)
            # PIERS cross-check, BOTH directions (the user 2026-07-28): read the
            # proposal's itemized lines and compare to the takeoff's PR band.
            if scope != "ftw" and k_src and str(k_src).lower().endswith(".pdf"):
                pl = read_proposal_lines(k_src)
                pr_cost = (bands or {}).get("pier_genuine") or 0
                if pl["piers"] and not pr_cost:
                    action.append("⚠ proposal quotes PIERS but no PR cost in the "
                                  "takeoff — budget is missing the piers")
                elif pr_cost and not pl["piers"]:
                    action.append(f"⚠ takeoff budgets PIERS (${pr_cost:,.0f}) but the "
                                  "proposal never quotes them — budget grabbed scope "
                                  "that is not sold")
                if pl["contingent"]:
                    action.append(f"note: ${pl['contingent']:,.0f} contingent "
                                  "'if needed' (rock beams/saw etc.) NOT in the contract")
            if co_amt:
                action.append(f"⚠ CHANGE ORDER ${co_amt:,.0f} found in the folder — "
                              "is it in the contract?")
            if provisional:
                action.append("NOT FINAL — takeoff bid used; no P.O. issued yet")
            if type_note:
                action.append(type_note)
            if not action:
                action.append("OK — priced" if easy else
                              (f"OK — verify ({n_prop} proposals / {n_tko} takeoffs)"
                               if (n_prop > 1 or n_tko > 1) else "OK"))
            if job not in gl_jobs:
                action.append("not in General Lista")

        return {
            "line": line, "job": job, "address": s["address"],
            "builder": s["builder"], "contract": contract, "etc": etc,
            "folder": folder, "easy": easy, "on_sched": on_sched,
            "in_gl": job in gl_jobs, "action": " · ".join(action),
            "billed": None, "costs": None, "qbo_url": None,
            "section": "cp" if is_cp else "main", "is_cp": is_cp,
            "k_src": k_src, "e_src": e_src, "provisional": provisional,
            "co_amt": co_amt or None,
            "co_file": (str(co_files[0]) if co_files else ""),
        }

    rows = [build_row(s) for s in sched]
    # CP jobs never render on the RP WIP — they get their own flagged section
    cp_rows = [r for r in rows if r.get("is_cp")]
    rows = [r for r in rows if not r.get("is_cp")]
    # de-dup by line, keep first
    seen, uniq = set(), []
    for r in rows:
        if r["line"] in seen:
            continue
        seen.add(r["line"])
        uniq.append(r)
    rows = uniq

    def _rank(r):
        """Ready-and-unambiguous first, then needs-a-fix (the user 2026-07-28)."""
        a = str(r.get("action") or "")
        needs_fix = a.startswith("Find") or "⚠" in a or "NOT FINAL" in a
        return (1 if needs_fix else 0,
                0 if r.get("easy") else 1,
                not str(a).startswith("OK"),
                r["line"])
    rows.sort(key=_rank)
    print(f"  WIP lines: {len(rows)}")

    # 3b) DROPPED off the schedule since the previous file --------------
    dropped = []
    prev = prior_schedule(RP.SCHEDULE_DIR, sched_path)
    if prev is not None:
        print(f"  drop check vs: {prev.name}")
        prev_sched = P.read_main_schedule(prev)
        cur_lines = {r["line"] for r in rows}
        seen_d = set()
        for s in prev_sched:
            job, scope = s["job"], s["scope"]
            line = job if (scope == "slab" or job.startswith("CP")) else f"{job}-FTW"
            if line in cur_lines or line in seen_d:
                continue
            seen_d.add(line)
            r = build_row(s, on_sched=False)
            if r.get("is_cp"):
                continue          # CP has its own flagged section
            r["section"] = "dropped"
            dropped.append(r)
        print(f"  dropped off the schedule: {len(dropped)} (unbilled ones kept)")

    # 3c) FTW BACKLOG from the General Lista ----------------------------
    # Rule (LOCKED — the user 2026-07-28): backlog = flatwork we EXPECT to do
    # but haven't started. Three tests: priced in the General Lista, NOT
    # AF=OTHER (another contractor won it), and NO -FTW scope on today's
    # schedule. A FOURTH test is applied after QBO below and is absolute:
    #   *** IF AN -FTW LINE HAS COSTS (or billing) IT IS NOT BACKLOG. ***
    # Costs mean the work started, so it is active — it must never sit in the
    # backlog section.
    sched_ftw = {s["job"] for s in sched if s["scope"] == "ftw"}
    backlog = []
    for rec in records:
        job = rec["job"]
        if job.startswith("CP"):
            continue
        if not (rec["flat_bid"] or rec["flat_cost"]):
            continue
        if rec.get("flat_other"):
            continue                       # AF=OTHER — excluded by rule
        if job in sched_ftw:
            continue                       # already on the schedule
        folders = sorted(rp_to_folders.get(job, ()),
                         key=lambda f: (f.parent.name, f.name))
        folder = folders[0] if folders else None
        n_prop, n_tko = _count_sources(folder)
        backlog.append({
            "line": f"{job}-FTW", "job": job,
            "address": str(rec.get("address") or ""),
            "builder": str(rec.get("builder") or ""),
            "contract": rec["flat_bid"], "etc": rec["flat_cost"],
            "folder": folder, "easy": (n_prop == 1 and n_tko == 1),
            "on_sched": False, "in_gl": True,
            "action": "Backlog — flatwork priced, not scheduled yet",
            "billed": None, "costs": None, "qbo_url": None,
            "section": "backlog",
        })
    print(f"  FTW backlog (GL, not on schedule): {len(backlog)}")

    # 4) QBO billed / costs -------------------------------------------
    print("  QBO: billed + costs per line …")
    all_rows = rows + cp_rows + dropped + backlog
    try:
        access, cid = qbo_api.load_credentials()
        pmap = qbo_api.build_project_customer_map(access, cid)
        start, end = "2019-01-01", dt.date.today().isoformat()
        for n, r in enumerate(all_rows, 1):
            cust = pmap.get(r["line"]) or pmap.get(r["job"])
            if not cust:
                continue
            try:
                pl = qbo_api.fetch_project_pl(access, cid, cust["id"], start, end)
                t = qbo_api.extract_pl_totals(pl)
                r["billed"] = t.get("income") or 0.0
                r["costs"] = (t.get("cogs") or 0.0) + (t.get("expenses") or 0.0)
                r["qbo_url"] = qbo_api.customer_url(cust["id"], cid)
            except Exception:
                pass
            if n % 20 == 0:
                print(f"    …{n}/{len(all_rows)}")
    except Exception as e:
        print(f"  ⚠ QBO unavailable ({type(e).__name__}) — billed/costs left blank")

    # 4a) ABSOLUTE RULE: an -FTW line with costs (or billing) has STARTED —
    # it is not backlog. Pull it out into its own flagged section.
    started = []
    still_backlog = []
    for r in backlog:
        c, b = r.get("costs"), r.get("billed")
        has_activity = (isinstance(c, (int, float)) and c > 0) or \
                       (isinstance(b, (int, float)) and b > 0)
        if has_activity:
            r["section"] = "started"
            bits = []
            if isinstance(c, (int, float)) and c > 0:
                bits.append(f"${c:,.0f} costs booked")
            if isinstance(b, (int, float)) and b > 0:
                bits.append(f"${b:,.0f} billed")
            r["action"] = ("⚠ NOT BACKLOG — work has started (" +
                           " · ".join(bits) + ") — get it on the schedule")
            started.append(r)
        else:
            still_backlog.append(r)
    backlog = still_backlog
    if started:
        print(f"  ⚠ FTW with costs/billing (NOT backlog): {len(started)}")
    print(f"  true FTW backlog (zero activity): {len(backlog)}")

    # 4b) a dropped job only matters if money is still on the table
    def _left(r):
        K, B = r.get("contract"), r.get("billed")
        if isinstance(K, (int, float)) and isinstance(B, (int, float)):
            return K - B
        return None
    kept = []
    for r in dropped:
        left = _left(r)
        costs = r.get("costs")
        unbilled_costs = (isinstance(r.get("billed"), (int, float))
                          and r["billed"] == 0
                          and isinstance(costs, (int, float)) and costs > 0)
        if (left is not None and left > 1.0) or unbilled_costs:
            r["action"] = ("⚠ DROPPED off the schedule — costs booked, NOTHING billed"
                           if unbilled_costs else
                           f"⚠ DROPPED off the schedule — ${left:,.0f} left to bill")
            kept.append(r)
    dropped = kept
    print(f"  dropped AND unbilled: {len(dropped)}")

    # 5) render --------------------------------------------------------
    def fill(ws, title):
        ws["A1"] = title
        ws["A1"].font = BOLD
        ws.append(HDR)
        for c in range(1, len(HDR) + 1):
            ws.cell(2, c).font = BOLD
            ws.cell(2, c).border = UNDER
            ws.cell(2, c).alignment = Alignment(horizontal="center",
                                                wrap_text=True)
        def emit(r):
            gp = None
            if isinstance(r["contract"], (int, float)) and r["contract"] and \
                    isinstance(r["etc"], (int, float)):
                gp = (r["contract"] - r["etc"]) / r["contract"]
            ws.append([r["line"], r["address"], r["builder"], r["contract"],
                       r["etc"], r["billed"], r["costs"], gp,
                       CHECK if r["on_sched"] else CROSS,
                       CHECK if r["in_gl"] else CROSS, r["action"],
                       r.get("co_amt"),
                       str(r["k_src"]) if r.get("k_src") else
                       (str(r["folder"]) if r["folder"] else ""),
                       str(r["e_src"]) if r.get("e_src") else
                       (str(r["folder"]) if r["folder"] else ""),
                       r.get("co_file") or "", "|"])
            i = ws.max_row
            money_font = GREEN if r["easy"] else BLACK
            link_font = LINKF_GREEN if r["easy"] else LINKF
            # NEVER a stored file:// hyperlink — Excel resolves those through
            # ScopedBookmarkAgent on load and beachballs on a network share
            # (the user 2026-07-28). The path goes in its own TEXT column.
            for c in (4, 5):
                ws.cell(i, c).number_format = CUR
                ws.cell(i, c).font = money_font
            if r.get("provisional"):
                ws.cell(i, 4).font = Font(italic=True, bold=r["easy"],
                                          color="006100" if r["easy"] else "000000")
            for c in (6, 7):
                ws.cell(i, c).number_format = CUR
                if r["qbo_url"] and ws.cell(i, c).value is not None:
                    ws.cell(i, c).hyperlink = r["qbo_url"]
                    ws.cell(i, c).font = LINKF
            ws.cell(i, 8).number_format = PCT
            ws.cell(i, 12).number_format = CUR
            if r.get("co_amt"):
                ws.cell(i, 12).font = Font(color="9C0006", bold=True)
            for c in (9, 10):
                ws.cell(i, c).alignment = Alignment(horizontal="center")
                ws.cell(i, c).font = (OKF if ws.cell(i, c).value == CHECK else NOF)

        for r in rows:
            emit(r)
        main_last = ws.max_row

        def band(text):
            ws.append([])
            ws.append([text])
            b = ws.max_row
            ws.cell(b, 1).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=b, start_column=1,
                           end_row=b, end_column=len(HDR))
            ws.cell(b, 1).border = UNDER

        if cp_rows:
            band(f"\u26a0 CP JOBS ON THE RP SCHEDULE ({len(cp_rows)}) \u2014 these belong in "
                 "the CP WIP; folder lives under AWARDED CP PROJECTS, not Residential")
            for r in cp_rows:
                emit(r)
        if dropped:
            band(f"\u26a0 DROPPED OFF THE SCHEDULE \u2014 STILL UNBILLED "
                 f"({len(dropped)}) \u2014 work left the board with money on the table")
            for r in dropped:
                emit(r)
        if started:
            band(f"\u26a0 FTW WITH COSTS \u2014 NOT BACKLOG ({len(started)}) \u2014 work has "
                 "started off-schedule; these belong on the schedule")
            for r in started:
                emit(r)
        if backlog:
            band(f"FTW BACKLOG \u2014 flatwork priced in the General Lista, not yet on "
                 f"the schedule ({len(backlog)}) \u2014 expected wins · zero costs, zero billing, excludes AF=OTHER")
            for r in backlog:
                emit(r)

        for n, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(n)].width = w
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HDR))}{max(2, main_last)}"

    title = (f"RP WIP — as of schedule {label}.  GREEN = one proposal + one "
             "takeoff in the folder (number is unambiguous).  Contract/ETC open "
             "the project FOLDER; Billed/Costs open the project in QBO.")
    easy_n = sum(1 for r in rows if r["easy"])
    print(f"  green (1 proposal + 1 takeoff): {easy_n}/{len(rows)}")

    if not args.commit:
        out = Path(args.out) if args.out else Path(
            os.getenv("RP_WIP_SIMPLE_XLSX",
                      str(Path.home() / "Downloads" / "RP WIP - Simple.xlsx")))
        lock = out.with_name("~$" + out.name)
        if lock.exists():
            raise SystemExit(f"{out.name} is open in Excel — close it first")
        wb = Workbook()
        fill(wb.active, title)
        wb.active.title = "RP WIP"
        wb.save(out)
        print(f"\n  ✓ PREVIEW → {out}")
        print("    re-run with --commit to write the 'Test - RP' tab")
        return 0

    # production write ------------------------------------------------
    GUARD.assert_write_allowed(TAB)
    path = CP.WIP_EXCEL_PATH
    lock = Path(path).with_name("~$" + Path(path).name)
    if lock.exists():
        raise SystemExit("The WIP master is open in Excel — close it first")
    wb = GUARD.open_wip_workbook_for_write(path)

    # HARVEST the user's cell comments before the rewrite, keyed by
    # (JOB #, column header) — the tab is rebuilt each run and comments would
    # otherwise be destroyed. Same contract as cp_wip_reader.write_test_cp.
    saved_comments = {}
    if TAB in wb.sheetnames:
        old = wb[TAB]
        labels = {c: str(old.cell(2, c).value).strip()
                  for c in range(1, old.max_column + 1)
                  if old.cell(2, c).value}
        for r in range(3, old.max_row + 1):
            job = old.cell(r, 1).value
            if not job:
                continue
            for c, label in labels.items():
                cm = old.cell(r, c).comment
                if cm is not None:
                    saved_comments[(str(job).strip(), label)] = (cm.text,
                                                                 cm.author)
        if saved_comments:
            print(f"  harvested {len(saved_comments)} user comment(s) to re-attach")
        del wb[TAB]

    ws = wb.create_sheet(TAB)
    fill(ws, title)

    # RE-ATTACH; anything that no longer has a row/column is reported loudly
    # rather than silently dropped.
    if saved_comments:
        from openpyxl.comments import Comment
        rowmap = {}
        for r in range(3, ws.max_row + 1):
            j = ws.cell(r, 1).value
            if j and not (isinstance(j, str) and
                          ("DROPPED" in j or "BACKLOG" in j)):
                rowmap.setdefault(str(j).strip(), r)
        colmap = {str(ws.cell(2, c).value).strip(): c
                  for c in range(1, len(HDR) + 1) if ws.cell(2, c).value}
        kept, orphans = 0, []
        for (job, label), (text, author) in saved_comments.items():
            r, c = rowmap.get(job), colmap.get(label)
            if r and c:
                ws.cell(r, c).comment = Comment(text, author or "")
                kept += 1
            else:
                orphans.append((job, label, text))
        print(f"  ✓ re-attached {kept} comment(s)")
        for job, label, text in orphans:
            print(f"  ⚠ COMMENT ORPHANED — {job} / {label}: "
                  f"{text[:70].replace(chr(10), ' ')}")

    wb.save(path)
    print(f"\n  ✓ Wrote {len(rows)} line(s) to '{TAB}' in {Path(path).name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
