#!/usr/bin/env python3
"""
concrete_cost_code_audit.py - audit that each supplier's bill lines carry the
RIGHT cost-code family for what they sell (the user 2026-08-25).

Cost codes live in the QBO Item name (SL1 / PV6 / CS1 …); the NUMBER is the cost
family (see shared/qbo_costs + the Cost Code Sheet): 1 Concrete (ready-mix) ·
2 Rebar · 3 Formwork/Lumber · 4 Aggregates · 5 Equip/51 Pump/52 Saw · 6 Labor ·
7 Specialty · 8 Fuel · 9 Supplies. Vendors fall into coding TYPES, each with a rule:

  • concrete supplier (ready-mix, e.g. Cowtown) → every line must be *1.
  • material supplier (e.g. RCI = lumber/rebar) → *2/*3/*4 only; NEVER *1
    (concrete), *5/*51/*52 (equipment), or *6 (labor).
  • both (e.g. Preferred Materials) → sells concrete AND material, so a line whose
    MEMO reads as concrete yardage / ready-mix MUST be *1, never another code.

The type is CAPTURED from the data (the *1 vs *2-4 split), because a correctly-run
vendor's coding makes it obvious. Tune with --threshold / --min-lines; force or
correct a type with an override JSON (kept OUTSIDE the repo).

AUTH   one keychain read via shared.qbo_api.load_credentials(). READ-ONLY on QBO.
OUTPUT OneDrive `Works In Progress/QBO Audits/Concrete Cost Code Audit.xlsx`
         • "Vendors"        - every captured vendor, its type + coding pattern
         • "Miscoded Lines" - every line that breaks its vendor's type rule
       Plain formatting (repo Excel rule); assert_clean is the LAST step.

Override JSON (--override, default <companyhealth>/concrete_suppliers.json):
    {"concrete": ["COWTOWN", …], "material": ["RCI", …],
     "both": ["PREFERRED MATERIALS", …], "exclude": ["SOME VENDOR", …]}
  Names match case-insensitively; an explicit type wins over auto-detection.
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Repo root on sys.path for shared/ (the one allowed path hack).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from shared import qbo_api  # noqa: E402
from shared import paths  # noqa: E402
from shared.xlsx_verify import assert_clean  # noqa: E402
# THE audit logic lives in shared/ so the bill-tracker's Audit - Cost Code sheet
# uses the exact same rules (repo rule: shared, not tool-to-tool imports).
from shared.cost_code_audit import (  # noqa: E402
    classify_vendors, flag_lines, load_override, code_families,
    TYPE_LABEL, TYPE_ORDER)

_PROJ_RE = re.compile(r"\b(?:MFD|CP|RP)\d+(?:-FTW)?\b", re.IGNORECASE)
QBO_BILL_URL = "https://qbo.intuit.com/app/bill?txnId={bill_id}"

DEFAULT_OUT = (paths.onedrive_base() / "Works In Progress" / "QBO Audits"
               / "Concrete Cost Code Audit.xlsx")
DEFAULT_OVERRIDE = paths.companyhealth_dir() / "concrete_suppliers.json"

DATE_FMT = "m/d/yyyy"
MONEY_FMT = '"$"#,##0.00'


# ─────────────────────── QBO pull ───────────────────────

def _leaf(name: str) -> str:
    """Cost code = leaf of a possibly-hierarchical item name ('Parent:SL1' → 'SL1')."""
    return (name or "").split(":")[-1].strip()


def _line_cost_code(line: dict) -> str:
    if line.get("DetailType") == "ItemBasedExpenseLineDetail":
        ref = (line.get("ItemBasedExpenseLineDetail") or {}).get("ItemRef") or {}
        return _leaf(ref.get("name", ""))
    return ""


def _line_account(line: dict, account_map: Dict[str, str]) -> str:
    if line.get("DetailType") == "AccountBasedExpenseLineDetail":
        ref = (line.get("AccountBasedExpenseLineDetail") or {}).get("AccountRef") or {}
        return account_map.get(ref.get("value", ""), ref.get("name", "")) or ""
    return ""


def _line_project(line: dict) -> str:
    det = line.get(line.get("DetailType", "")) or {}
    nm = (det.get("CustomerRef") or {}).get("name", "")
    m = _PROJ_RE.search(nm) if nm else None
    return m.group(0).upper() if m else ""


def pull_lines(access: str, cid: str, since: str) -> Tuple[List[dict], int]:
    """Every bill line since `since` → flat records with vendor + resolved cost code."""
    vendors = qbo_api.query_all(access, cid, "Vendor")
    vmap = {v["Id"]: (v.get("DisplayName") or v.get("CompanyName") or f"Vendor {v['Id']}")
            for v in vendors}
    accounts = qbo_api.query_all(access, cid, "Account")
    amap = {a["Id"]: (a.get("Name") or "") for a in accounts}

    bills = qbo_api.query_all(access, cid, "Bill", f"TxnDate >= '{since}'")
    rows: List[dict] = []
    for b in bills:
        vref = b.get("VendorRef") or {}
        vendor = vmap.get(vref.get("value", ""), vref.get("name", "?"))
        for ln in (b.get("Line") or []):
            if ln.get("DetailType") not in (
                    "ItemBasedExpenseLineDetail", "AccountBasedExpenseLineDetail"):
                continue
            code = _line_cost_code(ln)
            number, cost_name = code_families(code)
            rows.append({
                "vendor": vendor,
                "bill_id": b.get("Id", ""),
                "bill_doc": b.get("DocNumber", "") or "",
                "date": b.get("TxnDate", ""),
                "project": _line_project(ln),
                "cost_code": code,
                "number": number,
                "cost_name": cost_name,
                "account": _line_account(ln, amap),
                "amount": float(ln.get("Amount") or 0),
                "desc": (ln.get("Description") or "").strip(),
            })
    return rows, len(bills)


# ─────────────────────── Excel (plain) ───────────────────────

def _write_sheet(ws, headers: List[str], kinds: List[str], data: List[list],
                 widths: List[int]) -> None:
    """Plain sheet: bold header, freeze, AutoFilter, number formats. No fills
    (repo Excel rule)."""
    bold = Font(bold=True)
    left = Alignment(horizontal="left")
    for c, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 2
    for vals in data:
        for c, (val, kind) in enumerate(zip(vals, kinds), 1):
            cell = ws.cell(row=r, column=c, value=val)
            if kind == "money":
                cell.number_format = MONEY_FMT
            elif kind == "date":
                cell.number_format = DATE_FMT
            elif kind == "pct":
                cell.number_format = "0%"
            else:
                cell.alignment = left
        r += 1
    last = max(r - 1, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"


def _parse_date(s: str) -> Optional[dt.date]:
    try:
        return dt.date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


def build_workbook(agg: Dict[str, dict], vtype: Dict[str, str],
                   flags: List[dict], out: Path, meta: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 - captured vendors + their coding pattern; most-miscoded on top.
    ws1 = wb.create_sheet("Vendors")
    vend_rows = []
    for up, a in agg.items():
        t = vtype.get(up)
        if not t:
            continue
        vend_rows.append([
            a["vendor"], TYPE_LABEL[t], t, a["coded"],
            round(a.get("pct_c", 0.0), 4), round(a.get("pct_m", 0.0), 4),
            a["concrete"], a["material"], a["other"], a["nocode"],
            round(a["amount"], 2),
        ])
    # by type (concrete→material→both→review), then most off-pattern lines on top.
    vend_rows.sort(key=lambda v: (TYPE_ORDER.get(v[2], 3), -(v[8] + v[9])))
    vend_rows = [v[:2] + v[3:] for v in vend_rows]   # drop the sort-only type key
    _write_sheet(
        ws1,
        ["Vendor", "Type", "Coded Lines", "Concrete %", "Material %",
         "Concrete (1)", "Material (2-4)", "Other coded", "No-Code", "Total $"],
        ["text", "text", "text", "pct", "pct", "text", "text", "text", "text", "money"],
        vend_rows,
        [30, 16, 12, 11, 11, 12, 13, 12, 10, 15],
    )

    # Sheet 2 - the miscoded lines.
    ws2 = wb.create_sheet("Miscoded Lines")
    line_rows = []
    for f in flags:
        line_rows.append([
            f["vendor"], TYPE_LABEL.get(f["vtype"], ""), f["bill_doc"],
            _parse_date(f["date"]), f["project"], f["cost_code"],
            f["cost_name"] or "", round(f["amount"], 2), f["desc"],
            f["reason"], f["bill_id"],
        ])
    _write_sheet(
        ws2,
        ["Vendor", "Type", "Bill #", "Bill Date", "Project", "Cost Code",
         "Cost Name", "Amount", "Line Description", "Reason", "Open"],
        ["text", "text", "text", "date", "text", "text", "text", "money",
         "text", "text", "link"],
        line_rows,
        [26, 15, 12, 11, 12, 11, 18, 13, 32, 44, 6],
    )
    # Hyperlink the Open column → QBO bill.
    open_col = 11
    for i, f in enumerate(flags, start=2):
        if f["bill_id"]:
            cell = ws2.cell(row=i, column=open_col, value="↗")
            cell.hyperlink = QBO_BILL_URL.format(bill_id=f["bill_id"])
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")
    if not flags:
        ws2.cell(row=2, column=1,
                 value="✓ none found - every captured vendor codes to its type's rule")

    # Summary sheet (plain, label + value same row).
    ws0 = wb.create_sheet("Summary", 0)
    counts = {t: sum(1 for v in vtype.values() if v == t)
              for t in ("concrete", "material", "both", "hauler", "review")}
    summary = [
        ("Cost Code Audit - concrete / material / both / hauler", ""),
        ("Generated", meta["generated"]),
        ("Bills scanned (since " + meta["since"] + ")", meta["bills"]),
        ("Bill lines scanned", meta["lines"]),
        ("Concrete suppliers (→ all *1)", counts["concrete"]),
        ("Material suppliers (→ *2/*3/*4, no *1/*5/*6)", counts["material"]),
        ("Both suppliers (yardage memo → *1)", counts["both"]),
        ("Hauler vendors (haul-off *5 OK)", counts["hauler"]),
        ("Vendors to review (borderline)", counts["review"]),
        ("Miscoded lines flagged", len(flags)),
        ("Threshold / min coded lines", f'{meta["threshold"]:.0%} / {meta["min_lines"]}'),
        ("Override file", meta["override"]),
    ]
    for r, (label, val) in enumerate(summary, 1):
        c1 = ws0.cell(row=r, column=1, value=label)
        if r == 1:
            c1.font = Font(bold=True, size=13)
        else:
            c1.font = Font(bold=True)
        ws0.cell(row=r, column=2, value=val)
    ws0.column_dimensions["A"].width = 34
    ws0.column_dimensions["B"].width = 40

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    assert_clean(out)   # LAST step - never hand over a file that trips Excel repair


# ─────────────────────── main ───────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--since", default=None,
                    help="only bills on/after this date (default: 365 days ago)")
    ap.add_argument("--threshold", type=float, default=0.60,
                    help="min family share of coded lines to auto-type a vendor")
    ap.add_argument("--min-lines", type=int, default=3,
                    help="min coded lines before a vendor can auto-type")
    ap.add_argument("--review-floor", type=float, default=0.25,
                    help="family share above which an untyped vendor is surfaced for review")
    ap.add_argument("--override", type=Path, default=DEFAULT_OVERRIDE,
                    help="JSON: {concrete/material/both/exclude: [vendor names]}")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args()

    since = args.since or (dt.date.today() - dt.timedelta(days=365)).isoformat()

    print("→ authenticating to QBO …")
    access, cid = qbo_api.load_credentials()
    print("  ok.")

    print(f"→ pulling bills since {since} …")
    rows, n_bills = pull_lines(access, cid, since)
    print(f"  {n_bills} bills · {len(rows)} lines")

    override = load_override(args.override)
    n_over = sum(len(v) for v in override.values())
    if n_over:
        print("  override: " + " / ".join(f"{len(override[k])} {k}"
              for k in ("concrete", "material", "both", "hauler", "exclude")))

    agg, vtype = classify_vendors(rows, args.threshold, args.min_lines,
                                  args.review_floor, override)
    flags = flag_lines(rows, vtype)
    counts = {t: sum(1 for v in vtype.values() if v == t)
              for t in ("concrete", "material", "both", "hauler", "review")}
    print(f"  vendors: {counts['concrete']} concrete · {counts['material']} material · "
          f"{counts['both']} both · {counts['hauler']} hauler · {counts['review']} review"
          f"  →  {len(flags)} miscoded lines")

    meta = {
        "generated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "since": since, "bills": n_bills, "lines": len(rows),
        "threshold": args.threshold, "min_lines": args.min_lines,
        "override": str(args.override) if n_over else "(none)",
    }
    build_workbook(agg, vtype, flags, args.out, meta)
    print(f"  ✓ saved {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
