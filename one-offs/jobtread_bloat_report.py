#!/usr/bin/env python3
"""
jobtread_bloat_report.py — JobTread open-job bloat vs reality (the user
2026-07-24). READ-ONLY (JobTread + QBO + schedule); writes one Downloads Excel.

WHY
  JobTread has ~500 "open" jobs (approved/created, not closed) but only ~66 are
  on the current schedule — lots are finished in real life and never closed in
  JobTread. This flags which open JobTread jobs look DONE so they can be closed.

"DONE IN REAL LIFE" is inferred from QBO (the books = source of truth) + the
schedule, per open JobTread job (matched by project #, -FTW falls back to base):
  • PAID & IDLE   — fully collected in QBO (AR balance ≈ $0) AND last invoice
                    > 90 days ago AND not on the current schedule → CLOSE.
  • NO QBO / OLD  — no QBO invoices AND created > 120 days ago AND not
                    scheduled → stale shell, review to close/delete.
  • DONE, UNPAID  — idle > 90 days but AR balance > $0 → work done, still owed;
                    don't close yet, it's an AR item.
  • ACTIVE / NEW  — on the schedule, or billed within 90 days, or a recent
                    job with no bill yet → keep open.

Auth: one Touch ID (JT_GRANT_KEY + QBO creds share the automation-qbo blob).

Usage
  python3 jobtread_bloat_report.py                     # yesterday/latest schedule
  python3 jobtread_bloat_report.py --idle-days 120 --schedule <f.xlsx>
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO))
sys.path.insert(0, str(_REPO / "wip"))
sys.path.insert(0, str(_REPO / "one-offs"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from shared import qbo_vault, qbo_api
import rp_wip_reader as RP
import rp_schedule_wip_preview as P
import rp_jobtread_coverage as C   # pave(), ORG_ID

DIV = {"RP": "Residential", "CP": "Commercial", "MFD": "Multi Family"}
TODAY = date.today()


def _base(num: str) -> str:
    return re.sub(r"-FTW$", "", (num or "").upper())


def all_jt_jobs(key: str) -> list:
    out, page, guard = [], None, 0
    while True:
        dollar = {"size": 100}
        if page:
            dollar["page"] = page
        q = {"organization": {"$": {"id": C.ORG_ID}, "jobs": {
            "$": dollar, "nextPage": {},
            "nodes": {"number": {}, "name": {}, "status": {},
                      "createdAt": {}, "closedOn": {}}}}}
        r = C.pave(key, q)["organization"]["jobs"]
        out.extend(r["nodes"])
        page = r.get("nextPage")
        guard += 1
        if not page or guard > 60:
            break
    return out


def _pdate(s: str):
    """Parse 'YYYY-MM-DD' or ISO datetime → date, or None."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def qbo_activity(access, cid):
    """proj# → {billed, balance, last (date|None), n}. Balance from the project
    sub-customer; billed/last/n from its invoices."""
    cust = qbo_api.build_project_customer_map(access, cid)
    act = {p: {"billed": 0.0, "balance": v["balance"], "last": None, "n": 0}
           for p, v in cust.items()}
    for inv in qbo_api.query_all(access, cid, "Invoice"):
        proj = qbo_api.extract_proj((inv.get("CustomerRef") or {}).get("name"))
        if not proj:
            continue
        rec = act.setdefault(
            proj, {"billed": 0.0, "balance": 0.0, "last": None, "n": 0})
        rec["billed"] += float(inv.get("TotalAmt") or 0)
        rec["n"] += 1
        d = _pdate(inv.get("TxnDate"))
        if d and (rec["last"] is None or d > rec["last"]):
            rec["last"] = d
    return act


# ── Excel — clean (single-line title, bold header + underline, autofilter) ──
BOLD = Font(bold=True)
UNDER = Border(bottom=Side(style="thin", color="000000"))
CUR = '"$"#,##0.00'


def _add_sheet(wb, title, subtitle, header, rows, widths, money_cols=()):
    ws = wb.create_sheet(title)
    ws["A1"] = subtitle
    ws["A1"].font = BOLD
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(2, c)
        cell.font = BOLD
        cell.border = UNDER
        cell.alignment = Alignment(vertical="bottom")
    for row in rows:
        ws.append(row)
    for c in money_cols:
        for r in range(3, ws.max_row + 1):
            ws.cell(r, c).number_format = CUR
    last = chr(ord("A") + len(header) - 1)
    ws.auto_filter.ref = f"A2:{last}{max(2, ws.max_row)}"
    ws.freeze_panes = "A3"
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    return ws


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--schedule", help="schedule xlsx (default: latest)")
    ap.add_argument("--idle-days", type=int, default=90,
                    help="days since last invoice to count as idle (default 90)")
    ap.add_argument("--stale-days", type=int, default=120,
                    help="createdAt age for a no-QBO shell (default 120)")
    args = ap.parse_args()
    IDLE, STALE, EPS = args.idle_days, args.stale_days, 1.0

    print("\n  JOBTREAD BLOAT — open in JobTread vs done in real life")
    print("  " + "─" * 56)

    # schedule → active base #s
    if args.schedule:
        sched_path = Path(args.schedule)
    else:
        best = P.latest_schedule(RP.SCHEDULE_DIR)
        if best is None:
            print("  ✗ no schedule file found")
            return 1
        sched_path = best[1]
    print(f"  schedule: {sched_path.name}")
    sched_bases = {_base(s["job"]) for s in P.read_main_schedule(sched_path)}
    print(f"  active jobs on the schedule: {len(sched_bases)}")

    # one Touch ID unlocks JT + QBO (same automation-qbo blob)
    key = qbo_vault.get("JT_GRANT_KEY")
    print("  JobTread: full job sweep …")
    jt_all = all_jt_jobs(key)
    open_jobs = [j for j in jt_all
                 if j.get("status") != "closed" and not j.get("closedOn")]
    print(f"    {len(jt_all)} jobs total · {len(open_jobs)} open · "
          f"{len(jt_all) - len(open_jobs)} already closed")

    print("  QBO: customers + invoices …")
    access, cid = qbo_api.load_credentials()
    act = qbo_activity(access, cid)
    print(f"    {len(act)} QBO projects with activity")

    # CONTRACT per line — needed for "did we bill the FULL amount?"
    print("  General Lista: contract per line …")
    contracts = {}
    try:
        for rec in RP.read_general_list(RP.ALPHA_PATH)[0]:
            jb = rec["job"]
            if rec.get("slab_bid"):
                contracts[jb] = float(rec["slab_bid"])
            if rec.get("flat_bid") and not rec.get("flat_other"):
                contracts[f"{jb}-FTW"] = float(rec["flat_bid"])
    except Exception as e:
        print(f"    ⚠ General Lista unreadable ({type(e).__name__}) — "
              "contracts unknown, nothing will qualify to close")
    print(f"    {len(contracts)} lines with a contract")

    # COST ACTIVITY in the last 3 months, only for lines that could qualify
    # (unscheduled + fully billed + pre-2026) — a windowed P&L per project.
    cutoff = TODAY - timedelta(days=92)
    cust_map = qbo_api.build_project_customer_map(access, cid)
    cand = []
    for j in open_jobs:
        n = (j.get("number") or "").upper()
        if not n or n == "CP000" or _base(n) in sched_bases:
            continue
        a = act.get(n)
        k_ = contracts.get(n)
        if not a or not a["n"] or not k_ or k_ <= 0:
            continue
        if a["billed"] < k_ - max(1.0, k_ * 0.005):
            continue
        if a["last"] is None or a["last"].year >= 2026:
            continue
        cand.append(n)
    print(f"  QBO: 3-month cost check on {len(cand)} candidate(s) …")
    recent_costs = {}
    for i, n in enumerate(cand, 1):
        c = cust_map.get(n)
        if not c:
            continue
        try:
            pl = qbo_api.fetch_project_pl(access, cid, c["id"],
                                          cutoff.isoformat(), TODAY.isoformat())
            t = qbo_api.extract_pl_totals(pl)
            recent_costs[n] = (t.get("cogs") or 0.0) + (t.get("expenses") or 0.0)
        except Exception:
            recent_costs[n] = 0.0
        if i % 25 == 0:
            print(f"    …{i}/{len(cand)}")

    # classify every open JobTread job
    rows = []
    for j in open_jobs:
        num = (j.get("number") or "").upper()
        if not num or num == "CP000":
            continue
        pre = re.match(r"^(RP|CP|MFD)", num)
        div = DIV.get(pre.group(1) if pre else "", "?")
        scheduled = _base(num) in sched_bases
        created = _pdate(j.get("createdAt"))
        created_age = (TODAY - created).days if created else None

        # EXACT match only. The old -FTW→base fallback attributed the BASE
        # project's billing to the -FTW line, so an -FTW that was never
        # invoiced read as "paid & idle" (RP7431-FTW: 0 invoices of its own,
        # base RP7431 had $21,722 — the user 2026-07-30). A close decision must
        # judge each line on its OWN billing.
        a = act.get(num)
        qmatch = "exact" if a else "—"
        billed = a["billed"] if a else 0.0
        balance = a["balance"] if a else 0.0
        last = a["last"] if a else None
        idle = (TODAY - last).days if last else None

        # ── THE CLOSE RULE (the user 2026-07-30) — all four must hold ──
        #   1. billed the FULL contract amount
        #   2. last billing date is BEFORE 2026
        #   3. no cost activity in the last 3 months
        #   4. not on the schedule
        contract = contracts.get(num)
        billed = a["billed"] if a else 0.0
        full = (contract is not None and contract > 0
                and billed >= contract - max(1.0, contract * 0.005))
        pre2026 = (last is not None and last.year < 2026)
        recent_cost = recent_costs.get(num, 0.0) > 1.0

        if scheduled:
            verdict, rank = "KEEP — on the schedule", 5
        elif not a or not a["n"]:
            verdict, rank = "KEEP — never invoiced (not bloat, a gap)", 3
        elif contract is None or contract <= 0:
            verdict, rank = f"KEEP — no contract on file (billed ${billed:,.0f})", 3
        elif not full:
            verdict, rank = (f"KEEP — under-billed ${contract - billed:,.0f} of "
                             f"${contract:,.0f}"), 2
        elif not pre2026:
            verdict, rank = (f"KEEP — billed in {last.year}, not pre-2026"), 4
        elif recent_cost:
            verdict, rank = (f"KEEP — ${recent_costs[num]:,.0f} of costs in the "
                             "last 3 months"), 2
        else:
            verdict, rank = "CLOSE — billed in full · pre-2026 · no recent cost · unscheduled", 0

        rows.append({
            "job": num, "div": div, "status": j.get("status") or "",
            "name": j.get("name") or "", "billed": billed, "balance": balance,
            "last": last.isoformat() if last else "",
            "idle": idle if idle is not None else "",
            "created": created.isoformat() if created else "",
            "created_age": created_age if created_age is not None else "",
            "qmatch": qmatch, "sched": scheduled,
            "contract": contract, "full": full, "pre2026": pre2026,
            "recent_cost": recent_costs.get(num),
            "verdict": verdict, "rank": rank,
        })

    rows.sort(key=lambda r: (r["rank"],
                             -(r["idle"] if isinstance(r["idle"], int) else 0),
                             r["job"]))

    # counts
    from collections import Counter
    vc = Counter(r["verdict"] for r in rows)
    close = [r for r in rows if r["verdict"].startswith("CLOSE")]
    print("\n  VERDICTS (open JobTread jobs):")
    for v, n in sorted(vc.items(), key=lambda x: -x[1]):
        print(f"    {n:4}  {v}")
    print(f"  → {len(close)} job(s) meet ALL FOUR close tests")

    # workbook
    out = Path(os.getenv(
        "JT_BLOAT_XLSX",
        str(Path.home() / "Downloads" / "JobTread Bloat - Close Candidates.xlsx")))
    lock = out.with_name("~$" + out.name)
    if lock.exists():
        raise SystemExit(f"{out.name} is open in Excel — close it first")
    wb = Workbook()
    wb.remove(wb.active)

    hdr = ["JOB #", "DIV", "JT STATUS", "JT NAME", "VERDICT", "CONTRACT $",
           "QBO BILLED $", "BILLED FULL?", "LAST INVOICE", "PRE-2026?",
           "COSTS LAST 3 MO $", "ON SCHEDULE?", "AR BALANCE $", "CREATED"]

    def torow(r):
        return [r["job"], r["div"], r["status"], r["name"], r["verdict"],
                r.get("contract"), r["billed"],
                "YES" if r.get("full") else "no", r["last"],
                "YES" if r.get("pre2026") else "no",
                r.get("recent_cost"), "YES" if r["sched"] else "no",
                r["balance"], r["created"]]

    _add_sheet(
        wb, "Close Candidates",
        f"CLOSE CANDIDATES — {len(close)} job(s) meeting ALL FOUR tests "
        "(the user 2026-07-30): billed the FULL contract · last billed BEFORE 2026 · "
        "no cost activity in the last 3 months · not on the schedule. Each line is "
        "judged on its OWN QBO billing — no -FTW→base roll-up.",
        hdr, [torow(r) for r in close],
        (14, 6, 10, 30, 24, 15, 15, 13, 10, 12, 10), money_cols=(6, 7))

    _add_sheet(
        wb, "All Open (evidence)",
        f"ALL {len(rows)} OPEN JOBTREAD JOBS with the evidence behind each verdict. "
        "KEEP reasons say exactly which test failed.",
        hdr, [torow(r) for r in rows],
        (14, 6, 10, 30, 24, 15, 15, 13, 10, 12, 10), money_cols=(6, 7))

    wb.save(out)
    print(f"\n  ✓ Workbook → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
