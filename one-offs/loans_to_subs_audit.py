#!/usr/bin/env python3
"""
loans_to_subs_audit.py — Zero out the PARENT "Loans to Sub-Contractors"
(Other Current Asset) by moving every line posted directly to the parent
into the correct per-sub sub-account.

WHY
  The parent account is a header only — every dollar should live in a
  per-sub sub-account. Anything posted to the PARENT itself (bills, checks,
  vendor credits) is mis-coded and needs to move to the right sub-account.

TWO STEPS  (mirrors one-offs/qbo_recode_review.py)
  1. EXPORT (read-only — GET only, never writes to QBO)
       python3 one-offs/loans_to_subs_audit.py
     - One Touch ID via shared.qbo_api.load_credentials().
     - Finds the parent account by name (--account) and its sub-accounts
       (the recode targets).
     - Pulls the QBO GeneralLedger filtered to the PARENT id and keeps only
       lines posted to the parent itself (not to a sub-account).
     - Fetches each of those transactions once to capture its SyncToken and
       the id of the specific line posting to the parent (needed to apply a
       precise, reversible change later). Still read-only.
     - Writes a plain Excel to OneDrive "Works In Progress / QBO Audits":
       one row per parent line, a best-guess sub-account, a "Confirm
       Sub-Account" dropdown you fill in, and locked identity columns
       (Entity / Txn Id / Line Id / SyncToken) that drive the apply step.

  2. YOU review the workbook in Excel — set "Confirm Sub-Account" on every
     line you want moved (leave blank to skip). Sub-accounts that don't
     exist yet must be created in QBO first, then re-export so they appear.

  3. APPLY (gated — dry-run by default)
       python3 one-offs/loans_to_subs_audit.py --apply            # validate, no writes
       python3 one-offs/loans_to_subs_audit.py --apply --commit    # write approved lines
     For each confirmed line: re-resolve the sub name → id, re-fetch the txn,
     and (only its parent-posted line) set AccountRef to the target sub. Every
     other field is left untouched. Guards below.

SAFETY (non-negotiable)
  * EXPORT and "--apply without --commit" never write to QBO.
  * Only lines with a non-blank "Confirm Sub-Account" are ever touched.
  * Confirmed name must match a real sub-account name exactly, or skip+report.
  * SyncToken captured at export; if the txn changed since, skip+report.
  * Closed-period txns (<= QBO book-close date) skipped unless --include-closed.
  * The line is only moved if it still posts to the PARENT (else skip+report).
  * Per-txn JSON backups + a results CSV -> ~/Library/Logs/Proficient/.

DEPS: pip3 install --break-system-packages requests openpyxl
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import difflib
import json
import random
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

# Repo root on sys.path for shared/ (the one allowed path hack).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import requests  # noqa: E402

from shared import qbo_api  # noqa: E402
from shared import paths  # noqa: E402

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("pip3 install --break-system-packages openpyxl")

DEFAULT_ACCOUNT = "Loans to Sub-Contractors"
LOG_DIR = Path.home() / "Library" / "Logs" / "Proficient"
LOG_DIR.mkdir(parents=True, exist_ok=True)

# QBO GL "Transaction Type" string -> the API entity that owns it. Only these
# three carry AccountBasedExpenseLineDetail lines we can reclass; anything else
# (Journal Entry, Deposit, Bill Payment, Transfer, …) is left for a human.
TYPE_TO_ENTITY = {
    "bill": "Bill",
    "check": "Purchase",
    "expense": "Purchase",
    "cash expense": "Purchase",
    "credit card expense": "Purchase",
    "credit card charge": "Purchase",
    "credit card credit": "Purchase",
    "vendor credit": "VendorCredit",
    "supplier credit": "VendorCredit",
}
WRITABLE_ENTITIES = {"Bill", "Purchase", "VendorCredit"}

# Stopwords stripped before fuzzy-matching a vendor to a sub-account name.
_STOP = {"llc", "inc", "co", "company", "construction", "concrete", "the",
         "and", "&", "services", "service", "loan", "loans", "sub", "subs",
         "contractor", "contractors", "sub-contractor", "sub-contractors"}


def _norm(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    toks = [t for t in s.split() if t and t not in _STOP]
    return " ".join(toks)


def _norm_id(v) -> str:
    """Excel may read a digit-string id as a float — normalize back to a
    clean string ('12345.0' -> '12345', 3.0 -> '3')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _as_date(s):
    try:
        return dt.datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


# ───────────────────────── QBO entity get/post (writes) ─────────────────────
# shared/qbo_api gives us auth + query_all + report (reads). It has no
# single-entity fetch or POST, which is exactly what the reclass needs, so
# those two verbs live here — self-contained, same retry shape as _api_get.

def _qbo_request(method, path, access, cid, params=None, json_body=None):
    p = dict(params or {})
    p["minorversion"] = qbo_api.MINOR_VERSION
    url = f"{qbo_api.API_BASE}/v3/company/{cid}/{path}"
    for attempt in range(8):
        last = attempt == 7
        try:
            r = requests.request(
                method, url,
                headers={"Authorization": f"Bearer {access}",
                         "Accept": "application/json",
                         "Content-Type": "application/json"},
                params=p, json=json_body, timeout=120,
            )
        except (requests.exceptions.ReadTimeout,
                requests.exceptions.ConnectTimeout,
                requests.exceptions.ConnectionError) as e:
            if last:
                raise RuntimeError(f"{path} → network error: {e}")
            time.sleep(min(2 ** attempt, 30) + random.uniform(0, 1.0))
            continue
        if r.status_code == 200:
            return r.json()
        if r.status_code in (429, 500, 502, 503, 504) and not last:
            time.sleep(min(2 ** attempt, 30) + random.uniform(0, 1.0))
            continue
        raise RuntimeError(f"{path} → {r.status_code}: {r.text[:300]}")
    raise RuntimeError(f"{path} → unreachable")


def get_entity(entity, tid, access, cid):
    return _qbo_request("GET", f"{entity.lower()}/{tid}", access, cid).get(entity, {})


def update_entity(entity, body, access, cid):
    return _qbo_request("POST", entity.lower(), access, cid, json_body=body).get(entity, {})


def book_close_date(access, cid):
    prefs = _qbo_request("GET", "preferences", access, cid).get("Preferences", {})
    return prefs.get("AccountingInfoPrefs", {}).get("BookCloseDate")


# ───────────────────────── account lookup ─────────────────────────

def find_parent_and_subs(access, cid, account_name):
    """Return (parent_dict, [sub_dicts], by_id). Exact name first, then contains."""
    accts = qbo_api.query_all(access, cid, "Account")
    by_id = {a["Id"]: a for a in accts}
    want = account_name.strip().lower()

    parent = next((a for a in accts
                   if (a.get("Name") or "").strip().lower() == want), None)
    if not parent:
        cands = [a for a in accts
                 if want[:12] in (a.get("Name") or "").strip().lower()]
        if len(cands) == 1:
            parent = cands[0]
        elif cands:
            print("  Multiple accounts match — pass the exact name with "
                  "--account:")
            for a in cands:
                print(f"    - {a.get('Name')}  (id {a['Id']}, "
                      f"{a.get('AccountType')})")
            sys.exit(1)
    if not parent:
        print(f"✗  No account named like '{account_name}'. Some asset "
              f"accounts I can see:")
        for a in accts:
            if "asset" in (a.get("AccountType") or "").lower():
                print(f"    - {a.get('Name')}  (id {a['Id']})")
        sys.exit(1)

    pid = parent["Id"]
    subs = [a for a in accts
            if (a.get("ParentRef") or {}).get("value") == pid]
    for s in subs:
        s["_fq"] = s.get("FullyQualifiedName") or s.get("Name")
    return parent, subs, by_id


# ───────────────────────── GL report parsing ─────────────────────────

def _col_titles(report_data):
    """Map GL report column index -> lowercased title."""
    cols = (report_data.get("Columns") or {}).get("Column") or []
    return [(c.get("ColTitle") or c.get("ColType") or "").lower() for c in cols]


def _cell(coldata, idx):
    if idx is None or idx >= len(coldata):
        return ""
    return (coldata[idx] or {}).get("value", "") or ""


def _cell_id(coldata, idx):
    if idx is None or idx >= len(coldata):
        return ""
    return (coldata[idx] or {}).get("id", "") or ""


def _num(v):
    v = (v or "").replace(",", "").replace("$", "").strip()
    if not v:
        return 0.0
    neg = v.startswith("(") and v.endswith(")")
    v = v.strip("()")
    try:
        n = float(v)
    except ValueError:
        return 0.0
    return -n if neg else n


def parse_gl(report_data, parent_name, sub_names):
    """
    Walk the GeneralLedger report. Return the list of data rows that belong
    to the PARENT account section only (skip sub-account sections).
    """
    titles = _col_titles(report_data)

    def idx(*names):
        for n in names:
            for i, t in enumerate(titles):
                if n in t:
                    return i
        return None

    i_date = idx("date")
    i_type = idx("transaction type", "type")
    i_doc = idx("num", "doc")
    i_name = idx("name")
    i_memo = idx("memo", "description")
    i_split = idx("split", "account")   # offsetting account
    i_amt = idx("amount")
    i_debit = idx("debit")
    i_credit = idx("credit")
    i_bal = idx("balance")

    sub_lc = {s.strip().lower() for s in sub_names}
    parent_lc = parent_name.strip().lower()
    rows_out = []
    seen_sections = []

    def section_account_name(node):
        h = (node.get("Header") or {}).get("ColData") or []
        return (h[0].get("value") if h else "") or ""

    def walk(node, in_parent):
        rtype = node.get("type")
        if rtype == "Section":
            acct = section_account_name(node).strip()
            if acct:
                seen_sections.append(acct)
            acct_lc = acct.lower()
            this_in_parent = in_parent
            if acct_lc:
                this_in_parent = (acct_lc == parent_lc) and (acct_lc not in sub_lc)
            for child in (node.get("Rows") or {}).get("Row") or []:
                walk(child, this_in_parent)
        elif rtype == "Data":
            if not in_parent:
                return
            cd = node.get("ColData") or []
            date = _cell(cd, i_date)
            if not date:  # subtotal/blank line
                return
            debit = _num(_cell(cd, i_debit)) if i_debit is not None else 0.0
            credit = _num(_cell(cd, i_credit)) if i_credit is not None else 0.0
            amt = _num(_cell(cd, i_amt)) if i_amt is not None else (debit - credit)
            rows_out.append({
                "date": date,
                "type": _cell(cd, i_type),
                "doc": _cell(cd, i_doc),
                "name": _cell(cd, i_name),
                "memo": _cell(cd, i_memo),
                "split": _cell(cd, i_split),
                "debit": debit,
                "credit": credit,
                "amount": amt,
                "balance": _num(_cell(cd, i_bal)) if i_bal is not None else 0.0,
                "txn_id": _cell_id(cd, i_type) or _cell_id(cd, i_date),
            })
        else:
            for child in (node.get("Rows") or {}).get("Row") or []:
                walk(child, in_parent)

    for n in (report_data.get("Rows") or {}).get("Row") or []:
        walk(n, False)
    return rows_out, titles, seen_sections


# ───────────────────────── enrich: SyncToken + line id ─────────────────────

def parent_lines_of(txn, parent_id):
    """AccountBasedExpenseLineDetail lines whose AccountRef == the parent id."""
    out = []
    for line in txn.get("Line", []) or []:
        det = line.get("AccountBasedExpenseLineDetail")
        if det and str((det.get("AccountRef") or {}).get("value")) == str(parent_id):
            out.append({"line_id": str(line.get("Id", "")),
                        "amount": float(line.get("Amount") or 0)})
    return out


def enrich_rows(rows, parent_id, access, cid):
    """Fetch each unique writable txn once (read-only GET) and stamp each row
    with entity, sync_token, and the line_id of its parent posting — the
    identity the apply step needs. Non-writable / unfetchable rows get a note."""
    for x in rows:
        x["note"] = ""
        x["entity"] = TYPE_TO_ENTITY.get((x["type"] or "").strip().lower(), "")
        x["sync_token"] = ""
        x["line_id"] = ""
        if not x["entity"]:
            x["note"] = f"unsupported type '{x['type']}' — reclass by hand"

    groups = {}
    for x in rows:
        if x["entity"] and x["txn_id"]:
            groups.setdefault((x["entity"], x["txn_id"]), []).append(x)

    total = len(groups)
    print(f"→  Fetching {total} transaction(s) to capture SyncToken + line ids "
          f"(read-only)...")
    for i, ((entity, tid), grp) in enumerate(groups.items(), 1):
        try:
            txn = get_entity(entity, tid, access, cid)
        except Exception as e:  # read-only; a blip must not abort the audit
            for x in grp:
                x["note"] = f"could not fetch {entity} {tid}: {str(e)[:60]}"
            continue
        tok = str(txn.get("SyncToken", ""))
        pool = parent_lines_of(txn, parent_id)
        for x in grp:
            x["sync_token"] = tok
            want = abs(x["debit"]) or abs(x["credit"])
            match = next((pl for pl in pool
                          if abs(abs(pl["amount"]) - want) < 0.005), None)
            if match is None and pool:
                match = pool[0]
            if match:
                x["line_id"] = match["line_id"]
                pool.remove(match)
            elif not x["note"]:
                x["note"] = "no matching parent line in txn — reclass by hand"
        if i % 10 == 0 or i == total:
            print(f"    {i}/{total}")


# ───────────────────────── sub-account suggestion ─────────────────────────

def suggest_sub(vendor, subs):
    """Best-guess sub-account for a vendor name. Returns (name, confidence)."""
    if not vendor:
        return "", ""
    vn = _norm(vendor)
    if not vn:
        return "", ""
    best, best_score = "", 0.0
    for s in subs:
        sn = _norm(s.get("Name") or "")
        if not sn:
            continue
        vt, st = set(vn.split()), set(sn.split())
        overlap = len(vt & st) / max(1, len(vt | st))
        ratio = difflib.SequenceMatcher(None, vn, sn).ratio()
        score = max(overlap, ratio)
        if score > best_score:
            best, best_score = s.get("Name") or "", score
    if best_score >= 0.7:
        return best, "high"
    if best_score >= 0.45:
        return best, "maybe"
    return "", "review"


# ───────────────────────── deep links ─────────────────────────

def register_url(account_id, realm):
    return (f"https://qbo.intuit.com/app/login?pagereq="
            f"{quote(f'register?accountId={account_id}')}"
            f"&deeplinkcompanyid={realm}")


# ───────────────────────── excel ─────────────────────────

# Order: context (left) · decision (you edit) · identity (apply reads these).
COLUMNS = [
    ("Date", 12), ("Txn Type", 16), ("Doc #", 12), ("Vendor / Name", 26),
    ("Memo", 38), ("Offsetting Account", 26), ("Debit", 12), ("Credit", 12),
    ("Suggested Sub-Account", 26), ("Match", 8),
    ("Confirm Sub-Account", 26), ("Note", 26),
    ("Entity", 12), ("Txn Id", 12), ("Line Id", 10), ("SyncToken", 10),
]
EDIT_COL = "Confirm Sub-Account"
NUM_COLS = {"Debit", "Credit"}
TEXT_COLS = {"Doc #", "Confirm Sub-Account", "Note",
             "Entity", "Txn Id", "Line Id", "SyncToken"}


def _row_values(x):
    return {
        "Date": x["date"], "Txn Type": x["type"], "Doc #": x["doc"],
        "Vendor / Name": x["name"], "Memo": x["memo"],
        "Offsetting Account": x["split"], "Debit": x["debit"], "Credit": x["credit"],
        "Suggested Sub-Account": x.get("suggested", ""), "Match": x.get("match", ""),
        "Confirm Sub-Account": "", "Note": x.get("note", ""),
        "Entity": x.get("entity", ""), "Txn Id": x.get("txn_id", ""),
        "Line Id": x.get("line_id", ""), "SyncToken": x.get("sync_token", ""),
    }


def write_xlsx(path, parent, subs, rows, realm):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parent Direct Lines"
    bold = Font(bold=True)
    col_of = {title: i for i, (title, _w) in enumerate(COLUMNS, start=1)}

    r = 1
    ws.cell(r, 1, "Loans to Sub-Contractors — lines posted directly to the "
                  "PARENT (should be $0)").font = bold
    r += 1
    ws.cell(r, 1, f"Parent account: {parent.get('Name')}  (id {parent['Id']})")
    r += 1
    total_debit = sum(x["debit"] for x in rows)
    total_credit = sum(x["credit"] for x in rows)
    ws.cell(r, 1, f"Lines: {len(rows)}    Debits: {total_debit:,.2f}    "
                  f"Credits: {total_credit:,.2f}    "
                  f"Net in parent: {total_debit - total_credit:,.2f}")
    r += 1
    ws.cell(r, 1, "Open parent register in QBO:")
    ws.cell(r, 2, register_url(parent["Id"], realm))
    r += 2

    ws.cell(r, 1, "Fill 'Confirm Sub-Account' with the sub each line moves to "
                  "(dropdown). Leave blank to skip. Don't edit the id columns.").font = bold
    r += 1

    head_row = r
    for c, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(head_row, c, title)
        cell.font = bold
        ws.column_dimensions[get_column_letter(c)].width = width
    r += 1

    for x in rows:
        vals = _row_values(x)
        for title, c in col_of.items():
            cell = ws.cell(r, c, vals[title])
            if title in NUM_COLS:
                cell.number_format = "#,##0.00"
            elif title in TEXT_COLS:
                cell.number_format = "@"
        r += 1
    last_row = r - 1

    ws.freeze_panes = ws.cell(head_row + 1, 1)

    # Reference sheet + dropdown source: the sub-accounts (recode targets).
    ws2 = wb.create_sheet("Sub-Accounts (targets)")
    for c, (title, width) in enumerate(
            [("Sub-Account", 34), ("Account Id", 12), ("Current Balance", 16)],
            start=1):
        ws2.cell(1, c, title).font = bold
        ws2.column_dimensions[get_column_letter(c)].width = width
    rr = 2
    for s in sorted(subs, key=lambda a: (a.get("Name") or "")):
        ws2.cell(rr, 1, s.get("Name"))
        ws2.cell(rr, 2, s.get("Id"))
        ws2.cell(rr, 3, float(s.get("CurrentBalance") or 0)).number_format = "#,##0.00"
        rr += 1
    names_last = max(rr - 1, 2)

    # Dropdown on the Confirm column (advisory — typing a not-yet-created sub is
    # allowed; the apply step is the real name→id gate).
    if rows:
        dv = DataValidation(
            type="list",
            formula1=f"='Sub-Accounts (targets)'!$A$2:$A${names_last}",
            allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        cl = get_column_letter(col_of[EDIT_COL])
        dv.add(f"{cl}{head_row + 1}:{cl}{last_row}")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ───────────────────────── apply (writes, gated) ─────────────────────────

def _pick_sheet(wb):
    if "Parent Direct Lines" in wb.sheetnames:
        return wb["Parent Direct Lines"]
    return wb.active


def _locate_header(ws):
    """Find the header row (tolerates the preamble rows above it) and return
    (row_index, {lowercased header -> 1-based column})."""
    for r in range(1, min(ws.max_row, 40) + 1):
        vals = {str(c.value).strip().lower(): c.column
                for c in ws[r] if c.value is not None}
        if "txn id" in vals and "confirm sub-account" in vals:
            return r, vals
    sys.exit("✗  Could not find the header row (need 'Txn Id' and 'Confirm "
             "Sub-Account'). Is this the exported review workbook?")


def apply_reclass(path, access, cid, account_name, commit, include_closed):
    if not path.exists():
        sys.exit(f"✗  No such file: {path}")

    parent, subs, _ = find_parent_and_subs(access, cid, account_name)
    parent_id = str(parent["Id"])
    sub_by_name = {(s.get("Name") or "").strip().lower(): s for s in subs}
    close = _as_date(book_close_date(access, cid))
    print(f"✓  Parent: {parent.get('Name')} (id {parent_id}) — "
          f"{len(subs)} sub-account(s)")
    print(f"   Book-close date: {close or 'none'}")
    print(f"   Mode: {'COMMIT (writing to QBO)' if commit else 'DRY-RUN (no writes)'}")

    wb = load_workbook(path)
    ws = _pick_sheet(wb)
    head_row, col = _locate_header(ws)
    need = ["entity", "txn id", "line id", "synctoken", "confirm sub-account"]
    missing = [c for c in need if c not in col]
    if missing:
        sys.exit(f"✗  Workbook is missing columns {missing}. Re-export with the "
                 f"current script (it now writes identity columns), then --apply.")

    approved = []
    for r in range(head_row + 1, ws.max_row + 1):
        confirm = ws.cell(r, col["confirm sub-account"]).value
        confirm = str(confirm).strip() if confirm is not None else ""
        if not confirm:
            continue
        approved.append({
            "entity": _norm_id(ws.cell(r, col["entity"]).value),
            "tid": _norm_id(ws.cell(r, col["txn id"]).value),
            "lid": _norm_id(ws.cell(r, col["line id"]).value),
            "tok": _norm_id(ws.cell(r, col["synctoken"]).value),
            "confirm": confirm,
        })

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = LOG_DIR / "loans_reclass_backups" / stamp
    backup_dir.mkdir(parents=True, exist_ok=True)
    results = [("entity", "txn_id", "line_id", "target_sub", "result", "detail")]
    st = {"approved_rows": len(approved), "applied_lines": 0, "would_apply": 0,
          "skip_name": 0, "skip_stale": 0, "skip_closed": 0, "skip_unsupported": 0,
          "skip_no_line": 0, "skip_not_on_parent": 0, "err": 0}

    # Group by transaction — one fetch + one write per txn (a second write with
    # the pre-write SyncToken would fail).
    groups = {}
    for a in approved:
        groups.setdefault((a["entity"], a["tid"]), []).append(a)

    for (entity, tid), grp in groups.items():
        if entity not in WRITABLE_ENTITIES or not tid:
            for a in grp:
                results.append((entity, tid, a["lid"], a["confirm"],
                                "SKIP_UNSUPPORTED", f"entity not writable ({entity or '—'})"))
                st["skip_unsupported"] += 1
            continue

        targets = {}  # line_id -> (sub_id, sub_name)
        for a in grp:
            s = sub_by_name.get(a["confirm"].strip().lower())
            if not s:
                results.append((entity, tid, a["lid"], a["confirm"],
                                "SKIP_NAME_NOT_FOUND",
                                "confirmed sub-account not found — create it, re-export"))
                st["skip_name"] += 1
                continue
            if not a["lid"]:
                results.append((entity, tid, a["lid"], a["confirm"],
                                "SKIP_NO_LINE_ID",
                                "no line id captured — re-export or reclass by hand"))
                st["skip_no_line"] += 1
                continue
            targets[a["lid"]] = (str(s["Id"]), s.get("Name") or "")
        if not targets:
            continue

        try:
            fresh = get_entity(entity, tid, access, cid)
        except Exception as e:
            for lid, (sid, sn) in targets.items():
                results.append((entity, tid, lid, sn, "ERROR", f"fetch failed: {str(e)[:120]}"))
                st["err"] += 1
            continue

        stored_tok = grp[0]["tok"]
        if str(fresh.get("SyncToken", "")) != str(stored_tok):
            for lid, (sid, sn) in targets.items():
                results.append((entity, tid, lid, sn, "SKIP_STALE",
                                f"SyncToken {fresh.get('SyncToken')}≠export {stored_tok} — re-export"))
                st["skip_stale"] += 1
            continue

        tdate = _as_date(fresh.get("TxnDate", ""))
        if close and tdate and tdate <= close and not include_closed:
            for lid, (sid, sn) in targets.items():
                results.append((entity, tid, lid, sn, "SKIP_CLOSED", str(tdate)))
                st["skip_closed"] += 1
            continue

        (backup_dir / f"{entity}_{tid}.json").write_text(json.dumps(fresh, indent=1))

        change_lines = []  # (line_id, sub_id, sub_name)
        for line in fresh.get("Line", []) or []:
            lid = str(line.get("Id", ""))
            if lid not in targets:
                continue
            sid, sn = targets[lid]
            det = line.get("AccountBasedExpenseLineDetail")
            if not det:
                results.append((entity, tid, lid, sn, "SKIP_NOT_ACCOUNT_LINE",
                                "line is not an account-based expense line"))
                st["skip_not_on_parent"] += 1
                continue
            cur = str((det.get("AccountRef") or {}).get("value", ""))
            if cur != parent_id:
                results.append((entity, tid, lid, sn, "SKIP_NOT_ON_PARENT",
                                f"posts to {cur} not parent {parent_id} — already moved?"))
                st["skip_not_on_parent"] += 1
                continue
            det["AccountRef"] = {"value": sid, "name": sn}
            change_lines.append((lid, sid, sn))
        if not change_lines:
            continue

        if commit:
            fresh["sparse"] = False
            try:
                update_entity(entity, fresh, access, cid)
                time.sleep(0.25)
            except Exception as e:
                for lid, sid, sn in change_lines:
                    results.append((entity, tid, lid, sn, "ERROR", f"update failed: {str(e)[:120]}"))
                    st["err"] += 1
                continue
            for lid, sid, sn in change_lines:
                results.append((entity, tid, lid, sn, "APPLIED", f"{parent_id} → {sid}"))
                st["applied_lines"] += 1
        else:
            for lid, sid, sn in change_lines:
                results.append((entity, tid, lid, sn, "WOULD_APPLY", f"{parent_id} → {sid}"))
                st["would_apply"] += 1

    out_csv = backup_dir / "apply_results.csv"
    with open(out_csv, "w", newline="") as f:
        csv.writer(f).writerows(results)

    print()
    print(f"{'COMMIT' if commit else 'DRY-RUN'} summary:")
    for k, v in st.items():
        print(f"   {k:20s} {v}")
    print(f"→  Results CSV: {out_csv}")
    print(f"→  Per-txn backups: {backup_dir}")
    if not commit:
        print("   No writes made. Re-run with --commit to apply the approved lines.")


# ───────────────────────── export (read-only) ─────────────────────────

def _default_out_path():
    return (paths.get_path("ACB_AUDIT_OUT_DIR",
            paths.onedrive_base() / "Works In Progress" / "QBO Audits")
            / "Loans_to_SubContractors_Audit.xlsx")


def export(access, cid, account_name, start, end, out_override):
    parent, subs, _by_id = find_parent_and_subs(access, cid, account_name)
    print(f"✓  Parent: {parent.get('Name')} (id {parent['Id']}) — "
          f"{len(subs)} sub-account(s)")

    report = qbo_api.report(access, cid, "GeneralLedger", params={
        "start_date": start,
        "end_date": end,
        "accounting_method": "Accrual",
        "account": parent["Id"],
        "columns": ("tx_date,txn_type,doc_num,name,memo,"
                    "split_acc,debt_amt,credit_amt,subt_nat_amount"),
    })

    dump = LOG_DIR / f"loans_to_subs_GL_{dt.date.today().isoformat()}.json"
    dump.write_text(json.dumps(report, indent=2))

    sub_names = [s.get("Name") or "" for s in subs]
    rows, titles, sections = parse_gl(report, parent.get("Name"), sub_names)

    if not rows:
        print("⚠  No parent-only lines parsed. Column titles QBO returned: "
              f"{titles}")
        print(f"   Account sections seen: {sorted(set(sections))}")
        print(f"   Raw GL saved to: {dump}  (share this and I'll adjust the "
              f"parser)")

    for x in rows:
        s, m = suggest_sub(x["name"], subs)
        x["suggested"], x["match"] = s, m

    enrich_rows(rows, str(parent["Id"]), access, cid)

    out = Path(out_override).expanduser() if out_override else _default_out_path()
    write_xlsx(out, parent, subs, rows, cid)

    td = sum(x["debit"] for x in rows)
    tc = sum(x["credit"] for x in rows)
    print(f"✓  {len(rows)} line(s) posted directly to the parent")
    print(f"   Debits {td:,.2f}  Credits {tc:,.2f}  Net {td - tc:,.2f}")
    matched = sum(1 for x in rows if x["match"] in ("high", "maybe"))
    print(f"   Sub-account guessed for {matched}/{len(rows)} line(s) "
          f"(confirm the rest by hand)")
    writable = sum(1 for x in rows if x.get("entity") in WRITABLE_ENTITIES and x.get("line_id"))
    print(f"   Apply-ready (writable + line id): {writable}/{len(rows)} line(s)")
    unsupported = len(rows) - writable
    if unsupported:
        print(f"   {unsupported} line(s) need a human (unsupported type / no line id) "
              f"— see the Note column")
    print(f"→  {out}")
    print("   Next: review the workbook, fill 'Confirm Sub-Account', then:")
    print("        python3 one-offs/loans_to_subs_audit.py --apply            (dry-run)")
    print("        python3 one-offs/loans_to_subs_audit.py --apply --commit    (write)")


# ───────────────────────── main ─────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Reclass lines posted directly to the 'Loans to "
                    "Sub-Contractors' parent into per-sub sub-accounts "
                    "(export → you confirm → gated apply).")
    ap.add_argument("--account", default=DEFAULT_ACCOUNT,
                    help=f"Parent account name (default: {DEFAULT_ACCOUNT})")
    ap.add_argument("--start", default="2015-01-01", help="GL start date")
    ap.add_argument("--end", default=dt.date.today().isoformat(),
                    help="GL end date (default: today)")
    ap.add_argument("--out", default=None, help="Export xlsx path override")
    ap.add_argument("--apply", nargs="?", const="__DEFAULT__", default=None,
                    metavar="XLSX",
                    help="Apply mode: read the reviewed workbook and reclass "
                         "confirmed lines (default path if none given)")
    ap.add_argument("--commit", action="store_true",
                    help="with --apply: actually write to QBO (default: dry-run)")
    ap.add_argument("--include-closed", action="store_true",
                    help="with --apply --commit: also touch closed-period txns")
    a = ap.parse_args()

    access, cid = qbo_api.load_credentials()

    if a.apply is not None:
        path = _default_out_path() if a.apply == "__DEFAULT__" else Path(a.apply)
        apply_reclass(path, access, cid, a.account, a.commit, a.include_closed)
    else:
        export(access, cid, a.account, a.start, a.end, a.out)


if __name__ == "__main__":
    main()
