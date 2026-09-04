#!/usr/bin/env python3
"""
statement_reconciler.py — Vendor Statement vs QBO Reconciler

Takes a vendor statement PDF, identifies the vendor, pulls every open
Bill in QBO for that vendor, and reconciles the two side-by-side.

Categorizes each statement line into one of four buckets:
  ✓ MATCHED              — same Ref#, same amount (clerk OK)
  ⚠ VENDOR_TAX_VIOLATION — same Ref#, amount diff is exactly 8.25% (TX
                            sales tax) of the QBO amount → vendor billed
                            tax in violation of Proficient/vendor no-tax
                            agreement. Vendor action, not clerk action.
  ✗ CLERK_AMOUNT_MISMATCH— same Ref#, amount diff is some other value
                            → clerk entered wrong amount in QBO.
  ✗ MISSING_IN_QBO       — on statement, no Bill in QBO with that Ref#
                            → clerk has not entered the bill yet.
  ✗ MISSING_ON_STATEMENT — Bill exists open in QBO, but vendor doesn't
                            show it on the statement → stale unpaid bill
                            vendor may have credited / already received
                            payment for; needs the user's eyes.

Writes an Excel report (Summary + one sheet per category) named
  Statement_Reconciliation_<date>_<vendor>.xlsx

Manually-passed files get the SAME treatment as an --inbox sweep: each Excel
lands in the Synology Reconciliations folder, and a source file that already
lives in the Statement Inbox is archived to its DONE subfolder on success. If
the Synology share isn't mounted, output falls back to OUTDIR_DEFAULT (below)
and no file is moved. (--out is accepted but ignored in this mode.)

SUPPORTED PDF TEMPLATES (auto-detected by report-type signature, never by vendor name)
  • QuickBooks Statement                 — vendor-issued statement with "INV #<num>. Due <date>" lines
  • QuickBooks Customer Open Balance     — QBO Customer Open Balance report, columnar
  • QuickBooks Open Invoices             — QBO Open Invoices report, columnar (4-col)
  • Plus per-vendor layouts (White Cap, Bobcat, Bodin, BURNCO, Cintas, Cow Town,
    Sunbelt, Croell) and generic tabular/columnar — see TEMPLATE_LABELS.

USAGE
  python3 statement_reconciler.py /path/to/statement.pdf            # inbox-style: Excel → Reconciliations, source → DONE if in inbox
  python3 statement_reconciler.py /path/to/statement.pdf --vendor "Exact QBO Display Name"
  python3 statement_reconciler.py /path/to/statement.pdf --dry-run  # reconcile + print, write/move nothing
  python3 statement_reconciler.py /path/to/statement.pdf --yes      # skip prompts
  python3 statement_reconciler.py --inbox                           # sweep the whole Statement Inbox

INTERACTIVE FLOW
  Two Y/N prompts before any QBO call so a misread PDF never wastes API
  roundtrips:
    1. After parse → shows vendor / date / total / first+last line. Confirm.
    2. After QBO vendor lookup → shows matched vendor name. Confirm.
  Ends with an INBOX SUMMARY (reconciled / moved-to-DONE / left-for-a-human) and
  a clickable link to the Reconciliations folder. Does not auto-open the Excel.

DEPENDENCIES
  pip3 install --break-system-packages pdfplumber requests openpyxl
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.parse
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Tuple

# Allow import of qbo_vault from project root (same pattern as qbo_bill_tracker)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import requests
except ImportError:
    print("✗ pip3 install --break-system-packages requests")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("✗ pip3 install --break-system-packages pdfplumber")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ pip3 install --break-system-packages openpyxl")
    sys.exit(1)

from shared import qbo_vault as kc
from shared import paths

# ───────────────────────── constants ─────────────────────────

API_BASE = "https://quickbooks.api.intuit.com"
MINOR_VERSION = "70"
TX_SALES_TAX = 0.0825
TAX_TOLERANCE = 1.00  # dollars — accept within $1 of exact 8.25% match
EXACT_TOLERANCE = 0.01  # dollars — bill amount equality
LAG_LOOKBACK_DAYS = 60  # FLOOR for paid-bill lookback (days before stmt date)
LAG_LOOKBACK_STMT_PADDING_DAYS = 7  # extra padding before the statement's oldest line
# Effective cutoff = min(oldest_stmt_line_date - PADDING, stmt_date - LAG_LOOKBACK_DAYS)
# So statements with old bills extend the lookback as needed, but every run
# still pulls AT LEAST 60 days of paid history.
UNAPPROVED_TAG = "not approved"  # case-insensitive substring in QBO Bill Memo flags unapproved


def _is_approved(memo: str) -> bool:
    """Return False iff Memo STARTS WITH 'Not Approved' (case-insensitive,
    after stripping leading whitespace). Checks the beginning only so a
    memo like 'Approved 5/1 (was Not Approved prior)' is correctly treated
    as approved. Empty/missing memo → approved (default)."""
    return not (memo or "").lstrip().lower().startswith(UNAPPROVED_TAG)

OUTDIR_DEFAULT = paths.get_path(
    "ACB_RECON_OUT_DIR",
    paths.onedrive_base() / "Automations-/statement reconciles",
)
ALIAS_FILE = Path(__file__).resolve().parent / "vendor_aliases.json"
CLERK_PERF_CSV = OUTDIR_DEFAULT / "clerk_performance.csv"

# QBO deep-link to open a Bill in the browser. Same format used by the sibling
# bill tracker (excel_bill_sync.py). Rendered as a ↗ hyperlink in the report.
QBO_BILL_URL_TEMPLATE = "https://qbo.intuit.com/app/bill?txnId={bill_id}"

# ── Inbox automation (Synology) ────────────────────────────────
# Folder-driven workflow root on the Synology share (must be mounted). The
# sweep reads statements from the Inbox, writes each Excel to Reconciliations,
# and moves the processed source file into the Inbox's DONE subfolder.
INBOX_ROOT = Path("/Volumes/Accounting/Automations/Vendor Statements")
INBOX_SUPPORTED_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".heic", ".heif", ".xlsx", ".xls"}
STATEMENT_EMBED_MAX_PAGES = 20   # cap embedded statement pages to keep xlsx size sane
STATEMENT_EMBED_MAX_WIDTH = 900  # px — target on-sheet width per embedded page

# ── Template: QuickBooks Statement (vendor-issued) ─────────────
# Layout: a leading date and an "INV #<num>. Due <date>" anchor on every line.
# Example line:   02/25/2026 INV #589898. Due 03/15/2026 ...  80,620.00  84,770.31
QBO_STATEMENT_SIG = re.compile(r"INV\s*\#\d+\.\s*Due\s+\d", re.I)
STMT_LINE_RE = re.compile(
    r"""
    ^\s*
    (?P<date>\d{1,2}/\d{1,2}/\d{2,4})           # 02/25/2026
    \s+
    INV\s*\#?(?P<ref>\d+)\.?                    # INV #589898.
    .*?                                          # ... description ...
    (?P<amount>[\d,]+\.\d{2})                   # 80,620.00 (line amount)
    \s+
    (?P<balance>[\d,]+\.\d{2})\s*$              # 84,770.31 (running balance)
    """,
    re.VERBOSE | re.MULTILINE,
)

PO_RE = re.compile(r"PO\s*\#?\s*([^.]+?)\.", re.IGNORECASE)
ORIG_AMT_RE = re.compile(r"Orig\.\s*Amount\s*\$?([\d,]+\.\d{2})", re.IGNORECASE)
# Tried in order — first one with a hit wins. Most-specific label first.
STMT_DATE_PATTERNS = [
    # "Statement Date" / "Stmt Date" / "Statement Dt" label, then up to 40 chars to the date
    re.compile(r"(?:Statement|Stmt|Stat)\s*Da?te?[:\s][\s\S]{0,40}?(\d{1,2}/\d{1,2}/\d{2,4})", re.I),
    # Bare "Date" label, then up to 40 chars (vendor statement header noise between label and value)
    re.compile(r"\bDate\b[:\s][\s\S]{0,40}?(\d{1,2}/\d{1,2}/\d{2,4})", re.I),
    # Fallback: first standalone date on its own line — only used if labels missing
    re.compile(r"(?:^|\n)\s*(\d{1,2}/\d{1,2}/\d{4})\s*(?:\n|$)"),
]
# Amount Due: allow up to 80 chars (incl. newlines + "Amount Enc." label) between
# the "Amount Due" text and the dollar amount.
AMT_DUE_RE = re.compile(r"Amount\s+Due[\s\S]{0,80}?\$\s*([\d,]+\.\d{2})", re.IGNORECASE)

# ── Template: QuickBooks Customer Open Balance ─────────────────
# Vendor runs a "Customer Open Balance" report for Proficient Concrete from
# their own QBO. Layout (extracted text order):
#   <Vendor Name>                       ← vendor on its own line at top
#   9:35 AM                              ← timestamp
#   Customer Open Balance                ← report-title signature
#   05/15/26                             ← print date
#   Accrual Basis As of March 31, 2026   ← statement date
#   Type Date Num Memo Due Date Open Balance
#   PROFICIENT CONCRETE, LLC.            ← bill-to header
#   <SUB-CUSTOMER NAME>                  ← job
#   Invoice 3/18/2026 K733367 8811 OLD DECATUR RD 4/17/2026 269.95
#   Total <SUB-CUSTOMER>  269.95
#   ...
#   TOTAL 265,372.99
# Detection matches two title variants:
#   • "Customer Open Balance" (older QBO label)
#   • Header row "Type Date Num Memo Due Date Open Balance" (used by newer
#     QBO Statement reports titled "May 2026 Statement" or similar)
QBO_CUSTOMER_OPEN_BAL_SIG = re.compile(
    r"Customer\s+Open\s+Balance"
    r"|Type\s+Date\s+Num\s+Memo\s+Due\s+Date\s+Open\s+Balance",
    re.I,
)
QBO_AS_OF_RE = re.compile(
    r"As\s+of\s+([A-Za-z]+\s+\d{1,2},?\s+\d{4})", re.I)
# Matches EVERY transaction row in a Customer Open Balance report, not just
# invoices. Credit Memos and Payments carry NEGATIVE amounts (leading "-" or
# parentheses) and Payments have no due-date column — both must be parsed so
# the line-sum ties to the report's grand total. (Before 2026-08-12 only
# "Invoice" rows matched, so credits/payments were dropped, the line-sum came
# out too HIGH by the credit total, and every statement carrying a credit
# falsely failed the tie-out.) The due-date group is optional; the amount group
# accepts a sign or parentheses.
QBO_CUSTOMER_OPEN_BAL_LINE_RE = re.compile(
    r"""^\s*
    (?P<type>(?:Invoice|Credit(?:\s+Me\w*)?|Payment|Discount|Journal|
                 Deposit|Sales\s+Receipt|Check|Bill\s+Pmt|Transfer)
             (?:\s*\.\.\.)?)\s+        # QBO truncates a narrow Type cell to "...":
                                       #   "Credit ..." and "Credit Me..." both seen
    (?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<num>\S+)\s+
    (?P<memo>.+?)
    (?:\s+(?P<due>\d{1,2}/\d{1,2}/\d{2,4}))?\s+
    (?P<amount>\(?-?[\d,]+\.\d{2}\)?)\s*$""",
    re.VERBOSE | re.MULTILINE,
)
QBO_GRAND_TOTAL_RE = re.compile(r"^\s*TOTAL\s+([\d,]+\.\d{2})\s*$", re.MULTILINE)
# Allow asterisks/lowercase in body so QBO's "**EXEMPT" / "*EXEMPT" markers
# and " - Other" suffix on the parent customer header still register as
# sub-customers. (Example: "Dallas Area Habitat for Humanity **EXEMPT".)
QBO_SUBCUST_RE = re.compile(
    r"^\s*(?!Total\b|TOTAL\b|Invoice\b|Credit\b|Type\b|Accrual\b|Cash\b|Page\b)"
    r"([A-Z][A-Za-z0-9\s\(\)\.,&'\-/\*]{3,})\s*$",
    re.MULTILINE,
)

# ── Template: QuickBooks Open Invoices ─────────────────────────
# Vendor runs an "Open Invoices" report for Proficient Concrete from their
# own QBO. Layout (extracted text order — vendor appears at bottom because
# pdfplumber reads the header columns first):
#   Date Num Due Date Open Balance            ← column header
#   Proficient Concrete                        ← bill-to header
#   <SUB-CUSTOMER NAME> [**EXEMPT marker]      ← job
#   09/26/2025 D80025 10/26/2025 372.96       ← invoice line (4 cols)
#   Total <SUB-CUSTOMER> 372.96
#   ...
#   Total Proficient Concrete 90,507.30
#   TOTAL 90,507.30
#   1:27 PM <Vendor Name>                      ← timestamp + vendor SAME line
#   05/21/26 Open Invoices                     ← print-date + title SAME line
#   As of May 21, 2026                          ← statement date
#   Page 1
QBO_OPEN_INVOICES_SIG = re.compile(r"\bOpen\s+Invoices\b", re.I)
QBO_OPEN_INV_LINE_RE = re.compile(
    r"""^\s*
    (?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<num>\S+)\s+
    (?P<due>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<amount>[\d,]+\.\d{2})\s*$""",
    re.VERBOSE | re.MULTILINE,
)
# Timestamp pattern shared by both QBO report templates — used by the
# generic vendor finder.
QBO_TIMESTAMP_RE = re.compile(
    r"^\s*(?P<time>\d{1,2}:\d{2}\s*(?:AM|PM))\s*(?P<rest>.*)$", re.I)

# ── Template: Vendor Statement (tabular, CMC-style) ────────────
# Vendors using line-based statement layouts with the header:
#   "Date Invoice Due Date Amount Pymt Date Payment Amount Tp Balance"
# Each data row: <date> <inv#> <due-date> <amount> [optional payment fields] <balance>
# Sub-customer groupings ("SHIP TO XYZ") and aging rows are skipped.
VENDOR_STMT_TABULAR_SIG = re.compile(
    r"Date\s+Invoice\s+Due\s+Date\s+Amount\b.*?\bBalance\b",
    re.I | re.DOTALL,
)
VENDOR_STMT_TABULAR_LINE_RE = re.compile(
    r"""^\s*
    (?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<ref>[A-Za-z0-9]+)\s+
    (?P<due>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<amount>[\d,]+\.\d{2})
    .*?
    (?P<balance>[\d,]+\.\d{2})\s*$
    """,
    re.VERBOSE | re.MULTILINE,
)
# CMC sub-customer header lines (skip these)
VENDOR_STMT_SHIPTO_RE = re.compile(r"^\s*SHIP[- ]TO\b", re.I)

# ── Template: Vendor Statement (columnar — Preferred Materials / Sunrise) ──
# These statements have a visible table grid in the PDF, but pdfplumber's
# default text extraction returns each COLUMN as a vertical stack of values
# (Date column, Description column, Charge column, Balance column each as
# their own list of lines, NOT one line per row). Parsing this requires
# pdfplumber.extract_words() with x/y coordinates to reconstruct rows.
# Detection: title "Statement" + the column-header tokens "Description" and
# "Charge" appearing in the extracted text without the CMC-style header.
VENDOR_STMT_COLUMNAR_SIG = re.compile(
    r"\bStatement\b.*?\bDescription\b.*?\bCharge\b.*?\bBalance\b",
    re.I | re.DOTALL,
)

# ── Template: Vendor Statement (White Cap / Billtrust-style) ───
# White Cap-issued statements via Billtrust. Layout: Transaction Date,
# Transaction No., T (type code), Original Transaction, Balance Due.
# Type codes: I=Invoice, C=Credit Memo, R=Rental, D=Debit Memo,
# U=Unapplied Payment, *=In Review. pdfplumber.extract_tables() returns
# all rows of the data table as ONE cell of newline-separated lines —
# easy to split and parse line-by-line.
VENDOR_STMT_WHITECAP_SIG = re.compile(
    r"(?=[\s\S]*?White\s+Cap)"
    r"(?=[\s\S]*?CLOSING\s+DATE)"
    r"(?=[\s\S]*?BALANCE\s+DUE)",
    re.I,
)
# Line format in pdfplumber's text extraction:
#   <date> <ref> <type> <original> <balance>  [<ref-dup> <po> <balance-dup>]
# The right-side remittance copy gets merged onto the same line — we capture
# only the first 5 fields and ignore the rest.
WHITECAP_ROW_RE = re.compile(
    r"""^\s*
    (?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+
    (?P<ref>\d+)\s+
    (?P<tp>[A-Z]\*?)\s+
    (?P<orig>-?[\d,]+\.\d{2})\s+
    (?P<balance>-?[\d,]+\.\d{2})
    (?:\s+.*)?$    # optional trailing fields (right-side remittance copy)
    """,
    re.VERBOSE | re.MULTILINE,
)

# Strings that should NEVER be treated as a vendor name when extracted.
_VENDOR_NOISE_RE = re.compile(
    r"^(?:Page\s+\d+|Open\s+Invoices|Customer\s+Open\s+Balance|Accrual\s+Basis|Cash\s+Basis|"
    r"As\s+of\b|Date\b|Type\b|Total\b|TOTAL\b|Invoice\b|Credit\b|"
    r"Statement\s*$|Statement\s+Date\b|Bill\s+To\b|Amount\s+Due\b|"
    r"Currency\b|Subsidiary\b|Company\s*$|Description\b|Charge\b|Balance\s+Forward\b)",
    re.I,
)

# Styles
HEADER_FILL   = PatternFill("solid", start_color="1F4E78")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT    = Font(bold=True, name="Arial", size=14, color="1F4E78")
LABEL_FONT    = Font(bold=True, name="Arial", size=11)
BODY_FONT     = Font(name="Arial", size=10)
SUBTOTAL_FILL = PatternFill("solid", start_color="E7E6E6")
OK_FILL       = PatternFill("solid", start_color="C6EFCE")
WARN_FILL     = PatternFill("solid", start_color="FFEB9C")
BAD_FILL      = PatternFill("solid", start_color="FFC7CE")
THIN          = Side(border_style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT         = Alignment(horizontal="right",  vertical="center")
MONEY         = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

# ───────────────────────── terminal UI (stdlib only) ─────────────────────────

class _Term:
    """ANSI helpers. Auto-disables on non-tty or when NO_COLOR/--no-color is set."""
    enabled: bool = sys.stdout.isatty() and os.environ.get("NO_COLOR") is None
    G    = "\033[32m"  # green
    Y    = "\033[33m"  # yellow
    R    = "\033[31m"  # red
    B    = "\033[34m"  # blue
    C    = "\033[36m"  # cyan
    DIM  = "\033[2m"
    BOLD = "\033[1m"
    RESET    = "\033[0m"
    CLEARLINE = "\033[2K"  # clear entire line

    @classmethod
    def disable(cls) -> None:
        cls.enabled = False

    @classmethod
    def color(cls, code: str, text: str) -> str:
        if not cls.enabled:
            return text
        return f"{code}{text}{cls.RESET}"


def _width() -> int:
    return shutil.get_terminal_size((80, 24)).columns


def _phase(label: str) -> float:
    """Print a phase header → returns start time so the caller can stamp duration."""
    print(_Term.color(_Term.C, f"→ {label}"))
    return time.time()


def _done(t0: float, msg: str, marker: str = "✓", color: str = _Term.G) -> None:
    elapsed = time.time() - t0
    print(f"  {_Term.color(color, marker)} {msg}  {_Term.color(_Term.DIM, f'({elapsed*1000:.0f} ms)')}")


def _warn(msg: str) -> None:
    print(f"  {_Term.color(_Term.Y, '⚠')} {msg}")


def _fail(msg: str) -> None:
    print(f"  {_Term.color(_Term.R, '✗')} {msg}")


def _bar(current: int, total: int, suffix: str = "", width_target: int = 24) -> None:
    """Inline progress bar: redraws in place. No newline until caller adds one."""
    if total <= 0:
        return
    pct = max(0.0, min(1.0, current / total))
    filled = int(width_target * pct)
    bar = "█" * filled + "░" * (width_target - filled)
    line = f"\r  [{bar}] {current}/{total}  {suffix}"
    # Trim to terminal width so no wrap on tiny windows
    line = line[: max(20, _width() - 1)]
    sys.stdout.write(_Term.CLEARLINE + line)
    sys.stdout.flush()


def _bar_end() -> None:
    """Move to a fresh line after a progress bar finishes."""
    sys.stdout.write("\n")
    sys.stdout.flush()


# Category → display label + color for the live stream
_CAT_STYLE: Dict[str, Tuple[str, str]] = {
    "MATCHED":               ("✓ Match    ", _Term.G),
    "VENDOR_TAX_VIOLATION":  ("⚠ Tax viol ", _Term.Y),
    "CLERK_AMOUNT_MISMATCH": ("⚠ Mismatch ", _Term.Y),
    "LIKELY_VENDOR_LAG":     ("⊙ Paid lag ", _Term.B),
    "MISSING_IN_QBO":        ("✗ Not in QB", _Term.R),
    "MISSING_ON_STATEMENT":  ("✗ Not on St", _Term.R),
}


# ───────────────────────── data classes ─────────────────────────

@dataclass
class StmtLine:
    date: str           # YYYY-MM-DD
    ref: str            # invoice number
    amount: float       # line amount
    po: str = ""        # PO number from description
    address: str = ""   # job address text

@dataclass
class QboBill:
    bill_id: str
    doc_number: str
    txn_date: str
    open_balance: float
    total_amount: float
    memo: str = ""        # Bill memo from QBO. The QBO UI "Memo" field maps to
                          # the API `PrivateNote` field (verified via project
                          # script read_private_note.py). We also accept the
                          # rarer top-level `Memo` field as fallback. Contains
                          # "Not Approved" when AP/PM hasn't signed off.

    @property
    def is_approved(self) -> bool:
        return _is_approved(self.memo)

@dataclass
class ReconRow:
    category: str       # MATCHED | VENDOR_TAX_VIOLATION | CLERK_AMOUNT_MISMATCH | LIKELY_VENDOR_LAG | MISSING_IN_QBO | MISSING_ON_STATEMENT
    # Note: `date` and `ref` are kept for backward-compatibility with action-list,
    # CSV, and detail-sheet code. For MISSING_ON_STATEMENT they hold the QBO data.
    # The summary sheet uses the more explicit stmt_*/qbo_* fields below.
    date: str           # canonical date (stmt date if available, else QBO date)
    ref: str            # canonical ref (stmt ref if available, else QBO ref)
    stmt_amount: float
    qbo_amount: float
    po: str
    address: str
    notes: str = ""
    # Side-by-side fields for the new summary layout
    stmt_ref: str = ""
    qbo_ref: str = ""
    stmt_date: str = ""
    qbo_date: str = ""
    qbo_memo: str = ""    # bill-level Memo from QBO; "" when there is no QBO bill (MISSING_IN_QBO)
    qbo_bill_id: str = "" # QBO internal Bill Id, used to build the ↗ deep-link; "" when no QBO bill

# ───────────────────────── PDF parsing ─────────────────────────

def _pdf_text(pdf_path: Path) -> str:
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def _image_to_text(img_path: Path) -> str:
    """OCR an image (PNG/JPG/etc.) to plain text using Tesseract.
    Lazy-imports pytesseract + Pillow so installs aren't required for users
    who never pass image statements. Prints actionable install instructions
    on failure rather than a stack trace."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError as e:
        sys.exit(_Term.color(_Term.R,
            f"✗ Image OCR needs Tesseract + Python wrappers. Install once:\n"
            f"    brew install tesseract\n"
            f"    pip3 install pytesseract pillow pillow-heif\n"
            f"  (missing: {e.name})"))
    # Verify the tesseract binary itself is reachable
    try:
        pytesseract.get_tesseract_version()
    except Exception:
        sys.exit(_Term.color(_Term.R,
            "✗ Tesseract binary not found on PATH. Install once:\n"
            "    brew install tesseract"))
    # HEIC support (iPhone screenshots) needs the optional pillow-heif backend
    if img_path.suffix.lower() in (".heic", ".heif"):
        try:
            from pillow_heif import register_heif_opener
            register_heif_opener()
        except ImportError:
            sys.exit(_Term.color(_Term.R,
                "✗ HEIC images need the pillow-heif backend. Install once:\n"
                "    pip3 install pillow-heif"))
    try:
        img = Image.open(img_path)
        # --psm 6: assume a single uniform block of text. Critical for
        # statement tables — keeps each invoice row as a single line of OCR
        # output (default mode breaks columns into separate vertical stacks
        # which destroys the date-ref-amount alignment our parsers need).
        # preserve_interword_spaces=1 keeps wide-spaced columns readable.
        return pytesseract.image_to_string(
            img,
            config="--psm 6 -c preserve_interword_spaces=1"
        )
    except Exception as e:
        sys.exit(_Term.color(_Term.R, f"✗ OCR failed on {img_path.name}: {e}"))


TEMPLATE_LABELS = {
    "qbo_statement":             "QuickBooks Statement (PDF)",
    "qbo_customer_open_balance": "QuickBooks Customer Open Balance / Statement (PDF — Type Date Num Memo Due Date Open Balance header)",
    "qbo_open_invoices":         "QuickBooks Open Invoices (PDF)",
    "vendor_stmt_tabular":       "Vendor Statement, tabular (PDF — Date Invoice Due Date Amount ... Balance header, e.g. CMC)",
    "vendor_stmt_columnar":      "Vendor Statement, columnar (PDF — Date Description Charge Payment Balance grid, e.g. Preferred Materials / Sunrise)",
    "vendor_stmt_whitecap":      "Vendor Statement, White Cap / Billtrust (PDF — Transaction Date | Transaction No. | T | Original | Balance Due with I/C/R/D/U type codes)",
    "excel_columnar":            "Excel statement (.xlsx or .xls — columns: Inv Date | Invoice # | Original Invoice Amount | Balance | Due Date)",
    "vendor_bobcat":             "Vendor Statement, Bobcat/equipment (PDF — Invoice Number | Invoice Date | Due Date | Purchase Order | Balance)",
    "vendor_bodin":              "Vendor Statement, Bodin/concrete (PDF — Invoice | Date | Type | Reference | Yardage | Credit/Debit | Balance)",
    "vendor_burnco":             "Vendor Statement, BURNCO (PDF — Date | Number | Delivery Address | PO | Type | Original | Balance Due)",
    "vendor_cintas":             "Vendor Statement, Cintas (PDF — Date | Sold-To | Reference | Amount Due | Due Date)",
    "vendor_cowtown":            "Vendor Statement, past-due letter (PDF — Job | Inv. No. | Inv. Date | Due Date | Inv. Amount | Balance)",
    "vendor_sunbelt":            "Vendor Statement, Sunbelt Rentals (PDF — Date | Invoice | Job Description | Amount Due)",
    "vendor_croell":             "Vendor Statement, Croell Inc (PDF — Date | Cd | Invoice | Description | Amount | Balance doubled register/remittance layout)",
}


def detect_template(text: str) -> str:
    """Map a PDF's extracted text to one of the supported template keys.
    Templates are named by REPORT TYPE, never by vendor. Returns "" if no
    supported template matches."""
    # Order matters: most-specific signature first.
    # QBO Statement: "INV #N. Due N" line anchor
    if QBO_STATEMENT_SIG.search(text):
        return "qbo_statement"
    # QBO Customer Open Balance / new "Statement" variant
    if QBO_CUSTOMER_OPEN_BAL_SIG.search(text):
        return "qbo_customer_open_balance"
    # QBO Open Invoices
    if QBO_OPEN_INVOICES_SIG.search(text):
        return "qbo_open_invoices"
    # White Cap / Billtrust — check BEFORE the more-generic columnar/tabular
    # signatures since they could partially match a White Cap layout
    if VENDOR_STMT_WHITECAP_SIG.search(text):
        return "vendor_stmt_whitecap"
    # Specific vendor-statement layouts (2026-07-01 batch) — checked before the
    # generic tabular/columnar signatures so their distinct headers win.
    if BOBCAT_SIG.search(text):
        return "vendor_bobcat"
    if BODIN_SIG.search(text):
        return "vendor_bodin"
    if BURNCO_SIG.search(text):
        return "vendor_burnco"
    if CINTAS_SIG.search(text):
        return "vendor_cintas"
    if COWTOWN_SIG.search(text):
        return "vendor_cowtown"
    if SUNBELT_SIG.search(text):
        return "vendor_sunbelt"
    # Croell Inc — doubled register/remittance header; must beat the generic
    # tabular/columnar sigs (its "Finance Charge" wording trips columnar).
    if CROELL_SIG.search(text):
        return "vendor_croell"
    # Vendor tabular statement (Date Invoice Due Date Amount ... Balance header)
    if VENDOR_STMT_TABULAR_SIG.search(text):
        return "vendor_stmt_tabular"
    # Vendor columnar statement (Date / Description / Charge / Balance grid)
    if VENDOR_STMT_COLUMNAR_SIG.search(text):
        return "vendor_stmt_columnar"
    return ""


def _is_vendor_noise(s: str) -> bool:
    """True when a line is clearly NOT a vendor name (date, time, page, title,
    header row, our own company)."""
    if not s:
        return True
    if re.match(r"^\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?\s*$", s):
        return True   # pure date
    if re.match(r"^\d{1,2}:\d{2}\s*(?:AM|PM)\s*$", s, re.I):
        return True   # pure time
    if _VENDOR_NOISE_RE.match(s):
        return True   # known title / header keyword
    if "PROFICIENT" in s.upper():
        return True   # our own company — never the vendor we're reconciling against
    return False


def _find_qbo_report_vendor(text: str) -> str:
    """Extract vendor name from a QBO report (Customer Open Balance, Open
    Invoices, or any future QBO report). The vendor is paired with a HH:MM
    AM/PM timestamp the report-runner's QBO prints — either on the SAME line
    after the timestamp, or on the line IMMEDIATELY BEFORE it. Handles both
    layouts observed in pdfplumber-extracted text order."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        m = QBO_TIMESTAMP_RE.match(line)
        if not m:
            continue
        # Case A: vendor inline AFTER timestamp on same line
        #   "1:27 PM Post-Tension Services of Texas"
        rest = m.group("rest").strip()
        if rest and not _is_vendor_noise(rest):
            return rest
        # Case B: vendor on the line immediately BEFORE the timestamp
        #   "Ready Cable, Inc\n9:35 AM\n..."
        if i > 0:
            prev = lines[i - 1].strip()
            if prev and not _is_vendor_noise(prev):
                return prev
    return ""


def _find_qbo_statement_vendor(text: str) -> str:
    """Extract vendor from a QBO Statement (vendor-issued). Vendor appears
    prominently as a header line containing a corporate suffix. Skip our
    own header."""
    for line in text.splitlines():
        if re.search(r"\b(LLC|L\.L\.C\.|INC|CO\.|CORP|COMPANY)\b", line, re.I):
            cand = line.strip()
            if not _is_vendor_noise(cand):
                return cand
    return ""


def _paren_amount(raw: str) -> float:
    """Parse a money token that may be negative via a leading '-' OR parentheses.
    Credits/payments on statements print either way, and both must net out.
        '(1,483.03)' -> -1483.03 · '-491.73' -> -491.73 · '75.00' -> 75.00"""
    s = raw.strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "").replace("$", "").strip()
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return -v if neg else v


def parse_statement_qbo_customer_open_balance(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a QuickBooks Customer Open Balance report into the common shape."""
    vendor = _find_qbo_report_vendor(full_text)

    # Statement date — "As of <Month DD, YYYY>"
    stmt_date = ""
    m = QBO_AS_OF_RE.search(full_text)
    if m:
        raw = re.sub(r",", "", m.group(1))  # "March 31 2026"
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                stmt_date = dt.datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    # Grand total — "TOTAL <amount>" at end (not "Total <name> <amount>" — must be standalone TOTAL)
    amt_due = 0.0
    m = QBO_GRAND_TOTAL_RE.search(full_text)
    if m:
        try:
            amt_due = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # Line items — walk lines, tracking the current sub-customer for PO/job tagging
    lines: List[StmtLine] = []
    current_subcust = ""
    text_lines = full_text.splitlines()
    for line in text_lines:
        stripped = line.strip()
        if not stripped:
            continue
        # Transaction row FIRST (Invoice / Credit Memo / Payment / ...). Checked
        # before the sub-customer test because a Payment row has no due-date
        # column and would otherwise satisfy the caps-leading header pattern,
        # getting swallowed as a bogus sub-customer instead of counted.
        inv_m = QBO_CUSTOMER_OPEN_BAL_LINE_RE.match(line)
        if inv_m:
            memo = inv_m.group("memo").strip()
            # Truncated memos: QBO shows "..." when memo is cut off
            memo = memo.rstrip(".").strip() if memo.endswith("...") else memo
            lines.append(StmtLine(
                date=_norm_date(inv_m.group("date")),
                ref=inv_m.group("num").strip(),
                amount=_paren_amount(inv_m.group("amount")),
                po=current_subcust, address=memo))
            continue
        # Sub-customer header? (caps-leading line, not a txn/total/header)
        sub_m = QBO_SUBCUST_RE.match(line)
        if sub_m and not stripped.startswith(
                ("Invoice", "Credit", "Payment", "Total", "TOTAL", "Type", "Accrual", "Cash")):
            cand = sub_m.group(1).strip().rstrip(".").strip()
            # Filter out the parent customer header (our own company).
            # Keep sub-customers with the parent name + suffix ("... - Other", etc.).
            if cand.upper().rstrip(".").rstrip(",").strip() in (
                "PROFICIENT CONCRETE, LLC", "PROFICIENT CONCRETE LLC", "PROFICIENT CONCRETE"
            ):
                continue
            current_subcust = cand
            continue
    return vendor, stmt_date, amt_due, lines


def parse_statement_qbo_open_invoices(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a QuickBooks Open Invoices report into the common shape.
    4-column layout: Date | Num | Due Date | Open Balance. Grouped by
    sub-customer under the bill-to header ("Proficient Concrete")."""
    vendor = _find_qbo_report_vendor(full_text)

    # Statement date — "As of <Month DD, YYYY>" (same convention as Customer Open Balance)
    stmt_date = ""
    m = QBO_AS_OF_RE.search(full_text)
    if m:
        raw = re.sub(r",", "", m.group(1))
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                stmt_date = dt.datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    # Grand total
    amt_due = 0.0
    m = QBO_GRAND_TOTAL_RE.search(full_text)
    if m:
        try:
            amt_due = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # Line items
    lines: List[StmtLine] = []
    current_subcust = ""
    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        # Sub-customer header? (must NOT match an invoice-row pattern first)
        if QBO_OPEN_INV_LINE_RE.match(line):
            inv_m = QBO_OPEN_INV_LINE_RE.match(line)
            date_str = _norm_date(inv_m.group("date"))
            num = inv_m.group("num").strip()
            amount = float(inv_m.group("amount").replace(",", ""))
            lines.append(StmtLine(date=date_str, ref=num, amount=amount,
                                  po=current_subcust, address=""))
            continue
        sub_m = QBO_SUBCUST_RE.match(line)
        if sub_m and not stripped.startswith(("Invoice", "Total", "TOTAL", "Type", "Accrual", "Cash", "Page", "As ")):
            cand = sub_m.group(1).strip().rstrip(".").strip()
            # Skip the bill-to header (our own company, with or without LLC suffix)
            if cand.upper().rstrip(".").rstrip(",").strip() in (
                "PROFICIENT CONCRETE, LLC", "PROFICIENT CONCRETE LLC", "PROFICIENT CONCRETE"
            ):
                continue
            # Skip the column-header row
            if cand.upper().startswith("DATE NUM"):
                continue
            current_subcust = cand
            continue
    return vendor, stmt_date, amt_due, lines


# ── New vendor-statement templates (2026-07-01 batch) ─────────────────
# Each is keyed on a STRUCTURAL header signature, never the vendor name.
# All six were validated to tie their line-sum to the vendor's own stated
# total (Bobcat 19,395.99 · Bodin 169,039.44 · BURNCO 291,152.51 ·
# Cintas 389.08 · Cow Town 451,715.48 · Sunbelt 33,058.62).
BOBCAT_SIG  = re.compile(r"INVOICE NUMBER\s+INVOICE DATE\s+DUE DATE\s+PURCHASE ORDER\s+BALANCE", re.I)
BODIN_SIG   = re.compile(r"TYPE\s+REFERENCE\s+YARDAGE\s+CREDIT\s*/\s*DEBIT\s+BALANCE", re.I)
BURNCO_SIG  = re.compile(r"Delivery Address\s+PO Number\s+Type", re.I)
CINTAS_SIG  = re.compile(r"DATE\s+SOLD-TO\s+DESCRIPTION\s+REFERENCE\s+AMOUNT DUE\s+DUE DATE", re.I)
COWTOWN_SIG = re.compile(r"Inv\.\s*No\.\s+Inv\.\s*Date\s+Due Date", re.I)
SUNBELT_SIG = re.compile(r"DATE\s+INVOICE\s+JOB\s+DESCRIPTION\s+AMOUNT\s+DUE", re.I)

# ── Template: Croell Inc statement (added 2026-08-12) ─────────────────
# pdfplumber merges the left register and the right remittance stub onto one
# physical line, so each data row reads:
#   <date> <cd> <invoice> <description> <amount> <balance> <due date> \
#       <invoice-dup> <cd-dup> <amount-dup>
# We take the FIRST amount (the left "Amount" column) and ignore the duplicated
# remittance fields. Credits print in parentheses -> negative, so the line-sum
# nets to Balance Due. Cd codes seen: I=Invoice, F=Finance Charge.
# The doubled header is an unmistakable signature (checked before the generic
# tabular/columnar sigs, which the "Finance Charge" wording would otherwise trip).
CROELL_SIG = re.compile(
    r"Date\s+Cd\s+Invoice\s+Description\s+Amount\s+Balance\s+"
    r"Date\s+Due\s+Invoice\s+Cd\s+Amount", re.I)
CROELL_ROW_RE = re.compile(
    r"""^\s*
    (?P<date>\d{1,2}/\d{1,2}/\d{4})\s+
    (?P<cd>[A-Z])\s+
    (?P<num>\d+)\s+
    (?P<desc>.+?)\s+
    (?P<amount>\(?-?[\d,]+\.\d{2}\)?)\s+
    (?P<balance>\(?-?[\d,]+\.\d{2}\)?)\s+
    (?P<duedate>\d{1,2}/\d{1,2}/\d{4})\s+
    (?P<num2>\d+)\s+
    (?P<cd2>[A-Z])\s+
    (?P<amount2>\(?-?[\d,]+\.\d{2}\)?)\s*$""",
    re.VERBOSE | re.MULTILINE,
)


def _grab(pattern: str, text: str) -> str:
    m = re.search(pattern, text, re.I)
    return m.group(1) if m else ""


def _last_amt_after(label: str, text: str) -> float:
    """Last decimal amount on the block right after `label` (for aging-row totals)."""
    m = re.search(label + r"\s*[\r\n]+([$\d,. ]+)", text, re.I)
    if m:
        nums = re.findall(r"[\d,]+\.\d{2}", m.group(1))
        if nums:
            return float(nums[-1].replace(",", ""))
    return 0.0


def parse_statement_bobcat(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = _norm_date(_grab(r"\bDATE\s+(\d{1,2}/\d{1,2}/\d{2})\b", full_text))
    amt_due = _last_amt_after("TOTAL DUE", full_text)
    rx = re.compile(r"^(?P<ref>[0-9A-Z]{5,})\s+(?P<d>\d{1,2}/\d{1,2}/\d{2})\s+\d{1,2}/\d{1,2}/\d{2}\s+(?P<mid>.*?)\s*(?P<amt>[\d,]+\.\d{2})\s*$")
    lines: List[StmtLine] = []
    for ln in full_text.splitlines():
        m = rx.match(ln.strip())
        if not m:
            continue
        po = re.sub(r"\bAPPROVED\b", "", m["mid"], flags=re.I).strip()
        lines.append(StmtLine(date=_norm_date(m["d"]), ref=m["ref"],
                              amount=float(m["amt"].replace(",", "")), po=po))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_bodin(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = _norm_date(_grab(r"STATEMENT DATE[\s\S]{0,40}?(\d{1,2}/\d{1,2}/\d{4})", full_text))
    amt_due = 0.0
    m = re.search(r"THIS AMOUNT\s*[\r\n]+\s*([\d,]+\.\d{2})", full_text, re.I)
    if m:
        amt_due = float(m.group(1).replace(",", ""))
    rx = re.compile(r"^(?P<ref>\d+)\s+(?P<d>\d{1,2}/\d{1,2}/\d{4})\s+IN\s+(?P<mid>.*?)\s+[\d.]+\s*YDS\s+-?[\d,]+\.\d{2}\s+(?P<amt>-?[\d,]+\.\d{2})\s*$")
    lines: List[StmtLine] = []
    for ln in full_text.splitlines():
        m = rx.match(ln.strip())
        if not m:
            continue
        lines.append(StmtLine(date=_norm_date(m["d"]), ref=m["ref"],
                              amount=float(m["amt"].replace(",", "")), address=m["mid"].strip()))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_burnco(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = _norm_date(_grab(r"STATEMENT DATE:\s*(\d{2}-\d{2}-\d{4})", full_text).replace("-", "/"))
    amt_due = 0.0
    m = re.search(r"BALANCE DUE:\s*([\d,]+\.\d{2})", full_text, re.I)
    if m:
        amt_due = float(m.group(1).replace(",", ""))
    rx = re.compile(r"^(?P<d>\d{2}-\d{2}-\d{4})\s+(?P<ref>SA\d+)\s+(?P<mid>.*?)\s+(?:Invoice|Credit)\s+-?[\d,]+\.\d{2}\s+(?P<amt>-?[\d,]+\.\d{2})\s*$", re.I)
    lines: List[StmtLine] = []
    for ln in full_text.splitlines():
        m = rx.match(ln.strip())
        if not m:
            continue
        lines.append(StmtLine(date=_norm_date(m["d"].replace("-", "/")), ref=m["ref"],
                              amount=float(m["amt"].replace(",", "")), address=m["mid"].strip()))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_cintas(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = _norm_date(_grab(r"STATEMENT DATE\s+(\d{1,2}/\d{1,2}/\d{4})", full_text))
    amt_due = _last_amt_after("TOTAL DUE", full_text)
    rx = re.compile(r"^(?P<d>\d{1,2}/\d{1,2}/\d{4})\s+\d+\s+(?P<ref>\d+)\s+\$\s*(?P<amt>[\d,]+\.\d{2})\s+\d{1,2}/\d{1,2}/\d{4}\s*$")
    lines: List[StmtLine] = []
    for ln in full_text.splitlines():
        m = rx.match(ln.strip())
        if not m:
            continue
        lines.append(StmtLine(date=_norm_date(m["d"]), ref=m["ref"],
                              amount=float(m["amt"].replace(",", ""))))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_cowtown(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = ""
    m = re.search(r"([A-Z][a-z]+ \d{1,2}, \d{4})", full_text)
    if m:
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                stmt_date = dt.datetime.strptime(re.sub(r",", "", m.group(1)), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue
    amt_due = 0.0
    m = re.search(r"Past Due Amount:\s*\$?([\d,]+\.\d{2})", full_text, re.I)
    if m:
        amt_due = float(m.group(1).replace(",", ""))
    rx = re.compile(r"(?P<ref>\d{6})\s+(?P<d>\d{1,2}/\d{1,2}/\d{4})\s+\d{1,2}/\d{1,2}/\d{4}\s+\$(?P<inv>[\d,]+\.\d{2})\s+\$(?P<amt>[\d,]+\.\d{2})")
    lines: List[StmtLine] = []
    for m in rx.finditer(full_text):
        lines.append(StmtLine(date=_norm_date(m["d"]), ref=m["ref"],
                              amount=float(m["amt"].replace(",", ""))))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_sunbelt(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    stmt_date = _norm_date(_grab(r"[\r\n]\d{6,8}\s+(\d{1,2}/\d{1,2}/\d{2})\s+\d+", full_text))
    amt_due = 0.0
    m = re.search(r"TOTAL DUE\s*[\r\n]+\$?\s*([\d,]+\.\d{2})", full_text, re.I)
    if m:
        amt_due = float(m.group(1).replace(",", ""))
    rx = re.compile(r"^(?P<d>\d{1,2}/\d{1,2}/\d{2})\s+(?P<ref>\d{9}-\d{4})\s+(?P<desc>.*?)\s*(?P<amt>-?[\d,]+\.\d{2})\s*$")
    lines: List[StmtLine] = []
    for ln in full_text.splitlines():
        m = rx.match(ln.strip())
        if not m:
            continue
        lines.append(StmtLine(date=_norm_date(m["d"]), ref=m["ref"],
                              amount=float(m["amt"].replace(",", ""))))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return "", stmt_date, amt_due, lines


def parse_statement_croell(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a Croell Inc statement (doubled register/remittance layout).
    Takes the left 'Amount' column; credits in parentheses become negatives so
    the line-sum nets to Balance Due."""
    vendor = "Croell Inc"
    # The "Statement Date" header row is doubled and the value sits on the NEXT
    # line ("... Page\n08/08/2026 08/08/2026 ..."), so grab the first date there.
    stmt_date = _norm_date(
        _grab(r"Statement\s+Date[^\n]*\n\s*(\d{1,2}/\d{1,2}/\d{4})", full_text))

    # Balance Due = rightmost value on the aging footer — the last line that is
    # nothing but money tokens (Current / 1-30 / 31-60 / Over 60 / Bal Due).
    amt_due = 0.0
    money_line = re.compile(r"^\s*(?:\(?-?[\d,]+\.\d{2}\)?\s+){2,}\(?-?[\d,]+\.\d{2}\)?\s*$")
    for ln in reversed(full_text.splitlines()):
        if money_line.match(ln):
            amt_due = _paren_amount(re.findall(r"\(?-?[\d,]+\.\d{2}\)?", ln)[-1])
            break

    lines: List[StmtLine] = []
    for m in CROELL_ROW_RE.finditer(full_text):
        lines.append(StmtLine(date=_norm_date(m.group("date")), ref=m.group("num"),
                              amount=_paren_amount(m.group("amount")),
                              address=m.group("desc").strip()))
    if not amt_due:
        amt_due = round(sum(l.amount for l in lines), 2)
    return vendor, stmt_date, amt_due, lines


# ── Staple-vendor identity overrides ──────────────────────────────
# A few big, recurring vendors are impossible to identify from the generic
# body extraction alone: their name is either in a raster logo (no extractable
# text), absent from the letter body, or sits below OUR OWN bill-to name so the
# generic finder grabs "Proficient Concrete" instead. Left to the fuzzy
# filename/first-word fallback they collide with unrelated QBO names
# (BOB→'Bobby Tenison', COW→'COWBOY CHICKEN'). We hard-map any of a vendor's
# stable in-text markers to its EXACT QBO display name, applied across every
# template so it works regardless of which statement layout the vendor sends.
#
# Each entry: (exact QBO DisplayName, [lowercase marker substrings]).
# A marker must be specific enough that it can only mean this vendor. To add a
# staple: pick a token that always appears in its statements (an email domain,
# a letterhead street address, a distinctive brand spelling) and its QBO name.
_STAPLE_VENDORS: List[Tuple[str, List[str]]] = [
    # Bobcat is a multi-dealership brand (QBO has several "Bobcat of ..."); the
    # dealership logo is an image, but the code line ("Statement BOBNTXQ") and
    # credit-manager email ("...@bobcatntx.com") carry the region as text.
    ("Bobcat of North Texas", ["bobcatntx", "bobntx"]),
    # Cowtown Redi Mix: one statement variant is a past-due letter that only
    # shows its letterhead address; another is a QBO statement listing our own
    # name as bill-to. Match the address or the brand spelling.
    ("Cowtown Redi Mix Concrete", ["3400 bethlehem", "cowtown", "redi mix"]),
]


def _staple_vendor_override(full_text: str) -> str:
    """Return the exact QBO name for a known staple vendor if any of its
    markers appear in the statement text, else '' (no override)."""
    t = full_text.lower()
    for qbo_name, markers in _STAPLE_VENDORS:
        if any(marker in t for marker in markers):
            return qbo_name
    return ""


def parse_statement(path: Path) -> Tuple[str, str, float, List[StmtLine]]:
    """Returns (vendor_guess, stmt_date_YYYY-MM-DD, amount_due_total, lines).
    Dispatches by file extension first:
      • .xlsx/.xls/.xlsm → Excel parser
      • .png/.jpg/.jpeg/.tiff/.bmp/.heic → OCR to text, then run through PDF template detection
      • everything else → PDF text extraction + template detection
    For Excel/Image: vendor and stmt date may come back empty — caller fills
    them from --vendor / --stmt-date or the fallback chain in process_pdf.

    On the text path, a known staple-vendor marker (see _STAPLE_VENDORS)
    overrides whatever vendor the template parser extracted — the staple's exact
    QBO name is more reliable than generic body/filename extraction."""
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return parse_statement_excel(path)
    if ext in (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".heic", ".heif"):
        full_text = _image_to_text(path)
    else:
        full_text = _pdf_text(path)
    template = detect_template(full_text)
    if template == "qbo_customer_open_balance":
        result = parse_statement_qbo_customer_open_balance(full_text)
    elif template == "qbo_open_invoices":
        result = parse_statement_qbo_open_invoices(full_text)
    elif template == "qbo_statement":
        result = parse_statement_qbo_statement(full_text)
    elif template == "vendor_stmt_tabular":
        result = parse_statement_vendor_tabular(full_text)
    elif template == "vendor_stmt_columnar":
        # Columnar parser re-opens the PDF for positional extraction
        # (raw text loses row/column alignment in this layout)
        result = parse_statement_vendor_columnar(path)
    elif template == "vendor_stmt_whitecap":
        result = parse_statement_vendor_whitecap(full_text)
    elif template == "vendor_bobcat":
        result = parse_statement_bobcat(full_text)
    elif template == "vendor_bodin":
        result = parse_statement_bodin(full_text)
    elif template == "vendor_burnco":
        result = parse_statement_burnco(full_text)
    elif template == "vendor_cintas":
        result = parse_statement_cintas(full_text)
    elif template == "vendor_cowtown":
        result = parse_statement_cowtown(full_text)
    elif template == "vendor_sunbelt":
        result = parse_statement_sunbelt(full_text)
    elif template == "vendor_croell":
        result = parse_statement_croell(full_text)
    else:
        # No supported template detected — return empty so the caller surfaces
        # the unsupported-template error with the full list of supported formats.
        return "", "", 0.0, []
    # Staple-vendor identity wins over the parser's extracted vendor.
    override = _staple_vendor_override(full_text)
    if override:
        _vendor, stmt_date, amt_due, lines = result
        result = (override, stmt_date, amt_due, lines)
    return result


def _scan_excel_for_vendor(xlsx_path: Path) -> str:
    """Scan the first 5 rows × 5 columns of an Excel for text that looks
    like a vendor name. Skips header keywords, dates, numeric cells, our own
    company. Prefers cells with corporate suffix tokens (LLC/INC/CO/etc.) or
    industry-typical words. Returns first matching candidate or ''."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        return ""

    HEADER_KW = ("invoice", "inv date", "inv #", "due date", "balance",
                 "amount", "ref #", "ref#", "statement date", "page",
                 "total", "type", "date", "memo", "open", "as of")
    candidates: List[str] = []
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=5, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            if not text or len(text) < 4:
                continue
            text_low = text.lower()
            if any(kw in text_low for kw in HEADER_KW):
                continue
            if re.match(r"^\d{1,4}[/.\-]\d{1,2}", text):   # MM/DD or YYYY-MM (date-ish)
                continue
            if re.match(r"^\$?[\d,]+\.?\d*$", text):       # money/numeric
                continue
            if "proficient" in text_low:                    # our own company
                continue
            candidates.append(text)
    # Try wb.close() if read_only mode allocated handles
    try: wb.close()
    except Exception: pass

    # Prefer cells with strong vendor-name signals
    STRONG = re.compile(
        r"\b(LLC|L\.L\.C\.|INC|CO\.|CORP|COMPANY|LTD|MATERIALS|SERVICES|"
        r"MIX|CABLE|TENSION|CONCRETE|SUPPLY|READY|EQUIPMENT|HOLDINGS)\b",
        re.I,
    )
    for c in candidates:
        if STRONG.search(c):
            return c
    return candidates[0] if candidates else ""


def _vendor_from_filename(path: Path) -> str:
    """Derive a vendor-name hint from a filename. Strips dates (slash/dash/
    underscore-separated, month names, 4-digit years), generic words like
    'statement'/'invoice', and separators. Returns '' if nothing left."""
    stem = path.stem
    # Normalize separators to spaces first
    stem = re.sub(r"[_\-.]+", " ", stem)
    # Date patterns (now space-separated after normalization)
    stem = re.sub(r"\b\d{4}\s+\d{1,2}\s+\d{1,2}\b", " ", stem)   # YYYY MM DD
    stem = re.sub(r"\b\d{1,2}\s+\d{1,2}\s+\d{2,4}\b", " ", stem) # MM DD YYYY
    stem = re.sub(r"\b\d{4}\s+\d{1,2}\b", " ", stem)             # YYYY MM
    stem = re.sub(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", " ", stem)     # MM/DD/YY (original slashes)
    stem = re.sub(r"\b(?:19|20)\d{2}\b", " ", stem)              # standalone YYYY
    # Month names (full + abbreviated)
    stem = re.sub(r"\b(january|february|march|april|may|june|july|august|"
                  r"september|october|november|december|jan|feb|mar|apr|jun|"
                  r"jul|aug|sept?|oct|nov|dec)\b", " ", stem, flags=re.I)
    # Generic statement-related words
    stem = re.sub(r"\b(statement|stmt|invoice|invoices|bill|bills|report|open|"
                  r"aging|monthly|thru|through|as\s+of|past\s+due)\b",
                  " ", stem, flags=re.I)
    # Collapse whitespace
    stem = re.sub(r"\s+", " ", stem).strip()
    # Strip any remaining standalone leading/trailing digit tokens
    stem = re.sub(r"^(\d+\s+)+", "", stem)
    stem = re.sub(r"(\s+\d+)+$", "", stem)
    stem = stem.strip()
    # If what's left has no real letters (e.g., just orphan digits), return ''
    if not re.search(r"[A-Za-z]{2,}", stem):
        return ""
    return stem


def _excel_date_to_str(val) -> str:
    """Convert any Excel-readable date cell to YYYY-MM-DD, or '' if unparseable."""
    if val is None:
        return ""
    if isinstance(val, dt.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, dt.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        # Excel serial date (days since 1899-12-30)
        try:
            base = dt.datetime(1899, 12, 30)
            return (base + dt.timedelta(days=int(val))).strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            return ""
    if isinstance(val, str):
        return _norm_date(val.strip())
    return ""


def _xls_to_xlsx_temp(xls_path: Path) -> Path:
    """Convert a legacy .xls (Excel 97-2003) workbook to a temp .xlsx file
    and return the new path. openpyxl can't read .xls, so we use xlrd<2.0
    to read and rewrite as .xlsx. The temp file is created in the OS temp
    dir; macOS cleans these up automatically. Date cells (xlrd stores them
    as floats with a datemode flag) are converted to real datetime objects
    so the downstream parser sees them as dates not numbers.
    Lazy-imports xlrd so users who never see .xls files don't need it."""
    import tempfile
    try:
        import xlrd
    except ImportError:
        sys.exit(_Term.color(_Term.R,
            "✗ Legacy .xls support needs xlrd. Install once into the venv:\n"
            f"    cd '{Path(__file__).resolve().parent.parent / 'bill-tracker'}'\n"
            "    .venv/bin/python -m pip install 'xlrd<2'\n"
            "  (Pin to <2.0 — xlrd 2.0+ dropped .xls support for security reasons.)"))
    from openpyxl import Workbook

    try:
        xls = xlrd.open_workbook(str(xls_path))
    except Exception as e:
        sys.exit(_Term.color(_Term.R, f"✗ couldn't open .xls file: {e}"))

    sheet = xls.sheet_by_index(0)
    new_wb = Workbook()
    new_ws = new_wb.active
    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            v = cell.value
            # xlrd encodes dates as floats — convert back to datetime
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    v = xlrd.xldate.xldate_as_datetime(v, xls.datemode)
                except (xlrd.XLDateError, ValueError):
                    pass
            new_ws.cell(row=r + 1, column=c + 1, value=v)
    tmp = Path(tempfile.gettempdir()) / f"_reconcile_{xls_path.stem}.xlsx"
    new_wb.save(tmp)
    return tmp


def _find_excel_as_of_date(ws) -> str:
    """Scan the top 10 rows for an 'As of <Month DD, YYYY>' label and return
    the parsed YYYY-MM-DD. Same convention as the QBO PDF templates."""
    AS_OF = re.compile(r"As\s+of\s+([A-Za-z]+\s+\d{1,2},?\s+\d{4})", re.I)
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            m = AS_OF.search(cell)
            if not m:
                continue
            raw = re.sub(r",", "", m.group(1)).strip()
            for fmt in ("%B %d %Y", "%b %d %Y"):
                try:
                    return dt.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
    return ""


def _find_excel_grand_total(ws) -> float:
    """Look for a 'Total ...' row (e.g., 'Total - Cust00000058 ...') and grab
    its rightmost positive numeric — that's the grand total for tie-out.
    Uses the LAST matching 'total' row: on a multi-customer A/R aging export
    the per-customer subtotals are also labeled 'Total ...' and appear first,
    so taking the first one would return a subtotal. The grand total is last.
    Single-vendor statements have exactly one 'total' row, so this is a no-op
    for them."""
    grand = 0.0
    for row in ws.iter_rows(values_only=False):
        first_text = ""
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                first_text = cell.value.strip().lower()
                break
        if first_text.startswith("total"):
            nums = [c.value for c in row if isinstance(c.value, (int, float)) and c.value > 0]
            if nums:
                grand = float(max(nums))
    return grand


def _parse_excel_with_headers(ws) -> Tuple[List["StmtLine"], float]:
    """Header-based parser: scan first 10 rows for a header row with named
    columns (Invoice # + Original Invoice Amount, etc.). Map columns by name."""
    header_row = None
    col_map: Dict[str, int] = {}
    for row_idx in range(1, min(11, ws.max_row + 1)):
        mapping: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=c).value
            if not v:
                continue
            vs = str(v).lower().strip()
            if "inv date" == vs or "invoice date" in vs:
                mapping.setdefault("date", c)
            elif "invoice #" in vs or vs in ("inv #", "invoice number", "invoice no", "invoice #"):
                mapping.setdefault("ref", c)
            elif "original invoice amount" in vs or vs == "original amount":
                mapping["amount"] = c
            elif vs == "amount" and "amount" not in mapping:
                mapping["amount"] = c
            elif vs in ("balance", "open balance", "open bal"):
                mapping.setdefault("balance", c)
            elif "due date" in vs or vs == "due":
                mapping.setdefault("due_date", c)
        if "ref" in mapping and "amount" in mapping:
            header_row = row_idx
            col_map = mapping
            break

    if header_row is None:
        return [], 0.0

    lines: List[StmtLine] = []
    total = 0.0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ref_v = ws.cell(row=row_idx, column=col_map["ref"]).value
        amount_v = ws.cell(row=row_idx, column=col_map["amount"]).value
        if ref_v is None or amount_v is None:
            continue
        ref = str(ref_v).strip()
        if not ref:
            continue
        try:
            amount = float(amount_v)
        except (ValueError, TypeError):
            continue
        date_v = ws.cell(row=row_idx, column=col_map["date"]).value if "date" in col_map else None
        lines.append(StmtLine(date=_excel_date_to_str(date_v), ref=ref, amount=amount, po="", address=""))
        total += amount
    return lines, round(total, 2)


def _parse_excel_content_pattern(ws) -> Tuple[List["StmtLine"], float]:
    """Header-less parser for layouts like A/R Aging Detail reports.

    Identifies an invoice row by content alone:
      • at least one date cell (datetime or MM/DD/YY-style string)
      • at least one ref-like token (letters + digits like INV38014 / CV6415,
        or pure digits ≥3 chars) — taken from the FIRST whitespace-delimited
        token of any text cell
      • a rightmost positive numeric cell → treated as the amount

    Rows missing any of these (headers, totals, blank rows, customer-header
    rows like 'Cust00000058 Proficient Concrete LLC') are skipped."""
    REF_TOKEN = re.compile(r"^[A-Za-z]*\d{3,}$")
    DATE_STR  = re.compile(r"^\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}$")
    lines: List[StmtLine] = []
    total = 0.0
    for row in ws.iter_rows(values_only=False):
        date_val = None
        ref_val = None
        last_numeric_val = None
        last_numeric_col = -1
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, (dt.datetime, dt.date)):
                if date_val is None:
                    date_val = v
                continue
            if isinstance(v, str):
                vs = v.strip()
                if not vs:
                    continue
                if DATE_STR.match(vs) and date_val is None:
                    date_val = vs
                    continue
                if ref_val is None:
                    first_token = vs.split()[0] if vs.split() else ""
                    if REF_TOKEN.match(first_token):
                        ref_val = first_token
                continue
            if isinstance(v, (int, float)) and v > 0:
                if cell.column > last_numeric_col:
                    last_numeric_val = float(v)
                    last_numeric_col = cell.column
        if date_val and ref_val and last_numeric_val is not None:
            lines.append(StmtLine(
                date=_excel_date_to_str(date_val),
                ref=ref_val,
                amount=last_numeric_val,
                po="", address="",
            ))
            total += last_numeric_val
    return lines, round(total, 2)


def parse_statement_excel(xlsx_path: Path) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse an Excel statement. Two-pass detection:
       1. Header-based — looks for 'Invoice #' + 'Original Invoice Amount' (or
          similar) on a header row, then reads data rows by column position.
       2. Content-pattern (fallback) — header-less layouts like A/R Aging
          Detail reports. Detects invoice rows by content (date + ref# +
          rightmost numeric).

    Accepts both modern .xlsx and legacy .xls (Excel 97-2003). .xls files
    are transparently converted to a temp .xlsx first via xlrd.

    Statement date is auto-extracted from 'As of <Month DD, YYYY>' if present
    in the top rows. Grand total is taken from a 'Total ...' row when found,
    otherwise it's the sum of parsed line amounts.

    Vendor name is NOT extracted here — that's handled by the chain in
    process_pdf (--vendor flag → cached alias → Excel cell scan → cleaned
    filename → interactive prompt)."""
    # Legacy .xls → convert to temp .xlsx first so the rest of the code can
    # use openpyxl uniformly. _xls_to_xlsx_temp handles the conversion + date
    # cell normalization; openpyxl-only path stays unchanged.
    if xlsx_path.suffix.lower() == ".xls":
        xlsx_path = _xls_to_xlsx_temp(xlsx_path)
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    stmt_date = _find_excel_as_of_date(ws)

    # 1. Try header-based parser
    lines, total = _parse_excel_with_headers(ws)
    # 2. Fall through to content-pattern parser if no lines extracted
    if not lines:
        lines, total = _parse_excel_content_pattern(ws)

    # Prefer the "Total ..." row's amount if found — that's the report's own
    # grand total (more reliable than summing if any line was missed).
    grand_total = _find_excel_grand_total(ws)
    if grand_total > 0:
        total = grand_total

    return "", stmt_date, round(total, 2), lines


def parse_statement_vendor_tabular(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a tabular vendor statement (CMC-style). One line per invoice:
       <date> <ref> <due-date> <amount> [optional payment fields] <balance>
    Sub-customer headers (SHIP TO XYZ) and aging rows are skipped."""
    # Vendor — first non-empty line at the top usually has the company name.
    # CMC-style PDFs concatenate vendor name + "STATEMENT" on one line
    # ("CMC Construction Services STATEMENT"); strip the trailing word.
    vendor = ""
    for line in full_text.splitlines()[:5]:
        cand = line.strip()
        if not cand:
            continue
        # Strip trailing "STATEMENT" (case-insensitive) — common artifact
        cand = re.sub(r"\s+STATEMENT\s*$", "", cand, flags=re.I).strip()
        if not cand:
            continue
        if _is_vendor_noise(cand):
            continue
        # Skip address/state lines and phone numbers
        if re.match(r"^[\d\-\(\) ]+$", cand):
            continue
        vendor = cand
        break

    # Statement date — try labeled patterns first
    stmt_date = ""
    for pat in STMT_DATE_PATTERNS:
        m = pat.search(full_text)
        if m:
            stmt_date = _norm_date(m.group(1))
            if stmt_date:
                break

    # Amount due — look for "Total Due" label, then take the LAST dollar value
    # within the next ~200 chars (aging row format: labels on one line, totals
    # on next line; Total Due is the rightmost column).
    amt_due = 0.0
    m = re.search(r"Total\s+Due([\s\S]{0,200})", full_text, re.I)
    if m:
        nums = re.findall(r"[\d,]+\.\d{2}", m.group(1))
        if nums:
            try:
                amt_due = float(nums[-1].replace(",", ""))
            except ValueError:
                pass
    # Fallback to AMT_DUE_RE
    if amt_due == 0.0:
        m = AMT_DUE_RE.search(full_text)
        if m:
            try:
                amt_due = float(m.group(1).replace(",", ""))
            except ValueError:
                pass

    # Line items
    lines: List[StmtLine] = []
    current_subcust = ""
    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        # Track sub-customer headers like "SHIP TO BRIW BRIARWOOD"
        if VENDOR_STMT_SHIPTO_RE.match(line):
            # Pull the description after "SHIP TO" / "SHIP-TO"
            sub = re.sub(r"^\s*SHIP[- ]TO\s*", "", stripped, flags=re.I).strip()
            current_subcust = sub
            continue
        # Skip "SHIP-TO TOTAL" lines
        if re.match(r"^\s*SHIP[- ]TO\s+TOTAL\b", line, re.I):
            continue
        # Skip aging row (Current 31-60 ...)
        if re.match(r"^\s*Current\s+\d|^\s*\d+\.\d+\s+\d+\.\d+\s+\d+\.\d+", line):
            continue
        # Try to match an invoice row
        m = VENDOR_STMT_TABULAR_LINE_RE.match(line)
        if not m:
            continue
        try:
            amount = float(m.group("amount").replace(",", ""))
            balance = float(m.group("balance").replace(",", ""))
        except (ValueError, TypeError):
            continue
        # Use BALANCE as the open amount (more accurate than original amount —
        # accounts for partial payments). For unpaid bills balance == amount.
        # Skip zero-balance rows (fully paid).
        if balance <= 0:
            continue
        lines.append(StmtLine(
            date=_norm_date(m.group("date")),
            ref=m.group("ref"),
            amount=balance,
            po=current_subcust,
            address="",
        ))
    return vendor, stmt_date, amt_due, lines


def parse_statement_vendor_columnar(pdf_path: Path) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a columnar vendor statement (Preferred Materials / Sunrise style).
    These have a visible table grid but pdfplumber's default extraction
    flattens each column into a vertical stack. We re-extract with positional
    coordinates and group words by Y (row) — then map by X (column header)."""
    import pdfplumber
    lines: List[StmtLine] = []
    vendor = ""
    stmt_date = ""
    amt_due_text = ""

    with pdfplumber.open(pdf_path) as pdf:
        all_words: List[dict] = []
        for page_idx, page in enumerate(pdf.pages):
            page_words = page.extract_words(use_text_flow=False)
            # Tag each word with which page it came from (for Y-deduplication across pages)
            for w in page_words:
                w["page"] = page_idx
                all_words.append(w)

        # Vendor + stmt date from page-1 text (still useful since headers
        # come through as roughly-sequential lines). Strip trailing
        # "Date MM/DD/YYYY" / "STATEMENT" artifacts that get concatenated
        # onto the vendor name by pdfplumber's flat text extraction.
        page1_text = pdf.pages[0].extract_text() or ""
        for line in page1_text.splitlines()[:8]:
            cand = line.strip()
            # Strip trailing "Date MM/DD/YYYY" suffix
            cand = re.sub(r"\s+Date\s+\d{1,2}/\d{1,2}/\d{2,4}\s*$", "", cand, flags=re.I).strip()
            # Strip trailing STATEMENT word
            cand = re.sub(r"\s+STATEMENT\s*$", "", cand, flags=re.I).strip()
            if cand and not _is_vendor_noise(cand) and len(cand) > 3:
                vendor = cand
                break
        # Statement date — usually "Date MM/DD/YYYY"
        m = re.search(r"\bDate\s+(\d{1,2}/\d{1,2}/\d{2,4})", page1_text)
        if m:
            stmt_date = _norm_date(m.group(1))
        # Amount Due
        m = re.search(r"Amount\s+Due\s*\$?\s*([\d,]+\.\d{2})", page1_text, re.I)
        if m:
            try:
                amt_due_text = m.group(1).replace(",", "")
            except Exception:
                pass

    # Group all words across all pages into rows by (page, y-coordinate)
    # Y proximity threshold: 3 pixels (rows are usually 12-15px apart)
    rows_by_pos: Dict[Tuple[int, int], List[dict]] = {}
    for w in all_words:
        key = (w["page"], round(w["top"] / 3))  # bucket by 3-pixel y-band
        rows_by_pos.setdefault(key, []).append(w)

    # For each row, sort words left-to-right and extract:
    #   • First date (MM/DD/YYYY)
    #   • Description containing "Invoice #..." or similar ref pattern
    #   • Rightmost numeric (balance) — but we want CHARGE column not balance
    #   • Skip rows without a date
    # Require explicit "Invoice" keyword so Payment / Credit Memo / Balance
    # Forward rows aren't matched as invoices (they have their own descriptors
    # and would otherwise be picked up by a bare `#` alternation).
    REF_IN_DESC = re.compile(r"\bInvoice\s*#?\s*(\w+\d+)", re.I)
    AMOUNT_RE_LINE = re.compile(r"^-?[\d,]+\.\d{2}$")
    DATE_TOKEN = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$")
    total_charge = 0.0
    for key, words in sorted(rows_by_pos.items()):
        words.sort(key=lambda w: w["x0"])
        tokens = [w["text"] for w in words]
        # Find date as FIRST token matching MM/DD/YYYY
        date_token = next((t for t in tokens if DATE_TOKEN.match(t)), None)
        if not date_token:
            continue
        # Find ref# — look for "Invoice #XXX" patterns in concatenated tokens
        row_text = " ".join(tokens)
        ref_m = REF_IN_DESC.search(row_text)
        if not ref_m:
            continue
        # Find all numeric values on the row
        nums = [t for t in tokens if AMOUNT_RE_LINE.match(t.lstrip("-"))]
        if not nums:
            continue
        # Take the LEFTMOST positive numeric as the CHARGE (the original invoice amount).
        # Avoids picking running-balance which would over-count.
        charge_val = None
        for n in nums:
            try:
                v = float(n.replace(",", ""))
            except ValueError:
                continue
            if v > 0:
                charge_val = v
                break
        if charge_val is None or charge_val <= 0:
            continue
        lines.append(StmtLine(
            date=_norm_date(date_token),
            ref=ref_m.group(1),
            amount=charge_val,
            po="",
            address="",
        ))
        total_charge += charge_val

    amt_due = 0.0
    if amt_due_text:
        try:
            amt_due = float(amt_due_text)
        except ValueError:
            pass
    if amt_due == 0.0:
        amt_due = round(total_charge, 2)

    return vendor, stmt_date, amt_due, lines


def parse_statement_vendor_whitecap(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a White Cap / Billtrust statement. pdfplumber merges the left
    (statement) and right (remittance advice) copies onto a single line per
    row, so each data row in the extracted text looks like:
        <date> <transaction_no> <type> <original> <balance> <ref-dup> <po> <balance-dup>
    We capture the first 5 fields and ignore the duplicate trailing ones.

    Type codes: I=Invoice, C=Credit Memo, R=Rental, D=Debit Memo,
    U=Unapplied Payment, * suffix = In Review. We include I/R/D rows (real
    bills), skip C and U (vendor accounting adjustments, not real bills).
    Uses BALANCE DUE as the amount (already accounts for partial payments)."""
    vendor = "White Cap, L.P."

    # Statement date — "CLOSING DATE" label, then date on a nearby line
    stmt_date = ""
    m = re.search(r"CLOSING\s+DATE[\s\S]{0,40}?(\d{1,2}/\d{1,2}/\d{2,4})",
                  full_text, re.I)
    if m:
        stmt_date = _norm_date(m.group(1))

    # Total Due
    amt_due = 0.0
    m = re.search(r"TOTAL\s+DUE[\s\S]{0,40}?\$?([\d,]+\.\d{2})", full_text, re.I)
    if m:
        try:
            amt_due = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # Only skip U (Unapplied Payment — vendor accounting, no invoice ref to
    # match). Credit Memos (C) are kept as negative-amount lines so the sum
    # ties to Total Due — they're real items the vendor expects to net out,
    # and if QBO has a matching credit memo Bill with negative balance, the
    # reconciler will match it; if not, it'll surface as MISSING_IN_QBO so AP
    # knows to enter the credit.
    SKIP_TYPES = {"U"}
    lines: List[StmtLine] = []
    for m in WHITECAP_ROW_RE.finditer(full_text):
        tp = m.group("tp").rstrip("*")  # strip "*" in-review marker
        if tp in SKIP_TYPES:
            continue
        try:
            balance = float(m.group("balance").replace(",", ""))
        except (ValueError, TypeError):
            continue
        if balance == 0:
            continue
        lines.append(StmtLine(
            date=_norm_date(m.group("date")),
            ref=m.group("ref"),
            amount=balance,
            po="",
            address="",
        ))

    return vendor, stmt_date, amt_due, lines


def parse_statement_qbo_statement(full_text: str) -> Tuple[str, str, float, List[StmtLine]]:
    """Parse a QuickBooks Statement (vendor-issued statement with INV #N. Due N lines)."""
    vendor = _find_qbo_statement_vendor(full_text)

    # Statement date — try labeled patterns first, then standalone fallback
    stmt_date = ""
    for pat in STMT_DATE_PATTERNS:
        m = pat.search(full_text)
        if m:
            stmt_date = _norm_date(m.group(1))
            if stmt_date:
                break

    # Amount due — first match after "Amount Due" label
    amt_due = 0.0
    m = AMT_DUE_RE.search(full_text)
    if m:
        try:
            amt_due = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # Lines — match against full text (multiline)
    lines: List[StmtLine] = []
    for m in STMT_LINE_RE.finditer(full_text):
        date = _norm_date(m.group("date"))
        ref = m.group("ref")
        amount = float(m.group("amount").replace(",", ""))
        # Pull PO and address from the matched line+surrounding chars
        line_blob = full_text[max(0, m.start() - 20): m.end() + 200]
        po_m = PO_RE.search(line_blob)
        po = po_m.group(1).strip() if po_m else ""
        # Address: text between "Orig. Amount $X.XX." and the start of the
        # NEXT line (or end of table). PDF tables interject the line $amount
        # and running balance, and wrap long addresses across rows — both
        # need stripping.
        addr = ""
        orig_m = ORIG_AMT_RE.search(line_blob)
        if orig_m:
            tail = line_blob[orig_m.end():]
            # Cut at the next line boundary — whichever comes first:
            #   • next date (MM/DD/YYYY)
            #   • next "INV #..." marker
            #   • aging-bucket footer strings ("DAYS PAST", "CURRENT", "Amount Due")
            cutoff_patterns = [
                r"\d{1,2}/\d{1,2}/\d{2,4}",   # next line's date
                r"INV\s*\#",                  # next line's invoice marker
                r"\d{1,2}-\d{1,2}\b",         # aging range like "1-30", "31-60"
                r"DAYS\s+PAST",               # aging footer
                r"\bCURRENT\b",
                r"Amount\s+Due",
                r"OVER\s+\d+\s+DAYS",         # "OVER 90 DAYS PAST DUE"
            ]
            cutoff_re = re.compile("|".join(cutoff_patterns), re.I)
            m2 = cutoff_re.search(tail)
            if m2:
                tail = tail[:m2.start()]
            # Strip the "$N,NNN.NN $N,NNN.NN" pair (line amt + running balance)
            tail = re.sub(r"\s*[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s*", " ", tail)
            # Strip leading period+space left over from "Orig. Amount $X.XX. <addr>"
            tail = re.sub(r"^[\s.]+", "", tail)
            # Collapse newlines + multi-space to single space
            tail = re.sub(r"\s+", " ", tail).strip()
            # Truncate cleanly — first 50 chars, cut at last word boundary
            if len(tail) > 50:
                tail = tail[:50].rsplit(" ", 1)[0]
            addr = tail
        lines.append(StmtLine(date=date, ref=ref, amount=amount, po=po, address=addr))

    return vendor, stmt_date, amt_due, lines


def _norm_date(s: str) -> str:
    """MM/DD/YYYY or M/D/YY → YYYY-MM-DD."""
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s

# ───────────────────────── QBO auth + queries ─────────────────────────

def load_credentials() -> Tuple[str, str]:
    """Returns (access_token, company_id). Mirrors qbo_bill_tracker.py."""
    if not kc.has_credentials():
        sys.exit("✗ no credentials. Run: python3 setup_qbo.py")
    creds = kc.get_all()
    required = ["QBO_CLIENT_ID", "QBO_CLIENT_SECRET", "QBO_COMPANY_ID", "QBO_REFRESH_TOKEN"]
    if any(not creds.get(k) for k in required):
        sys.exit("✗ blob incomplete. Run: python3 setup_qbo.py")
    basic = base64.b64encode(
        f"{creds['QBO_CLIENT_ID']}:{creds['QBO_CLIENT_SECRET']}".encode()
    ).decode()
    r = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        headers={"Authorization": f"Basic {basic}",
                 "Content-Type": "application/x-www-form-urlencoded",
                 "Accept": "application/json"},
        data={"grant_type": "refresh_token",
              "refresh_token": creds["QBO_REFRESH_TOKEN"]},
        timeout=30,
    )
    if r.status_code != 200:
        sys.exit(f"✗ token refresh {r.status_code}: {r.text[:300]}")
    body = r.json()
    new_rt = body.get("refresh_token")
    if new_rt and new_rt != creds["QBO_REFRESH_TOKEN"]:
        try:
            kc.put("QBO_REFRESH_TOKEN", new_rt)
        except kc.SecretsError:
            pass
    return body["access_token"], creds["QBO_COMPANY_ID"]


def _api_get(access: str, path: str, params: Optional[dict] = None) -> dict:
    p = dict(params or {})
    p["minorversion"] = MINOR_VERSION
    r = requests.get(f"{API_BASE}{path}",
                     headers={"Authorization": f"Bearer {access}",
                              "Accept": "application/json"},
                     params=p, timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"GET {path} → {r.status_code}: {r.text[:300]}")
    return r.json()


def query(access: str, cid: str, q: str) -> dict:
    return _api_get(access, f"/v3/company/{cid}/query", {"query": q})


def query_all(access: str, cid: str, entity: str, where: str = "") -> List[dict]:
    out: List[dict] = []
    start = 1
    page = 500
    while True:
        q = f"SELECT * FROM {entity}"
        if where:
            q += f" WHERE {where}"
        q += f" STARTPOSITION {start} MAXRESULTS {page}"
        data = query(access, cid, q)
        batch = data.get("QueryResponse", {}).get(entity, [])
        if not batch:
            break
        out.extend(batch)
        if len(batch) < page:
            break
        start += page
    return out


# ───────────────────────── vendor alias cache ─────────────────────────

def _alias_key(s: str) -> str:
    """Normalize a vendor string for cache lookup. Preserves visual identity
    but ignores trailing whitespace/punctuation noise that varies by PDF run."""
    return re.sub(r"[\s\.,;]+$", "", (s or "").strip()).lower()


def load_aliases() -> Dict[str, dict]:
    """Return {alias_key: {'pdf_name', 'qbo_id', 'qbo_name', 'saved'}} or {} if missing/corrupt."""
    if not ALIAS_FILE.exists():
        return {}
    try:
        data = json.loads(ALIAS_FILE.read_text())
        if not isinstance(data, dict):
            return {}
        return data
    except (json.JSONDecodeError, OSError) as e:
        _warn(f"vendor_aliases.json unreadable ({e}); ignoring cache.")
        return {}


def save_aliases(aliases: Dict[str, dict]) -> None:
    try:
        ALIAS_FILE.write_text(json.dumps(aliases, indent=2, sort_keys=True))
    except OSError as e:
        _warn(f"could not save vendor_aliases.json ({e}); alias not persisted.")


def remember_vendor(pdf_name: str, qbo_id: str, qbo_name: str) -> None:
    aliases = load_aliases()
    key = _alias_key(pdf_name)
    aliases[key] = {
        "pdf_name": pdf_name,
        "qbo_id": qbo_id,
        "qbo_name": qbo_name,
        "saved": dt.date.today().isoformat(),
    }
    save_aliases(aliases)


def forget_vendor(needle: str) -> bool:
    """Remove one alias. Matches by alias_key OR by qbo_name (case-insensitive)."""
    aliases = load_aliases()
    target = _alias_key(needle)
    to_remove = [k for k, v in aliases.items()
                 if k == target or v.get("qbo_name", "").lower() == needle.lower()]
    for k in to_remove:
        del aliases[k]
    if to_remove:
        save_aliases(aliases)
    return bool(to_remove)


def find_alias_by_filename(filename: str) -> Optional[Tuple[str, dict, int, str]]:
    """Fuzzy-match a filename against cached aliases by token overlap.

    Used for Excel statements (where vendor isn't in the body): if the user
    names files like 'Ferguson_2026-03.xlsx' or 'Estrada March.xlsx', we can
    pick up the vendor from a previously-confirmed alias.

    Returns (alias_key, alias_data, score, matched_tokens_str) for the best
    confident match, or None if there's no clear winner (no matches, or a tie
    where we can't pick safely).

    Scoring: each alias-name token (>=3 chars) found in the normalized
    filename = +1. The first token (usually most distinctive — vendor's
    primary surname) earns a +1 bonus. A "clear winner" is the highest score
    AND strictly above the runner-up; otherwise we abstain."""
    aliases = load_aliases()
    if not aliases:
        return None

    # Normalize filename: strip extension, strip common date patterns,
    # collapse separators to spaces, lowercase.
    stem = Path(filename).stem
    stem = re.sub(r"\b\d{1,2}[/.\-_]\d{1,2}([/.\-_]\d{2,4})?\b", " ", stem)  # MM/DD or MM-DD-YYYY
    stem = re.sub(r"\b\d{4}[/.\-_]?\d{1,2}([/.\-_]?\d{1,2})?\b", " ", stem)  # YYYY-MM-DD
    stem = re.sub(r"[_\-.]+", " ", stem)
    fn_norm = stem.lower()

    scored: List[Tuple[int, str, dict, List[str]]] = []
    for key, data in aliases.items():
        # Prefer the QBO display name (canonical) over the PDF-extracted name.
        name = (data.get("qbo_name") or data.get("pdf_name") or "").lower()
        if not name:
            continue
        # Tokens >=3 chars from the alias name. Skip corporate-suffix noise
        # so 'LLC'/'INC'/'CO' don't accidentally match on filenames like
        # 'Inc_2026.xlsx' (which would never be a real vendor file).
        tokens = [t for t in re.split(r"[^a-z0-9]+", name)
                  if len(t) >= 3 and t not in {"llc", "inc", "corp", "company", "ltd", "the"}]
        if not tokens:
            continue
        matched = [t for t in tokens if t in fn_norm]
        if not matched:
            continue
        score = len(matched)
        if tokens[0] in fn_norm:
            score += 1  # first-token bonus (usually most distinctive)
        scored.append((score, key, data, matched))

    if not scored:
        return None
    scored.sort(reverse=True, key=lambda x: (x[0], -len(x[1])))  # higher score wins; tie → prefer longer key
    best_score, best_key, best_data, best_matched = scored[0]
    # Clear winner only if strictly higher than runner-up (avoid ambiguous matches).
    if len(scored) == 1 or scored[0][0] > scored[1][0]:
        return (best_key, best_data, best_score, ", ".join(best_matched))
    return None


def validate_cached_vendor(access: str, cid: str, qbo_id: str) -> Optional[str]:
    """Cheap SELECT-by-Id check. Returns current DisplayName if vendor still exists, else None."""
    try:
        rows = query(access, cid,
                     f"SELECT Id, DisplayName FROM Vendor WHERE Id = '{qbo_id}'"
                     ).get("QueryResponse", {}).get("Vendor", [])
        if rows:
            return rows[0].get("DisplayName", "")
    except RuntimeError:
        return None
    return None


def find_vendor_id_cached(access: str, cid: str, pdf_vendor: str,
                          override: str = "") -> Tuple[str, str, bool]:
    """Resolve vendor → (id, current_display_name, from_cache).
    - If --vendor override is set, use it directly (no cache lookup).
    - Else check alias cache; validate via QBO; on hit return (id, name, True).
    - On miss/invalid, fall back to LIKE search and return (id, name, False).
    Caller is responsible for saving the alias after user confirmation.
    """
    if override:
        return (*find_vendor_id(access, cid, override), False)
    aliases = load_aliases()
    cached = aliases.get(_alias_key(pdf_vendor))
    cached_id = cached.get("qbo_id") if cached else None
    if cached_id:
        current_name = validate_cached_vendor(access, cid, cached_id)
        if current_name:
            return cached_id, current_name, True
        _warn(f"cached vendor id {cached_id} ({cached.get('qbo_name')}) no longer in QBO — re-resolving.")
    return (*find_vendor_id(access, cid, pdf_vendor), False)


def find_vendor_id(access: str, cid: str, vendor_name_hint: str) -> Tuple[str, str]:
    """Returns (vendor_id, vendor_display_name).
    Tries progressively-relaxed LIKE searches against QBO Vendor.DisplayName:
      1. first 2 whitespace tokens with internal punctuation preserved
         ('Post-Tension Services' — handles hyphens/ampersands cleanly)
      2. first 1 token (handles vendors whose PDF shows only one word)
      3. hyphen-stripped variants of #1 and #2 (handles QBO names spelled
         without the hyphen — e.g., PDF 'Post-Tension' vs QBO 'Post Tension')
    Trailing punctuation like commas/periods is stripped from each token
    so 'Ready Cable, Inc' → tokens ['Ready', 'Cable', 'Inc']."""
    if not vendor_name_hint:
        sys.exit("✗ could not identify vendor from PDF. Pass --vendor explicitly.")
    raw_tokens = vendor_name_hint.split()
    tokens = [t.rstrip(",.;:") for t in raw_tokens if t.rstrip(",.;:")]
    if not tokens:
        sys.exit(f"✗ vendor hint '{vendor_name_hint}' contains no usable tokens.")

    # Build the ordered list of needles to try. Dedup while preserving order.
    attempts: List[Tuple[str, str]] = []
    seen: set = set()
    def _add(label: str, needle: str) -> None:
        n = needle.strip()
        if n and n not in seen:
            attempts.append((label, n))
            seen.add(n)

    # Full precise name first — a parser that identifies the exact entity
    # (e.g. 'Bobcat of North Texas') must beat the first-2-words reduction,
    # since 'Bobcat of' alone collides with 'Bobcat of Midland'. Only worth a
    # distinct attempt at 3+ tokens; at 1-2 tokens it equals the reductions
    # below and dedup skips it.
    if len(tokens) >= 3:
        _add("full name", " ".join(tokens))
    if len(tokens) >= 2:
        _add("first 2 words", f"{tokens[0]} {tokens[1]}")
    _add("first word", tokens[0])
    # Hyphen-stripped fallbacks — only if there's actually a hyphen
    if any("-" in t for t in tokens[:2]):
        if len(tokens) >= 2:
            _add("first 2 words (no hyphens)",
                 f"{tokens[0]} {tokens[1]}".replace("-", " "))
        _add("first word (no hyphens)", tokens[0].replace("-", " "))

    found_rows: List[dict] = []
    used_label = used_needle = ""
    for label, needle in attempts:
        safe = needle.replace("'", "''")
        rows = query(access, cid,
                     f"SELECT Id, DisplayName FROM Vendor WHERE DisplayName LIKE '%{safe}%'"
                     ).get("QueryResponse", {}).get("Vendor", [])
        if rows:
            found_rows, used_label, used_needle = rows, label, needle
            break

    if not found_rows:
        tried = ", ".join(repr(n) for _, n in attempts)
        sys.exit(f"✗ no QBO vendor matches. Tried: {tried}. "
                 "Pass --vendor with exact display name.")
    if len(found_rows) > 1:
        print(f"⚠ multiple vendor matches for {used_needle!r}: "
              f"{[r['DisplayName'] for r in found_rows]}")
        print(f"  using first: {found_rows[0]['DisplayName']}")
    if used_label != "first 2 words":
        print(f"  (matched via fallback: {used_label} → {used_needle!r})")
    return found_rows[0]["Id"], found_rows[0]["DisplayName"]


def _bills_to_objs(raw: List[dict]) -> List[QboBill]:
    out: List[QboBill] = []
    for b in raw:
        # QBO UI "Memo" on a Bill maps to API `PrivateNote`. A few bills
        # have a top-level `Memo` instead — accept either, prefer PrivateNote.
        memo_value = (str(b.get("PrivateNote") or "").strip()
                      or str(b.get("Memo") or "").strip())
        out.append(QboBill(
            bill_id=b["Id"],
            doc_number=str(b.get("DocNumber", "")).strip(),
            txn_date=b.get("TxnDate", ""),
            open_balance=float(b.get("Balance", 0)),
            total_amount=float(b.get("TotalAmt", 0)),
            memo=memo_value,
        ))
    return out


def get_open_bills_for_vendor(access: str, cid: str, vendor_id: str) -> List[QboBill]:
    """Pulls ALL open Bills (Balance > 0) regardless of past-due status."""
    raw = query_all(access, cid, "Bill",
                    where=f"VendorRef = '{vendor_id}' AND Balance > '0'")
    return _bills_to_objs(raw)


def get_recently_paid_bills_for_vendor(access: str, cid: str, vendor_id: str,
                                       since_date: str) -> List[QboBill]:
    """Pulls fully-paid Bills (Balance = 0) with TxnDate >= since_date.
    Used to catch vendor-lag: bills you paid before the statement was printed."""
    where = (f"VendorRef = '{vendor_id}' AND Balance = '0' "
             f"AND TxnDate >= '{since_date}'")
    raw = query_all(access, cid, "Bill", where=where)
    return _bills_to_objs(raw)

# ───────────────────────── matching ─────────────────────────

def _build_row(sl: StmtLine, bill: Optional[QboBill],
               paid_bill: Optional[QboBill] = None) -> ReconRow:
    """Pure classification for one statement line.
    Precedence: open Bill (MATCHED/TAX/MISMATCH) → paid Bill (LIKELY_VENDOR_LAG) → MISSING_IN_QBO."""
    if bill is None:
        if paid_bill is not None:
            return ReconRow(
                "LIKELY_VENDOR_LAG", sl.date, sl.ref,
                sl.amount, paid_bill.open_balance, sl.po, sl.address,
                f"Paid in QBO on {paid_bill.txn_date} (BillId={paid_bill.bill_id}). "
                "Statement likely printed before vendor received check. Verify check has cleared.",
                stmt_ref=sl.ref, qbo_ref=paid_bill.doc_number,
                stmt_date=sl.date, qbo_date=paid_bill.txn_date,
                qbo_memo=paid_bill.memo, qbo_bill_id=paid_bill.bill_id,
            )
        return ReconRow(
            "MISSING_IN_QBO", sl.date, sl.ref,
            sl.amount, 0.0, sl.po, sl.address, "ENTER BILL IN QBO",
            stmt_ref=sl.ref, qbo_ref="",
            stmt_date=sl.date, qbo_date="",
        )
    diff = round(sl.amount - bill.open_balance, 2)
    common = dict(po=sl.po, address=sl.address,
                  stmt_ref=sl.ref, qbo_ref=bill.doc_number,
                  stmt_date=sl.date, qbo_date=bill.txn_date,
                  qbo_memo=bill.memo, qbo_bill_id=bill.bill_id)
    if abs(diff) < EXACT_TOLERANCE:
        return ReconRow("MATCHED", sl.date, sl.ref, sl.amount,
                        bill.open_balance, **common)
    expected_tax = round(bill.open_balance * TX_SALES_TAX, 2)
    if abs(diff - expected_tax) <= TAX_TOLERANCE:
        return ReconRow("VENDOR_TAX_VIOLATION", sl.date, sl.ref,
                        sl.amount, bill.open_balance,
                        notes="Stmt = QBO × 1.0825 (TX sales tax). Per agreement: NO TAX. Vendor action.",
                        **common)
    return ReconRow("CLERK_AMOUNT_MISMATCH", sl.date, sl.ref,
                    sl.amount, bill.open_balance,
                    notes=f"Diff = ${diff:,.2f}. Investigate.", **common)


def reconcile_iter(lines: List[StmtLine],
                   bills: List[QboBill],
                   paid_bills: Optional[List[QboBill]] = None,
                   stmt_date: str = "",
                   ) -> Iterator[Tuple[int, int, ReconRow]]:
    """Match by Ref# = DocNumber. Yields (current_index, total, row) per result.

    Total is computed up front: len(lines) + #unmatched-in-QBO bills (i.e., the
    MISSING_ON_STATEMENT rows that will be emitted after the main loop).

    `stmt_date` is the statement's "as of" date (YYYY-MM-DD). When provided,
    QBO bills with a txn_date AFTER stmt_date are excluded from MISSING_ON_STATEMENT
    because they couldn't possibly have been on a statement that hadn't been
    printed yet. Without this filter, every fresh bill in QBO would false-flag.

    Reliability warnings are printed before iteration starts:
      • Duplicate DocNumber across QBO bills (silent last-wins).
      • Duplicate Ref# on the statement (vendor reused an invoice number).
      • Bills with empty DocNumber in QBO (cannot be matched).
    """
    # Detect duplicate DocNumbers in QBO
    seen_docs: Dict[str, int] = {}
    no_doc_bills: List[QboBill] = []
    for b in bills:
        if not b.doc_number:
            no_doc_bills.append(b)
            continue
        seen_docs[b.doc_number] = seen_docs.get(b.doc_number, 0) + 1
    dup_qbo = [d for d, n in seen_docs.items() if n > 1]
    if dup_qbo:
        _warn(f"QBO has duplicate DocNumber(s): {dup_qbo} — last-wins; review manually.")
    if no_doc_bills:
        _warn(f"{len(no_doc_bills)} open QBO bill(s) have empty DocNumber — cannot match by Ref#.")

    # Detect duplicate Ref# on the statement
    seen_stmt: Dict[str, int] = {}
    for l in lines:
        seen_stmt[l.ref] = seen_stmt.get(l.ref, 0) + 1
    dup_stmt = [r for r, n in seen_stmt.items() if n > 1]
    if dup_stmt:
        _warn(f"Statement has duplicate Ref#(s): {dup_stmt} — both lines will match the same QBO bill.")

    by_doc = {b.doc_number: b for b in bills if b.doc_number}
    paid_by_doc = {b.doc_number: b for b in (paid_bills or []) if b.doc_number}
    stmt_refs = {l.ref for l in lines}

    # MISSING_ON_STATEMENT: open in QBO, no matching ref on stmt, AND dated on
    # or before the statement's as-of date. Post-statement bills are excluded —
    # they couldn't have been on a statement that hadn't been printed yet.
    def _on_or_before_stmt(b: QboBill) -> bool:
        if not stmt_date or not b.txn_date:
            return True   # no cutoff info → don't filter (preserve old behavior)
        return b.txn_date <= stmt_date

    missing_on_stmt: List[QboBill] = []
    post_stmt_excluded = 0
    for b in bills:
        if not b.doc_number or b.doc_number in stmt_refs:
            continue
        if _on_or_before_stmt(b):
            missing_on_stmt.append(b)
        else:
            post_stmt_excluded += 1
    if post_stmt_excluded:
        _warn(f"excluded {post_stmt_excluded} QBO bill(s) dated AFTER stmt {stmt_date} "
              "from MISSING_ON_STATEMENT (can't be missing — statement is older).")

    no_doc_bills_filtered = [b for b in no_doc_bills if _on_or_before_stmt(b)]
    total = len(lines) + len(missing_on_stmt) + len(no_doc_bills_filtered)

    idx = 0
    for sl in lines:
        idx += 1
        open_match = by_doc.get(sl.ref)
        paid_match = paid_by_doc.get(sl.ref) if open_match is None else None
        yield idx, total, _build_row(sl, open_match, paid_match)

    # MISSING_ON_STATEMENT
    for b in missing_on_stmt:
        idx += 1
        yield idx, total, ReconRow(
            "MISSING_ON_STATEMENT", b.txn_date, b.doc_number,
            0.0, b.open_balance, "", "",
            "Open in QBO but not on statement. Vendor may have credited/applied payment.",
            stmt_ref="", qbo_ref=b.doc_number,
            stmt_date="", qbo_date=b.txn_date,
            qbo_memo=b.memo, qbo_bill_id=b.bill_id,
        )
    for b in no_doc_bills_filtered:
        idx += 1
        yield idx, total, ReconRow(
            "MISSING_ON_STATEMENT", b.txn_date, f"(no Ref#, BillId={b.bill_id})",
            0.0, b.open_balance, "", "",
            "QBO bill has no DocNumber — cannot match by Ref#. Add Ref # in QBO.",
            stmt_ref="", qbo_ref=f"(no Ref#, BillId={b.bill_id})",
            stmt_date="", qbo_date=b.txn_date,
            qbo_memo=b.memo, qbo_bill_id=b.bill_id,
        )


def reconcile(lines: List[StmtLine], bills: List[QboBill],
              paid_bills: Optional[List[QboBill]] = None,
              stmt_date: str = "") -> List[ReconRow]:
    """Backwards-compatible non-streaming wrapper."""
    return [row for _, _, row in reconcile_iter(lines, bills, paid_bills, stmt_date)]

# ───────────────────────── Excel writer ─────────────────────────

def _render_statement_pages(src: Path) -> List[Path]:
    """Render a statement file to PNG page images in a fresh temp dir.
    PDF → one PNG per page (via pypdfium2, ~150 DPI); image → one normalized
    PNG; Excel/other → [] (nothing to rasterize). Returns the list of PNG paths;
    caller owns the temp dir (paths[0].parent) and must clean it up after save."""
    ext = src.suffix.lower()
    tmpdir = Path(tempfile.mkdtemp(prefix="stmt_pages_"))
    imgs: List[Path] = []
    if ext == ".pdf":
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(str(src))
        try:
            n = min(len(pdf), STATEMENT_EMBED_MAX_PAGES)
            if len(pdf) > STATEMENT_EMBED_MAX_PAGES:
                _warn(f"statement has {len(pdf)} pages; embedding first "
                      f"{STATEMENT_EMBED_MAX_PAGES} only.")
            for i in range(n):
                page = pdf[i]
                pil = page.render(scale=150 / 72).to_pil()
                p = tmpdir / f"page_{i + 1:02d}.png"
                pil.save(p)
                imgs.append(p)
        finally:
            pdf.close()
    elif ext in (".png", ".jpg", ".jpeg", ".heic", ".heif"):
        from PIL import Image as PILImage
        if ext in (".heic", ".heif"):
            try:
                import pillow_heif
                pillow_heif.register_heif_opener()
            except Exception:
                pass
        with PILImage.open(src) as im:
            p = tmpdir / "page_01.png"
            im.convert("RGB").save(p)
            imgs.append(p)
    return imgs


def _embed_statement_tab(wb, src: Path) -> Optional[Path]:
    """Add a 'Statement' sheet with the source statement rendered as stacked
    page images. Returns the temp dir to clean up after wb.save (or None)."""
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    pages = _render_statement_pages(src)
    if not pages:
        return None
    ws = wb.create_sheet("Statement")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Source statement: {src.name}"
    ws["A1"].font = LABEL_FONT
    ws["A2"] = "Embedded image — travels inside this workbook."
    ws["A2"].font = Font(italic=True, name="Arial", size=10, color="808080")
    anchor_row = 4
    for p in pages:
        with PILImage.open(p) as im:
            w, h = im.size
        img = XLImage(str(p))
        if w > STATEMENT_EMBED_MAX_WIDTH:
            img.width = STATEMENT_EMBED_MAX_WIDTH
            img.height = int(h * (STATEMENT_EMBED_MAX_WIDTH / w))
        else:
            img.width, img.height = w, h
        ws.add_image(img, f"A{anchor_row}")
        anchor_row += max(2, int(img.height / 18)) + 2   # ~18px per sheet row + gap
    return pages[0].parent


def write_excel(out_path: Path, vendor: str, stmt_date: str, stmt_total: float,
                qbo_bills: List[QboBill], rows: List[ReconRow],
                statement_src: Optional[Path] = None) -> None:
    # QBO open as-of stmt_date: exclude post-statement bills from the displayed
    # total so the reconciliation math lines up with the statement snapshot.
    def _as_of(b: QboBill) -> bool:
        return (not stmt_date) or (not b.txn_date) or (b.txn_date <= stmt_date)
    qbo_bills_asof = [b for b in qbo_bills if _as_of(b)]
    post_stmt_bills = [b for b in qbo_bills if not _as_of(b)]
    qbo_total = round(sum(b.open_balance for b in qbo_bills_asof), 2)
    post_stmt_total = round(sum(b.open_balance for b in post_stmt_bills), 2)
    # Split MATCHED by approval status so unapproved-but-matched bills don't
    # hide inside the "clean" pile — they still need PM signoff before payment.
    cats = {
        "MATCHED_APPROVED":      [r for r in rows if r.category == "MATCHED" and _is_approved(r.qbo_memo)],
        "MATCHED_NOT_APPROVED":  [r for r in rows if r.category == "MATCHED" and not _is_approved(r.qbo_memo)],
        "VENDOR_TAX_VIOLATION":  [r for r in rows if r.category == "VENDOR_TAX_VIOLATION"],
        "CLERK_AMOUNT_MISMATCH": [r for r in rows if r.category == "CLERK_AMOUNT_MISMATCH"],
        "LIKELY_VENDOR_LAG":     [r for r in rows if r.category == "LIKELY_VENDOR_LAG"],
        "MISSING_IN_QBO":        [r for r in rows if r.category == "MISSING_IN_QBO"],
        "MISSING_ON_STATEMENT":  [r for r in rows if r.category == "MISSING_ON_STATEMENT"],
    }

    wb = Workbook()

    # ── Summary ──
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    s["A1"] = f"Statement Reconciliation — {vendor}"
    s["A1"].font = TITLE_FONT
    s["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    s.merge_cells("A1:L1")
    s.row_dimensions[1].height = 26
    s["A2"] = f"Statement date: {stmt_date}    |    Generated: {dt.date.today():%Y-%m-%d}"
    s["A2"].font = BODY_FONT
    s["A2"].alignment = Alignment(horizontal="left", vertical="center")
    s.merge_cells("A2:L2")
    s.row_dimensions[2].height = 18

    # Tie-out: label in merged A:C, dollar value in D. Wider label area so
    # "QBO open as of YYYY-MM-DD (NN bills)" fits without truncation.
    r = 4
    s.cell(row=r, column=1, value="TIE-OUT").font = HEADER_FONT
    s.cell(row=r, column=1).fill = HEADER_FILL
    s.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    def _tieout_row(row_idx: int, label: str, value, font, fill=None, formula=False):
        s.cell(row=row_idx, column=1, value=label).font = font
        s.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        s.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        c = s.cell(row=row_idx, column=5, value=value)
        c.number_format = MONEY; c.font = font; c.alignment = RIGHT
        if fill:
            for col in (1, 2, 3, 4, 5):
                s.cell(row=row_idx, column=col).fill = fill

    stmt_row = r
    _tieout_row(r, "Statement total", stmt_total, BODY_FONT); r += 1
    qbo_label = (f"QBO open as of {stmt_date} ({len(qbo_bills_asof)} bills)"
                 if stmt_date else f"QBO open ({len(qbo_bills_asof)} open Bills)")
    qbo_row = r
    _tieout_row(r, qbo_label, qbo_total, BODY_FONT); r += 1
    _tieout_row(r, "Reconciling difference", f"=E{stmt_row}-E{qbo_row}", LABEL_FONT, fill=SUBTOTAL_FILL); r += 1

    # FYI: post-statement bills excluded from the tie-out (wrap text, tall row)
    if post_stmt_bills:
        n = len(post_stmt_bills)
        info = (f"({n} QBO bill{'s' if n!=1 else ''} dated after {stmt_date}, "
                f"totaling ${post_stmt_total:,.2f}, "
                f"{'are' if n!=1 else 'is'} excluded from this tie-out — "
                f"{'they' if n!=1 else 'it'} couldn't have been on this statement.)")
        s.cell(row=r, column=1, value=info).font = Font(italic=True, name="Arial", size=10, color="808080")
        s.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        s.row_dimensions[r].height = 32
        r += 1

    # Bills pending approval in QBO — counted from in-scope (as-of) bills only
    unapproved = [b for b in qbo_bills_asof if not b.is_approved]
    if unapproved:
        unapproved_amt = round(sum(b.open_balance for b in unapproved), 2)
        msg = (f"⚠ Bills pending approval in QBO: {len(unapproved)}  "
               f"(${unapproved_amt:,.2f}) — see 'Approved?' column below "
               f"(red cells = needs PM signoff)")
        c = s.cell(row=r, column=1, value=msg)
        c.font = Font(bold=True, name="Arial", size=11, color="C62828")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", start_color="FFE5E5")
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        s.row_dimensions[r].height = 24
        r += 1
    r += 1

    # ── Per-category bill lists with collapsible groups (summary above) ──
    # Layout: leading ↗ bill-link column, then 9 columns side-by-side stmt vs
    # QBO with check marks, then Approved? + Note.
    #   A: ↗ (open bill in QBO)
    #   B: Stmt Ref#   C: QBO Ref#   D: ✓?
    #   E: Stmt $      F: QBO $      G: ✓?
    #   H: Stmt Date   I: QBO Date   J: ✓?
    # Group rows collapse under the category-header row. MATCHED collapsed by
    # default (boring); others expanded (actionable).
    s.sheet_properties.outlinePr.summaryBelow = False  # summary row is ABOVE detail

    LAG_FILL = PatternFill("solid", start_color="DDEBF7")
    CHECK_OK = "✓"
    CHECK_NO = "✗"
    CHECK_NA = "—"

    GREEN_FONT = Font(bold=True, name="Arial", size=10, color="2E7D32")
    RED_FONT   = Font(bold=True, name="Arial", size=10, color="C62828")
    DIM_FONT   = Font(name="Arial", size=10, color="808080")
    LINK_FONT  = Font(bold=True, name="Arial", size=11, color="0563C1", underline="single")

    CAT_DEFS = [
        ("MATCHED_APPROVED",      "✓ MATCHED — APPROVED",                       OK_FILL,   True),   # collapse by default (boring/clean)
        ("MATCHED_NOT_APPROVED",  "⚠ MATCHED but NOT APPROVED (chase PM)",      WARN_FILL, False),  # surface — needs PM signoff
        ("VENDOR_TAX_VIOLATION",  "⚠ VENDOR TAX VIOLATION (8.25%)",             WARN_FILL, False),
        ("CLERK_AMOUNT_MISMATCH", "⚠ CLERK AMOUNT MISMATCH",                    WARN_FILL, False),
        ("LIKELY_VENDOR_LAG",     "⊙ LIKELY VENDOR LAG (paid in QBO)",          LAG_FILL,  False),
        ("MISSING_IN_QBO",        "✗ MISSING IN QBO",                           BAD_FILL,  False),
        ("MISSING_ON_STATEMENT",  "✗ MISSING ON STATEMENT",                     BAD_FILL,  False),
    ]
    SUMMARY_NCOLS = 12

    def _section_header_text(label: str, count: int, stmt_sum: float, qbo_sum: float) -> str:
        diff = round(stmt_sum - qbo_sum, 2)
        if count == 0:
            return f"{label}  —  0 bills"
        if "TAX" in label or "MISMATCH" in label:
            return (f"{label}  —  {count} bill{'s' if count!=1 else ''}  ·  "
                    f"stmt ${stmt_sum:,.2f}  vs  QBO ${qbo_sum:,.2f}  (diff ${diff:,.2f})")
        if "MISSING ON STATEMENT" in label:
            return f"{label}  —  {count} bill{'s' if count!=1 else ''}  ·  ${qbo_sum:,.2f} open in QBO"
        if "VENDOR LAG" in label:
            return f"{label}  —  {count} bill{'s' if count!=1 else ''}  ·  ${stmt_sum:,.2f} on stmt"
        return f"{label}  —  {count} bill{'s' if count!=1 else ''}  ·  ${stmt_sum:,.2f}"

    def _write_section_header_row(row_idx: int, text: str, fill) -> None:
        s.cell(row=row_idx, column=1, value=text).font = LABEL_FONT
        s.cell(row=row_idx, column=1).alignment = LEFT
        for col in range(1, SUMMARY_NCOLS + 1):
            s.cell(row=row_idx, column=col).fill = fill
            s.cell(row=row_idx, column=col).border = BORDER
        s.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=SUMMARY_NCOLS)

    def _write_subhead_row(row_idx: int) -> None:
        headers = ["Bill", "Stmt Ref #", "QBO Ref #", "✓", "Stmt $", "QBO $", "✓",
                   "Stmt Date", "QBO Date", "✓", "Approved?", "Note"]
        for i, h in enumerate(headers, 1):
            c = s.cell(row=row_idx, column=i, value=h)
            c.font = Font(bold=True, name="Arial", size=10, color="404040")
            c.alignment = CENTER
            c.fill = SUBTOTAL_FILL
            c.border = BORDER

    # Approval-state cell styling
    UNAPPROVED_FILL = PatternFill("solid", start_color="FFE5E5")
    UNAPPROVED_FONT = Font(bold=True, name="Arial", size=10, color="C62828")
    APPROVED_FONT   = Font(bold=True, name="Arial", size=10, color="2E7D32")

    def _write_compare_row(row_idx: int, rr: ReconRow) -> None:
        # Ref check
        stmt_ref, qbo_ref = rr.stmt_ref, rr.qbo_ref
        ref_check = (CHECK_OK if (stmt_ref and qbo_ref and stmt_ref == qbo_ref)
                     else CHECK_NA if (not stmt_ref or not qbo_ref)
                     else CHECK_NO)
        # Amount check (within tolerance)
        amt_check = (CHECK_OK if abs(rr.stmt_amount - rr.qbo_amount) < EXACT_TOLERANCE
                     and (rr.stmt_amount > 0 or rr.qbo_amount > 0)
                     else CHECK_NA if (rr.stmt_amount == 0 or rr.qbo_amount == 0)
                          and rr.category in ("MISSING_IN_QBO", "MISSING_ON_STATEMENT")
                     else CHECK_NO)
        # Date check
        date_check = (CHECK_OK if (rr.stmt_date and rr.qbo_date and rr.stmt_date == rr.qbo_date)
                      else CHECK_NA if (not rr.stmt_date or not rr.qbo_date)
                      else CHECK_NO)
        # Approval state: only meaningful when there's a QBO bill (MISSING_IN_QBO has none)
        if rr.category == "MISSING_IN_QBO":
            approved_text = "—"
            approved_font = DIM_FONT
            approved_fill = None
        else:
            is_app = _is_approved(rr.qbo_memo)
            approved_text = "Yes" if is_app else "Not Approved"
            approved_font = APPROVED_FONT if is_app else UNAPPROVED_FONT
            approved_fill = None if is_app else UNAPPROVED_FILL

        # Column 1: ↗ deep-link that opens this Bill in QBO. Blank when there is
        # no QBO bill for the row (MISSING_IN_QBO). Ref# text stays non-clickable.
        link_cell = s.cell(row=row_idx, column=1, value=("↗" if rr.qbo_bill_id else ""))
        link_cell.alignment = CENTER
        link_cell.border = BORDER
        if rr.qbo_bill_id:
            link_cell.hyperlink = QBO_BILL_URL_TEMPLATE.format(bill_id=rr.qbo_bill_id)
            link_cell.font = LINK_FONT
        else:
            link_cell.font = BODY_FONT

        cells = [
            (rr.stmt_ref, LEFT, None),
            (rr.qbo_ref, LEFT, None),
            (ref_check, CENTER, GREEN_FONT if ref_check == CHECK_OK else (RED_FONT if ref_check == CHECK_NO else DIM_FONT)),
            (rr.stmt_amount if rr.stmt_amount else None, RIGHT, None),
            (rr.qbo_amount if rr.qbo_amount else None, RIGHT, None),
            (amt_check, CENTER, GREEN_FONT if amt_check == CHECK_OK else (RED_FONT if amt_check == CHECK_NO else DIM_FONT)),
            (rr.stmt_date, CENTER, None),
            (rr.qbo_date, CENTER, None),
            (date_check, CENTER, GREEN_FONT if date_check == CHECK_OK else (RED_FONT if date_check == CHECK_NO else DIM_FONT)),
            (approved_text, CENTER, approved_font),
            (rr.notes or "", LEFT, None),
        ]
        # Compare cells occupy columns 2..12 (column 1 is the ↗ link above).
        for i, (val, align, font) in enumerate(cells, 2):
            c = s.cell(row=row_idx, column=i, value=val)
            c.font = font or BODY_FONT
            c.alignment = align
            c.border = BORDER
            if i in (5, 6):
                c.number_format = MONEY
            # Highlight unapproved cell (Approved? column, now col 11) with red fill
            if i == 11 and approved_fill is not None:
                c.fill = approved_fill

    for key, label, fill, collapse_by_default in CAT_DEFS:
        bucket = cats[key]
        n = len(bucket)
        stmt_sum = round(sum(rr.stmt_amount for rr in bucket), 2)
        qbo_sum  = round(sum(rr.qbo_amount  for rr in bucket), 2)

        _write_section_header_row(r, _section_header_text(label, n, stmt_sum, qbo_sum), fill)
        r += 1
        if n == 0:
            continue

        # Sub-header and data rows — all at outline_level=1, grouped under the header
        sub_head_row = r
        _write_subhead_row(sub_head_row)
        s.row_dimensions[sub_head_row].outline_level = 1
        s.row_dimensions[sub_head_row].hidden = collapse_by_default
        r += 1

        for rr_row in bucket:
            _write_compare_row(r, rr_row)
            s.row_dimensions[r].outline_level = 1
            s.row_dimensions[r].hidden = collapse_by_default
            r += 1

        # Spacer row (not grouped) between categories
        r += 1

    # Column widths sized for the 12-col layout (↗ link + ref/amount/date
    # triplets + Approved? + Note)
    for col, w in {"A": 5, "B": 13, "C": 13, "D": 4, "E": 13, "F": 13, "G": 4,
                   "H": 12, "I": 12, "J": 4, "K": 14, "L": 36}.items():
        s.column_dimensions[col].width = w

    # Freeze the top rows (title + tie-out) so per-bill rows stay scrollable
    # under a fixed header. Freeze at row 4 + col A leaves the table header
    # area visible at all times.
    s.freeze_panes = "A4"

    # Embed the source statement as a self-contained "Statement" tab so the
    # workbook carries the original inside it — no external link to break when
    # the file is moved (Inbox → Reconciliations → Old-Done).
    cleanup_dir = None
    if statement_src is not None:
        try:
            cleanup_dir = _embed_statement_tab(wb, statement_src)
        except Exception as e:
            _warn(f"could not embed statement image ({e}); Excel saved without it.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    if cleanup_dir:
        shutil.rmtree(cleanup_dir, ignore_errors=True)

# ───────────────────────── interactive helpers ─────────────────────────

def _slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")[:40] or "Vendor"


def _confirm(prompt: str, default_yes: bool = True, skip: bool = False) -> bool:
    """Y/N confirmation. Returns True if user proceeds."""
    if skip:
        return True
    default_str = "[Y/n]" if default_yes else "[y/N]"
    try:
        ans = input(f"{prompt} {default_str} ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()  # newline after ^C
        return False
    if not ans:
        return default_yes
    return ans in ("y", "yes")


def _hr():
    print("─" * 60)


def _open_file(path: Path) -> None:
    """Open a file with the OS default app (macOS only — `open`)."""
    try:
        subprocess.run(["/usr/bin/open", str(path)], check=False)
    except Exception as e:
        print(f"  (could not auto-open: {e})")


def _term_link(label: str, path: Path) -> str:
    """OSC-8 terminal hyperlink. Click in macOS Terminal/iTerm2 to open in Finder.
    Falls back to plain `label  <url>` when ANSI is disabled / non-tty."""
    abs_path = path.resolve()
    url = "file://" + urllib.parse.quote(str(abs_path))
    if not _Term.enabled:
        return f"{label}  {url}"
    # OSC-8: ESC ] 8 ;; URL ST  text  ESC ] 8 ;; ST
    return f"\033]8;;{url}\033\\{label}\033]8;;\033\\"


def append_clerk_perf(rows: List["ReconRow"], vendor_name: str, stmt_date: str,
                      stmt_total: float) -> Optional[Path]:
    """Append one row to clerk_performance.csv. Focused on clerk-action metrics —
    vendor-side issues (tax violations) are EXCLUDED since the user reads those from QBO."""
    import csv as _csv

    by_cat: Dict[str, List["ReconRow"]] = {}
    for r in rows:
        by_cat.setdefault(r.category, []).append(r)

    missing  = by_cat.get("MISSING_IN_QBO", [])
    mismatch = by_cat.get("CLERK_AMOUNT_MISMATCH", [])
    matched  = by_cat.get("MATCHED", [])
    lag      = by_cat.get("LIKELY_VENDOR_LAG", [])

    # Oldest missing in days = (stmt_date) - (earliest stmt_date among missing)
    oldest_missing_days = ""
    if missing:
        try:
            stmt_dt = dt.date.fromisoformat(stmt_date)
            ages = []
            for r in missing:
                try:
                    ages.append((stmt_dt - dt.date.fromisoformat(r.date)).days)
                except ValueError:
                    pass
            if ages:
                oldest_missing_days = max(ages)
        except ValueError:
            pass

    row = {
        "reconcile_date":         dt.date.today().isoformat(),
        "vendor":                 vendor_name,
        "statement_date":         stmt_date,
        "statement_total":        f"{stmt_total:.2f}",
        "matched_count":          len(matched),
        "clerk_mismatch_count":   len(mismatch),
        "missing_in_qbo_count":   len(missing),
        "missing_in_qbo_amt":     f"{sum(r.stmt_amount for r in missing):.2f}",
        "oldest_missing_days":    oldest_missing_days,
        "likely_vendor_lag_count": len(lag),
    }

    CLERK_PERF_CSV.parent.mkdir(parents=True, exist_ok=True)
    is_new = not CLERK_PERF_CSV.exists()
    try:
        with CLERK_PERF_CSV.open("a", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=list(row.keys()))
            if is_new:
                w.writeheader()
            w.writerow(row)
        return CLERK_PERF_CSV
    except OSError as e:
        _warn(f"could not append clerk_performance.csv ({e})")
        return None


# ───────────────────────── main ─────────────────────────

def process_pdf(pdf_path: Path, args: argparse.Namespace,
                access: str, cid: str) -> Tuple[bool, Dict[str, int]]:
    """Run the full pipeline for one PDF. Returns (ok, counts).
    Assumes QBO is already authenticated; caller provides access + cid."""
    print(_Term.color(_Term.BOLD, "━" * min(60, _width())))
    print(_Term.color(_Term.BOLD, f"  STATEMENT RECONCILER  ·  {pdf_path.name}"))
    print(_Term.color(_Term.BOLD, "━" * min(60, _width())))

    # ── parse statement (PDF / Excel / Image) ───────────────
    ext = pdf_path.suffix.lower()
    is_excel = ext in (".xlsx", ".xls", ".xlsm")
    is_image = ext in (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".heic", ".heif")
    file_kind = "Excel" if is_excel else ("Image (OCR)" if is_image else "PDF")
    t0 = _phase(f"Reading {file_kind} {pdf_path.name}")
    vendor_guess, stmt_date, amt_due, lines = parse_statement(pdf_path)

    # Vendor fallback chain — runs for ALL file types (PDF, Excel, Image).
    # Step 2 (file body extraction) wins for PDFs where the vendor is in the
    # header/body. For Excel and OCR'd images where body extraction often
    # comes up empty, the chain falls through to filename / cell scan / prompt.
    #
    # Order rationale: explicit user signals (flag, alias, deliberate filename)
    # beat incidental file body data. Excel cells in particular often contain
    # OUR OWN company name as bill-to ("Proficient Concrete LLC") and the
    # vendor's full legal name with corp suffixes that fuzz the QBO LIKE
    # search — so filename (user-curated) beats cell scan (incidental).
    vendor_source = ""
    if args.vendor:
        # 1. Explicit --vendor flag — overrides everything
        vendor_guess = args.vendor
        vendor_source = "from --vendor flag"
    elif vendor_guess:
        # 2. File body extraction (PDF parser, OCR'd text via PDF parser).
        #    For Excel this is always empty; for PDFs this usually works.
        vendor_source = f"from {file_kind} body"

    # 3. Cached alias by filename (works for any extension)
    if not vendor_guess and not args.no_cache:
        match = find_alias_by_filename(pdf_path.name)
        if match:
            _key, data, _score, matched_tokens = match
            vendor_guess = data.get("qbo_name", "")
            vendor_source = f"from cached alias (filename tokens: {matched_tokens})"

    # 4. Cleaned filename (user-curated signal — trust it before incidental file body)
    if not vendor_guess:
        cleaned = _vendor_from_filename(pdf_path)
        if cleaned and len(cleaned) >= 3:
            vendor_guess = cleaned
            vendor_source = "from filename"

    # 5. Scan Excel cells (Excel only — last automated step before prompt)
    if not vendor_guess and is_excel:
        scanned = _scan_excel_for_vendor(pdf_path)
        if scanned:
            vendor_guess = scanned
            vendor_source = "from Excel cell"

    # 6. Interactive prompt — last resort, works for any file type
    if not vendor_guess:
        print()
        print(_Term.color(_Term.Y, "  ⚠ Couldn't auto-detect vendor (no flag, no cached alias, "
                                    "no usable filename, no body extraction)."))
        try:
            vendor_guess = input("  Vendor name to search QBO for: ").strip()
            vendor_source = "user-entered"
        except (EOFError, KeyboardInterrupt):
            print()
            _fail("vendor input required — aborted.")
            return False, {}

    if not vendor_guess:
        _fail("no vendor identified — aborted.")
        return False, {}

    # Always log which source the vendor came from (helps debug surprises)
    print(f"  {_Term.color(_Term.G, '✓')} Vendor candidate: "
          f"{_Term.color(_Term.C, vendor_guess)}  "
          f"{_Term.color(_Term.DIM, f'({vendor_source})')}")

    # Excel/Image — default stmt_date to today if the file didn't provide one
    if not stmt_date:
        if args.stmt_date:
            stmt_date = args.stmt_date
        elif is_excel or is_image:
            stmt_date = dt.date.today().isoformat()
            _warn(f"No statement date in {file_kind} — defaulting to today ({stmt_date}). "
                  f"Use --stmt-date YYYY-MM-DD if the statement is for a different as-of date.")

    vendor_hint = vendor_guess

    if not lines:
        _fail(f"No statement lines parsed from this {file_kind}.")
        print("    Supported formats:")
        for label in TEMPLATE_LABELS.values():
            print(f"      • {label}")
        print("    If this is a new format, share it so a parser can be added.")
        return False, {}

    line_sum = round(sum(l.amount for l in lines), 2)
    sum_matches = abs(line_sum - amt_due) <= 0.50
    _done(t0, f"Extracted {_Term.color(_Term.BOLD, f'{len(lines)} bill lines')} totaling "
              f"{_Term.color(_Term.BOLD, f'${line_sum:,.2f}')}")

    # ── confirm #1: parse sanity ─────────────────────────────
    print()
    _hr()
    print(_Term.color(_Term.BOLD, "  PARSE SANITY CHECK"))
    _hr()
    print(f"  Vendor:               {_Term.color(_Term.C, vendor_hint or '(unknown)')}")
    print(f"  Statement date:       {stmt_date or '(unknown)'}")
    print(f"  Amount Due (file):    ${amt_due:,.2f}")
    print(f"  Lines parsed:         {len(lines)}")
    sum_marker = _Term.color(_Term.G, '✓ matches Amount Due') if sum_matches else _Term.color(_Term.R, '⚠ DOES NOT MATCH')
    print(f"  Sum of line amounts:  ${line_sum:,.2f}  {sum_marker}")
    print()
    first = lines[0]; last = lines[-1]
    print(f"  First line:  {first.date}  Ref #{first.ref}  ${first.amount:,.2f}")
    print(f"  Last  line:  {last.date}  Ref #{last.ref}  ${last.amount:,.2f}")
    _hr()

    if not vendor_hint:
        _fail("could not identify vendor. Re-run with --vendor \"Exact Vendor Name\".")
        return False, {}
    if not stmt_date:
        _fail("could not identify statement date from PDF.")
        return False, {}
    if not sum_matches:
        _warn("line sum differs from Amount Due — parse may be incomplete.")

    if not _confirm("Parse looks correct — proceed to QBO lookup?", default_yes=sum_matches, skip=args.yes):
        print(_Term.color(_Term.R, "✗ aborted by user."))
        return False, {}

    # ── vendor resolve (uses alias cache when available) ────
    print()
    t0 = _phase("Resolving vendor in QBO")
    if args.no_cache:
        vendor_id, vendor_name = find_vendor_id(access, cid, args.vendor or vendor_hint)
        from_cache = False
    else:
        vendor_id, vendor_name, from_cache = find_vendor_id_cached(
            access, cid, vendor_hint, override=args.vendor)
    cache_marker = _Term.color(_Term.Y, " ★ from saved alias") if from_cache else ""
    _done(t0, f"Matched {_Term.color(_Term.BOLD, vendor_name)}  (id={vendor_id}){cache_marker}")

    # Unattended inbox mode: only auto-process vendors already confirmed in the
    # alias cache. A first-time vendor is left in the inbox so a human runs it
    # once (which caches it) — avoids reconciling against a wrong fuzzy match
    # with nobody watching.
    if getattr(args, "inbox_cached_only", False) and not from_cache and not args.vendor:
        _fail(f"vendor '{vendor_name}' is not in the alias cache yet — skipping in unattended "
              "inbox mode. Run it once manually (statement-reconcile <file>) to confirm + cache it.")
        return False, {}

    # ── confirm #2: vendor match — skipped on cache hit ─────
    if from_cache:
        print(f"  {_Term.color(_Term.DIM, '(skipping vendor confirmation — cached.)')}")
    else:
        print()
        _hr()
        print(_Term.color(_Term.BOLD, "  VENDOR MATCH CONFIRMATION"))
        _hr()
        print(f"  Statement says:  {vendor_hint}")
        print(f"  Matched QBO to:  {_Term.color(_Term.C, vendor_name)}  (id={vendor_id})")
        _hr()
        if not _confirm("Is this the correct vendor?", default_yes=True, skip=args.yes):
            print(_Term.color(_Term.R, "✗ aborted."))
            return False, {}
        if not args.no_cache:
            remember_vendor(vendor_hint, vendor_id, vendor_name)
            print(f"  {_Term.color(_Term.DIM, '(remembered this vendor — future statements will skip this step)')}")

    # ── pull bills ──────────────────────────────────────────
    print()
    t0 = _phase("Pulling open bills from QBO")
    bills = get_open_bills_for_vendor(access, cid, vendor_id)
    # As-of-stmt totals for the displayed tie-out (excludes post-stmt bills)
    bills_asof = [b for b in bills if not stmt_date or not b.txn_date or b.txn_date <= stmt_date]
    bills_post = [b for b in bills if stmt_date and b.txn_date and b.txn_date > stmt_date]
    qbo_total_all = sum(b.open_balance for b in bills)
    qbo_total = sum(b.open_balance for b in bills_asof)
    if bills_post:
        _done(t0, f"{_Term.color(_Term.BOLD, f'{len(bills)} open bills')} total "
                  f"${qbo_total_all:,.2f}  ·  "
                  f"as of {stmt_date}: {_Term.color(_Term.BOLD, f'{len(bills_asof)} bills, ${qbo_total:,.2f}')}  "
                  f"({len(bills_post)} dated after, ${sum(b.open_balance for b in bills_post):,.2f} excluded)")
    else:
        _done(t0, f"{_Term.color(_Term.BOLD, f'{len(bills_asof)} open bills')} totaling "
                  f"{_Term.color(_Term.BOLD, f'${qbo_total:,.2f}')}")

    if abs(len(bills_asof) - len(lines)) > max(5, len(lines) // 2):
        _warn(f"As-of bill counts differ a lot: stmt has {len(lines)}, QBO has {len(bills_asof)}. "
              "Could be normal (missing entries) or wrong vendor — verify before continuing.")

    # ── recently-paid bills (vendor-lag check) ──────────────
    # Data-driven lookback: extend far enough back to cover the OLDEST stmt
    # line (minus a week of padding), but never look back LESS than the
    # standard 60-day floor.
    floor_cutoff = dt.date.fromisoformat(stmt_date) - dt.timedelta(days=LAG_LOOKBACK_DAYS)
    # Only ISO-normalized dates are safe for fromisoformat(); _norm_date()
    # passes through unparseable strings unchanged, so filter those out here
    # rather than crash the run after QBO has already been queried.
    stmt_dates = [l.date for l in lines if l.date and re.match(r"^\d{4}-\d{2}-\d{2}$", l.date)]
    if stmt_dates:
        oldest_stmt = dt.date.fromisoformat(min(stmt_dates))
        padded_oldest = oldest_stmt - dt.timedelta(days=LAG_LOOKBACK_STMT_PADDING_DAYS)
        effective_cutoff = min(floor_cutoff, padded_oldest)
        cutoff_source = ("stmt-driven" if padded_oldest < floor_cutoff
                         else f"{LAG_LOOKBACK_DAYS}-day floor")
    else:
        effective_cutoff = floor_cutoff
        cutoff_source = f"{LAG_LOOKBACK_DAYS}-day floor"
    lag_cutoff = effective_cutoff.isoformat()
    t0 = _phase(f"Pulling paid bills since {lag_cutoff} ({cutoff_source}, vendor-lag check)")
    paid_bills = get_recently_paid_bills_for_vendor(access, cid, vendor_id, lag_cutoff)
    _done(t0, f"{len(paid_bills)} paid bills in lookback window", color=_Term.B)

    # ── live reconcile stream ───────────────────────────────
    print()
    t0 = _phase(f"Reconciling {len(lines)} statement lines vs {len(bills)} open + {len(paid_bills)} paid QBO bills")
    rows: List[ReconRow] = []
    counts: Dict[str, int] = {}
    for i, total, row in reconcile_iter(lines, bills, paid_bills, stmt_date):
        rows.append(row)
        counts[row.category] = counts.get(row.category, 0) + 1
        label, color = _CAT_STYLE.get(row.category, (row.category, _Term.RESET))
        running = (
            f"{_Term.color(_Term.G, '✓'+str(counts.get('MATCHED',0)))} "
            f"{_Term.color(_Term.Y, '⚠'+str(counts.get('VENDOR_TAX_VIOLATION',0)+counts.get('CLERK_AMOUNT_MISMATCH',0)))} "
            f"{_Term.color(_Term.B, '⊙'+str(counts.get('LIKELY_VENDOR_LAG',0)))} "
            f"{_Term.color(_Term.R, '✗'+str(counts.get('MISSING_IN_QBO',0)+counts.get('MISSING_ON_STATEMENT',0)))}"
        )
        suffix = f"{_Term.color(color, label)}  Ref #{row.ref:<10} {running}"
        _bar(i, total, suffix)
    _bar_end()
    _done(t0, f"Classified {len(rows)} bill comparisons")

    # ── summary ─────────────────────────────────────────────
    print()
    _hr()
    print(_Term.color(_Term.BOLD, "  RECONCILIATION SUMMARY"))
    _hr()
    print(f"  Statement total:           ${amt_due:,.2f}")
    qbo_label = f"QBO open as of {stmt_date}:" if stmt_date else "QBO open total:"
    print(f"  {qbo_label:25s}  ${qbo_total:,.2f}  ({len(bills_asof)} bills)")
    diff = amt_due - qbo_total
    diff_color = _Term.G if abs(diff) < 0.01 else (_Term.Y if abs(diff) < 100 else _Term.R)
    print(f"  Reconciling diff:          {_Term.color(diff_color, f'${diff:,.2f}')}")
    if bills_post:
        post_total = sum(b.open_balance for b in bills_post)
        print(f"  {_Term.color(_Term.DIM, f'(plus {len(bills_post)} QBO bill(s) dated after {stmt_date} totaling ${post_total:,.2f} — excluded from tie-out)')}")
    print()
    pretty: List[Tuple[str, str, str]] = [
        ("MATCHED",               "✓ Matched",                       _Term.G),
        ("VENDOR_TAX_VIOLATION",  "⚠ Vendor tax violation (8.25%)",  _Term.Y),
        ("CLERK_AMOUNT_MISMATCH", "⚠ Clerk amount mismatch",         _Term.Y),
        ("LIKELY_VENDOR_LAG",     "⊙ Likely vendor lag (paid)",      _Term.B),
        ("MISSING_IN_QBO",        "✗ Missing in QBO",                _Term.R),
        ("MISSING_ON_STATEMENT",  "✗ Missing on Statement",          _Term.R),
    ]
    for key, label, color in pretty:
        n = counts.get(key, 0)
        bucket_total = sum(r.stmt_amount or r.qbo_amount for r in rows if r.category == key)
        line = f"  {label:38s} {n:3d}   ${bucket_total:>12,.2f}"
        print(_Term.color(color, line) if n > 0 else _Term.color(_Term.DIM, line))
    _hr()

    if args.dry_run:
        print(_Term.color(_Term.DIM, "--dry-run set; no Excel written."))
        return True, counts

    # ── write Excel ─────────────────────────────────────────
    print()
    t0 = _phase("Writing Excel report")
    base_out_dir = getattr(args, "out_dir", None) or OUTDIR_DEFAULT
    out = args.out or (base_out_dir / f"Statement_Reconciliation_{stmt_date}_{_slug(vendor_name)}.xlsx")
    statement_src = pdf_path if getattr(args, "embed", False) else None
    write_excel(out, vendor_name, stmt_date, amt_due, bills, rows, statement_src=statement_src)
    _done(t0, f"Saved to {_Term.color(_Term.C, str(out))}")

    # ── clerk performance log (append one row) ──────────────
    t0 = _phase("Appending clerk-performance history")
    csv_path = append_clerk_perf(rows, vendor_name, stmt_date, amt_due)
    if csv_path:
        _done(t0, f"+1 row to {_Term.color(_Term.C, str(csv_path))}")

    # Clickable folder link (Cmd+Click in macOS Terminal/iTerm2 opens Finder)
    print(f"  {_Term.color(_Term.DIM, 'Outputs in:')} {_term_link(str(out.parent), out.parent)}")

    if not args.no_open:
        _open_file(out)
        print(f"  {_Term.color(_Term.DIM, '(Excel opened in default app)')}")
    return True, counts


def _resolve_workflow_dirs(base: Path, strict: bool = True
                           ) -> Tuple[Optional[Path], Optional[Path], Optional[Path]]:
    """Resolve (inbox, done, reconciliations) under `base`, tolerant of the
    exact folder spelling/spacing. With strict=True (the --inbox sweep) exits
    with a clear message if the share isn't mounted or the folders can't be
    found. With strict=False (a manually-passed file) returns (None, None, None)
    instead, so the caller can fall back to the default output dir and skip the
    DONE move rather than abort."""
    def _bail(msg: str) -> Tuple[None, None, None]:
        if strict:
            sys.exit(_Term.color(_Term.R, msg))
        return None, None, None

    if not base.exists():
        return _bail(f"✗ not found: {base}\n"
                     "  Is the Synology share mounted? (Finder → Go → Connect to Server)")
    entries = [d for d in base.iterdir() if d.is_dir()]

    def _find(*needles: str) -> Optional[Path]:
        for d in entries:
            norm = d.name.lower().strip().strip("-").strip()
            if any(n in norm for n in needles):
                return d
        return None

    inbox = _find("statement inbox", "inbox")
    recon = _find("reconciliation")
    if not inbox or not recon:
        found = ", ".join(sorted(d.name for d in entries)) or "(none)"
        return _bail(f"✗ couldn't find the Inbox / Reconciliations folders under {base}\n"
                     f"  Folders present: {found}")

    done = None
    for d in inbox.iterdir():
        if d.is_dir() and d.name.lower().strip() == "done":
            done = d
            break
    if done is None:
        done = inbox / "DONE"
        done.mkdir(exist_ok=True)   # safe: the archive subfolder within a confirmed inbox
    return inbox, done, recon


def _gather_inbox_files(inbox: Path) -> List[Path]:
    """Supported statement files sitting directly in the inbox (not DONE).
    Skips hidden files, Office lock files, and temp artifacts."""
    out: List[Path] = []
    for f in sorted(inbox.iterdir()):
        if not f.is_file():
            continue
        n = f.name
        if n.startswith(".") or n.startswith("~$") or n.endswith("#"):
            continue
        if f.suffix.lower() in INBOX_SUPPORTED_EXTS:
            out.append(f)
    return out


def _unique_dest(dest_dir: Path, name: str) -> Path:
    """A non-colliding path in dest_dir for `name` (adds ' (2)', ' (3)', …)."""
    if not (dest_dir / name).exists():
        return dest_dir / name
    stem, suf = Path(name).stem, Path(name).suffix
    i = 2
    while (dest_dir / f"{stem} ({i}){suf}").exists():
        i += 1
    return dest_dir / f"{stem} ({i}){suf}"


def run_inbox(args: argparse.Namespace, access: str, cid: str,
              inbox: Optional[Path], done: Optional[Path], recon: Path,
              files: List[Path]) -> int:
    """Reconcile each file → Excel into `recon` with the inbox presentation
    ([i/N] header + INBOX SUMMARY). A source file that lives in `inbox` is moved
    to `done` on success; a file passed from elsewhere is reconciled and left in
    place. Failures are never moved. Used by both the --inbox sweep and manual
    single/multi-file runs. Returns an exit code."""
    args.out = None
    args.out_dir = recon
    args.no_open = True
    args.embed = not getattr(args, "no_embed", False)
    # Interactive by default: a new/unknown vendor stops and asks the operator to
    # confirm the match right here. Only fully-unattended runs (--inbox --yes)
    # auto-skip unknowns (cached-only) so nothing is guessed with nobody watching.
    args.inbox_cached_only = args.yes

    inbox_r = inbox.resolve() if inbox else None

    moved: List[str] = []       # reconciled AND archived to DONE
    kept: List[str] = []        # reconciled but left in place (not an inbox file)
    failed: List[Tuple[str, str]] = []
    for i, f in enumerate(files, 1):
        print(_Term.color(_Term.BOLD, f"\n[{i}/{len(files)}] {f.name}"))
        try:
            ok, _counts = process_pdf(f, args, access, cid)
        except Exception as e:
            _fail(f"{f.name}: {e}")
            failed.append((f.name, str(e)))
            continue
        if not ok:
            failed.append((f.name, "skipped (see message above)"))
            continue
        # Archive to DONE only when the source actually lives in the inbox — and
        # never on a dry-run (nothing was written, so nothing should move).
        in_inbox = (done is not None and inbox_r is not None
                    and f.resolve().parent == inbox_r)
        if args.dry_run or not in_inbox:
            if not args.dry_run:
                print(_Term.color(_Term.DIM, "    ✓ reconciled — left in place (not in inbox)"))
            kept.append(f.name)
            continue
        dest = _unique_dest(done, f.name)
        try:
            shutil.move(str(f), str(dest))
            print(_Term.color(_Term.G, f"    ✓ done → {dest.name}"))
            moved.append(f.name)
        except Exception as e:
            _warn(f"reconciled but couldn't move to DONE: {e}")
            kept.append(f.name)

    print()
    _hr()
    print(_Term.color(_Term.BOLD, "  INBOX SUMMARY"))
    _hr()
    if moved:
        print(_Term.color(_Term.G, f"  Reconciled + moved to DONE:  {len(moved)}"))
    if kept:
        note = "reconciled (dry-run)" if args.dry_run else "reconciled, left in place"
        print(_Term.color(_Term.G, f"  {note+':':28} {len(kept)}"))
    if not moved and not kept:
        print(_Term.color(_Term.DIM, "  Reconciled:                  0"))
    if failed:
        print(_Term.color(_Term.R, f"  Left for a human:            {len(failed)}"))
        for name, why in failed:
            print(_Term.color(_Term.DIM, f"    · {name} — {why}"))
    print(f"\n  {_Term.color(_Term.BOLD, 'Excels:')} {_term_link(str(recon), recon)}")
    _hr()
    return 0 if not failed else 1


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("pdf", type=Path, nargs="*",
                   help="Path(s) to statement file(s) — accepts .pdf, .xlsx, .xls, or .png/.jpg "
                        "(images go through Tesseract OCR; .xls files auto-converted via xlrd)")
    p.add_argument("--vendor", default="",
                   help="Override vendor name. REQUIRED for .xlsx files (Excel has no vendor in body). "
                        "Otherwise guessed from PDF. In batch mode applies to ALL files — use with care.")
    p.add_argument("--stmt-date", default="", metavar="YYYY-MM-DD",
                   help="Statement as-of date. REQUIRED for .xlsx files if you want anything other "
                        "than today's date. PDFs read this from the body automatically. Format: YYYY-MM-DD.")
    p.add_argument("--out", type=Path, default=None, help="Override output xlsx path "
                                                          "(single-file mode only)")
    p.add_argument("--dry-run", action="store_true", help="Print findings without writing Excel")
    p.add_argument("--yes", action="store_true", help="Skip confirmation prompts")
    p.add_argument("--no-open", action="store_true", help="Do not auto-open the Excel report")
    p.add_argument("--no-color", action="store_true", help="Disable ANSI colors / progress bar")
    p.add_argument("--list-aliases", action="store_true", help="Print saved vendor aliases and exit")
    p.add_argument("--forget-vendor", default="", metavar="NAME",
                   help="Remove a saved alias by PDF name or QBO display name, then exit")
    p.add_argument("--no-cache", action="store_true", help="Do not use or update the vendor alias cache")
    p.add_argument("--inbox", action="store_true",
                   help="Automation: reconcile every statement in the Synology Statement Inbox, "
                        "write each Excel to Reconciliations, and move the source file to DONE.")
    p.add_argument("--inbox-root", type=Path, default=None,
                   help="Override the inbox workflow root (default: the Synology Vendor Statements folder).")
    p.add_argument("--embed", action="store_true",
                   help="Embed the source statement's pages as a 'Statement' tab (self-contained xlsx).")
    p.add_argument("--no-embed", action="store_true",
                   help="In --inbox mode, do NOT embed the statement image (smaller files).")
    args = p.parse_args()

    if args.no_color:
        _Term.disable()

    # ── alias-cache utility modes ───────────────────────────
    if args.list_aliases:
        aliases = load_aliases()
        if not aliases:
            print("(no saved aliases yet)")
            return 0
        print(_Term.color(_Term.BOLD, "SAVED VENDOR ALIASES"))
        for key, v in sorted(aliases.items()):
            print(f"  {v.get('pdf_name','?'):42s} → {_Term.color(_Term.C, v.get('qbo_name','?'))}  "
                  f"(id={v.get('qbo_id','?')}, saved {v.get('saved','?')})")
        return 0

    if args.forget_vendor:
        if forget_vendor(args.forget_vendor):
            print(_Term.color(_Term.G, f"✓ forgot alias for: {args.forget_vendor}"))
            return 0
        print(_Term.color(_Term.Y, f"⚠ no alias matched: {args.forget_vendor}"))
        return 1

    # ── inbox automation mode ───────────────────────────────
    if args.inbox:
        base = args.inbox_root or INBOX_ROOT
        inbox, done, recon = _resolve_workflow_dirs(base)
        files = _gather_inbox_files(inbox)
        _hr()
        print(_Term.color(_Term.BOLD, "  STATEMENT RECONCILER  ·  INBOX SWEEP"))
        _hr()
        print(f"  Inbox:  {inbox}")
        print(f"  Out:    {recon}")
        print(f"  Done:   {done}\n")
        if not files:
            print("  Inbox empty — nothing to reconcile.")
            return 0
        print(f"  Found {len(files)} file(s): {', '.join(f.name for f in files)}\n")
        if args.dry_run:
            for f in files:
                print(_Term.color(_Term.DIM,
                    f"  DRY-RUN {f.name}: would reconcile → {recon.name}/, "
                    f"then move source → {done.name}/"))
            print(_Term.color(_Term.DIM, "\n  (dry-run: no QBO calls, nothing written or moved.)"))
            return 0
        t0 = _phase("Authenticating to QBO (Touch ID may prompt)")
        access, cid = load_credentials()
        _done(t0, "Authenticated")
        return run_inbox(args, access, cid, inbox, done, recon, files)

    if not args.pdf:
        p.error("Statement file path required (.pdf / .xlsx / .png) — or use --list-aliases / --forget-vendor.")

    pdf_paths: List[Path] = [Path(p) for p in args.pdf]
    for pp in pdf_paths:
        if not pp.exists():
            sys.exit(_Term.color(_Term.R, f"✗ not found: {pp}"))

    # Manually-passed files get the SAME treatment as an --inbox sweep: each is
    # reconciled, its Excel written to the Reconciliations folder, and — if the
    # source actually lives in the inbox — the original archived to DONE. The run
    # ends with the INBOX SUMMARY. The inbox workflow dirs are resolved leniently
    # (strict=False): if the Synology share isn't mounted we fall back to the
    # default output dir and simply skip the archive move.
    if args.out:
        _warn("--out is ignored in inbox-style single-file mode; "
              "the Excel goes to the Reconciliations folder.")
    base = args.inbox_root or INBOX_ROOT
    inbox, done, recon = _resolve_workflow_dirs(base, strict=False)
    if recon is None:
        recon = OUTDIR_DEFAULT   # share not mounted → local default, no DONE move

    print(_Term.color(_Term.BOLD, "━" * min(60, _width())))
    if len(pdf_paths) == 1:
        print(_Term.color(_Term.BOLD, "  STATEMENT RECONCILER"))
    else:
        print(_Term.color(_Term.BOLD, f"  STATEMENT RECONCILER  ·  {len(pdf_paths)} FILES"))
    print(_Term.color(_Term.BOLD, "━" * min(60, _width())))
    print(f"  Out:   {recon}")
    if done:
        print(f"  Done:  {done}  (only files already in the inbox are moved)")
    print()
    if args.dry_run:
        print(_Term.color(_Term.DIM,
            "  (dry-run: reconciles and prints findings, but writes no Excel and moves nothing.)\n"))

    t0 = _phase("Authenticating to QBO (Touch ID may prompt)")
    access, cid = load_credentials()
    _done(t0, "Authenticated")

    return run_inbox(args, access, cid, inbox, done, recon, pdf_paths)



if __name__ == "__main__":
    sys.exit(main())
